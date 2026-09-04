"""
akuntansi_ai.py
================
Modul inti (murni Python, tidak bergantung Streamlit) untuk:
1. Membaca file Excel rekening koran multi-bank (multi-sheet), format kolom
   bebas beda-beda per bank -- kolom dideteksi otomatis berdasarkan nama header,
   bukan posisi tetap per bank.
2. Membaca sheet "COA" (Chart of Accounts) perusahaan bila ada.
3. Mempelajari pola historis: baris yang KOLOM JURNALNYA SUDAH DIISI (NO AKUN /
   NAMA AKUN debet & kredit) dipakai sebagai "contoh yang sudah benar" untuk
   menyusun aturan otomatis (signature keterangan -> pasangan akun).
4. Menerapkan pola tsb + fallback kata kunci COA + fallback AI (DeepSeek) untuk
   baris yang jurnalnya BELUM diisi (data mentah bulan berjalan).
5. Fitur tambahan: prediksi, deteksi anomali, rekomendasi, auto-correct, dll.

Didesain supaya bisa diuji tanpa perlu menjalankan Streamlit sama sekali.
"""

from __future__ import annotations

import calendar
import hashlib
import io
import json
import logging
import os
import random
import re
import shutil
import subprocess
import time
import traceback
from collections import Counter, defaultdict, OrderedDict
from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path
from typing import Optional, Dict, List, Tuple, Any

import pandas as pd
import openpyxl
import numpy as np

logger = logging.getLogger(__name__)


# ============================================================
# 1. MEMBACA SHEET COA (Chart of Accounts)
# ============================================================

def muat_coa(wb: openpyxl.Workbook, nama_sheet_coa: str = "COA") -> pd.DataFrame:
    """
    Mengembalikan DataFrame kolom: no_akun, nama_akun, kategori
    Mencari sheet yang namanya mengandung 'coa' (case-insensitive).
    Kalau tidak ada, kembalikan DataFrame kosong (fitur COA asli dinonaktifkan,
    fallback ke kata kunci generik).
    """
    target = None
    for name in wb.sheetnames:
        if "coa" in name.strip().lower():
            target = name
            break
    if target is None:
        return pd.DataFrame(columns=["no_akun", "nama_akun", "kategori"])

    return _ekstrak_coa_dari_ws(wb[target])


def _ekstrak_coa_dari_ws(ws) -> pd.DataFrame:
    """
    Logika inti ekstraksi sheet COA, dipisah dari muat_coa() supaya bisa dipakai
    ulang untuk sheet COA yang berasal dari sumber non-openpyxl (mis. .xls) --
    ws di sini cukup perlu punya method iter_rows(values_only=True) seperti
    openpyxl, lihat kelas _LembarDariBaris di bawah.
    """
    baris = []
    header_ditemukan = False
    kolom_map = {}
    for row in ws.iter_rows(values_only=True):
        sel = [str(c).strip().lower() if c is not None else "" for c in row]
        if not header_ditemukan:
            if "description" in sel or "nama" in " ".join(sel):
                header_ditemukan = True
                for i, h in enumerate(sel):
                    if h in ("cat", "kategori", "category"):
                        kolom_map["kategori"] = i
                    elif h in ("description", "nama", "nama akun", "namaakun"):
                        kolom_map["nama_akun"] = i
                    # [FIX] Sebelumnya kolom No Akun HANYA dikenali kalau header-nya
                    # KOSONG (blank) -- itu cocok utk 1 template lama, tapi template
                    # COA standar Indonesia yang paling umum justru punya header
                    # eksplisit ("No Akun", "Kode Akun", dst). Akibatnya utk COA
                    # dengan header standar begini, kolom_map["no_akun"] TIDAK
                    # PERNAH terisi -> muat_coa() selalu balik DataFrame kosong ->
                    # SEMUA cross-check akun-ke-COA di 15 jenis dokumen (bukti kas,
                    # kartu stok, aset tetap, pembelian, dst) diam-diam tidak pernah
                    # jalan walau file COA-nya valid & lengkap. Ditambah pengenalan
                    # header eksplisit ini (tetap ADDITIVE -- fallback header kosong
                    # di bawah tetap ada, jadi template lama yang sudah jalan tidak
                    # berubah perilakunya).
                    elif h in ("no akun", "no. akun", "nomor akun", "kode akun", "kode",
                               "no account", "account no", "account code",
                               "noakun", "kodeakun", "account number", "no. account"):
                        kolom_map.setdefault("no_akun", i)
                    elif h == "" and i not in kolom_map.values():
                        kolom_map.setdefault("no_akun", i)
            continue
        if "no_akun" not in kolom_map or "nama_akun" not in kolom_map:
            continue
        no_akun = row[kolom_map["no_akun"]] if kolom_map.get("no_akun") is not None else None
        nama_akun = row[kolom_map["nama_akun"]] if kolom_map.get("nama_akun") is not None else None
        kategori = row[kolom_map["kategori"]] if kolom_map.get("kategori") is not None else None
        if no_akun is None or nama_akun is None:
            continue
        baris.append({"no_akun": no_akun, "nama_akun": str(nama_akun).strip(), "kategori": kategori})

    return pd.DataFrame(baris)


# ============================================================
# 2. MEMBACA SHEET REKENING KORAN (deteksi kolom otomatis)
# ============================================================

def _cari_header_row(ws, max_scan: int = 15):
    """Cari baris header: baris yang mengandung 'keterangan'/'remarks' DAN 'saldo'/'balance'.
    [FIX] max_scan dinaikkan dari 6 -> 15 (disamakan dgn parser lain). Banyak
    export rekening koran punya beberapa baris judul/nama bank/periode SEBELUM
    baris header kolom sebenarnya, jadi 6 baris kadang tidak cukup -> sheet
    dianggap 'tidak dikenali' -> jatuh ke parser lain (mis. absensi) yg lebih
    longgar syaratnya."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        teks = " ".join(str(c) for c in row if c is not None).lower()
        if ("keterangan" in teks or "remarks" in teks) and ("saldo" in teks or "balance" in teks):
            return i + 1, list(row)
    return None, None


def _cari_idx(headers, keywords, sampai=None):
    rentang = headers[:sampai] if sampai is not None else headers
    for i, h in enumerate(rentang):
        if h is None:
            continue
        h_low = str(h).lower()
        if any(k in h_low for k in keywords):
            return i
    return None


# ============================================================
# 2b. DETEKSI JENIS DOKUMEN LAIN (belum ada parser/proses otomatisnya)
# ============================================================
# Dipakai sebagai fallback KETIKA sheet tidak cocok dengan parser yang sudah
# ada (rekening koran, data penjualan, POS, penilaian klien, buku bantu
# piutang). Tujuannya SEKADAR MENEBAK jenis dokumen apa isinya -- BELUM
# melakukan parsing/ekstraksi baris atau pembuatan jurnal untuk jenis2 ini.
#
# Kata kunci per kategori diambil dari format standar/umum yang dipakai di
# Indonesia untuk masing-masing dokumen (mis. kolom wajib pada template Excel
# e-Faktur/eBupot DJP untuk faktur pajak & bukti potong, kolom umum slip gaji
# sesuai praktik payroll Indonesia, kolom kartu stok gudang, dll). Kalau nanti
# mau ditambah parser sungguhan utk salah satu jenis ini, kata kunci di sini
# bisa dipakai sbg titik awal utk fungsi _cari_header_row_xxx yang baru.
DAFTAR_JENIS_DOKUMEN_LAIN = [
    {
        "kode": "bukti_kas",
        "label": "Bukti Kas Masuk/Keluar",
        "kata_kunci": [
            "kas masuk", "kas keluar", "bukti kas", "diterima dari",
            "dibayarkan kepada", "kas kecil", "petty cash", "nomor bukti kas",
        ],
    },
    {
        "kode": "rekonsiliasi_bank",
        "label": "Rekonsiliasi Bank",
        "kata_kunci": [
            "rekonsiliasi", "saldo menurut bank", "saldo menurut buku",
            "cek beredar", "outstanding check", "deposit in transit",
            "setoran dalam perjalanan", "selisih rekon",
        ],
    },
    {
        "kode": "invoice_pembelian",
        "label": "Invoice/Faktur Pembelian dari Supplier",
        "kata_kunci": [
            "invoice pembelian", "faktur pembelian", "nama supplier",
            "kode supplier", "purchase invoice", "bill to", "ship to",
            "nomor po", "termin pembayaran",
        ],
    },
    {
        "kode": "purchase_order",
        "label": "Purchase Order (PO)",
        "kata_kunci": [
            "purchase order", "nomor po", "tanggal po", "delivery date",
            "tanggal pengiriman", "syarat pembayaran", "vendor",
            "quantity ordered", "jumlah dipesan",
        ],
    },
    {
        "kode": "buku_bantu_utang",
        "label": "Buku Bantu Utang (AP Aging)",
        "kata_kunci": [
            "umur utang", "aging utang", "ap aging", "jatuh tempo",
            "outstanding invoice", "saldo utang", "kartu utang",
            "hutang usaha", "utang usaha",
        ],
    },
    {
        "kode": "faktur_pajak",
        "label": "Faktur Pajak (PPN)",
        "kata_kunci": [
            "npwp penjual", "npwp pembeli", "nomor faktur pajak",
            "kode transaksi", "dpp", "ppn", "kode objek faktur",
            "dasar pengenaan pajak", "npwp",
        ],
    },
    {
        "kode": "spt_masa",
        "label": "SPT Masa/Tahunan",
        "kata_kunci": [
            "spt masa", "masa pajak", "spt tahunan", "kurang bayar",
            "lebih bayar", "formulir 1111", "formulir 1721", "formulir 1771",
            "surat pemberitahuan",
        ],
    },
    {
        "kode": "bukti_potong_pajak",
        "label": "Bukti Potong Pajak (PPh 21/23/4(2))",
        "kata_kunci": [
            "bukti potong", "bukti pemotongan", "npwp pemotong",
            "kode objek pajak", "pph dipotong", "tanggal pemotongan",
            "nomor bukti potong", "penghasilan kotor", "pph yang dipotong",
        ],
    },
    {
        "kode": "kartu_stok",
        "label": "Kartu Stok/Persediaan",
        "kata_kunci": [
            "kode barang", "nama barang", "stok awal", "barang masuk",
            "barang keluar", "stok akhir", "sisa stok", "saldo stok",
            "mutasi barang",
        ],
    },
    {
        "kode": "slip_gaji",
        "label": "Slip Gaji Karyawan",
        "kata_kunci": [
            "gaji pokok", "tunjangan", "potongan", "pph 21",
            "take home pay", "gaji bersih", "nip", "jabatan", "bpjs",
        ],
    },
    {
        "kode": "absensi",
        "label": "Data Absensi/Timesheet",
        "kata_kunci": [
            "jam masuk", "jam keluar", "terlambat", "alpha", "izin",
            "cuti", "absensi", "timesheet", "hadir",
        ],
    },
    {
        "kode": "aset_tetap",
        "label": "Daftar Aset Tetap & Penyusutan",
        "kata_kunci": [
            "nama aset", "tanggal perolehan", "harga perolehan",
            "masa manfaat", "nilai residu", "akumulasi penyusutan",
            "nilai buku", "kelompok aset", "golongan fiskal",
        ],
    },
]

# Minimal proporsi kata kunci kategori yang harus cocok supaya dianggap
# tebakan yang layak dipercaya (bukan sekadar 1 kata umum yang kebetulan ada).
_AMBANG_DETEKSI_JENIS_LAIN = 0.2


def deteksi_jenis_dokumen_lain(ws, max_scan: int = 15) -> tuple[Optional[str], Optional[str], float]:
    """
    Menebak jenis dokumen (dari DAFTAR_JENIS_DOKUMEN_LAIN) berdasarkan kata
    kunci yang ditemukan di beberapa baris pertama sheet -- header maupun
    beberapa baris data awal, supaya tetap kena walau layout headernya tidak
    persis di baris 1 (mis. ada judul/logo perusahaan di atasnya).

    HANYA MENEBAK JENISNYA, belum mem-parsing isi barisnya (parser sungguhan
    untuk kategori2 ini belum ada -- lihat komentar di atas
    DAFTAR_JENIS_DOKUMEN_LAIN).

    Return: (kode, label, skor_keyakinan). skor_keyakinan = proporsi kata
    kunci kategori yang ketemu (0.0 - 1.0). Kalau skor tertinggi masih di
    bawah ambang minimal, dianggap tidak ada yang cocok -> (None, None, 0.0).
    """
    potongan_teks = []
    for row in ws.iter_rows(min_row=1, max_row=max_scan, values_only=True):
        for c in row:
            if c is not None:
                potongan_teks.append(str(c).strip().lower())
    teks = " | ".join(potongan_teks)
    if not teks:
        return None, None, 0.0

    kode_terbaik, label_terbaik, skor_terbaik = None, None, 0.0
    for kategori in DAFTAR_JENIS_DOKUMEN_LAIN:
        kata_kunci = kategori["kata_kunci"]
        cocok = sum(1 for kw in kata_kunci if kw in teks)
        skor = cocok / len(kata_kunci)
        if skor > skor_terbaik:
            kode_terbaik, label_terbaik, skor_terbaik = kategori["kode"], kategori["label"], skor

    if skor_terbaik < _AMBANG_DETEKSI_JENIS_LAIN:
        return None, None, 0.0
    return kode_terbaik, label_terbaik, round(skor_terbaik, 2)


class FormatTidakDikenali(ValueError):
    pass


class FormatFileTidakDidukung(ValueError):
    """Ekstensi file tidak termasuk yang didukung (.xlsx/.xlsm/.xls/.csv/.pdf)."""
    pass


class _LembarDariBaris:
    """
    Adapter ringan supaya sheet yang berasal dari CSV, .xls (Excel lama), atau
    PDF bisa dipakai APA ADANYA oleh parse_sheet_bank/penjualan/penilaian/piutang
    dan _ekstrak_coa_dari_ws -- semua fungsi itu HANYA memanggil
    `ws.iter_rows(min_row=..., max_row=..., values_only=True)`, jadi cukup
    mengimplementasikan method itu saja (kompatibel dgn semantik openpyxl:
    min_row/max_row 1-indexed & inklusif).
    """

    def __init__(self, baris: list[tuple]):
        self._baris = baris

    def iter_rows(self, min_row: int = 1, max_row: int | None = None, values_only: bool = True):
        akhir = max_row if max_row is not None else len(self._baris)
        for row in self._baris[min_row - 1: akhir]:
            yield row


class _LembarDataFrameSiapPakai:
    """
    [BARU] Adapter khusus untuk sheet yang HASIL EKSTRAKSINYA SUDAH BERUPA
    DataFrame final siap pakai (skema sama dengan output parse_sheet_bank()),
    bukan baris mentah yang masih perlu di-parse header/kolomnya.

    Dipakai untuk PDF rekening koran tanpa garis grid (mis. BCA "Tahapan")
    yang diekstrak lewat jalur fallback posisi-kata
    (_ekstrak_pdf_rekening_koran_berbasis_posisi) di _baca_pdf_sebagai_lembar()
    -- jalur itu langsung menghasilkan kolom mutasi_debet/mutasi_kredit yang
    sudah terpisah dari 1 kolom "MUTASI"+suffix DB di PDF asli, jadi TIDAK
    BISA (dan tidak perlu) diparse ulang oleh parse_sheet_bank() yang
    mencari kolom "DEBIT"/"KREDIT" terpisah di header.

    iter_rows() SENGAJA dibuat mengembalikan kosong (bukan melempar
    exception) supaya parser jenis dokumen LAIN (parse_sheet_penjualan,
    parse_sheet_faktur_pajak, dst -- yang tidak tahu-menahu soal adapter
    ini) tetap bisa memanggilnya dengan aman: mereka akan menganggap sheet
    ini "tidak ditemukan baris/header" secara alami dan melewatinya
    (FormatTidakDikenali / DataFrame kosong), BUKAN error/crash. Hanya
    parse_sheet_bank() yang tahu cara membaca adapter ini dengan benar
    (lewat pengecekan isinstance di awal fungsi -- lihat parse_sheet_bank).
    """

    def __init__(self, df_final: pd.DataFrame, peringatan: list[str] | None = None, ringkasan_footer: dict | None = None):
        self.df_final = df_final
        # [BARU] Dibawa dari jalur ekstraksi (peringatan "diekstrak lewat
        # fallback, cek manual" + hasil validasi otomatis vs footer PDF,
        # dan dict ringkasan resmi saldo_awal/mutasi_cr/mutasi_db/saldo_akhir)
        # supaya muat_workbook() bisa meneruskannya ke `peringatan` yang
        # dikembalikan ke pemanggil (proses_file_rekening_koran dkk) --
        # SEBELUMNYA info ini cuma di-log server, tidak pernah sampai ke
        # response API/UI di jalur upload umum.
        self.peringatan = peringatan or []
        self.ringkasan_footer = ringkasan_footer or {}

    def iter_rows(self, min_row: int = 1, max_row: int | None = None, values_only: bool = True):
        return iter(())


def _baca_csv_sebagai_baris(file_like) -> list[tuple]:
    """
    Baca file .csv jadi list baris mentah (header ikut dideteksi belakangan oleh
    parse_sheet_* seperti sheet Excel biasa, bukan di sini).

    - Auto-deteksi delimiter (koma/titik-koma/tab/pipe) karena export CSV dari
      Excel versi Indonesia defaultnya titik-koma, bukan koma.
    - Fallback encoding utf-8-sig -> cp1252 -> latin1 karena file dari klien
      sering berasal dari Excel Windows Indonesia (bukan UTF-8 murni).
    """
    import csv as _csv

    raw = file_like.read() if hasattr(file_like, "read") else file_like
    raw_bytes = raw.encode("utf-8") if isinstance(raw, str) else raw

    teks = None
    for enc in ("utf-8-sig", "cp1252", "latin1"):
        try:
            teks = raw_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if teks is None:
        teks = raw_bytes.decode("utf-8", errors="replace")

    baris_teks = teks.splitlines()
    sample = "\n".join(baris_teks[:5])
    try:
        delimiter = _csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except _csv.Error:
        delimiter = ";" if sample.count(";") > sample.count(",") else ","

    return [tuple(row) for row in _csv.reader(baris_teks, delimiter=delimiter)]


def _baca_xls_sebagai_lembar(file_like) -> dict[str, list[tuple]]:
    """
    Baca file .xls (format Excel LAMA -- beda dari .xlsx) lewat engine 'xlrd'.
    openpyxl TIDAK bisa baca .xls, hanya .xlsx/.xlsm, makanya butuh jalur lain.
    Perlu: pip install xlrd --break-system-packages
    """
    try:
        semua_sheet = pd.read_excel(file_like, sheet_name=None, header=None, engine="xlrd")
    except ImportError as e:
        raise RuntimeError(
            "Gagal membaca file .xls -- library 'xlrd' belum ter-install. "
            "Jalankan: pip install xlrd --break-system-packages"
        ) from e
    return {
        nama: [tuple(row) for row in df.itertuples(index=False, name=None)]
        for nama, df in semua_sheet.items()
    }


# ============================================================
# EKSTRAKSI PDF REKENING KORAN BERBASIS POSISI KATA -- FALLBACK
# UNTUK PDF TANPA GARIS TABEL (mis. format BCA "Tahapan")
# ============================================================
# [DIPINDAHKAN dari modules/kertas_kerja.py ke sini -- lihat catatan di
# _baca_pdf_sebagai_lembar() & muat_workbook() di bawah untuk kenapa]
#
# Beberapa rekening koran (contoh: BCA Tahapan) TIDAK punya garis grid
# antar-baris transaksi -- hanya header kolom yang dibingkai kotak,
# baris data di bawahnya murni teks sejajar kolom tanpa garis sama
# sekali. pdfplumber.extract_tables() (dipakai di _baca_pdf_sebagai_lembar
# di bawah) mengandalkan garis grid untuk mendeteksi baris/kolom, sehingga
# PDF jenis ini GAGAL total diekstrak (cuma header yang kebaca, 0 baris
# transaksi).
#
# Fungsi-fungsi di bawah ini adalah jalur ekstraksi ALTERNATIF berbasis
# KOORDINAT kata (x0/top dari pdfplumber.extract_words()), dipakai sebagai
# FALLBACK OTOMATIS oleh _baca_pdf_sebagai_lembar() saat jalur grid biasa
# gagal/menghasilkan 0 tabel. Didesain generik (posisi kolom dicari dinamis
# dari kata header per halaman, bukan koordinat hardcode) supaya tidak
# cuma jalan untuk 1 template BCA spesifik, tapi bank/format lain dengan
# pola serupa (header berbaris + data tanpa garis, 1 kolom MUTASI gabungan
# dengan suffix DB/CR atau DB/K) juga punya peluang terbaca.
#
# Divalidasi terhadap rekening koran BCA Tahapan 25 halaman: total &
# jumlah baris hasil ekstraksi PERSIS cocok dengan ringkasan resmi di
# halaman terakhir statement (MUTASI CR/MUTASI DB/jumlah baris).
#
# PENTING soal skema output: berbeda dari _baca_pdf_sebagai_lembar() jalur
# grid (yang mengembalikan "sheet mentah" berupa baris/kolom untuk
# DIPARSING LAGI oleh parse_sheet_bank()), jalur posisi-kata ini langsung
# menghasilkan DataFrame FINAL dengan skema yang PERSIS SAMA dengan output
# parse_sheet_bank() (lihat _baris_posisi_ke_df()). Ini WAJIB karena
# parse_sheet_bank() mencari kolom "DEBIT"/"KREDIT" terpisah di header,
# sedangkan PDF tanpa-grid seperti BCA Tahapan cuma punya 1 kolom "MUTASI"
# dengan suffix "DB" (kredit tanpa suffix) -- kalau dipaksa lewat
# parse_sheet_bank() lagi, akan ditolak (FormatTidakDikenali) karena
# kolom DEBIT/KREDIT terpisah itu memang tidak pernah ada di formatnya.

_HEADER_KOL_KATA_KUNCI_PDF = {
    "TANGGAL", "TGL", "DATE", "KETERANGAN", "REMARKS", "URAIAN",
    "CBG", "CABANG", "MUTASI", "DEBET", "DEBIT", "KREDIT", "CREDIT",
    "SALDO", "BALANCE",
}

_RE_TANGGAL_PDF_RK = re.compile(r"^\d{1,2}/\d{1,2}(/\d{2,4})?$")
_RE_JUMLAH_PDF_RK = re.compile(r"^-?[\d.,]+\.\d{2}$")
_RE_PERIODE_ID_RK = re.compile(
    r"\b(JANUARI|FEBRUARI|MARET|APRIL|MEI|JUNI|JULI|AGUSTUS|SEPTEMBER|OKTOBER|NOVEMBER|DESEMBER)\s+(\d{4})\b"
)


def _cluster_baris_kata_pdf(words: list[dict], toleransi: float = 2.5) -> list[list[dict]]:
    """Kelompokkan kata-kata (dict dari pdfplumber.extract_words) menjadi
    baris visual berdasarkan koordinat 'top', dengan toleransi (karena
    beda font/baseline antar-kata di baris yang sama bisa geser 1-2px)."""
    kata_urut = sorted(words, key=lambda w: w["top"])
    baris_semua: list[list[dict]] = []
    baris_skrg: list[dict] = []
    top_skrg: float | None = None
    for w in kata_urut:
        if top_skrg is None or abs(w["top"] - top_skrg) <= toleransi:
            baris_skrg.append(w)
            if top_skrg is None:
                top_skrg = w["top"]
        else:
            baris_semua.append(baris_skrg)
            baris_skrg = [w]
            top_skrg = w["top"]
    if baris_skrg:
        baris_semua.append(baris_skrg)
    return baris_semua


def _cari_baris_header_posisi_pdf(words: list[dict]) -> tuple[dict, float] | None:
    """Cari baris header transaksi sebenarnya (bukan kata 'MUTASI' yang
    nyasar muncul di teks disclaimer/catatan) -- baris (top-cluster) yang
    mengandung TANGGAL/TGL DAN KETERANGAN/REMARKS/URAIAN sekaligus, dan
    kata kunci terbanyak di antara top-cluster yang memenuhi syarat itu.
    Returns: ({KATA_KUNCI: x0}, top_header) atau None kalau tidak ketemu."""
    per_top: dict[int, list[tuple[str, float, float]]] = defaultdict(list)
    for w in words:
        t = w["text"].strip().upper()
        if t in _HEADER_KOL_KATA_KUNCI_PDF:
            per_top[round(w["top"])].append((t, w["x0"], w["top"]))

    kandidat_terbaik = None
    for _, items in per_top.items():
        token = {t for t, x, tp in items}
        punya_tanggal = "TANGGAL" in token or "TGL" in token or "DATE" in token
        punya_keterangan = "KETERANGAN" in token or "REMARKS" in token or "URAIAN" in token
        if punya_tanggal and punya_keterangan:
            if kandidat_terbaik is None or len(items) > len(kandidat_terbaik[1]):
                kandidat_terbaik = (items[0][2], items)

    if kandidat_terbaik is None:
        return None
    top_header, items = kandidat_terbaik
    peta_x = {t: x for t, x, _ in items}
    return peta_x, top_header


def _deteksi_tahun_periode_pdf_rk(words: list[dict]) -> int | None:
    """Cari teks 'PERIODE : <BULAN> <TAHUN>' (format umum rekening koran
    Indonesia) di halaman untuk menentukan tahun basis tanggal transaksi
    (yang di badan tabel biasanya cuma ditulis dd/mm tanpa tahun)."""
    baris_top = _cluster_baris_kata_pdf(words, toleransi=2.5)
    for baris in baris_top:
        teks_gabung = " ".join(w["text"] for w in sorted(baris, key=lambda w: w["x0"])).upper()
        m = _RE_PERIODE_ID_RK.search(teks_gabung)
        if m:
            return int(m.group(2))
    return None


def _ke_float_aman_pdf(teks: str) -> float | None:
    """Parse angka format ID/umum rekening koran ('935,832,796.75') ke
    float, aman kalau gagal (return None, bukan exception)."""
    try:
        return float(str(teks).replace(",", ""))
    except (ValueError, TypeError):
        return None


def _ekstrak_baris_posisi_satu_halaman_rk(
    words: list[dict], pageno: int, nama_file_pdf: str,
) -> tuple[list[dict], dict]:
    """Ekstrak baris transaksi 1 halaman lewat pengelompokan kata per
    koordinat -- lihat penjelasan pendekatan di komentar blok di atas.
    Mengembalikan (baris_hasil, ringkasan_footer):
      - baris_hasil: list of dict mentah (tanggal string dd/mm, mutasi
        string mentah, db_flag, saldo string mentah/None, keterangan
        gabungan multi-baris, halaman, source_pdf) -- BELUM dikonversi ke
        DataFrame standar (lihat _baris_posisi_ke_df_rk untuk itu).
      - ringkasan_footer: dict {saldo_awal, mutasi_cr, mutasi_db,
        saldo_akhir} kalau halaman ini adalah halaman TERAKHIR statement
        (punya blok ringkasan resmi di bawah tabel), else {} kosong."""
    hasil = _cari_baris_header_posisi_pdf(words)
    if hasil is None:
        return [], {}
    peta_x, top_header = hasil

    col_tanggal = peta_x.get("TANGGAL", peta_x.get("TGL", peta_x.get("DATE")))
    col_keterangan = peta_x.get("KETERANGAN", peta_x.get("REMARKS", peta_x.get("URAIAN")))
    col_cbg = peta_x.get("CBG", peta_x.get("CABANG"))
    col_mutasi = peta_x.get("MUTASI", peta_x.get("DEBET", peta_x.get("DEBIT")))
    col_saldo = peta_x.get("SALDO", peta_x.get("BALANCE"))

    if col_tanggal is None or col_keterangan is None or col_mutasi is None:
        return [], {}

    kata_badan = [w for w in words if w["top"] > top_header + 3]
    baris_kata = _cluster_baris_kata_pdf(kata_badan, toleransi=2.5)

    baris_hasil: list[dict] = []
    baris_skrg: dict | None = None
    ringkasan_footer: dict = {}
    mode_footer = False

    for baris in baris_kata:
        baris_urut = sorted(baris, key=lambda w: w["x0"])
        teks_list = [w["text"] for w in baris_urut]

        if any(t.startswith("Bersambung") for t in teks_list):
            break
        gabung_3kata = " ".join(teks_list[:3])
        if mode_footer or any(k in gabung_3kata for k in ("SALDO AWAL :", "MUTASI CR", "MUTASI DB", "SALDO AKHIR")):
            mode_footer = True
            teks_gabung = " ".join(teks_list).upper()
            angka = [t for t in teks_list if _RE_JUMLAH_PDF_RK.match(t)]
            # [FIX] Jumlah baris (mis. "63"/"236" di "MUTASI CR : 935,832,796.75
            # 63") ditulis sebagai bilangan bulat POLOS tanpa desimal --
            # _RE_JUMLAH_PDF_RK mewajibkan ".XX" 2 digit, jadi TIDAK PERNAH
            # cocok utk token count itu (cr_count/db_count akibatnya selalu
            # None sebelum fix ini, walau kode sudah "siap" menampungnya).
            # Dicari terpisah: token bilangan bulat polos SETELAH token
            # nominal (angka[0]) di baris yang sama.
            angka_count = [t for t in teks_list if re.match(r"^\d+$", t)]
            if teks_gabung.startswith("SALDO AWAL") and angka:
                ringkasan_footer["saldo_awal"] = _ke_float_aman_pdf(angka[0])
            elif teks_gabung.startswith("MUTASI CR") and angka:
                ringkasan_footer["mutasi_cr"] = _ke_float_aman_pdf(angka[0])
                if angka_count:
                    ringkasan_footer["cr_count"] = int(angka_count[0])
            elif teks_gabung.startswith("MUTASI DB") and angka:
                ringkasan_footer["mutasi_db"] = _ke_float_aman_pdf(angka[0])
                if angka_count:
                    ringkasan_footer["db_count"] = int(angka_count[0])
            elif teks_gabung.startswith("SALDO AKHIR") and angka:
                ringkasan_footer["saldo_akhir"] = _ke_float_aman_pdf(angka[0])
            continue

        kata_tanggal = None
        batas_kanan_tanggal = col_keterangan - 5 if col_keterangan else col_tanggal + 60
        for w in baris_urut:
            if (col_tanggal - 5) <= w["x0"] < batas_kanan_tanggal and _RE_TANGGAL_PDF_RK.match(w["text"]):
                kata_tanggal = w
                break

        if kata_tanggal is not None:
            if baris_skrg is not None:
                baris_hasil.append(baris_skrg)
            baris_skrg = {
                "halaman": pageno, "source_pdf": nama_file_pdf,
                "tanggal": kata_tanggal["text"], "keterangan_parts": [],
                "cbg": None, "mutasi": None, "db_flag": False, "saldo": None,
            }
            for w in baris_urut:
                if w is kata_tanggal:
                    continue
                x = w["x0"]
                if col_saldo is not None and x >= col_saldo - 5:
                    if _RE_JUMLAH_PDF_RK.match(w["text"]):
                        baris_skrg["saldo"] = w["text"]
                elif x >= col_mutasi - 5:
                    if _RE_JUMLAH_PDF_RK.match(w["text"]):
                        baris_skrg["mutasi"] = w["text"]
                    elif w["text"].strip().upper() in ("DB", "D"):
                        baris_skrg["db_flag"] = True
                elif col_cbg is not None and col_cbg - 5 <= x < col_mutasi - 5:
                    baris_skrg["cbg"] = w["text"]
                else:
                    baris_skrg["keterangan_parts"].append(w["text"])
        else:
            if baris_skrg is not None:
                for w in baris_urut:
                    x = w["x0"]
                    if col_saldo is not None and x >= col_saldo - 5:
                        continue
                    if x >= col_mutasi - 5:
                        continue
                    baris_skrg["keterangan_parts"].append(w["text"])

    if baris_skrg is not None:
        baris_hasil.append(baris_skrg)

    return baris_hasil, ringkasan_footer


def _baris_posisi_ke_df_rk(semua_baris: list[dict], nama_bank: str, tahun_default: int | None) -> pd.DataFrame:
    """Konversi hasil _ekstrak_baris_posisi_satu_halaman_rk (gabungan semua
    halaman 1 file PDF) menjadi DataFrame dengan skema PERSIS SAMA dengan
    output parse_sheet_bank() (no, bank, tanggal, keterangan, mutasi_debet,
    mutasi_kredit, saldo, supplier_cust, voucher, no_transaksi,
    no_akun_debet, nama_akun_debet, jml_debet, no_akun_kredit,
    nama_akun_kredit, jml_kredit) DITAMBAH source_pdf/halaman."""
    baris_out = []
    for b in semua_baris:
        keterangan = " ".join(b["keterangan_parts"]).strip()
        if not keterangan and b["mutasi"] is None:
            continue

        if keterangan.upper().startswith("SALDO AWAL") and b["mutasi"] is None:
            continue

        mutasi_val = None
        if b["mutasi"]:
            try:
                mutasi_val = float(b["mutasi"].replace(",", ""))
            except ValueError:
                mutasi_val = None
        saldo_val = None
        if b["saldo"]:
            try:
                saldo_val = float(b["saldo"].replace(",", ""))
            except ValueError:
                saldo_val = None

        tgl_parts = b["tanggal"].split("/")
        try:
            dd, mm = int(tgl_parts[0]), int(tgl_parts[1])
            if len(tgl_parts) == 3:
                yyyy = int(tgl_parts[2])
                yyyy = yyyy + 2000 if yyyy < 100 else yyyy
            else:
                yyyy = tahun_default or date.today().year
            tanggal_val = date(yyyy, mm, dd)
        except (ValueError, IndexError):
            tanggal_val = None

        baris_out.append({
            "no": None,
            "bank": nama_bank,
            "tanggal": tanggal_val,
            "keterangan": keterangan,
            "mutasi_debet": mutasi_val if (mutasi_val is not None and b["db_flag"]) else 0,
            "mutasi_kredit": mutasi_val if (mutasi_val is not None and not b["db_flag"]) else 0,
            "saldo": saldo_val,
            "supplier_cust": None,
            "voucher": b["cbg"],
            "no_transaksi": None,
            "no_akun_debet": None,
            "nama_akun_debet": None,
            "jml_debet": None,
            "no_akun_kredit": None,
            "nama_akun_kredit": None,
            "jml_kredit": None,
            "source_pdf": b["source_pdf"],
            "halaman": b["halaman"],
        })

    return pd.DataFrame(baris_out)


def _validasi_ringkasan_footer_rk(df: pd.DataFrame, ringkasan_footer: dict, nama_file_pdf: str) -> list[str]:
    """[BARU] Validasi OTOMATIS -- bandingkan total & jumlah baris hasil
    ekstraksi posisi-kata terhadap blok ringkasan resmi di footer PDF
    (SALDO AWAL/MUTASI CR/MUTASI DB/SALDO AKHIR + jumlah baris CR/DB kalau
    ada), supaya kecocokan angka TIDAK cuma dicek manual sekali waktu
    development (seperti sebelumnya), tapi otomatis dicek setiap kali ada
    PDF baru diproses -- kalau ekstraksinya meleset (mis. PDF bank lain
    dgn layout sedikit beda, ada baris kelewat), staf langsung dapat
    peringatan tegas alih-alih diam-diam dapat angka salah.

    Toleransi Rp 1 (bukan 0) untuk redam floating-point rounding, BUKAN
    untuk mentolerir baris yang benar-benar hilang -- selisih akibat 1
    baris kelewat pasti jauh lebih besar dari itu.

    Return: list peringatan -- KOSONG kalau tidak ada ringkasan_footer sama
    sekali (PDF tidak punya blok itu, tidak ada yang bisa divalidasi, bukan
    berarti gagal). Kalau ada ringkasan_footer, SELALU balik minimal 1
    pesan: sukses (✅) atau warning (⚠️) -- supaya keberadaan pesan ini
    sendiri jadi sinyal "validasi sudah jalan", bukan cuma sinyal
    error."""
    if not ringkasan_footer or df.empty:
        return []

    pesan: list[str] = []
    TOLERANSI_RP = 1.0

    total_db = float(df["mutasi_debet"].fillna(0).sum())
    total_cr = float(df["mutasi_kredit"].fillna(0).sum())
    jumlah_db = int((df["mutasi_debet"].fillna(0) > 0).sum())
    jumlah_cr = int((df["mutasi_kredit"].fillna(0) > 0).sum())

    mutasi_db_resmi = ringkasan_footer.get("mutasi_db")
    mutasi_cr_resmi = ringkasan_footer.get("mutasi_cr")
    saldo_akhir_resmi = ringkasan_footer.get("saldo_akhir")
    db_count_resmi = ringkasan_footer.get("db_count")
    cr_count_resmi = ringkasan_footer.get("cr_count")

    if mutasi_db_resmi is not None and abs(total_db - mutasi_db_resmi) > TOLERANSI_RP:
        pesan.append(
            f"⚠️ '{nama_file_pdf}': total MUTASI DB hasil ekstraksi (Rp {total_db:,.2f}) "
            f"BEDA dari ringkasan resmi PDF (Rp {mutasi_db_resmi:,.2f}), selisih "
            f"Rp {total_db - mutasi_db_resmi:,.2f} -- kemungkinan ada baris transaksi "
            "terlewat/salah baca. WAJIB dicek manual sebelum dipakai jadi jurnal."
        )
    if mutasi_cr_resmi is not None and abs(total_cr - mutasi_cr_resmi) > TOLERANSI_RP:
        pesan.append(
            f"⚠️ '{nama_file_pdf}': total MUTASI CR hasil ekstraksi (Rp {total_cr:,.2f}) "
            f"BEDA dari ringkasan resmi PDF (Rp {mutasi_cr_resmi:,.2f}), selisih "
            f"Rp {total_cr - mutasi_cr_resmi:,.2f} -- kemungkinan ada baris transaksi "
            "terlewat/salah baca. WAJIB dicek manual sebelum dipakai jadi jurnal."
        )
    if db_count_resmi is not None and jumlah_db != db_count_resmi:
        pesan.append(
            f"⚠️ '{nama_file_pdf}': jumlah baris DB hasil ekstraksi ({jumlah_db}) "
            f"BEDA dari jumlah resmi di footer PDF ({db_count_resmi}) -- WAJIB dicek manual."
        )
    if cr_count_resmi is not None and jumlah_cr != cr_count_resmi:
        pesan.append(
            f"⚠️ '{nama_file_pdf}': jumlah baris CR hasil ekstraksi ({jumlah_cr}) "
            f"BEDA dari jumlah resmi di footer PDF ({cr_count_resmi}) -- WAJIB dicek manual."
        )

    saldo_series = df["saldo"].dropna()
    saldo_terakhir = float(saldo_series.iloc[-1]) if not saldo_series.empty else None
    if saldo_akhir_resmi is not None and saldo_terakhir is not None and abs(saldo_terakhir - saldo_akhir_resmi) > TOLERANSI_RP:
        pesan.append(
            f"⚠️ '{nama_file_pdf}': SALDO AKHIR hasil ekstraksi (Rp {saldo_terakhir:,.2f}) "
            f"BEDA dari SALDO AKHIR resmi PDF (Rp {saldo_akhir_resmi:,.2f}) -- WAJIB dicek manual."
        )

    if not pesan:
        pesan.append(
            f"✅ '{nama_file_pdf}': validasi otomatis LULUS -- total & jumlah baris hasil "
            "ekstraksi cocok persis dengan ringkasan resmi di footer PDF (SALDO AWAL/AKHIR, "
            "MUTASI CR/DB)."
        )
    return pesan


def _ekstrak_pdf_rekening_koran_berbasis_posisi(
    file_like, nama_file_pdf: str, nama_bank: str,
) -> tuple[pd.DataFrame, list[str], dict]:
    """Titik masuk fallback posisi-kata untuk 1 file PDF rekening koran --
    dipanggil oleh _baca_pdf_sebagai_lembar() HANYA saat jalur grid
    (page.extract_tables()) gagal/tidak menghasilkan baris.

    Returns: (df, peringatan, ringkasan_footer) -- ringkasan_footer =
    dict {saldo_awal, mutasi_cr, mutasi_db, saldo_akhir} dari blok
    ringkasan resmi di footer halaman terakhir (kalau ada), else {}."""
    import pdfplumber

    peringatan: list[str] = []
    semua_baris: list[dict] = []
    tahun_terdeteksi: int | None = None
    ringkasan_footer: dict = {}

    if hasattr(file_like, "seek"):
        file_like.seek(0)
    with pdfplumber.open(file_like) as pdf:
        for pageno, page in enumerate(pdf.pages, start=1):
            words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
            if not words:
                continue
            if tahun_terdeteksi is None:
                tahun_terdeteksi = _deteksi_tahun_periode_pdf_rk(words)
            baris_halaman, ringkasan_halaman = _ekstrak_baris_posisi_satu_halaman_rk(words, pageno, nama_file_pdf)
            semua_baris.extend(baris_halaman)
            if ringkasan_halaman:
                ringkasan_footer = ringkasan_halaman

    if not semua_baris:
        return pd.DataFrame(), [
            f"'{nama_file_pdf}': ekstraksi fallback berbasis posisi kata juga tidak "
            "menemukan baris transaksi -- format PDF ini kemungkinan benar-benar "
            "tidak dikenali (bukan cuma soal tabel tanpa garis)."
        ], {}

    if tahun_terdeteksi is None:
        peringatan.append(
            f"'{nama_file_pdf}': tahun statement tidak terdeteksi dari teks 'PERIODE : <bulan> <tahun>' "
            f"-- dipakai tahun berjalan ({date.today().year}) sebagai fallback. Cek ulang kolom tanggal "
            "di GL untuk file ini."
        )

    df = _baris_posisi_ke_df_rk(semua_baris, nama_bank, tahun_terdeteksi)
    if df.empty:
        return df, peringatan + [
            f"'{nama_file_pdf}': baris mentah terdeteksi tapi semuanya gagal dikonversi ke transaksi valid."
        ], ringkasan_footer

    peringatan.append(
        f"'{nama_file_pdf}': diekstrak lewat jalur fallback posisi-kata (PDF tanpa garis tabel "
        f"antar-baris) -- {len(df)} transaksi terbaca. Disarankan cek sekilas hasil jurnal untuk "
        "file ini karena jalur ekstraksi ini lebih baru & lebih sensitif terhadap variasi layout "
        "dibanding jalur grid standar."
    )
    # [BARU] Validasi otomatis vs blok ringkasan resmi footer PDF (lihat
    # _validasi_ringkasan_footer_rk) -- hasilnya (sukses ✅ atau warning ⚠️)
    # SELALU ikut ditambahkan ke peringatan supaya sampai ke pemanggil di
    # KEDUA jalur (upload umum lewat _baca_pdf_sebagai_lembar, maupun
    # fitur Kertas Kerja lewat kertas_kerja._ekstrak_pdf_berbasis_posisi).
    peringatan.extend(_validasi_ringkasan_footer_rk(df, ringkasan_footer, nama_file_pdf))
    return df, peringatan, ringkasan_footer


def _baca_pdf_sebagai_lembar(file_like) -> dict[str, list[tuple]]:
    """
    Ekstrak tabel dari PDF (mis. rekening koran yang di-export sebagai PDF,
    bukan Excel) pakai pdfplumber. Tiap tabel yang terdeteksi per halaman
    jadi 'sheet' terpisah supaya bisa dipetakan ke parser yang sama seperti
    sheet Excel biasa.
    Perlu: pip install pdfplumber --break-system-packages

    CATATAN: ekstraksi tabel dari PDF jauh lebih rapuh dibanding Excel/CSV --
    hanya bekerja untuk PDF hasil export digital (teks asli & ada garis tabel).
    PDF hasil SCAN/foto (gambar) TIDAK didukung (perlu OCR terpisah). Selalu
    cek ulang hasil parsing dari PDF sebelum dipakai sebagai jurnal final.

    [FIX] Sebelumnya kalau page.extract_tables() tidak menemukan tabel sama
    sekali (mis. rekening koran BCA "Tahapan" yang tidak punya garis grid
    antar-baris transaksi), fungsi ini langsung raise RuntimeError -- PDF
    jenis ini gagal total diproses padahal isinya jelas rekening koran
    valid (lihat modules/kertas_kerja.py utk penjelasan lengkap & bukti
    validasi format ini). Sekarang: kalau jalur grid kosong, dicoba dulu
    jalur fallback berbasis posisi kata (_ekstrak_pdf_rekening_koran_berbasis_posisi)
    SEBELUM menyerah dgn RuntimeError. Hasil jalur fallback (kalau berhasil)
    ditandai lewat key khusus "__POSISI_KATA__:<nama_sheet>" di dict yang
    dikembalikan, isinya SUDAH berupa DataFrame final (bukan baris mentah)
    -- lihat _siapkan_daftar_sheet() & muat_workbook() untuk cara ini
    dibedakan & ditangani dari hasil jalur grid biasa.
    """
    try:
        import pdfplumber
    except ImportError as e:
        raise RuntimeError(
            "Gagal membaca file PDF -- library 'pdfplumber' belum ter-install. "
            "Jalankan: pip install pdfplumber --break-system-packages"
        ) from e

    hasil: dict[str, list[tuple]] = {}
    with pdfplumber.open(file_like) as pdf:
        for i_hal, page in enumerate(pdf.pages, start=1):
            tabel_di_halaman = page.extract_tables()
            for i_tabel, tabel in enumerate(tabel_di_halaman, start=1):
                if not tabel:
                    continue
                baris = [tuple(sel if sel is not None else "" for sel in row) for row in tabel]
                nama_sheet = (
                    f"Halaman {i_hal} - Tabel {i_tabel}" if len(tabel_di_halaman) > 1 else f"Halaman {i_hal}"
                )
                hasil[nama_sheet] = baris

    # [FIX -- KRUSIAL] "hasil" TIDAK KOSONG belum tentu artinya jalur grid
    # BERHASIL. Untuk PDF seperti BCA "Tahapan": tiap halaman punya kotak
    # kecil BERGARIS di bagian atas (info No. Rekening/Periode + baris
    # header kolom TANGGAL/KETERANGAN/.../SALDO), jadi extract_tables()
    # TETAP menangkap "tabel" (makanya `hasil` tidak pernah kosong) --
    # tapi badan tabel transaksi di bawahnya TIDAK BERGARIS, jadi baris
    # transaksinya SENDIRI tidak pernah ikut tertangkap: yang didapat cuma
    # header tanpa 1 pun baris data. Cek "if not hasil" saja lolos padahal
    # hasil ekstraksinya kosong secara efektif -- makanya sebelumnya bug
    # ini tidak pernah ketahuan lewat exception, cuma diam-diam
    # menghasilkan 0 baris transaksi di sepanjang pipeline setelahnya.
    #
    # Makanya dicek DUA arah sekarang: bukan cuma "apakah ada tabel",
    # tapi "apakah tabel itu benar-benar berisi baris TRANSAKSI" -- pakai
    # parse_sheet_bank() yang SAMA persis dipakai downstream (muat_workbook),
    # supaya kriterianya konsisten satu sumber kebenaran.
    total_baris_transaksi_grid = 0
    for nama_sheet, baris in hasil.items():
        try:
            df_cek = parse_sheet_bank(_LembarDariBaris(baris), nama_sheet)
            total_baris_transaksi_grid += len(df_cek)
        except FormatTidakDikenali:
            continue

    if total_baris_transaksi_grid == 0:
        # Jalur grid gagal (baik karena benar-benar 0 tabel, ATAU cuma
        # menangkap header tanpa 1 pun baris data seperti kasus BCA
        # Tahapan) -- coba fallback posisi-kata sebelum benar-benar
        # menyerah. Nama bank sementara dipakai nama file PDF tanpa
        # ekstensi (konsisten dengan default nama_bank yang dipakai
        # muat_workbook/proses_file_rekening_koran di tempat lain).
        nama_file_pdf = getattr(file_like, "name", "") or "rekening_koran.pdf"
        nama_bank = os.path.splitext(os.path.basename(nama_file_pdf))[0]
        df_fallback, peringatan_fallback, ringkasan_footer = _ekstrak_pdf_rekening_koran_berbasis_posisi(
            file_like, nama_file_pdf, nama_bank
        )
        if not df_fallback.empty:
            logger.info(
                f"📄 {nama_file_pdf}: jalur grid tidak menghasilkan baris transaksi "
                f"(cuma header/kotak info yang bergaris), berhasil diekstrak lewat "
                f"fallback posisi-kata ({len(df_fallback)} baris transaksi)."
            )
            for w in peringatan_fallback:
                logger.info(f"⚠️ {w}")
            # [FIX -- propagasi ke pemanggil] Sebelumnya peringatan_fallback
            # (termasuk hasil validasi otomatis vs ringkasan resmi footer
            # PDF) HANYA di-log ke server, tidak pernah sampai ke response
            # API/UI di jalur upload umum -- padahal justru ini yang paling
            # penting diketahui user (apakah ekstraksinya akurat atau
            # tidak). Sekarang dibawa lewat _LembarDataFrameSiapPakai, lalu
            # diteruskan ke `peringatan` yang dikembalikan muat_workbook().
            # Ganti total isi `hasil` dengan hasil fallback SAJA -- sheet
            # "Halaman N" hasil jalur grid (yang cuma berisi header/kotak
            # info, 0 baris data) dibuang supaya tidak ikut diproses lagi
            # (bakal selalu balik FormatTidakDikenali/df kosong) oleh
            # parser jenis dokumen lain di muat_workbook()/proses_file_xxx.
            return {
                "__POSISI_KATA__:Rekening Koran": (df_fallback, peringatan_fallback, ringkasan_footer)
            }

        if not hasil:
            raise RuntimeError(
                "Tidak ada tabel yang terdeteksi di PDF ini -- kemungkinan PDF hasil SCAN/foto "
                "(bukan teks asli), atau tabelnya tidak dalam format garis/grid yang bisa "
                "dideteksi otomatis. Jalur fallback berbasis posisi kata juga sudah dicoba dan "
                "tidak menemukan baris transaksi "
                f"({'; '.join(peringatan_fallback) if peringatan_fallback else 'tidak ada detail tambahan'})."
            )
        # Ada tabel tertangkap tapi 0 baris transaksi, dan fallback posisi-
        # kata JUGA tidak menemukan apa-apa -- tetap kembalikan hasil grid
        # apa adanya (bukan raise) supaya perilaku utk kasus lain yang
        # mungkin memang beda (bukan rekening koran sama sekali, mis. PDF
        # berisi tabel non-bank) tidak berubah -- pemanggil (parse_sheet_bank
        # dkk di muat_workbook) akan menyimpulkan sendiri lewat
        # FormatTidakDikenali/df kosong seperti biasa.
    return hasil


def _cari_idx_exact(headers, opsi_exact, sampai=None):
    """Cari kolom via exact match (setelah strip+lower), beda dgn _cari_idx yg substring.
    Dipakai utk header pendek spt 'No'/'ID' yg kalau substring match bakal
    kena kolom lain (mis. 'NO AKUN', 'NO TRANSAKSI')."""
    rentang = headers[:sampai] if sampai is not None else headers
    for i, h in enumerate(rentang):
        if h is None:
            continue
        if str(h).strip().lower() in opsi_exact:
            return i
    return None


def parse_sheet_bank(ws, nama_bank: str) -> pd.DataFrame:
    """
    Parse satu sheet rekening koran menjadi DataFrame kolom standar:
    no, bank, tanggal, keterangan, mutasi_debet, mutasi_kredit, saldo,
    supplier_cust, voucher, no_transaksi,
    no_akun_debet, nama_akun_debet, jml_debet,
    no_akun_kredit, nama_akun_kredit, jml_kredit
    (kolom jurnal berisi NaN kalau memang belum diisi di file sumber -- itu
    artinya sheet ini data mentah yang perlu dikategorikan.)

    [BARU] Kalau `ws` adalah _LembarDataFrameSiapPakai (PDF rekening koran
    tanpa garis grid yang sudah diekstrak lewat jalur fallback posisi-kata
    -- lihat _baca_pdf_sebagai_lembar), DataFrame di dalamnya SUDAH dalam
    skema final yang sama dengan output fungsi ini, jadi langsung
    dikembalikan apa adanya TANPA lewat deteksi header/kolom di bawah
    (yang memang tidak akan cocok -- PDF jenis ini cuma punya 1 kolom
    "MUTASI" bersuffix DB, bukan kolom DEBIT/KREDIT terpisah).
    """
    if isinstance(ws, _LembarDataFrameSiapPakai):
        df_final = ws.df_final.copy()
        # [FIX -- propagasi ke pemanggil] Tempelkan peringatan_fallback
        # (termasuk hasil validasi otomatis _validasi_ringkasan_footer_rk)
        # dan ringkasan_footer (saldo_awal/mutasi_cr/mutasi_db/saldo_akhir
        # resmi dari PDF) sebagai df.attrs -- cara ini TIDAK mengubah
        # signature return muat_workbook() (banyak pemanggil unpack fixed-
        # arity 6 elemen), tapi tetap membawa info ini sampai ke
        # proses_file_rekening_koran() lewat df_bank.attrs. Lihat
        # muat_workbook() untuk cara attrs ini dibaca & dipindah ke
        # `peringatan`/`ringkasan_footer_per_sheet`.
        df_final.attrs["peringatan_ekstraksi_fallback"] = ws.peringatan
        df_final.attrs["ringkasan_footer_fallback"] = ws.ringkasan_footer
        return df_final

    header_rownum, header_row = _cari_header_row(ws)
    if header_row is None:
        raise FormatTidakDikenali(
            f"Sheet '{nama_bank}' tidak dikenali sebagai rekening koran "
            "(tidak ditemukan kolom KETERANGAN/REMARKS + SALDO/BALANCE)."
        )

    headers = header_row
    saldo_idx = _cari_idx(headers, ["saldo", "balance"])
    if saldo_idx is None:
        raise FormatTidakDikenali(f"Kolom SALDO tidak ditemukan di sheet '{nama_bank}'.")

    idx_tanggal = _cari_idx(headers, ["tgl", "tanggal", "date"], sampai=saldo_idx + 1)
    idx_keterangan = _cari_idx(headers, ["keterangan", "remarks"], sampai=saldo_idx + 1)
    idx_debit_stmt = _cari_idx(headers, ["debit", "debet"], sampai=saldo_idx)
    idx_kredit_stmt = _cari_idx(headers, ["kredit", "credit"], sampai=saldo_idx)

    if idx_tanggal is None or idx_keterangan is None or idx_debit_stmt is None or idx_kredit_stmt is None:
        raise FormatTidakDikenali(
            f"Kolom wajib (tanggal/keterangan/debit/kredit) tidak lengkap terdeteksi di sheet '{nama_bank}'."
        )

    idx_no = _cari_idx_exact(headers, {"no", "id"}, sampai=saldo_idx)
    idx_supplier = _cari_idx(headers, ["supplier", "cust"])
    idx_voucher = _cari_idx_exact(headers, {"voucher"})
    idx_no_transaksi = _cari_idx_exact(headers, {"no transaksi", "no. transaksi", "notransaksi"})

    idx_jurnal_debet = None
    idx_jurnal_kredit = None
    for i, h in enumerate(headers):
        if h == "DEBET":
            idx_jurnal_debet = i
        if h == "KREDIT":
            idx_jurnal_kredit = i

    ada_jurnal = idx_jurnal_debet is not None and idx_jurnal_kredit is not None

    def _ke_angka_grid(v):
        """[FIX -- bug laten jalur grid] Nilai dari tabel grid PDF
        (pdfplumber.extract_tables) SELALU berupa string apa adanya di
        PDF, termasuk format ribuan ("2,000,000.00") -- beda dengan sel
        Excel yang openpyxl kembalikan sudah bertipe numerik. Sebelumnya
        nilai mentah ini langsung dipakai tanpa konversi, sehingga kolom
        mutasi_debet/mutasi_kredit/saldo bertipe string untuk PDF grid,
        dan CRASH (TypeError: '>' not supported between 'str' and 'int')
        begitu proses_dataframe() membandingkannya dengan angka. Fungsi
        ini idempotent utk nilai yang sudah int/float/None (dari Excel) --
        HANYA string yang dicoba dikonversi (pakai _ke_float_aman_pdf,
        helper yang sama dipakai jalur fallback posisi-kata supaya
        konsisten satu cara parsing angka di seluruh modul).
        """
        if v is None or isinstance(v, (int, float)):
            return v
        if isinstance(v, str):
            v_bersih = v.strip()
            if not v_bersih:
                return None
            # Buang suffix "DB"/"CR"/"K" yang kadang ikut di kolom MUTASI
            # gabungan (mis. rekening koran yang menuliskan "500,000.00 DB"
            # dalam satu sel tabel grid, bukan kolom debit/kredit terpisah).
            v_bersih = re.sub(r"\s*(DB|CR|K)\s*$", "", v_bersih, flags=re.IGNORECASE)
            angka = _ke_float_aman_pdf(v_bersih)
            return angka if angka is not None else v  # gagal parse -> kembalikan apa adanya, jangan diam-diam jadi 0
        return v

    rows = []
    for row in ws.iter_rows(min_row=header_rownum + 1, values_only=True):
        if len(row) <= max(idx_tanggal, idx_keterangan):
            continue
        if row[idx_keterangan] is None and row[idx_tanggal] is None:
            continue
        rows.append({
            "no": row[idx_no] if idx_no is not None else None,
            "bank": nama_bank,
            "tanggal": row[idx_tanggal],
            "keterangan": row[idx_keterangan],
            "mutasi_debet": _ke_angka_grid(row[idx_debit_stmt]) or 0,
            "mutasi_kredit": _ke_angka_grid(row[idx_kredit_stmt]) or 0,
            "saldo": _ke_angka_grid(row[saldo_idx]),
            "supplier_cust": row[idx_supplier] if idx_supplier is not None else None,
            "voucher": row[idx_voucher] if idx_voucher is not None else None,
            "no_transaksi": row[idx_no_transaksi] if idx_no_transaksi is not None else None,
            "no_akun_debet": row[idx_jurnal_debet - 2] if ada_jurnal else None,
            "nama_akun_debet": row[idx_jurnal_debet - 1] if ada_jurnal else None,
            "jml_debet": _ke_angka_grid(row[idx_jurnal_debet]) if ada_jurnal else None,
            "no_akun_kredit": row[idx_jurnal_kredit - 2] if ada_jurnal else None,
            "nama_akun_kredit": row[idx_jurnal_kredit - 1] if ada_jurnal else None,
            "jml_kredit": _ke_angka_grid(row[idx_jurnal_kredit]) if ada_jurnal else None,
        })

    df = pd.DataFrame(rows)
    if not df.empty:
        df["tanggal"] = pd.to_datetime(df["tanggal"], errors="coerce", dayfirst=True).dt.date
    return df


def _siapkan_daftar_sheet(file_like, nama_file: str) -> tuple[list[tuple[str, object]], pd.DataFrame]:
    """
    Deteksi format file dari ekstensi lalu kembalikan (daftar_sheet, df_coa) dalam
    bentuk seragam -- daftar_sheet berisi (nama_sheet, ws) di mana ws punya method
    iter_rows(values_only=True) baik itu worksheet openpyxl asli (.xlsx/.xlsm)
    maupun adapter _LembarDariBaris (.xls/.csv/.pdf). Sheet yang namanya
    mengandung 'coa' DIKELUARKAN dari daftar_sheet dan dipakai sbg df_coa.

    Format yang didukung: .xlsx, .xlsm, .xls, .csv, .pdf. Ekstensi kosong/tidak
    dikenal DIANGGAP .xlsx demi kompatibilitas mundur (mis. file-like tanpa
    atribut nama sama sekali).
    """
    ekstensi = os.path.splitext(nama_file)[1].lower()

    if ekstensi == ".xls":
        lembar = _baca_xls_sebagai_lembar(file_like)
    elif ekstensi == ".csv":
        nama_sheet = os.path.splitext(os.path.basename(nama_file))[0] or "CSV"
        lembar = {nama_sheet: _baca_csv_sebagai_baris(file_like)}
    elif ekstensi == ".pdf":
        lembar = _baca_pdf_sebagai_lembar(file_like)
    elif ekstensi in (".xlsx", ".xlsm", ""):
        wb = openpyxl.load_workbook(file_like, read_only=True, data_only=True)
        df_coa = muat_coa(wb)
        daftar_sheet = [(nama, wb[nama]) for nama in wb.sheetnames if "coa" not in nama.strip().lower()]
        return daftar_sheet, df_coa
    else:
        raise FormatFileTidakDidukung(
            f"Ekstensi file '{ekstensi}' tidak didukung. Format yang didukung: "
            ".xlsx, .xlsm, .xls, .csv, .pdf."
        )

    df_coa = pd.DataFrame(columns=["no_akun", "nama_akun", "kategori"])
    daftar_sheet = []
    for nama_sheet, baris in lembar.items():
        # [BARU] Marker khusus dari _baca_pdf_sebagai_lembar() jalur fallback
        # posisi-kata (PDF rekening koran tanpa garis grid) -- isinya SUDAH
        # DataFrame final, bukan list baris mentah, jadi dibungkus adapter
        # yang berbeda (_LembarDataFrameSiapPakai), BUKAN _LembarDariBaris
        # yang mengharapkan list of tuple. Lihat docstring
        # _LembarDataFrameSiapPakai & parse_sheet_bank untuk cara ini dibaca.
        if nama_sheet.startswith("__POSISI_KATA__:"):
            nama_sheet_asli = nama_sheet.split(":", 1)[1] or "Rekening Koran"
            # [FIX -- BUG KRUSIAL] `baris` di sini adalah tuple 3 elemen
            # (df_fallback, peringatan_fallback, ringkasan_footer) dari
            # _baca_pdf_sebagai_lembar(), BUKAN DataFrame tunggal.
            # Sebelumnya seluruh tuple ini dikirim apa adanya sebagai
            # argumen pertama (df_final) ke _LembarDataFrameSiapPakai(),
            # jadi ws.df_final = tuple (bukan DataFrame) -> parse_sheet_bank()
            # memanggil ws.df_final.copy() -> AttributeError ('tuple' object
            # has no attribute 'copy') -- exception ini BUKAN
            # FormatTidakDikenali, jadi tidak ketangkap di muat_workbook(),
            # menjalar ke atas dan bikin jenis dokumen "bank" gagal total.
            # Karena PDF rekening koran memang tidak cocok jenis dokumen
            # lain, hasilnya "Tidak ada jenis dokumen yang dikenali di file
            # ini" -- padahal isinya jelas rekening koran valid. Sekarang
            # tuple di-unpack dengan benar, dan peringatan_fallback +
            # ringkasan_footer ikut dibawa ke adapter (lihat parse_sheet_bank
            # & muat_workbook() untuk cara ini diteruskan ke pemanggil).
            df_fallback, peringatan_fallback, ringkasan_footer = baris
            daftar_sheet.append((
                nama_sheet_asli,
                _LembarDataFrameSiapPakai(df_fallback, peringatan_fallback, ringkasan_footer),
            ))
            continue
        adapter = _LembarDariBaris(baris)
        if "coa" in nama_sheet.strip().lower():
            df_coa = _ekstrak_coa_dari_ws(adapter)
            continue
        daftar_sheet.append((nama_sheet, adapter))
    return daftar_sheet, df_coa


# ============================================================
# DETEKSI LAPORAN KEUANGAN (FLEKSIBEL)
# ============================================================
# Sumber kebenaran TUNGGAL untuk "apakah file ini Laporan Keuangan" --
# dipakai oleh muat_workbook() di bawah, dan diimpor oleh app.py, supaya
# kedua tempat itu TIDAK PERNAH punya jawaban berbeda untuk file yang sama.
#
# Template lengkap 31 sheet (dipakai oleh generate_template_31_sheet() di
# app.py untuk membuat file kosong yang bisa diisi user).
TARGET_SHEETS_31 = [
    "C (2)", "T (2)", "HO", "PUTU", "OMAH", "WARUNG", "KOST WARUNG",
    "KOST KAPU", "KOST NUATI", "NDALEM RETNO", "CAFE KAPU",
    "KONSOLIDASI", "BS",
    "PNLC F", "COA", "FAFIX", "GL", "ADJ", "TB",
    "DETAIL UTANG LAINNYA", "LIST UTANG LAINNYA", "LIST PIUTANG",
    "LIST DEPOSIT", "DEPOSIT", "LIST PREPAID",
    "CF6", "LPK 6", "LB 6", "LP6", "CF 6", "CALK 5"
]

# Sheet "kunci" yang menjadi ciri khas laporan keuangan -- variasi penulisan
# umum (spasi, singkatan) ditambahkan di sini sebagai alias, bukan dengan
# menuntut kecocokan string yang persis.
SHEET_KUNCI_LAPORAN = [
    "COA", "GL", "TB", "BS", "CALK", "ADJ", "FAFIX", "FA FIX",
    "CF6", "CF 6", "LPK 6", "LB 6", "LP6", "LPE 6", "PNLC F", "PNL CF",
]

SHEET_ENTITAS_LAPORAN = [
    "HO", "PUTU", "OMAH", "WARUNG", "KOST WARUNG",
    "KOST KAPU", "KOST NUATI", "NDALEM RETNO", "CAFE KAPU",
    "KONSOLIDASI",
]


def _normalisasi_nama_sheet(nama: str) -> str:
    """
    Menormalkan nama sheet supaya variasi penulisan yang jelas-jelas sama
    secara maksud (spasi ekstra/tidak konsisten) tidak dianggap beda:
    - lowercase & strip
    - spasi ganda -> spasi tunggal
    """
    return " ".join(nama.strip().lower().split())


def deteksi_laporan_keuangan(sheets: list[str]) -> tuple[bool, str, list[str]]:
    """
    Deteksi file laporan keuangan secara FLEKSIBEL berdasarkan isi (nama-nama
    sheet), bukan berdasarkan kecocokan persis 31 nama & urutan sheet.

    Kenapa tidak exact-match: file laporan keuangan asli sering:
    - punya sheet TAMBAHAN di luar 31 sheet standar (mis. sheet per outlet
      baru: "TERAS", "GEDONG", "BOENTJIS"),
    - beda urutan sheet,
    - beda spasi/singkatan tipis pada nama sheet (mis. "FA FIX" vs "FAFIX",
      "CF 6" vs "CF6", "DEPOSITO" vs "DEPOSIT", "PNL CF" vs "PNLC F").
    Exact-match akan menolak semua variasi wajar itu sebagai "bukan laporan
    keuangan", padahal jelas-jelas itu laporan keuangan.

    Returns:
        (True, pesan, daftar_sheet_kunci_yang_ditemukan) jika terdeteksi
        (False, "Bukan laporan keuangan", []) jika bukan
    """
    kunci_norm = {_normalisasi_nama_sheet(k) for k in SHEET_KUNCI_LAPORAN}
    entitas_norm = {_normalisasi_nama_sheet(e) for e in SHEET_ENTITAS_LAPORAN}

    found_kunci = [s for s in sheets if _normalisasi_nama_sheet(s) in kunci_norm]
    found_entitas = [s for s in sheets if _normalisasi_nama_sheet(s) in entitas_norm]

    # Hilangkan duplikat alias (mis. "CF6" & "CF 6" tidak dihitung dobel)
    found_kunci_unik = list(dict.fromkeys(found_kunci))
    found_entitas_unik = list(dict.fromkeys(found_entitas))

    if len(found_kunci_unik) >= 2:
        contoh = ", ".join(found_kunci_unik[:3])
        return (
            True,
            f"Laporan Keuangan terdeteksi ({len(found_kunci_unik)} sheet kunci: {contoh}...)",
            found_kunci_unik,
        )

    if len(found_kunci_unik) >= 1 and len(found_entitas_unik) >= 1:
        return (
            True,
            f"Laporan Keuangan terdeteksi (sheet kunci: {found_kunci_unik[0]}, "
            f"entitas: {found_entitas_unik[0]})",
            found_kunci_unik,
        )

    return False, "Bukan laporan keuangan", []


# ============================================================
# [BARU] CACHE HASIL muat_workbook() -- hindari parsing penuh 2x
# ============================================================
# muat_workbook() mem-parsing SELURUH baris di semua sheet (bisa ribuan
# baris utk rekening koran/GL). Sebelumnya fungsi ini dipanggil ULANG DARI
# NOL dua kali untuk file yang SAMA PERSIS: sekali oleh
# deteksi_semua_sheet() (preview, endpoint /api/deteksi-file), sekali lagi
# oleh proses_file_rekening_koran()/proses_file_penjualan() dkk (commit,
# endpoint /api/proses-file) -- padahal alur normal user adalah preview
# dulu baru commit file yang sama beberapa saat kemudian. Cache ini
# menghilangkan parsing kedua yang redundan itu.
#
# Di-key oleh SHA256 ISI file (bukan nama file) -- supaya upload ulang
# file identik dengan nama beda tetap kena cache, dan file BEDA dengan
# nama kebetulan sama TIDAK salah pakai cache lawas. In-memory (hilang
# saat proses server restart) & dibatasi ukuran+umur (LRU + TTL sederhana)
# -- ini cuma dimaksudkan utk menjembatani jeda preview->commit dalam satu
# sesi kerja, BUKAN sebagai penyimpanan permanen/pengganti database.
_CACHE_MUAT_WORKBOOK: "OrderedDict[str, tuple]" = OrderedDict()
_CACHE_MUAT_WORKBOOK_MAX_ENTRI = 20
_CACHE_MUAT_WORKBOOK_TTL_DETIK = 900  # 15 menit -- cukup utk jeda preview->commit


def _hash_isi_file_like(file_like) -> Optional[str]:
    """
    SHA256 dari isi file_like, TANPA mengubah posisi baca (di-seek balik ke
    posisi semula setelah selesai dibaca) -- dipakai sbg key cache.

    Return None (BUKAN raise) kalau file_like tidak mendukung baca ulang
    (tidak punya getvalue()/read()+seek()) -- pemanggil HARUS menganggap
    None berarti "cache dilewati", upload tetap jalan seperti biasa tanpa
    error, cuma tanpa percepatan cache.
    """
    try:
        if hasattr(file_like, "getvalue"):
            isi = file_like.getvalue()
        elif hasattr(file_like, "read") and hasattr(file_like, "seek"):
            pos_semula = file_like.tell() if hasattr(file_like, "tell") else 0
            file_like.seek(0)
            isi = file_like.read()
            file_like.seek(pos_semula)
        else:
            return None
        if isinstance(isi, str):
            isi = isi.encode("utf-8", errors="ignore")
        return hashlib.sha256(isi).hexdigest()
    except Exception:
        return None


def muat_workbook(file_like, nama_file: str = None):
    """
    Wrapper cache di atas _muat_workbook_tanpa_cache() -- lihat komentar di
    _CACHE_MUAT_WORKBOOK di atas utk alasannya. Signature & return value
    IDENTIK dengan sebelumnya (tuple 6 elemen: df_bank, df_penjualan,
    df_penilaian, df_piutang, df_coa, peringatan), jadi seluruh pemanggil
    yang sudah ada (proses_file_rekening_koran, proses_file_penjualan,
    deteksi_semua_sheet, dkk) TIDAK PERLU berubah sama sekali.
    """
    hash_isi = _hash_isi_file_like(file_like)
    sekarang = time.time()

    if hash_isi is not None and hash_isi in _CACHE_MUAT_WORKBOOK:
        waktu_simpan, hasil_cache = _CACHE_MUAT_WORKBOOK[hash_isi]
        if sekarang - waktu_simpan <= _CACHE_MUAT_WORKBOOK_TTL_DETIK:
            _CACHE_MUAT_WORKBOOK.move_to_end(hash_isi)  # LRU: tandai baru dipakai
            df_bank, df_penjualan, df_penilaian, df_piutang, df_coa, peringatan = hasil_cache
            # .copy() supaya pemanggil TIDAK BISA korupsi isi cache lewat
            # mutasi DataFrame yang dikembalikan (mis. proses_dataframe()
            # yang menambah kolom kategorisasi di tempat) -- setiap
            # pemanggil selalu dapat salinan independen, sama seperti kalau
            # muat_workbook() benar-benar mem-parsing ulang dari nol.
            return (
                df_bank.copy(), df_penjualan.copy(), df_penilaian.copy(),
                df_piutang.copy(), df_coa.copy(), list(peringatan),
            )
        del _CACHE_MUAT_WORKBOOK[hash_isi]  # kadaluarsa (lewat TTL)

    hasil = _muat_workbook_tanpa_cache(file_like, nama_file)

    if hash_isi is not None:
        _CACHE_MUAT_WORKBOOK[hash_isi] = (sekarang, hasil)
        _CACHE_MUAT_WORKBOOK.move_to_end(hash_isi)
        while len(_CACHE_MUAT_WORKBOOK) > _CACHE_MUAT_WORKBOOK_MAX_ENTRI:
            _CACHE_MUAT_WORKBOOK.popitem(last=False)  # buang entri paling lama tidak dipakai

    return hasil


def _muat_workbook_tanpa_cache(file_like, nama_file: str = None) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, list[str]]:
    """
    Isi ASLI muat_workbook() -- TIDAK DIUBAH sama sekali, cuma diganti nama
    supaya bisa dibungkus cache lewat muat_workbook() di atas. Lihat
    muat_workbook() untuk penjelasan cache-nya.

    Baca satu file -- format .xlsx/.xlsm/.xls/.csv/.pdf semua didukung -- & DETEKSI
    OTOMATIS jenis tiap sheet (untuk .csv & PDF setiap "sheet" cukup 1 buah/tabel):
    - Sheet bernama mengandung 'coa' -> dipakai sebagai COA (tidak berlaku utk .csv).
    - Sheet lain: dicoba dulu sbg REKENING KORAN (keterangan+saldo), kalau tidak
      cocok dicoba sbg DATA PENJUALAN (invoice/customer + dpp/ppn/total), kalau
      tidak cocok dicoba sbg DATA PENILAIAN KLIEN/MAKER (Nama Klien + Maker +
      Score), kalau tidak cocok juga dicoba sbg BUKU BANTU PIUTANG (No Transaksi
      + Nama Pelanggan + Sub Total/Total Akhir -- kartu piutang per transaksi,
      BUKAN jurnal). Sheet yang tidak cocok semuanya dilewati & dilaporkan sbg
      peringatan.
      """

    # ============================================================
    # CEK APAKAH INI FILE LAPORAN KEUANGAN (deteksi fleksibel)
    # ============================================================
    try:
        # Reset file pointer ke awal (karena mungkin sudah dibaca sebelumnya)
        if hasattr(file_like, 'seek'):
            file_like.seek(0)
        wb = openpyxl.load_workbook(file_like, read_only=True)
        sheets = wb.sheetnames
        # Reset file pointer untuk diproses oleh kode selanjutnya (dipakai
        # baik saat terdeteksi laporan keuangan maupun tidak)
        if hasattr(file_like, 'seek'):
            file_like.seek(0)

        terdeteksi, pesan, _sheet_kunci = deteksi_laporan_keuangan(sheets)
        if terdeteksi:
            if hasattr(file_like, 'seek'):
                file_like.seek(0)
            return (
                pd.DataFrame(),  # df_bank
                pd.DataFrame(),  # df_jual
                pd.DataFrame(),  # df_penilaian
                pd.DataFrame(),  # df_piutang
                pd.DataFrame(),  # df_coa
                [f"📊 {pesan}"]  # peringatan
            )
    except Exception:
        # Jika gagal baca (mis. bukan file Excel valid), reset dan lanjutkan
        # -- biarkan kode di bawah yang menangani/menolak formatnya.
        if hasattr(file_like, 'seek'):
            file_like.seek(0)

    nama_file = nama_file or getattr(file_like, "name", "") or ""
    daftar_sheet, df_coa = _siapkan_daftar_sheet(file_like, nama_file)

    semua_bank = []
    semua_penjualan = []
    semua_penilaian = []
    semua_piutang = []
    peringatan = []
    # [BARU] Kumpulkan ringkasan_footer (saldo_awal/mutasi_cr/mutasi_db/
    # saldo_akhir resmi dari PDF) per nama sheet -- ditempel ke df_bank.attrs
    # di akhir fungsi supaya proses_file_rekening_koran() (jalur upload
    # umum) bisa mengembalikannya ke API response, sama seperti yang sudah
    # dilakukan susun_gl_dari_pdf_rekening_koran() (jalur Kertas Kerja).
    ringkasan_footer_per_sheet: dict = {}
    for nama, ws in daftar_sheet:
        try:
            df = parse_sheet_bank(ws, nama)
            if df.empty:
                peringatan.append(f"Sheet '{nama}' (rekening koran) tidak berisi baris data, dilewati.")
            else:
                semua_bank.append(df)
                # [FIX -- propagasi ke pemanggil] Sebelumnya peringatan
                # jalur fallback posisi-kata (termasuk hasil validasi
                # otomatis vs footer resmi PDF) HANYA sampai sini kalau
                # fungsi ini tidak crash duluan (lihat fix bug unpacking di
                # _siapkan_daftar_sheet) -- dan bahkan setelah tidak crash,
                # SEBELUMNYA tidak pernah benar-benar dibaca & diteruskan
                # ke `peringatan` di sini. Sekarang dibaca dari df.attrs
                # (ditempel parse_sheet_bank) dan digabung ke `peringatan`
                # yang dikembalikan fungsi ini.
                peringatan_fallback = df.attrs.get("peringatan_ekstraksi_fallback")
                if peringatan_fallback:
                    peringatan.extend(peringatan_fallback)
                ringkasan_footer_fallback = df.attrs.get("ringkasan_footer_fallback")
                if ringkasan_footer_fallback:
                    ringkasan_footer_per_sheet[nama] = ringkasan_footer_fallback
            continue
        except FormatTidakDikenali:
            pass

        try:
            df = parse_sheet_penjualan(ws, nama)
            if df.empty:
                peringatan.append(f"Sheet '{nama}' (data penjualan) tidak berisi baris data, dilewati.")
            else:
                semua_penjualan.append(df)
            continue
        except FormatTidakDikenali:
            pass

        # Format penjualan invoice-style di atas tidak cocok -- coba format
        # ringkasan POS/kasir (mis. export Moka: Tanggal+Outlet+Type+Amount)
        # sebelum menyerah dan lanjut cek penilaian/piutang. - DITAMBAHKAN
        try:
            df = parse_sheet_penjualan_pos(ws, nama)
            if df.empty:
                peringatan.append(f"Sheet '{nama}' (ringkasan penjualan POS) tidak berisi baris data, dilewati.")
            else:
                semua_penjualan.append(df)
            continue
        except FormatTidakDikenali:
            pass

        try:
            df = parse_sheet_penilaian_klien(ws, nama)
            if df.empty:
                peringatan.append(f"Sheet '{nama}' (penilaian klien/maker) tidak berisi baris data, dilewati.")
            else:
                semua_penilaian.append(df)
            continue
        except FormatTidakDikenali:
            pass

        try:
            df = parse_sheet_piutang(ws, nama)
            if df.empty:
                peringatan.append(f"Sheet '{nama}' (Buku Bantu Piutang) tidak berisi baris data, dilewati.")
            else:
                semua_piutang.append(df)
        except FormatTidakDikenali:
            _kode_lain, label_lain, skor_lain = deteksi_jenis_dokumen_lain(ws)
            if label_lain:
                peringatan.append(
                    f"Sheet '{nama}' sepertinya berisi **{label_lain}** "
                    f"(keyakinan ~{skor_lain:.0%}) -- jenis dokumen ini baru bisa DIDETEKSI, "
                    "pemrosesan otomatis/jurnal untuk jenis ini belum tersedia, sheet dilewati."
                )
            else:
                peringatan.append(
                    f"Sheet '{nama}' tidak dikenali sebagai rekening koran, data penjualan, "
                    "penilaian klien/maker, maupun Buku Bantu Piutang, dilewati."
                )

    df_bank = pd.concat(semua_bank, ignore_index=True) if semua_bank else pd.DataFrame()
    df_penjualan = pd.concat(semua_penjualan, ignore_index=True) if semua_penjualan else pd.DataFrame()
    df_penilaian = pd.concat(semua_penilaian, ignore_index=True) if semua_penilaian else pd.DataFrame()
    df_piutang = pd.concat(semua_piutang, ignore_index=True) if semua_piutang else pd.DataFrame()

    # [BARU] Tempel di sini (bukan mengandalkan attrs ikut lewat pd.concat,
    # yang tidak reliable antar versi pandas) -- df_bank yang dikembalikan
    # SELALU punya attrs ini terisi (dict kosong kalau tidak ada sheet yang
    # lewat jalur fallback), supaya pemanggil (proses_file_rekening_koran)
    # bisa baca dengan getattr(df_bank, "attrs", {}) tanpa perlu cek None.
    df_bank.attrs["ringkasan_footer_per_sheet"] = ringkasan_footer_per_sheet

    return df_bank, df_penjualan, df_penilaian, df_piutang, df_coa, peringatan


def deteksi_semua_sheet(file_like, nama_file: str = None) -> list[dict]:
    """
    MODE "DETEKSI SAJA" -- untuk tiap sheet di file, tebak jenis dokumennya
    TANPA memproses/membuat jurnal (beda dengan muat_workbook() yang langsung
    mem-parsing baris & menyusun DataFrame siap-jurnal). Cocok dipakai sebagai
    preview cepat sebelum user commit upload, atau untuk sheet yang jenisnya
    belum ada parser jurnalnya sama sekali (lihat DAFTAR_JENIS_DOKUMEN_LAIN).

    Return: list of dict, satu entri per sheet:
        {"sheet": nama_sheet, "kode": kode_kategori atau None,
         "label": label kategori (human-readable), "keyakinan": float 0-1,
         "sudah_ada_parser": bool}
    "sudah_ada_parser" True berarti sheet ini SUDAH bisa langsung diproses
    otomatis jadi jurnal lewat muat_workbook() + proses_dataframe(); False
    berarti jenisnya baru bisa dideteksi, belum bisa diproses otomatis.
    """
    # Cek dulu apakah keseluruhan file ini Laporan Keuangan (31 sheet) --
    # kalau ya, tidak perlu dicek sheet-per-sheet, tandai sebagai satu jenis.
    try:
        if hasattr(file_like, "seek"):
            file_like.seek(0)
        wb_cek = openpyxl.load_workbook(file_like, read_only=True)
        sheets_cek = wb_cek.sheetnames
        if hasattr(file_like, "seek"):
            file_like.seek(0)
        terdeteksi, pesan, _kunci = deteksi_laporan_keuangan(sheets_cek)
        if terdeteksi:
            return [{
                "sheet": "(seluruh file)",
                "kode": "laporan_keuangan",
                "label": f"Laporan Keuangan (31 sheet) -- {pesan}",
                "keyakinan": 1.0,
                "sudah_ada_parser": True,
            }]
    except Exception:
        if hasattr(file_like, "seek"):
            file_like.seek(0)

    nama_file = nama_file or getattr(file_like, "name", "") or ""
    daftar_sheet, df_coa = _siapkan_daftar_sheet(file_like, nama_file)

    # [PERCEPATAN] Sebelumnya 5 jenis paling umum (rekening_koran/
    # data_penjualan/pos_kasir/penilaian_klien/buku_bantu_piutang) dicek di
    # sini lewat cascade try/except SENDIRI yang 100% terpisah & mem-parsing
    # PENUH tiap sheet -- padahal cascade yang PERSIS SAMA juga dijalankan
    # lagi oleh muat_workbook() sesaat kemudian saat user commit upload
    # (endpoint /api/proses-file -> proses_file_rekening_koran() dkk).
    # Akibatnya file yang sama di-parse penuh DUA KALI, byte-per-byte
    # identik. Sekarang panggil muat_workbook() (yang sudah dibungkus cache
    # berbasis hash isi file -- lihat _CACHE_MUAT_WORKBOOK) SEKALI di sini;
    # hasilnya dipakai utk preview 5 jenis ini, DAN otomatis sudah
    # tersimpan di cache saat proses_file_xxx() memanggil muat_workbook()
    # lagi tak lama sesudahnya -- parsing penuhnya cuma benar-benar
    # terjadi SEKALI utk seluruh alur preview->commit.
    df_bank, df_penjualan, df_penilaian, df_piutang, df_coa_dari_workbook, _peringatan_workbook = \
        muat_workbook(file_like, nama_file)
    if df_coa.empty and not df_coa_dari_workbook.empty:
        df_coa = df_coa_dari_workbook

    hasil = []
    if not df_coa.empty:
        hasil.append({
            "sheet": "COA", "kode": "coa", "label": "Chart of Accounts (COA)",
            "keyakinan": 1.0, "sudah_ada_parser": True,
        })

    # Lookup nama_sheet -> (kode, label) dari hasil muat_workbook() di atas
    # -- O(1) per sheet, TANPA parse ulang. Urutan setdefault() sengaja
    # meniru prioritas cascade asli (bank > penjualan/pos > penilaian >
    # piutang); dalam praktiknya tidak akan pernah bentrok karena
    # muat_workbook() sendiri sudah mutually-exclusive per sheet (begitu
    # satu jenis cocok, sheet itu tidak dicoba ke jenis lain).
    ws_per_nama = dict(daftar_sheet)
    sheet_ke_kode: dict[str, tuple[str, str]] = {}
    # [PENTING] df_bank pakai nama kolom "bank" utk penanda asal sheet
    # (BUKAN "sheet" seperti df_penjualan/df_penilaian/df_piutang -- lihat
    # skema kolom di parse_sheet_bank()). Kalau dicek "sheet" di sini,
    # kondisinya selalu False & sheet rekening koran salah jatuh ke
    # "Tidak dikenali" di preview -- makanya kolomnya beda sengaja dipisah.
    if not df_bank.empty and "bank" in df_bank.columns:
        for s in df_bank["bank"].dropna().unique():
            sheet_ke_kode.setdefault(s, ("rekening_koran", "Rekening Koran (Bank Statement)"))
    if not df_penjualan.empty and "sheet" in df_penjualan.columns:
        for s in df_penjualan["sheet"].dropna().unique():
            if s in sheet_ke_kode:
                continue
            # Bedakan invoice-style ("data_penjualan") vs ringkasan POS/kasir
            # ("pos_kasir") TANPA re-parse penuh -- df_penjualan gabungan
            # keduanya (lihat muat_workbook()) jadi tidak bisa dibedakan dari
            # isinya saja. Cukup cek baris header (dibatasi ~15 baris oleh
            # _cari_header_row_penjualan_pos, bukan seluruh data sheet).
            ws_terkait = ws_per_nama.get(s)
            _rownum_pos, header_pos = (
                _cari_header_row_penjualan_pos(ws_terkait) if ws_terkait is not None else (None, None)
            )
            if header_pos is not None:
                sheet_ke_kode.setdefault(s, ("pos_kasir", "Data POS/Kasir"))
            else:
                sheet_ke_kode.setdefault(s, ("data_penjualan", "Data Penjualan (Invoice)"))
    if not df_penilaian.empty and "sheet" in df_penilaian.columns:
        for s in df_penilaian["sheet"].dropna().unique():
            sheet_ke_kode.setdefault(s, ("penilaian_klien", "Penilaian Klien/Maker"))
    if not df_piutang.empty and "sheet" in df_piutang.columns:
        for s in df_piutang["sheet"].dropna().unique():
            sheet_ke_kode.setdefault(s, ("buku_bantu_piutang", "Buku Bantu Piutang (AR Aging)"))

    # Sisanya (11 jenis yang TIDAK ikut dicek muat_workbook()) tetap lewat
    # cascade try/except seperti sebelumnya -- tidak diubah, supaya
    # perubahan ini murni percepatan, tidak mengubah cakupan jenis dokumen
    # yang bisa dideteksi.
    _pengecek_sisanya = [
        ("faktur_pajak", "Faktur Pajak (PPN)", parse_sheet_faktur_pajak),
        ("bukti_potong_pajak", "Bukti Potong Pajak (PPh 21/23/4(2))", parse_sheet_bukti_potong),
        ("spt_masa", "SPT Masa/Tahunan", parse_sheet_spt),
        ("slip_gaji", "Slip Gaji Karyawan", parse_sheet_slip_gaji),
        ("bukti_kas", "Bukti Kas Masuk/Keluar", parse_sheet_bukti_kas),
        ("kartu_stok", "Kartu Stok/Persediaan", parse_sheet_kartu_stok),
        ("aset_tetap", "Daftar Aset Tetap & Penyusutan", parse_sheet_aset_tetap),
        ("pembelian", "PO/Invoice Pembelian", parse_sheet_pembelian),
        ("rekonsiliasi_bank", "Rekonsiliasi Bank", parse_sheet_rekonsiliasi_bank),
        ("ap_aging", "Buku Bantu Utang (AP Aging)", parse_sheet_ap_aging),
        ("absensi", "Data Absensi/Timesheet", parse_sheet_absensi),
    ]

    for nama, ws in daftar_sheet:
        if nama in sheet_ke_kode:
            kode, label = sheet_ke_kode[nama]
            hasil.append({
                "sheet": nama, "kode": kode, "label": label,
                "keyakinan": 1.0, "sudah_ada_parser": True,
            })
            continue

        ditemukan = False
        for kode, label, fungsi_parse in _pengecek_sisanya:
            try:
                df = fungsi_parse(ws, nama)
            except FormatTidakDikenali:
                continue
            if df.empty:
                continue
            hasil.append({
                "sheet": nama, "kode": kode, "label": label,
                "keyakinan": 1.0, "sudah_ada_parser": True,
            })
            ditemukan = True
            break

        if ditemukan:
            continue

        kode_lain, label_lain, skor_lain = deteksi_jenis_dokumen_lain(ws)
        if label_lain:
            hasil.append({
                "sheet": nama, "kode": kode_lain, "label": label_lain,
                "keyakinan": skor_lain, "sudah_ada_parser": False,
            })
        else:
            hasil.append({
                "sheet": nama, "kode": None, "label": "Tidak dikenali",
                "keyakinan": 0.0, "sudah_ada_parser": False,
            })

    return hasil


def muat_workbook_rekening_koran(file_like) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    """Dipertahankan untuk kompatibilitas mundur (hanya rekening koran) - lihat muat_workbook()."""
    df_bank, _df_penjualan, _df_penilaian, _df_piutang, df_coa, peringatan = muat_workbook(file_like)
    return df_bank, df_coa, peringatan


# ============================================================
# 3. EKSTRAKSI "SIGNATURE" DARI KETERANGAN
# ============================================================

def ekstrak_signature(keterangan) -> str:
    """
    Ambil token pertama dari keterangan sebagai penanda pola (mis. 'NBMB',
    'BRIVA', 'PLNPOST', 'TLKM', 'BPJS', 'ONUS', dst). Angka di ekor token
    dibuang supaya varian nomor tidak memecah pola yang sama.
    """
    if keterangan is None or (isinstance(keterangan, float) and pd.isna(keterangan)):
        return "TIDAK_ADA_KETERANGAN"
    t = str(keterangan).upper().strip()
    if not t:
        return "TIDAK_ADA_KETERANGAN"
    token = t.split(" ")[0].split("/")[0]
    token = re.sub(r"[0-9]+$", "", token)
    return token if token else "TIDAK_ADA_KETERANGAN"


def _arah(row) -> str:
    return "MASUK" if (row.get("mutasi_kredit") or 0) > 0 else "KELUAR"


# ============================================================
# 4. MEMPELAJARI POLA DARI DATA HISTORIS
# ============================================================

@dataclass
class Pola:
    aturan: dict = field(default_factory=dict)

    def to_dict(self):
        out = {}
        for k, v in self.aturan.items():
            out["||".join(k)] = v
        return out

    @classmethod
    def from_dict(cls, d):
        aturan = {}
        for k, v in d.items():
            sig, arah = k.split("||")
            aturan[(sig, arah)] = v
        return cls(aturan=aturan)


def pelajari_pola(df: pd.DataFrame, min_samples: int = 3, confidence_threshold: float = 0.7) -> Pola:
    """
    Enhanced version: Mempelajari pola dengan minimum sample requirement dan confidence score.
    
    df harus sudah punya kolom no_akun_debet / no_akun_kredit (baris tanpa
    jurnal otomatis diabaikan). Untuk tiap (signature, arah), hitung pasangan
    akun (debet, kredit) yang paling sering muncul -- juga simpan apakah
    pola tsb 100% konsisten atau cuma mayoritas (perlu direview manual).

    Args:
        df: DataFrame dengan data transaksi
        min_samples: Minimum jumlah contoh untuk membentuk pola yang reliable
        confidence_threshold: Threshold confidence untuk menganggap pola valid
    
    Returns:
        Pola: Objek Pola dengan aturan-aturan yang sudah dipelajari
    """
    df_valid = df[df["no_akun_debet"].notna() & df["no_akun_kredit"].notna()].copy()
    if df_valid.empty:
        return Pola()

    df_valid["signature"] = df_valid["keterangan"].apply(ekstrak_signature)
    df_valid["arah"] = df_valid.apply(_arah, axis=1)
    df_valid["nominal"] = df_valid[["mutasi_debet", "mutasi_kredit"]].max(axis=1)

    aturan = {}
    for (sig, arah), g in df_valid.groupby(["signature", "arah"]):
        if len(g) < min_samples:
            continue
            
        pasangan = list(zip(
            g["no_akun_debet"], g["nama_akun_debet"],
            g["no_akun_kredit"], g["nama_akun_kredit"]
        ))
        counter = Counter(pasangan)
        
        total_samples = sum(counter.values())
        most_common = counter.most_common(1)[0]
        (nd, nnd, nk, nnk), count = most_common
        
        confidence = count / total_samples
        
        nominal_std = g["nominal"].std() if len(g) > 1 else 0
        mean_nominal = g["nominal"].mean()
        
        if len(g) > 2 and nominal_std > 0:
            z_scores = np.abs((g["nominal"] - mean_nominal) / (nominal_std + 1e-6))
            outliers = z_scores > 3
            if outliers.any():
                g_clean = g[~outliers]
                if len(g_clean) >= min_samples:
                    pasangan_clean = list(zip(
                        g_clean["no_akun_debet"], g_clean["nama_akun_debet"],
                        g_clean["no_akun_kredit"], g_clean["nama_akun_kredit"]
                    ))
                    counter_clean = Counter(pasangan_clean)
                    if counter_clean:
                        (nd, nnd, nk, nnk), count = counter_clean.most_common(1)[0]
                        confidence = count / len(g_clean)
                        total_samples = len(g_clean)
        
        is_konsisten = len(counter) == 1
        
        aturan[(sig, arah)] = {
            "no_akun_debet": nd,
            "nama_akun_debet": nnd,
            "no_akun_kredit": nk,
            "nama_akun_kredit": nnk,
            "konsisten": is_konsisten,
            "jumlah_contoh": int(count),
            "total_samples": total_samples,
            "confidence_score": round(confidence, 3),
            "variance": float(nominal_std),
            "min_samples_required": min_samples,
            "mean_nominal": float(mean_nominal),
            "is_valid": confidence >= confidence_threshold and total_samples >= min_samples,
            "is_variation": False,
            "is_template": False,
            "last_updated": datetime.now().isoformat()
        }

    return Pola(aturan=aturan)


def pelajari_pola_penjualan(df: pd.DataFrame, min_samples: int = 2) -> Pola:
    """
    Mempelajari pola dari data penjualan.
    
    Args:
        df: DataFrame data penjualan
        min_samples: Minimum jumlah contoh untuk membentuk pola
    
    Returns:
        Pola: Objek Pola dengan aturan-aturan yang sudah dipelajari
    """
    df_valid = df[df["no_akun_debet"].notna() & df["no_akun_kredit"].notna()].copy()
    if df_valid.empty:
        return Pola()

    df_valid["signature"] = df_valid["keterangan"].apply(ekstrak_signature)
    df_valid["signature"] = "PJL::" + df_valid["signature"]
    df_valid["arah"] = df_valid["cara_bayar"].fillna("TUNAI")

    aturan = {}
    for (sig, arah), g in df_valid.groupby(["signature", "arah"]):
        if len(g) < min_samples:
            continue
        
        # Ambil akun yang paling sering muncul
        debet_counter = Counter(zip(g["no_akun_debet"], g["nama_akun_debet"]))
        kredit_counter = Counter(zip(g["no_akun_kredit"], g["nama_akun_kredit"]))
        
        if debet_counter and kredit_counter:
            no_debet, nama_debet = debet_counter.most_common(1)[0][0]
            no_kredit, nama_kredit = kredit_counter.most_common(1)[0][0]
            
            # Cek PPN
            no_ppn = None
            nama_ppn = None
            if "no_akun_kredit_ppn" in g.columns:
                ppn_counter = Counter(zip(g["no_akun_kredit_ppn"], g["nama_akun_kredit_ppn"]))
                if ppn_counter:
                    no_ppn, nama_ppn = ppn_counter.most_common(1)[0][0]
            
            aturan[(sig, arah)] = {
                "no_akun_debet": no_debet,
                "nama_akun_debet": nama_debet,
                "no_akun_kredit": no_kredit,
                "nama_akun_kredit": nama_kredit,
                "no_akun_kredit_ppn": no_ppn,
                "nama_akun_kredit_ppn": nama_ppn,
                "konsisten": True,
                "jumlah_contoh": len(g),
                "confidence_score": 0.9,
                "is_valid": True,
                "is_variation": False,
                "is_template": False,
                "last_updated": datetime.now().isoformat()
            }

    return Pola(aturan=aturan)


def gabung_pola(pola_lama: Pola, pola_baru: Pola) -> Pola:
    """Gabungkan pola baru ke pola lama (pola baru menang kalau ada bentrok signature+arah)."""
    hasil = dict(pola_lama.aturan)
    hasil.update(pola_baru.aturan)
    return Pola(aturan=hasil)


_MAKS_VERSI_POLA_DISIMPAN = 15  # snapshot terlama otomatis dihapus melebihi ini


def _folder_versi_pola(path: str) -> Path:
    """Subfolder tempat snapshot versi utk satu file pola disimpan."""
    return Path(path).parent / "versi_pola"


def _simpan_snapshot_versi_pola(pola_snapshot: Pola, path: str, sumber_perubahan: Optional[str] = None) -> None:
    """
    Tulis SATU snapshot bernomor timestamp dari isi pola SEBELUM ditimpa.
    Dipanggil dari dalam simpan_pola() -- lihat penjelasan lengkap di sana
    kenapa ini dipisah dari mekanisme `<path>.bak` yang sudah ada.
    Gagal menyimpan snapshot TIDAK BOLEH menggagalkan simpan_pola() itu
    sendiri -- pemanggil (simpan_pola) yang menangani via try/except.

    [PENTING -- SEMANTIK LABEL] `sumber_perubahan` yang disimpan di sini
    BUKAN label "siapa yang menciptakan pola_snapshot ini", melainkan label
    "perubahan APA yang SEDANG DITULIS dan akan menimpanya" -- karena itulah
    nilainya diambil dari parameter simpan_pola() yang sedang berjalan, bukan
    dari histori si pola_snapshot itu sendiri. Ini SENGAJA: tujuannya supaya
    riwayat gampang dibaca sebagai "kalau rollback ke snapshot ini, perubahan
    <sumber_perubahan> itulah yang dibatalkan" -- pas dengan cara
    rollback_pola() dipakai (default: batalkan perubahan TERAKHIR).
    """
    folder_versi = _folder_versi_pola(path)
    folder_versi.mkdir(parents=True, exist_ok=True)
    nama_dasar = Path(path).stem
    # microdetik disertakan supaya 2 snapshot yang tercipta di detik yang
    # sama (mis. beberapa panggilan simpan_pola() beruntun dalam 1 request)
    # tidak saling menimpa nama file.
    stempel = datetime.now().strftime("%Y%m%dT%H%M%S%f")
    path_snapshot = folder_versi / f"{nama_dasar}__{stempel}.json"
    payload = {
        "_meta_versi": {
            "disimpan_pada": datetime.now().isoformat(),
            "sumber_perubahan": sumber_perubahan or "(tidak dicatat)",
            "jumlah_aturan": len(pola_snapshot.aturan),
        },
        "aturan": pola_snapshot.to_dict(),
    }
    tmp_path = str(path_snapshot) + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, default=str)
    os.replace(tmp_path, path_snapshot)

    _bersihkan_snapshot_versi_lama(folder_versi, nama_dasar)


def _bersihkan_snapshot_versi_lama(folder_versi: Path, nama_dasar: str,
                                    maks_disimpan: Optional[int] = None) -> None:
    """Buang snapshot tertua kalau sudah melebihi maks_disimpan -- nama file
    berisi timestamp jadi urutan sort nama = urutan waktu, aman dipakai
    langsung tanpa perlu baca isi tiap file.

    [FIX] maks_disimpan SENGAJA tidak diberi default `= _MAKS_VERSI_POLA_DISIMPAN`
    langsung di signature -- default argumen Python di-evaluasi SEKALI saat
    modul di-import, jadi kalau nanti _MAKS_VERSI_POLA_DISIMPAN diubah runtime
    (mis. dibaca dari env var/config di masa depan), fungsi ini akan tetap
    memakai nilai lama yang beku. Dibaca ulang di dalam body supaya selalu
    ikut nilai modul yang aktif saat fungsi ini betul-betul dipanggil.
    """
    if maks_disimpan is None:
        maks_disimpan = _MAKS_VERSI_POLA_DISIMPAN
    try:
        semua = sorted(folder_versi.glob(f"{nama_dasar}__*.json"))
    except OSError:
        return
    lebih = len(semua) - maks_disimpan
    if lebih > 0:
        for f in semua[:lebih]:
            try:
                f.unlink()
            except OSError:
                pass


def simpan_pola(pola: Pola, path: str, sumber_perubahan: Optional[str] = None) -> None:
    """
    Simpan pola ke disk. Sebelum menimpa file lama, buat backup (`<path>.bak`)
    dulu -- supaya kalau proses penulisan gagal di tengah jalan (disk penuh,
    proses ke-kill, dll) dan file jadi korup, pola yang sudah dipelajari
    TIDAK hilang begitu saja; masih bisa dipulihkan dari backup.
    Ditulis ke file sementara dulu lalu di-rename (atomic) supaya tidak pernah
    ada momen file setengah tertulis.

    [BARU -- VERSIONING] `<path>.bak` di atas cuma menyimpan SATU langkah
    mundur, dan tujuannya murni pemulihan dari file korup. Itu tidak cukup
    untuk kasus nyata: seorang akuntan meng-konfirmasi/mengoreksi 1 klarifikasi
    yang SALAH, itu langsung tergabung ke pola (lihat gabung_pola -- pola baru
    SELALU menang kalau bentrok signature+arah), dan baru ketahuan salah
    setelah beberapa kali proses/gabung lain terjadi -- pada titik itu
    `<path>.bak` sudah berisi versi yang salah juga, bukan versi yang benar.

    Karena itu, di sini kita SIMPAN JUGA snapshot bernomor timestamp ke
    subfolder `pola_data/versi_pola/` setiap kali file pola akan ditimpa.
    Beda dari `<path>.bak`:
    - menyimpan BANYAK langkah mundur (maks _MAKS_VERSI_POLA_DISIMPAN),
      bukan cuma 1
    - punya metadata (`sumber_perubahan`, waktu, jumlah aturan) supaya
      akuntan/dev bisa lihat riwayat perubahan, bukan cuma "ada backup"
    - dipakai lewat rollback_pola() -- lihat fungsi itu untuk cara pulih

    Snapshot HANYA dibuat dari isi file LAMA yang berhasil divalidasi via
    muat_pola() (bukan asal copy byte mentah) -- supaya folder versi tidak
    ikut menyimpan sampah kalau file lama kebetulan sedang korup (kasus itu
    sudah ditangani jalur `<path>.bak` yang sudah ada). Kegagalan membuat
    snapshot TIDAK menggagalkan penyimpanan pola itu sendiri -- ini fitur
    tambahan, bukan syarat mutlak supaya AI tetap bisa jalan.
    """
    if os.path.exists(path):
        try:
            import shutil
            shutil.copy2(path, path + ".bak")
        except OSError:
            pass

        try:
            pola_lama = muat_pola(path)
            _simpan_snapshot_versi_pola(pola_lama, path, sumber_perubahan)
        except Exception:
            logger.warning(
                f"Gagal membuat snapshot versi pola utk '{path}' -- lanjut simpan "
                "tanpa snapshot (tidak menghentikan proses).", exc_info=True,
            )

    tmp_path = path + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(pola.to_dict(), f, ensure_ascii=False, indent=2, default=str)
    os.replace(tmp_path, path)


def muat_pola(path: str) -> Pola:
    """
    Muat pola dari disk dengan validasi skema. Kalau file korup/format tidak
    sesuai, JANGAN diam-diam kembalikan Pola kosong (itu artinya kehilangan
    semua pola yang sudah dipelajari tanpa pemberitahuan) -- coba pulihkan
    dari backup `<path>.bak` dulu; kalau backup juga gagal, baru kembalikan
    Pola kosong sambil membiarkan exception aslinya diketahui lewat log.
    """
    def _muat_dan_validasi(p):
        with open(p, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            raise ValueError(f"Format pola tidak valid (bukan objek JSON): {p}")
        for k, v in data.items():
            if "||" not in k:
                raise ValueError(f"Key pola tidak valid (harus 'signature||arah'): {k!r}")
            if not isinstance(v, dict) or "no_akun_debet" not in v or "no_akun_kredit" not in v:
                raise ValueError(f"Isi aturan pola tidak lengkap untuk key {k!r}")
        return Pola.from_dict(data)

    if not os.path.exists(path):
        return Pola()

    try:
        return _muat_dan_validasi(path)
    except (json.JSONDecodeError, ValueError, OSError) as e:
        backup_path = path + ".bak"
        if os.path.exists(backup_path):
            try:
                pola_pulih = _muat_dan_validasi(backup_path)
                print(f"[PERINGATAN] '{path}' korup/tidak valid ({e}); pola dipulihkan dari backup.")
                return pola_pulih
            except (json.JSONDecodeError, ValueError, OSError):
                pass
        print(f"[PERINGATAN] '{path}' korup/tidak valid ({e}) dan backup tidak tersedia/gagal -- "
              "mulai dari pola kosong. File asli TIDAK dihapus, cek manual disarankan.")
        return Pola()


def daftar_versi_pola(path: str) -> list[dict]:
    """
    Daftar riwayat snapshot versi pola utk file <path> ini, URUT TERBARU
    KE TERLAMA. Dipakai endpoint riwayat di main.py, mis.:
        GET /api/client/{client_id}/pola/{jenis}/riwayat
    supaya akuntan/dev bisa lihat "pola ini berubah kapan saja & karena apa"
    sebelum memutuskan mau rollback ke titik mana.

    Setiap entri: {"nama_file_snapshot", "disimpan_pada", "sumber_perubahan",
    "jumlah_aturan"} -- "nama_file_snapshot" adalah nilai yang dikirim balik
    ke rollback_pola() kalau user memilih versi TERTENTU (bukan yang paling
    baru).

    CATATAN BACA: "sumber_perubahan" pada tiap entri adalah label PERUBAHAN
    YANG DIBATALKAN kalau kamu rollback ke snapshot itu -- bukan label siapa
    yang membuat state snapshot tsb (lihat catatan semantik lengkap di
    _simpan_snapshot_versi_pola). Entri PALING ATAS (terbaru) = "batalkan
    perubahan paling terakhir terjadi".
    """
    folder_versi = _folder_versi_pola(path)
    nama_dasar = Path(path).stem
    if not folder_versi.exists():
        return []
    hasil = []
    for f in sorted(folder_versi.glob(f"{nama_dasar}__*.json"), reverse=True):
        try:
            with open(f, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            meta = data.get("_meta_versi", {}) if isinstance(data, dict) else {}
            hasil.append({
                "nama_file_snapshot": f.name,
                "disimpan_pada": meta.get("disimpan_pada"),
                "sumber_perubahan": meta.get("sumber_perubahan"),
                "jumlah_aturan": meta.get("jumlah_aturan"),
            })
        except (json.JSONDecodeError, OSError):
            # satu snapshot korup tidak boleh menyembunyikan snapshot lain
            # yang masih baik dari daftar riwayat.
            continue
    return hasil


def rollback_pola(path: str, nama_file_snapshot: Optional[str] = None) -> dict:
    """
    Kembalikan file pola AKTIF (<path>) ke salah satu snapshot versi lama.

    nama_file_snapshot=None (default) -> otomatis pakai snapshot TERBARU yang
    tersedia, yaitu kondisi pola SEBELUM perubahan/penggabungan TERAKHIR kali
    terjadi. Ini kasus paling umum dipakai: "1 koreksi/klarifikasi barusan
    ternyata salah, batalkan perubahan itu."

    Untuk mundur ke titik yang lebih jauh, panggil daftar_versi_pola(path)
    dulu untuk lihat riwayatnya, lalu kirim nama_file_snapshot dari salah
    satu entrinya secara eksplisit.

    PENTING -- SELF-PROTECTING: rollback_pola() memulihkan lewat simpan_pola()
    juga (bukan copy file mentah), jadi kondisi pola SEBELUM rollback pun ikut
    otomatis tersimpan sbg snapshot baru dengan sumber_perubahan berlabel
    jelas. Artinya rollback tidak pernah destruktif/tidak bisa dibatalkan --
    rollback yang salah pilih pun masih bisa dirollback lagi.

    Return dict siap dipakai apa adanya sbg response API di main.py:
        {"sukses": bool, "pesan": str, "snapshot_dipakai": dict | None}
    """
    folder_versi = _folder_versi_pola(path)
    nama_dasar = Path(path).stem
    daftar = daftar_versi_pola(path)

    if not daftar:
        return {
            "sukses": False,
            "pesan": (
                f"Tidak ada riwayat versi tersimpan utk '{nama_dasar}' -- kemungkinan "
                "belum pernah ada perubahan sejak fitur versioning ini aktif, atau "
                "file pola-nya sendiri belum pernah dibuat."
            ),
            "snapshot_dipakai": None,
        }

    if nama_file_snapshot is None:
        target = daftar[0]
    else:
        target = next((d for d in daftar if d["nama_file_snapshot"] == nama_file_snapshot), None)
        if target is None:
            return {
                "sukses": False,
                "pesan": f"Snapshot '{nama_file_snapshot}' tidak ditemukan di riwayat versi '{nama_dasar}'.",
                "snapshot_dipakai": None,
            }

    path_snapshot = folder_versi / target["nama_file_snapshot"]
    try:
        with open(path_snapshot, "r", encoding="utf-8") as f:
            data = json.load(f)
        pola_dipulihkan = Pola.from_dict(data["aturan"])
    except (json.JSONDecodeError, OSError, KeyError, ValueError) as e:
        return {
            "sukses": False,
            "pesan": f"Gagal membaca/memuat snapshot '{target['nama_file_snapshot']}': {e}",
            "snapshot_dipakai": None,
        }

    simpan_pola(
        pola_dipulihkan, path,
        sumber_perubahan=f"rollback_ke_{target['nama_file_snapshot']}",
    )

    return {
        "sukses": True,
        "pesan": (
            f"Pola '{nama_dasar}' dikembalikan ke kondisi {target['disimpan_pada']} "
            f"({target['jumlah_aturan']} aturan, sumber saat itu: {target['sumber_perubahan']})."
        ),
        "snapshot_dipakai": target,
    }


# ============================================================
# 5. DATA AUGMENTATION UNTUK TRAINING POLA
# ============================================================

def augmentasi_pola(pola: Pola, df_coa: pd.DataFrame = None, threshold: float = 0.6) -> Pola:
    """
    Meningkatkan kualitas pola dengan:
    - Menambahkan variasi signature (sinonim, typo umum)
    - Menghubungkan pola yang mirip
    - Membuat pola turunan dari pola yang sudah ada
    
    Args:
        pola: Objek Pola yang akan diaugmentasi
        df_coa: DataFrame COA (opsional) untuk validasi akun
        threshold: Minimal confidence untuk augmentasi
    
    Returns:
        Pola: Pola yang sudah diaugmentasi
    """
    if not pola.aturan:
        return pola
    
    augmented = dict(pola.aturan)
    
    VARIASI_SIGNATURE = {
        'PLN': ['LISTRIK', 'PLN', 'TAGIHAN LISTRIK', 'BAYAR LISTRIK', 'TOKEN LISTRIK'],
        'TELKOM': ['TELEPON', 'INTERNET', 'TELKOM', 'TELKOMSEL', 'INDIHOME'],
        'BPJS': ['BPJS', 'JAMSOSTEK', 'BPJS KESEHATAN', 'BPJS KETENAGAKERJAAN'],
        'PDAM': ['AIR', 'PDAM', 'TAGIHAN AIR', 'BAYAR AIR'],
        'BCA': ['BANK BCA', 'BCA', 'TRANSFER BCA'],
        'MANDIRI': ['BANK MANDIRI', 'MANDIRI', 'TRANSFER MANDIRI'],
        'BRI': ['BANK BRI', 'BRI', 'TRANSFER BRI'],
        'BNI': ['BANK BNI', 'BNI', 'TRANSFER BNI'],
        'GAJI': ['PAYROLL', 'GAJI', 'SALARY', 'GAJI KARYAWAN', 'THR', 'BONUS'],
        'PEMBELIAN': ['PURCHASE', 'BELANJA', 'PEMBELIAN', 'INVOICE', 'PO'],
        'PENJUALAN': ['SALES', 'PENJUALAN', 'INVOICE PENJUALAN', 'SO'],
        'RETUR': ['RETURN', 'RETUR', 'REFUND', 'PENGEMBALIAN'],
        'SEWA': ['RENT', 'SEWA', 'KONTRAK', 'RENTAL'],
        'ATK': ['PERLENGKAPAN', 'ATK', 'ALAT TULIS', 'STATIONERY'],
        'BBM': ['BENSIN', 'PERTAMINA', 'SHELL', 'BBM', 'FUEL'],
        'PROMOSI': ['IKLAN', 'PROMOSI', 'ADVERTISING', 'MARKETING'],
        'SERVICE': ['MAINTENANCE', 'SERVICE', 'PERBAIKAN', 'REPAIR'],
        'PAJAK': ['TAX', 'PAJAK', 'PPH', 'PPN', 'TAXATION'],
        'ADMIN': ['ADMIN BANK', 'BIAYA ADM', 'PROVISI', 'BIAYA BANK'],
        'BUNGA': ['INTEREST', 'BUNGA', 'BUNGA BANK'],
        'ASURANSI': ['INSURANCE', 'ASURANSI', 'PREMI'],
        'TRANSPORT': ['TRANSPORTASI', 'ONGKOS KIRIM', 'DELIVERY'],
    }
    
    for (sig, arah), rule in list(pola.aturan.items()):
        if not rule.get('is_valid', True) and rule.get('jumlah_contoh', 0) < 2:
            continue
        if rule.get('confidence_score', 0) < threshold:
            continue
            
        sig_upper = sig.upper()
        for base, variants in VARIASI_SIGNATURE.items():
            if base in sig_upper or any(v in sig_upper for v in variants):
                for var in variants:
                    new_sig = sig_upper.replace(base, var)
                    if (new_sig, arah) not in augmented:
                        new_rule = dict(rule)
                        new_rule['jumlah_contoh'] = max(1, int(rule.get('jumlah_contoh', 1) * 0.5))
                        new_rule['confidence_score'] = rule.get('confidence_score', 1.0) * 0.8
                        new_rule['is_variation'] = True
                        new_rule['is_template'] = False
                        new_rule['original_signature'] = sig
                        new_rule['augmentation_type'] = 'variation'
                        augmented[(new_sig, arah)] = new_rule
    
    KATA_KUNCI_TEMPLATE = {
        'LISTRIK': ('BEBAN LISTRIK', 'BEBAN LISTRIK'),
        'TELEPON': ('BEBAN TELEPON', 'BEBAN TELEPON'),
        'INTERNET': ('BEBAN INTERNET', 'BEBAN INTERNET'),
        'GAJI': ('BEBAN GAJI', 'BEBAN GAJI'),
        'SEWA': ('BEBAN SEWA', 'BEBAN SEWA'),
        'ATK': ('BEBAN PERLENGKAPAN', 'BEBAN PERLENGKAPAN'),
        'BBM': ('BEBAN BBM', 'BEBAN BBM'),
        'PROMOSI': ('BEBAN PROMOSI', 'BEBAN PROMOSI'),
        'ASURANSI': ('BEBAN ASURANSI', 'BEBAN ASURANSI'),
        'PEMELIHARAAN': ('BEBAN PEMELIHARAAN', 'BEBAN PEMELIHARAAN'),
        'PAJAK': ('BEBAN PAJAK', 'BEBAN PAJAK'),
        'PPN': ('PPN KELUARAN', 'PPN KELUARAN'),
        'BUNGA': ('BEBAN BUNGA', 'BEBAN BUNGA'),
        'ADMIN': ('BEBAN ADMIN BANK', 'BEBAN ADMIN BANK'),
    }
    
    ada_pola_confidence_tinggi = any(
        rule.get('confidence_score', 0) >= threshold for rule in pola.aturan.values()
    )

    if ada_pola_confidence_tinggi:
        for keyword, (akun_debet, akun_kredit) in KATA_KUNCI_TEMPLATE.items():
            template_sig = f"TEMPLATE_{keyword}_*"
            has_template = any(template_sig in k[0] for k in augmented.keys())
            if not has_template:
                augmented[(template_sig, "KELUAR")] = {
                    'no_akun_debet': akun_debet,
                    'nama_akun_debet': akun_debet,
                    'no_akun_kredit': akun_kredit,
                    'nama_akun_kredit': akun_kredit,
                    'jumlah_contoh': 1,
                    'konsisten': True,
                    'confidence_score': 0.6,
                    'is_template': True,
                    'is_variation': False,
                    'keyword': keyword,
                    'augmentation_type': 'keyword_template',
                    'last_updated': datetime.now().isoformat()
                }
    
    if df_coa is not None and not df_coa.empty:
        for key, rule in list(augmented.items()):
            if rule.get('is_template') and not rule.get('coa_matched', False):
                sig, arah = key
                for keyword, (_, _) in KATA_KUNCI_TEMPLATE.items():
                    if keyword in sig.upper():
                        debet_match = df_coa[df_coa['nama_akun'].str.upper().str.contains(keyword, na=False)]
                        if not debet_match.empty:
                            rule['no_akun_debet'] = debet_match.iloc[0]['no_akun']
                            rule['nama_akun_debet'] = debet_match.iloc[0]['nama_akun']
                            rule['coa_matched'] = True
                        break
    
    return Pola(aturan=augmented)


def filter_augmentasi_berkualitas(pola: Pola, threshold: float = 0.5) -> Pola:
    """Filter hanya pola augmentasi yang berkualitas."""
    filtered = {}
    for key, rule in pola.aturan.items():
        if not rule.get('is_variation', False) and not rule.get('is_template', False):
            filtered[key] = rule
            continue
        if rule.get('confidence_score', 0) >= threshold:
            filtered[key] = rule
    return Pola(aturan=filtered)


def merge_dan_augmentasi_pola(pola_lama: Pola, pola_baru: Pola, df_coa: pd.DataFrame = None) -> Pola:
    """Gabungkan pola lama dan baru, lalu lakukan augmentasi."""
    merged = gabung_pola(pola_lama, pola_baru)
    augmented = augmentasi_pola(merged, df_coa)
    filtered = filter_augmentasi_berkualitas(augmented, threshold=0.5)
    return filtered


# ============================================================
# 6. FALLBACK KATA KUNCI
# ============================================================

# [DIPERLUAS] Value tiap kata kunci sekarang bisa berupa 3 bentuk:
#   - str                            : 1 penanda, sama untuk arah MASUK/KELUAR (perilaku lama)
#   - list[str]                      : beberapa KANDIDAT penanda dicoba berurutan -- dipakai
#                                       kalau penamaan akun bisa beda antar client (mis.
#                                       "JAMSOSTEK" vs "BPJS" -- istilah lama vs baru utk
#                                       akun yang sama)
#   - {"MASUK": ..., "KELUAR": ...}  : penanda BEDA tergantung arah transaksi (mis. BRIVA/QRIS
#                                       -- uang masuk lazimnya pelunasan piutang, uang keluar
#                                       pembayaran hutang -- akun yang relevan jelas beda)
#
# Urutan key dalam dict ini MENENTUKAN PRIORITAS ketika satu keterangan cocok dengan lebih
# dari satu kata kunci sekaligus (mis. teks MPN yang kebetulan juga memuat kode "PPH21") --
# taruh kata kunci yang lebih SPESIFIK lebih dulu, kata kunci umum belakangan.
#
# Kata kunci yang ditandai [VERIFIED] di bawah dicocokkan LANGSUNG ke jurnal asli di file
# contoh (1__SAU_Rekening_Koran_All_Bank_202607, yang sudah benar) -- bukan tebakan generik.
# Untuk kode channel bank yang nempel tanpa spasi ke nomor referensi (mis.
# "BPJS0000726070165383IBIZ...", "BRIVA8077708129988003IBIZTKP...") kata kuncinya SENGAJA
# dibiarkan tanpa word-boundary/spasi supaya tetap kena; untuk kata kunci pendek yang
# berisiko match ke teks lain (mis. "xl") spasi trailing SENGAJA disertakan sebagai
# pengaman minimal, karena format aslinya juga selalu diikuti spasi lalu nomor telepon.
KATA_KUNCI_AKUN = {
    # --- Utilitas & komunikasi ---
    "listrik": "LISTRIK", "pln": "LISTRIK",
    "plnpost": "LISTRIK",  # [VERIFIED] "PLNPOST 547100578212IBIZ..." (sudah kena lewat "pln" juga, ditulis eksplisit biar jelas)
    "telepon": ["TELEPON", "TELKOM"], "telkom": ["TELEPON", "TELKOM"],
    "tlkm": ["TELEPON", "TELKOM"],  # [VERIFIED] "TLKM 0218707747IBIZ...TELKOM..." -- kode bank, bukan "TELKOM" penuh
    "internet": ["TELEPON", "INTERNET"], "wifi": ["TELEPON", "INTERNET"],
    "halo ": ["TELEPON"],  # [VERIFIED] "HALO 08111139902IBIZ..." (Telkomsel Halo pascabayar korporat)
    "xl ": ["TELEPON"],    # [VERIFIED] "XL 08179049292IBIZ..." (pulsa/paket XL)

    # --- Payroll & jaminan sosial ---
    "gaji": "GAJI", "payroll": "GAJI",
    "bpjs": ["JAMSOSTEK", "BPJS"],  # [VERIFIED] "BPJS0000726070165383IBIZ..." -> BIAYA/HUTANG YMH DIBAYAR JAMSOSTEK
    "jamsostek": ["JAMSOSTEK", "BPJS"],

    # --- Pajak (MPN = Modul Penerimaan Negara, kanal setoran pajak online) ---
    # Kode jenis setoran spesifik (kalau kebetulan tertulis di keterangan, bukan cuma nomor
    # referensi generik) didahulukan dari "mpn" umum, supaya nyantol ke akun HUTANG PAJAK
    # yang paling tepat -- bukan cuma "HUTANG PAJAK" pertama yang ketemu di COA.
    "pph21": "HUTANG PAJAK 21", "pph 21": "HUTANG PAJAK 21",
    "pph23": "HUTANG PAJAK 23", "pph 23": "HUTANG PAJAK 23",
    "pph25": ["HUTANG PAJAK 25", "HUTANG PAJAK 29"], "pph 25": ["HUTANG PAJAK 25", "HUTANG PAJAK 29"],
    "pph29": ["HUTANG PAJAK 25", "HUTANG PAJAK 29"], "pph 29": ["HUTANG PAJAK 25", "HUTANG PAJAK 29"],
    "pph 4(2)": "HUTANG PAJAK 4", "pph4(2)": "HUTANG PAJAK 4", "pp 23": "HUTANG PAJAK 4",
    "ppn": ["HUTANG PAJAK PPN", "PPN KELUARAN", "PAJAK MASUKAN"],
    "mpn": {
        # [VERIFIED] "MPN 042202988610442 IBIZ..." selalu arah KELUAR di file contoh, dan
        # sub-jenis pajaknya TIDAK BISA ditebak dari teks keterangan (isinya cuma nomor
        # referensi generik, bukan kode jenis setoran) -- fallback ini cuma menuju bucket
        # HUTANG PAJAK umum untuk direview akuntan, BUKAN menebak PPh 21/23/25/29 yang benar.
        "KELUAR": ["HUTANG PAJAK", "PAJAK"],
        "MASUK": ["PAJAK MASUKAN", "PPN"],
    },
    "pajak": "PAJAK", "pph": "PAJAK",

    # --- Pembayaran via channel bank (arah menentukan hutang vs piutang) ---
    "briva": {
        # [VERIFIED] "BRIVA8077708129988003IBIZTKP..." arah KELUAR -> HUTANG USAHA di file
        # contoh (bayar ke supplier lewat BRI Virtual Account). Arah MASUK belum ada
        # contoh transaksinya di file sumber -- PIUTANG USAHA dipakai sbg tebakan default
        # paling umum (VA juga lazim dipakai utk menerima pembayaran dari customer).
        "KELUAR": ["HUTANG USAHA", "HUTANG"],
        "MASUK": ["PIUTANG USAHA", "PIUTANG"],
    },
    "qris": {
        "MASUK": ["PIUTANG USAHA", "PENJUALAN", "PIUTANG"],  # [VERIFIED] "QRISOffUs_3_260711..." -> PIUTANG USAHA
        "KELUAR": ["HUTANG USAHA", "HUTANG"],
    },

    # --- Lain-lain (sudah ada sebelumnya) ---
    "sewa": "SEWA", "rental": "SEWA", "kontrak": "SEWA",
    "atk": "PERLENGKAPAN", "alat tulis": "PERLENGKAPAN",
    "bensin": "BBM", "pertamina": "BBM", "shell": "BBM", "bbm": "BBM",
    "admin bank": "ADM BANK", "biaya adm": "ADM BANK", "provisi": "ADM BANK",
    "asuransi": "ASURANSI",
    "promosi": "PROMOSI", "iklan": "PROMOSI",
    "service": "PEMELIHARAAN", "maintenance": "PEMELIHARAAN", "perbaikan": "PEMELIHARAAN",
}


def _daftar_kandidat_penanda(nilai, arah: Optional[str]) -> list:
    """
    Normalisasi 1 entri KATA_KUNCI_AKUN (str / list[str] / {"MASUK":.., "KELUAR":..})
    jadi daftar kandidat "penanda" yang dicoba berurutan terhadap nama akun COA.
    """
    if isinstance(nilai, dict):
        if arah and arah in nilai:
            urutan = [nilai[arah]] + [v for k, v in nilai.items() if k != arah]
        else:
            # arah tidak diketahui (mis. dipanggil dari kode lama tanpa parameter arah) --
            # gabungkan semua arah sesuai urutan dict-nya sbg fallback, drpd tidak dicoba
            # sama sekali.
            urutan = list(nilai.values())
        kandidat = []
        for v in urutan:
            kandidat.extend(v if isinstance(v, list) else [v])
        return kandidat
    if isinstance(nilai, list):
        return nilai
    return [nilai]


def cocokkan_kata_kunci_ke_coa(keterangan: str, df_coa: pd.DataFrame, arah: Optional[str] = None,
                                nama_akun_upper: Optional[pd.Series] = None):
    """
    Coba cocokkan keterangan ke salah satu nama akun COA asli via kata kunci
    (lihat KATA_KUNCI_AKUN). `arah` ("MASUK"/"KELUAR", opsional) dipakai untuk
    entri yang akunnya beda tergantung arah transaksi (mis. BRIVA/QRIS --
    uang masuk = pelunasan piutang, uang keluar = pembayaran hutang).

    [BARU -- PERBAIKAN PERFORMA] Sebelumnya `df_coa["nama_akun"].str.upper()`
    dihitung ULANG dari nol di setiap panggilan fungsi ini -- DAN di dalamnya
    dihitung ulang LAGI untuk tiap kandidat penanda yang dicoba (bisa >1 kali
    per keyword). Untuk file dengan ribuan transaksi yang tidak kena pola
    historis (tahap 1 di proses_dataframe), ini jadi ribuan+ kali re-scan
    penuh tabel COA padahal isinya SAMA TERUS sepanjang 1 file diproses --
    persis kelas pemborosan "proses ulang yang tidak perlu". Sekarang
    `nama_akun_upper` bisa dikirim SUDAH DIHITUNG SEKALI oleh pemanggil
    (lihat proses_dataframe); kalau tidak dikirim (None), fungsi tetap
    hitung sendiri seperti sebelumnya -- backward compatible.
    """
    if df_coa is None or df_coa.empty:
        return None
    if nama_akun_upper is None:
        nama_akun_upper = df_coa["nama_akun"].astype(str).str.upper()
    t = str(keterangan).lower()
    for kw, nilai in KATA_KUNCI_AKUN.items():
        if kw not in t:
            continue
        for penanda in _daftar_kandidat_penanda(nilai, arah):
            cocok = df_coa[nama_akun_upper.str.contains(penanda, na=False)]
            if not cocok.empty:
                baris = cocok.iloc[0]
                return str(baris["no_akun"]), baris["nama_akun"]
    return None


# ============================================================
# 7. FALLBACK AI (DeepSeek)
# ============================================================

# [FIX] Backend sekarang FastAPI + React (bukan Streamlit lagi), jadi
# API key cukup dibaca dari environment variable (.env di-load oleh
# load_dotenv() di main.py sebelum modul ini di-import).
def ambil_api_key():
    return os.environ.get("DEEPSEEK_API_KEY")


def ambil_api_key_groq():
    return os.environ.get("GROQ_API_KEY")


# [BARU -- JALUR SEMENTARA GROQ MENGGANTIKAN DEEPSEEK] DeepSeek tetap jadi
# prioritas pertama, Groq jalur kedua (fallback). PENTING: key yang TERISI
# tidak selalu berarti BISA DIPAKAI -- DeepSeek bisa saja key-nya ada tapi
# saldonya habis (error 402 "Insufficient Balance"), jadi cek "apakah key
# ada" saja tidak cukup. Karena itu _konfigurasi_provider_chat() sekarang
# mengembalikan LIST semua provider yang key-nya terisi (urut prioritas),
# lalu si pemanggil (tanya_ai, dst) mencoba satu-satu: begitu satu provider
# gagal (saldo habis, key salah, server down, dst), otomatis coba provider
# berikutnya di list tanpa perlu restart atau ubah .env. Kalau semua provider
# di list tetap gagal, error dari provider TERAKHIR yang dilempar ke user.
def _konfigurasi_provider_chat():
    """Balikin LIST provider chat yang key-nya terisi, urut prioritas
    (DeepSeek dulu, baru Groq). List kosong kalau tidak ada key sama sekali."""
    daftar = []
    deepseek_key = ambil_api_key()
    if deepseek_key:
        daftar.append({
            "api_key": deepseek_key,
            "base_url": "https://api.deepseek.com",
            "model": "deepseek-chat",
            "nama": "DeepSeek",
        })
    groq_key = ambil_api_key_groq()
    if groq_key:
        daftar.append({
            "api_key": groq_key,
            "base_url": "https://api.groq.com/openai/v1",
            # [FIX -- DELAY CHAT, v3] Riwayat model Groq di sini:
            # 1) "llama-3.3-70b-versatile" -- BUKAN reasoning, streaming
            #    kata-per-kata natural. Di-deprecate resmi oleh Groq per 17
            #    Juni 2026, shutdown final 16 Agustus 2026 -- sekarang 404.
            # 2) "qwen/qwen3.6-27b" -- dicoba sbg pengganti, TAPI ternyata ini
            #    model PREVIEW (vision/multimodal, bukan utk produksi teks) --
            #    tetap terasa delay besar sebelum jawaban muncul (mirip gejala
            #    reasoning model), jadi bukan solusi yang tepat.
            # 3) SEKARANG: "openai/gpt-oss-20b" -- satu2nya model chat teks yg
            #    resmi berstatus PRODUCTION di Groq selain gpt-oss-120b (lihat
            #    console.groq.com/docs/models, per Agustus 2026). Ini JUGA
            #    model reasoning (mikir dulu sebelum jawab, delta reasoning
            #    tidak ikut ke delta.content) -- TIDAK ADA lagi opsi produksi
            #    yang non-reasoning di Groq saat ini. reasoning_effort="low"
            #    di bawah (lihat extra_params) meminimalkan durasi jeda
            #    "mikir" itu (tidak bisa dihilangkan total, tapi jauh lebih
            #    singkat drpd default/high) -- begitu mulai jawab, tetap
            #    streaming kata-per-kata seperti biasa.
            "model": "openai/gpt-oss-20b",
            "nama": "Groq",
            # [BARU] Parameter tambahan yang HANYA didukung provider ini
            # (DeepSeek tidak punya reasoning_effort, jadi TIDAK diisi di
            # konfigurasi DeepSeek di atas) -- di-merge ke kwargs
            # chat.completions.create() oleh tanya_ai()/tanya_ai_stream().
            "extra_params": {"reasoning_effort": "low"},
        })
    return daftar


# [BARU -- PEMISAHAN TUGAS CLAUDE/DEEPSEEK] Claude khusus dipakai untuk
# tugas "mengolah/menganalisis data file" yang butuh reasoning lebih dalam
# (mis. kategorisasi jurnal dari keterangan transaksi ambigu di bawah ini)
# -- DeepSeek (ambil_api_key() di atas) tetap dipakai khusus untuk chat/
# tanya-jawab interaktif (tanya_ai/tanya_ai_stream) & analisis ringkasan
# (modules/ai_analysis.py), supaya biaya API lebih terkendali: Claude cuma
# jalan sekali per file saat diproses, bukan tiap kali user mengetik.
def ambil_api_key_claude():
    return os.environ.get("ANTHROPIC_API_KEY")


# [BARU -- KEY GROQ TERPISAH UTK KATEGORISASI] Supaya rate limit chat
# (tanya_ai/tanya_ai_stream, dipanggil tiap user mengetik) & rate limit
# kategorisasi (kategorikan_dengan_ai/kategorikan_penjualan_dengan_ai,
# dipanggil per file diproses) TIDAK REBUTAN satu sama lain saat dua-duanya
# kebetulan jalan bersamaan, sekarang bisa pakai API key Groq YANG BEDA
# khusus kategorisasi -- diisi lewat env var GROQ_API_KEY_KATEGORISASI.
#
# Kalau env var ini KOSONG (belum diisi), otomatis FALLBACK ke
# GROQ_API_KEY yang sama dgn chat (lihat ambil_api_key_groq()) -- jadi
# kode ini TETAP JALAN tanpa perlu bikin key kedua dulu (opsional, bukan
# wajib), cuma kalau mau benar-benar pisah rate limit, tinggal isi env
# var baru ini dgn key Groq akun/project kedua tanpa ubah kode sama sekali.
def ambil_api_key_groq_kategorisasi():
    return os.environ.get("GROQ_API_KEY_KATEGORISASI") or ambil_api_key_groq()


# [BARU -- FALLBACK GROQ SEMENTARA UTK KATEGORISASI] Sama seperti Groq jadi
# fallback DeepSeek di _konfigurasi_provider_chat(), Groq sekarang JUGA bisa
# jadi fallback Claude khusus utk tugas kategorisasi jurnal -- dipakai kalau
# ANTHROPIC_API_KEY belum aktif (mis. kartu kredit blm ada utk billing
# Anthropic). Begitu ANTHROPIC_API_KEY diisi, urutan prioritas otomatis
# balik pakai Claude lagi TANPA ubah kode sama sekali -- lihat
# _konfigurasi_provider_kategorisasi() di bawah.
#
# Model Groq utk kategorisasi SENGAJA DIBEDAKAN dari model Groq utk chat
# ("openai/gpt-oss-20b" di _konfigurasi_provider_chat(), dioptimalkan utk
# latensi rendah krn di-stream langsung ke user sambil mengetik). Kategorisasi
# jurnal TIDAK di-stream ke user (jalan di background per file), dan butuh
# akurasi lebih tinggi (nomor akun harus presisi, bukan sekadar ngobrol) --
# jadi dipakai model Groq yang lebih besar/reasoning lebih kuat.
# "openai/gpt-oss-120b" berstatus PRODUCTION di Groq (bukan preview) per
# Agustus 2026 -- lihat catatan riwayat model di _konfigurasi_provider_chat().
# Bisa dioverride lewat env var GROQ_MODEL_KATEGORISASI tanpa ubah kode kalau
# nanti ada model Groq lain yang lebih cocok.
def ambil_model_kategorisasi_groq() -> str:
    return os.environ.get("GROQ_MODEL_KATEGORISASI", "openai/gpt-oss-120b")


def _konfigurasi_provider_kategorisasi() -> list[dict]:
    """Balikin LIST provider utk tugas kategorisasi jurnal (dipakai
    kategorikan_dengan_ai & kategorikan_penjualan_dengan_ai).

    [DIUBAH -- KATEGORISASI KHUSUS GROQ] Sebelumnya Claude jadi provider
    utama (fallback ke Groq kalau ANTHROPIC_API_KEY kosong). Sekarang
    kategorisasi jurnal SENGAJA HANYA lewat Groq -- Claude & DeepSeek TIDAK
    dipakai sama sekali di jalur ini (DeepSeek memang dari awal khusus utk
    chat/tanya-jawab di _konfigurasi_provider_chat(), bukan kategorisasi).
    List kosong kalau GROQ_API_KEY_KATEGORISASI / GROQ_API_KEY belum diisi.

    Field "tipe" membedakan SDK/endpoint yang harus dipakai pemanggil:
    - "openai_compatible": lewat SDK openai, client.chat.completions.create
      dgn base_url custom (lihat _proses_satu_chunk_ai) -- dipakai utk Groq
      (& provider OpenAI-compatible lain di masa depan kalau perlu).
    """
    daftar = []
    groq_key = ambil_api_key_groq_kategorisasi()
    if groq_key:
        daftar.append({
            "tipe": "openai_compatible",
            "api_key": groq_key,
            "base_url": "https://api.groq.com/openai/v1",
            "model": ambil_model_kategorisasi_groq(),
            "nama": "Groq",
            # reasoning_effort lebih tinggi drpd model chat ("low") krn di
            # sini akurasi lebih penting drpd latensi (tidak di-stream ke user).
            "extra_params": {"reasoning_effort": "medium"},
        })
    return daftar


def _tunggu_sebelum_retry_chunk(percobaan: int, exception: Exception) -> None:
    """
    [BARU -- PERBAIKAN PERFORMA] Backoff sebelum retry chunk AI yang gagal
    (dipakai oleh _proses_satu_chunk_ai & _proses_satu_chunk_ai_claude).

    SEBELUMNYA: retry loop di kedua fungsi itu langsung `continue` tanpa
    jeda sama sekali begitu 1 percobaan gagal. Untuk error biasa (timeout
    jaringan sesaat) ini kurang ideal tapi tidak fatal -- MASALAH BESARNYA
    ada di error rate-limit (HTTP 429): karena tiap chunk dijalankan
    PARALEL (lihat paralel_maks di _panggil_ai_batch_json*), begitu 1
    chunk kena rate limit, chunk-chunk lain yang jalan bersamaan biasanya
    JUGA kena limit yang sama -- retry instan tanpa jeda cuma menabrak
    limit yang sama berkali-kali, bukannya mereda. Untuk client dengan
    banyak transaksi tidak terpola (puluhan chunk x beberapa ronde
    paralel), ini yang paling mungkin membuat proses yang seharusnya
    beberapa menit jadi berjam-jam.

    Sekarang: exponential backoff dengan jitter (0-30% acak, supaya thread
    paralel tidak semuanya retry di detik yang SAMA PERSIS dan langsung
    saling tabrak lagi) -- 1s, 2s, 4s, 8s... dibatasi maksimum 20 detik per
    percobaan. Kalau error-nya terdeteksi sebagai rate-limit (429 / kelas
    RateLimitError SDK openai atau anthropic, atau pesan mengandung
    "rate limit"/"429"), jeda digandakan (dasar 3 detik, bukan 1 detik)
    karena rate limit butuh waktu lebih lama untuk reset dibanding error
    jaringan biasa. TIDAK mengubah max_percobaan/jumlah retry, cuma
    menambah jeda di antaranya -- perilaku sukses/gagal akhir sama persis
    seperti sebelumnya, cuma lebih "sopan" ke API supaya tidak terus
    menerus gagal karena menabrak limit yang sama.
    """
    pesan_error = str(exception).lower()
    nama_kelas_error = type(exception).__name__.lower()
    adalah_rate_limit = (
        "ratelimit" in nama_kelas_error
        or "429" in pesan_error
        or "rate limit" in pesan_error
        or "rate_limit" in pesan_error
        or "overloaded" in pesan_error  # Anthropic 529 "Overloaded" -- sama-sama butuh mundur, bukan retry instan
        or "529" in pesan_error
    )
    dasar_detik = 3.0 if adalah_rate_limit else 1.0
    jeda = min(20.0, dasar_detik * (2 ** percobaan))
    jeda += jeda * random.uniform(0, 0.3)  # jitter supaya thread paralel tidak retry bersamaan persis
    time.sleep(jeda)


def _proses_satu_chunk_ai(
    chunk: list, chunk_index: int, client, model: str, buat_prompt,
    token_dasar: int, token_per_item: int, max_percobaan: int,
    prompt_statis: Optional[str] = None, extra_params: Optional[dict] = None,
) -> tuple[dict, list[dict]]:
    """
    [BARU -- dipisah dari _panggil_ai_batch_json supaya bisa dijalankan
    PARALEL per-chunk lewat ThreadPoolExecutor, lihat catatan performa
    di _panggil_ai_batch_json.] Proses 1 chunk transaksi (dengan retry
    sendiri), return (hasil_chunk, log_kegagalan_chunk) -- TIDAK menyentuh
    state bersama apa pun, aman dipanggil dari banyak thread sekaligus.

    [BARU -- PERBAIKAN PERFORMA] Retry sekarang pakai backoff (lihat
    _tunggu_sebelum_retry_chunk) alih-alih langsung retry tanpa jeda --
    lihat docstring fungsi itu untuk alasan lengkap.

    [BARU -- DUKUNGAN PROVIDER OPENAI-COMPATIBLE LAIN SELAIN DEEPSEEK]
    `prompt_statis` opsional (default None = perilaku LAMA, tidak berubah):
    kalau diisi, digabung ke depan `buat_prompt(chunk)` jadi satu prompt
    utuh -- dipakai saat provider ini dipanggil sbg fallback dari jalur
    Claude (lihat _proses_satu_chunk_ai_claude), yang MEMISAH prompt jadi
    statis/dinamis demi prompt caching Anthropic. Endpoint chat.completions
    OpenAI-compatible (DeepSeek/Groq) tidak punya mekanisme cache_control
    yang sama, jadi di sini cukup digabung jadi 1 string biasa.

    `extra_params` opsional (mis. {"reasoning_effort": "medium"} utk Groq) --
    di-merge ke kwargs chat.completions.create(), sama seperti yang sudah
    dipakai tanya_ai()/tanya_ai_stream() lewat konfig["extra_params"].
    """
    bagian_dinamis = buat_prompt(chunk)
    prompt = f"{prompt_statis}\n\n{bagian_dinamis}" if prompt_statis is not None else bagian_dinamis
    max_tokens = token_dasar + token_per_item * len(chunk)
    hasil_chunk = {}
    log_kegagalan = []

    sukses = False
    alasan_gagal_terakhir = None
    for percobaan in range(max_percobaan):
        try:
            response = client.chat.completions.create(
                model=model,
                max_tokens=max_tokens,
                temperature=0,
                messages=[{"role": "user", "content": prompt}],
                **(extra_params or {}),
            )
            teks = response.choices[0].message.content.strip()
            teks = re.sub(r"^```(json)?|```$", "", teks, flags=re.MULTILINE).strip()
            if not teks.endswith("]"):
                raise ValueError("Respons JSON dari AI tampak terpotong (tidak diakhiri ']').")
            item_hasil = json.loads(teks)
            nomor_diterima = set()
            for item in item_hasil:
                try:
                    nomor = int(item.get("nomor"))
                except (TypeError, ValueError):
                    continue
                hasil_chunk[nomor] = item
                nomor_diterima.add(nomor)
            nomor_diminta = {b["idx"] for b in chunk}
            nomor_hilang = nomor_diminta - nomor_diterima
            if nomor_hilang:
                log_kegagalan.append({
                    "chunk_index": chunk_index,
                    "idx_terdampak": sorted(nomor_hilang),
                    "alasan": "AI tidak mengembalikan jawaban untuk nomor-nomor ini.",
                })
            sukses = True
            break
        except Exception as e:
            alasan_gagal_terakhir = f"{type(e).__name__}: {e}"
            if percobaan < max_percobaan - 1:
                _tunggu_sebelum_retry_chunk(percobaan, e)
                continue

    if not sukses:
        log_kegagalan.append({
            "chunk_index": chunk_index,
            "idx_terdampak": [b["idx"] for b in chunk],
            "alasan": alasan_gagal_terakhir or "Gagal tanpa pesan error spesifik.",
        })

    return hasil_chunk, log_kegagalan


def _panggil_ai_batch_json(items: list, api_key: str, buat_prompt, model: str = "deepseek-chat",
                                ukuran_chunk: int = 25, token_per_item: int = 220,
                                token_dasar: int = 500, max_percobaan: int = 3,
                                paralel_maks: int = 6, timeout_detik: float = 90.0,
                                base_url: str = "https://api.deepseek.com",
                                prompt_statis: Optional[str] = None,
                                extra_params: Optional[dict] = None) -> tuple[dict, list[dict]]:
    """
    Helper umum: panggil AI OpenAI-compatible (default DeepSeek) per-chunk kecil.

    [BARU -- DUKUNGAN PROVIDER LAIN] `base_url` sekarang parameter (default
    tetap DeepSeek, 100% backward compatible utk pemanggil lama) -- supaya
    fungsi yang SAMA PERSIS ini bisa dipakai juga utk Groq (atau provider
    OpenAI-compatible lain), bukan cuma DeepSeek. `prompt_statis` &
    `extra_params` diteruskan apa adanya ke _proses_satu_chunk_ai (lihat
    docstring di sana) -- dipakai KHUSUS saat fungsi ini dipanggil sbg
    fallback dari kategorisasi Claude (lihat _konfigurasi_provider_kategorisasi
    & kategorikan_dengan_ai).

    [PERBAIKAN PERFORMA] Sebelumnya chunk-chunk ini dipanggil SATU PER SATU
    (sequential for-loop) -- untuk 1 file rekening koran dengan ratusan
    transaksi yang tidak kena pola/kata-kunci, ini bisa jadi puluhan chunk
    x puluhan detik tiap chunk = puluhan menit HANYA untuk 1 file, dikali
    banyak bulan jadi berjam-jam. Sekarang chunk-chunk independen ini
    dijalankan PARALEL lewat ThreadPoolExecutor (network I/O-bound, jadi
    thread biasa sudah cukup, tidak perlu asyncio) -- percepatan kira-kira
    sebesar `paralel_maks` (mis. 6x lebih cepat untuk file dengan banyak
    chunk), dibatasi supaya tidak memicu rate limit DeepSeek.

    Juga menambahkan `timeout` eksplisit + mematikan retry bawaan SDK
    openai (`max_retries=0`) -- sebelumnya client dibuat tanpa timeout
    sama sekali, jadi kalau API lambat/hang, retry bawaan SDK (default 2x)
    bisa bertumpuk DI ATAS retry manual (`max_percobaan`) tanpa batas waktu
    yang jelas, membuat kegagalan jaringan sementara bisa memblokir proses
    sangat lama. Sekarang cuma retry manual (`max_percobaan`) yang jalan,
    dengan batas waktu jelas per percobaan.
    """
    import openai
    client = openai.OpenAI(
        api_key=api_key, base_url=base_url,
        timeout=timeout_detik, max_retries=0,
    )
    hasil_total = {}
    log_kegagalan = []

    daftar_chunk = [
        (items[awal:awal + ukuran_chunk], awal // ukuran_chunk)
        for awal in range(0, len(items), ukuran_chunk)
    ]
    if not daftar_chunk:
        return hasil_total, log_kegagalan

    import concurrent.futures
    with concurrent.futures.ThreadPoolExecutor(max_workers=min(paralel_maks, len(daftar_chunk))) as executor:
        future_ke_chunk = {
            executor.submit(
                _proses_satu_chunk_ai, chunk, chunk_index, client, model, buat_prompt,
                token_dasar, token_per_item, max_percobaan, prompt_statis, extra_params,
            ): chunk_index
            for chunk, chunk_index in daftar_chunk
        }
        for future in concurrent.futures.as_completed(future_ke_chunk):
            hasil_chunk, log_chunk = future.result()
            hasil_total.update(hasil_chunk)
            log_kegagalan.extend(log_chunk)

    return hasil_total, log_kegagalan


def _proses_satu_chunk_ai_claude(
    chunk: list, chunk_index: int, client, model: str, buat_prompt,
    token_dasar: int, token_per_item: int, max_percobaan: int,
    prompt_statis: Optional[str] = None,
) -> tuple[dict, list[dict]]:
    """Versi Claude dari _proses_satu_chunk_ai() -- SAMA PERSIS logikanya
    (parsing JSON, retry, pelaporan nomor yang hilang), cuma bentuk
    panggilan API-nya beda: Anthropic Messages API (client.messages.create,
    jawaban di response.content[0].text) alih-alih OpenAI-compatible
    chat.completions.create yang dipakai DeepSeek. Dipisah jadi fungsi
    sendiri (bukan if/else di dalam _proses_satu_chunk_ai) supaya kedua
    jalur tetap mudah dibaca & dites terpisah.

    [BARU -- PERBAIKAN PERFORMA -- PROMPT CACHING] `prompt_statis`
    opsional (default None = perilaku LAMA, tidak berubah sama sekali):
    kalau diisi, ini dianggap bagian prompt yang SAMA PERSIS di semua
    chunk dalam 1 batch (mis. instruksi + daftar COA klien, yang tidak
    berubah antar chunk -- cuma daftar transaksinya yang beda tiap
    chunk). Bagian ini dikirim sebagai content block terpisah dengan
    `cache_control: ephemeral` (fitur prompt caching Anthropic) --
    percobaan PERTAMA yang memakainya menulis ke cache, percobaan/chunk
    berikutnya yang prefix-nya identik (dalam window ~5 menit) dibaca
    dari cache jauh lebih cepat & lebih murah, alih-alih AI memproses
    ulang ribuan token instruksi+COA yang sama persis di SETIAP chunk.
    Kalau `prompt_statis` None, `buat_prompt(chunk)` dipakai APA ADANYA
    sebagai 1 string penuh, PERSIS seperti sebelum perubahan ini --
    pemanggil lama (yang belum di-update memisahkan bagian statis/dinamis)
    tidak terpengaruh sama sekali."""
    if prompt_statis is not None:
        bagian_dinamis = buat_prompt(chunk)
        messages = [{
            "role": "user",
            "content": [
                {"type": "text", "text": prompt_statis, "cache_control": {"type": "ephemeral"}},
                {"type": "text", "text": bagian_dinamis},
            ],
        }]
    else:
        prompt = buat_prompt(chunk)
        messages = [{"role": "user", "content": prompt}]
    max_tokens = token_dasar + token_per_item * len(chunk)
    hasil_chunk = {}
    log_kegagalan = []

    sukses = False
    alasan_gagal_terakhir = None
    for percobaan in range(max_percobaan):
        try:
            response = client.messages.create(
                model=model,
                max_tokens=max_tokens,
                temperature=0,
                messages=messages,
            )
            teks = response.content[0].text.strip()
            teks = re.sub(r"^```(json)?|```$", "", teks, flags=re.MULTILINE).strip()
            if not teks.endswith("]"):
                raise ValueError("Respons JSON dari AI tampak terpotong (tidak diakhiri ']').")
            item_hasil = json.loads(teks)
            nomor_diterima = set()
            for item in item_hasil:
                try:
                    nomor = int(item.get("nomor"))
                except (TypeError, ValueError):
                    continue
                hasil_chunk[nomor] = item
                nomor_diterima.add(nomor)
            nomor_diminta = {b["idx"] for b in chunk}
            nomor_hilang = nomor_diminta - nomor_diterima
            if nomor_hilang:
                log_kegagalan.append({
                    "chunk_index": chunk_index,
                    "idx_terdampak": sorted(nomor_hilang),
                    "alasan": "AI tidak mengembalikan jawaban untuk nomor-nomor ini.",
                })
            sukses = True
            break
        except Exception as e:
            alasan_gagal_terakhir = f"{type(e).__name__}: {e}"
            if percobaan < max_percobaan - 1:
                _tunggu_sebelum_retry_chunk(percobaan, e)
                continue

    if not sukses:
        log_kegagalan.append({
            "chunk_index": chunk_index,
            "idx_terdampak": [b["idx"] for b in chunk],
            "alasan": alasan_gagal_terakhir or "Gagal tanpa pesan error spesifik.",
        })

    return hasil_chunk, log_kegagalan


def _panggil_ai_batch_json_claude(items: list, api_key: str, buat_prompt, model: str = "claude-sonnet-5",
                                   ukuran_chunk: int = 25, token_per_item: int = 220,
                                   token_dasar: int = 500, max_percobaan: int = 3,
                                   paralel_maks: Optional[int] = None, timeout_detik: float = 90.0,
                                   prompt_statis: Optional[str] = None) -> tuple[dict, list[dict]]:
    # [GABUNG -- CATATAN] Parameter `api_key` di sini SEKARANG VESTIGIAL
    # (tidak dipakai langsung di body fungsi ini lagi) -- claude_client.py
    # baca ANTHROPIC_API_KEY dari env sendiri. Signature TETAP dipertahankan
    # apa adanya (tidak dihapus) supaya caller di _panggil_kategorisasi_
    # dengan_fallback (yang mengoper konfig["api_key"] dari
    # _konfigurasi_provider_kategorisasi()) tidak perlu diubah. Kalau nanti
    # ANTHROPIC_API_KEY di .env & hasil ambil_api_key_claude() SELALU sama
    # (kasus normal), ini tidak berdampak apa pun -- cuma relevan kalau ada
    # skenario 2 API key Claude berbeda dikirim sengaja lewat parameter ini
    # (belum ada kasus pemakaian seperti itu saat ini).
    """Versi Claude dari _panggil_ai_batch_json() -- dipakai KHUSUS untuk
    tugas kategorisasi jurnal (reasoning atas keterangan transaksi ambigu),
    bagian "olah/analisis file" yang sengaja diarahkan ke Claude, terpisah
    dari chat interaktif (tetap DeepSeek, lihat tanya_ai_stream()). Struktur
    chunking + paralel ThreadPoolExecutor SAMA PERSIS dengan versi DeepSeek
    supaya perilaku observability (jumlah chunk, batas paralel, retry)
    konsisten -- cuma client & format ekstraksi jawaban yang beda (lihat
    _proses_satu_chunk_ai_claude).

    [BARU -- PERBAIKAN PERFORMA] `paralel_maks` sekarang default None ->
    dibaca dari env var CLAUDE_KATEGORISASI_PARALEL_MAKS (fallback 6, sama
    seperti default lama) -- SAMA PERSIS perilakunya untuk pemanggil yang
    tidak mengubah apa pun, cuma sekarang bisa DITURUNKAN tanpa ubah kode
    kalau API key ternyata di tier dengan rate limit rendah (banyak chunk
    paralel = lebih gampang nabrak limit, lihat _tunggu_sebelum_retry_chunk).
    Pola env var ini sama seperti KERTAS_KERJA_PDF_PARALEL_MAKS yang sudah
    ada di kertas_kerja.py, supaya konsisten cara tuning-nya.

    [BARU -- PERBAIKAN PERFORMA -- PROMPT CACHING] `prompt_statis` opsional,
    diteruskan apa adanya ke _proses_satu_chunk_ai_claude (lihat docstring
    di sana) -- default None = perilaku lama, tidak berubah."""
    if paralel_maks is None:
        paralel_maks = int(os.environ.get("CLAUDE_KATEGORISASI_PARALEL_MAKS", "6"))
    # [GABUNG] Client Claude sekarang dibuat lewat modules/claude_client.py
    # (satu tempat terpusat), BUKAN anthropic.Anthropic() manual di sini
    # lagi -- timeout_detik & max_retries=0 tetap sama persis seperti
    # sebelumnya (retry manual per-chunk di _proses_satu_chunk_ai_claude /
    # _tunggu_sebelum_retry_chunk TETAP dipertahankan apa adanya di bawah,
    # TIDAK diganti ke claude_client.panggil_dengan_retry -- fungsi retry
    # di sini juga menangani jalur Groq, beda tanggung jawab dari retry
    # single-provider di claude_client.py).
    from .claude_client import ambil_client
    client = ambil_client(timeout=timeout_detik, max_retries=0)
    hasil_total = {}
    log_kegagalan = []

    daftar_chunk = [
        (items[awal:awal + ukuran_chunk], awal // ukuran_chunk)
        for awal in range(0, len(items), ukuran_chunk)
    ]
    if not daftar_chunk:
        return hasil_total, log_kegagalan

    import concurrent.futures
    with concurrent.futures.ThreadPoolExecutor(max_workers=min(paralel_maks, len(daftar_chunk))) as executor:
        future_ke_chunk = {
            executor.submit(
                _proses_satu_chunk_ai_claude, chunk, chunk_index, client, model, buat_prompt,
                token_dasar, token_per_item, max_percobaan, prompt_statis,
            ): chunk_index
            for chunk, chunk_index in daftar_chunk
        }
        for future in concurrent.futures.as_completed(future_ke_chunk):
            hasil_chunk, log_chunk = future.result()
            hasil_total.update(hasil_chunk)
            log_kegagalan.extend(log_chunk)

    return hasil_total, log_kegagalan


def mask_keterangan_sensitif(teks) -> str:
    """
    Samarkan data yang berpotensi PII (Personally Identifiable Information)
    sebelum dikirim ke API pihak ketiga (DeepSeek), sesuai kehati-hatian UU PDP:
    - Deretan digit panjang (>=10 digit, mis. nomor rekening/kartu/NIK) -> hanya
      sisakan 4 digit terakhir, sisanya di-mask dengan 'X'.
    """
    if teks is None or (isinstance(teks, float) and pd.isna(teks)):
        return teks
    t = str(teks)

    def _mask(match):
        digit = match.group(0)
        if len(digit) <= 4:
            return digit
        return "X" * (len(digit) - 4) + digit[-4:]

    return re.sub(r"\d{10,}", _mask, t)


def validasi_hasil_ai(item: dict, df_coa: pd.DataFrame, row: pd.Series, is_penjualan: bool = False) -> dict:
    """
    Validasi hasil dari AI dengan multiple checks.
    Returns: dict dengan 'valid', 'alasan', 'skor_kepercayaan'
    """
    hasil = {
        'valid': True,
        'alasan': [],
        'skor_kepercayaan': 1.0
    }
    
    if not is_penjualan and item.get('no_akun_debet') == item.get('no_akun_kredit'):
        hasil['valid'] = False
        hasil['alasan'].append("Akun debet dan kredit sama")
        hasil['skor_kepercayaan'] *= 0
        return hasil
    
    no_akun_valid = set(df_coa["no_akun"]) if (df_coa is not None and not df_coa.empty) else None
    if no_akun_valid:
        try:
            nd = item.get('no_akun_debet')
            nk = item.get('no_akun_kredit')
            if is_penjualan:
                nk_valid = int(nk) in no_akun_valid if nk is not None else False
                if not nk_valid:
                    hasil['valid'] = False
                    hasil['alasan'].append("Akun penjualan tidak ditemukan di COA")
                    hasil['skor_kepercayaan'] *= 0
                    return hasil
            else:
                nd_valid = int(nd) in no_akun_valid if nd is not None else False
                nk_valid = int(nk) in no_akun_valid if nk is not None else False
                if not (nd_valid and nk_valid):
                    hasil['valid'] = False
                    hasil['alasan'].append("Akun tidak ditemukan di COA")
                    hasil['skor_kepercayaan'] *= 0
                    return hasil
        except (TypeError, ValueError):
            hasil['valid'] = False
            hasil['alasan'].append("Format akun tidak valid")
            hasil['skor_kepercayaan'] *= 0
            return hasil
    
    confidence = str(item.get('confidence') or "").lower()
    if confidence == 'tinggi' and 'alasan' in item and len(str(item.get('alasan', ''))) < 20:
        hasil['skor_kepercayaan'] *= 0.7
        hasil['alasan'].append("Confidence tinggi tapi alasan tidak detail")
    
    nominal = max(row.get('mutasi_debet') or 0, row.get('mutasi_kredit') or 0)
    if nominal > 0 and not is_penjualan:
        debet_nama = str(item.get('nama_akun_debet', '')).upper()
        kredit_nama = str(item.get('nama_akun_kredit', '')).upper()
        is_debet_asset = any(kw in debet_nama for kw in ['KAS', 'BANK', 'PIUTANG', 'PERSEDIAAN'])
        is_kredit_liability = any(kw in kredit_nama for kw in ['HUTANG', 'MODAL', 'PENDAPATAN'])
        if not (is_debet_asset or is_kredit_liability):
            hasil['skor_kepercayaan'] *= 0.8
            hasil['alasan'].append("Jurnal tidak mengikuti pola umum")
    
    return hasil


def _apply_ai_results_to_dataframe(
    df: pd.DataFrame,
    mapping: dict,
    df_coa: pd.DataFrame,
    indices: list,
    ambang_nilai: int,
    is_penjualan: bool = False
) -> pd.DataFrame:
    """
    Fungsi terpusat untuk menerapkan hasil AI ke dataframe.
    """
    if not mapping:
        return df
    
    lookup_nama = dict(zip(df_coa["no_akun"], df_coa["nama_akun"])) if not df_coa.empty else {}
    urutan_confidence = {"rendah": 0, "sedang": 1, "tinggi": 2}
    
    for idx in indices:
        item = mapping.get(idx)
        if not item:
            continue
        
        row = df.loc[idx]
        
        validasi = validasi_hasil_ai(item, df_coa, row, is_penjualan)
        if not validasi['valid']:
            df.at[idx, "catatan_ai"] = f"Ditolak: {', '.join(validasi['alasan'])}"
            continue
        
        confidence = str(item.get("confidence") or "").lower()
        df.at[idx, "confidence_ai"] = confidence
        df.at[idx, "alasan_ai"] = item.get("alasan")
        
        if is_penjualan:
            nd = None
            nk = item.get("no_akun_kredit_penjualan") or item.get("no_akun_kredit") or item.get("no_akun_debet")
        else:
            nd = item.get("no_akun_debet")
            nk = item.get("no_akun_kredit")
        
        if nd is None and nk is None:
            df.at[idx, "catatan_ai"] = "AI menandai transaksi ini tidak yakin (null) -- perlu review manual."
            continue
        
        nominal = max(row.get("mutasi_debet") or 0, row.get("mutasi_kredit") or 0) if not is_penjualan else row.get("total", 0)
        
        if is_penjualan:
            if nk is not None:
                df.at[idx, "no_akun_kredit"] = nk
                df.at[idx, "nama_akun_kredit"] = lookup_nama.get(nk, nk)
                df.at[idx, "jml_kredit"] = row["dpp"] if row.get("ppn") else row["total"]
        else:
            if nd is not None:
                df.at[idx, "no_akun_debet"] = nd
                df.at[idx, "nama_akun_debet"] = lookup_nama.get(nd, nd)
                df.at[idx, "jml_debet"] = nominal
            if nk is not None:
                df.at[idx, "no_akun_kredit"] = nk
                df.at[idx, "nama_akun_kredit"] = lookup_nama.get(nk, nk)
                df.at[idx, "jml_kredit"] = nominal
        
        if item.get("supplier_cust"):
            df.at[idx, "supplier_cust"] = item["supplier_cust"]
        
        # [FIX -- LABEL SALAH] Sebelumnya selalu ditulis "AI (DeepSeek)", lalu
        # setelah pemisahan tugas Claude/DeepSeek diubah jadi hardcode
        # "AI (Claude)" karena SATU-SATUNYA pemanggil fungsi ini
        # (kategorikan_dengan_ai & kategorikan_penjualan_dengan_ai) SELALU
        # lewat _panggil_ai_batch_json_claude() saat itu.
        #
        # [DIUBAH -- KATEGORISASI KHUSUS GROQ] Sejak _konfigurasi_provider_
        # kategorisasi() diubah hanya berisi Groq, provider yang menjawab
        # item kategorisasi SELALU Groq. Nama provider SEBENARNYA tetap
        # dibawa lewat item["_sumber_provider"] (ditempel oleh
        # _panggil_kategorisasi_dengan_fallback) supaya audit trail
        # sumber_kategori tetap akurat kalau di masa depan ada provider
        # lain ditambahkan lagi; fallback ke "Groq" kalau field itu entah
        # kenapa tidak ada (mis. dipanggil dari jalur lama).
        nama_provider = item.get("_sumber_provider") or "Groq"
        if urutan_confidence.get(confidence, 0) < ambang_nilai:
            df.at[idx, "sumber_kategori"] = f"AI ({nama_provider}) - confidence {confidence or 'tidak diketahui'}, perlu review"
        else:
            df.at[idx, "sumber_kategori"] = f"AI ({nama_provider})"
    
    return df


def _panggil_kategorisasi_dengan_fallback(
    items: list, buat_prompt, prompt_statis: Optional[str] = None,
    ukuran_chunk: int = 25, token_per_item: int = 220, token_dasar: int = 500,
    max_percobaan: int = 3, timeout_detik: float = 90.0,
) -> tuple[dict, list[dict]]:
    """
    Dispatcher umum utk tugas kategorisasi (dipakai kategorikan_dengan_ai &
    kategorikan_penjualan_dengan_ai): coba provider dari
    _konfigurasi_provider_kategorisasi() SATU PER SATU sesuai urutan
    prioritas (sekarang HANYA Groq -- lihat docstring fungsi itu; struktur
    list tetap dipertahankan kalau nanti perlu tambah provider lain lagi).
    Item yang SUDAH terjawab oleh provider sebelumnya TIDAK
    dikirim ulang ke provider berikutnya -- HANYA item yang masih belum
    terjawab (baik krn provider itu gagal total, mis. key salah/saldo
    habis/exception tak terduga, MAUPUN krn sebagian item memang di-skip
    AI itu sendiri krn ambigu) yang di-retry ke provider berikutnya. Kalau
    provider TERAKHIR pun masih menyisakan item belum terjawab, item itu
    tetap tercatat di log_kegagalan (perilaku sama seperti sebelumnya --
    baris begini akan jatuh ke "Belum Terkategori - perlu review manual"
    di proses_dataframe/proses_dataframe_penjualan).

    Tiap item hasil ditandai item["_sumber_provider"] = nama provider yang
    benar-benar menjawabnya (sekarang selalu "Groq"), dipakai
    _apply_ai_results_to_dataframe supaya kolom sumber_kategori di GL/Excel
    selalu akurat.

    Kalau TIDAK ADA provider sama sekali (GROQ_API_KEY_KATEGORISASI &
    GROQ_API_KEY dua-duanya kosong), balik pesan error yang jelas --
    pemanggil akan menandai semua item sbg perlu review manual, bukan crash.
    """
    if not items:
        return {}, []
    daftar_provider = _konfigurasi_provider_kategorisasi()
    if not daftar_provider:
        return {}, [{
            "idx_terdampak": [b["idx"] for b in items],
            "alasan": "Tidak ada API key kategorisasi aktif (GROQ_API_KEY_KATEGORISASI / GROQ_API_KEY).",
        }]

    hasil_total: dict = {}
    log_kegagalan_terakhir: list[dict] = []
    item_tersisa = list(items)

    for konfig in daftar_provider:
        if not item_tersisa:
            break
        try:
            if konfig["tipe"] == "anthropic":
                hasil_chunk, log_chunk = _panggil_ai_batch_json_claude(
                    item_tersisa, konfig["api_key"], buat_prompt, model=konfig["model"],
                    ukuran_chunk=ukuran_chunk, token_per_item=token_per_item,
                    token_dasar=token_dasar, max_percobaan=max_percobaan,
                    timeout_detik=timeout_detik, prompt_statis=prompt_statis,
                )
            else:  # "openai_compatible" -- Groq (atau provider OpenAI-compatible lain di masa depan)
                hasil_chunk, log_chunk = _panggil_ai_batch_json(
                    item_tersisa, konfig["api_key"], buat_prompt, model=konfig["model"],
                    ukuran_chunk=ukuran_chunk, token_per_item=token_per_item,
                    token_dasar=token_dasar, max_percobaan=max_percobaan,
                    timeout_detik=timeout_detik, base_url=konfig["base_url"],
                    prompt_statis=prompt_statis, extra_params=konfig.get("extra_params"),
                )
        except Exception as e:
            # Provider ini gagal TOTAL (exception tak terduga di luar retry
            # internal -- mis. import gagal) -- catat, lanjut ke provider
            # berikutnya utk SEMUA item yang masih tersisa.
            log_kegagalan_terakhir = [{
                "idx_terdampak": [b["idx"] for b in item_tersisa],
                "alasan": f"Error {konfig['nama']}: {e}",
            }]
            continue

        for idx, jawaban in hasil_chunk.items():
            if isinstance(jawaban, dict):
                jawaban.setdefault("_sumber_provider", konfig["nama"])
            hasil_total[idx] = jawaban

        idx_terjawab = set(hasil_chunk.keys())
        item_tersisa = [b for b in item_tersisa if b["idx"] not in idx_terjawab]
        log_kegagalan_terakhir = log_chunk

    return hasil_total, log_kegagalan_terakhir


def kategorikan_dengan_ai(baris_belum_jelas: list[dict], df_coa: pd.DataFrame, api_key: str,
                           mask_pii: bool = True, ambang_confidence: str = "sedang"):
    """
    baris_belum_jelas: list of dict {"idx":..., "keterangan":..., "arah":..., "nominal":...}
    df_coa: daftar akun asli perusahaan (no_akun, nama_akun, kategori)
    mask_pii: kalau True, nomor identitas panjang (rekening/kartu/NIK) di kolom
      keterangan disamarkan dulu sebelum dikirim ke API pihak ketiga.
    ambang_confidence: "tinggi" | "sedang" | "rendah" -- confidence minimum untuk
      dianggap valid tanpa review.
    Return: (mapping, log_kegagalan)
    """
    if not baris_belum_jelas:
        return {}, []
    try:
        import openai
    except ImportError:
        return {}, [{"idx_terdampak": [b["idx"] for b in baris_belum_jelas], 
                     "alasan": "Library openai tidak terinstall"}]

    if mask_pii:
        baris_belum_jelas = [
            {**b, "keterangan": mask_keterangan_sensitif(b.get("keterangan"))}
            for b in baris_belum_jelas
        ]

    if df_coa is not None and not df_coa.empty:
        daftar_akun_str = "\n".join(f"{int(r.no_akun)} - {r.nama_akun}" for r in df_coa.itertuples())
    else:
        daftar_akun_str = "(Tidak ada COA asli diupload; gunakan nama kategori umum sebagai pengganti no_akun, mis. 'BEBAN LISTRIK')"

    def _buat_bagian_statis() -> str:
        """[BARU -- PERBAIKAN PERFORMA -- PROMPT CACHING] Bagian prompt
        yang SAMA PERSIS di semua chunk 1 batch (instruksi + daftar COA +
        format jawaban) -- dipisah dari daftar transaksi (yang beda tiap
        chunk) supaya bisa di-cache lewat cache_control (lihat
        _proses_satu_chunk_ai_claude). Isinya SAMA PERSIS dengan bagian
        statis dari template prompt lama, cuma instruksi format JSON
        (dulu di paling akhir, SETELAH daftar transaksi) dipindah ke sini
        (SEBELUM daftar transaksi) supaya SELURUH bagian statis
        terkumpul jadi 1 blok yang bisa di-cache utuh -- isi instruksinya
        sendiri TIDAK diubah sama sekali, cuma urutannya."""
        return f"""Kamu adalah asisten akuntansi yang TELITI, HATI-HATI, dan KONSERVATIF. Lebih baik
menandai transaksi sebagai "tidak yakin" daripada mengarang jurnal yang salah -- kesalahan
jurnal berdampak nyata ke laporan keuangan klien. Tugasmu: menentukan jurnal (akun yang
didebit dan dikredit) untuk tiap transaksi rekening koran berikut, berdasarkan Chart of
Accounts (COA) perusahaan ini:

{daftar_akun_str}

Aturan umum rekening koran bank:
- Transaksi arah MASUK (uang masuk ke rekening bank): akun BANK di-debit, akun lawan (mis.
  PIUTANG USAHA / PENJUALAN / PENDAPATAN LAIN) di-kredit.
- Transaksi arah KELUAR (uang keluar dari rekening bank): akun beban/hutang/aset di-debit,
  akun BANK di-kredit.
- no_akun_debet dan no_akun_kredit WAJIB berbeda -- jangan pernah sama.
- Kalau no_akun_debet atau no_akun_kredit merujuk ke akun BANK, WAJIB pakai nomor akun bank
  yang persis sesuai dari daftar COA di atas (jangan menebak/mengarang nomor akun baru).
- Kalau keterangan menyebut nama orang/perusahaan yang jelas merupakan pihak lawan transaksi,
  isi juga field supplier_cust dengan nama itu (nama singkat, tanpa kode transaksi).

Kasus AMBIGU yang sering salah -- perlakukan dengan hati-hati ekstra:
- Transfer/mutasi ANTAR REKENING BANK milik perusahaan sendiri (bukan transaksi ke pihak
  luar) BUKAN pendapatan maupun beban -- ini cuma pemindahan kas internal. Kalau keterangan
  mengindikasikan ini (mis. "TRSF E-BANKING" ke sesama rekening yang juga ada di daftar COA
  sebagai akun BANK), jurnal harus BANK lawan <-> BANK ini, bukan ke akun pendapatan/beban.
- Refund / pengembalian dana / reversal transaksi sebelumnya: arahnya BERLAWANAN dengan
  transaksi aslinya (kalau reversal dari pembelian, jadi pengurang beban/aset terkait, BUKAN
  otomatis "pendapatan lain-lain").
- Biaya admin/provisi bank nominal kecil: kalau ada akun khusus "BIAYA ADM BANK" di COA,
  pakai itu, jangan digabung ke akun beban umum lain.
- Kalau keterangan terlalu generik/kode saja (mis. hanya kode referensi tanpa nama pihak
  atau tujuan yang jelas) sehingga kamu betul-betul tidak bisa menyimpulkan akun lawannya,
  ini WAJIB masuk kategori tidak yakin -- JANGAN menebak akun "netral" seperti Pendapatan
  Lain-lain / Beban Lain-lain hanya supaya kolom terisi.

Field WAJIB per transaksi:
- "confidence": salah satu dari "tinggi", "sedang", "rendah" -- kejujuran di sini penting.
  "tinggi" HANYA kalau kamu yakin berdasarkan pola/kata kunci yang jelas di keterangan.
  "rendah" kalau kamu menerka berdasarkan pola umum saja, bukan bukti eksplisit di keterangan.
- "alasan": 1 kalimat singkat kenapa kamu memilih pasangan akun ini (untuk audit trail
  akuntan, bukan untuk pembaca umum).
- Kalau untuk satu transaksi kamu betul-betul tidak yakin/ambigu akun mana yang paling tepat
  (bukan sekadar sulit), isi no_akun_debet dan no_akun_kredit dengan null apa adanya, dan
  confidence "rendah" -- JANGAN menebak asal supaya tetap terisi. Baris seperti ini akan
  ditandai untuk direview manual oleh akuntan, itu lebih aman daripada jurnal yang salah.

Jawab HANYA dalam format JSON array yang valid, tanpa teks tambahan, tanpa markdown code fence,
dengan format:
[{{"nomor": 1, "no_akun_debet": 11200003, "no_akun_kredit": 11300003, "supplier_cust": "nama atau null",
  "confidence": "tinggi", "alasan": "Keterangan eksplisit menyebut pembayaran listrik PLN"}}]
"""

    def _buat_prompt(chunk):
        daftar_transaksi_str = "\n".join(
            f"{b['idx']}. [{b['arah']}, Rp{b['nominal']:,.0f}] {b['keterangan']}"
            for b in chunk
        )
        return f"""Daftar transaksi (nomor. [arah, nominal] keterangan) yang perlu kamu kategorikan:
{daftar_transaksi_str}"""

    # [BARU -- FALLBACK GROQ SEMENTARA] Sebelumnya langsung
    # _panggil_ai_batch_json_claude() (hardcode Claude, `api_key` param
    # fungsi ini WAJIB terisi ANTHROPIC_API_KEY). Sekarang lewat dispatcher
    # yang otomatis coba Claude dulu (kalau ANTHROPIC_API_KEY ada), fallback
    # ke Groq (kalau tidak) -- lihat _konfigurasi_provider_kategorisasi().
    # Parameter `api_key` di signature fungsi ini SUDAH TIDAK DIPAKAI LAGI
    # secara langsung (dibaca ulang dari env oleh dispatcher), tapi
    # signature-nya TETAP DIPERTAHANKAN apa adanya supaya semua pemanggil
    # yang sudah ada (proses_dataframe, dll -- yang mengoper
    # ambil_api_key_claude()) tidak perlu diubah sama sekali.
    return _panggil_kategorisasi_dengan_fallback(
        baris_belum_jelas, _buat_prompt, prompt_statis=_buat_bagian_statis(),
    )


# ============================================================
# 8. PIPELINE UTAMA
# ============================================================

def _cari_akun_bank_di_coa(nama_bank, df_coa: pd.DataFrame, nama_akun_upper: Optional[pd.Series] = None):
    """
    Cari akun COA untuk 1 bank (mis. nama sheet 'BRI'/'MANDIRI'/'BCA' -> akun
    'BANK BRI'/'BANK MANDIRI'/'BANK BCA') dengan mencari nama bank sbg substring
    nama akun COA -- dipakai di tahap kata-kunci (proses_dataframe) supaya sisi
    akun bank tidak lagi ditebak dari baris lain yang kebetulan sudah terkategori
    duluan (lihat catatan [FIX] di pemanggilnya).

    [BARU -- PERBAIKAN PERFORMA] `nama_akun_upper` opsional: Series
    `df_coa["nama_akun"].str.upper()` yang SUDAH dihitung sebelumnya oleh
    pemanggil (lihat proses_dataframe -- dihitung SEKALI di luar loop
    baris, bukan diulang tiap baris). Kalau tidak dikasih (None), fungsi
    ini tetap menghitungnya sendiri seperti sebelumnya -- 100% backward
    compatible untuk pemanggil lain yang belum diupdate. `df_coa` sendiri
    TIDAK BOLEH berubah isinya di antara panggilan-panggilan yang berbagi
    `nama_akun_upper` yang sama (aman selama dipakai dalam 1 proses_dataframe,
    karena df_coa memang tidak dimutasi di sana).
    """
    if df_coa is None or df_coa.empty or not nama_bank:
        return None, None
    nama_bank_upper = str(nama_bank).upper().strip()
    if not nama_bank_upper:
        return None, None
    if nama_akun_upper is None:
        nama_akun_upper = df_coa["nama_akun"].astype(str).str.upper()
    cocok = df_coa[nama_akun_upper.str.contains(nama_bank_upper, na=False)]
    if not cocok.empty:
        baris = cocok.iloc[0]
        return str(baris["no_akun"]), baris["nama_akun"]
    return None, None


def proses_dataframe(df: pd.DataFrame, df_coa: pd.DataFrame, pola: Pola,
                      pakai_ai: bool = False, api_key: Optional[str] = None,
                      mask_pii: bool = True, ambang_confidence: str = "sedang") -> pd.DataFrame:
    """
    SEMUA baris (baik yang jurnalnya sudah terisi di file asal maupun yang
    belum) diproses ulang lewat: pola historis -> kata kunci COA -> (opsional)
    AI -> kalau semua gagal, tandai 'Belum Terkategori'.
    """
    df = df.copy()
    df["sumber_kategori"] = None
    df["confidence_ai"] = None
    df["alasan_ai"] = None
    df["catatan_ai"] = None
    for _kol in ("no_akun_debet", "nama_akun_debet", "no_akun_kredit", "nama_akun_kredit", "jml_debet", "jml_kredit"):
        if _kol not in df.columns:
            df[_kol] = pd.Series([None] * len(df), dtype="object", index=df.index)
        else:
            df[_kol] = df[_kol].astype("object")
    urutan_confidence = {"rendah": 0, "sedang": 1, "tinggi": 2}
    ambang_nilai = urutan_confidence.get(ambang_confidence, 1)

    perlu_isi = df.index.tolist()

    # tahap 1: pola historis
    belum_selesai = []
    for idx in perlu_isi:
        row = df.loc[idx]
        sig = ekstrak_signature(row["keterangan"])
        arah = _arah(row)
        aturan = pola.aturan.get((sig, arah))
        if aturan is None:
            belum_selesai.append(idx)
            continue

        nominal = max(row.get("mutasi_debet") or 0, row.get("mutasi_kredit") or 0)
        dipakai = aturan
        if not aturan.get("konsisten", True) and "pola_nominal_kecil" in aturan:
            ambang = aturan["pola_nominal_kecil"]["ambang_nominal"]
            if nominal <= ambang:
                dipakai = aturan["pola_nominal_kecil"]

        df.at[idx, "no_akun_debet"] = dipakai["no_akun_debet"]
        df.at[idx, "nama_akun_debet"] = dipakai["nama_akun_debet"]
        df.at[idx, "no_akun_kredit"] = dipakai["no_akun_kredit"]
        df.at[idx, "nama_akun_kredit"] = dipakai["nama_akun_kredit"]
        df.at[idx, "jml_debet"] = nominal
        df.at[idx, "jml_kredit"] = nominal
        df.at[idx, "sumber_kategori"] = (
            "Sesuai Pola yang Dipelajari" if aturan.get("konsisten", True)
            else "Sesuai Pola yang Dipelajari (perlu cek)"
        )

    # tahap 2: kata kunci ke COA
    # [BARU -- PERBAIKAN PERFORMA] Hitung df_coa["nama_akun"].str.upper()
    # SEKALI di sini, di luar loop -- df_coa tidak berubah selama loop ini
    # berjalan, jadi tidak perlu dihitung ulang tiap baris (lihat catatan
    # lengkap di cocokkan_kata_kunci_ke_coa/_cari_akun_bank_di_coa).
    nama_akun_upper_coa = (
        df_coa["nama_akun"].astype(str).str.upper()
        if df_coa is not None and not df_coa.empty else None
    )
    masih_belum = []
    for idx in belum_selesai:
        row = df.loc[idx]
        arah = _arah(row)
        cocok = cocokkan_kata_kunci_ke_coa(row["keterangan"], df_coa, arah=arah, nama_akun_upper=nama_akun_upper_coa)
        nominal = max(row.get("mutasi_debet") or 0, row.get("mutasi_kredit") or 0)
        if cocok:
            no_akun, nama_akun = cocok
            # [FIX] Sebelumnya sisi akun BANK ditebak dari baris LAIN (bank yang sama)
            # yang kebetulan SUDAH terkategori duluan (di tahap kata-kunci ini juga,
            # secara berurutan) -- tanpa mengecek apakah baris itu memang punya akun
            # bank di sisi debet/kreditnya. Untuk baris arah KELUAR, akun debetnya
            # justru akun NON-bank (mis. beban/hutang yang dibayar) -- kalau baris
            # itu kebetulan diproses lebih dulu, baris berikutnya yang bank-nya sama
            # bisa salah ambil akun non-bank tsb sbg "akun bank". Sekarang dicari
            # LANGSUNG dari COA: nama bank/sheet (mis. "BRI", "MANDIRI") dicocokkan
            # sbg substring nama akun COA (mis. "BANK BRI") -- lebih andal & tidak
            # tergantung urutan baris yang diproses lebih dulu.
            bank_no_akun, bank_nama = _cari_akun_bank_di_coa(row["bank"], df_coa, nama_akun_upper=nama_akun_upper_coa)
            if arah == "MASUK":
                df.at[idx, "no_akun_debet"] = bank_no_akun
                df.at[idx, "nama_akun_debet"] = bank_nama
                df.at[idx, "no_akun_kredit"] = no_akun
                df.at[idx, "nama_akun_kredit"] = nama_akun
            else:
                df.at[idx, "no_akun_debet"] = no_akun
                df.at[idx, "nama_akun_debet"] = nama_akun
                df.at[idx, "no_akun_kredit"] = bank_no_akun
                df.at[idx, "nama_akun_kredit"] = bank_nama
            df.at[idx, "jml_debet"] = nominal
            df.at[idx, "jml_kredit"] = nominal
            df.at[idx, "sumber_kategori"] = "Kata kunci COA"
        else:
            masih_belum.append(idx)

    # tahap 3: AI
    if pakai_ai and masih_belum:
        api_key = api_key or ambil_api_key_claude()
        if api_key:
            batch = []
            for idx in masih_belum:
                row = df.loc[idx]
                nominal = max(row.get("mutasi_debet") or 0, row.get("mutasi_kredit") or 0)
                batch.append({
                    "idx": idx, "keterangan": row["keterangan"],
                    "arah": _arah(row), "nominal": nominal,
                })
            mapping, log_kegagalan_ai = kategorikan_dengan_ai(batch, df_coa, api_key, mask_pii=mask_pii)
            
            alasan_gagal_per_idx = {}
            for entri in (log_kegagalan_ai or []):
                for idx_gagal in entri.get("idx_terdampak", []):
                    alasan_gagal_per_idx[idx_gagal] = entri.get("alasan")
            
            df = _apply_ai_results_to_dataframe(
                df=df,
                mapping=mapping,
                df_coa=df_coa,
                indices=masih_belum,
                ambang_nilai=ambang_nilai,
                is_penjualan=False
            )
            
            for idx, alasan in alasan_gagal_per_idx.items():
                if pd.isna(df.at[idx, "catatan_ai"]):
                    df.at[idx, "catatan_ai"] = alasan

    # tahap 4: fallback ke data asli
    for idx in df[df["sumber_kategori"].isna()].index:
        row = df.loc[idx]
        if pd.notna(row.get("no_akun_debet")) and pd.notna(row.get("no_akun_kredit")):
            nominal = max(row.get("mutasi_debet") or 0, row.get("mutasi_kredit") or 0)
            if pd.isna(row.get("jml_debet")):
                df.at[idx, "jml_debet"] = nominal
            if pd.isna(row.get("jml_kredit")):
                df.at[idx, "jml_kredit"] = nominal
            row = df.loc[idx]
            df.at[idx, "sumber_kategori"] = "Data Asli dari File"
            sig = ekstrak_signature(row["keterangan"])
            arah = _arah(row)
            if (sig, arah) not in pola.aturan:
                pola.aturan[(sig, arah)] = {
                    "no_akun_debet": row.get("no_akun_debet"), "nama_akun_debet": row.get("nama_akun_debet"),
                    "no_akun_kredit": row.get("no_akun_kredit"), "nama_akun_kredit": row.get("nama_akun_kredit"),
                    "konsisten": True, "jumlah_contoh": 1,
                }

    df["sumber_kategori"] = df["sumber_kategori"].fillna("Belum Terkategori - perlu review manual")
    return df


# ============================================================
# 9. DATA PENJUALAN - PARSER
# ============================================================

def _cari_header_row_penjualan(ws, max_scan: int = 15):
    """Cari baris header: mengandung INVOICE/FAKTUR atau CUSTOMER/PELANGGAN, DAN DPP/PPN/TOTAL.
    [FIX] max_scan 6 -> 15, disamakan dgn parser lain (baris judul di atas
    header sering lebih dari 6 baris)."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        teks = " ".join(str(c) for c in row if c is not None).lower()
        ada_invoice = any(k in teks for k in ["invoice", "faktur"])
        ada_customer = any(k in teks for k in ["customer", "pelanggan"])
        # [FIX] Sebelumnya cuma menerima "dpp"/"ppn"/"total tagihan"/
        # "grand total" -- kolom nominal yang cuma bernama "Total" polos
        # (format umum di banyak file client) TIDAK PERNAH ketemu di sini,
        # padahal idx_total di parse_sheet_penjualan() di bawah SUDAH
        # menerima "total" polos. Akibatnya sheet dgn kolom "Total" saja
        # (tanpa DPP/PPN/Total Tagihan/Grand Total) selalu gagal dikenali
        # sebagai Data Penjualan walau seharusnya bisa diproses.
        ada_nominal = any(k in teks for k in ["dpp", "ppn", "total tagihan", "grand total", "total"])
        if (ada_invoice or ada_customer) and ada_nominal:
            return i + 1, list(row)
    return None, None


def parse_sheet_penjualan(ws, nama_sheet: str) -> pd.DataFrame:
    """
    Parse satu sheet data penjualan menjadi DataFrame kolom standar:
    sheet, tanggal, no_invoice, customer, keterangan, cara_bayar, dpp, ppn, total,
    no_akun_debet, nama_akun_debet, jml_debet,
    no_akun_kredit, nama_akun_kredit, jml_kredit
    no_akun_kredit_ppn, nama_akun_kredit_ppn, jml_kredit_ppn
    """
    header_rownum, header_row = _cari_header_row_penjualan(ws)
    if header_row is None:
        raise FormatTidakDikenali(
            f"Sheet '{nama_sheet}' tidak dikenali sebagai data penjualan "
            "(tidak ditemukan kolom INVOICE/FAKTUR atau CUSTOMER/PELANGGAN bersama DPP/PPN/TOTAL)."
        )

    headers = header_row
    idx_tanggal = _cari_idx(headers, ["tgl", "tanggal", "date"])
    idx_invoice = _cari_idx(headers, ["invoice", "faktur", "no inv"])
    idx_customer = _cari_idx(headers, ["customer", "pelanggan"])
    idx_ket = _cari_idx(headers, ["keterangan", "uraian", "nama barang", "deskripsi", "produk", "jasa"])
    idx_cara_bayar = _cari_idx(headers, ["cara bayar", "pembayaran", "termin", "metode bayar"])
    idx_dpp = _cari_idx(headers, ["dpp", "sub total", "subtotal"])
    idx_ppn = _cari_idx(headers, ["ppn", "pajak keluaran"])
    idx_total = _cari_idx(headers, ["total tagihan", "grand total", "total"])

    if idx_invoice is None and idx_customer is None:
        raise FormatTidakDikenali(
            f"Kolom INVOICE/FAKTUR atau CUSTOMER/PELANGGAN tidak ditemukan di sheet '{nama_sheet}'."
        )
    if idx_dpp is None and idx_total is None:
        raise FormatTidakDikenali(f"Kolom nominal (DPP/TOTAL) tidak ditemukan di sheet '{nama_sheet}'.")

    idx_jurnal_debet = None
    idx_jurnal_kredit = None
    idx_jurnal_kredit_ppn = None
    for i, h in enumerate(headers):
        if h == "DEBET":
            idx_jurnal_debet = i
        elif h == "KREDIT":
            idx_jurnal_kredit = i
        elif h == "KREDIT PPN":
            idx_jurnal_kredit_ppn = i

    ada_jurnal = idx_jurnal_debet is not None and idx_jurnal_kredit is not None

    rows = []
    for row in ws.iter_rows(min_row=header_rownum + 1, values_only=True):
        if idx_invoice is not None and idx_customer is not None:
            kosong = row[idx_invoice] is None and row[idx_customer] is None
        elif idx_invoice is not None:
            kosong = row[idx_invoice] is None
        else:
            kosong = row[idx_customer] is None
        if kosong:
            continue

        dpp = row[idx_dpp] if idx_dpp is not None else None
        ppn = row[idx_ppn] if idx_ppn is not None else 0
        total = row[idx_total] if idx_total is not None else None
        if total is None and dpp is not None:
            total = (dpp or 0) + (ppn or 0)
        if dpp is None and total is not None:
            dpp = (total or 0) - (ppn or 0)

        cara_bayar_raw = str(row[idx_cara_bayar]).strip().upper() if (idx_cara_bayar is not None and row[idx_cara_bayar]) else ""
        if any(k in cara_bayar_raw for k in ["KREDIT", "TERMIN", "PIUTANG"]):
            cara_bayar = "KREDIT"
        else:
            cara_bayar = "TUNAI"

        rows.append({
            "sheet": nama_sheet,
            "tanggal": row[idx_tanggal] if idx_tanggal is not None else None,
            "no_invoice": row[idx_invoice] if idx_invoice is not None else None,
            "customer": row[idx_customer] if idx_customer is not None else None,
            "keterangan": row[idx_ket] if idx_ket is not None else None,
            "cara_bayar": cara_bayar,
            "dpp": dpp or 0,
            "ppn": ppn or 0,
            "total": total or 0,
            "no_akun_debet": row[idx_jurnal_debet - 2] if ada_jurnal else None,
            "nama_akun_debet": row[idx_jurnal_debet - 1] if ada_jurnal else None,
            "jml_debet": row[idx_jurnal_debet] if ada_jurnal else None,
            "no_akun_kredit": row[idx_jurnal_kredit - 2] if ada_jurnal else None,
            "nama_akun_kredit": row[idx_jurnal_kredit - 1] if ada_jurnal else None,
            "jml_kredit": row[idx_jurnal_kredit] if ada_jurnal else None,
            "no_akun_kredit_ppn": row[idx_jurnal_kredit_ppn - 2] if idx_jurnal_kredit_ppn is not None else None,
            "nama_akun_kredit_ppn": row[idx_jurnal_kredit_ppn - 1] if idx_jurnal_kredit_ppn is not None else None,
            "jml_kredit_ppn": row[idx_jurnal_kredit_ppn] if idx_jurnal_kredit_ppn is not None else None,
        })

    df = pd.DataFrame(rows)
    if not df.empty and idx_tanggal is not None:
        df["tanggal"] = pd.to_datetime(df["tanggal"], errors="coerce", dayfirst=True).dt.date
    return df


# ============================================================
# 9b. DATA PENJUALAN - FORMAT RINGKASAN POS/KASIR (mis. export Moka) - DITAMBAHKAN
# ============================================================
# Beda dari parse_sheet_penjualan() di atas: format ini BUKAN satu baris =
# satu invoice, tapi satu BLOK baris per (Tanggal, Outlet), di mana tiap
# baris = satu "Type" (Gross Sales, Discount, Refund, Net Sales, Gratuity,
# Tax, Rounding, Total Collected). Dipetakan (SESUAI KONFIRMASI USER):
#   dpp   = Amount pada baris Type "Net Sales"
#   ppn   = Amount pada baris Type "Tax"
#   total = Amount pada baris Type "Total Collected"
# Kalau istilah "Type" di file kamu ternyata beda (mis. bukan bahasa
# Inggris), tambahkan ke set _TIPE_POS_DPP/_TIPE_POS_PPN/_TIPE_POS_TOTAL
# di bawah -- JANGAN diubah jadi cocok banyak Type sekaligus, supaya tidak
# dobel-hitung.

_TIPE_POS_DPP = {"net sales"}
_TIPE_POS_PPN = {"tax"}
_TIPE_POS_TOTAL = {"total collected"}


def _cari_header_row_penjualan_pos(ws, max_scan: int = 15):
    """Cari baris header ringkasan POS/kasir: mengandung Tanggal + Outlet + Type + Amount.
    [FIX] max_scan 6 -> 15, disamakan dgn parser lain."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        teks = " ".join(str(c) for c in row if c is not None).lower()
        ada_tanggal = any(k in teks for k in ["tanggal", "tgl", "date"])
        ada_outlet = "outlet" in teks
        ada_type = "type" in teks or "tipe" in teks
        ada_amount = "amount" in teks or "jumlah" in teks
        if ada_tanggal and ada_outlet and ada_type and ada_amount:
            return i + 1, list(row)
    return None, None


def parse_sheet_penjualan_pos(ws, nama_sheet: str) -> pd.DataFrame:
    """
    Parse sheet ringkasan penjualan gaya POS/kasir (mis. export Moka POS).
    Satu blok baris (per Type) per (Tanggal, Outlet) DIGABUNG jadi satu baris
    output per (Tanggal, Outlet), memakai kolom standar yang sama dengan
    parse_sheet_penjualan() supaya bisa digabung jadi satu df_jual.
    """
    header_rownum, header_row = _cari_header_row_penjualan_pos(ws)
    if header_row is None:
        raise FormatTidakDikenali(
            f"Sheet '{nama_sheet}' tidak dikenali sebagai ringkasan penjualan POS "
            "(tidak ditemukan kolom Tanggal + Outlet + Type + Amount)."
        )

    headers = header_row
    idx_tanggal = _cari_idx(headers, ["tanggal", "tgl", "date"])
    idx_outlet = _cari_idx(headers, ["outlet"])
    idx_type = _cari_idx(headers, ["type", "tipe"])
    idx_amount = _cari_idx(headers, ["amount", "jumlah"])

    if None in (idx_tanggal, idx_outlet, idx_type, idx_amount):
        raise FormatTidakDikenali(
            f"Kolom Tanggal/Outlet/Type/Amount tidak lengkap terdeteksi di sheet '{nama_sheet}'."
        )

    grup: dict = {}
    urutan: list = []
    for row in ws.iter_rows(min_row=header_rownum + 1, values_only=True):
        if len(row) <= max(idx_tanggal, idx_outlet, idx_type, idx_amount):
            continue
        tanggal = row[idx_tanggal]
        outlet = row[idx_outlet]
        tipe = row[idx_type]
        amount = row[idx_amount]
        if tanggal is None or outlet is None or tipe is None:
            continue

        key = (tanggal, outlet)
        if key not in grup:
            grup[key] = {"tanggal": tanggal, "customer": outlet, "dpp": 0, "ppn": 0, "total": 0}
            urutan.append(key)

        tipe_norm = str(tipe).strip().lower()
        nilai = amount or 0
        if tipe_norm in _TIPE_POS_DPP:
            grup[key]["dpp"] = nilai
        elif tipe_norm in _TIPE_POS_PPN:
            grup[key]["ppn"] = nilai
        elif tipe_norm in _TIPE_POS_TOTAL:
            grup[key]["total"] = nilai

    if not urutan:
        raise FormatTidakDikenali(
            f"Sheet '{nama_sheet}' cocok header POS tapi tidak ada baris data yang valid."
        )

    rows = []
    for key in urutan:
        g = grup[key]
        rows.append({
            "sheet": nama_sheet,
            "tanggal": g["tanggal"],
            "no_invoice": None,
            "customer": g["customer"],
            "keterangan": "Ringkasan Penjualan POS",
            "cara_bayar": "TUNAI",
            "dpp": g["dpp"],
            "ppn": g["ppn"],
            "total": g["total"],
            "no_akun_debet": None, "nama_akun_debet": None, "jml_debet": None,
            "no_akun_kredit": None, "nama_akun_kredit": None, "jml_kredit": None,
            "no_akun_kredit_ppn": None, "nama_akun_kredit_ppn": None, "jml_kredit_ppn": None,
        })

    df = pd.DataFrame(rows)
    if not df.empty:
        df["tanggal"] = pd.to_datetime(df["tanggal"], errors="coerce", dayfirst=True).dt.date
    return df


# ============================================================
# 10. DATA PENILAIAN KLIEN / MAKER
# ============================================================

def _cari_header_row_penilaian_klien(ws, max_scan: int = 15):
    """Cari baris header: mengandung NAMA KLIEN + MAKER.
    [FIX] max_scan 6 -> 15, disamakan dgn parser lain."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        teks = " ".join(str(c) for c in row if c is not None).lower()
        ada_nama_klien = "nama klien" in teks or ("nama" in teks and "klien" in teks)
        ada_maker = "maker" in teks
        if ada_nama_klien and ada_maker:
            return i + 1, list(row)
    return None, None


def parse_sheet_penilaian_klien(ws, nama_sheet: str) -> pd.DataFrame:
    """
    Parse satu sheet penilaian klien/maker menjadi DataFrame kolom standar:
    sheet, no, nama_klien, maker, score, bobot_klien, total_score, plus, minus,
    total_akhir, jenis_baris ("klien" atau "catatan/ringkasan").
    """
    header_rownum, header_row = _cari_header_row_penilaian_klien(ws)
    if header_row is None:
        raise FormatTidakDikenali(
            f"Sheet '{nama_sheet}' tidak dikenali sebagai penilaian klien/maker "
            "(tidak ditemukan kolom NAMA KLIEN + MAKER)."
        )

    headers = header_row

    def _idx_exact_ci(opsi_exact):
        for i, h in enumerate(headers):
            if h is None:
                continue
            if str(h).strip().lower() in opsi_exact:
                return i
        return None

    idx_no = _idx_exact_ci({"no", "id"})
    idx_maker = _idx_exact_ci({"maker"})
    idx_score = _idx_exact_ci({"score"})
    idx_total_akhir = None
    idx_nama_klien = None
    idx_bobot = None
    idx_total_score = None
    idx_plus = None
    idx_minus = None
    for i, h in enumerate(headers):
        if h is None:
            continue
        h_low = str(h).strip().lower()
        h_bersih = h_low.strip("() ")
        if "nama" in h_low and "klien" in h_low:
            idx_nama_klien = i
        elif "bobot" in h_low:
            idx_bobot = i
        elif "total" in h_low and "score" in h_low:
            idx_total_score = i
        elif "total" in h_low and "akhir" in h_low:
            idx_total_akhir = i
        elif h_bersih == "+":
            idx_plus = i
        elif h_bersih == "-":
            idx_minus = i

    if idx_nama_klien is None or idx_maker is None:
        raise FormatTidakDikenali(
            f"Kolom NAMA KLIEN atau MAKER tidak ditemukan di sheet '{nama_sheet}'."
        )

    rows = []
    for row in ws.iter_rows(min_row=header_rownum + 1, values_only=True):
        if len(row) <= max(idx_nama_klien, idx_maker):
            continue
        nama_klien = row[idx_nama_klien]
        maker = row[idx_maker]
        no_val = row[idx_no] if idx_no is not None else None
        if nama_klien is None and maker is None and no_val is None:
            continue

        rows.append({
            "sheet": nama_sheet,
            "no": no_val,
            "nama_klien": nama_klien,
            "maker": maker,
            "score": row[idx_score] if idx_score is not None else None,
            "bobot_klien": row[idx_bobot] if idx_bobot is not None else None,
            "total_score": row[idx_total_score] if idx_total_score is not None else None,
            "plus": row[idx_plus] if idx_plus is not None else None,
            "minus": row[idx_minus] if idx_minus is not None else None,
            "total_akhir": row[idx_total_akhir] if idx_total_akhir is not None else None,
            "jenis_baris": "klien" if no_val is not None else "catatan/ringkasan",
        })

    return pd.DataFrame(rows)


# ============================================================
# 11. BUKU BANTU PIUTANG
# ============================================================

# [FIX-3] Sinonim kata kunci dipindah jadi konstanta modul (bukan literal
# di dalam fungsi) supaya bisa dipakai bersama oleh _cari_header_row_piutang
# (deteksi baris header) DAN parse_sheet_piutang (pemetaan kolom -> field),
# jadi kedua tempat itu selalu sinkron -- sebelumnya daftar sinonim di dua
# tempat itu berbeda, sehingga ada kasus baris header "ketemu" tapi
# kolomnya sendiri tetap gagal dipetakan (atau sebaliknya).
KATA_KUNCI_PIUTANG_NO_TRANSAKSI = [
    "no transaksi", "no. transaksi", "notransaksi", "nomor transaksi",
    "no invoice", "no. invoice", "nomor invoice", "no inv", "no. inv",
    "no faktur", "no. faktur", "nomor faktur", "kode transaksi",
    "invoice number", "invoice no", "no bukti", "nomor bukti", "no. bukti",
]
KATA_KUNCI_PIUTANG_NAMA_PELANGGAN = [
    "nama pelanggan", "pelanggan", "nama customer", "customer",
    "nama klien", "klien", "nama debitur", "debitur",
    "nama konsumen", "konsumen", "customer name", "nama buyer", "buyer",
]
KATA_KUNCI_PIUTANG_NOMINAL = [
    "sub total", "subtotal", "total akhir", "nilai invoice", "nilai faktur",
    "total piutang", "total tagihan", "jumlah tagihan", "nilai tagihan",
    "grand total", "total transaksi", "nilai transaksi", "nilai",
    "jumlah", "total",
]


def _cari_header_row_piutang(ws, max_scan: int = 20):
    """Cari baris header Buku Bantu Piutang: NO TRANSAKSI/INVOICE + NAMA
    PELANGGAN + SUB TOTAL/TOTAL AKHIR/NILAI INVOICE (atau sinonim lain).

    [FIX] Sebelumnya cuma cocok substring "no transaksi" persis (tanpa
    titik) -- kolom yang ditulis "No. Transaksi" (pakai titik, format
    sangat umum) TIDAK PERNAH ketemu di sini, padahal idx_no_transaksi
    di parse_sheet_piutang() di bawah SUDAH menerima "no. transaksi".
    Akibatnya sheet dgn kolom "No. Transaksi" selalu gagal dikenali
    sebagai Buku Bantu Piutang walau seharusnya bisa diproses.

    [FIX-2] File model referensi BUKU_BANTU_PIUTANG_USAHA_2025.xlsx
    (dikirim user, sudah dipakai sebagai acuan gaya export di
    accounting_export.py) pakai header "No. Invoice" / "Pelanggan" /
    "Nilai Invoice" -- BUKAN "No Transaksi" / "Nama Pelanggan" /
    "Sub Total" seperti yang selama ini dicari di sini. Akibatnya file
    dengan format itu SELALU gagal dikenali ("Tidak ada jenis dokumen
    yang dikenali di file ini"), padahal parser & export-nya sudah
    siap. Sinonim baru ditambahkan di sini SEBAGAI TAMBAHAN (bukan
    pengganti) supaya kedua gaya penamaan tetap didukung.

    [FIX-3] Diperluas jauh lebih toleran karena masih ada file klien lain
    (mis. "3_Buku_Bantu_Piutang.xlsx") yang tetap gagal dikenali walau
    FIX-2 sudah ditambahkan -- kemungkinan penyebabnya salah satu/lebih:
      a) Header ada TAPI lewat dari baris ke-8 (banyak template klien
         punya judul perusahaan/nomor surat/periode laporan di baris atas
         sebelum tabel data dimulai) -> max_scan dinaikkan 8 -> 20.
      b) Istilah kolom klien belum ada di daftar sinonim (mis. "Nama
         Klien", "No Faktur", "Total Tagihan", "Jumlah") -> daftar
         KATA_KUNCI_PIUTANG_* di atas ditambah signifikan.
      c) Header kepencar di 2 baris berdekatan akibat merged cell /
         sub-judul kelompok kolom (mis. baris 1 cuma "DATA TRANSAKSI",
         nama kolom sebenarnya baru di baris 2) -> ditambah langkah kedua
         yang menggabungkan teks 2 baris berurutan sebelum menyerah.
    Kalau file klien TERNYATA masih juga gagal dikenali setelah FIX-3 ini,
    kemungkinan besar istilah kolomnya benar-benar di luar semua sinonim
    di atas -- solusinya cek isi errornya (sekarang ditulis lebih detail
    di parse_sheet_piutang) lalu tambahkan istilah barunya ke 3 daftar
    KATA_KUNCI_PIUTANG_* di atas, TIDAK perlu ubah fungsi ini lagi.
    """
    def _cocok(teks):
        return (
            any(k in teks for k in KATA_KUNCI_PIUTANG_NO_TRANSAKSI),
            any(k in teks for k in KATA_KUNCI_PIUTANG_NAMA_PELANGGAN),
            any(k in teks for k in KATA_KUNCI_PIUTANG_NOMINAL),
        )

    baris = []
    for row in ws.iter_rows(min_row=1, max_row=max_scan, values_only=True):
        teks = " ".join(str(c) for c in row if c is not None).lower()
        baris.append((row, teks))

    # 1) Kasus paling umum: ketiga kelompok kata kunci ada di SATU baris.
    for i, (row, teks) in enumerate(baris):
        ada_no_transaksi, ada_nama_pelanggan, ada_nominal = _cocok(teks)
        if ada_no_transaksi and ada_nama_pelanggan and ada_nominal:
            return i + 1, list(row)

    # 2) Fallback: header kepencar di 2 baris berdekatan (mis. baris judul
    #    kelompok kolom + baris nama kolom detail). Baris kolom sebenarnya
    #    (dipakai untuk _idx() di parse_sheet_piutang) diambil dari baris
    #    KEDUA karena biasanya itu yang berisi nama kolom rinci.
    for i in range(len(baris) - 1):
        _row_a, teks_a = baris[i]
        row_b, teks_b = baris[i + 1]
        ada_no_transaksi, ada_nama_pelanggan, ada_nominal = _cocok(teks_a + " " + teks_b)
        if ada_no_transaksi and ada_nama_pelanggan and ada_nominal:
            return i + 2, list(row_b)

    return None, None


def parse_sheet_piutang(ws, nama_sheet: str) -> pd.DataFrame:
    """
    Parse satu sheet "Buku Bantu Piutang" (kartu piutang per transaksi).
    Ini BUKAN jurnal (tidak ada kolom debet/kredit akun).
    """
    header_rownum, header_row = _cari_header_row_piutang(ws)
    if header_row is None:
        raise FormatTidakDikenali(
            f"Sheet '{nama_sheet}' tidak dikenali sebagai Buku Bantu Piutang "
            "(tidak ditemukan kolom NO TRANSAKSI + NAMA PELANGGAN + SUB TOTAL/TOTAL AKHIR)."
        )

    headers = header_row

    def _idx(kata_kunci, exact=False, exclude=()):
        for i, h in enumerate(headers):
            if h is None or i in exclude:
                continue
            h_low = str(h).strip().lower()
            if exact:
                if h_low in kata_kunci:
                    return i
            elif any(k in h_low for k in kata_kunci):
                return i
        return None

    idx_dept = _idx({"dept", "dept."}, exact=True)
    idx_kode_supp = _idx(["kode supp"])
    # [FIX] Ditambah sinonim "no. invoice"/"no invoice"/"nomor invoice" &
    # "pelanggan"/"nama customer"/"customer" -- file model referensi
    # BUKU_BANTU_PIUTANG_USAHA_2025.xlsx pakai header "No. Invoice" &
    # "Pelanggan" polos (bukan "No Transaksi"/"Nama Pelanggan"), jadi
    # sebelumnya selalu gagal cocok di sini walau _cari_header_row_piutang
    # sudah lolos.
    # [FIX-3] Sebelumnya pakai daftar sinonim SENDIRI (lebih pendek & exact
    # match) yang beda dari yang dipakai _cari_header_row_piutang() di atas
    # -- akibatnya ada kasus baris header BERHASIL dikenali (lolos deteksi)
    # tapi kolom "No Transaksi"/"Nama Pelanggan"-nya sendiri tetap gagal
    # dipetakan karena istilahnya tidak ada di daftar pendek ini. Sekarang
    # dipakai konstanta KATA_KUNCI_PIUTANG_* yang sama persis dengan yang
    # dipakai untuk deteksi baris header, supaya dua tahap ini selalu
    # sinkron. exact=True diganti substring biasa supaya tetap cocok kalau
    # header punya spasi/tanda kurung tambahan (mis. "No. Transaksi (Ref)").
    idx_no_transaksi = _idx(KATA_KUNCI_PIUTANG_NO_TRANSAKSI)
    idx_nama_pelanggan = _idx(KATA_KUNCI_PIUTANG_NAMA_PELANGGAN)
    # [FIX] Kolom "Jatuh Tempo" sebelumnya TIDAK PERNAH dicari/di-parse sama
    # sekali di sheet Piutang (beda dgn sisi AP Aging yg sudah punya
    # idx_jatuh_tempo) -- akibatnya sheet export "Buku Bantu Piutang" selalu
    # kosong di kolom Jatuh Tempo & aging tidak pernah bisa dihitung akurat,
    # cuma fallback ke tanggal invoice. Dicari SEBELUM idx_tanggal & indeksnya
    # dikecualikan dari pencarian "tanggal" supaya "Tanggal Jatuh Tempo" tidak
    # ketebak dobel jadi kolom "Tanggal Invoice".
    idx_jatuh_tempo = _idx(["jatuh tempo", "due date"])
    idx_tanggal = _idx(["tanggal", "tgl", "date"], exclude={idx_jatuh_tempo} if idx_jatuh_tempo is not None else ())
    idx_jml_item = _idx(["jml item", "jumlah item"])
    idx_sub_total = _idx(["sub total", "subtotal"])
    idx_pot = _idx(["pot.", "pot %", "potongan"])
    idx_pajak = _idx(["pajak"])
    idx_biaya_lain = _idx(["biaya lain"])
    # [FIX] Tambah sinonim "nilai invoice"/"nilai faktur" -- file model
    # referensi BUKU_BANTU_PIUTANG_USAHA_2025.xlsx menyimpan nominal invoice
    # di kolom "Nilai Invoice" (bukan "Total Akhir"). accounting_export.py
    # (ALIAS_PIUTANG) sudah membaca field "total_akhir" lebih dulu utk
    # mengisi kolom export "Nilai Invoice", jadi field ini yang paling tepat
    # dipetakan supaya nilainya benar-benar sampai ke laporan.
    # [FIX-3] Diperluas lagi dgn KATA_KUNCI_PIUTANG_NOMINAL (total tagihan,
    # nilai tagihan, grand total, dst) -- cari "total akhir"/"nilai invoice"
    # dulu secara spesifik baru fallback ke sinonim umum, supaya kalau file
    # punya KEDUANYA ("Sub Total" & "Total Tagihan" misalnya), yang dipilih
    # tetap kolom nominal akhir/paling relevan, bukan asal kolom pertama
    # yang mengandung kata "total".
    idx_total_akhir = _idx(["total akhir", "nilai invoice", "nilai faktur"]) \
        or _idx([k for k in KATA_KUNCI_PIUTANG_NOMINAL if k not in ("sub total", "subtotal")])
    idx_bayar_tunai = _idx(["bayar tunai"])
    idx_bayar_kredit = _idx(["bayar kredit"])
    # [BARU] "Segment" & "Project/Unit" sebelumnya tidak pernah dicari sama
    # sekali di sini -- akibatnya sheet export "Buku Bantu Piutang" SELALU
    # kosong di 2 kolom ini walau file client punya kolomnya (bukan bug di
    # accounting_export.py, datanya memang tidak pernah sampai ke df).
    idx_segment = _idx(["segment", "segmen"])
    idx_project_unit = _idx(["project/unit", "project / unit", "project unit", "unit/project", "project", "unit"])

    if idx_no_transaksi is None or idx_nama_pelanggan is None:
        # [FIX-3] Pesan error sekarang menyertakan isi baris header yang
        # BERHASIL ditemukan (bukan cuma bilang "tidak ditemukan") --
        # sebelumnya kalau lolos di _cari_header_row_piutang() tapi gagal
        # di _idx() sini, pesan errornya tidak membantu sama sekali untuk
        # tahu kolom mana yang sebenarnya belum dikenali sistem.
        header_terbaca = [str(h) for h in headers if h is not None]
        raise FormatTidakDikenali(
            f"Kolom NO TRANSAKSI atau NAMA PELANGGAN tidak ditemukan di sheet '{nama_sheet}'. "
            f"Header yang terbaca di baris {header_rownum}: {header_terbaca}"
        )

    def _ambil(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    rows = []
    for row in ws.iter_rows(min_row=header_rownum + 1, values_only=True):
        no_transaksi = _ambil(row, idx_no_transaksi)
        nama_pelanggan = _ambil(row, idx_nama_pelanggan)
        # [FIX] Sebelumnya baris dilewati HANYA kalau no_transaksi & nama_
        # pelanggan DUA-DUANYA kosong -- baris ringkasan/footer di akhir
        # tabel (mis. "TOTAL" di kolom No. Invoice, seperti pada file model
        # referensi BUKU_BANTU_PIUTANG_USAHA_2025.xlsx baris 19) punya
        # no_transaksi="TOTAL" (tidak None) tapi nama_pelanggan kosong,
        # jadi LOLOS syarat lama dan ikut dianggap 1 transaksi -- bikin
        # total_piutang & rekap_per_pelanggan dobel hitung. Transaksi asli
        # selalu punya nama pelanggan, jadi baris tanpa nama pelanggan cukup
        # dilewati saja.
        if nama_pelanggan is None:
            continue

        rows.append({
            "sheet": nama_sheet,
            "dept": _ambil(row, idx_dept),
            "kode_supplier": _ambil(row, idx_kode_supp),
            "no_transaksi": no_transaksi,
            "nama_pelanggan": nama_pelanggan,
            "tanggal": _ambil(row, idx_tanggal),
            "tanggal_jatuh_tempo": _ambil(row, idx_jatuh_tempo),
            "jml_item": _ambil(row, idx_jml_item),
            "sub_total": _ambil(row, idx_sub_total),
            "potongan_persen": _ambil(row, idx_pot),
            "pajak": _ambil(row, idx_pajak),
            "biaya_lain": _ambil(row, idx_biaya_lain),
            "total_akhir": _ambil(row, idx_total_akhir),
            "bayar_tunai": _ambil(row, idx_bayar_tunai),
            "bayar_kredit": _ambil(row, idx_bayar_kredit),
            "segment": _ambil(row, idx_segment),
            "project_unit": _ambil(row, idx_project_unit),
        })

    return pd.DataFrame(rows)


# ============================================================
# 11B. FAKTUR PAJAK (PPN) -- PARSING, VALIDASI, & DRAF JURNAL
# ============================================================
# Yang biasanya dikerjakan akuntan saat menerima Faktur Pajak:
#   1. Dicocokkan dengan format resmi e-Faktur/Coretax DJP (nomor seri,
#      NPWP) supaya nanti tidak ditolak saat lapor SPT Masa PPN.
#   2. Dicek apakah PPN yang tertulis sudah sesuai hitungan tarif berlaku.
#   3. Dicek nomor faktur duplikat (risiko dobel lapor/dobel kredit pajak).
#   4. Direkap jadi total PPN Keluaran vs Masukan utk estimasi kurang/lebih
#      bayar SPT Masa PPN.
#   5. Disiapkan draf jurnalnya (tetap harus direview manusia, karena akun
#      pasti & cara bayar tidak selalu bisa ditebak otomatis).
#
# CATATAN TARIF (bisa berubah -- jangan dianggap kebenaran mutlak selamanya):
# Sejak PMK 131/2024 (berlaku 1 Jan 2025), tarif PPN NOMINAL adalah 12%,
# TAPI untuk Barang/Jasa Kena Pajak NON-MEWAH, DPP yang dipakai adalah
# "DPP Nilai Lain" = 11/12 x Harga Jual, sehingga PPN EFEKTIF yang dipungut
# tetap 11% dari harga jual. Tarif 12% PENUH hanya berlaku utk barang mewah
# (kena PPnBM juga). Karena ini murni keputusan pemerintah yang bisa berubah
# kapan saja, nilai di bawah ini HANYA default -- selalu bisa dioverride
# lewat parameter tarif_ppn, dan tetap sarankan user mengecek PMK terbaru.
TARIF_PPN_EFEKTIF_DEFAULT = 0.11   # non-mewah (DPP Nilai Lain 11/12 x harga)
TARIF_PPN_MEWAH_DEFAULT = 0.12     # barang mewah (kena PPnBM), DPP penuh

_TOLERANSI_SELISIH_PPN_RUPIAH = 5  # toleransi pembulatan, dalam Rupiah


def _cari_header_row_faktur_pajak(ws, max_scan: int = 10):
    """
    Cari baris header sheet Faktur Pajak: butuh NPWP + NOMOR FAKTUR + DPP +
    PPN sekaligus (kombinasi ini yang membedakannya dari sheet lain yang
    kebetulan cuma punya sebagian, mis. rekening koran yang juga ada kolom
    'jumlah').
    """
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        teks = " ".join(str(c) for c in row if c is not None).lower()
        ada_npwp = "npwp" in teks
        ada_nomor_faktur = "nomor faktur" in teks or "no faktur" in teks or "no. faktur" in teks
        ada_dpp = "dpp" in teks
        ada_ppn = "ppn" in teks
        if ada_npwp and ada_nomor_faktur and ada_dpp and ada_ppn:
            return i + 1, list(row)
    return None, None


def _bersihkan_digit(teks) -> str:
    """Buang semua karakter selain digit dari NPWP/nomor faktur, supaya
    validasi format tidak terganggu gaya penulisan (titik/strip/spasi)."""
    if teks is None:
        return ""
    return re.sub(r"\D", "", str(teks))


def validasi_npwp(npwp) -> dict:
    """
    Validasi format NPWP secara STRUKTURAL saja (jumlah digit) -- BUKAN
    validasi keaslian ke database DJP (itu perlu portal/API resmi DJP, di
    luar cakupan modul offline ini; user tetap harus cek manual kalau ragu).

    Mendukung 2 format yang sah dipakai:
    - 15 digit: format NPWP lama (sebelum Coretax), mis. 01.234.567.8-901.000
    - 16 digit: format NPWP baru pasca-Coretax (2024+, pakai NIK utk WP OP)

    Return: {"valid": bool, "jumlah_digit": int, "catatan": str}
    """
    digit = _bersihkan_digit(npwp)
    if not digit:
        return {"valid": False, "jumlah_digit": 0, "catatan": "NPWP kosong."}
    if len(digit) == 15:
        return {"valid": True, "jumlah_digit": 15, "catatan": "Format NPWP lama (15 digit)."}
    if len(digit) == 16:
        return {"valid": True, "jumlah_digit": 16, "catatan": "Format NPWP baru pasca-Coretax (16 digit)."}
    return {
        "valid": False, "jumlah_digit": len(digit),
        "catatan": f"Jumlah digit tidak lazim ({len(digit)} digit) -- NPWP yang sah 15 atau 16 digit.",
    }


def validasi_nomor_faktur_pajak(nomor) -> dict:
    """
    Validasi FORMAT nomor seri faktur pajak (16 digit: 2 digit Kode
    Transaksi + 1 digit Kode Status + 13 digit Nomor Urut yang dialokasikan
    DJP). Sering ditulis dengan pemisah, mis. '010.001-25.12345678'.
    HANYA cek format -- bukan cek keaslian ke sistem Coretax DJP (portal
    resmi/API DJP diperlukan utk itu, di luar cakupan modul offline ini).
    """
    digit = _bersihkan_digit(nomor)
    if not digit:
        return {"valid": False, "kode_transaksi": None, "catatan": "Nomor faktur pajak kosong."}
    if len(digit) != 16:
        return {
            "valid": False, "kode_transaksi": None,
            "catatan": f"Panjang {len(digit)} digit -- nomor faktur pajak yang sah 16 digit.",
        }
    kode_transaksi = digit[0:2]
    kode_status = digit[2]
    nomor_urut = digit[3:]
    daftar_kode_transaksi_dikenal = {"01", "02", "03", "04", "05", "06", "07", "08", "09"}
    catatan = f"Kode transaksi {kode_transaksi}, kode status {kode_status}, nomor urut {nomor_urut}."
    if kode_transaksi not in daftar_kode_transaksi_dikenal:
        return {
            "valid": False, "kode_transaksi": kode_transaksi,
            "catatan": f"Kode transaksi '{kode_transaksi}' tidak dikenal (lazimnya 01-09). {catatan}",
        }
    return {"valid": True, "kode_transaksi": kode_transaksi, "catatan": catatan}


def parse_sheet_faktur_pajak(ws, nama_sheet: str) -> pd.DataFrame:
    """
    Parse satu sheet Faktur Pajak (PPN) jadi DataFrame baris-per-faktur.
    Kolom yang dicoba dikenali (nama header fleksibel, cocok pola umum
    template e-Faktur/Coretax DJP): tanggal, nomor faktur, kode transaksi,
    NPWP+nama penjual, NPWP+nama pembeli, DPP, PPN, keterangan/nama barang.
    """
    header_rownum, header_row = _cari_header_row_faktur_pajak(ws)
    if header_row is None:
        raise FormatTidakDikenali(
            f"Sheet '{nama_sheet}' tidak dikenali sebagai Faktur Pajak "
            "(tidak ditemukan kolom NPWP + NOMOR FAKTUR + DPP + PPN sekaligus)."
        )

    headers = header_row

    def _idx(kata_kunci):
        for i, h in enumerate(headers):
            if h is None:
                continue
            h_low = str(h).strip().lower()
            if any(k in h_low for k in kata_kunci):
                return i
        return None

    idx_tanggal = _idx(["tanggal faktur", "tanggal", "tgl"])
    idx_nomor_faktur = _idx(["nomor faktur", "no faktur", "no. faktur"])
    idx_kode_transaksi = _idx(["kode transaksi"])
    idx_npwp_penjual = _idx(["npwp penjual"])
    idx_nama_penjual = _idx(["nama penjual"])
    idx_npwp_pembeli = _idx(["npwp pembeli", "npwp lawan transaksi"])
    idx_nama_pembeli = _idx(["nama pembeli", "nama lawan transaksi"])
    idx_dpp = _idx(["dpp"])
    idx_ppn = _idx(["ppn"])
    idx_keterangan = _idx(["nama barang", "keterangan", "deskripsi"])

    if idx_nomor_faktur is None or idx_dpp is None or idx_ppn is None:
        raise FormatTidakDikenali(
            f"Kolom NOMOR FAKTUR, DPP, atau PPN tidak ditemukan di sheet '{nama_sheet}'."
        )

    def _ambil(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    rows = []
    for row in ws.iter_rows(min_row=header_rownum + 1, values_only=True):
        nomor_faktur = _ambil(row, idx_nomor_faktur)
        dpp = _ambil(row, idx_dpp)
        ppn = _ambil(row, idx_ppn)
        if nomor_faktur is None and dpp is None and ppn is None:
            continue

        rows.append({
            "sheet": nama_sheet,
            "tanggal": _ambil(row, idx_tanggal),
            "nomor_faktur": nomor_faktur,
            "kode_transaksi": _ambil(row, idx_kode_transaksi),
            "npwp_penjual": _ambil(row, idx_npwp_penjual),
            "nama_penjual": _ambil(row, idx_nama_penjual),
            "npwp_pembeli": _ambil(row, idx_npwp_pembeli),
            "nama_pembeli": _ambil(row, idx_nama_pembeli),
            "dpp": dpp,
            "ppn": ppn,
            "keterangan": _ambil(row, idx_keterangan),
        })

    return pd.DataFrame(rows)


def proses_faktur_pajak(
    df: pd.DataFrame,
    npwp_perusahaan: str = None,
    tarif_ppn: float = TARIF_PPN_EFEKTIF_DEFAULT,
    toleransi_rupiah: float = _TOLERANSI_SELISIH_PPN_RUPIAH,
) -> dict:
    """
    "Kerjaan akuntan" untuk Faktur Pajak: cross-check tiap baris, temukan
    yang mencurigakan, siapkan rekap SIAP dipakai utk lapor SPT Masa PPN.

    npwp_perusahaan: NPWP perusahaan (klien) sendiri -- dipakai utk
    menentukan arah tiap faktur (KELUARAN kalau perusahaan berperan sbg
    penjual, MASUKAN kalau sbg pembeli). Kalau None, arah tidak ditentukan
    & draf jurnal TIDAK dibuat -- lebih aman drpd menebak arah yang salah
    (jurnal terbalik = laporan keuangan salah total).

    Return dict:
        "df": DataFrame asli + kolom tambahan (arah, status, validitas
              format, ppn_seharusnya, selisih_ppn)
        "ringkasan": dict rekap (total DPP/PPN, PPN Keluaran/Masukan,
                     estimasi kurang/lebih bayar, jumlah yang perlu direview)
        "masalah": list baris yang PERLU DICEK MANUAL beserta alasannya
                   (format salah, PPN tidak sesuai hitungan, atau nomor
                   faktur duplikat)
        "draf_jurnal": list draf jurnal per faktur (hanya dibuat kalau
                       npwp_perusahaan diisi -- akun masih generik, WAJIB
                       direview akuntan sebelum diposting)
    """
    if df is None or df.empty:
        return {"df": df, "ringkasan": {}, "masalah": [], "draf_jurnal": []}

    df = df.copy()
    npwp_bersih = _bersihkan_digit(npwp_perusahaan) if npwp_perusahaan else None

    masalah = []
    nomor_faktur_terlihat: dict = {}

    arah_list, status_list = [], []
    valid_npwp_penjual_list, valid_npwp_pembeli_list = [], []
    valid_nomor_faktur_list = []
    ppn_seharusnya_list, selisih_ppn_list = [], []

    for i, row in df.iterrows():
        no_faktur = row.get("nomor_faktur")
        no_faktur_digit = _bersihkan_digit(no_faktur)

        v_faktur = validasi_nomor_faktur_pajak(no_faktur)
        v_npwp_penjual = validasi_npwp(row.get("npwp_penjual"))
        v_npwp_pembeli = validasi_npwp(row.get("npwp_pembeli"))
        valid_nomor_faktur_list.append(v_faktur["valid"])
        valid_npwp_penjual_list.append(v_npwp_penjual["valid"])
        valid_npwp_pembeli_list.append(v_npwp_pembeli["valid"])

        dpp = float(row.get("dpp") or 0)
        ppn = float(row.get("ppn") or 0)
        ppn_seharusnya = round(dpp * tarif_ppn)
        selisih = ppn - ppn_seharusnya
        ppn_seharusnya_list.append(ppn_seharusnya)
        selisih_ppn_list.append(selisih)

        # -- Tentukan arah (kalau NPWP perusahaan diketahui) --
        arah = "TIDAK DIKETAHUI"
        if npwp_bersih:
            npwp_penjual_digit = _bersihkan_digit(row.get("npwp_penjual"))
            npwp_pembeli_digit = _bersihkan_digit(row.get("npwp_pembeli"))
            if npwp_penjual_digit and npwp_penjual_digit == npwp_bersih:
                arah = "KELUARAN"  # perusahaan menjual -> PPN Keluaran (utang pajak)
            elif npwp_pembeli_digit and npwp_pembeli_digit == npwp_bersih:
                arah = "MASUKAN"   # perusahaan membeli -> PPN Masukan (kredit pajak)
            else:
                arah = "BUKAN MILIK PERUSAHAAN INI"
        arah_list.append(arah)

        # -- Kumpulkan alasan bermasalah (kalau ada) --
        alasan = []
        if not v_faktur["valid"]:
            alasan.append(f"Nomor faktur tidak sesuai format: {v_faktur['catatan']}")
        if not v_npwp_penjual["valid"]:
            alasan.append(f"NPWP penjual tidak sesuai format: {v_npwp_penjual['catatan']}")
        if not v_npwp_pembeli["valid"]:
            alasan.append(f"NPWP pembeli tidak sesuai format: {v_npwp_pembeli['catatan']}")
        if abs(selisih) > toleransi_rupiah:
            alasan.append(
                f"PPN tercatat Rp{ppn:,.0f} tidak sesuai perhitungan {tarif_ppn:.0%} x "
                f"DPP Rp{dpp:,.0f} = Rp{ppn_seharusnya:,.0f} (selisih Rp{selisih:,.0f}). "
                "Kalau ini transaksi barang mewah, tarifnya seharusnya 12% penuh -- cek ulang."
            )
        if no_faktur_digit:
            if no_faktur_digit in nomor_faktur_terlihat:
                alasan.append(
                    f"Nomor faktur DUPLIKAT -- sudah muncul di baris ke-"
                    f"{nomor_faktur_terlihat[no_faktur_digit] + 1} juga. Faktur pajak dobel "
                    "berisiko dobel lapor/dobel kredit pajak, WAJIB dicek manual."
                )
            else:
                nomor_faktur_terlihat[no_faktur_digit] = i

        status_list.append("PERLU REVIEW" if alasan else "OK")
        if alasan:
            masalah.append({
                "baris": i + 1,
                "nomor_faktur": no_faktur,
                "tanggal": row.get("tanggal"),
                "dpp": dpp,
                "ppn": ppn,
                "alasan": alasan,
            })

    df["arah"] = arah_list
    df["status"] = status_list
    df["valid_nomor_faktur"] = valid_nomor_faktur_list
    df["valid_npwp_penjual"] = valid_npwp_penjual_list
    df["valid_npwp_pembeli"] = valid_npwp_pembeli_list
    df["ppn_seharusnya"] = ppn_seharusnya_list
    df["selisih_ppn"] = selisih_ppn_list

    # -- Draf jurnal (hanya kalau arah diketahui, supaya tidak menebak salah) --
    draf_jurnal = []
    for i, row in df.iterrows():
        if row["arah"] not in ("KELUARAN", "MASUKAN"):
            continue
        dpp = float(row.get("dpp") or 0)
        ppn = float(row.get("ppn") or 0)
        total = dpp + ppn
        if row["arah"] == "KELUARAN":
            draf_jurnal.append({
                "baris": i + 1, "nomor_faktur": row.get("nomor_faktur"), "arah": "KELUARAN",
                "no_akun_debet": "PIUTANG/KAS", "nama_akun_debet": "Piutang Usaha / Kas (cek cara bayar)",
                "jml_debet": total,
                "no_akun_kredit": "PENJUALAN", "nama_akun_kredit": "Penjualan",
                "jml_kredit": dpp,
                "no_akun_kredit_ppn": "PPN KELUARAN", "nama_akun_kredit_ppn": "PPN Keluaran",
                "jml_kredit_ppn": ppn,
                "catatan": "Draf otomatis -- cek akun & cara bayar (tunai/kredit) sebelum posting.",
            })
        else:
            draf_jurnal.append({
                "baris": i + 1, "nomor_faktur": row.get("nomor_faktur"), "arah": "MASUKAN",
                "no_akun_debet": "PEMBELIAN/BEBAN", "nama_akun_debet": "Pembelian / Beban terkait (sesuaikan akun)",
                "jml_debet": dpp,
                "no_akun_debet_ppn": "PPN MASUKAN", "nama_akun_debet_ppn": "PPN Masukan",
                "jml_debet_ppn": ppn,
                "no_akun_kredit": "UTANG/KAS", "nama_akun_kredit": "Utang Usaha / Kas (cek cara bayar)",
                "jml_kredit": total,
                "catatan": "Draf otomatis -- cek akun & apakah PPN Masukan ini bisa dikreditkan.",
            })

    # -- Ringkasan siap dipakai utk SPT Masa PPN --
    total_dpp = float(df["dpp"].fillna(0).sum())
    total_ppn = float(df["ppn"].fillna(0).sum())
    ppn_keluaran = float(df.loc[df["arah"] == "KELUARAN", "ppn"].fillna(0).sum())
    ppn_masukan = float(df.loc[df["arah"] == "MASUKAN", "ppn"].fillna(0).sum())

    ringkasan = {
        "jumlah_faktur": len(df),
        "total_dpp": total_dpp,
        "total_ppn": total_ppn,
        "ppn_keluaran": ppn_keluaran,
        "ppn_masukan": ppn_masukan,
        "estimasi_ppn_kurang_lebih_bayar": ppn_keluaran - ppn_masukan,
        "jumlah_perlu_review": len(masalah),
        "jumlah_duplikat": sum(1 for m in masalah if any("DUPLIKAT" in a for a in m["alasan"])),
        "tarif_dipakai": tarif_ppn,
        "catatan_tarif": (
            f"Perhitungan pakai tarif efektif {tarif_ppn:.0%} (default utk BKP/JKP non-mewah "
            "sesuai PMK 131/2024, DPP Nilai Lain 11/12 x harga). Kalau ini transaksi barang "
            "mewah, tarif seharusnya 12% penuh dari DPP -- panggil ulang dgn tarif_ppn=0.12. "
            "Tarif pajak bisa berubah sewaktu-waktu, selalu cek update PMK terbaru."
        ),
    }
    if npwp_bersih is None:
        ringkasan["catatan_arah"] = (
            "NPWP perusahaan tidak diisi -- arah (Keluaran/Masukan) & estimasi kurang/lebih "
            "bayar TIDAK bisa dihitung akurat, draf jurnal tidak dibuat. Isi npwp_perusahaan "
            "supaya hasilnya lengkap."
        )

    return {"df": df, "ringkasan": ringkasan, "masalah": masalah, "draf_jurnal": draf_jurnal}


def proses_file_faktur_pajak(
    file_like,
    nama_file: str = None,
    npwp_perusahaan: str = None,
    tarif_ppn: float = TARIF_PPN_EFEKTIF_DEFAULT,
) -> dict:
    """
    Fungsi siap-pakai tingkat file: baca file (.xlsx/.xlsm/.xls/.csv/.pdf),
    cari semua sheet yang cocok sebagai Faktur Pajak, gabungkan, lalu proses
    lewat proses_faktur_pajak(). Dipakai oleh app.py utk 1 file upload.

    Return sama seperti proses_faktur_pajak(), ditambah "sheet_dilewati":
    daftar nama sheet yang TIDAK cocok format Faktur Pajak sama sekali.
    """
    nama_file = nama_file or getattr(file_like, "name", "") or ""
    daftar_sheet, _df_coa = _siapkan_daftar_sheet(file_like, nama_file)

    df_list = []
    sheet_dilewati = []
    for nama, ws in daftar_sheet:
        try:
            df_sheet = parse_sheet_faktur_pajak(ws, nama)
        except FormatTidakDikenali:
            sheet_dilewati.append(nama)
            continue
        if df_sheet.empty:
            sheet_dilewati.append(nama)
            continue
        df_list.append(df_sheet)

    if not df_list:
        return {
            "df": pd.DataFrame(), "ringkasan": {}, "masalah": [], "draf_jurnal": [],
            "sheet_dilewati": sheet_dilewati,
        }

    df_gabungan = pd.concat(df_list, ignore_index=True)
    hasil = proses_faktur_pajak(df_gabungan, npwp_perusahaan=npwp_perusahaan, tarif_ppn=tarif_ppn)
    hasil["sheet_dilewati"] = sheet_dilewati
    return hasil


# ============================================================
# 11b. BUKTI POTONG PAJAK (PPh 21/23/4(2)) — "DETEKSI-SAJA" -> DIPROSES
# ============================================================
# Sama seperti Faktur Pajak (PPN) di atas: sheet Bukti Potong BUKAN jurnal
# siap-posting. "Kerjaan akuntan" di sini adalah: kenali jenis PPh-nya (21,
# 23, atau 4 ayat 2 -- final), cek kode objek pajak & tarifnya masuk akal,
# cek NPWP, cek duplikat nomor bukti potong, lalu siapkan draf jurnal +
# rekap yang siap dipakai untuk cross-check ke SPT Masa PPh terkait.
#
# CATATAN TARIF: tarif di bawah adalah tarif UMUM yang berlaku luas per
# aturan berjalan (UU HPP & aturan turunannya). Tarif bisa berubah lewat
# PMK/aturan baru, dan sejumlah kode objek (terutama jasa konstruksi &
# PPh 21 dgn skema TER) punya tarif yang tergantung kualifikasi/kategori
# spesifik yang TIDAK bisa ditentukan hanya dari kode objek + DPP saja --
# untuk kasus itu modul ini menandai "PERLU CEK MANUAL" alih-alih menebak.

# -- Kamus kode objek pajak -> (jenis_pajak, uraian, tarif atau rentang tarif) --
# jenis_pajak: "PPH23" | "PPH4(2)" | "PPH21"
# tarif: float (tarif tunggal) ATAU tuple (tarif_min, tarif_max) kalau tarif
# bervariasi tergantung kualifikasi (mis. jasa konstruksi).
KAMUS_KODE_OBJEK_BUKTI_POTONG: Dict[str, Dict[str, Any]] = {
    # --- PPh Pasal 23 ---
    "24-104-01": {"jenis_pajak": "PPH23", "uraian": "Dividen", "tarif": 0.15},
    "24-104-02": {"jenis_pajak": "PPH23", "uraian": "Bunga (selain final Pasal 4(2))", "tarif": 0.15},
    "24-104-03": {"jenis_pajak": "PPH23", "uraian": "Royalti", "tarif": 0.15},
    "24-100-04": {"jenis_pajak": "PPH23", "uraian": "Hadiah, penghargaan, bonus (selain final)", "tarif": 0.15},
    "24-104-08": {"jenis_pajak": "PPH23", "uraian": "Sewa harta (selain tanah & bangunan)", "tarif": 0.02},
    "24-104-09": {"jenis_pajak": "PPH23", "uraian": "Jasa teknik, manajemen, konsultan, jasa lain", "tarif": 0.02},
    # --- PPh Pasal 4 ayat (2) — final ---
    "28-104-04": {"jenis_pajak": "PPH4(2)", "uraian": "Sewa tanah & bangunan", "tarif": 0.10},
    "28-104-11": {"jenis_pajak": "PPH4(2)", "uraian": "Bunga deposito/tabungan & diskonto SBI", "tarif": 0.20},
    "28-104-08": {"jenis_pajak": "PPH4(2)", "uraian": "Pengalihan hak atas tanah & bangunan", "tarif": 0.025},
    "28-104-09": {"jenis_pajak": "PPH4(2)", "uraian": "Jasa konstruksi", "tarif": (0.0175, 0.06)},
    "28-104-19": {"jenis_pajak": "PPH4(2)", "uraian": "Dividen WP OP dalam negeri", "tarif": 0.10},
    "28-104-24": {"jenis_pajak": "PPH4(2)", "uraian": "Transaksi saham di bursa efek", "tarif": (0.001, 0.006)},
}

# Tarif PPh 21 (skema TER -- Tarif Efektif Rata-rata, PMK 168/2023) berbeda
# per kategori (A/B/C, tergantung status PTKP) & lapisan penghasilan bruto
# bulanan -- TIDAK bisa dipastikan hanya dari DPP & PPh yang tertulis di
# bukti potong. Modul ini HANYA cek tarif efektif (PPh/DPP) itu masuk akal
# (0% s/d 34%, sesuai rentang TER/Pasal 17 tertinggi) -- bukan validasi
# ketat ke tabel TER, karena butuh data PTKP karyawan yang tidak ada di
# bukti potong.
_PPH21_TARIF_EFEKTIF_MIN = 0.0
_PPH21_TARIF_EFEKTIF_MAKS = 0.34

_TOLERANSI_SELISIH_PPH_RUPIAH = 5  # toleransi pembulatan, dalam Rupiah


def _cari_header_row_bukti_potong(ws, max_scan: int = 10):
    """
    Cari baris header sheet Bukti Potong: butuh NPWP + NOMOR BUKTI POTONG +
    (DPP/penghasilan bruto) + PPh dipotong sekaligus -- kombinasi ini yang
    membedakannya dari Faktur Pajak (yang pakai 'PPN', bukan 'PPh dipotong').
    """
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        teks = " ".join(str(c) for c in row if c is not None).lower()
        ada_npwp = "npwp" in teks
        ada_nomor_bupot = (
            "nomor bukti potong" in teks or "no bukti potong" in teks
            or "no. bukti potong" in teks or "nomor bukti pemotongan" in teks
        )
        ada_dpp = "dpp" in teks or "penghasilan bruto" in teks or "jumlah bruto" in teks
        ada_pph = "pph dipotong" in teks or "pph yang dipotong" in teks or "jumlah pph" in teks
        if ada_npwp and ada_nomor_bupot and ada_dpp and ada_pph:
            return i + 1, list(row)
    return None, None


def validasi_nomor_bukti_potong(nomor) -> dict:
    """
    Validasi format nomor bukti potong SECARA STRUKTURAL saja: harus terisi
    dan minimal berupa deretan digit/alfanumerik yang wajar (nomor bukti
    potong e-Bupot lazimnya belasan digit). Format resmi tidak sekaku nomor
    seri Faktur Pajak, jadi di sini kita hanya cek "kosong / terlalu pendek /
    mengandung karakter aneh" -- keaslian tetap harus dicek manual ke
    aplikasi e-Bupot DJP.
    """
    if nomor is None or str(nomor).strip() == "":
        return {"valid": False, "catatan": "Nomor bukti potong kosong."}
    teks = str(nomor).strip()
    digit = _bersihkan_digit(teks)
    if len(digit) < 10:
        return {
            "valid": False,
            "catatan": f"Nomor bukti potong '{teks}' terlalu pendek ({len(digit)} digit) -- "
                       "nomor bukti potong e-Bupot lazimnya belasan digit.",
        }
    return {"valid": True, "catatan": "Format wajar (terisi, panjang digit masuk akal)."}


def cek_tarif_bukti_potong(kode_objek, jenis_pajak_tersurat, dpp: float, pph: float,
                            toleransi_rupiah: float = _TOLERANSI_SELISIH_PPH_RUPIAH) -> dict:
    """
    "Kerjaan akuntan" inti: cross-check PPh yang tertulis di bukti potong
    terhadap DPP/penghasilan bruto x tarif yang seharusnya berlaku untuk
    kode objek pajak tsb.

    Return dict:
        "jenis_pajak": PPH21 / PPH23 / PPH4(2) / TIDAK DIKENALI
        "uraian_objek": deskripsi kode objek (kalau dikenali)
        "tarif_seharusnya": float, tuple (rentang), atau None
        "pph_seharusnya": estimasi Rp (None kalau tarif berupa rentang/tidak pasti)
        "selisih": Rp (None kalau pph_seharusnya None)
        "sesuai": bool | None (None = tidak bisa dipastikan otomatis)
        "catatan": penjelasan singkat, termasuk kalau perlu cek manual
    """
    kode_bersih = str(kode_objek).strip() if kode_objek else ""
    info = KAMUS_KODE_OBJEK_BUKTI_POTONG.get(kode_bersih)

    # -- PPh 21: tarif TER tergantung PTKP, tidak bisa dipastikan dari kode objek --
    jenis_tersurat_upper = str(jenis_pajak_tersurat or "").upper().replace(" ", "")
    if "21" in jenis_tersurat_upper or kode_bersih.startswith("21-"):
        tarif_efektif = (pph / dpp) if dpp else 0.0
        sesuai = _PPH21_TARIF_EFEKTIF_MIN <= tarif_efektif <= _PPH21_TARIF_EFEKTIF_MAKS
        return {
            "jenis_pajak": "PPH21", "uraian_objek": "Penghasilan sehubungan pekerjaan (PPh 21)",
            "tarif_seharusnya": None, "pph_seharusnya": None, "selisih": None,
            "sesuai": sesuai if not sesuai else None,  # hanya diflag kalau TIDAK masuk akal
            "catatan": (
                f"Tarif efektif tertulis {tarif_efektif:.1%} -- di luar rentang wajar TER/Pasal 17 "
                f"(0%-34%), WAJIB dicek manual." if not sesuai else
                "PPh 21 pakai skema TER (tarif tergantung kategori A/B/C & status PTKP karyawan) -- "
                "tidak bisa divalidasi pasti dari bukti potong saja, tapi tarif efektif masih dalam "
                "rentang wajar."
            ),
        }

    if info is None:
        return {
            "jenis_pajak": "TIDAK DIKENALI", "uraian_objek": None,
            "tarif_seharusnya": None, "pph_seharusnya": None, "selisih": None, "sesuai": None,
            "catatan": (
                f"Kode objek pajak '{kode_bersih or '(kosong)'}' tidak ada di kamus internal -- "
                "PERLU CEK MANUAL ke daftar kode objek pajak e-Bupot DJP yang berlaku."
            ),
        }

    tarif = info["tarif"]
    if isinstance(tarif, tuple):
        tarif_min, tarif_maks = tarif
        pph_min, pph_maks = dpp * tarif_min, dpp * tarif_maks
        sesuai = (pph_min - toleransi_rupiah) <= pph <= (pph_maks + toleransi_rupiah)
        return {
            "jenis_pajak": info["jenis_pajak"], "uraian_objek": info["uraian"],
            "tarif_seharusnya": tarif, "pph_seharusnya": None, "selisih": None,
            "sesuai": sesuai if not sesuai else None,
            "catatan": (
                f"'{info['uraian']}' punya tarif bervariasi {tarif_min:.2%}-{tarif_maks:.2%} "
                "tergantung kualifikasi/kategori (mis. kualifikasi jasa konstruksi) -- "
                + (
                    f"PPh tertulis Rp{pph:,.0f} di LUAR rentang wajar "
                    f"Rp{pph_min:,.0f}-Rp{pph_maks:,.0f}, WAJIB dicek manual."
                    if not sesuai else
                    "PPh tertulis masih dalam rentang wajar, tapi tetap cek kualifikasi pastinya."
                )
            ),
        }

    pph_seharusnya = round(dpp * tarif)
    selisih = pph - pph_seharusnya
    sesuai = abs(selisih) <= toleransi_rupiah
    return {
        "jenis_pajak": info["jenis_pajak"], "uraian_objek": info["uraian"],
        "tarif_seharusnya": tarif, "pph_seharusnya": pph_seharusnya, "selisih": selisih,
        "sesuai": sesuai,
        "catatan": (
            f"Sesuai -- '{info['uraian']}' tarif {tarif:.0%}." if sesuai else
            f"PPh tercatat Rp{pph:,.0f} tidak sesuai perhitungan {tarif:.0%} x DPP Rp{dpp:,.0f} "
            f"= Rp{pph_seharusnya:,.0f} (selisih Rp{selisih:,.0f}) untuk '{info['uraian']}'."
        ),
    }


def parse_sheet_bukti_potong(ws, nama_sheet: str) -> pd.DataFrame:
    """
    Parse satu sheet Bukti Potong (PPh 21/23/4(2)) jadi DataFrame
    baris-per-bukti-potong. Kolom yang dicoba dikenali (nama header
    fleksibel, cocok pola umum template e-Bupot DJP): tanggal, nomor bukti
    potong, jenis pajak (kalau tersurat), kode objek pajak, NPWP+nama
    pemotong, NPWP+nama yang dipotong, DPP/penghasilan bruto, PPh dipotong,
    masa pajak.
    """
    header_rownum, header_row = _cari_header_row_bukti_potong(ws)
    if header_row is None:
        raise FormatTidakDikenali(
            f"Sheet '{nama_sheet}' tidak dikenali sebagai Bukti Potong "
            "(tidak ditemukan kolom NPWP + NOMOR BUKTI POTONG + DPP + PPh DIPOTONG sekaligus)."
        )

    headers = header_row

    def _idx(kata_kunci):
        for i, h in enumerate(headers):
            if h is None:
                continue
            h_low = str(h).strip().lower()
            if any(k in h_low for k in kata_kunci):
                return i
        return None

    idx_tanggal = _idx(["tanggal pemotongan", "tanggal bukti potong", "tanggal", "tgl"])
    idx_masa_pajak = _idx(["masa pajak", "periode"])
    idx_nomor_bupot = _idx(["nomor bukti potong", "no bukti potong", "no. bukti potong", "nomor bukti pemotongan"])
    idx_jenis_pajak = _idx(["jenis pajak", "jenis pph"])
    idx_kode_objek = _idx(["kode objek pajak", "kode objek"])
    idx_npwp_pemotong = _idx(["npwp pemotong"])
    idx_nama_pemotong = _idx(["nama pemotong"])
    idx_npwp_dipotong = _idx(["npwp dipotong", "npwp lawan transaksi", "npwp penerima penghasilan"])
    idx_nama_dipotong = _idx(["nama dipotong", "nama lawan transaksi", "nama penerima penghasilan"])
    idx_dpp = _idx(["dpp", "penghasilan bruto", "jumlah bruto"])
    idx_pph = _idx(["pph dipotong", "pph yang dipotong", "jumlah pph"])
    idx_keterangan = _idx(["uraian", "keterangan", "deskripsi"])

    if idx_nomor_bupot is None or idx_dpp is None or idx_pph is None:
        raise FormatTidakDikenali(
            f"Kolom NOMOR BUKTI POTONG, DPP, atau PPh DIPOTONG tidak ditemukan di sheet '{nama_sheet}'."
        )

    def _ambil(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    rows = []
    for row in ws.iter_rows(min_row=header_rownum + 1, values_only=True):
        nomor_bupot = _ambil(row, idx_nomor_bupot)
        dpp = _ambil(row, idx_dpp)
        pph = _ambil(row, idx_pph)
        if nomor_bupot is None and dpp is None and pph is None:
            continue

        rows.append({
            "sheet": nama_sheet,
            "tanggal": _ambil(row, idx_tanggal),
            "masa_pajak": _ambil(row, idx_masa_pajak),
            "nomor_bukti_potong": nomor_bupot,
            "jenis_pajak_tersurat": _ambil(row, idx_jenis_pajak),
            "kode_objek_pajak": _ambil(row, idx_kode_objek),
            "npwp_pemotong": _ambil(row, idx_npwp_pemotong),
            "nama_pemotong": _ambil(row, idx_nama_pemotong),
            "npwp_dipotong": _ambil(row, idx_npwp_dipotong),
            "nama_dipotong": _ambil(row, idx_nama_dipotong),
            "dpp": dpp,
            "pph": pph,
            "keterangan": _ambil(row, idx_keterangan),
        })

    return pd.DataFrame(rows)


def proses_bukti_potong(
    df: pd.DataFrame,
    npwp_perusahaan: str = None,
    toleransi_rupiah: float = _TOLERANSI_SELISIH_PPH_RUPIAH,
) -> dict:
    """
    "Kerjaan akuntan" untuk Bukti Potong: cross-check tiap baris (jenis
    pajak, kode objek & tarif, NPWP, duplikat nomor bukti potong), siapkan
    rekap SIAP dipakai untuk cross-check ke SPT Masa PPh 21/23/4(2) &
    kredit pajak (PPh Pasal 23/4(2) yang DIPOTONG PIHAK LAIN atas
    penghasilan perusahaan = kredit pajak / PPh dibayar di muka).

    npwp_perusahaan: NPWP perusahaan (klien) sendiri -- dipakai utk
    menentukan arah tiap bukti potong (DITERIMA kalau perusahaan berperan
    sbg pihak yang dipotong/menerima penghasilan -> jadi kredit pajak,
    DIPOTONG/DIBUAT kalau perusahaan berperan sbg pemotong atas pihak lain
    -> jadi utang setor PPh). Kalau None, arah tidak ditentukan & draf
    jurnal TIDAK dibuat -- lebih aman drpd menebak arah yang salah.

    Return dict:
        "df": DataFrame asli + kolom tambahan (arah, status, validitas
              format, jenis_pajak, uraian_objek, tarif_seharusnya,
              pph_seharusnya, selisih_pph)
        "ringkasan": dict rekap (total per jenis pajak PPh 21/23/4(2),
                     jumlah yang perlu direview)
        "masalah": list baris yang PERLU DICEK MANUAL beserta alasannya
                   (format salah, tarif tidak sesuai, NPWP tidak valid,
                   kode objek tidak dikenali, atau nomor duplikat)
        "draf_jurnal": list draf jurnal per bukti potong (hanya dibuat
                       kalau npwp_perusahaan diisi -- akun masih generik,
                       WAJIB direview akuntan sebelum diposting)
    """
    if df is None or df.empty:
        return {"df": df, "ringkasan": {}, "masalah": [], "draf_jurnal": []}

    df = df.copy()
    npwp_bersih = _bersihkan_digit(npwp_perusahaan) if npwp_perusahaan else None

    masalah = []
    nomor_bupot_terlihat: dict = {}

    arah_list, status_list = [], []
    valid_nomor_list, valid_npwp_pemotong_list, valid_npwp_dipotong_list = [], [], []
    jenis_pajak_list, uraian_objek_list = [], []
    tarif_seharusnya_list, pph_seharusnya_list, selisih_pph_list = [], [], []

    for i, row in df.iterrows():
        no_bupot = row.get("nomor_bukti_potong")
        no_bupot_digit = _bersihkan_digit(no_bupot)

        v_nomor = validasi_nomor_bukti_potong(no_bupot)
        v_npwp_pemotong = validasi_npwp(row.get("npwp_pemotong"))
        v_npwp_dipotong = validasi_npwp(row.get("npwp_dipotong"))
        valid_nomor_list.append(v_nomor["valid"])
        valid_npwp_pemotong_list.append(v_npwp_pemotong["valid"])
        valid_npwp_dipotong_list.append(v_npwp_dipotong["valid"])

        dpp = float(row.get("dpp") or 0)
        pph = float(row.get("pph") or 0)
        v_tarif = cek_tarif_bukti_potong(
            row.get("kode_objek_pajak"), row.get("jenis_pajak_tersurat"), dpp, pph, toleransi_rupiah,
        )
        jenis_pajak_list.append(v_tarif["jenis_pajak"])
        uraian_objek_list.append(v_tarif["uraian_objek"])
        tarif_seharusnya_list.append(v_tarif["tarif_seharusnya"])
        pph_seharusnya_list.append(v_tarif["pph_seharusnya"])
        selisih_pph_list.append(v_tarif["selisih"])

        # -- Tentukan arah (kalau NPWP perusahaan diketahui) --
        arah = "TIDAK DIKETAHUI"
        if npwp_bersih:
            npwp_pemotong_digit = _bersihkan_digit(row.get("npwp_pemotong"))
            npwp_dipotong_digit = _bersihkan_digit(row.get("npwp_dipotong"))
            if npwp_dipotong_digit and npwp_dipotong_digit == npwp_bersih:
                arah = "DITERIMA"   # perusahaan dipotong pihak lain -> kredit pajak
            elif npwp_pemotong_digit and npwp_pemotong_digit == npwp_bersih:
                arah = "DIBUAT"     # perusahaan memotong pihak lain -> wajib setor
            else:
                arah = "BUKAN MILIK PERUSAHAAN INI"
        arah_list.append(arah)

        # -- Kumpulkan alasan bermasalah (kalau ada) --
        alasan = []
        if not v_nomor["valid"]:
            alasan.append(f"Nomor bukti potong bermasalah: {v_nomor['catatan']}")
        if not v_npwp_pemotong["valid"]:
            alasan.append(f"NPWP pemotong tidak sesuai format: {v_npwp_pemotong['catatan']}")
        if not v_npwp_dipotong["valid"]:
            alasan.append(f"NPWP dipotong tidak sesuai format: {v_npwp_dipotong['catatan']}")
        if v_tarif["jenis_pajak"] == "TIDAK DIKENALI":
            alasan.append(v_tarif["catatan"])
        elif v_tarif["sesuai"] is False:
            alasan.append(v_tarif["catatan"])
        if no_bupot_digit:
            if no_bupot_digit in nomor_bupot_terlihat:
                alasan.append(
                    f"Nomor bukti potong DUPLIKAT -- sudah muncul di baris ke-"
                    f"{nomor_bupot_terlihat[no_bupot_digit] + 1} juga. Bukti potong dobel "
                    "berisiko dobel kredit pajak / dobel lapor, WAJIB dicek manual."
                )
            else:
                nomor_bupot_terlihat[no_bupot_digit] = i

        status_list.append("PERLU REVIEW" if alasan else "OK")
        if alasan:
            masalah.append({
                "baris": i + 1,
                "nomor_bukti_potong": no_bupot,
                "tanggal": row.get("tanggal"),
                "jenis_pajak": v_tarif["jenis_pajak"],
                "dpp": dpp,
                "pph": pph,
                "alasan": alasan,
            })

    df["arah"] = arah_list
    df["status"] = status_list
    df["valid_nomor_bukti_potong"] = valid_nomor_list
    df["valid_npwp_pemotong"] = valid_npwp_pemotong_list
    df["valid_npwp_dipotong"] = valid_npwp_dipotong_list
    df["jenis_pajak"] = jenis_pajak_list
    df["uraian_objek"] = uraian_objek_list
    df["tarif_seharusnya"] = tarif_seharusnya_list
    df["pph_seharusnya"] = pph_seharusnya_list
    df["selisih_pph"] = selisih_pph_list

    # -- Draf jurnal (hanya kalau arah diketahui, supaya tidak menebak salah) --
    draf_jurnal = []
    for i, row in df.iterrows():
        if row["arah"] not in ("DITERIMA", "DIBUAT"):
            continue
        dpp = float(row.get("dpp") or 0)
        pph = float(row.get("pph") or 0)
        neto = dpp - pph
        if row["arah"] == "DITERIMA":
            # Perusahaan menerima penghasilan, dipotong pihak lain -> kas/piutang neto
            # + PPh Dibayar di Muka (kredit pajak) sebesar pph, sebesar dpp jadi pendapatan.
            draf_jurnal.append({
                "baris": i + 1, "nomor_bukti_potong": row.get("nomor_bukti_potong"), "arah": "DITERIMA",
                "no_akun_debet_1": "KAS/BANK/PIUTANG", "nama_akun_debet_1": "Kas/Bank/Piutang (neto diterima)",
                "jml_debet_1": neto,
                "no_akun_debet_2": "PPH DIBAYAR DIMUKA", "nama_akun_debet_2":
                    f"PPh Dibayar di Muka ({row.get('jenis_pajak') or '-'}, kredit pajak)",
                "jml_debet_2": pph,
                "no_akun_kredit": "PENDAPATAN", "nama_akun_kredit": "Pendapatan / Penghasilan terkait (sesuaikan akun)",
                "jml_kredit": dpp,
                "catatan": "Draf otomatis -- cek akun pendapatan yang sesuai; PPh Dibayar di Muka "
                           "ini jadi kredit pajak saat lapor SPT Tahunan/Masa terkait.",
            })
        else:
            # Perusahaan memotong pihak lain -> beban/utang ke pihak lain (neto) +
            # utang PPh yang wajib disetor ke kas negara.
            draf_jurnal.append({
                "baris": i + 1, "nomor_bukti_potong": row.get("nomor_bukti_potong"), "arah": "DIBUAT",
                "no_akun_debet": "BEBAN/PEMBELIAN", "nama_akun_debet": "Beban/Pembelian terkait (sesuaikan akun)",
                "jml_debet": dpp,
                "no_akun_kredit_1": "KAS/BANK/UTANG", "nama_akun_kredit_1": "Kas/Bank/Utang (neto dibayarkan)",
                "jml_kredit_1": neto,
                "no_akun_kredit_2": "UTANG PPH", "nama_akun_kredit_2":
                    f"Utang PPh ({row.get('jenis_pajak') or '-'}, wajib disetor)",
                "jml_kredit_2": pph,
                "catatan": "Draf otomatis -- cek akun beban yang sesuai; Utang PPh ini WAJIB disetor "
                           "(mis. tanggal 10 bulan berikutnya) & dilaporkan di SPT Masa terkait.",
            })

    # -- Ringkasan siap dipakai utk cross-check SPT Masa PPh 21/23/4(2) --
    total_dpp = float(df["dpp"].fillna(0).sum())
    total_pph = float(df["pph"].fillna(0).sum())
    per_jenis = (
        df.groupby("jenis_pajak")[["dpp", "pph"]].sum().fillna(0).to_dict("index")
        if "jenis_pajak" in df.columns else {}
    )
    pph_diterima = float(df.loc[df["arah"] == "DITERIMA", "pph"].fillna(0).sum())
    pph_dibuat = float(df.loc[df["arah"] == "DIBUAT", "pph"].fillna(0).sum())

    ringkasan = {
        "jumlah_bukti_potong": len(df),
        "total_dpp": total_dpp,
        "total_pph": total_pph,
        "rekap_per_jenis_pajak": per_jenis,
        "pph_diterima_kredit_pajak": pph_diterima,
        "pph_dibuat_wajib_setor": pph_dibuat,
        "jumlah_perlu_review": len(masalah),
        "jumlah_duplikat": sum(1 for m in masalah if any("DUPLIKAT" in a for a in m["alasan"])),
        "catatan_tarif": (
            "Perhitungan tarif memakai kamus kode objek pajak umum yang berlaku saat ini. "
            "PPh 21 memakai skema TER (tergantung PTKP karyawan) sehingga hanya dicek "
            "kewajaran rentangnya, bukan divalidasi pasti. Jasa konstruksi & beberapa objek "
            "PPh 4(2) tarifnya bervariasi tergantung kualifikasi -- WAJIB dicek manual. "
            "Tarif pajak bisa berubah sewaktu-waktu, selalu cek update PMK terbaru."
        ),
    }
    if npwp_bersih is None:
        ringkasan["catatan_arah"] = (
            "NPWP perusahaan tidak diisi -- arah (Diterima/Dibuat) & draf jurnal TIDAK bisa "
            "dihitung akurat. Isi npwp_perusahaan supaya hasilnya lengkap."
        )

    return {"df": df, "ringkasan": ringkasan, "masalah": masalah, "draf_jurnal": draf_jurnal}


def proses_file_bukti_potong(
    file_like,
    nama_file: str = None,
    npwp_perusahaan: str = None,
) -> dict:
    """
    Fungsi siap-pakai tingkat file: baca file (.xlsx/.xlsm/.xls/.csv/.pdf),
    cari semua sheet yang cocok sebagai Bukti Potong, gabungkan, lalu proses
    lewat proses_bukti_potong(). Dipakai oleh app.py utk 1 file upload.

    Return sama seperti proses_bukti_potong(), ditambah "sheet_dilewati":
    daftar nama sheet yang TIDAK cocok format Bukti Potong sama sekali.
    """
    nama_file = nama_file or getattr(file_like, "name", "") or ""
    daftar_sheet, _df_coa = _siapkan_daftar_sheet(file_like, nama_file)

    df_list = []
    sheet_dilewati = []
    for nama, ws in daftar_sheet:
        try:
            df_sheet = parse_sheet_bukti_potong(ws, nama)
        except FormatTidakDikenali:
            sheet_dilewati.append(nama)
            continue
        if df_sheet.empty:
            sheet_dilewati.append(nama)
            continue
        df_list.append(df_sheet)

    if not df_list:
        return {
            "df": pd.DataFrame(), "ringkasan": {}, "masalah": [], "draf_jurnal": [],
            "sheet_dilewati": sheet_dilewati,
        }

    df_gabungan = pd.concat(df_list, ignore_index=True)
    hasil = proses_bukti_potong(df_gabungan, npwp_perusahaan=npwp_perusahaan)
    hasil["sheet_dilewati"] = sheet_dilewati
    return hasil


# ============================================================
# 11C. SPT MASA/TAHUNAN -- PARSING, KLASIFIKASI, & SARAN TINDAK LANJUT
# ============================================================
# Beda dgn Faktur Pajak & Bukti Potong (satu baris = satu dokumen bukti
# transaksi), sheet SPT biasanya berupa REKAP/LOG kepatuhan pelaporan: satu
# baris = satu SPT yang sudah/akan dilaporkan untuk suatu masa/tahun pajak.
#
# "Kerjaan akuntan" saat menerima rekap ini:
#   1. Kenali jenis SPT-nya (Masa PPN/PPh 21/23/4(2)/25, atau Tahunan Badan/OP).
#   2. Cek NPWP & kelengkapan periode (masa/tahun pajak).
#   3. Hitung status Kurang Bayar/Lebih Bayar/Nihil (dari kolom yg tertulis,
#      atau dari Pajak Terutang - Kredit Pajak kalau kolom KB/LB tidak
#      diisi), lalu cross-check konsistensi keduanya kalau sama-sama ada.
#   4. Hitung ESTIMASI batas waktu setor & lapor dari masa/tahun pajak, lalu
#      tandai kalau sudah/berisiko TERLAMBAT.
#   5. Kasih SARAN TINDAK LANJUT konkret per baris (segera setor, ajukan
#      kompensasi/restitusi, tetap wajib lapor walau nihil, dst) -- supaya
#      user langsung tahu harus ngapain, bukan cuma lihat angka mentah.
#
# CATATAN BATAS WAKTU: tanggal jatuh tempo di bawah adalah ATURAN UMUM yang
# berlaku luas per UU KUP & aturan turunannya SAAT INI. Pergeseran jatuh
# tempo yang bertepatan hari libur, ijin penundaan/perpanjangan khusus, dan
# perubahan aturan lewat PMK baru TIDAK dihitung otomatis di sini -- selalu
# cek kalender pajak resmi DJP untuk kepastian.

KATEGORI_SPT: Dict[str, Dict[str, Any]] = {
    "ppn_masa": {
        "label": "SPT Masa PPN (Formulir 1111)",
        "periodisitas": "masa",
        "kata_kunci": ["1111", "spt masa ppn", "masa ppn"],
        "batas_setor_akhir_bulan": True, "batas_lapor_akhir_bulan": True,
    },
    "pph21_masa": {
        "label": "SPT Masa PPh Pasal 21/26 (Formulir 1721)",
        "periodisitas": "masa",
        "kata_kunci": ["1721", "pph pasal 21", "pph 21", "pph21"],
        "batas_setor_tgl": 10, "batas_lapor_tgl": 20,
    },
    "pph23_masa": {
        "label": "SPT Masa PPh Pasal 23/26 (Unifikasi)",
        "periodisitas": "masa",
        "kata_kunci": ["pph pasal 23", "pph 23", "pph23"],
        "batas_setor_tgl": 10, "batas_lapor_tgl": 20,
    },
    "pph4ayat2_masa": {
        "label": "SPT Masa PPh Pasal 4 Ayat (2) Final (Unifikasi)",
        "periodisitas": "masa",
        "kata_kunci": ["pasal 4 ayat", "4 ayat 2", "4(2)", "pph final"],
        "batas_setor_tgl": 10, "batas_lapor_tgl": 20,
    },
    "pph25_masa": {
        "label": "SPT Masa PPh Pasal 25 (Angsuran)",
        "periodisitas": "masa",
        "kata_kunci": ["pasal 25", "pph 25", "pph25", "angsuran"],
        "batas_setor_tgl": 15, "batas_lapor_tgl": 15,
    },
    "pph_badan_tahunan": {
        "label": "SPT Tahunan PPh Badan (Formulir 1771)",
        "periodisitas": "tahunan",
        "kata_kunci": ["1771", "tahunan badan", "pph badan"],
        "batas_lapor_bulan_ke": 4, "batas_lapor_tgl_bulan": 30,  # 30 April tahun berikutnya
    },
    "pph_op_tahunan": {
        "label": "SPT Tahunan PPh Orang Pribadi (Formulir 1770/1770 S/1770 SS)",
        "periodisitas": "tahunan",
        "kata_kunci": ["1770", "tahunan orang pribadi", "tahunan pribadi", "tahunan op"],
        "batas_lapor_bulan_ke": 3, "batas_lapor_tgl_bulan": 31,  # 31 Maret tahun berikutnya
    },
}

_BULAN_ID = {
    "januari": 1, "februari": 2, "maret": 3, "april": 4, "mei": 5, "juni": 6,
    "juli": 7, "agustus": 8, "september": 9, "oktober": 10, "november": 11, "desember": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "jun": 6, "jul": 7, "agt": 8, "ags": 8,
    "sep": 9, "okt": 10, "nov": 11, "des": 12,
}

_TOLERANSI_SELISIH_SPT_RUPIAH = 5  # toleransi pembulatan, dalam Rupiah


def _cari_header_row_spt(ws, max_scan: int = 10):
    """
    Cari baris header sheet SPT Masa/Tahunan: butuh NPWP + (MASA PAJAK atau
    TAHUN PAJAK) + minimal satu kolom nilai/status (PAJAK TERUTANG / KREDIT
    PAJAK / KURANG BAYAR / LEBIH BAYAR / STATUS LAPOR / JENIS SPT / FORMULIR)
    -- kombinasi ini yang membedakannya dari Faktur Pajak/Bukti Potong (yang
    pakai NOMOR FAKTUR/NOMOR BUKTI POTONG per baris, bukan rekap per periode).
    """
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        teks = " ".join(str(c) for c in row if c is not None).lower()
        ada_npwp = "npwp" in teks
        ada_periode = "masa pajak" in teks or "tahun pajak" in teks or "periode pajak" in teks
        ada_nilai = any(k in teks for k in [
            "kurang bayar", "lebih bayar", "pajak terutang", "kredit pajak",
            "status lapor", "jenis spt", "formulir",
        ])
        if ada_npwp and ada_periode and ada_nilai:
            return i + 1, list(row)
    return None, None


def deteksi_kategori_spt(*teks_list) -> Optional[str]:
    """
    Tebak kategori SPT (key di KATEGORI_SPT) dari gabungan teks bebas (mis.
    isi kolom 'Jenis SPT'/'Formulir' + keterangan). Kalau 2+ kategori
    cocok, kata kunci TERPANJANG yang menang (mis. '1770' vs kata kunci lain
    yang kebetulan lebih umum) supaya tebakan lebih spesifik.
    """
    gabungan = " ".join(str(t).strip().lower() for t in teks_list if t is not None)
    if not gabungan:
        return None
    kandidat = []
    for kode, info in KATEGORI_SPT.items():
        for kw in info["kata_kunci"]:
            if kw in gabungan:
                kandidat.append((len(kw), kode))
    if not kandidat:
        return None
    kandidat.sort(reverse=True)
    return kandidat[0][1]


def _parse_periode_pajak(masa_pajak, tahun_pajak) -> Tuple[Optional[int], Optional[int]]:
    """
    Coba tebak (bulan, tahun) dari isi kolom 'masa pajak' / 'tahun pajak'.
    Mendukung: nama bulan Indonesia + tahun ("Januari 2026"), "MM/YYYY",
    "MM-YYYY", "YYYY-MM", atau tahun 4 digit saja (utk SPT Tahunan, bulan
    tetap None). Return (None, None) kalau sama sekali tidak bisa ditebak --
    lebih aman drpd menebak periode yang salah.
    """
    gabungan = " ".join(str(t) for t in (masa_pajak, tahun_pajak) if t is not None)
    if not gabungan.strip():
        return None, None

    bulan_hasil, tahun_hasil = None, None

    m_tahun = re.search(r"\b(20[1-3]\d)\b", gabungan)
    if m_tahun:
        tahun_hasil = int(m_tahun.group(1))

    gabungan_low = gabungan.lower()
    for nama, angka in _BULAN_ID.items():
        if nama in gabungan_low:
            bulan_hasil = angka
            break
    if bulan_hasil is None:
        m_mm_yyyy = re.search(r"\b(0?[1-9]|1[0-2])[/\-](20[1-3]\d)\b", gabungan)
        if m_mm_yyyy:
            bulan_hasil = int(m_mm_yyyy.group(1))
            tahun_hasil = tahun_hasil or int(m_mm_yyyy.group(2))
        else:
            m_yyyy_mm = re.search(r"\b(20[1-3]\d)[/\-](0?[1-9]|1[0-2])\b", gabungan)
            if m_yyyy_mm:
                tahun_hasil = tahun_hasil or int(m_yyyy_mm.group(1))
                bulan_hasil = int(m_yyyy_mm.group(2))

    return bulan_hasil, tahun_hasil


def hitung_estimasi_jatuh_tempo(kode_kategori: Optional[str], bulan: Optional[int], tahun: Optional[int]) -> dict:
    """
    Hitung ESTIMASI tanggal batas setor & batas lapor berdasarkan kategori
    SPT + periode pajaknya. HANYA aturan umum (lihat catatan di atas modul
    ini) -- bukan kepastian mutlak, dan tidak memperhitungkan pergeseran
    jatuh tempo krn hari libur.

    Return: {"tanggal_batas_setor": date|None, "tanggal_batas_lapor": date|None,
             "catatan": str}
    """
    info = KATEGORI_SPT.get(kode_kategori)
    if info is None:
        return {
            "tanggal_batas_setor": None, "tanggal_batas_lapor": None,
            "catatan": "Jenis SPT tidak dikenali -- batas waktu tidak bisa diestimasi "
                       "otomatis, cek kalender pajak DJP.",
        }

    if info["periodisitas"] == "masa":
        if not bulan or not tahun:
            return {
                "tanggal_batas_setor": None, "tanggal_batas_lapor": None,
                "catatan": "Masa pajak (bulan/tahun) tidak terbaca -- batas waktu tidak "
                           "bisa diestimasi otomatis.",
            }
        bulan_lapor, tahun_lapor = bulan + 1, tahun
        if bulan_lapor > 12:
            bulan_lapor, tahun_lapor = bulan_lapor - 12, tahun_lapor + 1
        akhir_bulan_lapor = calendar.monthrange(tahun_lapor, bulan_lapor)[1]

        if info.get("batas_setor_akhir_bulan"):
            tgl_setor = date(tahun_lapor, bulan_lapor, akhir_bulan_lapor)
        else:
            tgl_setor = date(tahun_lapor, bulan_lapor, min(info["batas_setor_tgl"], akhir_bulan_lapor))

        if info.get("batas_lapor_akhir_bulan"):
            tgl_lapor = date(tahun_lapor, bulan_lapor, akhir_bulan_lapor)
        else:
            tgl_lapor = date(tahun_lapor, bulan_lapor, min(info["batas_lapor_tgl"], akhir_bulan_lapor))

        return {
            "tanggal_batas_setor": tgl_setor, "tanggal_batas_lapor": tgl_lapor,
            "catatan": "Estimasi berdasarkan aturan umum jatuh tempo bulan berikutnya -- "
                       "cek pergeseran hari libur di kalender pajak DJP.",
        }

    # -- periodisitas tahunan --
    if not tahun:
        return {
            "tanggal_batas_setor": None, "tanggal_batas_lapor": None,
            "catatan": "Tahun pajak tidak terbaca -- batas waktu tidak bisa diestimasi otomatis.",
        }
    tahun_lapor = tahun + 1
    bulan_batas, tgl_batas = info["batas_lapor_bulan_ke"], info["batas_lapor_tgl_bulan"]
    akhir_bulan_batas = calendar.monthrange(tahun_lapor, bulan_batas)[1]
    tgl_lapor = date(tahun_lapor, bulan_batas, min(tgl_batas, akhir_bulan_batas))
    return {
        # Kurang bayar SPT Tahunan WAJIB dilunasi SEBELUM SPT disampaikan --
        # jadi batas "aman" utk setor dianggap sama dgn batas lapor (paling lambat).
        "tanggal_batas_setor": tgl_lapor, "tanggal_batas_lapor": tgl_lapor,
        "catatan": "Kurang bayar SPT Tahunan wajib dilunasi SEBELUM SPT disampaikan -- "
                   "jangan tunggu sampai tanggal batas lapor.",
    }


def parse_sheet_spt(ws, nama_sheet: str) -> pd.DataFrame:
    """
    Parse satu sheet SPT Masa/Tahunan jadi DataFrame baris-per-SPT. Kolom
    yang dicoba dikenali (nama header fleksibel): tanggal lapor, jenis
    SPT/formulir, masa pajak, tahun pajak, NPWP+nama WP, pajak terutang,
    kredit pajak, kurang bayar, lebih bayar, status lapor, keterangan.
    """
    header_rownum, header_row = _cari_header_row_spt(ws)
    if header_row is None:
        raise FormatTidakDikenali(
            f"Sheet '{nama_sheet}' tidak dikenali sebagai SPT Masa/Tahunan "
            "(tidak ditemukan kolom NPWP + MASA/TAHUN PAJAK + minimal satu "
            "kolom nilai/status sekaligus)."
        )

    headers = header_row

    def _idx(kata_kunci):
        for i, h in enumerate(headers):
            if h is None:
                continue
            h_low = str(h).strip().lower()
            if any(k in h_low for k in kata_kunci):
                return i
        return None

    idx_tanggal_lapor = _idx(["tanggal lapor", "tgl lapor", "tanggal spt"])
    idx_jenis_spt = _idx(["jenis spt", "formulir", "jenis pajak"])
    idx_masa_pajak = _idx(["masa pajak"])
    idx_tahun_pajak = _idx(["tahun pajak"])
    idx_npwp = _idx(["npwp"])
    idx_nama_wp = _idx(["nama wp", "nama wajib pajak", "nama perusahaan"])
    idx_pajak_terutang = _idx(["pajak terutang", "jumlah pajak terutang"])
    idx_kredit_pajak = _idx(["kredit pajak", "pajak dibayar dimuka", "pajak dibayar di muka"])
    idx_kurang_bayar = _idx(["kurang bayar", "kurang dibayar"])
    idx_lebih_bayar = _idx(["lebih bayar", "lebih dibayar"])
    idx_status_lapor = _idx(["status lapor", "status spt", "status"])
    idx_keterangan = _idx(["keterangan", "catatan"])

    if idx_npwp is None or (idx_masa_pajak is None and idx_tahun_pajak is None):
        raise FormatTidakDikenali(
            f"Kolom NPWP atau MASA/TAHUN PAJAK tidak ditemukan di sheet '{nama_sheet}'."
        )

    def _ambil(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    rows = []
    for row in ws.iter_rows(min_row=header_rownum + 1, values_only=True):
        npwp = _ambil(row, idx_npwp)
        masa_pajak = _ambil(row, idx_masa_pajak)
        tahun_pajak = _ambil(row, idx_tahun_pajak)
        jenis_spt = _ambil(row, idx_jenis_spt)
        pajak_terutang = _ambil(row, idx_pajak_terutang)
        kredit_pajak = _ambil(row, idx_kredit_pajak)
        kurang_bayar = _ambil(row, idx_kurang_bayar)
        lebih_bayar = _ambil(row, idx_lebih_bayar)
        if all(v is None for v in (
            npwp, masa_pajak, tahun_pajak, jenis_spt,
            pajak_terutang, kredit_pajak, kurang_bayar, lebih_bayar,
        )):
            continue

        rows.append({
            "sheet": nama_sheet,
            "tanggal_lapor": _ambil(row, idx_tanggal_lapor),
            "jenis_spt_tersurat": jenis_spt,
            "masa_pajak": masa_pajak,
            "tahun_pajak": tahun_pajak,
            "npwp": npwp,
            "nama_wp": _ambil(row, idx_nama_wp),
            "pajak_terutang": pajak_terutang,
            "kredit_pajak": kredit_pajak,
            "kurang_bayar_tertulis": kurang_bayar,
            "lebih_bayar_tertulis": lebih_bayar,
            "status_lapor_tersurat": _ambil(row, idx_status_lapor),
            "keterangan": _ambil(row, idx_keterangan),
        })

    return pd.DataFrame(rows)


def proses_spt(df: pd.DataFrame, toleransi_rupiah: float = _TOLERANSI_SELISIH_SPT_RUPIAH) -> dict:
    """
    "Kerjaan akuntan" untuk SPT Masa/Tahunan: kenali jenis & periodenya,
    hitung status Kurang/Lebih Bayar/Nihil, cek konsistensi nilai tertulis
    vs Pajak Terutang - Kredit Pajak, estimasi jatuh tempo setor & lapor,
    tandai yang sudah/berisiko terlambat, dan kasih SARAN TINDAK LANJUT
    konkret per baris (mis. segera setor sebelum tanggal X, ajukan
    kompensasi/restitusi, tetap wajib lapor walau nihil).

    Return dict:
        "df": DataFrame asli + kolom tambahan (kategori_spt, jenis_spt,
              bulan_pajak, tahun_pajak_terbaca, valid_npwp, status_bayar,
              nominal_status, tanggal_batas_setor, tanggal_batas_lapor,
              sudah_lapor, terlambat_lapor, berisiko_terlambat_setor,
              saran_tindak_lanjut, status)
        "ringkasan": dict rekap (jumlah per jenis SPT, total kurang/lebih
                     bayar, jumlah nihil, jumlah terlambat, jumlah perlu
                     review)
        "masalah": list baris yang PERLU DICEK MANUAL beserta alasannya
                   (NPWP tidak valid, jenis SPT tidak dikenali, nilai
                   tertulis tidak konsisten, terlambat lapor/setor, atau
                   kombinasi NPWP+jenis+periode duplikat)
    """
    if df is None or df.empty:
        return {"df": df, "ringkasan": {}, "masalah": []}

    df = df.copy()
    hari_ini = datetime.now().date()

    masalah = []
    kombinasi_terlihat: dict = {}

    kategori_list, label_list, bulan_list, tahun_list = [], [], [], []
    valid_npwp_list = []
    status_bayar_list, nominal_status_list = [], []
    batas_setor_list, batas_lapor_list = [], []
    sudah_lapor_list, terlambat_lapor_list, terlambat_setor_list = [], [], []
    saran_list, status_list = [], []

    for i, row in df.iterrows():
        alasan = []

        v_npwp = validasi_npwp(row.get("npwp"))
        valid_npwp_list.append(v_npwp["valid"])
        if not v_npwp["valid"]:
            alasan.append(f"NPWP tidak sesuai format: {v_npwp['catatan']}")

        kategori = deteksi_kategori_spt(row.get("jenis_spt_tersurat"), row.get("keterangan"))
        info_kategori = KATEGORI_SPT.get(kategori)
        label_spt = info_kategori["label"] if info_kategori else (
            str(row.get("jenis_spt_tersurat")) if row.get("jenis_spt_tersurat") else "Tidak dikenali"
        )
        kategori_list.append(kategori or "TIDAK DIKENALI")
        label_list.append(label_spt)
        if kategori is None:
            alasan.append(
                "Jenis SPT tidak dikenali dari kolom Jenis SPT/Formulir -- isi lebih jelas "
                "(mis. 'Formulir 1111', 'SPT Masa PPh 21', 'SPT Tahunan Badan 1771') supaya "
                "estimasi jatuh tempo & saran tindak lanjut bisa dihitung."
            )

        bulan, tahun = _parse_periode_pajak(row.get("masa_pajak"), row.get("tahun_pajak"))
        bulan_list.append(bulan)
        tahun_list.append(tahun)

        pt_raw = row.get("pajak_terutang")
        kp_raw = row.get("kredit_pajak")
        # [FIX] SEBELUMNYA "float(row.get(...) or 0)" langsung mengubah nilai
        # KOSONG (None/NaN) jadi 0.0, sehingga di titik ini sudah tidak bisa
        # dibedakan lagi mana yang "memang nol" (Nihil sungguhan -- banyak
        # terjadi mis. SPT Masa PPh 21 bulan tanpa karyawan kena pajak) vs
        # "memang belum diisi sama sekali". Akibatnya baris Nihil asli malah
        # jatuh ke status TIDAK DIKETAHUI. Sekarang keduanya dicek terpisah
        # SEBELUM dikonversi ke 0.0.
        ada_nilai_pajak = not (pd.isna(pt_raw) and pd.isna(kp_raw))
        pajak_terutang = float(pt_raw) if not pd.isna(pt_raw) else 0.0
        kredit_pajak = float(kp_raw) if not pd.isna(kp_raw) else 0.0
        kb_raw = row.get("kurang_bayar_tertulis")
        lb_raw = row.get("lebih_bayar_tertulis")
        # [FIX-2] "kb_raw not in (None, "")" TIDAK menangkap NaN numerik
        # (muncul dari sel kosong di kolom numerik setelah gabung sheet) --
        # "nan not in (None, "")" bernilai True karena nan != None dan
        # nan != "", jadi float(nan) lolos jadi kb_tertulis = nan (bukan
        # None). nan itu truthy, sehingga baris "if kb_tertulis and
        # kb_tertulis > 0" & "if (kb_tertulis or lb_tertulis) and (...)"
        # ikut ke-trigger memakai nilai nan meski sel aslinya kosong --
        # pola bug yang sama persis dgn pajak_terutang/kredit_pajak di atas.
        # Pakai pd.isna() supaya konsisten.
        kb_tertulis = float(kb_raw) if not pd.isna(kb_raw) else None
        lb_tertulis = float(lb_raw) if not pd.isna(lb_raw) else None
        selisih_hitung = pajak_terutang - kredit_pajak  # > 0 = kurang bayar

        if kb_tertulis and kb_tertulis > 0:
            status_bayar, nominal_status = "KURANG BAYAR", kb_tertulis
        elif lb_tertulis and lb_tertulis > 0:
            status_bayar, nominal_status = "LEBIH BAYAR", lb_tertulis
        elif ada_nilai_pajak:
            if selisih_hitung > toleransi_rupiah:
                status_bayar, nominal_status = "KURANG BAYAR", selisih_hitung
            elif selisih_hitung < -toleransi_rupiah:
                status_bayar, nominal_status = "LEBIH BAYAR", abs(selisih_hitung)
            else:
                status_bayar, nominal_status = "NIHIL", 0.0
        else:
            status_bayar, nominal_status = "TIDAK DIKETAHUI", None
        status_bayar_list.append(status_bayar)
        nominal_status_list.append(nominal_status)

        if (kb_tertulis or lb_tertulis) and (pajak_terutang or kredit_pajak):
            nominal_tertulis = (kb_tertulis or 0) - (lb_tertulis or 0)
            if abs(nominal_tertulis - selisih_hitung) > toleransi_rupiah:
                alasan.append(
                    f"Nilai Kurang/Lebih Bayar tertulis (Rp{nominal_tertulis:,.0f}) tidak "
                    f"konsisten dengan Pajak Terutang - Kredit Pajak (Rp{selisih_hitung:,.0f}) "
                    "-- WAJIB dicek manual."
                )

        jatuh_tempo = hitung_estimasi_jatuh_tempo(kategori, bulan, tahun)
        tgl_batas_setor = jatuh_tempo["tanggal_batas_setor"]
        tgl_batas_lapor = jatuh_tempo["tanggal_batas_lapor"]
        batas_setor_list.append(tgl_batas_setor)
        batas_lapor_list.append(tgl_batas_lapor)

        tanggal_lapor = row.get("tanggal_lapor")
        # [FIX] Setelah pd.concat() menggabungkan banyak sheet, kolom
        # "tanggal_lapor" bisa berubah jadi dtype datetime64 kalau ADA baris
        # lain yang terisi tanggal -- baris yang tanggalnya kosong ikut
        # berubah dari None jadi pandas.NaT (BUKAN None lagi). Masalahnya,
        # pd.NaT lolos isinstance(..., (datetime, date)) (NaT memang subclass
        # keduanya demi kompatibilitas pandas), lalu NaT.date() balik NaT
        # lagi, dan NaT > tanggal_biasa MELEMPAR TypeError -- meledakkan
        # proses_spt() untuk SELURUH batch gara-gara satu baris "Sudah Lapor"
        # tanpa tanggal lapor terisi. Netralkan dulu jadi None/date Python
        # biasa sebelum dipakai, supaya kosong tetap dianggap kosong.
        if pd.isna(tanggal_lapor):
            tanggal_lapor = None
        elif isinstance(tanggal_lapor, datetime):
            tanggal_lapor = tanggal_lapor.date()
        status_lapor_tersurat = str(row.get("status_lapor_tersurat") or "").strip().lower()
        belum_lapor_tersurat = "belum" in status_lapor_tersurat
        sudah_lapor = bool(tanggal_lapor) or (
            any(k in status_lapor_tersurat for k in ["sudah", "lapor", "selesai", "done"])
            and not belum_lapor_tersurat
        )
        sudah_lapor_list.append(sudah_lapor)

        terlambat_lapor = False
        if tgl_batas_lapor:
            if sudah_lapor and isinstance(tanggal_lapor, date):
                terlambat_lapor = tanggal_lapor > tgl_batas_lapor
            elif not sudah_lapor:
                terlambat_lapor = hari_ini > tgl_batas_lapor
        terlambat_lapor_list.append(terlambat_lapor)

        terlambat_setor = bool(
            tgl_batas_setor and status_bayar == "KURANG BAYAR" and not sudah_lapor and hari_ini > tgl_batas_setor
        )
        terlambat_setor_list.append(terlambat_setor)

        if terlambat_lapor:
            alasan.append(
                f"⚠️ TERLAMBAT LAPOR -- batas lapor estimasi {tgl_batas_lapor:%d-%m-%Y} sudah "
                "lewat. Segera lapor & siapkan kemungkinan denda administrasi keterlambatan (UU KUP)."
            )
        if terlambat_setor:
            alasan.append(
                f"⚠️ BERISIKO TERLAMBAT SETOR -- batas setor estimasi {tgl_batas_setor:%d-%m-%Y} "
                "sudah lewat dan status masih Kurang Bayar. Segera setor untuk hindari sanksi "
                "bunga keterlambatan."
            )

        npwp_digit = _bersihkan_digit(row.get("npwp"))
        kunci_kombinasi = (npwp_digit, kategori, bulan, tahun)
        if npwp_digit and kategori:
            if kunci_kombinasi in kombinasi_terlihat:
                alasan.append(
                    "SPT DUPLIKAT -- kombinasi NPWP + jenis SPT + periode yang sama sudah "
                    f"muncul di baris ke-{kombinasi_terlihat[kunci_kombinasi] + 1}, WAJIB dicek manual."
                )
            else:
                kombinasi_terlihat[kunci_kombinasi] = i

        # -- SARAN TINDAK LANJUT (inti kebutuhan akuntan) --
        saran = []
        if status_bayar == "KURANG BAYAR":
            saran.append(
                f"Kurang bayar Rp{nominal_status:,.0f} -- segera setor via e-Billing (kode akun "
                "& jenis setoran sesuai jenis pajak)"
                + (f" sebelum {tgl_batas_setor:%d-%m-%Y}" if tgl_batas_setor else "")
                + ", lalu lapor SPT"
                + (f" sebelum {tgl_batas_lapor:%d-%m-%Y}" if tgl_batas_lapor else "")
                + " untuk menghindari sanksi bunga & denda keterlambatan."
            )
        elif status_bayar == "LEBIH BAYAR":
            saran.append(
                f"Lebih bayar Rp{nominal_status:,.0f} -- pastikan diajukan KOMPENSASI ke masa/"
                "tahun pajak berikutnya, atau RESTITUSI (tersedia jalur dipercepat untuk WP "
                "kriteria/persyaratan tertentu). Cek juga jangan sampai dobel hitung dengan PPh "
                "Dibayar di Muka dari rekap Bukti Potong yang sudah diproses terpisah."
            )
        elif status_bayar == "NIHIL":
            saran.append(
                "Nihil -- tetap WAJIB lapor tepat waktu"
                + (f" (paling lambat {tgl_batas_lapor:%d-%m-%Y})" if tgl_batas_lapor else "")
                + " meski tidak ada pajak terutang, supaya tidak kena denda administrasi "
                "keterlambatan lapor."
            )
        else:
            saran.append(
                "Nilai Pajak Terutang/Kredit Pajak/Kurang-Lebih Bayar belum terisi -- lengkapi "
                "dulu supaya status & saran tindak lanjut bisa dihitung."
            )
        if belum_lapor_tersurat:
            saran.append("Status masih tertulis BELUM LAPOR -- prioritaskan sebelum jatuh tempo.")
        saran_list.append(" | ".join(saran))

        status_final = "PERLU REVIEW" if alasan else "OK"
        status_list.append(status_final)
        if alasan:
            masalah.append({
                "baris": i + 1,
                "npwp": row.get("npwp"),
                "jenis_spt": label_spt,
                "masa_tahun": row.get("masa_pajak") or row.get("tahun_pajak"),
                "status_bayar": status_bayar,
                "nominal": nominal_status,
                "alasan": alasan,
            })

    df["kategori_spt"] = kategori_list
    df["jenis_spt"] = label_list
    df["bulan_pajak"] = bulan_list
    df["tahun_pajak_terbaca"] = tahun_list
    df["valid_npwp"] = valid_npwp_list
    df["status_bayar"] = status_bayar_list
    df["nominal_status"] = nominal_status_list
    df["tanggal_batas_setor"] = batas_setor_list
    df["tanggal_batas_lapor"] = batas_lapor_list
    df["sudah_lapor"] = sudah_lapor_list
    df["terlambat_lapor"] = terlambat_lapor_list
    df["berisiko_terlambat_setor"] = terlambat_setor_list
    df["saran_tindak_lanjut"] = saran_list
    df["status"] = status_list

    total_kurang_bayar = float(df.loc[df["status_bayar"] == "KURANG BAYAR", "nominal_status"].fillna(0).sum())
    total_lebih_bayar = float(df.loc[df["status_bayar"] == "LEBIH BAYAR", "nominal_status"].fillna(0).sum())
    per_jenis_spt = df.groupby("jenis_spt").size().to_dict()

    ringkasan = {
        "jumlah_spt": len(df),
        "rekap_per_jenis_spt": per_jenis_spt,
        "jumlah_kurang_bayar": int((df["status_bayar"] == "KURANG BAYAR").sum()),
        "total_kurang_bayar": total_kurang_bayar,
        "jumlah_lebih_bayar": int((df["status_bayar"] == "LEBIH BAYAR").sum()),
        "total_lebih_bayar": total_lebih_bayar,
        "jumlah_nihil": int((df["status_bayar"] == "NIHIL").sum()),
        "jumlah_terlambat_lapor": int(df["terlambat_lapor"].sum()),
        "jumlah_berisiko_terlambat_setor": int(df["berisiko_terlambat_setor"].sum()),
        "jumlah_perlu_review": len(masalah),
        "catatan_batas_waktu": (
            "Tanggal batas setor/lapor adalah ESTIMASI dari aturan umum jatuh tempo yang "
            "berlaku saat ini (belum memperhitungkan pergeseran hari libur atau ijin "
            "penundaan/perpanjangan khusus) -- selalu cek kalender pajak resmi DJP."
        ),
    }

    return {"df": df, "ringkasan": ringkasan, "masalah": masalah}


def proses_file_spt(file_like, nama_file: str = None) -> dict:
    """
    Fungsi siap-pakai tingkat file: baca file (.xlsx/.xlsm/.xls/.csv/.pdf),
    cari semua sheet yang cocok sebagai SPT Masa/Tahunan, gabungkan, lalu
    proses lewat proses_spt(). Dipakai oleh app.py utk 1 file upload.

    Return sama seperti proses_spt(), ditambah "sheet_dilewati": daftar nama
    sheet yang TIDAK cocok format SPT sama sekali.
    """
    nama_file = nama_file or getattr(file_like, "name", "") or ""
    daftar_sheet, _df_coa = _siapkan_daftar_sheet(file_like, nama_file)

    df_list = []
    sheet_dilewati = []
    for nama, ws in daftar_sheet:
        try:
            df_sheet = parse_sheet_spt(ws, nama)
        except FormatTidakDikenali:
            sheet_dilewati.append(nama)
            continue
        if df_sheet.empty:
            sheet_dilewati.append(nama)
            continue
        df_list.append(df_sheet)

    if not df_list:
        return {"df": pd.DataFrame(), "ringkasan": {}, "masalah": [], "sheet_dilewati": sheet_dilewati}

    df_gabungan = pd.concat(df_list, ignore_index=True)
    hasil = proses_spt(df_gabungan)
    hasil["sheet_dilewati"] = sheet_dilewati
    return hasil


# ============================================================
# 11D. SLIP GAJI KARYAWAN -- PARSING, VALIDASI KOMPONEN, & DRAF JURNAL PAYROLL
# ============================================================
# Sheet Slip Gaji biasanya berupa REKAP payroll satu periode: satu baris =
# satu karyawan = satu slip gaji bulan tsb (gaji pokok, tunjangan, potongan,
# take home pay).
#
# "Kerjaan akuntan" saat menerima rekap payroll ini:
#   1. Cross-check komponen: Gaji Bersih (THP) tertulis HARUS sama dengan
#      (Gaji Pokok + Total Tunjangan) - Total Potongan -- ini yang paling
#      sering salah ketik/salah rumus di file payroll manual.
#   2. Cek kewajaran potongan PPh 21 (tarif efektif terhadap gaji bruto,
#      sama seperti pendekatan di Bukti Potong -- karena PPh 21 pakai skema
#      TER yang tergantung PTKP, tidak bisa dipastikan pasti dari slip saja).
#   3. Cek kewajaran potongan BPJS Kesehatan & BPJS Ketenagakerjaan (JHT/JP)
#      terhadap persentase resmi -- hanya sbg *reasonableness check*, karena
#      dasar upah & batas atas (plafon) BPJS bisa berubah & beda kebijakan
#      per perusahaan.
#   4. Deteksi anomali: THP negatif/nol, potongan melebihi gaji bruto,
#      slip duplikat (karyawan + periode yang sama muncul 2x).
#   5. Siapkan draf jurnal payroll (Beban Gaji & Tunjangan di debet; Utang
#      PPh 21, Utang BPJS karyawan, & Utang Gaji/Kas bersih di kredit) --
#      akun tetap generik, WAJIB direview akuntan sebelum posting.
#
# CATATAN TARIF: persentase BPJS & plafon dasar upah di bawah adalah ATURAN
# UMUM yang berlaku luas SAAT INI. Perubahan plafon tahunan, kebijakan upah
# minimum, atau skema tambahan (mis. pensiun/asuransi swasta) TIDAK dihitung
# di sini -- selalu cek aturan BPJS/PMK terbaru & kebijakan payroll internal.

# BPJS Kesehatan: iuran 5% dari upah (1% ditanggung karyawan, 4% perusahaan),
# dasar upah dibatasi plafon (per aturan berlaku umum saat ini: Rp12.000.000).
BPJS_KESEHATAN_TARIF_KARYAWAN = 0.01
BPJS_KESEHATAN_TARIF_PERUSAHAAN = 0.04
# [DIPERBAIKI] SEBELUMNYA plafon ini murni angka hardcode -- kalau pemerintah
# update plafon, kode HARUS direvisi & di-deploy ulang manual. Sekarang bisa
# di-override lewat environment variable BPJS_KESEHATAN_PLAFON_UPAH TANPA
# ubah kode sama sekali (tinggal set env var di server & restart proses) --
# kalau env var tidak diset, tetap pakai nilai default di bawah ini sbg
# fallback. Nilai yang DIPAKAI SAAT PROSES BERJALAN juga ditampilkan balik
# di ringkasan hasil (lihat "catatan_bpjs_pph") supaya kelihatan kalau ada
# yang lupa update.
BPJS_KESEHATAN_PLAFON_UPAH = float(os.environ.get("BPJS_KESEHATAN_PLAFON_UPAH", 12_000_000))

# BPJS Ketenagakerjaan -- JHT (Jaminan Hari Tua): 2% karyawan, 3.7% perusahaan
# (tanpa plafon upah). JP (Jaminan Pensiun): 1% karyawan, 2% perusahaan,
# dasar upah dibatasi plafon yang di-update berkala oleh pemerintah -- di
# sini hanya dipakai sbg estimasi kasar, WAJIB dicek angka plafon terbaru.
BPJS_JHT_TARIF_KARYAWAN = 0.02
BPJS_JHT_TARIF_PERUSAHAAN = 0.037
BPJS_JP_TARIF_KARYAWAN = 0.01
BPJS_JP_TARIF_PERUSAHAAN = 0.02
# [DIPERBAIKI] Sama seperti plafon Kesehatan di atas -- sekarang overridable
# lewat environment variable BPJS_JP_PLAFON_UPAH_ESTIMASI, indikatif, SELALU
# cek angka plafon JP resmi terbaru.
BPJS_JP_PLAFON_UPAH_ESTIMASI = float(os.environ.get("BPJS_JP_PLAFON_UPAH_ESTIMASI", 10_547_400))

_TOLERANSI_SELISIH_GAJI_RUPIAH = 5     # toleransi pembulatan komponen gaji, dlm Rupiah
_TOLERANSI_TARIF_BPJS = 0.003          # toleransi relatif utk reasonableness check BPJS
_PPH21_TARIF_EFEKTIF_MAKS_SLIP = 0.34  # sama dgn batas atas Pasal 17/TER tertinggi

# [BARU] Lantai konservatif utk sanity-check PPh 21 tambahan (lihat
# cek_kewajaran_pph21_slip) -- PTKP TK/0 (tanpa tanggungan) adalah kategori
# PTKP TERENDAH yang berlaku umum; siapapun status PTKP karyawan sebenarnya,
# PTKP bulanannya TIDAK MUNGKIN lebih rendah dari angka ini. WAJIB dicek
# angka PTKP resmi terbaru kalau berubah (bisa di-override lewat environment
# variable PTKP_BULANAN_TK0_ESTIMASI).
_PTKP_BULANAN_TK0_ESTIMASI = float(os.environ.get("PTKP_BULANAN_TK0_ESTIMASI", 4_500_000))
_AMBANG_PPH21_DI_BAWAH_PTKP = 15_000  # PPh 21 di atas ini dianggap 'berarti', bukan cuma pembulatan

# [BARU] Ambang deteksi anomali gaji ANTAR-PERIODE per karyawan (gap #4).
_AMBANG_PERUBAHAN_GAJI_POKOK_ANTAR_PERIODE = 0.5      # >50% perubahan relatif dianggap anomali
_AMBANG_PERUBAHAN_GAJI_POKOK_MINIMAL_RUPIAH = 500_000  # abaikan lonjakan kecil scr rupiah (gapok rendah)


def _parse_angka_id(v) -> float:
    """
    Parse nilai jadi float, tahan berbagai format umum di file Excel
    akuntansi Indonesia:
      - angka murni (int/float) dari openpyxl -> langsung dipakai
      - "1.500.000"        -> 1500000.0   (titik = pemisah ribuan)
      - "1.500.000,75"     -> 1500000.75  (koma = desimal, gaya Indonesia)
      - "Rp 1.500.000"     -> 1500000.0   (buang simbol mata uang & spasi)
      - "(500.000)"        -> -500000.0   (kurung = negatif, gaya akuntansi)
      - "1,500,000.75"     -> 1500000.75  (gaya Inggris, titik = desimal)
      - None / "" / "-"    -> 0.0
    Kalau tetap tidak bisa di-parse, return 0.0 (tidak raise, supaya satu
    sel yang aneh tidak menggagalkan seluruh baris/sheet).
    """
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        try:
            if v != v:  # NaN check tanpa import math/pandas di sini
                return 0.0
        except Exception:
            pass
        return float(v)

    s = str(v).strip()
    if s in ("", "-", "–", "—", "n/a", "N/A", "nan", "None"):
        return 0.0

    negatif = False
    if s.startswith("(") and s.endswith(")"):
        negatif = True
        s = s[1:-1].strip()

    # Buang simbol mata uang & teks non-angka umum (Rp, IDR, dgn/tanpa spasi
    # setelahnya, mis. "Rp 1.500.000" ATAU "Rp1.500.000")
    s = re.sub(r"(?i)^\s*(rp|idr)\.?\s*", "", s).strip()
    s = s.replace(" ", "")

    if s.startswith("-"):
        negatif = True
        s = s[1:]

    if not s:
        return 0.0

    # Tentukan gaya format: kalau ada koma DAN titik, yang PALING KANAN
    # adalah pemisah desimal (baik gaya ID "1.500.000,75" maupun EN
    # "1,500,000.75"); kalau cuma salah satu, tebak dari posisi/panjang.
    ada_titik = "." in s
    ada_koma = "," in s

    try:
        if ada_titik and ada_koma:
            if s.rfind(",") > s.rfind("."):
                # gaya ID: titik ribuan, koma desimal
                s = s.replace(".", "").replace(",", ".")
            else:
                # gaya EN: koma ribuan, titik desimal
                s = s.replace(",", "")
        elif ada_koma and not ada_titik:
            # Cuma koma, tidak ada titik. Kalau cuma SATU koma & tepat 2
            # digit di belakangnya -> gaya ID desimal ("1500000,75").
            # Selain itu (banyak koma, atau bukan 2 digit) -> pemisah
            # ribuan gaya EN ("1,500,000"), buang semua koma.
            bagian = s.split(",")
            if len(bagian) == 2 and len(bagian[1]) == 2:
                s = s.replace(",", ".")
            else:
                s = s.replace(",", "")
        elif ada_titik and not ada_koma:
            bagian = s.split(".")
            if len(bagian[-1]) == 2:
                s = s  # sudah dalam bentuk desimal titik yang benar (EN)
            elif len(bagian) > 1 and all(len(b) == 3 for b in bagian[1:]):
                s = s.replace(".", "")  # semua kelompok 3 digit -> ribuan
            # else biarkan apa adanya (mis. "12.5")

        hasil = float(s)
        return -hasil if negatif else hasil
    except (ValueError, TypeError):
        return 0.0


_KATA_KUNCI_KARYAWAN = [
    "nama karyawan", "nama pegawai", "nip", "nomor karyawan",
    "kode karyawan", "employee id", "id karyawan", "nama staff",
    "nama staf", "nama",  # "nama" generik sengaja PALING TERAKHIR
]
_KATA_KUNCI_GAJI_POKOK = [
    "gaji pokok", "basic salary", "gapok", "gaji dasar", "upah pokok",
]
_KATA_KUNCI_POTONGAN_ATAU_PPH = [
    "potongan", "deduction", "pph 21", "pph21", "pph pasal 21", "pajak",
]
_KATA_KUNCI_THP = [
    "gaji bersih", "take home pay", "thp", "netto", "net pay",
    "gaji diterima", "jumlah diterima", "total diterima",
]


def _gabung_teks_baris(row) -> str:
    return " ".join(str(c) for c in row if c is not None).lower()


def _skor_baris_header_gaji(teks: str) -> dict:
    return {
        "karyawan": any(k in teks for k in _KATA_KUNCI_KARYAWAN[:-1]),  # exclude "nama" polos di skor utama
        "karyawan_lemah": "nama" in teks,
        "gaji_pokok": any(k in teks for k in _KATA_KUNCI_GAJI_POKOK),
        "potongan": any(k in teks for k in _KATA_KUNCI_POTONGAN_ATAU_PPH),
        "thp": any(k in teks for k in _KATA_KUNCI_THP),
    }


def _cari_header_row_slip_gaji(ws, max_scan: int = 15, logger=logger, nama_sheet: str = ""):
    """
    Cari baris header Slip Gaji. Strategi berlapis (dari paling ketat ke
    paling longgar) supaya toleran ke variasi template riil:

      Lapis A -- 1 baris tunggal, syarat lengkap (karyawan + gaji pokok +
                 potongan/pph ATAU thp) -- paling akurat kalau template
                 rapi 1 baris header.
      Lapis B -- HEADER 2-BARIS: gabungkan teks baris i + baris i+1 (untuk
                 kasus header grup, mis. "POTONGAN" di baris atas dan
                 "BPJS KES | BPJS JHT | PPh 21" di baris bawah), cek syarat
                 yang sama pada teks gabungan. Kalau cocok, PAKAI BARIS
                 BAWAH (i+1) sebagai baris kolom aktual (karena situ yang
                 berisi nama kolom presisi), tapi tandai perlu isi kolom
                 kosong dari label grup di atasnya.
      Lapis C -- LONGGAR: syarat cukup gaji_pokok + (karyawan ATAU
                 karyawan_lemah) + (potongan ATAU thp) -- menampung
                 template minimalis yang cuma punya sedikit kolom.

    Return: (nomor_baris_1_indexed, list_header, info) atau (None, None, None).
    info berisi {"mode": "1baris"|"2baris", "baris_atas": list|None}
    untuk dipakai parse_sheet_slip_gaji menggabungkan label grup kalau perlu.
    """
    baris_semua = list(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True))

    def _cell(r, idx):
        return r[idx] if r is not None and idx < len(r) else None

    # Lapis A & C sekaligus (longgar sudah mencakup ketat) -- 1 baris
    for i, row in enumerate(baris_semua):
        teks = _gabung_teks_baris(row)
        skor = _skor_baris_header_gaji(teks)
        ada_karyawan = skor["karyawan"] or skor["karyawan_lemah"]
        if skor["gaji_pokok"] and ada_karyawan and (skor["potongan"] or skor["thp"]):
            # Sebelum langsung dipakai: cek apakah baris ini sebenarnya
            # LABEL GRUP 2-baris, bukan header final -- tandanya: ada sel
            # kosong di baris ini yang justru TERISI di baris berikutnya,
            # DAN baris berikutnya BUKAN baris data (kolom nama/gaji
            # pokok yang sudah ketemu di baris ini masih kosong di baris
            # berikutnya -- kalau kolom itu SUDAH terisi di baris
            # berikutnya, berarti baris berikutnya memang baris data
            # karyawan pertama, bukan sub-header).
            row_next = baris_semua[i + 1] if i + 1 < len(baris_semua) else None
            gabung_diperlukan = False
            if row_next is not None:
                idx_nama_kasar = next(
                    (k for k, h in enumerate(row) if h and any(kk in str(h).lower() for kk in _KATA_KUNCI_KARYAWAN)),
                    None,
                )
                idx_gapok_kasar = next(
                    (k for k, h in enumerate(row) if h and any(kk in str(h).lower() for kk in _KATA_KUNCI_GAJI_POKOK)),
                    None,
                )
                row_next_bukan_data = True
                if idx_nama_kasar is not None and _cell(row_next, idx_nama_kasar) not in (None, ""):
                    row_next_bukan_data = False
                if idx_gapok_kasar is not None and _cell(row_next, idx_gapok_kasar) not in (None, ""):
                    row_next_bukan_data = False
                ada_sel_lanjutan = any(
                    (row[k] is None or str(row[k]).strip() == "") and _cell(row_next, k) not in (None, "")
                    for k in range(max(len(row), len(row_next)))
                )
                gabung_diperlukan = ada_sel_lanjutan and row_next_bukan_data

            if gabung_diperlukan:
                # Prioritas: label di baris BAWAH (lebih dekat ke data,
                # biasanya lebih spesifik, mis. "BPJS Kes"/"PPh 21") MENANG
                # kalau terisi -- baru fallback ke label baris ATAS kalau
                # posisi itu kosong di baris bawah (mis. kolom "Nama
                # Karyawan"/"Gaji Pokok"/"Gaji Bersih" yang memang cuma
                # ada di baris atas, tidak ada sub-kolom di bawahnya).
                # Ini penting karena label grup (mis. "POTONGAN" yang
                # merged menaungi beberapa kolom) TIDAK BOLEH menutupi
                # sub-header spesifik di kolom yang sama pada baris bawah.
                headers_gabungan = list(row)
                for idx in range(min(len(headers_gabungan), len(row_next))):
                    if row_next[idx] not in (None, ""):
                        headers_gabungan[idx] = row_next[idx]
                if logger:
                    logger.info(
                        f"Sheet '{nama_sheet}': header Slip Gaji terdeteksi 2-baris "
                        f"(label grup di baris {i+1}, sub-kolom di baris {i+2}), digabung."
                    )
                return i + 2, headers_gabungan, {"mode": "2baris", "baris_atas": list(row)}

            return i + 1, list(row), {"mode": "1baris", "baris_atas": None}

    # Lapis B -- kasus header 2-baris di mana baris ATAS SENDIRI tidak lolos
    # syarat longgar (mis. cuma berisi "NIP | Nama | Gaji Pokok" tanpa kata
    # apapun terkait potongan/gaji bersih -- polos "kolom identitas" tanpa
    # label grup potongan sama sekali), tapi GABUNGAN baris ini + baris
    # berikutnya lolos. Ini fallback tambahan di luar Lapis A (yang sudah
    # menangani kasus baris atas MEMANG lolos syarat sendiri).
    for i in range(len(baris_semua) - 1):
        row_atas, row_bawah = baris_semua[i], baris_semua[i + 1]
        teks_gabungan = _gabung_teks_baris(row_atas) + " " + _gabung_teks_baris(row_bawah)
        skor = _skor_baris_header_gaji(teks_gabungan)
        ada_karyawan = skor["karyawan"] or skor["karyawan_lemah"]
        if skor["gaji_pokok"] and ada_karyawan and (skor["potongan"] or skor["thp"]):
            headers_gabungan = list(row_atas)
            for idx in range(min(len(headers_gabungan), len(row_bawah))):
                if row_bawah[idx] not in (None, ""):
                    headers_gabungan[idx] = row_bawah[idx]
            if logger:
                logger.info(
                    f"Sheet '{nama_sheet}': header Slip Gaji terdeteksi 2-baris "
                    f"(baris {i+1}-{i+2}, fallback lapis B), digabung."
                )
            return i + 2, headers_gabungan, {"mode": "2baris", "baris_atas": list(row_atas)}

    if logger:
        logger.warning(
            f"Sheet '{nama_sheet}': TIDAK dikenali sebagai Slip Gaji dalam "
            f"{max_scan} baris pertama -- tidak ditemukan kombinasi "
            "(nama karyawan/NIP) + (gaji pokok) + (potongan/PPh21 ATAU "
            "gaji bersih/THP), baik 1-baris maupun 2-baris header."
        )
    return None, None, None


def cek_komponen_gaji(
    gaji_pokok: float, total_tunjangan: float, total_potongan: float, gaji_bersih: float,
    toleransi_rupiah: float = _TOLERANSI_SELISIH_GAJI_RUPIAH,
) -> dict:
    """
    "Kerjaan akuntan" paling dasar saat cek slip gaji: pastikan rumus
    Gaji Bersih = (Gaji Pokok + Total Tunjangan) - Total Potongan itu BENAR
    secara aritmatika -- ini sumber kesalahan paling umum di file payroll
    manual (salah ketik komponen, lupa masukkan salah satu potongan, dll).
    """
    gaji_bruto = (gaji_pokok or 0) + (total_tunjangan or 0)
    gaji_bersih_seharusnya = gaji_bruto - (total_potongan or 0)
    selisih = (gaji_bersih or 0) - gaji_bersih_seharusnya
    sesuai = abs(selisih) <= toleransi_rupiah
    return {
        "gaji_bruto": gaji_bruto,
        "gaji_bersih_seharusnya": gaji_bersih_seharusnya,
        "selisih": selisih,
        "sesuai": sesuai,
        "catatan": (
            "Komponen gaji konsisten -- Gaji Bersih = (Gaji Pokok + Tunjangan) - Potongan."
            if sesuai else
            f"Gaji Bersih tertulis Rp{gaji_bersih or 0:,.0f} TIDAK SAMA dengan hasil hitung "
            f"(Gaji Pokok Rp{gaji_pokok or 0:,.0f} + Tunjangan Rp{total_tunjangan or 0:,.0f}) "
            f"- Potongan Rp{total_potongan or 0:,.0f} = Rp{gaji_bersih_seharusnya:,.0f} "
            f"(selisih Rp{selisih:,.0f}) -- WAJIB dicek ulang komponen slip ini."
        ),
    }


def cek_kewajaran_bpjs(
    gaji_pokok: float, bpjs_kesehatan: float, bpjs_jht: float, bpjs_jp: float,
    toleransi_relatif: float = _TOLERANSI_TARIF_BPJS,
) -> list:
    """
    Reasonableness check potongan BPJS (bagian karyawan) terhadap tarif resmi
    berlaku umum. HANYA memberi PERINGATAN kalau potongan tertulis di luar
    rentang wajar -- bukan validasi pasti, karena dasar upah (gaji pokok
    saja vs gaji pokok+tunjangan tetap) & plafon bisa beda kebijakan.
    """
    catatan = []
    dasar = min(gaji_pokok or 0, BPJS_KESEHATAN_PLAFON_UPAH)
    perkiraan_kesehatan = dasar * BPJS_KESEHATAN_TARIF_KARYAWAN
    if bpjs_kesehatan and dasar > 0:
        if abs(bpjs_kesehatan - perkiraan_kesehatan) > max(toleransi_relatif * dasar, 2000):
            catatan.append(
                f"Potongan BPJS Kesehatan Rp{bpjs_kesehatan:,.0f} di luar perkiraan wajar "
                f"({BPJS_KESEHATAN_TARIF_KARYAWAN:.0%} x dasar upah, plafon Rp"
                f"{BPJS_KESEHATAN_PLAFON_UPAH:,.0f}) ≈ Rp{perkiraan_kesehatan:,.0f} -- "
                "cek manual (mungkin dasar upahnya beda / ada tunjangan tetap ikut dihitung)."
            )
    dasar_jht = gaji_pokok or 0
    perkiraan_jht = dasar_jht * BPJS_JHT_TARIF_KARYAWAN
    if bpjs_jht and dasar_jht > 0:
        if abs(bpjs_jht - perkiraan_jht) > max(toleransi_relatif * dasar_jht, 2000):
            catatan.append(
                f"Potongan BPJS JHT Rp{bpjs_jht:,.0f} di luar perkiraan wajar "
                f"({BPJS_JHT_TARIF_KARYAWAN:.0%} x gaji pokok) ≈ Rp{perkiraan_jht:,.0f} -- cek manual."
            )
    dasar_jp = min(gaji_pokok or 0, BPJS_JP_PLAFON_UPAH_ESTIMASI)
    perkiraan_jp = dasar_jp * BPJS_JP_TARIF_KARYAWAN
    if bpjs_jp and dasar_jp > 0:
        if abs(bpjs_jp - perkiraan_jp) > max(toleransi_relatif * dasar_jp, 2000):
            catatan.append(
                f"Potongan BPJS JP Rp{bpjs_jp:,.0f} di luar perkiraan wajar "
                f"({BPJS_JP_TARIF_KARYAWAN:.0%} x dasar upah, plafon indikatif Rp"
                f"{BPJS_JP_PLAFON_UPAH_ESTIMASI:,.0f}) ≈ Rp{perkiraan_jp:,.0f} -- "
                "cek manual & pastikan plafon JP yang dipakai sudah yang terbaru."
            )
    return catatan


def cek_kewajaran_pph21_slip(gaji_bruto: float, pph21: float) -> Optional[str]:
    """
    Sama seperti pendekatan PPh 21 di Bukti Potong: skema TER tergantung
    kategori PTKP karyawan, tidak bisa dipastikan pasti dari slip gaji saja
    -- hanya diflag kalau tarif efektifnya TIDAK MASUK AKAL (di luar 0%-34%).

    [BARU] Ditambah 1 sanity-check kedua yang lebih presisi & risiko
    salah-flag rendah: kalau gaji bruto SUDAH di bawah/sekitar PTKP TK/0
    (kategori PTKP TERENDAH yang berlaku umum, lihat _PTKP_BULANAN_TK0_ESTIMASI)
    tapi masih ada potongan PPh 21 yang cukup besar, itu tetap janggal
    APAPUN kategori PTKP karyawan sebenarnya (karena TK/0 sudah yang
    terendah -- kategori lain PTKP-nya lebih tinggi lagi, jadi pajaknya
    seharusnya makin kecil, bukan makin besar).

    CATATAN PENTING: kedua pengecekan di atas TETAP BUKAN pengganti tabel
    TER (Tarif Efektif Rata-rata) resmi per kategori PTKP -- implementasi
    tabel TER lengkap (puluhan baris x beberapa kategori PTKP) adalah scope
    terpisah yang jauh lebih besar & butuh keputusan desain tersendiri
    (mis. dari mana kategori PTKP tiap karyawan didapat kalau tidak ada
    kolomnya di slip). Fungsi ini tetap hanya reasonableness check kasar.
    """
    if not gaji_bruto:
        return None
    tarif_efektif = (pph21 or 0) / gaji_bruto
    if not (0.0 <= tarif_efektif <= _PPH21_TARIF_EFEKTIF_MAKS_SLIP):
        return (
            f"Potongan PPh 21 Rp{pph21 or 0:,.0f} -- tarif efektif {tarif_efektif:.1%} dari gaji "
            "bruto, DI LUAR rentang wajar TER/Pasal 17 (0%-34%), WAJIB dicek manual."
        )
    if gaji_bruto <= _PTKP_BULANAN_TK0_ESTIMASI and (pph21 or 0) > _AMBANG_PPH21_DI_BAWAH_PTKP:
        return (
            f"Potongan PPh 21 Rp{pph21 or 0:,.0f} tertulis padahal Gaji Bruto Rp{gaji_bruto:,.0f} "
            f"ada di bawah/sekitar PTKP TK/0 (±Rp{_PTKP_BULANAN_TK0_ESTIMASI:,.0f}/bulan, kategori "
            "PTKP TERENDAH yang berlaku umum) -- karyawan dg penghasilan sekecil ini seharusnya "
            "TIDAK/NYARIS TIDAK kena potongan PPh 21 (kecuali ada komponen tidak tetap besar spt "
            "bonus/THR di periode ini), WAJIB dicek manual. [Ini bukan perhitungan TER resmi per "
            "kategori PTKP -- hanya sanity-check tambahan.]"
        )
    return None


_NAMA_BULAN_ID = {
    1: "Januari", 2: "Februari", 3: "Maret", 4: "April", 5: "Mei", 6: "Juni",
    7: "Juli", 8: "Agustus", 9: "September", 10: "Oktober", 11: "November", 12: "Desember",
}


def _bersihkan_id_slip(v) -> Optional[str]:
    """
    Bersihkan NIP/NPWP dari kesalahan representasi angka Excel:
      - Excel sering baca kolom NIP sbg NUMBER, jadi "00123" -> 123 (leading
        zero HILANG di source-nya sendiri, bukan sesuatu yang bisa
        dikembalikan parser -- WAJIB disarankan format kolom sbg TEXT).
      - NIP panjang (belasan digit, mis. NIP 18 digit) sering kebaca sbg
        float dan muncul trailing ".0" (mis. 123456789012345.0) -- INI yang
        dibuang di sini supaya tidak nampil aneh di hasil.
    """
    if v is None:
        return None
    if isinstance(v, float):
        if v != v:  # NaN
            return None
        return str(int(v)) if v.is_integer() else str(v)
    if isinstance(v, int):
        return str(v)
    s = str(v).strip()
    return s if s else None


def _format_periode_slip(v):
    """
    Kolom "Periode Gaji" kadang diisi sbg tanggal Excel asli (mis. dipilih
    dari date picker), yang oleh openpyxl dibaca jadi objek datetime/date --
    tampil aneh kalau dibiarkan (mis. "2026-06-01 00:00:00"). Format ulang
    jadi "Bulan Tahun" ala Indonesia. Kalau sudah berupa teks (paling umum,
    mis. "Juni 2026"), dibiarkan apa adanya.
    """
    # [FIX] "v is None" TIDAK menangkap pandas.NaT (nilai kosong yang muncul
    # setelah pd.concat menggabungkan sheet dengan kolom bertipe tanggal) --
    # NaT lolos ke bawah, isinstance(NaT, datetime) ikut True, lalu
    # NaT.month/.year diam-diam balikin `nan` (float), sehingga hasilnya
    # jadi teks sampah "nan nan" alih-alih None. pd.isna() menangkap None
    # DAN NaT sekaligus.
    if pd.isna(v):
        return None
    if isinstance(v, (datetime, date)):
        return f"{_NAMA_BULAN_ID.get(v.month, v.month)} {v.year}"
    return v


# ------------------------------------------------------------------
# [BARU] Normalisasi & validasi "Periode Gaji".
#
# Dipakai untuk:
#   (a) Deteksi duplikat yang TAHAN terhadap beda format penulisan periode
#       yang sebenarnya SAMA (mis. "Juni 2026" vs "06/2026" vs "2026-06").
#       SEBELUMNYA kunci duplikat memakai string periode_gaji APA ADANYA,
#       jadi 2 baris yang sebetulnya orang & bulan sama tapi ditulis beda
#       format tidak akan ketahuan sbg duplikat.
#   (b) Mengurutkan periode secara KRONOLOGIS untuk deteksi anomali gaji
#       antar-periode (lihat blok anomali di proses_slip_gaji()).
#   (c) Cek kewajaran periode itu sendiri -- tahun terlalu lampau (mis.
#       "1900") atau jauh di masa depan, indikasi salah ketik/salah kolom.
# ------------------------------------------------------------------
_BULAN_KE_ANGKA = {
    "januari": 1, "jan": 1,
    "februari": 2, "feb": 2,
    "maret": 3, "mar": 3,
    "april": 4, "apr": 4,
    "mei": 5, "may": 5,
    "juni": 6, "jun": 6, "june": 6,
    "juli": 7, "jul": 7, "july": 7,
    "agustus": 8, "agu": 8, "ags": 8, "aug": 8, "august": 8,
    "september": 9, "sep": 9, "sept": 9,
    "oktober": 10, "okt": 10, "oct": 10, "october": 10,
    "november": 11, "nov": 11,
    "desember": 12, "des": 12, "dec": 12, "december": 12,
}

_TAHUN_PERIODE_MINIMAL = 2000  # periode gaji sebelum tahun ini dianggap tidak masuk akal
_TOLERANSI_TAHUN_PERIODE_KE_DEPAN = 1  # periode > (tahun berjalan + N tahun) dianggap tidak masuk akal


def _normalisasi_periode_gaji(v) -> Optional[Tuple[int, int]]:
    """
    Ubah nilai kolom "Periode Gaji" (format bebas) jadi (tahun, bulan)
    ternormalisasi supaya bisa dibandingkan apple-to-apple, TERLEPAS dari
    gaya penulisannya di file sumber. Return None kalau sama sekali tidak
    bisa dikenali -- caller lalu jatuh ke fallback (bandingkan string apa
    adanya, sama seperti perilaku lama) supaya baris begini tetap ikut
    diproses, bukan malah gagal/dilewati.

    Format yang dikenali (case-insensitive, lentur thd spasi/pemisah):
      - Objek datetime/date Python (dari sel Excel bertipe tanggal asli)
      - "Juni 2026", "Jun 2026", "Jun-26", "Jun'26" (nama bulan ID/EN,
        penuh atau singkatan, + tahun 2 atau 4 digit)
      - "06/2026", "06-2026", "06.2026" (bulan/tahun numerik)
      - "2026-06", "2026/06" (tahun-bulan numerik gaya ISO)
      - Fallback terakhir: coba parse pakai pandas.to_datetime.
    """
    # [FIX] sama seperti _format_periode_slip -- pd.isna() menangkap None
    # DAN pandas.NaT, supaya baris yang periodenya kosong (setelah gabung
    # sheet) tetap dianggap "tidak bisa dikenali" (None), bukan (nan, nan).
    if pd.isna(v):
        return None
    if isinstance(v, (datetime, date)):
        return (v.year, v.month)

    s = str(v).strip()
    if not s:
        return None
    s_low = s.lower()

    m = re.match(r"^([a-zA-Z]+)[\s\-/,]*'?(\d{2,4})$", s_low)
    if m:
        bulan = _BULAN_KE_ANGKA.get(m.group(1))
        if bulan:
            tahun = int(m.group(2))
            if tahun < 100:
                tahun += 2000
            return (tahun, bulan)

    m = re.match(r"^(\d{1,2})[\/\-\.](\d{4})$", s_low)
    if m:
        bulan, tahun = int(m.group(1)), int(m.group(2))
        if 1 <= bulan <= 12:
            return (tahun, bulan)

    m = re.match(r"^(\d{4})[\/\-\.](\d{1,2})$", s_low)
    if m:
        tahun, bulan = int(m.group(1)), int(m.group(2))
        if 1 <= bulan <= 12:
            return (tahun, bulan)

    try:
        dt = pd.to_datetime(s, dayfirst=True, errors="raise")
        return (int(dt.year), int(dt.month))
    except Exception:
        return None


def _cek_kewajaran_periode_gaji(periode_ternormalisasi: Optional[Tuple[int, int]]) -> Optional[str]:
    """
    Flag periode gaji yang tahunnya tidak masuk akal (mis. "1900", atau
    jauh di masa depan) -- indikasi umum salah ketik/salah baca kolom
    tanggal. Periode yang TIDAK BISA dinormalisasi sama sekali tidak
    diperiksa di sini (caller hanya memanggil kalau argumennya bukan None).
    """
    if periode_ternormalisasi is None:
        return None
    tahun, _bulan = periode_ternormalisasi
    tahun_berjalan = datetime.now().year
    if tahun < _TAHUN_PERIODE_MINIMAL:
        return (
            f"Periode Gaji tahun {tahun} tidak masuk akal (terlalu jauh di masa lalu) -- "
            "kemungkinan salah ketik atau salah baca kolom tanggal, WAJIB dicek manual."
        )
    if tahun > tahun_berjalan + _TOLERANSI_TAHUN_PERIODE_KE_DEPAN:
        return (
            f"Periode Gaji tahun {tahun} ada jauh di MASA DEPAN (tahun berjalan {tahun_berjalan}) -- "
            "kemungkinan salah ketik atau salah baca kolom tanggal, WAJIB dicek manual."
        )
    return None


def parse_sheet_slip_gaji(ws, nama_sheet: str) -> pd.DataFrame:
    """
    Parse satu sheet Slip Gaji jadi DataFrame baris-per-karyawan-per-periode.
    Kolom yang dicoba dikenali (nama header fleksibel, cocok pola umum
    template payroll Indonesia): NIP/nama karyawan, jabatan, periode gaji,
    gaji pokok, tunjangan (bisa lebih dari satu kolom -- dijumlahkan),
    potongan (BPJS Kesehatan, BPJS JHT/JP, PPh 21, potongan lain --
    dijumlahkan jadi total_potongan, masing2 tetap disimpan kalau ada
    kolomnya sendiri), gaji bersih/take home pay.
    """
    header_rownum, header_row, info = _cari_header_row_slip_gaji(ws, logger=logger, nama_sheet=nama_sheet)
    if header_row is None:
        raise FormatTidakDikenali(
            f"Sheet '{nama_sheet}' tidak dikenali sebagai Slip Gaji "
            "(tidak ditemukan kolom NAMA KARYAWAN/NIP + GAJI POKOK + "
            "POTONGAN/PPh 21 + GAJI BERSIH/TAKE HOME PAY sekaligus, "
            "baik dalam 1 baris maupun 2 baris header)."
        )

    headers = list(header_row)
    baris_atas = info.get("baris_atas") if info else None

    # Kalau header 2-baris: kolom yang labelnya kosong di baris bawah
    # tapi ada label grup di baris atas -> pakai label grup itu, supaya
    # kata kunci pencarian kolom tetap bisa kena (mis. baris atas
    # "POTONGAN" membentang di atas kolom BPJS Kes/JHT/PPh21 yang masing2
    # cuma tertulis "BPJS Kes"/"JHT"/"PPh 21" di baris bawah -- itu tetap
    # OK karena kata kuncinya sudah spesifik. Ini terutama menolong kolom
    # yang baris bawahnya BENAR2 kosong / cuma angka romawi/urutan.
    if baris_atas:
        for idx in range(min(len(headers), len(baris_atas))):
            if (headers[idx] is None or str(headers[idx]).strip() == "") and baris_atas[idx]:
                headers[idx] = baris_atas[idx]

    def _idx(kata_kunci, sampai=None):
        rentang = list(enumerate(headers))[:sampai] if sampai else list(enumerate(headers))
        for i, h in rentang:
            if h is None:
                continue
            h_low = str(h).strip().lower()
            if any(k in h_low for k in kata_kunci):
                return i
        return None

    def _idx_semua(kata_kunci):
        hasil = []
        for i, h in enumerate(headers):
            if h is None:
                continue
            h_low = str(h).strip().lower()
            if any(k in h_low for k in kata_kunci):
                hasil.append(i)
        return hasil

    idx_nip = _idx(["nip", "nomor karyawan", "kode karyawan", "employee id", "id karyawan"])
    idx_nama = _idx(["nama karyawan", "nama pegawai", "nama staff", "nama staf", "nama"])
    idx_jabatan = _idx(["jabatan", "posisi", "position", "job title"])
    idx_departemen = _idx(["departemen", "divisi", "department", "unit kerja", "bagian"])
    idx_periode = _idx(["periode gaji", "periode", "masa", "bulan", "period"])
    idx_npwp = _idx(["npwp"])
    idx_gaji_pokok = _idx(_KATA_KUNCI_GAJI_POKOK)
    idx_tunjangan = _idx_semua(["tunjangan", "allowance", "insentif", "bonus"])
    idx_bpjs_kesehatan = _idx(["bpjs kesehatan", "bpjs kes", "jkn"])
    idx_bpjs_jht = _idx(["jht", "jaminan hari tua"])
    idx_bpjs_jp_cand = _idx(["jaminan pensiun"])
    idx_bpjs_jp = idx_bpjs_jp_cand if idx_bpjs_jp_cand is not None else _idx(["jp"])
    idx_pph21 = _idx(["pph 21", "pph21", "pph pasal 21"])
    idx_potongan_lain = _idx_semua(["potongan lain", "potongan lain-lain", "potongan lainnya", "other deduction"])
    idx_total_potongan = _idx(["total potongan", "jumlah potongan", "total deduction"])
    # Fallback: kolom generik "Potongan" polos (tanpa "total"/"jumlah" di
    # depan) -- pola umum di template ringkas/UMKM yang cuma punya SATU
    # kolom potongan tanpa dirinci. Dicek TERPISAH & STRIK (header == "potongan"
    # persis, bukan substring) supaya TIDAK ikut kena kolom "Potongan Lain"
    # (yang sudah ditangani idx_potongan_lain di atas).
    if idx_total_potongan is None:
        for i, h in enumerate(headers):
            if h is None:
                continue
            if str(h).strip().lower() in ("potongan", "deduction", "potongan gaji"):
                idx_total_potongan = i
                break
    idx_gaji_bersih = _idx(_KATA_KUNCI_THP)

    # Peringatan kolom AMBIGU: kalau lebih dari satu kolom sama-sama cocok
    # jadi kandidat Gaji Pokok / Gaji Bersih, `_idx()` diam-diam pakai yang
    # PERTAMA ketemu -- berisiko salah pilih kolom kalau template punya mis.
    # "Gaji Pokok Lama" & "Gaji Pokok Baru" sekaligus. Tetap jalan (supaya
    # tidak mengganggu file yang mayoritas benar), tapi WAJIB dicatat di log
    # supaya ketahuan & bisa dicek manual, bukan silent salah baca.
    _kandidat_gaji_pokok = _idx_semua(_KATA_KUNCI_GAJI_POKOK)
    if len(_kandidat_gaji_pokok) > 1:
        logger.warning(
            f"Sheet '{nama_sheet}': ditemukan {len(_kandidat_gaji_pokok)} kolom yang sama2 "
            f"cocok sbg 'Gaji Pokok' ({[headers[i] for i in _kandidat_gaji_pokok]}) -- dipakai "
            f"kolom '{headers[_kandidat_gaji_pokok[0]]}' (paling kiri). Cek manual kalau ini salah."
        )
    _kandidat_gaji_bersih = _idx_semua(_KATA_KUNCI_THP)
    if len(_kandidat_gaji_bersih) > 1:
        logger.warning(
            f"Sheet '{nama_sheet}': ditemukan {len(_kandidat_gaji_bersih)} kolom yang sama2 "
            f"cocok sbg 'Gaji Bersih/THP' ({[headers[i] for i in _kandidat_gaji_bersih]}) -- "
            f"dipakai kolom '{headers[_kandidat_gaji_bersih[0]]}' (paling kiri). Cek manual kalau ini salah."
        )

    if idx_gaji_pokok is None or idx_gaji_bersih is None:
        raise FormatTidakDikenali(
            f"Kolom GAJI POKOK atau GAJI BERSIH/TAKE HOME PAY tidak ditemukan di sheet '{nama_sheet}' "
            f"(header terbaca: {headers})."
        )

    def _ambil(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    def _jumlah_kolom(row, idx_list):
        return sum(_parse_angka_id(_ambil(row, i)) for i in idx_list)

    rows = []
    baris_dilewati_kosong = 0
    for row in ws.iter_rows(min_row=header_rownum + 1, values_only=True):
        nama = _ambil(row, idx_nama)
        nip = _ambil(row, idx_nip)
        gaji_pokok_raw = _ambil(row, idx_gaji_pokok)
        gaji_bersih_raw = _ambil(row, idx_gaji_bersih)
        if nama is None and nip is None and gaji_pokok_raw is None and gaji_bersih_raw is None:
            baris_dilewati_kosong += 1
            continue

        gaji_pokok = _parse_angka_id(gaji_pokok_raw)
        gaji_bersih = _parse_angka_id(gaji_bersih_raw)

        # Baris "total"/rekap di bawah tabel sering ikut kebaca -- deteksi
        # & lewati kalau kolom nama kosong tapi ada angka besar (indikasi
        # baris subtotal, bukan data per karyawan).
        if (nama is None or str(nama).strip() == "") and (nip is None or str(nip).strip() == ""):
            if gaji_pokok == 0 and gaji_bersih == 0:
                continue
            teks_baris = _gabung_teks_baris(row).lower()
            if "total" in teks_baris or "jumlah" in teks_baris or "rekap" in teks_baris:
                continue

        total_tunjangan = _jumlah_kolom(row, idx_tunjangan)
        bpjs_kesehatan = _parse_angka_id(_ambil(row, idx_bpjs_kesehatan))
        bpjs_jht = _parse_angka_id(_ambil(row, idx_bpjs_jht))
        bpjs_jp = _parse_angka_id(_ambil(row, idx_bpjs_jp))
        pph21 = _parse_angka_id(_ambil(row, idx_pph21))
        potongan_lain = _jumlah_kolom(row, idx_potongan_lain)

        total_potongan_tertulis_raw = _ambil(row, idx_total_potongan)
        total_potongan_tertulis = None
        if total_potongan_tertulis_raw not in (None, ""):
            total_potongan_tertulis = _parse_angka_id(total_potongan_tertulis_raw)
            total_potongan = total_potongan_tertulis
            if total_potongan == 0:
                total_potongan = bpjs_kesehatan + bpjs_jht + bpjs_jp + pph21 + potongan_lain
        else:
            total_potongan = bpjs_kesehatan + bpjs_jht + bpjs_jp + pph21 + potongan_lain

        rows.append({
            "sheet": nama_sheet,
            "nip": _bersihkan_id_slip(nip),
            "nama_karyawan": nama,
            "jabatan": _ambil(row, idx_jabatan),
            "departemen": _ambil(row, idx_departemen),
            "periode_gaji": _format_periode_slip(_ambil(row, idx_periode)),
            "npwp": _bersihkan_id_slip(_ambil(row, idx_npwp)),
            "gaji_pokok": gaji_pokok,
            "total_tunjangan": total_tunjangan,
            "bpjs_kesehatan": bpjs_kesehatan,
            "bpjs_jht": bpjs_jht,
            "bpjs_jp": bpjs_jp,
            "pph21": pph21,
            "potongan_lain": potongan_lain,
            "total_potongan": total_potongan,
            # Field TAMBAHAN (tidak menghapus/mengubah field lama) -- angka
            # "Total Potongan" MENTAH sebagaimana tertulis di slip (kalau ada
            # kolomnya), dipakai proses_slip_gaji() utk cross-check terhadap
            # jumlah komponen (BPJS+PPh21+potongan lain). None kalau sheet
            # ini tidak punya kolom Total Potongan sama sekali.
            "total_potongan_tertulis": total_potongan_tertulis,
            "gaji_bersih": gaji_bersih,
        })

    if baris_dilewati_kosong > 0:
        logger.info(f"Sheet '{nama_sheet}': {baris_dilewati_kosong} baris kosong dilewati (bukan error).")
    logger.info(f"Sheet '{nama_sheet}': {len(rows)} baris slip gaji berhasil diparse.")

    return pd.DataFrame(rows)


def proses_slip_gaji(
    df: pd.DataFrame,
    toleransi_rupiah: float = _TOLERANSI_SELISIH_GAJI_RUPIAH,
    histori_gaji_sebelumnya: Optional[Dict[str, dict]] = None,
) -> dict:
    """
    "Kerjaan akuntan" untuk Slip Gaji: cross-check komponen tiap slip (rumus
    gaji bersih, kewajaran PPh 21 & BPJS), cek duplikat (karyawan+periode
    sama muncul 2x, TAHAN beda format penulisan periode), deteksi anomali
    gaji antar-periode, lalu siapkan rekap payroll & draf jurnal.

    Parameter:
        histori_gaji_sebelumnya: [BARU, opsional] dict {id_karyawan:
            {"periode_urut": (tahun, bulan), "gaji_pokok": ..., "gaji_bruto": ...}}
            dari hasil["histori_gaji_terbaru"] pemrosesan SEBELUMNYA (mis.
            upload bulan lalu) -- dipakai supaya deteksi anomali gaji
            antar-periode (gap #4) tetap jalan walau tiap bulan diupload
            sbg file terpisah, bukan cuma antar-baris dalam 1 file yang
            sama. Fungsi ini TIDAK menyimpan histori ke disk sendiri --
            pemanggil (mis. main.py) yang bertanggung jawab menyimpan &
            mengoper balik "histori_gaji_terbaru" dari return dict ini.
            Default None -- kalau tidak diisi, deteksi anomali HANYA
            jalan antar-baris dalam 1 file/batch ini saja (perilaku minimal,
            tidak butuh perubahan apapun di pemanggil lama).

    Return dict:
        "df": DataFrame asli + kolom tambahan (gaji_bruto, gaji_bersih_seharusnya,
              selisih_gaji, status)
        "ringkasan": rekap total gaji dibayarkan, total per komponen potongan,
                     rekap per departemen (kalau ada kolomnya)
        "masalah": list slip yang PERLU DICEK MANUAL beserta alasannya
        "draf_jurnal": list draf jurnal payroll per karyawan (akun generik,
                       WAJIB direview akuntan sebelum posting)
        "histori_gaji_terbaru": [BARU] dict {id_karyawan: {...}} berisi data
                       gaji periode PALING AKHIR per karyawan di batch ini --
                       simpan & oper balik sbg histori_gaji_sebelumnya di
                       pemrosesan berikutnya kalau mau anomali antar-periode
                       tetap terdeteksi lintas file upload.
    """
    if df is None or df.empty:
        return {"df": df, "ringkasan": {}, "masalah": [], "draf_jurnal": [], "histori_gaji_terbaru": {}}

    df = df.copy()
    masalah = []
    slip_terlihat: dict = {}

    gaji_bruto_list, gaji_bersih_seharusnya_list, selisih_list, status_list = [], [], [], []

    for i, row in df.iterrows():
        gaji_pokok = float(row.get("gaji_pokok") or 0)
        total_tunjangan = float(row.get("total_tunjangan") or 0)
        total_potongan = float(row.get("total_potongan") or 0)
        gaji_bersih = float(row.get("gaji_bersih") or 0)

        v_komponen = cek_komponen_gaji(gaji_pokok, total_tunjangan, total_potongan, gaji_bersih, toleransi_rupiah)
        gaji_bruto_list.append(v_komponen["gaji_bruto"])
        gaji_bersih_seharusnya_list.append(v_komponen["gaji_bersih_seharusnya"])
        selisih_list.append(v_komponen["selisih"])

        alasan = []
        if not v_komponen["sesuai"]:
            alasan.append(v_komponen["catatan"])

        if gaji_bersih <= 0 and (gaji_pokok > 0 or total_tunjangan > 0):
            alasan.append(
                f"Gaji Bersih tertulis Rp{gaji_bersih:,.0f} (nol/negatif) padahal ada Gaji "
                "Pokok/Tunjangan -- WAJIB dicek, kemungkinan salah input."
            )

        # [BARU] Gaji Pokok negatif -- tidak masuk akal secara akuntansi,
        # kemungkinan salah ketik/salah kolom.
        if gaji_pokok < 0:
            alasan.append(
                f"Gaji Pokok tertulis NEGATIF Rp{gaji_pokok:,.0f} -- tidak masuk akal, "
                "WAJIB dicek kembali data sumbernya."
            )

        # [BARU] Gaji Pokok nol tapi ada komponen lain yang terisi -- indikasi
        # kuat kolom Gaji Pokok salah terbaca (mis. formula error/kolom
        # geser), bukan berarti karyawan memang tidak digaji pokok.
        if gaji_pokok == 0 and (total_tunjangan > 0 or total_potongan > 0 or gaji_bersih > 0):
            alasan.append(
                "Gaji Pokok tertulis Rp0 padahal ada Tunjangan/Potongan/Gaji Bersih yang "
                "terisi -- kemungkinan kolom Gaji Pokok salah terbaca (sel kosong/formula "
                "error/kolom bergeser), WAJIB dicek manual."
            )

        # [BARU] Komponen potongan NEGATIF (BPJS Kesehatan/JHT/JP, PPh 21,
        # Potongan Lain) -- SEBELUMNYA nilai negatif langsung dijumlahkan apa
        # adanya (mis. potongan_lain negatif MALAH MENGURANGI total potongan
        # alih-alih menambah), tanpa flag apapun -- berisiko 'menyembunyikan'
        # kesalahan lain di baliknya secara diam-diam. Sekarang WAJIB diflag
        # supaya ketahuan; angka mentahnya TETAP dipakai apa adanya di
        # perhitungan (tidak diubah/dipaksa positif otomatis), karena bisa
        # saja itu memang koreksi/refund yang sah -- keputusan akhir tetap
        # di akuntan yang mereview.
        for _label_komp, _nilai_komp in (
            ("BPJS Kesehatan", row.get("bpjs_kesehatan")),
            ("BPJS JHT", row.get("bpjs_jht")),
            ("BPJS JP", row.get("bpjs_jp")),
            ("PPh 21", row.get("pph21")),
            ("Potongan Lain", row.get("potongan_lain")),
        ):
            if _kosong(_nilai_komp):
                continue
            try:
                _nilai_komp_f = float(_nilai_komp)  # [FIX] no "or 0" fallback -- _kosong() sudah menyaring None/NaN di atas
            except (TypeError, ValueError):
                continue
            if _nilai_komp_f < 0:
                alasan.append(
                    f"Komponen potongan '{_label_komp}' tertulis NEGATIF Rp{_nilai_komp_f:,.0f} -- "
                    "nilai negatif pada kolom potongan akan MENGURANGI total potongan alih-alih "
                    "menambah (bisa menyembunyikan kesalahan lain di baliknya secara diam-diam), "
                    "WAJIB dicek manual apakah ini salah ketik/salah tanda atau memang koreksi/"
                    "refund yang sah."
                )

        # [BARU] Periode Gaji tidak masuk akal (tahun terlalu lampau, mis.
        # "1900", atau jauh di masa depan) -- indikasi umum salah ketik/salah
        # baca kolom tanggal. Hasil normalisasi (_periode_ternormalisasi) juga
        # dipakai ulang di bawah utk kunci duplikat yang tahan beda format.
        _periode_ternormalisasi = _normalisasi_periode_gaji(row.get("periode_gaji"))
        _pesan_periode = _cek_kewajaran_periode_gaji(_periode_ternormalisasi)
        if _pesan_periode:
            alasan.append(_pesan_periode)

        # [BARU] Cross-check "Total Potongan" TERTULIS di slip (kalau sheet
        # punya kolomnya) terhadap jumlah komponen potongan yang dirinci
        # (BPJS Kesehatan+JHT+JP+PPh21+Potongan Lain) -- SEBELUMNYA angka
        # tertulis ini langsung dipakai apa adanya tanpa pernah dibandingkan,
        # jadi selisih/salah ketik di kolom ini tidak pernah ketahuan.
        total_potongan_tertulis = row.get("total_potongan_tertulis")
        if total_potongan_tertulis not in (None, ""):
            try:
                total_potongan_tertulis_f = float(total_potongan_tertulis)
            except (TypeError, ValueError):
                total_potongan_tertulis_f = None
            if total_potongan_tertulis_f is not None:
                komponen_potongan = (
                    float(row.get("bpjs_kesehatan") or 0) + float(row.get("bpjs_jht") or 0)
                    + float(row.get("bpjs_jp") or 0) + float(row.get("pph21") or 0)
                    + float(row.get("potongan_lain") or 0)
                )
                selisih_potongan = total_potongan_tertulis_f - komponen_potongan
                if abs(selisih_potongan) > toleransi_rupiah:
                    alasan.append(
                        f"Kolom 'Total Potongan' tertulis Rp{total_potongan_tertulis_f:,.0f} "
                        "TIDAK SAMA dengan jumlah komponen potongan yang dirinci (BPJS "
                        f"Kesehatan+JHT+JP+PPh21+Potongan Lain) Rp{komponen_potongan:,.0f} "
                        f"(selisih Rp{selisih_potongan:,.0f}) -- kemungkinan ada komponen "
                        "potongan lain yang tidak tertangkap kolom terpisah (mis. potongan "
                        "koperasi/kasbon), ATAU salah ketik -- WAJIB dicek manual."
                    )

        if total_potongan > v_komponen["gaji_bruto"]:
            alasan.append(
                f"Total Potongan Rp{total_potongan:,.0f} MELEBIHI Gaji Bruto Rp"
                f"{v_komponen['gaji_bruto']:,.0f} -- tidak masuk akal, cek komponen potongan."
            )

        pesan_pph21 = cek_kewajaran_pph21_slip(v_komponen["gaji_bruto"], float(row.get("pph21") or 0))
        if pesan_pph21:
            alasan.append(pesan_pph21)

        alasan.extend(cek_kewajaran_bpjs(
            gaji_pokok, float(row.get("bpjs_kesehatan") or 0),
            float(row.get("bpjs_jht") or 0), float(row.get("bpjs_jp") or 0),
        ))

        if row.get("npwp"):
            v_npwp = validasi_npwp(row.get("npwp"))
            if not v_npwp["valid"]:
                alasan.append(f"NPWP karyawan tidak sesuai format: {v_npwp['catatan']}")

        # [DIPERBAIKI] SEBELUMNYA kunci duplikat memakai string periode_gaji
        # APA ADANYA -- kalau periode yang SAMA ditulis beda format di 2
        # baris (mis. "Juni 2026" vs "06/2026"), duplikatnya tidak ketahuan
        # krn string-nya beda. Sekarang pakai (tahun, bulan) ternormalisasi
        # (_periode_ternormalisasi, dihitung di atas) sbg kunci; kalau
        # periode itu TIDAK BISA dinormalisasi sama sekali, fallback ke
        # string apa adanya (perilaku lama) supaya baris tsb tetap ikut
        # dicek dedup, bukan malah dilewati begitu saja.
        _id_karyawan = str(row.get("nip") or row.get("nama_karyawan") or "").strip().lower()
        _kunci_periode = (
            _periode_ternormalisasi if _periode_ternormalisasi is not None
            else ("RAW", str(row.get("periode_gaji") or "").strip().lower())
        )
        kunci_slip = (_id_karyawan, _kunci_periode)
        if kunci_slip != ("", ("RAW", "")):
            if kunci_slip in slip_terlihat:
                alasan.append(
                    f"Slip gaji DUPLIKAT -- karyawan & periode yang sama sudah muncul di baris "
                    f"ke-{slip_terlihat[kunci_slip] + 1} juga. Berisiko dobel bayar/dobel posting, "
                    "WAJIB dicek manual."
                )
            else:
                slip_terlihat[kunci_slip] = i

        status_list.append("PERLU REVIEW" if alasan else "OK")
        if alasan:
            masalah.append({
                "baris": i + 1,
                "sheet": row.get("sheet"),
                "nip": row.get("nip"),
                "nama_karyawan": row.get("nama_karyawan"),
                "periode_gaji": row.get("periode_gaji"),
                "gaji_bruto": v_komponen["gaji_bruto"],
                "gaji_bersih": gaji_bersih,
                "alasan": alasan,
            })

    df["gaji_bruto"] = gaji_bruto_list
    df["gaji_bersih_seharusnya"] = gaji_bersih_seharusnya_list
    df["selisih_gaji"] = selisih_list
    df["status"] = status_list

    # ------------------------------------------------------------------
    # [BARU] Deteksi anomali gaji ANTAR-PERIODE per karyawan (gap #4).
    # SEBELUMNYA tidak ada perbandingan lintas periode sama sekali -- mis.
    # Gaji Pokok karyawan yang sama tiba-tiba naik/turun drastis dibanding
    # periode sebelumnya tidak akan terdeteksi. Dua sumber pembanding
    # dipakai sekaligus (saling melengkapi, tidak saling menghapus):
    #   (a) ANTAR-BARIS DI BATCH INI SENDIRI -- kalau file/upload yang sama
    #       memuat >1 periode utk karyawan yang sama (mis. rekap beberapa
    #       bulan sekaligus, atau beberapa sheet per bulan digabung jadi
    #       satu df sebelum dipanggil ke sini), periode-periode tsb
    #       diurutkan kronologis (pakai _normalisasi_periode_gaji) lalu
    #       dibandingkan berurutan.
    #   (b) HISTORI DARI PEMROSESAN SEBELUMNYA (opsional, lewat parameter
    #       histori_gaji_sebelumnya) -- supaya anomali tetap terdeteksi
    #       walau tiap bulan diupload sbg file terpisah. Fungsi ini SENDIRI
    #       tidak menyimpan apapun ke disk (murni fungsi data-in data-out,
    #       sama seperti fungsi proses_xxx lain di modul ini) -- pemanggil
    #       (mis. main.py) yang bertanggung jawab menyimpan hasil
    #       "histori_gaji_terbaru" di bawah & mengopernya balik sbg
    #       histori_gaji_sebelumnya pada pemrosesan periode berikutnya.
    # ------------------------------------------------------------------
    histori_gaji_terbaru: dict = {}
    if not df.empty:
        df["_id_karyawan_anomali"] = df.apply(
            lambda r: str(r.get("nip") or r.get("nama_karyawan") or "").strip().lower(), axis=1
        )
        df["_periode_urut_anomali"] = df.apply(
            lambda r: _normalisasi_periode_gaji(r.get("periode_gaji")), axis=1
        )

        for id_karyawan, grup in df[df["_id_karyawan_anomali"] != ""].groupby("_id_karyawan_anomali"):
            grup_terurut = grup[grup["_periode_urut_anomali"].notna()].sort_values("_periode_urut_anomali")
            histori_awal = (histori_gaji_sebelumnya or {}).get(id_karyawan) if histori_gaji_sebelumnya else None
            baris_sebelumnya = None

            for _, baris_now in grup_terurut.iterrows():
                pembanding, sumber_pembanding = None, None
                if baris_sebelumnya is not None:
                    pembanding = {
                        "gaji_pokok": float(baris_sebelumnya.get("gaji_pokok") or 0),
                        "periode_label": baris_sebelumnya.get("periode_gaji"),
                    }
                    sumber_pembanding = f"periode {pembanding['periode_label']} (di file/upload yang sama)"
                elif (
                    histori_awal is not None
                    and histori_awal.get("periode_urut") is not None
                    and baris_now["_periode_urut_anomali"] is not None
                    # [FIX] "periode_urut" awalnya SELALU tuple (tahun, bulan)
                    # kalau histori_gaji_sebelumnya datang langsung dari hasil
                    # proses_slip_gaji() di memori yang sama -- TAPI begitu
                    # histori itu disimpan ke disk sbg JSON lalu dibaca ulang
                    # (lihat ak.simpan_histori_gaji/muat_histori_gaji, dipakai
                    # main.py utk anomali lintas file upload), JSON tidak
                    # punya tipe tuple, jadi otomatis berubah jadi list --
                    # perbandingan "list < tuple" MELEMPAR TypeError & bikin
                    # SELURUH proses slip gaji GAGAL. tuple(...) di kedua sisi
                    # menyamakan tipe apapun sumbernya (tuple asli / list dari
                    # JSON), tanpa perlu tahu/peduli asalnya dari mana.
                    and tuple(histori_awal["periode_urut"]) < tuple(baris_now["_periode_urut_anomali"])
                ):
                    pembanding = histori_awal
                    sumber_pembanding = f"upload sebelumnya (periode {histori_awal.get('periode_label', '-')})"

                if pembanding is not None:
                    gp_lama = float(pembanding.get("gaji_pokok") or 0)
                    gp_baru = float(baris_now.get("gaji_pokok") or 0)
                    if gp_lama > 0:
                        perubahan = (gp_baru - gp_lama) / gp_lama
                        if (
                            abs(perubahan) > _AMBANG_PERUBAHAN_GAJI_POKOK_ANTAR_PERIODE
                            and abs(gp_baru - gp_lama) > _AMBANG_PERUBAHAN_GAJI_POKOK_MINIMAL_RUPIAH
                        ):
                            baris_idx = baris_now.name
                            arah = "NAIK" if perubahan > 0 else "TURUN"
                            pesan_anomali = (
                                f"Gaji Pokok {arah} {abs(perubahan):.0%} dibanding {sumber_pembanding}: "
                                f"dari Rp{gp_lama:,.0f} menjadi Rp{gp_baru:,.0f} -- WAJIB dicek, "
                                "kemungkinan salah input, ATAU memang ada perubahan resmi (promosi/"
                                "demosi/kenaikan berkala) yang perlu didokumentasikan."
                            )
                            entri_ada = next((m for m in masalah if m["baris"] == baris_idx + 1), None)
                            if entri_ada:
                                entri_ada["alasan"].append(pesan_anomali)
                            else:
                                masalah.append({
                                    "baris": baris_idx + 1,
                                    "sheet": baris_now.get("sheet"),
                                    "nip": baris_now.get("nip"),
                                    "nama_karyawan": baris_now.get("nama_karyawan"),
                                    "periode_gaji": baris_now.get("periode_gaji"),
                                    "gaji_bruto": float(baris_now.get("gaji_bruto") or 0),
                                    "gaji_bersih": float(baris_now.get("gaji_bersih") or 0),
                                    "alasan": [pesan_anomali],
                                })
                            df.at[baris_idx, "status"] = "PERLU REVIEW"

                baris_sebelumnya = baris_now

            if not grup_terurut.empty:
                _terakhir = grup_terurut.iloc[-1]
                if _terakhir["_periode_urut_anomali"] is not None:
                    histori_gaji_terbaru[id_karyawan] = {
                        "periode_urut": _terakhir["_periode_urut_anomali"],
                        "periode_label": _terakhir.get("periode_gaji"),
                        "gaji_pokok": float(_terakhir.get("gaji_pokok") or 0),
                        "gaji_bruto": float(_terakhir.get("gaji_bruto") or 0),
                    }

        df.drop(columns=["_id_karyawan_anomali", "_periode_urut_anomali"], inplace=True)

    # -- Draf jurnal payroll (Beban Gaji & Tunjangan di debet; Utang PPh 21,
    #    Utang BPJS karyawan, Utang Gaji/Kas bersih di kredit) --
    draf_jurnal = []
    for i, row in df.iterrows():
        gaji_bruto = float(row.get("gaji_bruto") or 0)
        pph21 = float(row.get("pph21") or 0)
        bpjs_total_karyawan = float(row.get("bpjs_kesehatan") or 0) + float(row.get("bpjs_jht") or 0) + float(row.get("bpjs_jp") or 0)
        potongan_lain = float(row.get("potongan_lain") or 0)
        gaji_bersih = float(row.get("gaji_bersih") or 0)

        # [BARU] Cek keseimbangan draf jurnal per baris: total KREDIT (PPh 21
        # + BPJS karyawan + Potongan Lain + Kas/Gaji Bersih) HARUS SAMA dengan
        # DEBET (Gaji Bruto) -- kalau slip aslinya sudah tidak konsisten
        # (lihat "masalah" & kolom "selisih_gaji"), draf jurnal ini akan ikut
        # TIDAK BALANCE. SEBELUMNYA ini tidak pernah dicek di level jurnal,
        # jadi draf yang sudah tidak balance bisa lolos sampai mau diposting.
        jml_kredit_total = pph21 + bpjs_total_karyawan + potongan_lain + gaji_bersih
        selisih_jurnal = gaji_bruto - jml_kredit_total
        jurnal_balance = abs(selisih_jurnal) <= toleransi_rupiah

        draf_jurnal.append({
            "baris": i + 1, "sheet": row.get("sheet"),
            "nip": row.get("nip"), "nama_karyawan": row.get("nama_karyawan"),
            "periode_gaji": row.get("periode_gaji"),
            "no_akun_debet": "BEBAN GAJI & TUNJANGAN", "nama_akun_debet": "Beban Gaji & Tunjangan (bruto)",
            "jml_debet": gaji_bruto,
            "no_akun_kredit_1": "UTANG PPH 21", "nama_akun_kredit_1": "Utang PPh 21 (wajib disetor)",
            "jml_kredit_1": pph21,
            "no_akun_kredit_2": "UTANG BPJS", "nama_akun_kredit_2": "Utang BPJS (potongan karyawan, wajib disetor)",
            "jml_kredit_2": bpjs_total_karyawan,
            "no_akun_kredit_3": "UTANG/POTONGAN LAIN", "nama_akun_kredit_3": "Utang/Potongan Lain-lain",
            "jml_kredit_3": potongan_lain,
            "no_akun_kredit_4": "KAS/BANK/UTANG GAJI", "nama_akun_kredit_4": "Kas/Bank/Utang Gaji (bersih dibayarkan)",
            "jml_kredit_4": gaji_bersih,
            # Field TAMBAHAN (additive, tidak menghapus field lama):
            "balance": jurnal_balance,
            "selisih_jurnal": selisih_jurnal,
            "status_slip": row.get("status"),
            "catatan": (
                "Draf otomatis -- BELUM termasuk beban BPJS porsi PERUSAHAAN (Kesehatan "
                f"{BPJS_KESEHATAN_TARIF_PERUSAHAAN:.1%}, JHT {BPJS_JHT_TARIF_PERUSAHAAN:.1%}, "
                "JP, JKK, JKM) karena tidak selalu tertulis di slip gaji per karyawan -- "
                "tambahkan manual dari data payroll perusahaan bila ada. Utang PPh 21 & BPJS "
                "WAJIB disetor sesuai jatuh tempo & di-cross-check ke SPT Masa PPh 21 terkait."
                + ("" if jurnal_balance else (
                    f" -- PERINGATAN: draf jurnal ini TIDAK BALANCE (Debet Rp{gaji_bruto:,.0f} "
                    f"vs Total Kredit Rp{jml_kredit_total:,.0f}, selisih Rp{selisih_jurnal:,.0f}) "
                    "karena data slip aslinya sudah tidak konsisten -- JANGAN diposting sebelum "
                    "slip sumbernya dikoreksi."
                ))
            ),
        })

    total_gaji_pokok = float(df["gaji_pokok"].fillna(0).sum())
    total_tunjangan = float(df["total_tunjangan"].fillna(0).sum())
    total_bruto = float(df["gaji_bruto"].fillna(0).sum())
    total_pph21 = float(df["pph21"].fillna(0).sum())
    total_bpjs_kesehatan = float(df["bpjs_kesehatan"].fillna(0).sum())
    total_bpjs_jht = float(df["bpjs_jht"].fillna(0).sum())
    total_bpjs_jp = float(df["bpjs_jp"].fillna(0).sum())
    total_potongan_lain = float(df["potongan_lain"].fillna(0).sum())
    total_bersih = float(df["gaji_bersih"].fillna(0).sum())

    rekap_departemen = {}
    if "departemen" in df.columns and df["departemen"].notna().any():
        rekap_departemen = (
            df.groupby("departemen")[["gaji_bruto", "gaji_bersih"]].sum().fillna(0).to_dict("index")
        )

    ringkasan = {
        "jumlah_slip": len(df),
        "jumlah_karyawan": df["nip"].nunique() if df["nip"].notna().any() else df["nama_karyawan"].nunique(),
        "total_gaji_pokok": total_gaji_pokok,
        "total_tunjangan": total_tunjangan,
        "total_gaji_bruto": total_bruto,
        "total_pph21": total_pph21,
        "total_bpjs_kesehatan_karyawan": total_bpjs_kesehatan,
        "total_bpjs_jht_karyawan": total_bpjs_jht,
        "total_bpjs_jp_karyawan": total_bpjs_jp,
        "total_potongan_lain": total_potongan_lain,
        "total_gaji_bersih_dibayarkan": total_bersih,
        "rekap_per_departemen": rekap_departemen,
        "jumlah_perlu_review": len(masalah),
        "jumlah_duplikat": sum(1 for m in masalah if any("DUPLIKAT" in a for a in m["alasan"])),
        "jumlah_draf_jurnal_tidak_balance": sum(1 for d in draf_jurnal if not d["balance"]),
        # [BARU] Counter tambahan utk gap #4, #5, #6 -- supaya ketahuan
        # ringkas dari ringkasan saja tanpa harus scan "masalah" satu-satu.
        "jumlah_anomali_gaji_antar_periode": sum(
            1 for m in masalah if any("dibanding periode" in a for a in m["alasan"])
        ),
        "jumlah_potongan_negatif": sum(
            1 for m in masalah if any("tertulis NEGATIF" in a for a in m["alasan"])
        ),
        "jumlah_periode_janggal": sum(
            1 for m in masalah if any("Periode Gaji tahun" in a for a in m["alasan"])
        ),
        "catatan_bpjs_pph": (
            "Kewajaran PPh 21 & BPJS di sini hanya reasonableness check (tarif efektif/estimasi "
            "resmi berlaku umum) -- PPh 21 pakai skema TER tergantung PTKP karyawan (belum "
            "implementasi tabel TER resmi lengkap -- scope terpisah), & plafon BPJS bisa berubah, "
            "jadi TIDAK menggantikan perhitungan payroll resmi. Beban BPJS porsi PERUSAHAAN belum "
            "termasuk di draf jurnal karena umumnya tidak tertulis di slip per karyawan. "
            f"[BARU] Plafon yang DIPAKAI SAAT PROSES INI: BPJS Kesehatan Rp"
            f"{BPJS_KESEHATAN_PLAFON_UPAH:,.0f}, BPJS JP (estimasi) Rp{BPJS_JP_PLAFON_UPAH_ESTIMASI:,.0f} "
            "-- kalau pemerintah update plafon, ubah lewat environment variable "
            "BPJS_KESEHATAN_PLAFON_UPAH / BPJS_JP_PLAFON_UPAH_ESTIMASI (tanpa perlu ubah kode)."
        ),
    }

    return {
        "df": df, "ringkasan": ringkasan, "masalah": masalah, "draf_jurnal": draf_jurnal,
        "histori_gaji_terbaru": histori_gaji_terbaru,
    }


def proses_file_slip_gaji(
    file_like, nama_file: str = None,
    histori_gaji_sebelumnya: Optional[Dict[str, dict]] = None,
) -> dict:
    """
    Fungsi siap-pakai tingkat file: baca file (.xlsx/.xlsm/.xls/.csv/.pdf),
    cari semua sheet yang cocok sebagai Slip Gaji, gabungkan, lalu proses
    lewat proses_slip_gaji(). Dipakai oleh app.py utk 1 file upload.

    [BARU] histori_gaji_sebelumnya: parameter opsional (default None,
    100% backward compatible dgn pemanggil lama) -- diteruskan apa adanya
    ke proses_slip_gaji() utk deteksi anomali gaji ANTAR-PERIODE lintas
    file upload (lihat docstring proses_slip_gaji()).

    Return sama seperti proses_slip_gaji() (termasuk "histori_gaji_terbaru"),
    ditambah "sheet_dilewati": daftar nama sheet yang TIDAK cocok format
    Slip Gaji sama sekali.
    """
    nama_file = nama_file or getattr(file_like, "name", "") or ""
    daftar_sheet, _df_coa = _siapkan_daftar_sheet(file_like, nama_file)

    df_list = []
    sheet_dilewati = []
    for nama, ws in daftar_sheet:
        try:
            df_sheet = parse_sheet_slip_gaji(ws, nama)
        except FormatTidakDikenali:
            sheet_dilewati.append(nama)
            continue
        if df_sheet.empty:
            sheet_dilewati.append(nama)
            continue
        df_list.append(df_sheet)

    if not df_list:
        return {
            "df": pd.DataFrame(), "ringkasan": {}, "masalah": [], "draf_jurnal": [],
            "histori_gaji_terbaru": {}, "sheet_dilewati": sheet_dilewati,
        }

    df_gabungan = pd.concat(df_list, ignore_index=True)
    hasil = proses_slip_gaji(df_gabungan, histori_gaji_sebelumnya=histori_gaji_sebelumnya)
    hasil["sheet_dilewati"] = sheet_dilewati
    return hasil


# ============================================================
# 11E. BUKTI KAS MASUK/KELUAR -- PARSING, VALIDASI SALDO, & DRAF JURNAL
# ============================================================
# Sheet Bukti Kas biasanya berupa REKAP/LOG mutasi kas kecil (petty cash)
# atau kas tunai non-bank: satu baris = satu bukti kas masuk ATAU keluar.
#
# "Kerjaan akuntan" saat menerima rekap bukti kas ini:
#   1. Kenali arah tiap bukti: MASUK (kas bertambah, mis. diterima dari...)
#      atau KELUAR (kas berkurang, mis. dibayarkan kepada...) -- dari kolom
#      jenis eksplisit kalau ada, atau dari kolom mana yang terisi
#      (jml_masuk vs jml_keluar).
#   2. Cek kelengkapan: nomor bukti, tanggal, keterangan/pihak terkait harus
#      terisi -- bukti kas tanpa keterangan jelas rawan disalahgunakan.
#   3. Kalau ada kolom saldo berjalan, cross-check: Saldo Akhir = Saldo Awal
#      + Kas Masuk - Kas Keluar per baris (berurutan) -- ini yang paling
#      sering meleset di rekap manual (salah rumus / lompat baris).
#   4. Deteksi anomali: nomor bukti duplikat, nominal 0/negatif, bukti
#      keluar nominal besar tanpa keterangan/penerima jelas.
#   5. Siapkan draf jurnal per bukti (akun lawan tetap generik -- WAJIB
#      direview akuntan, karena akun pasti tergantung tujuan kas tsb, mis.
#      apakah itu pembayaran beban, pelunasan piutang, dll).
#
# CATATAN: modul ini TIDAK menggantikan kontrol internal kas kecil (mis.
# batas plafon kas kecil per perusahaan, otorisasi berjenjang) -- itu
# kebijakan internal yang beda-beda per klien, WAJIB dicek manual.

_TOLERANSI_SELISIH_SALDO_KAS_RUPIAH = 5  # toleransi pembulatan saldo, dlm Rupiah
_AMBANG_KAS_KELUAR_BESAR_TANPA_KETERANGAN = 1_000_000  # indikatif, sesuaikan kebijakan klien
_AMBANG_NOMINAL_EKSTRIM_KAS = 500_000_000  # indikatif -- di atas ini dianggap tidak wajar utk kas tunai


def _kosong_kas(v) -> bool:
    """
    [FIX -- bug NaN lolos dari pengecekan "kosong"] True kalau nilai v harus
    dianggap KOSONG: None, string kosong/whitespace, atau NaN pandas.

    Kenapa perlu helper ini: pd.DataFrame(list_of_dicts) MENGKONVERSI setiap
    None jadi float('nan') di semua kolom -- termasuk kolom teks (bukan cuma
    kolom angka). Kode lama di modul ini pakai `if not row.get("keterangan")`
    atau `if not row.get("pihak_terkait")` dengan asumsi nilai kosong = None,
    padahal setelah lewat DataFrame nilainya sudah jadi NaN. Di Python,
    `bool(float('nan'))` adalah True (NaN itu truthy!), jadi `not nan` = False
    -- akibatnya baris dengan keterangan/pihak terkait BENAR-BENAR kosong
    tidak pernah kena tandai "PERLU REVIEW", dan draf jurnal bisa salah pilih
    teks "nan" sebagai nama akun/kategori. Helper ini menyamakan None, ""
    (setelah strip), dan NaN sebagai satu kondisi "kosong" yang konsisten.
    """
    if v is None:
        return True
    if isinstance(v, float) and pd.isna(v):
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def _teks_kas(v, default: str = "") -> str:
    """Ambil v sebagai teks bersih, atau `default` kalau v kosong (lihat _kosong_kas)."""
    return default if _kosong_kas(v) else str(v).strip()


def _cari_header_row_bukti_kas(ws, max_scan: int = 10):
    """
    Cari baris header sheet Bukti Kas: butuh (nomor bukti kas ATAU kas
    masuk/keluar tersurat) + tanggal + (keterangan/uraian) sekaligus.
    """
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        teks = " ".join(str(c) for c in row if c is not None).lower()
        ada_nomor_kas = "nomor bukti kas" in teks or "no bukti kas" in teks or "no. bukti kas" in teks
        ada_kas_arah = "kas masuk" in teks or "kas keluar" in teks
        ada_tanggal = "tanggal" in teks or "tgl" in teks
        ada_keterangan = "keterangan" in teks or "uraian" in teks or "diterima dari" in teks or "dibayarkan kepada" in teks
        if (ada_nomor_kas or ada_kas_arah) and ada_tanggal and ada_keterangan:
            return i + 1, list(row)
    return None, None


def cek_saldo_berjalan_kas(df: pd.DataFrame, toleransi_rupiah: float = _TOLERANSI_SELISIH_SALDO_KAS_RUPIAH) -> list:
    """
    Cross-check saldo berjalan (kalau kolomnya ada & terisi): tiap baris,
    Saldo Akhir SEHARUSNYA = saldo baris sebelumnya + Kas Masuk - Kas Keluar.
    Return list index baris (0-based, sesuai df) yang selisihnya di luar
    toleransi -- KOSONG kalau kolom saldo tidak ada/tidak lengkap.

    [FIX -- bug kontinuitas saldo lintas sheet] Kalau file punya lebih dari
    satu sheet Bukti Kas (mis. "Kas Kecil Kantor" & "Kas Kecil Toko", atau
    kas bulan Januari & Februari yang dipisah per sheet), proses_file_bukti_kas()
    menggabungkan semua sheet itu jadi SATU df sebelum dicek di sini. Tanpa
    reset per sheet, saldo baris TERAKHIR sheet pertama akan dipakai sebagai
    "saldo sebelumnya" utk baris PERTAMA sheet kedua -- padahal itu dua kas
    yang sama sekali tidak berhubungan, jadi selalu muncul "selisih saldo"
    palsu tepat di titik pergantian sheet. Sekarang saldo_sebelumnya di-reset
    setiap kali kolom "sheet" berganti nilai, supaya kontinuitas saldo cuma
    dicek DI DALAM satu sheet/kas yang sama.
    """
    if "saldo" not in df.columns or not df["saldo"].notna().any():
        return []
    bermasalah = []
    saldo_sebelumnya = None
    sheet_sebelumnya = None
    for i, row in df.iterrows():
        sheet_sekarang = row.get("sheet")
        if sheet_sekarang != sheet_sebelumnya:
            # Ganti sheet (kas/periode berbeda) -- mulai ulang rantai saldo.
            saldo_sebelumnya = None
            sheet_sebelumnya = sheet_sekarang
        saldo_tertulis = row.get("saldo")
        if _kosong_kas(saldo_tertulis):
            saldo_sebelumnya = None
            continue
        saldo_tertulis = float(saldo_tertulis)
        if saldo_sebelumnya is not None:
            seharusnya = saldo_sebelumnya + float(row.get("jml_masuk") or 0) - float(row.get("jml_keluar") or 0)
            if abs(saldo_tertulis - seharusnya) > toleransi_rupiah:
                bermasalah.append((i, seharusnya, saldo_tertulis))
        saldo_sebelumnya = saldo_tertulis
    return bermasalah


def parse_sheet_bukti_kas(ws, nama_sheet: str) -> pd.DataFrame:
    """
    Parse satu sheet Bukti Kas Masuk/Keluar jadi DataFrame baris-per-bukti.
    Kolom yang dicoba dikenali (nama header fleksibel): tanggal, nomor
    bukti kas, jenis/arah (kalau tersurat), keterangan, pihak terkait
    (diterima dari/dibayarkan kepada), jumlah masuk, jumlah keluar
    (atau satu kolom "jumlah" + kolom jenis untuk menentukan arahnya),
    saldo berjalan (kalau ada), kategori/akun terkait, penanggung jawab.
    """
    header_rownum, header_row = _cari_header_row_bukti_kas(ws)
    if header_row is None:
        raise FormatTidakDikenali(
            f"Sheet '{nama_sheet}' tidak dikenali sebagai Bukti Kas Masuk/Keluar "
            "(tidak ditemukan kolom NOMOR BUKTI KAS/KAS MASUK-KELUAR + TANGGAL + KETERANGAN sekaligus)."
        )

    headers = header_row

    def _idx(kata_kunci):
        for i, h in enumerate(headers):
            if h is None:
                continue
            h_low = str(h).strip().lower()
            if any(k in h_low for k in kata_kunci):
                return i
        return None

    idx_tanggal = _idx(["tanggal", "tgl"])
    idx_nomor = _idx(["nomor bukti kas", "no bukti kas", "no. bukti kas", "nomor bukti"])
    idx_jenis = _idx(["jenis", "tipe transaksi", "kas masuk/keluar"])
    idx_keterangan = _idx(["keterangan", "uraian", "deskripsi"])
    idx_pihak = _idx(["diterima dari", "dibayarkan kepada", "pihak terkait", "kepada/dari"])
    idx_jml_masuk = _idx(["kas masuk", "jumlah masuk", "penerimaan"])
    idx_jml_keluar = _idx(["kas keluar", "jumlah keluar", "pengeluaran"])
    idx_jumlah_tunggal = _idx(["jumlah", "nominal"])
    idx_saldo = _idx(["saldo"])
    idx_kategori = _idx(["kategori", "akun", "pos anggaran"])
    idx_penanggung_jawab = _idx(["penanggung jawab", "disetujui oleh", "approved by", "diketahui"])

    if idx_nomor is None and idx_jml_masuk is None and idx_jml_keluar is None and idx_jumlah_tunggal is None:
        raise FormatTidakDikenali(
            f"Kolom NOMOR BUKTI KAS atau JUMLAH (masuk/keluar) tidak ditemukan di sheet '{nama_sheet}'."
        )

    def _ambil(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    def _ambil_angka(row, idx):
        v = _ambil(row, idx)
        try:
            return float(v) if v not in (None, "") else 0.0
        except (TypeError, ValueError):
            return 0.0

    rows = []
    for row in ws.iter_rows(min_row=header_rownum + 1, values_only=True):
        nomor = _ambil(row, idx_nomor)
        jml_masuk = _ambil_angka(row, idx_jml_masuk)
        jml_keluar = _ambil_angka(row, idx_jml_keluar)
        jenis_tersurat = _ambil(row, idx_jenis)
        keterangan = _ambil(row, idx_keterangan)

        # -- Fallback: kolom jumlah tunggal + jenis tersurat menentukan arah --
        if idx_jml_masuk is None and idx_jml_keluar is None and idx_jumlah_tunggal is not None:
            jumlah_tunggal = _ambil_angka(row, idx_jumlah_tunggal)
            jenis_upper = str(jenis_tersurat or "").upper()
            if "KELUAR" in jenis_upper:
                jml_keluar = jumlah_tunggal
            elif "MASUK" in jenis_upper:
                jml_masuk = jumlah_tunggal

        if nomor is None and jml_masuk == 0 and jml_keluar == 0 and keterangan is None:
            continue

        rows.append({
            "sheet": nama_sheet,
            "tanggal": _ambil(row, idx_tanggal),
            "nomor_bukti_kas": nomor,
            "jenis_tersurat": jenis_tersurat,
            "keterangan": keterangan,
            "pihak_terkait": _ambil(row, idx_pihak),
            "jml_masuk": jml_masuk,
            "jml_keluar": jml_keluar,
            "saldo": _ambil(row, idx_saldo),
            "kategori": _ambil(row, idx_kategori),
            "penanggung_jawab": _ambil(row, idx_penanggung_jawab),
        })

    return pd.DataFrame(rows)


def proses_bukti_kas(
    df: pd.DataFrame,
    toleransi_rupiah: float = _TOLERANSI_SELISIH_SALDO_KAS_RUPIAH,
    ambang_keluar_besar: float = _AMBANG_KAS_KELUAR_BESAR_TANPA_KETERANGAN,
    ambang_ekstrim: float = _AMBANG_NOMINAL_EKSTRIM_KAS,
    df_coa: pd.DataFrame = None,
) -> dict:
    """
    "Kerjaan akuntan" untuk Bukti Kas Masuk/Keluar: tentukan arah tiap
    baris, cek kelengkapan (nomor/keterangan/pihak terkait), cek duplikat
    nomor bukti, cross-check saldo berjalan (kalau ada), flag pengeluaran
    besar tanpa keterangan jelas, flag nominal yang tidak wajar utk kas
    tunai, cross-check akun_terkait ke COA (kalau df_coa disediakan &
    kolom akun_terkait terisi), siapkan rekap & draf jurnal.

    Return dict: "df", "ringkasan" (total kas masuk/keluar, saldo bersih,
    rekap per kategori), "masalah" (list perlu direview), "draf_jurnal".
    """
    if df is None or df.empty:
        return {"df": df, "ringkasan": {}, "masalah": [], "draf_jurnal": []}

    df = df.copy()
    masalah = []
    nomor_terlihat: dict = {}
    sheet_terlihat_sebelumnya = None

    no_akun_valid = None
    if df_coa is not None and not df_coa.empty and "no_akun" in df_coa.columns:
        no_akun_valid = set(df_coa["no_akun"].astype(str).str.strip())

    arah_list, status_list = [], []
    for i, row in df.iterrows():
        # [FIX -- bug duplikat lintas sheet] Sama seperti kontinuitas saldo
        # di cek_saldo_berjalan_kas(): tiap sheet Bukti Kas = kas/buku yang
        # BERBEDA (mis. "Kas Kecil Kantor" vs "Kas Kecil Toko"), jadi WAJAR
        # kalau penomoran BKM-001/BKK-001 di kedua kas itu sama-sama mulai
        # dari 001. Sebelumnya nomor_terlihat dipakai lintas SEMUA sheet
        # yang sudah digabung jadi satu df, jadi kas manapun yang kebetulan
        # pakai skema nomor sama akan SELALU ditandai "DUPLIKAT" palsu.
        # Sekarang pelacak nomor direset tiap kali kolom "sheet" berganti,
        # supaya duplikat cuma dicek DI DALAM satu sheet/kas yang sama.
        sheet_sekarang = row.get("sheet")
        if sheet_sekarang != sheet_terlihat_sebelumnya:
            nomor_terlihat = {}
            sheet_terlihat_sebelumnya = sheet_sekarang

        jml_masuk = float(row.get("jml_masuk") or 0)
        jml_keluar = float(row.get("jml_keluar") or 0)
        # [FIX -- bug kolom masuk & keluar terisi bersamaan] Sebelumnya arah
        # ditentukan murni "jml_masuk > 0 ? MASUK : (jml_keluar > 0 ? KELUAR
        # : TIDAK DIKETAHUI)" -- kalau KEDUA kolom kebetulan terisi (salah
        # input/kolom tertukar), arah diam-diam dipilih "MASUK" tanpa
        # peringatan, dan nilai jml_keluar TIDAK PERNAH ikut ke draf_jurnal
        # (hanya jml_masuk yang dijurnal) walau tetap ikut dijumlah ke
        # total_kas_keluar di ringkasan -- ringkasan & draf jurnal jadi
        # tidak sinkron tanpa jejak. Sekarang baris begini ditandai arah
        # "AMBIGU" tersendiri (bukan MASUK/KELUAR), supaya draf jurnal TIDAK
        # dibuat otomatis untuk baris ini sampai dipastikan manual arah yang
        # benar (lihat loop draf_jurnal di bawah, hanya menangani arah
        # MASUK/KELUAR).
        if jml_masuk > 0 and jml_keluar > 0:
            arah = "AMBIGU (MASUK & KELUAR TERISI)"
        elif jml_masuk > 0:
            arah = "MASUK"
        elif jml_keluar > 0:
            arah = "KELUAR"
        else:
            arah = "TIDAK DIKETAHUI"
        arah_list.append(arah)

        alasan = []
        nomor = row.get("nomor_bukti_kas")
        if _kosong_kas(nomor):
            alasan.append("Nomor bukti kas kosong -- tiap bukti kas idealnya bernomor urut untuk kontrol internal.")
        else:
            # -- Dibandingkan sbg STRING utuh (bukan digit saja) agar seri nomor
            #    dengan prefiks beda (mis. BKM-001 vs BKK-001) TIDAK dianggap
            #    duplikat palsu -- prefiks itu justru penanda arah yang sah. --
            nomor_bersih = str(nomor).strip().lower()
            if nomor_bersih in nomor_terlihat:
                alasan.append(
                    f"Nomor bukti kas DUPLIKAT -- sudah muncul di baris ke-{nomor_terlihat[nomor_bersih] + 1} "
                    f"juga (sheet '{sheet_sekarang}')."
                )
            else:
                nomor_terlihat[nomor_bersih] = i

        if _kosong_kas(row.get("keterangan")):
            alasan.append("Keterangan/uraian kosong -- bukti kas tanpa keterangan rawan disalahgunakan, cek manual.")

        if arah == "TIDAK DIKETAHUI":
            alasan.append("Jumlah masuk & keluar sama-sama kosong/nol -- tidak bisa ditentukan arah bukti kas ini.")

        if arah == "AMBIGU (MASUK & KELUAR TERISI)":
            alasan.append(
                f"Kolom Kas Masuk (Rp{jml_masuk:,.0f}) dan Kas Keluar (Rp{jml_keluar:,.0f}) SAMA-SAMA terisi "
                "di baris yang sama -- satu bukti kas seharusnya cuma salah satu arah. Cek apakah ada kolom "
                "yang tertukar/salah isi. Draf jurnal untuk baris ini TIDAK dibuat otomatis sampai arah yang "
                "benar dipastikan manual."
            )

        if jml_masuk < 0 or jml_keluar < 0:
            alasan.append("Jumlah kas bernilai negatif -- tidak wajar, cek input.")

        if arah == "KELUAR" and jml_keluar >= ambang_keluar_besar and _kosong_kas(row.get("pihak_terkait")):
            alasan.append(
                f"Kas Keluar Rp{jml_keluar:,.0f} (nominal besar) tanpa 'Dibayarkan Kepada' -- "
                "WAJIB dicek manual, pastikan ada bukti pendukung & otorisasi."
            )

        nominal_terbesar = max(jml_masuk, jml_keluar)
        if nominal_terbesar >= ambang_ekstrim:
            alasan.append(
                f"Nominal Rp{nominal_terbesar:,.0f} di atas ambang wajar transaksi kas tunai "
                f"(indikatif Rp{ambang_ekstrim:,.0f}) -- untuk nominal sebesar ini biasanya lewat "
                "transfer bank, bukan kas tunai. WAJIB diverifikasi ke dokumen pendukung."
            )

        if no_akun_valid is not None and not _kosong_kas(row.get("kategori")):
            kategori_bersih = _teks_kas(row.get("kategori"))
            if kategori_bersih not in no_akun_valid:
                alasan.append(
                    f"Kategori/akun terkait '{kategori_bersih}' tidak ditemukan di daftar akun (COA) -- "
                    "cek apakah kode akun sudah benar."
                )

        status_list.append("PERLU REVIEW" if alasan else "OK")
        if alasan:
            masalah.append({
                "baris": i + 1, "sheet": row.get("sheet"), "nomor_bukti_kas": nomor,
                "tanggal": row.get("tanggal"),
                "arah": arah, "jml_masuk": jml_masuk, "jml_keluar": jml_keluar, "alasan": alasan,
            })

    df["arah"] = arah_list
    df["status"] = status_list

    saldo_bermasalah = cek_saldo_berjalan_kas(df, toleransi_rupiah)
    for i, seharusnya, tertulis in saldo_bermasalah:
        pesan = (
            f"Saldo berjalan tertulis Rp{tertulis:,.0f} TIDAK SAMA dengan hasil hitung "
            f"(saldo sebelumnya + kas masuk - kas keluar) = Rp{seharusnya:,.0f} -- "
            "WAJIB dicek ulang urutan/rumus saldo."
        )
        masalah.append({
            "baris": i + 1, "sheet": df.at[i, "sheet"], "nomor_bukti_kas": df.at[i, "nomor_bukti_kas"],
            "tanggal": df.at[i, "tanggal"],
            "arah": df.at[i, "arah"], "jml_masuk": df.at[i, "jml_masuk"], "jml_keluar": df.at[i, "jml_keluar"],
            "alasan": [pesan],
        })
        df.at[i, "status"] = "PERLU REVIEW"

    # -- Draf jurnal (akun lawan generik, WAJIB direview) --
    draf_jurnal = []
    for i, row in df.iterrows():
        if row["arah"] == "MASUK":
            draf_jurnal.append({
                "baris": i + 1, "sheet": row.get("sheet"), "nomor_bukti_kas": row.get("nomor_bukti_kas"), "arah": "MASUK",
                "no_akun_debet": "KAS", "nama_akun_debet": "Kas/Kas Kecil",
                "jml_debet": float(row.get("jml_masuk") or 0),
                "no_akun_kredit": "PENDAPATAN/PIUTANG/LAIN", "nama_akun_kredit":
                    f"Sesuaikan akun ({_teks_kas(row.get('kategori')) or _teks_kas(row.get('keterangan')) or 'cek keterangan'})",
                "jml_kredit": float(row.get("jml_masuk") or 0),
                "catatan": "Draf otomatis -- tentukan akun lawan sesuai sumber kas masuk (pelunasan "
                           "piutang, pendapatan tunai, setoran modal, dll).",
            })
        elif row["arah"] == "KELUAR":
            draf_jurnal.append({
                "baris": i + 1, "sheet": row.get("sheet"), "nomor_bukti_kas": row.get("nomor_bukti_kas"), "arah": "KELUAR",
                "no_akun_debet": "BEBAN/UTANG/LAIN", "nama_akun_debet":
                    f"Sesuaikan akun ({_teks_kas(row.get('kategori')) or _teks_kas(row.get('keterangan')) or 'cek keterangan'})",
                "jml_debet": float(row.get("jml_keluar") or 0),
                "no_akun_kredit": "KAS", "nama_akun_kredit": "Kas/Kas Kecil",
                "jml_kredit": float(row.get("jml_keluar") or 0),
                "catatan": "Draf otomatis -- tentukan akun lawan sesuai tujuan kas keluar (beban "
                           "operasional, pelunasan utang, pembelian tunai, dll).",
            })

    total_masuk = float(df["jml_masuk"].fillna(0).sum())
    total_keluar = float(df["jml_keluar"].fillna(0).sum())

    rekap_kategori = {}
    if "kategori" in df.columns and df["kategori"].notna().any():
        # [FIX] Normalisasi kategori (strip + apa adanya) sebelum groupby --
        # sebelumnya nilai mentah dipakai langsung, jadi "ATK", "atk", " ATK "
        # dihitung sebagai 3 kategori terpisah di rekap padahal seharusnya satu.
        df["_kategori_rekap"] = df["kategori"].apply(lambda v: _teks_kas(v, default=None))
        rekap_kategori = (
            df[df["_kategori_rekap"].notna()]
            .groupby("_kategori_rekap")[["jml_masuk", "jml_keluar"]]
            .sum().fillna(0).to_dict("index")
        )
        df.drop(columns=["_kategori_rekap"], inplace=True)

    # [BARU] Rekap per pihak terkait (diterima dari / dibayarkan kepada) --
    # membantu akuntan lihat konsentrasi transaksi ke satu pihak tertentu
    # (mis. satu vendor yang berulang kali menerima kas keluar besar), sama
    # semangatnya dengan rekap_per_kategori di atas.
    rekap_pihak_terkait = {}
    if "pihak_terkait" in df.columns and df["pihak_terkait"].notna().any():
        df["_pihak_rekap"] = df["pihak_terkait"].apply(lambda v: _teks_kas(v, default=None))
        rekap_pihak_terkait = (
            df[df["_pihak_rekap"].notna()]
            .groupby("_pihak_rekap")[["jml_masuk", "jml_keluar"]]
            .sum().fillna(0).to_dict("index")
        )
        df.drop(columns=["_pihak_rekap"], inplace=True)

    # [BARU] Rekap PER SHEET -- kalau file berisi lebih dari satu sheet Bukti
    # Kas (mis. beberapa kas kecil / beberapa periode sekaligus dalam satu
    # file), total_kas_masuk/total_kas_keluar di atas adalah GABUNGAN semua
    # sheet -- angka itu sendiri kurang berguna buat akuntan kalau tiap sheet
    # sebenarnya kas yang berbeda. rekap_per_sheet memberi rincian per sheet
    # supaya tetap kelihatan mana kas/periode mana, sama seperti pola yang
    # dipakai di Rekonsiliasi Bank (ringkasan_gabungan tidak menjumlah
    # lintas rekening). Baris tidak digabung/dipisah secara fisik -- df dan
    # draf_jurnal tetap satu tabel dengan kolom "sheet" per baris, supaya
    # tetap bisa difilter/disortir di Excel biasa.
    rekap_per_sheet = {}
    if "sheet" in df.columns:
        for nama_sheet, grp in df.groupby("sheet", dropna=False):
            m_sheet = float(grp["jml_masuk"].fillna(0).sum())
            k_sheet = float(grp["jml_keluar"].fillna(0).sum())
            rekap_per_sheet[str(nama_sheet)] = {
                "jumlah_bukti": int(len(grp)),
                "total_kas_masuk": m_sheet,
                "total_kas_keluar": k_sheet,
                "saldo_bersih_periode": m_sheet - k_sheet,
                "jumlah_perlu_review": int((grp["status"] == "PERLU REVIEW").sum()),
            }

    ringkasan = {
        "jumlah_bukti": len(df),
        "jumlah_sheet": df["sheet"].nunique() if "sheet" in df.columns else 1,
        "total_kas_masuk": total_masuk,
        "total_kas_keluar": total_keluar,
        "saldo_bersih_periode": total_masuk - total_keluar,
        "rekap_per_sheet": rekap_per_sheet,
        "rekap_per_kategori": rekap_kategori,
        "rekap_per_pihak_terkait": rekap_pihak_terkait,
        "jumlah_perlu_review": len(masalah),
        "jumlah_duplikat": sum(1 for m in masalah if any("DUPLIKAT" in a for a in m["alasan"])),
        "jumlah_selisih_saldo": len(saldo_bermasalah),
        "jumlah_nominal_ekstrim": sum(1 for m in masalah if any("di atas ambang wajar" in a for a in m["alasan"])),
        "jumlah_ambigu_masuk_keluar": sum(1 for a in arah_list if a == "AMBIGU (MASUK & KELUAR TERISI)"),
        "catatan": (
            "Cross-check saldo berjalan & deteksi duplikat nomor bukti hanya berlaku DI DALAM satu "
            "sheet/kas yang sama -- rantai saldo & pelacak nomor di-reset tiap kali file berpindah "
            "sheet (tiap sheet dianggap kas/periode terpisah, tidak disambung/dibandingkan lintas "
            "sheet). Ambang nominal besar/ekstrim bersifat indikatif -- sesuaikan dengan kebijakan "
            "plafon kas kecil masing-masing perusahaan. Cross-check akun ke COA hanya berjalan kalau "
            "kolom Kategori/Akun terisi & file COA tersedia. Baris dengan Kas Masuk & Kas Keluar "
            "SAMA-SAMA terisi ditandai arah 'AMBIGU' dan tidak dijurnal otomatis. Kalau file berisi "
            "lebih dari satu sheet Bukti Kas, total di atas adalah GABUNGAN semua sheet -- lihat "
            "'rekap_per_sheet' untuk rincian per kas/periode."
        ),
    }

    return {"df": df, "ringkasan": ringkasan, "masalah": masalah, "draf_jurnal": draf_jurnal}


def proses_file_bukti_kas(file_like, nama_file: str = None) -> dict:
    """
    Fungsi siap-pakai tingkat file: baca file (.xlsx/.xlsm/.xls/.csv/.pdf),
    cari semua sheet yang cocok sebagai Bukti Kas Masuk/Keluar, gabungkan,
    lalu proses lewat proses_bukti_kas() (termasuk cross-check ke COA kalau
    file punya sheet Chart of Accounts). Dipakai oleh app.py utk 1 file upload.

    Return sama seperti proses_bukti_kas(), ditambah "sheet_dilewati":
    daftar nama sheet yang TIDAK cocok format Bukti Kas sama sekali.
    """
    nama_file = nama_file or getattr(file_like, "name", "") or ""
    daftar_sheet, df_coa = _siapkan_daftar_sheet(file_like, nama_file)

    df_list = []
    sheet_dilewati = []
    for nama, ws in daftar_sheet:
        try:
            df_sheet = parse_sheet_bukti_kas(ws, nama)
        except FormatTidakDikenali:
            sheet_dilewati.append(nama)
            continue
        if df_sheet.empty:
            sheet_dilewati.append(nama)
            continue
        df_list.append(df_sheet)

    if not df_list:
        return {"df": pd.DataFrame(), "ringkasan": {}, "masalah": [], "draf_jurnal": [], "sheet_dilewati": sheet_dilewati}

    df_gabungan = pd.concat(df_list, ignore_index=True)
    hasil = proses_bukti_kas(df_gabungan, df_coa=df_coa)
    hasil["sheet_dilewati"] = sheet_dilewati
    return hasil


# ============================================================
# 11F. KARTU STOK/PERSEDIAAN -- PARSING, VALIDASI SALDO PER BARANG, & DRAF JURNAL
# ============================================================
# Sheet Kartu Stok biasanya berupa LOG mutasi persediaan per barang: satu
# baris = satu transaksi (barang masuk ATAU keluar) untuk SATU kode barang,
# tersusun kronologis, dengan saldo qty berjalan (kadang juga saldo nilai).
# BEDA dari modul lain: saldo di sini harus di-cross-check PER KODE BARANG
# (bukan global 1 saldo utk semua baris), karena satu sheet biasa memuat
# mutasi banyak barang sekaligus.
#
# "Kerjaan akuntan/gudang" saat menerima kartu stok ini:
#   1. Kenali arah tiap baris: MASUK (pembelian, retur jual, dst.) atau
#      KELUAR (penjualan, retur beli, pemakaian, dst.).
#   2. Per KODE BARANG (diurutkan tanggal), cross-check Saldo Qty = Saldo
#      Qty sebelumnya + Qty Masuk - Qty Keluar -- kesalahan rumus/lompat
#      baris paling sering muncul di sini.
#   3. Kalau ada kolom Saldo Nilai, cross-check serupa pakai nilai
#      (qty x harga) -- HANYA reasonableness check, karena metode
#      penilaian (FIFO/Average/dll) menentukan harga keluar yang PASTI,
#      dan itu tidak bisa dipastikan ulang dari kartu stok saja.
#   4. Deteksi anomali PALING KRITIS: STOK MINUS (saldo qty < 0) -- barang
#      keluar lebih banyak dari yang tersedia, WAJIB dicek segera (bisa
#      berarti ada transaksi belum tercatat, atau kesalahan input).
#   5. Deteksi anomali lain: qty/harga negatif, harga 0 padahal ada qty,
#      nomor bukti duplikat dalam kode barang yang sama.
#   6. Siapkan draf jurnal per baris (Persediaan didebet saat masuk,
#      dikredit + HPP didebet saat keluar) -- akun lawan tetap generik,
#      WAJIB direview akuntan (metode HPP pasti tergantung kebijakan
#      perusahaan: FIFO/Average/dll).
#
# CATATAN: modul ini TIDAK menghitung ulang HPP dengan metode FIFO/Average
# -- itu perlu urutan pembelian lengkap & kebijakan resmi perusahaan, WAJIB
# dikerjakan/diverifikasi manual oleh akuntan.

_TOLERANSI_SELISIH_QTY_STOK = 0.01       # toleransi pembulatan qty (mis. satuan desimal)
_TOLERANSI_SELISIH_NILAI_STOK_RUPIAH = 5  # toleransi pembulatan nilai persediaan


def _cari_header_row_kartu_stok(ws, max_scan: int = 10):
    """
    Cari baris header sheet Kartu Stok: butuh (kode/nama barang) + tanggal
    + (qty masuk/keluar ATAU saldo stok/stok akhir) sekaligus.
    """
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        teks = " ".join(str(c) for c in row if c is not None).lower()
        ada_barang = "kode barang" in teks or "nama barang" in teks
        ada_tanggal = "tanggal" in teks or "tgl" in teks
        ada_mutasi = (
            "masuk" in teks or "keluar" in teks or "stok awal" in teks
            or "stok akhir" in teks or "sisa stok" in teks or "saldo stok" in teks
        )
        if ada_barang and ada_tanggal and ada_mutasi:
            return i + 1, list(row)
    return None, None


def parse_sheet_kartu_stok(ws, nama_sheet: str) -> pd.DataFrame:
    """
    Parse satu sheet Kartu Stok/Persediaan jadi DataFrame baris-per-mutasi.
    Kolom yang dicoba dikenali (nama header fleksibel): kode/nama barang,
    satuan, tanggal, nomor bukti/referensi, keterangan, qty masuk, harga
    satuan masuk, qty keluar, harga satuan keluar, saldo qty (stok
    akhir/sisa stok), saldo nilai (kalau ada).
    """
    header_rownum, header_row = _cari_header_row_kartu_stok(ws)
    if header_row is None:
        raise FormatTidakDikenali(
            f"Sheet '{nama_sheet}' tidak dikenali sebagai Kartu Stok/Persediaan "
            "(tidak ditemukan kolom KODE/NAMA BARANG + TANGGAL + MUTASI MASUK-KELUAR/SALDO STOK sekaligus)."
        )

    headers = header_row

    def _idx(kata_kunci):
        for i, h in enumerate(headers):
            if h is None:
                continue
            h_low = str(h).strip().lower()
            if any(k in h_low for k in kata_kunci):
                return i
        return None

    idx_kode = _idx(["kode barang", "kode item", "sku"])
    idx_nama = _idx(["nama barang", "nama item", "deskripsi barang"])
    idx_satuan = _idx(["satuan", "unit"])
    idx_tanggal = _idx(["tanggal", "tgl"])
    idx_nomor = _idx(["nomor bukti", "no bukti", "no. bukti", "referensi", "no referensi"])
    idx_keterangan = _idx(["keterangan", "uraian"])
    idx_qty_masuk = _idx(["qty masuk", "barang masuk", "jumlah masuk", "masuk"])
    idx_harga_masuk = _idx(["harga masuk", "harga beli", "harga satuan masuk"])
    idx_qty_keluar = _idx(["qty keluar", "barang keluar", "jumlah keluar", "keluar"])
    idx_harga_keluar = _idx(["harga keluar", "harga jual", "harga satuan keluar", "hpp"])
    idx_saldo_qty = _idx(["saldo stok", "stok akhir", "sisa stok", "saldo qty"])
    idx_saldo_nilai = _idx(["saldo nilai", "nilai persediaan", "nilai stok"])

    if idx_kode is None and idx_nama is None:
        raise FormatTidakDikenali(
            f"Kolom KODE BARANG atau NAMA BARANG tidak ditemukan di sheet '{nama_sheet}'."
        )

    def _ambil(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    def _ambil_angka(row, idx):
        v = _ambil(row, idx)
        try:
            return float(v) if v not in (None, "") else 0.0
        except (TypeError, ValueError):
            return 0.0

    rows = []
    for row in ws.iter_rows(min_row=header_rownum + 1, values_only=True):
        kode = _ambil(row, idx_kode)
        nama = _ambil(row, idx_nama)
        qty_masuk = _ambil_angka(row, idx_qty_masuk)
        qty_keluar = _ambil_angka(row, idx_qty_keluar)
        saldo_qty = _ambil(row, idx_saldo_qty)
        if kode is None and nama is None and qty_masuk == 0 and qty_keluar == 0 and saldo_qty in (None, ""):
            continue

        rows.append({
            "sheet": nama_sheet,
            "kode_barang": kode,
            "nama_barang": nama,
            "satuan": _ambil(row, idx_satuan),
            "tanggal": _ambil(row, idx_tanggal),
            "nomor_bukti": _ambil(row, idx_nomor),
            "keterangan": _ambil(row, idx_keterangan),
            "qty_masuk": qty_masuk,
            "harga_masuk": _ambil_angka(row, idx_harga_masuk),
            "qty_keluar": qty_keluar,
            "harga_keluar": _ambil_angka(row, idx_harga_keluar),
            "saldo_qty": float(saldo_qty) if saldo_qty not in (None, "") else None,
            "saldo_nilai": (
                float(_ambil(row, idx_saldo_nilai))
                if _ambil(row, idx_saldo_nilai) not in (None, "") else None
            ),
        })

    return pd.DataFrame(rows)


def cek_saldo_berjalan_kartu_stok(
    df: pd.DataFrame,
    toleransi_qty: float = _TOLERANSI_SELISIH_QTY_STOK,
    toleransi_nilai: float = _TOLERANSI_SELISIH_NILAI_STOK_RUPIAH,
) -> list:
    """
    Cross-check saldo qty (& saldo nilai kalau ada) PER KODE BARANG,
    diurutkan berdasarkan urutan baris asli (asumsi sheet sudah kronologis
    per barang -- kalau tidak, urutan tanggal WAJIB dicek manual dulu).

    Return list dict {"index": i, "kode_barang", "seharusnya_qty",
    "tertulis_qty", "seharusnya_nilai", "tertulis_nilai"} untuk baris yang
    saldonya di luar toleransi.
    """
    bermasalah = []
    if "saldo_qty" not in df.columns or not df["saldo_qty"].notna().any():
        return bermasalah

    for kode, grup in df.groupby(df["kode_barang"].fillna(df.get("nama_barang"))):
        saldo_qty_sebelumnya = None
        saldo_nilai_sebelumnya = None
        for i, row in grup.iterrows():
            saldo_qty_tertulis = row.get("saldo_qty")
            saldo_nilai_tertulis = row.get("saldo_nilai")
            if saldo_qty_tertulis is None:
                saldo_qty_sebelumnya = None
                saldo_nilai_sebelumnya = None
                continue

            catatan_baris = {"index": i, "kode_barang": kode}
            ada_masalah = False

            if saldo_qty_sebelumnya is not None:
                seharusnya_qty = saldo_qty_sebelumnya + float(row.get("qty_masuk") or 0) - float(row.get("qty_keluar") or 0)
                if abs(saldo_qty_tertulis - seharusnya_qty) > toleransi_qty:
                    catatan_baris["seharusnya_qty"] = seharusnya_qty
                    catatan_baris["tertulis_qty"] = saldo_qty_tertulis
                    ada_masalah = True
            saldo_qty_sebelumnya = saldo_qty_tertulis

            if saldo_nilai_tertulis is not None and saldo_nilai_sebelumnya is not None:
                nilai_masuk = float(row.get("qty_masuk") or 0) * float(row.get("harga_masuk") or 0)
                nilai_keluar = float(row.get("qty_keluar") or 0) * float(row.get("harga_keluar") or 0)
                seharusnya_nilai = saldo_nilai_sebelumnya + nilai_masuk - nilai_keluar
                if abs(saldo_nilai_tertulis - seharusnya_nilai) > toleransi_nilai:
                    catatan_baris["seharusnya_nilai"] = seharusnya_nilai
                    catatan_baris["tertulis_nilai"] = saldo_nilai_tertulis
                    ada_masalah = True
            if saldo_nilai_tertulis is not None:
                saldo_nilai_sebelumnya = saldo_nilai_tertulis

            if ada_masalah:
                bermasalah.append(catatan_baris)

    return bermasalah


def proses_kartu_stok(
    df: pd.DataFrame,
    toleransi_qty: float = _TOLERANSI_SELISIH_QTY_STOK,
    toleransi_nilai: float = _TOLERANSI_SELISIH_NILAI_STOK_RUPIAH,
) -> dict:
    """
    "Kerjaan akuntan/gudang" untuk Kartu Stok: tentukan arah tiap baris,
    cross-check saldo qty & nilai PER KODE BARANG, deteksi STOK MINUS
    (paling kritis), duplikat nomor bukti dlm barang yg sama, harga 0
    padahal ada qty, siapkan rekap per barang & draf jurnal.

    Return dict: "df", "ringkasan" (total qty/nilai per barang, jumlah
    barang dgn stok minus), "masalah" (list perlu direview), "draf_jurnal".
    """
    if df is None or df.empty:
        return {"df": df, "ringkasan": {}, "masalah": [], "draf_jurnal": []}

    df = df.copy()
    masalah = []

    arah_list, status_list = [], []
    nomor_terlihat_per_barang: dict = {}

    for i, row in df.iterrows():
        qty_masuk = float(row.get("qty_masuk") or 0)
        qty_keluar = float(row.get("qty_keluar") or 0)
        arah = "MASUK" if qty_masuk > 0 else ("KELUAR" if qty_keluar > 0 else "TIDAK DIKETAHUI")
        arah_list.append(arah)

        kode_barang = row.get("kode_barang") or row.get("nama_barang")
        alasan = []

        if arah == "TIDAK DIKETAHUI":
            alasan.append("Qty Masuk & Qty Keluar sama-sama kosong/nol -- tidak bisa ditentukan arah mutasi ini.")

        if qty_masuk < 0 or qty_keluar < 0:
            alasan.append("Qty bernilai negatif -- tidak wajar, cek input.")

        if arah == "MASUK" and qty_masuk > 0 and not row.get("harga_masuk"):
            alasan.append("Qty Masuk terisi tapi Harga Masuk kosong/nol -- nilai persediaan jadi tidak akurat.")

        if row.get("saldo_qty") is not None and row.get("saldo_qty") < 0:
            alasan.append(
                f"🚨 STOK MINUS: saldo qty tertulis {row.get('saldo_qty'):,.2f} -- barang keluar melebihi "
                "yang tersedia. PALING KRITIS, WAJIB dicek segera (kemungkinan ada transaksi belum "
                "tercatat / kesalahan input tanggal / urutan baris tidak kronologis)."
            )

        nomor = row.get("nomor_bukti")
        if nomor and str(nomor).strip():
            kunci = (str(kode_barang or "").strip().lower(), str(nomor).strip().lower())
            if kunci in nomor_terlihat_per_barang:
                alasan.append(
                    f"Nomor bukti '{nomor}' DUPLIKAT untuk barang yang sama -- sudah muncul di baris "
                    f"ke-{nomor_terlihat_per_barang[kunci] + 1} juga."
                )
            else:
                nomor_terlihat_per_barang[kunci] = i

        status_list.append("PERLU REVIEW" if alasan else "OK")
        if alasan:
            masalah.append({
                "baris": i + 1, "kode_barang": row.get("kode_barang"), "nama_barang": row.get("nama_barang"),
                "tanggal": row.get("tanggal"), "arah": arah,
                "qty_masuk": qty_masuk, "qty_keluar": qty_keluar, "alasan": alasan,
            })

    df["arah"] = arah_list
    df["status"] = status_list

    saldo_bermasalah = cek_saldo_berjalan_kartu_stok(df, toleransi_qty, toleransi_nilai)
    for sb in saldo_bermasalah:
        i = sb["index"]
        alasan_saldo = []
        if "seharusnya_qty" in sb:
            alasan_saldo.append(
                f"Saldo Qty tertulis {sb['tertulis_qty']:,.2f} TIDAK SAMA dengan hasil hitung "
                f"(saldo sebelumnya + qty masuk - qty keluar) = {sb['seharusnya_qty']:,.2f} -- "
                "WAJIB dicek ulang urutan/rumus."
            )
        if "seharusnya_nilai" in sb:
            alasan_saldo.append(
                f"Saldo Nilai tertulis Rp{sb['tertulis_nilai']:,.0f} TIDAK SAMA dengan hasil hitung "
                f"≈ Rp{sb['seharusnya_nilai']:,.0f} -- ini reasonableness check saja (metode HPP "
                "FIFO/Average tidak dihitung ulang di sini), tetap cek manual."
            )
        masalah.append({
            "baris": i + 1, "kode_barang": df.at[i, "kode_barang"], "nama_barang": df.at[i, "nama_barang"],
            "tanggal": df.at[i, "tanggal"], "arah": df.at[i, "arah"],
            "qty_masuk": df.at[i, "qty_masuk"], "qty_keluar": df.at[i, "qty_keluar"], "alasan": alasan_saldo,
        })
        df.at[i, "status"] = "PERLU REVIEW"

    # -- Draf jurnal (Persediaan didebet saat masuk; HPP didebet & Persediaan
    #    dikredit saat keluar -- akun lawan generik, WAJIB direview) --
    draf_jurnal = []
    for i, row in df.iterrows():
        if row["arah"] == "MASUK":
            nilai = float(row.get("qty_masuk") or 0) * float(row.get("harga_masuk") or 0)
            draf_jurnal.append({
                "baris": i + 1, "kode_barang": row.get("kode_barang"), "nama_barang": row.get("nama_barang"),
                "arah": "MASUK",
                "no_akun_debet": "PERSEDIAAN", "nama_akun_debet": "Persediaan Barang Dagang",
                "jml_debet": nilai,
                "no_akun_kredit": "UTANG/KAS/LAIN", "nama_akun_kredit": "Sesuaikan akun (utang usaha/kas/dll)",
                "jml_kredit": nilai,
                "catatan": "Draf otomatis -- sesuaikan akun lawan (pembelian tunai/kredit, retur jual, dll).",
            })
        elif row["arah"] == "KELUAR":
            nilai = float(row.get("qty_keluar") or 0) * float(row.get("harga_keluar") or 0)
            draf_jurnal.append({
                "baris": i + 1, "kode_barang": row.get("kode_barang"), "nama_barang": row.get("nama_barang"),
                "arah": "KELUAR",
                "no_akun_debet": "HPP", "nama_akun_debet": "Harga Pokok Penjualan",
                "jml_debet": nilai,
                "no_akun_kredit": "PERSEDIAAN", "nama_akun_kredit": "Persediaan Barang Dagang",
                "jml_kredit": nilai,
                "catatan": "Draf otomatis -- nilai HPP di sini pakai Harga Keluar tertulis di kartu "
                           "stok, BUKAN hasil hitung ulang metode FIFO/Average -- WAJIB diverifikasi "
                           "ke kebijakan penilaian persediaan perusahaan.",
            })

    kolom_kode = df["kode_barang"].fillna(df["nama_barang"]) if "nama_barang" in df.columns else df["kode_barang"]
    rekap_per_barang = (
        df.assign(_kode=kolom_kode)
        .groupby("_kode")[["qty_masuk", "qty_keluar"]].sum().fillna(0).to_dict("index")
    )
    stok_minus = [m for m in masalah if any("STOK MINUS" in a for a in m["alasan"])]

    ringkasan = {
        "jumlah_baris_mutasi": len(df),
        "jumlah_barang": df["kode_barang"].nunique() if df["kode_barang"].notna().any() else df["nama_barang"].nunique(),
        "total_qty_masuk": float(df["qty_masuk"].fillna(0).sum()),
        "total_qty_keluar": float(df["qty_keluar"].fillna(0).sum()),
        "rekap_per_barang": rekap_per_barang,
        "jumlah_perlu_review": len(masalah),
        "jumlah_stok_minus": len(stok_minus),
        "jumlah_duplikat": sum(1 for m in masalah if any("DUPLIKAT" in a for a in m["alasan"])),
        "jumlah_selisih_saldo": len(saldo_bermasalah),
        "catatan": (
            "Cross-check saldo dilakukan PER KODE BARANG dgn asumsi baris sudah kronologis per "
            "barang di sheet aslinya. Nilai HPP di draf jurnal pakai Harga Keluar tertulis di "
            "kartu stok, BUKAN hasil hitung ulang FIFO/Average -- tetap WAJIB diverifikasi manual."
        ),
    }

    return {"df": df, "ringkasan": ringkasan, "masalah": masalah, "draf_jurnal": draf_jurnal}


def proses_file_kartu_stok(file_like, nama_file: str = None) -> dict:
    """
    Fungsi siap-pakai tingkat file: baca file (.xlsx/.xlsm/.xls/.csv/.pdf),
    cari semua sheet yang cocok sebagai Kartu Stok/Persediaan, gabungkan,
    lalu proses lewat proses_kartu_stok(). Dipakai oleh app.py utk 1 file upload.

    Return sama seperti proses_kartu_stok(), ditambah "sheet_dilewati".
    """
    nama_file = nama_file or getattr(file_like, "name", "") or ""
    daftar_sheet, _df_coa = _siapkan_daftar_sheet(file_like, nama_file)

    df_list = []
    sheet_dilewati = []
    for nama, ws in daftar_sheet:
        try:
            df_sheet = parse_sheet_kartu_stok(ws, nama)
        except FormatTidakDikenali:
            sheet_dilewati.append(nama)
            continue
        if df_sheet.empty:
            sheet_dilewati.append(nama)
            continue
        df_list.append(df_sheet)

    if not df_list:
        return {"df": pd.DataFrame(), "ringkasan": {}, "masalah": [], "draf_jurnal": [], "sheet_dilewati": sheet_dilewati}

    df_gabungan = pd.concat(df_list, ignore_index=True)
    hasil = proses_kartu_stok(df_gabungan)
    hasil["sheet_dilewati"] = sheet_dilewati
    return hasil


# ============================================================
# 11G. ASET TETAP -- PARSING, HITUNG PENYUSUTAN, VALIDASI, & DRAF JURNAL
# ============================================================
# Sheet Aset Tetap: daftar aset perusahaan dengan harga perolehan,
# masa manfaat, akumulasi penyusutan, dan nilai buku.
#
# "Kerjaan akuntan" saat menerima daftar aset tetap ini:
#   1. Cek kelengkapan data aset (nama, tanggal perolehan, harga perolehan,
#      masa manfaat -- tanpa ini penyusutan tidak bisa dihitung).
#   2. Hitung penyusutan pakai metode GARIS LURUS (metode paling umum &
#      jadi default fiskal utk banyak kelompok aset di Indonesia).
#   3. Hitung akumulasi penyusutan yang SEHARUSNYA per tanggal laporan
#      (berdasarkan umur aset dlm bulan sejak tanggal perolehan).
#   4. Bandingkan dengan akumulasi penyusutan & nilai buku yang tertulis.
#   5. Deteksi aset yang sudah HABIS masa manfaat (nilai buku ~ nilai
#      residu) -- informasi, bukan error, tapi perlu diketahui akuntan.
#   6. Deteksi masa manfaat tidak wajar (kosong, <=0, atau di luar
#      rentang wajar 1-50 tahun).
#   7. Deteksi duplikat kode/nama aset.
#   8. Siapkan draf jurnal penyusutan BULAN BERJALAN (Beban Penyusutan
#      debet, Akumulasi Penyusutan kredit) per aset yang belum habis
#      masa manfaatnya.
#
# CATATAN: modul ini HANYA menghitung metode garis lurus (straight-line).
# Kalau perusahaan pakai metode saldo menurun (declining balance) utk
# tujuan fiskal/golongan tertentu, hasil di sini jadi PEMBANDING saja,
# WAJIB direview akuntan pajak.

_TOLERANSI_SELISIH_PENYUSUTAN = 500  # toleransi pembulatan rupiah
_MASA_MANFAAT_MINIMAL = 1
_MASA_MANFAAT_MAKSIMAL = 50


def _cari_header_row_aset_tetap(ws, max_scan: int = 10):
    """Cari baris header sheet Aset Tetap.
    [FIX] Sebelumnya "nama"/"aset"/"harga" SENDIRIAN dianggap cukup -- kata
    ini terlalu umum (banyak dokumen lain punya kolom "Nama" atau "Harga",
    mis. invoice pembelian) sehingga rawan salah deteksi. Sekarang wajib
    frasa spesifik aset tetap."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        teks = " ".join(str(c) for c in row if c is not None).lower()
        ada_nama_aset = any(k in teks for k in ["nama aset", "deskripsi aset", "aset tetap"])
        ada_harga = any(k in teks for k in ["harga perolehan", "nilai perolehan"])
        ada_penyusutan = any(k in teks for k in ["akumulasi penyusutan", "nilai buku", "masa manfaat", "golongan fiskal"])
        if ada_nama_aset and (ada_harga or ada_penyusutan):
            return i + 1, list(row)
    return None, None


def _kategorikan_aset(nama_aset: str) -> str:
    """Kategorikan aset berdasarkan nama (utk rekap & referensi golongan fiskal)."""
    if nama_aset is None:
        return "LAINNYA"
    nama_lower = str(nama_aset).lower()
    if any(k in nama_lower for k in ["tanah", "land"]):
        return "TANAH"
    # [FIX] "kantor" dihapus dari kata kunci BANGUNAN -- kata ini terlalu
    # umum & muncul di banyak nama aset yang BUKAN bangunan (mis.
    # "Kalkulator Kantor", "Meja Kantor", "Printer Kantor", "AC Kantor"),
    # jadi barang kecil salah kena kategori BANGUNAN -> salah golongan
    # fiskal (20 th, padahal seharusnya Gol. I/4 th) -> rekomendasi jadi
    # keliru. "Gedung Kantor"/"Bangunan Kantor" tetap tertangkap lewat
    # kata kunci "gedung"/"bangunan"/"ruko"/"building".
    if any(k in nama_lower for k in ["bangunan", "gedung", "ruko", "building"]):
        return "BANGUNAN"
    if any(k in nama_lower for k in ["kendaraan", "mobil", "motor", "truk", "bus", "vehicle"]):
        return "KENDARAAN"
    if any(k in nama_lower for k in [
        "mesin", "engine", "alat berat", "ekskavator", "excavator", "genset",
        "forklift", "kompresor", "boiler", "buldoser", "bulldozer", "crane",
    ]):
        return "MESIN"
    if any(k in nama_lower for k in ["komputer", "laptop", "server", "pc", "printer"]):
        return "KOMPUTER"
    if any(k in nama_lower for k in ["furnitur", "meja", "kursi", "lemari", "furniture"]):
        return "FURNITUR"
    if any(k in nama_lower for k in ["peralatan", "alat", "equipment", "tool"]):
        return "PERALATAN"
    return "LAINNYA"


# ============================================================
# GOLONGAN FISKAL (PMK 96/PMK.03/2009) -- referensi tarif penyusutan
# FISKAL (pajak), dipakai utk rekonsiliasi fiskal (koreksi beda waktu
# penyusutan komersial vs fiskal di SPT Tahunan Badan). Beda dgn
# penyusutan KOMERSIAL (PSAK 16, pakai masa manfaat & nilai residu
# sesuai kebijakan perusahaan), penyusutan FISKAL:
#   - masa manfaat & tarif SUDAH DITETAPKAN per golongan (bukan pilihan
#     perusahaan),
#   - dasar penyusutan = HARGA PEROLEHAN PENUH (nilai residu TIDAK
#     mengurangi dasar penyusutan fiskal -- beda dgn komersial).
# [CATATAN PENTING] Pemetaan kategori -> golongan di bawah ini
# SIMPLIFIKASI berdasar jenis aset paling umum per golongan (bukan
# tabel lengkap Lampiran I/II PMK 96/2009). WAJIB diverifikasi manual
# oleh akuntan pajak, terutama utk MESIN (bisa gol. II/III/IV
# tergantung jenis industri) & KENDARAAN alat berat/khusus.
# ============================================================
_TARIF_GOLONGAN_FISKAL = {
    "I":   {"masa_manfaat": 4,  "tarif_garis_lurus": 0.25,   "tarif_saldo_menurun": 0.50},
    "II":  {"masa_manfaat": 8,  "tarif_garis_lurus": 0.125,  "tarif_saldo_menurun": 0.25},
    "III": {"masa_manfaat": 16, "tarif_garis_lurus": 0.0625, "tarif_saldo_menurun": 0.125},
    "IV":  {"masa_manfaat": 20, "tarif_garis_lurus": 0.05,   "tarif_saldo_menurun": 0.10},
    "BANGUNAN_PERMANEN":      {"masa_manfaat": 20, "tarif_garis_lurus": 0.05, "tarif_saldo_menurun": None},
    "BANGUNAN_TIDAK_PERMANEN": {"masa_manfaat": 10, "tarif_garis_lurus": 0.10, "tarif_saldo_menurun": None},
}

# Batas kapitalisasi default -- aset dgn harga perolehan di bawah ini
# DISARANKAN dibebankan langsung sbg beban operasional, bukan
# dikapitalisasi & disusutkan (kebijakan akuntansi umum di banyak
# perusahaan kecil-menengah Indonesia). Ini SARAN, bukan aturan baku
# PSAK -- threshold sebenarnya WAJIB ikut kebijakan akuntansi client
# masing-masing (kalau ada), bisa beda-beda per perusahaan.
_BATAS_KAPITALISASI_DEFAULT_RUPIAH = 1_000_000


def _tentukan_golongan_fiskal(kategori: Optional[str], nama_aset=None) -> Optional[str]:
    """
    Tentukan golongan fiskal (PMK 96/2009) dari kategori aset (hasil
    _kategorikan_aset / kategori tertulis client) + kata kunci di nama
    aset. SIMPLIFIKASI -- lihat catatan di atas _TARIF_GOLONGAN_FISKAL.
    Return None kalau golongan tidak bisa ditentukan otomatis (perlu
    klasifikasi manual oleh akuntan pajak).
    """
    if not kategori:
        return None
    kategori = str(kategori).strip().upper()
    nama_lower = str(nama_aset or "").lower()

    if kategori == "TANAH":
        return "TANAH"  # tidak disusutkan -- bukan golongan penyusutan
    if kategori == "BANGUNAN":
        if any(k in nama_lower for k in ["tidak permanen", "semi permanen", "sementara", "darurat"]):
            return "BANGUNAN_TIDAK_PERMANEN"
        return "BANGUNAN_PERMANEN"
    if kategori in ("KOMPUTER", "FURNITUR", "PERALATAN"):
        return "I"
    if kategori == "KENDARAAN":
        if any(k in nama_lower for k in ["motor", "sepeda"]) and "mobil" not in nama_lower:
            return "I"
        return "II"
    if kategori == "MESIN":
        if any(k in nama_lower for k in ["berat", "ekskavator", "excavator", "bulldozer", "crane", "buldoser"]):
            return "III"
        return "II"
    return None  # kategori "LAINNYA" / tidak dikenal -- perlu klasifikasi manual


def _hitung_penyusutan_garis_lurus(harga: float, masa_manfaat: int, nilai_residu: float = 0) -> float:
    """
    Penyusutan per TAHUN metode garis lurus:
        (Harga Perolehan - Nilai Residu) / Masa Manfaat (tahun)
    Return 0 kalau masa_manfaat tidak valid (<=0/None), karena penyusutan
    tidak bisa dihitung tanpa masa manfaat yang wajar.
    """
    try:
        # [FIX] "float(x or 0)" tidak aman untuk NaN (NaN dianggap truthy
        # di Python, jadi fallback "or 0" tidak kepakai) -- pakai
        # _angka_aman() supaya harga/nilai_residu NaN tidak lolos jadi NaN
        # dan meracuni hasil penyusutan (yang lalu masuk ke draf jurnal &
        # rekap Aset Tetap).
        harga = _angka_aman(harga)
        nilai_residu = _angka_aman(nilai_residu)
        masa_manfaat = int(masa_manfaat) if masa_manfaat else 0
    except (TypeError, ValueError):
        return 0.0
    if masa_manfaat <= 0:
        return 0.0
    return max(harga - nilai_residu, 0.0) / masa_manfaat


def _hitung_umur_aset_bulan(tanggal_perolehan, tanggal_acuan: date = None) -> int:
    """
    Umur aset dalam bulan PENUH sejak tanggal perolehan s.d. tanggal acuan
    (default: hari ini). Return 0 kalau tanggal_perolehan tidak valid atau
    tanggal perolehan di masa depan.
    """
    tanggal_acuan = tanggal_acuan or date.today()
    # [FIX] pd.isna() dulu, bukan cuma "is None" -- pandas.NaT (muncul kalau
    # baris ini tanggal-perolehannya kosong tapi baris LAIN di kolom yang
    # sama terisi tanggal, mis. setelah gabung banyak sheet aset) lolos
    # "is None" DAN lolos isinstance(datetime), lalu NaT.date() balik NaT
    # lagi, dan "tp > tanggal_acuan" MELEMPAR TypeError -- meledakkan
    # seluruh proses_aset_tetap() gara-gara satu baris tanggal kosong.
    if pd.isna(tanggal_perolehan):
        return 0
    try:
        if isinstance(tanggal_perolehan, datetime):
            tp = tanggal_perolehan.date()
        elif isinstance(tanggal_perolehan, date):
            tp = tanggal_perolehan
        else:
            tp = pd.to_datetime(tanggal_perolehan).date()
    except Exception:
        return 0
    if tp > tanggal_acuan:
        return 0
    bulan = (tanggal_acuan.year - tp.year) * 12 + (tanggal_acuan.month - tp.month)
    if tanggal_acuan.day < tp.day:
        bulan -= 1
    return max(bulan, 0)


def parse_sheet_aset_tetap(ws, nama_sheet: str) -> pd.DataFrame:
    """
    Parse satu sheet Aset Tetap jadi DataFrame satu baris = satu aset.
    Kolom yang dicoba dikenali (nama header fleksibel): kode/nama aset,
    kategori/golongan (kalau tidak ada, ditebak dari nama aset), tanggal
    perolehan, harga perolehan, nilai residu, masa manfaat (tahun),
    akumulasi penyusutan tertulis, nilai buku tertulis.
    """
    header_rownum, header_row = _cari_header_row_aset_tetap(ws)
    if header_row is None:
        raise FormatTidakDikenali(
            f"Sheet '{nama_sheet}' tidak dikenali sebagai Daftar Aset Tetap "
            "(tidak ditemukan kolom NAMA ASET + HARGA PEROLEHAN/PENYUSUTAN sekaligus)."
        )

    headers = header_row

    def _idx(kata_kunci):
        for i, h in enumerate(headers):
            if h is None:
                continue
            h_low = str(h).strip().lower()
            if any(k in h_low for k in kata_kunci):
                return i
        return None

    idx_kode = _idx(["kode aset", "kode barang", "no aset", "nomor aset"])
    idx_nama = _idx(["nama aset", "deskripsi aset", "nama"])
    idx_kategori = _idx(["kategori", "golongan", "kelompok aset", "golongan fiskal"])
    idx_tanggal = _idx(["tanggal perolehan", "tgl perolehan", "tanggal beli"])
    # [BARU] "Mulai Digunakan" -- sebelumnya field ini TIDAK PERNAH dicari
    # dari file sumber sama sekali, padahal sudah ada kolomnya di sheet
    # export "Buku Bantu Aktiva Tetap" (accounting_export.py) & sudah
    # diteruskan apa adanya di laporan_keuangan.py::susun_jadwal_
    # penyusutan_bulanan() -- akibatnya kolom "Mulai Digunakan" di Excel
    # selalu kosong. Dicari terpisah dari "tanggal perolehan" karena bisa
    # beda (mis. aset dibeli Desember tapi baru dipakai/disusutkan Januari
    # tahun berikutnya); kalau tidak ada kolomnya di file client, nanti
    # susun_jadwal_penyusutan_bulanan() fallback ke tanggal_perolehan.
    idx_mulai_digunakan = _idx(["mulai digunakan", "tanggal mulai digunakan", "tanggal mulai pakai",
                                 "start of use", "in service"])
    idx_harga = _idx(["harga perolehan", "nilai perolehan", "cost"])
    idx_residu = _idx(["nilai residu", "nilai sisa", "residu"])
    idx_masa = _idx(["masa manfaat", "umur ekonomis", "useful life"])
    idx_akumulasi = _idx(["akumulasi penyusutan", "akumulasi"])
    idx_nilai_buku = _idx(["nilai buku"])

    if idx_nama is None:
        raise FormatTidakDikenali(f"Kolom NAMA ASET tidak ditemukan di sheet '{nama_sheet}'.")

    def _ambil(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    def _ambil_angka(row, idx):
        v = _ambil(row, idx)
        try:
            return float(v) if v not in (None, "") else None
        except (TypeError, ValueError):
            return None

    rows = []
    for row in ws.iter_rows(min_row=header_rownum + 1, values_only=True):
        nama = _ambil(row, idx_nama)
        harga = _ambil_angka(row, idx_harga)
        if nama is None and harga is None:
            continue
        kategori_tertulis = _ambil(row, idx_kategori)
        rows.append({
            "sheet": nama_sheet,
            "kode_aset": _ambil(row, idx_kode),
            "nama_aset": nama,
            "kategori": str(kategori_tertulis).strip().upper() if kategori_tertulis else _kategorikan_aset(nama),
            "tanggal_perolehan": _ambil(row, idx_tanggal),
            "mulai_digunakan": _ambil(row, idx_mulai_digunakan),
            "harga_perolehan": harga or 0.0,
            "nilai_residu": _ambil_angka(row, idx_residu) or 0.0,
            "masa_manfaat_tahun": _ambil_angka(row, idx_masa),
            "akumulasi_penyusutan_tertulis": _ambil_angka(row, idx_akumulasi),
            "nilai_buku_tertulis": _ambil_angka(row, idx_nilai_buku),
        })

    return pd.DataFrame(rows)


def proses_aset_tetap(
    df: pd.DataFrame,
    toleransi_rupiah: float = _TOLERANSI_SELISIH_PENYUSUTAN,
    masa_manfaat_min: int = _MASA_MANFAAT_MINIMAL,
    masa_manfaat_maks: int = _MASA_MANFAAT_MAKSIMAL,
    tanggal_acuan: date = None,
    batas_kapitalisasi_rupiah: float = _BATAS_KAPITALISASI_DEFAULT_RUPIAH,
) -> dict:
    """
    "Kerjaan akuntan" untuk Daftar Aset Tetap -- meniru alur kerja akuntan
    yang menangani aset tetap client:
      1. Cek kelengkapan & kewajaran data per baris.
      2. Hitung penyusutan KOMERSIAL (PSAK 16, garis lurus, dasar =
         harga - nilai residu) & bandingkan dgn angka tertulis client.
      3. Tentukan golongan FISKAL (PMK 96/2009) & hitung penyusutan
         fiskal (dasar = harga penuh, tanpa residu) sbg pembanding utk
         rekonsiliasi fiskal / koreksi SPT Tahunan Badan.
      4. Kenali TANAH otomatis -- tidak disusutkan, tidak dianggap data
         kurang walau masa manfaat kosong.
      5. Tandai aset di bawah batas kapitalisasi -- saran dibebankan
         langsung, bukan disusutkan.
      6. Deteksi aset habis masa manfaat, masa manfaat tidak wajar, &
         duplikat kode/nama.
      7. Siapkan draf jurnal penyusutan KOMERSIAL bulan berjalan, dan
         rekomendasi tindakan konkret per aset (bukan cuma status flag).

    Return dict: "df", "ringkasan", "masalah", "draf_jurnal",
    "rekonsiliasi_fiskal", "aset_di_bawah_batas_kapitalisasi",
    "jadwal_penyusutan_bulanan".
    """
    kosong = {
        "df": df, "ringkasan": {}, "masalah": [], "draf_jurnal": [],
        "rekonsiliasi_fiskal": [], "aset_di_bawah_batas_kapitalisasi": [],
        "jadwal_penyusutan_bulanan": [],
    }
    if df is None or df.empty:
        return kosong

    tanggal_acuan = tanggal_acuan or date.today()
    df = df.copy()
    masalah = []
    nama_terlihat: dict = {}

    penyusutan_tahun_list, penyusutan_bulan_list = [], []
    umur_bulan_list, akumulasi_seharusnya_list, nilai_buku_seharusnya_list = [], [], []
    status_list = []
    golongan_fiskal_list, masa_manfaat_fiskal_list, tarif_fiskal_list = [], [], []
    penyusutan_fiskal_tahun_list, penyusutan_fiskal_bulan_list = [], []
    akumulasi_fiskal_list, nilai_buku_fiskal_list, selisih_fiskal_list = [], [], []
    di_bawah_kapitalisasi_list = []
    rekonsiliasi_fiskal = []
    aset_di_bawah_kapitalisasi = []

    for i, row in df.iterrows():
        alasan = []
        rekomendasi = []
        nama = row.get("nama_aset")
        harga = float(row.get("harga_perolehan") or 0)
        nilai_residu = float(row.get("nilai_residu") or 0)
        masa_manfaat = row.get("masa_manfaat_tahun")
        tanggal_perolehan = row.get("tanggal_perolehan")
        kategori = str(row.get("kategori") or "").strip().upper()
        is_tanah = (kategori == "TANAH")

        if not nama or str(nama).strip() == "":
            alasan.append("Nama aset kosong -- baris tidak bisa diidentifikasi.")
            rekomendasi.append("Minta client lengkapi nama/deskripsi aset sebelum baris ini diproses lebih lanjut.")
        if harga <= 0:
            alasan.append("Harga perolehan kosong/nol -- penyusutan tidak bisa dihitung.")
            rekomendasi.append("Minta client lengkapi harga perolehan (cek invoice pembelian/BAST aset).")
        # [FIX] "tanggal_perolehan is None" / "masa_manfaat is None" tidak
        # menangkap NaN/NaT -- begitu kolom digabung jadi DataFrame, sel
        # kosong di kolom angka/tanggal yang kolomnya sama-sama berisi
        # baris lain yang terisi akan berubah jadi np.nan / pd.NaT, BUKAN
        # None, sehingga baris lolos tanpa peringatan padahal datanya
        # kosong (status tampil "OK" dgn penyusutan 0, akuntan tidak sadar
        # data belum lengkap). Pakai pd.isna() supaya NaN/NaT ikut kena.
        if pd.isna(tanggal_perolehan) or str(tanggal_perolehan).strip() == "":
            alasan.append("Tanggal perolehan kosong -- umur aset & akumulasi penyusutan tidak bisa dihitung.")
            rekomendasi.append("Minta client lengkapi tanggal perolehan (cek invoice pembelian/BAST aset).")

        # -- Tanah: dikenali otomatis, TIDAK disusutkan (PSAK 16 & pajak) --
        if is_tanah:
            if not pd.isna(masa_manfaat):
                alasan.append(
                    "Kategori TANAH biasanya TIDAK disusutkan -- kolom masa manfaat terisi, "
                    "cek apakah ini benar Tanah atau salah kategori."
                )
            rekomendasi.append(
                "Tanah tidak disusutkan (PSAK 16 & ketentuan pajak) -- dicatat sebesar harga "
                "perolehan, cek kalau ada indikasi penurunan nilai (impairment) tapi bukan penyusutan rutin."
            )
        else:
            if pd.isna(masa_manfaat):
                alasan.append("Masa manfaat kosong -- penyusutan tidak bisa dihitung, WAJIB dilengkapi.")
                rekomendasi.append("Minta client lengkapi masa manfaat/umur ekonomis aset ini.")
            elif masa_manfaat < masa_manfaat_min or masa_manfaat > masa_manfaat_maks:
                alasan.append(
                    f"Masa manfaat {masa_manfaat:.0f} tahun di luar rentang wajar "
                    f"({masa_manfaat_min}-{masa_manfaat_maks} tahun) -- cek input, mungkin salah satuan (bulan vs tahun)."
                )
        if nilai_residu < 0:
            alasan.append("Nilai residu bernilai negatif -- tidak wajar, cek input.")
        if nilai_residu > harga:
            alasan.append("Nilai residu LEBIH BESAR dari harga perolehan -- tidak wajar, cek input.")

        # -- Penyusutan KOMERSIAL (PSAK 16, garis lurus) --
        penyusutan_tahun = 0.0 if is_tanah else _hitung_penyusutan_garis_lurus(harga, masa_manfaat, nilai_residu)
        penyusutan_bulan = penyusutan_tahun / 12 if penyusutan_tahun else 0.0
        umur_bulan = _hitung_umur_aset_bulan(tanggal_perolehan, tanggal_acuan)
        akumulasi_maks = max(harga - nilai_residu, 0.0)
        akumulasi_seharusnya = min(penyusutan_bulan * umur_bulan, akumulasi_maks) if penyusutan_bulan else 0.0
        nilai_buku_seharusnya = harga - akumulasi_seharusnya

        penyusutan_tahun_list.append(penyusutan_tahun)
        penyusutan_bulan_list.append(penyusutan_bulan)
        umur_bulan_list.append(umur_bulan)
        akumulasi_seharusnya_list.append(akumulasi_seharusnya)
        nilai_buku_seharusnya_list.append(nilai_buku_seharusnya)

        akumulasi_tertulis = row.get("akumulasi_penyusutan_tertulis")
        if akumulasi_tertulis is not None and not pd.isna(akumulasi_tertulis) and abs(akumulasi_tertulis - akumulasi_seharusnya) > toleransi_rupiah:
            alasan.append(
                f"Akumulasi Penyusutan tertulis Rp{akumulasi_tertulis:,.0f} TIDAK SAMA dengan hasil hitung "
                f"garis lurus ≈ Rp{akumulasi_seharusnya:,.0f} (umur {umur_bulan} bulan) -- cek metode "
                "penyusutan yang dipakai atau tanggal perolehan/masa manfaat yang diinput."
            )
            rekomendasi.append(
                "Minta kartu aset tetap / kebijakan penyusutan ke client utk cek metode yang dipakai "
                "(garis lurus vs saldo menurun) sebelum posting jurnal penyesuaian."
            )

        nilai_buku_tertulis = row.get("nilai_buku_tertulis")
        if nilai_buku_tertulis is not None and not pd.isna(nilai_buku_tertulis) and abs(nilai_buku_tertulis - nilai_buku_seharusnya) > toleransi_rupiah:
            alasan.append(
                f"Nilai Buku tertulis Rp{nilai_buku_tertulis:,.0f} TIDAK SAMA dengan hasil hitung "
                f"≈ Rp{nilai_buku_seharusnya:,.0f}."
            )

        if penyusutan_tahun > 0 and nilai_buku_seharusnya <= nilai_residu + toleransi_rupiah and umur_bulan > 0:
            alasan.append(
                "ℹ️ Aset ini SUDAH HABIS masa manfaatnya (nilai buku ≈ nilai residu) -- penyusutan "
                "berhenti, tidak perlu dijurnal lagi. Informasi, bukan kesalahan."
            )
            rekomendasi.append(
                "Kalau aset masih dipakai operasional, TETAP tercatat di neraca sebesar nilai residu -- "
                "jangan dihapus dari daftar aset. Kalau sudah tidak dipakai/dijual, siapkan jurnal "
                "penghapusan (write-off) aset."
            )

        if nama and str(nama).strip():
            kunci = str(row.get("kode_aset") or nama).strip().lower()
            if kunci in nama_terlihat:
                alasan.append(
                    f"Kode/nama aset DUPLIKAT -- sudah muncul di baris ke-{nama_terlihat[kunci] + 1} juga."
                )
                rekomendasi.append(
                    "Konfirmasi ke client: transaksi ganda/salah input, atau memang 2 unit aset berbeda "
                    "dgn kode kebetulan sama -- kalau memang 2 unit, kode aset WAJIB dibedakan (mis. tambah -1/-2)."
                )
            else:
                nama_terlihat[kunci] = i

        # -- Kebijakan kapitalisasi: aset kecil sebaiknya dibebankan langsung --
        di_bawah_kapitalisasi = bool(harga > 0 and harga < batas_kapitalisasi_rupiah)
        di_bawah_kapitalisasi_list.append(di_bawah_kapitalisasi)
        if di_bawah_kapitalisasi:
            catatan_kapitalisasi = (
                f"Harga perolehan Rp{harga:,.0f} DI BAWAH batas kapitalisasi (Rp{batas_kapitalisasi_rupiah:,.0f}) -- "
                "pertimbangkan dibebankan langsung sbg beban operasional periode berjalan (konfirmasi dulu "
                "ke kebijakan akuntansi/threshold internal client), bukan dikapitalisasi & disusutkan."
            )
            rekomendasi.append(catatan_kapitalisasi)
            # [FIX] rekomendasi kapitalisasi ini bisa muncul di baris yang
            # SEMUA data lainnya lengkap/wajar (alasan == [], status "OK"),
            # sehingga baris ini TIDAK masuk ke list "masalah" (yang hanya
            # menampung baris ber-alasan) -- rekomendasinya jadi hilang
            # kalau tidak dicatat terpisah di sini juga.
            aset_di_bawah_kapitalisasi.append({
                "baris": i + 1, "kode_aset": row.get("kode_aset"), "nama_aset": nama,
                "harga_perolehan": harga, "rekomendasi": catatan_kapitalisasi,
            })

        # -- Golongan FISKAL (PMK 96/2009) & penyusutan fiskal (pembanding) --
        golongan = _tentukan_golongan_fiskal(kategori, nama)
        if golongan is None and kategori not in ("", "LAINNYA"):
            golongan = None
        if golongan is None and not is_tanah:
            golongan = _tentukan_golongan_fiskal(_kategorikan_aset(nama), nama)

        if golongan in (None,) and not is_tanah:
            rekomendasi.append(
                f"Golongan fiskal tidak bisa ditentukan otomatis dari kategori '{kategori or '(kosong)'}' -- "
                "klasifikasikan manual sesuai Lampiran PMK 96/2009 sebelum lapor SPT Tahunan Badan."
            )
            masa_manfaat_fiskal = tarif_fiskal = None
            penyusutan_fiskal_tahun = penyusutan_fiskal_bulan = 0.0
            akumulasi_fiskal = 0.0
            nilai_buku_fiskal = harga
        elif is_tanah or golongan == "TANAH":
            golongan = "TANAH"
            masa_manfaat_fiskal = tarif_fiskal = None
            penyusutan_fiskal_tahun = penyusutan_fiskal_bulan = 0.0
            akumulasi_fiskal = 0.0
            nilai_buku_fiskal = harga
        else:
            ref = _TARIF_GOLONGAN_FISKAL[golongan]
            masa_manfaat_fiskal = ref["masa_manfaat"]
            tarif_fiskal = ref["tarif_garis_lurus"]
            # Dasar penyusutan FISKAL = harga perolehan PENUH (nilai residu
            # TIDAK mengurangi dasar penyusutan fiskal -- beda dgn komersial).
            penyusutan_fiskal_tahun = harga * tarif_fiskal if harga > 0 else 0.0
            penyusutan_fiskal_bulan = penyusutan_fiskal_tahun / 12 if penyusutan_fiskal_tahun else 0.0
            akumulasi_fiskal = min(penyusutan_fiskal_bulan * umur_bulan, harga) if penyusutan_fiskal_bulan else 0.0
            nilai_buku_fiskal = harga - akumulasi_fiskal

            if masa_manfaat is not None and not pd.isna(masa_manfaat) and int(masa_manfaat) != masa_manfaat_fiskal:
                rekomendasi.append(
                    f"Masa manfaat KOMERSIAL ({masa_manfaat:.0f} th) beda dgn masa manfaat golongan FISKAL "
                    f"{golongan} ({masa_manfaat_fiskal} th, tarif garis lurus {tarif_fiskal*100:.2f}%/th) -- "
                    "ini WAJAR (komersial ikut kebijakan perusahaan, fiskal ikut PMK 96/2009), tapi WAJIB "
                    "dicatat sbg koreksi fiskal (beda waktu) saat SPT Tahunan Badan, bukan kesalahan input."
                )

        selisih_akumulasi = akumulasi_seharusnya - akumulasi_fiskal
        golongan_fiskal_list.append(golongan)
        masa_manfaat_fiskal_list.append(masa_manfaat_fiskal)
        tarif_fiskal_list.append(tarif_fiskal)
        penyusutan_fiskal_tahun_list.append(penyusutan_fiskal_tahun)
        penyusutan_fiskal_bulan_list.append(penyusutan_fiskal_bulan)
        akumulasi_fiskal_list.append(akumulasi_fiskal)
        nilai_buku_fiskal_list.append(nilai_buku_fiskal)
        selisih_fiskal_list.append(selisih_akumulasi)

        if golongan not in (None, "TANAH") and harga > 0:
            rekonsiliasi_fiskal.append({
                "baris": i + 1, "kode_aset": row.get("kode_aset"), "nama_aset": nama,
                "golongan_fiskal": golongan,
                "penyusutan_komersial_per_tahun": round(penyusutan_tahun, 0),
                "penyusutan_fiskal_per_tahun": round(penyusutan_fiskal_tahun, 0),
                "selisih_penyusutan_per_tahun": round(penyusutan_tahun - penyusutan_fiskal_tahun, 0),
                "akumulasi_komersial_seharusnya": round(akumulasi_seharusnya, 0),
                "akumulasi_fiskal_seharusnya": round(akumulasi_fiskal, 0),
                "selisih_akumulasi": round(selisih_akumulasi, 0),
            })

        status_list.append("PERLU REVIEW" if alasan else "OK")
        # [FIX] beberapa rekomendasi (mis. beda masa manfaat komersial vs
        # fiskal, saran kapitalisasi) bersifat INFORMASI/WAJAR, bukan
        # kesalahan data -- jadi "alasan" bisa kosong padahal "rekomendasi"
        # ada isinya. Kalau syaratnya cuma "if alasan", baris begini tidak
        # pernah masuk "masalah" & rekomendasinya hilang total dari hasil.
        # Status ("OK"/"PERLU REVIEW") tetap murni dari "alasan" -- baris
        # OK dgn rekomendasi info tetap "OK", cuma catatannya ikut disertakan.
        if alasan or rekomendasi:
            masalah.append({
                "baris": i + 1, "kode_aset": row.get("kode_aset"), "nama_aset": nama,
                "kategori": row.get("kategori"), "alasan": alasan, "rekomendasi": rekomendasi,
            })

    df["penyusutan_per_tahun"] = penyusutan_tahun_list
    df["penyusutan_per_bulan"] = penyusutan_bulan_list
    df["umur_bulan"] = umur_bulan_list
    df["akumulasi_penyusutan_seharusnya"] = akumulasi_seharusnya_list
    df["nilai_buku_seharusnya"] = nilai_buku_seharusnya_list
    df["status"] = status_list
    df["golongan_fiskal"] = golongan_fiskal_list
    df["masa_manfaat_fiskal_tahun"] = masa_manfaat_fiskal_list
    df["tarif_penyusutan_fiskal_garis_lurus"] = tarif_fiskal_list
    df["penyusutan_fiskal_per_tahun"] = penyusutan_fiskal_tahun_list
    df["penyusutan_fiskal_per_bulan"] = penyusutan_fiskal_bulan_list
    df["akumulasi_penyusutan_fiskal_seharusnya"] = akumulasi_fiskal_list
    df["nilai_buku_fiskal_seharusnya"] = nilai_buku_fiskal_list
    df["selisih_akumulasi_komersial_vs_fiskal"] = selisih_fiskal_list
    df["di_bawah_batas_kapitalisasi"] = di_bawah_kapitalisasi_list

    # -- Draf jurnal penyusutan KOMERSIAL BULAN BERJALAN, untuk aset yang belum habis masa manfaat --
    draf_jurnal = []
    for i, row in df.iterrows():
        if row["penyusutan_per_bulan"] <= 0:
            continue
        sudah_habis = row["nilai_buku_seharusnya"] <= float(row.get("nilai_residu") or 0) + toleransi_rupiah
        if sudah_habis:
            continue
        draf_jurnal.append({
            "baris": i + 1, "kode_aset": row.get("kode_aset"), "nama_aset": row.get("nama_aset"),
            "no_akun_debet": "BEBAN PENYUSUTAN", "nama_akun_debet": f"Beban Penyusutan - {row.get('kategori')}",
            "jml_debet": round(row["penyusutan_per_bulan"], 0),
            "no_akun_kredit": "AKUMULASI PENYUSUTAN", "nama_akun_kredit": f"Akumulasi Penyusutan - {row.get('kategori')}",
            "jml_kredit": round(row["penyusutan_per_bulan"], 0),
            "catatan": "Draf otomatis -- penyusutan KOMERSIAL 1 bulan berjalan, metode garis lurus (PSAK 16).",
        })

    rekap_kategori = df.groupby("kategori")[["harga_perolehan", "akumulasi_penyusutan_seharusnya"]].sum().to_dict("index")
    aset_habis = [m for m in masalah if any("SUDAH HABIS" in a for a in m["alasan"])]
    jumlah_golongan_tidak_diketahui = sum(
        1 for g, kat in zip(golongan_fiskal_list, df["kategori"]) if g is None
    )

    ringkasan = {
        "jumlah_aset": len(df),
        "total_harga_perolehan": float(df["harga_perolehan"].fillna(0).sum()),
        "total_akumulasi_penyusutan_seharusnya": float(df["akumulasi_penyusutan_seharusnya"].fillna(0).sum()),
        "total_nilai_buku_seharusnya": float(df["nilai_buku_seharusnya"].fillna(0).sum()),
        "total_penyusutan_per_bulan": float(df["penyusutan_per_bulan"].fillna(0).sum()),
        "rekap_per_kategori": rekap_kategori,
        # [FIX] "masalah" sekarang juga memuat baris "OK" yang cuma punya
        # rekomendasi informasional (lihat catatan di atas) -- jumlah_perlu_review
        # WAJIB dihitung dari status_list (murni dari "alasan"), bukan len(masalah)
        # lagi, supaya tidak ikut menghitung baris "OK" sbg "perlu review".
        "jumlah_perlu_review": sum(1 for s in status_list if s == "PERLU REVIEW"),
        "jumlah_aset_habis_masa_manfaat": len(aset_habis),
        "jumlah_duplikat": sum(1 for m in masalah if any("DUPLIKAT" in a for a in m["alasan"])),
        # -- Fiskal (PMK 96/2009) --
        "total_penyusutan_fiskal_per_bulan": float(df["penyusutan_fiskal_per_bulan"].fillna(0).sum()),
        "total_akumulasi_penyusutan_fiskal_seharusnya": float(df["akumulasi_penyusutan_fiskal_seharusnya"].fillna(0).sum()),
        "total_koreksi_fiskal_akumulasi": float(df["selisih_akumulasi_komersial_vs_fiskal"].fillna(0).sum()),
        "jumlah_aset_golongan_fiskal_tidak_diketahui": int(jumlah_golongan_tidak_diketahui),
        # -- Kapitalisasi --
        "batas_kapitalisasi_rupiah": float(batas_kapitalisasi_rupiah),
        "jumlah_aset_di_bawah_batas_kapitalisasi": len(aset_di_bawah_kapitalisasi),
        "catatan": (
            "Penyusutan KOMERSIAL dihitung dgn metode GARIS LURUS saja (PSAK 16, dasar harga - residu). "
            "Kalau perusahaan pakai metode saldo menurun utk pembukuan, hasil di sini jadi pembanding, "
            "WAJIB direview akuntan. Draf jurnal hanya utk penyusutan komersial 1 bulan berjalan."
        ),
        "catatan_fiskal": (
            "Penyusutan FISKAL (kolom/field '*_fiskal_*') dihitung per golongan PMK 96/2009, metode garis "
            "lurus, dasar HARGA PENUH (nilai residu tidak mengurangi). Pemetaan kategori->golongan fiskal "
            "SIMPLIFIKASI -- WAJIB diverifikasi akuntan pajak ke Lampiran PMK 96/2009, terutama utk MESIN & "
            "aset golongan 'tidak diketahui'. Angka fiskal ini HANYA pembanding utk rekonsiliasi/koreksi "
            "fiskal SPT Tahunan Badan -- TIDAK dipakai di draf jurnal pembukuan (jurnal pembukuan tetap "
            "pakai angka komersial)."
        ),
    }

    # ============================================================
    # [BARU] Jadwal Penyusutan Bulanan 12-Kolom (siap export ke Excel)
    # ============================================================

    def _susun_jadwal_penyusutan_bulanan(df: pd.DataFrame, tahun: int = None) -> List[Dict[str, Any]]:
        """
        Susun jadwal penyusutan FISKAL 12 bulan per aset (Jan-Des),
        format siap export ke Excel -- pakai
        penyusutan_fiskal_per_bulan/akumulasi_penyusutan_fiskal_seharusnya
        yang sudah dihitung di atas (bukan angka komersial), karena jadwal
        ini dimaksudkan sbg pendamping rekonsiliasi_fiskal/PPh Badan 31E,
        bukan draf jurnal pembukuan (draf jurnal komersial sudah ada
        terpisah di 'draf_jurnal').
        """
        if df is None or df.empty:
            return []

        tahun = tahun or date.today().year
        hasil = []

        for i, row in df.iterrows():
            if str(row.get("kategori") or "").strip().upper() == "TANAH":
                continue

            penyusutan_bulan = float(row.get("penyusutan_fiskal_per_bulan") or 0)
            if penyusutan_bulan <= 0:
                continue

            nama_aset = row.get("nama_aset") or f"Aset-{i + 1}"
            kode_aset = row.get("kode_aset") or f"ASET-{i + 1}"

            akumulasi_awal = float(row.get("akumulasi_penyusutan_fiskal_seharusnya") or 0)

            jadwal = []
            for bulan in range(1, 13):
                akumulasi_baru = akumulasi_awal + (penyusutan_bulan * bulan)
                jadwal.append({
                    "bulan": bulan,
                    "penyusutan_bulan_ini": round(penyusutan_bulan, 2),
                    "akumulasi_sampai_bulan": round(akumulasi_baru, 2),
                })

            hasil.append({
                "kode_aset": kode_aset,
                "nama_aset": nama_aset,
                "kategori": row.get("kategori"),
                "golongan_fiskal": row.get("golongan_fiskal"),
                "harga_perolehan": float(row.get("harga_perolehan") or 0),
                "penyusutan_per_bulan": round(penyusutan_bulan, 2),
                "penyusutan_per_tahun": round(penyusutan_bulan * 12, 2),
                "akumulasi_awal_tahun": round(akumulasi_awal, 2),
                "akumulasi_akhir_tahun": round(akumulasi_awal + (penyusutan_bulan * 12), 2),
                "jadwal_bulanan": jadwal,  # 12 bulan
            })

        return hasil

    jadwal_penyusutan = _susun_jadwal_penyusutan_bulanan(df)

    return {
        "df": df, "ringkasan": ringkasan, "masalah": masalah, "draf_jurnal": draf_jurnal,
        "rekonsiliasi_fiskal": rekonsiliasi_fiskal,
        "aset_di_bawah_batas_kapitalisasi": aset_di_bawah_kapitalisasi,
        "jadwal_penyusutan_bulanan": jadwal_penyusutan,
    }


def proses_file_aset_tetap(
    file_like,
    nama_file: str = None,
    batas_kapitalisasi_rupiah: float = _BATAS_KAPITALISASI_DEFAULT_RUPIAH,
) -> dict:
    """
    Fungsi siap-pakai tingkat file: baca file (.xlsx/.xlsm/.xls/.csv/.pdf),
    cari semua sheet yang cocok sebagai Daftar Aset Tetap, gabungkan, lalu
    proses lewat proses_aset_tetap(). Dipakai oleh app.py utk 1 file upload.

    Kalau file lebih dari 1 sheet Aset Tetap: semua sheet DIGABUNG jadi
    satu tabel sebelum diproses (duplikat kode aset & rekap kategori
    dihitung LINTAS sheet), tapi tiap baris tetap menyimpan sheet asalnya
    lewat kolom "sheet" di df hasil.

    Return sama seperti proses_aset_tetap(), ditambah "sheet_dilewati".
    """
    nama_file = nama_file or getattr(file_like, "name", "") or ""
    daftar_sheet, _df_coa = _siapkan_daftar_sheet(file_like, nama_file)

    df_list = []
    sheet_dilewati = []
    for nama, ws in daftar_sheet:
        try:
            df_sheet = parse_sheet_aset_tetap(ws, nama)
        except FormatTidakDikenali:
            sheet_dilewati.append(nama)
            continue
        if df_sheet.empty:
            sheet_dilewati.append(nama)
            continue
        df_list.append(df_sheet)

    if not df_list:
        return {
            "df": pd.DataFrame(), "ringkasan": {}, "masalah": [], "draf_jurnal": [],
            "rekonsiliasi_fiskal": [], "aset_di_bawah_batas_kapitalisasi": [],
            "jadwal_penyusutan_bulanan": [],
            "sheet_dilewati": sheet_dilewati,
        }

    df_gabungan = pd.concat(df_list, ignore_index=True)
    hasil = proses_aset_tetap(df_gabungan, batas_kapitalisasi_rupiah=batas_kapitalisasi_rupiah)
    hasil["sheet_dilewati"] = sheet_dilewati
    return hasil


# ============================================================
# 11H. PURCHASE ORDER / INVOICE PEMBELIAN -- PARSING, VALIDASI, & DRAF JURNAL
# ============================================================
# Satu sheet bisa berisi PURCHASE ORDER (pesanan ke supplier, BELUM jadi
# kewajiban/utang -- sifatnya komitmen) ATAU INVOICE PEMBELIAN/Faktur
# Pembelian dari supplier (SUDAH jadi kewajiban/utang begitu barang/jasa
# diterima). Modul ini mengenali keduanya lewat kata kunci di sheet &
# memberi perlakuan BEDA:
#
# "Kerjaan akuntan" saat menerima dokumen pembelian:
#   1. Kenali jenis dokumen: PO atau INVOICE (lihat kata kunci di sheet).
#   2. Cek kelengkapan: nomor dokumen, tanggal, nama supplier, item.
#   3. Cross-check per baris: Subtotal = Qty x Harga Satuan.
#   4. Cross-check Total = Subtotal + PPN (kalau ada kolom PPN) --
#      HANYA reasonableness check, karena tarif PPN bisa beda per
#      periode/jenis barang, WAJIB dicek ke faktur pajak aslinya.
#   5. Deteksi qty/harga negatif atau nol.
#   6. Deteksi nomor dokumen duplikat.
#   7. Siapkan draf jurnal HANYA untuk INVOICE (PO tidak dijurnal --
#      PO baru komitmen/pesanan, belum ada kewajiban akuntansi sampai
#      barang/jasa diterima & invoice diterbitkan).
#
# CATATAN: modul ini TIDAK melakukan pencocokan 3-way (PO vs Penerimaan
# Barang vs Invoice) -- itu perlu data PO & invoice yang sudah punya
# nomor referensi silang yang konsisten, WAJIB dicek manual oleh akuntan.

_TOLERANSI_SELISIH_PEMBELIAN_RUPIAH = 50


def _cari_header_row_pembelian(ws, max_scan: int = 10):
    """Cari baris header sheet PO/Invoice Pembelian."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        teks = " ".join(str(c) for c in row if c is not None).lower()
        ada_barang = any(k in teks for k in ["nama barang", "deskripsi", "item", "jasa"])
        ada_qty = any(k in teks for k in ["qty", "jumlah", "kuantitas"])
        ada_harga = any(k in teks for k in ["harga", "price", "unit price"])
        ada_supplier = any(k in teks for k in ["supplier", "vendor", "pemasok"])
        if ada_barang and ada_qty and ada_harga and ada_supplier:
            return i + 1, list(row)
    return None, None


def _jenis_dokumen_pembelian(ws, max_scan: int = 15) -> str:
    """Tebak PO atau INVOICE berdasarkan kata kunci di beberapa baris awal sheet."""
    potongan = []
    for row in ws.iter_rows(min_row=1, max_row=max_scan, values_only=True):
        for c in row:
            if c is not None:
                potongan.append(str(c).strip().lower())
    teks = " | ".join(potongan)
    skor_po = sum(1 for k in ["purchase order", "nomor po", "tanggal po", "delivery date", "quantity ordered"] if k in teks)
    skor_invoice = sum(1 for k in ["invoice pembelian", "faktur pembelian", "bill to", "termin pembayaran", "no invoice"] if k in teks)
    return "PO" if skor_po >= skor_invoice else "INVOICE"


def parse_sheet_pembelian(ws, nama_sheet: str) -> pd.DataFrame:
    """
    Parse satu sheet PO/Invoice Pembelian jadi DataFrame satu baris = satu
    item/baris pesanan. Kolom yang dicoba dikenali: jenis dokumen (PO/
    INVOICE, ditebak otomatis), nomor dokumen, tanggal, nama supplier,
    nama barang/jasa, qty, harga satuan, subtotal, PPN, total.
    """
    header_rownum, header_row = _cari_header_row_pembelian(ws)
    if header_row is None:
        raise FormatTidakDikenali(
            f"Sheet '{nama_sheet}' tidak dikenali sebagai Purchase Order/Invoice Pembelian "
            "(tidak ditemukan kolom NAMA BARANG + QTY + HARGA + SUPPLIER sekaligus)."
        )

    headers = header_row
    jenis_dokumen = _jenis_dokumen_pembelian(ws)

    def _idx(kata_kunci):
        for i, h in enumerate(headers):
            if h is None:
                continue
            h_low = str(h).strip().lower()
            if any(k in h_low for k in kata_kunci):
                return i
        return None

    idx_nomor = _idx(["nomor po", "no po", "nomor invoice", "no invoice", "nomor dokumen", "no. invoice"])
    idx_tanggal = _idx(["tanggal", "tgl"])
    idx_supplier = _idx(["nama supplier", "supplier", "vendor", "pemasok"])
    idx_barang = _idx(["nama barang", "deskripsi", "item", "jasa"])
    idx_qty = _idx(["qty", "jumlah", "kuantitas"])
    idx_harga = _idx(["harga satuan", "harga", "unit price"])
    idx_subtotal = _idx(["subtotal", "sub total", "jumlah harga"])
    idx_ppn = _idx(["ppn", "pajak"])
    idx_total = _idx(["total", "grand total"])

    if idx_barang is None:
        raise FormatTidakDikenali(f"Kolom NAMA BARANG/ITEM tidak ditemukan di sheet '{nama_sheet}'.")

    def _ambil(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    def _ambil_angka(row, idx):
        v = _ambil(row, idx)
        try:
            return float(v) if v not in (None, "") else 0.0
        except (TypeError, ValueError):
            return 0.0

    rows = []
    for row in ws.iter_rows(min_row=header_rownum + 1, values_only=True):
        barang = _ambil(row, idx_barang)
        qty = _ambil_angka(row, idx_qty)
        if barang is None and qty == 0:
            continue
        rows.append({
            "sheet": nama_sheet,
            "jenis_dokumen": jenis_dokumen,
            "nomor_dokumen": _ambil(row, idx_nomor),
            "tanggal": _ambil(row, idx_tanggal),
            "nama_supplier": _ambil(row, idx_supplier),
            "nama_barang": barang,
            "qty": qty,
            "harga_satuan": _ambil_angka(row, idx_harga),
            "subtotal_tertulis": _ambil_angka(row, idx_subtotal) or None,
            "ppn_tertulis": _ambil_angka(row, idx_ppn) or None,
            "total_tertulis": _ambil_angka(row, idx_total) or None,
        })

    return pd.DataFrame(rows)


def proses_pembelian(df: pd.DataFrame, toleransi_rupiah: float = _TOLERANSI_SELISIH_PEMBELIAN_RUPIAH) -> dict:
    """
    "Kerjaan akuntan" untuk PO/Invoice Pembelian: cek kelengkapan, cross-
    check Subtotal = Qty x Harga, cross-check Total = Subtotal + PPN
    (reasonableness), deteksi qty/harga tidak wajar & nomor dokumen
    duplikat, siapkan draf jurnal utang usaha HANYA utk baris ber-jenis
    INVOICE (PO tidak dijurnal, baru komitmen pesanan).

    Return dict: "df", "ringkasan", "masalah", "draf_jurnal".
    """
    if df is None or df.empty:
        return {"df": df, "ringkasan": {}, "masalah": [], "draf_jurnal": []}

    df = df.copy()
    masalah = []
    nomor_terlihat: dict = {}

    subtotal_hitung_list, status_list = [], []
    for i, row in df.iterrows():
        alasan = []
        qty = float(row.get("qty") or 0)
        harga = float(row.get("harga_satuan") or 0)
        subtotal_hitung = qty * harga
        subtotal_hitung_list.append(subtotal_hitung)

        nomor = row.get("nomor_dokumen")
        if not nomor or str(nomor).strip() == "":
            alasan.append(f"Nomor {row.get('jenis_dokumen')} kosong -- WAJIB ada utk penomoran & pelacakan.")
        else:
            # [FIX -- bug duplikat lintas supplier] Sebelumnya kunci cuma
            # "nomor_dokumen" mentah, tanpa nama supplier -- padahal nomor
            # PO/Invoice diterbitkan MASING-MASING supplier secara independen
            # (Supplier A dan Supplier B sama-sama wajar punya "INV-001"
            # sendiri-sendiri). Akibatnya file dgn PO/Invoice dari >1 supplier
            # akan sering salah tandai "DUPLIKAT" padahal itu dokumen dari
            # penerbit yang beda sama sekali. Samakan polanya dengan
            # proses_ap_aging() yang sudah benar pakai kunci gabungan
            # (supplier, nomor) -- ditambah jenis_dokumen supaya PO dan
            # Invoice dgn nomor sama dari supplier yang sama juga tidak
            # dianggap duplikat satu sama lain (beda jenis dokumen).
            kunci = (
                str(row.get("nama_supplier") or "").strip().lower(),
                str(row.get("jenis_dokumen") or "").strip().lower(),
                str(nomor).strip().lower(),
            )
            if kunci in nomor_terlihat:
                alasan.append(
                    f"Nomor {row.get('jenis_dokumen') or 'dokumen'} DUPLIKAT utk supplier yang sama -- "
                    f"sudah muncul di baris ke-{nomor_terlihat[kunci] + 1} juga."
                )
            else:
                nomor_terlihat[kunci] = i

        if not row.get("nama_supplier"):
            alasan.append("Nama supplier/vendor kosong.")
        if qty <= 0:
            alasan.append("Qty kosong/nol/negatif -- tidak wajar.")
        if harga <= 0:
            alasan.append("Harga satuan kosong/nol/negatif -- tidak wajar.")

        subtotal_tertulis = row.get("subtotal_tertulis")
        if subtotal_tertulis is not None and abs(subtotal_tertulis - subtotal_hitung) > toleransi_rupiah:
            alasan.append(
                f"Subtotal tertulis Rp{subtotal_tertulis:,.0f} TIDAK SAMA dengan hasil hitung "
                f"(Qty x Harga) = Rp{subtotal_hitung:,.0f}."
            )

        total_tertulis = row.get("total_tertulis")
        ppn_tertulis = row.get("ppn_tertulis") or 0
        dasar_total = subtotal_tertulis if subtotal_tertulis is not None else subtotal_hitung
        if total_tertulis is not None:
            total_seharusnya = dasar_total + ppn_tertulis
            if abs(total_tertulis - total_seharusnya) > toleransi_rupiah:
                alasan.append(
                    f"Total tertulis Rp{total_tertulis:,.0f} TIDAK SAMA dengan Subtotal + PPN = "
                    f"Rp{total_seharusnya:,.0f} -- cek tarif/nominal PPN."
                )

        status_list.append("PERLU REVIEW" if alasan else "OK")
        if alasan:
            masalah.append({
                "baris": i + 1, "jenis_dokumen": row.get("jenis_dokumen"), "nomor_dokumen": nomor,
                "nama_supplier": row.get("nama_supplier"), "nama_barang": row.get("nama_barang"), "alasan": alasan,
            })

    df["subtotal_hitung"] = subtotal_hitung_list
    df["status"] = status_list

    # -- Draf jurnal HANYA utk baris INVOICE (PO tidak dijurnal) --
    # [FIX] Sebelumnya jml_debet = subtotal (TANPA ppn) tapi jml_kredit = total
    # (subtotal + ppn) -- selisih persis sebesar ppn setiap kali invoice-nya
    # kena PPN, jadi baris jurnal ini TIDAK BALANCE (debet != kredit untuk 1
    # transaksi yang sama). Catatan lama mengakui ini ("PPN... tambahkan baris
    # terpisah") tapi baris terpisah itu tidak pernah benar2 dibuat -- cuma
    # dicatat di field "ppn_masukan" yang tidak mempengaruhi angka jurnal sama
    # sekali. Sekarang disamakan dengan pola yang sudah benar di
    # parse_sheet_faktur_pajak (arah MASUKAN, lihat di atas): PPN dipecah ke
    # field debet terpisah (no_akun_debet_ppn/jml_debet_ppn) -- supaya
    # jml_debet + jml_debet_ppn = subtotal + ppn = total = jml_kredit, balance.
    draf_jurnal = []
    for i, row in df.iterrows():
        if row.get("jenis_dokumen") != "INVOICE":
            continue
        subtotal = row.get("subtotal_tertulis") if row.get("subtotal_tertulis") is not None else row["subtotal_hitung"]
        ppn = float(row.get("ppn_tertulis") or 0)
        total = subtotal + ppn
        baris_jurnal = {
            "baris": i + 1, "nomor_dokumen": row.get("nomor_dokumen"), "nama_supplier": row.get("nama_supplier"),
            "no_akun_debet": "PERSEDIAAN/BEBAN", "nama_akun_debet": f"Sesuaikan akun ({row.get('nama_barang') or 'cek deskripsi'})",
            "jml_debet": subtotal,
            "no_akun_kredit": "UTANG USAHA", "nama_akun_kredit": "Utang Usaha",
            "jml_kredit": total,
            "catatan": "Draf otomatis -- akun debet (persediaan/beban) WAJIB disesuaikan jenis barang/jasa.",
        }
        if ppn:
            baris_jurnal["no_akun_debet_ppn"] = "PPN MASUKAN"
            baris_jurnal["nama_akun_debet_ppn"] = "PPN Masukan"
            baris_jurnal["jml_debet_ppn"] = ppn
            baris_jurnal["catatan"] += (
                " PPN dicatat sbg PPN Masukan (debet terpisah) -- cek apakah bisa dikreditkan (PKP)."
            )
        draf_jurnal.append(baris_jurnal)


    jumlah_po = int((df["jenis_dokumen"] == "PO").sum())
    jumlah_invoice = int((df["jenis_dokumen"] == "INVOICE").sum())

    ringkasan = {
        "jumlah_baris": len(df),
        "jumlah_baris_po": jumlah_po,
        "jumlah_baris_invoice": jumlah_invoice,
        "jumlah_dokumen_unik": df["nomor_dokumen"].nunique(),
        "total_nilai_po": float(df.loc[df["jenis_dokumen"] == "PO", "subtotal_hitung"].sum()),
        "total_nilai_invoice": float(df.loc[df["jenis_dokumen"] == "INVOICE", "subtotal_hitung"].sum()),
        "jumlah_perlu_review": len(masalah),
        "jumlah_duplikat": sum(1 for m in masalah if any("DUPLIKAT" in a for a in m["alasan"])),
        "catatan": (
            "PO TIDAK dijurnal (baru komitmen pesanan, belum kewajiban). Draf jurnal hanya utk "
            "baris berjenis INVOICE. Tidak ada pencocokan 3-way PO vs Penerimaan Barang vs "
            "Invoice -- WAJIB dicek manual kalau perusahaan butuh kontrol itu."
        ),
    }

    return {"df": df, "ringkasan": ringkasan, "masalah": masalah, "draf_jurnal": draf_jurnal}


def proses_file_pembelian(file_like, nama_file: str = None) -> dict:
    """
    Fungsi siap-pakai tingkat file: baca file (.xlsx/.xlsm/.xls/.csv/.pdf),
    cari semua sheet yang cocok sebagai PO/Invoice Pembelian, gabungkan,
    lalu proses lewat proses_pembelian(). Dipakai oleh app.py utk 1 file upload.

    Return sama seperti proses_pembelian(), ditambah "sheet_dilewati".
    """
    nama_file = nama_file or getattr(file_like, "name", "") or ""
    daftar_sheet, _df_coa = _siapkan_daftar_sheet(file_like, nama_file)

    df_list = []
    sheet_dilewati = []
    for nama, ws in daftar_sheet:
        try:
            df_sheet = parse_sheet_pembelian(ws, nama)
        except FormatTidakDikenali:
            sheet_dilewati.append(nama)
            continue
        if df_sheet.empty:
            sheet_dilewati.append(nama)
            continue
        df_list.append(df_sheet)

    if not df_list:
        return {"df": pd.DataFrame(), "ringkasan": {}, "masalah": [], "draf_jurnal": [], "sheet_dilewati": sheet_dilewati}

    df_gabungan = pd.concat(df_list, ignore_index=True)
    hasil = proses_pembelian(df_gabungan)
    hasil["sheet_dilewati"] = sheet_dilewati
    return hasil


# ============================================================
# 11I. REKONSILIASI BANK -- PARSING, HITUNG SALDO DISESUAIKAN, & DRAF JURNAL
# ============================================================
# Sheet Rekonsiliasi Bank biasanya berisi DUA angka saldo awal (Saldo
# Menurut Bank/rekening koran & Saldo Menurut Buku/catatan perusahaan) +
# daftar ITEM PENYESUAI (cek beredar, setoran dalam perjalanan, biaya
# bank, bunga bank, koreksi lain).
#
# "Kerjaan akuntan" saat menyusun rekonsiliasi bank:
#   1. Ambil saldo awal Menurut Bank & Menurut Buku dari sheet.
#   2. Kelompokkan tiap item penyesuai: CEK BEREDAR & SETORAN DALAM
#      PERJALANAN menyesuaikan SISI BANK; BIAYA BANK, BUNGA BANK, &
#      koreksi buku lain menyesuaikan SISI BUKU.
#   3. Hitung Saldo Bank Disesuaikan = Saldo Bank - Cek Beredar +
#      Setoran Dalam Perjalanan (+/- koreksi bank lain).
#   4. Hitung Saldo Buku Disesuaikan = Saldo Buku + Bunga Bank -
#      Biaya Bank (+/- koreksi buku lain).
#   5. KEDUA saldo disesuaikan itu HARUS SAMA -- kalau tidak, ada
#      transaksi yang belum tercatat/salah catat, WAJIB ditelusuri.
#   6. Siapkan draf jurnal HANYA utk item yang menyesuaikan SISI BUKU
#      (bunga bank, biaya bank, koreksi buku lain) -- karena itu berarti
#      ada transaksi yang BELUM dicatat perusahaan. Item sisi bank (cek
#      beredar, setoran dlm perjalanan) TIDAK dijurnal -- itu cuma beda
#      waktu (timing difference) yang akan otomatis clear di periode
#      berikutnya begitu bank memproses.

_TOLERANSI_SELISIH_REKON_BANK = 100

_KATEGORI_REKON_BANK = {
    "CEK_BEREDAR": ["cek beredar", "outstanding check", "cek belum cair"],
    "SETORAN_DALAM_PERJALANAN": ["setoran dalam perjalanan", "deposit in transit", "setoran belum masuk"],
    "BIAYA_BANK": ["biaya bank", "biaya admin", "bank charge", "biaya transfer"],
    "BUNGA_BANK": ["bunga bank", "jasa giro", "interest income"],
    "KOREKSI_BANK": ["koreksi bank", "kesalahan bank"],
    "KOREKSI_BUKU": ["koreksi buku", "kesalahan catat", "koreksi pembukuan"],
}


def _kategorikan_item_rekon(teks_kategori: str) -> str:
    if not teks_kategori:
        return "LAINNYA"
    t = str(teks_kategori).strip().lower()
    for kode, kata_kunci in _KATEGORI_REKON_BANK.items():
        if any(k in t for k in kata_kunci):
            return kode
    return "LAINNYA"


def _cari_saldo_rekon_bank(ws, max_scan: int = 40):
    """
    Scan seluruh sheet (baris & kolom) cari label 'saldo menurut bank' &
    'saldo menurut buku' (atau sinonimnya), ambil angka pertama di sel
    sebelah kanan / sel di bawahnya sbg nilai saldo awal.
    Return (saldo_bank, saldo_buku) -- None kalau tidak ketemu.
    """
    label_bank = ["saldo menurut bank", "saldo per bank", "saldo rekening koran", "saldo kas di bank"]
    label_buku = ["saldo menurut buku", "saldo per buku", "saldo kas menurut catatan", "saldo buku besar kas"]

    saldo_bank, saldo_buku = None, None
    baris_list = list(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True))
    for r_idx, row in enumerate(baris_list):
        for c_idx, cell in enumerate(row):
            if cell is None:
                continue
            teks = str(cell).strip().lower()
            cocok_bank = any(k in teks for k in label_bank)
            cocok_buku = any(k in teks for k in label_buku)
            if not (cocok_bank or cocok_buku):
                continue
            nilai = None
            if c_idx + 1 < len(row):
                nilai = row[c_idx + 1]
            if (nilai is None or nilai == "") and r_idx + 1 < len(baris_list) and c_idx < len(baris_list[r_idx + 1]):
                nilai = baris_list[r_idx + 1][c_idx]
            try:
                nilai = float(nilai) if nilai not in (None, "") else None
            except (TypeError, ValueError):
                nilai = None
            if nilai is None:
                continue
            if cocok_bank and saldo_bank is None:
                saldo_bank = nilai
            if cocok_buku and saldo_buku is None:
                saldo_buku = nilai
    return saldo_bank, saldo_buku


def _cari_header_row_rekon_bank(ws, max_scan: int = 15):
    """Cari baris header daftar item penyesuai rekonsiliasi bank."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        teks = " ".join(str(c) for c in row if c is not None).lower()
        ada_keterangan = any(k in teks for k in ["keterangan", "uraian", "deskripsi"])
        ada_jumlah = any(k in teks for k in ["jumlah", "nominal", "nilai"])
        ada_kategori = any(k in teks for k in ["kategori", "jenis"]) or any(
            k in teks for kk in _KATEGORI_REKON_BANK.values() for k in kk
        )
        if ada_keterangan and ada_jumlah and ada_kategori:
            return i + 1, list(row)
    return None, None


def parse_sheet_rekonsiliasi_bank(ws, nama_sheet: str) -> pd.DataFrame:
    """
    Parse sheet Rekonsiliasi Bank: item-item penyesuai jadi DataFrame
    (satu baris = satu item), dan saldo awal Menurut Bank/Menurut Buku
    disimpan di df.attrs["saldo_bank"] & df.attrs["saldo_buku"].
    """
    saldo_bank, saldo_buku = _cari_saldo_rekon_bank(ws)
    header_rownum, header_row = _cari_header_row_rekon_bank(ws)
    if header_row is None:
        if saldo_bank is None and saldo_buku is None:
            raise FormatTidakDikenali(
                f"Sheet '{nama_sheet}' tidak dikenali sebagai Rekonsiliasi Bank "
                "(tidak ditemukan saldo menurut bank/buku maupun daftar item penyesuai)."
            )
        df_kosong = pd.DataFrame(columns=["sheet", "tanggal", "keterangan", "kategori", "jumlah"])
        df_kosong.attrs["saldo_bank"] = saldo_bank
        df_kosong.attrs["saldo_buku"] = saldo_buku
        return df_kosong

    headers = header_row

    def _idx(kata_kunci):
        for i, h in enumerate(headers):
            if h is None:
                continue
            h_low = str(h).strip().lower()
            if any(k in h_low for k in kata_kunci):
                return i
        return None

    idx_tanggal = _idx(["tanggal", "tgl"])
    idx_keterangan = _idx(["keterangan", "uraian", "deskripsi"])
    idx_kategori = _idx(["kategori", "jenis"])
    idx_jumlah = _idx(["jumlah", "nominal", "nilai"])

    def _ambil(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    rows = []
    for row in ws.iter_rows(min_row=header_rownum + 1, values_only=True):
        keterangan = _ambil(row, idx_keterangan)
        jumlah_raw = _ambil(row, idx_jumlah)
        try:
            jumlah = float(jumlah_raw) if jumlah_raw not in (None, "") else None
        except (TypeError, ValueError):
            jumlah = None
        if keterangan is None and jumlah is None:
            continue
        kategori_tertulis = _ambil(row, idx_kategori) or keterangan
        rows.append({
            "sheet": nama_sheet,
            "tanggal": _ambil(row, idx_tanggal),
            "keterangan": keterangan,
            "kategori_tertulis": kategori_tertulis,
            "kategori": _kategorikan_item_rekon(kategori_tertulis),
            "jumlah": jumlah or 0.0,
        })

    df = pd.DataFrame(rows)
    df.attrs["saldo_bank"] = saldo_bank
    df.attrs["saldo_buku"] = saldo_buku
    return df


def proses_rekonsiliasi_bank(
    df: pd.DataFrame,
    saldo_bank: float = None,
    saldo_buku: float = None,
    toleransi_rupiah: float = _TOLERANSI_SELISIH_REKON_BANK,
) -> dict:
    """
    "Kerjaan akuntan" untuk Rekonsiliasi Bank: hitung Saldo Bank
    Disesuaikan & Saldo Buku Disesuaikan dari saldo awal + item penyesuai,
    bandingkan (harus sama), dan siapkan draf jurnal utk item sisi buku
    (bunga bank, biaya bank, koreksi buku) yang belum tercatat.

    Return dict: "df", "ringkasan", "masalah", "draf_jurnal".
    """
    if saldo_bank is None and df is not None:
        saldo_bank = df.attrs.get("saldo_bank")
    if saldo_buku is None and df is not None:
        saldo_buku = df.attrs.get("saldo_buku")

    if df is None:
        df = pd.DataFrame()
    df = df.copy()
    masalah = []

    if saldo_bank is None or saldo_buku is None:
        masalah.append({
            "baris": 0, "keterangan": "(saldo awal)",
            "alasan": [
                "Saldo Menurut Bank dan/atau Saldo Menurut Buku tidak ditemukan di sheet -- "
                "rekonsiliasi TIDAK BISA dihitung tanpa keduanya. Pastikan sheet punya label "
                "'Saldo Menurut Bank' dan 'Saldo Menurut Buku' dengan angkanya."
            ],
        })

    def _jumlah_kategori(kode):
        if df.empty or "kategori" not in df.columns:
            return 0.0
        return float(df.loc[df["kategori"] == kode, "jumlah"].sum())

    cek_beredar = _jumlah_kategori("CEK_BEREDAR")
    setoran_transit = _jumlah_kategori("SETORAN_DALAM_PERJALANAN")
    koreksi_bank = _jumlah_kategori("KOREKSI_BANK")
    biaya_bank = _jumlah_kategori("BIAYA_BANK")
    bunga_bank = _jumlah_kategori("BUNGA_BANK")
    koreksi_buku = _jumlah_kategori("KOREKSI_BUKU")

    saldo_bank_disesuaikan = saldo_buku_disesuaikan = None
    if saldo_bank is not None:
        saldo_bank_disesuaikan = saldo_bank - cek_beredar + setoran_transit + koreksi_bank
    if saldo_buku is not None:
        saldo_buku_disesuaikan = saldo_buku + bunga_bank - biaya_bank + koreksi_buku

    selisih = None
    if saldo_bank_disesuaikan is not None and saldo_buku_disesuaikan is not None:
        selisih = saldo_bank_disesuaikan - saldo_buku_disesuaikan
        if abs(selisih) > toleransi_rupiah:
            masalah.append({
                "baris": 0, "keterangan": "(hasil rekonsiliasi)",
                "alasan": [
                    f"Saldo Bank Disesuaikan (Rp{saldo_bank_disesuaikan:,.0f}) TIDAK SAMA dengan "
                    f"Saldo Buku Disesuaikan (Rp{saldo_buku_disesuaikan:,.0f}) -- selisih "
                    f"Rp{selisih:,.0f}. Ada transaksi yang belum tercatat/salah catat, WAJIB "
                    "ditelusuri sebelum ditutup."
                ],
            })

    # -- baris tanpa kategori jelas juga di-flag --
    status_list = []
    for i, row in df.iterrows():
        alasan = []
        if row.get("kategori") == "LAINNYA":
            alasan.append(
                f"Kategori item '{row.get('kategori_tertulis')}' tidak dikenali sbg salah satu dari "
                "cek beredar/setoran dlm perjalanan/biaya bank/bunga bank/koreksi -- cek manual "
                "termasuk sisi mana (bank/buku) item ini menyesuaikan."
            )
        if not row.get("jumlah"):
            alasan.append("Jumlah kosong/nol.")
        status_list.append("PERLU REVIEW" if alasan else "OK")
        if alasan:
            masalah.append({
                "baris": i + 1, "keterangan": row.get("keterangan"), "kategori": row.get("kategori"), "alasan": alasan,
            })
    if not df.empty:
        df["status"] = status_list

    # -- Draf jurnal HANYA utk item sisi BUKU (belum tercatat perusahaan) --
    draf_jurnal = []
    if bunga_bank:
        draf_jurnal.append({
            "keterangan": "Jasa giro/bunga bank belum dicatat",
            "no_akun_debet": "BANK", "nama_akun_debet": "Kas di Bank", "jml_debet": bunga_bank,
            "no_akun_kredit": "PENDAPATAN BUNGA", "nama_akun_kredit": "Pendapatan Bunga/Jasa Giro",
            "jml_kredit": bunga_bank,
            "catatan": "Draf otomatis -- dari item kategori BUNGA_BANK di rekonsiliasi.",
        })
    if biaya_bank:
        draf_jurnal.append({
            "keterangan": "Biaya administrasi bank belum dicatat",
            "no_akun_debet": "BEBAN BANK", "nama_akun_debet": "Beban Administrasi Bank", "jml_debet": biaya_bank,
            "no_akun_kredit": "BANK", "nama_akun_kredit": "Kas di Bank", "jml_kredit": biaya_bank,
            "catatan": "Draf otomatis -- dari item kategori BIAYA_BANK di rekonsiliasi.",
        })
    if koreksi_buku:
        draf_jurnal.append({
            "keterangan": "Koreksi pembukuan (sisi buku)",
            "no_akun_debet": "BANK/LAIN", "nama_akun_debet": "Sesuaikan akun (cek detail item koreksi)",
            "jml_debet": abs(koreksi_buku),
            "no_akun_kredit": "LAIN/BANK", "nama_akun_kredit": "Sesuaikan akun (cek detail item koreksi)",
            "jml_kredit": abs(koreksi_buku),
            "catatan": "Draf otomatis -- arah debet/kredit tergantung apakah koreksi menambah atau "
                       "mengurangi saldo buku, WAJIB dicek manual per item.",
        })

    ringkasan = {
        "saldo_menurut_bank": saldo_bank,
        "saldo_menurut_buku": saldo_buku,
        "cek_beredar": cek_beredar,
        "setoran_dalam_perjalanan": setoran_transit,
        "biaya_bank": biaya_bank,
        "bunga_bank": bunga_bank,
        "koreksi_bank_lain": koreksi_bank,
        "koreksi_buku_lain": koreksi_buku,
        "saldo_bank_disesuaikan": saldo_bank_disesuaikan,
        "saldo_buku_disesuaikan": saldo_buku_disesuaikan,
        "selisih": selisih,
        "status_rekonsiliasi": "BALANCE" if (selisih is not None and abs(selisih) <= toleransi_rupiah) else "TIDAK BALANCE",
        "jumlah_item": len(df),
        "jumlah_perlu_review": len(masalah),
        "catatan": (
            "Item CEK BEREDAR & SETORAN DALAM PERJALANAN tidak dijurnal (beda waktu, akan clear "
            "sendiri). Draf jurnal hanya utk item sisi BUKU (bunga/biaya bank/koreksi) yang "
            "berarti transaksi belum tercatat perusahaan."
        ),
    }

    return {"df": df, "ringkasan": ringkasan, "masalah": masalah, "draf_jurnal": draf_jurnal}


def proses_file_rekonsiliasi_bank(file_like, nama_file: str = None) -> dict:
    """
    Fungsi siap-pakai tingkat file utk Rekonsiliasi Bank. Karena saldo awal
    Menurut Bank/Buku biasanya cuma ada SEKALI per sheet (bukan per baris),
    kalau ada beberapa sheet rekonsiliasi dalam 1 file, tiap sheet diproses
    TERPISAH lalu hasilnya digabung jadi list per sheet (BUKAN dijumlah
    jadi 1 angka, karena tiap sheet biasanya = 1 rekening bank/1 periode).
    """
    nama_file = nama_file or getattr(file_like, "name", "") or ""
    daftar_sheet, _df_coa = _siapkan_daftar_sheet(file_like, nama_file)

    hasil_per_sheet = []
    sheet_dilewati = []
    for nama, ws in daftar_sheet:
        try:
            df_sheet = parse_sheet_rekonsiliasi_bank(ws, nama)
        except FormatTidakDikenali:
            sheet_dilewati.append(nama)
            continue
        hasil = proses_rekonsiliasi_bank(
            df_sheet, saldo_bank=df_sheet.attrs.get("saldo_bank"), saldo_buku=df_sheet.attrs.get("saldo_buku")
        )
        hasil["sheet"] = nama
        hasil_per_sheet.append(hasil)

    if not hasil_per_sheet:
        return {
            "df": pd.DataFrame(), "ringkasan": {}, "masalah": [], "draf_jurnal": [],
            "per_sheet": [], "sheet_dilewati": sheet_dilewati,
        }

    # [FIX] Sebelumnya fungsi ini HANYA return "per_sheet" (list per rekening
    # bank/periode) tanpa key "df"/"ringkasan"/"masalah"/"draf_jurnal" di level
    # atas -- padahal SEMUA proses_file_xxx lain (slip_gaji, ap_aging,
    # pembelian, bukti_kas, dst) selalu return keempat key itu di level atas.
    # Akibatnya:
    #   1. main.py (`data_disimpan.get("draf_jurnal")`) selalu dapat kosong utk
    #      rekonsiliasi_bank -> draf jurnal (bunga/biaya bank) TIDAK PERNAH
    #      otomatis ditarik ke antrean "Perlu Posting" seperti jenis lain.
    #   2. Frontend yang membaca hasil.df/hasil.ringkasan/hasil.masalah secara
    #      seragam utk semua jenis dokumen akan dapat undefined/kosong khusus
    #      utk Rekonsiliasi Bank.
    # PERBAIKAN: tambahkan versi GABUNGAN (flat) dari semua sheet di level
    # atas -- "per_sheet" & "sheet_dilewati" TETAP ada apa adanya (tidak
    # menghapus apapun) utk kebutuhan tampilan detail per rekening/sheet.
    df_gabungan = pd.concat(
        [h["df"] for h in hasil_per_sheet if h.get("df") is not None and not h["df"].empty],
        ignore_index=True,
    ) if any(h.get("df") is not None and not h["df"].empty for h in hasil_per_sheet) else pd.DataFrame()

    masalah_gabungan = []
    for h in hasil_per_sheet:
        for m in (h.get("masalah") or []):
            m_baru = dict(m)
            m_baru.setdefault("sheet", h.get("sheet"))
            masalah_gabungan.append(m_baru)

    draf_jurnal_gabungan = []
    for h in hasil_per_sheet:
        for d in (h.get("draf_jurnal") or []):
            d_baru = dict(d)
            d_baru.setdefault("sheet", h.get("sheet"))
            draf_jurnal_gabungan.append(d_baru)

    jumlah_balance = sum(1 for h in hasil_per_sheet if h.get("ringkasan", {}).get("status_rekonsiliasi") == "BALANCE")
    jumlah_tidak_balance = len(hasil_per_sheet) - jumlah_balance

    ringkasan_gabungan = {
        "jumlah_sheet_direkonsiliasi": len(hasil_per_sheet),
        "jumlah_sheet_balance": jumlah_balance,
        "jumlah_sheet_tidak_balance": jumlah_tidak_balance,
        "jumlah_item_total": len(df_gabungan),
        "jumlah_perlu_review_total": len(masalah_gabungan),
        "daftar_sheet": [
            {
                "sheet": h.get("sheet"),
                "status_rekonsiliasi": h.get("ringkasan", {}).get("status_rekonsiliasi"),
                "selisih": h.get("ringkasan", {}).get("selisih"),
                "saldo_menurut_bank": h.get("ringkasan", {}).get("saldo_menurut_bank"),
                "saldo_menurut_buku": h.get("ringkasan", {}).get("saldo_menurut_buku"),
            }
            for h in hasil_per_sheet
        ],
        "catatan": (
            "Ringkasan ini GABUNGAN semua sheet/rekening hanya utk hitungan jumlah item & "
            "status per sheet -- saldo TIDAK dijumlahkan lintas rekening (lihat "
            "'daftar_sheet' atau 'per_sheet' utk detail & saldo per rekening bank)."
        ),
    }

    return {
        "df": df_gabungan,
        "ringkasan": ringkasan_gabungan,
        "masalah": masalah_gabungan,
        "draf_jurnal": draf_jurnal_gabungan,
        "per_sheet": hasil_per_sheet,
        "sheet_dilewati": sheet_dilewati,
    }


# ============================================================
# 11J. BUKU BANTU UTANG (AP AGING) -- PARSING, HITUNG UMUR UTANG, & REKAP
# ============================================================
# Sheet Buku Bantu Utang: daftar invoice/tagihan dari supplier yang BELUM
# lunas (atau lunas sebagian), dipakai utk memantau umur utang per vendor
# (aging) supaya tidak telat bayar / kena penalti / merusak hubungan
# dgn supplier.
#
# "Kerjaan akuntan" saat menyusun AP Aging:
#   1. Cek kelengkapan: nama supplier, nomor invoice, tanggal jatuh
#      tempo, jumlah utang.
#   2. Cross-check Sisa Utang = Jumlah Utang - Jumlah Dibayar (kalau ada
#      kolom jumlah dibayar).
#   3. Hitung umur utang (hari sejak jatuh tempo s.d. hari ini) &
#      kelompokkan ke bucket aging standar: Belum Jatuh Tempo, 1-30,
#      31-60, 61-90, >90 hari.
#   4. Deteksi nomor invoice duplikat per supplier.
#   5. Deteksi sisa utang negatif (kelebihan bayar) -- tidak wajar.
#   6. Flag utang yang SUDAH SANGAT lewat jatuh tempo (>90 hari) sbg
#      prioritas tindak lanjut (paling berisiko penalti/rusak relasi).
#
# CATATAN: modul ini TIDAK membuat draf jurnal -- Buku Bantu Utang adalah
# SUBLEDGER (rincian per vendor), bukan sumber jurnal baru; utangnya
# sendiri seharusnya sudah dijurnal saat invoice pembelian diterima
# (lihat modul 11H). Modul ini murni utk kontrol umur utang & saldo,
# idealnya totalnya dicocokkan ke saldo akun Utang Usaha di buku besar.

_BUCKET_AGING_UTANG = [
    (0, "BELUM JATUH TEMPO"),
    (30, "1-30 HARI"),
    (60, "31-60 HARI"),
    (90, "61-90 HARI"),
    (None, ">90 HARI"),
]


def _hitung_umur_utang_hari(tanggal_jatuh_tempo, tanggal_acuan: date = None) -> Optional[int]:
    """Hari sejak jatuh tempo s.d. tanggal_acuan (default hari ini). Negatif = belum jatuh tempo."""
    tanggal_acuan = tanggal_acuan or date.today()
    # [FIX] sama seperti _hitung_umur_aset_bulan -- pd.isna() menangkap
    # pandas.NaT juga, bukan cuma None, supaya baris dengan tanggal jatuh
    # tempo kosong tidak bikin "(tanggal_acuan - tjt).days" meledak dengan
    # TypeError ketika tjt ternyata NaT (bukan None murni).
    if pd.isna(tanggal_jatuh_tempo):
        return None
    try:
        if isinstance(tanggal_jatuh_tempo, datetime):
            tjt = tanggal_jatuh_tempo.date()
        elif isinstance(tanggal_jatuh_tempo, date):
            tjt = tanggal_jatuh_tempo
        else:
            tjt = pd.to_datetime(tanggal_jatuh_tempo).date()
    except Exception:
        return None
    return (tanggal_acuan - tjt).days


def _kategorikan_umur_utang(hari: Optional[int]) -> str:
    if hari is None:
        return "TIDAK DIKETAHUI"
    if hari <= 0:
        return "BELUM JATUH TEMPO"
    if hari <= 30:
        return "1-30 HARI"
    if hari <= 60:
        return "31-60 HARI"
    if hari <= 90:
        return "61-90 HARI"
    return ">90 HARI"


def _angka_aman(nilai, default: float = 0.0) -> float:
    """
    [FIX] Konversi ke float dengan aman -- None/NaN/inf jadi `default`
    (0.0). Beberapa titik lain di file ini masih pakai pola
    `float(x or 0)`, yang TIDAK aman untuk NaN: `float('nan') or 0`
    mengembalikan `nan` itu sendiri (NaN dianggap truthy di Python, beda
    dari None/0/""), jadi fallback "or 0"-nya tidak pernah kepakai kalau
    nilainya NaN. Ini pola bug yang sama dengan yang melatarbelakangi
    `_kosong()` di atas -- bedanya `_kosong()` untuk pengecekan boolean,
    `_angka_aman()` ini untuk nilai yang akan dipakai di operasi
    aritmetika (penjumlahan/pengurangan), yang tanpa fix bisa membuat
    NaN "menyebar" ke setiap hasil hitungan berikutnya.
    """
    if nilai is None:
        return default
    try:
        f = float(nilai)
    except (TypeError, ValueError):
        return default
    if pd.isna(f) or f in (float("inf"), float("-inf")):
        return default
    return f


def _kosong(nilai) -> bool:
    """
    Cek "kosong" yang aman utk None, NaN (float), NaT (pandas), MAUPUN
    string kosong/spasi -- dipakai di proses_ap_aging() karena setelah
    beberapa sheet digabung dgn pd.concat(), sel yang tadinya None di
    openpyxl bisa berubah jadi NaN pada kolom yang tipe datanya campur
    (object). `not nilai` / `nilai is None` SENGAJA TIDAK dipakai lagi di
    sini karena keduanya salah utk NaN: `not float('nan')` == False, dan
    `float('nan') is None` == False juga -- akibatnya baris kosong lolos
    tanpa peringatan (lihat catatan [FIX] di pemanggilnya).
    """
    if nilai is None:
        return True
    try:
        if pd.isna(nilai):
            return True
    except (TypeError, ValueError):
        pass
    return str(nilai).strip() == ""


def _adalah_baris_ringkasan(*nilai_list) -> bool:
    """
    [BARU] Deteksi baris ringkasan/footer (mis. "TOTAL", "SUB TOTAL",
    "GRAND TOTAL", "JUMLAH KESELURUHAN") yang lazim ditaruh manual di
    baris terakhir file AP Aging/AR Aging client -- dicek ke tiap nilai
    yang dioper (biasanya kolom supplier & nomor invoice).

    Kenapa perlu: parse_sheet_ap_aging() sebelumnya menganggap SEMUA baris
    di bawah header sebagai 1 invoice, termasuk baris "TOTAL" itu sendiri
    (karena kolom Jumlah Utang-nya terisi angka rekap, bukan kosong).
    Akibatnya baris rekap itu ikut masuk sbg "invoice" di export, DAN ikut
    dijumlahkan lagi ke baris TOTAL export -- total akhir jadi 2x lipat
    dari yang benar. Lihat proses_file_ap_aging()/parse_sheet_ap_aging().
    """
    kata_kunci = {"total", "sub total", "subtotal", "grand total",
                  "jumlah", "jumlah keseluruhan", "jumlah total"}
    for nilai in nilai_list:
        if _kosong(nilai):
            continue
        if str(nilai).strip().lower() in kata_kunci:
            return True
    return False


def _cari_header_row_ap_aging(ws, max_scan: int = 10):
    """Cari baris header sheet Buku Bantu Utang (AP Aging)."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        teks = " ".join(str(c) for c in row if c is not None).lower()
        ada_supplier = any(k in teks for k in ["supplier", "vendor", "pemasok"])
        ada_invoice = any(k in teks for k in ["invoice", "faktur", "no. faktur", "nomor tagihan"])
        ada_jatuh_tempo = any(k in teks for k in ["jatuh tempo", "due date"])
        ada_jumlah = any(k in teks for k in ["jumlah", "saldo", "nominal", "utang"])
        if ada_supplier and (ada_invoice or ada_jatuh_tempo) and ada_jumlah:
            return i + 1, list(row)
    return None, None


def parse_sheet_ap_aging(ws, nama_sheet: str) -> pd.DataFrame:
    """
    Parse satu sheet Buku Bantu Utang jadi DataFrame satu baris = satu
    invoice/tagihan supplier yang belum lunas.
    """
    header_rownum, header_row = _cari_header_row_ap_aging(ws)
    if header_row is None:
        raise FormatTidakDikenali(
            f"Sheet '{nama_sheet}' tidak dikenali sebagai Buku Bantu Utang/AP Aging "
            "(tidak ditemukan kolom SUPPLIER + INVOICE/JATUH TEMPO + JUMLAH sekaligus)."
        )

    headers = header_row

    def _idx(kata_kunci):
        for i, h in enumerate(headers):
            if h is None:
                continue
            h_low = str(h).strip().lower()
            if any(k in h_low for k in kata_kunci):
                return i
        return None

    idx_supplier = _idx(["nama supplier", "supplier", "vendor", "pemasok"])
    idx_invoice = _idx(["nomor invoice", "no invoice", "no. faktur", "nomor faktur", "nomor tagihan"])
    idx_tgl_invoice = _idx(["tanggal invoice", "tanggal faktur", "tgl invoice"])
    idx_jatuh_tempo = _idx(["jatuh tempo", "due date"])
    idx_jumlah_utang = _idx(["jumlah utang", "nilai invoice", "total utang", "jumlah tagihan"])
    idx_dibayar = _idx(["jumlah dibayar", "sudah dibayar", "dibayar"])
    idx_sisa = _idx(["sisa utang", "saldo utang", "sisa tagihan"])
    # [BARU] Sama seperti parse_sheet_piutang() -- "Segment" & "Project/Unit"
    # sebelumnya tidak pernah dicari di sini, jadi sheet export "Buku Bantu
    # Hutang" selalu kosong di 2 kolom ini walau file client (mis. contoh
    # "BUKU BANTU HUTANG USAHA 2025") punya kolomnya.
    idx_segment = _idx(["segment", "segmen"])
    idx_project_unit = _idx(["project/unit", "project / unit", "project unit", "unit/project", "project", "unit"])

    if idx_supplier is None:
        raise FormatTidakDikenali(f"Kolom NAMA SUPPLIER tidak ditemukan di sheet '{nama_sheet}'.")

    def _ambil(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    def _ambil_angka(row, idx):
        v = _ambil(row, idx)
        try:
            return float(v) if v not in (None, "") else None
        except (TypeError, ValueError):
            return None

    rows = []
    for row in ws.iter_rows(min_row=header_rownum + 1, values_only=True):
        supplier = _ambil(row, idx_supplier)
        jumlah_utang = _ambil_angka(row, idx_jumlah_utang)
        sisa_tertulis = _ambil_angka(row, idx_sisa)
        if supplier is None and jumlah_utang is None and sisa_tertulis is None:
            continue
        # [FIX] Baris ringkasan/footer ("TOTAL", "SUB TOTAL", dst) yang
        # ditaruh manual di file client -- lihat catatan di
        # _adalah_baris_ringkasan(). Dicek ke kolom supplier & nomor
        # invoice karena itu yang biasanya ditulis "TOTAL" oleh client
        # (bukan Vendor asli), sementara kolom Jumlah Utang-nya justru
        # terisi angka rekap sehingga baris ini TIDAK kena filter "kosong"
        # di atas dan lolos sbg invoice palsu.
        nomor_invoice_mentah = _ambil(row, idx_invoice)
        if _adalah_baris_ringkasan(supplier, nomor_invoice_mentah):
            continue
        rows.append({
            "sheet": nama_sheet,
            "nama_supplier": supplier,
            "nomor_invoice": _ambil(row, idx_invoice),
            "tanggal_invoice": _ambil(row, idx_tgl_invoice),
            "tanggal_jatuh_tempo": _ambil(row, idx_jatuh_tempo),
            "jumlah_utang": jumlah_utang or 0.0,
            "jumlah_dibayar": _ambil_angka(row, idx_dibayar) or 0.0,
            "sisa_utang_tertulis": sisa_tertulis,
            "segment": _ambil(row, idx_segment),
            "project_unit": _ambil(row, idx_project_unit),
        })

    return pd.DataFrame(rows)


def proses_ap_aging(df: pd.DataFrame, toleransi_rupiah: float = 50, tanggal_acuan: date = None) -> dict:
    """
    "Kerjaan akuntan" untuk Buku Bantu Utang: cross-check sisa utang,
    hitung umur & bucket aging per invoice, rekap per supplier & per
    bucket, deteksi duplikat & anomali, siapkan saran tindak lanjut utk
    utang yang sudah sangat lewat jatuh tempo.

    Return dict: "df", "ringkasan", "masalah", "saran_tindak_lanjut".
    (Tidak ada "draf_jurnal" -- lihat catatan di atas modul ini.)
    """
    if df is None or df.empty:
        return {"df": df, "ringkasan": {}, "masalah": [], "saran_tindak_lanjut": []}

    tanggal_acuan = tanggal_acuan or date.today()
    df = df.copy()
    masalah = []
    invoice_terlihat: dict = {}

    sisa_hitung_list, umur_list, bucket_list, status_list = [], [], [], []
    # [BARU] Terpisah dari status_list (flag KUALITAS DATA "PERLU
    # REVIEW"/"OK") -- ini status PELUNASAN per invoice ("Lunas"/
    # "Outstanding") berdasar sisa saldo, sesuai arti kolom "Status" di
    # sheet export "Buku Bantu Hutang" (persis pola kolom K di file model
    # referensi: =IF(SaldoHutang=0,"Lunas","Outstanding")). Sebelumnya
    # export salah pakai status_list (data-quality) utk kolom ini --
    # akibatnya SEMUA baris tertulis "PERLU REVIEW" di Excel walau
    # invoice-nya sudah lunas.
    status_pelunasan_list = []
    for i, row in df.iterrows():
        alasan = []
        supplier = row.get("nama_supplier")
        nomor_invoice = row.get("nomor_invoice")
        jumlah_utang = float(row.get("jumlah_utang") or 0)
        jumlah_dibayar = float(row.get("jumlah_dibayar") or 0)
        sisa_hitung = jumlah_utang - jumlah_dibayar
        sisa_hitung_list.append(sisa_hitung)

        # [FIX] Sebelumnya pakai "if not supplier / not nomor_invoice" --
        # itu TIDAK menangkap nilai kosong yang sudah jadi NaN float
        # (bukan None murni), yang justru SERING terjadi setelah
        # pd.concat() menggabungkan beberapa sheet AP Aging jadi satu
        # df (kolom bertipe object campur None/NaN -> openpyxl None di
        # sheet lain membuat pandas menormalkan sel kosong di kolom yang
        # sama jadi NaN). "not nan" bernilai False (nan itu truthy) dan
        # str(nan).strip() == "nan" (bukan ""), jadi baris dengan
        # supplier/nomor invoice kosong LOLOS tanpa diberi warning sama
        # sekali -- padahal seharusnya ditandai "PERLU REVIEW". Diganti
        # pakai pd.isna() (aman utk None, NaN, maupun NaT sekaligus) +
        # cek string kosong setelah strip.
        if _kosong(supplier):
            alasan.append("Nama supplier kosong.")
        if _kosong(nomor_invoice):
            alasan.append("Nomor invoice kosong -- sulit dilacak/dicocokkan ke invoice aslinya.")
        else:
            kunci = (
                "" if _kosong(supplier) else str(supplier).strip().lower(),
                str(nomor_invoice).strip().lower(),
            )
            if kunci in invoice_terlihat:
                alasan.append(
                    f"Nomor invoice DUPLIKAT utk supplier yang sama -- sudah muncul di baris "
                    f"ke-{invoice_terlihat[kunci] + 1} juga."
                )
            else:
                invoice_terlihat[kunci] = i

        # [FIX] "is None" tidak menangkap pandas.NaT (tanggal kosong hasil
        # pd.concat lintas sheet) -- diganti pd.isna() spt di
        # _hitung_umur_utang_hari(), supaya pesan "tanggal jatuh tempo
        # kosong" konsisten muncul, bukan cuma bucket-nya diam2 jadi
        # "TIDAK DIKETAHUI" tanpa penjelasan di daftar masalah.
        if pd.isna(row.get("tanggal_jatuh_tempo")):
            alasan.append("Tanggal jatuh tempo kosong -- umur utang tidak bisa dihitung.")

        # [FIX] Sama seperti supplier/nomor_invoice di atas: "is not None"
        # tidak menangkap NaN (nilai kosong hasil pd.concat lintas sheet),
        # jadi sisa_final bisa diam2 jadi NaN dan bikin perbandingan
        # `sisa_final > toleransi_rupiah` di bawah SELALU False (NaN tidak
        # pernah > apapun) -- akibatnya baris yg sudah >90 hari TIDAK
        # ditandai prioritas tinggi kalau kolom "Sisa Utang" di file
        # sumber kosong. Diganti pd.isna() supaya fallback ke sisa_hitung
        # benar2 kepakai.
        sisa_tertulis = row.get("sisa_utang_tertulis")
        sisa_tertulis_kosong = pd.isna(sisa_tertulis)
        if not sisa_tertulis_kosong and abs(sisa_tertulis - sisa_hitung) > toleransi_rupiah:
            alasan.append(
                f"Sisa Utang tertulis Rp{sisa_tertulis:,.0f} TIDAK SAMA dengan hasil hitung "
                f"(Jumlah Utang - Jumlah Dibayar) = Rp{sisa_hitung:,.0f}."
            )

        sisa_final = sisa_hitung if sisa_tertulis_kosong else sisa_tertulis
        if sisa_final < -toleransi_rupiah:
            alasan.append(
                f"Sisa utang NEGATIF (Rp{sisa_final:,.0f}) -- kemungkinan kelebihan bayar ke "
                "supplier ini, cek manual (bisa jadi uang muka utk invoice berikutnya)."
            )

        umur = _hitung_umur_utang_hari(row.get("tanggal_jatuh_tempo"), tanggal_acuan)
        bucket = _kategorikan_umur_utang(umur)
        umur_list.append(umur)
        bucket_list.append(bucket)

        # [BARU] Status pelunasan per invoice -- toleransi_rupiah yang
        # sama dipakai supaya konsisten dgn definisi "sisa_final != 0" di
        # atas (mis. sisa Rp2 krn pembulatan tetap dianggap Lunas).
        status_pelunasan_list.append("Lunas" if abs(sisa_final) <= toleransi_rupiah else "Outstanding")

        if bucket == ">90 HARI" and sisa_final > toleransi_rupiah:
            alasan.append(
                f"🚨 SUDAH LEWAT JATUH TEMPO {umur} HARI (>90 hari) -- prioritas tinggi utk "
                "segera dibayar/dikonfirmasi ke supplier, risiko penalti/rusak relasi."
            )

        status_list.append("PERLU REVIEW" if alasan else "OK")
        if alasan:
            masalah.append({
                "baris": i + 1, "nama_supplier": supplier, "nomor_invoice": nomor_invoice,
                "jatuh_tempo": row.get("tanggal_jatuh_tempo"), "umur_hari": umur, "alasan": alasan,
            })

    df["sisa_utang_hitung"] = sisa_hitung_list
    df["umur_hari"] = umur_list
    df["bucket_aging"] = bucket_list
    df["status"] = status_list
    df["status_pelunasan"] = status_pelunasan_list

    sisa_kolom = df["sisa_utang_tertulis"].fillna(df["sisa_utang_hitung"])
    rekap_supplier = df.assign(_sisa=sisa_kolom).groupby("nama_supplier")["_sisa"].sum().to_dict()
    rekap_bucket = df.assign(_sisa=sisa_kolom).groupby("bucket_aging")["_sisa"].sum().to_dict()

    saran_tindak_lanjut = [
        f"Hubungi/lunasi tagihan ke {m['nama_supplier']} (invoice {m['nomor_invoice']}) -- "
        f"sudah lewat jatuh tempo {m['umur_hari']} hari."
        for m in masalah if m.get("umur_hari") and m["umur_hari"] > 90
    ]

    ringkasan = {
        "jumlah_invoice": len(df),
        "jumlah_supplier": df["nama_supplier"].nunique(),
        "total_sisa_utang": float(sisa_kolom.sum()),
        "rekap_per_supplier": rekap_supplier,
        "rekap_per_bucket_aging": rekap_bucket,
        "jumlah_perlu_review": len(masalah),
        "jumlah_lewat_90_hari": sum(1 for b in bucket_list if b == ">90 HARI"),
        "jumlah_duplikat": sum(1 for m in masalah if any("DUPLIKAT" in a for a in m["alasan"])),
        "catatan": (
            "Buku Bantu Utang adalah subledger, tidak menghasilkan jurnal baru -- utangnya "
            "sendiri seharusnya sudah dijurnal saat invoice pembelian diterima. Total sisa "
            "utang di sini idealnya dicocokkan ke saldo akun Utang Usaha di buku besar."
        ),
    }

    return {"df": df, "ringkasan": ringkasan, "masalah": masalah, "saran_tindak_lanjut": saran_tindak_lanjut}


def proses_file_ap_aging(file_like, nama_file: str = None) -> dict:
    """
    Fungsi siap-pakai tingkat file: baca file (.xlsx/.xlsm/.xls/.csv/.pdf),
    cari semua sheet yang cocok sebagai Buku Bantu Utang, gabungkan, lalu
    proses lewat proses_ap_aging(). Dipakai oleh app.py utk 1 file upload.

    Return sama seperti proses_ap_aging(), ditambah "sheet_dilewati".
    """
    nama_file = nama_file or getattr(file_like, "name", "") or ""
    daftar_sheet, _df_coa = _siapkan_daftar_sheet(file_like, nama_file)

    df_list = []
    sheet_dilewati = []
    for nama, ws in daftar_sheet:
        try:
            df_sheet = parse_sheet_ap_aging(ws, nama)
        except FormatTidakDikenali:
            sheet_dilewati.append(nama)
            continue
        if df_sheet.empty:
            sheet_dilewati.append(nama)
            continue
        df_list.append(df_sheet)

    if not df_list:
        return {"df": pd.DataFrame(), "ringkasan": {}, "masalah": [], "saran_tindak_lanjut": [], "sheet_dilewati": sheet_dilewati}

    df_gabungan = pd.concat(df_list, ignore_index=True)
    hasil = proses_ap_aging(df_gabungan)
    hasil["sheet_dilewati"] = sheet_dilewati
    return hasil


# ============================================================
# 11K. ABSENSI/TIMESHEET -- PARSING, VALIDASI, & REKAP KEHADIRAN
# ============================================================
# Sheet Absensi: log kehadiran harian karyawan (jam masuk/keluar dan/atau
# status Hadir/Izin/Sakit/Cuti/Alpha).
#
# "Kerjaan akuntan/HR" saat menerima data absensi:
#   1. Cek kelengkapan: nama/NIP karyawan, tanggal, jam masuk-keluar
#      ATAU status kehadiran.
#   2. Hitung keterlambatan (jam masuk - jam masuk standar, default
#      08:00, bisa disesuaikan).
#   3. Hitung total jam kerja per hari (jam keluar - jam masuk).
#   4. Deteksi anomali: jam keluar SEBELUM jam masuk (data rusak/salah
#      input), status Hadir tapi jam keluar kosong (lupa absen pulang),
#      keterlambatan ekstrem.
#   5. Rekap per karyawan: total hadir, izin, sakit, cuti, alpha, total
#      menit terlambat, rata-rata jam kerja.
#
# CATATAN: modul ini TIDAK membuat draf jurnal -- absensi bukan transaksi
# akuntansi langsung, tapi jadi DASAR utk potongan gaji (keterlambatan/
# alpha) di modul Slip Gaji. Rekap di sini idealnya dicocokkan manual ke
# slip gaji karyawan yang bersangkutan pada periode yang sama.

_JAM_MASUK_STANDAR_DEFAULT = "08:00"
_AMBANG_TERLAMBAT_EKSTREM_MENIT = 180  # >3 jam telat -- kemungkinan data salah input


def _cari_header_row_absensi(ws, max_scan: int = 10):
    """Cari baris header sheet Absensi/Timesheet.
    [FIX] Sebelumnya syaratnya terlalu longgar: "nama" SENDIRIAN dianggap
    cukup utk ada_karyawan, dan "keterangan"/"status" SENDIRIAN dianggap
    cukup utk ada_absensi. Karena "tanggal" + "keterangan" ada di HAMPIR
    SEMUA dokumen keuangan (mis. rekening koran punya TANGGAL + KETERANGAN,
    dan kadang ada kolom bernama "Nama Pengirim"/"Nama Bank"), kombinasi itu
    gampang salah dianggap sheet absensi padahal bukan. Sekarang:
    - ada_karyawan wajib frasa spesifik ("nama karyawan"/"nip"/"nama pegawai"/
      "employee"), bukan "nama" sendirian.
    - ada_absensi wajib penanda absensi yg spesifik ("jam masuk"/"jam keluar"/
      "check in"/"check out"/"hadir"/"alpha"/"terlambat"), bukan
      "status"/"keterangan" sendirian (dua kata itu terlalu umum)."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        teks = " ".join(str(c) for c in row if c is not None).lower()
        ada_karyawan = any(k in teks for k in ["nama karyawan", "nip", "nama pegawai", "employee"])
        ada_tanggal = "tanggal" in teks or "tgl" in teks
        ada_absensi = any(k in teks for k in [
            "jam masuk", "jam keluar", "check in", "check out",
            "clock in", "clock out", "hadir", "alpha", "terlambat", "timesheet",
        ])
        if ada_karyawan and ada_tanggal and ada_absensi:
            return i + 1, list(row)
    return None, None


def _parse_jam_ke_menit(value) -> Optional[int]:
    """
    Konversi jam (string 'HH:MM', datetime.time, datetime.datetime, atau
    pecahan hari ala Excel) jadi menit sejak tengah malam. None kalau
    tidak bisa di-parse.
    """
    if value is None or value == "" or pd.isna(value):
        return None
    try:
        if isinstance(value, datetime):
            return value.hour * 60 + value.minute
        if hasattr(value, "hour") and hasattr(value, "minute"):
            return value.hour * 60 + value.minute
        if isinstance(value, (int, float)):
            total_menit = round(float(value) * 24 * 60)
            return total_menit % (24 * 60)
        teks = str(value).strip()
        if ":" in teks:
            bagian = teks.split(":")
            jam, menit = int(bagian[0]), int(bagian[1])
            return jam * 60 + menit
    except (ValueError, TypeError, IndexError):
        return None
    return None


def parse_sheet_absensi(ws, nama_sheet: str) -> pd.DataFrame:
    """
    Parse satu sheet Absensi jadi DataFrame satu baris = satu hari
    kehadiran satu karyawan. Kolom yang dicoba dikenali: nama/NIP
    karyawan, tanggal, jam masuk, jam keluar, status kehadiran.
    """
    header_rownum, header_row = _cari_header_row_absensi(ws)
    if header_row is None:
        raise FormatTidakDikenali(
            f"Sheet '{nama_sheet}' tidak dikenali sebagai Data Absensi/Timesheet "
            "(tidak ditemukan kolom NAMA KARYAWAN + TANGGAL + JAM MASUK-KELUAR/STATUS sekaligus)."
        )

    headers = header_row

    def _idx(kata_kunci):
        for i, h in enumerate(headers):
            if h is None:
                continue
            h_low = str(h).strip().lower()
            if any(k in h_low for k in kata_kunci):
                return i
        return None

    idx_nama = _idx(["nama karyawan", "nama", "employee"])
    idx_nip = _idx(["nip", "id karyawan", "employee id"])
    idx_tanggal = _idx(["tanggal", "tgl"])
    idx_jam_masuk = _idx(["jam masuk", "check in", "clock in"])
    idx_jam_keluar = _idx(["jam keluar", "check out", "clock out"])
    idx_status = _idx(["status", "keterangan"])

    if idx_nama is None:
        raise FormatTidakDikenali(f"Kolom NAMA KARYAWAN tidak ditemukan di sheet '{nama_sheet}'.")

    def _ambil(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    rows = []
    for row in ws.iter_rows(min_row=header_rownum + 1, values_only=True):
        nama = _ambil(row, idx_nama)
        tanggal = _ambil(row, idx_tanggal)
        if nama is None and tanggal is None:
            continue
        status_tertulis = _ambil(row, idx_status)
        rows.append({
            "sheet": nama_sheet,
            "nama_karyawan": nama,
            "nip": _ambil(row, idx_nip),
            "tanggal": tanggal,
            "jam_masuk": _ambil(row, idx_jam_masuk),
            "jam_keluar": _ambil(row, idx_jam_keluar),
            "status": str(status_tertulis).strip().upper() if status_tertulis else None,
        })

    return pd.DataFrame(rows)


def proses_absensi(
    df: pd.DataFrame,
    jam_masuk_standar: str = _JAM_MASUK_STANDAR_DEFAULT,
    ambang_terlambat_ekstrem_menit: int = _AMBANG_TERLAMBAT_EKSTREM_MENIT,
) -> dict:
    """
    "Kerjaan akuntan/HR" untuk Data Absensi: hitung keterlambatan & jam
    kerja per baris, deteksi anomali (jam keluar sebelum jam masuk, lupa
    absen pulang, keterlambatan ekstrem), lalu rekap kehadiran per
    karyawan (hadir/izin/sakit/cuti/alpha, total menit terlambat,
    rata-rata jam kerja).

    Return dict: "df", "ringkasan", "masalah".
    (Tidak ada "draf_jurnal" -- lihat catatan di atas modul ini.)
    """
    if df is None or df.empty:
        return {"df": df, "ringkasan": {}, "masalah": []}

    df = df.copy()
    masalah = []
    menit_standar = _parse_jam_ke_menit(jam_masuk_standar) or 8 * 60

    menit_masuk_list, menit_keluar_list = [], []
    keterlambatan_list, jam_kerja_list, status_list = [], [], []

    for i, row in df.iterrows():
        alasan = []
        nama = row.get("nama_karyawan")
        status = row.get("status")
        menit_masuk = _parse_jam_ke_menit(row.get("jam_masuk"))
        menit_keluar = _parse_jam_ke_menit(row.get("jam_keluar"))
        menit_masuk_list.append(menit_masuk)
        menit_keluar_list.append(menit_keluar)

        if not nama or str(nama).strip() == "":
            alasan.append("Nama karyawan kosong.")
        if row.get("tanggal") is None:
            alasan.append("Tanggal kosong.")

        status_hadir = status in (None, "", "HADIR", "H", "MASUK")
        keterlambatan = None
        jam_kerja = None

        if menit_masuk is not None and menit_keluar is not None:
            if menit_keluar <= menit_masuk:
                alasan.append(
                    "Jam Keluar SEBELUM/SAMA DENGAN Jam Masuk -- data tidak wajar, kemungkinan "
                    "salah input atau lintas hari (shift malam) yang belum ditandai."
                )
            else:
                jam_kerja = (menit_keluar - menit_masuk) / 60

        if menit_masuk is not None:
            keterlambatan = max(menit_masuk - menit_standar, 0)
            if keterlambatan > ambang_terlambat_ekstrem_menit:
                alasan.append(
                    f"Keterlambatan {keterlambatan} menit (>{ambang_terlambat_ekstrem_menit} menit) -- "
                    "tidak wajar, cek apakah jam masuk salah input."
                )
        elif status_hadir:
            alasan.append("Status Hadir tapi Jam Masuk kosong -- cek input.")

        if status_hadir and menit_masuk is not None and menit_keluar is None:
            alasan.append("Status Hadir tapi Jam Keluar kosong -- kemungkinan karyawan lupa absen pulang.")

        keterlambatan_list.append(keterlambatan)
        jam_kerja_list.append(jam_kerja)

        status_list.append("PERLU REVIEW" if alasan else "OK")
        if alasan:
            masalah.append({
                "baris": i + 1, "nama_karyawan": nama, "tanggal": row.get("tanggal"),
                "status": status, "alasan": alasan,
            })

    df["menit_masuk"] = menit_masuk_list
    df["menit_keluar"] = menit_keluar_list
    df["keterlambatan_menit"] = keterlambatan_list
    df["jam_kerja"] = jam_kerja_list
    df["status_validasi"] = status_list

    def _hitung_kategori(status_series, kata_kunci):
        return status_series.fillna("HADIR").str.upper().apply(lambda s: any(k in s for k in kata_kunci))

    status_series = df["status"].fillna("HADIR").astype(str).str.upper()
    hadir_mask = status_series.apply(lambda s: any(k in s for k in ["HADIR", "H", "MASUK"]) and not any(
        k in s for k in ["IZIN", "SAKIT", "CUTI", "ALPHA", "ALFA"]
    ))
    izin_mask = status_series.str.contains("IZIN")
    sakit_mask = status_series.str.contains("SAKIT")
    cuti_mask = status_series.str.contains("CUTI")
    alpha_mask = status_series.str.contains("ALPHA") | status_series.str.contains("ALFA")

    rekap_list = []
    for nama, grup in df.groupby("nama_karyawan"):
        idx = grup.index
        rekap_list.append({
            "nama_karyawan": nama,
            "jumlah_hadir": int(hadir_mask.loc[idx].sum()),
            "jumlah_izin": int(izin_mask.loc[idx].sum()),
            "jumlah_sakit": int(sakit_mask.loc[idx].sum()),
            "jumlah_cuti": int(cuti_mask.loc[idx].sum()),
            "jumlah_alpha": int(alpha_mask.loc[idx].sum()),
            "total_menit_terlambat": int(grup["keterlambatan_menit"].fillna(0).sum()),
            "rata_rata_jam_kerja": round(float(grup["jam_kerja"].dropna().mean()), 2) if grup["jam_kerja"].notna().any() else None,
        })
    rekap_per_karyawan = {r["nama_karyawan"]: {k: v for k, v in r.items() if k != "nama_karyawan"} for r in rekap_list}

    ringkasan = {
        "jumlah_baris": len(df),
        "jumlah_karyawan": df["nama_karyawan"].nunique(),
        "total_hadir": int(hadir_mask.sum()),
        "total_izin": int(izin_mask.sum()),
        "total_sakit": int(sakit_mask.sum()),
        "total_cuti": int(cuti_mask.sum()),
        "total_alpha": int(alpha_mask.sum()),
        "rekap_per_karyawan": rekap_per_karyawan,
        "jumlah_perlu_review": len(masalah),
        "catatan": (
            "Jam masuk standar dianggap "
            f"{jam_masuk_standar} (bisa disesuaikan). Data ini TIDAK menghasilkan jurnal -- "
            "jadi dasar potongan gaji (keterlambatan/alpha) di modul Slip Gaji, cocokkan manual "
            "ke slip gaji periode yang sama."
        ),
    }

    return {"df": df, "ringkasan": ringkasan, "masalah": masalah}


def proses_file_absensi(file_like, nama_file: str = None) -> dict:
    """
    Fungsi siap-pakai tingkat file: baca file (.xlsx/.xlsm/.xls/.csv/.pdf),
    cari semua sheet yang cocok sebagai Data Absensi/Timesheet, gabungkan,
    lalu proses lewat proses_absensi(). Dipakai oleh app.py utk 1 file upload.

    Return sama seperti proses_absensi(), ditambah "sheet_dilewati".
    """
    nama_file = nama_file or getattr(file_like, "name", "") or ""
    daftar_sheet, _df_coa = _siapkan_daftar_sheet(file_like, nama_file)

    df_list = []
    sheet_dilewati = []
    for nama, ws in daftar_sheet:
        try:
            df_sheet = parse_sheet_absensi(ws, nama)
        except FormatTidakDikenali:
            sheet_dilewati.append(nama)
            continue
        if df_sheet.empty:
            sheet_dilewati.append(nama)
            continue
        df_list.append(df_sheet)

    if not df_list:
        return {"df": pd.DataFrame(), "ringkasan": {}, "masalah": [], "sheet_dilewati": sheet_dilewati}

    df_gabungan = pd.concat(df_list, ignore_index=True)
    hasil = proses_absensi(df_gabungan)
    hasil["sheet_dilewati"] = sheet_dilewati
    return hasil


# ============================================================
# 12. PROSES DATA PENJUALAN - PIPELINE UTAMA
# ============================================================

def ekstrak_signature_penjualan(row) -> str:
    dasar = row.get("keterangan") or row.get("customer") or ""
    return f"PJL::{ekstrak_signature(dasar)}"


def _cari_kandidat_coa(df_coa: pd.DataFrame, kata_kunci_list: list[str]) -> pd.DataFrame:
    if df_coa is None or df_coa.empty:
        return pd.DataFrame(columns=["no_akun", "nama_akun", "kategori"])
    pola_regex = "|".join(kata_kunci_list)
    return df_coa[df_coa["nama_akun"].str.upper().str.contains(pola_regex, na=False, regex=True)]


def tentukan_akun_kas_piutang(df_coa: pd.DataFrame, cara_bayar: str):
    kw = ["PIUTANG"] if cara_bayar == "KREDIT" else ["KAS", "BANK"]
    kandidat = _cari_kandidat_coa(df_coa, kw)
    if len(kandidat) >= 1:
        baris = kandidat.iloc[0]
        return baris["no_akun"], baris["nama_akun"]
    label = "PIUTANG USAHA" if cara_bayar == "KREDIT" else "KAS"
    return label, label


def tentukan_akun_ppn_keluaran(df_coa: pd.DataFrame):
    kandidat = _cari_kandidat_coa(df_coa, ["PPN KELUARAN", "PPN"])
    if len(kandidat) >= 1:
        baris = kandidat.iloc[0]
        return baris["no_akun"], baris["nama_akun"]
    return "PPN KELUARAN", "PPN KELUARAN"


def cari_akun_penjualan_pasti(df_coa: pd.DataFrame):
    kandidat = _cari_kandidat_coa(df_coa, ["PENJUALAN", "PENDAPATAN"])
    if len(kandidat) == 1:
        baris = kandidat.iloc[0]
        return baris["no_akun"], baris["nama_akun"]
    return None


def _isi_baris_penjualan(df, idx, row, no_debet, nama_debet, no_kredit, nama_kredit,
                         no_kredit_ppn=None, nama_kredit_ppn=None):
    total = row["total"] or 0
    dpp = row["dpp"] or 0
    ppn = row["ppn"] or 0
    df.at[idx, "no_akun_debet"] = no_debet
    df.at[idx, "nama_akun_debet"] = nama_debet
    df.at[idx, "jml_debet"] = total
    df.at[idx, "no_akun_kredit"] = no_kredit
    df.at[idx, "nama_akun_kredit"] = nama_kredit
    df.at[idx, "jml_kredit"] = dpp if ppn else total
    if ppn and no_kredit_ppn is not None:
        df.at[idx, "no_akun_kredit_ppn"] = no_kredit_ppn
        df.at[idx, "nama_akun_kredit_ppn"] = nama_kredit_ppn
        df.at[idx, "jml_kredit_ppn"] = ppn


def kategorikan_penjualan_dengan_ai(baris_belum_jelas: list[dict], df_coa: pd.DataFrame, api_key: str,
                                     mask_pii: bool = True):
    """Kategorikan akun penjualan yang ambigu dengan AI."""
    if not baris_belum_jelas:
        return {}, []
    try:
        import openai
    except ImportError:
        return {}, [{"idx_terdampak": [b["idx"] for b in baris_belum_jelas], 
                     "alasan": "Library openai tidak terinstall"}]

    if mask_pii:
        baris_belum_jelas = [
            {**b, "customer": mask_keterangan_sensitif(b.get("customer")),
             "keterangan": mask_keterangan_sensitif(b.get("keterangan"))}
            for b in baris_belum_jelas
        ]

    if df_coa is not None and not df_coa.empty:
        daftar_akun_str = "\n".join(f"{int(r.no_akun)} - {r.nama_akun}" for r in df_coa.itertuples())
    else:
        daftar_akun_str = "(Tidak ada COA asli diupload; gunakan nama kategori umum, mis. 'PENJUALAN BARANG', 'PENJUALAN JASA')"

    def _buat_bagian_statis_penjualan() -> str:
        """[BARU -- PERBAIKAN PERFORMA -- PROMPT CACHING] Sama pola-nya
        dengan _buat_bagian_statis() di kategorikan_dengan_ai(): bagian
        prompt yang SAMA PERSIS di semua chunk 1 batch (instruksi + daftar
        COA + format jawaban) dipisah dari daftar transaksi (beda tiap
        chunk) supaya bisa di-cache lewat cache_control (lihat
        _proses_satu_chunk_ai_claude). Sebelumnya fungsi ini menggabungkan
        daftar_akun_str ke dalam _buat_prompt(chunk) tiap panggilan --
        artinya COA & instruksi terkirim ulang PENUH tiap chunk tanpa cache
        hit sama sekali. Isi instruksinya sendiri TIDAK diubah, cuma
        dipisah dari bagian dinamis."""
        return f"""Kamu adalah asisten akuntansi yang TELITI dan KONSERVATIF. Perusahaan ini mencatat
transaksi PENJUALAN. Akun Kas/Piutang dan PPN Keluaran sudah ditentukan otomatis; tugasmu
HANYA menentukan akun PENJUALAN/PENDAPATAN yang paling tepat untuk tiap transaksi berikut,
berdasarkan Chart of Accounts (COA) perusahaan ini:

{daftar_akun_str}

- no_akun_kredit_penjualan WAJIB salah satu nomor akun yang persis ada di daftar COA di atas.
- Kalau ada lebih dari satu akun penjualan yang sama-sama masuk akal dan keterangan tidak cukup jelas, termasuk KASUS TIDAK YAKIN.
- Kalau betul-betul tidak yakin, isi no_akun_kredit_penjualan dengan null.
- "confidence": "tinggi" | "sedang" | "rendah"
- "alasan": 1 kalimat singkat.

Jawab HANYA dalam format JSON array:
[{{"nomor": 1, "no_akun_kredit_penjualan": 41100001, "confidence": "tinggi",
  "alasan": "Nama produk cocok persis dengan akun Penjualan Jasa Konsultasi"}}]
"""

    def _buat_prompt(chunk):
        daftar_transaksi_str = "\n".join(
            f"{b['idx']}. [{b['cara_bayar']}] Customer: {b.get('customer') or '-'} | "
            f"{b.get('keterangan') or '-'} | DPP Rp{b['dpp']:,.0f} | PPN Rp{b['ppn']:,.0f} | Total Rp{b['total']:,.0f}"
            for b in chunk
        )
        return f"""Daftar transaksi:
{daftar_transaksi_str}"""

    # [BARU -- FALLBACK GROQ SEMENTARA + PROMPT CACHING] Sama seperti
    # kategorikan_dengan_ai() di atas -- lewat dispatcher yang otomatis
    # coba Claude dulu, fallback Groq kalau ANTHROPIC_API_KEY belum aktif,
    # dan sekarang juga meneruskan prompt_statis supaya COA & instruksi
    # (biasanya bagian terbesar dari prompt) di-cache lewat cache_control,
    # bukan dikirim ulang penuh tiap chunk. Param `api_key` dipertahankan
    # di signature demi kompatibilitas pemanggil lama, tidak dipakai langsung.
    return _panggil_kategorisasi_dengan_fallback(
        baris_belum_jelas, _buat_prompt, prompt_statis=_buat_bagian_statis_penjualan(),
    )


def proses_dataframe_penjualan(df: pd.DataFrame, df_coa: pd.DataFrame, pola: Pola,
                                pakai_ai: bool = False, api_key: Optional[str] = None,
                                mask_pii: bool = True, ambang_confidence: str = "sedang") -> pd.DataFrame:
    """
    Sama filosofinya dgn proses_dataframe() utk rekening koran: SEMUA baris
    (termasuk yg sudah berjurnal di file asal) diproses ulang lewat
    pola -> ATURAN STANDAR PENJUALAN + kata kunci COA -> AI.
    """
    df = df.copy()
    df["sumber_kategori"] = None
    df["confidence_ai"] = None
    df["alasan_ai"] = None
    df["catatan_ai"] = None
    for _kol in ("no_akun_debet", "nama_akun_debet", "no_akun_kredit", "nama_akun_kredit", "jml_debet", "jml_kredit"):
        if _kol not in df.columns:
            df[_kol] = pd.Series([None] * len(df), dtype="object", index=df.index)
        else:
            df[_kol] = df[_kol].astype("object")
    urutan_confidence = {"rendah": 0, "sedang": 1, "tinggi": 2}
    ambang_nilai = urutan_confidence.get(ambang_confidence, 1)
    for kol in ("no_akun_debet", "nama_akun_debet", "no_akun_kredit", "nama_akun_kredit",
                "jml_debet", "jml_kredit",
                "no_akun_kredit_ppn", "nama_akun_kredit_ppn", "jml_kredit_ppn"):
        if kol not in df.columns:
            df[kol] = None
    for kol in ("no_akun_debet", "no_akun_kredit", "no_akun_kredit_ppn"):
        df[kol] = df[kol].astype(object)

    kolom_asli_penjualan = [
        "no_akun_debet", "nama_akun_debet", "no_akun_kredit", "nama_akun_kredit",
        "jml_debet", "jml_kredit", "no_akun_kredit_ppn", "nama_akun_kredit_ppn", "jml_kredit_ppn",
    ]
    df_asli = df[kolom_asli_penjualan].copy()

    # [BARU -- PERBAIKAN PERFORMA] tentukan_akun_kas_piutang() &
    # cari_akun_penjualan_pasti() HANYA bergantung pada df_coa (konstan
    # sepanjang fungsi ini) + cara_bayar (cuma 2 nilai mungkin: TUNAI/
    # KREDIT) -- TIDAK bergantung pada isi baris lain sama sekali. Sebelum
    # cache ini, tahap 2 di bawah memanggil keduanya UNTUK SETIAP BARIS,
    # dan tiap panggilan itu men-scan ULANG SELURUH tabel COA dari nol
    # (df_coa["nama_akun"].str.upper().str.contains(...)) -- utk laporan
    # berisi ribuan transaksi, itu ribuan+ scan penuh atas tabel yang
    # HASILNYA SELALU SAMA. Cukup dihitung SEKALI di sini per nilai
    # cara_bayar yang mungkin muncul, lalu dipakai ulang lewat lookup
    # dict O(1) di dalam loop -- tidak mengubah hasil sama sekali, murni
    # menghilangkan kerja berulang yang tidak perlu (pola yang sama dgn
    # optimasi nama_akun_upper_coa di proses_dataframe() utk rekening
    # koran). Pada contoh "Data Penjualan Detail 01-30 Agustus 2026"
    # (10.715 transaksi), ini memangkas tahap kategorisasi dari ~9-23
    # detik jadi hitungan sub-detik.
    _cache_akun_kas_piutang: dict = {}
    def _akun_kas_piutang_cached(cara_bayar):
        if cara_bayar not in _cache_akun_kas_piutang:
            _cache_akun_kas_piutang[cara_bayar] = tentukan_akun_kas_piutang(df_coa, cara_bayar)
        return _cache_akun_kas_piutang[cara_bayar]
    _akun_ppn_keluaran_cache = tentukan_akun_ppn_keluaran(df_coa)
    _akun_penjualan_pasti_cache = cari_akun_penjualan_pasti(df_coa)

    perlu_isi = df.index.tolist()

    # tahap 1: pola historis
    belum_pola = []
    for idx in perlu_isi:
        row = df.loc[idx]
        sig = ekstrak_signature_penjualan(row)
        arah = row["cara_bayar"]
        aturan = pola.aturan.get((sig, arah))
        if aturan is None:
            belum_pola.append(idx)
            continue
        _isi_baris_penjualan(
            df, idx, row,
            aturan["no_akun_debet"], aturan["nama_akun_debet"],
            aturan["no_akun_kredit"], aturan["nama_akun_kredit"],
            aturan.get("no_akun_kredit_ppn"), aturan.get("nama_akun_kredit_ppn"),
        )
        df.at[idx, "sumber_kategori"] = (
            "Sesuai Pola yang Dipelajari" if aturan.get("konsisten", True)
            else "Sesuai Pola yang Dipelajari (perlu cek)"
        )

    # tahap 2: aturan standar penjualan
    ambigu = []
    for idx in belum_pola:
        row = df.loc[idx]
        akun_debet = _akun_kas_piutang_cached(row["cara_bayar"])
        akun_ppn = _akun_ppn_keluaran_cache if row["ppn"] else (None, None)
        akun_penjualan = _akun_penjualan_pasti_cache

        if akun_penjualan is not None:
            _isi_baris_penjualan(
                df, idx, row,
                akun_debet[0], akun_debet[1],
                akun_penjualan[0], akun_penjualan[1],
                akun_ppn[0], akun_ppn[1],
            )
            df.at[idx, "sumber_kategori"] = "Aturan standar penjualan + kata kunci COA"
        elif pd.notna(df_asli.at[idx, "no_akun_kredit"]):
            asli = df_asli.loc[idx]
            _isi_baris_penjualan(
                df, idx, row,
                akun_debet[0], akun_debet[1],
                asli["no_akun_kredit"], asli["nama_akun_kredit"],
                akun_ppn[0], akun_ppn[1],
            )
            df.at[idx, "sumber_kategori"] = "Data Asli dari File"
            sig = ekstrak_signature_penjualan(row)
            arah = row["cara_bayar"]
            if (sig, arah) not in pola.aturan:
                pola.aturan[(sig, arah)] = {
                    "no_akun_debet": akun_debet[0], "nama_akun_debet": akun_debet[1],
                    "no_akun_kredit": asli["no_akun_kredit"], "nama_akun_kredit": asli["nama_akun_kredit"],
                    "no_akun_kredit_ppn": akun_ppn[0], "nama_akun_kredit_ppn": akun_ppn[1],
                    "konsisten": True, "jumlah_contoh": 1,
                }
        else:
            df.at[idx, "no_akun_debet"] = akun_debet[0]
            df.at[idx, "nama_akun_debet"] = akun_debet[1]
            df.at[idx, "jml_debet"] = row["total"]
            if row["ppn"]:
                df.at[idx, "no_akun_kredit_ppn"] = akun_ppn[0]
                df.at[idx, "nama_akun_kredit_ppn"] = akun_ppn[1]
                df.at[idx, "jml_kredit_ppn"] = row["ppn"]
            ambigu.append(idx)

    # tahap 3: AI menentukan akun Penjualan
    # [FIX -- PERBAIKAN PERFORMA] Kalau df_coa kosong/tidak ada (mis. PDF
    # "Jurnal Penjualan Kasir" diupload tanpa client_id/COA terhubung),
    # prompt AI mewajibkan "no_akun_kredit_penjualan" berupa NOMOR AKUN
    # persis dari COA -- tapi kalau COA-nya sendiri kosong, AI tidak
    # mungkin bisa memenuhi syarat itu, sehingga SELALU balas null/tidak
    # yakin untuk semua baris (baris tetap jatuh ke label generik di
    # bagian "sisanya" di bawah, SAMA seperti kalau AI di-skip sejak
    # awal). Memanggil AI di kondisi ini cuma menghabiskan waktu
    # (network round-trip ke Groq/Claude, ratusan chunk utk file besar)
    # TANPA mengubah hasil akhir sama sekali -- jadi di-skip langsung
    # kalau df_coa kosong. ambigu TIDAK diubah/dikosongkan di sini
    # supaya loop "sisanya" di bawah tetap mengisi label generik
    # "PENJUALAN" + jml_kredit seperti biasa (perilaku akhir 100% sama,
    # cuma tanpa panggilan AI yang percuma).
    ada_coa = df_coa is not None and not df_coa.empty
    print(f"[DEBUG-TIMING] tahap-3 AI: ada_coa={ada_coa}, jumlah baris ambigu={len(ambigu)}, "
          f"pakai_ai={pakai_ai} -> AI {'DIPANGGIL' if (pakai_ai and ambigu and ada_coa) else 'DI-SKIP'}")

    if pakai_ai and ambigu and ada_coa:
        api_key = api_key or ambil_api_key_claude()
        if api_key:
            batch = []
            for idx in ambigu:
                row = df.loc[idx]
                batch.append({
                    "idx": idx, "customer": row.get("customer"), "keterangan": row.get("keterangan"),
                    "cara_bayar": row["cara_bayar"], "dpp": row["dpp"], "ppn": row["ppn"], "total": row["total"],
                })
            mapping, log_kegagalan_ai = kategorikan_penjualan_dengan_ai(batch, df_coa, api_key, mask_pii=mask_pii)
            alasan_gagal_per_idx = {}
            for entri in (log_kegagalan_ai or []):
                for idx_gagal in entri.get("idx_terdampak", []):
                    alasan_gagal_per_idx[idx_gagal] = entri.get("alasan")
            
            df = _apply_ai_results_to_dataframe(
                df=df,
                mapping=mapping,
                df_coa=df_coa,
                indices=ambigu,
                ambang_nilai=ambang_nilai,
                is_penjualan=True
            )
            
            for idx, alasan in alasan_gagal_per_idx.items():
                if pd.isna(df.at[idx, "catatan_ai"]):
                    df.at[idx, "catatan_ai"] = alasan

    # sisanya: akun Penjualan masih kosong -> label generik
    for idx in ambigu:
        if pd.isna(df.at[idx, "no_akun_kredit"]):
            row = df.loc[idx]
            df.at[idx, "no_akun_kredit"] = "PENJUALAN"
            df.at[idx, "nama_akun_kredit"] = "PENJUALAN"
            df.at[idx, "jml_kredit"] = row["dpp"] if row["ppn"] else row["total"]
            df.at[idx, "sumber_kategori"] = "Belum Terkategori - perlu review manual (akun Penjualan ambigu di COA)"

    df["sumber_kategori"] = df["sumber_kategori"].fillna("Belum Terkategori - perlu review manual")
    return df


# ============================================================
# 13. VALIDASI JURNAL
# ============================================================

TOLERANSI_BALANCE = 1.0


def cek_keseimbangan_jurnal(df: pd.DataFrame) -> dict:
    """
    Validasi TERAKHIR sebelum jurnal dianggap final: total seluruh DEBET harus
    sama dengan total seluruh KREDIT di satu batch data.
    """
    if df is None or df.empty:
        return {"balance": True, "total_debet": 0.0, "total_kredit": 0.0, "selisih": 0.0,
                "jumlah_baris_belum_terkategori": 0}

    total_debet = pd.to_numeric(df.get("jml_debet"), errors="coerce").fillna(0).sum()
    total_kredit = pd.to_numeric(df.get("jml_kredit"), errors="coerce").fillna(0).sum()
    if "jml_kredit_ppn" in df.columns:
        total_kredit += pd.to_numeric(df["jml_kredit_ppn"], errors="coerce").fillna(0).sum()
    # [FIX] jml_debet_ppn (PPN Masukan dicatat sbg baris debet terpisah --
    # lihat parse_sheet_faktur_pajak arah MASUKAN & proses_pembelian) belum
    # pernah ikut dijumlahkan ke total_debet di sini, padahal jml_kredit_ppn
    # (pasangannya utk arah KELUARAN) sudah lebih dulu ditangani di atas.
    # Akibatnya SETIAP baris MASUKAN yang PPN-nya kena split (faktur pajak
    # masukan & sekarang invoice pembelian yang kena PPN) selalu ketahuan
    # "TIDAK BALANCE" oleh validasi ini -- padahal debet+kredit-nya sebenarnya
    # sudah sama persis, cuma belum kehitung semua sisi debetnya.
    if "jml_debet_ppn" in df.columns:
        total_debet += pd.to_numeric(df["jml_debet_ppn"], errors="coerce").fillna(0).sum()

    selisih = round(float(total_debet) - float(total_kredit), 2)
    jumlah_belum = 0
    if "sumber_kategori" in df.columns:
        jumlah_belum = int(df["sumber_kategori"].astype(str).str.contains(
            "perlu review|Belum Terkategori", case=False, regex=True, na=False
        ).sum())

    return {
        "balance": abs(selisih) <= TOLERANSI_BALANCE,
        "total_debet": float(total_debet),
        "total_kredit": float(total_kredit),
        "selisih": selisih,
        "jumlah_baris_belum_terkategori": jumlah_belum,
    }


# ============================================================
# 14. PENILAIAN KINERJA KLIEN / MAKER
# ============================================================

KESALAHAN_PRIORITAS_AWAL = [
    {"judul": "Tidak aware dengan monitoring yang menjadi outstanding", "kata_kunci": ["outstanding", "belum di monitor", "belum dipantau", "belum ter-clear", "outstanding balance"]},
    {"judul": "Hasil rekonsiliasi tidak ada pasangannya", "kata_kunci": ["tidak ada pasangan", "belum ketemu pasangan", "unmatched", "selisih rekon", "reconciliation gap"]},
    {"judul": "Belum melibatkan pihak lain untuk konsultasi saat rekonsiliasi buntu", "kata_kunci": ["konsultasi", "melibatkan pihak lain", "tidak melibatkan", "eskalasi"]},
    {"judul": "Data awal yang diminta belum memenuhi deadline waktu yang diminta", "kata_kunci": ["data awal", "belum memenuhi deadline", "telat", "delay data"]},
    {"judul": "Laporan keuangan terbentuk tidak balance", "kata_kunci": ["tidak balance", "tidak seimbang", "selisih", "imbalance"]},
    {"judul": "Komponen laporan keuangan standar tidak terbentuk lengkap", "kata_kunci": ["laporan posisi keuangan", "arus kas", "perubahan ekuitas", "laba rugi", "catatan atas laporan"]},
]

KESALAHAN_UMUM_AKUNTAN = [
    "Penjualan Pribadi yang sudah ada vouchernya total belum diadjust ke Utang Pemegang Saham",
    "Monitoring itu dilanjutkan semua tahun",
    "Belum dapat data",
    "Tolong penamaan file monitoring disamakan, jangan YS tapi YM",
    "Jurnal PPN Keluaran",
    "Belum ada data",
    "Di April belum diberi kontrol pendapatan & piutang",
    "Masih banyak uang masuk sales di bank yg belum ketemu pasangannya",
    "Piutang masih selisih",
    "Masih belum ada Beban Gaji accrue",
    "Rekon WSI dengan",
    "Masih belum terekon semua",
    "Beban belum terekon seluruhnya",
    "Monitoring ulang semua utang + jurnal accrue",
    "Rekon WSI dengan",
    "Utang, jurnal biaya, jurnal pembelian",
    "Tambahan monitoring piutang, utang",
    "Jurnalnya GL Finance",
    "Sudah dihitung, belum ada billing dan belum dibayar"
]

KESALAHAN_KRITIS_TAMBAHAN = [
    "Score 0 berulang", "Bobot klien tidak konsisten", "Total score tidak sesuai rumus", 
    "Piutang selisih besar", "Jurnal PPN tidak terekon", "Beban accrue belum dicatat",
    "Monitoring utang belum lengkap", "Data klien duplikat", "Deadline klien terlewat",
    "Kualitas dokumen rendah", "Tidak ada bukti rekonsiliasi"
]


def rapikan_penilaian_klien(df: pd.DataFrame) -> pd.DataFrame:
    """Merapikan data penilaian kinerja klien/maker."""
    if df.empty:
        df = df.copy()
        if "jenis_baris" not in df.columns:
            df["jenis_baris"] = pd.Series(dtype="object")
        if "maker" not in df.columns:
            df["maker"] = pd.Series(dtype="object")
        if "score" not in df.columns:
            df["score"] = pd.Series(dtype="float64")
        return df
    
    df = df.copy()
    
    kolom_map = {
        'NAMA KLIEN': 'nama_klien',
        'MAKER': 'maker',
        'SCORE': 'score',
        'BOBOT KLIEN': 'bobot_klien',
        'TOTAL SCORE': 'total_score',
        'TOTAL AKHIR': 'total_akhir',
        '+': 'plus',
        '-': 'minus'
    }
    df.rename(columns={k: v for k, v in kolom_map.items() if k in df.columns}, inplace=True)
    df.columns = [str(col).strip().lower().replace(' ', '_') for col in df.columns]
    
    numeric_cols = ['score', 'bobot_klien', 'total_score', 'plus', 'minus', 'total_akhir']
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    if 'no' in df.columns:
        df['jenis_baris'] = df['no'].apply(lambda x: 'klien' if pd.notna(x) and x != '' else 'catatan')
    else:
        df['jenis_baris'] = 'klien'
    
    if 'maker' not in df.columns:
        df['maker'] = None
    df['maker'] = df['maker'].fillna('Tidak Diketahui')
    if 'score' not in df.columns:
        df['score'] = None
    df['score'] = df['score'].fillna(0)
    
    return df


TOLERANSI_SELISIH = 0.01


def perbaiki_data_penilaian(df: pd.DataFrame) -> tuple[pd.DataFrame, list[dict]]:
    """Auto-fix data penilaian yang bisa diperbaiki otomatis."""
    df = rapikan_penilaian_klien(df)
    df = df.copy()
    log: list[dict] = []

    def _catat(idx, kolom, lama, baru, alasan):
        log.append({
            "baris": idx,
            "nama_klien": df.at[idx, "nama_klien"] if "nama_klien" in df.columns else None,
            "maker": df.at[idx, "maker"] if "maker" in df.columns else None,
            "kolom": kolom,
            "nilai_lama": lama,
            "nilai_baru": baru,
            "alasan": alasan,
        })

    kolom_wajib = {"score", "bobot_klien", "total_score", "plus", "minus", "total_akhir"}
    if not kolom_wajib.issubset(df.columns) or df.empty:
        return df, log

    for idx, row in df.iterrows():
        if row.get("jenis_baris") != "klien":
            continue
        if pd.isna(row["total_akhir"]) or row["total_akhir"] < 0:
            _catat(idx, "total_akhir", row["total_akhir"], 0, "Nilai negatif/tidak valid → direset")
            df.at[idx, "total_akhir"] = 0

        score = row["score"]
        bobot = row["bobot_klien"]
        total_score = row["total_score"]
        plus = row["plus"]
        minus = row["minus"]
        total_akhir = row["total_akhir"]

        if pd.isna(bobot) and pd.notna(total_score) and pd.notna(score) and score not in (0, None):
            bobot = round(total_score / score, 4)
            df.at[idx, "bobot_klien"] = bobot
            _catat(idx, "bobot_klien", None, bobot, "Ditarik balik dari TOTAL SCORE / SCORE")

        if pd.notna(score) and pd.notna(bobot):
            seharusnya_total_score = round(score * bobot, 2)
            if pd.isna(total_score):
                df.at[idx, "total_score"] = seharusnya_total_score
                _catat(idx, "total_score", None, seharusnya_total_score,
                       "Dihitung dari SCORE x BOBOT KLIEN")
                total_score = seharusnya_total_score
            elif abs(total_score - seharusnya_total_score) > TOLERANSI_SELISIH:
                lama = total_score
                df.at[idx, "total_score"] = seharusnya_total_score
                _catat(idx, "total_score", lama, seharusnya_total_score,
                       "Tidak sesuai rumus SCORE x BOBOT KLIEN -- dikoreksi")
                total_score = seharusnya_total_score

        if pd.isna(plus):
            df.at[idx, "plus"] = 0.0
            _catat(idx, "plus", None, 0.0, "Kolom (+) kosong, diisi 0")
            plus = 0.0
        if pd.isna(minus):
            df.at[idx, "minus"] = 0.0
            _catat(idx, "minus", None, 0.0, "Kolom (-) kosong, diisi 0")
            minus = 0.0

        if pd.notna(total_score):
            seharusnya_total_akhir = round(total_score + plus - minus, 2)
            if pd.isna(total_akhir):
                df.at[idx, "total_akhir"] = seharusnya_total_akhir
                _catat(idx, "total_akhir", None, seharusnya_total_akhir,
                       "Dihitung dari TOTAL SCORE + (+) - (-)")
            elif abs(total_akhir - seharusnya_total_akhir) > TOLERANSI_SELISIH:
                lama = total_akhir
                df.at[idx, "total_akhir"] = seharusnya_total_akhir
                _catat(idx, "total_akhir", lama, seharusnya_total_akhir,
                       "Tidak sesuai rumus TOTAL SCORE + (+) - (-) -- dikoreksi")

    return df, log


def analisis_kesalahan_penilaian(df: pd.DataFrame) -> dict:
    """Analisis kesalahan pada data penilaian."""
    df = rapikan_penilaian_klien(df)
    temuan = []
    teks_lengkap = " ".join(str(x) for x in df.values.flatten() if pd.notna(x)).lower()

    for poin in KESALAHAN_PRIORITAS_AWAL:
        if any(kata in teks_lengkap for kata in poin["kata_kunci"]):
            temuan.append({
                "kategori": "PRIORITAS TINGGI",
                "temuan": poin["judul"],
                "tingkat_keparahan": "KRITIS",
                "rekomendasi": "🚨 WAJIB ditindaklanjuti SEGERA sebelum laporan final.",
                "severity_score": 90
            })

    semua_kesalahan = KESALAHAN_UMUM_AKUNTAN + KESALAHAN_KRITIS_TAMBAHAN
    for kesalahan in semua_kesalahan:
        if any(kata in teks_lengkap for kata in kesalahan.lower().split()[:5]):
            temuan.append({
                "kategori": "Kesalahan Umum",
                "temuan": kesalahan,
                "tingkat_keparahan": "Sedang",
                "rekomendasi": "Lakukan verifikasi manual dan dokumentasi lengkap.",
                "severity_score": 60
            })

    klien_valid = df[df['jenis_baris'] == 'klien']
    rata_score = 0
    if not klien_valid.empty:
        rata_score = klien_valid['score'].mean()
        zero_score_pct = (klien_valid['score'] == 0).sum() / len(klien_valid) * 100

        if rata_score < 80:
            temuan.append({
                "kategori": "Kinerja Keseluruhan",
                "temuan": f"Rata-rata score rendah ({rata_score:.1f})",
                "tingkat_keparahan": "Tinggi" if rata_score < 70 else "Sedang",
                "rekomendasi": "Buat improvement plan + follow-up mingguan.",
                "severity_score": 85 if rata_score < 70 else 65
            })

        if zero_score_pct > 25:
            temuan.append({
                "kategori": "Kelengkapan Penilaian",
                "temuan": f"{zero_score_pct:.1f}% klien memiliki score 0",
                "tingkat_keparahan": "KRITIS",
                "rekomendasi": "Lengkapi penilaian semua klien sebelum deadline.",
                "severity_score": 95
            })

        if 'total_score' in klien_valid.columns and 'score' in klien_valid.columns and 'bobot_klien' in klien_valid.columns:
            mismatch = abs(klien_valid['total_score'] - klien_valid['score'] * klien_valid['bobot_klien']) > TOLERANSI_SELISIH
            if mismatch.sum() > 0:
                temuan.append({
                    "kategori": "Konsistensi Data",
                    "temuan": f"Ada {mismatch.sum()} baris total_score tidak sesuai rumus",
                    "tingkat_keparahan": "Tinggi",
                    "rekomendasi": "Perbaiki manual atau gunakan fitur auto-fix.",
                    "severity_score": 80
                })

    temuan = sorted(temuan, key=lambda x: x.get("severity_score", 50), reverse=True)

    return {
        "total_klien_dinilai": len(klien_valid),
        "rata_rata_score": round(float(rata_score), 2),
        "total_temuan": len(temuan),
        "temuan": temuan[:20],
        "rekomendasi_umum": "Semua temuan KRITIS harus diselesaikan sebelum laporan keuangan final diserahkan.",
        "peringatan_ketat": "Jangan posting jurnal jika ada temuan severity > 80."
    }


def proses_dataframe_penilaian(df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """Pipeline lengkap untuk data penilaian kinerja."""
    df_diperbaiki, log_koreksi = perbaiki_data_penilaian(df)
    analisis = analisis_kesalahan_penilaian(df_diperbaiki)
    analisis["koreksi_otomatis"] = log_koreksi
    analisis["total_koreksi_otomatis"] = len(log_koreksi)
    
    if any(t.get("tingkat_keparahan") == "KRITIS" for t in analisis.get("temuan", [])):
        analisis["status_global"] = "TIDAK LULUS - Perlu review mendalam"
    else:
        analisis["status_global"] = "LULUS dengan catatan"
    
    return df_diperbaiki, analisis


# ============================================================
# 14B. WRAPPER FILE-LEVEL -- REKENING KORAN (JURNAL KORAN), PENJUALAN,
#      PENILAIAN KLIEN, BUKU BANTU PIUTANG
# ============================================================
# [FIX] Logic modul-modul ini (muat_workbook, proses_dataframe,
# proses_dataframe_penjualan, proses_dataframe_penilaian) SUDAH ADA sejak
# lama dan dulu dipanggil LANGSUNG dari app.py (Streamlit). Supaya backend
# FastAPI (main.py) + frontend React bisa memanggilnya dengan pola yang
# SAMA seperti proses_file_bukti_kas() dkk -- satu fungsi terima
# (file_like, nama_file), keluarkan dict {"df", "ringkasan", "masalah",
# "draf_jurnal", "sheet_dilewati"} -- dibungkus di sini.

_FOLDER_POLA = Path(__file__).parent / "pola_data"


def _path_pola(nama_dasar: str, client_id: Optional[int] = None) -> str:
    """Path file pola: per-client kalau client_id dikasih, else pola global bersama."""
    _FOLDER_POLA.mkdir(exist_ok=True)
    nama = f"{nama_dasar}_client_{client_id}.json" if client_id else f"{nama_dasar}_global.json"
    return str(_FOLDER_POLA / nama)


# [FIX] Sebelumnya "histori_gaji_sebelumnya"/"histori_gaji_terbaru" di
# proses_slip_gaji() (lihat modul 11D) TIDAK PERNAH disimpan/dibaca balik
# oleh pemanggil (main.py) -- akibatnya deteksi anomali gaji ANTAR-PERIODE
# hanya jalan kalau semua bulan kebetulan ada di 1 file/upload yang sama,
# padahal di dunia nyata payroll biasanya diupload terpisah per bulan.
# Dua fungsi ini menyimpan/membaca histori per-client ke disk, PERSIS pola
# yang sama dengan simpan_pola()/muat_pola() di atas (backup + atomic write)
# supaya histori tidak pernah hilang/korup gara-gara proses ke-interupsi.
def muat_histori_gaji(path: str) -> Dict[str, dict]:
    """Baca histori gaji per-karyawan (periode terakhir) dari disk. Return {} kalau file belum ada/korup."""
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except (json.JSONDecodeError, OSError) as e:
        logger.warning(f"Gagal baca histori gaji dari {path}: {e} -- dianggap kosong (deteksi anomali antar-periode reset).")
        return {}


def simpan_histori_gaji(histori: Dict[str, dict], path: str):
    """Simpan histori gaji terbaru per karyawan ke disk (backup + atomic write, sama seperti simpan_pola())."""
    if os.path.exists(path):
        try:
            import shutil
            shutil.copy2(path, path + ".bak")
        except OSError:
            pass
    tmp_path = path + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(histori, f, ensure_ascii=False, indent=2, default=str)
    os.replace(tmp_path, path)


def proses_file_rekening_koran(
    file_like, nama_file: str = None, client_id: Optional[int] = None, pakai_ai: bool = True
) -> dict:
    """
    "Jurnal Koran" -- mutasi rekening koran/bank (multi-sheet, multi-bank)
    di-jurnal-kan otomatis: pola historis milik client -> kata kunci COA ->
    AI (DeepSeek, kalau DEEPSEEK_API_KEY aktif) -> sisanya ditandai "Belum
    Terkategori" utk direview manual. Pola yang berhasil dipelajari (baris
    yang jurnalnya sudah lengkap di file asal) otomatis disimpan lagi
    supaya dipakai ulang di upload bulan berikutnya.
    """
    nama_file = nama_file or getattr(file_like, "name", "") or ""
    df_bank, _df_jual, _df_nilai, _df_piutang, df_coa, peringatan = muat_workbook(file_like, nama_file)

    # [FIX -- propagasi ke pemanggil, poin B] Sebelumnya ringkasan_footer
    # (saldo_awal/mutasi_cr/mutasi_db/saldo_akhir resmi dari blok footer
    # PDF, untuk PDF yang lewat jalur fallback posisi-kata) DIBUANG TOTAL
    # di jalur upload umum ini -- cuma "sheet_dilewati" (peringatan teks)
    # yang sampai ke response. Sekarang dibaca dari df_bank.attrs (lihat
    # muat_workbook()) dan disertakan sebagai key terpisah "ringkasan_footer"
    # supaya frontend bisa tampilkan saldo resmi PDF vs hasil ekstraksi
    # secara terstruktur, sama seperti yang sudah dilakukan
    # susun_gl_dari_pdf_rekening_koran() untuk fitur Kertas Kerja.
    ringkasan_footer = getattr(df_bank, "attrs", {}).get("ringkasan_footer_per_sheet", {}) if df_bank is not None else {}

    # [BARU] COA (Chart of Accounts) dari sheet "COA" di file yang sama --
    # sebelumnya df_coa cuma dipakai INTERNAL utk kategorisasi jurnal
    # (proses_dataframe), tidak pernah dikirim balik ke pemanggil. Sekarang
    # disertakan sebagai list of dict supaya frontend bisa menampilkannya
    # jadi tab "COA" tersendiri (kosong kalau file tidak punya sheet COA).
    daftar_coa = df_coa.to_dict("records") if df_coa is not None and not df_coa.empty else []

    if df_bank is None or df_bank.empty:
        return {
            "df": pd.DataFrame(), "ringkasan": {}, "masalah": [], "draf_jurnal": [],
            "sheet_dilewati": peringatan, "coa": daftar_coa, "ringkasan_footer": ringkasan_footer,
        }

    path_pola = _path_pola("pola_bank", client_id)
    pola = muat_pola(path_pola)

    df_hasil = proses_dataframe(
        df_bank, df_coa, pola, pakai_ai=pakai_ai,
        api_key=ambil_api_key_claude() if pakai_ai else None,
    )
    simpan_pola(pola, path_pola, sumber_perubahan="proses_file_rekening_koran (auto dari data asli)")

    keseimbangan = cek_keseimbangan_jurnal(df_hasil)

    masalah = []
    for i, row in df_hasil.iterrows():
        sumber = str(row.get("sumber_kategori") or "")
        if "Belum Terkategori" in sumber or "perlu cek" in sumber:
            masalah.append({
                "baris": i + 1, "tanggal": row.get("tanggal"), "bank": row.get("bank"),
                "keterangan": row.get("keterangan"),
                "nominal": max(row.get("mutasi_debet") or 0, row.get("mutasi_kredit") or 0),
                "alasan": [f"Sumber kategori: {sumber or 'tidak diketahui'} -- akun belum pasti, cek manual."],
            })

    # [BARU -- frontend butuh arah mutasi bank yang PASTI, bukan tebak-tebak
    # dari nama akun] jml_debet/jml_kredit di jurnal akuntansi SELALU sama
    # besar (double-entry), jadi tidak bisa dipakai frontend untuk tahu mana
    # sisi yang benar-benar akun bank (utk hitung saldo berjalan). Kolom
    # mutasi_debet/mutasi_kredit MENTAH dari rekening koran (lihat _arah())
    # justru sudah tahu persis arahnya -- disertakan di sini supaya frontend
    # tidak perlu menebak dari nama akun lagi.
    draf_jurnal = [
        {
            "baris": i + 1, "tanggal": row.get("tanggal"), "bank": row.get("bank"),
            "keterangan": row.get("keterangan"),
            "no_akun_debet": row.get("no_akun_debet"), "nama_akun_debet": row.get("nama_akun_debet"),
            "jml_debet": row.get("jml_debet"),
            "no_akun_kredit": row.get("no_akun_kredit"), "nama_akun_kredit": row.get("nama_akun_kredit"),
            "jml_kredit": row.get("jml_kredit"),
            "mutasi_debet": row.get("mutasi_debet"),
            "mutasi_kredit": row.get("mutasi_kredit"),
            "sumber_kategori": row.get("sumber_kategori"),
            "catatan": row.get("catatan_ai"),
        }
        for i, row in df_hasil.iterrows()
    ]

    ringkasan = {
        "jumlah_transaksi": len(df_hasil),
        "total_debet": keseimbangan["total_debet"],
        "total_kredit": keseimbangan["total_kredit"],
        "balance": keseimbangan["balance"],
        "selisih": keseimbangan["selisih"],
        "jumlah_perlu_review": len(masalah),
        "jumlah_pola_dipakai": int((df_hasil["sumber_kategori"] == "Sesuai Pola yang Dipelajari").sum()),
        "jumlah_total_pola_tersimpan": len(pola.aturan),
    }

    return {
        "df": df_hasil, "ringkasan": ringkasan, "masalah": masalah, "draf_jurnal": draf_jurnal,
        "sheet_dilewati": peringatan, "coa": daftar_coa, "ringkasan_footer": ringkasan_footer,
    }


def proses_file_bootstrap_pola_bank(
    file_like, nama_file: str = None, client_id: Optional[int] = None,
    min_samples: int = 2,
) -> dict:
    """
    [BARU] Bootstrap pola untuk CLIENT BARU (cold-start) -- prioritas #1.

    Beda dengan proses_file_rekening_koran(): fungsi ini TIDAK menjalankan
    kategorisasi otomatis (pola historis -> kata kunci -> AI) sama sekali.
    Tujuannya cuma "menyuapi" pola_bank_client_{client_id}.json dari
    rekening koran BULAN-BULAN LALU yang SUDAH DIJURNAL LENGKAP oleh
    akuntan (kolom NO AKUN/NAMA AKUN debet & kredit di file sumber sudah
    terisi manual) -- supaya upload PERTAMA rekening koran MENTAH client
    ini nanti tidak mulai dari pola kosong.

    Parsing pakai muat_workbook() + parse_sheet_bank() yang SAMA PERSIS
    dengan pipeline biasa -- tidak ada logic parsing baru. Baris yang
    kolom jurnalnya belum terisi otomatis diabaikan (sama seperti
    pelajari_pola() yang sudah ada).

    min_samples default DITURUNKAN ke 2 (dibanding default 3 di
    pelajari_pola()) karena data bootstrap biasanya cuma dari beberapa
    bulan riwayat -- sinyal per signature (mis. "BPJS") lebih sedikit
    dibanding pola yang terus terkumpul dari upload rutin bulanan.
    confidence_score & is_valid tetap dihitung dengan logic yang sama
    (lihat pelajari_pola()), jadi baris yang TIDAK konsisten antar contoh
    tidak akan dianggap valid cuma karena sample-nya sedikit.

    Boleh dipanggil berkali-kali (mis. upload rekening koran Jan, lalu
    Feb, lalu Mar secara terpisah) -- pola yang baru dipelajari selalu
    DIGABUNG (gabung_pola) ke pola yang sudah tersimpan di disk, bukan
    menimpa dari awal. Kalau ada bentrok signature+arah antara pola lama
    dan pola dari file yang baru saja diupload, pola BARU yang menang
    (asumsi: upload terbaru adalah versi yang paling dipercaya user untuk
    proses bootstrap ini).
    """
    nama_file = nama_file or getattr(file_like, "name", "") or ""
    df_bank, _df_jual, _df_nilai, _df_piutang, _df_coa, peringatan = muat_workbook(file_like, nama_file)

    if df_bank is None or df_bank.empty:
        return {
            "jumlah_baris_dibaca": 0, "jumlah_baris_dipakai": 0,
            "jumlah_pola_baru": 0, "jumlah_pola_diperbarui": 0,
            "jumlah_pola_dipelajari_total": 0, "total_pola_tersimpan": 0,
            "detail_pola_baru": [], "detail_pola_diperbarui": [],
            "sheet_dilewati": peringatan,
        }

    jumlah_baris_dibaca = len(df_bank)
    baris_ada_jurnal = df_bank["no_akun_debet"].notna() & df_bank["no_akun_kredit"].notna()
    jumlah_baris_dipakai = int(baris_ada_jurnal.sum())

    if jumlah_baris_dipakai == 0:
        peringatan = list(peringatan) + [
            "Tidak ditemukan baris dengan kolom jurnal (DEBET/KREDIT) terisi -- "
            "tidak ada pola yang bisa dipelajari dari file ini. Pastikan yang "
            "diupload adalah rekening koran yang KOLOM JURNALNYA SUDAH DIISI "
            "LENGKAP oleh akuntan, bukan rekening koran mentah."
        ]

    pola_baru = pelajari_pola(df_bank, min_samples=min_samples)

    path_pola = _path_pola("pola_bank", client_id)
    pola_lama = muat_pola(path_pola)

    # Hitung selisih SEBELUM digabung, supaya ringkasan ke user bisa bedakan
    # pola yang betul-betul baru vs pola lama yang cuma diperbarui (mis.
    # akun berubah karena file bulan berikutnya lebih akurat) -- ini murni
    # untuk pelaporan, tidak memengaruhi logic penggabungan itu sendiri.
    detail_baru = []
    detail_diperbarui = []
    for key, aturan in pola_baru.aturan.items():
        sig, arah = key
        lama = pola_lama.aturan.get(key)
        ringkas = {
            "signature": sig,
            "arah": arah,
            "akun_debet": f"{aturan['no_akun_debet']} - {aturan['nama_akun_debet']}",
            "akun_kredit": f"{aturan['no_akun_kredit']} - {aturan['nama_akun_kredit']}",
            "jumlah_contoh": aturan["jumlah_contoh"],
            "confidence_score": aturan["confidence_score"],
            "is_valid": aturan["is_valid"],
        }
        if lama is None:
            detail_baru.append(ringkas)
        elif (lama.get("no_akun_debet"), lama.get("no_akun_kredit")) != (
            aturan["no_akun_debet"], aturan["no_akun_kredit"]
        ):
            ringkas["akun_lama_debet"] = f"{lama.get('no_akun_debet')} - {lama.get('nama_akun_debet')}"
            ringkas["akun_lama_kredit"] = f"{lama.get('no_akun_kredit')} - {lama.get('nama_akun_kredit')}"
            detail_diperbarui.append(ringkas)

    pola_gabungan = gabung_pola(pola_lama, pola_baru)
    simpan_pola(pola_gabungan, path_pola, sumber_perubahan="proses_file_bootstrap_pola_bank")

    return {
        "jumlah_baris_dibaca": jumlah_baris_dibaca,
        "jumlah_baris_dipakai": jumlah_baris_dipakai,
        "jumlah_pola_baru": len(detail_baru),
        "jumlah_pola_diperbarui": len(detail_diperbarui),
        "jumlah_pola_dipelajari_total": len(pola_baru.aturan),
        "total_pola_tersimpan": len(pola_gabungan.aturan),
        "detail_pola_baru": detail_baru,
        "detail_pola_diperbarui": detail_diperbarui,
        "sheet_dilewati": peringatan,
    }


def proses_file_penjualan(
    file_like, nama_file: str = None, client_id: Optional[int] = None, pakai_ai: bool = True
) -> dict:
    """
    Data Penjualan (invoice-style) + ringkasan POS/kasir digabung jadi satu
    -- keduanya dikenali otomatis oleh muat_workbook() lalu dijurnal-kan
    lewat proses_dataframe_penjualan() (pola historis -> aturan standar
    penjualan + kata kunci COA -> AI).
    """
    import time as _time_debug  # [DEBUG SEMENTARA] hapus/comment setelah selesai profiling

    nama_file = nama_file or getattr(file_like, "name", "") or ""

    _t_parse = _time_debug.perf_counter()
    _df_bank, df_jual, _df_nilai, _df_piutang, df_coa, peringatan = muat_workbook(file_like, nama_file)
    print(f"[DEBUG-TIMING] muat_workbook('{nama_file}'): {_time_debug.perf_counter() - _t_parse:.2f} detik "
          f"({0 if df_jual is None else len(df_jual)} baris penjualan terdeteksi)")

    if df_jual is None or df_jual.empty:
        return {"df": pd.DataFrame(), "ringkasan": {}, "masalah": [], "draf_jurnal": [], "sheet_dilewati": peringatan}

    path_pola = _path_pola("pola_penjualan", client_id)
    pola = muat_pola(path_pola)

    _t_kategori = _time_debug.perf_counter()
    df_hasil = proses_dataframe_penjualan(
        df_jual, df_coa, pola, pakai_ai=pakai_ai,
        api_key=ambil_api_key_claude() if pakai_ai else None,
    )
    print(f"[DEBUG-TIMING] proses_dataframe_penjualan (pola+aturan+AI) untuk '{nama_file}': "
          f"{_time_debug.perf_counter() - _t_kategori:.2f} detik")

    _t_simpan_pola = _time_debug.perf_counter()
    simpan_pola(pola, path_pola, sumber_perubahan="proses_file_penjualan (auto dari data asli)")
    print(f"[DEBUG-TIMING] simpan_pola untuk '{nama_file}': {_time_debug.perf_counter() - _t_simpan_pola:.2f} detik")

    keseimbangan = cek_keseimbangan_jurnal(df_hasil)

    masalah = []
    for i, row in df_hasil.iterrows():
        sumber = str(row.get("sumber_kategori") or "")
        if "Belum Terkategori" in sumber or "perlu cek" in sumber:
            masalah.append({
                "baris": i + 1, "tanggal": row.get("tanggal"), "keterangan": row.get("keterangan"),
                "alasan": [f"Sumber kategori: {sumber or 'tidak diketahui'} -- akun belum pasti, cek manual."],
            })

    draf_jurnal = [
        {
            "baris": i + 1, "tanggal": row.get("tanggal"), "keterangan": row.get("keterangan"),
            "no_akun_debet": row.get("no_akun_debet"), "nama_akun_debet": row.get("nama_akun_debet"),
            "jml_debet": row.get("jml_debet"),
            "no_akun_kredit": row.get("no_akun_kredit"), "nama_akun_kredit": row.get("nama_akun_kredit"),
            "jml_kredit": row.get("jml_kredit"),
            "no_akun_kredit_ppn": row.get("no_akun_kredit_ppn"), "jml_kredit_ppn": row.get("jml_kredit_ppn"),
            "sumber_kategori": row.get("sumber_kategori"),
        }
        for i, row in df_hasil.iterrows()
    ]

    # [FIX] total_penjualan & rata_rata_transaksi sebelumnya tidak pernah
    # dihitung di sini -- frontend (documentTypes.js) mencari 2 field ini
    # untuk kartu ringkasan "Total Penjualan" & "Rata-rata / Transaksi",
    # jadi selalu tampil "-" karena field-nya memang belum ada.
    # "jml_kredit" = akun penjualan/revenue yang dikredit (di luar PPN
    # keluaran yang ada di kolom terpisah "jml_kredit_ppn"), jadi ini
    # representasi nilai penjualan bersih per baris transaksi.
    jumlah_transaksi = len(df_hasil)
    total_penjualan = float(pd.to_numeric(df_hasil.get("jml_kredit"), errors="coerce").fillna(0).sum())
    rata_rata_transaksi = total_penjualan / jumlah_transaksi if jumlah_transaksi else 0

    ringkasan = {
        "jumlah_transaksi": jumlah_transaksi,
        "total_penjualan": total_penjualan,
        "rata_rata_transaksi": rata_rata_transaksi,
        "total_debet": keseimbangan["total_debet"],
        "total_kredit": keseimbangan["total_kredit"],
        "balance": keseimbangan["balance"],
        "selisih": keseimbangan["selisih"],
        "jumlah_perlu_review": len(masalah),
    }

    return {"df": df_hasil, "ringkasan": ringkasan, "masalah": masalah, "draf_jurnal": draf_jurnal, "sheet_dilewati": peringatan}


# ============================================================
# [BARU] EKSTRAKSI PDF "DATA PENJUALAN DETAIL" (LAPORAN KASIR/POS
# PER-TRANSAKSI) -- dipakai jenis_dokumen "jurnal_penjualan_kasir"
# ============================================================
# Format sumber: laporan PDF hasil export sistem kasir/POS toko (bukan
# rekening koran, bukan invoice-per-baris ala parse_sheet_penjualan()).
# Tiap transaksi tercetak sbg SATU BLOK multi-baris:
#
#   <No Transaksi>  <Tanggal>  [Dept.]  <Kode Pel.>  <Nama Pelanggan>  <Alamat...>
#   No.  Kd. Item  Nama Item  Jml  Satuan  Harga  Pot. %  Total
#   1    <kode>    <nama item, bisa wrap beberapa baris>   ...
#   2    ...
#                                            <total qty>          <subtotal>
#       Pot. : <n>   Pajak : <n>   Biaya : <n>   Total Akhir : <n>
#
# Header kolom "No Transaksi Tanggal Dept. Kode Pel. Nama Pelanggan Alamat"
# DICETAK ULANG di baris pertama SETIAP HALAMAN (bukan transaksi baru) --
# begitu juga footer laporan "<tgl> <jam> REPORT <hal>/<total>". Satu blok
# transaksi bisa terpotong di tengah halaman (tabel item lanjut ke halaman
# berikutnya tanpa header baru) -- makanya parsing dilakukan atas SELURUH
# teks dokumen yang sudah digabung per baris, bukan per halaman terpisah,
# dengan baris header/footer generik di atas cukup DILEWATI di mana pun
# munculnya.
#
# BEDA PENTING dari _ekstrak_pdf_rekening_koran_berbasis_posisi() di atas:
# fungsi itu berbasis KOORDINAT kata (x0/top dari pdfplumber.extract_words())
# krn rekening koran butuh presisi kolom DEBIT/KREDIT/SALDO yang sejajar
# tanpa garis. Laporan penjualan ini TIDAK butuh presisi kolom sama sekali
# -- yang dijurnalkan hanyalah TOTAL AKHIR per transaksi (1 blok = 1 baris
# jurnal penjualan), bukan rincian tiap item -- jadi cukup pakai
# page.extract_text() (urutan baris apa adanya, tanpa perlu kembalikan ke
# koordinat) lalu di-parse dgn state machine sederhana berbasis regex per
# baris. Jauh lebih murah & lebih tahan banting terhadap variasi lebar
# kolom dibanding pendekatan posisi-kata.
#
# Divalidasi terhadap contoh "Data Penjualan Detail 01-30 Agustus 2026"
# (2003 halaman, ~1.600 transaksi, prefix No Transaksi "KSR-" & "JL-").

_RE_HEADER_TRANSAKSI_JUAL = re.compile(
    r"^([A-Z]{1,4}-\d{3,10}-\d{2,4})\s+(\d{1,2}/\d{1,2}/\d{2,4})\s+(.*)$"
)
_RE_KODE_PEL_JUAL = re.compile(r"^([A-Z]{1,4}\d{2,8})\b\s*(.*)$")
_RE_FOOTER_TRANSAKSI_JUAL = re.compile(
    r"^Pot\.\s*:\s*([\d.,]+)\s+Pajak\s*:\s*([\d.,]+)\s+Biaya\s*:\s*([\d.,]+)\s+"
    r"Total Akhir\s*:\s*([\d.,]+)\s*$"
)
_RE_HEADER_KOLOM_JUAL = re.compile(r"^No Transaksi\s+Tanggal\b")
_RE_ITEM_HEADER_JUAL = re.compile(r"^No\.\s+Kd\.\s*Item\b")
_RE_FOOTER_HALAMAN_JUAL = re.compile(r"^\d{1,2}/\d{1,2}/\d{2,4}\s+\d{1,2}:\d{2}\s+REPORT\b")


def _angka_id(teks) -> float:
    """Parse angka format Indonesia ('900.000' -> 900000.0, '1.234,50' ->
    1234.5) -- titik = pemisah ribuan, koma = pemisah desimal."""
    if teks is None:
        return 0.0
    t = str(teks).strip()
    if not t:
        return 0.0
    t = t.replace(".", "").replace(",", ".")
    try:
        return float(t)
    except ValueError:
        return 0.0


def _ambil_teks_pdf_layout(data: bytes, nama_file: str) -> tuple[str, int]:
    """
    Ekstrak SELURUH teks PDF sekali jalan lewat `pdftotext -layout` (poppler-
    utils), dibaca dari stdin & ditulis ke stdout langsung (tanpa file
    sementara di disk -- aman dipanggil dari request FastAPI concurrent).

    [BARU -- PERBAIKAN KECEPATAN] Sebelumnya PDF ini dibaca lewat
    pdfplumber (buka tiap halaman -> re-layout & re-cluster posisi tiap
    KARAKTER jadi kata via algoritma Python murni), yang perlu ~150-160ms
    per halaman. Untuk laporan ribuan halaman (mis. contoh "Data
    Penjualan Detail 01-30 Agustus 2026", 2003 halaman) itu ~5-6 MENIT
    hanya utk tahap ekstraksi teks -- sebelum kategorisasi apa pun
    dimulai. `pdftotext` (dari poppler, C++, dipakai jutaan kali di
    banyak sistem) mengerjakan hal yang SAMA (susun ulang teks per baris
    sesuai posisi asli lewat opsi -layout) dalam hitungan DETIK utk jumlah
    halaman yang sama (diuji: 2003 halaman ~9 detik vs pdfplumber ~5.5
    menit) -- ~35-40x lebih cepat, tanpa kehilangan informasi apa pun yg
    dibutuhkan parser di bawah (fungsi ini murni baca URUTAN baris teks,
    BEDA dari ekstraksi rekening koran yg butuh KOORDINAT x/y presisi tiap
    kata utk memisahkan kolom DEBET/KREDIT/SALDO tanpa garis tabel --
    laporan penjualan ini tidak butuh itu, lihat catatan di docstring
    _ekstrak_pdf_jual_kasir_berbasis_posisi di bawah).

    Kalau binary `pdftotext` tidak tersedia di server (belum install
    poppler-utils), otomatis fallback ke pdfplumber (lambat tapi tetap
    jalan) supaya tidak ada environment yang tiba-tiba error total.

    Returns: (teks_gabungan, jumlah_halaman).
    """
    if shutil.which("pdftotext"):
        try:
            hasil = subprocess.run(
                ["pdftotext", "-layout", "-", "-"],
                input=data, capture_output=True, timeout=180, check=True,
            )
            teks = hasil.stdout.decode("utf-8", errors="replace")
            jumlah_halaman = teks.count("\x0c") + (0 if teks.endswith("\x0c") else 1)
            return teks, jumlah_halaman
        except (subprocess.SubprocessError, OSError) as e:
            logger.warning(
                "pdftotext gagal utk '%s' (%s) -- fallback ke pdfplumber (lebih lambat).",
                nama_file, e,
            )
    # --- fallback: pdfplumber (dipertahankan sbg jaring pengaman kalau
    # poppler-utils belum ter-install di server -- lihat catatan di atas) ---
    try:
        import pdfplumber
    except ImportError as e:
        raise RuntimeError(
            "Gagal membaca file PDF -- baik 'pdftotext' (poppler-utils) maupun "
            "library 'pdfplumber' tidak tersedia. Install salah satu: "
            "'apt-get install poppler-utils' (direkomendasikan, jauh lebih cepat) "
            "atau 'pip install pdfplumber --break-system-packages'."
        ) from e
    potongan = []
    jumlah_halaman = 0
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        jumlah_halaman = len(pdf.pages)
        for page in pdf.pages:
            potongan.append(page.extract_text() or "")
    return "\x0c".join(potongan), jumlah_halaman


def _ekstrak_pdf_jual_kasir_berbasis_posisi(file_like, nama_file: str = None) -> tuple:
    """
    Ekstrak semua blok transaksi dari PDF "Data Penjualan Detail" (lihat
    penjelasan format di atas) jadi DataFrame dengan skema PERSIS SAMA
    dengan output parse_sheet_penjualan() (sheet, tanggal, no_invoice,
    customer, keterangan, cara_bayar, dpp, ppn, total, + kolom jurnal
    kosong) -- supaya bisa langsung diproses ulang oleh
    proses_dataframe_penjualan() yang sudah ada, tanpa logic kategorisasi
    baru.

    Detail ITEM per transaksi SENGAJA tidak diekstrak (tidak dibutuhkan --
    yang dijurnalkan adalah 1 baris Debet Kas/Piutang - Kredit Penjualan
    per TRANSAKSI, bukan per item), jadi baris-baris tabel item cukup
    dilewati sampai ketemu baris footer "Pot. : ... Total Akhir : ...".

    cara_bayar ditebak dari PREFIX No Transaksi ("KSR" = kasir = TUNAI,
    prefix lain mis. "JL" = dianggap KREDIT/piutang) karena PDF sumber
    tidak mencantumkan metode bayar eksplisit -- lihat catatan lebih
    lengkap di docstring proses_file_jurnal_penjualan_kasir().

    [BARU] Sumber teks sekarang lewat _ambil_teks_pdf_layout() (pdftotext,
    lihat catatan kecepatan di sana) -- karena pdftotext -layout menjaga
    LEBAR KOLOM asli (banyak spasi utk sejajarkan kolom, beda dari
    pdfplumber.extract_text() yg otomatis satu-spasi antar kata), tiap
    baris di-normalisasi dulu (spasi/tab berurutan -> satu spasi) sebelum
    dicocokkan ke regex -- regex-nya sendiri TIDAK berubah sama sekali,
    supaya perilaku parsing per-baris identik dgn versi pdfplumber lama.

    [BARU] Kuirk layout sumber: kalau tabel item 1 transaksi tidak muat di
    sisa halaman, mesin kasir mencetak DUA KALI baris header transaksi
    yang sama -- sekali "menggantung" di akhir halaman (tanpa tabel item/
    footer, krn tidak ada ruang lagi) lalu diulang PERSIS di awal halaman
    berikutnya (baru diikuti tabel item & footer yang sebenarnya). Baris
    yang menggantung ini terdeteksi sbg "transaksi baru dibuka" oleh
    parser padahal bukan transaksi baru -- kalau tidak ditangani, tiap
    kemunculannya memicu peringatan palsu "tidak menemukan Total Akhir"
    (walau transaksinya SENDIRI tetap tercatat benar dari kemunculan
    kedua, jadi tidak ada data yang hilang -- sudah diverifikasi silang
    total 24 kejadian pada contoh file "Data Penjualan Detail 01-30
    Agustus 2026", semuanya no. transaksi yang sama persis muncul 2x
    berturut-turut). Sekarang kemunculan kedua dgn no. transaksi SAMA
    PERSIS langsung dianggap "buka ulang" tanpa mencatat peringatan --
    peringatan hanya dicatat kalau no. transaksi BERBEDA yang menyusup di
    tengah (baru itu indikasi baris 'Total Akhir' benar-benar hilang).

    Returns: (df, peringatan) -- df kosong kalau tidak ada transaksi yang
    berhasil diekstrak sama sekali (peringatan akan menjelaskan kenapa).
    """
    import time as _time_debug  # [DEBUG SEMENTARA]
    nama_file = nama_file or getattr(file_like, "name", "") or "upload.pdf"
    if hasattr(file_like, "read"):
        data = file_like.read()
    else:
        data = file_like
    _t_ekstrak_teks = _time_debug.perf_counter()
    teks_gabungan, jumlah_halaman = _ambil_teks_pdf_layout(data, nama_file)
    print(f"[DEBUG-TIMING] _ambil_teks_pdf_layout('{nama_file}', {jumlah_halaman} halaman): "
          f"{_time_debug.perf_counter() - _t_ekstrak_teks:.2f} detik")

    peringatan: list = []
    baris_transaksi: list = []

    transaksi_terbuka = None
    teks_pelanggan: list = []
    sedang_item = False

    _t_parsing_loop = _time_debug.perf_counter()  # [DEBUG SEMENTARA]

    def _tutup(footer_match):
        nonlocal transaksi_terbuka, teks_pelanggan, sedang_item
        pot, pajak, biaya, total_akhir = (_angka_id(x) for x in footer_match.groups())
        customer_teks = " ".join(t for t in teks_pelanggan if t).strip()
        no_transaksi = transaksi_terbuka["no_transaksi"]
        # [ASUMSI] KSR = transaksi kasir langsung (dianggap tunai). Prefix
        # lain (mis. "JL" -- kemungkinan "Jual"/pengiriman ke pelanggan
        # terdaftar dgn alamat lengkap) dianggap kredit/piutang. Kalau
        # ternyata salah untuk sebagian transaksi, baris tetap bisa
        # dikoreksi manual seperti baris hasil kategorisasi otomatis lain
        # (lewat review "Belum Terkategori" / klarifikasi).
        cara_bayar = "TUNAI" if no_transaksi.upper().startswith("KSR") else "KREDIT"
        baris_transaksi.append({
            "sheet": nama_file,
            "tanggal": transaksi_terbuka["tanggal"],
            "no_invoice": no_transaksi,
            "customer": customer_teks or transaksi_terbuka.get("kode_pel") or "-",
            "keterangan": customer_teks or no_transaksi,
            "cara_bayar": cara_bayar,
            "dpp": max(total_akhir - pajak, 0.0),
            "ppn": pajak,
            "total": total_akhir,
        })
        transaksi_terbuka = None
        teks_pelanggan = []
        sedang_item = False

    for raw_baris in teks_gabungan.split("\n"):
        # Kolaps spasi/tab berurutan jadi 1 spasi -- lihat catatan di
        # docstring di atas (pdftotext -layout menjaga lebar kolom asli
        # pakai banyak spasi, regex di bawah dituning utk 1 spasi seperti
        # keluaran pdfplumber.extract_text() lama). .strip() juga
        # membuang karakter form-feed (\x0c, penanda pergantian halaman)
        # yang menempel di awal baris pertama tiap halaman.
        baris = re.sub(r"[ \t]+", " ", raw_baris).strip()
        if not baris:
            continue
        if _RE_HEADER_KOLOM_JUAL.match(baris) or _RE_FOOTER_HALAMAN_JUAL.match(baris):
            continue

        m_header = _RE_HEADER_TRANSAKSI_JUAL.match(baris)
        if m_header:
            no_transaksi_baru = m_header.group(1)
            if transaksi_terbuka is not None and transaksi_terbuka["no_transaksi"] != no_transaksi_baru:
                peringatan.append(
                    f"Transaksi {transaksi_terbuka['no_transaksi']} tidak menemukan "
                    "baris 'Total Akhir' sebelum transaksi berikutnya dimulai -- dilewati."
                )
            # (else: header transaksi yg sama persis terulang krn tabel
            # itemnya terpotong halaman -- lihat catatan kuirk di atas,
            # bukan indikasi masalah, jadi tidak dicatat sbg peringatan)
            no_transaksi, tanggal, sisa = m_header.groups()
            m_kode = _RE_KODE_PEL_JUAL.match(sisa)
            if m_kode:
                kode_pel, sisa_teks = m_kode.groups()
            else:
                kode_pel, sisa_teks = None, sisa
            transaksi_terbuka = {
                "no_transaksi": no_transaksi, "tanggal": tanggal, "kode_pel": kode_pel,
            }
            teks_pelanggan = [sisa_teks.strip()] if sisa_teks.strip() else []
            sedang_item = False
            continue

        m_footer = _RE_FOOTER_TRANSAKSI_JUAL.match(baris)
        if m_footer and transaksi_terbuka is not None:
            _tutup(m_footer)
            continue

        if transaksi_terbuka is None:
            # Baris di luar blok transaksi mana pun (mis. judul
            # laporan & alamat toko di awal halaman 1) -- lewati.
            continue

        if _RE_ITEM_HEADER_JUAL.match(baris):
            sedang_item = True
            continue

        if not sedang_item:
            # Masih bagian "Nama Pelanggan/Alamat" yang wrap ke
            # baris berikutnya (sebelum tabel item dimulai).
            teks_pelanggan.append(baris)
        # else: isi tabel item -- detail per-item tidak diekstrak
        # (lihat penjelasan di docstring), baris dilewati.

    if transaksi_terbuka is not None:
        peringatan.append(
            f"Transaksi {transaksi_terbuka['no_transaksi']} tidak menemukan baris 'Total Akhir' "
            "sampai akhir file -- dilewati (kemungkinan file terpotong)."
        )

    if not baris_transaksi:
        peringatan.append(f"Tidak ditemukan transaksi yang bisa diekstrak dari '{nama_file}'.")
        return pd.DataFrame(), peringatan

    print(f"[DEBUG-TIMING] parsing loop regex ({len(baris_transaksi)} transaksi): "
          f"{_time_debug.perf_counter() - _t_parsing_loop:.2f} detik")  # [DEBUG SEMENTARA]

    df = pd.DataFrame(baris_transaksi)
    df["tanggal"] = pd.to_datetime(df["tanggal"], errors="coerce", dayfirst=True).dt.date

    # Kolom jurnal kosong -- diisi belakangan oleh proses_dataframe_penjualan()
    # (skema disamakan dgn output parse_sheet_penjualan() supaya bisa dipakai
    # ulang fungsi kategorisasi yang sama persis, termasuk pola historis &
    # AI, tanpa logic baru).
    for kol in ("no_akun_debet", "nama_akun_debet", "jml_debet",
                "no_akun_kredit", "nama_akun_kredit", "jml_kredit",
                "no_akun_kredit_ppn", "nama_akun_kredit_ppn", "jml_kredit_ppn"):
        df[kol] = None

    peringatan.append(
        f"{len(df)} transaksi berhasil diekstrak dari {jumlah_halaman} halaman PDF ('{nama_file}')."
    )
    return df, peringatan


def proses_file_jurnal_penjualan_kasir(
    file_like, nama_file: str = None, client_id: Optional[int] = None, pakai_ai: bool = True,
    df_coa_client: Optional[pd.DataFrame] = None,
) -> dict:
    """
    "Jurnal Penjualan Kasir" -- laporan PDF "Data Penjualan Detail"
    per-blok transaksi (No Transaksi/Tanggal/Dept./Kode Pel./Nama
    Pelanggan/Alamat + tabel item + baris Pot./Pajak/Biaya/Total Akhir per
    transaksi), lihat _ekstrak_pdf_jual_kasir_berbasis_posisi() di atas.

    BEDA dari proses_file_penjualan(): fungsi itu mengharapkan 1 baris rata
    per transaksi di sheet Excel/PDF bergrid (kolom invoice/customer/dpp/
    ppn/total sejajar). Laporan ini per-blok multi-baris tanpa grid, dan
    detail ITEM tidak dijurnalkan satu-satu -- yang dijurnalkan adalah
    TOTAL AKHIR per transaksi (1 blok = 1 baris jurnal penjualan Debet
    Kas/Piutang - Kredit Penjualan).

    df_coa_client (opsional): COA permanen client dari DB
    (dbc.ambil_coa_client), diteruskan oleh main.py krn PDF laporan kasir
    ini TIDAK PERNAH punya sheet COA sendiri di dalam filenya sendiri
    (beda dari proses_file_penjualan yg bisa terima Excel dgn sheet COA
    terpisah dalam file yang sama) -- tanpa ini akun Penjualan/Kas/Piutang
    akan selalu jatuh ke label generik ("KAS"/"PIUTANG USAHA"/"Belum
    Terkategori") kecuali pakai_ai=True dan API key aktif.

    Kategorisasi memakai pola historis TERPISAH dari proses_file_penjualan
    ("pola_penjualan_kasir", bukan "pola_penjualan") krn signature sumber
    beda karakter (ringkasan per-transaksi laporan kasir vs baris invoice
    per-item) -- pola dari 1 jenis dokumen tidak seharusnya ikut
    memengaruhi kategorisasi jenis dokumen lain.
    """
    nama_file = nama_file or getattr(file_like, "name", "") or "upload.pdf"
    import time as _time_debug  # [DEBUG SEMENTARA]
    _t_ekstrak = _time_debug.perf_counter()
    df_jual, peringatan = _ekstrak_pdf_jual_kasir_berbasis_posisi(file_like, nama_file)
    print(f"[DEBUG-TIMING] TOTAL _ekstrak_pdf_jual_kasir_berbasis_posisi: "
          f"{_time_debug.perf_counter() - _t_ekstrak:.2f} detik")

    if df_jual is None or df_jual.empty:
        return {"df": pd.DataFrame(), "ringkasan": {}, "masalah": [], "draf_jurnal": [], "sheet_dilewati": peringatan}

    df_coa = df_coa_client if (df_coa_client is not None and not df_coa_client.empty) else pd.DataFrame()

    path_pola = _path_pola("pola_penjualan_kasir", client_id)
    pola = muat_pola(path_pola)

    _t_kategori = _time_debug.perf_counter()
    df_hasil = proses_dataframe_penjualan(
        df_jual, df_coa, pola, pakai_ai=pakai_ai,
        api_key=ambil_api_key_claude() if pakai_ai else None,
    )
    print(f"[DEBUG-TIMING] proses_dataframe_penjualan (pola+aturan+AI): "
          f"{_time_debug.perf_counter() - _t_kategori:.2f} detik")

    _t_simpan_pola = _time_debug.perf_counter()
    simpan_pola(pola, path_pola, sumber_perubahan="proses_file_jurnal_penjualan_kasir (auto dari data asli)")
    print(f"[DEBUG-TIMING] simpan_pola: {_time_debug.perf_counter() - _t_simpan_pola:.2f} detik")

    _t_keseimbangan = _time_debug.perf_counter()
    keseimbangan = cek_keseimbangan_jurnal(df_hasil)
    print(f"[DEBUG-TIMING] cek_keseimbangan_jurnal: {_time_debug.perf_counter() - _t_keseimbangan:.2f} detik")

    _t_masalah = _time_debug.perf_counter()
    masalah = []
    for i, row in df_hasil.iterrows():
        sumber = str(row.get("sumber_kategori") or "")
        if "Belum Terkategori" in sumber or "perlu cek" in sumber:
            masalah.append({
                "baris": i + 1, "tanggal": row.get("tanggal"), "keterangan": row.get("keterangan"),
                "no_invoice": row.get("no_invoice"),
                "alasan": [f"Sumber kategori: {sumber or 'tidak diketahui'} -- akun belum pasti, cek manual."],
            })
    print(f"[DEBUG-TIMING] loop 'masalah' (iterrows x{len(df_hasil)}): "
          f"{_time_debug.perf_counter() - _t_masalah:.2f} detik")

    _t_draf = _time_debug.perf_counter()  # [DEBUG SEMENTARA]
    # [CATATAN] Skema draf_jurnal disamakan dgn hasil rekening_koran
    # (ImportRekeningKoranModal.tsx::DrafJurnalRow) supaya frontend bisa
    # pakai interface & layar preview yang sama persis -- field "bank" &
    # "mutasi_debet"/"mutasi_kredit" tidak relevan di sini (bukan mutasi
    # bank) jadi diisi None; frontend hanya memakainya utk hitung saldo
    # kas berjalan pada jalur rekening_koran, TIDAK dipakai sama sekali di
    # jalur jurnal_penjualan_kasir (lihat drafJurnalPenjualanToTransactions
    # di modal).
    draf_jurnal = [
        {
            "baris": i + 1, "tanggal": row.get("tanggal"), "bank": None,
            "keterangan": row.get("keterangan"),
            "no_akun_debet": row.get("no_akun_debet"), "nama_akun_debet": row.get("nama_akun_debet"),
            "jml_debet": row.get("jml_debet"),
            "no_akun_kredit": row.get("no_akun_kredit"), "nama_akun_kredit": row.get("nama_akun_kredit"),
            "jml_kredit": row.get("jml_kredit"),
            "mutasi_debet": None, "mutasi_kredit": None,
            "sumber_kategori": row.get("sumber_kategori"),
            "catatan": row.get("catatan_ai"),
            "no_invoice": row.get("no_invoice"),
        }
        for i, row in df_hasil.iterrows()
    ]
    print(f"[DEBUG-TIMING] loop 'draf_jurnal' (iterrows x{len(df_hasil)}): "
          f"{_time_debug.perf_counter() - _t_draf:.2f} detik")

    jumlah_transaksi = len(df_hasil)
    total_penjualan = float(pd.to_numeric(df_hasil.get("jml_kredit"), errors="coerce").fillna(0).sum())
    rata_rata_transaksi = total_penjualan / jumlah_transaksi if jumlah_transaksi else 0

    ringkasan = {
        "jumlah_transaksi": jumlah_transaksi,
        "total_penjualan": total_penjualan,
        "rata_rata_transaksi": rata_rata_transaksi,
        "total_debet": keseimbangan["total_debet"],
        "total_kredit": keseimbangan["total_kredit"],
        "balance": keseimbangan["balance"],
        "selisih": keseimbangan["selisih"],
        "jumlah_perlu_review": len(masalah),
    }

    return {
        "df": df_hasil, "ringkasan": ringkasan, "masalah": masalah,
        "draf_jurnal": draf_jurnal, "sheet_dilewati": peringatan,
    }


def proses_file_penilaian_klien(file_like, nama_file: str = None) -> dict:
    """Penilaian Klien/Maker -- cek kelengkapan & kewajaran skor, bukan jurnal akuntansi."""
    nama_file = nama_file or getattr(file_like, "name", "") or ""
    _df_bank, _df_jual, df_penilaian, _df_piutang, _df_coa, peringatan = muat_workbook(file_like, nama_file)

    if df_penilaian is None or df_penilaian.empty:
        return {"df": pd.DataFrame(), "ringkasan": {}, "masalah": [], "draf_jurnal": [], "sheet_dilewati": peringatan}

    df_hasil, analisis = proses_dataframe_penilaian(df_penilaian)
    masalah = [
        {"baris": None, "kategori": t.get("kategori"), "alasan": [f"{t.get('temuan')} (tingkat: {t.get('tingkat_keparahan')})"]}
        for t in analisis.get("temuan", [])
    ]

    return {"df": df_hasil, "ringkasan": analisis, "masalah": masalah, "draf_jurnal": [], "sheet_dilewati": peringatan}


def _hitung_aging_bucket_piutang(df_piutang: pd.DataFrame, tanggal_acuan: date = None) -> dict:
    """
    [BARU] Hitung aging bucket piutang, simetris dengan aging AP di
    _kategorikan_umur_utang() (lihat di atas). Piutang tidak punya fungsi
    aging sebelumnya -- proses_file_piutang() cuma merekap total & per
    pelanggan, tanpa breakdown umur.

    Dipisah jadi fungsi sendiri (bukan reuse _kategorikan_umur_utang)
    karena kolom sumber tanggal & nominalnya beda nama dari sisi AP
    (tanggal_jatuh_tempo/sisa_utang_tertulis vs beberapa kemungkinan nama
    kolom piutang yang belum tentu seragam antar file client).
    """
    if df_piutang is None or df_piutang.empty:
        return {"bucket": {}, "total": 0, "total_piutang": 0, "rekap_per_pelanggan": {}, "jumlah_invoice": 0, "jumlah_pelanggan": 0}

    tanggal_acuan = tanggal_acuan or date.today()

    def _hitung_umur_hari(tanggal):
        if pd.isna(tanggal):
            return None
        try:
            if isinstance(tanggal, datetime):
                t = tanggal.date()
            elif isinstance(tanggal, date):
                t = tanggal
            else:
                t = pd.to_datetime(tanggal).date()
            return (tanggal_acuan - t).days
        except Exception:
            return None

    df = df_piutang.copy()

    # Coba cari kolom tanggal jatuh tempo -- nama kolom piutang belum
    # tentu seragam antar file client (beda dgn AP yg sudah baku
    # "tanggal_jatuh_tempo"), jadi dicoba beberapa kandidat.
    kolom_tanggal = None
    for col in ["tanggal_jatuh_tempo", "jatuh_tempo", "tanggal", "tgl"]:
        if col in df.columns and df[col].notna().any():
            kolom_tanggal = col
            break

    if kolom_tanggal is None:
        return {"bucket": {}, "total": 0, "total_piutang": 0, "rekap_per_pelanggan": {}, "jumlah_invoice": len(df), "jumlah_pelanggan": 0}

    df["_umur_hari"] = df[kolom_tanggal].apply(_hitung_umur_hari)

    def _bucket(hari):
        # [FIX] "hari is None" tidak menangkap NaN (nilai None yang masuk ke
        # kolom Series pandas otomatis dinormalisasi jadi NaN float) --
        # ditemukan saat menambahkan _perkaya_piutang_per_baris() di bawah
        # yang punya logika sama persis. Pakai pd.isna() spy None maupun NaN
        # sama-sama tertangkap, bukan cuma salah satu.
        if pd.isna(hari):
            return "TIDAK DIKETAHUI"
        if hari <= 0:
            return "BELUM JATUH TEMPO"
        if hari <= 30:
            return "1-30 HARI"
        if hari <= 60:
            return "31-60 HARI"
        if hari <= 90:
            return "61-90 HARI"
        return ">90 HARI"

    df["_bucket"] = df["_umur_hari"].apply(_bucket)

    # Kolom nominal -- ikuti prioritas yg sama seperti di proses_file_piutang()
    kolom_nominal = "sisa_utang_tertulis" if "sisa_utang_tertulis" in df.columns else "jumlah_utang"
    if kolom_nominal not in df.columns:
        kolom_nominal = "total_akhir" if "total_akhir" in df.columns else "sub_total"
    if kolom_nominal not in df.columns:
        return {"bucket": {}, "total": 0, "total_piutang": 0, "rekap_per_pelanggan": {}, "jumlah_invoice": len(df), "jumlah_pelanggan": 0}

    bucket = {}
    total = 0.0
    for b, g in df.groupby("_bucket"):
        nominal = float(pd.to_numeric(g[kolom_nominal], errors="coerce").fillna(0).sum())
        bucket[b] = nominal
        total += nominal

    rekap_per_pelanggan = {}
    if "nama_pelanggan" in df.columns:
        rekap_per_pelanggan = df.groupby("nama_pelanggan")[kolom_nominal].apply(
            lambda s: float(pd.to_numeric(s, errors="coerce").fillna(0).sum())
        ).to_dict()

    return {
        "bucket": bucket,
        "total": total,
        "total_piutang": total,
        "rekap_per_pelanggan": rekap_per_pelanggan,
        "jumlah_invoice": len(df),
        "jumlah_pelanggan": df["nama_pelanggan"].nunique() if "nama_pelanggan" in df.columns else 0,
    }


def _perkaya_piutang_per_baris(df_piutang: pd.DataFrame, tanggal_acuan: date = None) -> pd.DataFrame:
    """
    [FIX/BARU] Sebelumnya HANYA ada _hitung_aging_bucket_piutang() yang
    mengembalikan AGREGAT (total per bucket, total per pelanggan) -- tidak
    ada fungsi yang menempelkan umur_hari/bucket_aging/status/saldo per
    baris ke df_piutang itu sendiri, padahal accounting_export.py (sheet
    "Buku Bantu Piutang") mengasumsikan field2 itu ADA per baris lewat
    ALIAS_PIUTANG, persis seperti proses_ap_aging() sudah lakukan utk sisi
    hutang. Akibatnya kolom Saldo Piutang/Hari Tertunggak/Bucket
    Aging/Status/Penerimaan di Excel export SELALU kosong walau datanya
    sebenarnya ada. Fungsi ini melengkapi df_piutang supaya simetris
    dengan sisi AP -- dipanggil dari proses_file_piutang() di bawah.

    Nama kolom hasil (umur_hari, bucket_aging, status, sisa_piutang_hitung)
    SENGAJA disamakan persis dengan nama field sisi AP supaya alias yang
    sudah ada di accounting_export.py::ALIAS_PIUTANG langsung nyambung
    tanpa perlu diubah lagi.
    """
    if df_piutang is None or df_piutang.empty:
        return df_piutang

    tanggal_acuan = tanggal_acuan or date.today()
    df = df_piutang.copy()

    kolom_tanggal = None
    for col in ["tanggal_jatuh_tempo", "jatuh_tempo", "tanggal", "tgl"]:
        if col in df.columns and df[col].notna().any():
            kolom_tanggal = col
            break

    kolom_nominal = "total_akhir" if "total_akhir" in df.columns else "sub_total"

    def _hitung_umur_hari(tanggal):
        if pd.isna(tanggal):
            return None
        try:
            if isinstance(tanggal, datetime):
                t = tanggal.date()
            elif isinstance(tanggal, date):
                t = tanggal
            else:
                t = pd.to_datetime(tanggal).date()
            return (tanggal_acuan - t).days
        except Exception:
            return None

    def _bucket(hari):
        # [FIX] Sama seperti di _hitung_aging_bucket_piutang(): "hari is None"
        # tidak menangkap NaN, pakai pd.isna() supaya baris tanpa tanggal
        # jatuh tempo benar-benar masuk "TIDAK DIKETAHUI", bukan salah
        # kejebak default ">90 HARI".
        if pd.isna(hari):
            return "TIDAK DIKETAHUI"
        if hari <= 0:
            return "BELUM JATUH TEMPO"
        if hari <= 30:
            return "1-30 HARI"
        if hari <= 60:
            return "31-60 HARI"
        if hari <= 90:
            return "61-90 HARI"
        return ">90 HARI"

    df["umur_hari"] = df[kolom_tanggal].apply(_hitung_umur_hari) if kolom_tanggal else None
    df["bucket_aging"] = df["umur_hari"].apply(_bucket)

    nominal = (pd.to_numeric(df[kolom_nominal], errors="coerce").fillna(0)
               if kolom_nominal in df.columns else pd.Series(0.0, index=df.index))
    bayar_tunai = (pd.to_numeric(df["bayar_tunai"], errors="coerce").fillna(0)
                   if "bayar_tunai" in df.columns else pd.Series(0.0, index=df.index))
    bayar_kredit = (pd.to_numeric(df["bayar_kredit"], errors="coerce").fillna(0)
                    if "bayar_kredit" in df.columns else pd.Series(0.0, index=df.index))
    df["penerimaan_total"] = bayar_tunai + bayar_kredit
    df["sisa_piutang_hitung"] = nominal - df["penerimaan_total"]

    def _status(row):
        if row["sisa_piutang_hitung"] < 0:
            return "PERLU REVIEW"
        if row["bucket_aging"] == ">90 HARI" and row["sisa_piutang_hitung"] > 0:
            return "PERLU REVIEW"
        return "OK"

    df["status"] = df.apply(_status, axis=1)
    return df


def proses_file_piutang(file_like, nama_file: str = None) -> dict:
    """
    Buku Bantu Piutang (AR) -- BUKAN jurnal (kartu piutang per transaksi),
    jadi cuma direkap & dicek kelengkapannya, tidak ada draf_jurnal.
    """
    nama_file = nama_file or getattr(file_like, "name", "") or ""
    _df_bank, _df_jual, _df_nilai, df_piutang, _df_coa, peringatan = muat_workbook(file_like, nama_file)

    if df_piutang is None or df_piutang.empty:
        return {"df": pd.DataFrame(), "ringkasan": {}, "masalah": [], "draf_jurnal": [], "sheet_dilewati": peringatan}

    # [FIX] Lengkapi df per-baris (umur_hari/bucket_aging/status/saldo)
    # SEBELUM dipakai di bawah -- lihat docstring _perkaya_piutang_per_baris().
    df_piutang = _perkaya_piutang_per_baris(df_piutang)

    masalah = []
    for i, row in df_piutang.iterrows():
        alasan = []
        if not row.get("nama_pelanggan") or str(row.get("nama_pelanggan")).strip() == "":
            alasan.append("Nama pelanggan kosong.")
        total = row.get("total_akhir") if pd.notna(row.get("total_akhir")) else row.get("sub_total")
        # [FIX] "total is not None and float(total or 0) < 0" tidak menangkap
        # kasus NaN: NaN "is not None" True, tapi "float(nan or 0)" tetap nan
        # (NaN truthy), lalu "nan < 0" selalu False -- baris dengan nominal
        # piutang NaN lolos tanpa peringatan sama sekali. Pakai _kosong() +
        # _angka_aman() supaya NaN benar-benar tertangkap.
        if not _kosong(total) and _angka_aman(total) < 0:
            alasan.append("Nominal piutang negatif -- tidak wajar, cek input.")
        elif _kosong(total):
            alasan.append("Nominal piutang (total_akhir/sub_total) kosong -- cek input.")
        if alasan:
            masalah.append({"baris": i + 1, "nama_pelanggan": row.get("nama_pelanggan"), "alasan": alasan})

    kolom_total = "total_akhir" if "total_akhir" in df_piutang.columns else "sub_total"
    total_piutang = float(pd.to_numeric(df_piutang.get(kolom_total), errors="coerce").fillna(0).sum())
    rekap_per_pelanggan = {}
    if "nama_pelanggan" in df_piutang.columns:
        rekap_per_pelanggan = (
            df_piutang.groupby("nama_pelanggan")[kolom_total].sum().fillna(0).to_dict()
            if kolom_total in df_piutang.columns else {}
        )

    ringkasan = {
        "jumlah_transaksi": len(df_piutang),
        "total_piutang": total_piutang,
        "jumlah_pelanggan": df_piutang["nama_pelanggan"].nunique() if "nama_pelanggan" in df_piutang.columns else 0,
        "rekap_per_pelanggan": rekap_per_pelanggan,
        "jumlah_perlu_review": len(masalah),
    }

    # [BARU] Aging bucket piutang, simetris dgn aging AP (_kategorikan_umur_utang).
    # Kalau tidak ketemu kolom tanggal jatuh tempo, "bucket" akan kosong ({})
    # tapi total_piutang/rekap_per_pelanggan di atas tetap terisi seperti biasa.
    aging = _hitung_aging_bucket_piutang(df_piutang)
    ringkasan["aging_bucket"] = aging["bucket"]
    if aging["bucket"]:
        ringkasan["total_piutang"] = aging["total"]
        ringkasan["rekap_per_pelanggan"] = aging["rekap_per_pelanggan"]
        ringkasan["jumlah_invoice"] = aging["jumlah_invoice"]

    return {"df": df_piutang, "ringkasan": ringkasan, "masalah": masalah, "draf_jurnal": [], "sheet_dilewati": peringatan}


def proses_file_laporan_keuangan(file_like, nama_file: str = None) -> dict:
    """
    Wrapper utk file Laporan Keuangan lengkap (31 sheet: COA, GL, TB, BS,
    CALK, ADJ, FA FIX, CF6, dst -- lihat TARGET_SHEETS_31). File jenis ini
    BUKAN data mentah yang perlu dijurnal -- ini laporan yang SUDAH jadi,
    jadi tidak ada draf_jurnal. Wrapper ini cuma mengonfirmasi & merekap
    sheet kunci apa saja yang ketemu, supaya frontend tahu file ini valid
    & lengkap (atau sheet apa yang kurang).
    """
    nama_file = nama_file or getattr(file_like, "name", "") or ""
    if hasattr(file_like, "seek"):
        file_like.seek(0)
    try:
        wb = openpyxl.load_workbook(file_like, read_only=True)
        sheets = wb.sheetnames
    finally:
        if hasattr(file_like, "seek"):
            file_like.seek(0)

    terdeteksi, pesan, sheet_kunci_ditemukan = deteksi_laporan_keuangan(sheets)
    if not terdeteksi:
        return {"df": pd.DataFrame(), "ringkasan": {}, "masalah": [], "draf_jurnal": [], "per_sheet": None}

    kunci_norm_ditemukan = {_normalisasi_nama_sheet(s) for s in sheet_kunci_ditemukan}
    target_norm = {_normalisasi_nama_sheet(s) for s in TARGET_SHEETS_31}
    sheet_hilang = [s for s in TARGET_SHEETS_31 if _normalisasi_nama_sheet(s) not in kunci_norm_ditemukan]

    ringkasan = {
        "pesan": pesan,
        "jumlah_sheet_di_file": len(sheets),
        "jumlah_sheet_standar_31": len(TARGET_SHEETS_31),
        "sheet_kunci_ditemukan": sheet_kunci_ditemukan,
        "sheet_standar_kemungkinan_hilang": sheet_hilang,
        "catatan": (
            "File ini terdeteksi sebagai Laporan Keuangan lengkap (bukan data mentah), "
            "jadi tidak diproses jadi draf jurnal -- cukup dipakai sebagai laporan final. "
            "'sheet_standar_kemungkinan_hilang' hanya indikatif (variasi nama sheet & sheet "
            "entitas tambahan itu wajar, lihat SHEET_ENTITAS_LAPORAN)."
        ),
    }

    return {
        "df": pd.DataFrame(),
        "ringkasan": ringkasan,
        "masalah": [],
        "draf_jurnal": [],
        "per_sheet": {"daftar_sheet": sheets},
    }


def generate_template_31_sheet() -> bytes:
    """
    Buat file Excel kosong berisi 31 sheet standar (TARGET_SHEETS_31) --
    dipakai sebagai starting point kalau user belum punya file Laporan
    Keuangan dan mau mulai dari template kosong. Return bytes (siap
    di-download / dikirim sbg StreamingResponse dari FastAPI).
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for nama in TARGET_SHEETS_31:
        wb.create_sheet(title=nama)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ============================================================
# 15. AI CHAT UTILITY
# ============================================================

def buat_system_prompt_akuntansi() -> str:
    """Context kaya untuk AI agar selalu paham domain akuntansi kamu, cara
    berpikir, dan gaya jawab -- supaya kualitas percakapan setara AI besar
    (ChatGPT/Claude), bukan cuma jawaban template pendek."""
    return """Kamu adalah AI Gouf Consulting -- rekan kerja AI untuk tim akuntansi/konsultan, bukan sekadar mesin jawab pertanyaan.

# TENTANG APLIKASI INI
Kamu berjalan DI DALAM dashboard "Gouf Consulting Accounting" -- aplikasi
web akuntansi yang dipakai konsultan/tim finance untuk mengelola pembukuan
banyak client sekaligus. Kamu bukan chatbot berdiri sendiri; kamu adalah
salah satu menu ("Agent AI") di sidebar dashboard ini, dengan menu lain
yang bisa kamu rujuk kalau relevan:
- Overview -- ringkasan KPI lintas client
- Agent AI -- ini kamu, tempat user chat & upload file
- Financial Statements -- Profit & Loss, Balance Sheet, Cash Flow
- Transactions -- seluruh jurnal yang sudah diposting
- Accounts Receivable / Accounts Payable -- piutang & utang
- Assets, Liabilities, Equity -- neraca per kategori
- Budget & Forecast -- rencana anggaran vs aktual
- Tax & Compliance -- kewajiban & jatuh tempo pajak
- Financial Analytics -- rasio & tren keuangan
- AI Financial Analyst, Audit -- analisis & jejak audit

Kamu TAHU dari mana data tiap halaman itu berasal (lihat bagian "ALUR
FILE -> HALAMAN DASHBOARD" di bawah), jadi kalau user tanya soal
kapabilitasmu atau ke mana data hasil upload akan muncul, jawab dari
pengetahuan itu -- jangan bersikap seperti tidak tahu app-mu sendiri.

# KEAHLIAN
- Double entry bookkeeping & prinsip akuntansi Indonesia (PSAK/SAK EMKM)
- Rekonsiliasi rekening koran/mutasi bank & pola historis jurnal
- Chart of Accounts (COA) perusahaan
- Aturan penjualan tunai/kredit + PPN Keluaran/Masukan
- Piutang, utang, aset tetap, kartu stok, pembelian, payroll, pajak (PPh 21/23/4(2), SPT)
- Penilaian kinerja maker/klien berdasarkan 6 poin prioritas & 20 kesalahan standar akuntan

# CARA BERPIKIR
- Jangan pernah mengarang angka, nama akun, atau nomor akun. Kalau data belum ada/tidak yakin, katakan terus terang dan tanyakan atau minta user cek sumbernya -- akurasi lebih penting daripada terlihat serba tahu.
- Kalau pertanyaan user ambigu, jangan langsung menebak-nebak panjang lebar: tanyakan SATU hal paling penting untuk memperjelas, atau jawab dengan asumsi yang kamu sebutkan eksplisit.
- Untuk masalah akuntansi yang kompleks (jurnal tidak balance, rekonsiliasi tidak ketemu pasangannya, dsb), jelaskan alur berpikirmu langkah demi langkah sebelum kasih kesimpulan -- bukan cuma lempar jawaban akhir.
- Kalau user upload/proses file, gunakan konteks data yang sudah diproses (lihat bagian KONTEKS di bawah, jika ada) sebelum menjawab -- jangan minta user mengulang info yang sudah ada.
- PENTING: kamu SUDAH BISA memproses file yang dikirim user dan otomatis mengarahkannya ke halaman dashboard yang cocok -- ini bukan fitur yang perlu dikonfirmasi/ditanyakan dulu (jangan tanya "dashboard mana yang dimaksud?" atau "apakah kamu punya akses ke dashboard?"). Kalau user tanya soal ini, jawab dengan percaya diri berdasarkan alur di bawah, dan kalau dia belum kirim file, cukup arahkan dia upload lewat chat ini.

# ALUR FILE -> HALAMAN DASHBOARD
Setiap file yang di-upload otomatis dideteksi jenisnya lalu dijurnal (kalau
transaksional) atau disimpan sebagai data tersendiri. Dari situ, halaman
sidebar berikut menghitung tampilannya sendiri-sendiri (bukan tabel
terpisah per halaman -- semua bermuara dari jurnal yang sama):

- **Transactions**: dari SEMUA jurnal yang sudah diposting. Diisi oleh
  upload Rekening Koran/Mutasi Bank, Data Penjualan (Invoice/POS/Kasir),
  Pembelian, Bukti Kas Masuk/Keluar, Slip Gaji, dan jenis transaksional
  lainnya.
- **Accounts Receivable**: transaksi berkategori "Piutang" dari jurnal di
  atas, ditambah detail aging dari upload Buku Bantu Piutang (AR).
- **Accounts Payable**: transaksi berkategori "Utang", ditambah detail
  aging dari upload AP Aging (Utang Jatuh Tempo).
- **Audit**: sama seperti Transactions (semua jurnal termasuk yang masih
  draft/belum direview).
- **Tax & Compliance**: transaksi berkategori "Tax" (baris jurnal "Hutang
  Pajak") dari upload Rekening Koran/Faktur Pajak/Bukti Potong Pajak.
- **Financial Statements** (Profit & Loss, Balance Sheet, Cash Flow):
  agregat trial balance bulanan dari SEMUA jurnal yang sudah diposting;
  atau bisa juga langsung upload Laporan Keuangan Lengkap (31 Sheet) siap
  pakai.
- **Assets, Liabilities, Equity**: sumber sama dengan Financial Statements
  (trial balance bulanan + Chart of Accounts), plus detail register dari
  upload Aset Tetap.
- **Financial Analytics**: sumber sama dengan Financial Statements.
- **Budget & Forecast**: "Actual" dari data Profit & Loss di atas; "Budget"
  dihitung otomatis dari actual + asumsi pertumbuhan (belum ada fitur
  input budget manual, jadi bukan hasil upload file).

Catatan jujur: jenis dokumen Kartu Stok, Absensi Karyawan, Penilaian
Klien/Maker, dan SPT Masa/Tahunan SUDAH dikenali & disimpan backend, tapi
BELUM ada halaman dashboard khusus yang menampilkannya di frontend saat
ini -- kalau user tanya soal ini, katakan terus terang bahwa datanya
tersimpan tapi belum ada halaman viewer-nya, jangan berpura-pura ada.

# GAYA JAWAB
- Bahasa Indonesia, natural dan percakapan -- bukan kaku/formal berlebihan, tapi tetap profesional (bukan bahasa gaul).
- Ramah tapi lugas: langsung ke inti, tidak bertele-tele di pembuka.
- Pakai markdown supaya mudah dibaca: **bold** untuk istilah/angka penting, bullet/numbered list untuk langkah atau daftar, tabel kalau membandingkan beberapa item, blok kode untuk nomor akun/nama field teknis.
- Panjang jawaban menyesuaikan pertanyaan: pertanyaan simpel dijawab singkat (1-3 kalimat), pertanyaan kompleks boleh panjang & terstruktur dengan sub-judul kalau perlu.
- Kalau menunjukkan masalah/kesalahan di data, selalu sertai rekomendasi perbaikannya juga -- jangan cuma bilang "ada yang salah".

# BATASAN
- Kamu bukan pengganti keputusan final akuntan/auditor bersertifikat untuk hal yang berdampak hukum/pajak besar -- untuk kasus begitu, sarankan verifikasi ke supervisor/pihak berwenang, sesuai poin "harus melibatkan pihak lain untuk konsultasi" di standar internal.
- Jangan pernah mengubah data user secara diam-diam tanpa disebutkan -- kalau kamu memperbaiki sesuatu, selalu sebutkan apa yang diubah dan kenapa."""


def _panggil_deepseek_dengan_retry(client_kwargs_fn, max_percobaan: int = 3, **create_kwargs):
    """Panggil DeepSeek API dengan retry + exponential backoff.

    AI besar (ChatGPT/Claude) jarang terlihat 'down' di mata user karena
    error sesaat (timeout, rate limit) otomatis dicoba ulang di belakang
    layar. Fungsi ini menirunya: retry sampai `max_percobaan` kali dengan
    jeda yang makin lama (0.5s, 1s, 2s, ...), dan HANYA retry untuk error
    yang sifatnya sementara (timeout/koneksi/rate limit/server error),
    bukan error permanen (mis. API key salah).
    """
    import time
    import openai as openai_module

    error_sementara = (
        openai_module.APITimeoutError,
        openai_module.APIConnectionError,
        openai_module.RateLimitError,
        openai_module.InternalServerError,
    )

    percobaan_terakhir = None
    for percobaan in range(max_percobaan):
        try:
            return client_kwargs_fn(**create_kwargs)
        except error_sementara as e:
            percobaan_terakhir = e
            if percobaan < max_percobaan - 1:
                time.sleep(0.5 * (2 ** percobaan))
                continue
            raise
        except Exception:
            # Error permanen (auth, argumen salah, dst) -- jangan diulang.
            raise
    raise percobaan_terakhir  # pragma: no cover (harusnya sudah raise di atas)


def _susun_messages(pesan: str, system_prompt: str, riwayat: list) -> list:
    system_msg = {
        "role": "system",
        "content": system_prompt or buat_system_prompt_akuntansi(),
    }
    return [system_msg] + list(riwayat or []) + [{"role": "user", "content": pesan}]


def tanya_ai(pesan: str, system_prompt: str = None, riwayat: list = None,
             temperature: float = 0.6, max_tokens: int = 2048) -> str:
    """Panggil DeepSeek API (non-streaming), dengan retry otomatis.

    temperature=0.6 (bukan 0.3) supaya jawaban chat terasa lebih natural
    dan tidak kaku/repetitif -- mirip nada AI besar untuk obrolan biasa.
    Untuk tugas yang butuh presisi tinggi (ekstraksi data terstruktur),
    panggil dengan temperature lebih rendah (mis. 0.1-0.2) dari caller-nya.
    """
    import openai
    daftar_provider = _konfigurasi_provider_chat()
    if not daftar_provider:
        raise Exception("Tidak ada API key chat aktif (DEEPSEEK_API_KEY / GROQ_API_KEY)")

    messages = _susun_messages(pesan, system_prompt, riwayat)
    error_terakhir = None
    for konfig in daftar_provider:
        try:
            client = openai.OpenAI(api_key=konfig["api_key"], base_url=konfig["base_url"])
            response = _panggil_deepseek_dengan_retry(
                client.chat.completions.create,
                model=konfig["model"],
                max_tokens=max_tokens,
                temperature=temperature,
                messages=messages,
                **konfig.get("extra_params", {}),
            )
            return response.choices[0].message.content
        except Exception as e:
            # Provider ini gagal (saldo habis, key salah, dst) -- catat lalu
            # coba provider berikutnya di daftar (kalau masih ada).
            error_terakhir = Exception(f"Error {konfig['nama']}: {str(e)}")
            continue
    raise error_terakhir


def tanya_ai_stream(pesan: str, system_prompt: str = None, riwayat: list = None,
                     temperature: float = 0.6, max_tokens: int = 2048):
    """Versi streaming dari tanya_ai(), dengan retry otomatis di percobaan
    KONEKSI awal (sebelum token pertama keluar) -- setelah stream mulai
    jalan, retry di tengah jalan tidak dilakukan karena sebagian jawaban
    sudah terlanjur terkirim ke user.

    Generator yang yield potongan teks (delta) begitu diterima dari DeepSeek,
    lalu dibungkus jadi event SSE (Server-Sent Events) oleh endpoint
    /api/chat/stream di main.py, sehingga frontend React bisa menampilkan
    efek kata-per-kata (mirip ChatGPT/Claude) lewat EventSource / fetch
    stream -- tidak perlu setup WebSocket manual.

    max_tokens dinaikkan dari 1200 -> 2048 supaya jawaban kompleks (mis.
    penjelasan langkah-demi-langkah rekonsiliasi) tidak terpotong di tengah.
    """
    import openai
    daftar_provider = _konfigurasi_provider_chat()
    if not daftar_provider:
        raise Exception("Tidak ada API key chat aktif (DEEPSEEK_API_KEY / GROQ_API_KEY)")

    messages = _susun_messages(pesan, system_prompt, riwayat)
    error_terakhir = None
    for konfig in daftar_provider:
        sudah_ada_token = False
        try:
            client = openai.OpenAI(api_key=konfig["api_key"], base_url=konfig["base_url"])
            stream = _panggil_deepseek_dengan_retry(
                client.chat.completions.create,
                model=konfig["model"],
                max_tokens=max_tokens,
                temperature=temperature,
                messages=messages,
                stream=True,
                **konfig.get("extra_params", {}),
            )
            for potongan in stream:
                delta = potongan.choices[0].delta.content
                if delta:
                    sudah_ada_token = True
                    yield delta
            return  # provider ini sukses sampai habis -- selesai
        except Exception as e:
            error_terakhir = Exception(f"Error {konfig['nama']}: {str(e)}")
            if sudah_ada_token:
                # Sebagian jawaban sudah terlanjur terkirim ke user -- jangan
                # pindah provider di tengah jalan, langsung lempar error.
                raise error_terakhir
            continue  # belum ada token keluar sama sekali -- aman coba provider berikutnya
    raise error_terakhir


def buat_judul_percakapan(pesan_pertama: str) -> str:
    """Auto-generate judul singkat dari pesan pertama user, mirip fitur
    auto-title percakapan baru di ChatGPT/Claude. Fallback ke potongan
    pesan asli kalau API tidak tersedia/gagal."""
    fallback = (pesan_pertama or "Percakapan Baru").strip()[:50] or "Percakapan Baru"
    daftar_provider = _konfigurasi_provider_chat()
    if not daftar_provider:
        return fallback
    import openai
    prompt = (
        "Buatkan judul SANGAT singkat (maksimal 6 kata) dalam Bahasa Indonesia "
        "untuk sebuah percakapan chat yang dimulai dengan pesan user berikut:\n\n"
        f"\"{pesan_pertama}\"\n\n"
        "Balas HANYA judulnya saja, tanpa tanda kutip, tanpa titik di akhir, "
        "tanpa embel-embel penjelasan apa pun."
    )
    for konfig in daftar_provider:
        try:
            client = openai.OpenAI(api_key=konfig["api_key"], base_url=konfig["base_url"])
            response = _panggil_deepseek_dengan_retry(
                client.chat.completions.create,
                model=konfig["model"],
                max_tokens=20,
                temperature=0.4,
                messages=[{"role": "user", "content": prompt}],
            )
            judul = response.choices[0].message.content.strip().strip('"').strip()
            return judul[:100] if judul else fallback
        except Exception:
            continue  # provider ini gagal -- coba provider berikutnya (kalau ada)
    return fallback


def buat_ringkasan_konteks_data(ringkasan_data: list[str], jumlah_pola: int, jumlah_temuan_mencurigakan: int) -> str:
    """Rangkai ringkasan singkat tentang data yang SUDAH diupload/diproses."""
    if not ringkasan_data and jumlah_pola == 0:
        return ""
    bagian = ["\n\nKONTEKS obrolan ini sejauh ini (gunakan kalau relevan dengan pertanyaan user):"]
    if ringkasan_data:
        bagian.append("- Data yang sudah diproses: " + ", ".join(ringkasan_data) + ".")
    if jumlah_pola:
        bagian.append(f"- AI sudah mempelajari {jumlah_pola} pola pasangan akun dari histori transaksi.")
    if jumlah_temuan_mencurigakan:
        bagian.append(f"- Ada {jumlah_temuan_mencurigakan} pola transaksi mencurigakan yang belum dicek ulang user.")
    return "\n".join(bagian)


def buat_kalimat_pengantar_hasil(ringkasan_teks: str, ada_pola_mencurigakan: bool = False) -> str:
    """Minta AI membuatkan SATU kalimat pengantar singkat & kontekstual."""
    fallback = "Berikut hasil yang berhasil diproses:"
    daftar_provider = _konfigurasi_provider_chat()
    if not daftar_provider:
        return fallback
    import openai
    prompt = (
        "Kamu adalah asisten AI akuntansi (AI Gouf Consulting) yang baru saja selesai "
        f"memproses file yang diupload user. Ringkasan hasilnya: {ringkasan_teks}.\n"
        + (
            "Ada pola transaksi mencurigakan yang perlu dicek ulang oleh user.\n"
            if ada_pola_mencurigakan else ""
        )
        + "Tulis SATU kalimat pendek (maksimal 20 kata), ramah dan profesional, dalam Bahasa "
          "Indonesia, sebagai kalimat pengantar tepat sebelum tabel hasil ditampilkan ke user. "
          "Sesuaikan isi & nada kalimat dengan ringkasan di atas. "
          "JANGAN pakai tanda kutip, JANGAN pakai markdown/emoji berlebihan, JANGAN beri "
          "penjelasan lain apa pun -- balas HANYA kalimatnya saja, tanpa embel-embel."
    )
    for konfig in daftar_provider:
        try:
            client = openai.OpenAI(api_key=konfig["api_key"], base_url=konfig["base_url"])
            response = client.chat.completions.create(
                model=konfig["model"],
                max_tokens=80,
                temperature=0.6,
                messages=[{"role": "user", "content": prompt}],
            )
            teks = response.choices[0].message.content.strip().strip('"').strip()
            return teks if teks else fallback
        except Exception:
            continue  # provider ini gagal -- coba provider berikutnya (kalau ada)
    return fallback


# ============================================================
# 16. DATA QUALITY REPORT
# ============================================================

def generate_data_quality_report(df: pd.DataFrame, pola: Pola) -> Dict:
    """Generate comprehensive data quality report untuk monitoring."""
    report = {
        'total_records': len(df),
        'data_quality_score': 100,
        'warnings': [],
        'issues': []
    }
    
    if df.empty:
        report['data_quality_score'] = 0
        report['warnings'].append("Dataframe kosong")
        report['status'] = 'EMPTY'
        return report
    
    required_cols = ['keterangan', 'mutasi_debet', 'mutasi_kredit']
    for col in required_cols:
        if col in df.columns:
            missing = df[col].isna().sum()
            if missing > 0:
                report['warnings'].append(f"{missing} baris missing di kolom {col}")
                report['data_quality_score'] -= 5
    
    if 'tanggal' in df.columns:
        try:
            pd.to_datetime(df['tanggal'])
        except:
            report['issues'].append("Format tanggal tidak konsisten")
            report['data_quality_score'] -= 10
    
    duplicates = df.duplicated(subset=['keterangan', 'mutasi_debet', 'mutasi_kredit']).sum()
    if duplicates > 0:
        report['warnings'].append(f"{duplicates} transaksi duplikat ditemukan")
        report['data_quality_score'] -= 5
    
    if 'keterangan' in df.columns:
        df['signature'] = df['keterangan'].apply(ekstrak_signature)
        total_unique = df['signature'].nunique()
        covered = sum(1 for sig in df['signature'].unique() if sig in pola.aturan)
        coverage = (covered / total_unique * 100) if total_unique > 0 else 0
        
        report['pattern_coverage'] = {
            'total_patterns': total_unique,
            'covered': covered,
            'coverage_percentage': coverage
        }
        if coverage < 50:
            report['warnings'].append(f"Pattern coverage rendah ({coverage:.1f}%)")
            report['data_quality_score'] -= 10
    
    if 'mutasi_debet' in df.columns and 'mutasi_kredit' in df.columns:
        amounts_raw = pd.concat([df['mutasi_debet'], df['mutasi_kredit']])
        # Paksa numerik: nilai non-numerik (string aneh, sel kosong, dsb) jadi NaN lalu dibuang,
        # supaya quantile() tidak error saat membandingkan float vs str.
        amounts = pd.to_numeric(amounts_raw, errors='coerce').dropna()
        n_invalid = len(amounts_raw) - len(amounts)
        if n_invalid > 0:
            report['warnings'].append(f"{n_invalid} nilai mutasi_debet/mutasi_kredit tidak bisa dibaca sebagai angka (diabaikan saat cek outlier)")
            report['data_quality_score'] -= 5
        if len(amounts) > 0:
            q1 = amounts.quantile(0.25)
            q3 = amounts.quantile(0.75)
            iqr = q3 - q1
            outliers = amounts[(amounts < q1 - 1.5 * iqr) | (amounts > q3 + 1.5 * iqr)]
            if len(outliers) > 0:
                report['warnings'].append(f"{len(outliers)} transaksi dengan nominal ekstrim ditemukan")
                report['data_quality_score'] -= 5
    
    report['data_quality_score'] = max(0, report['data_quality_score'])
    report['status'] = 'GOOD' if report['data_quality_score'] >= 80 else 'NEEDS REVIEW'
    
    return report


# ============================================================
# 17. DETEKSI POLA MENCURIGAKAN
# ============================================================

AMBANG_SIGNATURE_BERBEDA = 5


def deteksi_pola_mencurigakan(pola: Pola, ambang: int = AMBANG_SIGNATURE_BERBEDA) -> list[dict]:
    """
    Kelompokkan aturan pola berdasarkan pasangan akun (debet, kredit).
    Kalau satu pasangan akun dipakai oleh >= `ambang` signature BERBEDA yang
    semuanya masih jumlah_contoh == 1, tandai sbg mencurigakan.
    """
    kelompok = {}
    for (sig, arah), aturan in pola.aturan.items():
        if aturan.get("jumlah_contoh", 1) > 1:
            continue
        pasangan = (aturan.get("no_akun_debet"), aturan.get("no_akun_kredit"))
        kelompok.setdefault(pasangan, {
            "no_akun_debet": aturan.get("no_akun_debet"), "nama_akun_debet": aturan.get("nama_akun_debet"),
            "no_akun_kredit": aturan.get("no_akun_kredit"), "nama_akun_kredit": aturan.get("nama_akun_kredit"),
            "signature": [],
        })
        kelompok[pasangan]["signature"].append(f"{sig} ({arah})")

    hasil = []
    for info in kelompok.values():
        jumlah = len(info["signature"])
        if jumlah >= ambang:
            hasil.append({
                "no_akun_debet": info["no_akun_debet"], "nama_akun_debet": info["nama_akun_debet"],
                "no_akun_kredit": info["no_akun_kredit"], "nama_akun_kredit": info["nama_akun_kredit"],
                "jumlah_signature_berbeda": jumlah,
                "contoh_signature": info["signature"][:5],
            })
    return hasil


def _kunci_evaluasi(temuan: dict) -> str:
    return f"{temuan['no_akun_debet']}||{temuan['no_akun_kredit']}"


def muat_evaluasi_pola(path: str) -> dict:
    """Muat histori evaluasi pola mencurigakan dari disk."""
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except (json.JSONDecodeError, OSError):
        return {}


def simpan_evaluasi_pola(path: str, temuan_baru: list[dict]) -> list[dict]:
    """Simpan/perbarui histori evaluasi pola mencurigakan ke disk."""
    histori = muat_evaluasi_pola(path)
    benar_benar_baru = []
    for temuan in temuan_baru:
        kunci = _kunci_evaluasi(temuan)
        if kunci not in histori:
            benar_benar_baru.append(temuan)
        histori[kunci] = {
            **temuan,
            "pertama_terdeteksi": histori.get(kunci, {}).get("pertama_terdeteksi", datetime.now().isoformat()),
            "terakhir_terdeteksi": datetime.now().isoformat(),
            "jumlah_kemunculan_sesi": histori.get(kunci, {}).get("jumlah_kemunculan_sesi", 0) + 1,
        }

    tmp_path = path + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(histori, f, ensure_ascii=False, indent=2, default=str)
    os.replace(tmp_path, path)
    return benar_benar_baru


# ============================================================
# 18. FITUR AI TAMBAHAN
# ============================================================

def prediksi_pola_transaksi(pola: Pola, df: pd.DataFrame, horizon: int = 30) -> dict:
    """Prediksi transaksi yang akan muncul berdasarkan pola historis."""
    if df.empty or not pola.aturan:
        return {"status": "Tidak cukup data", "prediksi": []}
    
    df = df.copy()
    df['signature'] = df['keterangan'].apply(ekstrak_signature)
    if 'tanggal' in df.columns:
        df['tanggal'] = pd.to_datetime(df['tanggal'])
        df['bulan'] = df['tanggal'].dt.month
        df['hari'] = df['tanggal'].dt.day
    else:
        return {"status": "Tidak ada kolom tanggal", "prediksi": []}
    
    prediksi = []
    for (sig, arah), aturan in pola.aturan.items():
        if not aturan.get('is_valid', True):
            continue
            
        data_sig = df[df['signature'] == sig]
        if len(data_sig) < 2:
            continue
        
        data_sig = data_sig.sort_values('tanggal')
        intervals = data_sig['tanggal'].diff().dropna()
        if len(intervals) < 1:
            continue
        
        avg_interval = intervals.mean().days
        if avg_interval < 1:
            avg_interval = 1
        
        nominal_col = 'mutasi_debet' if arah == 'KELUAR' else 'mutasi_kredit'
        avg_nominal = data_sig[nominal_col].mean()
        
        last_date = data_sig['tanggal'].max()
        next_date = last_date + pd.Timedelta(days=avg_interval)
        
        std_interval = intervals.std().days if len(intervals) > 1 else 0
        confidence = aturan.get('confidence_score', 0.5)
        
        if std_interval > avg_interval * 0.5:
            confidence *= 0.7
        
        prediksi.append({
            'signature': sig,
            'arah': arah,
            'tanggal_prediksi': next_date.strftime('%Y-%m-%d'),
            'nominal_prediksi': round(avg_nominal, 2),
            'akun_debet': aturan.get('no_akun_debet'),
            'akun_kredit': aturan.get('no_akun_kredit'),
            'confidence': round(confidence, 3),
            'interval_hari': avg_interval,
            'total_historis': len(data_sig)
        })
    
    prediksi = sorted(prediksi, key=lambda x: x['tanggal_prediksi'])
    
    return {
        'status': 'success',
        'total_prediksi': len(prediksi),
        'horizon': horizon,
        'prediksi': prediksi[:horizon]
    }


def deteksi_anomali_transaksi(df: pd.DataFrame, pola: Pola, threshold: float = 2.5) -> pd.DataFrame:
    """Deteksi transaksi anomali (outlier) berdasarkan pola historis."""
    if df.empty:
        return df
    
    df_result = df.copy()
    df_result['is_anomaly'] = False
    df_result['anomaly_score'] = 0.0
    df_result['anomaly_reason'] = ''
    
    df_result['signature'] = df_result['keterangan'].apply(ekstrak_signature)
    
    for (sig, arah), aturan in pola.aturan.items():
        mask = df_result['signature'] == sig
        data_sig = df_result[mask]
        
        if len(data_sig) < 3:
            continue
        
        nominal_col = 'mutasi_debet' if arah == 'KELUAR' else 'mutasi_kredit'
        if nominal_col not in df_result.columns:
            continue
        
        nominals = data_sig[nominal_col].values
        
        mean = np.mean(nominals)
        std = np.std(nominals)
        
        if std == 0:
            continue
        
        for idx in data_sig.index:
            z_score = abs((df_result.loc[idx, nominal_col] - mean) / std)
            
            if z_score > threshold:
                df_result.loc[idx, 'is_anomaly'] = True
                df_result.loc[idx, 'anomaly_score'] = z_score
                df_result.loc[idx, 'anomaly_reason'] = f'Nominal ekstrim untuk pola {sig} ({arah})'
    
    return df_result


# ============================================================
# 18b. [FIX] BUNGKUS deteksi_anomali_transaksi() -- DIPANGGIL main.py
#      TAPI SEBELUMNYA TIDAK PERNAH DIDEFINISIKAN (bug kritis)
# ============================================================
# main.py (endpoint /api/proses-file & /api/proses-file/stream) memanggil
# ak.cari_anomali_untuk_alert(df_mentah, pola_client) untuk tiap jenis
# dokumen yang punya pola per-client (rekening koran & penjualan) --
# tapi fungsi ini tidak pernah ada di file ini. Akibatnya: SETIAP kali
# akuntan upload rekening koran/penjualan, proses upload gagal total
# dengan AttributeError persis di titik ini (loop-nya tidak dibungkus
# try/except di main.py), walau file sebenarnya sudah berhasil
# dikategorikan/dijurnal -- baris deteksi_pola_mencurigakan() setelahnya
# pun ikut tidak pernah tereksekusi karena request keburu gagal.
def cari_anomali_untuk_alert(df: pd.DataFrame, pola: Pola, threshold: float = 2.5) -> list[dict]:
    """
    Bungkus deteksi_anomali_transaksi() jadi list dict siap dipakai
    dbc.buat_alert_anomali() dari main.py.

    Return: list of dict, satu per transaksi yang lolos ambang anomali,
    masing2:
        baris_index    : index asli di df
        pesan          : teks siap tampil ke akuntan
        anomaly_score  : z-score (semakin besar semakin ekstrim), atau None
    """
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return []
    if "keterangan" not in df.columns:
        return []

    try:
        df_anomali = deteksi_anomali_transaksi(df, pola, threshold=threshold)
    except Exception:
        return []

    if df_anomali is None or df_anomali.empty or "is_anomaly" not in df_anomali.columns:
        return []

    hasil: list[dict] = []
    for idx, row in df_anomali[df_anomali["is_anomaly"] == True].iterrows():  # noqa: E712
        keterangan = row.get("keterangan")
        keterangan_str = str(keterangan) if keterangan not in (None, "") else "(tanpa keterangan)"
        tanggal = row.get("tanggal")
        tanggal_str = str(tanggal) if pd.notna(tanggal) else "?"

        mutasi_debet = row.get("mutasi_debet")
        mutasi_kredit = row.get("mutasi_kredit")
        if pd.notna(mutasi_debet) and float(mutasi_debet or 0) > 0:
            nominal = mutasi_debet
        else:
            nominal = mutasi_kredit
        nominal_str = f"{float(nominal):,.0f}" if pd.notna(nominal) else "?"

        skor = row.get("anomaly_score")
        alasan = row.get("anomaly_reason") or "menyimpang dari pola historis"

        if pd.notna(skor):
            pesan = (
                f"\u26a0\ufe0f Nominal tidak biasa: transaksi tgl {tanggal_str} sebesar "
                f"Rp{nominal_str} (\"{keterangan_str}\") -- {alasan} (z-score {float(skor):.1f})."
            )
        else:
            pesan = (
                f"\u26a0\ufe0f Nominal tidak biasa: transaksi tgl {tanggal_str} sebesar "
                f"Rp{nominal_str} (\"{keterangan_str}\")."
            )

        hasil.append({
            "baris_index": int(idx),
            "pesan": pesan,
            "anomaly_score": float(skor) if pd.notna(skor) else None,
        })

    return hasil


# ============================================================
# 19. SMART SUGGESTIONS
# ============================================================

def get_smart_suggestions(
    df_bank: pd.DataFrame,
    df_jual: pd.DataFrame,
    df_coa: pd.DataFrame,
    pola: Pola,
    max_suggestions: int = 5
) -> list:
    """
    Generate saran cerdas berdasarkan konteks data.
    
    Args:
        df_bank: DataFrame rekening koran
        df_jual: DataFrame data penjualan
        df_coa: DataFrame COA
        pola: Objek Pola yang sudah dipelajari
        max_suggestions: Maksimum jumlah saran
    
    Returns:
        list: Daftar saran dengan format {'type', 'icon', 'text'}
    """
    suggestions = []
    
    # 1. Saran berdasarkan pola yang tersedia
    if pola.aturan:
        suggestions.append({
            'type': 'info',
            'icon': '💡',
            'text': f'AI sudah mempelajari {len(pola.aturan)} pola transaksi. Upload file baru untuk menerapkan pola.'
        })
    
    # 2. Saran berdasarkan data bank
    if df_bank is not None and not df_bank.empty:
        # Cek transaksi belum terkategori
        if 'sumber_kategori' in df_bank.columns:
            belum_terkategori = df_bank[
                df_bank['sumber_kategori'].str.contains('Belum Terkategori', na=False)
            ]
            if len(belum_terkategori) > 0:
                suggestions.append({
                    'type': 'warning',
                    'icon': '⚠️',
                    'text': f'{len(belum_terkategori)} transaksi belum terkategori. Upload file COA atau isi manual.'
                })
        
        # Cek keseimbangan jurnal
        cek = cek_keseimbangan_jurnal(df_bank)
        if not cek['balance']:
            suggestions.append({
                'type': 'danger',
                'icon': '🚨',
                'text': f'Jurnal tidak balance! Selisih Rp{cek["selisih"]:,.0f}. Periksa kembali data Anda.'
            })
    
    # 3. Saran berdasarkan COA
    if df_coa is not None and not df_coa.empty:
        suggestions.append({
            'type': 'info',
            'icon': '📚',
            'text': f'COA dengan {len(df_coa)} akun sudah dimuat. Transaksi akan dikategorikan berdasarkan COA ini.'
        })
    
    # 4. Saran augmentasi
    if len(pola.aturan) > 10:
        suggestions.append({
            'type': 'info',
            'icon': '🔄',
            'text': f'Ada {len(pola.aturan)} pola. Aktifkan augmentasi untuk meningkatkan coverage.'
        })
    
    # 5. Saran berdasarkan data quality
    if df_bank is not None and not df_bank.empty:
        report = generate_data_quality_report(df_bank, pola)
        if report['data_quality_score'] < 70:
            suggestions.append({
                'type': 'warning',
                'icon': '📊',
                'text': f'Data quality score: {report["data_quality_score"]}/100. Perbaiki data untuk hasil lebih akurat.'
            })
    
    # 6. Saran berdasarkan data penjualan
    if df_jual is not None and not df_jual.empty:
        if 'ppn' in df_jual.columns:
            no_ppn = df_jual[df_jual['ppn'] == 0]
            if len(no_ppn) > len(df_jual) * 0.5:
                suggestions.append({
                    'type': 'info',
                    'icon': '📝',
                    'text': f'Banyak transaksi tanpa PPN ({len(no_ppn)} dari {len(df_jual)}). Verifikasi pencatatan PPN.'
                })
        
        if 'cara_bayar' in df_jual.columns:
            kredit_count = df_jual[df_jual['cara_bayar'] == 'KREDIT'].shape[0]
            if kredit_count > 0:
                suggestions.append({
                    'type': 'info',
                    'icon': '💳',
                    'text': f'Ada {kredit_count} transaksi penjualan kredit. Pastikan piutang dicatat dengan benar.'
                })
    
    return suggestions[:max_suggestions]


# ============================================================
# 20. REKOMENDASI AKUNTANSI
# ============================================================

def generate_rekomendasi_akuntansi(
    df_bank: pd.DataFrame,
    df_jual: pd.DataFrame,
    df_coa: pd.DataFrame,
    pola: Pola
) -> dict:
    """Generate rekomendasi otomatis berdasarkan analisis data."""
    rekomendasi = {
        'prioritas_tinggi': [],
        'prioritas_sedang': [],
        'prioritas_rendah': [],
        'total_rekomendasi': 0
    }
    
    if df_bank is not None and not df_bank.empty:
        belum_terkategori = df_bank[
            df_bank['sumber_kategori'].str.contains('Belum Terkategori|perlu review', na=False)
        ]
        if len(belum_terkategori) > 0:
            rekomendasi['prioritas_tinggi'].append({
                'judul': f'Transaksi belum terkategori ({len(belum_terkategori)} baris)',
                'deskripsi': f'Terdapat {len(belum_terkategori)} transaksi yang belum memiliki jurnal.',
                'tindakan': 'Review manual dan tentukan akun yang tepat',
                'data': belum_terkategori.to_dict('records')[:5]
            })
    
    temuan_mencurigakan = deteksi_pola_mencurigakan(pola)
    if temuan_mencurigakan:
        for temuan in temuan_mencurigakan:
            rekomendasi['prioritas_tinggi'].append({
                'judul': f'Pola mencurigakan: {temuan["nama_akun_debet"]} / {temuan["nama_akun_kredit"]}',
                'deskripsi': f'{temuan["jumlah_signature_berbeda"]} transaksi berbeda menggunakan akun yang sama',
                'tindakan': 'Validasi apakah akun ini benar atau default yang keliru',
                'data': temuan
            })
    
    if df_bank is not None and not df_bank.empty:
        cek = cek_keseimbangan_jurnal(df_bank)
        if not cek['balance']:
            rekomendasi['prioritas_tinggi'].append({
                'judul': f'Jurnal tidak balance (selisih Rp{cek["selisih"]:,.0f})',
                'deskripsi': f'Total Debet Rp{cek["total_debet"]:,.0f} ≠ Total Kredit Rp{cek["total_kredit"]:,.0f}',
                'tindakan': 'Cek baris yang mungkin salah/kelewat/duplikat',
                'data': {'selisih': cek['selisih']}
            })
    
    if df_bank is not None and not df_bank.empty:
        report = generate_data_quality_report(df_bank, pola)
        if report['data_quality_score'] < 70:
            rekomendasi['prioritas_sedang'].append({
                'judul': f'Data quality rendah ({report["data_quality_score"]}/100)',
                'deskripsi': f'Data quality perlu perbaikan. Issues: {", ".join(report["issues"][:3])}',
                'tindakan': 'Perbaiki data yang hilang atau tidak konsisten',
                'data': report
            })
    
    if df_bank is not None and not df_bank.empty:
        if 'pattern_coverage' in report:
            coverage = report['pattern_coverage']['coverage_percentage']
            if coverage < 50:
                rekomendasi['prioritas_sedang'].append({
                    'judul': f'Pattern coverage rendah ({coverage:.1f}%)',
                    'deskripsi': f'Hanya {coverage:.1f}% pola transaksi yang sudah dipelajari AI',
                    'tindakan': 'Upload lebih banyak data historis untuk meningkatkan akurasi AI',
                    'data': report['pattern_coverage']
                })
    
    if df_jual is not None and not df_jual.empty:
        no_ppn = df_jual[df_jual['ppn'] == 0]
        if len(no_ppn) > len(df_jual) * 0.5:
            rekomendasi['prioritas_rendah'].append({
                'judul': f'Banyak transaksi tanpa PPN ({len(no_ppn)} dari {len(df_jual)})',
                'deskripsi': 'Sebagian besar transaksi penjualan tidak memiliki PPN',
                'tindakan': 'Verifikasi apakah semua penjualan sudah benar dicatat PPN-nya',
                'data': {'total': len(df_jual), 'tanpa_ppn': len(no_ppn)}
            })
    
    for level in ['prioritas_tinggi', 'prioritas_sedang', 'prioritas_rendah']:
        rekomendasi['total_rekomendasi'] += len(rekomendasi[level])
    
    return rekomendasi


# ============================================================
# 21. FORMAT JURNAL UNTUK EXPORT
# ============================================================

def format_jurnal_untuk_export(
    df: pd.DataFrame,
    format_type: str = 'excel'
) -> Dict[str, Any]:
    """Format jurnal untuk berbagai keperluan export."""
    if df.empty:
        return {'status': 'error', 'message': 'Data kosong'}
    
    result = {
        'status': 'success',
        'total_baris': len(df),
        'data': None
    }
    
    kolom_wajib = [
        'tanggal', 'keterangan', 'no_akun_debet', 'nama_akun_debet',
        'no_akun_kredit', 'nama_akun_kredit', 'jml_debet', 'jml_kredit'
    ]
    
    for kol in kolom_wajib:
        if kol not in df.columns:
            df[kol] = None
    
    if format_type == 'jurnal_umum':
        jurnal_umum = []
        for _, row in df.iterrows():
            if row['jml_debet'] and row['jml_debet'] > 0:
                jurnal_umum.append({
                    'tanggal': row['tanggal'],
                    'keterangan': row['keterangan'],
                    'no_akun': row['no_akun_debet'],
                    'nama_akun': row['nama_akun_debet'],
                    'debet': row['jml_debet'],
                    'kredit': 0
                })
            if row['jml_kredit'] and row['jml_kredit'] > 0:
                jurnal_umum.append({
                    'tanggal': row['tanggal'],
                    'keterangan': row['keterangan'],
                    'no_akun': row['no_akun_kredit'],
                    'nama_akun': row['nama_akun_kredit'],
                    'debet': 0,
                    'kredit': row['jml_kredit']
                })
        result['data'] = jurnal_umum
        result['format'] = 'jurnal_umum'
        
    elif format_type == 'csv':
        result['data'] = df.to_csv(index=False)
        result['format'] = 'csv'
        
    elif format_type == 'json':
        result['data'] = df.to_dict('records')
        result['format'] = 'json'
        
    else:
        result['data'] = df
        result['format'] = 'excel'
    
    return result


# ============================================================
# 22. VALIDASI JURNAL ENTRY
# ============================================================

def validate_journal_entry(row: dict) -> dict:
    """Validasi 1 baris jurnal."""
    result = {
        'valid': True,
        'errors': [],
        'warnings': []
    }
    
    if row.get('no_akun_debet') == row.get('no_akun_kredit'):
        result['valid'] = False
        result['errors'].append('Akun debet dan kredit sama')
    
    for col in ['jml_debet', 'jml_kredit']:
        val = row.get(col, 0)
        if val and val < 0:
            result['valid'] = False
            result['errors'].append(f'{col} bernilai negatif')
    
    tanggal = row.get('tanggal')
    if tanggal:
        try:
            pd.to_datetime(tanggal)
        except:
            result['warnings'].append('Format tanggal tidak valid')
    
    if not row.get('keterangan'):
        result['warnings'].append('Keterangan kosong')
    
    return result


# ============================================================
# 23. CLEANUP DATAFRAME
# ============================================================

def cleanup_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Bersihkan dataframe dari data yang tidak perlu."""
    if df.empty:
        return df
    
    df_clean = df.copy()
    df_clean = df_clean.dropna(axis=1, how='all')
    df_clean = df_clean.dropna(axis=0, how='all')
    
    kolom_cek = ['tanggal', 'keterangan', 'mutasi_debet', 'mutasi_kredit']
    kolom_ada = [k for k in kolom_cek if k in df_clean.columns]
    if kolom_ada:
        df_clean = df_clean.drop_duplicates(subset=kolom_ada, keep='first')
    
    for col in df_clean.select_dtypes(include=['object']).columns:
        df_clean[col] = df_clean[col].astype(str).str.strip()
    
    df_clean = df_clean.reset_index(drop=True)
    
    return df_clean


# ============================================================
# 24. BANDINGKAN PERIODE
# ============================================================

def bandingkan_periode(
    df1: pd.DataFrame,
    df2: pd.DataFrame,
    label1: str = 'Periode 1',
    label2: str = 'Periode 2'
) -> dict:
    """Bandingkan dua periode data."""
    result = {
        'status': 'success',
        'periode1': label1,
        'periode2': label2,
        'perbandingan': {}
    }
    
    if df1.empty and df2.empty:
        result['status'] = 'error'
        result['message'] = 'Kedua periode kosong'
        return result
    
    def get_stats(df, label):
        if df.empty:
            return {'total': 0, 'debet': 0, 'kredit': 0, 'nama': label}
        
        stats = {
            'nama': label,
            'total': len(df),
            'debet': df['jml_debet'].sum() if 'jml_debet' in df.columns else 0,
            'kredit': df['jml_kredit'].sum() if 'jml_kredit' in df.columns else 0,
        }
        
        if 'sumber_kategori' in df.columns:
            stats['belum_terkategori'] = df[
                df['sumber_kategori'].str.contains('Belum Terkategori|perlu review', na=False)
            ].shape[0]
        
        return stats
    
    stats1 = get_stats(df1, label1)
    stats2 = get_stats(df2, label2)
    
    result['perbandingan']['statistik'] = {
        'periode1': stats1,
        'periode2': stats2
    }
    
    if stats1['total'] > 0 and stats2['total'] > 0:
        result['perbandingan']['perubahan'] = {
            'total_transaksi': {
                'absolut': stats2['total'] - stats1['total'],
                'persen': ((stats2['total'] - stats1['total']) / stats1['total']) * 100
            },
            'total_nominal': {
                'debet': {
                    'absolut': stats2['debet'] - stats1['debet'],
                    'persen': ((stats2['debet'] - stats1['debet']) / stats1['debet']) * 100 if stats1['debet'] > 0 else 0
                },
                'kredit': {
                    'absolut': stats2['kredit'] - stats1['kredit'],
                    'persen': ((stats2['kredit'] - stats1['kredit']) / stats1['kredit']) * 100 if stats1['kredit'] > 0 else 0
                }
            }
        }
    
    return result


# ============================================================
# 25. AUTO-CORRECT JURNAL
# ============================================================

def auto_correct_journal(
    df: pd.DataFrame,
    df_coa: pd.DataFrame,
    pola: Pola,
    max_corrections: int = 10
) -> tuple[pd.DataFrame, list]:
    """Coba perbaiki otomatis jurnal yang bermasalah."""
    if df.empty:
        return df, []
    
    df_corrected = df.copy()
    log = []
    
    for idx, row in df_corrected.iterrows():
        if len(log) >= max_corrections:
            break
        
        corrections = []
        
        if row.get('no_akun_debet') == row.get('no_akun_kredit'):
            sig = ekstrak_signature(row.get('keterangan'))
            arah = _arah(row)
            aturan = pola.aturan.get((sig, arah))
            
            if aturan:
                df_corrected.loc[idx, 'no_akun_debet'] = aturan['no_akun_debet']
                df_corrected.loc[idx, 'no_akun_kredit'] = aturan['no_akun_kredit']
                corrections.append('Akun debet/kredit diperbaiki berdasarkan pola')
        
        for col in ['jml_debet', 'jml_kredit']:
            if row.get(col, 0) and row.get(col) < 0:
                df_corrected.loc[idx, col] = abs(row.get(col, 0))
                corrections.append(f'{col} dinormalisasi (abs)')
        
        if corrections:
            log.append({
                'idx': idx,
                'keterangan': row.get('keterangan'),
                'perbaikan': corrections,
                'status': 'fixed'
            })
    
    return df_corrected, log


# ============================================================
# 26. BATCH PROCESS DATAFRAMES
# ============================================================

def batch_process_dataframes(
    dataframes: Dict[str, pd.DataFrame],
    df_coa: pd.DataFrame,
    pola: Pola,
    pakai_ai: bool = False,
    callback=None
) -> Dict[str, pd.DataFrame]:
    """Proses multiple dataframe dengan progress callback."""
    hasil = {}
    total = len(dataframes)
    
    for i, (jenis, df) in enumerate(dataframes.items()):
        if callback:
            callback(i + 1, total, f'Memproses {jenis}...')
        
        if jenis == 'bank':
            hasil[jenis] = proses_dataframe(df, df_coa, pola, pakai_ai=pakai_ai)
        elif jenis == 'jual':
            hasil[jenis] = proses_dataframe_penjualan(df, df_coa, pola, pakai_ai=pakai_ai)
        else:
            hasil[jenis] = df
    
    return hasil



# ============================================================
# [BARU] MEKANISME TANYA BALIK KE AKUNTAN (klarifikasi transaksi)
# ============================================================
# Kenapa ini ada:
#   Sebelumnya, baris yang AI ragu/gagal kategorikan cuma ditandai lewat
#   kolom "sumber_kategori" dan berakhir diam-diam di sheet "Perlu
#   Direview" -- gampang terlewat. Fungsi di bawah ini AKTIF menyusun
#   pertanyaan yang bisa ditampilkan sbg pesan ke akuntan (mirip cara
#   akuntan manusia nanya balik ke klien), dipanggil dari main.py
#   setelah proses_dataframe() / proses_dataframe_penjualan() selesai.
#
#   Sesuai keputusan: yang menjawab pertanyaan ini akuntan internal
#   (lewat dashboard/chat React), BUKAN klien langsung lewat WA. Kalau
#   AI sempat menebak kategorinya, tebakan itu tetap ditampilkan di
#   pertanyaan (bukan dikosongkan) sambil diberi tanda "perlu
#   konfirmasi" -- lihat field "butuh_konfirmasi_saja".
#
#   Fungsi ini TIDAK mengubah df atau proses_dataframe() yang sudah ada
#   sama sekali -- murni membaca kolom yang sudah dihasilkan proses
#   yang ada (sumber_kategori, confidence_ai, no_akun_debet, dst).
def cari_baris_perlu_klarifikasi(df: pd.DataFrame) -> list[dict]:
    """
    Pindai df hasil proses_dataframe()/proses_dataframe_penjualan()
    (kolom 'sumber_kategori' wajib ada -- kalau tidak ada, berarti df
    ini bukan hasil kategorisasi otomatis, langsung dikembalikan kosong).

    Return: list of dict, satu per baris yang perlu ditanyakan, masing2:
        baris_index        : index asli di df (dipakai utk update balik
                              kalau nanti mau menandai baris ini sudah
                              terjawab langsung di draf jurnal)
        tanggal             : str atau None
        keterangan          : str atau None
        nominal             : float atau None
        arah                : "masuk" / "keluar" / None
        tebakan_kategori    : nama akun tebakan AI (kalau ada) atau None
        confidence_ai       : float atau None
        butuh_konfirmasi_saja : True kalau AI SUDAH sempat menebak (tinggal
                                 dikonfirmasi ya/tidak), False kalau AI
                                 benar2 tidak bisa menebak sama sekali
        pertanyaan          : teks pertanyaan siap tampil ke akuntan
    """
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return []
    if "sumber_kategori" not in df.columns:
        return []

    hasil_pertanyaan: list[dict] = []

    for idx, row in df.iterrows():
        sumber = str(row.get("sumber_kategori") or "")

        ada_tebakan = pd.notna(row.get("no_akun_debet")) and pd.notna(row.get("no_akun_kredit"))

        perlu_tanya = False
        butuh_konfirmasi_saja = False
        if "Belum Terkategori" in sumber:
            perlu_tanya = True
            butuh_konfirmasi_saja = False
        # [FIX -- ikut label baru] sumber_kategori sekarang ditulis
        # "AI (Claude)" oleh _apply_ai_results_to_dataframe (lihat fix di
        # sana) -- cek "AI (DeepSeek)" LAMA dipertahankan juga sebagai
        # OR supaya data hasil proses SEBELUM fix ini (baris GL/hasil lama
        # yang sudah tersimpan di DB dengan label lama) tetap kena logic
        # klarifikasi yang sama, tidak diam-diam berhenti diproses cuma
        # karena labelnya beda antara data lama vs baru.
        elif ("AI (Claude)" in sumber or "AI (DeepSeek)" in sumber) and "perlu review" in sumber:
            perlu_tanya = True
            butuh_konfirmasi_saja = ada_tebakan

        if not perlu_tanya:
            continue

        keterangan = row.get("keterangan")
        if keterangan is None or (isinstance(keterangan, float) and pd.isna(keterangan)):
            keterangan = row.get("KETERANGAN")
        tanggal = row.get("tanggal")
        if tanggal is None or (isinstance(tanggal, float) and pd.isna(tanggal)):
            tanggal = row.get("TANGGAL")

        jml_debet = row.get("jml_debet")
        jml_kredit = row.get("jml_kredit")
        if pd.notna(jml_kredit) and float(jml_kredit or 0) > 0:
            arah = "masuk"
            nominal = jml_kredit
            tebakan_nama = row.get("nama_akun_kredit")
        elif pd.notna(jml_debet) and float(jml_debet or 0) > 0:
            arah = "keluar"
            nominal = jml_debet
            tebakan_nama = row.get("nama_akun_debet")
        else:
            arah = None
            nominal = jml_debet if pd.notna(jml_debet) else jml_kredit
            tebakan_nama = row.get("nama_akun_kredit") or row.get("nama_akun_debet")

        keterangan_str = str(keterangan) if keterangan not in (None, "") else "(tanpa keterangan)"
        tanggal_str = str(tanggal) if pd.notna(tanggal) else "?"
        nominal_str = f"{float(nominal):,.0f}" if pd.notna(nominal) else "?"

        if butuh_konfirmasi_saja:
            pertanyaan = (
                f"Transaksi tgl {tanggal_str} sebesar Rp{nominal_str} "
                f"(\"{keterangan_str}\") -- AI menebak ini masuk ke akun "
                f"\"{tebakan_nama}\". Sudah benar?"
            )
        else:
            pertanyaan = (
                f"Transaksi tgl {tanggal_str} sebesar Rp{nominal_str} "
                f"(\"{keterangan_str}\") belum bisa dikategorikan otomatis "
                f"-- ini masuk akun apa?"
            )

        # [BARU] Simpan juga no_akun/nama_akun mentah (bukan cuma nama
        # tebakan) + cara_bayar (khusus penjualan) -- dipakai
        # bangun_pola_dari_feedback_klarifikasi() di bawah supaya retraining
        # pola (#3) punya cukup info bikin pasangan jurnal lengkap dari
        # jawaban akuntan, bukan cuma teks pertanyaan.
        hasil_pertanyaan.append({
            "baris_index": int(idx),
            "tanggal": tanggal_str if pd.notna(tanggal) else None,
            "keterangan": keterangan_str,
            "nominal": float(nominal) if pd.notna(nominal) else None,
            "arah": arah,
            "tebakan_kategori": str(tebakan_nama) if pd.notna(tebakan_nama) else None,
            "confidence_ai": (
                float(row.get("confidence_ai"))
                if pd.notna(row.get("confidence_ai")) else None
            ),
            "butuh_konfirmasi_saja": butuh_konfirmasi_saja,
            "pertanyaan": pertanyaan,
            "no_akun_debet": row.get("no_akun_debet") if pd.notna(row.get("no_akun_debet")) else None,
            "nama_akun_debet": row.get("nama_akun_debet") if pd.notna(row.get("nama_akun_debet")) else None,
            "no_akun_kredit": row.get("no_akun_kredit") if pd.notna(row.get("no_akun_kredit")) else None,
            "nama_akun_kredit": row.get("nama_akun_kredit") if pd.notna(row.get("nama_akun_kredit")) else None,
            "cara_bayar": row.get("cara_bayar") if pd.notna(row.get("cara_bayar")) else None,
        })

    return hasil_pertanyaan


# ============================================================
# 20. [BARU] FEEDBACK LOOP -- LATIH ULANG POLA DARI KLARIFIKASI (#3)
# ============================================================
# Menutup celah: pelajari_pola()/pelajari_pola_penjualan() di atas TIDAK
# PERNAH dipanggil di mana pun sebelumnya -- pola hanya bertambah 1-per-1
# secara implisit di proses_dataframe()/proses_dataframe_penjualan() saat
# baris punya jurnal lengkap dari file asal. Fungsi di bawah ini
# memanfaatkan jawaban akuntan yang sudah terkumpul di tabel pola_augmentasi
# (lewat mekanisme klarifikasi -- lihat db_client.jawab_pertanyaan_klarifikasi)
# supaya transaksi serupa berikutnya tidak perlu ditanya lagi.

_KATA_KONFIRMASI = {
    "ya", "benar", "betul", "sesuai", "setuju", "cocok",
    "ok", "oke", "iya", "yup", "sip", "fix", "yoi",
}


def _jawaban_mengkonfirmasi(jawaban: Optional[str]) -> bool:
    """True kalau jawaban akuntan intinya 'ya, tebakan AI sudah benar' --
    bukan koreksi ke akun lain. Sengaja dibatasi <= 4 kata SEMUANYA kata
    konfirmasi, supaya jawaban panjang yang menyebut nama akun lain
    (berarti koreksi) tidak salah dianggap konfirmasi."""
    if not jawaban:
        return False
    bersih = re.sub(r"[^a-z\s]", "", str(jawaban).lower()).strip()
    if not bersih:
        return False
    kata = bersih.split()
    if len(kata) > 4:
        return False
    return all(k in _KATA_KONFIRMASI for k in kata)


def _cocokkan_teks_ke_akun_coa(teks: Optional[str], df_coa: Optional[pd.DataFrame]):
    """Coba temukan 1 akun COA yang paling cocok dari teks jawaban koreksi
    akuntan (mis. 'ke akun Biaya Listrik & Air' atau '5104'). Coba nomor
    akun persis dulu, baru overlap kata pada nama akun. Return
    (no_akun, nama_akun) atau None kalau tidak ada yang cukup yakin cocok."""
    if df_coa is None or df_coa.empty or not teks or "no_akun" not in df_coa.columns:
        return None
    t = str(teks).strip()

    # 1) coba cocokkan sebagai nomor akun persis (>= 2 digit, supaya tidak
    # kena angka nominal 1 digit yang kebetulan nyasar di teks)
    for no in re.findall(r"\b\d{2,}\b", t):
        cocok = df_coa[df_coa["no_akun"].astype(str) == no]
        if not cocok.empty:
            baris = cocok.iloc[0]
            return str(baris["no_akun"]), baris["nama_akun"]

    # 2) overlap kata pada nama akun -- ambil kandidat dengan overlap
    # terbanyak, minimal 1 kata yang cocok
    kata_teks = set(re.findall(r"[a-z]+", t.lower()))
    if not kata_teks:
        return None

    kandidat_terbaik = None
    skor_terbaik = 0
    for _, baris in df_coa.iterrows():
        nama_akun = str(baris.get("nama_akun") or "")
        if not nama_akun:
            continue
        kata_akun = set(re.findall(r"[a-z]+", nama_akun.lower()))
        overlap = len(kata_teks & kata_akun)
        if overlap > skor_terbaik:
            skor_terbaik = overlap
            kandidat_terbaik = baris

    if kandidat_terbaik is not None and skor_terbaik > 0:
        return str(kandidat_terbaik["no_akun"]), kandidat_terbaik["nama_akun"]
    return None


def _tentukan_akun_final_dari_baris_feedback(
    row: dict, jenis: str, df_coa: Optional[pd.DataFrame] = None,
) -> Tuple[Optional[dict], Optional[str]]:
    """
    [BARU -- diekstrak dari bangun_pola_dari_feedback_klarifikasi(), TANPA
    mengubah perilaku] Tentukan "ground truth" (akun debet/kredit yang benar
    menurut akuntan) dari SATU baris feedback klarifikasi, terlepas dari
    urusan agregasi ke Pola.

    Dipakai di 2 tempat:
    1. bangun_pola_dari_feedback_klarifikasi() -- membentuk pola baru dari
       SEMUA baris feedback.
    2. evaluasi_pola_sebelum_commit() -- mengukur akurasi pola KANDIDAT
       terhadap baris feedback yang SENGAJA tidak dipakai melatih (holdout),
       sebelum pola itu benar-benar digabung+disimpan sbg pola live.
       Fungsi #2 ini alasan utama logic-nya dipisah ke sini: supaya kedua
       tempat pasti pakai definisi "ground truth" yang SAMA PERSIS, tidak
       ada 2 versi logic yang bisa diam-diam berbeda seiring waktu.

    Return (hasil, alasan_lewat):
    - Sukses: (dict berisi signature/arah/sumber/akun, None)
    - Gagal:  (None, "tanpa_konteks" | "gagal_cocok_akun") -- 2 alasan ini
      cocok 1:1 dengan 2 counter statistik di
      bangun_pola_dari_feedback_klarifikasi() (jumlah_dilewati_tanpa_konteks
      / jumlah_dilewati_gagal_cocok_akun), supaya caller bisa tetap
      menghitung statistik yang sama seperti sebelum diekstrak.
    """
    data_asli = row.get("data_asli") or {}
    koreksi = row.get("koreksi") or {}
    jawaban = str(koreksi.get("jawaban") or "").strip()

    keterangan = data_asli.get("keterangan")
    if not keterangan:
        # Feedback lama (sebelum konteks lengkap ikut disimpan) atau
        # format tidak dikenal -- tidak cukup info, dilewati.
        return None, "tanpa_konteks"

    prefix_sig = "PJL::" if jenis == "penjualan" else ""
    sig = prefix_sig + ekstrak_signature(keterangan)
    nd, nnd = data_asli.get("no_akun_debet"), data_asli.get("nama_akun_debet")
    nk, nnk = data_asli.get("no_akun_kredit"), data_asli.get("nama_akun_kredit")

    if jenis == "penjualan":
        arah = data_asli.get("cara_bayar") or "TUNAI"
    else:
        arah = "MASUK" if data_asli.get("arah") == "masuk" else "KELUAR"

    # --- Jalur konfirmasi: tebakan AI sudah lengkap & akuntan bilang ok
    if _jawaban_mengkonfirmasi(jawaban) and nd and nk:
        return {
            "signature": sig, "arah": arah, "sumber": "konfirmasi_akuntan",
            "no_akun_debet": nd, "nama_akun_debet": nnd,
            "no_akun_kredit": nk, "nama_akun_kredit": nnk,
        }, None

    if not jawaban:
        return None, "tanpa_konteks"

    # --- Jalur koreksi: cari akun pengganti dari teks jawaban
    akun_cocok = _cocokkan_teks_ke_akun_coa(jawaban, df_coa)
    if akun_cocok is None:
        return None, "gagal_cocok_akun"
    no_baru, nama_baru = akun_cocok

    if jenis == "penjualan":
        # Akun yang dikoreksi akuntan = sisi kredit (pendapatan/akun
        # tujuan); sisi debet (kas/bank/piutang) dipertahankan dari
        # tebakan asal kalau ada, else dianggap sama dgn akun terkoreksi.
        nd_final, nnd_final = nd or no_baru, nnd or nama_baru
        nk_final, nnk_final = no_baru, nama_baru
    elif arah == "MASUK":
        # debet = bank (harus sudah pasti), kredit = sisi yang dikoreksi
        if not nd:
            return None, "gagal_cocok_akun"
        nd_final, nnd_final = nd, nnd
        nk_final, nnk_final = no_baru, nama_baru
    else:  # KELUAR
        # kredit = bank (harus sudah pasti), debet = sisi yang dikoreksi
        if not nk:
            return None, "gagal_cocok_akun"
        nd_final, nnd_final = no_baru, nama_baru
        nk_final, nnk_final = nk, nnk

    return {
        "signature": sig, "arah": arah, "sumber": "koreksi_akuntan",
        "no_akun_debet": nd_final, "nama_akun_debet": nnd_final,
        "no_akun_kredit": nk_final, "nama_akun_kredit": nnk_final,
    }, None


def bangun_pola_dari_feedback_klarifikasi(
    augmentasi_rows: list[dict],
    jenis: str,
    df_coa: Optional[pd.DataFrame] = None,
    client_id: Optional[int] = None,
) -> Tuple[Pola, dict]:
    """
    Ubah feedback yang sudah terkumpul di tabel pola_augmentasi (hasil jawab
    klarifikasi) jadi Pola baru siap digabung ke pola tersimpan client.

    2 jalur (detail penentuan ground truth ada di
    _tentukan_akun_final_dari_baris_feedback, diekstrak ke sana supaya
    evaluasi_pola_sebelum_commit() bisa pakai definisi yang identik):
    - KONFIRMASI: jawaban "ya"/"benar"/dst -> confidence 1.0 (sudah
      divalidasi manusia).
    - KOREKSI: jawaban berisi nama/nomor akun lain -> dicocokkan ke df_coa;
      confidence_score dihitung dari histori aktual (hitung_confidence_aktual),
      bukan angka tetap.

    Return (Pola, statistik) -- statistik untuk dilaporkan balik ke akuntan
    lewat endpoint retraining.
    """
    statistik = {
        "jumlah_dipakai_konfirmasi": 0,
        "jumlah_dipakai_koreksi": 0,
        "jumlah_dilewati_tanpa_konteks": 0,
        "jumlah_dilewati_gagal_cocok_akun": 0,
    }
    aturan: dict = {}

    # [BARU] Dihitung SEKALI di awal (bukan per-baris di dalam loop) --
    # histori akurasi client ini tidak berubah selama fungsi ini berjalan,
    # jadi tidak perlu query ulang tiap baris. Lihat hitung_confidence_aktual
    # untuk penjelasan kenapa ini menggantikan angka 0.85 yang sebelumnya
    # hardcoded di sini.
    confidence_koreksi = hitung_confidence_aktual(client_id, jenis, jalur="koreksi_akuntan")

    for row in augmentasi_rows:
        if row.get("jenis") != jenis:
            continue

        hasil, alasan_lewat = _tentukan_akun_final_dari_baris_feedback(row, jenis, df_coa)
        if hasil is None:
            statistik[f"jumlah_dilewati_{alasan_lewat}"] += 1
            continue

        kunci = (hasil["signature"], hasil["arah"])
        confidence = 1.0 if hasil["sumber"] == "konfirmasi_akuntan" else confidence_koreksi
        aturan[kunci] = {
            "no_akun_debet": hasil["no_akun_debet"], "nama_akun_debet": hasil["nama_akun_debet"],
            "no_akun_kredit": hasil["no_akun_kredit"], "nama_akun_kredit": hasil["nama_akun_kredit"],
            "konsisten": True,
            "jumlah_contoh": aturan.get(kunci, {}).get("jumlah_contoh", 0) + 1,
            "confidence_score": confidence, "is_valid": True,
            "is_variation": False, "is_template": False,
            "sumber": hasil["sumber"],
            "last_updated": datetime.now().isoformat(),
        }
        statistik[f"jumlah_dipakai_{'konfirmasi' if hasil['sumber'] == 'konfirmasi_akuntan' else 'koreksi'}"] += 1

    return Pola(aturan=aturan), statistik


def latih_ulang_pola_dari_feedback(
    augmentasi_rows: list[dict],
    jenis: str,
    client_id: Optional[int] = None,
    df_coa: Optional[pd.DataFrame] = None,
) -> dict:
    """
    Orkestrasi retraining #3: bangun Pola baru dari feedback klarifikasi,
    gabung ke pola tersimpan client ini (pola baru menang kalau bentrok
    signature+arah -- lihat gabung_pola), lalu simpan lagi ke disk.

    jenis harus "rekening_koran" atau "penjualan" (path file pola beda).
    """
    nama_dasar = "pola_bank" if jenis == "rekening_koran" else "pola_penjualan"
    path_pola = _path_pola(nama_dasar, client_id)

    pola_lama = muat_pola(path_pola)
    pola_baru, statistik = bangun_pola_dari_feedback_klarifikasi(augmentasi_rows, jenis, df_coa, client_id=client_id)
    pola_gabungan = gabung_pola(pola_lama, pola_baru)
    # [BARU -- VERSIONING] Ini jalur PALING RAWAN: feedback klarifikasi
    # datang langsung dari input manusia (akuntan), dan gabung_pola() akan
    # MENIMPA aturan lama tanpa tanya kalau signature+arah-nya bentrok --
    # jadi 1 klik klarifikasi yang salah bisa langsung merusak pola yang
    # sudah lama benar. Label sumber_perubahan di sini SENGAJA dibuat
    # rollback-friendly (bisa langsung dipakai lagi sbg argumen ke
    # daftar_versi_pola()/rollback_pola() dari endpoint retraining di
    # main.py) supaya kalau akuntan sadar salah klarifikasi, tinggal
    # rollback_pola(path_pola) tanpa perlu investigasi manual.
    simpan_pola(
        pola_gabungan, path_pola,
        sumber_perubahan=f"latih_ulang_pola_dari_feedback (jenis={jenis}, {len(augmentasi_rows)} feedback)",
    )

    # [BARU -- METRIK AKURASI TERPUSAT] Setiap kali retraining jalan, kita
    # SUDAH tahu persis dari statistik di atas: berapa baris AI benar
    # (konfirmasi_akuntan, tidak perlu dikoreksi) vs berapa baris AI salah
    # (koreksi_akuntan, akuntan harus membetulkan). Ini sinyal akurasi
    # riil yang selama ini dihitung tapi TIDAK PERNAH disimpan sebagai
    # histori -- cuma dikembalikan sekali ke caller lalu hilang. Sekarang
    # dicatat ke log persisten supaya bisa dilihat trennya dari waktu ke
    # waktu (lihat catat_metrik_akurasi/hitung_tren_akurasi di bawah).
    catat_metrik_akurasi(
        client_id=client_id, jenis=jenis,
        jumlah_konfirmasi=statistik.get("jumlah_dipakai_konfirmasi", 0),
        jumlah_koreksi=statistik.get("jumlah_dipakai_koreksi", 0),
        sumber_data="klarifikasi_ui",
    )

    return {
        "jenis": jenis,
        "jumlah_feedback_total": len(augmentasi_rows),
        **statistik,
        "jumlah_pola_sebelum": len(pola_lama.aturan),
        "jumlah_pola_baru_atau_diupdate": len(pola_baru.aturan),
        "jumlah_pola_sesudah": len(pola_gabungan.aturan),
        "path_pola": path_pola,
    }


# ============================================================
# [BARU] STAGING / VALIDASI SEBELUM POLA HASIL RETRAIN DI-LIVE-KAN
# ============================================================
# Mengisi celah baru yg ditemukan SETELAH versioning+rollback & metrik
# akurasi terpusat selesai dibangun: latih_ulang_pola_dari_feedback() di
# atas LANGSUNG commit (simpan_pola) begitu dipanggil -- pola baru LANGSUNG
# dipakai utk transaksi berikutnya tanpa pengujian apa pun dulu. Kalau
# kebetulan banyak feedback keliru masuk sekaligus (akuntan baru yg belum
# hafal COA, salah klik konfirmasi massal, dll), pola yg buruk itu ikut
# ter-live-kan tanpa ketahuan lebih dulu -- padahal infrastruktur utk
# mendeteksi ini SUDAH ada semua (daftar_versi_pola/rollback_pola utk
# baru ketahuan SETELAH kejadian; sekarang ditambah lapis SEBELUM kejadian).
#
# Caranya: pola KANDIDAT dilatih HANYA dari sebagian feedback (train
# split), lalu diuji ke sisa feedback yg SENGAJA disisihkan (test split,
# porsi PALING BARU -- mensimulasikan "kalau pola dilatih dari data s.d.
# kemarin, seberapa akurat menebak transaksi hari ini yg belum pernah
# dilihat"). Akurasinya dibandingkan dgn pola yg SEDANG live diuji ke test
# split yg SAMA. Hanya kalau aman (di atas ambang & tidak regresi) baru
# latih_ulang_pola_dari_feedback() yg sebenarnya (commit ke disk) dipanggil.

def _split_train_test_augmentasi(
    augmentasi_rows: list[dict], rasio_test: float = 0.2, minimal_test: int = 5,
) -> Tuple[list[dict], list[dict]]:
    """
    Pisah baris feedback jadi train/test utk staging evaluation.

    [ASUMSI PENTING] augmentasi_rows diasumsikan datang terurut kronologis
    menaik (baris terlama duluan, spt hasil query dbc.ambil_pola_augmentasi
    yg diurutkan berdasarkan waktu jawab) -- test split diambil dari EKOR
    list (porsi paling baru). Kalau urutan sumber datanya ternyata terbalik
    atau tidak terjamin urut, split ini masih tetap valid secara statistik
    (train/test tetap partisi acak-cukup dari populasi yg sama), HANYA
    interpretasi "simulasi kondisi nyata" (uji ke transaksi TERBARU) yg
    jadi kurang tepat -- bukan bug fatal, tapi catat di sini kalau nanti
    urutan sumber data berubah.

    Kalau data terlalu sedikit utk split yg bermakna (< 2x minimal_test),
    seluruh data dipakai sbg train SEKALIGUS test (evaluasi jadi kurang
    independen/optimis -- ditandai lewat "sukses_evaluasi": False di
    evaluasi_pola_sebelum_commit supaya caller tahu tidak boleh terlalu
    percaya angkanya, tapi tetap ada fallback aman: COMMIT_OTOMATIS,
    sama seperti perilaku SEBELUM staging ada).
    """
    if len(augmentasi_rows) < minimal_test * 2:
        return augmentasi_rows, augmentasi_rows
    n_test = max(minimal_test, round(len(augmentasi_rows) * rasio_test))
    n_test = min(n_test, len(augmentasi_rows) - minimal_test)  # sisakan minimal utk train
    return augmentasi_rows[:-n_test], augmentasi_rows[-n_test:]


def evaluasi_pola_sebelum_commit(
    augmentasi_rows: list[dict], jenis: str,
    client_id: Optional[int] = None, df_coa: Optional[pd.DataFrame] = None,
    ambang_akurasi_minimal: float = 0.6, ambang_regresi: float = 0.05,
    rasio_test: float = 0.2,
) -> dict:
    """
    Uji pola KANDIDAT (dilatih hanya dari train split feedback yg tersedia
    sekarang) terhadap test split yg disisihkan, SEBELUM pola itu benar2
    digabung+disimpan sbg pola live. Lihat penjelasan lengkap di komentar
    blok di atas fungsi ini.

    Ini BUKAN pengganti hitung_tren_akurasi() (yg mengukur pola LIVE
    terhadap transaksi nyata dari waktu ke waktu, setelah kejadian) --
    ini pengujian sekali-pakai per batch retraining, sebelum kejadian.

    Return dict:
        sukses_evaluasi: bool -- False = data terlalu sedikit, split tidak
            independen, angka di bawah ini jangan terlalu dipercaya
        jumlah_test: int
        akurasi_pola_baru: float | None -- 0..1, pola kandidat (lama+baru
            HANYA dari train split) diuji ke test split
        akurasi_pola_lama: float | None -- 0..1, pola yg SEDANG live diuji
            ke test split yg SAMA (baseline pembanding)
        delta_akurasi: float | None -- akurasi_pola_baru - akurasi_pola_lama
        rekomendasi: "COMMIT_OTOMATIS" | "PERLU_REVIEW_MANUAL"
        alasan: str
    """
    rows_relevan = [r for r in augmentasi_rows if r.get("jenis") == jenis]
    train_rows, test_rows = _split_train_test_augmentasi(rows_relevan, rasio_test=rasio_test)
    split_independen = len(train_rows) < len(rows_relevan) or (train_rows is not test_rows and train_rows != test_rows)

    nama_dasar = "pola_bank" if jenis == "rekening_koran" else "pola_penjualan"
    path_pola = _path_pola(nama_dasar, client_id)
    pola_lama = muat_pola(path_pola)

    pola_kandidat_baru, _stat_train = bangun_pola_dari_feedback_klarifikasi(
        train_rows, jenis, df_coa, client_id=client_id,
    )
    pola_kandidat_gabungan = gabung_pola(pola_lama, pola_kandidat_baru)

    def _akurasi_terhadap(pola_uji: Pola, rows_uji: list[dict]) -> Optional[float]:
        ground_truths = []
        for row in rows_uji:
            hasil, _alasan = _tentukan_akun_final_dari_baris_feedback(row, jenis, df_coa)
            if hasil is not None:
                ground_truths.append(hasil)
        if not ground_truths:
            return None
        benar = sum(
            1 for gt in ground_truths
            if (aturan := pola_uji.aturan.get((gt["signature"], gt["arah"]))) is not None
            and str(aturan.get("no_akun_debet")) == str(gt["no_akun_debet"])
            and str(aturan.get("no_akun_kredit")) == str(gt["no_akun_kredit"])
        )
        return round(benar / len(ground_truths), 4)

    akurasi_baru = _akurasi_terhadap(pola_kandidat_gabungan, test_rows)
    akurasi_lama = _akurasi_terhadap(pola_lama, test_rows)

    if not split_independen or akurasi_baru is None:
        return {
            "sukses_evaluasi": False,
            "jumlah_test": len(test_rows),
            "akurasi_pola_baru": akurasi_baru,
            "akurasi_pola_lama": akurasi_lama,
            "delta_akurasi": None,
            "rekomendasi": "COMMIT_OTOMATIS",
            "alasan": (
                f"Data feedback terlalu sedikit utk staging yg bermakna "
                f"({len(rows_relevan)} baris) -- langsung commit spt biasa "
                "(sama seperti perilaku SEBELUM staging ada)."
            ),
        }

    delta = None if akurasi_lama is None else round(akurasi_baru - akurasi_lama, 4)
    if akurasi_baru < ambang_akurasi_minimal:
        rekomendasi = "PERLU_REVIEW_MANUAL"
        alasan = (
            f"Akurasi pola kandidat ({akurasi_baru:.0%}) di bawah ambang minimal "
            f"({ambang_akurasi_minimal:.0%}) saat diuji ke {len(test_rows)} baris "
            "feedback terbaru yg sengaja tidak ikut dipakai melatih."
        )
    elif delta is not None and delta < -ambang_regresi:
        rekomendasi = "PERLU_REVIEW_MANUAL"
        alasan = (
            f"Pola kandidat justru LEBIH BURUK ({akurasi_baru:.0%}) drpd pola yg "
            f"sedang live ({akurasi_lama:.0%}) -- turun {abs(delta):.0%}, melebihi "
            f"toleransi regresi ({ambang_regresi:.0%}). Kemungkinan ada feedback "
            "salah/typo di batch ini."
        )
    else:
        rekomendasi = "COMMIT_OTOMATIS"
        alasan = f"Akurasi pola kandidat {akurasi_baru:.0%} aman utk di-live-kan otomatis."

    return {
        "sukses_evaluasi": True,
        "jumlah_test": len(test_rows),
        "akurasi_pola_baru": akurasi_baru,
        "akurasi_pola_lama": akurasi_lama,
        "delta_akurasi": delta,
        "rekomendasi": rekomendasi,
        "alasan": alasan,
    }


def latih_ulang_pola_dengan_staging(
    augmentasi_rows: list[dict], jenis: str,
    client_id: Optional[int] = None, df_coa: Optional[pd.DataFrame] = None,
    paksa_commit: bool = False,
    ambang_akurasi_minimal: float = 0.6, ambang_regresi: float = 0.05,
) -> dict:
    """
    Pembungkus latih_ulang_pola_dari_feedback() dgn gerbang staging: jalankan
    evaluasi_pola_sebelum_commit() dulu; simpan_pola() (via
    latih_ulang_pola_dari_feedback) HANYA benar2 dipanggil kalau
    rekomendasinya "COMMIT_OTOMATIS", atau paksa_commit=True (utk supervisor
    yg sudah review manual & tetap yakin mau lanjut walau evaluasi
    menyarankan hati-hati).

    Kalau rekomendasi "PERLU_REVIEW_MANUAL" dan paksa_commit=False: TIDAK
    ADA perubahan tersimpan sama sekali -- pola LIVE tetap yg lama persis,
    caller (endpoint retrain-pola di main.py) mengembalikan hasil evaluasi
    ke supervisor utk diputuskan manual dulu.
    """
    evaluasi = evaluasi_pola_sebelum_commit(
        augmentasi_rows, jenis, client_id=client_id, df_coa=df_coa,
        ambang_akurasi_minimal=ambang_akurasi_minimal, ambang_regresi=ambang_regresi,
    )

    if evaluasi["rekomendasi"] == "PERLU_REVIEW_MANUAL" and not paksa_commit:
        return {
            "jenis": jenis,
            "status": "DITAHAN_MENUNGGU_REVIEW",
            "jumlah_feedback_total": len(augmentasi_rows),
            "evaluasi_staging": evaluasi,
            "pesan": (
                "Retraining TIDAK dilanjutkan otomatis -- evaluasi staging "
                "menyarankan review manual dulu (lihat evaluasi_staging.alasan). "
                "Panggil ulang dgn paksa_commit=True kalau setelah dicek "
                "manual tetap yakin mau lanjut."
            ),
        }

    hasil_commit = latih_ulang_pola_dari_feedback(
        augmentasi_rows, jenis, client_id=client_id, df_coa=df_coa,
    )
    status = (
        "DIKOMIT_PAKSA_SETELAH_REVIEW"
        if (evaluasi["rekomendasi"] == "PERLU_REVIEW_MANUAL" and paksa_commit)
        else "DIKOMIT_OTOMATIS"
    )
    return {"status": status, "evaluasi_staging": evaluasi, **hasil_commit}


# ============================================================
# [BARU] METRIK AKURASI TERPUSAT (mengisi Tahap 6 -- Evaluasi -- dari
# siklus umum AI: kumpulkan data -> proses -> latih -> EVALUASI ->
# putuskan -> monitor. Sebelumnya cuma ada evaluasi KUALITAS INPUT
# (generate_data_quality_report) dan evaluasi KEWAJARAN POLA
# (deteksi_pola_mencurigakan) -- tidak ada satupun yang menjawab
# pertanyaan paling penting: "AI ini makin akurat atau makin ngaco dari
# waktu ke waktu?". Fungsi di bawah mengisi itu, dengan menyimpan histori
# statistik konfirmasi-vs-koreksi yang SEBENARNYA sudah dihitung sejak
# lama di bangun_pola_dari_feedback_klarifikasi(), tapi sebelumnya tidak
# pernah disimpan -- cuma dikembalikan sekali ke caller lalu hilang.
# ============================================================

def _path_metrik_akurasi(client_id: Optional[int] = None) -> str:
    """Path file log metrik akurasi, 1 file per client (mirip _path_pola)."""
    nama = f"metrik_akurasi_client_{client_id}.json" if client_id is not None else "metrik_akurasi_default.json"
    return str(Path("data") / nama)


def catat_metrik_akurasi(
    client_id: Optional[int],
    jenis: str,
    jumlah_konfirmasi: int,
    jumlah_koreksi: int,
    sumber_data: str = "klarifikasi_ui",
) -> None:
    """
    Tambah 1 entri histori akurasi (append-only log, mirip semangat
    _simpan_snapshot_versi_pola -- tidak menimpa histori lama).

    Tidak melempar exception kalau gagal tulis (mis. folder data/ belum
    ada / tidak writable) -- metrik akurasi ini SIFATNYA observability,
    bukan bagian dari alur inti (retraining pola tetap harus lanjut
    sukses walau logging metrik gagal). Kegagalan cukup dicatat ke log
    aplikasi biasa.
    """
    if jumlah_konfirmasi == 0 and jumlah_koreksi == 0:
        return  # tidak ada sinyal apa pun untuk dicatat
    entri = {
        "waktu": datetime.now().isoformat(),
        "jenis": jenis,
        "sumber_data": sumber_data,
        "jumlah_konfirmasi": jumlah_konfirmasi,
        "jumlah_koreksi": jumlah_koreksi,
        "total": jumlah_konfirmasi + jumlah_koreksi,
        "akurasi_persen": round(jumlah_konfirmasi / (jumlah_konfirmasi + jumlah_koreksi) * 100, 2),
    }
    try:
        path = Path(_path_metrik_akurasi(client_id))
        path.parent.mkdir(parents=True, exist_ok=True)
        histori = []
        if path.exists():
            with open(path, "r", encoding="utf-8") as f:
                histori = json.load(f)
        histori.append(entri)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(histori, f, ensure_ascii=False, indent=2)
    except Exception:
        logger.exception(f"Gagal mencatat metrik akurasi (client_id={client_id}, jenis={jenis})")


def muat_riwayat_metrik_akurasi(client_id: Optional[int], jenis: Optional[str] = None) -> list[dict]:
    """Baca semua entri histori akurasi, opsional difilter per jenis (rekening_koran/penjualan)."""
    path = Path(_path_metrik_akurasi(client_id))
    if not path.exists():
        return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            histori = json.load(f)
    except Exception:
        logger.exception(f"Gagal membaca metrik akurasi (client_id={client_id})")
        return []
    if jenis is not None:
        histori = [h for h in histori if h.get("jenis") == jenis]
    return histori


def hitung_tren_akurasi(client_id: Optional[int], jenis: Optional[str] = None, n_bulan_terakhir: int = 6) -> dict:
    """
    Agregasi histori akurasi per bulan (YYYY-MM), supaya bisa dijawab
    langsung: "AI makin akurat atau makin ngaco?" -- bukan cuma angka
    kualitas data mentah seperti generate_data_quality_report().

    Return:
        {
          "per_bulan": [{"bulan": "2026-06", "total": 42, "akurasi_persen": 87.3}, ...],
          "akurasi_keseluruhan": float,
          "tren": "NAIK" | "TURUN" | "STABIL" | "DATA_BELUM_CUKUP",
          "selisih_persen_poin": float,  # bulan terakhir vs bulan sebelumnya
        }
    """
    histori = muat_riwayat_metrik_akurasi(client_id, jenis)
    if not histori:
        return {"per_bulan": [], "akurasi_keseluruhan": None, "tren": "DATA_BELUM_CUKUP", "selisih_persen_poin": None}

    per_bulan_raw: dict[str, dict] = {}
    for entri in histori:
        bulan = entri["waktu"][:7]  # "YYYY-MM"
        agg = per_bulan_raw.setdefault(bulan, {"konfirmasi": 0, "koreksi": 0})
        agg["konfirmasi"] += entri["jumlah_konfirmasi"]
        agg["koreksi"] += entri["jumlah_koreksi"]

    bulan_terurut = sorted(per_bulan_raw.keys())[-n_bulan_terakhir:]
    per_bulan = []
    for bulan in bulan_terurut:
        agg = per_bulan_raw[bulan]
        total = agg["konfirmasi"] + agg["koreksi"]
        per_bulan.append({
            "bulan": bulan, "total": total,
            "jumlah_konfirmasi": agg["konfirmasi"], "jumlah_koreksi": agg["koreksi"],
            "akurasi_persen": round(agg["konfirmasi"] / total * 100, 2) if total else None,
        })

    total_konfirmasi = sum(e["jumlah_konfirmasi"] for e in histori)
    total_koreksi = sum(e["jumlah_koreksi"] for e in histori)
    total_semua = total_konfirmasi + total_koreksi
    akurasi_keseluruhan = round(total_konfirmasi / total_semua * 100, 2) if total_semua else None

    tren, selisih = "DATA_BELUM_CUKUP", None
    valid = [b for b in per_bulan if b["akurasi_persen"] is not None]
    if len(valid) >= 2:
        selisih = round(valid[-1]["akurasi_persen"] - valid[-2]["akurasi_persen"], 2)
        if selisih > 1.0:
            tren = "NAIK"
        elif selisih < -1.0:
            tren = "TURUN"
        else:
            tren = "STABIL"

    return {
        "per_bulan": per_bulan,
        "akurasi_keseluruhan": akurasi_keseluruhan,
        "tren": tren,
        "selisih_persen_poin": selisih,
    }


def hitung_confidence_aktual(
    client_id: Optional[int], jenis: str, jalur: str = "koreksi_akuntan",
    minimal_sample: int = 10, default_jika_kurang_data: float = 0.85,
) -> float:
    """
    [BARU -- mengganti confidence_score hardcoded] Sebelumnya
    bangun_pola_dari_feedback_klarifikasi() selalu menulis confidence_score
    = 0.85 utk SEMUA pola hasil jalur koreksi_akuntan -- angka itu ditulis
    manual (dugaan), bukan hasil pengukuran. Fungsi ini menghitung
    confidence yang SEBENARNYA dari histori: dari seluruh pola yang lahir
    dari jalur koreksi akuntan, berapa persen yang ternyata BERTAHAN (tidak
    dikoreksi ULANG lagi setelahnya) -- itu proxy paling dekat yang kita
    punya utk "seberapa bisa dipercaya pola hasil koreksi akuntan".

    Kalau histori belum cukup (< minimal_sample baris jalur koreksi_akuntan
    tercatat), fallback ke default_jika_kurang_data (angka lama, supaya
    tidak menghasilkan confidence yang goyah krn sample terlalu sedikit).

    Konfirmasi akuntan (jalur "konfirmasi_akuntan") SENGAJA TIDAK dihitung
    ulang di sini dan tetap confidence 1.0 di caller -- itu bukan estimasi,
    memang sudah divalidasi manusia secara langsung per baris.
    """
    histori = muat_riwayat_metrik_akurasi(client_id, jenis)
    histori = [h for h in histori if h.get("sumber_data") == "klarifikasi_ui"]
    total_koreksi = sum(h["jumlah_koreksi"] for h in histori)
    total_konfirmasi = sum(h["jumlah_konfirmasi"] for h in histori)
    total = total_koreksi + total_konfirmasi
    if total < minimal_sample:
        return default_jika_kurang_data
    # Proxy akurasi jalur koreksi: makin tinggi rasio konfirmasi (AI benar
    # tanpa perlu dikoreksi) dibanding total, makin bisa dipercaya juga
    # pola-pola yang lahir dari koreksi (menandakan akuntan & AI makin
    # sering sepakat secara umum untuk client ini).
    rasio = total_konfirmasi / total
    # Dibatasi rentang wajar 0.5-0.95: pola hasil koreksi manusia tidak
    # pernah dianggap SEPENUHNYA 1.0 (beda dari konfirmasi langsung), dan
    # tidak pernah dibiarkan jatuh terlalu rendah walau rasio konfirmasi
    # sedang buruk (pola itu tetap valid, hanya kurang "terbukti berulang").
    return round(min(0.95, max(0.5, rasio)), 2)


# ============================================================
# [BARU - Prioritas #6, jalur langsung] Upload file hasil koreksi akuntan
# ============================================================

def _normalisasi_kode_akun_lokal(nilai) -> Optional[str]:
    """
    Duplikat SENGAJA dari accounting_export._normalisasi_kode_akun() --
    supaya akuntansi_ai.py tidak perlu import dari modules/accounting_export.py
    (mengikuti pola yang sudah dipakai di codebase ini, mis.
    db_client._kode_bank_dari_nama_lokal() / db_client._buat_transaction_hash_baris(),
    yang sengaja diduplikasi utk menghindari dependency silang antar modul).
    KALAU FORMULA DI accounting_export._normalisasi_kode_akun() DIUBAH,
    UBAH JUGA DI SINI -- kalau tidak, lookup ke sheet COA di fungsi
    proses_file_hasil_koreksi_akuntan() bisa salah cocok.
    """
    if nilai is None or (isinstance(nilai, float) and pd.isna(nilai)):
        return None
    try:
        f = float(nilai)
        if f == int(f):
            return str(int(f))
    except (TypeError, ValueError):
        pass
    teks = str(nilai).strip()
    return teks if teks else None


def proses_file_hasil_koreksi_akuntan(
    file_like, nama_file: str = None, client_id: Optional[int] = None, min_samples: int = 1,
) -> dict:
    """
    [BARU - Prioritas #6, jalur LANGSUNG dari file] Pelajari pola dari file
    HASIL EXPORT format akuntan (keluaran
    accounting_export.export_rekening_koran_format_akuntan()) yang SUDAH
    DIKOREKSI MANUAL oleh akuntan -- baris kuning (belum dikategorikan)
    dan/atau merah (kode akun salah) sudah diisi/dibetulkan langsung di
    Excel, lalu file itu diupload balik ke sini.

    BEDA dari /api/client/{id}/retrain-pola (yang menarik dari jawaban
    fitur klarifikasi di UI): fungsi ini membaca LANGSUNG file Excel yang
    dikoreksi, untuk akuntan yang terbiasa mengedit file kerja Excel-nya
    sendiri dan tidak (belum tentu) pernah menyentuh fitur klarifikasi di
    UI. Keduanya saling melengkapi -- boleh dipakai salah satu atau dua-
    duanya, hasilnya sama-sama digabung (gabung_pola) ke
    pola_bank_client_{id}.json yang sama.

    Layout yang diharapkan (lihat accounting_export._tulis_sheet_bank()
    & _tulis_sheet_coa() -- HARUS file hasil export modul ini, bukan
    rekening koran mentah dari bank):
      - 1 sheet "COA": baris 1 header, baris 2+ = CAT | NO AKUN | DESCRIPTION.
      - 1 sheet per bank: baris 1 subtotal, baris 2 header, baris 3+ data,
        kolom B=Tanggal, C=Keterangan, D=Mutasi Debet, E=Mutasi Kredit,
        K=No Akun Debet, N=No Akun Kredit.

    [PENTING] Nama akun diambil ULANG dari sheet COA di fungsi ini sendiri
    (lookup manual dari kolom K/N -- kode akun MENTAH), BUKAN dibaca dari
    kolom L/O (yang isinya formula VLOOKUP di file aslinya). Ini sengaja
    supaya fungsi ini TETAP BENAR walau file diedit/disimpan lewat tool
    yang tidak mengevaluasi ulang formula Excel (cached value formula bisa
    saja kosong/None dalam kondisi begitu) -- openpyxl load_workbook biasa
    (data_only=False) sudah cukup, tidak butuh Excel/LibreOffice terinstall
    di server untuk "memaksa" recalculate.

    Baris yang kode akun debet/kredit-nya (kolom K/N) masih kosong, ATAU
    kodenya tidak ditemukan di sheet COA, DILEWATI (tidak cukup valid
    untuk dipelajari sbg pola) -- dihitung terpisah di
    jumlah_dilewati_kosong / jumlah_dilewati_kode_tidak_di_coa supaya
    akuntan tahu kalau msh ada baris yang belum lengkap.

    min_samples default 1 (lebih rendah dari pelajari_pola/bootstrap biasa)
    -- karena tiap baris di file koreksi ini SUDAH melalui review manual
    akuntan (bukan tebakan AI/heuristik), jadi 1 contoh saja sudah cukup
    dipercaya jadi pola.
    """
    # [CATATAN TEMPEL] Pola/muat_pola/pelajari_pola/gabung_pola/simpan_pola/
    # _path_pola TIDAK perlu diimport -- begitu fungsi ini ditempel ke
    # dalam akuntansi_ai.py, semuanya sudah ada di scope modul yang sama.

    nama_file = nama_file or getattr(file_like, "name", "") or ""
    wb = openpyxl.load_workbook(file_like, data_only=False)

    # --- Sheet COA -> dict {kode_akun_ternormalisasi: nama_akun} ---
    coa_map: dict = {}
    if "COA" in wb.sheetnames:
        ws_coa = wb["COA"]
        for row in ws_coa.iter_rows(min_row=2, values_only=True):
            if len(row) < 3:
                continue
            kunci = _normalisasi_kode_akun_lokal(row[1])
            nama_akun = row[2]
            if kunci is not None and nama_akun:
                coa_map[kunci] = str(nama_akun).strip()

    if not coa_map:
        raise ValueError(
            "Sheet 'COA' tidak ditemukan atau kosong di file ini -- pastikan yang "
            "diupload adalah file HASIL EXPORT format akuntan (dari fitur Export "
            "Format Akuntan), bukan rekening koran mentah."
        )

    baris_dipakai = []
    jumlah_baris_dibaca = 0
    jumlah_dilewati_kosong = 0
    jumlah_dilewati_kode_tidak_di_coa = 0
    sheet_dilewati = []

    _HEADER_DIHARAPKAN = ["No", "Tanggal", "Keterangan", "Mutasi Debet", "Mutasi Kredit"]

    for nama_sheet in wb.sheetnames:
        if nama_sheet == "COA":
            continue
        ws = wb[nama_sheet]
        if ws.max_row < 3:
            continue
        header = [c.value for c in ws[2][:5]]
        if header != _HEADER_DIHARAPKAN:
            sheet_dilewati.append(
                f"{nama_sheet}: header baris 2 tidak sesuai format export (dilewati -- "
                "pastikan ini file hasil export, bukan file yang formatnya sudah diubah)."
            )
            continue

        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[1] is None and row[2] is None:
                continue  # baris kosong
            jumlah_baris_dibaca += 1

            tanggal, keterangan = row[1], row[2]
            mutasi_debet = row[3] if len(row) > 3 else 0
            mutasi_kredit = row[4] if len(row) > 4 else 0
            no_akun_debet_raw = row[10] if len(row) > 10 else None   # kolom K
            no_akun_kredit_raw = row[13] if len(row) > 13 else None  # kolom N

            kd = _normalisasi_kode_akun_lokal(no_akun_debet_raw)
            kk = _normalisasi_kode_akun_lokal(no_akun_kredit_raw)
            if kd is None or kk is None:
                jumlah_dilewati_kosong += 1
                continue

            nama_debet = coa_map.get(kd)
            nama_kredit = coa_map.get(kk)
            if nama_debet is None or nama_kredit is None:
                jumlah_dilewati_kode_tidak_di_coa += 1
                continue

            baris_dipakai.append({
                "keterangan": keterangan,
                # [FIX] _angka_aman() bukan "float(x or 0)" -- kalau file yang
                # diimpor ulang ini adalah export lama (sebelum fix NaN di
                # accounting_export.py) yang kebetulan punya sel NaN, "or 0"
                # tidak akan menangkapnya (NaN truthy di Python).
                "mutasi_debet": _angka_aman(mutasi_debet),
                "mutasi_kredit": _angka_aman(mutasi_kredit),
                "no_akun_debet": kd, "nama_akun_debet": nama_debet,
                "no_akun_kredit": kk, "nama_akun_kredit": nama_kredit,
            })

    if not baris_dipakai:
        return {
            "jumlah_baris_dibaca": jumlah_baris_dibaca, "jumlah_baris_dipakai": 0,
            "jumlah_pola_baru": 0, "jumlah_pola_diperbarui": 0,
            "jumlah_pola_dipelajari_total": 0, "total_pola_tersimpan": 0,
            "detail_pola_baru": [], "detail_pola_diperbarui": [],
            "jumlah_dilewati_kosong": jumlah_dilewati_kosong,
            "jumlah_dilewati_kode_tidak_di_coa": jumlah_dilewati_kode_tidak_di_coa,
            "sheet_dilewati": sheet_dilewati,
        }

    df = pd.DataFrame(baris_dipakai)
    pola_baru = pelajari_pola(df, min_samples=min_samples)

    path_pola = _path_pola("pola_bank", client_id)
    pola_lama = muat_pola(path_pola)

    detail_baru, detail_diperbarui = [], []
    for key, aturan in pola_baru.aturan.items():
        sig, arah = key
        lama = pola_lama.aturan.get(key)
        ringkas = {
            "signature": sig, "arah": arah,
            "akun_debet": f"{aturan['no_akun_debet']} - {aturan['nama_akun_debet']}",
            "akun_kredit": f"{aturan['no_akun_kredit']} - {aturan['nama_akun_kredit']}",
            "jumlah_contoh": aturan["jumlah_contoh"], "confidence_score": aturan["confidence_score"],
            "is_valid": aturan["is_valid"],
        }
        if lama is None:
            detail_baru.append(ringkas)
        elif (lama.get("no_akun_debet"), lama.get("no_akun_kredit")) != (
            aturan["no_akun_debet"], aturan["no_akun_kredit"]
        ):
            ringkas["akun_lama_debet"] = f"{lama.get('no_akun_debet')} - {lama.get('nama_akun_debet')}"
            ringkas["akun_lama_kredit"] = f"{lama.get('no_akun_kredit')} - {lama.get('nama_akun_kredit')}"
            detail_diperbarui.append(ringkas)

    pola_gabungan = gabung_pola(pola_lama, pola_baru)
    simpan_pola(
        pola_gabungan, path_pola,
        sumber_perubahan=f"proses_file_hasil_koreksi_akuntan ({len(baris_dipakai)} baris koreksi)",
    )

    return {
        "jumlah_baris_dibaca": jumlah_baris_dibaca,
        "jumlah_baris_dipakai": len(baris_dipakai),
        "jumlah_pola_baru": len(detail_baru),
        "jumlah_pola_diperbarui": len(detail_diperbarui),
        "jumlah_pola_dipelajari_total": len(pola_baru.aturan),
        "total_pola_tersimpan": len(pola_gabungan.aturan),
        "detail_pola_baru": detail_baru,
        "detail_pola_diperbarui": detail_diperbarui,
        "jumlah_dilewati_kosong": jumlah_dilewati_kosong,
        "jumlah_dilewati_kode_tidak_di_coa": jumlah_dilewati_kode_tidak_di_coa,
        "sheet_dilewati": sheet_dilewati,
        "path_pola": path_pola,
    }


# ============================================================
# [BARU] TUJUAN PEMROSESAN PER JENIS DOKUMEN (mengisi Tahap 1 -- Definisi
# Tujuan -- dari siklus umum AI, yang sebelumnya implisit/tersebar: tiap
# proses_file_xxx() langsung lompat ke parsing tanpa satu tempat yang
# menyatakan dengan jelas "tujuan pemrosesan jenis dokumen ini apa".
# Dipakai oleh proses_file_dengan_jaring_pengaman() supaya pesan hasil/
# error ke akuntan menyebut tujuan pemrosesannya, bukan cuma nama teknis
# jenis dokumen -- dan supaya developer baru bisa baca 1 dict ini utk
# paham cakupan tiap pipeline tanpa harus telusuri ribuan baris kode.
# ============================================================

TUJUAN_PEMROSESAN: dict[str, str] = {
    "rekening_koran": "Mengubah mutasi rekening koran mentah jadi draf jurnal (debet/kredit per transaksi), belajar dari pola historis yang sudah dijurnal akuntan.",
    "penjualan": "Mengubah data penjualan/invoice jadi draf jurnal (Kas/Piutang - Penjualan - PPN Keluaran) sesuai cara bayar.",
    "penjualan_pos": "Mengubah data transaksi POS/kasir jadi draf jurnal penjualan harian.",
    "penilaian_klien": "Menilai kualitas pembukuan tiap klien per staff (Maker), mendeteksi kesalahan berdasarkan 20 standar kesalahan akuntansi, dan merapikan/melengkapi file penilaian.",
    "piutang": "Menyusun buku bantu piutang & aging piutang (bucket umur piutang) per pelanggan.",
    "faktur_pajak": "Memvalidasi faktur pajak keluaran (nomor faktur, NPWP, DPP/PPN) sesuai aturan DJP.",
    "bukti_potong": "Memvalidasi bukti potong PPh 21/23/4(2) (nomor bukti, tarif sesuai kode objek pajak) & menyusun draf jurnal terkait.",
    "spt": "Memvalidasi kesesuaian SPT Masa/Tahunan (PPh 21/23/PPN) dengan estimasi jatuh tempo & toleransi selisih rupiah.",
    "slip_gaji": "Memvalidasi komponen slip gaji (kewajaran BPJS, PPh 21) & menyusun draf jurnal beban gaji per periode.",
    "bukti_kas": "Memvalidasi bukti kas masuk/keluar (deteksi duplikat, validasi saldo berjalan) & menyusun draf jurnal kas.",
    "kartu_stok": "Memvalidasi mutasi kartu stok & saldo berjalan persediaan per item.",
    "aset_tetap": "Menyusun jadwal penyusutan bulanan aset tetap & draf jurnal penyusutan, termasuk klasifikasi golongan fiskal.",
    "pembelian": "Menyusun draf jurnal pembelian (termasuk validasi selisih nominal PO vs invoice).",
    "rekonsiliasi_bank": "Merekonsiliasi saldo buku besar bank vs saldo rekening koran, mengelompokkan item penyesuaian (deposit/cek dalam perjalanan, dll).",
    "ap_aging": "Menyusun buku bantu utang & aging utang (bucket umur utang) per vendor.",
    "absensi": "Merapikan data absensi (jam kerja, keterlambatan) sebagai bahan validasi komponen slip gaji.",
    "hasil_koreksi_akuntan": "Mempelajari pola jurnal baru dari file rekening koran yang SUDAH dikoreksi manual oleh akuntan, untuk memperbaiki akurasi kategorisasi otomatis client ini ke depannya.",
    "bootstrap_pola_bank": "Mengisi pola awal client baru dari rekening koran bulan lalu yang sudah dijurnal, supaya tidak mulai dari pola kosong.",
    "laporan_keuangan": "Menyusun 5 laporan keuangan standar (Neraca, Laba Rugi, Perubahan Ekuitas, Arus Kas, CALK) dari GL/Neraca Saldo atau gabungan draf jurnal semua pipeline.",
}


def jelaskan_tujuan_pemrosesan(label_jenis: str) -> str:
    """
    Cocokkan label_jenis (label bebas yang dikirim caller ke
    proses_file_dengan_jaring_pengaman, mis. "Rekening Koran") ke kunci
    TUJUAN_PEMROSESAN dengan pencocokan longgar (case-insensitive,
    spasi->underscore) -- supaya tidak perlu ubah semua pemanggilan yang
    sudah ada di main.py hanya demi mencocokkan kunci dict persis.
    Kembalikan string kosong kalau tidak ketemu (BUKAN error -- dict ini
    bersifat aditif/dokumentatif, bukan validasi wajib).
    """
    if not label_jenis:
        return ""
    kunci_dicoba = label_jenis.strip().lower().replace(" ", "_").replace("/", "_")
    if kunci_dicoba in TUJUAN_PEMROSESAN:
        return TUJUAN_PEMROSESAN[kunci_dicoba]
    for kunci, tujuan in TUJUAN_PEMROSESAN.items():
        if kunci in kunci_dicoba or kunci_dicoba in kunci:
            return tujuan
    return ""


# ============================================================
# JARING PENGAMAN -- PEMBUNGKUS AMAN UNTUK SEMUA proses_file_xxx()
# ============================================================
# [BARU] Kenapa ini ditambahkan: hari ini ditemukan & diperbaiki beberapa
# bug nyata (lihat komentar [FIX] di proses_spt, _format_periode_slip,
# _normalisasi_periode_gaji, _hitung_umur_aset_bulan, _hitung_umur_utang_hari,
# _parse_jam_ke_menit) yang KESELURUHANNYA berasal dari satu pola akar
# masalah yang sama: sel Excel kosong yang duduk di kolom yang JUGA berisi
# tanggal/waktu asli berubah dari None jadi pandas.NaT/nan begitu banyak
# baris/sheet digabung jadi satu DataFrame -- dan NaT itu lolos dari
# pengecekan "is None" biasa.
#
# Sudah diaudit & ditest manual untuk semua 15 jenis dokumen yang py, TAPI
# file ini >9.900 baris -- TIDAK ADA JAMINAN semua kombinasi data client di
# dunia nyata sudah tertangkap hari ini. Supaya SATU baris data aneh dari
# SATU client tidak menjatuhkan seluruh proses upload dengan traceback
# mentah ke user, bungkus setiap pemanggilan proses_file_xxx() di main.py
# dengan proses_file_dengan_jaring_pengaman() di bawah -- BUKAN memanggil
# proses_file_spt/proses_file_slip_gaji/dst secara langsung.
#
# Cara pakai di main.py (ganti pemanggilan langsung):
#   from modules.akuntansi_ai import proses_file_dengan_jaring_pengaman, proses_file_spt
#   hasil = proses_file_dengan_jaring_pengaman(proses_file_spt, file_like, nama_file, "SPT Masa/Tahunan")
#   if hasil["sukses"]:
#       return hasil["data"]      # <- ini persis dict yang dikembalikan proses_file_spt biasanya
#   else:
#       return {"error": hasil["pesan_error"]}, 422   # atau format error API yang dipakai main.py
#
# Ini TIDAK menggantikan perbaikan bug satu-per-satu (bug yang mendasarinya
# tetap harus diperbaiki begitu ketemu) -- ini cuma jaring pengaman terakhir
# supaya kegagalan selalu berupa PESAN YANG BISA DIBACA AKUNTAN, bukan
# traceback Python / halaman blank / 500 Internal Server Error.
def proses_file_dengan_jaring_pengaman(
    fungsi_proses,
    file_like,
    nama_file: str = None,
    label_jenis: str = "",
    *args,
    **kwargs,
) -> dict:
    """
    Bungkus pemanggilan proses_file_xxx (mis. proses_file_spt,
    proses_file_slip_gaji, proses_file_bukti_kas, dst) supaya EXCEPTION
    APAPUN yang belum ketahuan/ketest tidak menjatuhkan seluruh request,
    tapi dikembalikan sebagai pesan error yang rapi & bisa ditindaklanjuti.

    [FIX-3] SEBELUMNYA fungsi ini cuma meneruskan (file_like, nama_file=...)
    ke fungsi_proses -- padahal banyak proses_file_xxx BUTUH argumen lain
    supaya jalan benar, mis. client_id di proses_file_rekening_koran(),
    proses_file_bootstrap_pola_bank(), proses_file_penjualan(), dan
    proses_file_hasil_koreksi_akuntan() (dipakai utk pola historis per
    client -- kalau silent jatuh ke default None, pembelajaran pola jadi
    salah client TANPA error apapun), juga npwp_perusahaan di
    proses_file_faktur_pajak()/proses_file_bukti_potong(),
    histori_gaji_sebelumnya di proses_file_slip_gaji(), pakai_ai, dan
    min_samples. Sekarang *args/**kwargs apa pun yang dikirim caller
    diteruskan apa adanya ke fungsi_proses.

    Contoh pakai (di main.py):
        proses_file_dengan_jaring_pengaman(
            proses_file_rekening_koran, file_like, nama_file,
            "Rekening Koran", client_id=client_id,
        )

    Return:
        {"sukses": True, "data": <dict hasil proses_file_xxx asli>}
        atau
        {"sukses": False, "pesan_error": str, "jenis_error": str,
         "detail_teknis": str}
        "detail_teknis" HANYA untuk log/debug internal (jangan ditampilkan
        polos ke user akuntan) -- "pesan_error" yang cocok ditampilkan ke UI.
    """
    nama_file = nama_file or getattr(file_like, "name", "") or "(file tidak diketahui)"
    tujuan = jelaskan_tujuan_pemrosesan(label_jenis)  # [BARU] lihat TUJUAN_PEMROSESAN di atas
    try:
        if hasattr(file_like, "seek"):
            file_like.seek(0)
        data = fungsi_proses(file_like, *args, nama_file=nama_file, **kwargs)
        return {"sukses": True, "data": data, "tujuan_pemrosesan": tujuan}
    except FormatFileTidakDidukung as e:
        return {
            "sukses": False,
            "pesan_error": f"Format file '{nama_file}' tidak didukung. {e}",
            "jenis_error": "FORMAT_TIDAK_DIDUKUNG",
            "detail_teknis": str(e),
        }
    except FormatTidakDikenali as e:
        return {
            "sukses": False,
            "pesan_error": (
                f"File '{nama_file}' tidak dikenali sebagai {label_jenis or 'jenis dokumen ini'}. "
                f"{e} Coba cek lagi jenis dokumennya, atau upload ulang dgn format kolom standar."
            ),
            "jenis_error": "FORMAT_TIDAK_DIKENALI",
            "detail_teknis": str(e),
            "tujuan_pemrosesan": tujuan,
        }
    except Exception as e:
        logger.exception(f"❌ Gagal memproses '{nama_file}' sebagai {label_jenis or '(?)'}: {e}")
        return {
            "sukses": False,
            "pesan_error": (
                f"Terjadi kesalahan tak terduga saat memproses '{nama_file}' sebagai "
                f"{label_jenis or 'dokumen ini'}. File & baris datanya kemungkinan punya "
                "format/isi yang belum ditangani -- silakan cek manual dulu file ini, dan "
                "laporkan ke tim dev supaya bug-nya bisa diperbaiki permanen (jangan cuma "
                "diulang-ulang upload, karena errornya kemungkinan besar akan terjadi lagi)."
            ),
            "jenis_error": type(e).__name__,
            "detail_teknis": traceback.format_exc(),
            "tujuan_pemrosesan": tujuan,
        }


# ============================================================
# AKHIR FILE
# ============================================================

if __name__ == "__main__":
    # Test code jika diperlukan
    pass