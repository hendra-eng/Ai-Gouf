#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
uji_akurasi_kategorisasi.py
=============================
Skrip evaluasi end-to-end (Prioritas #1) untuk mengukur:
  - Berapa % baris berhasil terkategori OTOMATIS, dipecah per SUMBER
    (pola historis / kata kunci COA / AI / data asli dari file) vs
    berapa % jatuh ke "Belum Terkategori".
  - Kalau dijalankan dengan --strip: berapa % dari yang "berhasil
    terkategori" itu SUNGGUH BENAR (dibanding jurnal asli yang sudah
    dikonfirmasi akuntan) -- bukan cuma "terisi tapi belum tentu benar".

Skrip ini TIDAK reimplementasi logika kategorisasi. Ia memanggil langsung
`akuntansi_ai.proses_file_rekening_koran()` -- fungsi yang sama persis yang
dipakai endpoint FastAPI produksi -- supaya angka yang dihasilkan benar-benar
mencerminkan pipeline yang sesungguhnya.

CARA PAKAI
----------
1) Kalau kamu SUDAH punya rekening koran MENTAH (belum ada jurnal sama
   sekali) sebagai file uji:

     python uji_akurasi_kategorisasi.py --file mentah_juli.xlsx \
         --client-id 123

2) Kalau kamu BELUM punya file mentah dan cuma punya contoh file yang
   SUDAH dijurnal lengkap (seperti file "SAU Rekening Koran ... JULY"
   yang kamu kirim) -- pakai --strip. Skrip akan:
     a. Membaca jurnal asli (VOUCHER/NO AKUN/NAMA AKUN/DEBET/KREDIT) dari
        file itu sebagai GROUND TRUTH.
     b. Membuat salinan file dengan kolom-kolom jurnal itu DIKOSONGKAN
        (mensimulasikan file mentah).
     c. Menjalankan proses_file_rekening_koran() pada salinan kosong itu.
     d. Membandingkan hasil prediksi vs ground truth per baris -> akurasi
        SUNGGUH BENAR, bukan cuma "ada isinya".

     python uji_akurasi_kategorisasi.py \
         --file "1__SAU_Rekening_Koran_All_Bank_202607__JULY_.xlsx" \
         --strip --client-id EVALTEST --pola-bersih

FLAG
----
  --file PATH         Path file .xlsx yang diuji (wajib)
  --client-id ID       ID client (dipakai untuk path pola_bank_client_{id}.json).
                        Pola historis milik client ini yang akan dipakai/ditambah.
  --strip              Buat "file mentah" otomatis dari file yang sudah dijurnal
                        (lihat penjelasan di atas), dan hitung akurasi vs ground
                        truth. Kalau tidak dipakai, skrip HANYA menghitung %
                        fill-rate per sumber (tidak ada ground truth untuk cek benar/salah).
  --tanpa-ai            Matikan tahap AI (uji baseline pola + kata kunci saja).
  --pola-bersih         Hapus dulu pola_bank_client_{id}.json sebelum uji (simulasi
                        client BENAR-BENAR baru / cold start). Tanpa flag ini,
                        pola historis yang sudah ada (kalau ada) akan tetap dipakai
                        DAN akan bertambah dari baris yang berhasil dikategorikan.
  --output-dir PATH    Folder untuk file Excel hasil (default: ./hasil_uji)

OUTPUT
------
  1. Ringkasan di terminal: breakdown per sumber kategori (+ akurasi kalau --strip).
  2. <output-dir>/hasil_<nama_file>.xlsx -- hasil lengkap dengan highlight:
       - KUNING: baris "Belum Terkategori" (perlu review manual)
       - MERAH:  baris yang salah dibanding ground truth (hanya muncul kalau --strip)
       - HIJAU muda: baris "perlu cek" (pola tidak 100% konsisten / AI confidence rendah)
"""

from __future__ import annotations

import argparse
import io
import shutil
import sys
from collections import Counter
from pathlib import Path

import akuntansi_ai as ak
import openpyxl
from openpyxl.styles import PatternFill

FILL_KUNING = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
FILL_MERAH = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
FILL_HIJAU = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")


# ============================================================
# 1. MEMBUAT "FILE MENTAH" DARI FILE YANG SUDAH DIJURNAL (--strip)
# ============================================================

def buat_file_mentah_dan_ground_truth(path_asli: Path, path_keluaran: Path):
    """
    Baca file asli (sudah dijurnal), simpan ground truth (no_akun_debet/kredit
    asli per baris, per sheet bank), lalu tulis salinan file dengan kolom
    jurnal (NO AKUN/NAMA AKUN/DEBET/KREDIT, dan VOUCHER) DIKOSONGKAN --
    mensimulasikan file mentah yang belum pernah disentuh akuntan.

    Deteksi kolom jurnal PERSIS sama seperti parse_sheet_bank() di
    akuntansi_ai.py (header row dicari otomatis, header "DEBET"/"KREDIT"
    exact match, 2 kolom sebelumnya = NO AKUN / NAMA AKUN).

    Return: dict {nama_sheet: list of {"no_akun_debet", "no_akun_kredit", ...}}
            berurutan sesuai baris data (index sejajar dengan hasil parse
            pipeline nanti, SELAMA baris di file tidak diacak).
    """
    wb = openpyxl.load_workbook(path_asli, data_only=True)
    ground_truth: dict[str, list[dict]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_rownum, header_row = ak._cari_header_row(ws)
        if header_row is None:
            continue  # bukan sheet rekening koran (mis. sheet COA) -> lewati

        saldo_idx = ak._cari_idx(header_row, ["saldo", "balance"])
        if saldo_idx is None:
            continue

        idx_debit_stmt = ak._cari_idx(header_row, ["debit", "debet"], sampai=saldo_idx)
        idx_kredit_stmt = ak._cari_idx(header_row, ["kredit", "credit"], sampai=saldo_idx)
        if idx_debit_stmt is None or idx_kredit_stmt is None:
            continue  # bukan sheet rekening koran

        idx_tanggal = ak._cari_idx(header_row, ["tgl", "tanggal", "date"], sampai=saldo_idx + 1)
        idx_keterangan = ak._cari_idx(header_row, ["keterangan", "remarks"], sampai=saldo_idx + 1)
        if idx_tanggal is None or idx_keterangan is None:
            continue  # sama seperti parse_sheet_bank(): tanpa ini baris tidak akan ke-parse

        idx_jurnal_debet = idx_jurnal_kredit = idx_voucher = None
        for i, h in enumerate(header_row):
            if h == "DEBET":
                idx_jurnal_debet = i
            if h == "KREDIT":
                idx_jurnal_kredit = i
            if h == "VOUCHER":
                idx_voucher = i

        if idx_jurnal_debet is None or idx_jurnal_kredit is None:
            continue  # sheet ini belum ada kolom jurnal -> sudah "mentah", lewati

        baris_gt = []
        kolom_kosongkan = set()
        for excel_row in ws.iter_rows(min_row=header_rownum + 1):
            if excel_row[idx_keterangan].value is None and excel_row[idx_tanggal].value is None:
                continue

            no_akun_debet = excel_row[idx_jurnal_debet - 2].value
            nama_akun_debet = excel_row[idx_jurnal_debet - 1].value
            no_akun_kredit = excel_row[idx_jurnal_kredit - 2].value
            nama_akun_kredit = excel_row[idx_jurnal_kredit - 1].value

            baris_gt.append({
                "no_akun_debet": no_akun_debet,
                "nama_akun_debet": nama_akun_debet,
                "no_akun_kredit": no_akun_kredit,
                "nama_akun_kredit": nama_akun_kredit,
            })

            for offset in (idx_jurnal_debet - 2, idx_jurnal_debet - 1, idx_jurnal_debet,
                           idx_jurnal_kredit - 2, idx_jurnal_kredit - 1, idx_jurnal_kredit):
                kolom_kosongkan.add(offset)
            if idx_voucher is not None:
                kolom_kosongkan.add(idx_voucher)

            for offset in kolom_kosongkan:
                excel_row[offset].value = None

        ground_truth[sheet_name] = baris_gt

    wb.save(path_keluaran)
    return ground_truth


# ============================================================
# 2. HIGHLIGHT HASIL DI EXCEL
# ============================================================

def tulis_excel_hasil(df_hasil, ground_truth_flat: list | None, path_keluaran: Path):
    """
    Tulis df_hasil (kolom sumber_kategori sudah ada) ke Excel dengan warna:
      - kuning: Belum Terkategori
      - hijau : perlu cek (pola tidak konsisten / AI confidence rendah)
      - merah : salah dibanding ground_truth_flat (list sejajar index df_hasil,
                None kalau tidak ada ground truth utk baris itu / mode tanpa --strip)
    """
    import pandas as pd

    kolom = ["no", "bank", "tanggal", "keterangan", "mutasi_debet", "mutasi_kredit",
             "no_akun_debet", "nama_akun_debet", "jml_debet",
             "no_akun_kredit", "nama_akun_kredit", "jml_kredit",
             "sumber_kategori", "confidence_ai", "alasan_ai", "catatan_ai"]
    kolom = [k for k in kolom if k in df_hasil.columns]
    df_out = df_hasil[kolom].copy()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hasil Uji"
    ws.append(kolom)

    idx_sumber = kolom.index("sumber_kategori")

    for i, (_, row) in enumerate(df_out.iterrows()):
        nilai_baris = [row[k] for k in kolom]
        nilai_baris = [v.isoformat() if hasattr(v, "isoformat") else v for v in nilai_baris]
        ws.append(nilai_baris)

        excel_row_idx = i + 2
        sumber = str(row["sumber_kategori"] or "")

        salah = False
        if ground_truth_flat is not None and i < len(ground_truth_flat) and ground_truth_flat[i] is not None:
            salah = ground_truth_flat[i] is False  # False = prediksi tidak cocok ground truth

        if "Belum Terkategori" in sumber:
            fill = FILL_KUNING
        elif salah:
            fill = FILL_MERAH
        elif "perlu cek" in sumber or "perlu review" in sumber:
            fill = FILL_HIJAU
        else:
            fill = None

        if fill:
            for c in range(1, len(kolom) + 1):
                ws.cell(row=excel_row_idx, column=c).fill = fill

    for col_cells in ws.columns:
        panjang = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(panjang + 2, 10), 45)

    wb.save(path_keluaran)


# ============================================================
# 3. MAIN
# ============================================================

def _norm_akun(v):
    """Normalisasi no_akun untuk dibandingkan (biar '11200003' == 11200003.0 == 11200003)."""
    if v is None:
        return None
    try:
        return str(int(float(v)))
    except (ValueError, TypeError):
        return str(v).strip().upper() or None


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--file", required=True, help="Path file .xlsx yang diuji")
    parser.add_argument("--client-id", default="EVALTEST", help="ID client (default: EVALTEST, terisolasi dari data produksi)")
    parser.add_argument("--strip", action="store_true", help="Buat file mentah otomatis dari file yang sudah dijurnal + hitung akurasi vs ground truth")
    parser.add_argument("--tanpa-ai", action="store_true", help="Matikan tahap AI (uji baseline pola + kata kunci saja)")
    parser.add_argument("--pola-bersih", action="store_true", help="Hapus dulu pola tersimpan utk client-id ini sebelum uji (simulasi cold start)")
    parser.add_argument("--output-dir", default="./hasil_uji", help="Folder output (default: ./hasil_uji)")
    args = parser.parse_args()

    path_asli = Path(args.file)
    if not path_asli.exists():
        print(f"❌ File tidak ditemukan: {path_asli}", file=sys.stderr)
        sys.exit(1)

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if args.pola_bersih:
        path_pola = Path(ak._path_pola("pola_bank", args.client_id))
        for p in (path_pola, Path(str(path_pola) + ".bak")):
            if p.exists():
                p.unlink()
                print(f"🗑️  Pola lama dihapus: {p}")

    ground_truth = None
    path_uji = path_asli

    if args.strip:
        path_mentah = output_dir / f"mentah_{path_asli.name}"
        print(f"🔧 Membuat versi 'mentah' dari {path_asli.name} (kolom jurnal dikosongkan)...")
        ground_truth = buat_file_mentah_dan_ground_truth(path_asli, path_mentah)
        total_gt_rows = sum(len(v) for v in ground_truth.values())
        print(f"   -> ground truth diambil dari {total_gt_rows} baris, {len(ground_truth)} sheet bank.")
        path_uji = path_mentah

    print(f"\n🚀 Menjalankan proses_file_rekening_koran() (client_id={args.client_id!r}, "
          f"pakai_ai={not args.tanpa_ai})...\n")

    with open(path_uji, "rb") as f:
        buf = io.BytesIO(f.read())
    buf.name = path_uji.name

    hasil = ak.proses_file_rekening_koran(
        buf, nama_file=path_uji.name, client_id=args.client_id, pakai_ai=not args.tanpa_ai,
    )

    df_hasil = hasil["df"]
    if df_hasil.empty:
        print("❌ Tidak ada baris rekening koran yang berhasil dibaca dari file ini.")
        if hasil.get("sheet_dilewati"):
            print("   Sheet dilewati / peringatan:")
            for p in hasil["sheet_dilewati"]:
                print(f"     - {p}")
        sys.exit(1)

    # --- Breakdown sumber kategori ---
    def kelompok_sumber(s: str) -> str:
        s = str(s or "")
        if "Belum Terkategori" in s:
            return "Belum Terkategori"
        if s.startswith("Sesuai Pola"):
            return "Pola historis (perlu cek)" if "perlu cek" in s else "Pola historis"
        if s == "Kata kunci COA":
            return "Kata kunci COA"
        if s.startswith("AI"):
            return "AI (perlu review)" if "perlu review" in s else "AI"
        if s == "Data Asli dari File":
            return "Data asli dari file (sudah terisi sebelumnya)"
        return "Lainnya/tidak diketahui"

    counter = Counter(kelompok_sumber(s) for s in df_hasil["sumber_kategori"])
    total = len(df_hasil)

    print("=" * 72)
    print(f"  RINGKASAN KATEGORISASI OTOMATIS -- {path_asli.name}")
    print("=" * 72)
    print(f"  Total baris: {total}\n")

    urutan = ["Pola historis", "Pola historis (perlu cek)", "Kata kunci COA",
              "AI", "AI (perlu review)", "Data asli dari file (sudah terisi sebelumnya)",
              "Belum Terkategori", "Lainnya/tidak diketahui"]
    total_otomatis = 0
    for k in urutan:
        n = counter.get(k, 0)
        if n == 0:
            continue
        pct = n / total * 100
        print(f"  {k:<48} {n:>6}  ({pct:5.1f}%)")
        if k != "Belum Terkategori":
            total_otomatis += n

    print("-" * 72)
    print(f"  {'TOTAL OTOMATIS (bukan Belum Terkategori)':<48} {total_otomatis:>6}  "
          f"({total_otomatis/total*100:5.1f}%)")
    print(f"  {'Belum Terkategori (perlu review manual)':<48} "
          f"{counter.get('Belum Terkategori', 0):>6}  "
          f"({counter.get('Belum Terkategori', 0)/total*100:5.1f}%)")
    print("=" * 72)

    # --- Akurasi vs ground truth (kalau --strip) ---
    ground_truth_flat = None
    if ground_truth is not None:
        # Susun ground truth sejajar urutan baris di df_hasil. muat_workbook()
        # menggabungkan sheet dgn urutan yg sama seperti wb.sheetnames, dan
        # dalam satu sheet urutan baris tidak diubah oleh pipeline -- jadi
        # ground_truth per-sheet TETAP bisa disejajarkan selama kita
        # mengurutkan berdasarkan (bank, urutan_asli).
        gt_per_bank = {bank: list(rows) for bank, rows in ground_truth.items()}
        pointer = {bank: 0 for bank in gt_per_bank}
        ground_truth_flat = []
        cocok = salah = tanpa_gt = 0
        cocok_per_sumber = Counter()
        salah_per_sumber = Counter()

        for _, row in df_hasil.iterrows():
            bank = row.get("bank")
            gt_list = gt_per_bank.get(bank)
            if gt_list is None or pointer[bank] >= len(gt_list):
                ground_truth_flat.append(None)
                tanpa_gt += 1
                continue
            gt = gt_list[pointer[bank]]
            pointer[bank] += 1

            gt_debet = _norm_akun(gt["no_akun_debet"])
            gt_kredit = _norm_akun(gt["no_akun_kredit"])
            if gt_debet is None and gt_kredit is None:
                # baris asli juga tidak ada jurnalnya (mis. baris kosong/summary) -> skip dari perhitungan akurasi
                ground_truth_flat.append(None)
                tanpa_gt += 1
                continue

            pred_debet = _norm_akun(row.get("no_akun_debet"))
            pred_kredit = _norm_akun(row.get("no_akun_kredit"))
            sumber = kelompok_sumber(row.get("sumber_kategori"))

            if sumber == "Belum Terkategori":
                ground_truth_flat.append(None)  # tidak dihitung benar/salah, memang belum ditebak
                continue

            benar = (pred_debet == gt_debet and pred_kredit == gt_kredit)
            ground_truth_flat.append(benar)
            if benar:
                cocok += 1
                cocok_per_sumber[sumber] += 1
            else:
                salah += 1
                salah_per_sumber[sumber] += 1

        total_dinilai = cocok + salah
        print(f"\n  AKURASI DIBANDING JURNAL ASLI (ground truth, --strip)")
        print("-" * 72)
        if total_dinilai == 0:
            print("  (Tidak ada baris yang bisa dinilai -- cek apakah file sumber memang punya jurnal terisi)")
        else:
            print(f"  Baris yang ditebak pipeline (bukan 'Belum Terkategori'): {total_dinilai}")
            print(f"  Benar : {cocok:>6}  ({cocok/total_dinilai*100:5.1f}%)")
            print(f"  Salah : {salah:>6}  ({salah/total_dinilai*100:5.1f}%)")
            print()
            print("  Rincian akurasi per sumber:")
            for k in urutan:
                c, s = cocok_per_sumber.get(k, 0), salah_per_sumber.get(k, 0)
                if c + s == 0:
                    continue
                print(f"    {k:<46} benar {c:>5} / {c+s:<5} ({c/(c+s)*100:5.1f}%)")
        print("=" * 72)

    # --- Tulis Excel hasil dengan highlight ---
    path_keluaran = output_dir / f"hasil_{path_asli.stem}.xlsx"
    tulis_excel_hasil(df_hasil, ground_truth_flat, path_keluaran)
    print(f"\n📄 Hasil detail + highlight disimpan ke: {path_keluaran}")
    print(f"   (kuning = Belum Terkategori, hijau = perlu cek, merah = salah vs ground truth)")

    if hasil.get("sheet_dilewati"):
        print(f"\n⚠️  Sheet dilewati / peringatan:")
        for p in hasil["sheet_dilewati"]:
            print(f"   - {p}")


if __name__ == "__main__":
    main()