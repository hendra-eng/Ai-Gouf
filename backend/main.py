"""
main.py
=======
Backend FastAPI untuk AI Gouf Consulting.

CARA MENJALANKAN (development):
    cd backend
    python -m venv venv
    source venv/bin/activate        # Windows: venv\\Scripts\\activate
    pip install -r requirements.txt
    # Buat file .env di folder ini (lihat .env.example) sebelum jalan --
    # isinya DATABASE_URL (Supabase), JWT_SECRET_KEY, DEEPSEEK_API_KEY.
    uvicorn main:app --reload --port 8000

============================================================
[FIX v4] Tambah load_dotenv() di baris PALING ATAS, SEBELUM import
akuntansi_ai / db_client / modules.auth. Tanpa ini, isi file .env
(DATABASE_URL ke Supabase, JWT_SECRET_KEY, DEEPSEEK_API_KEY) TIDAK
PERNAH terbaca otomatis -- padahal python-dotenv sudah ada di
requirements.txt sejak awal, cuma belum pernah benar-benar dipanggil.
Dulu di Streamlit ini otomatis kebaca lewat st.secrets/secrets.toml;
sekarang gantinya file .env + load_dotenv() ini.

URUTAN PENTING: load_dotenv() harus dipanggil SEBELUM `import db_client
as dbc` dan `from modules import auth`, karena db_client.py membaca
os.environ.get("DATABASE_URL") saat pertama kali di-import (jadi
koneksi ke Supabase dibuat saat itu juga), dan modules/auth.py membaca
JWT_SECRET_KEY dengan cara yang sama. Kalau load_dotenv() dipanggil
sesudah kedua import itu, sudah kelambatan -- keduanya akan pakai
default yang salah (sqlite lokal / secret key tidak aman).
============================================================
"""

from dotenv import load_dotenv
from pathlib import Path as _PathAwal

# [FIX v5] Sebelumnya load_dotenv() dipanggil tanpa argumen, yang berarti
# python-dotenv cuma mencari file .env mulai dari CURRENT WORKING
# DIRECTORY (folder tempat command `uvicorn` diketik) ke atas. Kalau
# kamu jalankan uvicorn dari folder lain -- misal root "migrasi-react"
# alih-alih "migrasi-react\backend" (gampang kejadian di terminal VS
# Code, yang defaultnya buka di root workspace) -- file .env di folder
# backend TIDAK PERNAH ketemu. load_dotenv() lalu diam-diam tidak
# melakukan apa-apa (tanpa error), sehingga DATABASE_URL/JWT_SECRET_KEY/
# DEEPSEEK_API_KEY semuanya balik ke default (DATABASE_URL jatuh ke
# fallback "sqlite:///ai_gouf.db" di db_client.py) -- data pun nyasar
# ke SQLite lokal, bukan Supabase, TANPA ada error yang kelihatan.
#
# Sekarang path .env dihitung dari lokasi file main.py ini SENDIRI
# (bukan dari cwd), jadi hasilnya konsisten mau uvicorn dijalankan dari
# folder mana pun.
_ENV_PATH = _PathAwal(__file__).resolve().parent / ".env"
load_dotenv(dotenv_path=_ENV_PATH)  # [FIX v5] WAJIB paling atas, sebelum import db_client / modules.auth

if not _ENV_PATH.exists():
    # Jangan diam saja kalau file .env-nya sendiri tidak ketemu --
    # ini kemungkinan besar kesalahan setup (file belum dibuat / typo
    # nama file), bukan cuma soal cwd.
    print(f"[PERINGATAN] File .env tidak ditemukan di: {_ENV_PATH}")
    print("DATABASE_URL, JWT_SECRET_KEY, dan DEEPSEEK_API_KEY kemungkinan akan pakai nilai default yang salah.")

import asyncio
import base64
import hashlib
import io
import json
import math
import os
import queue
import threading
import uuid
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
from apscheduler.schedulers.background import BackgroundScheduler
from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel

import akuntansi_ai as ak
import db_client as dbc
from modules import (
    accounting_export, ai_analysis, auth, calk_aset_tetap, calk_export,
    cross_matching, dashboard, dedup_transaksi,
    deteksi_kesalahan_pembelian as dkp, history,
    laporan_keuangan as lapkeu, notifikasi,
    kertas_kerja,  # [BARU] Generator Kertas Kerja Laporan Keuangan dari PDF rekening koran
    ai_file_reader,  # [BARU] Kirim file (teks/gambar/PDF) langsung ke Claude API, tanpa parsing manual
    cache_cleanup,  # [BARU -- POIN 1] Pembersihan terjadwal cache ekstraksi PDF & Office (TTL + LRU)
    excel_export_worker,  # [BARU -- POINT 2] Generate Excel hasil proses lewat ProcessPoolExecutor
    claude_client,  # [BARU] Narasi AI (CALK/Asumsi/Ringkasan) di export 18-sheet, lihat _lengkapi_narasi_ai_export_18_sheet()
)
from modules.kertas_kerja_router import router as kertas_kerja_router  # [BARU] Endpoint Kertas Kerja Laporan Keuangan
from modules.tax_router import router as tax_router
from modules.tax_case_router import router as tax_case_router
from modules import tax_scheduler

# [FIX v5] Konfirmasi eksplisit di terminal, database mana yang BENAR-BENAR
# kepakai saat startup -- supaya "diam-diam jatuh ke sqlite lokal" tidak
# bisa lolos tanpa ketahuan lagi. Password/detail koneksi disensor,
# cukup tunjukkan jenis DB + host-nya saja.
_db_url_terpakai = os.environ.get("DATABASE_URL", "sqlite:///ai_gouf.db")
if _db_url_terpakai.startswith("sqlite"):
    print(f"[DB] Memakai SQLite LOKAL: {_db_url_terpakai}  <-- BUKAN Supabase! Cek .env kalau ini tidak diinginkan.")
else:
    # Contoh: postgresql://postgres:xxxx@host.supabase.co:5432/postgres
    # -> tampilkan cuma bagian setelah "@" (host + db), sensor user:password.
    _bagian_setelah_at = _db_url_terpakai.split("@")[-1] if "@" in _db_url_terpakai else "(format tidak dikenali)"
    print(f"[DB] Memakai Postgres/Supabase, host: {_bagian_setelah_at}")

app = FastAPI(title="AI Gouf Consulting API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "http://localhost:5174",
        "http://127.0.0.1:5174",
        # [BARU] Frontend "Agent AI" sekarang di-port ke dalam Dashboard
        # Next.js (default dev server port 3000), bukan lagi project Vite
        # yang lama (5173/5174) -- tanpa origin ini browser akan menolak
        # semua request dari Dashboard ke backend ini (CORS error di
        # console, walau backend-nya hidup & benar).
        "http://localhost:3000",
        "http://127.0.0.1:3000",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

FOLDER_HASIL = Path(__file__).parent / "hasil_output"
FOLDER_HASIL.mkdir(exist_ok=True)

# ============================================================
# [BARU -- Point 3] CACHE HASIL GENERATE LAPORAN 18-SHEET
# ============================================================
# Masalah yang diperbaiki: sebelum ini, tiap kali user klik "Generate"
# (baik unduh Excel maupun preview JSON) untuk client+tahun yang SAMA,
# _susun_data_export_18_sheet() menghitung ULANG SEMUANYA dari nol --
# query jurnal ribuan baris, susun lampiran SPT rinci, susun tren
# piutang/utang, dst -- padahal kalau tidak ada data baru masuk sejak
# generate terakhir, hasilnya PASTI SAMA PERSIS.
#
# Cache di sini SENGAJA disimpan per-proses (dict biasa + lock), BUKAN
# lewat file/Redis -- cukup untuk deployment 1 proses uvicorn seperti
# sekarang (lihat catatan cara jalan di kepala file ini). Kalau nanti
# di-scale ke banyak worker/proses, cache ini perlu dipindah ke
# penyimpanan bersama (mis. tabel DB atau Redis) supaya konsisten
# antar proses -- dicatat di sini supaya tidak lupa kalau saatnya tiba.
#
# Validasi cache-hit BUKAN pakai TTL waktu (mis. "cache berlaku 5
# menit") -- itu rawan menyajikan laporan basi kalau kebetulan ada
# upload/koreksi PERSIS di jendela waktu itu. Sebagai gantinya dipakai
# SIGNATURE data (lihat db_client.hitung_signature_data_laporan()):
# hash ringan dari COUNT+MAX(timestamp) jurnal_posting & COA client
# ini. Selama signature-nya SAMA dengan saat hasil terakhir dihitung,
# datanya dijamin belum berubah -- cache aman dipakai berapa lama pun.
_CACHE_EXPORT_18_SHEET: Dict[str, Dict[str, Any]] = {}
_LOCK_CACHE_EXPORT_18_SHEET = threading.Lock()

# Batas jumlah entri cache supaya dict ini tidak tumbuh tanpa henti
# kalau banyak client/tahun berbeda-beda dipakai (memory leak). Entri
# TERLAMA (berdasar waktu terakhir ditulis) dibuang duluan kalau
# kepenuhan -- cukup sebagai pagar kasar, bukan LRU presisi.
_BATAS_ENTRI_CACHE_EXPORT_18_SHEET = 200


def _kunci_cache_export_18_sheet(client_id: int, req: "Export18SheetRequest") -> str:
    """
    Kunci cache = client_id + SEMUA parameter request yang memengaruhi
    hasil (bukan cuma tahun) -- req.tahun_sebelumnya, metode_penyusutan,
    prive_atau_dividen, dst semuanya ikut menentukan hasil akhir, jadi
    dua request dengan tahun sama tapi parameter lain beda HARUS
    dianggap kunci cache berbeda (kalau tidak, user bisa dapat hasil
    generate dengan parameter yang salah karena "kena" cache request
    sebelumnya).
    """
    bahan = json.dumps(
        {"client_id": client_id, **req.dict()}, sort_keys=True, default=str,
    )
    return hashlib.sha256(bahan.encode("utf-8")).hexdigest()[:24]


def _ambil_cache_export_18_sheet(kunci: str, signature_saat_ini: str) -> Optional[Any]:
    """Cache-hit hanya kalau kunci ADA dan signature datanya masih sama."""
    with _LOCK_CACHE_EXPORT_18_SHEET:
        entri = _CACHE_EXPORT_18_SHEET.get(kunci)
        if entri and entri.get("signature") == signature_saat_ini:
            return entri.get("hasil")
    return None


def _simpan_cache_export_18_sheet(kunci: str, signature_saat_ini: str, hasil: Any) -> None:
    with _LOCK_CACHE_EXPORT_18_SHEET:
        if len(_CACHE_EXPORT_18_SHEET) >= _BATAS_ENTRI_CACHE_EXPORT_18_SHEET and kunci not in _CACHE_EXPORT_18_SHEET:
            kunci_terlama = min(
                _CACHE_EXPORT_18_SHEET, key=lambda k: _CACHE_EXPORT_18_SHEET[k]["waktu"],
            )
            del _CACHE_EXPORT_18_SHEET[kunci_terlama]
        _CACHE_EXPORT_18_SHEET[kunci] = {
            "signature": signature_saat_ini, "hasil": hasil, "waktu": datetime.now(),
        }

app.include_router(tax_router, prefix="/tax", tags=["tax-research"])
app.include_router(tax_case_router, prefix="/tax/cases", tags=["tax-case-law"])
app.include_router(kertas_kerja_router, prefix="/kertas-kerja", tags=["kertas-kerja"])  # [BARU]


@app.on_event("startup")
def _startup_buat_tabel_db():
    # [FIX] init_db() (Base.metadata.create_all) sebelumnya tidak pernah
    # dipanggil di mana pun -- kalau database masih kosong/tabel baru
    # (mis. "percakapan", "pesan_chat") belum ada, semua endpoint yang
    # menyentuhnya akan error. create_all() aman dipanggil berkali-kali:
    # tabel yang sudah ada tidak akan diubah/dihapus.
    try:
        dbc.init_db()
    except Exception as e:  # noqa: BLE001
        print(f"[PERINGATAN] Gagal inisialisasi tabel database saat startup: {e}")


# [BARU] Sistem reminder/deadline proaktif SPT -- lihat modules/notifikasi.py.
# proses_spt() di akuntansi_ai.py sudah menghitung jumlah_terlambat_lapor &
# jumlah_berisiko_terlambat_setor per baris; sebelumnya angka itu cuma
# muncul di response upload/Excel (dilaporkan SETELAH data diupload).
# Scheduler ini jalan tiap hari (jam bisa diatur lewat REMINDER_JAM_CEK di
# .env, default 07:00) utk mengecek tabel reminder_deadline_spt dan kirim
# notifikasi in-app + WA SEBELUM jatuh tempo (H-3/H-1), bukan cuma setelah
# terlambat.
_scheduler = BackgroundScheduler(timezone="Asia/Jakarta")


@app.on_event("startup")
def _startup_scheduler_reminder():
    jam_str = os.environ.get("REMINDER_JAM_CEK", "07:00")
    try:
        jam, menit = [int(x) for x in jam_str.split(":")]
    except ValueError:
        jam, menit = 7, 0
    try:
        _scheduler.add_job(
            notifikasi.jalankan_pengecekan_reminder_spt,
            "cron",
            hour=jam,
            minute=menit,
            id="cek_reminder_deadline_spt",
            replace_existing=True,
        )
        # [BARU -- POIN 1] Daftarkan job pembersihan cache ekstraksi
        # (PDF di kertas_kerja.py + Office di ai_file_reader.py) ke
        # SCHEDULER YANG SAMA -- 1 BackgroundScheduler proses cukup utk
        # semua job terjadwal, tidak perlu scheduler APScheduler baru.
        # Jam berbeda (default 03:00, di luar jam kantor) dari reminder
        # SPT (default 07:00) supaya tidak numpuk I/O disk di jam yang
        # sama -- bisa dioverride lewat CACHE_CLEANUP_JAM/CACHE_CLEANUP_MENIT
        # di .env kalau perlu.
        jam_cache = int(os.environ.get("CACHE_CLEANUP_JAM", "3"))
        menit_cache = int(os.environ.get("CACHE_CLEANUP_MENIT", "0"))
        cache_cleanup.daftarkan_job_pembersihan_cache(_scheduler, jam=jam_cache, menit=menit_cache)
        _scheduler.start()
        print(f"[notifikasi] Scheduler reminder deadline SPT aktif, jalan tiap hari jam {jam:02d}:{menit:02d} WIB.")
    except Exception as e:  # noqa: BLE001
        print(f"[PERINGATAN] Gagal menyalakan scheduler reminder deadline: {e}")


@app.on_event("shutdown")
def _shutdown_scheduler_reminder():
    try:
        _scheduler.shutdown(wait=False)
    except Exception:
        pass


# [BARU] Scheduler tugas latar belakang fitur riset pajak (mis. cek
# ulang peraturan berstatus belum diverifikasi) -- lihat
# modules/tax_scheduler.py. Sebelumnya modul ini sudah ditulis lengkap
# tapi belum pernah dipanggil dari main.py sama sekali, jadi job-nya
# tidak pernah jalan. Dipisah dari _scheduler (reminder SPT) di atas
# karena memang dua scheduler APScheduler yang berbeda tujuan.
@app.on_event("startup")
def _startup_tax_scheduler():
    try:
        tax_scheduler.start_scheduler()
    except Exception as e:  # noqa: BLE001
        print(f"[PERINGATAN] Gagal menyalakan tax scheduler: {e}")


@app.on_event("shutdown")
def _shutdown_tax_scheduler():
    try:
        tax_scheduler.stop_scheduler()
    except Exception:
        pass


class PesanRiwayat(BaseModel):
    role: str
    content: str


class ChatRequest(BaseModel):
    pesan: str
    riwayat: Optional[List[PesanRiwayat]] = None
    # [BARU] Konteks data yang sudah diproses di percakapan ini, supaya AI
    # tidak "buta" soal hasil upload sebelumnya. Lihat ChatPage.jsx --
    # ringkasan_data dibangun dari resultsByCategory tiap kategori yang
    # sudah punya hasil. client_id (opsional) dipakai untuk mengambil
    # jumlah pola & temuan mencurigakan MILIK CLIENT ITU dari file pola
    # yang sudah tersimpan (lihat akuntansi_ai.muat_pola/_path_pola).
    ringkasan_data: Optional[List[str]] = None
    client_id: Optional[int] = None
    # [BARU] Isi kalau percakapan ini di jalur "esb_account" (spesifik soal
    # 1 akun ESB) -- dipakai untuk kasih AI konteks detail akun itu (nama,
    # tipe, status aktif), bukan cuma status ESB client secara umum.
    esb_account_id: Optional[int] = None
    # [BARU] Id percakapan (dari /api/percakapan) supaya pesan user & balasan
    # AI disimpan permanen ke database, bukan cuma hidup di state React.
    percakapan_id: Optional[int] = None


@app.get("/api/health")
def health():
    return {
        "status": "ok",
        # [FIX -- JALUR SEMENTARA GROQ] Sebelumnya cuma cek DEEPSEEK_API_KEY,
        # jadi kalau DeepSeek kosong/gagal tapi GROQ_API_KEY aktif (jalur
        # sementara, lihat _konfigurasi_provider_chat() di akuntansi_ai.py),
        # endpoint ini salah lapor ai_aktif=False padahal chat tetap jalan.
        "ai_aktif": bool(ak._konfigurasi_provider_chat()),
        # [DIUBAH -- KATEGORISASI KHUSUS GROQ] Sebelumnya key ini bernama
        # "claude_aktif" karena kategorisasi (dipakai kertas_kerja/
        # kategorikan_dengan_ai, dst) tadinya lewat Claude dulu baru fallback
        # Groq. Sekarang _konfigurasi_provider_kategorisasi() di
        # akuntansi_ai.py HANYA berisi Groq (Claude & DeepSeek sudah tidak
        # dipakai sama sekali di jalur kategorisasi), jadi key-nya diganti
        # nama supaya tidak menyesatkan. Nilainya true kalau
        # GROQ_API_KEY_KATEGORISASI atau GROQ_API_KEY terisi.
        "kategorisasi_aktif": bool(ak._konfigurasi_provider_kategorisasi()),
        "database_aktif": dbc.cek_koneksi(),  # [FIX v4] cek cepat Supabase konek atau tidak
    }


@app.post("/api/login")
def login(username: str = Form(...), password: str = Form(...)):
    user = auth.authenticate(username, password)
    if not user:
        raise HTTPException(status_code=401, detail="Username atau password salah.")
    token = auth.buat_token(user)
    return {
        "token": token,
        "username": user["username"],
        "role": user["role"],
        "nama": user.get("nama"),
    }


@app.get("/api/client")
def api_daftar_client(
    tipe: Optional[str] = None,
    punya_esb: Optional[bool] = None,
    user: dict = Depends(auth.require_level(3)), 
):
    """[BARU] Tambah query param `punya_esb` (true/false) untuk filter
    client yang sudah/belum punya integrasi ESB, mis:
        GET /api/client?punya_esb=true   -> client yang sudah ada akun ESB
        GET /api/client?punya_esb=false  -> client yang belum ada akun ESB
    """
    return {"clients": dbc.daftar_client(tipe, punya_esb=punya_esb)}


@app.post("/api/client")
def api_tambah_client(
    nama: str = Form(...),
    lokasi: Optional[str] = Form(None),
    tipe: str = Form("accounting"),
    # [BARU] nomor_wa/email opsional saat bikin client -- dipakai sistem
    # reminder deadline SPT utk kirim notifikasi WA/email. Bisa juga diisi
    # belakangan lewat PUT /api/client/{client_id}/kontak.
    nomor_wa: Optional[str] = Form(None),
    email: Optional[str] = Form(None),
    user: dict = Depends(auth.require_level(3)),  
):
    client_id = dbc.tambah_client(nama, lokasi, tipe, nomor_wa=nomor_wa, email=email)
    if client_id is None:
        raise HTTPException(status_code=500, detail="Gagal menambah client.")
    return {"id": client_id, "nama": nama, "lokasi": lokasi, "tipe": tipe, "nomor_wa": nomor_wa, "email": email}


class UpdateKontakClientRequest(BaseModel):
    nomor_wa: Optional[str] = None  # format internasional mis. "6281234567890"
    email: Optional[str] = None


@app.put("/api/client/{client_id}/kontak")
def api_update_kontak_client(
    client_id: int,
    req: UpdateKontakClientRequest,
    user: dict = Depends(auth.require_level(3)),
):
    """Isi/ubah nomor WA & email client -- wajib diisi supaya client ini
    bisa dapat reminder deadline SPT lewat WA (in-app tetap jalan tanpa
    ini, tapi WA tidak akan terkirim kalau nomor_wa kosong)."""
    berhasil = dbc.update_kontak_client(client_id, nomor_wa=req.nomor_wa, email=req.email)
    if not berhasil:
        raise HTTPException(status_code=404, detail="Client tidak ditemukan.")
    return {"berhasil": True}


# ============================================================
# [BARU] AKUN ESB (integrasi API POS/kasir) per client
# ============================================================

class TambahEsbAccountRequest(BaseModel):
    account_name: str
    esb_type: Optional[str] = None
    api_base_url: Optional[str] = None
    consumer_key: Optional[str] = None
    consumer_secret: Optional[str] = None
    is_active: bool = True
    is_default: bool = False
    auto_discover: bool = False


@app.get("/api/client/{client_id}/esb-accounts")
def api_esb_accounts_client(client_id: int, user: dict = Depends(auth.get_current_user)):
    """List akun ESB milik satu client. consumer_secret dikirim ter-mask
    (mis. '••••6321'), TIDAK PERNAH dalam bentuk asli -- lihat
    db_client._mask_secret()."""
    return {"esb_accounts": dbc.ambil_esb_accounts_client(client_id)}


@app.post("/api/client/{client_id}/esb-accounts")
def api_tambah_esb_account(
    client_id: int,
    req: TambahEsbAccountRequest,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    akun_id = dbc.tambah_esb_account(
        client_id=client_id, account_name=req.account_name, esb_type=req.esb_type,
        api_base_url=req.api_base_url, consumer_key=req.consumer_key,
        consumer_secret=req.consumer_secret, is_active=req.is_active,
        is_default=req.is_default, auto_discover=req.auto_discover,
    )
    if akun_id is None:
        raise HTTPException(status_code=500, detail="Gagal menambah akun ESB.")
    return {"id": akun_id, "client_id": client_id, "account_name": req.account_name}


@app.delete("/api/client/{client_id}/esb-accounts/{esb_account_id}")
def api_hapus_esb_account(client_id: int, esb_account_id: int, user: dict = Depends(auth.get_current_user)):
    berhasil = dbc.hapus_esb_account(esb_account_id)
    if not berhasil:
        raise HTTPException(status_code=404, detail="Akun ESB tidak ditemukan.")
    return {"berhasil": True}


@app.get("/api/client/{client_id}/riwayat")
def api_riwayat_client(client_id: int, user: dict = Depends(auth.get_current_user)):
    """Riwayat hasil UMUM CLIENT (tabel 'hasil'). Untuk riwayat akun ESB,
    pakai /api/esb-account/{esb_account_id}/riwayat."""
    hasil = dbc.ambil_hasil_client(client_id)
    riwayat = [
        {
            "jenis_dokumen": h["jenis"],
            "hasil": h["data"],
            "nama_file": h["data"].get("nama_file", "-") if isinstance(h["data"], dict) else "-",
            "tanggal": h["dibuat_at"],
        }
        for h in hasil
    ]
    return {"riwayat": riwayat}


@app.get("/api/esb-account/{esb_account_id}/riwayat")
def api_riwayat_esb_account(esb_account_id: int, user: dict = Depends(auth.get_current_user)):
    """Riwayat hasil khusus 1 akun ESB (tabel 'hasil_esb', terpisah dari
    hasil umum client di tabel 'hasil')."""
    hasil = dbc.ambil_hasil_esb(esb_account_id)
    riwayat = [
        {
            "jenis_dokumen": h["jenis"],
            "hasil": h["data"],
            "nama_file": h["data"].get("nama_file", "-") if isinstance(h["data"], dict) else "-",
            "tanggal": h["dibuat_at"],
            "esb_account_id": h["esb_account_id"],
        }
        for h in hasil
    ]
    return {"riwayat": riwayat}


@app.get("/api/client/{client_id}/audit-log")
def api_audit_log_client(
    client_id: int,
    limit: int = 200,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """
    [BARU] Audit trail: riwayat siapa-mengubah-apa-kapan untuk 1 client --
    mencakup auto-fix data (aksi="auto_fix_data"), perubahan COA
    (tambah/update/hapus_akun_coa), jawaban klarifikasi, posting/tolak
    jurnal, dan generate laporan keuangan. Diurutkan dari yang terbaru.
    Lihat modules/history.py & db_client.py::log_audit/get_audit_history.
    """
    return {"audit_log": history.ambil_riwayat(client_id=client_id, limit=limit)}


@app.get("/api/client/{client_id}/dashboard")
def api_dashboard_client(client_id: int, user: dict = Depends(auth.get_current_user)):
    """[BARU] Live Dashboard per client -- dihitung dari riwayat hasil proses
    yang sudah tersimpan di database (lihat modules/dashboard.py::
    ringkas_dashboard_dari_riwayat untuk penjelasan kenapa tidak memakai
    get_live_stats() versi Streamlit lama)."""
    hasil = dbc.ambil_hasil_client(client_id)
    riwayat = [
        {"jenis_dokumen": h["jenis"], "hasil": h["data"]}
        for h in hasil
    ]
    return dashboard.ringkas_dashboard_dari_riwayat(riwayat)


@app.get("/api/client/{client_id}/rekonsiliasi-lintas-dokumen")
def api_rekonsiliasi_lintas_dokumen(
    client_id: int,
    npwp_perusahaan: Optional[str] = None,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """[BARU] Rekonsiliasi lintas-dokumen (cross-matching) -- lihat
    modules/cross_matching.py. Menjalankan 3 pencocokan sekaligus dari data
    yang sudah tersimpan utk client ini: Bank vs Piutang, PPN Faktur vs SPT
    Masa PPN, dan Slip Gaji vs Absensi. Semua rule-based (bukan AI
    generatif) -- hasil "TIDAK_KETEMU"/"PERLU_DICEK" tetap wajib direview
    manusia, endpoint ini cuma mempercepat proses cari.

    npwp_perusahaan (opsional): kalau diisi, hanya faktur pajak dengan
    npwp_penjual == ini yang dihitung sbg PPN Keluaran perusahaan.
    """
    hasil = cross_matching.jalankan_rekonsiliasi_lintas_dokumen(
        client_id, dbc, npwp_perusahaan=npwp_perusahaan,
    )
    return _bersihkan_untuk_json(hasil)


class DeteksiKesalahanPembelianRequest(BaseModel):
    checks: List[str] = []  # kosong = jalankan semua 7 pengecekan


@app.post("/api/client/{client_id}/deteksi-kesalahan-pembelian")
def api_deteksi_kesalahan_pembelian(
    client_id: int,
    req: DeteksiKesalahanPembelianRequest,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """[BARU] "Deteksi & pencegahan kesalahan" untuk data Pembelian
    (PO/Invoice) -- lihat modules/deteksi_kesalahan_pembelian.py. 7
    pengecekan rule-based (bukan AI generatif), user pilih salah satu/
    beberapa/semua lewat `checks` (kosong = semua):
      po_invoice, pph23_jasa, harga_tidak_wajar, supplier_baru,
      validasi_tanggal, rekap_supplier, cross_check_ap_aging.
    Hasil "PERLU_DICEK"/"PERLU REVIEW"/"SELISIH" tetap wajib direview
    manusia -- endpoint ini mempercepat proses cari, bukan menggantikan
    keputusan akuntan."""
    hasil = dkp.jalankan_deteksi_kesalahan_pembelian(client_id, dbc, checks=req.checks)
    return _bersihkan_untuk_json(hasil)


# ============================================================
# ANALISIS AI (Claude) -- tabel hasil_analisis
# [UBAH] Sebelumnya DeepSeek -- lihat catatan migrasi di masing-masing
# endpoint di bawah (api_buat_analisis_ai / api_buat_ringkasan_eksekutif).
# ============================================================

@app.post("/api/client/{client_id}/analisis-ai")
def api_buat_analisis_ai(
    client_id: int,
    esb_account_id: Optional[int] = None,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """[UBAH -- pindah DeepSeek ke Claude] Minta Claude generate ringkasan
    & temuan dari data hasil yang sudah tersimpan untuk client ini, lalu
    simpan ke tabel hasil_analisis. esb_account_id opsional -- kalau
    diisi, analisis dari tabel hasil_esb (khusus akun ESB itu); kalau
    tidak, dari tabel hasil (umum client).

    [CATATAN MIGRASI] Sebelumnya lewat ai_analysis.analisis_ringkasan_keuangan()
    (DeepSeek) -- sekarang lewat claude_client.analisis_ringkasan_keuangan_claude(),
    fungsi Claude yang SUDAH ada di claude_client.py (dibuat "setara"
    fungsi DeepSeek lama, tool_schema-nya identik: ringkasan/temuan_penting/
    potensi_masalah) -- format `hasil` yang tersimpan ke DB & dikembalikan
    ke frontend TIDAK BERUBAH, jadi tidak perlu ubah apa pun di sisi
    frontend. ai_analysis.py TIDAK dihapus -- masih dipakai modul lain
    (mis. kertas_kerja)."""
    if esb_account_id is not None:
        riwayat = dbc.ambil_hasil_esb(esb_account_id)
    else:
        riwayat = dbc.ambil_hasil_client(client_id)

    if not riwayat:
        raise HTTPException(status_code=400, detail="Belum ada data hasil untuk client ini, tidak bisa dianalisis.")

    try:
        hasil_claude = claude_client.analisis_ringkasan_keuangan_claude(
            riwayat, client_id=str(client_id),
        )
    except claude_client.ClaudeError as e:
        raise HTTPException(status_code=502, detail=str(e))

    analisis_id = dbc.simpan_hasil_analisis(
        client_id=client_id,
        jenis_analisis="ringkasan_keuangan",
        hasil=hasil_claude,
        prompt=f"Ringkasan keuangan (Claude) dari {len(riwayat)} data hasil proses.",
        model_ai=claude_client.MODEL_DEFAULT,
        esb_account_id=esb_account_id,
    )

    return {"id": analisis_id, "hasil": hasil_claude, "model_ai": claude_client.MODEL_DEFAULT}


@app.get("/api/client/{client_id}/analisis-ai")
def api_riwayat_analisis_ai(
    client_id: int,
    jenis_analisis: Optional[str] = None,
    esb_account_id: Optional[int] = None,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """Ambil riwayat analisis AI yang sudah pernah dibuat untuk client ini."""
    return {"riwayat": dbc.ambil_hasil_analisis_client(
        client_id, jenis_analisis=jenis_analisis, esb_account_id=esb_account_id,
    )}


# ============================================================
# [BARU] RINGKASAN EKSEKUTIF (#4) -- versi ringkas laporan utk klien
# NON-AKUNTAN, terpisah dari detail teknis. Gabungan kartu angka (dihitung
# langsung dari data tersimpan, GRATIS/real-time) + narasi AI (opsional,
# [UBAH] manggil Claude -- sebelumnya DeepSeek, lihat catatan migrasi di
# api_buat_ringkasan_eksekutif() di bawah -- makanya dipisah jadi endpoint POST sendiri, bukan
# otomatis tiap kali kartu di-load lewat GET). Narasinya disimpan ke tabel
# hasil_analisis yang SAMA dengan analisis-ai di atas lewat
# jenis_analisis="ringkasan_eksekutif", jadi riwayatnya otomatis bisa
# diambil juga lewat GET /api/client/{client_id}/analisis-ai?jenis_analisis=
# ringkasan_eksekutif kalau suatu saat dibutuhkan, tanpa endpoint terpisah.
# ============================================================

def _hitung_angka_ringkasan_eksekutif(client_id: int) -> dict:
    hasil = dbc.ambil_hasil_client(client_id)
    riwayat = [
        {"jenis_dokumen": h["jenis"], "hasil": h["data"], "tanggal": h["dibuat_at"]}
        for h in hasil
    ]
    return dashboard.ringkas_eksekutif_dari_riwayat(riwayat)


@app.get("/api/client/{client_id}/ringkasan-eksekutif")
def api_ringkasan_eksekutif(client_id: int, user: dict = Depends(auth.get_current_user)):
    """Kartu angka utama, dihitung real-time dari data tersimpan (tanpa
    panggil AI, jadi selalu boleh diakses berkali-kali) + narasi AI
    TERAKHIR yang pernah digenerate (kalau ada), supaya klien tetap lihat
    sesuatu tanpa harus menunggu POST baru setiap buka halaman."""
    angka = _hitung_angka_ringkasan_eksekutif(client_id)
    riwayat_narasi = dbc.ambil_hasil_analisis_client(
        client_id, jenis_analisis="ringkasan_eksekutif", limit=1,
    )
    narasi_terakhir = riwayat_narasi[0] if riwayat_narasi else None
    return {"angka": angka, "narasi_terakhir": narasi_terakhir}


@app.post("/api/client/{client_id}/ringkasan-eksekutif")
def api_buat_ringkasan_eksekutif(client_id: int, user: dict = Depends(auth.get_current_user)):
    """[UBAH -- pindah DeepSeek ke Claude] Generate ULANG narasi AI dari
    angka terkini & simpan sbg riwayat baru. Dipisah dari GET di atas krn
    ini yang benar-benar memanggil Claude (ada biaya/kuota) -- jadi harus
    eksplisit diminta akuntan/klien lewat tombol, bukan otomatis jalan
    tiap kartu di-load.

    [CATATAN MIGRASI] Sebelumnya lewat ai_analysis.buat_ringkasan_eksekutif()
    (DeepSeek) -- sekarang lewat claude_client.generate_ringkasan_eksekutif_claude(),
    fungsi baru khusus dibuat untuk endpoint ini (bahasa awam utk klien
    non-akuntan, bukan istilah teknis akuntansi) -- lihat docstring-nya di
    claude_client.py. Response shape TIDAK berubah, frontend tidak perlu
    diubah."""
    angka = _hitung_angka_ringkasan_eksekutif(client_id)
    if angka["total_dokumen_diproses"] == 0:
        raise HTTPException(
            status_code=400,
            detail="Belum ada data hasil untuk client ini, tidak bisa dibuat ringkasan.",
        )

    try:
        narasi_claude = claude_client.generate_ringkasan_eksekutif_claude(
            angka["kartu_utama"], angka["per_kategori"], client_id=str(client_id),
        )
    except claude_client.ClaudeError as e:
        raise HTTPException(status_code=502, detail=str(e))

    analisis_id = dbc.simpan_hasil_analisis(
        client_id=client_id,
        jenis_analisis="ringkasan_eksekutif",
        hasil=narasi_claude,
        prompt="Ringkasan eksekutif (Claude) dari kartu angka & data per kategori terkini.",
        model_ai=claude_client.MODEL_DEFAULT,
    )

    return {
        "id": analisis_id,
        "angka": angka,
        "narasi": narasi_claude,
        "model_ai": claude_client.MODEL_DEFAULT,
    }


# ============================================================
# [BARU] RIWAYAT PERCAKAPAN (sidebar chat history, mirip ChatGPT/Claude)
# ============================================================

class BuatPercakapanRequest(BaseModel):
    client_id: Optional[int] = None
    # [BARU] Isi ini kalau percakapan spesifik soal 1 akun ESB tertentu --
    # jalurnya otomatis kepisah dari percakapan umum client (lihat jalur
    # di api_daftar_percakapan di bawah).
    esb_account_id: Optional[int] = None
    judul: Optional[str] = None


@app.post("/api/percakapan")
def api_buat_percakapan(req: BuatPercakapanRequest, user: dict = Depends(auth.get_current_user)):
    """Mulai sesi percakapan baru. Judul sementara "Percakapan Baru" --
    akan diganti otomatis begitu pesan pertama user terkirim (lihat
    chat_stream di bawah), sama seperti ChatGPT/Claude auto-title chat baru."""
    percakapan_id = dbc.buat_percakapan(
        username=user["username"],
        client_id=req.client_id,
        esb_account_id=req.esb_account_id,
        judul=req.judul or "Percakapan Baru",
    )
    if percakapan_id is None:
        raise HTTPException(status_code=500, detail="Gagal membuat percakapan baru.")
    return {"id": percakapan_id, "judul": req.judul or "Percakapan Baru"}


@app.get("/api/percakapan")
def api_daftar_percakapan(
    client_id: Optional[int] = None,
    esb_account_id: Optional[int] = None,
    jalur: Optional[str] = None,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """List percakapan milik user yang sedang login, terbaru dulu.

    [BARU] `jalur` memisahkan 2 tab riwayat di sidebar, mis:
        GET /api/percakapan?jalur=client       -> tab "Percakapan Client"
        GET /api/percakapan?jalur=esb_account  -> tab "Percakapan Akun ESB"
    """
    return {"percakapan": dbc.daftar_percakapan(
        user["username"], client_id=client_id, esb_account_id=esb_account_id, jalur=jalur,
    )}


@app.get("/api/percakapan/{percakapan_id}/pesan")
def api_pesan_percakapan(percakapan_id: int, user: dict = Depends(auth.get_current_user)):
    """Ambil seluruh isi chat dari satu percakapan, untuk direstore ke UI
    saat user membuka kembali percakapan lama dari sidebar."""
    return {"pesan": dbc.ambil_pesan_percakapan(percakapan_id)}


@app.delete("/api/percakapan/{percakapan_id}")
def api_hapus_percakapan(percakapan_id: int, user: dict = Depends(auth.get_current_user)):
    berhasil = dbc.hapus_percakapan(percakapan_id)
    if not berhasil:
        raise HTTPException(status_code=404, detail="Percakapan tidak ditemukan.")
    return {"berhasil": True}


def _format_sse(data: str) -> str:
    return f"data: {json.dumps({'delta': data})}\n\n"


@app.post("/api/chat/stream")
def chat_stream(req: ChatRequest, user: dict = Depends(auth.get_current_user)):
    # [FIX -- DELAY CHAT] Sebelumnya bagian konteks (pola, ESB, KPI) di bawah
    # ini dijalankan BERURUTAN satu-satu -- tiap panggilan ke Supabase (ESB,
    # KPI) adalah round-trip jaringan terpisah, jadi totalnya bisa numpuk
    # jadi 1-3 detik jeda SEBELUM AI (Groq) sempat mulai menjawab, walau
    # Groq sendiri sangat cepat begitu dipanggil. Ketiga bagian ini saling
    # independen (tidak saling butuh hasil satu sama lain), jadi sekarang
    # dijalankan PARALEL lewat ThreadPoolExecutor -- totalnya jadi sekitar
    # selama panggilan TERLAMA saja, bukan jumlah semuanya.
    def _ambil_konteks_pola():
        jumlah_pola_ = 0
        jumlah_temuan_mencurigakan_ = 0
        if req.client_id is not None:
            try:
                pola_bank = ak.muat_pola(ak._path_pola("pola_bank", req.client_id))
                pola_penjualan = ak.muat_pola(ak._path_pola("pola_penjualan", req.client_id))
                jumlah_pola_ = len(pola_bank.aturan) + len(pola_penjualan.aturan)
                jumlah_temuan_mencurigakan_ = len(ak.deteksi_pola_mencurigakan(pola_bank))
            except Exception as e:  # noqa: BLE001
                print(f"[PERINGATAN] Gagal ambil konteks pola untuk chat: {e}")
        return jumlah_pola_, jumlah_temuan_mencurigakan_

    def _ambil_info_esb():
        # [BARU] Kasih tahu AI status integrasi ESB, dengan 2 mode:
        # - kalau req.esb_account_id diisi (percakapan jalur "esb_account"):
        #   kasih detail akun ESB itu spesifik.
        # - kalau tidak, tapi ada req.client_id: kasih ringkasan umum status
        #   ESB client tersebut (perilaku lama, tetap dipertahankan).
        if req.esb_account_id is not None:
            try:
                akun_esb_client = dbc.ambil_esb_accounts_client(req.client_id) if req.client_id else []
                akun = next((a for a in akun_esb_client if a["id"] == req.esb_account_id), None)
                if akun:
                    status = "aktif" if akun.get("is_active") else "TIDAK aktif"
                    return (
                        f"\n- Percakapan ini spesifik soal akun ESB '{akun['account_name']}' "
                        f"(tipe: {akun.get('esb_type') or '-'}, status: {status})."
                    )
            except Exception as e:  # noqa: BLE001
                print(f"[PERINGATAN] Gagal ambil detail akun ESB untuk chat: {e}")
        elif req.client_id is not None:
            try:
                akun_esb = dbc.ambil_esb_accounts_client(req.client_id)
                if akun_esb:
                    aktif = [a for a in akun_esb if a.get("is_active")]
                    nama_akun = ", ".join(a["account_name"] for a in akun_esb)
                    return (
                        f"\n- Integrasi ESB: client ini SUDAH punya {len(akun_esb)} akun ESB "
                        f"({len(aktif)} aktif) -- {nama_akun}."
                    )
                return "\n- Integrasi ESB: client ini BELUM punya akun ESB sama sekali."
            except Exception as e:  # noqa: BLE001
                print(f"[PERINGATAN] Gagal ambil info ESB untuk chat: {e}")
        return ""

    def _ambil_konteks_kpi():
        # ============================================================
        # [BARU] Tanya-jawab natural language ke data yang sudah tersimpan
        # ============================================================
        # Sebelumnya AI cuma tahu HITUNGAN (mis. "Piutang: 12 baris diproses,
        # 2 perlu direview" dari ringkasan_data) -- bukan ANGKA ASLI (total
        # piutang berapa rupiah, dst), jadi tidak bisa jawab pertanyaan spt
        # "berapa total piutang klien X bulan ini?". Konteks di bawah ini
        # menyuntikkan angka ASLI dari ringkasan tersimpan (lihat
        # modules/dashboard.py::bangun_kpi_kunci_dari_riwayat /
        # bangun_ringkasan_lintas_client untuk detail & asumsi datanya).
        if req.client_id is not None:
            # Mode 1: percakapan sedang fokus ke SATU client -- kasih semua
            # angka terbaru per jenis dokumen client itu.
            try:
                riwayat_lengkap = dbc.ambil_hasil_client(req.client_id)
                riwayat_utk_kpi = [
                    {"jenis_dokumen": h["jenis"], "hasil": h["data"]} for h in riwayat_lengkap
                ]
                kpi_per_jenis = dashboard.bangun_kpi_kunci_dari_riwayat(riwayat_utk_kpi)
                if kpi_per_jenis:
                    return (
                        "\n- Data angka TERBARU per jenis dokumen client ini (format JSON, "
                        "field sama seperti di tampilan hasil). PAKAI ANGKA INI untuk jawab "
                        "pertanyaan spesifik soal jumlah/total -- JANGAN mengarang angka lain, "
                        "dan kalau field yang ditanya tidak ada di sini, bilang terus terang "
                        "belum ada datanya:\n"
                        + json.dumps(kpi_per_jenis, ensure_ascii=False, default=str)
                    )
            except Exception as e:  # noqa: BLE001
                print(f"[PERINGATAN] Gagal ambil konteks KPI client untuk chat: {e}")
        else:
            # Mode 2: TIDAK ada client aktif -- kemungkinan pertanyaan LINTAS
            # CLIENT (mis. "klien mana yang score-nya di bawah 70?"). Baru
            # dukung jenis "penilaian" dulu (lihat
            # dashboard.FIELD_HEADLINE_LINTAS_CLIENT utk nambah jenis lain).
            try:
                semua_client = dbc.daftar_client()
                ringkasan_lintas = dashboard.bangun_ringkasan_lintas_client(
                    semua_client, dbc.ambil_hasil_client, jenis="penilaian"
                )
                if ringkasan_lintas:
                    return (
                        "\n- Ringkasan Penilaian Klien/Maker LINTAS CLIENT (format JSON), "
                        "dipakai kalau user tanya perbandingan antar client (mis. 'klien mana "
                        "yang score-nya di bawah 70?'). JANGAN mengarang client/angka di luar "
                        "daftar ini:\n"
                        + json.dumps(ringkasan_lintas, ensure_ascii=False, default=str)
                    )
            except Exception as e:  # noqa: BLE001
                print(f"[PERINGATAN] Gagal ambil konteks KPI lintas-client untuk chat: {e}")
        return ""

    # [BARU -- TRANSPARANSI PROSES] Sebelumnya bagian konteks (pola, ESB,
    # KPI) dibangun DI SINI secara blocking, SEBELUM StreamingResponse
    # dibuka -- user tidak melihat apa-apa sampai semuanya selesai (1-3
    # detik diam). Sekarang dipindah KE DALAM event_generator() supaya
    # setiap tahap bisa mengirim event "progress" (step mulai -> step
    # selesai) ke frontend SEBELUM AI mulai menjawab -- pola yang SAMA
    # dengan _format_sse_progress() yang sudah dipakai di
    # /api/proses-file/stream & /generate-kertas-kerja/stream. Frontend
    # (ChatBubble + ProcessingSteps) menampilkan event ini sebagai daftar
    # langkah yang sedang/sudah dikerjakan AI, mirip panel "Menjalankan N
    # perintah..." ala Claude Code.
    #
    # [FIX -- DELAY CHAT] Generate judul (buat_judul_percakapan, panggilan
    # AI TERPISAH) TETAP dijalankan di THREAD TERPISAH ("fire and forget")
    # sebelum stream dimulai -- bukan sesuatu yang user tunggu aktif (cuma
    # label sidebar), jadi tidak perlu jadi salah satu step yang ditampilkan.
    percakapan_baru = req.percakapan_id is not None and not (req.riwayat or [])
    if percakapan_baru:
        def _buat_judul_di_background():
            try:
                judul = ak.buat_judul_percakapan(req.pesan)
                dbc.ubah_judul_percakapan(req.percakapan_id, judul)
            except Exception as e:  # noqa: BLE001
                print(f"[PERINGATAN] Gagal auto-generate judul percakapan: {e}")
        threading.Thread(target=_buat_judul_di_background, daemon=True).start()

    # [BARU] Simpan pesan user ke database SEBELUM memanggil AI, supaya
    # pesan tidak hilang meski panggilan AI di bawah gagal total.
    if req.percakapan_id is not None:
        dbc.simpan_pesan_chat(req.percakapan_id, "user", req.pesan)

    def event_generator():
        jawaban_lengkap = []
        try:
            # [BARU] Step 1 -- kumpulkan konteks (pola, ESB, KPI) paralel
            # lewat ThreadPoolExecutor, TAPI sekarang di dalam generator
            # supaya event "processing" bisa dikirim SEBELUM mulai, dan
            # "done" SETELAH selesai -- ganti label di bawah ini sesuka
            # kamu, ini yang muncul di UI (ProcessingSteps).
            yield _format_sse_progress(
                type="progress", step="konteks",
                label="Membaca pola transaksi & data client",
                status="processing",
            )
            with ThreadPoolExecutor(max_workers=3) as executor:
                future_pola = executor.submit(_ambil_konteks_pola)
                future_esb = executor.submit(_ambil_info_esb)
                future_kpi = executor.submit(_ambil_konteks_kpi)

                jumlah_pola, jumlah_temuan_mencurigakan = future_pola.result()
                info_esb = future_esb.result()
                konteks_data_kpi = future_kpi.result()
            yield _format_sse_progress(
                type="progress", step="konteks",
                label="Membaca pola transaksi & data client",
                status="done",
            )

            # [BARU] Step 2 -- susun system prompt dari konteks yang
            # sudah dikumpulkan.
            yield _format_sse_progress(
                type="progress", step="susun_prompt",
                label="Menyusun konteks jawaban",
                status="processing",
            )
            konteks = ak.buat_ringkasan_konteks_data(
                req.ringkasan_data or [], jumlah_pola, jumlah_temuan_mencurigakan
            )
            system_prompt = ak.buat_system_prompt_akuntansi() + konteks + info_esb + konteks_data_kpi
            riwayat_untuk_ai = [{"role": p.role, "content": p.content} for p in (req.riwayat or [])]
            yield _format_sse_progress(
                type="progress", step="susun_prompt",
                label="Menyusun konteks jawaban",
                status="done",
            )

            # [BARU] Step 3 -- panggil AI & mulai stream token jawaban.
            # Event "processing" dikirim SEKALI di awal (bukan per-token);
            # frontend menandai step ini "done" begitu token PERTAMA
            # diterima (lihat catatan di ChatPage.jsx).
            yield _format_sse_progress(
                type="progress", step="jawab",
                label="Menyusun jawaban",
                status="processing",
            )
            for potongan in ak.tanya_ai_stream(
                req.pesan, system_prompt=system_prompt, riwayat=riwayat_untuk_ai
            ):
                jawaban_lengkap.append(potongan)
                yield _format_sse(potongan)
        except Exception as e:
            yield f"data: {json.dumps({'error': str(e)})}\n\n"
        finally:
            # [BARU] Simpan balasan AI (sepanjang apa pun yang berhasil
            # terkirim sebelum error, kalau ada) ke database.
            if req.percakapan_id is not None and jawaban_lengkap:
                dbc.simpan_pesan_chat(req.percakapan_id, "assistant", "".join(jawaban_lengkap))
            yield "data: [DONE]\n\n"

    return StreamingResponse(event_generator(), media_type="text/event-stream")


def _bersihkan_untuk_json(nilai: Any) -> Any:
    if isinstance(nilai, pd.DataFrame):
        return _bersihkan_untuk_json(nilai.to_dict(orient="records"))
    if isinstance(nilai, dict):
        return {str(k): _bersihkan_untuk_json(v) for k, v in nilai.items()}
    if isinstance(nilai, (list, tuple)):
        return [_bersihkan_untuk_json(v) for v in nilai]
    if isinstance(nilai, (pd.Timestamp, datetime, date)):
        return str(nilai)
    if isinstance(nilai, float) and (math.isnan(nilai) or math.isinf(nilai)):
        return None
    if isinstance(nilai, pd.Series):
        return _bersihkan_untuk_json(nilai.tolist())
    try:
        import numpy as np
        if isinstance(nilai, np.generic):
            return nilai.item()
    except ImportError:
        pass
    if pd.isna(nilai) if not isinstance(nilai, (list, dict)) else False:
        return None
    return nilai


def _ada_isi(hasil: dict) -> bool:
    df_hasil = hasil.get("df")
    if df_hasil is not None and not df_hasil.empty:
        return True
    per_sheet = hasil.get("per_sheet")
    if per_sheet:
        return True
    return False


def _buat_excel_hasil(hasil_per_jenis: dict) -> Path:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for kode, hasil in hasil_per_jenis.items():
        label = _PEMROSES_DOKUMEN.get(kode, (kode, None))[0]

        ws_ringkasan = wb.create_sheet(f"Ringkasan-{kode}"[:31])
        ws_ringkasan.append(["Ringkasan", label])
        for k, v in (hasil.get("ringkasan") or {}).items():
            ws_ringkasan.append([k, str(v)])

        masalah = hasil.get("masalah") or []
        ws_masalah = wb.create_sheet(f"Perlu Review-{kode}"[:31])
        # [FIX] kolom "Rekomendasi" ditambahkan -- proses_aset_tetap() (dan
        # jenis dokumen lain ke depannya) mengisi "rekomendasi" per baris
        # masalah (saran tindakan konkret, bukan cuma alasan), tapi sheet
        # export ini sebelumnya cuma baca "alasan" -- rekomendasinya jadi
        # tidak pernah sampai ke file Excel yang didownload staf/client,
        # padahal itu justru bagian paling berguna buat akuntan.
        ws_masalah.append(["Baris", "Alasan", "Rekomendasi"])
        for m in masalah:
            baris = m.get("baris")
            alasan_gabungan = " | ".join(m.get("alasan", []))
            rekomendasi_gabungan = " | ".join(m.get("rekomendasi", []))
            ws_masalah.append([baris, alasan_gabungan, rekomendasi_gabungan])

        draf = hasil.get("draf_jurnal") or []
        if draf:
            ws_jurnal = wb.create_sheet(f"Draf Jurnal-{kode}"[:31])
            kolom = list(draf[0].keys())
            ws_jurnal.append(kolom)
            for baris_jurnal in draf:
                ws_jurnal.append([str(baris_jurnal.get(k, "")) for k in kolom])

        # [BARU] Sheet tambahan generik -- dibaca kalau proses_xxx() menyediakan
        # key ini di hasil (saat ini baru proses_aset_tetap yang mengisi).
        # Ditulis generik (bukan khusus "if kode == 'aset_tetap'") supaya
        # jenis dokumen lain otomatis ikut ter-export kalau nanti diisi juga.
        rekon_fiskal = hasil.get("rekonsiliasi_fiskal") or []
        if rekon_fiskal:
            ws_fiskal = wb.create_sheet(f"Rekon Fiskal-{kode}"[:31])
            kolom = list(rekon_fiskal[0].keys())
            ws_fiskal.append(kolom)
            for baris_fiskal in rekon_fiskal:
                ws_fiskal.append([str(baris_fiskal.get(k, "")) for k in kolom])

        di_bawah_kapitalisasi = hasil.get("aset_di_bawah_batas_kapitalisasi") or []
        if di_bawah_kapitalisasi:
            ws_kap = wb.create_sheet(f"Batas Kapitalisasi-{kode}"[:31])
            kolom = list(di_bawah_kapitalisasi[0].keys())
            ws_kap.append(kolom)
            for baris_kap in di_bawah_kapitalisasi:
                ws_kap.append([str(baris_kap.get(k, "")) for k in kolom])

    nama_unik = f"{uuid.uuid4().hex}.xlsx"
    path_file = FOLDER_HASIL / nama_unik
    wb.save(path_file)
    return path_file


_PEMROSES_DOKUMEN = {
    # [FIX] 4 jenis baru -- sebelumnya cuma bisa jalan lewat app.py (Streamlit),
    # sekarang dibungkus proses_file_xxx() yang seragam di akuntansi_ai.py
    # supaya bisa dipanggil dari React/n8n lewat FastAPI juga.
    "rekening_koran": ("Rekening Koran / Mutasi Bank (Jurnal Koran)", ak.proses_file_rekening_koran),
    "penjualan": ("Data Penjualan (Invoice & POS/Kasir)", ak.proses_file_penjualan),
    "penilaian_klien": ("Penilaian Klien/Maker", ak.proses_file_penilaian_klien),
    "buku_bantu_piutang": ("Buku Bantu Piutang (AR)", ak.proses_file_piutang),
    "laporan_keuangan": ("Laporan Keuangan Lengkap (31 Sheet)", ak.proses_file_laporan_keuangan),
    "faktur_pajak": ("Faktur Pajak (PPN)", ak.proses_file_faktur_pajak),
    "bukti_potong_pajak": ("Bukti Potong Pajak (PPh 21/23/4(2))", ak.proses_file_bukti_potong),
    "spt_masa": ("SPT Masa/Tahunan", ak.proses_file_spt),
    "slip_gaji": ("Slip Gaji Karyawan", ak.proses_file_slip_gaji),
    "bukti_kas": ("Bukti Kas Masuk/Keluar", ak.proses_file_bukti_kas),
    "kartu_stok": ("Kartu Stok", ak.proses_file_kartu_stok),
    "aset_tetap": ("Aset Tetap", ak.proses_file_aset_tetap),
    "pembelian": ("Pembelian", ak.proses_file_pembelian),
    "rekonsiliasi_bank": ("Rekonsiliasi Bank", ak.proses_file_rekonsiliasi_bank),
    "ap_aging": ("AP Aging (Utang Jatuh Tempo)", ak.proses_file_ap_aging),
    "absensi": ("Absensi Karyawan", ak.proses_file_absensi),
}


@app.get("/api/jenis-dokumen")
def daftar_jenis_dokumen():
    return {
        "jenis_didukung": [
            {"kode": kode, "label": label} for kode, (label, _fn) in _PEMROSES_DOKUMEN.items()
        ]
    }


@app.post("/api/deteksi-file")
async def deteksi_file(file: UploadFile = File(...), user: dict = Depends(auth.get_current_user)):
    try:
        isi = await file.read()
        buf = io.BytesIO(isi)
        buf.name = file.filename or "upload.xlsx"
        # [FIX -- GAP EVENT LOOP] ak.deteksi_semua_sheet() sync & CPU-bound
        # (parsing openpyxl penuh, bisa ribuan baris) -- dipanggil langsung
        # di sini akan MEMBLOKIR seluruh event loop FastAPI selama proses
        # berjalan, sehingga SEMUA user lain (bukan cuma yang deteksi file
        # ini) ikut menunggu request apa pun ke server. asyncio.to_thread()
        # melempar eksekusinya ke thread pool terpisah supaya event loop
        # tetap bebas melayani request lain. Pola identik dgn
        # /api/ai-baca-file (lihat modules/ai_file_reader.py).
        hasil = await asyncio.to_thread(ak.deteksi_semua_sheet, buf, file.filename or "upload.xlsx")
        return {"nama_file": file.filename, "terdeteksi": _bersihkan_untuk_json(hasil)}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Gagal mendeteksi file: {e}") from e


# ============================================================
# [BARU] AI FILE READER -- kirim file (teks/gambar/PDF) LANGSUNG ke Claude
# API, tanpa parsing manual (pdfplumber/pytesseract/dll) di sisi backend.
# Lihat modules/ai_file_reader.py untuk detail & batasan (ukuran file,
# format didukung). Endpoint ini TERPISAH dari pipeline akuntansi_ai.py/
# kertas_kerja.py yang sudah ada -- tidak menggantikan atau mengubah
# apa pun di pipeline itu, murni fitur baru berdiri sendiri.
# ============================================================


def _cek_content_length_awal(request: Request) -> None:
    """
    [BARU] Pagar PERTAMA sebelum `await file.read()`/`await f.read()`
    dipanggil -- baca header `Content-Length` (ukuran TOTAL body request,
    termasuk overhead multipart/boundary utk endpoint banyak-file) dan
    tolak SEGERA (413) kalau sudah jelas jauh melebihi limit terbesar
    yang kita punya (lihat ai_file_reader.MAX_UPLOAD_BYTES_PRACHECK) --
    TANPA membaca body request ke memory sama sekali.

    Ini BUKAN pengganti validasi per-file yang sudah ada (_validasi_pdf/
    _validasi_gambar/_validasi_office di ai_file_reader.py, yang jalan
    SETELAH tipe file diketahui & isinya sudah di-read) -- itu tetap jalan
    seperti biasa & lebih presisi (per jenis file). Ini cuma penjaga awal
    murah utk kasus upload yang SUDAH PASTI kelewat besar (mis. client
    salah pilih file 500MB) -- tanpa pagar ini, server tetap akan
    membaca 500MB itu penuh ke memory dulu sebelum akhirnya menolak,
    yang boros & rawan dipakai utk isi memory server (bentuk DoS ringan)
    kalau banyak upload besar dikirim bersamaan.

    Kalau header Content-Length tidak ada (klien tidak mengirimnya, atau
    pakai chunked transfer) -- SENGAJA dilewati (bukan ditolak), karena
    tidak semua klien HTTP mengirim header ini; validasi presisi per-file
    setelah baca tetap jadi penjaga akhir yang sebenarnya.
    """
    content_length = request.headers.get("content-length")
    if content_length is None:
        return
    try:
        ukuran = int(content_length)
    except ValueError:
        return
    if ukuran > ai_file_reader.MAX_UPLOAD_BYTES_PRACHECK:
        raise HTTPException(
            status_code=413,
            detail=(
                f"Ukuran upload ({ukuran / (1024*1024):.1f} MB) melebihi batas "
                f"{ai_file_reader.MAX_UPLOAD_BYTES_PRACHECK // (1024*1024)} MB -- "
                "ditolak sebelum file dibaca ke memory. Pecah file jadi beberapa "
                "bagian lebih kecil."
            ),
        )

@app.post("/api/ai-baca-file")
async def api_ai_baca_file(
    request: Request,
    file: UploadFile = File(...),
    pertanyaan: str = Form(...),
    user: dict = Depends(auth.get_current_user),
):
    """
    [BARU] Upload 1 file (teks: .md/.txt/.csv/.json/.html, gambar:
    .png/.jpg/.jpeg/.gif/.webp, PDF: .pdf, atau Office: .xlsx/.xlsm/.docx/
    .xls/.doc -- lihat modules/ai_file_reader.py) + pertanyaan bebas.
    Untuk teks/gambar/PDF: file diteruskan APA ADANYA ke Claude API (base64
    utk gambar/PDF, teks polos utk teks), Claude yang membaca langsung.
    Untuk Office (xlsx/xlsm/docx/xls/doc): file TIDAK bisa dikirim mentah
    ke Claude API (bukan format yang didukung), jadi diparsing dulu secara
    LOKAL (openpyxl/python-docx utk format modern, xlrd/textract utk .xls/
    .doc lama) -- hasil ekstraksinya (bukan file binernya) yang dikirim ke
    Claude. Hasil parsing Office di-cache di disk per hash isi file, jadi
    upload ulang file yang sama tidak diparsing dari nol lagi.

    Args (multipart/form-data):
        file: 1 file, maks ukuran mengikuti batas Anthropic API (PDF
            32MB/100 halaman, gambar 5MB) utk teks/gambar/PDF, atau limit
            kita sendiri 50MB utk file Office (lihat modules/ai_file_reader.py).
        pertanyaan: instruksi bebas, mis. "Ekstrak semua transaksi jadi
            tabel" atau "Ringkas dokumen ini".

    Returns JSON: {"nama_file": str, "jawaban": str}

    Error:
        400 kalau tipe file tidak didukung (lihat
            ai_file_reader.deteksi_tipe_file).
        500 kalau panggilan ke Claude API gagal (mis. API key belum
            diset, rate limit, network error).
    """
    nama_file = file.filename or "upload"
    _cek_content_length_awal(request)
    isi = await file.read()

    tipe = ai_file_reader.deteksi_tipe_file(nama_file)
    if tipe == "tidak_didukung":
        raise HTTPException(
            status_code=400,
            detail=(
                f"Tipe file '{Path(nama_file).suffix}' tidak didukung oleh AI File Reader. "
                "Didukung: .md/.txt/.csv/.json/.html (teks), "
                ".png/.jpg/.jpeg/.gif/.webp (gambar), .pdf (dokumen), "
                ".xlsx/.xlsm/.docx/.pptx (Office modern), .xls/.doc (Office lama)."
            ),
        )

    try:
        # [FIX -- GAP EVENT LOOP] kirim_file_ke_ai() SYNC & CPU/network-bound
        # (parsing openpyxl/pandas + panggilan blocking ke Anthropic API) --
        # dipanggil langsung di sini akan MEMBLOKIR seluruh event loop
        # FastAPI selama proses berjalan, sehingga SEMUA user lain (bukan
        # cuma yang upload file) ikut menunggu request apa pun ke server.
        # asyncio.to_thread() melempar eksekusinya ke thread pool terpisah
        # supaya event loop tetap bebas melayani request lain secara
        # bersamaan. Exception dari dalam fungsi tetap diteruskan apa
        # adanya (ValueError/RuntimeError/dll) sehingga blok except di
        # bawah ini tidak perlu berubah.
        jawaban = await asyncio.to_thread(
            ai_file_reader.kirim_file_ke_ai, isi, nama_file, pertanyaan
        )
    except ValueError as e:
        # [FIX -- GAP] File melanggar limit resmi Claude API (PDF >32MB/
        # >100 halaman, gambar >5MB) -- ini kesalahan INPUT user, bukan
        # error server, jadi 400 dengan pesan jelas (lihat validasi di
        # ai_file_reader._validasi_pdf/_validasi_gambar), bukan 500 generik.
        raise HTTPException(status_code=400, detail=str(e))
    except RuntimeError as e:
        # ANTHROPIC_API_KEY belum diset -- lihat _ambil_client() di ai_file_reader.py
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:  # noqa: BLE001 -- mis. error dari Anthropic API (rate limit, dll)
        raise HTTPException(status_code=500, detail=f"Gagal memproses file lewat AI: {e}")

    return {"nama_file": nama_file, "jawaban": jawaban}


# [BARU] Versi STREAMING dari /api/ai-baca-file -- jawaban dikirim
# bertahap (potongan teks) begitu diterima dari Claude API, bukan
# ditunggu sampai selesai penuh baru dikembalikan sekaligus. Berguna
# utk pertanyaan panjang/kompleks di atas file besar, supaya UI bisa
# menampilkan progress "mengetik" (mirip Claude AI sendiri) daripada
# spinner diam yang bisa terasa lama tanpa kepastian.
@app.post("/api/ai-baca-file-stream")
async def api_ai_baca_file_stream(
    request: Request,
    file: UploadFile = File(...),
    pertanyaan: str = Form(...),
    user: dict = Depends(auth.get_current_user),
):
    """
    Sama seperti /api/ai-baca-file (format file didukung sama persis),
    tapi response berupa STREAM teks polos (`text/plain`, chunked) --
    frontend baca `response.body` sebagai ReadableStream & tampilkan tiap
    potongan begitu tiba, bukan `await response.json()` yang nunggu
    semuanya kelar.

    [PENTING -- urutan validasi] Ekstraksi file (kalau Office) & validasi
    ukuran/tipe file dilakukan DI SINI, SEBELUM StreamingResponse dibuka --
    lihat catatan di ai_file_reader.siapkan_konten_pesan_dari_file(). Ini
    supaya file tidak valid tetap balik HTTPException 400/500 yang benar
    (bukan status 200 yang sudah kadung terkirim lalu stream putus di
    tengah tanpa status code yang masuk akal).

    Errors:
        400 kalau tipe file tidak didukung / melanggar limit ukuran.
        413 kalau Content-Length request sudah jelas kelewat besar
            (lihat _cek_content_length_awal) -- ditolak sebelum file
            dibaca ke memory sama sekali.
        500 kalau API key belum diset. Error yang terjadi SETELAH stream
            mulai jalan (rate limit/overload di tengah) TIDAK bisa lagi
            jadi HTTPException (status 200 sudah terkirim) -- stream akan
            terhenti begitu saja; frontend yang harus deteksi ini (sama
            pola dgn KertasKerjaPage.jsx: "stream berakhir tanpa selesai"
            -> tampilkan pesan error, minta coba lagi).
    """
    nama_file = file.filename or "upload"
    _cek_content_length_awal(request)
    isi = await file.read()

    tipe = ai_file_reader.deteksi_tipe_file(nama_file)
    if tipe == "tidak_didukung":
        raise HTTPException(
            status_code=400,
            detail=(
                f"Tipe file '{Path(nama_file).suffix}' tidak didukung oleh AI File Reader. "
                "Didukung: .md/.txt/.csv/.json/.html (teks), "
                ".png/.jpg/.jpeg/.gif/.webp (gambar), .pdf (dokumen), "
                ".xlsx/.xlsm/.docx/.pptx (Office modern), .xls/.doc (Office lama)."
            ),
        )

    try:
        # [PENTING] Disiapkan (termasuk ekstraksi Office & validasi ukuran)
        # SEBELUM StreamingResponse dibuka -- lihat docstring di atas.
        # [FIX -- GAP EVENT LOOP] sama seperti /api/ai-baca-file: ini fungsi
        # sync CPU-bound (openpyxl/python-docx/pandas) -- dilempar ke
        # thread pool via asyncio.to_thread() supaya tidak memblokir event
        # loop selama parsing file Office besar berlangsung.
        content = await asyncio.to_thread(
            ai_file_reader.siapkan_konten_pesan_dari_file, isi, nama_file, pertanyaan
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except RuntimeError as e:
        # mis. library python-pptx/xlrd/textract belum terinstall
        raise HTTPException(status_code=500, detail=str(e))

    def _generator_teks():
        try:
            # [FIX -- SEMENTARA] Routing Claude/Groq sekarang otomatis
            # berdasar tipe `content` (string vs list) di dalam
            # kirim_file_ke_ai_stream itu sendiri -- lihat docstring-nya.
            yield from ai_file_reader.kirim_file_ke_ai_stream(content)
        except Exception as e:  # noqa: BLE001 -- lihat catatan error di docstring
            ai_file_reader.logger.error(f"❌ Stream ai-baca-file terhenti di tengah untuk '{nama_file}': {e}")
            # Tetap yield pesan error sebagai teks biasa di akhir stream --
            # lebih baik daripada koneksi terputus tanpa penjelasan sama
            # sekali (frontend tetap bisa tampilkan potongan ini ke user).
            yield f"\n\n[Terputus: {e}]"

    return StreamingResponse(_generator_teks(), media_type="text/plain; charset=utf-8")


# [BARU] Versi banyak file dalam 1 pertanyaan (mis. "bandingkan file A dan
# B", "rekap semua rekening koran ini") -- lihat
# ai_file_reader.kirim_banyak_file_ke_ai untuk alasan & batasan lengkap
# (limit gambar per request, limit payload gabungan 32MB).
@app.post("/api/ai-baca-banyak-file")
async def api_ai_baca_banyak_file(
    request: Request,
    files: List[UploadFile] = File(...),
    pertanyaan: str = Form(...),
    user: dict = Depends(auth.get_current_user),
):
    """
    Upload BEBERAPA file (boleh campur teks/gambar/PDF/Office) + 1
    pertanyaan yang berlaku untuk semua file itu sekaligus -- berguna
    untuk kasus lintas-file yang tidak bisa dijawab dari 1 file saja
    (mis. "file mana yang datanya beda dengan yang lain", "gabungkan
    semua transaksi berikut", "bandingkan sheet penjualan bulan ini vs
    bulan lalu"). File Office (xlsx/xlsm/docx/xls/doc) diparsing lokal
    dulu (sama seperti /api/ai-baca-file, hasilnya di-cache) -- hasil
    ekstraksinya digabung ke prompt teks bareng file teks lain.

    Returns JSON: {"nama_file": [str, ...], "jawaban": str}

    Error:
        400 kalau ada file dengan tipe tidak didukung, atau ukuran/jumlah
            file melanggar limit Claude API (lihat
            ai_file_reader.kirim_banyak_file_ke_ai).
        500 kalau panggilan ke Claude API gagal.
    """
    if not files:
        raise HTTPException(status_code=400, detail="Tidak ada file yang diupload.")

    _cek_content_length_awal(request)

    daftar_file: List[Tuple[bytes, str]] = []
    for f in files:
        nama_file = f.filename or "upload"
        isi = await f.read()
        daftar_file.append((isi, nama_file))

    try:
        # [FIX -- GAP EVENT LOOP] sama alasannya dgn /api/ai-baca-file:
        # kirim_banyak_file_ke_ai() sync, di dalamnya ada ThreadPoolExecutor
        # sendiri utk paralel ANTAR file -- tapi pemanggilan fungsi ini
        # SENDIRI ke thread tsb tetap BLOCKING dari sudut pandang event
        # loop (thread pemanggil menunggu executor.map selesai). to_thread()
        # di sini melempar seluruh pekerjaan (termasuk nunggu ThreadPoolExecutor
        # internal) ke luar event loop.
        jawaban = await asyncio.to_thread(
            ai_file_reader.kirim_banyak_file_ke_ai, daftar_file, pertanyaan
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Gagal memproses file lewat AI: {e}")

    return {"nama_file": [nama for _, nama in daftar_file], "jawaban": jawaban}


# [FIX] Kode jenis dokumen yang fungsinya mendukung pembelajaran pola
# per-client (client_id diteruskan supaya pola tersimpan terpisah per client,
# bukan tercampur jadi satu pola global).
_JENIS_DENGAN_POLA_PER_CLIENT = {"rekening_koran", "penjualan"}

# [BARU] Nama dasar file pola per jenis dokumen (dipakai ak._path_pola()) --
# sama persis dengan yang dipakai di dalam ak.proses_file_rekening_koran /
# ak.proses_file_penjualan, supaya alert anomali di bawah membaca pola yang
# BARU SAJA dipelajari & disimpan dari upload file ini.
_POLA_PER_JENIS = {
    "rekening_koran": "pola_bank",
    "penjualan": "pola_penjualan",
}


def _coba_simpan_sebagai_coa(isi: bytes, nama_file: str, client_id: Optional[int]) -> Optional[dict]:
    """
    [BARU] Coba baca file sbg sheet COA (ak.muat_coa) & langsung simpan ke
    COA permanen client (dbc.simpan_coa_bulk) -- supaya file COA yang
    diupload lewat KOTAK CHAT (bukan lewat panel terpisah) tetap tersimpan,
    bukan berakhir "tidak dikenali".

    COA beda dari 15 jenis dokumen transaksi di _PEMROSES_DOKUMEN: dia
    bukan dokumen yang menghasilkan draf jurnal, cuma data referensi
    (dipakai buat cross-check akun di dokumen LAIN) -- makanya ditangani
    terpisah di sini, DICOBA DULUAN sebelum file dicoba ke semua
    proses_file_xxx yang lain.

    Return None kalau file ini TIDAK mengandung sheet COA (berarti bukan
    file COA -- lanjut ke alur deteksi 15 jenis dokumen seperti biasa).
    Return dict kalau file ini COA -- baik berhasil disimpan, maupun gagal
    karena alasan spesifik ke COA (mis. belum ada client aktif), supaya
    pesan errornya jelas & tertuju, bukan "tidak dikenali" yang membingungkan.
    """
    try:
        buf = io.BytesIO(isi)
        wb = openpyxl.load_workbook(buf, data_only=True, read_only=True)
    except Exception:
        return None

    try:
        df_coa = ak.muat_coa(wb)
    finally:
        try:
            wb.close()
        except Exception:
            pass

    if df_coa.empty:
        return None  # tidak ada sheet COA di file ini -- bukan urusan fungsi ini

    if client_id is None:
        return {
            "sukses": False,
            "pesan": "File ini terdeteksi sebagai COA, tapi belum ada client aktif -- "
                     "pilih client dulu (dropdown 'Pilih Klien' di kiri atas), lalu upload ulang.",
        }

    akun = [
        {
            "no_akun": str(row["no_akun"]).strip(),
            "nama_akun": str(row["nama_akun"]).strip(),
            "kategori": (str(row["kategori"]).strip()
                         if row.get("kategori") not in (None, "") else None),
            "sub_kategori": None,
            "normal_saldo": None,
            "saldo_awal": 0,
        }
        for _, row in df_coa.iterrows()
        if str(row.get("no_akun") or "").strip() and str(row.get("nama_akun") or "").strip()
    ]
    if not akun:
        return None

    # ganti_semua=False -- upload cepat lewat chat cuma NAMBAH/PERBARUI akun
    # yang ada di file ini, TIDAK menghapus akun COA lain yang sudah
    # tersimpan.
    jumlah = dbc.simpan_coa_bulk(client_id, akun, ganti_semua=False)
    return {
        "sukses": True,
        "jumlah_akun_tersimpan": jumlah,
        "pesan": f"{jumlah} akun COA berhasil disimpan/diperbarui untuk client ini.",
    }


def _proses_semua_jenis(
    isi: bytes,
    nama_file: str,
    jenis_dokumen: Optional[str],
    client_id: Optional[int] = None,
    on_progress: Optional[Callable[..., None]] = None,
    pakai_ai: bool = True,
):
    """
    [FIX] Tambah parameter on_progress opsional (default None) -- dipanggil
    tiap kali satu jenis dokumen SELESAI dicoba (berhasil/gagal/dilewati),
    dengan signature on_progress(kode, label, status, **extra). Dipakai
    oleh /api/proses-file/stream utk kirim event SSE per jenis dokumen;
    pemanggil lama (/api/proses-file, /api/proses-dan-buat-excel) tidak
    perlu berubah sama sekali krn default-nya None (tidak melapor apa-apa,
    perilaku identik dengan sebelum fitur ini ada).
    """
    def _lapor(kode, label, status, **extra):
        if on_progress is not None:
            try:
                on_progress(kode, label, status, **extra)
            except Exception as e:  # noqa: BLE001
                # Progress reporting TIDAK BOLEH menggagalkan pemrosesan
                # file itu sendiri -- kalau gagal lapor progress, cukup
                # dicatat di terminal, jangan sampai melempar exception ke
                # atas dan membatalkan proses file yg sedang berjalan.
                print(f"[PERINGATAN] Gagal lapor progress ({kode}): {e}")

    def _buat_buffer():
        b = io.BytesIO(isi)
        b.name = nama_file
        return b

    def _panggil(fungsi, kode):
        if kode in _JENIS_DENGAN_POLA_PER_CLIENT:
            # [BARU] pakai_ai diteruskan ke rekening_koran & penjualan --
            # dua-duanya sama-sama terima kwarg ini (lihat
            # ak.proses_file_rekening_koran / ak.proses_file_penjualan).
            return fungsi(_buat_buffer(), nama_file, client_id=client_id, pakai_ai=pakai_ai)
        if kode == "slip_gaji":
            # [FIX] Sebelumnya histori_gaji_sebelumnya/histori_gaji_terbaru
            # (utk deteksi anomali gaji ANTAR-PERIODE lintas upload bulan
            # berbeda -- lihat docstring ak.proses_slip_gaji()) tidak pernah
            # disambungkan di sini, jadi anomali antar-bulan hanya kedeteksi
            # kalau semua bulan kebetulan ada dalam SATU file/upload yang
            # sama. Sekarang histori per-client dibaca sebelum proses, lalu
            # digabung & disimpan lagi setelah proses -- persis pola yang
            # sudah dipakai utk pola_bank/pola_penjualan di atas. Kalau
            # client_id None (mis. proses tanpa konteks client), histori
            # jatuh ke file global bersama (ak._path_pola sudah handle ini),
            # tetap tidak error, cuma tidak per-client.
            path_histori = ak._path_pola("histori_gaji", client_id)
            histori_lama = ak.muat_histori_gaji(path_histori)
            hasil = fungsi(_buat_buffer(), nama_file, histori_gaji_sebelumnya=histori_lama or None)
            histori_baru = hasil.get("histori_gaji_terbaru") or {}
            if histori_baru:
                # Gabung (bukan timpa total) -- karyawan yang TIDAK ada di
                # file bulan ini (mis. resign/belum digaji bulan ini) tetap
                # mempertahankan histori terakhirnya, bukan hilang.
                histori_gabungan = {**histori_lama, **histori_baru}
                try:
                    ak.simpan_histori_gaji(histori_gabungan, path_histori)
                except OSError as e:  # noqa: BLE001
                    # Kegagalan simpan histori TIDAK BOLEH menggagalkan hasil
                    # proses slip gaji yang sudah berhasil dihitung -- cukup
                    # dicatat, anomali antar-periode bulan depan saja yang
                    # kena dampak (reset), bukan proses upload ini.
                    print(f"[PERINGATAN] Gagal simpan histori gaji ke {path_histori}: {e}")
            return hasil
        return fungsi(_buat_buffer(), nama_file)

    hasil_semua: dict = {}
    error_per_jenis: dict = {}

    # [BARU] Cek dulu apakah ini file COA -- lihat penjelasan lengkap di
    # _coba_simpan_sebagai_coa(). Cuma dicek kalau user TIDAK secara
    # eksplisit minta jenis_dokumen lain (mis. upload lewat panel khusus
    # yang sudah menentukan jenisnya sendiri).
    if not jenis_dokumen:
        hasil_coa = _coba_simpan_sebagai_coa(isi, nama_file, client_id)
        if hasil_coa is not None:
            if hasil_coa["sukses"]:
                _lapor("coa", "Chart of Accounts (COA)", "done", pesan=hasil_coa["pesan"])
                return {"coa": hasil_coa}, {}
            else:
                _lapor("coa", "Chart of Accounts (COA)", "error", pesan=hasil_coa["pesan"])
                return {}, {"coa": hasil_coa["pesan"]}

    if jenis_dokumen:
        if jenis_dokumen not in _PEMROSES_DOKUMEN:
            raise HTTPException(
                status_code=400,
                detail=f"jenis_dokumen '{jenis_dokumen}' tidak dikenali. "
                       f"Lihat /api/jenis-dokumen untuk daftar yang valid.",
            )
        label, fungsi = _PEMROSES_DOKUMEN[jenis_dokumen]
        _lapor(jenis_dokumen, label, "processing")
        try:
            hasil_semua[jenis_dokumen] = _panggil(fungsi, jenis_dokumen)
            _lapor(jenis_dokumen, label, "done")
        except Exception as e:  # noqa: BLE001
            _lapor(jenis_dokumen, label, "error", pesan=str(e))
            raise HTTPException(
                status_code=500, detail=f"Gagal memproses sebagai {label}: {e}"
            ) from e
    else:
        for kode, (label, fungsi) in _PEMROSES_DOKUMEN.items():
            _lapor(kode, label, "processing")
            try:
                hasil = _panggil(fungsi, kode)
                if _ada_isi(hasil):
                    hasil_semua[kode] = hasil
                    _lapor(kode, label, "done")
                else:
                    _lapor(kode, label, "skip")
            except Exception as e:  # noqa: BLE001
                error_per_jenis[kode] = f"{type(e).__name__}: {e}"
                _lapor(kode, label, "skip", pesan=str(e))

    return hasil_semua, error_per_jenis


def _proses_dan_simpan_satu_file(
    isi: bytes,
    nama_file: str,
    jenis_dokumen: Optional[str],
    client_id: Optional[int],
    conv_id: Optional[str],
    esb_account_id: Optional[int],
    konfirmasi_duplikat: bool,
    user: dict,
    pakai_ai: bool = True,
) -> dict:
    """
    [REFACTOR - batch upload] Badan asli /api/proses-file (deteksi -> simpan
    hasil -> tarik draf jurnal ke posting -> reminder SPT -> klarifikasi ->
    alert anomali) DIPINDAH ke sini apa adanya, supaya bisa dipanggil BERKALI-
    KALI dari /api/client/{id}/proses-file-batch (satu file, satu panggilan)
    TANPA duplikasi logic. /api/proses-file sendiri sekarang tinggal baca
    file lalu delegasikan ke fungsi ini -- perilakunya untuk 1 file identik
    dengan sebelum refactor ini (tidak ada logic yang diubah, cuma dipindah).

    pakai_ai default True (perilaku lama) -- diteruskan apa adanya ke
    _proses_semua_jenis, lihat catatan di endpoint /api/proses-file.
    """
    hasil_semua, error_per_jenis = _proses_semua_jenis(isi, nama_file, jenis_dokumen, client_id, pakai_ai=pakai_ai)

    if not hasil_semua:
        return {
            "nama_file": nama_file,
            "hasil": {},
            "tidak_terdeteksi": True,
            "pesan": "Tidak ada jenis dokumen yang dikenali di file ini.",
            "detail_error": error_per_jenis or None,
        }

    hasil_json = _bersihkan_untuk_json(hasil_semua)

    if client_id is not None:
        conv_id_final = conv_id or datetime.now().isoformat()
        for kode, hasil in hasil_json.items():
            data_disimpan = dict(hasil)
            data_disimpan["nama_file"] = nama_file

            # [BARU] Audit trail: kalau pemroses jenis dokumen ini melakukan
            # auto-fix (mis. proses_file_penilaian_klien -> koreksi_otomatis
            # dari perbaiki_data_penilaian), catat sebagai satu entri audit
            # per upload -- bukan cuma balik ke response lalu hilang begitu
            # tab ditutup. Detail berisi daftar lengkap perubahan per baris
            # (kolom, nilai lama -> baru, alasan) supaya bisa ditelusuri
            # siapa (via user login yg upload) - apa - kapan.
            koreksi_otomatis = hasil.get("koreksi_otomatis") or []
            if koreksi_otomatis:
                dbc.log_audit(
                    client_id=client_id,
                    user=user.get("username", "unknown"),
                    aksi="auto_fix_data",
                    detail={
                        "jenis_dokumen": kode,
                        "nama_file": nama_file,
                        "total_koreksi": len(koreksi_otomatis),
                        "koreksi": koreksi_otomatis[:200],  # batasi ukuran 1 entri log
                    },
                )

            if esb_account_id is not None:
                # Hasil spesifik 1 akun ESB -> tabel terpisah hasil_esb
                dbc.simpan_hasil_esb(client_id, esb_account_id, conv_id_final, kode, data_disimpan)
            else:
                # Hasil umum client -> tabel hasil
                dbc.simpan_hasil(client_id, conv_id_final, kode, data_disimpan)

                # [BARU] Tarik draf_jurnal (kalau ada) ke antrean review
                # jurnal_posting -- supaya muncul di layar "Perlu Posting"
                # sebelum bisa dihitung ke 5 Laporan Keuangan Standar.
                # Tidak dilakukan untuk hasil_esb (belum ada alur review
                # terpisah utk hasil per akun ESB).
                draf_jurnal = data_disimpan.get("draf_jurnal") or []
                if draf_jurnal:
                    hasil_tersimpan = dbc.ambil_hasil_client(client_id, jenis=kode, limit=1)
                    hasil_id = hasil_tersimpan[0]["id"] if hasil_tersimpan else None

                    # [BARU - dedup upload] KHUSUS rekening_koran: cek dulu
                    # apakah file/baris ini indikasi upload ulang atau
                    # revisi dari upload sebelumnya, SEBELUM ditarik ke
                    # jurnal_posting -- supaya tidak dobel hitung & tidak
                    # membakar nomor voucher baru utk transaksi yang
                    # sebenarnya sudah pernah masuk sistem.
                    if kode == "rekening_koran":
                        file_hash = dedup_transaksi.hitung_file_hash(isi)
                        evaluasi = dedup_transaksi.evaluasi_upload_rekening_koran(
                            client_id=client_id, draf_jurnal=draf_jurnal, file_hash=file_hash,
                        )

                        if evaluasi.perlu_konfirmasi and not konfirmasi_duplikat:
                            # DITAHAN -- jangan tarik ke posting sama sekali,
                            # simpan snapshot supaya bisa ditarik belakangan
                            # kalau akuntan konfirmasi lewat endpoint terpisah.
                            for kel in evaluasi.kelompok:
                                dbc.catat_upload_batch(
                                    client_id=client_id, kode_bank=kel.kode_bank, periode=kel.periode,
                                    status="menunggu_konfirmasi", hasil_id=hasil_id, nama_file=nama_file,
                                    file_hash=file_hash, jumlah_baris_total=kel.jumlah_baris_total,
                                    jumlah_baris_baru=kel.jumlah_baris_baru,
                                    jumlah_baris_overlap=kel.jumlah_baris_overlap,
                                    status_deteksi=kel.status,
                                    draf_jurnal=[b for b in draf_jurnal if
                                                 dedup_transaksi._kode_bank_dari_nama(b.get("bank") or "BANK") == kel.kode_bank
                                                 and dedup_transaksi._periode_dari_tanggal(b.get("tanggal")) == kel.periode],
                                    diupload_oleh=user.get("username", "unknown"),
                                )
                            hasil_json[kode]["draf_jurnal"] = dedup_transaksi.hapus_kolom_internal(draf_jurnal)
                            hasil_json[kode]["duplikat"] = evaluasi.to_dict()
                            # TIDAK dipanggil: dbc.tarik_draf_jurnal_ke_posting(...)
                            continue

                        # BARU, atau sudah eksplisit dikonfirmasi lanjut --
                        # tarik seperti biasa, lalu catat batch sbg 'aktif'
                        # (jadi acuan pembanding utk upload berikutnya).
                        #
                        # [CATATAN] Kelompok (bank, periode) dihitung ULANG
                        # di sini via kelompokkan_draf_jurnal() alih-alih
                        # memakai evaluasi.kelompok langsung, karena utk
                        # kasus FILE_IDENTIK yang tetap dikonfirmasi lanjut
                        # (konfirmasi_duplikat=True), evaluasi.kelompok
                        # sengaja KOSONG (evaluasi_upload_rekening_koran
                        # return awal sebelum sempat menghitung per baris --
                        # lihat "Lapis 1" di dedup_transaksi.py). Dengan
                        # dihitung ulang di sini, kasus itu tetap tercatat
                        # rapi sbg batch 'aktif' baru (hanya overlap stats-
                        # nya tidak terisi, krn memang tidak dihitung ulang).
                        dbc.tarik_draf_jurnal_ke_posting(client_id, hasil_id, kode, draf_jurnal)
                        info_kelompok = {(k.kode_bank, k.periode): k for k in evaluasi.kelompok}
                        for (kb, per), baris_list in dedup_transaksi.kelompokkan_draf_jurnal(draf_jurnal).items():
                            kel = info_kelompok.get((kb, per))
                            batch_baru_id = dbc.catat_upload_batch(
                                client_id=client_id, kode_bank=kb, periode=per,
                                status="aktif", hasil_id=hasil_id, nama_file=nama_file, file_hash=file_hash,
                                jumlah_baris_total=len(baris_list),
                                jumlah_baris_baru=(kel.jumlah_baris_baru if kel else len(baris_list)),
                                jumlah_baris_overlap=(kel.jumlah_baris_overlap if kel else 0),
                                status_deteksi=(kel.status if kel else evaluasi.status_keseluruhan),
                                diupload_oleh=user.get("username", "unknown"),
                            )
                            batch_lama = kel.batch_sebelumnya if kel else dbc.ambil_batch_aktif(client_id, kb, per)
                            if batch_lama and konfirmasi_duplikat:
                                dbc.tandai_batch_diganti(batch_lama["id"], batch_baru_id)
                        hasil_json[kode]["duplikat"] = evaluasi.to_dict()
                    else:
                        dbc.tarik_draf_jurnal_ke_posting(client_id, hasil_id, kode, draf_jurnal)

        # [BARU] Reminder/deadline proaktif SPT: proses_spt() sudah
        # menghitung tanggal_batas_lapor/tanggal_batas_setor & status
        # sudah_lapor per baris -- ekstrak jadi kewajiban lapor/setor dan
        # simpan ke reminder_deadline_spt supaya scheduler harian (lihat
        # modules/notifikasi.py) bisa mengingatkan LEWAT WA/in-app SEBELUM
        # jatuh tempo, bukan cuma dilaporkan di response upload ini. Pakai
        # hasil_semua (bukan hasil_json) krn butuh df pandas asli dgn
        # tanggal masih berupa objek date, bukan string hasil JSON-clean.
        hasil_mentah_spt = hasil_semua.get("spt_masa")
        if hasil_mentah_spt:
            df_spt_mentah = hasil_mentah_spt.get("df")
            item_reminder = notifikasi.ekstrak_item_reminder_dari_df(df_spt_mentah)
            if item_reminder:
                dbc.simpan_reminder_deadline_spt(client_id, item_reminder)

        # [BARU] Mekanisme tanya balik ke akuntan: untuk jenis dokumen yang
        # dikategorikan lewat proses_dataframe()/proses_dataframe_penjualan()
        # (rekening koran & penjualan -- keduanya sudah punya kolom
        # sumber_kategori/confidence_ai per baris), cari baris yang AI
        # ragu/gagal kategorikan, lalu simpan sbg pertanyaan berstatus
        # "pending" supaya muncul di dashboard klarifikasi -- bukan cuma
        # terkubur diam-diam di sheet "Perlu Direview". Pakai hasil_semua
        # (bukan hasil_json) krn butuh df pandas asli, bukan yg sudah
        # di-JSON-kan oleh _bersihkan_untuk_json().
        for kode in _JENIS_DENGAN_POLA_PER_CLIENT:
            hasil_mentah = hasil_semua.get(kode)
            if not hasil_mentah:
                continue
            df_mentah = hasil_mentah.get("df")
            for item in ak.cari_baris_perlu_klarifikasi(df_mentah):
                dbc.buat_pertanyaan_klarifikasi(
                    client_id=client_id,
                    jenis=kode,
                    pertanyaan=item["pertanyaan"],
                    conv_id=conv_id_final,
                    baris_index=item["baris_index"],
                    konteks=item,
                    tebakan_kategori=item.get("tebakan_kategori"),
                    butuh_konfirmasi_saja=item.get("butuh_konfirmasi_saja", False),
                )

        # [BARU] Alert anomali ke akuntan: sama pola dengan klarifikasi di
        # atas, tapi sumbernya ak.deteksi_anomali_transaksi() (nominal
        # transaksi janggal dibanding pola historis client) & ak.deteksi_
        # pola_mencurigakan() (pasangan akun dipakai banyak pola berbeda
        # yang masing2 baru 1x muncul). Pola di-reload dari disk krn sudah
        # sempat diperbarui & disimpan di dalam ak.proses_file_xxx() di
        # atas -- sama seperti cara chat_stream() mengambil konteks pola.
        for kode in _JENIS_DENGAN_POLA_PER_CLIENT:
            hasil_mentah = hasil_semua.get(kode)
            nama_pola = _POLA_PER_JENIS.get(kode)
            if not hasil_mentah or not nama_pola:
                continue
            try:
                pola_client = ak.muat_pola(ak._path_pola(nama_pola, client_id))
            except Exception as e:  # noqa: BLE001
                print(f"[PERINGATAN] Gagal muat pola utk deteksi anomali ({kode}): {e}")
                continue

            df_mentah = hasil_mentah.get("df")
            for item in ak.cari_anomali_untuk_alert(df_mentah, pola_client):
                dbc.buat_alert_anomali(
                    client_id=client_id,
                    jenis=kode,
                    tipe_alert="nominal_ekstrim",
                    pesan=item["pesan"],
                    conv_id=conv_id_final,
                    baris_index=item["baris_index"],
                    konteks=item,
                    skor=item.get("anomaly_score"),
                )

            # [BARU] Deteksi pola mencurigakan di-dedup lewat evaluasi_pola
            # (ak.muat_evaluasi_pola/simpan_evaluasi_pola, sebelumnya sudah
            # ada di akuntansi_ai.py tapi belum pernah dipanggil dari mana
            # pun) -- supaya akuntan TIDAK dialert ulang utk pasangan akun
            # yang sama di setiap upload file berikutnya, hanya saat
            # pertama kali terdeteksi.
            temuan_mencurigakan = ak.deteksi_pola_mencurigakan(pola_client)
            if temuan_mencurigakan:
                path_evaluasi = ak._path_pola(f"evaluasi_{nama_pola}", client_id)
                temuan_benar_baru = ak.simpan_evaluasi_pola(path_evaluasi, temuan_mencurigakan)
                for temuan in temuan_benar_baru:
                    pesan = (
                        f"Akun \"{temuan['nama_akun_debet']}\" (debet) / "
                        f"\"{temuan['nama_akun_kredit']}\" (kredit) dipakai oleh "
                        f"{temuan['jumlah_signature_berbeda']} pola transaksi berbeda "
                        f"yang masing2 baru muncul 1x -- cek apakah kategorisasi ini benar."
                    )
                    dbc.buat_alert_anomali(
                        client_id=client_id,
                        jenis=kode,
                        tipe_alert="pola_mencurigakan",
                        pesan=pesan,
                        conv_id=conv_id_final,
                        konteks=temuan,
                        skor=float(temuan["jumlah_signature_berbeda"]),
                    )

    return {
        "nama_file": nama_file,
        "hasil": hasil_json,
        "tidak_terdeteksi": False,
    }


@app.post("/api/proses-file")
async def proses_file(
    file: UploadFile = File(...),
    # [FIX] Sebelumnya jenis_dokumen/client_id/conv_id/esb_account_id
    # dideklarasikan sebagai default biasa (Optional[str] = None), yang
    # membuat FastAPI menganggapnya QUERY PARAMETER (?client_id=1&...),
    # BUKAN bagian dari form-data multipart. Padahal frontend React kirim
    # semua field ini lewat FormData.append(...) bersamaan dgn file-nya --
    # akibatnya field-field ini tidak pernah kebaca oleh backend (selalu
    # None), dan kalau frontend mewajibkan salah satunya, request malah
    # bisa gagal total dengan 422 Unprocessable Entity tanpa pesan jelas
    # di UI ("tidak menampilkan apa-apa"). Form(None) memaksa FastAPI
    # membaca field ini dari body form-data yang sama dengan file.
    jenis_dokumen: Optional[str] = Form(None),
    client_id: Optional[int] = Form(None),
    conv_id: Optional[str] = Form(None),
    esb_account_id: Optional[int] = Form(None),
    # [BARU - dedup upload] Kalau True, akuntan SUDAH melihat
    # peringatan duplikat/revisi (dari respons upload sebelumnya
    # yang punya konfirmasi_diperlukan=True) dan memilih tetap
    # LANJUTKAN SEMUA baris apa adanya. Default False -- upload
    # normal (tanpa duplikat) tidak terpengaruh sama sekali.
    konfirmasi_duplikat: bool = Form(False),
    # [BARU] Toggle kategorisasi AI (Claude/Groq) untuk baris yang tidak
    # kecocokan pola historis maupun kata kunci COA. Default True (perilaku
    # lama, tidak berubah untuk pemanggil yang sudah ada). Kalau False,
    # baris begitu langsung ditandai "Belum Terkategori - perlu review
    # manual" TANPA memanggil API AI sama sekali -- berguna untuk instalasi
    # yang sengaja tidak mau bergantung ke API key pihak ketiga, atau untuk
    # upload besar supaya tidak menunggu lama/timeout menunggu banyak
    # panggilan AI berurutan (lihat _panggil_kategorisasi_dengan_fallback).
    pakai_ai: bool = Form(True),
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    isi = await file.read()
    nama_file = file.filename or "upload.xlsx"
    # [FIX -- GAP EVENT LOOP] _proses_dan_simpan_satu_file() sync & berat
    # (parsing penuh + beberapa write ke DB) -- lihat catatan identik di
    # /api/deteksi-file di atas. Dibungkus asyncio.to_thread supaya tidak
    # memblokir event loop -- endpoint ini yang PALING SERING dipakai utk
    # upload harian, jadi paling penting dibenerin duluan.
    return await asyncio.to_thread(
        _proses_dan_simpan_satu_file,
        isi, nama_file, jenis_dokumen, client_id, conv_id, esb_account_id,
        konfirmasi_duplikat, user, pakai_ai,
    )


# ============================================================
# [BARU] UPLOAD BANYAK FILE SEKALIGUS (BATCH)
# ============================================================
# Beda dari /api/proses-file (1 file): endpoint ini terima BEBERAPA file
# dalam satu request, lalu:
#   1. TAHAP DETEKSI -- tiap file discan dulu (ak.deteksi_semua_sheet(),
#      TANPA efek samping/simpan apa pun) supaya tahu jenis dokumen apa
#      yang ada di masing-masing file SEBELUM memutuskan apa-apa.
#   2. TAHAP RENCANA -- dari hasil deteksi semua file, disusun urutan
#      pemrosesan (data referensi/master seperti Aset Tetap & Piutang/
#      Utang lebih dulu, baru transaksi seperti Rekening Koran/Penjualan
#      yang bisa dicocokkan ke data referensi itu) + catatan kalau ada
#      jenis dokumen yang sama muncul di lebih dari 1 file (potensi
#      duplikat/revisi -- akuntan perlu tahu sebelum lanjut).
#   3. TAHAP EKSEKUSI -- file diproses satu-satu SESUAI URUTAN RENCANA
#      lewat _proses_dan_simpan_satu_file() (logic yang sama dengan
#      /api/proses-file, jadi hasil per file 100% konsisten) -- satu file
#      gagal TIDAK menggagalkan file lain dalam batch yang sama.
#   4. LINTAS FILE -- kalau batch ini menghasilkan Rekening Koran DAN
#      Buku Bantu Piutang sekaligus, otomatis dicocokkan
#      (cross_matching.cocokkan_bank_piutang) supaya penerimaan di bank
#      langsung ditautkan ke invoice yang sesuai, tanpa akuntan harus
#      upload ulang atau minta manual.

_URUTAN_PRIORITAS_JENIS = {
    # Angka lebih kecil = diproses lebih dulu. Data referensi/master
    # (belum tentu butuh dicocokkan ke apa pun) didahulukan; data
    # transaksional yang PALING diuntungkan oleh cross-matching
    # (rekening_koran <-> piutang) diletakkan paling akhir supaya semua
    # kandidat pencocokannya sudah tersimpan lebih dulu.
    "aset_tetap": 0,
    "buku_bantu_piutang": 1,
    "ap_aging": 1,
    "pembelian": 2,
    "penjualan": 3,
    "rekening_koran": 4,
}


def _susun_rencana_batch(deteksi_per_file: List[dict]) -> List[dict]:
    """
    deteksi_per_file: [{"nama_file": str, "sheets": list-hasil-deteksi_semua_sheet}]
    Return: list langkah terurut siap dieksekusi:
        [{"urutan": int, "nama_file": str, "jenis_terdeteksi": [kode, ...],
          "alasan": str}]
    plus catatan jenis yang muncul di >1 file (potensi duplikat).
    """
    kemunculan_jenis: Dict[str, List[str]] = {}
    langkah = []
    for entri in deteksi_per_file:
        jenis_di_file = sorted({
            s["kode"] for s in entri["sheets"]
            if s.get("kode") and s.get("sudah_ada_parser")
        })
        for kode in jenis_di_file:
            kemunculan_jenis.setdefault(kode, []).append(entri["nama_file"])
        if not jenis_di_file:
            alasan = "Tidak ada sheet yang dikenali/punya parser -- akan dilewati saat eksekusi."
        else:
            label_list = ", ".join(jenis_di_file)
            alasan = f"Terdeteksi sebagai: {label_list}."
        prioritas = min((_URUTAN_PRIORITAS_JENIS.get(k, 99) for k in jenis_di_file), default=99)
        langkah.append({
            "nama_file": entri["nama_file"],
            "jenis_terdeteksi": jenis_di_file,
            "alasan": alasan,
            "_prioritas": prioritas,
        })

    langkah.sort(key=lambda x: x["_prioritas"])
    for i, l in enumerate(langkah, 1):
        l["urutan"] = i
        del l["_prioritas"]

    duplikat_lintas_file = {
        kode: files for kode, files in kemunculan_jenis.items() if len(files) > 1
    }
    if duplikat_lintas_file:
        for l in langkah:
            for kode in l["jenis_terdeteksi"]:
                if kode in duplikat_lintas_file:
                    l["alasan"] += (
                        f" PERHATIAN: jenis '{kode}' juga ada di file lain dalam batch ini "
                        f"({', '.join(f for f in duplikat_lintas_file[kode] if f != l['nama_file'])}) "
                        f"-- cek apakah ini duplikat/revisi sebelum konfirmasi."
                    )
    return langkah


def _tahun_dari_draf_jurnal_list(daftar_draf_jurnal) -> set:
    """Ekstrak tahun (int) dari kolom 'tanggal' tiap baris draf_jurnal."""
    tahun_set = set()
    for baris in daftar_draf_jurnal or []:
        tanggal = baris.get("tanggal") if isinstance(baris, dict) else None
        if not tanggal:
            continue
        try:
            tahun_set.add(int(pd.to_datetime(tanggal).year))
        except Exception:  # noqa: BLE001
            continue
    return tahun_set


def _tahun_dari_hasil_json(hasil_json: dict) -> set:
    """hasil_json: {kode_jenis: {..., 'draf_jurnal': [...]}} -- hasil 1 file."""
    tahun_set = set()
    for data in (hasil_json or {}).values():
        if isinstance(data, dict):
            tahun_set |= _tahun_dari_draf_jurnal_list(data.get("draf_jurnal"))
    return tahun_set


def _tahun_dari_hasil_batch(hasil_per_file: List[dict]) -> set:
    """hasil_per_file: hasil dari proses_file_batch() -- gabungan banyak file."""
    tahun_set = set()
    for entri in hasil_per_file:
        tahun_set |= _tahun_dari_hasil_json(entri.get("hasil") or {})
    return tahun_set


# [BARU] 7 jenis dokumen yang berperan dalam laporan 18-sheet -- COA +
# 6 jenis file transaksi/pendukung. Dipakai oleh _cek_kelengkapan_
# dokumen_18_sheet() untuk menghitung berapa dari 7 ini yang sudah
# tersedia utk client ybs, BUKAN untuk memaksa ke-7 nya wajib ada.
_JENIS_DOKUMEN_18_SHEET: Dict[str, str] = {
    "coa": "Chart of Account (COA)",
    "rekening_koran": "Rekening Koran / Mutasi Bank",
    "penjualan": "Data Penjualan (Invoice & POS/Kasir)",
    "pembelian": "Pembelian",
    "aset_tetap": "Aset Tetap",
    "buku_bantu_piutang": "Buku Bantu Piutang (AR)",
    "ap_aging": "AP Aging (Utang Jatuh Tempo)",
}

# [BARU] Ambang jumlah jenis dokumen (dari 7 di atas) yang harus SUDAH
# TERSEDIA (baik dari upload lama yang tersimpan di database, maupun
# dari batch yang baru saja diproses) sebelum laporan 18-sheet otomatis
# digenerate. AI ini tidak cuma dipakai untuk laporan keuangan -- jadi
# begitu minimal jumlah ini terpenuhi, laporan TETAP dibuat apa adanya
# (bagian dari jenis yang belum diupload dibiarkan kosong di file-nya),
# bukan menunggu ke-7 jenis lengkap dulu.
_AMBANG_JUMLAH_JENIS_UNTUK_AUTO_18_SHEET = 3


def _cek_kelengkapan_dokumen_18_sheet(client_id: int) -> Dict[str, Any]:
    """
    [BARU] Mengecek berapa dari 7 jenis dokumen (lihat
    _JENIS_DOKUMEN_18_SHEET) yang sudah tersedia untuk client ini di
    database -- dipakai _auto_generate_laporan_18_sheet() untuk
    menentukan (a) apakah laporan 18-sheet layak digenerate otomatis
    sekarang (lihat _AMBANG_JUMLAH_JENIS_UNTUK_AUTO_18_SHEET), dan (b)
    jenis apa saja yang masih kosong -- supaya bisa disampaikan balik
    ke chat sebagai pemberitahuan kelengkapan, BUKAN error yang
    menahan laporan.

    Returns:
        dict: {"jumlah_terpenuhi": int, "terpenuhi": List[str] (label),
        "kurang": List[str] (label)}.
    """
    terpenuhi: List[str] = []
    kurang: List[str] = []

    coa_client = dbc.ambil_coa_client(client_id)
    if coa_client:
        terpenuhi.append(_JENIS_DOKUMEN_18_SHEET["coa"])
    else:
        kurang.append(_JENIS_DOKUMEN_18_SHEET["coa"])

    for kode in ("rekening_koran", "penjualan", "pembelian", "aset_tetap", "buku_bantu_piutang", "ap_aging"):
        hasil = dbc.ambil_hasil_client(client_id, jenis=kode, limit=1)
        if hasil:
            terpenuhi.append(_JENIS_DOKUMEN_18_SHEET[kode])
        else:
            kurang.append(_JENIS_DOKUMEN_18_SHEET[kode])

    return {"jumlah_terpenuhi": len(terpenuhi), "terpenuhi": terpenuhi, "kurang": kurang}


def _auto_generate_laporan_18_sheet(
    client_id: Optional[int], tahun_set: set, user: dict,
    on_progress: Optional[Callable[..., None]] = None,
) -> List[dict]:
    """
    [BARU] Dipakai bersama oleh proses_file_batch() (upload banyak file)
    dan proses_file_stream() (upload 1 file dari chat) supaya begitu file
    dikirim, laporan 18-sheet LANGSUNG keluar tanpa akuntan harus koreksi
    atau posting jurnal draft satu-satu dulu (_bangun_export_18_sheet
    sudah jalan dgn hanya_terposting=False -- baris yang akunnya masih
    perlu dikoreksi tetap ditandai lewat kolom "Status Validasi" di sheet
    GL, BUKAN dengan menahan laporan).

    [UBAH] Sebelumnya diblokir kalau COA kosong. Sekarang dipakai aturan
    "minimal N dari 7 jenis dokumen" (lihat _cek_kelengkapan_dokumen_
    18_sheet & _AMBANG_JUMLAH_JENIS_UNTUK_AUTO_18_SHEET) -- begitu
    ambang itu terpenuhi, laporan TETAP digenerate apa adanya (jenis
    yang belum diupload otomatis kosong di sheet terkait, karena
    accounting_export.py sudah pakai .get(...) or {}/[] di semua key).
    Jenis yang masih kurang TIDAK menggagalkan laporan -- cuma
    dilaporkan balik lewat field "pesan_kelengkapan" supaya bisa
    ditampilkan sebagai pemberitahuan di chat setelah file keluar.

    [BARU] on_progress: callback opsional (step, label, status, pesan) --
    diteruskan APA ADANYA ke _bangun_export_18_sheet() (yang lalu
    meneruskannya lagi ke _susun_data_export_18_sheet()) supaya SETIAP
    sub-tahap penyusunan laporan (COA, jurnal, laporan keuangan, lampiran
    SPT, laporan bulanan, aset tetap, PPh Badan, piutang/hutang, tren
    saldo, narasi AI, susun file Excel) ikut terlapor -- dipakai endpoint
    SSE /api/client/{client_id}/proses-file-batch/stream supaya user
    lihat proses ini step-by-step, bukan cuma satu baris besar "generate
    laporan 18-sheet".
    """
    laporan_18_sheet: List[dict] = []
    if client_id is None or not tahun_set:
        return laporan_18_sheet

    kelengkapan = _cek_kelengkapan_dokumen_18_sheet(client_id)
    if kelengkapan["jumlah_terpenuhi"] < _AMBANG_JUMLAH_JENIS_UNTUK_AUTO_18_SHEET:
        laporan_18_sheet.append({
            "status": "perlu_file",
            "pesan": (
                f"Laporan 18-sheet belum digenerate otomatis -- baru "
                f"{kelengkapan['jumlah_terpenuhi']} dari 7 jenis dokumen yang "
                f"terdeteksi (minimal {_AMBANG_JUMLAH_JENIS_UNTUK_AUTO_18_SHEET} "
                f"dibutuhkan). Yang masih kurang: {', '.join(kelengkapan['kurang'])}."
            ),
        })
        return laporan_18_sheet

    for tahun in sorted(tahun_set):
        try:
            isi_excel = _bangun_export_18_sheet(
                client_id, Export18SheetRequest(tahun=tahun), user, on_progress=on_progress,
            )
            pesan_kelengkapan = None
            if kelengkapan["kurang"]:
                pesan_kelengkapan = (
                    "Laporan 18-sheet berhasil dibuat. File yang Anda kirim masih "
                    f"kurang di bagian: {', '.join(kelengkapan['kurang'])} -- "
                    "sheet yang terkait bagian tersebut dikosongkan di file ini."
                )
            laporan_18_sheet.append({
                "status": "berhasil",
                "tahun": tahun,
                "nama_file": f"Laporan_Keuangan_{tahun}_18_Sheet.xlsx",
                "file_base64": base64.b64encode(isi_excel).decode("ascii"),
                "pesan_kelengkapan": pesan_kelengkapan,
            })
        except HTTPException as e:
            laporan_18_sheet.append({
                "status": "perlu_file", "tahun": tahun,
                "pesan": e.detail if isinstance(e.detail, str) else str(e.detail),
            })
        except Exception as e:  # noqa: BLE001
            laporan_18_sheet.append({
                "status": "gagal", "tahun": tahun,
                "pesan": f"Gagal membuat laporan 18-sheet otomatis untuk tahun {tahun}: {e}",
            })
    return laporan_18_sheet


@app.post("/api/client/{client_id}/proses-file-batch")
async def proses_file_batch(
    client_id: int,
    files: List[UploadFile] = File(...),
    jenis_dokumen: Optional[str] = Form(None),
    conv_id: Optional[str] = Form(None),
    konfirmasi_duplikat: bool = Form(False),
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    if not files:
        raise HTTPException(status_code=400, detail="Tidak ada file yang diupload.")

    conv_id_final = conv_id or datetime.now().isoformat()

    # -- Baca semua file dulu (butuh isinya utuh utk tahap deteksi & eksekusi) --
    file_bytes: Dict[str, bytes] = {}
    for f in files:
        nama = f.filename or f"upload_{len(file_bytes) + 1}.xlsx"
        file_bytes[nama] = await f.read()

    # [BARU] PDF rekening koran -> working paper (Kertas Kerja), dipisah
    # dari batch SEBELUM tahap deteksi. Alasan: ak.deteksi_semua_sheet()
    # (tahap deteksi di bawah) HANYA paham struktur sheet Excel -- PDF
    # selalu dilabeli "tidak ada sheet dikenali" (tampil sbg "tidak
    # dikenali" di rencana) padahal sebenarnya BISA diproses lewat jalur
    # PDF khusus kertas_kerja.py. Supaya rencana tidak lagi salah tampil,
    # DAN supaya PDF rekening koran menghasilkan working paper (bukan ikut
    # numpuk ke _auto_generate_laporan_18_sheet di step 5), file .pdf
    # dikeluarkan dari file_bytes di sini dan diproses terpisah lewat
    # kertas_kerja.generate_kertas_kerja() (SEKALIGUS semua PDF dalam
    # batch ini, konsisten dgn desain kertas_kerja.py yg memang mendukung
    # multi bulan/bank sekaligus digabung jadi satu GL/working paper).
    nama_file_pdf_batch = [n for n in file_bytes if n.lower().endswith(".pdf")]
    file_bytes_pdf = {n: file_bytes.pop(n) for n in nama_file_pdf_batch}

    kertas_kerja_hasil = None
    if file_bytes_pdf:
        try:
            df_coa_kk, peringatan_coa_kk = _siapkan_df_coa_untuk_kertas_kerja(client_id, None, None)
            daftar_file_pdf_kk: List[Tuple[Any, str]] = []
            for nama_file_pdf, isi_pdf in file_bytes_pdf.items():
                buf = io.BytesIO(isi_pdf)
                buf.name = nama_file_pdf
                daftar_file_pdf_kk.append((buf, nama_file_pdf))
            kertas_kerja_hasil = await asyncio.to_thread(
                _jalankan_generate_kertas_kerja,
                client_id, daftar_file_pdf_kk, df_coa_kk, peringatan_coa_kk,
                [], True, user,
            )
        except HTTPException as e:
            kertas_kerja_hasil = {
                "status": "gagal",
                "pesan": e.detail if isinstance(e.detail, str) else str(e.detail),
                "file": nama_file_pdf_batch,
            }
        except Exception as e:  # noqa: BLE001
            kertas_kerja_hasil = {
                "status": "gagal", "pesan": f"Gagal generate kertas kerja: {e}",
                "file": nama_file_pdf_batch,
            }

    if not file_bytes:
        # Semua file dalam batch ini adalah PDF (sudah ditangani di atas
        # via kertas_kerja) -- tidak ada file lain yang perlu lewat jalur
        # deteksi/eksekusi/laporan 18-sheet di bawah, langsung kembalikan.
        dbc.log_audit(
            client_id=client_id, user=user.get("username", "unknown"),
            aksi="proses_file_batch",
            detail={"jumlah_file": len(files), "jumlah_file_pdf_kertas_kerja": len(nama_file_pdf_batch)},
        )
        return {
            "jumlah_file": len(files),
            "rencana": [],
            "hasil_per_file": [],
            "cross_matching": None,
            "laporan_18_sheet": [],
            "kertas_kerja": kertas_kerja_hasil,
        }

    # -- 1. TAHAP DETEKSI (tanpa efek samping) --
    # [FIX -- GAP EVENT LOOP] Sama alasannya dgn /api/deteksi-file --
    # ak.deteksi_semua_sheet() sync/berat, dipanggil BERKALI-KALI di loop
    # ini (1x per file dalam batch). Tiap panggilan dibungkus to_thread
    # SATU-SATU (bukan seluruh loop sekaligus jadi 1 thread) supaya
    # penanganan error per-file tetap presisi sama seperti sebelumnya.
    deteksi_per_file = []
    for nama_file, isi in file_bytes.items():
        buf = io.BytesIO(isi)
        buf.name = nama_file
        try:
            sheets = await asyncio.to_thread(ak.deteksi_semua_sheet, buf, nama_file)
        except Exception as e:  # noqa: BLE001
            sheets = []
            print(f"[PERINGATAN] Gagal deteksi {nama_file}: {e}")
        deteksi_per_file.append({"nama_file": nama_file, "sheets": sheets})

    # -- 2. TAHAP RENCANA --
    rencana = _susun_rencana_batch(deteksi_per_file)

    # -- 3. TAHAP EKSEKUSI, sesuai urutan rencana --
    hasil_per_file = []
    hasil_mentah_per_file: Dict[str, dict] = {}
    for langkah in rencana:
        nama_file = langkah["nama_file"]
        isi = file_bytes[nama_file]
        try:
            # [FIX -- GAP EVENT LOOP] Sama pola dgn /api/proses-file --
            # dibungkus to_thread PER FILE (bukan seluruh loop sekaligus)
            # supaya try/except per-file di bawah ini tetap menangkap error
            # 1 file tanpa menggagalkan file lain dalam batch, persis
            # seperti perilaku sebelumnya.
            hasil = await asyncio.to_thread(
                _proses_dan_simpan_satu_file,
                isi, nama_file, jenis_dokumen, client_id, conv_id_final, None,
                konfirmasi_duplikat, user,
            )
            hasil_per_file.append({"nama_file": nama_file, "urutan": langkah["urutan"], **hasil})
            hasil_mentah_per_file[nama_file] = hasil
        except Exception as e:  # noqa: BLE001
            # [PENTING] Satu file gagal TIDAK menggagalkan file lain dalam
            # batch -- dicatat sebagai error per file, batch tetap lanjut.
            hasil_per_file.append({
                "nama_file": nama_file, "urutan": langkah["urutan"],
                "error": str(e), "tidak_terdeteksi": True,
            })

    # -- 4. LINTAS FILE: cocokkan bank <-> piutang otomatis kalau keduanya ada --
    cross_matching_hasil = None
    hasil_bank = next(
        (h["hasil"].get("rekening_koran") for h in hasil_per_file
         if h.get("hasil") and h["hasil"].get("rekening_koran")), None,
    )
    hasil_piutang = next(
        (h["hasil"].get("buku_bantu_piutang") for h in hasil_per_file
         if h.get("hasil") and h["hasil"].get("buku_bantu_piutang")), None,
    )
    if hasil_bank and hasil_piutang:
        try:
            df_bank = hasil_bank.get("df")
            df_piutang = hasil_piutang.get("df")
            # [FIX -- BUG "'list' object has no attribute 'empty'"]
            # df_bank/df_piutang di sini adalah hasil serialisasi JSON
            # (list of dict, lihat catatan di fiscal_reconciliation.py),
            # BUKAN pandas DataFrame lagi -- padahal
            # cross_matching.cocokkan_bank_piutang() mengasumsikan
            # DataFrame (memanggil .empty, .copy(), dst). Convert balik
            # ke DataFrame di sini sebelum dipanggil.
            if not isinstance(df_bank, pd.DataFrame):
                df_bank = pd.DataFrame(df_bank or [])
            if not isinstance(df_piutang, pd.DataFrame):
                df_piutang = pd.DataFrame(df_piutang or [])
            # [FIX -- GAP EVENT LOOP] cocokkan_bank_piutang() sync, bisa
            # berat kalau df_bank/df_piutang besar (cross-join/matching
            # baris demi baris) -- dibungkus to_thread sama alasannya
            # dgn pemanggilan lain di endpoint ini.
            hasil_cocok = await asyncio.to_thread(cross_matching.cocokkan_bank_piutang, df_bank, df_piutang)
            # [FIX] cocokkan_bank_piutang() mengembalikan DICT
            # {"hasil": [...], "mutasi_bank_masuk_belum_terpakai": [...],
            # "ringkasan": {...}} -- BUKAN DataFrame -- versi lama di sini
            # memperlakukannya seolah DataFrame (df_cocok["cocok_piutang"],
            # .columns) sehingga jumlah_cocok selalu diam-diam bernilai 0.
            daftar_hasil_cocok = (hasil_cocok or {}).get("hasil", [])
            jumlah_cocok = sum(1 for r in daftar_hasil_cocok if r.get("status") == "MATCHED")
            cross_matching_hasil = {
                "dilakukan": True,
                "jumlah_baris_bank_cocok_ke_piutang": jumlah_cocok,
                "ringkasan": (hasil_cocok or {}).get("ringkasan"),
            }
        except Exception as e:  # noqa: BLE001
            cross_matching_hasil = {"dilakukan": False, "error": str(e)}

    # -- 5. [BARU] AUTO-GENERATE LAPORAN 14-SHEET (lihat docstring
    # _auto_generate_laporan_18_sheet) --
    # [FIX -- GAP EVENT LOOP] Generate workbook 18-sheet itu kerja berat
    # (openpyxl susun banyak sheet dari data DB) -- dibungkus to_thread
    # sama alasannya dgn pemanggilan lain di endpoint batch ini.
    laporan_18_sheet = await asyncio.to_thread(
        _auto_generate_laporan_18_sheet,
        client_id, _tahun_dari_hasil_batch(hasil_per_file), user,
    )

    dbc.log_audit(
        client_id=client_id, user=user.get("username", "unknown"),
        aksi="proses_file_batch",
        detail={
            "jumlah_file": len(files),
            "jumlah_file_pdf_kertas_kerja": len(nama_file_pdf_batch),
            "rencana": [{"urutan": l["urutan"], "nama_file": l["nama_file"],
                         "jenis_terdeteksi": l["jenis_terdeteksi"]} for l in rencana],
            "tahun_laporan_auto": [l.get("tahun") for l in laporan_18_sheet if l.get("status") == "berhasil"],
        },
    )

    return {
        "jumlah_file": len(files),
        "rencana": rencana,
        "hasil_per_file": hasil_per_file,
        "cross_matching": cross_matching_hasil,
        "laporan_18_sheet": laporan_18_sheet,
        "kertas_kerja": kertas_kerja_hasil,
    }


# ============================================================
# [BARU] VERSI STREAMING (SSE) DARI /proses-file-batch
# ============================================================
# Dipakai ChatPage.jsx (lihat frontend/src/lib/api.js::
# prosesFileBatchStream()) supaya SETIAP tahap -- mulai dari file yang
# BARU SAJA diterima dari user, sampai ke sub-tahap internal generate
# laporan 18-sheet (COA, jurnal, laporan keuangan, dst) -- tampil
# sebagai daftar langkah di chat (komponen ProcessingSteps.jsx), mirip
# panel "Menjalankan N perintah..." ala Claude Code. Alur logic PERSIS
# SAMA dengan proses_file_batch() di atas (sengaja TIDAK diduplikasi
# beda logic-nya, cuma dibungkus ulang jadi sinkron + lapor progress per
# tahap) -- kalau proses_file_batch() di atas diubah, endpoint ini juga
# perlu disesuaikan.
@app.post("/api/client/{client_id}/proses-file-batch/stream")
async def proses_file_batch_stream(
    client_id: int,
    files: List[UploadFile] = File(...),
    jenis_dokumen: Optional[str] = Form(None),
    conv_id: Optional[str] = Form(None),
    konfirmasi_duplikat: bool = Form(False),
    user: dict = Depends(auth.require_level(3)),
):
    """
    Versi streaming (SSE) dari POST .../proses-file-batch. Urutan step
    yang dikirim (field "step" tiap event "progress"):

        1. "baca_file"        -- file yang diterima dari user (PALING AWAL)
        2. "kertas_kerja"     -- HANYA kalau ada PDF rekening koran di batch
        3. "deteksi"          -- deteksi jenis dokumen tiap file
        4. "rencana"          -- urutan pemrosesan disusun
        5. "eksekusi:<nama_file>" -- satu step PER FILE, diklasifikasi & disimpan
        6. "cross_matching"   -- HANYA kalau ada Rekening Koran + Piutang
        7. "18sheet:<sub_step>" -- sub-tahap internal generate 18-sheet,
           diteruskan APA ADANYA dari _susun_data_export_18_sheet() /
           _bangun_export_18_sheet() (on_progress) -- mis.
           "18sheet:coa", "18sheet:jurnal", "18sheet:laporan_keuangan",
           "18sheet:lampiran_spt", "18sheet:laporan_bulanan",
           "18sheet:aset_tetap", "18sheet:pph_badan",
           "18sheet:piutang_hutang", "18sheet:tren_saldo",
           "18sheet:narasi_ai", "18sheet:generate_excel" (atau
           "18sheet:cache" kalau hasil diambil dari cache, lihat
           _bangun_export_18_sheet).

    Tiap step muncul 2x (status "processing" lalu "done"/"skip"/"error")
    KECUALI beberapa yang instan (mis. "rencana"). Ganti teks "label" di
    fungsi ini / _susun_data_export_18_sheet() / _auto_generate_
    laporan_18_sheet() sesuai nama yang kamu mau tampilkan ke user --
    "step" (id) sengaja dipisah dari "label" (teks) supaya bisa diganti
    bebas tanpa mengubah logic frontend.

    Event terakhir sebelum "[DONE]" bertipe "result", skemanya PERSIS
    SAMA dengan response /api/client/{client_id}/proses-file-batch biasa
    ({jumlah_file, rencana, hasil_per_file, cross_matching,
    laporan_18_sheet, kertas_kerja}) -- frontend cukup ganti CARA
    MEMANGGIL (baca event SSE bertahap lewat prosesFileBatchStream()),
    logika MEMBACA hasil akhir TIDAK berubah dari prosesFileBatch().
    """
    if not files:
        raise HTTPException(status_code=400, detail="Tidak ada file yang diupload.")

    conv_id_final = conv_id or datetime.now().isoformat()

    # File dibaca (await, di event loop) SEBELUM masuk ke thread --
    # UploadFile/SpooledTemporaryFile FastAPI tidak aman dipakai dari
    # thread lain sekaligus event loop async secara bersamaan, sama pola
    # dengan endpoint stream lain (lihat api_generate_kertas_kerja_stream).
    file_bytes: Dict[str, bytes] = {}
    for f in files:
        nama = f.filename or f"upload_{len(file_bytes) + 1}.xlsx"
        file_bytes[nama] = await f.read()

    q: "queue.Queue" = queue.Queue()

    def jalankan():
        try:
            # -- 0. [BARU] Step PALING AWAL yang diminta: laporkan file apa
            # saja yang baru saja diterima dari user, SEBELUM diproses
            # apa pun.
            daftar_nama_file = list(file_bytes.keys())
            q.put({
                "type": "progress", "step": "baca_file",
                "label": f"Membaca {len(daftar_nama_file)} file dari Anda: {', '.join(daftar_nama_file)}",
                "status": "processing",
            })
            fb = dict(file_bytes)  # copy lokal -- aman diubah (di-pop) di thread ini
            q.put({
                "type": "progress", "step": "baca_file",
                "label": f"Membaca {len(daftar_nama_file)} file dari Anda",
                "status": "done",
            })

            # -- 0.5. PDF rekening koran -> Kertas Kerja, dipisah dulu dari
            # batch (sama seperti proses_file_batch(), lihat komentar
            # panjang di sana untuk alasannya).
            nama_file_pdf_batch = [n for n in fb if n.lower().endswith(".pdf")]
            file_bytes_pdf = {n: fb.pop(n) for n in nama_file_pdf_batch}

            kertas_kerja_hasil = None
            if file_bytes_pdf:
                q.put({
                    "type": "progress", "step": "kertas_kerja",
                    "label": f"Menyusun Kertas Kerja dari {len(file_bytes_pdf)} PDF rekening koran",
                    "status": "processing",
                })
                try:
                    df_coa_kk, peringatan_coa_kk = _siapkan_df_coa_untuk_kertas_kerja(client_id, None, None)
                    daftar_file_pdf_kk: List[Tuple[Any, str]] = []
                    for nama_file_pdf, isi_pdf in file_bytes_pdf.items():
                        buf = io.BytesIO(isi_pdf)
                        buf.name = nama_file_pdf
                        daftar_file_pdf_kk.append((buf, nama_file_pdf))
                    kertas_kerja_hasil = _jalankan_generate_kertas_kerja(
                        client_id, daftar_file_pdf_kk, df_coa_kk, peringatan_coa_kk,
                        [], True, user,
                    )
                    q.put({
                        "type": "progress", "step": "kertas_kerja",
                        "label": f"Menyusun Kertas Kerja dari {len(file_bytes_pdf)} PDF rekening koran",
                        "status": "done",
                    })
                except HTTPException as e:
                    pesan_gagal = e.detail if isinstance(e.detail, str) else str(e.detail)
                    kertas_kerja_hasil = {"status": "gagal", "pesan": pesan_gagal, "file": nama_file_pdf_batch}
                    q.put({
                        "type": "progress", "step": "kertas_kerja",
                        "label": "Menyusun Kertas Kerja", "status": "error", "pesan": pesan_gagal,
                    })
                except Exception as e:  # noqa: BLE001
                    kertas_kerja_hasil = {
                        "status": "gagal", "pesan": f"Gagal generate kertas kerja: {e}",
                        "file": nama_file_pdf_batch,
                    }
                    q.put({
                        "type": "progress", "step": "kertas_kerja",
                        "label": "Menyusun Kertas Kerja", "status": "error", "pesan": str(e),
                    })

            if not fb:
                # Semua file dalam batch ini PDF (sudah ditangani di atas) --
                # tidak ada jalur deteksi/eksekusi/18-sheet, langsung kirim
                # hasil akhir.
                dbc.log_audit(
                    client_id=client_id, user=user.get("username", "unknown"),
                    aksi="proses_file_batch_stream",
                    detail={"jumlah_file": len(files), "jumlah_file_pdf_kertas_kerja": len(nama_file_pdf_batch)},
                )
                q.put({
                    "type": "result",
                    "jumlah_file": len(files), "rencana": [], "hasil_per_file": [],
                    "cross_matching": None, "laporan_18_sheet": [], "kertas_kerja": kertas_kerja_hasil,
                })
                return

            # -- 1. TAHAP DETEKSI --
            q.put({
                "type": "progress", "step": "deteksi",
                "label": f"Mendeteksi jenis dokumen dari {len(fb)} file",
                "status": "processing",
            })
            deteksi_per_file = []
            for nama_file, isi in fb.items():
                buf = io.BytesIO(isi)
                buf.name = nama_file
                try:
                    sheets = ak.deteksi_semua_sheet(buf, nama_file)
                except Exception as e:  # noqa: BLE001
                    sheets = []
                    print(f"[PERINGATAN] Gagal deteksi {nama_file}: {e}")
                deteksi_per_file.append({"nama_file": nama_file, "sheets": sheets})
            q.put({
                "type": "progress", "step": "deteksi",
                "label": f"Mendeteksi jenis dokumen dari {len(fb)} file",
                "status": "done",
            })

            # -- 2. TAHAP RENCANA --
            q.put({
                "type": "progress", "step": "rencana",
                "label": "Menyusun rencana urutan pemrosesan",
                "status": "processing",
            })
            rencana = _susun_rencana_batch(deteksi_per_file)
            q.put({
                "type": "progress", "step": "rencana",
                "label": "Menyusun rencana urutan pemrosesan",
                "status": "done",
            })

            # -- 3. TAHAP EKSEKUSI, sesuai urutan rencana -- [BARU] satu
            # step PER FILE (bukan 1 step besar) supaya user lihat file
            # mana yang sedang diproses.
            hasil_per_file = []
            for langkah in rencana:
                nama_file = langkah["nama_file"]
                isi = fb[nama_file]
                q.put({
                    "type": "progress", "step": f"eksekusi:{nama_file}",
                    "label": f'Memproses "{nama_file}"', "status": "processing",
                })
                try:
                    hasil = _proses_dan_simpan_satu_file(
                        isi, nama_file, jenis_dokumen, client_id, conv_id_final, None,
                        konfirmasi_duplikat, user,
                    )
                    hasil_per_file.append({"nama_file": nama_file, "urutan": langkah["urutan"], **hasil})
                    q.put({
                        "type": "progress", "step": f"eksekusi:{nama_file}",
                        "label": f'Memproses "{nama_file}"', "status": "done",
                    })
                except Exception as e:  # noqa: BLE001
                    # [PENTING] Satu file gagal TIDAK menggagalkan file lain
                    # dalam batch -- dicatat sebagai error per file, batch
                    # tetap lanjut (sama seperti proses_file_batch()).
                    hasil_per_file.append({
                        "nama_file": nama_file, "urutan": langkah["urutan"],
                        "error": str(e), "tidak_terdeteksi": True,
                    })
                    q.put({
                        "type": "progress", "step": f"eksekusi:{nama_file}",
                        "label": f'Memproses "{nama_file}"', "status": "error", "pesan": str(e),
                    })

            # -- 4. LINTAS FILE: cocokkan bank <-> piutang otomatis kalau
            # keduanya ada di batch ini --
            cross_matching_hasil = None
            hasil_bank = next(
                (h["hasil"].get("rekening_koran") for h in hasil_per_file
                 if h.get("hasil") and h["hasil"].get("rekening_koran")), None,
            )
            hasil_piutang_cm = next(
                (h["hasil"].get("buku_bantu_piutang") for h in hasil_per_file
                 if h.get("hasil") and h["hasil"].get("buku_bantu_piutang")), None,
            )
            if hasil_bank and hasil_piutang_cm:
                q.put({
                    "type": "progress", "step": "cross_matching",
                    "label": "Mencocokkan Rekening Koran dengan Buku Bantu Piutang",
                    "status": "processing",
                })
                try:
                    df_bank = hasil_bank.get("df")
                    df_piutang = hasil_piutang_cm.get("df")
                    # [FIX -- BUG "'list' object has no attribute 'empty'"]
                    # df_bank/df_piutang di sini adalah hasil serialisasi
                    # JSON (list of dict), BUKAN pandas DataFrame lagi --
                    # padahal cross_matching.cocokkan_bank_piutang()
                    # mengasumsikan DataFrame (memanggil .empty, .copy(),
                    # dst). Convert balik ke DataFrame sebelum dipanggil.
                    if not isinstance(df_bank, pd.DataFrame):
                        df_bank = pd.DataFrame(df_bank or [])
                    if not isinstance(df_piutang, pd.DataFrame):
                        df_piutang = pd.DataFrame(df_piutang or [])
                    hasil_cocok = cross_matching.cocokkan_bank_piutang(df_bank, df_piutang)
                    # [FIX] cocokkan_bank_piutang() mengembalikan DICT
                    # {"hasil": [...], "mutasi_bank_masuk_belum_terpakai":
                    # [...], "ringkasan": {...}} -- BUKAN DataFrame --
                    # versi lama memperlakukannya seolah DataFrame
                    # (df_cocok["cocok_piutang"], .columns) sehingga
                    # jumlah_cocok selalu diam-diam bernilai 0.
                    daftar_hasil_cocok = (hasil_cocok or {}).get("hasil", [])
                    jumlah_cocok = sum(1 for r in daftar_hasil_cocok if r.get("status") == "MATCHED")
                    cross_matching_hasil = {
                        "dilakukan": True,
                        "jumlah_baris_bank_cocok_ke_piutang": jumlah_cocok,
                        "ringkasan": (hasil_cocok or {}).get("ringkasan"),
                    }
                    q.put({
                        "type": "progress", "step": "cross_matching",
                        "label": f"Mencocokkan Rekening Koran dengan Buku Bantu Piutang ({jumlah_cocok} baris cocok)",
                        "status": "done",
                    })
                except Exception as e:  # noqa: BLE001
                    cross_matching_hasil = {"dilakukan": False, "error": str(e)}
                    q.put({
                        "type": "progress", "step": "cross_matching",
                        "label": "Mencocokkan Rekening Koran dengan Buku Bantu Piutang",
                        "status": "error", "pesan": str(e),
                    })

            # -- 5. AUTO-GENERATE LAPORAN 18-SHEET -- [BARU] on_progress
            # diteruskan ke _auto_generate_laporan_18_sheet() supaya
            # SEMUA sub-tahap internalnya (COA, jurnal, laporan keuangan,
            # lampiran SPT, laporan bulanan, aset tetap, PPh Badan,
            # piutang/hutang, tren saldo, narasi AI, susun file Excel --
            # lihat _susun_data_export_18_sheet() & _bangun_export_18_sheet())
            # ikut tampil sebagai step SENDIRI ("18sheet:<sub_step>"),
            # bukan cuma 1 baris besar "generate 18-sheet".
            def on_progress_18_sheet(step, label, status, pesan=None):
                q.put({
                    "type": "progress", "step": f"18sheet:{step}",
                    "label": label, "status": status, "pesan": pesan,
                })

            tahun_set = _tahun_dari_hasil_batch(hasil_per_file)
            laporan_18_sheet = _auto_generate_laporan_18_sheet(
                client_id, tahun_set, user, on_progress=on_progress_18_sheet,
            )

            dbc.log_audit(
                client_id=client_id, user=user.get("username", "unknown"),
                aksi="proses_file_batch_stream",
                detail={
                    "jumlah_file": len(files),
                    "jumlah_file_pdf_kertas_kerja": len(nama_file_pdf_batch),
                    "rencana": [{"urutan": l["urutan"], "nama_file": l["nama_file"],
                                 "jenis_terdeteksi": l["jenis_terdeteksi"]} for l in rencana],
                    "tahun_laporan_auto": [l.get("tahun") for l in laporan_18_sheet if l.get("status") == "berhasil"],
                },
            )

            q.put({
                "type": "result",
                "jumlah_file": len(files),
                "rencana": rencana,
                "hasil_per_file": hasil_per_file,
                "cross_matching": cross_matching_hasil,
                "laporan_18_sheet": laporan_18_sheet,
                "kertas_kerja": kertas_kerja_hasil,
            })
        except Exception as e:  # noqa: BLE001
            q.put({"type": "error", "pesan": str(e)})
        finally:
            q.put(None)  # sinyal: tidak ada event lagi

    threading.Thread(target=jalankan, daemon=True).start()

    def event_generator():
        while True:
            item = q.get()
            if item is None:
                break
            yield _format_sse_progress(**item)
        yield "data: [DONE]\n\n"

    return StreamingResponse(event_generator(), media_type="text/event-stream")


@app.post("/api/upload-batch/{batch_id}/konfirmasi")
async def konfirmasi_upload_batch(
    batch_id: int,
    aksi: str = Form(...),  # "lanjutkan_semua" | "hanya_baris_baru" | "batalkan"
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas, sama seperti /api/proses-file
):
    """
    [BARU - dedup upload] Tindak lanjut atas batch berstatus
    'menunggu_konfirmasi' (dibuat oleh /api/proses-file saat
    terdeteksi indikasi duplikat/revisi rekening koran).

    aksi:
      "lanjutkan_semua"   -- tetap tarik SEMUA baris ke jurnal_posting
                              apa adanya (dipakai kalau overlap
                              ternyata memang kebetulan/sah, BUKAN
                              duplikat sungguhan).
      "hanya_baris_baru"  -- tarik HANYA baris yang fingerprint-nya
                              belum pernah ada (kasus revisi yang sah:
                              buang baris lama yang sudah tercatat,
                              simpan baris tambahan yang genuinely baru).
      "batalkan"          -- tolak batch ini sepenuhnya, tidak ada
                              baris yang ditarik ke jurnal_posting.
    """
    batch = dbc.ambil_upload_batch_by_id(batch_id)
    if batch is None:
        raise HTTPException(status_code=404, detail="Batch upload tidak ditemukan.")
    if batch["status"] != "menunggu_konfirmasi":
        raise HTTPException(
            status_code=400,
            detail=f"Batch ini sudah diproses sebelumnya (status saat ini: {batch['status']}).",
        )
    if aksi not in ("lanjutkan_semua", "hanya_baris_baru", "batalkan"):
        raise HTTPException(status_code=422, detail="aksi harus salah satu: lanjutkan_semua, hanya_baris_baru, batalkan.")

    draf_jurnal = batch.get("draf_jurnal") or []

    if aksi == "batalkan":
        dbc.perbarui_status_upload_batch(batch_id, "dibatalkan", user=user.get("username", "unknown"))
        history.catat_riwayat(
            client_id=batch["client_id"], user=user.get("username", "unknown"),
            aksi="batalkan_upload_duplikat",
            detail={"batch_id": batch_id, "nama_file": batch.get("nama_file"), "kode_bank": batch["kode_bank"], "periode": batch["periode"]},
        )
        return {"batch_id": batch_id, "status": "dibatalkan", "jumlah_baris_ditarik": 0}

    # [FIX -- GAP EVENT LOOP] hash_lama bisa berisi ribuan fingerprint (dari
    # dbc.ambil_hash_transaksi_aktif) & draf_jurnal juga bisa ribuan baris --
    # list comprehension buat_signature_baris per baris + query DB ini sync,
    # dibungkus to_thread biar tidak memblokir event loop.
    def _susun_draf_dipakai() -> list:
        if aksi == "hanya_baris_baru":
            hash_lama = dbc.ambil_hash_transaksi_aktif(batch["client_id"], batch["kode_bank"], batch["periode"])
            return [b for b in draf_jurnal if dedup_transaksi.buat_signature_baris(b) not in hash_lama]
        return draf_jurnal  # lanjutkan_semua

    draf_dipakai = await asyncio.to_thread(_susun_draf_dipakai)
    jumlah = await asyncio.to_thread(
        dbc.tarik_draf_jurnal_ke_posting, batch["client_id"], batch["hasil_id"], "rekening_koran", draf_dipakai
    )

    dbc.perbarui_status_upload_batch(batch_id, "aktif", user=user.get("username", "unknown"))
    if batch["status_deteksi"] in ("REVISI_SEBAGIAN", "DUPLIKAT_PENUH"):
        batch_lama = dbc.ambil_batch_aktif(batch["client_id"], batch["kode_bank"], batch["periode"])
        if batch_lama and batch_lama["id"] != batch_id:
            dbc.tandai_batch_diganti(batch_lama["id"], batch_id)

    history.catat_riwayat(
        client_id=batch["client_id"], user=user.get("username", "unknown"),
        aksi="konfirmasi_upload_duplikat",
        detail={
            "batch_id": batch_id, "aksi": aksi, "nama_file": batch.get("nama_file"),
            "kode_bank": batch["kode_bank"], "periode": batch["periode"], "jumlah_baris_ditarik": jumlah,
        },
    )
    return {"batch_id": batch_id, "status": "aktif", "aksi": aksi, "jumlah_baris_ditarik": jumlah}


@app.get("/api/client/{client_id}/upload-batch")
async def daftar_upload_batch(
    client_id: int,
    limit: int = 100,
    user: dict = Depends(auth.get_current_user),
):
    """[BARU - dedup upload] Riwayat upload rekening koran per
    client -- termasuk yang masih 'menunggu_konfirmasi' (utk UI
    menampilkan badge "ada upload perlu ditinjau")."""
    return {"batch": dbc.daftar_upload_batch_client(client_id, limit=limit)}


@app.post("/api/client/{client_id}/bootstrap-pola-bank")
async def api_bootstrap_pola_bank(
    client_id: int,
    files: List[UploadFile] = File(...),
    min_samples: int = Form(2),
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """
    [BARU] "Suapi" pola_bank_client_{client_id}.json dari rekening koran
    bulan-bulan lalu yang SUDAH DIJURNAL LENGKAP oleh akuntan -- lihat
    ak.proses_file_bootstrap_pola_bank() untuk detail lengkap.

    Beda dari /api/proses-file:
    - TIDAK menjalankan kategorisasi/AI sama sekali (murni belajar pola).
    - TIDAK disimpan ke tabel `hasil` (bukan "hasil upload" biasa) --
      output-nya cuma pola_bank_client_{client_id}.json + 1 entri audit log.
    - Boleh terima BEBERAPA file sekaligus (mis. rekening koran Jan+Feb+Mar
      dalam satu request) -- diproses berurutan, tiap file menggabung ke
      pola yang sudah tersimpan dari file sebelumnya (bukan saling menimpa).
    """
    if not files:
        raise HTTPException(status_code=400, detail="Tidak ada file yang diupload.")

    ringkasan_per_file = []
    total_pola_baru = 0
    total_pola_diperbarui = 0
    total_baris_dibaca = 0
    total_baris_dipakai = 0

    for file in files:
        isi = await file.read()
        nama_file = file.filename or "upload.xlsx"
        buf = io.BytesIO(isi)
        buf.name = nama_file
        try:
            # [FIX -- GAP EVENT LOOP] proses_file_bootstrap_pola_bank() sync
            # & berat (parsing rekening koran penuh), dipanggil berkali-kali
            # di loop ini (1x per file). Dibungkus to_thread per file, sama
            # pola dgn /api/proses-file-batch di atas.
            hasil = await asyncio.to_thread(
                ak.proses_file_bootstrap_pola_bank,
                buf, nama_file, client_id=client_id, min_samples=min_samples,
            )
        except Exception as e:  # noqa: BLE001
            ringkasan_per_file.append({"nama_file": nama_file, "error": str(e)})
            continue

        total_pola_baru += hasil["jumlah_pola_baru"]
        total_pola_diperbarui += hasil["jumlah_pola_diperbarui"]
        total_baris_dibaca += hasil["jumlah_baris_dibaca"]
        total_baris_dipakai += hasil["jumlah_baris_dipakai"]
        ringkasan_per_file.append({
            "nama_file": nama_file,
            "jumlah_baris_dibaca": hasil["jumlah_baris_dibaca"],
            "jumlah_baris_dipakai": hasil["jumlah_baris_dipakai"],
            "jumlah_pola_baru": hasil["jumlah_pola_baru"],
            "jumlah_pola_diperbarui": hasil["jumlah_pola_diperbarui"],
            "detail_pola_baru": hasil["detail_pola_baru"],
            "detail_pola_diperbarui": hasil["detail_pola_diperbarui"],
            "sheet_dilewati": hasil["sheet_dilewati"],
        })

    # Audit trail -- satu entri per REQUEST (bisa berisi banyak file
    # sekaligus), bukan per file, supaya tidak membanjiri log audit client.
    dbc.log_audit(
        client_id=client_id,
        user=user.get("username", "unknown"),
        aksi="bootstrap_pola_bank",
        detail={
            "jumlah_file": len(files),
            "total_pola_baru": total_pola_baru,
            "total_pola_diperbarui": total_pola_diperbarui,
            "total_baris_dibaca": total_baris_dibaca,
            "total_baris_dipakai": total_baris_dipakai,
            "per_file": [
                {k: v for k, v in f.items() if k not in ("detail_pola_baru", "detail_pola_diperbarui")}
                for f in ringkasan_per_file
            ],
        },
    )

    pola_sekarang = ak.muat_pola(ak._path_pola("pola_bank", client_id))

    return {
        "client_id": client_id,
        "total_pola_baru": total_pola_baru,
        "total_pola_diperbarui": total_pola_diperbarui,
        "total_baris_dibaca": total_baris_dibaca,
        "total_baris_dipakai": total_baris_dipakai,
        "total_pola_tersimpan_sekarang": len(pola_sekarang.aturan),
        "per_file": ringkasan_per_file,
    }


# ============================================================
# [BARU] KERTAS KERJA LAPORAN KEUANGAN (working paper) DARI PDF
# REKENING KORAN -- LANGKAH 1 dari alur "kertas kerja dulu, baru user
# konfirmasi generate laporan 18-sheet" (lihat modules/kertas_kerja.py
# untuk detail pipeline PDF -> GL -> Bank_Control -> Bank_Posting_Summary
# -> TB/BS/PNL_Monthly -> file .xlsx).
#
# Endpoint ini BARU membuat working paper-nya saja & mengembalikan
# ringkasan status (jumlah transaksi per confidence, status per bulan) --
# BELUM memanggil generate laporan 18-sheet final. Langkah konfirmasi
# user + endpoint generate laporan final menyusul terpisah.
# ============================================================

def _siapkan_df_coa_untuk_kertas_kerja(
    client_id: int, coa_bytes: Optional[bytes], nama_coa_file: Optional[str],
) -> Tuple[pd.DataFrame, List[str]]:
    """
    Sumber COA untuk kertas kerja, urutan prioritas:
      1. File Excel COA skema kertas kerja diupload di request ini
         (kolom Account No./Account Name/Normal Balance/Statement/
         FS Group/Notes -- lihat kertas_kerja.muat_coa_kertas_kerja).
      2. Fallback: COA client yang sudah tersimpan di database (skema
         lebih sederhana, dipetakan best-effort -- lihat
         kertas_kerja.bangun_coa_kertas_kerja_dari_db, PASTI menghasilkan
         peringatan karena kolom Statement/staging cuma ditebak).
    """
    if coa_bytes:
        try:
            wb_coa = openpyxl.load_workbook(io.BytesIO(coa_bytes), data_only=True)
        except Exception as e:  # noqa: BLE001
            raise HTTPException(
                status_code=400,
                detail=f"File COA '{nama_coa_file}' gagal dibaca sebagai Excel: {e}",
            )
        df_coa = kertas_kerja.muat_coa_kertas_kerja(wb_coa)
        if df_coa.empty:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Sheet 'COA' tidak ditemukan atau kosong di file '{nama_coa_file}'. "
                    "Pastikan ada sheet bernama mengandung 'COA' dengan header "
                    "Account No./Account Name/Normal Balance/Statement/FS Group/Notes."
                ),
            )
        return df_coa, []

    daftar_akun_db = dbc.ambil_coa_client(client_id)
    df_coa, peringatan = kertas_kerja.bangun_coa_kertas_kerja_dari_db(daftar_akun_db)
    if df_coa.empty:
        raise HTTPException(
            status_code=400,
            detail=(
                "COA client ini masih kosong (baik di database maupun tidak ada file "
                "COA yang diupload di request ini). Upload COA dulu (menu COA, atau "
                "sertakan file Excel COA skema kertas kerja lewat parameter coa_file)."
            ),
        )
    return df_coa, peringatan


@app.post("/api/client/{client_id}/generate-kertas-kerja")
async def api_generate_kertas_kerja(
    client_id: int,
    files: List[UploadFile] = File(...),
    coa_file: Optional[UploadFile] = File(None),
    pakai_ai: bool = Form(True),
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """
    [BARU] Generate "Kertas Kerja Laporan Keuangan" (working paper Excel)
    dari PDF rekening koran client -- LANGKAH 1 dari alur konfirmasi
    (lihat komentar di atas blok endpoint ini).

    Args (multipart/form-data):
        files: 1 atau banyak file PDF rekening koran (boleh lintas
            bulan/bank sekaligus, akan digabung jadi satu GL).
        coa_file: opsional, file Excel COA skema kertas kerja. Kalau
            tidak diisi, COA diambil dari database client (fallback,
            lihat _siapkan_df_coa_untuk_kertas_kerja).
        pakai_ai: default True -- teruskan ke akuntansi_ai.proses_dataframe
            untuk klasifikasi transaksi bank (AI + pola historis). Set
            False kalau hanya ingin memakai pola historis client saja
            (lebih cepat/murah, cocok untuk uji coba awal).

    Returns JSON:
        client_id, tahun (ditebak dari transaksi GL), nama_file,
        file_base64 (working paper .xlsx), ringkasan (lihat
        kertas_kerja.ringkasan_status_kertas_kerja), peringatan
        (gabungan semua peringatan proses, termasuk dari sumber COA).
    """
    if not files:
        raise HTTPException(status_code=400, detail="Tidak ada file PDF rekening koran yang diupload.")

    client = dbc.ambil_client(client_id)
    if not client:
        raise HTTPException(status_code=404, detail=f"Client dengan id {client_id} tidak ditemukan.")

    coa_bytes = await coa_file.read() if coa_file is not None else None
    df_coa, peringatan_coa = _siapkan_df_coa_untuk_kertas_kerja(
        client_id, coa_bytes, coa_file.filename if coa_file else None,
    )

    daftar_file_pdf: List[Tuple[Any, str]] = []
    nama_file_ditolak: List[str] = []
    for f in files:
        nama_file = f.filename or "upload.pdf"
        if not nama_file.lower().endswith(".pdf"):
            nama_file_ditolak.append(nama_file)
            continue
        isi = await f.read()
        buf = io.BytesIO(isi)
        buf.name = nama_file
        daftar_file_pdf.append((buf, nama_file))

    if not daftar_file_pdf:
        raise HTTPException(
            status_code=400,
            detail=(
                "Tidak ada file PDF yang valid untuk diproses "
                f"(file ditolak karena bukan .pdf: {nama_file_ditolak})."
            ),
        )

    # [FIX -- DEDUP] Badan logic (generate + tulis Excel + log audit) sekarang
    # ada di _jalankan_generate_kertas_kerja() (lihat blok versi stream di
    # bawah), dipakai ulang persis sama di sini -- sebelumnya kode ini
    # ter-duplikasi di 2 tempat, berisiko drift diam-diam kalau salah satu
    # diperbaiki tapi yang lain lupa.
    # [FIX -- GAP EVENT LOOP] _jalankan_generate_kertas_kerja() sync & berat
    # (ekstraksi PDF + klasifikasi AI + tulis Excel, bisa beberapa menit --
    # lihat catatan panjang di atas fungsinya). Versi /stream di bawah SUDAH
    # menjalankan ini di thread terpisah demi SSE; versi blocking ini
    # sebelumnya TIDAK, padahal fungsinya sendiri sudah didesain aman
    # dipanggil dari thread ("aman dipanggil dari thread biasa" -- lihat
    # docstring _jalankan_generate_kertas_kerja). Dibungkus to_thread di sini
    # supaya endpoint non-stream ini juga tidak memblokir event loop.
    return await asyncio.to_thread(
        _jalankan_generate_kertas_kerja,
        client_id, daftar_file_pdf, df_coa, peringatan_coa,
        nama_file_ditolak, pakai_ai, user,
    )


# ============================================================
# [BARU] VERSI STREAMING (SSE) DARI /generate-kertas-kerja
# ============================================================
# MASALAH: PDF rekening koran multi-halaman/multi-bulan bisa makan waktu
# beberapa menit (ekstraksi PDF + klasifikasi AI) -- endpoint blocking di
# atas (api_generate_kertas_kerja) menahan 1 koneksi HTTP selama itu, dan
# reverse proxy/browser BIASANYA timeout di 60-120 detik untuk koneksi yang
# tidak ada aktivitas data. Solusinya PERSIS pola yang sudah dipakai di
# /api/proses-file/stream di atas (lihat komentar lengkap di sana): jalankan
# kertas_kerja.generate_kertas_kerja() (kode SINKRON/blocking -- pandas,
# pdfplumber, requests ke DeepSeek) di THREAD terpisah, kirim event progress
# lewat queue ke browser sebagai SSE selagi thread masih jalan. Karena selalu
# ada byte yang dikirim tiap event progress, koneksi tidak pernah "diam" --
# ini yang membuat reverse proxy tidak menganggapnya timeout, BUKAN karena
# prosesnya jadi lebih cepat (durasi total tetap sama, cuma tidak lagi
# 1 request blocking tanpa respons apa pun sampai selesai).
#
# Endpoint BLOCKING lama (api_generate_kertas_kerja) SENGAJA TIDAK dihapus/
# diubah -- frontend yang belum sempat pindah ke versi stream tetap jalan
# seperti biasa. Badan logic (persiapan COA, validasi file, generate,
# tulis Excel, log audit) SENGAJA disatukan ke _jalankan_generate_kertas_kerja()
# di bawah supaya TIDAK ada logic yang di-duplikasi antara 2 endpoint ini --
# keduanya cuma beda cara mengirim hasil balik ke browser (JSON sekali vs
# event SSE bertahap).
def _jalankan_generate_kertas_kerja(
    client_id: int,
    daftar_file_pdf: List[Tuple[Any, str]],
    df_coa: pd.DataFrame,
    peringatan_coa: List[str],
    nama_file_ditolak: List[str],
    pakai_ai: bool,
    user: dict,
    progress_callback: Optional[Any] = None,
    tahun_override: Optional[int] = None,
) -> Dict[str, Any]:
    """Badan logic generate kertas kerja, dipakai ulang oleh endpoint
    blocking (api_generate_kertas_kerja) maupun versi stream (di bawah).
    Melempar HTTPException persis seperti sebelumnya kalau gagal -- ini
    aman dipanggil dari thread biasa (bukan cuma dari request handler
    FastAPI), HTTPException di sini cuma dipakai sebagai exception class
    yang sudah bawa status_code+detail, bukan benar2 di-raise ke FastAPI.

    Args tambahan:
        progress_callback: [BARU] diteruskan apa adanya ke
            kertas_kerja.generate_kertas_kerja() -- dipakai
            api_generate_kertas_kerja_per_file_stream() di bawah untuk
            melaporkan progress SSE per file PDF. None (default) untuk
            2 endpoint lama yang tidak butuh progress per-file.
        tahun_override: [BARU] kalau diisi, dipakai APA ADANYA sebagai
            tahun kertas kerja, menimpa tebakan otomatis dari
            tentukan_tahun_dari_gl() -- tebakan otomatis TETAP dijalankan
            (bukan dilewati) supaya peringatan "transaksi lintas tahun"
            tetap muncul kalau relevan, cuma nilai tahun akhirnya yang
            ditimpa oleh pilihan user.
    """
    try:
        hasil, peringatan = kertas_kerja.generate_kertas_kerja(
            daftar_file_pdf, df_coa, client_id=client_id, pakai_ai=pakai_ai,
            peringatan_awal=peringatan_coa, progress_callback=progress_callback,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Gagal generate kertas kerja: {e}")

    if nama_file_ditolak:
        hasil.peringatan.append(
            f"File berikut dilewati karena bukan PDF: {nama_file_ditolak}."
        )

    tahun, peringatan_tahun = kertas_kerja.tentukan_tahun_dari_gl(hasil.gl)
    hasil.peringatan.extend(peringatan_tahun)
    if tahun_override:
        tahun = tahun_override

    isi_excel = kertas_kerja.tulis_kertas_kerja_excel(hasil, tahun, identitas={})
    ringkasan = kertas_kerja.ringkasan_status_kertas_kerja(hasil)
    nama_file_output = f"Kertas_Kerja_Laporan_Keuangan_{tahun}.xlsx"

    dbc.log_audit(
        client_id=client_id,
        user=user.get("username", "unknown"),
        aksi="generate_kertas_kerja",
        detail={
            "tahun": tahun,
            "jumlah_file_pdf": len(daftar_file_pdf),
            "jumlah_transaksi": ringkasan["jumlah_transaksi"],
            "confidence_count": ringkasan["confidence_count"],
            "jumlah_peringatan": ringkasan["jumlah_peringatan"],
        },
    )

    return {
        "client_id": client_id,
        "tahun": tahun,
        "nama_file": nama_file_output,
        "file_base64": base64.b64encode(isi_excel).decode("ascii"),
        "ringkasan": ringkasan,
        "peringatan": hasil.peringatan,
    }


@app.post("/api/client/{client_id}/generate-kertas-kerja/stream")
async def api_generate_kertas_kerja_stream(
    client_id: int,
    files: List[UploadFile] = File(...),
    coa_file: Optional[UploadFile] = File(None),
    pakai_ai: bool = Form(True),
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """
    [BARU] Versi streaming (SSE) dari POST .../generate-kertas-kerja --
    lihat komentar blok di atas untuk alasan lengkap. Event terakhir
    sebelum "[DONE]" bertipe "result" dan skemanya PERSIS SAMA dengan
    response endpoint blocking (client_id, tahun, nama_file, file_base64,
    ringkasan, peringatan) -- jadi frontend cukup ganti CARA MEMANGGIL
    (baca event SSE bertahap), logika MEMBACA hasil akhir tidak berubah.

    File PDF & COA dibaca (await f.read()) DI SINI, SEBELUM masuk ke
    thread -- UploadFile/SpooledTemporaryFile FastAPI tidak aman dipakai
    dari thread lain sekaligus async event loop, jadi semua isi file
    sudah harus jadi bytes biasa dulu sebelum thread mulai (sama seperti
    pola di proses_file_stream di atas).
    """
    if not files:
        raise HTTPException(status_code=400, detail="Tidak ada file PDF rekening koran yang diupload.")

    client = dbc.ambil_client(client_id)
    if not client:
        raise HTTPException(status_code=404, detail=f"Client dengan id {client_id} tidak ditemukan.")

    coa_bytes = await coa_file.read() if coa_file is not None else None
    nama_coa_file = coa_file.filename if coa_file else None

    daftar_isi_pdf: List[Tuple[bytes, str]] = []
    nama_file_ditolak: List[str] = []
    for f in files:
        nama_file = f.filename or "upload.pdf"
        if not nama_file.lower().endswith(".pdf"):
            nama_file_ditolak.append(nama_file)
            continue
        isi = await f.read()
        daftar_isi_pdf.append((isi, nama_file))

    if not daftar_isi_pdf:
        raise HTTPException(
            status_code=400,
            detail=(
                "Tidak ada file PDF yang valid untuk diproses "
                f"(file ditolak karena bukan .pdf: {nama_file_ditolak})."
            ),
        )

    q: "queue.Queue" = queue.Queue()

    def jalankan():
        try:
            q.put({"type": "progress", "step": "coa", "label": "Menyiapkan Chart of Accounts", "status": "processing"})
            try:
                df_coa, peringatan_coa = _siapkan_df_coa_untuk_kertas_kerja(client_id, coa_bytes, nama_coa_file)
            except HTTPException as e:
                q.put({"type": "error", "pesan": str(e.detail)})
                return
            q.put({"type": "progress", "step": "coa", "label": "Menyiapkan Chart of Accounts", "status": "done"})

            # File dibungkus ulang jadi BytesIO baru per file di sini (bukan
            # di request handler) -- BytesIO murni in-memory, aman dipakai
            # lintas thread selama tidak diakses 2 thread BERSAMAAN, dan di
            # sini cuma thread ini yang menyentuhnya.
            daftar_file_pdf: List[Tuple[Any, str]] = []
            for isi, nama_file in daftar_isi_pdf:
                buf = io.BytesIO(isi)
                buf.name = nama_file
                daftar_file_pdf.append((buf, nama_file))

            q.put({
                "type": "progress", "step": "ekstraksi_klasifikasi",
                "label": f"Mengekstrak & mengklasifikasi transaksi dari {len(daftar_file_pdf)} file PDF"
                         + (" (dibantu AI)" if pakai_ai else " (tanpa AI, pola & kata kunci saja)"),
                "status": "processing",
            })
            try:
                hasil_akhir = _jalankan_generate_kertas_kerja(
                    client_id, daftar_file_pdf, df_coa, peringatan_coa,
                    nama_file_ditolak, pakai_ai, user,
                )
            except HTTPException as e:
                q.put({"type": "error", "pesan": str(e.detail)})
                return
            q.put({
                "type": "progress", "step": "ekstraksi_klasifikasi",
                "label": "Mengekstrak & mengklasifikasi transaksi", "status": "done",
            })

            q.put({"type": "progress", "step": "excel", "label": "Menyusun file Excel kertas kerja", "status": "done"})
            q.put({"type": "result", **hasil_akhir})
        except Exception as e:  # noqa: BLE001
            q.put({"type": "error", "pesan": str(e)})
        finally:
            q.put(None)  # sinyal: tidak ada event lagi

    threading.Thread(target=jalankan, daemon=True).start()

    def event_generator():
        while True:
            item = q.get()
            if item is None:
                break
            yield _format_sse_progress(**item)
        yield "data: [DONE]\n\n"

    return StreamingResponse(event_generator(), media_type="text/event-stream")


# ============================================================
# [BARU] VERSI STREAMING (SSE) PER-FILE DARI /generate-kertas-kerja
# ============================================================
# Dipakai KertasKerjaPage.jsx (lihat frontend/src/lib/api.js::
# generateKertasKerjaStream()) -- BEDA dari /generate-kertas-kerja/stream
# di atas: endpoint di atas melapor progress PER TAHAP (COA -> ekstraksi
# -> excel), endpoint ini melapor progress PER FILE PDF (queued ->
# processing -> done/cache_hit/error per file), karena
# susun_gl_dari_pdf_rekening_koran() memproses semua file itu PARALEL --
# untuk batch banyak bulan/bank, progress per-tahap terasa "diam" lama di
# 1 tahap ("ekstraksi_klasifikasi") tanpa user tahu file mana yang sudah/
# belum selesai.
#
# Path & skema event SENGAJA disamakan PERSIS dengan yang sudah ditulis
# di frontend (lihat catatan integrasi di api.js) supaya tidak perlu
# ubah frontend sama sekali:
#   {"type": "progress", "file": nama, "status": "queued"|"processing"|
#    "done"|"cache_hit"|"error", "pesan"?: str}
#   {"type": "result", client_id, tahun, nama_file, file_base64,
#    ringkasan, peringatan}   -- skema IDENTIK dgn endpoint blocking
#   {"type": "error", "pesan": str}
#
# Badan logic (COA, validasi file, generate, tulis Excel, log audit) tetap
# lewat _jalankan_generate_kertas_kerja() yang sama (lihat blok di atas) --
# cuma di sini progress_callback (dan tahun_override, lihat Form "tahun")
# diisi supaya event per-file ikut mengalir ke `q` selagi
# kertas_kerja.generate_kertas_kerja() masih berjalan di thread terpisah.
@app.post("/api/client/{client_id}/kertas-kerja/generate/stream")
async def api_generate_kertas_kerja_per_file_stream(
    client_id: int,
    files: List[UploadFile] = File(...),
    coa_file: Optional[UploadFile] = File(None),
    tahun: Optional[int] = Form(None),
    pakai_ai: bool = Form(True),
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """
    [BARU] Versi SSE dari generate kertas kerja dengan progress PER FILE
    PDF -- lihat komentar blok di atas untuk skema event & alasan lengkap.

    Args (multipart/form-data): sama seperti /generate-kertas-kerja,
    ditambah `tahun` opsional (int) -- kalau diisi, dipakai apa adanya
    sebagai tahun kertas kerja (menimpa tebakan otomatis dari
    tentukan_tahun_dari_gl, lihat _jalankan_generate_kertas_kerja).
    """
    if not files:
        raise HTTPException(status_code=400, detail="Tidak ada file PDF rekening koran yang diupload.")

    client = dbc.ambil_client(client_id)
    if not client:
        raise HTTPException(status_code=404, detail=f"Client dengan id {client_id} tidak ditemukan.")

    coa_bytes = await coa_file.read() if coa_file is not None else None
    nama_coa_file = coa_file.filename if coa_file else None

    daftar_isi_pdf: List[Tuple[bytes, str]] = []
    nama_file_ditolak: List[str] = []
    for f in files:
        nama_file = f.filename or "upload.pdf"
        if not nama_file.lower().endswith(".pdf"):
            nama_file_ditolak.append(nama_file)
            continue
        isi = await f.read()
        daftar_isi_pdf.append((isi, nama_file))

    if not daftar_isi_pdf:
        raise HTTPException(
            status_code=400,
            detail=(
                "Tidak ada file PDF yang valid untuk diproses "
                f"(file ditolak karena bukan .pdf: {nama_file_ditolak})."
            ),
        )

    q: "queue.Queue" = queue.Queue()

    def jalankan():
        try:
            # Lapor "queued" untuk SEMUA file PDF yang valid di depan, sebelum
            # COA/thread pool mulai -- supaya UI langsung menampilkan daftar
            # lengkap file dengan status awal, bukan muncul satu-satu belakangan.
            for _, nama_file in daftar_isi_pdf:
                q.put({"type": "progress", "file": nama_file, "status": "queued"})

            # [FIX -- GAP] File yang ditolak (bukan .pdf) sebelumnya TIDAK
            # pernah dilaporkan lewat SSE sama sekali -- di frontend
            # (KertasKerjaPage.jsx) file itu sudah ditampilkan dengan status
            # awal "queued" (di-set lokal begitu tombol Generate ditekan,
            # untuk SEMUA file yang dipilih user termasuk yang bukan PDF),
            # dan karena tidak ada event progress lanjutan untuk namanya,
            # baris itu akan macet selamanya di "Menunggu" walau proses lain
            # sudah selesai. Sekarang dilaporkan "error" segera di sini,
            # sebelum tahap COA/ekstraksi mulai.
            for nama_file in nama_file_ditolak:
                q.put({
                    "type": "progress", "file": nama_file, "status": "error",
                    "pesan": "Bukan file PDF -- hanya rekening koran PDF yang diterima.",
                })

            try:
                df_coa, peringatan_coa = _siapkan_df_coa_untuk_kertas_kerja(client_id, coa_bytes, nama_coa_file)
            except HTTPException as e:
                q.put({"type": "error", "pesan": str(e.detail)})
                return

            # File dibungkus ulang jadi BytesIO baru per file DI SINI (bukan
            # di request handler) -- sama seperti alasan di endpoint stream
            # per-tahap di atas: BytesIO murni in-memory, aman lintas thread
            # selama tidak diakses 2 thread BERSAMAAN.
            daftar_file_pdf: List[Tuple[Any, str]] = []
            for isi, nama_file in daftar_isi_pdf:
                buf = io.BytesIO(isi)
                buf.name = nama_file
                daftar_file_pdf.append((buf, nama_file))

            def _lapor_progress_file(nama_file_pdf: str, status: str, pesan: Optional[str] = None) -> None:
                # Dipanggil dari thread WORKER milik ThreadPoolExecutor di
                # dalam susun_gl_dari_pdf_rekening_koran (PDF_PARALEL_MAKS
                # worker sekaligus, BUKAN cuma thread `jalankan` ini) --
                # queue.Queue.put() thread-safe, jadi aman dipanggil dari
                # banyak worker bersamaan tanpa lock tambahan di sini.
                item: Dict[str, Any] = {"type": "progress", "file": nama_file_pdf, "status": status}
                if pesan:
                    item["pesan"] = pesan
                q.put(item)

            try:
                hasil_akhir = _jalankan_generate_kertas_kerja(
                    client_id, daftar_file_pdf, df_coa, peringatan_coa,
                    nama_file_ditolak, pakai_ai, user,
                    progress_callback=_lapor_progress_file, tahun_override=tahun,
                )
            except HTTPException as e:
                q.put({"type": "error", "pesan": str(e.detail)})
                return

            q.put({"type": "result", **hasil_akhir})
        except Exception as e:  # noqa: BLE001
            q.put({"type": "error", "pesan": str(e)})
        finally:
            q.put(None)  # sinyal: tidak ada event lagi

    threading.Thread(target=jalankan, daemon=True).start()

    def event_generator():
        while True:
            item = q.get()
            if item is None:
                break
            yield _format_sse_progress(**item)
        yield "data: [DONE]\n\n"

    return StreamingResponse(event_generator(), media_type="text/event-stream")


# ============================================================
# [BARU] LANGKAH 2: KONFIRMASI KERTAS KERJA -> LAPORAN 18-SHEET
# ============================================================
# Menutup gap #1 (jembatan konfirmasi -> 18-sheet, lihat komentar blok
# /generate-kertas-kerja di atas) DAN gap #2 (re-upload kertas kerja yang
# sudah dikoreksi) SEKALIGUS lewat SATU endpoint stateless: terima file
# kertas kerja .xlsx (boleh dikirim balik apa adanya oleh frontend begitu
# user klik "Ya, lanjut", ATAU versi yang sudah dikoreksi user di sheet
# Adjustments/Opening_Balance -- endpoint ini SELALU baca ulang dari file,
# jadi kedua kasus otomatis diperlakukan sama), baca ulang 4 sheet-nya
# lewat fungsi kertas_kerja.baca_*_dari_kertas_kerja() yang sudah ada,
# susun data lewat kertas_kerja.susun_data_export_18_sheet_dari_kertas_kerja()
# (jembatan yang sudah ada, TIDAK diubah), lalu serialize ke laporan
# 18-sheet final lewat accounting_export.export_18_sheet_lengkap()/
# export_18_sheet_sebagai_json() -- TIDAK ada perubahan sama sekali di
# accounting_export.py, karena bentuk data_export sudah persis sama
# dengan yang dipakai _susun_data_export_18_sheet() versi database.
#
# Desain sengaja stateless (tidak simpan HasilKertasKerja di memori server
# antar-request /generate-kertas-kerja -> endpoint ini) -- lihat catatan
# di docstring kertas_kerja.baca_gl_dari_kertas_kerja().

def _baca_kertas_kerja_untuk_bridge(
    isi_file: bytes, nama_file: str,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, int, List[str]]:
    """
    Baca 4 sheet (COA, GL, Adjustments, Opening_Balance) dari 1 file kertas
    kerja yang diupload, plus tebak tahun dari GL -- dipakai bareng oleh
    endpoint Excel & JSON di bawah (pola sama seperti
    _susun_data_export_18_sheet dipakai bareng _bangun_export_18_sheet &
    _bangun_preview_18_sheet_json).

    Setiap fungsi baca_*_dari_kertas_kerja() (kecuali muat_coa_kertas_kerja,
    yang menerima objek Workbook) membuka workbook-nya SENDIRI dari
    file_like yang diberikan (lihat modules/kertas_kerja.py) -- di sini
    masing-masing sengaja dikasih io.BytesIO(isi_file) BARU (bukan 1
    buffer yang di-seek ulang) supaya tidak saling mengganggu posisi baca.
    """
    try:
        wb_coa = openpyxl.load_workbook(io.BytesIO(isi_file), data_only=True)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(
            status_code=400,
            detail=f"File '{nama_file}' gagal dibaca sebagai Excel: {e}",
        )

    df_coa = kertas_kerja.muat_coa_kertas_kerja(wb_coa)
    if df_coa.empty:
        raise HTTPException(
            status_code=400,
            detail="Sheet 'COA' tidak ditemukan atau kosong di file kertas kerja yang diupload.",
        )

    try:
        df_gl = kertas_kerja.baca_gl_dari_kertas_kerja(io.BytesIO(isi_file))
        df_adjustments = kertas_kerja.baca_adjustments_dari_kertas_kerja(io.BytesIO(isi_file))
        df_opening = kertas_kerja.baca_opening_balance_dari_kertas_kerja(io.BytesIO(isi_file))
    except ValueError as e:
        # ValueError = sheet wajib tidak ditemukan/kosong -- pesan sudah
        # jelas dari kertas_kerja.py sendiri, teruskan apa adanya sbg 400
        # (sama seperti pola di /generate-kertas-kerja).
        raise HTTPException(status_code=400, detail=str(e))

    tahun, peringatan_tahun = kertas_kerja.tentukan_tahun_dari_gl(df_gl)
    return df_gl, df_coa, df_adjustments, df_opening, tahun, peringatan_tahun


class KonfirmasiKertasKerjaKe18SheetRequest(BaseModel):
    """
    [KETERBATASAN -- SENGAJA] Field yang ADA di Export18SheetRequest tapi
    TIDAK didukung di sini (kompensasi_kerugian_fiskal, kredit_pajak,
    skema_pajak, tambahan_peredaran_bruto_lainnya,
    retur_pengurangan_peredaran_bruto, keterangan_peredaran_bruto,
    metode_penyusutan, tahun_sebelumnya) memang belum diekspos --
    kertas_kerja.susun_data_export_18_sheet_dari_kertas_kerja() belum
    menerima parameter itu, dan PPh Badan 31E dihitung dengan koreksi
    fiskal default 0 (lihat KETERBATASAN di docstring fungsi itu).
    Akuntan WAJIB cek ulang pph_hasil secara manual sebelum SPT
    difinalkan -- sama seperti disclaimer yang sudah ada di kertas kerja
    bank-only ini.
    """
    nama_perusahaan: Optional[str] = None
    prive_atau_dividen: float = 0
    setoran_modal_baru: float = 0
    penyesuaian_ekuitas_manual: float = 0


def _bangun_data_export_18_sheet_dari_kertas_kerja(
    client_id: int, isi_file: bytes, nama_file: str,
    req: "KonfirmasiKertasKerjaKe18SheetRequest", user: dict,
) -> Tuple[dict, List[str]]:
    """Badan logic bareng utk endpoint Excel & JSON di bawah -- baca file,
    susun data_export lewat jembatan yang sudah ada, log audit."""
    df_gl, df_coa, df_adjustments, df_opening, tahun, peringatan_tahun = _baca_kertas_kerja_untuk_bridge(
        isi_file, nama_file,
    )

    data_export = kertas_kerja.susun_data_export_18_sheet_dari_kertas_kerja(
        df_gl=df_gl, df_coa=df_coa, df_adjustments=df_adjustments, df_opening=df_opening,
        tahun=tahun, nama_perusahaan=req.nama_perusahaan,
        prive_atau_dividen=req.prive_atau_dividen, setoran_modal_baru=req.setoran_modal_baru,
        penyesuaian_ekuitas_manual=req.penyesuaian_ekuitas_manual,
    )

    dbc.log_audit(
        client_id=client_id, user=user.get("username", "unknown"),
        aksi="konfirmasi_kertas_kerja_ke_18_sheet",
        detail={
            "tahun": tahun,
            "nama_file_kertas_kerja": nama_file,
            "jumlah_baris_jurnal": len(data_export.get("jurnal") or []),
            "jumlah_akun_coa": len(data_export.get("coa") or []),
            "jumlah_peringatan_tahun": len(peringatan_tahun),
        },
    )
    return data_export, peringatan_tahun


@app.post("/api/client/{client_id}/kertas-kerja/konfirmasi-ke-18-sheet")
async def api_konfirmasi_kertas_kerja_ke_18_sheet(
    client_id: int,
    file: UploadFile = File(...),
    nama_perusahaan: Optional[str] = Form(None),
    prive_atau_dividen: float = Form(0),
    setoran_modal_baru: float = Form(0),
    penyesuaian_ekuitas_manual: float = Form(0),
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas -- sama seperti generate-kertas-kerja
):
    """
    [BARU] LANGKAH 2 dari alur "kertas kerja dulu, baru konfirmasi generate
    18-sheet" (lihat komentar blok endpoint /generate-kertas-kerja di atas).

    Menerima file kertas kerja .xlsx -- BOLEH file yang barusan didownload
    dari /generate-kertas-kerja lalu dikirim balik apa adanya (user klik
    "Ya, lanjut" tanpa edit apapun), ATAU versi yang sudah dikoreksi user
    di sheet Adjustments/Opening_Balance (menutup gap #2 sekaligus, tanpa
    endpoint terpisah) -- endpoint ini SELALU baca ulang dari file, jadi
    kedua kasus diperlakukan identik.

    Tidak butuh state dari request /generate-kertas-kerja sebelumnya
    (stateless -- lihat catatan di kertas_kerja.baca_gl_dari_kertas_kerja).

    Args (multipart/form-data):
        file: file kertas kerja .xlsx (wajib ada sheet COA/GL/
            Adjustments/Opening_Balance -- persis hasil /generate-kertas-kerja
            atau /generate-kertas-kerja yang sudah dikoreksi manual).
        nama_perusahaan, prive_atau_dividen, setoran_modal_baru,
            penyesuaian_ekuitas_manual: sama seperti field bernama sama di
            Export18SheetRequest (lihat KonfirmasiKertasKerjaKe18SheetRequest
            utk field yang SENGAJA belum didukung).

    Catatan: client_id di URL TIDAK divalidasi terhadap isi file kertas
    kerja (cuma dipakai utk cek client ada & log audit) -- pastikan file
    yang diupload memang milik client yang benar.

    Returns: file .xlsx laporan 18-sheet final siap download. Kalau
    transaksi GL ternyata mencakup >1 tahun, peringatannya dikirim lewat
    header response `X-Peringatan-Tahun` (opsional dibaca frontend).
    """
    client = dbc.ambil_client(client_id)
    if not client:
        raise HTTPException(status_code=404, detail=f"Client dengan id {client_id} tidak ditemukan.")

    nama_file = file.filename or "kertas_kerja.xlsx"
    if not nama_file.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="File yang diupload harus berformat .xlsx.")
    isi_file = await file.read()

    req = KonfirmasiKertasKerjaKe18SheetRequest(
        nama_perusahaan=nama_perusahaan, prive_atau_dividen=prive_atau_dividen,
        setoran_modal_baru=setoran_modal_baru, penyesuaian_ekuitas_manual=penyesuaian_ekuitas_manual,
    )

    try:
        # [FIX -- GAP EVENT LOOP] Dua tahap berat di sini: (1) baca+susun
        # data export dari kertas kerja, (2) generate workbook 18-sheet
        # (accounting_export.export_18_sheet_lengkap -- openpyxl susun
        # banyak sheet). Keduanya sync, dibungkus to_thread masing-masing
        # supaya event loop tidak terblokir selama proses ini jalan.
        data_export, peringatan_tahun = await asyncio.to_thread(
            _bangun_data_export_18_sheet_dari_kertas_kerja,
            client_id, isi_file, nama_file, req, user,
        )
    except HTTPException:
        raise
    except Exception as e:  # noqa: BLE001
        raise HTTPException(
            status_code=500,
            detail=f"Gagal menyusun laporan 18-sheet dari kertas kerja: {e}",
        )

    isi_excel = await asyncio.to_thread(accounting_export.export_18_sheet_lengkap, data_export)
    tahun = data_export.get("periode", "")

    headers = {
        "Content-Disposition": f'attachment; filename="Laporan_Keuangan_{tahun}_18_Sheet.xlsx"',
    }
    if peringatan_tahun:
        # Header non-standar, opsional dibaca frontend -- aman diabaikan
        # kalau frontend belum baca header ini.
        headers["X-Peringatan-Tahun"] = " | ".join(peringatan_tahun)

    return StreamingResponse(
        io.BytesIO(isi_excel),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.post("/api/client/{client_id}/kertas-kerja/konfirmasi-ke-18-sheet-json")
async def api_konfirmasi_kertas_kerja_ke_18_sheet_json(
    client_id: int,
    file: UploadFile = File(...),
    nama_perusahaan: Optional[str] = Form(None),
    prive_atau_dividen: float = Form(0),
    setoran_modal_baru: float = Form(0),
    penyesuaian_ekuitas_manual: float = Form(0),
    user: dict = Depends(auth.require_level(3)),
):
    """
    [BARU] Versi JSON dari POST .../kertas-kerja/konfirmasi-ke-18-sheet --
    supaya frontend bisa tampilkan preview ke-18 sheet LANGSUNG DI LAYAR
    dulu sebelum user download file-nya (pola sama dgn export-18-sheet-json,
    lihat _bangun_preview_18_sheet_json). Parameter sama persis dengan
    versi Excel di atas.
    """
    client = dbc.ambil_client(client_id)
    if not client:
        raise HTTPException(status_code=404, detail=f"Client dengan id {client_id} tidak ditemukan.")

    nama_file = file.filename or "kertas_kerja.xlsx"
    if not nama_file.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="File yang diupload harus berformat .xlsx.")
    isi_file = await file.read()

    req = KonfirmasiKertasKerjaKe18SheetRequest(
        nama_perusahaan=nama_perusahaan, prive_atau_dividen=prive_atau_dividen,
        setoran_modal_baru=setoran_modal_baru, penyesuaian_ekuitas_manual=penyesuaian_ekuitas_manual,
    )

    try:
        # [FIX -- GAP EVENT LOOP] Sama alasannya dgn versi Excel di atas.
        data_export, peringatan_tahun = await asyncio.to_thread(
            _bangun_data_export_18_sheet_dari_kertas_kerja,
            client_id, isi_file, nama_file, req, user,
        )
    except HTTPException:
        raise
    except Exception as e:  # noqa: BLE001
        raise HTTPException(
            status_code=500,
            detail=f"Gagal menyusun laporan 18-sheet dari kertas kerja: {e}",
        )

    hasil_json = await asyncio.to_thread(accounting_export.export_18_sheet_sebagai_json, data_export)
    hasil_json["peringatan_tahun"] = peringatan_tahun
    return hasil_json


# ============================================================
# [BARU - Prioritas #6, jalur langsung] UPLOAD FILE HASIL KOREKSI
# ============================================================
# Melengkapi /api/client/{id}/retrain-pola (yang menarik dari jawaban
# fitur klarifikasi di UI): endpoint ini membaca LANGSUNG file Excel
# hasil export format akuntan yang sudah dikoreksi manual oleh akuntan
# (kolom NO AKUN di baris kuning/merah sudah diisi/dibetulkan), untuk
# akuntan yang terbiasa mengedit file kerja Excel-nya sendiri.

@app.post("/api/client/{client_id}/upload-hasil-koreksi")
async def api_upload_hasil_koreksi(
    client_id: int,
    file: UploadFile = File(...),
    min_samples: int = Form(1),
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """
    Upload file HASIL EXPORT format akuntan (dari endpoint
    export-format-akuntan) yang sudah dikoreksi manual oleh akuntan --
    baris kuning/merah sudah diisi/dibetulkan kode akunnya. Baris yang
    valid (kode akun ada & ditemukan di sheet COA file itu sendiri)
    dipelajari jadi pola baru, digabung ke pola_bank_client_{id}.json
    yang sudah ada -- SAMA seperti /api/client/{id}/retrain-pola, cuma
    sumber datanya file Excel langsung, bukan jawaban klarifikasi di UI.
    """
    isi = await file.read()
    nama_file = file.filename or "hasil_koreksi.xlsx"
    buf = io.BytesIO(isi)
    buf.name = nama_file
    try:
        # [FIX -- GAP EVENT LOOP] proses_file_hasil_koreksi_akuntan() sync
        # & berat (parsing Excel + pelajari pola) -- dibungkus to_thread
        # sama alasannya dgn endpoint upload lain.
        hasil = await asyncio.to_thread(
            ak.proses_file_hasil_koreksi_akuntan,
            buf, nama_file, client_id=client_id, min_samples=min_samples,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    dbc.log_audit(
        client_id=client_id,
        user=user.get("username", "unknown"),
        aksi="upload_hasil_koreksi_akuntan",
        detail={
            "nama_file": nama_file,
            "jumlah_baris_dibaca": hasil["jumlah_baris_dibaca"],
            "jumlah_baris_dipakai": hasil["jumlah_baris_dipakai"],
            "jumlah_pola_baru": hasil["jumlah_pola_baru"],
            "jumlah_pola_diperbarui": hasil["jumlah_pola_diperbarui"],
        },
    )
    return {"client_id": client_id, "nama_file": nama_file, **hasil}


# ============================================================
# [BARU - Prioritas #3] METRIK KATEGORISASI OTOMATIS
# ============================================================
# Supaya "seberapa dekat hasil otomatis ke hasil manual akuntan" bisa
# diukur LANGSUNG begitu rekening koran mentah diupload -- tanpa
# akuntan/kamu perlu hitung baris satu-satu.

@app.get("/api/client/{client_id}/hasil/{hasil_id}/metrik-kategorisasi")
def api_metrik_kategorisasi(
    client_id: int, hasil_id: int,
    user: dict = Depends(auth.get_current_user),
):
    """
    Hitung % baris yang berhasil otomatis dikategorikan (dipecah per
    sumber: pola historis / kata kunci COA / AI / data asli dari file /
    belum terkategori) untuk SATU hasil upload rekening koran (hasil_id
    dari response /api/proses-file). Pakai ini setelah upload rekening
    koran MENTAH client baru untuk dapat angka konkret Prioritas #3,
    bukan tebak-tebakan.
    """
    hasil = dbc.ambil_hasil_by_id(hasil_id)
    if hasil is None or hasil.get("client_id") != client_id:
        raise HTTPException(status_code=404, detail="Hasil tidak ditemukan untuk client ini.")
    data = hasil.get("data") or {}
    df_hasil = pd.DataFrame(data.get("df") or [])
    metrik = accounting_export.hitung_metrik_kategorisasi(df_hasil)
    return {"client_id": client_id, "hasil_id": hasil_id, "nama_file": data.get("nama_file"), **metrik}


@app.get("/api/client/{client_id}/pola-bank")
def api_lihat_pola_bank(client_id: int, user: dict = Depends(auth.get_current_user)):
    """
    [BARU] Lihat isi pola_bank_client_{client_id}.json apa adanya -- untuk
    verifikasi hasil bootstrap (item di atas) atau pola yang terkumpul dari
    upload rutin, tanpa perlu upload file apa pun.
    """
    pola = ak.muat_pola(ak._path_pola("pola_bank", client_id))
    daftar = [
        {
            "signature": sig, "arah": arah,
            "akun_debet": f"{a['no_akun_debet']} - {a['nama_akun_debet']}",
            "akun_kredit": f"{a['no_akun_kredit']} - {a['nama_akun_kredit']}",
            "jumlah_contoh": a.get("jumlah_contoh"),
            "confidence_score": a.get("confidence_score"),
            "is_valid": a.get("is_valid"),
            "last_updated": a.get("last_updated"),
        }
        for (sig, arah), a in pola.aturan.items()
    ]
    daftar.sort(key=lambda x: x["signature"])
    return {"client_id": client_id, "jumlah_pola": len(daftar), "pola": daftar}


# ============================================================
# [BARU - Prioritas #6] LATIH ULANG POLA DARI FEEDBACK KLARIFIKASI
# ============================================================
# Menutup loop feedback: akuntan menjawab pertanyaan klarifikasi lewat
# /api/klarifikasi/{id}/jawab (jawaban otomatis tercatat ke tabel
# pola_augmentasi -- lihat dbc.jawab_pertanyaan_klarifikasi), tapi
# ak.latih_ulang_pola_dari_feedback() yang MENGOLAH feedback itu jadi
# pola baru belum pernah dipanggil dari endpoint mana pun. Semua fungsi
# berat SUDAH ada (ak.bangun_pola_dari_feedback_klarifikasi,
# ak.latih_ulang_pola_dari_feedback, dbc.ambil_pola_augmentasi,
# dbc.ambil_coa_client) -- endpoint ini murni menyambungkannya, tidak
# menambah logic baru di akuntansi_ai.py/db_client.py.

@app.post("/api/client/{client_id}/retrain-pola")
def api_retrain_pola(
    client_id: int,
    jenis: Optional[str] = None,  # "rekening_koran" | "penjualan" | None (proses keduanya)
    limit: int = 500,
    paksa_commit: bool = False,  # [BARU] lanjutkan commit walau evaluasi staging menyarankan review manual
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """
    Latih ulang pola_bank_client_{id}.json / pola_penjualan_client_{id}.json
    dari feedback klarifikasi yang sudah terkumpul (konfirmasi "ya benar"
    atau koreksi ke akun lain) -- supaya transaksi serupa berikutnya bisa
    langsung dikenali otomatis, tidak perlu ditanya manual lagi ke akuntan.

    [BARU -- STAGING] Sebelum benar-benar commit, batch feedback diuji dulu
    lewat ak.evaluasi_pola_sebelum_commit() (pola kandidat dilatih dari
    sebagian feedback, diuji ke sisanya yang sengaja disisihkan). Kalau
    hasilnya "PERLU_REVIEW_MANUAL" (akurasi terlalu rendah atau justru lebih
    buruk dari pola yang sedang live), retraining DITAHAN -- tidak ada
    perubahan tersimpan -- dan response berisi detail evaluasi supaya
    supervisor bisa putuskan manual. Kirim ulang dengan paksa_commit=true
    kalau setelah dicek tetap mau lanjut.

    jenis=None (default) memproses "rekening_koran" DAN "penjualan"
    sekaligus dalam 1 panggilan. Aman dipanggil berkali-kali/dijadwalkan
    rutin -- feedback yang sudah pernah dipakai tetap ada di tabel
    pola_augmentasi (tidak dihapus), gabung_pola() di akuntansi_ai.py akan
    menimpa pola lama dengan yang baru kalau signature+arah sama, bukan
    dobel-tambah.
    """
    daftar_jenis = [jenis] if jenis else sorted(_JENIS_DENGAN_POLA_PER_CLIENT)
    tidak_dikenal = [j for j in daftar_jenis if j not in _JENIS_DENGAN_POLA_PER_CLIENT]
    if tidak_dikenal:
        raise HTTPException(
            status_code=400,
            detail=f"jenis tidak dikenal: {tidak_dikenal}. Pilih dari {sorted(_JENIS_DENGAN_POLA_PER_CLIENT)}.",
        )

    df_coa = pd.DataFrame(dbc.ambil_coa_client(client_id))

    hasil_per_jenis: Dict[str, Any] = {}
    for j in daftar_jenis:
        augmentasi_rows = dbc.ambil_pola_augmentasi(client_id=client_id, jenis=j, limit=limit)
        if not augmentasi_rows:
            hasil_per_jenis[j] = {
                "jenis": j,
                "jumlah_feedback_total": 0,
                "pesan": "Belum ada feedback klarifikasi tersimpan untuk jenis ini.",
            }
            continue
        hasil_per_jenis[j] = ak.latih_ulang_pola_dengan_staging(
            augmentasi_rows, jenis=j, client_id=client_id, df_coa=df_coa,
            paksa_commit=paksa_commit,
        )

    dbc.log_audit(
        client_id=client_id,
        user=user.get("username", "unknown"),
        aksi="retrain_pola_dari_feedback",
        detail={
            "jenis_diproses": daftar_jenis, "paksa_commit": paksa_commit,
            "hasil": hasil_per_jenis,
        },
    )

    return {"client_id": client_id, "hasil": hasil_per_jenis}


# ============================================================
# [BARU] RIWAYAT VERSI & ROLLBACK POLA
# ============================================================
# Melengkapi endpoint retrain-pola di atas: setiap kali pola_bank/
# pola_penjualan client ini DITIMPA -- baik lewat retrain-pola di atas,
# bootstrap, atau upload rutin (proses_file_rekening_koran/penjualan) --
# ak.simpan_pola() di akuntansi_ai.py OTOMATIS menyimpan snapshot versi
# sebelumnya ke pola_data/versi_pola/. 2 endpoint di bawah ini HANYA
# menyambungkan fungsi yang sudah ada di sana (ak.daftar_versi_pola,
# ak.rollback_pola) -- tidak ada logic baru di akuntansi_ai.py.
#
# Kegunaan utama: kalau 1 jawaban klarifikasi/koreksi akuntan yang barusan
# di-retrain ternyata SALAH (typo, salah pilih akun), supervisor bisa
# membatalkannya tanpa perlu investigasi manual atau kehilangan pola lain
# yang sudah benar sebelumnya.

@app.get("/api/client/{client_id}/pola/{jenis}/riwayat")
def api_riwayat_pola(
    client_id: int,
    jenis: str,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """
    Daftar riwayat perubahan pola_bank/pola_penjualan client ini, TERBARU
    DULU -- dipakai UI utk menampilkan histori sebelum supervisor memutuskan
    mau rollback ke titik mana (lihat endpoint rollback di bawah).

    jenis harus salah satu dari _JENIS_DENGAN_POLA_PER_CLIENT
    ("rekening_koran" atau "penjualan").

    CATATAN BACA (diteruskan dari ak.daftar_versi_pola docstring): field
    "sumber_perubahan" pada tiap entri adalah label PERUBAHAN YANG DIBATALKAN
    kalau rollback dilakukan ke snapshot itu -- BUKAN label siapa yang
    membuat state snapshot tsb. Entri PALING ATAS (terbaru) = "batalkan
    perubahan paling terakhir terjadi".
    """
    if jenis not in _JENIS_DENGAN_POLA_PER_CLIENT:
        raise HTTPException(
            status_code=400,
            detail=f"jenis tidak dikenal: {jenis!r}. Pilih dari {sorted(_JENIS_DENGAN_POLA_PER_CLIENT)}.",
        )
    path_pola = ak._path_pola(_POLA_PER_JENIS[jenis], client_id)
    riwayat = ak.daftar_versi_pola(path_pola)
    return {"client_id": client_id, "jenis": jenis, "jumlah_versi": len(riwayat), "riwayat": riwayat}


class RollbackPolaRequest(BaseModel):
    nama_file_snapshot: Optional[str] = None  # None = otomatis batalkan perubahan terakhir


@app.post("/api/client/{client_id}/pola/{jenis}/rollback")
def api_rollback_pola(
    client_id: int,
    jenis: str,
    req: RollbackPolaRequest,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas -- sama dgn retrain-pola
):
    """
    Kembalikan pola_bank/pola_penjualan client ini ke salah satu titik di
    riwayat (GET /riwayat di atas).

    req.nama_file_snapshot=None (default) = batalkan HANYA perubahan
    TERAKHIR yang terjadi -- kasus paling umum: 1 klarifikasi/koreksi
    barusan ternyata salah input. Untuk mundur lebih jauh, ambil salah
    satu "nama_file_snapshot" dari GET /riwayat lalu kirim eksplisit di
    sini.

    Rollback SENDIRI otomatis tercatat lagi sbg snapshot baru (lihat
    ak.rollback_pola docstring di akuntansi_ai.py) -- jadi salah pilih versi
    saat rollback pun masih bisa dibatalkan lagi, tidak destruktif.
    """
    if jenis not in _JENIS_DENGAN_POLA_PER_CLIENT:
        raise HTTPException(
            status_code=400,
            detail=f"jenis tidak dikenal: {jenis!r}. Pilih dari {sorted(_JENIS_DENGAN_POLA_PER_CLIENT)}.",
        )
    path_pola = ak._path_pola(_POLA_PER_JENIS[jenis], client_id)
    hasil = ak.rollback_pola(path_pola, nama_file_snapshot=req.nama_file_snapshot)
    if not hasil["sukses"]:
        # Riwayat kosong / snapshot tidak ditemukan -- kesalahan permintaan
        # user (client_id/jenis/nama_file_snapshot tidak cocok apapun),
        # bukan error server.
        raise HTTPException(status_code=404, detail=hasil["pesan"])

    dbc.log_audit(
        client_id=client_id,
        user=user.get("username", "unknown"),
        aksi="rollback_pola",
        detail={
            "jenis": jenis,
            "nama_file_snapshot_diminta": req.nama_file_snapshot,
            "hasil": hasil,
        },
    )

    return {"client_id": client_id, "jenis": jenis, **hasil}


# ============================================================
# [BARU] METRIK AKURASI TERPUSAT -- lihat ak.hitung_tren_akurasi /
# ak.catat_metrik_akurasi di akuntansi_ai.py.
# ============================================================
# Endpoint ini MURNI baca (tidak menulis apa pun) -- data yang dibaca
# sudah otomatis tercatat sejak sekarang setiap kali endpoint
# /api/client/{id}/retrain-pola di atas dipanggil (ak.latih_ulang_pola_
# dari_feedback sudah memanggil ak.catat_metrik_akurasi di dalamnya).
# Tidak perlu retroaktif -- histori mulai terbentuk dari titik ini ke
# depan; retrain-pola yang sudah pernah dijalankan SEBELUM fungsi ini
# ada tidak tercatat (datanya sudah hilang, tidak ada cara mengambil lagi).
#
# Dipakai UI utk menjawab "AI-nya makin akurat atau makin ngaco bulan
# ini?" per client -- sebelumnya pertanyaan ini tidak bisa dijawab sama
# sekali, cuma ada skor kualitas DATA (metrik-kategorisasi di atas),
# bukan skor akurasi KEPUTUSAN AI dari waktu ke waktu.

@app.get("/api/client/{client_id}/metrik-akurasi")
def api_metrik_akurasi(
    client_id: int,
    jenis: Optional[str] = None,  # "rekening_koran" | "penjualan" | None (gabungan keduanya)
    n_bulan_terakhir: int = 6,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas -- sama seperti retrain-pola
):
    """
    Tren akurasi kategorisasi AI per bulan untuk client ini, dihitung dari
    histori konfirmasi ("AI benar, akuntan tidak perlu koreksi") vs koreksi
    ("AI salah, akuntan membetulkan") yang terkumpul tiap kali retrain-pola
    dijalankan.

    jenis=None (default) menggabungkan rekening_koran + penjualan jadi satu
    tren -- kirim jenis eksplisit kalau supervisor mau lihat per jenis
    dokumen terpisah.

    Response "tren" bisa "NAIK" / "TURUN" / "STABIL" (dibanding bulan
    sebelumnya) atau "DATA_BELUM_CUKUP" kalau belum ada 2 bulan data.
    """
    if jenis is not None and jenis not in _JENIS_DENGAN_POLA_PER_CLIENT:
        raise HTTPException(
            status_code=400,
            detail=f"jenis tidak dikenal: {jenis!r}. Pilih dari {sorted(_JENIS_DENGAN_POLA_PER_CLIENT)}.",
        )

    if jenis is not None:
        hasil = ak.hitung_tren_akurasi(client_id, jenis=jenis, n_bulan_terakhir=n_bulan_terakhir)
        return {"client_id": client_id, "jenis": jenis, **hasil}

    # jenis=None -- gabungkan rekening_koran + penjualan jadi satu tren
    # bulanan (dijumlahkan per bulan), plus detail per-jenis terpisah utk
    # supervisor yang mau drill-down.
    per_jenis = {
        j: ak.hitung_tren_akurasi(client_id, jenis=j, n_bulan_terakhir=n_bulan_terakhir)
        for j in sorted(_JENIS_DENGAN_POLA_PER_CLIENT)
    }
    gabungan_per_bulan: Dict[str, Dict[str, int]] = {}
    for hasil_jenis in per_jenis.values():
        for entri in hasil_jenis["per_bulan"]:
            agg = gabungan_per_bulan.setdefault(entri["bulan"], {"jumlah_konfirmasi": 0, "jumlah_koreksi": 0})
            agg["jumlah_konfirmasi"] += entri["jumlah_konfirmasi"]
            agg["jumlah_koreksi"] += entri["jumlah_koreksi"]

    per_bulan_gabungan = []
    for bulan in sorted(gabungan_per_bulan.keys()):
        agg = gabungan_per_bulan[bulan]
        total = agg["jumlah_konfirmasi"] + agg["jumlah_koreksi"]
        per_bulan_gabungan.append({
            "bulan": bulan, "total": total,
            "jumlah_konfirmasi": agg["jumlah_konfirmasi"], "jumlah_koreksi": agg["jumlah_koreksi"],
            "akurasi_persen": round(agg["jumlah_konfirmasi"] / total * 100, 2) if total else None,
        })

    tren_gabungan, selisih_gabungan = "DATA_BELUM_CUKUP", None
    valid = [b for b in per_bulan_gabungan if b["akurasi_persen"] is not None]
    if len(valid) >= 2:
        selisih_gabungan = round(valid[-1]["akurasi_persen"] - valid[-2]["akurasi_persen"], 2)
        tren_gabungan = "NAIK" if selisih_gabungan > 1.0 else ("TURUN" if selisih_gabungan < -1.0 else "STABIL")

    total_konfirmasi_semua = sum(b["jumlah_konfirmasi"] for b in per_bulan_gabungan)
    total_koreksi_semua = sum(b["jumlah_koreksi"] for b in per_bulan_gabungan)
    total_semua = total_konfirmasi_semua + total_koreksi_semua

    return {
        "client_id": client_id,
        "jenis": None,
        "per_bulan": per_bulan_gabungan,
        "akurasi_keseluruhan": round(total_konfirmasi_semua / total_semua * 100, 2) if total_semua else None,
        "tren": tren_gabungan,
        "selisih_persen_poin": selisih_gabungan,
        "per_jenis": per_jenis,
    }


def _format_sse_progress(**kwargs) -> str:
    """Bungkus satu event progress jadi 1 baris SSE. Field 'type' selalu
    ada: 'progress' (satu langkah update), 'result' (hasil akhir, skema
    IDENTIK dengan response /api/proses-file biasa), atau 'error' (gagal
    total, mis. file tidak bisa dibuka sama sekali)."""
    return f"data: {json.dumps(kwargs, ensure_ascii=False, default=str)}\n\n"


@app.post("/api/proses-file/stream")
async def proses_file_stream(
    file: UploadFile = File(...),
    jenis_dokumen: Optional[str] = Form(None),
    client_id: Optional[int] = Form(None),
    conv_id: Optional[str] = Form(None),
    esb_account_id: Optional[int] = Form(None),
    user: dict = Depends(auth.require_level(3)),
):
    """
    [BARU] Versi streaming (SSE) dari /api/proses-file -- supaya frontend
    bisa menampilkan progress step-by-step ("Membaca file...", "Mendeteksi
    Rekening Koran...", "Menyimpan hasil...", dst) alih-alih loading kosong
    lalu hasil muncul sekaligus di akhir.

    Event terakhir sebelum "[DONE]" bertipe "result" dan skemanya PERSIS
    SAMA dengan response /api/proses-file biasa ({nama_file, hasil,
    tidak_terdeteksi}) -- jadi frontend cukup ganti CARA MEMANGGIL endpoint
    ini (baca event SSE, bukan satu Promise), logika MEMBACA hasil di akhir
    tidak perlu berubah.

    Kenapa pakai thread terpisah (bukan langsung jalan di event loop
    FastAPI): _proses_semua_jenis, penyimpanan ke DB (db_client.py), dan
    deteksi anomali semuanya kode SINKRON/blocking (pandas, requests ke AI,
    query DB biasa) -- BUKAN async. Kalau dijalankan langsung di sini,
    Python akan menjalankannya sampai selesai dulu sebelum sempat yield
    event apa pun ke browser -- persis masalah "loading diam lalu muncul
    sekaligus" yang mau diperbaiki. Callback on_progress menaruh tiap event
    ke queue (q); event_generator() di bawah membaca queue itu terus-
    menerus SELAGI thread masih jalan, sehingga event benar-benar terkirim
    real-time ke browser.
    """
    isi = await file.read()
    nama_file = file.filename or "upload.xlsx"
    q: "queue.Queue" = queue.Queue()

    def jalankan():
        def on_progress(kode, label, status, **extra):
            q.put({"type": "progress", "step": kode, "label": label, "status": status, **extra})

        try:
            q.put({"type": "progress", "step": "baca_file", "label": "Membaca file", "status": "done"})

            # [BARU] PDF rekening koran -> working paper (Kertas Kerja),
            # BUKAN alur "proses semua jenis dokumen" + auto laporan
            # 18-sheet di bawah. Sebelumnya endpoint ini selalu memicu
            # _auto_generate_laporan_18_sheet() untuk SEMUA jenis dokumen
            # begitu tahun & COA client tersedia -- padahal untuk PDF
            # rekening koran akuntan butuh working paper (GL + Bank_Control
            # + Bank_Posting_Summary + TB/BS/PNL_Monthly, lihat
            # modules/kertas_kerja.py) supaya bisa DIKOREKSI DULU, bukan
            # laporan final langsung jadi dari klasifikasi mentah.
            # Dicek dari EKSTENSI FILE saja (bukan hasil deteksi jenis
            # dokumen) supaya tidak perlu jalankan _proses_semua_jenis()
            # dulu baru ketahuan ini rekening koran -- PDF yang diupload
            # lewat kotak chat SELALU diasumsikan rekening koran, konsisten
            # dengan satu-satunya jenis dokumen PDF yang didukung
            # kertas_kerja.py saat ini. Hanya berlaku kalau ada client_id
            # aktif (working paper butuh COA client); kalau tidak ada
            # client_id, tetap jatuh ke alur lama di bawah.
            if nama_file.lower().endswith(".pdf") and client_id is not None:
                q.put({
                    "type": "progress", "step": "kertas_kerja_coa",
                    "label": "Menyiapkan COA untuk kertas kerja", "status": "processing",
                })
                df_coa, peringatan_coa = _siapkan_df_coa_untuk_kertas_kerja(client_id, None, None)
                q.put({
                    "type": "progress", "step": "kertas_kerja_coa",
                    "label": "Menyiapkan COA untuk kertas kerja", "status": "done",
                })

                buf = io.BytesIO(isi)
                buf.name = nama_file

                def on_progress_pdf(nama_file_pdf, status, pesan=None):
                    q.put({
                        "type": "progress", "step": "kertas_kerja_pdf",
                        "label": f"Ekstraksi {nama_file_pdf}", "status": status,
                        "pesan": pesan,
                    })

                # [FIX -- DEDUP] Pakai ulang _jalankan_generate_kertas_kerja()
                # yang sudah ada (dipakai juga oleh /generate-kertas-kerja
                # dan /generate-kertas-kerja/stream) -- supaya logic generate
                # + tulis Excel + log audit TIDAK ter-duplikasi di 3 tempat.
                hasil_kk = _jalankan_generate_kertas_kerja(
                    client_id, [(buf, nama_file)], df_coa, peringatan_coa,
                    nama_file_ditolak=[], pakai_ai=True, user=user,
                    progress_callback=on_progress_pdf,
                )
                q.put({
                    "type": "result", "nama_file": nama_file,
                    "hasil": {}, "tidak_terdeteksi": False,
                    "kertas_kerja": hasil_kk,
                })
                return

            hasil_semua, error_per_jenis = _proses_semua_jenis(
                isi, nama_file, jenis_dokumen, client_id, on_progress=on_progress
            )

            if not hasil_semua:
                q.put({
                    "type": "result",
                    "nama_file": nama_file,
                    "hasil": {},
                    "tidak_terdeteksi": True,
                    "pesan": "Tidak ada jenis dokumen yang dikenali di file ini.",
                    "detail_error": error_per_jenis or None,
                })
                return

            hasil_json = _bersihkan_untuk_json(hasil_semua)

            if client_id is not None:
                q.put({"type": "progress", "step": "simpan", "label": "Menyimpan hasil ke riwayat client", "status": "processing"})

                conv_id_final = conv_id or datetime.now().isoformat()
                for kode, hasil in hasil_json.items():
                    data_disimpan = dict(hasil)
                    data_disimpan["nama_file"] = nama_file

                    koreksi_otomatis = hasil.get("koreksi_otomatis") or []
                    if koreksi_otomatis:
                        dbc.log_audit(
                            client_id=client_id,
                            user=user.get("username", "unknown"),
                            aksi="auto_fix_data",
                            detail={
                                "jenis_dokumen": kode,
                                "nama_file": nama_file,
                                "total_koreksi": len(koreksi_otomatis),
                                "koreksi": koreksi_otomatis[:200],
                            },
                        )

                    if esb_account_id is not None:
                        dbc.simpan_hasil_esb(client_id, esb_account_id, conv_id_final, kode, data_disimpan)
                    else:
                        dbc.simpan_hasil(client_id, conv_id_final, kode, data_disimpan)

                        draf_jurnal = data_disimpan.get("draf_jurnal") or []
                        if draf_jurnal:
                            hasil_tersimpan = dbc.ambil_hasil_client(client_id, jenis=kode, limit=1)
                            hasil_id = hasil_tersimpan[0]["id"] if hasil_tersimpan else None
                            dbc.tarik_draf_jurnal_ke_posting(client_id, hasil_id, kode, draf_jurnal)

                q.put({"type": "progress", "step": "simpan", "label": "Menyimpan hasil ke riwayat client", "status": "done"})

                q.put({"type": "progress", "step": "reminder_spt", "label": "Mengecek reminder deadline SPT", "status": "processing"})
                hasil_mentah_spt = hasil_semua.get("spt_masa")
                if hasil_mentah_spt:
                    df_spt_mentah = hasil_mentah_spt.get("df")
                    item_reminder = notifikasi.ekstrak_item_reminder_dari_df(df_spt_mentah)
                    if item_reminder:
                        dbc.simpan_reminder_deadline_spt(client_id, item_reminder)
                q.put({"type": "progress", "step": "reminder_spt", "label": "Mengecek reminder deadline SPT", "status": "done"})

                q.put({"type": "progress", "step": "klarifikasi", "label": "Mencari baris yang perlu klarifikasi", "status": "processing"})
                for kode in _JENIS_DENGAN_POLA_PER_CLIENT:
                    hasil_mentah = hasil_semua.get(kode)
                    if not hasil_mentah:
                        continue
                    df_mentah = hasil_mentah.get("df")
                    for item in ak.cari_baris_perlu_klarifikasi(df_mentah):
                        dbc.buat_pertanyaan_klarifikasi(
                            client_id=client_id,
                            jenis=kode,
                            pertanyaan=item["pertanyaan"],
                            conv_id=conv_id_final,
                            baris_index=item["baris_index"],
                            konteks=item,
                            tebakan_kategori=item.get("tebakan_kategori"),
                            butuh_konfirmasi_saja=item.get("butuh_konfirmasi_saja", False),
                        )
                q.put({"type": "progress", "step": "klarifikasi", "label": "Mencari baris yang perlu klarifikasi", "status": "done"})

                q.put({"type": "progress", "step": "anomali", "label": "Mendeteksi anomali & pola mencurigakan", "status": "processing"})
                for kode in _JENIS_DENGAN_POLA_PER_CLIENT:
                    hasil_mentah = hasil_semua.get(kode)
                    nama_pola = _POLA_PER_JENIS.get(kode)
                    if not hasil_mentah or not nama_pola:
                        continue
                    try:
                        pola_client = ak.muat_pola(ak._path_pola(nama_pola, client_id))
                    except Exception as e:  # noqa: BLE001
                        print(f"[PERINGATAN] Gagal muat pola utk deteksi anomali ({kode}): {e}")
                        continue

                    df_mentah = hasil_mentah.get("df")
                    for item in ak.cari_anomali_untuk_alert(df_mentah, pola_client):
                        dbc.buat_alert_anomali(
                            client_id=client_id,
                            jenis=kode,
                            tipe_alert="nominal_ekstrim",
                            pesan=item["pesan"],
                            conv_id=conv_id_final,
                            baris_index=item["baris_index"],
                            konteks=item,
                            skor=item.get("anomaly_score"),
                        )

                    temuan_mencurigakan = ak.deteksi_pola_mencurigakan(pola_client)
                    if temuan_mencurigakan:
                        path_evaluasi = ak._path_pola(f"evaluasi_{nama_pola}", client_id)
                        temuan_benar_baru = ak.simpan_evaluasi_pola(path_evaluasi, temuan_mencurigakan)
                        for temuan in temuan_benar_baru:
                            pesan = (
                                f"Akun \"{temuan['nama_akun_debet']}\" (debet) / "
                                f"\"{temuan['nama_akun_kredit']}\" (kredit) dipakai oleh "
                                f"{temuan['jumlah_signature_berbeda']} pola transaksi berbeda "
                                f"yang masing2 baru muncul 1x -- cek apakah kategorisasi ini benar."
                            )
                            dbc.buat_alert_anomali(
                                client_id=client_id,
                                jenis=kode,
                                tipe_alert="pola_mencurigakan",
                                pesan=pesan,
                                conv_id=conv_id_final,
                                konteks=temuan,
                                skor=float(temuan["jumlah_signature_berbeda"]),
                            )
                q.put({"type": "progress", "step": "anomali", "label": "Mendeteksi anomali & pola mencurigakan", "status": "done"})

                # [BARU] AUTO-GENERATE LAPORAN 14-SHEET -- lihat docstring
                # _auto_generate_laporan_18_sheet(): begitu file dari chat
                # ini selesai diproses, laporan 18-sheet langsung disusun
                # kalau tahun pajaknya kebaca dari draf_jurnal & COA client
                # sudah ada -- TANPA menunggu koreksi/posting manual dulu.
                q.put({"type": "progress", "step": "laporan_18_sheet", "label": "Menyusun laporan 18-sheet", "status": "processing"})
                laporan_18_sheet = _auto_generate_laporan_18_sheet(
                    client_id, _tahun_dari_hasil_json(hasil_json), user,
                )
                q.put({"type": "progress", "step": "laporan_18_sheet", "label": "Menyusun laporan 18-sheet", "status": "done"})
            else:
                laporan_18_sheet = []

            q.put({
                "type": "result", "nama_file": nama_file, "hasil": hasil_json,
                "tidak_terdeteksi": False, "laporan_18_sheet": laporan_18_sheet,
            })
        except HTTPException as e:
            q.put({"type": "error", "pesan": str(e.detail)})
        except Exception as e:  # noqa: BLE001
            q.put({"type": "error", "pesan": str(e)})
        finally:
            q.put(None)  # sinyal: tidak ada event lagi

    threading.Thread(target=jalankan, daemon=True).start()

    def event_generator():
        while True:
            item = q.get()
            if item is None:
                break
            yield _format_sse_progress(**item)
        yield "data: [DONE]\n\n"

    return StreamingResponse(event_generator(), media_type="text/event-stream")


# ============================================================
# [BARU] MEKANISME TANYA BALIK -- endpoint klarifikasi
# ============================================================
# Dijawab oleh akuntan internal lewat dashboard/chat React (bukan klien
# langsung lewat WA -- sesuai keputusan). Pertanyaannya sendiri dibuat
# otomatis di dalam /api/proses-file di atas, lewat
# ak.cari_baris_perlu_klarifikasi().

@app.get("/api/klarifikasi")
def api_daftar_klarifikasi(
    client_id: Optional[int] = None,
    status: Optional[str] = "pending",
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """Daftar pertanyaan klarifikasi. Default cuma yang 'pending' --
    kirim status=None (atau status kosong) utk lihat semua termasuk yg
    sudah 'answered'."""
    return {
        "pertanyaan": dbc.daftar_pertanyaan_klarifikasi(client_id=client_id, status=status)
    }


class JawabKlarifikasiRequest(BaseModel):
    jawaban: str


@app.post("/api/klarifikasi/{pertanyaan_id}/jawab")
def api_jawab_klarifikasi(
    pertanyaan_id: int,
    req: JawabKlarifikasiRequest,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """Simpan jawaban akuntan atas 1 pertanyaan klarifikasi. Begitu
    dijawab, otomatis juga tercatat sbg feedback koreksi di tabel
    pola_augmentasi (lihat dbc.jawab_pertanyaan_klarifikasi) supaya
    transaksi serupa berikutnya tidak perlu ditanya lagi."""
    berhasil = dbc.jawab_pertanyaan_klarifikasi(
        pertanyaan_id=pertanyaan_id,
        jawaban=req.jawaban,
        username=user.get("username", "unknown"),
    )
    if not berhasil:
        raise HTTPException(
            status_code=404,
            detail="Pertanyaan tidak ditemukan atau gagal menyimpan jawaban.",
        )
    return {"berhasil": True}


# ============================================================
# [BARU] ALERT ANOMALI -- endpoint (mirip mekanisme klarifikasi di atas)
# ============================================================
# Ditinjau oleh akuntan internal lewat dashboard React. Alert-nya sendiri
# dibuat otomatis di dalam /api/proses-file di atas, lewat
# ak.cari_anomali_untuk_alert() & ak.deteksi_pola_mencurigakan().

@app.get("/api/client/{client_id}/reminder-spt")
def api_reminder_spt_client(
    client_id: int,
    hanya_belum_selesai: bool = True,
    user: dict = Depends(auth.get_current_user),
):
    """Kalender kewajiban lapor/setor SPT 1 client (SEMUA yang belum
    selesai, bukan cuma yang sudah/lagi diingatkan) -- beda dari
    /api/alert-anomali yang isinya notifikasi yang SUDAH terkirim."""
    return {"reminder": dbc.daftar_reminder_spt_client(client_id, hanya_belum_selesai=hanya_belum_selesai)}


@app.post("/api/notifikasi/jalankan-sekarang")
def api_jalankan_reminder_sekarang(user: dict = Depends(auth.require_level(3))):
    """Jalankan pengecekan reminder deadline SPT SEKARANG JUGA (tidak
    nunggu jadwal harian) -- utk testing manual bahwa WA/in-app benar2
    terkirim setelah setting FONNTE_TOKEN & nomor_wa client."""
    return notifikasi.jalankan_pengecekan_reminder_spt()


@app.get("/api/alert-anomali")
def api_daftar_alert_anomali(
    client_id: Optional[int] = None,
    status: Optional[str] = "baru",
    tipe_alert: Optional[str] = None,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """Daftar alert anomali. Default cuma yang 'baru' -- kirim status=None
    (atau status kosong) utk lihat semua termasuk yg sudah 'dilihat'/
    'diabaikan'. tipe_alert opsional: 'nominal_ekstrim' atau
    'pola_mencurigakan'."""
    return {
        "alert": dbc.daftar_alert_anomali(
            client_id=client_id, status=status, tipe_alert=tipe_alert,
        )
    }


class TandaiAlertAnomaliRequest(BaseModel):
    status: str  # "dilihat" atau "diabaikan"


@app.post("/api/alert-anomali/{alert_id}/tandai")
def api_tandai_alert_anomali(
    alert_id: int,
    req: TandaiAlertAnomaliRequest,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """Tandai 1 alert sbg 'dilihat' atau 'diabaikan' oleh akuntan yang login."""
    berhasil = dbc.tandai_alert_anomali(
        alert_id=alert_id,
        status=req.status,
        username=user.get("username", "unknown"),
    )
    if not berhasil:
        raise HTTPException(
            status_code=404,
            detail="Alert tidak ditemukan atau status tidak valid ('dilihat'/'diabaikan').",
        )
    return {"berhasil": True}


@app.post("/api/proses-dan-buat-excel")
async def proses_dan_buat_excel(
    file: UploadFile = File(...),
    # [FIX] Sama seperti /api/proses-file di atas -- jenis_dokumen harus
    # Form(None) supaya kebaca dari body form-data yang sama dengan file,
    # bukan dianggap query parameter oleh FastAPI.
    jenis_dokumen: Optional[str] = Form(None),
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    isi = await file.read()
    nama_file = file.filename or "upload.xlsx"

    # [FIX -- GAP EVENT LOOP] _proses_semua_jenis() sync & berat (parsing
    # penuh) -- dibungkus to_thread sama alasannya dgn endpoint upload lain.
    hasil_semua, error_per_jenis = await asyncio.to_thread(
        _proses_semua_jenis, isi, nama_file, jenis_dokumen
    )

    if not hasil_semua:
        return {
            "nama_file": nama_file,
            "berhasil": False,
            "pesan": "Tidak ada jenis dokumen yang dikenali di file ini.",
            "detail_error": error_per_jenis or None,
        }

    # [FIX -- GAP EVENT LOOP + POINT 2] _buat_excel_hasil() (lama, di bawah)
    # murni CPU-bound (openpyxl susun banyak sheet, iter ribuan baris) --
    # THREAD saja TIDAK bikin ini benar-benar paralel (GIL-bound, lihat
    # benchmark di ai_file_reader.py: paralel-thread vs sequential nyaris
    # sama). Dipindah ke modules/excel_export_worker.py (modul LEAF, tanpa
    # import FastAPI/DB) supaya AMAN dijalankan lewat ProcessPoolExecutor --
    # lihat catatan lengkap di modul itu soal KENAPA tidak cukup didekor
    # to_thread di tempat, harus dipindah ke modul terpisah.
    # label_per_kode dikirim eksplisit (bukan dibaca ulang dari
    # _PEMROSES_DOKUMEN di dalam modul worker) supaya excel_export_worker.py
    # tidak perlu mengimpor apa pun dari main.py -- menghindari duplikasi
    # sekaligus risiko drift kalau _PEMROSES_DOKUMEN berubah.
    label_per_kode = {kode: lbl for kode, (lbl, _fn) in _PEMROSES_DOKUMEN.items()}
    path_file_hasil = await asyncio.to_thread(
        excel_export_worker.jalankan_buat_excel_hasil_di_proses,
        hasil_semua, str(FOLDER_HASIL), label_per_kode,
    )

    return {
        "nama_file": nama_file,
        "berhasil": True,
        "ringkasan_per_jenis": _bersihkan_untuk_json(
            {kode: h.get("ringkasan") for kode, h in hasil_semua.items()}
        ),
        "jumlah_perlu_review_per_jenis": {
            kode: len(h.get("masalah") or []) for kode, h in hasil_semua.items()
        },
        "nama_file_hasil": path_file_hasil.name,
        "path_unduh": f"/api/unduh/{path_file_hasil.name}",
    }


# ============================================================
# [BARU] EXPORT REKENING KORAN -- FORMAT KERJA AKUNTAN
# ============================================================
# Beda dari /api/proses-dan-buat-excel di atas (yang menulis Ringkasan/
# Perlu Review/Draf Jurnal sebagai teks datar lewat _buat_excel_hasil):
# endpoint ini menghasilkan file Excel format kerja akuntan yang SAMA
# persis strukturnya dengan file kerja rekonsiliasi bank yang biasa
# dipakai tim akuntan -- 1 sheet COA + 1 sheet per bank, nomor voucher
# berurutan, formula VLOOKUP ke COA, dan formula cek saldo berjalan.
# Lihat modules/accounting_export.py -> export_rekening_koran_format_akuntan().

@app.get("/api/client/{client_id}/rekening-koran/export-format-akuntan/{hasil_id}")
async def api_export_rekening_koran_format_akuntan(
    client_id: int,
    hasil_id: int,
    file_piutang: Optional[UploadFile] = File(None),
    user: dict = Depends(auth.get_current_user),
):
    """
    [BERUBAH - Prioritas #7] Sebelumnya endpoint ini menerima UPLOAD FILE
    dan memanggil ak.proses_file_rekening_koran() dari nol -- artinya
    file yang sama bisa di-parse & dipanggil ke AI DUA KALI kalau
    akuntan juga sudah upload lewat /api/proses-file (yang otomatis
    menarik hasilnya ke antrean jurnal_posting). Sekarang endpoint ini
    TIDAK menerima file lagi -- ia membaca ULANG hasil yang SUDAH
    tersimpan dari upload sebelumnya (hasil_id, didapat dari response
    /api/proses-file atau /api/client/{client_id}/hasil), digabung
    dengan voucher & status posting terkini dari tabel jurnal_posting.

    Alur yang benar sekarang:
      1. Upload rekening koran SEKALI lewat /api/proses-file (Supervisor+).
         -> tersimpan ke tabel hasil, DAN baris2nya otomatis masuk
         antrean jurnal_posting (status draft, voucher SUDAH digenerate
         saat itu juga -- lihat dbc.tarik_draf_jurnal_ke_posting()).
      2. Panggil endpoint INI kapan saja dengan hasil_id dari langkah 1
         untuk mengunduh Excel format kerja akuntan. Voucher di Excel
         akan SAMA PERSIS setiap kali di-export ulang (idempoten), dan
         kolom "Status Posting" menunjukkan draft/terposting terkini --
         jadi Excel yang sama bisa dipakai baik SEBELUM maupun SESUDAH
         supervisor menekan "Posting" di layar review, tanpa perlu
         parse ulang file atau bikin endpoint terpisah untuk masing2.
      3. (Opsional, sebelum atau sesudah unduh Excel) Supervisor
         mengonfirmasi baris2 di jurnal_posting -- satu-satu lewat
         /api/client/{client_id}/jurnal-posting/{posting_id}/konfirmasi,
         atau sekaligus lewat endpoint konfirmasi-semua di bawah utk
         baris yang akunnya bukan placeholder.

    file_piutang tetap opsional: kalau dikasih, dipakai untuk auto-isi
    kolom Supplier/Cust pada baris uang masuk yang belum ketahuan
    pasangannya (lihat _cocokkan_supplier_opsional di accounting_export.py).
    """
    hasil_row = dbc.ambil_hasil_by_id(hasil_id)
    if hasil_row is None or hasil_row["client_id"] != client_id or hasil_row["jenis"] != "rekening_koran":
        raise HTTPException(
            status_code=404,
            detail="Hasil rekening koran dengan id ini tidak ditemukan untuk client tersebut. "
                   "Upload dulu lewat /api/proses-file, lalu pakai hasil_id dari response-nya.",
        )

    data = hasil_row["data"] or {}
    df_hasil = pd.DataFrame(data.get("df") or [])
    if df_hasil.empty:
        raise HTTPException(
            status_code=400,
            detail="Hasil tersimpan ini tidak punya baris rekening koran (df kosong).",
        )
    df_hasil = df_hasil.reset_index(drop=True)

    # --- Gabungkan voucher & status posting terkini dari jurnal_posting ---
    # Dicocokkan via posisi baris asli (baris_asal = index+1, lihat catatan
    # di db_client.JurnalPosting.baris_asal) -- BUKAN via isi baris, karena
    # rekening koran sering punya transaksi identik (nominal & keterangan
    # sama persis) yang bikin pencocokan berbasis konten ambigu.
    posting_rows = dbc.ambil_jurnal_posting_by_hasil(client_id, hasil_id)
    peta_posting = {p["baris_asal"]: p for p in posting_rows if p.get("baris_asal")}

    label_status = {
        "draft": "Draft (belum diposting)",
        "terposting": "Terposting",
        "ditolak": "Ditolak",
    }
    daftar_voucher, daftar_status = [], []
    for i in range(len(df_hasil)):
        p = peta_posting.get(i + 1)  # baris_asal 1-based, sama seperti "baris": i+1 di draf_jurnal
        if p:
            daftar_voucher.append(p.get("voucher"))
            daftar_status.append(label_status.get(p.get("status"), p.get("status")))
        else:
            daftar_voucher.append(None)
            daftar_status.append("Belum Terkategori (belum masuk antrean posting)")
    df_hasil["voucher"] = daftar_voucher
    df_hasil["status_posting"] = daftar_status

    df_coa = pd.DataFrame(data.get("coa") or [])

    df_piutang = None
    if file_piutang is not None:
        isi_piutang = await file_piutang.read()
        buf_piutang = io.BytesIO(isi_piutang)
        buf_piutang.name = file_piutang.filename or "piutang.xlsx"
        hasil_piutang = ak.proses_file_piutang(buf_piutang, buf_piutang.name)
        df_piutang = hasil_piutang.get("df")

    # [FIX -- GAP EVENT LOOP + POINT 2] Sebelumnya endpoint ini memanggil
    # accounting_export.export_rekening_koran_format_akuntan() LANGSUNG
    # secara sync -- tanpa to_thread maupun ProcessPoolExecutor sama
    # sekali -- padahal fungsi ini openpyxl susun sheet per bank + formula
    # VLOOKUP/saldo berjalan, murni CPU-bound persis seperti kasus di
    # /api/proses-dan-buat-excel (Point 2). THREAD saja tidak cukup
    # (GIL-bound), jadi dipakai worker yang sama di
    # modules/excel_export_worker.py lewat ProcessPoolExecutor, dibungkus
    # asyncio.to_thread di sini supaya event loop juga tidak ikut terblokir
    # selama menunggu hasil dari proses worker.
    try:
        isi_excel = await asyncio.to_thread(
            excel_export_worker.jalankan_export_rekening_koran_format_akuntan_di_proses,
            df_hasil, df_coa, df_piutang,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e

    nama_file_asal = data.get("nama_file") or "rekening_koran.xlsx"
    nama_unduh = f"Rekening_Koran_{nama_file_asal.rsplit('.', 1)[0]}.xlsx"
    return StreamingResponse(
        io.BytesIO(isi_excel),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nama_unduh}"'},
    )


@app.post("/api/client/{client_id}/jurnal-posting/hasil/{hasil_id}/konfirmasi-semua")
async def api_konfirmasi_posting_massal(
    client_id: int,
    hasil_id: int,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas, sama seperti /api/proses-file
):
    """
    [BARU - Prioritas #7] Konfirmasi SEKALIGUS semua baris jurnal_posting
    berstatus 'draft' milik satu hasil_id (satu file upload) -- supaya
    supervisor tidak perlu klik "Posting" satu-satu untuk rekening koran
    yang bisa berisi ratusan/ribuan baris. Baris dengan akun placeholder
    (butuh keputusan manual akun lawannya) TETAP dilewati & harus
    dikonfirmasi satu-satu lewat endpoint /konfirmasi yang sudah ada,
    setelah akuntan mengisi akun yang benar -- lihat
    dbc.konfirmasi_posting_massal() untuk detail.
    """
    hasil_row = dbc.ambil_hasil_by_id(hasil_id)
    if hasil_row is None or hasil_row["client_id"] != client_id:
        raise HTTPException(status_code=404, detail="Hasil dengan id ini tidak ditemukan untuk client tersebut.")

    ringkasan = dbc.konfirmasi_posting_massal(client_id, hasil_id, user.get("username", "unknown"))
    return {
        "pesan": f"{ringkasan['diposting']} baris berhasil diposting, "
                 f"{ringkasan['dilewati_placeholder']} baris dilewati (akun masih placeholder, perlu isi manual).",
        **ringkasan,
    }


# ============================================================
# [BARU] CHART OF ACCOUNTS (COA) PERMANEN PER CLIENT
# ============================================================

class AkunCoaRequest(BaseModel):
    no_akun: str
    nama_akun: str
    kategori: Optional[str] = None  # ASET/LIABILITAS/EKUITAS/PENDAPATAN/BEBAN
    sub_kategori: Optional[str] = None
    normal_saldo: Optional[str] = None  # DEBET/KREDIT
    saldo_awal: float = 0
    segment: Optional[str] = None       # [BARU] mis. "Excavator"/"Scaffolding"/"Umum"/"Semua"
    arus_kas: Optional[str] = None      # [BARU] "Operasi"/"Investasi"/"Pendanaan"/"Nonkas"/dst
    keterangan: Optional[str] = None    # [BARU] catatan bebas per akun
    # [BARU] Dipakai khusus sheet "Neraca Saldo Awal" -- kolom "Lawan
    # Transaksi" & "Project/Asset Unit" per akun, supaya baris saldo awal
    # tidak hardcode "Pemilik"/"HO" utk semua akun (mis. akun excavator
    # bisa punya project_unit "EXC-01 & EXC-02", akun modal Tuan A bisa
    # punya lawan_transaksi "Tuan A"). Opsional -- kalau kosong, sheet
    # export tetap jalan dengan fallback "-".
    lawan_transaksi_saldo_awal: Optional[str] = None
    project_unit_saldo_awal: Optional[str] = None


class CoaBulkRequest(BaseModel):
    akun: List[AkunCoaRequest]
    ganti_semua: bool = True


@app.get("/api/client/{client_id}/coa")
def api_ambil_coa(client_id: int, user: dict = Depends(auth.get_current_user)):
    """Ambil seluruh Chart of Accounts (COA) permanen milik satu client."""
    return {"coa": dbc.ambil_coa_client(client_id)}


@app.post("/api/client/{client_id}/coa")
def api_simpan_coa_bulk(client_id: int, req: CoaBulkRequest, user: dict = Depends(auth.get_current_user)):
    """
    Simpan COA client sekaligus (dari form input manual atau hasil impor
    sheet 'COA' file Excel yang sudah diparse frontend). Default
    ganti_semua=True (replace total) -- kirim ganti_semua=false kalau
    cuma mau menambah/memperbarui sebagian akun.
    """
    jumlah = dbc.simpan_coa_bulk(
        client_id, [a.model_dump() for a in req.akun], ganti_semua=req.ganti_semua
    )
    dbc.log_audit(
        client_id=client_id, user=user.get("username", "unknown"),
        aksi="simpan_coa_bulk",
        detail={"ganti_semua": req.ganti_semua, "jumlah_akun": jumlah},
    )
    return {"berhasil": True, "jumlah_akun_tersimpan": jumlah}


@app.post("/api/client/{client_id}/coa/akun")
def api_tambah_akun_coa(client_id: int, req: AkunCoaRequest, user: dict = Depends(auth.get_current_user)):
    """Tambah satu akun COA baru untuk client."""
    berhasil = dbc.tambah_akun_coa(
        client_id, req.no_akun, req.nama_akun, req.kategori,
        req.sub_kategori, req.normal_saldo, req.saldo_awal,
        segment=req.segment, arus_kas=req.arus_kas, keterangan=req.keterangan,
        lawan_transaksi_saldo_awal=req.lawan_transaksi_saldo_awal,  # [BARU]
        project_unit_saldo_awal=req.project_unit_saldo_awal,  # [BARU]
    )
    if not berhasil:
        raise HTTPException(status_code=500, detail="Gagal menambah akun COA.")
    dbc.log_audit(
        client_id=client_id, user=user.get("username", "unknown"),
        aksi="tambah_akun_coa",
        detail={"no_akun": req.no_akun, "nama_akun": req.nama_akun, "kategori": req.kategori},
    )
    return {"berhasil": True}


@app.put("/api/client/{client_id}/coa/akun/{akun_id}")
def api_update_akun_coa(client_id: int, akun_id: int, req: AkunCoaRequest, user: dict = Depends(auth.get_current_user)):
    """Perbarui satu akun COA (mis. mengisi kategori yang tadinya kosong)."""
    sebelum = dbc.ambil_akun_coa_by_id(akun_id)
    berhasil = dbc.update_akun_coa(
        akun_id, no_akun=req.no_akun, nama_akun=req.nama_akun, kategori=req.kategori,
        sub_kategori=req.sub_kategori, normal_saldo=req.normal_saldo, saldo_awal=req.saldo_awal,
        segment=req.segment, arus_kas=req.arus_kas, keterangan=req.keterangan,
        lawan_transaksi_saldo_awal=req.lawan_transaksi_saldo_awal,  # [BARU]
        project_unit_saldo_awal=req.project_unit_saldo_awal,  # [BARU]
    )
    if not berhasil:
        raise HTTPException(status_code=404, detail="Akun COA tidak ditemukan.")
    dbc.log_audit(
        client_id=client_id, user=user.get("username", "unknown"),
        aksi="update_akun_coa",
        detail={
            "akun_id": akun_id,
            "sebelum": sebelum,
            "sesudah": req.model_dump(exclude_none=True),
        },
    )
    return {"berhasil": True}


@app.delete("/api/client/{client_id}/coa/akun/{akun_id}")
def api_hapus_akun_coa(client_id: int, akun_id: int, user: dict = Depends(auth.get_current_user)):
    """Nonaktifkan (soft-delete) satu akun COA."""
    sebelum = dbc.ambil_akun_coa_by_id(akun_id)
    berhasil = dbc.hapus_akun_coa(akun_id)
    if not berhasil:
        raise HTTPException(status_code=404, detail="Akun COA tidak ditemukan.")
    dbc.log_audit(
        client_id=client_id, user=user.get("username", "unknown"),
        aksi="hapus_akun_coa",
        detail={"akun_id": akun_id, "sebelum": sebelum},
    )
    return {"berhasil": True}


# ============================================================
# [BARU] REVIEW & POSTING JURNAL (draf placeholder -> siap laporan)
# ============================================================

@app.get("/api/client/{client_id}/jurnal-posting")
def api_daftar_jurnal_posting(
    client_id: int,
    status: Optional[str] = "draft",
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """
    Daftar baris jurnal yang perlu direview/diposting (default status=draft).
    Pakai ?status=terposting atau ?status=ditolak untuk lihat histori,
    atau ?status= (kosong) untuk semua status.
    """
    status_final = status if status else None
    return {"jurnal": dbc.daftar_jurnal_posting(client_id, status=status_final)}


class KonfirmasiPostingRequest(BaseModel):
    no_akun_debet: Optional[str] = None
    nama_akun_debet: Optional[str] = None
    no_akun_kredit: Optional[str] = None
    nama_akun_kredit: Optional[str] = None
    tanggal: Optional[str] = None
    keterangan: Optional[str] = None
    # [BARU - fix GL 2025] field tambahan supaya akuntan bisa mengisi
    # kolom Lawan Transaksi/Project/Unit/Invoice-Referensi/Jatuh Tempo
    # sheet "GL 2025" saat konfirmasi posting -- sebelumnya tidak ada
    # jalur sama sekali untuk mengisi field-field ini.
    lawan_transaksi: Optional[str] = None
    no_dokumen: Optional[str] = None
    project_unit: Optional[str] = None
    jatuh_tempo: Optional[str] = None


@app.post("/api/client/{client_id}/jurnal-posting/{posting_id}/konfirmasi")
def api_konfirmasi_posting(
    client_id: int, posting_id: int, req: KonfirmasiPostingRequest,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """
    Konfirmasi satu baris jurnal jadi 'terposting'. Kalau akun asalnya
    masih placeholder (mis. "KAS", "PIUTANG/KAS"), WAJIB isi no_akun_debet
    & no_akun_kredit dengan nomor akun COA yang sebenarnya lewat body
    request ini -- endpoint ini menolak dengan 400 kalau akun akhir masih
    mengandung "/" (tanda placeholder yang belum diisi).
    """
    daftar = dbc.daftar_jurnal_posting(client_id, status=None)
    baris = next((j for j in daftar if j["id"] == posting_id), None)
    if baris is None:
        raise HTTPException(status_code=404, detail="Baris jurnal tidak ditemukan.")

    no_debet_final = req.no_akun_debet or baris["no_akun_debet"]
    no_kredit_final = req.no_akun_kredit or baris["no_akun_kredit"]
    if "/" in no_debet_final or "/" in no_kredit_final:
        raise HTTPException(
            status_code=400,
            detail=(
                "Akun debet/kredit masih placeholder (mis. 'KAS/PIUTANG'). "
                "Isi no_akun_debet & no_akun_kredit dengan nomor akun COA yang "
                "sebenarnya sebelum diposting."
            ),
        )

    berhasil = dbc.konfirmasi_posting_jurnal(
        posting_id, user.get("username", "unknown"),
        no_akun_debet=req.no_akun_debet, nama_akun_debet=req.nama_akun_debet,
        no_akun_kredit=req.no_akun_kredit, nama_akun_kredit=req.nama_akun_kredit,
        tanggal=req.tanggal, keterangan=req.keterangan,
        lawan_transaksi=req.lawan_transaksi, no_dokumen=req.no_dokumen,
        project_unit=req.project_unit, jatuh_tempo=req.jatuh_tempo,
    )
    if not berhasil:
        raise HTTPException(status_code=500, detail="Gagal memposting jurnal.")

    dbc.log_audit(
        client_id=client_id, user=user.get("username", "unknown"),
        aksi="posting_jurnal", detail={"posting_id": posting_id, **req.model_dump(exclude_none=True)},
    )
    return {"berhasil": True}


class TolakPostingRequest(BaseModel):
    alasan: Optional[str] = None


@app.post("/api/client/{client_id}/jurnal-posting/{posting_id}/tolak")
def api_tolak_posting(
    client_id: int, posting_id: int, req: TolakPostingRequest,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """Tolak satu baris jurnal (mis. duplikat/salah deteksi) -- tidak akan masuk laporan keuangan."""
    berhasil = dbc.tolak_posting_jurnal(posting_id, user.get("username", "unknown"), req.alasan)
    if not berhasil:
        raise HTTPException(status_code=404, detail="Baris jurnal tidak ditemukan.")
    dbc.log_audit(
        client_id=client_id, user=user.get("username", "unknown"),
        aksi="tolak_jurnal", detail={"posting_id": posting_id, "alasan": req.alasan},
    )
    return {"berhasil": True}


# ============================================================
# [BARU] 5 LAPORAN KEUANGAN STANDAR
# ============================================================

class GenerateLaporanKeuanganRequest(BaseModel):
    periode: str  # mis. "2026-07"
    tanggal_mulai: Optional[str] = None
    tanggal_akhir: Optional[str] = None
    prive_atau_dividen: float = 0
    setoran_modal_baru: float = 0
    penyesuaian_ekuitas_manual: float = 0


@app.post("/api/client/{client_id}/laporan-keuangan/generate")
def api_generate_laporan_keuangan(
    client_id: int, req: GenerateLaporanKeuanganRequest,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """
    Generate 5 Laporan Keuangan Standar (Neraca, Laba Rugi, Perubahan
    Ekuitas, Arus Kas, CALK) dari jurnal + COA client, lalu simpan sebagai
    snapshot baru (histori tidak ditimpa -- generate ulang untuk periode
    yang sama akan membuat snapshot baru).

    [BARU] hanya_terposting=False -- endpoint ini TIDAK LAGI mewajibkan
    akuntan mengonfirmasi-posting jurnal draft satu-satu dulu sebelum
    laporan bisa dibuat (pola yang sama dgn export-18-sheet). Baris draft
    (termasuk yang akunnya masih placeholder) ikut apa adanya; status per
    akun ditandai lewat field "keterangan_perlu_dikoreksi" yang sekarang
    disertakan tiap baris Neraca/Laba Rugi/Perubahan Ekuitas (lihat
    hitung_saldo_per_akun() di laporan_keuangan.py) -- bukan lewat gate
    status database seperti sebelumnya.
    """
    coa = dbc.ambil_coa_client(client_id)
    jurnal = dbc.ambil_jurnal_terposting(client_id, req.tanggal_mulai, req.tanggal_akhir, hanya_terposting=False)

    hasil = lapkeu.generate_5_laporan_keuangan(
        jurnal, coa, req.periode,
        prive_atau_dividen=req.prive_atau_dividen,
        setoran_modal_baru=req.setoran_modal_baru,
        penyesuaian_ekuitas_manual=req.penyesuaian_ekuitas_manual,
    )

    lap_id = dbc.simpan_laporan_keuangan(
        client_id, req.periode, hasil, dibuat_oleh=user.get("username", "unknown"),
        tanggal_mulai=req.tanggal_mulai, tanggal_akhir=req.tanggal_akhir,
    )

    dbc.log_audit(
        client_id=client_id, user=user.get("username", "unknown"),
        aksi="generate_laporan_keuangan",
        detail={"periode": req.periode, "laporan_id": lap_id, "balance": hasil["neraca"]["balance"]},
    )

    return {"laporan_id": lap_id, "laporan": hasil}


@app.get("/api/client/{client_id}/laporan-keuangan")
def api_ambil_laporan_keuangan(
    client_id: int, periode: Optional[str] = None,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """
    Ambil snapshot laporan keuangan. Kalau ?periode=2026-07 diisi, return
    snapshot TERBARU untuk periode itu. Kalau tidak diisi, return daftar
    riwayat semua snapshot (tanpa isi lengkap -- panggil lagi dengan
    ?periode=... untuk ambil isinya).
    """
    if periode:
        lap = dbc.ambil_laporan_keuangan_terbaru(client_id, periode)
        if lap is None:
            raise HTTPException(status_code=404, detail=f"Belum ada laporan keuangan untuk periode {periode}.")
        return lap
    return {"riwayat": dbc.daftar_riwayat_laporan_keuangan(client_id)}


# ============================================================
# [BARU] LAMPIRAN SPT TAHUNAN BADAN (A01-A09 / L01-L05 / E01-E04)
# ============================================================

@app.get("/api/client/{client_id}/laporan-keuangan/{periode}/lampiran-spt")
def api_lampiran_spt(
    client_id: int,
    periode: str,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """
    Susun Lampiran SPT Tahunan Badan (A01-A09 Neraca, L01-L05 Laba Rugi,
    E01-E04 Perubahan Ekuitas) dari snapshot laporan keuangan yang SUDAH
    digenerate untuk periode ini (lihat POST .../laporan-keuangan/generate).
    Tidak menyimpan snapshot baru -- murni turunan dari data yang sudah ada.
    """
    lap = dbc.ambil_laporan_keuangan_terbaru(client_id, periode)
    if not lap:
        raise HTTPException(404, f"Belum ada laporan keuangan untuk periode {periode}")

    return lapkeu.susun_lampiran_spt_lengkap(lap["data"])


# ============================================================
# [BARU] PPh BADAN PASAL 31E - ENDPOINT
# ============================================================

class GeneratePPhBadanRequest(BaseModel):
    # [FIX] PPh Badan adalah pajak TAHUNAN -- field "periode" (string bebas,
    # dulu dicontohkan "2026-07" seperti periode BULANAN laporan keuangan)
    # DIHAPUS supaya tidak ada jalan bagi caller utk secara tidak sengaja
    # menghitung PPh Badan dari data 1 bulan. "tahun_pajak" sekarang WAJIB
    # dan jadi SATU-SATUNYA sumber kebenaran periode -- dipakai untuk
    # membangun rentang tanggal 1 Jan - 31 Des tahun tsb sendiri di bawah.
    tahun_pajak: int  # 2026
    nama_perusahaan: Optional[str] = None
    kompensasi_kerugian_fiskal: float = 0
    kredit_pajak: Optional[Dict[str, float]] = None  # pph_22, pph_23, pph_24, angsuran_pph_25
    # [BARU] Skema pajak client -- HANYA "Tarif Umum Pasal 17/31E" (default,
    # nilai persis harus sama dgn pph_badan.SKEMA_TARIF_UMUM_31E) yang
    # berhak atas fasilitas 31E. Ubah kalau client sebenarnya pakai PPh
    # Final UMKM -- fasilitas 31E akan otomatis di-nolkan dgn peringatan.
    skema_pajak: str = "Tarif Umum Pasal 17/31E"
    # [BARU] 2 komponen tambahan Peredaran Bruto, di luar "Peredaran Bruto
    # Usaha dari PNL" (yang otomatis diambil dari total_pendapatan laporan
    # keuangan tahunan) -- lihat pph_badan.hitung_total_peredaran_bruto().
    tambahan_peredaran_bruto_lainnya: float = 0
    retur_pengurangan_peredaran_bruto: float = 0
    # [BARU] Label bebas ttg sumber peredaran bruto usaha, beda per
    # perusahaan (mis. "Pendapatan sewa dan mobilisasi", "Pendapatan jasa
    # konstruksi") -- ditulis di kolom "Status/Keterangan" sheet export.
    keterangan_peredaran_bruto: Optional[str] = None


@app.post("/api/client/{client_id}/pph-badan/generate")
def api_generate_pph_badan(
    client_id: int,
    req: GeneratePPhBadanRequest,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """
    Generate PPh Badan 31E untuk SATU TAHUN PAJAK penuh (1 Jan - 31 Des):
    1. Ambil laba_rugi_bersih & total_pendapatan SETAHUN PENUH -- pakai
       snapshot laporan_keuangan yang periode-nya = str(tahun_pajak) kalau
       sudah pernah digenerate; kalau belum ada, generate OTOMATIS di sini
       dari jurnal terposting tanggal_mulai={tahun}-01-01 s/d
       {tahun}-12-31, supaya endpoint ini TIDAK PERNAH diam-diam memakai
       laporan bulanan yang salah lingkup (ini yang jadi bug sebelumnya:
       req.periode bebas bisa saja "2026-07" / 1 bulan, padahal PPh Badan
       butuh angka setahun).
    2. Ambil rekonsiliasi_fiskal dari hasil proses_aset_tetap() yang
       tersimpan -- [FIX] sekarang dicocokkan ke tahun_pajak yang diminta
       (via tanggal upload), bukan asal ambil upload Aset Tetap TERBARU
       apa pun tahunnya seperti sebelumnya. Kalau tidak ada yang cocok,
       tetap fallback ke upload terbaru TAPI hasil diberi flag peringatan
       eksplisit supaya akuntan sadar & mengecek manual.
    3. Hitung dengan pph_badan.hitung_pph_pasal_31e()
    """
    from modules import pph_badan
    from modules import fiscal_reconciliation

    tahun = req.tahun_pajak
    periode_tahunan = str(tahun)
    tanggal_mulai = f"{tahun}-01-01"
    tanggal_akhir = f"{tahun}-12-31"

    # -- 1. Laporan keuangan SETAHUN PENUH --
    lap = dbc.ambil_laporan_keuangan_terbaru(client_id, periode_tahunan)
    if not lap:
        # Belum pernah ada snapshot laporan tahunan untuk tahun ini --
        # generate otomatis dari jurnal terposting sepanjang tahun tsb
        # (pola yang sama dgn endpoint laporan-bulanan/generate di bawah),
        # supaya user tidak perlu tahu harus generate laporan-keuangan
        # dgn periode tahunan secara manual dulu sebelum bisa hitung PPh.
        coa = dbc.ambil_coa_client(client_id)
        # [BARU] hanya_terposting=False -- lihat catatan yang sama di
        # laporan-keuangan/generate di atas.
        jurnal = dbc.ambil_jurnal_terposting(client_id, tanggal_mulai, tanggal_akhir, hanya_terposting=False)
        if not jurnal:
            raise HTTPException(
                404,
                f"Belum ada laporan keuangan maupun jurnal untuk tahun pajak {tahun}. "
                f"Pastikan ada data yang sudah diproses (draft maupun terposting) antara "
                f"{tanggal_mulai} dan {tanggal_akhir} sebelum generate PPh Badan.",
            )
        data_laporan = lapkeu.generate_5_laporan_keuangan(jurnal, coa, periode_tahunan)
        lap_id = dbc.simpan_laporan_keuangan(
            client_id, periode_tahunan, data_laporan,
            dibuat_oleh=user.get("username", "unknown"),
            tanggal_mulai=tanggal_mulai, tanggal_akhir=tanggal_akhir,
        )
        lap = {"id": lap_id, "periode": periode_tahunan, "data": data_laporan}

    laba_bersih = lap["data"]["laba_rugi"]["laba_rugi_bersih"]
    total_pendapatan = lap["data"]["laba_rugi"]["total_pendapatan"]

    # -- 2. Rekonsiliasi fiskal dari Aset Tetap, dicocokkan ke tahun_pajak --
    hasil_aset_semua = dbc.ambil_hasil_client(client_id, jenis="aset_tetap", limit=100)
    hasil_aset_tahun_ini = [
        h for h in hasil_aset_semua
        if h.get("dibuat_at") and str(h["dibuat_at"]).startswith(periode_tahunan)
    ]

    rekon_fiskal: Dict[str, Any] = {}
    peringatan_aset: Optional[str] = None
    if hasil_aset_tahun_ini:
        data = hasil_aset_tahun_ini[0]["data"]
        rekon_fiskal = fiscal_reconciliation.ringkas_rekonsiliasi_fiskal_dari_aset_tetap(data)
    elif hasil_aset_semua:
        # [FIX] Fallback ke upload Aset Tetap TERBARU apa pun tahunnya
        # (perilaku lama) -- tapi sekarang diberi peringatan eksplisit,
        # bukan diam-diam dipakai seolah datanya memang utk tahun ini.
        data = hasil_aset_semua[0]["data"]
        rekon_fiskal = fiscal_reconciliation.ringkas_rekonsiliasi_fiskal_dari_aset_tetap(data)
        peringatan_aset = (
            f"Tidak ditemukan upload Aset Tetap yang bertanggal tahun {tahun} -- "
            f"koreksi fiskal di bawah memakai upload Aset Tetap TERBARU yang tersedia "
            f"(diupload {hasil_aset_semua[0].get('dibuat_at')}). Mohon verifikasi manual "
            f"apakah data ini memang mewakili tahun pajak {tahun}."
        )
    # Kalau tidak ada data Aset Tetap sama sekali, rekon_fiskal tetap {}
    # (perilaku lama) -- koreksi fiskal dianggap 0, PKP = laba komersial saja.

    hasil = pph_badan.hitung_pph_pasal_31e(
        peredaran_bruto=total_pendapatan,
        laba_bersih_komersial=laba_bersih,
        tambahan_peredaran_bruto_lainnya=req.tambahan_peredaran_bruto_lainnya or 0,
        retur_pengurangan_peredaran_bruto=req.retur_pengurangan_peredaran_bruto or 0,
        koreksi_fiskal_positif=rekon_fiskal.get("koreksi_fiskal_positif", 0),
        koreksi_fiskal_negatif=rekon_fiskal.get("koreksi_fiskal_negatif", 0),
        kompensasi_kerugian_fiskal=req.kompensasi_kerugian_fiskal or 0,
        kredit_pajak=req.kredit_pajak,
        tahun_pajak=tahun,
        nama_perusahaan=req.nama_perusahaan,
        skema_pajak=req.skema_pajak or "Tarif Umum Pasal 17/31E",
        keterangan_peredaran_bruto=req.keterangan_peredaran_bruto,
    )
    if peringatan_aset:
        hasil["peringatan_data_aset_tetap"] = peringatan_aset

    analisis_id = dbc.simpan_hasil_analisis(
        client_id=client_id,
        jenis_analisis="pph_badan_31e",
        hasil=hasil,
        prompt=f"Tahun pajak {tahun}, peredaran bruto {total_pendapatan:,.0f}",
        model_ai="rule_based",
    )

    dbc.log_audit(
        client_id=client_id,
        user=user.get("username", "unknown"),
        aksi="generate_pph_badan",
        detail={"tahun_pajak": tahun, "pph_terutang": hasil["pph_badan_terutang"]},
    )

    return {"laporan_id": analisis_id, "hasil": hasil}


@app.get("/api/client/{client_id}/pph-badan/riwayat")
def api_riwayat_pph_badan(
    client_id: int,
    tahun_pajak: Optional[int] = None,
    user: dict = Depends(auth.require_level(3)),
):
    """
    Ambil riwayat perhitungan PPh Badan yang sudah pernah dibuat.
    [BARU] filter opsional ?tahun_pajak=2026 -- karena satu client bisa
    punya riwayat PPh Badan dari beberapa tahun pajak sekaligus di bawah
    jenis_analisis yang sama ("pph_badan_31e"), tanpa filter ini caller
    harus menyaring sendiri dari field "hasil.tahun_pajak" di tiap baris.
    """
    riwayat = dbc.ambil_hasil_analisis_client(client_id, jenis_analisis="pph_badan_31e")
    if tahun_pajak is not None:
        riwayat = [r for r in riwayat if (r.get("hasil") or {}).get("tahun_pajak") == tahun_pajak]
    return {"riwayat": riwayat}


@app.get("/api/client/{client_id}/pph-badan/export/{analisis_id}")
def api_export_pph_badan_excel(
    client_id: int,
    analisis_id: int,
    user: dict = Depends(auth.require_level(3)),
):
    """Export hasil perhitungan PPh Badan 31E (yang sudah pernah digenerate) ke Excel."""
    riwayat = dbc.ambil_hasil_analisis_client(client_id, jenis_analisis="pph_badan_31e", limit=100)
    analisis = next((r for r in riwayat if r["id"] == analisis_id), None)
    if not analisis:
        raise HTTPException(404, "Analisis PPh Badan tidak ditemukan")

    hasil = analisis["hasil"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PPh Badan 31E"

    ws.append(["LAPORAN PPH BADAN PASAL 31E"])
    ws.append([f"Periode/Tahun Pajak: {hasil.get('tahun_pajak', '')}"])
    ws.append([f"Nama Perusahaan: {hasil.get('nama_perusahaan', '')}"])
    ws.append([])

    ws.append(["A. REKONSILIASI FISKAL"])
    rekon = hasil.get("rekonsiliasi_fiskal", {})
    ws.append(["Laba Bersih Komersial", rekon.get("laba_bersih_komersial", 0)])
    ws.append(["Koreksi Fiskal Positif", rekon.get("koreksi_fiskal_positif", 0)])
    ws.append(["Koreksi Fiskal Negatif", rekon.get("koreksi_fiskal_negatif", 0)])
    ws.append(["Penghasilan Neto Fiskal", rekon.get("penghasilan_neto_fiskal", 0)])
    ws.append(["Kompensasi Kerugian Fiskal", rekon.get("kompensasi_kerugian_fiskal", 0)])
    ws.append(["PKP (Pembulatan Ribuan)", rekon.get("penghasilan_kena_pajak", 0)])
    ws.append([])

    ws.append(["B. FASILITAS PASAL 31E"])
    fasilitas = hasil.get("fasilitas_31e", {})
    ws.append(["Peredaran Bruto", fasilitas.get("peredaran_bruto", 0)])
    ws.append(["PKP Mendapat Fasilitas", fasilitas.get("pkp_mendapat_fasilitas", 0)])
    ws.append(["PKP Tidak Mendapat Fasilitas", fasilitas.get("pkp_tidak_mendapat_fasilitas", 0)])
    ws.append(["Status Fasilitas", fasilitas.get("status_fasilitas", "")])
    ws.append([])

    ws.append(["C. PPH TERUTANG"])
    ws.append(["PPH atas PKP Fasilitas (11%)", hasil.get("pph_atas_pkp_fasilitas", 0)])
    ws.append(["PPH atas PKP Non-Fasilitas (22%)", hasil.get("pph_atas_pkp_nonfasilitas", 0)])
    ws.append(["PPH Badan Terutang", hasil.get("pph_badan_terutang", 0)])
    ws.append(["PPH Tanpa Fasilitas 31E", hasil.get("pph_tanpa_fasilitas_31e", 0)])
    ws.append(["Penghematan Pajak 31E", hasil.get("penghematan_pajak_pasal_31e", 0)])
    ws.append(["Tarif Efektif Riil", hasil.get("tarif_pajak_efektif_riil", 0)])
    ws.append([])

    ws.append(["D. KREDIT PAJAK"])
    kredit = hasil.get("kredit_pajak", {})
    ws.append(["PPH 22", kredit.get("pph_22", 0)])
    ws.append(["PPH 23", kredit.get("pph_23", 0)])
    ws.append(["PPH 24", kredit.get("pph_24", 0)])
    ws.append(["Angsuran PPH 25", kredit.get("angsuran_pph_25", 0)])
    ws.append(["Total Kredit Pajak", kredit.get("total", 0)])
    ws.append([])

    ws.append(["E. STATUS"])
    ws.append(["PPH Pasal 29 (Kurang Bayar)", hasil.get("pph_pasal_29_kurang_bayar", 0)])
    ws.append(["PPH Pasal 28a (Lebih Bayar)", hasil.get("pph_pasal_28a_lebih_bayar", 0)])
    ws.append(["Status", hasil.get("status", "")])
    ws.append([])
    ws.append(["Catatan:", hasil.get("catatan", "")])

    # [BARU] Kalau data Aset Tetap yang dipakai bukan dari tahun pajak yang
    # diminta (lihat peringatan_aset di endpoint generate), tampilkan juga
    # di file Excel -- supaya peringatan ini tidak hilang begitu saja kalau
    # yang dibaca cuma file export-nya, bukan response JSON generate-nya.
    if hasil.get("peringatan_data_aset_tetap"):
        ws.append([])
        ws.append(["⚠️ PERINGATAN DATA ASET TETAP", hasil.get("peringatan_data_aset_tetap")])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="PPh_Badan_{hasil.get("tahun_pajak", "")}.xlsx"'},
    )


# ============================================================
# [FASE 5 -- roadmap CALK] CATATAN ATAS LAPORAN KEUANGAN (CALK)
# ============================================================
# [KEPUTUSAN poin 13 roadmap] Client model (db_client.py) TIDAK ditambah
# tabel/kolom baru utk data profil CALK (akta, notaris, susunan
# komisaris/direksi) -- dipakai ULANG tabel `hasil_analisis` yang SUDAH
# ADA (jenis_analisis="calk_profil"), pola sama persis dengan
# jenis_analisis="pph_badan_31e" yang sudah jalan. Alasan: field ini
# spesifik utk 1 fitur (CALK), jarang berubah, dan tabel generik ini
# sudah tepat guna (key jenis_analisis + JSON hasil) -- migrasi tabel
# baru cuma menambah risiko tanpa manfaat nyata dibanding dipakai ulang.
# Kalau nanti field ini dipakai fitur LAIN juga (bukan cuma CALK), baru
# pertimbangkan naik kelas jadi kolom permanen di tabel clients.

class CalkProfilRequest(BaseModel):
    """Field PERSIS sama dengan CONTOH_PROFIL di modules/calk_export.py --
    lihat docstring tulis_note_1_umum()/tulis_note_2_kebijakan_akuntansi()
    utk detail tiap field dipakai di kalimat mana. Semua Optional supaya
    bisa diisi bertahap (field kosong otomatis jadi placeholder
    "-- lengkapi data --" saat CALK digenerate, TIDAK error)."""
    nama_perusahaan: Optional[str] = None
    nomor_akta_pendirian: Optional[str] = None
    tanggal_akta_pendirian: Optional[str] = None
    nama_notaris: Optional[str] = None
    no_sk_kemenkumham: Optional[str] = None
    tanggal_sk: Optional[str] = None
    nomor_akta_perubahan_terakhir: Optional[str] = None
    tanggal_akta_perubahan_terakhir: Optional[str] = None
    no_sk_perubahan_terakhir: Optional[str] = None
    tanggal_sk_perubahan: Optional[str] = None
    bidang_usaha_id: Optional[str] = None
    bidang_usaha_en: Optional[str] = None
    domisili_id: Optional[str] = None
    domisili_en: Optional[str] = None
    tahun_mulai_operasi: Optional[str] = None
    # Tiap item: [jabatan_id, nama_orang, jabatan_en] -- lihat
    # tulis_note_1_umum() di calk_export.py, dioper apa adanya.
    komisaris: List[Tuple[str, str, str]] = []
    direksi: List[Tuple[str, str, str]] = []
    kepala_cabang: List[Tuple[str, str, str]] = []
    jumlah_karyawan_lalu: Optional[int] = None
    tahun_karyawan_lalu: Optional[str] = None
    jumlah_karyawan_now: Optional[int] = None
    tahun_karyawan_now: Optional[str] = None
    umur_manfaat_inventaris: Optional[str] = None
    umur_manfaat_bangunan: Optional[str] = None


@app.post("/api/client/{client_id}/calk/profil")
def api_simpan_calk_profil(
    client_id: int, req: CalkProfilRequest,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """
    Simpan/update profil Note 1 (Umum) & Note 2 (Kebijakan Akuntansi)
    CALK untuk client ini. Histori TIDAK ditimpa -- tiap panggilan bikin
    entri baru (pola sama dgn laporan-keuangan/generate); endpoint GET
    di bawah selalu ambil yang TERBARU. Field kosong tetap disimpan apa
    adanya (None) -- validasi kelengkapan bukan tanggung jawab endpoint
    ini, tapi tulis_note_1_umum() (placeholder "-- lengkapi data --").
    """
    profil_id = dbc.simpan_hasil_analisis(
        client_id=client_id, jenis_analisis="calk_profil",
        hasil=req.model_dump(), model_ai="manual_input",
    )
    if profil_id is None:
        raise HTTPException(500, "Gagal menyimpan profil CALK.")

    dbc.log_audit(
        client_id=client_id, user=user.get("username", "unknown"),
        aksi="simpan_calk_profil", detail={"profil_id": profil_id},
    )
    return {"profil_id": profil_id, "profil": req.model_dump()}


@app.get("/api/client/{client_id}/calk/profil")
def api_ambil_calk_profil(
    client_id: int, user: dict = Depends(auth.require_level(3)),
):
    """Ambil profil CALK TERBARU client ini. Balik dict kosong (BUKAN
    404) kalau belum pernah diisi -- supaya frontend bisa langsung
    tampilkan form kosong tanpa perlu menangani error khusus dulu."""
    riwayat = dbc.ambil_hasil_analisis_client(client_id, jenis_analisis="calk_profil", limit=1)
    if not riwayat:
        return {"profil_id": None, "profil": {}}
    return {"profil_id": riwayat[0]["id"], "profil": riwayat[0]["hasil"]}


def _ambil_atau_generate_laporan_keuangan(
    client_id: int, periode: str,
    tanggal_mulai: Optional[str], tanggal_akhir: Optional[str],
    user: dict,
) -> Dict[str, Any]:
    """
    [FASE 5] Ambil snapshot laporan_keuangan utk 1 periode; kalau belum
    pernah digenerate, generate OTOMATIS dari jurnal terposting lalu
    simpan sbg snapshot baru -- pola SAMA PERSIS dgn fallback auto-
    generate di endpoint pph-badan/generate (lihat komentar di sana).
    Dipakai 2x oleh calk/generate (utk periode "now" & "lalu").

    Melempar HTTPException(404) kalau snapshot belum ada DAN jurnal utk
    tanggal_mulai/tanggal_akhir juga tidak diisi/tidak ada datanya --
    caller wajib tahu, bukan diam-diam dilewati.
    """
    lap = dbc.ambil_laporan_keuangan_terbaru(client_id, periode)
    if lap:
        return lap

    if not tanggal_mulai or not tanggal_akhir:
        raise HTTPException(
            404,
            f'Belum ada laporan keuangan utk periode "{periode}", dan '
            f"tanggal_mulai/tanggal_akhir tidak diisi utk generate otomatis. "
            f"Generate dulu lewat POST .../laporan-keuangan/generate, atau "
            f"isi tanggal_mulai_now/lalu & tanggal_akhir_now/lalu di request ini.",
        )

    coa = dbc.ambil_coa_client(client_id)
    jurnal = dbc.ambil_jurnal_terposting(client_id, tanggal_mulai, tanggal_akhir, hanya_terposting=False)
    if not jurnal:
        raise HTTPException(
            404,
            f'Belum ada laporan keuangan maupun jurnal utk periode "{periode}" '
            f"({tanggal_mulai} s/d {tanggal_akhir}). Pastikan ada data yang sudah "
            f"diproses (draft maupun terposting) sebelum generate CALK.",
        )
    data_laporan = lapkeu.generate_5_laporan_keuangan(jurnal, coa, periode)
    lap_id = dbc.simpan_laporan_keuangan(
        client_id, periode, data_laporan, dibuat_oleh=user.get("username", "unknown"),
        tanggal_mulai=tanggal_mulai, tanggal_akhir=tanggal_akhir,
    )
    return {"id": lap_id, "periode": periode, "data": data_laporan}


def _ambil_aset_tetap_untuk_calk(
    client_id: int, tanggal_lalu: date, tanggal_now: date,
) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    """
    [FASE 5] Ambil upload Aset Tetap TERBARU client ini (tabel `hasil`,
    jenis="aset_tetap" -- hasil proses_aset_tetap() yg disimpan lewat
    upload/proses-file, lihat akuntansi_ai.py) dan susun jadi format
    mutasi siap pakai Note 7 lewat calk_aset_tetap.siapkan_aset_tetap_untuk_calk().

    proses_aset_tetap() menyimpan DataFrame mentah per-aset di key "df"
    -- setelah lewat _bersihkan_untuk_json() (main.py) sebelum disimpan,
    key itu jadi LIST OF DICT (df.to_dict("records")), jadi di sini
    tinggal pd.DataFrame(...) lagi utk direkonstruksi. Kalau bentuknya
    ternyata beda dari yang diharapkan (mis. karena versi lama data
    sebelum "df" ikut disimpan), TIDAK melempar error ke caller --
    return (None, pesan_peringatan) supaya CALK tetap bisa digenerate
    TANPA Note 7 (skip graceful, sama seperti kalau aset_tetap=None
    dikirim ke susun_dan_tulis_semua_note_calk()), bukan gagal total.

    Returns:
        (aset_tetap_calk atau None, pesan_peringatan atau None)
    """
    hasil_aset = dbc.ambil_hasil_client(client_id, jenis="aset_tetap", limit=1)
    if not hasil_aset:
        return None, "Tidak ada data Aset Tetap yang pernah diupload -- Note Aset Tetap dilewati."

    data = hasil_aset[0]["data"] or {}
    df_records = data.get("df")
    if not df_records:
        return None, (
            f'Data Aset Tetap terbaru (diupload {hasil_aset[0].get("dibuat_at")}) tidak '
            f'punya field "df" (data mentah per-aset) -- kemungkinan format lama. '
            f"Note Aset Tetap dilewati, upload ulang file Aset Tetap kalau perlu Note ini."
        )

    try:
        df_aset_tetap = pd.DataFrame(df_records)
        aset_tetap_calk = calk_aset_tetap.siapkan_aset_tetap_untuk_calk(
            df_aset_tetap, tanggal_lalu=tanggal_lalu, tanggal_now=tanggal_now,
        )
    except Exception as e:  # noqa: BLE001
        # [Fase 4 poin 12 -- konsisten] error di sumber data OPSIONAL ini
        # TIDAK boleh menggagalkan generate CALK secara keseluruhan --
        # dicatat sbg peringatan, Note Aset Tetap dilewati.
        return None, f"Gagal menyusun data Aset Tetap utk CALK ({type(e).__name__}: {e}) -- Note Aset Tetap dilewati."

    if aset_tetap_calk.get("peringatan"):
        # Peringatan internal calk_aset_tetap (mis. kategori tidak dikenal,
        # indikasi aset dilepas) digabung ke pesan yg dibalik ke caller,
        # TAPI aset_tetap_calk TETAP dipakai (bukan None) -- ini
        # peringatan kualitas data, bukan kegagalan.
        pesan = "Peringatan data Aset Tetap: " + "; ".join(aset_tetap_calk["peringatan"])
        return aset_tetap_calk, pesan
    return aset_tetap_calk, None


class CalkGenerateRequest(BaseModel):
    periode_now: str  # mis. "2026-07" -- key snapshot laporan keuangan "sekarang"
    tanggal_mulai_now: Optional[str] = None  # dipakai HANYA kalau snapshot belum ada
    tanggal_akhir_now: str  # mis. "2026-07-31" -- SELALU dipakai sbg tanggal_now CALK
    periode_lalu: str  # mis. "2025" -- key snapshot pembanding
    tanggal_mulai_lalu: Optional[str] = None
    tanggal_akhir_lalu: str  # mis. "2025-12-31" -- SELALU dipakai sbg tanggal_lalu CALK
    # PPh Badan -- opsional, kalau tidak diisi otomatis dicari dari riwayat
    # pph-badan/generate tahun yg sama dgn tanggal_akhir_now (lihat di bawah).
    pph_badan_analisis_id: Optional[int] = None
    # PPh Final UMKM -- opsional, dipakai kalau client skema PP 55/2022
    # (bukan Tarif Umum Pasal 17/31E). Isi HANYA SALAH SATU dgn
    # pph_badan_analisis_id -- kalau dua-duanya kosong, Note Perpajakan
    # dilewati (lihat susun_dan_tulis_semua_note_calk()).
    pph_final_umkm: Optional[Dict[str, Any]] = None
    pihak_berelasi: Optional[Dict[str, List[str]]] = None
    peristiwa_setelah_neraca: Optional[Dict[str, List[str]]] = None
    tanggal_persetujuan: Optional[str] = None
    nama_penanggung_jawab_id: str = "Direksi"
    nama_penanggung_jawab_en: str = "Board of Directors"
    # [BARU] Grouping manual Note 4 Piutang Usaha (cabang/channel), by
    # no_akun -- opsional, lihat calk_mapping.susun_grouping_piutang_usaha()
    # utk format persis. None/tidak diisi = Note 4 tetap flat (perilaku
    # lama, tidak berubah).
    grouping_piutang_usaha: Optional[Dict[str, Any]] = None


@app.post("/api/client/{client_id}/calk/generate")
def api_generate_calk(
    client_id: int, req: CalkGenerateRequest,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """
    [FASE 5 poin 14] Generate CALK lengkap (docx + pdf): kumpulkan
    neraca+laba_rugi 2 periode (auto-generate kalau snapshot belum ada,
    lihat _ambil_atau_generate_laporan_keuangan()) + profil (Note 1/2) +
    Aset Tetap (Note 7, best-effort -- lihat _ambil_aset_tetap_untuk_calk())
    + PPh Badan/PPh Final UMKM (Note 15) -> panggil
    calk_export.export_calk() (Fase 4) -> simpan metadata hasil (path
    file, peringatan QA) ke tabel hasil_analisis -> log audit.

    TIDAK menyimpan isi file (docx/pdf) sbg BLOB di database -- file
    fisik tersimpan di FOLDER_HASIL (disk), yang disimpan di DB cuma
    NAMA filenya (lihat catatan keamanan path di endpoint download di
    bawah, pola sama dgn /api/unduh/{nama_file} yg sudah ada).
    """
    tanggal_now = datetime.strptime(req.tanggal_akhir_now, "%Y-%m-%d").date()
    tanggal_lalu = datetime.strptime(req.tanggal_akhir_lalu, "%Y-%m-%d").date()

    lap_now = _ambil_atau_generate_laporan_keuangan(
        client_id, req.periode_now, req.tanggal_mulai_now, req.tanggal_akhir_now, user,
    )
    lap_lalu = _ambil_atau_generate_laporan_keuangan(
        client_id, req.periode_lalu, req.tanggal_mulai_lalu, req.tanggal_akhir_lalu, user,
    )

    # --- Profil (Note 1 & 2) ---
    riwayat_profil = dbc.ambil_hasil_analisis_client(client_id, jenis_analisis="calk_profil", limit=1)
    profil = riwayat_profil[0]["hasil"] if riwayat_profil else {}

    # --- Aset Tetap (Note 7, best-effort) ---
    aset_tetap_calk, peringatan_aset = _ambil_aset_tetap_untuk_calk(client_id, tanggal_lalu, tanggal_now)

    # --- PPh Badan / PPh Final UMKM (Note 15) ---
    hasil_pph_badan = None
    peringatan_pph = None
    if req.pph_final_umkm:
        pass  # dipakai apa adanya, lihat kwargs_orchestrator di bawah
    elif req.pph_badan_analisis_id:
        rec = dbc.ambil_hasil_analisis_by_id(req.pph_badan_analisis_id)
        if rec and rec["jenis_analisis"] == "pph_badan_31e":
            hasil_pph_badan = rec["hasil"]
        else:
            peringatan_pph = (
                f"pph_badan_analisis_id={req.pph_badan_analisis_id} tidak ditemukan/bukan "
                f"analisis PPh Badan -- Note Perpajakan dicari otomatis dari riwayat tahun "
                f"{tanggal_now.year} sbg fallback."
            )
    if hasil_pph_badan is None and not req.pph_final_umkm:
        # Fallback: cari otomatis dari riwayat pph-badan/generate tahun yg
        # sama dgn tanggal_now (pola sama dgn pph-badan/generate sendiri
        # saat mencari data Aset Tetap yg cocok tahunnya).
        riwayat_pph = dbc.ambil_hasil_analisis_client(client_id, jenis_analisis="pph_badan_31e", limit=50)
        cocok = next((r for r in riwayat_pph if r["hasil"].get("tahun_pajak") == tanggal_now.year), None)
        if cocok:
            hasil_pph_badan = cocok["hasil"]
        elif not peringatan_pph:
            peringatan_pph = (
                f"Tidak ditemukan riwayat PPh Badan utk tahun {tanggal_now.year} -- "
                f"Note Perpajakan dilewati. Generate dulu lewat POST .../pph-badan/generate, "
                f"atau isi pph_final_umkm kalau client skema PPh Final UMKM."
            )

    tanggal_persetujuan = (
        datetime.strptime(req.tanggal_persetujuan, "%Y-%m-%d").date()
        if req.tanggal_persetujuan else None
    )

    # [FIX -- BUG NYATA] Sebelumnya panggilan ini TIDAK dibungkus try/except
    # sama sekali -- kalau calk_export.export_calk() melempar RuntimeError
    # (soffice tidak ditemukan/gagal convert/timeout, lihat penanganan error
    # yang SUDAH rapi & jelas di calk_export._convert_docx_ke_pdf()), error
    # itu naik TANPA ketangkep dan FastAPI membalasnya sbg 500 generik ke
    # frontend -- membuang percuma pesan error jelas yang sudah dibangun di
    # lapisan bawah. Melanggar checklist skill pdf-writing ("Error soffice
    # tidak ditemukan/gagal/timeout ditangani terpisah dengan pesan yang
    # bisa ditindaklanjuti, bukan 500 generik"). Disamakan di sini dgn pola
    # yang SUDAH benar di kertas_kerja_router.py (try/except spesifik,
    # detail pesan asli diteruskan ke HTTPException, bukan dibungkam).
    #
    # 502 Bad Gateway dipakai (bukan 500) -- kegagalan ini berasal dari
    # dependency eksternal (binary LibreOffice di server), bukan bug logic
    # aplikasi ini sendiri; docx tetap tersimpan di disk meski convert PDF
    # gagal (lihat catatan export_calk()), jadi pesan error menyebutkan itu
    # supaya akuntan/supervisor tahu docx-nya masih bisa diambil manual.
    try:
        hasil_export = calk_export.export_calk(
            output_dir=str(FOLDER_HASIL),
            nama_file_dasar=f"CALK_client{client_id}_{req.periode_now}",
            profil=profil,
            neraca_now=lap_now["data"]["neraca"], neraca_lalu=lap_lalu["data"]["neraca"],
            laba_rugi_now=lap_now["data"]["laba_rugi"], laba_rugi_lalu=lap_lalu["data"]["laba_rugi"],
            tanggal_now=tanggal_now, tanggal_lalu=tanggal_lalu,
            aset_tetap=aset_tetap_calk,
            hasil_pph_badan=hasil_pph_badan,
            pph_final_umkm=req.pph_final_umkm,
            pihak_berelasi=req.pihak_berelasi,
            peristiwa_setelah_neraca=req.peristiwa_setelah_neraca,
            tanggal_persetujuan=tanggal_persetujuan,
            nama_penanggung_jawab_id=req.nama_penanggung_jawab_id,
            nama_penanggung_jawab_en=req.nama_penanggung_jawab_en,
            grouping_piutang_usaha=req.grouping_piutang_usaha,
        )
    except RuntimeError as e:
        # [FIX] main.py tidak punya variabel `logger` sendiri -- pakai
        # logger modul calk_export.py (sudah ada, get_module_logger("calk_export"))
        # drpd membuat logger baru/bare-name yang tidak ada di namespace ini.
        calk_export.logger.error(f"❌ Gagal generate CALK utk client {client_id}: {e}")
        raise HTTPException(
            status_code=502,
            detail=(
                f"Gagal membuat file CALK (docx/PDF): {e} -- dokumen docx "
                f"kemungkinan tetap tersimpan di server meski convert PDF "
                f"gagal (lihat pesan di atas), hubungi admin/cek instalasi "
                f"LibreOffice di server kalau error berulang."
            ),
        )
    except Exception as e:  # noqa: BLE001 -- kegagalan tak terduga lain (mis. permission folder)
        calk_export.logger.error(f"❌ Gagal generate CALK utk client {client_id}: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Gagal generate CALK utk client {client_id}: {e}",
        )

    peringatan_gabungan = list(hasil_export["hasil_orchestrator"].get("peringatan") or [])
    if peringatan_aset:
        peringatan_gabungan.append(peringatan_aset)
    if peringatan_pph:
        peringatan_gabungan.append(peringatan_pph)

    metadata = {
        "periode_now": req.periode_now, "periode_lalu": req.periode_lalu,
        "tanggal_now": req.tanggal_akhir_now, "tanggal_lalu": req.tanggal_akhir_lalu,
        "docx_filename": os.path.basename(hasil_export["docx"]),
        "pdf_filename": os.path.basename(hasil_export["pdf"]),
        "nomor_note_terakhir": hasil_export["hasil_orchestrator"]["nomor_note_terakhir"],
        "daftar_note_ditulis": hasil_export["hasil_orchestrator"]["daftar_note_ditulis"],
        "peringatan": peringatan_gabungan,
    }
    calk_id = dbc.simpan_hasil_analisis(
        client_id=client_id, jenis_analisis="calk_generate",
        hasil=metadata, model_ai="rule_based",
    )

    dbc.log_audit(
        client_id=client_id, user=user.get("username", "unknown"),
        aksi="generate_calk",
        detail={
            "calk_id": calk_id, "periode_now": req.periode_now, "periode_lalu": req.periode_lalu,
            "jumlah_note": hasil_export["hasil_orchestrator"]["nomor_note_terakhir"],
            "jumlah_peringatan": len(peringatan_gabungan),
        },
    )

    return {"calk_id": calk_id, **metadata}


@app.get("/api/client/{client_id}/calk/riwayat")
def api_riwayat_calk(
    client_id: int, user: dict = Depends(auth.require_level(3)),
):
    """[FASE 5 poin 15] Riwayat semua CALK yang pernah digenerate client
    ini, terbaru dulu -- pola sama dgn GET .../pph-badan/riwayat."""
    return {"riwayat": dbc.ambil_hasil_analisis_client(client_id, jenis_analisis="calk_generate", limit=100)}


@app.get("/api/client/{client_id}/calk/{calk_id}/download")
def api_download_calk(
    client_id: int, calk_id: int, format: str = "pdf",
    user: dict = Depends(auth.get_current_user),
):
    """
    [FASE 5 poin 15] Serve file CALK (docx/pdf) yang sudah digenerate,
    lewat FileResponse -- pola sama dgn /api/unduh/{nama_file} yg sudah
    ada (termasuk pengamanan path yg sama: Path(...).name membuang
    komponen direktori dari nama file, jadi tidak bisa dipakai utk
    keluar dari FOLDER_HASIL walau nama file di DB entah bagaimana
    berisi "../").

    Args:
        format: "pdf" (default) atau "docx".
    """
    if format not in ("pdf", "docx"):
        raise HTTPException(400, 'Parameter format harus "pdf" atau "docx".')

    rec = dbc.ambil_hasil_analisis_by_id(calk_id)
    if not rec or rec["jenis_analisis"] != "calk_generate" or rec["client_id"] != client_id:
        raise HTTPException(404, "Riwayat CALK tidak ditemukan.")

    nama_file = rec["hasil"].get(f"{format}_filename")
    if not nama_file:
        raise HTTPException(404, f"File {format} tidak tercatat utk CALK id={calk_id}.")

    nama_aman = Path(nama_file).name
    path_file = FOLDER_HASIL / nama_aman
    if not path_file.exists():
        raise HTTPException(404, "File tidak ditemukan (mungkin sudah dihapus dari server).")

    media_type = (
        "application/pdf" if format == "pdf"
        else "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )
    return FileResponse(path_file, media_type=media_type, filename=nama_aman)


# ============================================================
# [BARU] LAPORAN BULANAN SETAHUN - ENDPOINT
# ============================================================

class GenerateLaporanBulananRequest(BaseModel):
    tahun: int  # 2026


@app.post("/api/client/{client_id}/laporan-bulanan/generate")
def api_generate_laporan_bulanan(
    client_id: int,
    req: GenerateLaporanBulananRequest,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """
    Generate Trial Balance, Laba Rugi, dan Balance Sheet bulanan
    Jan-Des dalam SATU tabel per laporan (12 kolom bulan).

    [BARU] hanya_terposting=False -- lihat catatan yang sama di
    laporan-keuangan/generate di atas.
    """
    jurnal = dbc.ambil_jurnal_terposting(
        client_id,
        tanggal_mulai=f"{req.tahun}-01-01",
        tanggal_akhir=f"{req.tahun}-12-31",
        hanya_terposting=False,
    )

    if not jurnal:
        raise HTTPException(404, f"Belum ada jurnal untuk tahun {req.tahun}")

    coa = dbc.ambil_coa_client(client_id)

    # [FIX -- POINT 3] sertakan_saldo_per_bulan=True: hasil hitung_saldo_per_akun()
    # utk tiap 12 bulan (yang tetap dihitung DI DALAM fungsi ini) ikut dikembalikan
    # lewat "_saldo_per_akun_per_bulan", jadi loop di bawah TIDAK perlu menghitung
    # ulang dari nol dengan filter tanggal yang malah berbeda (string compare vs
    # _tanggal_jurnal()/_akhir_bulan() yang dipakai internal) -- sumber duplikasi
    # & potensi hasil beda-tipis sudah dihapus di sini.
    hasil = lapkeu.susun_laporan_bulanan_setahun(jurnal, coa, req.tahun, sertakan_saldo_per_bulan=True)
    per_bulan_saldo = hasil.pop("_saldo_per_akun_per_bulan", [])

    # [BARU] Simpan snapshot saldo per akun untuk TIAP bulan ke
    # riwayat_saldo_bulanan -- ini satu-satunya tempat snapshot bulanan
    # dibuat (upload AR/AP aging manual TIDAK menyimpan histori bulanan,
    # cuma data upload TERBARU). Sheet "Ringkasan" pada export 18-sheet
    # (tren Piutang/Utang per bulan) bergantung pada data ini, jadi
    # generate laporan bulanan HARUS dijalankan dulu sebelum tren bisa
    # tampil.
    baris_tersimpan = 0
    for bulan in range(1, 13):
        saldo_per_akun = per_bulan_saldo[bulan - 1] if bulan - 1 < len(per_bulan_saldo) else {}
        baris_tersimpan += dbc.simpan_riwayat_saldo_bulanan(
            client_id=client_id, saldo_per_akun=saldo_per_akun, tahun=req.tahun, bulan=bulan,
        )
    hasil.setdefault("meta", {})["riwayat_saldo_tersimpan"] = baris_tersimpan

    analisis_id = dbc.simpan_hasil_analisis(
        client_id=client_id,
        jenis_analisis=f"laporan_bulanan_{req.tahun}",
        hasil=hasil,
        prompt=f"Laporan bulanan tahun {req.tahun}",
        model_ai="rule_based",
    )

    dbc.log_audit(
        client_id=client_id,
        user=user.get("username", "unknown"),
        aksi="generate_laporan_bulanan",
        detail={"tahun": req.tahun, "riwayat_saldo_tersimpan": baris_tersimpan},
    )

    return {"laporan_id": analisis_id, "hasil": hasil}


# ============================================================
# [BARU] RIWAYAT SALDO BULANAN - ENDPOINT (tren Piutang/Utang, dst)
# ============================================================

@app.get("/api/client/{client_id}/riwayat-saldo")
def api_ambil_riwayat_saldo(
    client_id: int,
    tahun: int,
    no_akun: Optional[str] = None,
    kategori: Optional[str] = None,
    user: dict = Depends(auth.get_current_user),
):
    """
    Ambil riwayat saldo bulanan (snapshot per akun per bulan, lihat
    endpoint POST .../laporan-bulanan/generate yang mengisinya).
    Kalau `no_akun` diisi, hanya 1 akun; kalau tidak, semua akun (bisa
    difilter `kategori`).
    """
    if no_akun:
        return {"tahun": tahun, "no_akun": no_akun, "data": dbc.ambil_riwayat_saldo_bulanan(client_id, no_akun, tahun)}
    return {"tahun": tahun, "data": dbc.ambil_riwayat_saldo_bulanan_client(client_id, tahun, kategori=kategori)}


@app.get("/api/client/{client_id}/riwayat-saldo/tren")
def api_ambil_tren_saldo(
    client_id: int,
    tahun: int,
    pola_no_akun: Optional[str] = None,
    kategori: Optional[str] = None,
    user: dict = Depends(auth.get_current_user),
):
    """
    Ambil tren saldo bulanan dikelompokkan per akun -- `pola_no_akun`
    dukung wildcard SQL LIKE (mis. "11%%" untuk semua akun Piutang
    dengan prefix "11").
    """
    hasil = dbc.ambil_riwayat_saldo_bulanan_akun_tren(client_id, tahun, pola_no_akun=pola_no_akun, kategori=kategori)
    return {"tahun": tahun, "jumlah_akun": len(hasil), "data": hasil}


@app.get("/api/client/{client_id}/riwayat-saldo/ringkasan")
def api_ringkasan_riwayat_saldo(
    client_id: int,
    tahun: int,
    user: dict = Depends(auth.get_current_user),
):
    """Ambil ringkasan tren saldo dikelompokkan per KATEGORI (ASET/LIABILITAS/dst) per bulan, untuk dashboard."""
    data = dbc.ambil_riwayat_saldo_bulanan_client(client_id, tahun)
    ringkasan: Dict[str, Dict[int, float]] = {}
    for item in data:
        kat = item.get("kategori") or "TIDAK_DIKENALI"
        ringkasan.setdefault(kat, {b: 0.0 for b in range(1, 13)})
        ringkasan[kat][item["bulan"]] += item.get("saldo_akhir", 0)

    hasil = [
        {
            "kategori": kat,
            "per_bulan": [round(per_bulan[b], 2) for b in range(1, 13)],
            "total": round(sum(per_bulan.values()), 2),
        }
        for kat, per_bulan in ringkasan.items()
    ]
    return {"tahun": tahun, "data": hasil}


@app.get("/api/client/{client_id}/laporan-bulanan/{tahun}")
def api_ambil_laporan_bulanan(
    client_id: int,
    tahun: int,
    user: dict = Depends(auth.require_level(3)),
):
    """Ambil laporan bulanan yang sudah pernah digenerate."""
    riwayat = dbc.ambil_hasil_analisis_client(
        client_id, jenis_analisis=f"laporan_bulanan_{tahun}", limit=1
    )
    if not riwayat:
        raise HTTPException(404, f"Belum ada laporan bulanan untuk tahun {tahun}")
    return riwayat[0]



# ============================================================
# [BARU] JADWAL PENYUSUTAN 12 BULAN - EXPORT
# ============================================================

@app.get("/api/client/{client_id}/jadwal-penyusutan/export")
def api_export_jadwal_penyusutan(
    client_id: int,
    tahun: int,
    metode: str = "komersial",  # "komersial" atau "fiskal"
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """
    Export jadwal penyusutan 12 bulan (sheet "Buku Bantu Aktiva Tetap")
    ke Excel, dari upload Aset Tetap TERBARU milik client.
    """
    if metode not in ("komersial", "fiskal"):
        raise HTTPException(400, "Parameter 'metode' harus 'komersial' atau 'fiskal'.")

    hasil_aset = dbc.ambil_hasil_client(client_id, jenis="aset_tetap", limit=1)
    if not hasil_aset:
        raise HTTPException(404, "Belum ada data Aset Tetap untuk client ini -- upload Daftar Aset Tetap terlebih dahulu.")

    records = hasil_aset[0]["data"].get("df") or []
    jadwal = lapkeu.susun_jadwal_penyusutan_bulanan(records, tahun, metode=metode)
    if not jadwal.get("aset"):
        raise HTTPException(404, f"Tidak ada aset yang disusutkan (metode {metode}) untuk tahun {tahun}.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Jadwal Penyusutan {metode.capitalize()}"

    ws.append([f"JADWAL PENYUSUTAN {metode.upper()} -- TAHUN {tahun}"])
    ws.append([f"Jumlah Aset: {jadwal['jumlah_aset']}"])
    ws.append([f"Total Penyusutan per Tahun: {jadwal['total_per_tahun']:,.0f}"])
    ws.append([])

    headers = (
        ["No", "Kode Aset", "Nama Aset", "Kategori", "Golongan", "Harga Perolehan",
         "Penyusutan/Bulan", "Penyusutan/Tahun", "Akumulasi Awal"]
        + [f"Bln {i}" for i in range(1, 13)]
        + ["Akumulasi Akhir", "Nilai Buku Akhir"]
    )
    baris_header = ws.max_row + 1
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=baris_header, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for idx, aset in enumerate(jadwal["aset"], 1):
        baris = [
            idx, aset.get("kode_aset"), aset.get("nama_aset"), aset.get("kategori"),
            aset.get("golongan_fiskal"), aset.get("harga_perolehan", 0),
            aset.get("penyusutan_per_bulan", 0), aset.get("penyusutan_per_tahun", 0),
            aset.get("akumulasi_awal_tahun", 0),
        ]
        baris += [b.get("penyusutan_bulan_ini", 0) for b in aset.get("jadwal_bulanan", [])]
        baris += [aset.get("akumulasi_akhir_tahun", 0), aset.get("nilai_buku_akhir_tahun", 0)]
        ws.append(baris)

    for row in range(baris_header + 1, ws.max_row + 1):
        for col in [6, 7, 8, 9] + list(range(10, 22)):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    for column in ws.columns:
        lebar = max((len(str(c.value)) for c in column if c.value is not None), default=0)
        ws.column_dimensions[column[0].column_letter].width = min(lebar + 2, 30)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Jadwal_Penyusutan_{metode}_{tahun}.xlsx"'},
    )


# ============================================================
# [BARU] EXPORT 14-SHEET LENGKAP
# ============================================================

class Export18SheetRequest(BaseModel):
    tahun: int  # 2026
    tahun_sebelumnya: Optional[int] = None
    metode_penyusutan: str = "komersial"  # "komersial" atau "fiskal"
    # Dipakai HANYA kalau laporan-keuangan/PPh Badan untuk tahun ini
    # belum pernah digenerate -- akan digenerate otomatis (sama seperti
    # endpoint pph-badan/generate & laporan-keuangan/generate).
    prive_atau_dividen: float = 0
    setoran_modal_baru: float = 0
    penyesuaian_ekuitas_manual: float = 0
    nama_perusahaan: Optional[str] = None
    kompensasi_kerugian_fiskal: float = 0
    kredit_pajak: Optional[Dict[str, float]] = None
    # [BARU] sama seperti di GeneratePPhBadanRequest -- dipakai HANYA kalau
    # PPh Badan tahun ini belum pernah digenerate lewat endpoint
    # pph-badan/generate (lihat blok "5. PPh Badan 31E" di bawah).
    skema_pajak: str = "Tarif Umum Pasal 17/31E"
    tambahan_peredaran_bruto_lainnya: float = 0
    retur_pengurangan_peredaran_bruto: float = 0
    keterangan_peredaran_bruto: Optional[str] = None


def _bangun_export_18_sheet(
    client_id: int, req: "Export18SheetRequest", user: dict,
    on_progress: Optional[Callable[..., None]] = None,
) -> bytes:
    """
    Export 18-sheet Excel lengkap (COA, Buku Bantu Piutang/Hutang/Aktiva
    Tetap, Trial Balance/Laba Rugi/Balance Sheet Bulanan, PPh Badan 31E,
    Lampiran SPT BS/PNL/Ekuitas rinci per kode akun, GL, Neraca Saldo
    Awal, Ringkasan) untuk SATU TAHUN PAJAK penuh.

    [REFACTOR] Bagian "kumpulkan & hitung data" dipindah ke
    _susun_data_export_18_sheet() (dipakai bareng oleh endpoint JSON
    preview di bawah) -- fungsi ini sekarang tinggal panggil itu lalu
    serialize ke .xlsx lewat modules.accounting_export.export_18_sheet_lengkap().

    Melempar HTTPException(404, ...) kalau data yang dibutuhkan (jurnal/
    laporan keuangan) belum ada sama sekali untuk tahun ini -- pemanggil
    (endpoint manual maupun auto-generate di proses_file_batch) yang
    memutuskan cara menampilkan pesan itu ke user.

    [BARU] on_progress: callback opsional dipanggil dgn (step: str,
    label: str, status: "processing"|"done", pesan: Optional[str]=None)
    di tiap tahap -- dipakai endpoint SSE
    /api/client/{client_id}/export-18-sheet/stream (lihat di bawah) untuk
    mengalirkan progress ke chat, PERSIS pola yang sama dengan
    _jalankan_generate_kertas_kerja(). Kalau None (dipanggil endpoint
    blocking lama / auto-generate di proses_file_batch), perilaku SAMA
    PERSIS seperti sebelumnya -- tidak ada efek samping tambahan.
    """
    def _lapor(step: str, label: str, status: str, pesan: Optional[str] = None) -> None:
        if on_progress:
            on_progress(step, label, status, pesan)

    # [BARU -- Point 3] Cek cache dulu -- kalau signature data client+tahun
    # ini belum berubah sejak generate terakhir, langsung pakai file Excel
    # yang sudah pernah dibuat, skip total (query DB + hitung ulang semua
    # sheet + serialize openpyxl). Lihat catatan struktur cache di atas
    # (dekat FOLDER_HASIL).
    kunci_cache = _kunci_cache_export_18_sheet(client_id, req)
    signature = dbc.hitung_signature_data_laporan(client_id, req.tahun)
    isi_excel = _ambil_cache_export_18_sheet(kunci_cache, signature)
    if isi_excel is not None:
        _lapor("cache", "Memakai laporan yang sudah pernah dibuat (belum ada perubahan data)", "done")
        dbc.log_audit(
            client_id=client_id, user=user.get("username", "unknown"),
            aksi="export_18_sheet_cache_hit",
            detail={"tahun": req.tahun},
        )
        return isi_excel

    data_export = _susun_data_export_18_sheet(client_id, req, user, on_progress=on_progress)
    _lapor("generate_excel", "Menyusun file Excel 18-sheet", "processing")
    isi_excel = accounting_export.export_18_sheet_lengkap(data_export)
    _lapor("generate_excel", "Menyusun file Excel 18-sheet", "done")
    _simpan_cache_export_18_sheet(kunci_cache, signature, isi_excel)

    dbc.log_audit(
        client_id=client_id, user=user.get("username", "unknown"),
        aksi="export_18_sheet",
        detail={
            "tahun": req.tahun,
            "jumlah_baris_jurnal": len(data_export.get("jurnal") or []),
            "jumlah_akun_coa": len(data_export.get("coa") or []),
        },
    )

    return isi_excel


def _bangun_preview_18_sheet_json(client_id: int, req: "Export18SheetRequest", user: dict) -> dict:
    """
    [BARU] Versi JSON dari _bangun_export_18_sheet() -- dipakai endpoint
    GET .../export-18-sheet-json supaya ke-18 sheet bisa ditampilkan
    LANGSUNG DI LAYAR (bukan cuma lewat file .xlsx yang di-download).

    Memakai _susun_data_export_18_sheet() yang SAMA PERSIS dipakai versi
    Excel -- data yang tampil di layar dijamin sinkron dengan yang ada di
    file .xlsx kalau di-download, karena berasal dari sumber yang sama.
    """
    # [BARU -- Point 3] Kunci cache SENGAJA dipisah dari versi Excel di atas
    # (prefix "json:") -- isinya beda bentuk (dict vs bytes), jangan
    # sampai versi JSON kebagian bytes Excel dari cache atau sebaliknya.
    # Signature-nya boleh sama (data sumbernya identik), cuma kunci &
    # slot cache-nya yang dipisah.
    kunci_cache = "json:" + _kunci_cache_export_18_sheet(client_id, req)
    signature = dbc.hitung_signature_data_laporan(client_id, req.tahun)
    hasil_json = _ambil_cache_export_18_sheet(kunci_cache, signature)
    if hasil_json is not None:
        dbc.log_audit(
            client_id=client_id, user=user.get("username", "unknown"),
            aksi="preview_export_18_sheet_json_cache_hit",
            detail={"tahun": req.tahun},
        )
        return hasil_json

    data_export = _susun_data_export_18_sheet(client_id, req, user)
    hasil_json = accounting_export.export_18_sheet_sebagai_json(data_export)
    _simpan_cache_export_18_sheet(kunci_cache, signature, hasil_json)

    dbc.log_audit(
        client_id=client_id, user=user.get("username", "unknown"),
        aksi="preview_export_18_sheet_json",
        detail={
            "tahun": req.tahun,
            "jumlah_baris_jurnal": len(data_export.get("jurnal") or []),
            "jumlah_akun_coa": len(data_export.get("coa") or []),
        },
    )

    return hasil_json


def _lengkapi_narasi_ai_export_18_sheet(
    calk: Dict[str, Any],
    asumsi: Dict[str, Any],
    neraca: Dict[str, Any],
    laba_rugi: Dict[str, Any],
    client_id: int,
) -> Dict[str, Any]:
    """
    [BARU] Lengkapi data export 18-sheet dengan narasi hasil Claude API:
    - calk["narasi_catatan"]  -> lewat claude_client.generate_narasi_calk_claude()
    - asumsi["narasi_ai"]     -> lewat claude_client.generate_narasi_asumsi_claude()
    - ringkasan_analisis      -> lewat claude_client.analisis_ringkasan_keuangan_claude()
      (dipakai sheet "Ringkasan", lihat return value fungsi ini)

    `calk`/`asumsi` dimutasi LANGSUNG (in-place) karena keduanya sudah
    dict yang akan diteruskan apa adanya ke accounting_export -- lebih
    simpel drpd bikin salinan yang isinya sama.

    SENGAJA best-effort per panggilan (try/except terpisah tiap fungsi,
    BUKAN 1 try/except besar): kalau salah satu panggilan Claude gagal
    (timeout/rate limit/dll), 2 panggilan lainnya TETAP jalan, dan
    export 18-sheet TETAP lanjut tanpa narasi yang gagal itu saja --
    bukan gagal total. Ini konsisten dgn cara accounting_export.py
    didesain: field2 narasi ini opsional, sheet tetap ter-generate
    normal kalau kosong (lihat komentar "[BARU -- integrasi Claude API]"
    di modules/accounting_export.py).

    Dipanggil SEBELUM data di-cache (lihat _susun_data_export_18_sheet)
    supaya panggilan Claude API cuma terjadi sekali per signature data,
    bukan tiap kali endpoint dipanggil.
    """
    ringkasan_analisis: Dict[str, Any] = {}
    cid = str(client_id)

    if calk.get("kerangka_catatan"):
        try:
            calk["narasi_catatan"] = claude_client.generate_narasi_calk_claude(
                calk["kerangka_catatan"], neraca, laba_rugi, asumsi, client_id=cid,
            )
        except Exception as e:
            claude_client.logger.warning(
                f"⚠️ Gagal generate narasi CALK (export 18-sheet) utk client {client_id}: {e}"
            )

    try:
        asumsi["narasi_ai"] = claude_client.generate_narasi_asumsi_claude(asumsi, client_id=cid)
    except Exception as e:
        claude_client.logger.warning(
            f"⚠️ Gagal generate narasi asumsi (export 18-sheet) utk client {client_id}: {e}"
        )

    try:
        ringkasan_analisis = claude_client.analisis_ringkasan_keuangan_claude(
            [{"neraca": neraca, "laba_rugi": laba_rugi}], client_id=cid,
        )
    except Exception as e:
        claude_client.logger.warning(
            f"⚠️ Gagal generate ringkasan analisis (export 18-sheet) utk client {client_id}: {e}"
        )

    return ringkasan_analisis


def _susun_data_export_18_sheet(
    client_id: int, req: "Export18SheetRequest", user: dict,
    on_progress: Optional[Callable[..., None]] = None,
) -> dict:
    """
    [REFACTOR] Badan asli dari _bangun_export_18_sheet() DIPINDAH ke sini
    apa adanya (logic "kumpulkan & hitung data" tidak diubah sama sekali)
    supaya bisa dipanggil dari 2 tempat: versi Excel (_bangun_export_18_sheet,
    dipakai endpoint download & auto-generate proses_file_batch) DAN versi
    JSON preview (_bangun_preview_18_sheet_json, dipakai endpoint
    .../export-18-sheet-json) -- tanpa duplikasi logic pengumpulan data,
    cuma beda di langkah terakhir (serialize ke .xlsx vs ke JSON).

    Menggabungkan data dari beberapa sumber yang sudah ada (menghitung
    ulang / menggenerate otomatis kalau snapshot untuk tahun ini belum
    pernah dibuat, mengikuti pola endpoint laporan-keuangan/generate dan
    pph-badan/generate).

    Melempar HTTPException(404, ...) kalau data yang dibutuhkan (jurnal/
    laporan keuangan) belum ada sama sekali untuk tahun ini -- pemanggil
    yang memutuskan cara menampilkan pesan itu ke user.

    Returns:
        dict: "data_export" siap dikirim ke
        modules.accounting_export.export_18_sheet_lengkap() ATAU
        modules.accounting_export.export_18_sheet_sebagai_json().

    [BARU] on_progress: callback opsional (step, label, status, pesan) --
    lihat docstring _bangun_export_18_sheet() untuk detail. Nama step
    ("coa", "jurnal", dst) & label bisa diganti bebas di sisi pemanggil
    (endpoint SSE) kalau mau teks lain di chat -- di sini cuma dipanggil
    apa adanya per tahap.
    """
    def _lapor(step: str, label: str, status: str, pesan: Optional[str] = None) -> None:
        if on_progress:
            on_progress(step, label, status, pesan)

    from modules import pph_badan, fiscal_reconciliation

    tahun = req.tahun
    periode_tahunan = str(tahun)
    tanggal_mulai = f"{tahun}-01-01"
    tanggal_akhir = f"{tahun}-12-31"

    _lapor("coa", "Membaca Chart of Account (COA)", "processing")
    coa = dbc.ambil_coa_client(client_id)
    _lapor("coa", "Membaca Chart of Account (COA)", "done")

    # [BARU] hanya_terposting=False -- endpoint ini TIDAK LAGI mewajibkan
    # akuntan mengonfirmasi-posting jurnal draft satu-satu dulu sebelum
    # laporan 18-sheet bisa dibuat. Baris draft (termasuk yang akunnya
    # masih placeholder) ikut apa adanya; status per baris ditandai lewat
    # kolom "Status Validasi" di sheet GL <tahun> (lihat
    # modules/accounting_export.py) -- bukan lewat gate status database
    # seperti sebelumnya. 3 endpoint lain (laporan-keuangan/generate,
    # pph-badan/generate, laporan-bulanan/generate) SENGAJA TIDAK diubah
    # -- tetap hanya_terposting=True (default) supaya laporan resmi yang
    # sudah disimpan/dikonfirmasi manual di halaman lain tidak terpengaruh.
    _lapor("jurnal", "Mengambil jurnal transaksi tahun berjalan", "processing")
    jurnal = dbc.ambil_jurnal_terposting(client_id, tanggal_mulai, tanggal_akhir, hanya_terposting=False)
    _lapor("jurnal", "Mengambil jurnal transaksi tahun berjalan", "done")

    # -- 1. Laporan keuangan (Neraca/Laba Rugi/Perubahan Ekuitas) setahun penuh --
    _lapor("laporan_keuangan", "Menyusun Neraca, Laba Rugi & Perubahan Ekuitas", "processing")
    lap = dbc.ambil_laporan_keuangan_terbaru(client_id, periode_tahunan)
    if not lap:
        if not jurnal:
            _lapor(
                "laporan_keuangan", "Menyusun Neraca, Laba Rugi & Perubahan Ekuitas", "error",
                f"Belum ada laporan keuangan maupun jurnal untuk tahun {tahun}.",
            )
            raise HTTPException(
                404,
                f"Belum ada laporan keuangan maupun jurnal untuk tahun {tahun}. "
                f"Pastikan ada data yang sudah diproses (draft maupun terposting) antara {tanggal_mulai} dan {tanggal_akhir}.",
            )
        data_laporan = lapkeu.generate_5_laporan_keuangan(
            jurnal, coa, periode_tahunan,
            prive_atau_dividen=req.prive_atau_dividen,
            setoran_modal_baru=req.setoran_modal_baru,
            penyesuaian_ekuitas_manual=req.penyesuaian_ekuitas_manual,
        )
        lap_id = dbc.simpan_laporan_keuangan(
            client_id, periode_tahunan, data_laporan, dibuat_oleh=user.get("username", "unknown"),
            tanggal_mulai=tanggal_mulai, tanggal_akhir=tanggal_akhir,
        )
        lap = {"id": lap_id, "periode": periode_tahunan, "data": data_laporan}
    _lapor("laporan_keuangan", "Menyusun Neraca, Laba Rugi & Perubahan Ekuitas", "done")

    neraca = lap["data"]["neraca"]
    laba_rugi = lap["data"]["laba_rugi"]
    perubahan_ekuitas = lap["data"]["perubahan_ekuitas"]
    # [FIX] Sebelumnya "arus_kas"/"calk" TIDAK PERNAH diteruskan ke
    # data_export padahal lapkeu.generate_5_laporan_keuangan() sudah
    # menghitung keduanya (lihat laporan_keuangan.py) -- akibatnya sheet
    # 12 "Laporan Arus Kas" & 13 "CALK" tidak pernah punya data untuk
    # ditampilkan sama sekali walau sudah dihitung. .get() dgn fallback {}
    # supaya tetap aman kalau ada snapshot laporan_keuangan LAMA yang
    # tersimpan sebelum "arus_kas"/"calk" ada di generate_5_laporan_keuangan().
    arus_kas = lap["data"].get("arus_kas") or {}
    calk = lap["data"].get("calk") or {}

    # -- 2. Lampiran SPT rinci per kode akun --
    _lapor("lampiran_spt", "Menyusun Lampiran SPT rinci per kode akun", "processing")
    lampiran_rinci = lapkeu.susun_lampiran_spt_lengkap_rinci(
        lap["data"], tahun_sebelumnya=req.tahun_sebelumnya, coa=coa,
    )
    _lapor("lampiran_spt", "Menyusun Lampiran SPT rinci per kode akun", "done")

    # -- 3. Laporan bulanan (Trial Balance/Laba Rugi/Balance Sheet 12 kolom) --
    _lapor("laporan_bulanan", "Menyusun Trial Balance/Laba Rugi/Balance Sheet bulanan", "processing")
    # Digenerate otomatis (dan disimpan, TERMASUK snapshot riwayat_saldo_
    # bulanan) kalau belum pernah ada untuk tahun ini -- supaya tren
    # Piutang/Utang di sheet Ringkasan selalu terisi walau user belum
    # pernah memanggil endpoint laporan-bulanan/generate secara terpisah.
    riwayat_lb = dbc.ambil_hasil_analisis_client(client_id, jenis_analisis=f"laporan_bulanan_{tahun}", limit=1)
    if riwayat_lb:
        laporan_bulanan = riwayat_lb[0]["hasil"]
    else:
        # [FIX -- POINT 3] sama seperti api_generate_laporan_bulanan() di atas --
        # pakai_ulang hasil per-bulan yang sudah dihitung, bukan hitung_saldo_per_akun()
        # ulang 12x dengan filter tanggal yang beda dari logic internalnya.
        laporan_bulanan = lapkeu.susun_laporan_bulanan_setahun(jurnal, coa, tahun, sertakan_saldo_per_bulan=True)
        per_bulan_saldo = laporan_bulanan.pop("_saldo_per_akun_per_bulan", [])
        for bulan in range(1, 13):
            saldo_per_akun = per_bulan_saldo[bulan - 1] if bulan - 1 < len(per_bulan_saldo) else {}
            dbc.simpan_riwayat_saldo_bulanan(client_id=client_id, saldo_per_akun=saldo_per_akun, tahun=tahun, bulan=bulan)
        dbc.simpan_hasil_analisis(
            client_id=client_id, jenis_analisis=f"laporan_bulanan_{tahun}",
            hasil=laporan_bulanan, prompt=f"Laporan bulanan tahun {tahun} (auto, dari export 18-sheet)",
            model_ai="rule_based",
        )
    _lapor("laporan_bulanan", "Menyusun Trial Balance/Laba Rugi/Balance Sheet bulanan", "done")

    # -- 4. Jadwal penyusutan (sheet "Buku Bantu Aktiva Tetap") --
    _lapor("aset_tetap", "Menghitung jadwal penyusutan Aktiva Tetap", "processing")
    hasil_aset = dbc.ambil_hasil_client(client_id, jenis="aset_tetap", limit=1)
    jadwal_aset: Dict[str, Any] = {}
    if hasil_aset:
        records_aset = hasil_aset[0]["data"].get("df") or []
        jadwal_aset = lapkeu.susun_jadwal_penyusutan_bulanan(records_aset, tahun, metode=req.metode_penyusutan)
    _lapor(
        "aset_tetap", "Menghitung jadwal penyusutan Aktiva Tetap",
        "done" if hasil_aset else "skip",
        None if hasil_aset else "Belum ada file Aset Tetap diupload -- sheet ini dikosongkan.",
    )

    # -- 5. PPh Badan 31E -- pakai riwayat yang cocok tahun_pajak, kalau
    # belum ada dihitung otomatis (sama seperti endpoint pph-badan/generate).
    _lapor("pph_badan", "Menghitung PPh Badan Pasal 31E", "processing")
    riwayat_pph = dbc.ambil_hasil_analisis_client(client_id, jenis_analisis="pph_badan_31e", limit=100)
    pph_cocok = next((r for r in riwayat_pph if (r.get("hasil") or {}).get("tahun_pajak") == tahun), None)
    if pph_cocok:
        pph_hasil = pph_cocok["hasil"]
    else:
        rekon_fiskal: Dict[str, Any] = {}
        if hasil_aset:
            rekon_fiskal = fiscal_reconciliation.ringkas_rekonsiliasi_fiskal_dari_aset_tetap(hasil_aset[0]["data"])
        pph_hasil = pph_badan.hitung_pph_pasal_31e(
            peredaran_bruto=laba_rugi.get("total_pendapatan", 0),
            laba_bersih_komersial=laba_rugi.get("laba_rugi_bersih", 0),
            tambahan_peredaran_bruto_lainnya=req.tambahan_peredaran_bruto_lainnya or 0,
            retur_pengurangan_peredaran_bruto=req.retur_pengurangan_peredaran_bruto or 0,
            koreksi_fiskal_positif=rekon_fiskal.get("koreksi_fiskal_positif", 0),
            koreksi_fiskal_negatif=rekon_fiskal.get("koreksi_fiskal_negatif", 0),
            kompensasi_kerugian_fiskal=req.kompensasi_kerugian_fiskal,
            kredit_pajak=req.kredit_pajak,
            tahun_pajak=tahun,
            nama_perusahaan=req.nama_perusahaan,
            skema_pajak=req.skema_pajak or "Tarif Umum Pasal 17/31E",
            keterangan_peredaran_bruto=req.keterangan_peredaran_bruto,
        )
        dbc.simpan_hasil_analisis(
            client_id=client_id, jenis_analisis="pph_badan_31e", hasil=pph_hasil,
            prompt=f"Tahun pajak {tahun} (auto, dari export 18-sheet)", model_ai="rule_based",
        )
    # [FIX] Sebelumnya TIDAK ADA _lapor(..., "done") di sini -- step
    # "pph_badan" jadi tersangkut selamanya di status "processing" di
    # tampilan chat (ProcessingSteps.jsx), walau perhitungannya sendiri
    # sudah selesai baris di atas.
    _lapor("pph_badan", "Menghitung PPh Badan Pasal 31E", "done")

    # -- 6. Buku Bantu Piutang / Hutang (upload TERBARU) --
    _lapor("piutang_hutang", "Membaca Buku Bantu Piutang & Hutang", "processing")
    _lapor("piutang_hutang", "Membaca Buku Bantu Piutang & Hutang", "processing")
    hasil_piutang = dbc.ambil_hasil_client(client_id, jenis="buku_bantu_piutang", limit=1)
    df_piutang = hasil_piutang[0]["data"].get("df") if hasil_piutang else None
    hasil_hutang = dbc.ambil_hasil_client(client_id, jenis="ap_aging", limit=1)
    df_hutang = hasil_hutang[0]["data"].get("df") if hasil_hutang else None
    _lapor(
        "piutang_hutang", "Membaca Buku Bantu Piutang & Hutang",
        "done" if (hasil_piutang or hasil_hutang) else "skip",
        None if (hasil_piutang or hasil_hutang) else "Belum ada file Piutang/Hutang diupload -- sheet terkait dikosongkan.",
    )

    # -- 7. Tren Piutang/Utang per bulan (dari riwayat_saldo_bulanan) --
    _lapor("tren_saldo", "Menghitung tren saldo Piutang/Utang per bulan", "processing")
    akun_piutang = [
        a.get("no_akun") for a in coa
        if a.get("kategori") == "ASET" and (
            a.get("sub_kategori") == "Piutang" or "PIUTANG" in str(a.get("nama_akun") or "").upper()
        )
    ]
    akun_utang = [
        a.get("no_akun") for a in coa
        if a.get("kategori") == "LIABILITAS" and (
            a.get("sub_kategori") == "Utang" or "UTANG" in str(a.get("nama_akun") or "").upper()
            or "HUTANG" in str(a.get("nama_akun") or "").upper()
        )
    ]
    riwayat_tren = dbc.ambil_riwayat_saldo_bulanan_akun_tren(client_id, tahun)
    tren_piutang = accounting_export.get_tren_saldo_per_bulan(riwayat_tren, akun_piutang)
    tren_utang = accounting_export.get_tren_saldo_per_bulan(riwayat_tren, akun_utang)
    _lapor("tren_saldo", "Menghitung tren saldo Piutang/Utang per bulan", "done")

    # -- 8. Petunjuk & Asumsi --
    # [FIX] Sebelumnya key "asumsi" TIDAK PERNAH dikirim ke data_export,
    # padahal accounting_export.export_18_sheet_lengkap() membaca
    # data.get("asumsi") untuk mengisi sheet "Petunjuk & Asumsi" -- akibatnya
    # Periode Awal/Periode Akhir selalu kosong (padahal tanggal_mulai/
    # tanggal_akhir sudah dihitung di atas), dan Nama Perusahaan cuma
    # kebetulan terisi kalau sudah nempel di cache pph_hasil dari generate
    # sebelumnya (tidak konsisten). Nama perusahaan sekarang diambil
    # langsung dari data client (sumber yang selalu ada), req.nama_perusahaan
    # dipakai sebagai override kalau eksplisit dikirim di request.
    info_client = dbc.ambil_client(client_id) or {}
    asumsi = {
        "nama_perusahaan": req.nama_perusahaan or info_client.get("nama") or "",
        "periode_awal": tanggal_mulai,
        "periode_akhir": tanggal_akhir,
        # [BARU] Sebelumnya tidak dikirim -- accounting_export.py sudah
        # sedia baris "Tanggal Laporan / Aging" tapi selalu kosong.
        # Dipakai tanggal_akhir (akhir periode laporan) sebagai cutoff
        # aging piutang/hutang, konsisten dgn field yg sama dipakai di
        # proses_ap_aging()/proses_piutang() saat generate laporan bulanan.
        "tanggal_laporan": tanggal_akhir,
    }

    # -- 9. Narasi AI (CALK naratif, catatan asumsi, ringkasan analisis) --
    # [BARU] Best-effort, lihat _lengkapi_narasi_ai_export_18_sheet() --
    # dipanggil di sini (SEBELUM return, SEBELUM hasil di-cache oleh
    # pemanggil) supaya Claude API cuma dipanggil sekali per signature
    # data client+tahun, bukan tiap kali endpoint diakses selama cache
    # masih valid.
    _lapor("narasi_ai", "Menyusun narasi CALK & ringkasan analisis (AI)", "processing")
    ringkasan_analisis = _lengkapi_narasi_ai_export_18_sheet(
        calk, asumsi, neraca, laba_rugi, client_id,
    )
    _lapor("narasi_ai", "Menyusun narasi CALK & ringkasan analisis (AI)", "done")

    # -- 10. Kembalikan data siap-pakai (pemanggil yang memutuskan Excel/JSON) --
    return {
        "periode": periode_tahunan,
        "tahun_sebelumnya": req.tahun_sebelumnya,
        "coa": coa,
        "jurnal": jurnal,
        "df_piutang": df_piutang,
        "df_hutang": df_hutang,
        "jadwal_aset": jadwal_aset,
        "laporan_bulanan": laporan_bulanan,
        "pph_hasil": pph_hasil,
        "neraca": neraca,
        "laba_rugi": laba_rugi,
        "perubahan_ekuitas": perubahan_ekuitas,
        "arus_kas": arus_kas,
        "calk": calk,
        "lampiran_rinci": lampiran_rinci,
        "tren_piutang": tren_piutang,
        "tren_utang": tren_utang,
        "asumsi": asumsi,
        "ringkasan_analisis": ringkasan_analisis,
    }


@app.post("/api/client/{client_id}/export-18-sheet")
def api_export_18_sheet(
    client_id: int,
    req: Export18SheetRequest,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas
):
    """Endpoint tipis -- badan logic ada di _bangun_export_18_sheet()
    supaya bisa dipakai ulang dari proses_file_batch() (auto-generate)."""
    isi_excel = _bangun_export_18_sheet(client_id, req, user)
    return StreamingResponse(
        io.BytesIO(isi_excel),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Laporan_Keuangan_{req.tahun}_18_Sheet.xlsx"'},
    )


@app.post("/api/client/{client_id}/export-18-sheet-json")
def api_export_18_sheet_json(
    client_id: int,
    req: Export18SheetRequest,
    user: dict = Depends(auth.require_level(3)),  # Supervisor ke atas -- sama seperti versi Excel
):
    """
    [BARU] Versi JSON dari POST .../export-18-sheet -- supaya ke-18 sheet
    (Petunjuk & Asumsi, COA, Neraca Saldo Awal, GL, Buku Bantu Piutang/
    Hutang/Aktiva Tetap, Trial Balance/Laba Rugi/Balance Sheet Bulanan,
    Ringkasan, BS/PNL Lampiran SPT, PPh Badan 31E) bisa ditampilkan
    LANGSUNG DI LAYAR (tab per-sheet), bukan cuma lewat file .xlsx yang
    di-download. Menerima parameter request yang SAMA PERSIS dengan
    endpoint Excel-nya (tahun, tahun_sebelumnya, dst).

    Badan logic ada di _bangun_preview_18_sheet_json() -- memakai ulang
    _susun_data_export_18_sheet() yang sama dengan versi Excel, supaya
    data yang tampil di layar selalu sinkron dengan file yang di-download.
    """
    return _bangun_preview_18_sheet_json(client_id, req, user)


@app.get("/api/template-laporan-keuangan")
def unduh_template_laporan_keuangan():
    """[FIX] Endpoint baru: download template Excel kosong 31 sheet standar
    laporan keuangan, untuk user yang belum punya file & mau mulai dari nol."""
    isi_bytes = ak.generate_template_31_sheet()
    return StreamingResponse(
        io.BytesIO(isi_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_laporan_keuangan_31_sheet.xlsx"},
    )


@app.get("/api/unduh/{nama_file}")
def unduh_file_hasil(nama_file: str, user: dict = Depends(auth.get_current_user)):
    nama_aman = Path(nama_file).name
    path_file = FOLDER_HASIL / nama_aman
    if not path_file.exists():
        raise HTTPException(status_code=404, detail="File tidak ditemukan (mungkin sudah dihapus).")
    return FileResponse(
        path_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="Hasil_Proses_Akuntansi.xlsx",
    )