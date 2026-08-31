"""

test_akuntansi_ai.py

=====================

Unit test NYATA untuk logika inti akuntansi_ai.py -- semuanya jalan LOKAL,

tanpa API key dan tanpa koneksi internet, jadi bisa (dan sebaiknya) dijalankan

di setiap perubahan kode sebelum dipakai ke data klien sungguhan.



Cara jalankan:

    pip install pytest pandas openpyxl --break-system-packages

    pytest test_akuntansi_ai.py -v



Untuk smoke test koneksi API (butuh DEEPSEEK_API_KEY), lihat test_koneksi_api.py

-- file itu terpisah supaya test logika inti tidak pernah gagal gara-gara

masalah jaringan/kuota API.

"""

import json

import math

import os

import tempfile



import pandas as pd

import pytest



import akuntansi_ai as ak





# ============================================================

# ekstrak_signature -- dasar dari pembelajaran pola

# ============================================================



def test_ekstrak_signature_ambil_token_pertama():

    assert ak.ekstrak_signature("NBMB 123456 an. Budi") == "NBMB"





def test_ekstrak_signature_buang_angka_di_ekor_token():

    assert ak.ekstrak_signature("BRIVA12345") == "BRIVA"

    assert ak.ekstrak_signature("BRIVA99999") == "BRIVA"





def test_ekstrak_signature_kosong_atau_none():

    assert ak.ekstrak_signature(None) == "TIDAK_ADA_KETERANGAN"

    assert ak.ekstrak_signature("") == "TIDAK_ADA_KETERANGAN"

    assert ak.ekstrak_signature(float("nan")) == "TIDAK_ADA_KETERANGAN"





# ============================================================

# mask_keterangan_sensitif -- privasi data sebelum ke AI

# ============================================================



def test_mask_menyamarkan_digit_panjang_sisakan_4_akhir():

    hasil = ak.mask_keterangan_sensitif("TRSF ke rekening 1234567890123 an. Budi")

    assert "1234567890123" not in hasil

    assert "0123" in hasil





def test_mask_tidak_menyentuh_digit_pendek():

    hasil = ak.mask_keterangan_sensitif("Bayar 50000 kode 12345")

    assert hasil == "Bayar 50000 kode 12345"





def test_mask_none_dan_nan_aman():

    assert ak.mask_keterangan_sensitif(None) is None

    assert pd.isna(ak.mask_keterangan_sensitif(float("nan")))





# ============================================================

# cek_keseimbangan_jurnal -- validasi double-entry global

# ============================================================



def test_keseimbangan_balance():

    df = pd.DataFrame({

        "jml_debet": [100000, 50000],

        "jml_kredit": [100000, 50000],

        "sumber_kategori": ["AI (DeepSeek)", "Kata kunci COA"],

    })

    hasil = ak.cek_keseimbangan_jurnal(df)

    assert hasil["balance"] is True

    assert hasil["selisih"] == 0

    assert hasil["jumlah_baris_belum_terkategori"] == 0





def test_keseimbangan_tidak_balance_terdeteksi():

    df = pd.DataFrame({

        "jml_debet": [100000, 50000],

        "jml_kredit": [100000, 40000],

        "sumber_kategori": ["AI (DeepSeek)", "Kata kunci COA"],

    })

    hasil = ak.cek_keseimbangan_jurnal(df)

    assert hasil["balance"] is False

    assert hasil["selisih"] == pytest.approx(10000)





def test_keseimbangan_hitung_jml_kredit_ppn():

    df = pd.DataFrame({

        "jml_debet": [110000],

        "jml_kredit": [100000],

        "jml_kredit_ppn": [10000],

        "sumber_kategori": ["Aturan standar penjualan + kata kunci COA"],

    })

    hasil = ak.cek_keseimbangan_jurnal(df)

    assert hasil["balance"] is True





def test_keseimbangan_menghitung_baris_perlu_review():

    df = pd.DataFrame({

        "jml_debet": [100000, 50000],

        "jml_kredit": [100000, 50000],

        "sumber_kategori": [

            "AI (DeepSeek) - confidence rendah, perlu review",

            "Belum Terkategori - perlu review manual",

        ],

    })

    hasil = ak.cek_keseimbangan_jurnal(df)

    assert hasil["jumlah_baris_belum_terkategori"] == 2





def test_keseimbangan_dataframe_kosong():

    hasil = ak.cek_keseimbangan_jurnal(pd.DataFrame())

    assert hasil["balance"] is True

    assert hasil["total_debet"] == 0.0





# ============================================================

# pelajari_pola -- deteksi pola konsisten vs tidak konsisten

# ============================================================



def _baris_bank(keterangan, debet, kredit, no_d, nama_d, no_k, nama_k):

    return {

        "keterangan": keterangan, "mutasi_debet": debet, "mutasi_kredit": kredit,

        "no_akun_debet": no_d, "nama_akun_debet": nama_d,

        "no_akun_kredit": no_k, "nama_akun_kredit": nama_k,

    }





def test_pelajari_pola_konsisten_kalau_selalu_sama():

    df = pd.DataFrame([

        _baris_bank("PLNPOST001", 500000, 0, "5100", "BEBAN LISTRIK", "1100", "BANK"),

        _baris_bank("PLNPOST002", 520000, 0, "5100", "BEBAN LISTRIK", "1100", "BANK"),

    ])

    pola = ak.pelajari_pola(df, min_samples=2)

    aturan = pola.aturan[("PLNPOST", "KELUAR")]

    assert aturan["konsisten"] is True

    assert aturan["no_akun_debet"] == "5100"





def test_pelajari_pola_tidak_konsisten_kalau_pasangan_akun_beda():

    df = pd.DataFrame([

        _baris_bank("TRSF001", 500000, 0, "5100", "BEBAN A", "1100", "BANK"),

        _baris_bank("TRSF002", 500000, 0, "5200", "BEBAN B", "1100", "BANK"),

        _baris_bank("TRSF003", 500000, 0, "5200", "BEBAN B", "1100", "BANK"),

    ])

    pola = ak.pelajari_pola(df)

    aturan = pola.aturan[("TRSF", "KELUAR")]

    assert aturan["konsisten"] is False





# ============================================================

# TEST FUNGSI BARU: pelajari_pola_enhanced

# ============================================================



def test_pelajari_pola_enhanced_min_samples():

    df = pd.DataFrame([

        _baris_bank("UNIK001", 100000, 0, "5100", "BEBAN", "1100", "BANK"),

        _baris_bank("UNIK002", 200000, 0, "5200", "BEBAN", "1100", "BANK"),

    ])

    pola = ak.pelajari_pola(df, min_samples=3)

    assert ("UNIK", "KELUAR") not in pola.aturan





def test_pelajari_pola_enhanced_confidence_score():

    df = pd.DataFrame([

        _baris_bank("TEST001", 100000, 0, "5100", "BEBAN A", "1100", "BANK"),

        _baris_bank("TEST002", 100000, 0, "5100", "BEBAN A", "1100", "BANK"),

        _baris_bank("TEST003", 100000, 0, "5200", "BEBAN B", "1100", "BANK"),

    ])

    pola = ak.pelajari_pola(df, min_samples=2)

    aturan = pola.aturan[("TEST", "KELUAR")]

    assert aturan["confidence_score"] == pytest.approx(0.667, rel=0.01)

    assert aturan["jumlah_contoh"] == 2





# ============================================================

# TEST FUNGSI BARU: augmentasi_pola

# ============================================================



def test_augmentasi_pola_variasi():

    pola = ak.Pola()

    pola.aturan[("PLN", "KELUAR")] = {

        "no_akun_debet": "5100",

        "nama_akun_debet": "BEBAN LISTRIK",

        "no_akun_kredit": "1100",

        "nama_akun_kredit": "BANK",

        "konsisten": True,

        "jumlah_contoh": 3,

        "confidence_score": 0.9,

        "is_valid": True

    }

    

    augmented = ak.augmentasi_pola(pola)

    assert len(augmented.aturan) > 1

    has_variation = any('LISTRIK' in key[0] for key in augmented.aturan.keys())

    assert has_variation





def test_augmentasi_pola_filter_confidence():

    pola = ak.Pola()

    pola.aturan[("PLN", "KELUAR")] = {

        "no_akun_debet": "5100",

        "nama_akun_debet": "BEBAN LISTRIK",

        "no_akun_kredit": "1100",

        "nama_akun_kredit": "BANK",

        "konsisten": True,

        "jumlah_contoh": 1,

        "confidence_score": 0.3,

        "is_valid": False

    }

    

    augmented = ak.augmentasi_pola(pola, threshold=0.6)

    assert len(augmented.aturan) == 1





# ============================================================

# TEST FUNGSI BARU: pelajari_pola_penjualan

# ============================================================



def test_pelajari_pola_penjualan_dasar():

    df = pd.DataFrame([

        {

            "keterangan": "PENJUALAN BARANG A",

            "customer": "PT ABC",

            "cara_bayar": "TUNAI",

            "dpp": 100000,

            "ppn": 11000,

            "total": 111000,

            "no_akun_debet": "1100",

            "nama_akun_debet": "KAS",

            "jml_debet": 111000,

            "no_akun_kredit": "4100",

            "nama_akun_kredit": "PENJUALAN",

            "jml_kredit": 100000,

            "no_akun_kredit_ppn": "2100",

            "nama_akun_kredit_ppn": "PPN KELUARAN",

            "jml_kredit_ppn": 11000,

        },

    ])

    

    pola = ak.pelajari_pola_penjualan(df, min_samples=1)

    assert len(pola.aturan) > 0

    assert any('PENJUALAN' in key[0] for key in pola.aturan.keys())





# ============================================================

# TEST FUNGSI BARU: deteksi_pola_mencurigakan

# ============================================================



def test_deteksi_pola_mencurigakan():

    pola = ak.Pola()

    for i in range(10):

        pola.aturan[(f"SIG{i}", "MASUK")] = {

            "no_akun_debet": "1100",

            "nama_akun_debet": "KAS",

            "no_akun_kredit": "9999",

            "nama_akun_kredit": "AKUN DEFAULT",

            "jumlah_contoh": 1,

        }

    

    pola.aturan[("PLN", "KELUAR")] = {

        "no_akun_debet": "5100",

        "nama_akun_debet": "BEBAN LISTRIK",

        "no_akun_kredit": "1100",

        "nama_akun_kredit": "KAS",

        "jumlah_contoh": 3,

    }

    

    temuan = ak.deteksi_pola_mencurigakan(pola, ambang=5)

    assert len(temuan) >= 1

    found = any(t["no_akun_debet"] == "1100" and t["no_akun_kredit"] == "9999" for t in temuan)

    assert found





# ============================================================

# TEST FUNGSI BARU: generate_rekomendasi_akuntansi

# ============================================================



def test_generate_rekomendasi_akuntansi():

    df_bank = pd.DataFrame({

        "keterangan": ["PLN 001", "PLN 002", "TRSF 001"],

        "sumber_kategori": ["Belum Terkategori - perlu review manual", 

                           "Sesuai Pola", 

                           "Belum Terkategori - perlu review manual"],

        "mutasi_debet": [100000, 0, 200000],

        "mutasi_kredit": [0, 50000, 0],

        "jml_debet": [100000, None, 200000],

        "jml_kredit": [None, 50000, None],

    })

    

    df_jual = pd.DataFrame()

    df_coa = pd.DataFrame({"no_akun": [1100, 5100], "nama_akun": ["KAS", "BEBAN LISTRIK"]})

    pola = ak.Pola()

    

    rekomendasi = ak.generate_rekomendasi_akuntansi(df_bank, df_jual, df_coa, pola)

    assert rekomendasi["total_rekomendasi"] > 0

    assert "prioritas_tinggi" in rekomendasi

    assert len(rekomendasi["prioritas_tinggi"]) >= 1





# ============================================================

# TEST FUNGSI BARU: get_smart_suggestions

# ============================================================



def test_get_smart_suggestions():

    df_bank = pd.DataFrame({

        "keterangan": ["PLN 001"],

        "sumber_kategori": ["Belum Terkategori - perlu review manual"],

        "mutasi_debet": [100000],

        "mutasi_kredit": [0],

        "jml_debet": [None],

        "jml_kredit": [None],

    })

    df_jual = pd.DataFrame()

    df_coa = pd.DataFrame({"no_akun": [1100, 5100], "nama_akun": ["KAS", "BEBAN LISTRIK"]})

    pola = ak.Pola()

    

    suggestions = ak.get_smart_suggestions(df_bank, df_jual, df_coa, pola)

    assert len(suggestions) > 0

    has_warning = any("belum terkategori" in s["text"].lower() for s in suggestions)

    assert has_warning





# ============================================================

# TEST FUNGSI BARU: auto_correct_journal

# ============================================================



def test_auto_correct_journal():

    df = pd.DataFrame({

        "keterangan": ["PLN 001"],

        "no_akun_debet": ["5100"],

        "no_akun_kredit": ["5100"],

        "jml_debet": [100000],

        "jml_kredit": [100000],

        "mutasi_debet": [100000],

        "mutasi_kredit": [0],

    })

    

    df_coa = pd.DataFrame({"no_akun": [5100, 1100], "nama_akun": ["BEBAN LISTRIK", "KAS"]})

    pola = ak.Pola()

    pola.aturan[("PLN", "KELUAR")] = {

        "no_akun_debet": "5100",

        "nama_akun_debet": "BEBAN LISTRIK",

        "no_akun_kredit": "1100",

        "nama_akun_kredit": "KAS",

        "konsisten": True,

        "jumlah_contoh": 1,

    }

    

    df_fix, log = ak.auto_correct_journal(df, df_coa, pola)

    assert len(log) > 0

    assert "Akun debet/kredit diperbaiki berdasarkan pola" in log[0]["perbaikan"][0]





# ============================================================

# TEST FUNGSI BARU: batch_process_dataframes

# ============================================================



def test_batch_process_dataframes():

    df_bank = pd.DataFrame({

        "keterangan": ["PLN 001"],

        "mutasi_debet": [100000],

        "mutasi_kredit": [0],

        "no_akun_debet": ["5100"],

        "no_akun_kredit": ["1100"],

    })

    

    df_jual = pd.DataFrame({

        "keterangan": ["PENJUALAN"],

        "customer": ["PT ABC"],

        "cara_bayar": ["TUNAI"],

        "dpp": [100000],

        "ppn": [0],

        "total": [100000],

    })

    

    dataframes = {

        "bank": df_bank,

        "jual": df_jual,

    }

    

    df_coa = pd.DataFrame()

    pola = ak.Pola()

    

    progress_log = []

    def callback(i, total, msg):

        progress_log.append((i, total, msg))

    

    hasil = ak.batch_process_dataframes(dataframes, df_coa, pola, pakai_ai=False, callback=callback)

    assert "bank" in hasil

    assert "jual" in hasil

    assert len(progress_log) == 2





# ============================================================

# TEST FUNGSI BARU: format_jurnal_untuk_export

# ============================================================



def test_format_jurnal_untuk_export():

    df = pd.DataFrame({

        "tanggal": ["2024-01-01"],

        "keterangan": ["PLN 001"],

        "no_akun_debet": ["5100"],

        "nama_akun_debet": ["BEBAN LISTRIK"],

        "no_akun_kredit": ["1100"],

        "nama_akun_kredit": ["KAS"],

        "jml_debet": [100000],

        "jml_kredit": [100000],

    })

    

    result = ak.format_jurnal_untuk_export(df, format_type='jurnal_umum')

    assert result["status"] == "success"

    assert result["format"] == "jurnal_umum"

    assert len(result["data"]) == 2

    

    result = ak.format_jurnal_untuk_export(df, format_type='csv')

    assert result["format"] == "csv"

    assert isinstance(result["data"], str)

    

    result = ak.format_jurnal_untuk_export(df, format_type='json')

    assert result["format"] == "json"

    assert isinstance(result["data"], list)





# ============================================================

# TEST FUNGSI BARU: cleanup_dataframe

# ============================================================



def test_cleanup_dataframe():

    df = pd.DataFrame({

        "a": [1, 2, None],

        "b": ["x", None, "z"],

        "c": [None, None, None],

    })

    

    df_clean = ak.cleanup_dataframe(df)

    assert "c" not in df_clean.columns

    assert len(df_clean) > 0





# ============================================================

# TEST FUNGSI BARU: bandingkan_periode

# ============================================================



def test_bandingkan_periode():

    df1 = pd.DataFrame({

        "jml_debet": [100000, 50000],

        "jml_kredit": [100000, 50000],

        "sumber_kategori": ["Sesuai Pola", "Sesuai Pola"],

    })

    

    df2 = pd.DataFrame({

        "jml_debet": [150000, 75000],

        "jml_kredit": [150000, 75000],

        "sumber_kategori": ["Sesuai Pola", "Sesuai Pola"],

    })

    

    hasil = ak.bandingkan_periode(df1, df2, "Januari", "Februari")

    assert hasil["status"] == "success"

    assert "perbandingan" in hasil

    assert "perubahan" in hasil["perbandingan"]





# ============================================================

# TEST FUNGSI BARU: validate_journal_entry

# ============================================================



def test_validate_journal_entry():

    row = {

        "no_akun_debet": "5100",

        "no_akun_kredit": "1100",

        "jml_debet": 100000,

        "jml_kredit": 100000,

        "tanggal": "2024-01-01",

        "keterangan": "PLN 001",

    }

    result = ak.validate_journal_entry(row)

    assert result["valid"] is True

    assert len(result["errors"]) == 0

    

    row_invalid = {

        "no_akun_debet": "5100",

        "no_akun_kredit": "5100",

        "jml_debet": 100000,

        "jml_kredit": 100000,

    }

    result = ak.validate_journal_entry(row_invalid)

    assert result["valid"] is False

    assert len(result["errors"]) > 0





# ============================================================

# muat_pola / simpan_pola -- ketahanan terhadap file korup

# ============================================================



def test_simpan_dan_muat_pola_roundtrip(tmp_path):

    path = str(tmp_path / "pola.json")

    pola = ak.Pola(aturan={("SIG", "MASUK"): {

        "no_akun_debet": "1100", "nama_akun_debet": "BANK",

        "no_akun_kredit": "4100", "nama_akun_kredit": "PENJUALAN",

        "konsisten": True, "jumlah_contoh": 3,

        "confidence_score": 0.95,

        "is_valid": True

    }})

    ak.simpan_pola(pola, path)

    pola_dimuat = ak.muat_pola(path)

    assert ("SIG", "MASUK") in pola_dimuat.aturan

    assert pola_dimuat.aturan[("SIG", "MASUK")]["no_akun_debet"] == "1100"

    assert pola_dimuat.aturan[("SIG", "MASUK")].get("confidence_score") == 0.95





def test_muat_pola_file_tidak_ada_kembalikan_kosong(tmp_path):

    path = str(tmp_path / "tidak_ada.json")

    pola = ak.muat_pola(path)

    assert pola.aturan == {}





def test_muat_pola_korup_pulih_dari_backup(tmp_path, capsys):

    path = str(tmp_path / "pola.json")

    pola_baik = ak.Pola(aturan={("SIG", "MASUK"): {

        "no_akun_debet": "1100", "nama_akun_debet": "BANK",

        "no_akun_kredit": "4100", "nama_akun_kredit": "PENJUALAN",

        "konsisten": True, "jumlah_contoh": 1,

    }})

    ak.simpan_pola(pola_baik, path)

    ak.simpan_pola(pola_baik, path)



    with open(path, "w", encoding="utf-8") as f:

        f.write("{ bukan json valid !!!")



    pola_pulih = ak.muat_pola(path)

    assert ("SIG", "MASUK") in pola_pulih.aturan

    tangkapan = capsys.readouterr()

    assert "PERINGATAN" in tangkapan.out





def test_muat_pola_korup_tanpa_backup_kembalikan_kosong(tmp_path, capsys):

    path = str(tmp_path / "pola.json")

    with open(path, "w", encoding="utf-8") as f:

        f.write("bukan json sama sekali")

    pola = ak.muat_pola(path)

    assert pola.aturan == {}

    tangkapan = capsys.readouterr()

    assert "PERINGATAN" in tangkapan.out

    assert (tmp_path / "pola.json").exists()





# ============================================================

# perbaiki_data_penilaian -- auto-fix + log koreksi

# ============================================================



def test_perbaiki_data_penilaian_hitung_total_score_kosong():

    df = pd.DataFrame([{

        "no": 1, "nama_klien": "Klien A", "maker": "Maker A",

        "score": 80, "bobot_klien": 0.5,

        "total_score": None, "plus": None, "minus": None, "total_akhir": None,

        "jenis_baris": "klien",

    }])

    df_baru, log = ak.perbaiki_data_penilaian(df)

    assert df_baru.at[0, "total_score"] == 40.0

    assert df_baru.at[0, "total_akhir"] == 40.0

    assert len(log) >= 2

    kolom_yang_diubah = {l["kolom"] for l in log}

    assert "total_score" in kolom_yang_diubah

    assert "total_akhir" in kolom_yang_diubah





def test_perbaiki_data_penilaian_koreksi_total_score_salah_rumus():

    df = pd.DataFrame([{

        "no": 1, "nama_klien": "Klien B", "maker": "Maker B",

        "score": 80, "bobot_klien": 0.5,

        "total_score": 999,

        "plus": 0, "minus": 0, "total_akhir": None,

        "jenis_baris": "klien",

    }])

    df_baru, log = ak.perbaiki_data_penilaian(df)

    assert df_baru.at[0, "total_score"] == 40.0

    alasan = [l["alasan"] for l in log if l["kolom"] == "total_score"][0]

    assert "dikoreksi" in alasan.lower()





def test_perbaiki_data_penilaian_tidak_menyentuh_baris_catatan():

    df = pd.DataFrame([{

        "no": None, "nama_klien": None, "maker": None,

        "score": None, "bobot_klien": None,

        "total_score": None, "plus": None, "minus": None, "total_akhir": None,

        "jenis_baris": "catatan/ringkasan",

    }])

    df_baru, log = ak.perbaiki_data_penilaian(df)

    assert len(log) == 0





# ============================================================

# _LembarDariBaris / muat_workbook -- dukungan multi-format

# ============================================================



def test_lembar_dari_baris_iter_rows_seperti_openpyxl():

    adapter = ak._LembarDariBaris([("a", "b"), ("c", "d"), ("e", "f")])

    assert list(adapter.iter_rows()) == [("a", "b"), ("c", "d"), ("e", "f")]

    assert list(adapter.iter_rows(min_row=2)) == [("c", "d"), ("e", "f")]

    assert list(adapter.iter_rows(min_row=1, max_row=2)) == [("a", "b"), ("c", "d")]





def test_baca_csv_deteksi_delimiter_titik_koma():

    import io

    csv_text = "Tanggal;Keterangan;Saldo\n01/01/2026;PLNPOST001;1000000\n"

    baris = ak._baca_csv_sebagai_baris(io.BytesIO(csv_text.encode("utf-8")))

    assert baris[0] == ("Tanggal", "Keterangan", "Saldo")

    assert baris[1][1] == "PLNPOST001"





def test_muat_workbook_ekstensi_tidak_didukung_ditolak():

    import io

    f = io.BytesIO(b"apa saja")

    f.name = "berkas.docx"

    with pytest.raises(ak.FormatFileTidakDidukung):

        ak.muat_workbook(f)





# ============================================================

# deteksi_jenis_dokumen_lain -- deteksi jenis dokumen di luar 5 jenis

# yang sudah ada parser jurnalnya (rekening koran/penjualan/POS/

# penilaian/piutang). Sengaja dites TANPA API key / internet, murni

# pencocokan kata kunci terhadap header standar dokumen Indonesia.

# ============================================================



def _ws_dari_header(headers, jumlah_baris_contoh=1):

    baris = [tuple(headers)] + [tuple(headers) for _ in range(jumlah_baris_contoh)]

    return ak._LembarDariBaris(baris)





def test_deteksi_slip_gaji():

    ws = _ws_dari_header([

        "NIP", "Nama Karyawan", "Jabatan", "Gaji Pokok",

        "Tunjangan", "Potongan", "PPh 21", "Take Home Pay",

    ])

    kode, label, skor = ak.deteksi_jenis_dokumen_lain(ws)

    assert kode == "slip_gaji"

    assert "Slip Gaji" in label

    assert skor > 0





def test_deteksi_kartu_stok():

    ws = _ws_dari_header([

        "Kode Barang", "Nama Barang", "Satuan",

        "Stok Awal", "Barang Masuk", "Barang Keluar", "Stok Akhir",

    ])

    kode, label, _skor = ak.deteksi_jenis_dokumen_lain(ws)

    assert kode == "kartu_stok"





def test_deteksi_faktur_pajak():

    ws = _ws_dari_header([

        "Nomor Faktur Pajak", "NPWP Penjual", "NPWP Pembeli",

        "Kode Transaksi", "DPP", "PPN",

    ])

    kode, label, _skor = ak.deteksi_jenis_dokumen_lain(ws)

    assert kode == "faktur_pajak"





def test_deteksi_bukti_potong_pajak():

    ws = _ws_dari_header([

        "Nomor Bukti Potong", "NPWP Pemotong", "Kode Objek Pajak",

        "Penghasilan Kotor", "Tarif", "PPh Dipotong", "Tanggal Pemotongan",

    ])

    kode, label, _skor = ak.deteksi_jenis_dokumen_lain(ws)

    assert kode == "bukti_potong_pajak"





def test_deteksi_aset_tetap():

    ws = _ws_dari_header([

        "Nama Aset", "Tanggal Perolehan", "Harga Perolehan",

        "Masa Manfaat", "Nilai Residu", "Akumulasi Penyusutan", "Nilai Buku",

    ])

    kode, label, _skor = ak.deteksi_jenis_dokumen_lain(ws)

    assert kode == "aset_tetap"





def test_deteksi_purchase_order():

    ws = _ws_dari_header([

        "Nomor PO", "Tanggal PO", "Vendor", "Delivery Date", "Syarat Pembayaran",

    ])

    kode, label, _skor = ak.deteksi_jenis_dokumen_lain(ws)

    assert kode == "purchase_order"





def test_deteksi_tidak_ada_yang_cocok():

    ws = _ws_dari_header(["Kolom A", "Kolom B", "Kolom C"])

    kode, label, skor = ak.deteksi_jenis_dokumen_lain(ws)

    assert kode is None

    assert label is None

    assert skor == 0.0





def test_deteksi_semua_sheet_gabungan_dikenal_dan_tidak_dikenal():

    """muat_workbook mengenali sheet rekening koran; sheet slip gaji di file

    yang sama harus tetap dilewati TAPI dilaporkan lewat 'peringatan' dengan

    tebakan jenisnya, bukan sekadar 'tidak dikenali'."""

    import io

    import openpyxl



    wb = openpyxl.Workbook()

    ws_bank = wb.active

    ws_bank.title = "Bank BCA"

    ws_bank.append(["Tanggal", "Keterangan", "Debit", "Kredit", "Saldo"])

    ws_bank.append(["01/01/2026", "Transfer masuk", 0, 500000, 500000])



    ws_gaji = wb.create_sheet("Payroll Januari")

    ws_gaji.append([

        "NIP", "Nama Karyawan", "Jabatan", "Gaji Pokok",

        "Tunjangan", "Potongan", "PPh 21", "Take Home Pay",

    ])

    ws_gaji.append(["001", "Budi", "Staff", 5000000, 500000, 200000, 100000, 5200000])



    buf = io.BytesIO()

    wb.save(buf)

    buf.seek(0)

    buf.name = "contoh.xlsx"



    df_bank, df_jual, df_penilaian, df_piutang, df_coa, peringatan = ak.muat_workbook(buf)



    assert not df_bank.empty

    assert any("Slip Gaji" in p for p in peringatan)





def test_muat_workbook_csv_rekening_koran_end_to_end():

    import io

    csv_text = (

        "Tanggal;Keterangan;Mutasi Debet;Mutasi Kredit;Saldo\n"

        "01/01/2026;PLNPOST001;500000;0;1000000\n"

        "02/01/2026;PLNPOST002;520000;0;480000\n"

    )

    f = io.BytesIO(csv_text.encode("utf-8"))

    f.name = "mutasi_bank_januari.csv"

    df_bank, df_jual, df_penilaian, df_piutang, df_coa, peringatan = ak.muat_workbook(f)

    assert not df_bank.empty

    assert len(df_bank) == 2





# ============================================================

# analisis_kesalahan_penilaian

# ============================================================



def test_analisis_kritis_score_rendah():

    df = pd.DataFrame([{"nama_klien": "Test", "score": 45, "bobot_klien": 1.0, "jenis_baris": "klien"}])

    analisis = ak.analisis_kesalahan_penilaian(df)

    assert analisis["total_temuan"] > 0

    assert any("rendah" in t["temuan"].lower() for t in analisis["temuan"])





# ============================================================

# merge_dan_augmentasi_pola

# ============================================================



def test_merge_dan_augmentasi_pola():

    pola_lama = ak.Pola()

    pola_lama.aturan[("PLN", "KELUAR")] = {

        "no_akun_debet": "5100",

        "nama_akun_debet": "BEBAN LISTRIK",

        "no_akun_kredit": "1100",

        "nama_akun_kredit": "BANK",

        "konsisten": True,

        "jumlah_contoh": 3,

        "confidence_score": 0.9,

        "is_valid": True

    }

    

    pola_baru = ak.Pola()

    pola_baru.aturan[("TELKOM", "KELUAR")] = {

        "no_akun_debet": "5200",

        "nama_akun_debet": "BEBAN TELEPON",

        "no_akun_kredit": "1100",

        "nama_akun_kredit": "BANK",

        "konsisten": True,

        "jumlah_contoh": 2,

        "confidence_score": 0.8,

        "is_valid": True

    }

    

    merged = ak.merge_dan_augmentasi_pola(pola_lama, pola_baru)

    assert ("PLN", "KELUAR") in merged.aturan

    assert ("TELKOM", "KELUAR") in merged.aturan

    assert len(merged.aturan) >= 3





# ============================================================

# TES FAKTUR PAJAK (PPN)

# ============================================================



def test_validasi_npwp_format_lama_15_digit():

    hasil = ak.validasi_npwp("01.234.567.8-901.000")

    assert hasil["valid"] is True

    assert hasil["jumlah_digit"] == 15





def test_validasi_npwp_format_baru_16_digit():

    hasil = ak.validasi_npwp("0123456789012345")

    assert hasil["valid"] is True

    assert hasil["jumlah_digit"] == 16





def test_validasi_npwp_jumlah_digit_tidak_lazim():

    hasil = ak.validasi_npwp("12345")

    assert hasil["valid"] is False

    assert hasil["jumlah_digit"] == 5





def test_validasi_npwp_kosong():

    hasil = ak.validasi_npwp(None)

    assert hasil["valid"] is False

    assert hasil["jumlah_digit"] == 0





def test_validasi_nomor_faktur_pajak_valid():

    hasil = ak.validasi_nomor_faktur_pajak("010.001-25.00000001")

    assert hasil["valid"] is True

    assert hasil["kode_transaksi"] == "01"





def test_validasi_nomor_faktur_pajak_panjang_salah():

    hasil = ak.validasi_nomor_faktur_pajak("123456")

    assert hasil["valid"] is False





def test_validasi_nomor_faktur_pajak_kode_transaksi_tidak_dikenal():

    hasil = ak.validasi_nomor_faktur_pajak("990001250000001")  # 15 digit, kode 99

    assert hasil["valid"] is False





def _df_faktur_pajak_contoh():

    return pd.DataFrame([

        {

            "sheet": "Faktur", "tanggal": "2026-01-05",

            "nomor_faktur": "010.001-25.00000001", "kode_transaksi": "01",

            "npwp_penjual": "01.234.567.8-901.000", "nama_penjual": "PT Perusahaanku",

            "npwp_pembeli": "02.345.678.9-012.000", "nama_pembeli": "PT Pembeli A",

            "dpp": 10_000_000, "ppn": 1_100_000, "keterangan": "Jasa konsultasi",

        },

        {

            # PPN tidak sesuai hitungan (harusnya 550rb @11%, tercatat 600rb)

            "sheet": "Faktur", "tanggal": "2026-01-06",

            "nomor_faktur": "010.001-25.00000002", "kode_transaksi": "01",

            "npwp_penjual": "01.234.567.8-901.000", "nama_penjual": "PT Perusahaanku",

            "npwp_pembeli": "03.456.789.0-123.000", "nama_pembeli": "PT Pembeli B",

            "dpp": 5_000_000, "ppn": 600_000, "keterangan": "Barang X",

        },

        {

            # Duplikat nomor faktur dari baris pertama

            "sheet": "Faktur", "tanggal": "2026-01-07",

            "nomor_faktur": "010.001-25.00000001", "kode_transaksi": "01",

            "npwp_penjual": "01.234.567.8-901.000", "nama_penjual": "PT Perusahaanku",

            "npwp_pembeli": "04.567.890.1-234.000", "nama_pembeli": "PT Pembeli C",

            "dpp": 2_000_000, "ppn": 220_000, "keterangan": "Barang Y",

        },

        {

            # Perusahaan sbg pembeli -> PPN Masukan

            "sheet": "Faktur", "tanggal": "2026-01-08",

            "nomor_faktur": "010.001-25.00000003", "kode_transaksi": "01",

            "npwp_penjual": "05.678.901.2-345.000", "nama_penjual": "PT Supplier",

            "npwp_pembeli": "01.234.567.8-901.000", "nama_pembeli": "PT Perusahaanku",

            "dpp": 3_000_000, "ppn": 330_000, "keterangan": "Bahan baku",

        },

    ])





def test_proses_faktur_pajak_deteksi_ppn_tidak_sesuai():

    df = _df_faktur_pajak_contoh()

    hasil = ak.proses_faktur_pajak(df)

    alasan_gabungan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert any("tidak sesuai perhitungan" in a for a in alasan_gabungan)





def test_proses_faktur_pajak_deteksi_duplikat():

    df = _df_faktur_pajak_contoh()

    hasil = ak.proses_faktur_pajak(df)

    alasan_gabungan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert any("DUPLIKAT" in a for a in alasan_gabungan)

    assert hasil["ringkasan"]["jumlah_duplikat"] == 1





def test_proses_faktur_pajak_tanpa_npwp_tidak_ada_draf_jurnal():

    df = _df_faktur_pajak_contoh()

    hasil = ak.proses_faktur_pajak(df)  # npwp_perusahaan=None

    assert hasil["draf_jurnal"] == []

    assert "catatan_arah" in hasil["ringkasan"]





def test_proses_faktur_pajak_dengan_npwp_hitung_arah_dan_jurnal():

    df = _df_faktur_pajak_contoh()

    hasil = ak.proses_faktur_pajak(df, npwp_perusahaan="01.234.567.8-901.000")

    assert hasil["ringkasan"]["ppn_keluaran"] == pytest.approx(1_100_000 + 600_000 + 220_000)

    assert hasil["ringkasan"]["ppn_masukan"] == pytest.approx(330_000)

    assert len(hasil["draf_jurnal"]) == 4

    arah_set = {d["arah"] for d in hasil["draf_jurnal"]}

    assert arah_set == {"KELUARAN", "MASUKAN"}





def test_proses_faktur_pajak_dataframe_kosong():

    hasil = ak.proses_faktur_pajak(pd.DataFrame())

    assert hasil["masalah"] == []

    assert hasil["draf_jurnal"] == []





def test_parse_sheet_faktur_pajak_kolom_tidak_lengkap_gagal():

    wb = __import__("openpyxl").Workbook()

    ws = wb.active

    ws.append(["Tanggal", "Keterangan", "Jumlah"])  # bukan format faktur pajak

    ws.append(["2026-01-01", "Test", 1000])

    with pytest.raises(ak.FormatTidakDikenali):

        ak.parse_sheet_faktur_pajak(ws, "Sheet1")





# ============================================================

# TES SLIP GAJI KARYAWAN

# ============================================================



def test_cek_komponen_gaji_sesuai():

    hasil = ak.cek_komponen_gaji(

        gaji_pokok=8_000_000, total_tunjangan=2_000_000,

        total_potongan=1_500_000, gaji_bersih=8_500_000,

    )

    assert hasil["sesuai"] is True

    assert hasil["gaji_bruto"] == 10_000_000





def test_cek_komponen_gaji_tidak_sesuai():

    hasil = ak.cek_komponen_gaji(

        gaji_pokok=8_000_000, total_tunjangan=2_000_000,

        total_potongan=1_500_000, gaji_bersih=9_000_000,  # salah, harusnya 8.5jt

    )

    assert hasil["sesuai"] is False

    assert hasil["selisih"] == pytest.approx(500_000)





def test_cek_kewajaran_pph21_slip_wajar():

    # 10% dari gaji bruto -- masih dalam rentang wajar TER

    assert ak.cek_kewajaran_pph21_slip(10_000_000, 1_000_000) is None





def test_cek_kewajaran_pph21_slip_tidak_wajar():

    # 50% dari gaji bruto -- di luar rentang wajar (maks 34%)

    pesan = ak.cek_kewajaran_pph21_slip(10_000_000, 5_000_000)

    assert pesan is not None

    assert "DI LUAR rentang wajar" in pesan





def test_cek_kewajaran_bpjs_wajar():

    gaji_pokok = 10_000_000

    catatan = ak.cek_kewajaran_bpjs(

        gaji_pokok,

        bpjs_kesehatan=gaji_pokok * ak.BPJS_KESEHATAN_TARIF_KARYAWAN,

        bpjs_jht=gaji_pokok * ak.BPJS_JHT_TARIF_KARYAWAN,

        bpjs_jp=gaji_pokok * ak.BPJS_JP_TARIF_KARYAWAN,

    )

    assert catatan == []





def test_cek_kewajaran_bpjs_tidak_wajar():

    catatan = ak.cek_kewajaran_bpjs(

        10_000_000, bpjs_kesehatan=900_000, bpjs_jht=0, bpjs_jp=0,  # jauh dari 1% wajar

    )

    assert any("BPJS Kesehatan" in c for c in catatan)





def _df_slip_gaji_contoh():

    return pd.DataFrame([

        {

            "sheet": "Payroll", "nip": "001", "nama_karyawan": "Budi Santoso",

            "jabatan": "Staff", "departemen": "Finance", "periode_gaji": "Juni 2026",

            "npwp": "01.234.567.8-901.000",

            "gaji_pokok": 8_000_000, "total_tunjangan": 1_000_000,

            "bpjs_kesehatan": 80_000, "bpjs_jht": 160_000, "bpjs_jp": 80_000,

            "pph21": 300_000, "potongan_lain": 0,

            "total_potongan": 620_000, "gaji_bersih": 8_380_000,

        },

        {

            # Komponen tidak konsisten -- gaji bersih salah hitung

            "sheet": "Payroll", "nip": "002", "nama_karyawan": "Siti Aminah",

            "jabatan": "Supervisor", "departemen": "Operasional", "periode_gaji": "Juni 2026",

            "npwp": "02.345.678.9-012.000",

            "gaji_pokok": 10_000_000, "total_tunjangan": 1_500_000,

            "bpjs_kesehatan": 100_000, "bpjs_jht": 200_000, "bpjs_jp": 100_000,

            "pph21": 500_000, "potongan_lain": 0,

            "total_potongan": 900_000, "gaji_bersih": 12_000_000,  # seharusnya 10.6jt

        },

        {

            # Duplikat dari baris pertama (NIP & periode sama)

            "sheet": "Payroll", "nip": "001", "nama_karyawan": "Budi Santoso",

            "jabatan": "Staff", "departemen": "Finance", "periode_gaji": "Juni 2026",

            "npwp": "01.234.567.8-901.000",

            "gaji_pokok": 8_000_000, "total_tunjangan": 1_000_000,

            "bpjs_kesehatan": 80_000, "bpjs_jht": 160_000, "bpjs_jp": 80_000,

            "pph21": 300_000, "potongan_lain": 0,

            "total_potongan": 620_000, "gaji_bersih": 8_380_000,

        },

    ])





def test_proses_slip_gaji_deteksi_komponen_tidak_sesuai():

    df = _df_slip_gaji_contoh()

    hasil = ak.proses_slip_gaji(df)

    alasan_gabungan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert any("TIDAK SAMA dengan hasil hitung" in a for a in alasan_gabungan)





def test_proses_slip_gaji_deteksi_duplikat():

    df = _df_slip_gaji_contoh()

    hasil = ak.proses_slip_gaji(df)

    alasan_gabungan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert any("DUPLIKAT" in a for a in alasan_gabungan)

    assert hasil["ringkasan"]["jumlah_duplikat"] == 1





def test_proses_slip_gaji_ringkasan_total():

    df = _df_slip_gaji_contoh()

    hasil = ak.proses_slip_gaji(df)

    assert hasil["ringkasan"]["jumlah_slip"] == 3

    assert hasil["ringkasan"]["total_gaji_bersih_dibayarkan"] == pytest.approx(8_380_000 + 12_000_000 + 8_380_000)





def test_proses_slip_gaji_draf_jurnal_dibuat_per_baris():

    df = _df_slip_gaji_contoh()

    hasil = ak.proses_slip_gaji(df)

    assert len(hasil["draf_jurnal"]) == 3

    assert hasil["draf_jurnal"][0]["jml_debet"] == 9_000_000  # gaji pokok + tunjangan





def test_proses_slip_gaji_dataframe_kosong():

    hasil = ak.proses_slip_gaji(pd.DataFrame())

    assert hasil["masalah"] == []

    assert hasil["draf_jurnal"] == []





def test_parse_sheet_slip_gaji_kolom_tidak_lengkap_gagal():

    wb = __import__("openpyxl").Workbook()

    ws = wb.active

    ws.append(["Tanggal", "Keterangan", "Jumlah"])  # bukan format slip gaji

    ws.append(["2026-01-01", "Test", 1000])

    with pytest.raises(ak.FormatTidakDikenali):

        ak.parse_sheet_slip_gaji(ws, "Sheet1")





def test_parse_sheet_slip_gaji_end_to_end():

    wb = __import__("openpyxl").Workbook()

    ws = wb.active

    ws.append([

        "NIP", "Nama Karyawan", "Jabatan", "Departemen", "Periode Gaji", "NPWP",

        "Gaji Pokok", "Tunjangan Transport", "Tunjangan Makan",

        "BPJS Kesehatan", "JHT", "Jaminan Pensiun", "PPh 21", "Gaji Bersih",

    ])

    ws.append([

        "001", "Budi Santoso", "Staff", "Finance", "Juni 2026", "01.234.567.8-901.000",

        8_000_000, 500_000, 500_000,

        80_000, 160_000, 80_000, 300_000, 8_380_000,

    ])

    df = ak.parse_sheet_slip_gaji(ws, "Payroll")

    assert len(df) == 1

    assert df.iloc[0]["gaji_pokok"] == 8_000_000

    assert df.iloc[0]["total_tunjangan"] == 1_000_000

    hasil = ak.proses_slip_gaji(df)

    assert hasil["masalah"] == []  # semua komponen konsisten & wajar





# ============================================================

# TES BUKTI KAS MASUK/KELUAR

# ============================================================



def _df_bukti_kas_contoh():

    return pd.DataFrame([

        {"sheet": "Kas", "tanggal": "2026-06-01", "nomor_bukti_kas": "BKM-001",

         "jenis_tersurat": "Masuk", "keterangan": "Setoran modal tunai",

         "pihak_terkait": "Pemilik", "jml_masuk": 5_000_000, "jml_keluar": 0,

         "saldo": None, "kategori": "3-1000", "penanggung_jawab": "Kasir A"},

        {"sheet": "Kas", "tanggal": "2026-06-02", "nomor_bukti_kas": "BKK-001",

         "jenis_tersurat": "Keluar", "keterangan": "Beli ATK",

         "pihak_terkait": "Toko ABC", "jml_masuk": 0, "jml_keluar": 250_000,

         "saldo": None, "kategori": "6-1000", "penanggung_jawab": "Kasir A"},

    ])





def test_proses_bukti_kas_nomor_beda_prefiks_tidak_dianggap_duplikat():

    # BKM-001 dan BKK-001 punya "001" yang sama tapi PREFIKS beda (arah beda)

    # -- ini bukan duplikat, harus lolos tanpa flag DUPLIKAT.

    df = _df_bukti_kas_contoh()

    hasil = ak.proses_bukti_kas(df)

    alasan_gabungan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert not any("DUPLIKAT" in a for a in alasan_gabungan)





def test_proses_bukti_kas_deteksi_duplikat_nomor_sama_persis():

    df = _df_bukti_kas_contoh()

    df = pd.concat([df, df.iloc[[0]]], ignore_index=True)  # duplikasi baris pertama

    hasil = ak.proses_bukti_kas(df)

    alasan_gabungan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert any("DUPLIKAT" in a for a in alasan_gabungan)





def test_proses_bukti_kas_arah_terdeteksi_benar():

    df = _df_bukti_kas_contoh()

    hasil = ak.proses_bukti_kas(df)

    assert list(hasil["df"]["arah"]) == ["MASUK", "KELUAR"]





def test_proses_bukti_kas_nominal_ekstrim_terdeteksi():

    df = _df_bukti_kas_contoh()

    df.loc[len(df)] = {

        "sheet": "Kas", "tanggal": "2026-06-03", "nomor_bukti_kas": "BKM-002",

        "jenis_tersurat": "Masuk", "keterangan": "Setoran besar",

        "pihak_terkait": "Investor", "jml_masuk": 600_000_000, "jml_keluar": 0,

        "saldo": None, "kategori": "3-1000", "penanggung_jawab": "Kasir A",

    }

    hasil = ak.proses_bukti_kas(df)

    alasan_gabungan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert any("di atas ambang wajar" in a for a in alasan_gabungan)





def test_proses_bukti_kas_keluar_besar_tanpa_pihak_terkait():

    df = pd.DataFrame([{

        "sheet": "Kas", "tanggal": "2026-06-01", "nomor_bukti_kas": "BKK-010",

        "jenis_tersurat": "Keluar", "keterangan": "Pembayaran",

        "pihak_terkait": None, "jml_masuk": 0, "jml_keluar": 2_000_000,

        "saldo": None, "kategori": None, "penanggung_jawab": None,

    }])

    hasil = ak.proses_bukti_kas(df)

    alasan_gabungan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert any("Dibayarkan Kepada" in a for a in alasan_gabungan)





def test_proses_bukti_kas_saldo_berjalan_konsisten():

    df = pd.DataFrame([

        {"sheet": "Kas", "tanggal": "2026-06-01", "nomor_bukti_kas": "BKM-001",

         "jenis_tersurat": "Masuk", "keterangan": "Setoran", "pihak_terkait": "A",

         "jml_masuk": 1_000_000, "jml_keluar": 0, "saldo": 1_000_000,

         "kategori": None, "penanggung_jawab": None},

        {"sheet": "Kas", "tanggal": "2026-06-02", "nomor_bukti_kas": "BKK-001",

         "jenis_tersurat": "Keluar", "keterangan": "Beli ATK", "pihak_terkait": "B",

         "jml_masuk": 0, "jml_keluar": 200_000, "saldo": 800_000,

         "kategori": None, "penanggung_jawab": None},

    ])

    hasil = ak.proses_bukti_kas(df)

    alasan_gabungan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert not any("Saldo berjalan" in a for a in alasan_gabungan)





def test_proses_bukti_kas_saldo_berjalan_tidak_konsisten_dilaporkan_semua():

    # Dua baris SEKALIGUS salah saldonya -- pastikan KEDUANYA dilaporkan

    # (bukan cuma yang pertama seperti bug di kode versi lama).

    df = pd.DataFrame([

        {"sheet": "Kas", "tanggal": "2026-06-01", "nomor_bukti_kas": "BKM-001",

         "jenis_tersurat": "Masuk", "keterangan": "Setoran", "pihak_terkait": "A",

         "jml_masuk": 1_000_000, "jml_keluar": 0, "saldo": 1_000_000,

         "kategori": None, "penanggung_jawab": None},

        {"sheet": "Kas", "tanggal": "2026-06-02", "nomor_bukti_kas": "BKK-001",

         "jenis_tersurat": "Keluar", "keterangan": "Beli ATK", "pihak_terkait": "B",

         "jml_masuk": 0, "jml_keluar": 200_000, "saldo": 999_000,  # salah, harusnya 800rb

         "kategori": None, "penanggung_jawab": None},

        {"sheet": "Kas", "tanggal": "2026-06-03", "nomor_bukti_kas": "BKK-002",

         "jenis_tersurat": "Keluar", "keterangan": "Beli galon", "pihak_terkait": "C",

         "jml_masuk": 0, "jml_keluar": 50_000, "saldo": 700_000,  # salah, harusnya 949rb dari saldo tertulis sblmnya

         "kategori": None, "penanggung_jawab": None},

    ])

    hasil = ak.proses_bukti_kas(df)

    baris_bermasalah_saldo = [m["baris"] for m in hasil["masalah"] if any("Saldo berjalan" in a for a in m["alasan"])]

    assert set(baris_bermasalah_saldo) == {2, 3}





def test_proses_bukti_kas_cross_check_coa():

    df = _df_bukti_kas_contoh()

    df_coa = pd.DataFrame([

        {"no_akun": "3-1000", "nama_akun": "Modal", "kategori": "Ekuitas"},

        # sengaja TIDAK sertakan "6-1000" supaya baris ke-2 diflag

    ])

    hasil = ak.proses_bukti_kas(df, df_coa=df_coa)

    alasan_gabungan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert any("tidak ditemukan di daftar akun" in a for a in alasan_gabungan)





def test_proses_bukti_kas_ringkasan_total():

    df = _df_bukti_kas_contoh()

    hasil = ak.proses_bukti_kas(df)

    assert hasil["ringkasan"]["total_kas_masuk"] == 5_000_000

    assert hasil["ringkasan"]["total_kas_keluar"] == 250_000

    assert hasil["ringkasan"]["saldo_bersih_periode"] == 4_750_000





def test_proses_bukti_kas_draf_jurnal_arah_benar():

    df = _df_bukti_kas_contoh()

    hasil = ak.proses_bukti_kas(df)

    assert hasil["draf_jurnal"][0]["no_akun_debet"] == "KAS"       # bukti masuk: Kas di debet

    assert hasil["draf_jurnal"][1]["no_akun_kredit"] == "KAS"      # bukti keluar: Kas di kredit





def test_proses_bukti_kas_dataframe_kosong():

    hasil = ak.proses_bukti_kas(pd.DataFrame())

    assert hasil["masalah"] == []

    assert hasil["draf_jurnal"] == []





def test_parse_sheet_bukti_kas_kolom_tidak_lengkap_gagal():

    wb = __import__("openpyxl").Workbook()

    ws = wb.active

    ws.append(["Nama Karyawan", "Gaji Pokok"])  # bukan format bukti kas

    ws.append(["Budi", 5_000_000])

    with pytest.raises(ak.FormatTidakDikenali):

        ak.parse_sheet_bukti_kas(ws, "Sheet1")





def test_parse_sheet_bukti_kas_end_to_end():

    wb = __import__("openpyxl").Workbook()

    ws = wb.active

    ws.append(["Nomor Bukti Kas", "Tanggal", "Jenis", "Keterangan", "Diterima Dari", "Jumlah"])

    ws.append(["BKM-001", "2026-06-01", "Masuk", "Setoran modal", "Pemilik", 5_000_000])

    df = ak.parse_sheet_bukti_kas(ws, "Kas")

    assert len(df) == 1

    assert df.iloc[0]["jml_masuk"] == 5_000_000

    hasil = ak.proses_bukti_kas(df)

    assert hasil["df"].iloc[0]["arah"] == "MASUK"





# ============================================================

# [BARU] TES REGRESI -- BUG DUPLIKAT LINTAS SHEET & KOLOM MASUK/KELUAR

# AMBIGU (lihat catatan [FIX] di proses_bukti_kas())

# ============================================================



def test_proses_bukti_kas_nomor_sama_beda_sheet_tidak_dianggap_duplikat():

    # Dua kas kecil yang BERBEDA (mis. "Kas Kantor" vs "Kas Toko") sama-sama

    # wajar mulai penomoran dari BKM-001 -- ini BUKAN duplikat krn beda sheet.

    df = pd.DataFrame([

        {"sheet": "Kas Kantor", "tanggal": "2026-06-01", "nomor_bukti_kas": "BKM-001",

         "jenis_tersurat": "Masuk", "keterangan": "Setoran modal", "pihak_terkait": "Pemilik",

         "jml_masuk": 1_000_000, "jml_keluar": 0, "saldo": None, "kategori": None, "penanggung_jawab": None},

        {"sheet": "Kas Toko", "tanggal": "2026-06-01", "nomor_bukti_kas": "BKM-001",

         "jenis_tersurat": "Masuk", "keterangan": "Setoran modal toko", "pihak_terkait": "Pemilik",

         "jml_masuk": 500_000, "jml_keluar": 0, "saldo": None, "kategori": None, "penanggung_jawab": None},

    ])

    hasil = ak.proses_bukti_kas(df)

    alasan_gabungan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert not any("DUPLIKAT" in a for a in alasan_gabungan)





def test_proses_bukti_kas_nomor_sama_dalam_sheet_sama_tetap_dianggap_duplikat():

    # Kebalikan dari test di atas -- kalau dalam SATU sheet yang sama, nomor

    # sama harus tetap kedeteksi duplikat (regresi bug lama tidak boleh terjadi).

    df = pd.DataFrame([

        {"sheet": "Kas Kantor", "tanggal": "2026-06-01", "nomor_bukti_kas": "BKM-001",

         "jenis_tersurat": "Masuk", "keterangan": "Setoran modal", "pihak_terkait": "Pemilik",

         "jml_masuk": 1_000_000, "jml_keluar": 0, "saldo": None, "kategori": None, "penanggung_jawab": None},

        {"sheet": "Kas Kantor", "tanggal": "2026-06-02", "nomor_bukti_kas": "BKM-001",

         "jenis_tersurat": "Masuk", "keterangan": "Setoran modal lagi", "pihak_terkait": "Pemilik",

         "jml_masuk": 500_000, "jml_keluar": 0, "saldo": None, "kategori": None, "penanggung_jawab": None},

    ])

    hasil = ak.proses_bukti_kas(df)

    alasan_gabungan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert any("DUPLIKAT" in a for a in alasan_gabungan)





def test_proses_bukti_kas_masuk_dan_keluar_terisi_bersamaan_ditandai_ambigu():

    df = pd.DataFrame([{

        "sheet": "Kas", "tanggal": "2026-06-01", "nomor_bukti_kas": "BKX-001",

        "jenis_tersurat": None, "keterangan": "Salah input kolom", "pihak_terkait": "A",

        "jml_masuk": 300_000, "jml_keluar": 150_000, "saldo": None, "kategori": None, "penanggung_jawab": None,

    }])

    hasil = ak.proses_bukti_kas(df)

    assert hasil["df"].iloc[0]["arah"] == "AMBIGU (MASUK & KELUAR TERISI)"

    alasan_gabungan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert any("SAMA-SAMA terisi" in a for a in alasan_gabungan)

    # Baris ambigu TIDAK BOLEH menghasilkan draf jurnal otomatis

    assert hasil["draf_jurnal"] == []

    assert hasil["ringkasan"]["jumlah_ambigu_masuk_keluar"] == 1





def test_proses_bukti_kas_rekap_per_pihak_terkait():

    df = _df_bukti_kas_contoh()

    hasil = ak.proses_bukti_kas(df)

    rekap = hasil["ringkasan"]["rekap_per_pihak_terkait"]

    assert rekap["Pemilik"]["jml_masuk"] == 5_000_000

    assert rekap["Toko ABC"]["jml_keluar"] == 250_000





# ============================================================

# [BARU] TES PEMBELIAN (PO/INVOICE) -- termasuk regresi bug duplikat

# lintas supplier (lihat catatan [FIX] di proses_pembelian())

# ============================================================



def _df_pembelian_contoh():

    return pd.DataFrame([

        {"nomor_dokumen": "INV-001", "jenis_dokumen": "INVOICE", "nama_supplier": "Supplier A",

         "qty": 10, "harga_satuan": 50_000, "subtotal_tertulis": 500_000,

         "ppn_tertulis": 55_000, "total_tertulis": 555_000},

        {"nomor_dokumen": "INV-001", "jenis_dokumen": "INVOICE", "nama_supplier": "Supplier B",

         "qty": 5, "harga_satuan": 20_000, "subtotal_tertulis": 100_000,

         "ppn_tertulis": 11_000, "total_tertulis": 111_000},

    ])





def test_proses_pembelian_nomor_sama_beda_supplier_tidak_dianggap_duplikat():

    # Supplier A dan Supplier B menerbitkan invoice masing-masing secara

    # independen -- kebetulan sama-sama "INV-001" BUKAN duplikat.

    df = _df_pembelian_contoh()

    hasil = ak.proses_pembelian(df)

    alasan_gabungan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert not any("DUPLIKAT" in a for a in alasan_gabungan)





def test_proses_pembelian_nomor_sama_supplier_sama_tetap_dianggap_duplikat():

    df = _df_pembelian_contoh()

    df.loc[1, "nama_supplier"] = "Supplier A"  # sekarang supplier & nomor sama-sama sama

    hasil = ak.proses_pembelian(df)

    alasan_gabungan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert any("DUPLIKAT" in a for a in alasan_gabungan)





def test_proses_pembelian_subtotal_dan_total_dihitung_benar():

    df = _df_pembelian_contoh()

    hasil = ak.proses_pembelian(df)

    alasan_gabungan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert not any("TIDAK SAMA" in a for a in alasan_gabungan)





# ============================================================

# TES KARTU STOK/PERSEDIAAN

# ============================================================



def _df_kartu_stok_dua_barang():

    # Dua kode barang berbeda, saldo qty berjalan KONSISTEN utk keduanya --

    # membuktikan cross-check saldo PER BARANG (bukan global) tidak salah

    # flag walau baris-baris diselang-seling antar barang.

    return pd.DataFrame([

        {"sheet": "Stok", "kode_barang": "BRG-A", "nama_barang": "Barang A", "satuan": "pcs",

         "tanggal": "2026-06-01", "nomor_bukti": "IN-001", "keterangan": "Pembelian",

         "qty_masuk": 100, "harga_masuk": 10_000, "qty_keluar": 0, "harga_keluar": 0,

         "saldo_qty": 100, "saldo_nilai": None},

        {"sheet": "Stok", "kode_barang": "BRG-B", "nama_barang": "Barang B", "satuan": "pcs",

         "tanggal": "2026-06-01", "nomor_bukti": "IN-001", "keterangan": "Pembelian",  # nomor sama, barang beda -> BUKAN duplikat

         "qty_masuk": 50, "harga_masuk": 20_000, "qty_keluar": 0, "harga_keluar": 0,

         "saldo_qty": 50, "saldo_nilai": None},

        {"sheet": "Stok", "kode_barang": "BRG-A", "nama_barang": "Barang A", "satuan": "pcs",

         "tanggal": "2026-06-02", "nomor_bukti": "OUT-001", "keterangan": "Penjualan",

         "qty_masuk": 0, "harga_masuk": 0, "qty_keluar": 30, "harga_keluar": 10_000,

         "saldo_qty": 70, "saldo_nilai": None},

        {"sheet": "Stok", "kode_barang": "BRG-B", "nama_barang": "Barang B", "satuan": "pcs",

         "tanggal": "2026-06-02", "nomor_bukti": "OUT-001", "keterangan": "Penjualan",  # nomor sama, barang beda

         "qty_masuk": 0, "harga_masuk": 0, "qty_keluar": 10, "harga_keluar": 20_000,

         "saldo_qty": 40, "saldo_nilai": None},

    ])





def test_proses_kartu_stok_nomor_bukti_sama_barang_beda_tidak_duplikat():

    df = _df_kartu_stok_dua_barang()

    hasil = ak.proses_kartu_stok(df)

    alasan_gabungan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert not any("DUPLIKAT" in a for a in alasan_gabungan)





def test_proses_kartu_stok_saldo_per_barang_konsisten_walau_diselang_seling():

    df = _df_kartu_stok_dua_barang()

    hasil = ak.proses_kartu_stok(df)

    alasan_gabungan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert not any("Saldo Qty tertulis" in a for a in alasan_gabungan)





def test_proses_kartu_stok_deteksi_stok_minus():

    df = pd.DataFrame([

        {"sheet": "Stok", "kode_barang": "BRG-C", "nama_barang": "Barang C", "satuan": "pcs",

         "tanggal": "2026-06-01", "nomor_bukti": "IN-010", "keterangan": "Pembelian",

         "qty_masuk": 10, "harga_masuk": 5_000, "qty_keluar": 0, "harga_keluar": 0,

         "saldo_qty": 10, "saldo_nilai": None},

        {"sheet": "Stok", "kode_barang": "BRG-C", "nama_barang": "Barang C", "satuan": "pcs",

         "tanggal": "2026-06-02", "nomor_bukti": "OUT-010", "keterangan": "Penjualan berlebih",

         "qty_masuk": 0, "harga_masuk": 0, "qty_keluar": 25, "harga_keluar": 5_000,

         "saldo_qty": -15, "saldo_nilai": None},  # stok minus

    ])

    hasil = ak.proses_kartu_stok(df)

    alasan_gabungan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert any("STOK MINUS" in a for a in alasan_gabungan)

    assert hasil["ringkasan"]["jumlah_stok_minus"] == 1





def test_proses_kartu_stok_saldo_salah_hitung_per_barang_terdeteksi():

    df = _df_kartu_stok_dua_barang()

    df.loc[2, "saldo_qty"] = 999  # rusak saldo baris ke-3 (BRG-A, harusnya 70)

    hasil = ak.proses_kartu_stok(df)

    alasan_gabungan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert any("Saldo Qty tertulis 999" in a for a in alasan_gabungan)





def test_proses_kartu_stok_harga_masuk_kosong_terdeteksi():

    df = pd.DataFrame([{

        "sheet": "Stok", "kode_barang": "BRG-D", "nama_barang": "Barang D", "satuan": "pcs",

        "tanggal": "2026-06-01", "nomor_bukti": "IN-020", "keterangan": "Pembelian",

        "qty_masuk": 20, "harga_masuk": 0, "qty_keluar": 0, "harga_keluar": 0,

        "saldo_qty": None, "saldo_nilai": None,

    }])

    hasil = ak.proses_kartu_stok(df)

    alasan_gabungan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert any("Harga Masuk kosong" in a for a in alasan_gabungan)





def test_proses_kartu_stok_ringkasan_rekap_per_barang():

    df = _df_kartu_stok_dua_barang()

    hasil = ak.proses_kartu_stok(df)

    assert hasil["ringkasan"]["jumlah_barang"] == 2

    assert hasil["ringkasan"]["total_qty_masuk"] == 150

    assert hasil["ringkasan"]["total_qty_keluar"] == 40





def test_proses_kartu_stok_draf_jurnal_arah_benar():

    df = _df_kartu_stok_dua_barang()

    hasil = ak.proses_kartu_stok(df)

    baris_masuk = [d for d in hasil["draf_jurnal"] if d["arah"] == "MASUK"]

    baris_keluar = [d for d in hasil["draf_jurnal"] if d["arah"] == "KELUAR"]

    assert all(d["no_akun_debet"] == "PERSEDIAAN" for d in baris_masuk)

    assert all(d["no_akun_kredit"] == "PERSEDIAAN" for d in baris_keluar)

    assert all(d["no_akun_debet"] == "HPP" for d in baris_keluar)





def test_proses_kartu_stok_dataframe_kosong():

    hasil = ak.proses_kartu_stok(pd.DataFrame())

    assert hasil["masalah"] == []

    assert hasil["draf_jurnal"] == []





def test_parse_sheet_kartu_stok_kolom_tidak_lengkap_gagal():

    wb = __import__("openpyxl").Workbook()

    ws = wb.active

    ws.append(["Nomor Bukti Kas", "Tanggal", "Jumlah"])  # bukan format kartu stok

    ws.append(["BKM-001", "2026-06-01", 5_000_000])

    with pytest.raises(ak.FormatTidakDikenali):

        ak.parse_sheet_kartu_stok(ws, "Sheet1")





def test_parse_sheet_kartu_stok_end_to_end():

    wb = __import__("openpyxl").Workbook()

    ws = wb.active

    ws.append([

        "Kode Barang", "Nama Barang", "Satuan", "Tanggal", "Nomor Bukti", "Keterangan",

        "Qty Masuk", "Harga Masuk", "Qty Keluar", "Harga Keluar", "Saldo Stok",

    ])

    ws.append(["BRG-A", "Barang A", "pcs", "2026-06-01", "IN-001", "Pembelian",

               100, 10_000, 0, 0, 100])

    df = ak.parse_sheet_kartu_stok(ws, "Stok")

    assert len(df) == 1

    assert df.iloc[0]["qty_masuk"] == 100

    hasil = ak.proses_kartu_stok(df)

    assert hasil["df"].iloc[0]["arah"] == "MASUK"





# ============================================================

# TES TAMBAHAN -- PENGUATAN SLIP GAJI (gap #1-#6)

# ============================================================



def test_normalisasi_periode_gaji_format_beda_dikenali_sama():

    # "Juni 2026" vs "06/2026" vs "2026-06" vs datetime harus dikenali

    # sbg periode yang SAMA (tahun, bulan) meski format tulisannya beda.

    assert ak._normalisasi_periode_gaji("Juni 2026") == (2026, 6)

    assert ak._normalisasi_periode_gaji("06/2026") == (2026, 6)

    assert ak._normalisasi_periode_gaji("6/2026") == (2026, 6)

    assert ak._normalisasi_periode_gaji("2026-06") == (2026, 6)

    assert ak._normalisasi_periode_gaji("Jun-26") == (2026, 6)

    import datetime as _dt

    assert ak._normalisasi_periode_gaji(_dt.date(2026, 6, 15)) == (2026, 6)





def test_proses_slip_gaji_duplikat_terdeteksi_walau_format_periode_beda():

    # [Gap #1] Karyawan & periode sama, tapi format penulisan periode beda

    # ("Juni 2026" vs "06/2026") -- SEBELUMNYA tidak ketahuan sbg duplikat.

    df = pd.DataFrame([

        {"nip": "001", "nama_karyawan": "Budi", "periode_gaji": "Juni 2026",

         "gaji_pokok": 8_000_000, "total_tunjangan": 0, "total_potongan": 0,

         "gaji_bersih": 8_000_000, "bpjs_kesehatan": 0, "bpjs_jht": 0,

         "bpjs_jp": 0, "pph21": 0, "potongan_lain": 0},

        {"nip": "001", "nama_karyawan": "Budi", "periode_gaji": "06/2026",

         "gaji_pokok": 8_000_000, "total_tunjangan": 0, "total_potongan": 0,

         "gaji_bersih": 8_000_000, "bpjs_kesehatan": 0, "bpjs_jht": 0,

         "bpjs_jp": 0, "pph21": 0, "potongan_lain": 0},

    ])

    hasil = ak.proses_slip_gaji(df)

    alasan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert any("DUPLIKAT" in a for a in alasan)

    assert hasil["ringkasan"]["jumlah_duplikat"] == 1





def test_cek_kewajaran_pph21_slip_di_bawah_ptkp_tetap_dipotong_besar():

    # [Gap #2] Gaji bruto di bawah PTKP TK/0 tapi masih ada potongan PPh 21

    # yang berarti -- janggal terlepas dari kategori PTKP karyawan sebenarnya.

    pesan = ak.cek_kewajaran_pph21_slip(3_000_000, 200_000)

    assert pesan is not None

    assert "PTKP" in pesan

    # Tarif efektif kecil (10%) di atas PTKP tetap tidak diflag (perilaku lama).

    assert ak.cek_kewajaran_pph21_slip(10_000_000, 1_000_000) is None





def test_bpjs_plafon_bisa_dioverride_env_var(monkeypatch):

    # [Gap #3] Plafon BPJS sekarang bisa diubah lewat environment variable

    # tanpa mengedit kode.

    monkeypatch.setenv("BPJS_KESEHATAN_PLAFON_UPAH", "13500000")

    import importlib

    ak_reload = importlib.reload(ak)

    assert ak_reload.BPJS_KESEHATAN_PLAFON_UPAH == 13_500_000

    monkeypatch.delenv("BPJS_KESEHATAN_PLAFON_UPAH", raising=False)

    importlib.reload(ak)  # kembalikan ke default utk tes lain





def test_proses_slip_gaji_deteksi_anomali_gaji_antar_periode_dalam_batch():

    # [Gap #4] Gaji Pokok karyawan yang sama naik >50% dibanding periode

    # sebelumnya DALAM FILE/BATCH yang sama.

    df = pd.DataFrame([

        {"nip": "005", "nama_karyawan": "Joko", "periode_gaji": "Mei 2026",

         "gaji_pokok": 8_000_000, "total_tunjangan": 0, "total_potongan": 0,

         "gaji_bersih": 8_000_000, "bpjs_kesehatan": 0, "bpjs_jht": 0,

         "bpjs_jp": 0, "pph21": 0, "potongan_lain": 0},

        {"nip": "005", "nama_karyawan": "Joko", "periode_gaji": "Juni 2026",

         "gaji_pokok": 32_000_000, "total_tunjangan": 0, "total_potongan": 0,

         "gaji_bersih": 32_000_000, "bpjs_kesehatan": 0, "bpjs_jht": 0,

         "bpjs_jp": 0, "pph21": 0, "potongan_lain": 0},

    ])

    hasil = ak.proses_slip_gaji(df)

    alasan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert any("dibanding periode" in a for a in alasan)

    assert hasil["ringkasan"]["jumlah_anomali_gaji_antar_periode"] == 1

    assert hasil["histori_gaji_terbaru"]["005"]["gaji_pokok"] == 32_000_000





def test_proses_slip_gaji_deteksi_anomali_gaji_lintas_file_pakai_histori():

    # [Gap #4] Anomali tetap terdeteksi walau periode sebelumnya berasal

    # dari upload/file terpisah (dioper lewat histori_gaji_sebelumnya).

    histori_awal = {

        "005": {"periode_urut": (2026, 5), "periode_label": "Mei 2026",

                "gaji_pokok": 8_000_000, "gaji_bruto": 8_000_000},

    }

    df = pd.DataFrame([

        {"nip": "005", "nama_karyawan": "Joko", "periode_gaji": "Juni 2026",

         "gaji_pokok": 32_000_000, "total_tunjangan": 0, "total_potongan": 0,

         "gaji_bersih": 32_000_000, "bpjs_kesehatan": 0, "bpjs_jht": 0,

         "bpjs_jp": 0, "pph21": 0, "potongan_lain": 0},

    ])

    hasil = ak.proses_slip_gaji(df, histori_gaji_sebelumnya=histori_awal)

    alasan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert any("upload sebelumnya" in a for a in alasan)





def test_proses_slip_gaji_kenaikan_kecil_tidak_diflag_anomali():

    # Kenaikan wajar (mis. kenaikan berkala kecil) TIDAK boleh diflag,

    # supaya tidak banjir false-positive.

    df = pd.DataFrame([

        {"nip": "006", "nama_karyawan": "Rani", "periode_gaji": "Mei 2026",

         "gaji_pokok": 8_000_000, "total_tunjangan": 0, "total_potongan": 0,

         "gaji_bersih": 8_000_000, "bpjs_kesehatan": 0, "bpjs_jht": 0,

         "bpjs_jp": 0, "pph21": 0, "potongan_lain": 0},

        {"nip": "006", "nama_karyawan": "Rani", "periode_gaji": "Juni 2026",

         "gaji_pokok": 8_300_000, "total_tunjangan": 0, "total_potongan": 0,

         "gaji_bersih": 8_300_000, "bpjs_kesehatan": 0, "bpjs_jht": 0,

         "bpjs_jp": 0, "pph21": 0, "potongan_lain": 0},

    ])

    hasil = ak.proses_slip_gaji(df)

    alasan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert not any("dibanding periode" in a for a in alasan)





def test_proses_slip_gaji_potongan_lain_negatif_diflag():

    # [Gap #5] potongan_lain negatif SEBELUMNYA hanya diam-diam mengurangi

    # total potongan tanpa flag apapun.

    df = pd.DataFrame([

        {"nip": "007", "nama_karyawan": "Dewi", "periode_gaji": "Juli 2026",

         "gaji_pokok": 8_000_000, "total_tunjangan": 0, "total_potongan": -100_000,

         "gaji_bersih": 8_100_000, "bpjs_kesehatan": 0, "bpjs_jht": 0,

         "bpjs_jp": 0, "pph21": 0, "potongan_lain": -100_000},

    ])

    hasil = ak.proses_slip_gaji(df)

    alasan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert any("tertulis NEGATIF" in a for a in alasan)

    assert hasil["ringkasan"]["jumlah_potongan_negatif"] == 1





def test_proses_slip_gaji_periode_tahun_janggal_diflag():

    # [Gap #6] Periode gaji dgn tahun tidak masuk akal (mis. "1900" atau

    # jauh di masa depan) -- indikasi salah ketik/salah kolom tanggal.

    df = pd.DataFrame([

        {"nip": "008", "nama_karyawan": "Agus", "periode_gaji": "Januari 1900",

         "gaji_pokok": 8_000_000, "total_tunjangan": 0, "total_potongan": 0,

         "gaji_bersih": 8_000_000, "bpjs_kesehatan": 0, "bpjs_jht": 0,

         "bpjs_jp": 0, "pph21": 0, "potongan_lain": 0},

        {"nip": "009", "nama_karyawan": "Wati", "periode_gaji": "Januari 2099",

         "gaji_pokok": 8_000_000, "total_tunjangan": 0, "total_potongan": 0,

         "gaji_bersih": 8_000_000, "bpjs_kesehatan": 0, "bpjs_jht": 0,

         "bpjs_jp": 0, "pph21": 0, "potongan_lain": 0},

    ])

    hasil = ak.proses_slip_gaji(df)

    alasan = [a for m in hasil["masalah"] for a in m["alasan"]]

    assert sum("Periode Gaji tahun" in a for a in alasan) == 2

    assert hasil["ringkasan"]["jumlah_periode_janggal"] == 2





if __name__ == "__main__":

    raise SystemExit(pytest.main([__file__, "-v"]))