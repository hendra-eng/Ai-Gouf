"""
tests/test_accounting_export.py
================================
Unit test untuk export format akuntansi standar (Jurnal Umum, Buku Besar,
Neraca Saldo, Export Paket Lengkap, Export 14-Sheet Lengkap) di
modules/accounting_export.py.
"""

import io

import openpyxl
import pandas as pd
import pytest

from modules.accounting_export import (
    generate_jurnal_umum,
    generate_buku_besar,
    generate_neraca_saldo,
    cek_neraca_saldo_balance,
    export_paket_akuntansi_lengkap,
    export_14_sheet_lengkap,
)


@pytest.fixture
def df_jurnal():
    return pd.DataFrame({
        "tanggal": ["2026-07-01", "2026-07-02"],
        "keterangan": ["Bayar listrik", "Terima piutang"],
        "no_akun_debet": ["5100", "1100"],
        "nama_akun_debet": ["BEBAN LISTRIK", "KAS"],
        "no_akun_kredit": ["1100", "1200"],
        "nama_akun_kredit": ["KAS", "PIUTANG USAHA"],
        "jml_debet": [100000, 200000],
        "jml_kredit": [100000, 200000],
    })


class TestJurnalUmum:
    def test_jurnal_umum_dua_baris_per_transaksi(self, df_jurnal):
        hasil = generate_jurnal_umum(df_jurnal)
        assert len(hasil) == 4  # 2 transaksi x (1 debet + 1 kredit)
        assert set(hasil["no_bukti"]) == {"JU-00001", "JU-00002"}

    def test_jurnal_umum_balance(self, df_jurnal):
        hasil = generate_jurnal_umum(df_jurnal)
        assert abs(hasil["debet"].sum() - hasil["kredit"].sum()) < 1.0

    def test_jurnal_umum_data_kosong(self):
        hasil = generate_jurnal_umum(pd.DataFrame())
        assert hasil.empty


class TestBukuBesar:
    def test_buku_besar_per_akun(self, df_jurnal):
        buku_besar = generate_buku_besar(df_jurnal)
        assert "5100" in buku_besar
        assert "1100" in buku_besar
        assert "saldo_berjalan" in buku_besar["1100"]["data"].columns

    def test_buku_besar_kosong(self):
        assert generate_buku_besar(pd.DataFrame()) == {}


class TestNeracaSaldo:
    def test_neraca_saldo_balance(self, df_jurnal):
        neraca = generate_neraca_saldo(df_jurnal)
        assert cek_neraca_saldo_balance(neraca) is True

    def test_neraca_saldo_ada_baris_total(self, df_jurnal):
        neraca = generate_neraca_saldo(df_jurnal)
        assert "TOTAL" in neraca["no_akun"].values

    def test_neraca_saldo_tidak_balance_terdeteksi(self):
        df_tidak_balance = pd.DataFrame({
            "tanggal": ["2026-07-01"],
            "keterangan": ["Transaksi aneh"],
            "no_akun_debet": ["5100"], "nama_akun_debet": ["BEBAN"],
            "no_akun_kredit": ["1100"], "nama_akun_kredit": ["KAS"],
            "jml_debet": [100000], "jml_kredit": [90000],
        })
        neraca = generate_neraca_saldo(df_tidak_balance)
        assert cek_neraca_saldo_balance(neraca) is False


class TestExportPaketLengkap:
    def test_export_menghasilkan_bytes_excel(self, df_jurnal):
        hasil = export_paket_akuntansi_lengkap(df_jurnal)
        assert isinstance(hasil, bytes)
        assert len(hasil) > 0


# ---------------------------------------------------------------------------
# Export 14-Sheet Lengkap
# ---------------------------------------------------------------------------

SHEET_AKTIVA = "Buku Bantu Aktiva Tetap"

# Kolom (1-indexed, openpyxl) di sheet Buku Bantu Aktiva Tetap untuk periode
# 1 tahun (12 bulan): 13 kolom info dasar + 12 kolom bulan + 3 kolom ringkasan.
COL_HARGA_PEROLEHAN = 8
COL_AKUM_AWAL = 13
COL_BULAN_PERTAMA = 14   # Jan
COL_BULAN_TERAKHIR = 25  # Dec
COL_PENYUSUTAN_TAHUN = 26
COL_AKUM_AKHIR = 27
COL_NILAI_BUKU = 28

# [UPDATE] Layout sheet "Buku Bantu Aktiva Tetap" sekarang (lihat
# accounting_export.py sekitar baris 1359-1391) punya baris judul + baris
# kosong + baris tanggal akhir bulan SEBELUM baris header teks -- disesuaikan
# supaya sama persis dengan template referensi
# (Model_Laporan_Keuangan_SPT_PPh31E_2025.xlsx):
#   row 1  -> judul "BUKU BANTU AKTIVA TETAP & JADWAL PENYUSUTAN <tahun>"
#   row 2  -> kosong
#   row 3  -> baris tanggal akhir tiap bulan (31 Jan, 28/29 Feb, dst)
#   row 4  -> header teks ("Asset ID", "Jan-26", ..., "Penyusutan 2026", ...)
#   row 5+ -> data per aset, lalu baris TOTAL setelahnya
ROW_HEADER = 4
ROW_DATA_PERTAMA = 5
ROW_DATA_KEDUA = 6
ROW_TOTAL = 7
COL_TOTAL_LABEL = 0  # kolom A ("Asset ID") -- label TOTAL dipindah ke sini,
                     # bukan kolom B ("Nama Aset") lagi, sesuai template.


@pytest.fixture
def jadwal_aset_dua_aset():
    """
    Dua aset dengan skenario beda:
      - FA-001: aset baru, mulai disusutkan Juli (bulan 7) -- 6 bulan awal
        harus 0, akumulasi awal tahun 0.
      - FA-002: aset lama, sudah punya akumulasi awal tahun, disusutkan
        rata setahun penuh.
    """
    return {
        "metode": "Garis Lurus",
        "aset": [
            {
                "kode_aset": "FA-001", "nama_aset": "Mobil Operasional",
                "kode_akun_aset": "1510", "kode_akum_penyusutan": "1511",
                "kode_beban_penyusutan": "6100",
                "tanggal_perolehan": "2026-07-01", "mulai_digunakan": "2026-07-01",
                "harga_perolehan": 5_000_000, "nilai_residu": 0, "umur_tahun": 4,
                "metode": "Garis Lurus", "penyusutan_per_bulan": 100_000,
                "akumulasi_awal_tahun": 0,
                "jadwal_bulanan": [
                    {"bulan": i + 1, "penyusutan_bulan_ini": 0 if i < 6 else 100_000}
                    for i in range(12)
                ],
            },
            {
                "kode_aset": "FA-002", "nama_aset": "Mesin Produksi",
                "kode_akun_aset": "1520", "kode_akum_penyusutan": "1521",
                "kode_beban_penyusutan": "6100",
                "tanggal_perolehan": "2020-01-01", "mulai_digunakan": "2020-01-01",
                "harga_perolehan": 10_000_000, "nilai_residu": 0, "umur_tahun": 5,
                "metode": "Garis Lurus", "penyusutan_per_bulan": 200_000,
                "akumulasi_awal_tahun": 1_000_000,
                "jadwal_bulanan": [
                    {"bulan": i + 1, "penyusutan_bulan_ini": 200_000} for i in range(12)
                ],
            },
        ],
    }


def _load_sheet(hasil_bytes: bytes, nama_sheet: str):
    wb = openpyxl.load_workbook(io.BytesIO(hasil_bytes))
    return wb, wb[nama_sheet]


class TestExport14SheetLengkap:
    def test_menghasilkan_bytes_excel_dengan_14_sheet(self, jadwal_aset_dua_aset):
        hasil = export_14_sheet_lengkap({
            "periode": "2026", "coa": [], "jurnal": [], "jadwal_aset": jadwal_aset_dua_aset,
        })
        assert isinstance(hasil, bytes)
        assert len(hasil) > 0
        wb = openpyxl.load_workbook(io.BytesIO(hasil))
        assert len(wb.sheetnames) == 14
        assert SHEET_AKTIVA in wb.sheetnames
        assert "GL 2026" in wb.sheetnames

    def test_aktiva_tetap_kolom_ringkasan_ada_di_header(self, jadwal_aset_dua_aset):
        hasil = export_14_sheet_lengkap({
            "periode": "2026", "coa": [], "jurnal": [], "jadwal_aset": jadwal_aset_dua_aset,
        })
        _, ws = _load_sheet(hasil, SHEET_AKTIVA)
        header = [c.value for c in ws[ROW_HEADER]]
        assert "Penyusutan 2026" in header
        assert "Akum. Penyusutan Akhir" in header
        assert "Nilai Buku 31/12/2026" in header

    def test_aktiva_tetap_aset_baru_tengah_tahun(self, jadwal_aset_dua_aset):
        """FA-001 baru disusutkan mulai Juli -> Jan-Jun harus 0."""
        hasil = export_14_sheet_lengkap({
            "periode": "2026", "coa": [], "jurnal": [], "jadwal_aset": jadwal_aset_dua_aset,
        })
        _, ws = _load_sheet(hasil, SHEET_AKTIVA)
        baris_fa001 = ws[ROW_DATA_PERTAMA]
        bulan_values = [c.value for c in baris_fa001[COL_BULAN_PERTAMA - 1:COL_BULAN_TERAKHIR]]
        assert bulan_values[:6] == [0, 0, 0, 0, 0, 0]
        assert bulan_values[6:] == [100_000] * 6
        assert baris_fa001[COL_PENYUSUTAN_TAHUN - 1].value == 600_000
        assert baris_fa001[COL_AKUM_AKHIR - 1].value == 600_000  # akum awal 0 + 600rb
        assert baris_fa001[COL_NILAI_BUKU - 1].value == 4_400_000  # 5jt - 600rb

    def test_aktiva_tetap_aset_lama_dengan_akum_awal(self, jadwal_aset_dua_aset):
        """FA-002 sudah disusutkan sebelumnya (akumulasi_awal_tahun > 0)."""
        hasil = export_14_sheet_lengkap({
            "periode": "2026", "coa": [], "jurnal": [], "jadwal_aset": jadwal_aset_dua_aset,
        })
        _, ws = _load_sheet(hasil, SHEET_AKTIVA)
        baris_fa002 = ws[ROW_DATA_KEDUA]
        assert baris_fa002[COL_PENYUSUTAN_TAHUN - 1].value == 2_400_000  # 12 x 200rb
        assert baris_fa002[COL_AKUM_AKHIR - 1].value == 3_400_000  # 1jt + 2.4jt
        assert baris_fa002[COL_NILAI_BUKU - 1].value == 6_600_000  # 10jt - 3.4jt

    def test_aktiva_tetap_baris_total(self, jadwal_aset_dua_aset):
        hasil = export_14_sheet_lengkap({
            "periode": "2026", "coa": [], "jurnal": [], "jadwal_aset": jadwal_aset_dua_aset,
        })
        _, ws = _load_sheet(hasil, SHEET_AKTIVA)
        baris_total = ws[ROW_TOTAL]
        assert baris_total[COL_TOTAL_LABEL].value == "TOTAL"  # kolom Asset ID
        assert baris_total[COL_HARGA_PEROLEHAN - 1].value == 15_000_000
        assert baris_total[COL_PENYUSUTAN_TAHUN - 1].value == 3_000_000
        assert baris_total[COL_AKUM_AKHIR - 1].value == 4_000_000
        assert baris_total[COL_NILAI_BUKU - 1].value == 11_000_000

    def test_aktiva_tetap_konsistensi_nilai_buku(self, jadwal_aset_dua_aset):
        """Nilai Buku = Harga Perolehan - Akum. Penyusutan Akhir, untuk tiap baris aset."""
        hasil = export_14_sheet_lengkap({
            "periode": "2026", "coa": [], "jurnal": [], "jadwal_aset": jadwal_aset_dua_aset,
        })
        _, ws = _load_sheet(hasil, SHEET_AKTIVA)
        for baris in (ws[ROW_DATA_PERTAMA], ws[ROW_DATA_KEDUA]):
            harga = baris[COL_HARGA_PEROLEHAN - 1].value
            akum_akhir = baris[COL_AKUM_AKHIR - 1].value
            nilai_buku = baris[COL_NILAI_BUKU - 1].value
            assert nilai_buku == harga - akum_akhir

    def test_aktiva_tetap_data_kosong(self):
        """Belum ada aset -> header tetap ditulis, lalu pesan placeholder
        di baris berikutnya, bukan error/crash."""
        hasil = export_14_sheet_lengkap({
            "periode": "2026", "coa": [], "jurnal": [], "jadwal_aset": {},
        })
        _, ws = _load_sheet(hasil, SHEET_AKTIVA)
        assert ws.cell(row=ROW_HEADER, column=1).value == "Asset ID"
        isi_baris_placeholder = ws.cell(row=ROW_HEADER + 1, column=1).value
        assert "Belum ada data Aset Tetap" in isi_baris_placeholder

    def test_export_tanpa_data_sama_sekali_tidak_error(self):
        """data minimal (semua kosong) tidak boleh membuat fungsi crash."""
        hasil = export_14_sheet_lengkap({"periode": "2026"})
        assert isinstance(hasil, bytes)
        assert len(hasil) > 0