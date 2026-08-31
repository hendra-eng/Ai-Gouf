"""
modules/accounting_export.py
=============================
Export ke format akuntansi standar:
- Jurnal Umum (General Journal) dengan nomor bukti berurutan
- Buku Besar (General Ledger) per akun dengan saldo berjalan
- Neraca Saldo (Trial Balance) per akun

Input yang diharapkan: DataFrame jurnal dengan kolom minimal
tanggal, keterangan, no_akun_debet, nama_akun_debet, no_akun_kredit,
nama_akun_kredit, jml_debet, jml_kredit (format yang sama seperti
`st.session_state.hasil_bank` di app.py / output akuntansi_ai.py).
"""

from __future__ import annotations

import calendar
import io
import re
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Dict, List, Optional

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties

from .cross_matching import cocokkan_bank_piutang
from .logging_config import get_module_logger
from .laporan_keuangan import (
    susun_neraca_lampiran_spt_baku,
    susun_laba_rugi_rekonsiliasi_fiskal_lampiran_spt_baku,
    _tanggal_untuk_jadwal as _tanggal_untuk_jadwal_aset,
)

logger = get_module_logger("accounting_export")


def _angka(v) -> float:
    """
    [FIX] Konversi nilai ke float dengan aman -- None/NaN/inf semua
    dianggap 0.0. Kode lama di modul ini pakai pola
    `float(row.get("mutasi_debet") or 0)`, yang TIDAK aman untuk NaN:
    `float('nan') or 0` mengembalikan `nan` itu sendiri (NaN truthy di
    Python), jadi fallback "or 0"-nya tidak kepakai. Ini bisa menulis nilai
    NaN langsung ke sel Excel (baris saldo_awal/mutasi di
    _tulis_sheet_bank) -- Excel/openpyxl tidak menerima NaN dengan baik
    (bisa gagal saat dibuka atau export error), dan formula SUBTOTAL/cek
    saldo berjalan ikut rusak begitu 1 sel NaN muncul. Pola bug yang sama
    dengan yang sudah diperbaiki di akuntansi_ai.py.
    """
    if v is None:
        return 0.0
    try:
        f = float(v)
    except (TypeError, ValueError):
        return 0.0
    if pd.isna(f) or f in (float("inf"), float("-inf")):
        return 0.0
    return f

# [BERUBAH - Prioritas #7] accounting_export.py TIDAK LAGI mengambil nomor
# voucher dari database sendiri (sebelumnya via db_client.ambil_blok_nomor_voucher(),
# makanya dulu perlu `import db_client as dbc` + parameter client_id di sini).
# Sejak keputusan integrasi #7, penomoran voucher rekening_koran dipindah
# SELURUHNYA ke saat upload -- lihat db_client.tarik_draf_jurnal_ke_posting()
# -- supaya ada SATU sumber nomor voucher yang konsisten dgn status posting
# di tabel jurnal_posting. Modul ini sekarang murni menulis apa pun voucher
# yang sudah ada di df_hasil["voucher"] (diisi oleh main.py dari jurnal_posting
# sebelum memanggil export_rekening_koran_format_akuntan()); baris yang belum
# punya voucher (belum masuk antrean posting -- lihat PLACEHDR_VOUCHER di
# bawah) TIDAK digenerate nomor baru di sini, supaya tidak menarik dobel dari
# counter yang sama dan supaya nomor voucher di Excel selalu identik dengan
# yang tercatat di jurnal_posting.
PLACEHOLDER_VOUCHER = "(belum diposting)"

# [BARU - Prioritas #5] Warna highlight baris yang perlu direview manual
# akuntan, supaya bisa langsung di-scroll/filter -- bukan baca satu-satu
# di antara ribuan baris. Dua kategori beda tingkat urgensi:
#   - KUNING : sisi debet dan/atau kredit BELUM ada kode akun sama sekali
#              (belum berhasil dikategorikan oleh pola/AI) -- kasus paling
#              umum, akuntan tinggal isi manual.
#   - MERAH  : kode akun SUDAH ada tapi tidak ditemukan di sheet COA --
#              berarti VLOOKUP bakal gagal (muncul "BELUM TERKATEGORI" di
#              kolom Nama Akun) -- ini masalah data (kode salah ketik/COA
#              belum lengkap), lebih mendesak drpd sekadar belum dikategorikan.
FILL_BELUM_DIKATEGORI = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
FILL_KODE_TIDAK_DITEMUKAN = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
JUMLAH_KOLOM_SHEET_BANK = 16  # A..P


def generate_jurnal_umum(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ubah jurnal 1-baris-per-transaksi (debet & kredit di kolom terpisah)
    menjadi format Jurnal Umum standar: 1 baris per SISI (debet/kredit),
    dengan nomor bukti berurutan per transaksi.
    """
    if df is None or df.empty:
        return pd.DataFrame(columns=[
            "no_bukti", "tanggal", "keterangan", "no_akun", "nama_akun", "debet", "kredit"
        ])

    baris = []
    for i, (_, row) in enumerate(df.iterrows(), start=1):
        no_bukti = f"JU-{i:05d}"
        tanggal = row.get("tanggal")
        keterangan = row.get("keterangan")

        # Transaksi yang belum terkategorisasi (salah satu/kedua akun masih
        # kosong) dilewati -- kalau tetap dimasukkan, hanya salah satu sisi
        # (debet/kredit) yang akan tercatat sehingga jurnal jadi tidak balance.
        if pd.isna(row.get("no_akun_debet")) or pd.isna(row.get("no_akun_kredit")):
            continue

        # [FIX -- BUG NYATA, pola NaN-truthy yang PERSIS dijelaskan di
        # docstring _angka() paling atas file ini, tapi kelewat dipakai
        # di sini] Sebelumnya `row.get("jml_debet") or 0` -- kalau
        # "jml_debet" NaN (bukan None/0, mis. sel kosong yang lewat
        # pandas), `NaN or 0` mengembalikan NaN itu sendiri (NaN truthy
        # di Python) -- lalu `if debet:` JUGA True untuk NaN, jadi baris
        # dgn nilai NaN tetap DITULIS ke Jurnal Umum (bukan dilewati
        # seperti niat aslinya "cuma tulis kalau ada nilai"). NaN yang
        # nyasar ke sel Excel TIDAK memicu error saat ditulis (openpyxl
        # diam saja), tapi Excel akan menampilkannya sbg cell rusak/
        # #NUM! saat dibuka client -- dan baris itu OTOMATIS ikut
        # menyandera Buku Besar & Neraca Saldo (generate_buku_besar/
        # generate_neraca_saldo, keduanya dibangun dari fungsi ini).
        # Fix: pakai _angka() (sudah ada, dibuat khusus utk mencegah pola
        # ini) -- NaN sekarang benar-benar jadi 0.0, baris itu otomatis
        # ikut terlewati oleh `if debet:`/`if kredit:` di bawah, PERSIS
        # niat aslinya (cuma tulis sisi yang punya nilai sungguhan).
        debet = _angka(row.get("jml_debet"))
        if debet:
            baris.append({
                "no_bukti": no_bukti, "tanggal": tanggal, "keterangan": keterangan,
                "no_akun": row.get("no_akun_debet"), "nama_akun": row.get("nama_akun_debet"),
                "debet": debet, "kredit": 0,
            })

        kredit = _angka(row.get("jml_kredit"))
        if kredit:
            baris.append({
                "no_bukti": no_bukti, "tanggal": tanggal, "keterangan": keterangan,
                "no_akun": row.get("no_akun_kredit"), "nama_akun": row.get("nama_akun_kredit"),
                "debet": 0, "kredit": kredit,
            })

    hasil = pd.DataFrame(baris)
    logger.info(f"✅ Jurnal Umum dibuat: {len(hasil)} baris dari {len(df)} transaksi")
    return hasil


def generate_buku_besar(df: pd.DataFrame, df_coa: Optional[pd.DataFrame] = None) -> dict:
    """
    Buat Buku Besar: satu DataFrame per no_akun, dengan saldo berjalan.

    Returns:
        dict {no_akun: {"nama_akun": str, "data": DataFrame}}
    """
    jurnal_umum = generate_jurnal_umum(df)
    if jurnal_umum.empty:
        return {}

    jurnal_umum = jurnal_umum.sort_values(["no_akun", "tanggal"], na_position="last")

    buku_besar = {}
    for no_akun, grup in jurnal_umum.groupby("no_akun"):
        if no_akun is None:
            continue
        grup = grup.copy().sort_values("tanggal", na_position="last")
        grup["saldo_berjalan"] = (grup["debet"] - grup["kredit"]).cumsum()
        nama_akun = grup["nama_akun"].dropna().iloc[0] if grup["nama_akun"].notna().any() else ""
        buku_besar[no_akun] = {
            "nama_akun": nama_akun,
            "data": grup[["tanggal", "no_bukti", "keterangan", "debet", "kredit", "saldo_berjalan"]],
        }

    logger.info(f"✅ Buku Besar dibuat untuk {len(buku_besar)} akun")
    return buku_besar


def generate_neraca_saldo(df: pd.DataFrame, df_coa: Optional[pd.DataFrame] = None) -> pd.DataFrame:
    """
    Buat Neraca Saldo (Trial Balance): total debet & kredit per akun,
    plus saldo akhir (debet jika positif, kredit jika negatif).
    """
    jurnal_umum = generate_jurnal_umum(df)
    if jurnal_umum.empty:
        return pd.DataFrame(columns=["no_akun", "nama_akun", "total_debet", "total_kredit", "saldo_debet", "saldo_kredit"])

    agg = jurnal_umum.groupby("no_akun").agg(
        nama_akun=("nama_akun", lambda s: s.dropna().iloc[0] if s.notna().any() else ""),
        total_debet=("debet", "sum"),
        total_kredit=("kredit", "sum"),
    ).reset_index()

    if df_coa is not None and not df_coa.empty and "no_akun" in df_coa.columns:
        coa_map = df_coa.set_index(df_coa["no_akun"].astype(str))["kategori"].to_dict() \
            if "kategori" in df_coa.columns else {}
        agg["kategori"] = agg["no_akun"].astype(str).map(coa_map)

    selisih = agg["total_debet"] - agg["total_kredit"]
    agg["saldo_debet"] = selisih.clip(lower=0)
    agg["saldo_kredit"] = (-selisih).clip(lower=0)

    agg = agg.sort_values("no_akun")

    total_row = pd.DataFrame([{
        "no_akun": "TOTAL", "nama_akun": "",
        "total_debet": agg["total_debet"].sum(), "total_kredit": agg["total_kredit"].sum(),
        "saldo_debet": agg["saldo_debet"].sum(), "saldo_kredit": agg["saldo_kredit"].sum(),
    }])
    hasil = pd.concat([agg, total_row], ignore_index=True)

    logger.info(f"✅ Neraca Saldo dibuat untuk {len(agg)} akun")
    return hasil


def cek_neraca_saldo_balance(neraca: pd.DataFrame, toleransi: float = 1.0) -> bool:
    """Cek apakah total saldo debet = total saldo kredit (baris TOTAL)."""
    if neraca is None or neraca.empty:
        return False
    baris_total = neraca[neraca["no_akun"] == "TOTAL"]
    if baris_total.empty:
        return False
    selisih = abs(baris_total["saldo_debet"].iloc[0] - baris_total["saldo_kredit"].iloc[0])
    return bool(selisih <= toleransi)


def hitung_metrik_kategorisasi(df_hasil: pd.DataFrame) -> Dict[str, object]:
    """
    [BARU - Prioritas #3] Ringkas berapa % baris rekening koran yang
    berhasil otomatis dikategorikan, dipecah per SUMBER (pola historis /
    kata kunci COA / AI / data asli dari file / belum terkategori) --
    supaya begitu rekening koran MENTAH diupload, angka "seberapa dekat
    ke hasil manual akuntan" langsung kelihatan, tanpa perlu dihitung
    manual baris per baris.

    df_hasil: keluaran ak.proses_file_rekening_koran()["df"] (wajib punya
    kolom "sumber_kategori" -- lihat proses_dataframe() di akuntansi_ai.py
    untuk daftar lengkap nilai yang mungkin muncul di kolom itu).

    Bucket dicocokkan dari POTONGAN teks sumber_kategori (bukan exact
    match) supaya tahan terhadap variasi suffix, mis. "AI (DeepSeek) -
    confidence tinggi, perlu review" tetap masuk bucket "ai", dan "Sesuai
    Pola yang Dipelajari (perlu cek)" tetap masuk bucket "pola_historis"
    SEKALIGUS terhitung di jumlah_perlu_review.
    """
    if df_hasil is None or df_hasil.empty or "sumber_kategori" not in df_hasil.columns:
        return {
            "jumlah_transaksi": 0, "jumlah_otomatis_terkategori": 0,
            "persentase_otomatis": 0.0, "jumlah_perlu_review": 0,
            "persentase_perlu_review": 0.0, "per_sumber": {},
        }

    total = len(df_hasil)
    sumber = df_hasil["sumber_kategori"].fillna("Belum Terkategori - perlu review manual").astype(str)

    def _termasuk(potongan: str) -> pd.Series:
        return sumber.str.contains(potongan, case=False, na=False, regex=True)

    bucket = {
        "pola_historis": _termasuk("Pola yang Dipelajari"),
        "kata_kunci_coa": _termasuk("Kata kunci COA"),
        "ai": _termasuk(r"AI \(DeepSeek\)"),
        "data_asli_dari_file": _termasuk("Data Asli dari File"),
        "belum_terkategori": _termasuk("Belum Terkategori"),
    }
    # Baris yang sumber_kategori-nya tidak cocok bucket manapun di atas
    # (mis. versi kode berikutnya menambah sumber baru tapi lupa
    # diupdate di sini) tetap dihitung terpisah, supaya totalnya selalu
    # pas -- bukan diam-diam hilang dari ringkasan.
    sudah_kehitung = pd.concat(bucket.values(), axis=1).any(axis=1)
    bucket["lainnya_tidak_dikenali"] = ~sudah_kehitung

    per_sumber = {
        nama: {
            "jumlah": int(mask.sum()),
            "persentase": round(float(mask.sum()) / total * 100, 1) if total else 0.0,
        }
        for nama, mask in bucket.items()
    }

    jumlah_belum = per_sumber["belum_terkategori"]["jumlah"]
    jumlah_otomatis = total - jumlah_belum

    # "perlu review" = union SEMUA baris yang tetap butuh mata manusia:
    # belum terkategori sama sekali, ATAU sumbernya eksplisit menandai
    # "perlu cek"/"perlu review" (pola tidak 100% konsisten antar contoh,
    # atau AI confidence rendah/sedang) -- beda dari "belum_terkategori"
    # yang HANYA baris kosong total.
    perlu_review_mask = _termasuk("Belum Terkategori") | _termasuk("perlu cek") | _termasuk("perlu review")
    jumlah_perlu_review = int(perlu_review_mask.sum())

    return {
        "jumlah_transaksi": total,
        "jumlah_otomatis_terkategori": jumlah_otomatis,
        "persentase_otomatis": round(jumlah_otomatis / total * 100, 1) if total else 0.0,
        "jumlah_perlu_review": jumlah_perlu_review,
        "persentase_perlu_review": round(jumlah_perlu_review / total * 100, 1) if total else 0.0,
        "per_sumber": per_sumber,
    }


def export_paket_akuntansi_lengkap(df: pd.DataFrame, df_coa: Optional[pd.DataFrame] = None) -> bytes:
    """
    Export satu file Excel berisi 3 sheet: Jurnal Umum, Neraca Saldo,
    dan ringkasan Buku Besar (semua akun digabung dalam 1 sheet, dipisah
    per akun dengan baris header).

    Returns:
        bytes file Excel siap didownload.
    """
    buffer = io.BytesIO()

    jurnal_umum = generate_jurnal_umum(df)
    neraca_saldo = generate_neraca_saldo(df, df_coa)
    buku_besar = generate_buku_besar(df, df_coa)

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        jurnal_umum.to_excel(writer, sheet_name="Jurnal Umum", index=False)
        neraca_saldo.to_excel(writer, sheet_name="Neraca Saldo", index=False)

        # Buku besar: tulis semua akun ke satu sheet, dipisah dengan baris judul
        baris_gabungan = []
        for no_akun, info in buku_besar.items():
            baris_gabungan.append(pd.DataFrame([{
                "tanggal": f"=== AKUN {no_akun} - {info['nama_akun']} ===",
                "no_bukti": "", "keterangan": "", "debet": "", "kredit": "", "saldo_berjalan": "",
            }]))
            baris_gabungan.append(info["data"])
        if baris_gabungan:
            pd.concat(baris_gabungan, ignore_index=True).to_excel(writer, sheet_name="Buku Besar", index=False)
        else:
            pd.DataFrame().to_excel(writer, sheet_name="Buku Besar", index=False)

    buffer.seek(0)
    logger.info("✅ Paket akuntansi lengkap (Jurnal Umum + Buku Besar + Neraca Saldo) berhasil dibuat")
    return buffer.getvalue()


# ============================================================
# [BARU] EXPORT REKENING KORAN -- FORMAT KERJA AKUNTAN
# ============================================================
# Beda dari fungsi-fungsi di atas (yang menghasilkan Jurnal Umum/Buku
# Besar/Neraca Saldo generik): fungsi di bawah ini meniru PERSIS format
# file kerja rekonsiliasi bank yang biasa dipakai tim akuntan -- 1 sheet
# COA + 1 sheet per bank, dengan:
#   - nomor voucher berurutan per bank (mis. "BRI-0726-1")
#   - kolom NO AKUN diisi kode akun mentah, NAMA AKUN pakai formula
#     VLOOKUP ke sheet COA (bukan nilai yang sudah "dihitung" di Python)
#   - kolom CEK saldo berjalan (formula, bukan angka jadi) supaya
#     kalau data mentahnya berubah, akuntan bisa recalculate dan
#     langsung ketahuan kalau ada baris yang saldonya tidak nyambung
#
# Input df_hasil: DataFrame keluaran ak.proses_file_rekening_koran()["df"]
# (fungsi ini SUDAH melalui parse_sheet_bank() + proses_dataframe() di
# akuntansi_ai.py, jadi kolom no_akun_debet/no_akun_kredit dsb sudah ada
# -- walau bisa saja masih NaN untuk baris yang belum berhasil
# dikategorikan, itu wajar & akan tampil kosong di sheet hasil, tinggal
# diisi manual oleh akuntan).

def _kode_bank_dari_nama(nama_bank: str) -> str:
    """
    Kode singkat dari nama bank untuk prefix nomor voucher, mis.
    "BANK BRI" -> "BRI", "Mandiri" -> "MANDIRI" -- ambil kata TERAKHIR
    supaya hasilnya konsisten dengan pola voucher yang lazim dipakai
    akuntan (mis. "BRI-0726-1"), bukan "BANK-BRI-0726-1".
    """
    kata = str(nama_bank).strip().upper().split()
    return kata[-1] if kata else "BANK"


def _normalisasi_kode_akun(nilai) -> Optional[str]:
    """
    [BARU - Prioritas #5] Normalisasi kode akun ke string SEBELUM
    dibandingkan, supaya tidak salah tandai "kode tidak ditemukan" gara-gara
    representasi angka beda (bug yang sempat ketemu saat testing: kolom
    pandas yang campur int & None otomatis naik jadi float64, jadi
    11200003 (int, di COA) vs 11200003.0 (float, di df_hasil kalau
    kolomnya sempat diisi None di baris lain) -- str() keduanya beda
    ("11200003" vs "11200003.0") padahal kodenya SAMA. Fungsi ini
    menyeragamkan keduanya jadi "11200003" apapun representasi awalnya.
    """
    if nilai is None or (isinstance(nilai, float) and pd.isna(nilai)):
        return None
    try:
        f = float(nilai)
        if f == int(f):
            return str(int(f))
    except (TypeError, ValueError):
        pass
    return str(nilai).strip()


def _tulis_sheet_coa(wb: openpyxl.Workbook, df_coa: Optional[pd.DataFrame]):
    """
    Tulis sheet "COA" dengan layout CAT | NO AKUN | DESCRIPTION -- kolom B
    (NO AKUN) & C (DESCRIPTION) sengaja di posisi ini karena formula
    VLOOKUP di sheet bank mereferensikan COA!$B:$C secara eksplisit.
    """
    ws = wb.create_sheet("COA")
    ws.append(["CAT", "NO AKUN", "DESCRIPTION"])
    if df_coa is not None and not df_coa.empty:
        for _, r in df_coa.iterrows():
            ws.append([r.get("kategori"), r.get("no_akun"), r.get("nama_akun")])
    return ws


def _cocokkan_supplier_opsional(df_bank: pd.DataFrame, df_piutang: Optional[pd.DataFrame]) -> pd.DataFrame:
    """
    Isi kolom supplier_cust (dan voucher, kalau masih kosong) dengan hasil
    pencocokan ke Buku Bantu Piutang (kalau df_piutang dikasih).

    [BERUBAH - Prioritas #8] Sebelumnya fungsi ini punya implementasi lokal
    sendiri (cuma cek nominal + kemiripan nama, TANPA cek tanggal walau
    tanggalnya sempat dihitung) sebagai jalan pintas, karena
    cross_matching.cocokkan_bank_piutang() saat itu punya bug: memperlakukan
    df_bank["mutasi_debet"] sebagai "uang MASUK", padahal berdasarkan
    konvensi parse_sheet_bank() di akuntansi_ai.py, uang masuk dari customer
    tercatat di mutasi_kredit. Bug itu SUDAH diperbaiki di cross_matching.py,
    jadi fungsi ini sekarang memanggil cocokkan_bank_piutang() langsung --
    lebih kuat (cek tanggal + nominal + nama, bukan cuma nominal + nama)
    dan menghindari 2 implementasi pencocokan bank<->piutang yang bisa
    berbeda hasil satu sama lain.
    """
    df_bank = df_bank.copy()
    if df_piutang is None or df_piutang.empty or "nama_pelanggan" not in df_piutang.columns:
        return df_bank

    # Hanya baris yang supplier_cust-nya BELUM terisi yang ikut dilombakan
    # utk dicocokkan -- baris yang sudah ada isinya (mis. dari hasil AI /
    # pola belajar sebelumnya) tidak boleh "direbut" kandidatnya, dan tidak
    # perlu diproses ulang. Filter berbasis boolean mask begini (bukan
    # df_bank.reset_index()) SENGAJA dipertahankan supaya index baris yang
    # dikirim ke cocokkan_bank_piutang() tetap SAMA dengan index di df_bank
    # -- cocokkan_bank_piutang() sekarang mengembalikan bank_index berbasis
    # index asli df_bank yang di-passing (lihat catatan di cross_matching.py),
    # jadi bisa langsung dipakai utk df_bank.at[idx, ...] di bawah tanpa
    # pemetaan index tambahan.
    supplier_kosong = df_bank["supplier_cust"].isna() | (
        df_bank["supplier_cust"].astype(str).str.strip() == ""
    )
    df_bank_kandidat = df_bank[supplier_kosong]

    hasil_cocok = cocokkan_bank_piutang(df_bank_kandidat, df_piutang)
    for h in hasil_cocok["hasil"]:
        if h["status"] != "MATCHED" or h["bank_index"] is None:
            continue
        idx = h["bank_index"]
        if idx not in df_bank.index:
            continue  # jaga-jaga kalau index tidak valid, tidak seharusnya terjadi
        df_bank.at[idx, "supplier_cust"] = h["nama_pelanggan"]
        voucher_sekarang = df_bank.at[idx, "voucher"]
        if voucher_sekarang in (None, "") or pd.isna(voucher_sekarang):
            df_bank.at[idx, "voucher"] = h["no_transaksi"]

    return df_bank


def _tulis_sheet_bank(
    wb: openpyxl.Workbook,
    nama_bank: str,
    df_bank: pd.DataFrame,
    kode_akun_valid: Optional[set] = None,
) -> Dict[str, int]:
    """
    Tulis satu sheet bank dengan struktur baku (sama untuk semua bank,
    tidak meniru posisi kolom khas tiap bank sumber -- ini file HASIL yang
    kita generate sendiri, jadi layoutnya konsisten):

    A No | B Tanggal | C Keterangan | D Mutasi Debet | E Mutasi Kredit |
    F Saldo Akhir | G Cek Saldo Berjalan (formula) | H Selisih Cek |
    I Supplier/Cust | J Voucher | K No Akun Debet | L Nama Akun Debet
    (VLOOKUP) | M Jumlah Debet | N No Akun Kredit | O Nama Akun Kredit
    (VLOOKUP) | P Jumlah Kredit

    [BARU - Prioritas #5] Baris yang perlu direview manual akuntan
    di-highlight (lihat FILL_BELUM_DIKATEGORI / FILL_KODE_TIDAK_DITEMUKAN
    di atas), header di-freeze, dan auto-filter dinyalakan supaya akuntan
    bisa filter kolom "Nama Akun" utk isi kosong/"BELUM TERKATEGORI" tanpa
    scroll manual di antara ribuan baris.

    kode_akun_valid: set berisi (str) semua no_akun yang ADA di sheet COA
    -- dipakai untuk deteksi kode akun "nyasar" (sudah diisi tapi tidak
    ketemu di COA, yg nanti bikin VLOOKUP gagal) SEBELUM file dibuka di
    Excel, bukan menunggu akuntan buka file & baru ketahuan ada
    "BELUM TERKATEGORI". Kalau None (COA kosong/tidak dikasih), highlight
    kategori MERAH ini dilewati (tidak ada cara mengecek validitasnya).

    [BERUBAH - Prioritas #7] Fungsi ini TIDAK LAGI generate nomor voucher
    baru -- voucher diasumsikan sudah ada di kolom df_bank["voucher"]
    (diisi pemanggil dari jurnal_posting). Baris yang voucher-nya masih
    kosong ditulis dengan PLACEHOLDER_VOUCHER, bukan nomor baru, supaya
    tidak menarik dobel dari counter persisten (lihat catatan di kepala
    modul ini & di export_rekening_koran_format_akuntan()).

    Return dict ringkasan {"belum_dikategorikan": int, "kode_tidak_ditemukan": int}
    supaya bisa dilaporkan balik ke pemanggil (endpoint/log).
    """
    ws = wb.create_sheet(str(nama_bank)[:31])
    n = len(df_bank)
    baris_terakhir = n + 2  # header di baris 2, data mulai baris 3

    # --- Baris 1: total SUBTOTAL (tetap benar walau baris disortir/difilter) ---
    ws.append([
        None, None, None,
        f"=SUBTOTAL(9,D3:D{baris_terakhir})", f"=SUBTOTAL(9,E3:E{baris_terakhir})",
        None, None, None, None, None,
        "VOUCHER & FORMULA CEK -- matikan Sort/Filter dulu sebelum diurutkan ulang",
        None, None, None, None, None,
    ])

    # --- Baris 2: header + saldo awal (seed utk formula cek baris pertama) ---
    saldo_awal = 0
    if not df_bank.empty and "saldo" in df_bank.columns:
        baris0 = df_bank.iloc[0]
        s0 = baris0.get("saldo")
        if pd.notna(s0):
            d0 = _angka(baris0.get("mutasi_debet"))  # [FIX] NaN-safe
            k0 = _angka(baris0.get("mutasi_kredit"))  # [FIX] NaN-safe
            saldo_awal = float(s0) - k0 + d0

    ws.append([
        "No", "Tanggal", "Keterangan", "Mutasi Debet", "Mutasi Kredit", "Saldo Akhir",
        saldo_awal, "Cek", "Supplier/Cust", "Voucher",
        "No Akun", "Nama Akun", "Debet", "No Akun", "Nama Akun", "Kredit",
    ])

    jumlah_belum_dikategorikan = 0
    jumlah_kode_tidak_ditemukan = 0

    for i, (_, row) in enumerate(df_bank.iterrows(), start=1):
        r = i + 2  # nomor baris Excel (data mulai baris 3)

        voucher = row.get("voucher")
        if voucher in (None, "") or (isinstance(voucher, float) and pd.isna(voucher)):
            # [BERUBAH - Prioritas #7] TIDAK generate nomor baru di sini.
            # Baris tanpa voucher berarti belum ditarik ke antrean
            # jurnal_posting (lihat db_client.tarik_draf_jurnal_ke_posting,
            # biasanya karena no_akun_debet/no_akun_kredit masih kosong) --
            # akan otomatis dapat nomor permanen begitu baris itu berhasil
            # dikategorikan & masuk antrean posting; export berikutnya akan
            # menampilkan nomor aslinya.
            voucher = PLACEHOLDER_VOUCHER

        no_akun_d = row.get("no_akun_debet")
        no_akun_k = row.get("no_akun_kredit")
        ada_d = pd.notna(no_akun_d)
        ada_k = pd.notna(no_akun_k)

        saldo_val = row.get("saldo")

        ws.append([
            "=ROW()-ROW($A$2)",
            row.get("tanggal"),
            row.get("keterangan"),
            _angka(row.get("mutasi_debet")),  # [FIX] NaN-safe
            _angka(row.get("mutasi_kredit")),  # [FIX] NaN-safe
            float(saldo_val) if pd.notna(saldo_val) else None,
            f"=G{r - 1}+E{r}-D{r}",
            f"=F{r}-G{r}",
            row.get("supplier_cust") or "",
            voucher,
            no_akun_d if ada_d else None,
            f'=IFERROR(VLOOKUP(K{r},COA!$B:$C,2,0),"BELUM TERKATEGORI")' if ada_d else "",
            f"=D{r}+E{r}",
            no_akun_k if ada_k else None,
            f'=IFERROR(VLOOKUP(N{r},COA!$B:$C,2,0),"BELUM TERKATEGORI")' if ada_k else "",
            f"=M{r}",
        ])

        # --- [BARU - Prioritas #5] Tentukan status & highlight baris ---
        kode_d_nyasar = (
            ada_d and kode_akun_valid is not None and _normalisasi_kode_akun(no_akun_d) not in kode_akun_valid
        )
        kode_k_nyasar = (
            ada_k and kode_akun_valid is not None and _normalisasi_kode_akun(no_akun_k) not in kode_akun_valid
        )

        if kode_d_nyasar or kode_k_nyasar:
            status_fill = FILL_KODE_TIDAK_DITEMUKAN
            jumlah_kode_tidak_ditemukan += 1
        elif not ada_d or not ada_k:
            status_fill = FILL_BELUM_DIKATEGORI
            jumlah_belum_dikategorikan += 1
        else:
            status_fill = None

        if status_fill is not None:
            for c in range(1, JUMLAH_KOLOM_SHEET_BANK + 1):
                ws.cell(row=r, column=c).fill = status_fill

    for r in range(3, baris_terakhir + 1):
        ws.cell(row=r, column=2).number_format = "DD/MM/YYYY"
        for c in (4, 5, 6, 7, 8, 13, 16):
            ws.cell(row=r, column=c).number_format = "#,##0"

    # --- [BARU - Prioritas #5] Freeze header + auto-filter, supaya akuntan
    # bisa filter kolom Nama Akun (kosong / "BELUM TERKATEGORI") atau warna
    # tanpa scroll manual satu-satu di antara ribuan baris.
    ws.freeze_panes = "A3"
    if n > 0:
        ws.auto_filter.ref = f"A2:P{baris_terakhir}"

    # --- Update catatan di baris 1 dengan ringkasan jumlah baris bermasalah ---
    ws.cell(row=1, column=11).value = (
        "VOUCHER & FORMULA CEK -- matikan Sort/Filter dulu sebelum diurutkan ulang | "
        f"{jumlah_belum_dikategorikan} baris kuning (belum dikategorikan), "
        f"{jumlah_kode_tidak_ditemukan} baris merah (kode akun tidak ditemukan di COA)"
    )

    return {
        "belum_dikategorikan": jumlah_belum_dikategorikan,
        "kode_tidak_ditemukan": jumlah_kode_tidak_ditemukan,
    }


def export_rekening_koran_format_akuntan(
    df_hasil: pd.DataFrame,
    df_coa: Optional[pd.DataFrame] = None,
    df_piutang: Optional[pd.DataFrame] = None,
) -> bytes:
    """
    Export df_hasil (keluaran ak.proses_file_rekening_koran()["df"]) jadi
    file Excel format kerja akuntan: 1 sheet "COA" + 1 sheet per bank,
    dengan nomor voucher, formula VLOOKUP ke COA, dan formula cek saldo
    berjalan -- meniru format file kerja rekonsiliasi bank yang dipakai
    tim akuntan (bukan dump data mentah seperti _buat_excel_hasil di
    main.py, yang menulis Ringkasan/Perlu Review/Draf Jurnal sebagai
    teks datar tanpa formula).

    df_hasil wajib punya kolom (lihat parse_sheet_bank() / proses_dataframe()
    di akuntansi_ai.py): bank, tanggal, keterangan, mutasi_debet,
    mutasi_kredit, saldo, supplier_cust, voucher, no_akun_debet,
    nama_akun_debet, no_akun_kredit, nama_akun_kredit.

    [BERUBAH - Prioritas #7] Fungsi ini TIDAK LAGI menerima client_id dan
    TIDAK LAGI mengambil/menggenerate nomor voucher dari database.
    Penomoran voucher rekening_koran sekarang SELURUHNYA terjadi sekali,
    saat upload, di db_client.tarik_draf_jurnal_ke_posting() (dipanggil
    main.py setelah /api/proses-file), lalu disimpan permanen di tabel
    jurnal_posting. Pemanggil fungsi INI (lihat endpoint
    /api/client/{client_id}/rekening-koran/export-format-akuntan/{hasil_id}
    di main.py) bertanggung jawab mengisi df_hasil["voucher"] dari
    jurnal_posting SEBELUM memanggil fungsi ini -- baris yang voucher-nya
    masih kosong di sini (belum masuk antrean posting) akan ditulis dengan
    PLACEHOLDER_VOUCHER, BUKAN nomor baru, supaya:
      1. tidak ada 2 sumber independen yang menarik dari counter
         voucher_counter yang sama (yang tadinya bisa bikin nomor beda
         antara Excel hasil export vs tabel jurnal_posting), dan
      2. file Excel yang sama bisa diexport ulang kapan saja (idempoten)
         tanpa efek samping ke database.

    df_piutang opsional: kalau dikasih (DataFrame keluaran
    ak.proses_file_piutang()["df"]), baris uang MASUK (mutasi_kredit > 0)
    yang belum ada nama supplier/cust-nya akan dicoba dicocokkan ke Buku
    Bantu Piutang berdasar nominal + kemiripan nama (lihat
    _cocokkan_supplier_opsional -- perhatikan catatan soal bug
    cross_matching.cocokkan_bank_piutang() di sana).

    Return bytes file Excel siap didownload/dikirim sbg StreamingResponse.
    """
    if df_hasil is None or df_hasil.empty:
        raise ValueError("df_hasil kosong -- tidak ada data rekening koran untuk diexport.")
    if "bank" not in df_hasil.columns:
        raise ValueError("df_hasil tidak punya kolom 'bank' -- pastikan ini keluaran proses_file_rekening_koran().")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _tulis_sheet_coa(wb, df_coa)

    # [BARU - Prioritas #5] Kumpulkan semua no_akun yang ADA di COA (sbg
    # string, supaya perbandingan tidak gagal gara-gara int vs str) --
    # dipakai _tulis_sheet_bank() utk deteksi kode akun "nyasar" (highlight
    # merah) SEBELUM file dibuka di Excel.
    kode_akun_valid = None
    if df_coa is not None and not df_coa.empty and "no_akun" in df_coa.columns:
        kode_akun_valid = {
            k for k in (_normalisasi_kode_akun(v) for v in df_coa["no_akun"]) if k is not None
        }

    total_belum_dikategorikan = 0
    total_kode_tidak_ditemukan = 0
    total_voucher_kosong = 0
    for nama_bank, df_bank in df_hasil.groupby("bank", sort=False):
        df_bank_final = _cocokkan_supplier_opsional(df_bank, df_piutang)

        voucher_kosong = df_bank_final["voucher"].isna() | (
            df_bank_final["voucher"].astype(str).str.strip() == ""
        )
        total_voucher_kosong += int(voucher_kosong.sum())

        ringkasan_sheet = _tulis_sheet_bank(wb, nama_bank, df_bank_final, kode_akun_valid)
        total_belum_dikategorikan += ringkasan_sheet["belum_dikategorikan"]
        total_kode_tidak_ditemukan += ringkasan_sheet["kode_tidak_ditemukan"]

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    logger.info(
        f"✅ Export format akuntan (rekening koran): "
        f"{df_hasil['bank'].nunique()} bank, {len(df_hasil)} transaksi; "
        f"{total_voucher_kosong} baris belum punya voucher permanen (masih '{PLACEHOLDER_VOUCHER}', "
        "belum masuk antrean jurnal_posting); "
        f"{total_belum_dikategorikan} baris belum dikategorikan (kuning), "
        f"{total_kode_tidak_ditemukan} baris kode akun tidak ditemukan di COA (merah)"
    )
    return buffer.getvalue()

# ============================================================
# [BARU] EXPORT 14 SHEET LENGKAP (model referensi export akuntansi)
# ============================================================
# Sheet-sheet di bawah ini dibangun MURNI dari data yang sudah diambil/
# dihitung sebelumnya (di endpoint POST
# /api/client/{client_id}/export-14-sheet, main.py) -- modul ini sengaja
# TIDAK mengakses db_client sendiri, konsisten dengan desain modul ini
# di tempat lain (lihat catatan soal voucher rekening_koran di atas):
# data masuk lewat parameter, file Excel (bytes) keluar, tidak ada efek
# samping ke database.

_HEADER_FONT_14SHEET = Font(bold=True, size=11)
_HEADER_FILL_14SHEET = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_SUBTOTAL_FONT_14SHEET = Font(bold=True)
_FORMAT_RUPIAH_14SHEET = "#,##0"

# [BARU] Gaya khusus sheet "Petunjuk & Asumsi" -- meniru PERSIS file model
# referensi (MODEL_LAPORAN_KEUANGAN_OTOMATIS...xlsx): judul & baris header
# fill navy #17365D + font putih bold center+wrap, cell input (Modal/Kas/
# Bank/Aset/Parameter) fill kuning #FFF2CC + font biru (sesuai legenda
# "Warna biru = input"), formula sheet-lokal font hitam (sesuai legenda
# "Warna hitam"), judul "TAMBAHAN UNTUK SPT..." fill biru muda #D9EAF7 +
# font navy. SENGAJA dipisah dari _HEADER_FONT_14SHEET/_HEADER_FILL_14SHEET
# (dipakai sheet 14-sheet lain seperti COA) supaya perubahan di sini TIDAK
# memengaruhi sheet lain -- hanya sheet "Petunjuk & Asumsi" yang diminta
# disamakan persis dengan file model referensi.
_TITLE_FONT_ASUMSI = Font(name="Carlito", bold=True, size=15, color="FFFFFFFF")
_TITLE_FILL_ASUMSI = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_TITLE_ALIGN_ASUMSI = Alignment(horizontal="left", vertical="center")
_HEADER_FONT_ASUMSI = Font(name="Carlito", bold=True, size=11, color="FFFFFFFF")
_HEADER_FILL_ASUMSI = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_HEADER_ALIGN_ASUMSI = Alignment(horizontal="center", vertical="center", wrap_text=True)
_INPUT_FONT_ASUMSI = Font(name="Carlito", size=11, color="FF0000FF")
_INPUT_FILL_ASUMSI = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
_FORMULA_FONT_ASUMSI = Font(name="Carlito", size=11, color="FF000000")
_LABEL_ALIGN_ASUMSI = Alignment(vertical="center", wrap_text=True)
_DESC_ALIGN_ASUMSI = Alignment(vertical="center", wrap_text=True)
_SUBTITLE_FONT_ASUMSI = Font(name="Carlito", bold=True, size=11, color="FF17365D")
_SUBTITLE_FILL_ASUMSI = PatternFill(start_color="FFD9EAF7", end_color="FFD9EAF7", fill_type="solid")
_FORMAT_RUPIAH_ASUMSI = r"#,##0;[Red]\(#,##0\);\-"
_FORMAT_TANGGAL_ASUMSI = r"dd\-mmm\-yyyy"
_FORMAT_INT_ASUMSI = "0"
_LEBAR_KOLOM_ASUMSI = {"A": 84.57, "B": 41.0, "C": 42.29, "D": 42.57, "E": 44.14, "F": 61.29, "G": 46.71, "H": 62.14}
# [BARU] Garis kotak tipis abu-abu (#A6A6A6) yang mengelilingi tiap tabel
# di sheet ini -- ditemukan dgn audit ulang file model referensi
# (sebelumnya TIDAK ADA border sama sekali di helper2 sheet ini). Pola di
# file model: baris header dapat top+bottom di semua kolomnya + left di
# kolom pertama + right di kolom terakhir; baris isi/data cuma dapat left
# di kolom pertama + right di kolom terakhir (dinding kiri-kanan kotak,
# tanpa langit2/lantai); baris TERAKHIR suatu tabel dapat tambahan bottom
# di kolom pertama & terakhir utk menutup kotaknya.
_GARIS_ASUMSI = Side(style="thin", color="FFA6A6A6")

# [BARU] Gaya khusus sheet 12 "BS Lampiran SPT" ("NERACA -- LAMPIRAN SPT
# TAHUNAN BADAN (DALAM RUPIAH)") -- diaudit LANGSUNG dari file model
# referensi (NERACA___LAMPIRAN_SPT_TAHUNAN.xlsx) sel demi sel: font Carlito
# di seluruh sel angka & header, tapi kolom Kode/Uraian/Keterangan pada
# baris item biasa pakai Calibri (BUKAN Carlito -- beda dari sheet lain).
# SENGAJA dipisah dari _HEADER_FONT_14SHEET/dst (dipakai sheet 14-sheet
# lain) & dari _TITLE_FONT_ASUMSI/dst (dipakai sheet "Petunjuk & Asumsi")
# supaya perubahan di sini TIDAK memengaruhi sheet lain -- hanya sheet 12
# yang diminta disamakan persis 100% dgn file referensi.
_NERACA_TITLE_FONT = Font(name="Carlito", bold=True, size=15, color="FFFFFFFF")
_NERACA_TITLE_FILL = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_NERACA_TITLE_ALIGN = Alignment(horizontal="left", vertical="center")

_NERACA_HEADER_FONT = Font(name="Carlito", bold=True, size=11, color="FFFFFFFF")
_NERACA_HEADER_FILL = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_NERACA_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)

_NERACA_KATEGORI_FONT = Font(name="Carlito", bold=True, size=11, color="FF17365D")
_NERACA_KATEGORI_FILL = PatternFill(start_color="FFD9EAF7", end_color="FFD9EAF7", fill_type="solid")

_NERACA_SUBTOTAL_FONT = Font(name="Carlito", bold=True, size=11, color="FF000000")
_NERACA_SUBTOTAL_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")

_NERACA_ITEM_FONT_LABEL = Font(name="Calibri", size=11)  # kolom A/B/E baris item biasa
_NERACA_ITEM_FONT_HITUNG = Font(name="Carlito", size=11, color="FF008000")  # nilai dari perhitungan (hijau)
_NERACA_ITEM_FONT_MANUAL = Font(name="Carlito", size=11, color="FF0000FF")  # nilai input manual (biru)
_NERACA_ITEM_FILL_MANUAL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")

_NERACA_FORMAT_RUPIAH = r"#,##0;[Red]\(#,##0\);\-"
_NERACA_LEBAR_KOLOM = {"A": 6.43, "B": 37.29, "C": 19.71, "D": 32.71, "E": 37.43}

_GARIS_NERACA_THIN = Side(style="thin", color="FF000000")
_GARIS_NERACA_DOUBLE = Side(style="double", color="FF000000")


def _border_neraca_header(kolom: int) -> Border:
    """Border baris header (row 3): top+bottom thin semua kolom, + left di
    kolom pertama (A) & right di kolom terakhir (E), persis file referensi."""
    return Border(
        top=_GARIS_NERACA_THIN, bottom=_GARIS_NERACA_THIN,
        left=_GARIS_NERACA_THIN if kolom == 1 else None,
        right=_GARIS_NERACA_THIN if kolom == 5 else None,
    )


def _border_neraca_kategori(kolom: int) -> Border:
    """Border baris kategori/header_pertama: bottom thin semua kolom, + left
    di kolom A, right di kolom E."""
    return Border(
        bottom=_GARIS_NERACA_THIN,
        left=_GARIS_NERACA_THIN if kolom == 1 else None,
        right=_GARIS_NERACA_THIN if kolom == 5 else None,
    )


def _border_neraca_subtotal(kolom: int) -> Border:
    """Border baris subtotal/JUMLAH/CHECK BALANCE: top thin + bottom double
    semua kolom, + left di kolom A, right di kolom E."""
    return Border(
        top=_GARIS_NERACA_THIN, bottom=_GARIS_NERACA_DOUBLE,
        left=_GARIS_NERACA_THIN if kolom == 1 else None,
        right=_GARIS_NERACA_THIN if kolom == 5 else None,
    )


def _border_neraca_item(kolom: int) -> Border:
    """Border baris item biasa: cuma dinding kiri (kolom A) & kanan (kolom E),
    tanpa top/bottom -- persis file referensi."""
    if kolom == 1:
        return Border(left=_GARIS_NERACA_THIN)
    if kolom == 5:
        return Border(right=_GARIS_NERACA_THIN)
    return Border()


# [BARU] Gaya khusus sheet 14 "PPh Badan 31E" -- diaudit LANGSUNG dari file
# referensi (PERHITUNGAN_PPh_BADAN___TARIF_PASAL_17_DAN_FASILITAS_PASAL_31E.xlsx)
# sel demi sel: judul merge A1:F1 navy putih, header row (baris 3) navy putih,
# baris "section header" (Tarif PPh Badan Umum / Peredaran Bruto Usaha dari
# PNL / Laba Bersih Komersial / Kredit Pajak PPh Pasal 22) navy bold fill biru
# muda #D9EAF7, baris subtotal (TOTAL PEREDARAN BRUTO dst) bold hitam fill
# #D9E1F2 (BEDA dari #D9EAF7 section header di atas -- jangan tertukar) +
# border atas tipis/bawah ganda (bawah tipis saja di baris PENUTUP terakhir,
# row 41), sel input manual (Tahun Pajak/Skema Pajak/Tambahan Peredaran Bruto
# Lainnya/Retur/Kredit Pajak Ps.23/Ps.24/Angsuran Ps.25) font biru + fill
# kuning #FFF2CC, sel hasil formula/link (dari sheet lain atau dari sel lain
# sheet ini) font hijau. SENGAJA dipisah dari _HEADER_FONT_14SHEET/dst &
# _NERACA_*/_TITLE_FONT_ASUMSI (dipakai sheet lain) supaya perubahan di sini
# TIDAK memengaruhi sheet lain -- hanya sheet 14 yang diminta disamakan
# persis dengan file referensi.
_PPH31E_TITLE_FONT = Font(name="Carlito", bold=True, size=15, color="FFFFFFFF")
_PPH31E_TITLE_FILL = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_PPH31E_TITLE_ALIGN = Alignment(horizontal="left", vertical="center")

_PPH31E_HEADER_FONT = Font(name="Carlito", bold=True, size=11, color="FFFFFFFF")
_PPH31E_HEADER_FILL = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_PPH31E_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_PPH31E_SECTION_FONT = Font(name="Carlito", bold=True, size=11, color="FF17365D")
_PPH31E_SECTION_FILL = PatternFill(start_color="FFD9EAF7", end_color="FFD9EAF7", fill_type="solid")

_PPH31E_SUBTOTAL_FONT = Font(name="Carlito", bold=True, size=11, color="FF000000")
_PPH31E_SUBTOTAL_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")

_PPH31E_LABEL_FONT = Font(name="Calibri", size=11)  # kolom A/E/F baris item biasa
_PPH31E_LINK_FONT = Font(name="Carlito", size=11, color="FF008000")  # hijau -- hasil formula/link
_PPH31E_INPUT_FONT = Font(name="Carlito", size=11, color="FF0000FF")  # biru -- input manual
_PPH31E_INPUT_FILL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
_PPH31E_STATUS_FONT = Font(name="Calibri", size=11)  # baris "Status Fasilitas Pasal 31E" -- plain, tidak diwarnai
_PPH31E_WRAP_ALIGN = Alignment(wrap_text=True, vertical="center")

_PPH31E_FORMAT_RUPIAH = r"#,##0;[Red]\(#,##0\);\-"
_PPH31E_FORMAT_PERSEN = "0.00%"
_PPH31E_LEBAR_KOLOM = {"A": 58.57, "B": 39.43, "C": 17.29, "D": 15.14, "E": 59.29, "F": 106.57}

_GARIS_PPH31E_THIN = Side(style="thin", color="FF000000")
_GARIS_PPH31E_DOUBLE = Side(style="double", color="FF000000")

# [BARU] Nama skema pajak yang berhak atas fasilitas Pasal 17/31E -- HARUS
# sama persis dengan modules.pph_badan.SKEMA_TARIF_UMUM_31E (disalin sbg
# string literal, bukan import, supaya modul ini tetap murni/tidak
# menambah dependency baru -- lihat catatan "modul ini sengaja TIDAK
# mengakses db_client sendiri" di atas). Dipakai dalam formula IF() sel
# B25/B26/B27/D31 persis seperti file referensi.
_PPH31E_SKEMA_TARIF_UMUM = "Tarif Umum Pasal 17/31E"


def _border_pph31e_header(kolom: int) -> Border:
    """Border baris header (row 3): top+bottom thin semua kolom, + left di
    kolom pertama (A) & right di kolom terakhir (F), persis file referensi."""
    return Border(
        top=_GARIS_PPH31E_THIN, bottom=_GARIS_PPH31E_THIN,
        left=_GARIS_PPH31E_THIN if kolom == 1 else None,
        right=_GARIS_PPH31E_THIN if kolom == 6 else None,
    )


def _border_pph31e_item(kolom: int) -> Border:
    """Border baris item biasa: dinding kiri (kolom A) & kanan (kolom F),
    tanpa top/bottom -- persis file referensi."""
    if kolom == 1:
        return Border(left=_GARIS_PPH31E_THIN)
    if kolom == 6:
        return Border(right=_GARIS_PPH31E_THIN)
    return Border()


def _border_pph31e_section(kolom: int) -> Border:
    """Border baris section header (Tarif PPh Badan Umum dst): bottom thin
    semua kolom, + left di A, right di F."""
    return Border(
        bottom=_GARIS_PPH31E_THIN,
        left=_GARIS_PPH31E_THIN if kolom == 1 else None,
        right=_GARIS_PPH31E_THIN if kolom == 6 else None,
    )


def _border_pph31e_subtotal(kolom: int, terakhir: bool = False) -> Border:
    """Border baris subtotal (TOTAL PEREDARAN BRUTO dst): top thin + bottom
    double, KECUALI baris penutup terakhir (row 41, PPh Pasal 28A) yang
    cuma top+bottom thin -- persis file referensi. + left di A, right di F."""
    return Border(
        top=_GARIS_PPH31E_THIN,
        bottom=_GARIS_PPH31E_THIN if terakhir else _GARIS_PPH31E_DOUBLE,
        left=_GARIS_PPH31E_THIN if kolom == 1 else None,
        right=_GARIS_PPH31E_THIN if kolom == 6 else None,
    )


def _tulis_sheet_pph_badan_31e(ws, pph_hasil: Dict[str, Any]) -> None:
    """[BARU] Penulis khusus sheet 14 "PPh Badan 31E" -- meniru PERSIS
    struktur & gaya file referensi (PERHITUNGAN_PPh_BADAN___TARIF_PASAL_17_
    DAN_FASILITAS_PASAL_31E.xlsx) sel demi sel, TERMASUK menulis FORMULA
    Excel hidup (bukan angka mentah hasil hitung Python) untuk setiap sel
    yang murni turunan dari sel LAIN DI SHEET INI SENDIRI -- tarif efektif
    fasilitas (B10), total peredaran bruto (B17), penghasilan neto fiskal
    (B21), PKP sebelum pembulatan (B23), PKP dibulatkan (B24), status &
    pembagian fasilitas 31E (B25/B26/B27), PPh atas PKP fasilitas/non (D28/
    D29), PPh badan terutang (D30), PPh tanpa fasilitas (D31), penghematan
    pajak (D32), tarif efektif riil (D33), total kredit pajak (D39), PPh
    Pasal 29/28A (D40/D41) -- sesuai aturan "gunakan formula, jangan angka
    mentah" utk model keuangan: kalau akuntan mengedit sel input manual
    (kuning) langsung di Excel, seluruh sheet ikut menghitung ulang dengan
    benar, persis seperti file referensi yang formula-nya diaudit.

    Nilai yang SUMBER ASLINYA dari LUAR sheet ini (Peredaran Bruto Usaha
    dari PNL, Laba Bersih Komersial, Koreksi Fiskal Positif/Negatif,
    Kompensasi Kerugian Fiskal, Kredit Pajak PPh Pasal 22 -- semua sudah
    dihitung modul lain & diterima di sini murni lewat parameter pph_hasil,
    BUKAN oleh rumus milik sheet ini) ditulis sebagai nilai + font hijau
    (konvensi "link" file referensi), KECUALI kalau memang sel input manual
    akuntan (font biru + fill kuning #FFF2CC).

    [FIX] Dua koreksi disengaja terhadap file referensi (bukan salah tik):
    (1) B11/B12 (Batas Peredaran Bruto ...) di file referensi punya
    number_format "0.00%" yang salah tempel dari baris di atasnya (tampil
    "480000000000,00%" -- rusak/tidak terbaca) -- di sini diberi format
    rupiah yang benar, sama seperti C11/C12 di sebelahnya.
    (2) baris "PPh Pasal 29"/"PPh Pasal 28a" di kode LAMA sudah ada duluan
    (label sedikit beda kapitalisasi) -- di sini label disamakan PERSIS ke
    file referensi ("PPh PASAL 29 – KURANG BAYAR" / "PPh PASAL 28A – LEBIH
    BAYAR", huruf besar semua + em dash, gaya subtotal fill #D9E1F2)."""
    fasilitas = pph_hasil.get("fasilitas_31e", {}) or {}
    rekon = pph_hasil.get("rekonsiliasi_fiskal", {}) or {}
    kredit = pph_hasil.get("kredit_pajak", {}) or {}
    peredaran_detail = pph_hasil.get("peredaran_bruto_detail", {}) or {}
    skema_pajak = pph_hasil.get("skema_pajak") or _PPH31E_SKEMA_TARIF_UMUM
    persentase_pengurangan = pph_hasil.get("persentase_pengurangan_pasal_31e", 0.5)
    tarif_umum = pph_hasil.get("tarif_pph_badan_umum", 0.22)
    batas1 = pph_hasil.get("batas_peredaran_bruto_fasilitas_penuh", 4_800_000_000)
    batas2 = pph_hasil.get("batas_maks_peredaran_bruto_fasilitas", 50_000_000_000)
    keterangan_bruto = pph_hasil.get("keterangan_peredaran_bruto") or "Pendapatan usaha setahun (dari Laba Rugi)"

    SUMBER_UU = "https://www.pajak.go.id/id/uu-hpp"
    SUMBER_31E = "https://www.pajak.go.id/id/artikel/lima-jenis-tarif-pph-badan-yang-wajib-diperhatikan"
    SUMBER_LAMPIRAN_VIII = "https://www.pajak.go.id/sites/default/files/2019-03/Lampiran%20VIII%20Petunjuk%20Pengisian%201771.pdf"
    SUMBER_UU_HPP_SALINAN = "https://www.pajak.go.id/sites/default/files/2021-12/Salinan%20UU%20Nomor%207%20Tahun%202021.pdf"

    def _c(row, col, value=None, font=None, fill=None, numfmt=None, align=None, border=None):
        cell = ws.cell(row=row, column=col)
        if value is not None:
            cell.value = value
        if font is not None:
            cell.font = font
        if fill is not None:
            cell.fill = fill
        if numfmt is not None:
            cell.number_format = numfmt
        if align is not None:
            cell.alignment = align
        if border is not None:
            cell.border = border
        return cell

    def _section_row(row, label, value_b=None, e=None, f=None, numfmt_b=_PPH31E_FORMAT_RUPIAH):
        for col in range(1, 7):
            _c(row, col, font=_PPH31E_SECTION_FONT, fill=_PPH31E_SECTION_FILL,
               border=_border_pph31e_section(col))
        _c(row, 1, label)
        if value_b is not None:
            _c(row, 2, value_b, numfmt=numfmt_b)
        if e is not None:
            _c(row, 5, e, align=_PPH31E_WRAP_ALIGN)
        if f is not None:
            _c(row, 6, f, align=_PPH31E_WRAP_ALIGN)

    def _subtotal_row(row, label, col_hasil, formula, terakhir=False, f=None):
        for col in range(1, 7):
            _c(row, col, font=_PPH31E_SUBTOTAL_FONT, fill=_PPH31E_SUBTOTAL_FILL,
               border=_border_pph31e_subtotal(col, terakhir))
        _c(row, 1, label)
        _c(row, col_hasil, formula, numfmt=_PPH31E_FORMAT_RUPIAH)
        if f is not None:
            _c(row, 6, f, align=_PPH31E_WRAP_ALIGN)

    def _item_border(row):
        _c(row, 1, border=_border_pph31e_item(1))
        _c(row, 6, border=_border_pph31e_item(6))

    # ---- Judul (row 1, merge A1:F1) ----
    ws.merge_cells("A1:F1")
    _c(1, 1, "PERHITUNGAN PPh BADAN – TARIF PASAL 17 DAN FASILITAS PASAL 31E",
       font=_PPH31E_TITLE_FONT, fill=_PPH31E_TITLE_FILL, align=_PPH31E_TITLE_ALIGN)
    ws.row_dimensions[1].height = 27.95

    # row 2: kosong

    # ---- Header (row 3) ----
    headers = ["Parameter / Perhitungan", "Nilai", "Tarif / Batas", "Hasil",
               "Status / Keterangan", "Sumber Resmi"]
    for col, h in enumerate(headers, 1):
        _c(3, col, h, font=_PPH31E_HEADER_FONT, fill=_PPH31E_HEADER_FILL,
           align=_PPH31E_HEADER_ALIGN, border=_border_pph31e_header(col))
    ws.row_dimensions[3].height = 30

    # ---- Identitas (4-6) ----
    _c(4, 1, "Nama Perusahaan")
    _c(4, 2, pph_hasil.get("nama_perusahaan", ""), font=_PPH31E_LINK_FONT)
    _item_border(4)

    _c(5, 1, "Tahun Pajak")
    _c(5, 2, pph_hasil.get("tahun_pajak", ""), font=_PPH31E_INPUT_FONT, fill=_PPH31E_INPUT_FILL)
    _item_border(5)

    _c(6, 1, "Skema Pajak")
    _c(6, 2, skema_pajak, font=_PPH31E_INPUT_FONT, fill=_PPH31E_INPUT_FILL)
    _c(6, 5, "Ubah jika perusahaan masih menggunakan PPh Final")
    _item_border(6)
    # row 7: kosong

    # ---- Tarif (8-12) ----
    _section_row(8, "Tarif PPh Badan Umum", "=C8", "Tarif umum", SUMBER_UU, numfmt_b=_PPH31E_FORMAT_PERSEN)
    _c(8, 3, tarif_umum, numfmt=_PPH31E_FORMAT_PERSEN)

    _c(9, 1, "Persentase Pengurangan Tarif Pasal 31E")
    _c(9, 2, "=C9", font=_PPH31E_LINK_FONT, numfmt=_PPH31E_FORMAT_PERSEN)
    _c(9, 3, persentase_pengurangan, numfmt=_PPH31E_FORMAT_PERSEN)
    _c(9, 5, "Pengurangan 50% dari tarif umum")
    _c(9, 6, SUMBER_31E, align=_PPH31E_WRAP_ALIGN)
    _item_border(9)

    _c(10, 1, "Tarif Efektif Fasilitas")
    _c(10, 2, "=B8*B9", font=_PPH31E_LINK_FONT, numfmt=_PPH31E_FORMAT_PERSEN)
    _c(10, 5, f"{persentase_pengurangan * 100:.0f}% × {tarif_umum * 100:.0f}% = "
              f"{persentase_pengurangan * tarif_umum * 100:.0f}%")
    _c(10, 6, SUMBER_31E, align=_PPH31E_WRAP_ALIGN)
    _item_border(10)

    # [FIX] B11/B12: rupiah yang benar (lihat catatan FIX di docstring), bukan
    # "0.00%" yang rusak seperti di file referensi.
    _c(11, 1, "Batas Peredaran Bruto yang Mendapat Fasilitas")
    _c(11, 2, "=C11", font=_PPH31E_LINK_FONT, numfmt=_PPH31E_FORMAT_RUPIAH)
    _c(11, 3, batas1, numfmt=_PPH31E_FORMAT_RUPIAH)
    _c(11, 5, "Bagian omzet sampai Rp4,8 miliar")
    _c(11, 6, SUMBER_31E, align=_PPH31E_WRAP_ALIGN)
    _item_border(11)

    _c(12, 1, "Batas Maksimum Peredaran Bruto WP Penerima Fasilitas")
    _c(12, 2, "=C12", font=_PPH31E_LINK_FONT, numfmt=_PPH31E_FORMAT_RUPIAH)
    _c(12, 3, batas2, numfmt=_PPH31E_FORMAT_RUPIAH)
    _c(12, 5, "Di atas Rp50 miliar tidak mendapat fasilitas")
    _c(12, 6, SUMBER_31E, align=_PPH31E_WRAP_ALIGN)
    _item_border(12)
    # row 13: kosong

    # ---- Peredaran bruto (14-17) ----
    _section_row(14, "Peredaran Bruto Usaha dari PNL",
                 peredaran_detail.get("peredaran_bruto_usaha", 0), e=keterangan_bruto)

    _c(15, 1, "Tambahan Peredaran Bruto Lainnya untuk Pasal 31E")
    _c(15, 2, peredaran_detail.get("tambahan_peredaran_bruto_lainnya", 0),
       font=_PPH31E_INPUT_FONT, fill=_PPH31E_INPUT_FILL, numfmt=_PPH31E_FORMAT_RUPIAH)
    _c(15, 5, "Input jika ada penghasilan lain yang termasuk peredaran bruto")
    _item_border(15)

    _c(16, 1, "Retur/Pengurangan Peredaran Bruto")
    _c(16, 2, peredaran_detail.get("retur_pengurangan_peredaran_bruto", 0),
       font=_PPH31E_INPUT_FONT, fill=_PPH31E_INPUT_FILL, numfmt=_PPH31E_FORMAT_RUPIAH)
    _c(16, 5, "Input sebagai nilai positif")
    _item_border(16)

    _subtotal_row(17, "TOTAL PEREDARAN BRUTO", 2, "=MAX(0,B14+B15-B16)")
    ws.row_dimensions[17].height = 15.75

    # ---- Laba fiskal (18-24) ----
    _section_row(18, "Laba Bersih Komersial", rekon.get("laba_bersih_komersial", 0),
                 f=SUMBER_LAMPIRAN_VIII)
    ws.row_dimensions[18].height = 15.75

    _c(19, 1, "Koreksi Fiskal Positif")
    _c(19, 2, rekon.get("koreksi_fiskal_positif", 0), font=_PPH31E_LINK_FONT, numfmt=_PPH31E_FORMAT_RUPIAH)
    _c(19, 6, SUMBER_UU_HPP_SALINAN, align=_PPH31E_WRAP_ALIGN)
    _item_border(19)

    _c(20, 1, "Koreksi Fiskal Negatif")
    _c(20, 2, rekon.get("koreksi_fiskal_negatif", 0), font=_PPH31E_LINK_FONT, numfmt=_PPH31E_FORMAT_RUPIAH)
    _c(20, 6, SUMBER_UU_HPP_SALINAN, align=_PPH31E_WRAP_ALIGN)
    _item_border(20)

    _subtotal_row(21, "PENGHASILAN NETO FISKAL", 2, "=B18+B19-B20", f=SUMBER_UU_HPP_SALINAN)
    ws.row_dimensions[21].height = 30.75

    _c(22, 1, "Kompensasi Kerugian Fiskal")
    _c(22, 2, rekon.get("kompensasi_kerugian_fiskal", 0), font=_PPH31E_LINK_FONT, numfmt=_PPH31E_FORMAT_RUPIAH)
    _c(22, 5, "Terhubung dari rekonsiliasi fiskal")
    _c(22, 6, SUMBER_LAMPIRAN_VIII, align=_PPH31E_WRAP_ALIGN)
    _item_border(22)
    ws.row_dimensions[22].height = 15.75

    _c(23, 1, "PKP Sebelum Pembulatan")
    _c(23, 2, "=MAX(0,B21-B22)", font=_PPH31E_LINK_FONT, numfmt=_PPH31E_FORMAT_RUPIAH)
    _c(23, 6, SUMBER_LAMPIRAN_VIII, align=_PPH31E_WRAP_ALIGN)
    _item_border(23)

    _subtotal_row(24, "PENGHASILAN KENA PAJAK – RIBUAN PENUH", 2, "=INT(B23/1000)*1000",
                  f=SUMBER_LAMPIRAN_VIII)
    ws.row_dimensions[24].height = 30.75

    # ---- Fasilitas 31E (25-27) ----
    status_formula = (
        f'=IF(B6<>"{_PPH31E_SKEMA_TARIF_UMUM}","Tidak dihitung – periksa skema pajak",'
        f'IF(B17<=C11,"Seluruh PKP mendapat fasilitas 50% tarif",'
        f'IF(B17<=C12,"Fasilitas proporsional atas bagian omzet Rp4,8 miliar",'
        f'"Tidak memperoleh fasilitas Pasal 31E")))'
    )
    _c(25, 1, "Status Fasilitas Pasal 31E")
    _c(25, 2, status_formula, font=_PPH31E_STATUS_FONT)
    _item_border(25)
    ws.row_dimensions[25].height = 15.75

    fasilitas_formula = (
        f'=IF(B6<>"{_PPH31E_SKEMA_TARIF_UMUM}",0,'
        f'IF(B17<=C11,B24,IF(B17<=C12,INT(B24*C11/B17),0)))'
    )
    _c(26, 1, "PKP yang Mendapat Fasilitas")
    _c(26, 2, fasilitas_formula, font=_PPH31E_LINK_FONT, numfmt=_PPH31E_FORMAT_RUPIAH)
    _item_border(26)

    nonfasilitas_formula = f'=IF(B6<>"{_PPH31E_SKEMA_TARIF_UMUM}",0,MAX(0,B24-B26))'
    _c(27, 1, "PKP yang Tidak Mendapat Fasilitas")
    _c(27, 2, nonfasilitas_formula, font=_PPH31E_LINK_FONT, numfmt=_PPH31E_FORMAT_RUPIAH)
    _item_border(27)

    # ---- PPh terutang (28-33) ----
    _c(28, 1, "PPh atas PKP Fasilitas")
    _c(28, 3, pph_hasil.get("tarif_efektif_fasilitas", 0.11), numfmt=_PPH31E_FORMAT_PERSEN)
    _c(28, 4, "=B26*C28", font=_PPH31E_LINK_FONT, numfmt=_PPH31E_FORMAT_RUPIAH)
    _c(28, 5, "Tarif 11%")
    _c(28, 6, SUMBER_31E, align=_PPH31E_WRAP_ALIGN)
    _item_border(28)

    _c(29, 1, "PPh atas PKP Nonfasilitas")
    _c(29, 3, tarif_umum, numfmt=_PPH31E_FORMAT_PERSEN)
    _c(29, 4, "=B27*C29", font=_PPH31E_LINK_FONT, numfmt=_PPH31E_FORMAT_RUPIAH)
    _c(29, 5, "Tarif 22%")
    _c(29, 6, SUMBER_31E, align=_PPH31E_WRAP_ALIGN)
    _item_border(29)

    _subtotal_row(30, "PPh BADAN TERUTANG", 4, "=D28+D29", f=SUMBER_31E)
    ws.row_dimensions[30].height = 15.75

    _c(31, 1, "PPh Tanpa Fasilitas Pasal 31E")
    _c(31, 3, tarif_umum, numfmt=_PPH31E_FORMAT_PERSEN)
    _c(31, 4, f'=IF(B6<>"{_PPH31E_SKEMA_TARIF_UMUM}",0,B24*C31)', font=_PPH31E_LINK_FONT,
       numfmt=_PPH31E_FORMAT_RUPIAH)
    _c(31, 5, "Sebagai perbandingan")
    _c(31, 6, SUMBER_31E, align=_PPH31E_WRAP_ALIGN)
    _item_border(31)
    ws.row_dimensions[31].height = 15.75

    _subtotal_row(32, "PENGHEMATAN PAJAK PASAL 31E", 4, "=MAX(0,D31-D30)", f=SUMBER_31E)
    ws.row_dimensions[32].height = 15.75

    _c(33, 1, "Tarif Pajak Efektif")
    _c(33, 4, "=IF(B24=0,0,D30/B24)", font=_PPH31E_LINK_FONT, numfmt=_PPH31E_FORMAT_PERSEN)
    _c(33, 6, SUMBER_31E, align=_PPH31E_WRAP_ALIGN)
    _item_border(33)
    ws.row_dimensions[33].height = 15.75
    # row 34: kosong

    # ---- Kredit pajak & SPT akhir (35-41) ----
    _section_row(35, "Kredit Pajak PPh Pasal 22", kredit.get("pph_22", 0),
                 e="Input sesuai bukti pungut")

    _c(36, 1, "Kredit Pajak PPh Pasal 23")
    _c(36, 2, kredit.get("pph_23", 0), font=_PPH31E_INPUT_FONT, fill=_PPH31E_INPUT_FILL,
       numfmt=_PPH31E_FORMAT_RUPIAH)
    _c(36, 5, "Input sesuai bukti potong")
    _item_border(36)

    _c(37, 1, "Kredit Pajak PPh Pasal 24")
    _c(37, 2, kredit.get("pph_24", 0), font=_PPH31E_INPUT_FONT, fill=_PPH31E_INPUT_FILL,
       numfmt=_PPH31E_FORMAT_RUPIAH)
    _c(37, 5, "Input jika ada kredit pajak luar negeri")
    _item_border(37)

    _c(38, 1, "Angsuran PPh Pasal 25")
    _c(38, 2, kredit.get("angsuran_pph_25", 0), font=_PPH31E_INPUT_FONT, fill=_PPH31E_INPUT_FILL,
       numfmt=_PPH31E_FORMAT_RUPIAH)
    _c(38, 5, "Input jumlah pembayaran selama tahun berjalan")
    _item_border(38)

    _subtotal_row(39, "TOTAL KREDIT PAJAK", 4, "=SUM(B35:B38)")
    ws.row_dimensions[39].height = 15.75

    _subtotal_row(40, "PPh PASAL 29 – KURANG BAYAR", 4, "=MAX(0,D30-D39)")
    ws.row_dimensions[40].height = 16.5

    _subtotal_row(41, "PPh PASAL 28A – LEBIH BAYAR", 4, "=MAX(0,D39-D30)", terakhir=True)
    ws.row_dimensions[41].height = 15.75

    # ---- Lebar kolom (persis file referensi, bukan autofit) ----
    for kolom, lebar in _PPH31E_LEBAR_KOLOM.items():
        ws.column_dimensions[kolom].width = lebar
    ws.sheet_view.zoomScale = 75


def _tulis_sheet_neraca_lampiran_spt_baku(ws_bs, neraca_baku: Dict[str, Any], tahun: Any) -> None:
    """[BARU] Penulis khusus sheet 12 "BS Lampiran SPT" -- meniru PERSIS
    struktur & gaya file referensi "NERACA -- LAMPIRAN SPT TAHUNAN BADAN
    (DALAM RUPIAH)" sel demi sel (lihat konstanta _NERACA_* di atas):
    5 kolom (Kode/Uraian/Tahun ini/Tahun lalu-Saldo Awal/Keterangan), judul
    merge A1:E1 fill navy, header row fill navy, baris kategori & item
    pertama liabilitas/ekuitas fill biru muda bold, baris subtotal fill
    biru muda-2 bold border ganda, baris item biasa font Calibri + kolom
    nilai hijau (hitung)/biru+kuning (manual). Dipanggil dari sheet 12 di
    export_14_sheet_lengkap() -- GANTI total penulisan lama yang masih
    pakai gaya generik _HEADER_FONT_14SHEET/_SUBTOTAL_FONT_14SHEET (tidak
    sama dengan file referensi).
    """
    tahun_sebelumnya = neraca_baku.get("tahun_sebelumnya")
    label_kolom_lalu = (
        f"31 Desember {tahun_sebelumnya} / Saldo Awal" if tahun_sebelumnya
        else "Tahun Lalu / Saldo Awal"
    )
    headers = ["Kode", "Uraian", f"31 Desember {tahun}", label_kolom_lalu, "Keterangan"]

    # --- Judul (baris 1, merge A1:E1) ---
    ws_bs.append([f"NERACA \u2013 LAMPIRAN SPT TAHUNAN BADAN (DALAM RUPIAH)"] + [None] * 4)
    ws_bs.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    cell_judul = ws_bs.cell(row=1, column=1)
    cell_judul.font = _NERACA_TITLE_FONT
    cell_judul.fill = _NERACA_TITLE_FILL
    cell_judul.alignment = _NERACA_TITLE_ALIGN
    ws_bs.row_dimensions[1].height = 27.95

    # --- Baris 2 kosong (spt file referensi) ---
    ws_bs.append([None] * 5)

    # --- Header (baris 3) ---
    ws_bs.append(headers)
    for c in range(1, 6):
        cell = ws_bs.cell(row=3, column=c)
        cell.font = _NERACA_HEADER_FONT
        cell.fill = _NERACA_HEADER_FILL
        cell.alignment = _NERACA_HEADER_ALIGN
        cell.border = _border_neraca_header(c)
    ws_bs.row_dimensions[3].height = 30

    # --- Baris isi ---
    for b in neraca_baku["baris"]:
        tipe = b["tipe"]
        r = ws_bs.max_row + 1
        if tipe == "judul":
            ws_bs.append(["", b["uraian"], "", "", ""])
            for c in range(1, 6):
                cell = ws_bs.cell(row=r, column=c)
                cell.font = _NERACA_KATEGORI_FONT
                cell.fill = _NERACA_KATEGORI_FILL
                cell.border = _border_neraca_kategori(c)
        elif tipe in ("subtotal", "check") or b.get("gaya") == "subtotal":
            # [FIX] Baris JUMLAH ASET LANCAR/JUMLAH ASET/JUMLAH LIABILITAS/
            # JUMLAH EKUITAS/JUMLAH LIABILITAS DAN EKUITAS/CHECK BALANCE di
            # file referensi kolom Kode-nya KOSONG (bukan "-") -- yang punya
            # kode cuma A08 "Aset Tetap Neto" (gaya="subtotal" tapi tipe
            # "akun", jadi memang sudah punya field "kode").
            kode = b.get("kode", "")
            ws_bs.append([kode, b["uraian"], b["nilai_ini"], b["nilai_lalu"], b.get("keterangan", "")])
            for c in range(1, 6):
                cell = ws_bs.cell(row=r, column=c)
                cell.font = _NERACA_SUBTOTAL_FONT
                cell.fill = _NERACA_SUBTOTAL_FILL
                cell.border = _border_neraca_subtotal(c)
                if c in (3, 4):
                    cell.number_format = _NERACA_FORMAT_RUPIAH
        elif b.get("gaya") == "header_pertama":
            # [FIX] L01/E01: kolom A/B/E pakai gaya kategori (fill+bold+navy),
            # TAPI kolom nilai C/D tetap ikut warna hijau/biru sesuai sumber
            # data (bukan navy) -- persis file referensi (C19/D19 di sana
            # hijau #008000 walau baris L01 fill-nya biru muda kategori).
            ws_bs.append([b["kode"], b["uraian"], b["nilai_ini"], b["nilai_lalu"], b.get("keterangan", "")])
            manual = b.get("sumber") == "manual"
            for c in (1, 2, 5):
                cell = ws_bs.cell(row=r, column=c)
                cell.font = _NERACA_KATEGORI_FONT
                cell.fill = _NERACA_KATEGORI_FILL
                cell.border = _border_neraca_kategori(c)
            for c in (3, 4):
                cell = ws_bs.cell(row=r, column=c)
                cell.font = Font(
                    name="Carlito", bold=True,
                    color="FF0000FF" if manual else "FF008000",
                )
                cell.fill = _NERACA_ITEM_FILL_MANUAL if manual else _NERACA_KATEGORI_FILL
                cell.border = _border_neraca_kategori(c)
                cell.number_format = _NERACA_FORMAT_RUPIAH
        else:  # baris "akun" biasa
            ws_bs.append([b["kode"], b["uraian"], b["nilai_ini"], b["nilai_lalu"], b.get("keterangan", "")])
            manual = b.get("sumber") == "manual"
            for c in (1, 2, 5):
                cell = ws_bs.cell(row=r, column=c)
                cell.font = _NERACA_ITEM_FONT_LABEL
                cell.border = _border_neraca_item(c)
            for c in (3, 4):
                cell = ws_bs.cell(row=r, column=c)
                cell.font = _NERACA_ITEM_FONT_MANUAL if manual else _NERACA_ITEM_FONT_HITUNG
                cell.number_format = _NERACA_FORMAT_RUPIAH
                cell.border = _border_neraca_item(c)
                if manual:
                    cell.fill = _NERACA_ITEM_FILL_MANUAL

    for kolom, lebar in _lebar_kolom_dari_isi(ws_bs).items():
        ws_bs.column_dimensions[kolom].width = lebar


# [BARU] Gaya khusus sheet 13 "PNL Lampiran SPT" ("LABA RUGI &
# REKONSILIASI FISKAL -- LAMPIRAN SPT TAHUNAN BADAN") -- diaudit LANGSUNG
# dari file model referensi (LABA_RUGI___REKONSILIASI.xlsx) sel demi sel,
# pola SAMA seperti _NERACA_* di atas (sheet 12) tapi 7 kolom, bukan 5,
# dan subtotal row TERAKHIR ("PKP Ribuan Penuh") menutup kotak dgn border
# bottom THIN (bukan double) -- lihat _tulis_sheet_pnl_lampiran_spt_baku().
# SENGAJA dipisah dari _NERACA_*/_HEADER_FONT_14SHEET/dst supaya perubahan
# di sini TIDAK memengaruhi sheet lain -- hanya sheet 13 yang diminta
# disamakan persis 100% dgn file referensi ini.
_PNL_TITLE_FONT = Font(name="Carlito", bold=True, size=15, color="FFFFFFFF")
_PNL_TITLE_FILL = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_PNL_TITLE_ALIGN = Alignment(horizontal="left", vertical="center")

_PNL_HEADER_FONT = Font(name="Carlito", bold=True, size=11, color="FFFFFFFF")
_PNL_HEADER_FILL = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_PNL_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)

_PNL_JUDUL_FONT = Font(name="Carlito", bold=True, size=11, color="FF17365D")
_PNL_JUDUL_FILL = PatternFill(start_color="FFD9EAF7", end_color="FFD9EAF7", fill_type="solid")

_PNL_SUBTOTAL_FONT = Font(name="Carlito", bold=True, size=11, color="FF000000")
_PNL_SUBTOTAL_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")

_PNL_ITEM_FONT_LABEL = Font(name="Calibri", size=11, color="FF000000")  # kolom A/B/G baris akun
_PNL_ITEM_FONT_HIJAU = Font(name="Carlito", size=11, color="FF008000")  # kolom C (komersial, dari perhitungan)
_PNL_ITEM_FONT_HITAM = Font(name="Carlito", size=11, color="FF000000")  # kolom D/E/F
_PNL_ITEM_FILL_PUTIH = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")

_PNL_FORMAT_RUPIAH = r"#,##0;[Red]\(#,##0\);\-"
_PNL_LEBAR_KOLOM = {"A": 6.43, "B": 57.86, "C": 16.43, "D": 16.0, "E": 16.86, "F": 12.43, "G": 75.29}

_GARIS_PNL_THIN = Side(style="thin", color="FF000000")
_GARIS_PNL_DOUBLE = Side(style="double", color="FF000000")


def _border_pnl_header(kolom: int) -> Border:
    """Baris header (row 3): top+bottom thin semua kolom, + left di kolom
    pertama (A/1) & right di kolom terakhir (G/7), persis file referensi."""
    return Border(
        top=_GARIS_PNL_THIN, bottom=_GARIS_PNL_THIN,
        left=_GARIS_PNL_THIN if kolom == 1 else None,
        right=_GARIS_PNL_THIN if kolom == 7 else None,
    )


def _border_pnl_judul(kolom: int) -> Border:
    """Baris judul seksi (mis. 'BEBAN LANGSUNG'): bottom thin semua kolom
    (TANPA top), + left di kolom A, right di kolom G."""
    return Border(
        bottom=_GARIS_PNL_THIN,
        left=_GARIS_PNL_THIN if kolom == 1 else None,
        right=_GARIS_PNL_THIN if kolom == 7 else None,
    )


def _border_pnl_subtotal(kolom: int, tutup_akhir: bool = False) -> Border:
    """Baris subtotal/JUMLAH/EBITDA/dst: top thin + bottom DOUBLE semua
    kolom, + left di kolom A, right di kolom G. tutup_akhir=True utk baris
    PALING BAWAH tabel ('PKP Ribuan Penuh') -- bottom-nya cuma thin (nutup
    kotak), bukan double, persis file referensi."""
    return Border(
        top=_GARIS_PNL_THIN,
        bottom=_GARIS_PNL_THIN if tutup_akhir else _GARIS_PNL_DOUBLE,
        left=_GARIS_PNL_THIN if kolom == 1 else None,
        right=_GARIS_PNL_THIN if kolom == 7 else None,
    )


def _border_pnl_item(kolom: int) -> Border:
    """Baris akun biasa: cuma dinding kiri (kolom A) & kanan (kolom G),
    tanpa top/bottom -- persis file referensi."""
    if kolom == 1:
        return Border(left=_GARIS_PNL_THIN)
    if kolom == 7:
        return Border(right=_GARIS_PNL_THIN)
    return Border()


def _tulis_sheet_pnl_lampiran_spt_baku(ws_pnl, pnl_baku: Dict[str, Any], tahun: Any) -> Dict[str, Any]:
    """[BARU] Penulis khusus sheet 13 "PNL Lampiran SPT" -- meniru PERSIS
    struktur & gaya file referensi "LABA RUGI & REKONSILIASI FISKAL --
    LAMPIRAN SPT TAHUNAN BADAN" sel demi sel (lihat konstanta _PNL_* di
    atas): 7 kolom (Kode/Uraian/Komersial/Koreksi Positif/Koreksi Negatif/
    Fiskal/Keterangan), judul merge A1:G1 fill navy, header row 3 fill
    navy, baris judul seksi fill biru muda bold navy, baris subtotal fill
    biru-abu bold border ganda, baris akun font Calibri (kolom
    A/B/G)+Carlito (kolom C hijau, D/E/F hitam, D/E fill putih).

    Formula (bukan angka statis) ditulis SEBISA mungkin persis pola file
    referensi -- subtotal = SUM() baris akun di atasnya, LABA KOTOR/EBITDA/
    LABA USAHA = selisih dua subtotal, F33=D33/F34=E34/F36=C36, PKP
    SEBELUM PEMBULATAN = MAX(0,F35-F36), PKP RIBUAN PENUH =
    INT(F37/1000)*1000 -- SATU-satunya bagian yang TETAP angka statis
    (bukan formula) adalah kolom Komersial per akun (datang dari hasil
    laporan_keuangan yang sudah dihitung, bukan link ke sheet lain seperti
    file referensi -- lihat catatan di modul laporan_keuangan.py) dan
    TOTAL KOREKSI FISKAL POSITIF/NEGATIF (datang dari parameter
    rekonsiliasi_pkp/pph_badan, sesuai desain yang sudah ada -- lihat
    catatan section 10c di laporan_keuangan.py)."""
    tahun_label = tahun

    # ---- Judul (row 1, merge A1:G1) ----
    ws_pnl.append([f"LABA RUGI & REKONSILIASI FISKAL \u2013 LAMPIRAN SPT TAHUNAN BADAN"])
    cell_judul = ws_pnl.cell(row=1, column=1)
    cell_judul.font = _PNL_TITLE_FONT
    cell_judul.fill = _PNL_TITLE_FILL
    cell_judul.alignment = _PNL_TITLE_ALIGN
    ws_pnl.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws_pnl.row_dimensions[1].height = 27.95

    # ---- Baris 2 kosong (persis file referensi) ----
    ws_pnl.append([None])

    # ---- Header (row 3) ----
    header_row = 3
    headers = ["Kode", "Uraian", f"Komersial {tahun_label}", "Koreksi Positif",
               "Koreksi Negatif", f"Fiskal {tahun_label}", "Keterangan"]
    for col, h in enumerate(headers, 1):
        cell = ws_pnl.cell(row=header_row, column=col, value=h)
        cell.font = _PNL_HEADER_FONT
        cell.fill = _PNL_HEADER_FILL
        cell.alignment = _PNL_HEADER_ALIGN
        cell.border = _border_pnl_header(col)
    ws_pnl.row_dimensions[header_row].height = 30
    # ---- Baris data ----
    # Lacak baris awal/akhir tiap seksi akun (utk formula SUM subtotal) dan
    # baris subtotal kunci (utk formula selisih LABA KOTOR/EBITDA/dst),
    # persis pola file referensi (mis. C8='=SUM(C5:C7)', C16='=C8-C15').
    seksi_range: Dict[str, tuple] = {}       # nama seksi -> (row_awal, row_akhir) baris akun
    subtotal_row: Dict[str, int] = {}        # label subtotal -> nomor baris
    seksi_awal_row: Optional[int] = None

    # [FIX] Range baris akun "Pendapatan Lain-lain" & "Beban Lain-lain" --
    # dua seksi ini TIDAK punya baris "judul" tersendiri di data (langsung
    # muncul sbg baris akun sesudah subtotal "LABA USAHA", persis file
    # referensi), jadi dilacak lewat field "arah" tiap baris akun begitu
    # posisi sudah lewat subtotal LABA USAHA (pendapatan_lain muncul lebih
    # dulu, baru beban_lain -- urutan tetap dari laporan_keuangan.py).
    setelah_laba_usaha = False
    pl_start = pl_end = bl_start = bl_end = None

    baris_list = pnl_baku["baris"]
    n = len(baris_list)

    for idx, b in enumerate(baris_list):
        tipe = b["tipe"]
        r = ws_pnl.max_row + 1
        is_baris_terakhir = (idx == n - 1)

        if tipe == "judul":
            ws_pnl.append([None, b["uraian"], None, None, None, None, b.get("keterangan", "")])
            for c in range(1, 8):
                cell = ws_pnl.cell(row=r, column=c)
                cell.font = _PNL_JUDUL_FONT
                cell.fill = _PNL_JUDUL_FILL
                cell.border = _border_pnl_judul(c)
                if c in (3, 4, 5, 6):  # kolom angka tetap diberi number_format walau kosong, persis file referensi
                    cell.number_format = _PNL_FORMAT_RUPIAH
            # Mulai lacak seksi akun baru sesudah baris judul ini
            seksi_awal_row = r + 1

        elif tipe == "akun":
            arah = b.get("arah", "beban")
            ws_pnl.append([b["kode"], b["uraian"], b["komersial"], b["koreksi_positif"],
                            b["koreksi_negatif"], None, b.get("keterangan", "")])
            for c in (1, 2, 7):
                cell = ws_pnl.cell(row=r, column=c)
                cell.font = _PNL_ITEM_FONT_LABEL
                cell.border = _border_pnl_item(c)
            cell_c = ws_pnl.cell(row=r, column=3)
            cell_c.font = _PNL_ITEM_FONT_HIJAU
            cell_c.number_format = _PNL_FORMAT_RUPIAH
            cell_c.border = _border_pnl_item(3)
            for c in (4, 5):
                cell = ws_pnl.cell(row=r, column=c)
                cell.font = _PNL_ITEM_FONT_HITAM
                cell.fill = _PNL_ITEM_FILL_PUTIH
                cell.number_format = _PNL_FORMAT_RUPIAH
                cell.border = _border_pnl_item(c)
            # Fiskal = Komersial + Positif - Negatif (pendapatan) atau
            # Komersial - Positif + Negatif (beban) -- persis pola file
            # referensi (baris 4101 vs baris 5101).
            if arah == "pendapatan":
                formula_f = f"=C{r}+D{r}-E{r}"
            else:
                formula_f = f"=C{r}-D{r}+E{r}"
            cell_f = ws_pnl.cell(row=r, column=6, value=formula_f)
            cell_f.font = _PNL_ITEM_FONT_HITAM
            cell_f.number_format = _PNL_FORMAT_RUPIAH
            cell_f.border = _border_pnl_item(6)

            # Lacak range Pendapatan Lain-lain / Beban Lain-lain (dipakai
            # formula "LABA BERSIH KOMERSIAL" di bawah) -- lihat catatan
            # `setelah_laba_usaha` di atas.
            if setelah_laba_usaha:
                if arah == "pendapatan":
                    pl_start = r if pl_start is None else pl_start
                    pl_end = r
                else:
                    bl_start = r if bl_start is None else bl_start
                    bl_end = r

        else:  # "subtotal"
            uraian = b["uraian"]
            d_val = b.get("koreksi_positif")
            e_val = b.get("koreksi_negatif")

            # ---- Tentukan formula kolom C & F sesuai posisi baris ----
            formula_c: Any = b["komersial"]
            formula_f: Any = b["fiskal"]

            if uraian.startswith("JUMLAH ") and seksi_awal_row is not None and seksi_awal_row <= r - 1:
                seksi_range[uraian] = (seksi_awal_row, r - 1)
                formula_c = f"=SUM(C{seksi_awal_row}:C{r - 1})"
                formula_f = f"=SUM(F{seksi_awal_row}:F{r - 1})"
            elif uraian == "LABA KOTOR":
                rp, rb = subtotal_row["JUMLAH PENDAPATAN USAHA"], subtotal_row["JUMLAH BEBAN LANGSUNG"]
                formula_c, formula_f = f"=C{rp}-C{rb}", f"=F{rp}-F{rb}"
            elif uraian == "EBITDA":
                rk, ro = subtotal_row["LABA KOTOR"], subtotal_row["JUMLAH BEBAN OPERASIONAL"]
                formula_c, formula_f = f"=C{rk}-C{ro}", f"=F{rk}-F{ro}"
            elif uraian == "LABA USAHA":
                re_, rp_ = subtotal_row["EBITDA"], subtotal_row["JUMLAH BEBAN PENYUSUTAN"]
                formula_c, formula_f = f"=C{re_}-C{rp_}", f"=F{re_}-F{rp_}"
            elif uraian == "LABA BERSIH KOMERSIAL":
                r_usaha = subtotal_row["LABA USAHA"]
                pl_c = f"SUM(C{pl_start}:C{pl_end})" if pl_start else "0"
                pl_f = f"SUM(F{pl_start}:F{pl_end})" if pl_start else "0"
                bl_c = f"SUM(C{bl_start}:C{bl_end})" if bl_start else "0"
                bl_f = f"SUM(F{bl_start}:F{bl_end})" if bl_start else "0"
                formula_c = f"=C{r_usaha}+{pl_c}-{bl_c}"
                formula_f = f"=F{r_usaha}+{pl_f}-{bl_f}"
            elif uraian == "TOTAL KOREKSI FISKAL POSITIF":
                formula_f = f"=D{r}"  # F33 = D33, persis file referensi
            elif uraian == "TOTAL KOREKSI FISKAL NEGATIF":
                formula_f = f"=E{r}"  # F34 = E34, persis file referensi
            elif uraian == "PENGHASILAN NETO FISKAL":
                r_lbk = subtotal_row["LABA BERSIH KOMERSIAL"]
                r_pos = subtotal_row["TOTAL KOREKSI FISKAL POSITIF"]
                r_neg = subtotal_row["TOTAL KOREKSI FISKAL NEGATIF"]
                formula_f = f"=F{r_lbk}+F{r_pos}-F{r_neg}"
            elif uraian == "KOMPENSASI KERUGIAN FISKAL":
                formula_f = f"=C{r}"  # F36 = C36, persis file referensi
            elif uraian == "PENGHASILAN KENA PAJAK SEBELUM PEMBULATAN":
                r_neto = subtotal_row["PENGHASILAN NETO FISKAL"]
                r_komp = subtotal_row["KOMPENSASI KERUGIAN FISKAL"]
                formula_f = f"=MAX(0,F{r_neto}-F{r_komp})"
            elif uraian == "PENGHASILAN KENA PAJAK -- RIBUAN PENUH":
                r_sblm = subtotal_row["PENGHASILAN KENA PAJAK SEBELUM PEMBULATAN"]
                formula_f = f"=INT(F{r_sblm}/1000)*1000"

            ws_pnl.append([None, uraian, formula_c, d_val, e_val, formula_f, b.get("keterangan", "")])
            for c in range(1, 8):
                cell = ws_pnl.cell(row=r, column=c)
                cell.font = _PNL_SUBTOTAL_FONT
                cell.fill = _PNL_SUBTOTAL_FILL
                cell.border = _border_pnl_subtotal(c, tutup_akhir=is_baris_terakhir)
                if c in (3, 4, 5, 6):
                    cell.number_format = _PNL_FORMAT_RUPIAH
            subtotal_row[uraian] = r

            # Sesudah baris "LABA USAHA" ditulis, baris akun berikutnya
            # (sebelum "LABA BERSIH KOMERSIAL") adalah Pendapatan Lain-lain
            # lalu Beban Lain-lain -- mulai lacak range-nya lewat flag ini.
            if uraian == "LABA USAHA":
                setelah_laba_usaha = True
            elif uraian == "LABA BERSIH KOMERSIAL":
                setelah_laba_usaha = False

    for kolom, lebar in _lebar_kolom_dari_isi(ws_pnl).items():
        ws_pnl.column_dimensions[kolom].width = lebar

    # [BARU] Return info baris (nomor baris tiap subtotal kunci + range
    # baris akun Pendapatan/Beban Lain-lain) -- dipakai sheet 17
    # "Rekonsiliasi Fiskal" (_tulis_sheet_rekonsiliasi_fiskal()) supaya
    # blok ringkasan REV/DIR/GP/OPEX/EBITDA/DEP/OI/OE/NP-nya bisa dibuat
    # FORMULA LINTAS-SHEET hidup (mis. ='PNL Lampiran SPT'!C13) yang
    # nempel langsung ke baris subtotal sheet ini -- BUKAN angka Python
    # beku terpisah -- supaya kedua sheet dijamin tie-out otomatis kalau
    # salah satu sisi diedit manual di Excel (sebelumnya fungsi ini tidak
    # mengembalikan apa pun, caller di 6284 mengabaikan return value).
    return {
        "sheet_nama": ws_pnl.title,
        "subtotal_row": subtotal_row,
        "pl_range": (pl_start, pl_end),
        "bl_range": (bl_start, bl_end),
    }


def _border_header_asumsi(kolom_pertama: int, kolom_terakhir: int, kolom: int) -> Border:
    return Border(
        top=_GARIS_ASUMSI, bottom=_GARIS_ASUMSI,
        left=_GARIS_ASUMSI if kolom == kolom_pertama else None,
        right=_GARIS_ASUMSI if kolom == kolom_terakhir else None,
    )


def _border_isi_asumsi(kolom_pertama: int, kolom_terakhir: int, kolom: int, tutup_bawah: bool = False) -> Border:
    return Border(
        left=_GARIS_ASUMSI if kolom == kolom_pertama else None,
        right=_GARIS_ASUMSI if kolom == kolom_terakhir else None,
        bottom=_GARIS_ASUMSI if tutup_bawah and kolom in (kolom_pertama, kolom_terakhir) else None,
    )


def _tulis_header_asumsi(ws, headers: List[str], row: int, lebar_kotak: Optional[int] = None) -> None:
    """lebar_kotak: jumlah kolom yang dianggap "milik" kotak border tabel
    ini -- default sama dgn jumlah header, tapi tabel "Sheet/Fungsi/Input
    Biru/Formula Hijau/Catatan" di file model referensi border-nya
    selebar 8 kolom (A-H) walau isi labelnya cuma 5 -- lihat pemanggilnya."""
    n = lebar_kotak or len(headers)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _HEADER_FONT_ASUMSI
        cell.fill = _HEADER_FILL_ASUMSI
        cell.alignment = _HEADER_ALIGN_ASUMSI
    for col in range(1, n + 1):
        ws.cell(row=row, column=col).border = _border_header_asumsi(1, n, col)
    ws.row_dimensions[row].height = 30


def _ke_tanggal_asumsi(v: Any) -> Any:
    """Parse string tanggal ISO ("YYYY-MM-DD" dst, format yg dikirim
    main.py) jadi objek datetime asli -- supaya Excel mengenalinya sbg
    tanggal (format dd-mmm-yyyy, rata kanan) sesuai file model referensi,
    bukan teks biasa. Kalau sudah date/datetime atau tidak bisa di-parse,
    dikembalikan apa adanya."""
    if isinstance(v, (date, datetime)):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(v, fmt)
            except ValueError:
                continue
    return v


def _tulis_input_asumsi(ws, row: int, label: str, nilai: Any, numfmt: Optional[str] = None) -> None:
    """Tulis 1 baris [label, nilai] dgn label kolom A polos (default,
    sama seperti file model referensi) dan nilai kolom B bergaya input
    (font biru, fill kuning) -- sesuai legenda "Warna biru = input/
    hardcode yang dapat diubah" yang ada di sheet ini sendiri."""
    cell_label = ws.cell(row=row, column=1, value=label)
    cell_label.alignment = _LABEL_ALIGN_ASUMSI
    cell_label.border = _border_isi_asumsi(1, 2, 1)
    cell_nilai = ws.cell(row=row, column=2, value=nilai)
    cell_nilai.font = _INPUT_FONT_ASUMSI
    cell_nilai.fill = _INPUT_FILL_ASUMSI
    cell_nilai.alignment = _LABEL_ALIGN_ASUMSI
    cell_nilai.border = _border_isi_asumsi(1, 2, 2)
    if numfmt:
        cell_nilai.number_format = numfmt


def _tulis_formula_asumsi(ws, row: int, label: str, formula: Any, numfmt: Optional[str] = None, tutup_bawah: bool = False) -> None:
    """Tulis 1 baris [label, formula] dgn nilai kolom B bergaya formula
    (font hitam, tanpa fill) -- beda dari _tulis_input_asumsi, sesuai
    legenda "Warna hitam = formula/perhitungan dalam sheet yang sama".
    tutup_bawah=True utk baris TERAKHIR suatu tabel (nutup kotak border
    dgn garis bawah), sesuai file model referensi."""
    cell_label = ws.cell(row=row, column=1, value=label)
    cell_label.alignment = _LABEL_ALIGN_ASUMSI
    cell_label.border = _border_isi_asumsi(1, 2, 1, tutup_bawah)
    cell_nilai = ws.cell(row=row, column=2, value=formula)
    cell_nilai.font = _FORMULA_FONT_ASUMSI
    cell_nilai.alignment = _LABEL_ALIGN_ASUMSI
    cell_nilai.border = _border_isi_asumsi(1, 2, 2, tutup_bawah)
    if numfmt:
        cell_nilai.number_format = numfmt


# [BARU] Gaya khusus sheet "Trial Balance Bulanan" -- meniru PERSIS file
# model referensi (judul besar merged di baris 1, header 2-baris dgn fill
# biru dongker + font putih, sel "Kode Akun"/"Nama Akun" merge vertikal,
# nama bulan merge horizontal di atas pasangan Debit/Kredit-nya).
_TITLE_FONT_TB_BULANAN = Font(bold=True, size=15)
_HEADER_FONT_TB_BULANAN = Font(bold=True, color="FFFFFFFF")
_HEADER_FILL_TB_BULANAN = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_ALIGN_CENTER_TB_BULANAN = Alignment(horizontal="center", vertical="center")
# [BARU] Format angka akuntansi persis file referensi: positif polos,
# negatif merah dalam kurung, nol ditampilkan sebagai "-".
_FORMAT_AKUNTANSI_TB_BULANAN = r"#,##0;[Red]\(#,##0\);\-"
# Baris TOTAL di paling bawah sheet Trial Balance Bulanan: bold + fill
# biru muda (SAMA seperti _HEADER_FILL_14SHEET di sheet 14-sheet lain --
# file referensi juga pakai warna ini utk baris TOTAL, BUKAN dongker
# yang dipakai header kolom bulan).
_TOTAL_FONT_TB_BULANAN = Font(bold=True)
_TOTAL_FILL_TB_BULANAN = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")

# [FIX Sheet 8] Konstanta KHUSUS sheet "Trial Balance Bulanan" (Sheet 8)
# supaya visualnya 100% sama dengan file model referensi
# (TRIAL_BALANCE_BULANAN___JANUARI_S_D__DESEMBER_2025.xlsx): font Carlito
# di semua sel (bukan default Calibri), judul rata KIRI + fill dongker
# (bukan center tanpa fill), header wrap_text, baris data Kode/Nama Akun
# biru #0000FF dan kolom Debit/Kredit hijau #008000, baris TOTAL label
# hitam + angka hijau, lebar kolom manual persis referensi. Sengaja
# dipisah dari _TITLE_FONT_TB_BULANAN/_HEADER_FONT_TB_BULANAN/
# _ALIGN_CENTER_TB_BULANAN/_TOTAL_FONT_TB_BULANAN/_TOTAL_FILL_TB_BULANAN
# di atas -- konstanta lama itu TETAP dipakai apa adanya oleh sheet
# "Laba Rugi Bulanan" & "Balance Sheet Bulanan" supaya keduanya TIDAK
# ikut berubah.
_TITLE_FONT_TB8 = Font(name="Carlito", bold=True, size=15, color="FFFFFFFF")
_TITLE_FILL_TB8 = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_TITLE_ALIGN_TB8 = Alignment(horizontal="left", vertical="center")
_HEADER_FONT_TB8 = Font(name="Carlito", size=11, bold=True, color="FFFFFFFF")
_HEADER_ALIGN_TB8 = Alignment(horizontal="center", vertical="center", wrap_text=False)
_DATA_FONT_KODE_NAMA_TB8 = Font(name="Carlito", size=11, color="FF0000FF")
_DATA_FONT_ANGKA_TB8 = Font(name="Carlito", size=11, color="FF008000")
_TOTAL_FONT_LABEL_TB8 = Font(name="Carlito", bold=True, color="FF000000")
_TOTAL_FONT_ANGKA_TB8 = Font(name="Carlito", bold=True, color="FF008000")
# Semua baris data & TOTAL rata tengah VERTIKAL (bukan horizontal) --
# persis file referensi.
_DATA_ALIGN_TB8 = Alignment(vertical="center")
_LEBAR_KOLOM_TB8 = {"A": 12.0, "B": 42.140625, "C": 14.28515625}
# [FIX Sheet 8] Border header (baris 3-4) & baris TOTAL -- SEBELUMNYA tidak
# ada border sama sekali (ikut default tanpa garis), sedangkan file model
# referensi punya garis tipis abu #A6A6A6 mengelilingi tiap sel header
# (kotak per-kolom "Kode Akun"/"Nama Akun" gabung vertikal, kotak per
# pasangan Debit/Kredit tiap bulan -- TANPA garis pemisah antara kolom
# Debit & Kredit yang sama bulannya) dan garis atas tipis + bawah GANDA
# navy #17365D di baris TOTAL (persis pola _TOTAL_BORDER_PIUTANG yang
# sudah dipakai sheet lain, dipisah jadi konstanta sendiri di sini karena
# warnanya berbeda dari _GARIS_ASUMSI abu yang dipakai header GL/Asumsi).
_GARIS_HEADER_TB8 = Side(style="thin", color="FFA6A6A6")
_BORDER_HEADER_TB8_KIRI = Border(top=_GARIS_HEADER_TB8, bottom=_GARIS_HEADER_TB8, left=_GARIS_HEADER_TB8)
_BORDER_HEADER_TB8_KANAN = Border(top=_GARIS_HEADER_TB8, bottom=_GARIS_HEADER_TB8, right=_GARIS_HEADER_TB8)
# [BARU] Border TENGAH -- utk kolom di tengah blok 3-kolom per bulan
# (Debit | Kredit | Ending Balance): kolom Kredit ada DI TENGAH, jadi
# tanpa garis kiri/kanan sama sekali (garis kiri/kanan cuma di kolom
# pertama & terakhir tiap blok bulan), cuma ikut garis atas/bawah header.
_BORDER_HEADER_TB8_TENGAH = Border(top=_GARIS_HEADER_TB8, bottom=_GARIS_HEADER_TB8)
_BORDER_TOTAL_TB8 = Border(top=Side(style="thin", color="FF17365D"), bottom=Side(style="double", color="FF17365D"))

# [BARU] Gaya khusus sheet "Laba Rugi Bulanan" -- meniru PERSIS file model
# referensi (LAPORAN_LABA_RUGI_BULANAN___2025.xlsx):
#   - Judul & header kolom bulan: fill dongker #17365D, font putih bold.
#   - Baris kategori (PENDAPATAN/BEBAN LANGSUNG/BEBAN OPERASIONAL/
#     PENYUSUTAN): bold, fill biru muda #D9EAF7, font navy #17365D.
#   - Baris TOTAL/subtotal (LABA KOTOR/EBITDA/LABA USAHA/LABA BERSIH
#     BULANAN/LABA BERSIH YTD): bold, fill #D9E1F2 (SAMA seperti baris
#     TOTAL sheet Trial Balance Bulanan -- reuse konstanta yang sudah ada).
#   - Baris item "Pendapatan Lain-lain"/"Beban Lain-lain" (di luar
#     kategori manapun): font biru #0000FF, tidak bold, tanpa fill.
_HEADER_FONT_LR_BULANAN = Font(name="Carlito", bold=True, color="FFFFFFFF")
_HEADER_FILL_LR_BULANAN = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_TITLE_FONT_LR_BULANAN = Font(name="Carlito", bold=True, size=15, color="FFFFFFFF")
_KATEGORI_FONT_LR_BULANAN = Font(name="Carlito", bold=True, color="FF17365D")
_KATEGORI_FILL_LR_BULANAN = PatternFill(start_color="FFD9EAF7", end_color="FFD9EAF7", fill_type="solid")
# [FIX] Sebelumnya baris item Pendapatan/Beban Lain-lain pakai SATU warna
# biru di semua kolom (A-N), dan baris item kategori normal (PENDAPATAN/
# BEBAN LANGSUNG/dst) TIDAK diberi font sama sekali (jatuh ke default
# hitam Calibri) -- padahal file referensi konsisten: kolom Kode & Uraian
# SELALU biru #0000FF, kolom nilai (C-N) SELALU hijau #008000, di SEMUA
# baris item termasuk Pendapatan/Beban Lain-lain. Dipecah jadi 2 font
# terpisah supaya bisa diterapkan persis sesuai kolomnya.
_KODE_FONT_LR_BULANAN = Font(name="Carlito", color="FF0000FF")
_VALUE_FONT_LR_BULANAN = Font(name="Carlito", color="FF008000")
# [FIX] Baris TOTAL/subtotal sebelumnya reuse _TOTAL_FONT_TB_BULANAN yang
# tidak set nama font (default Calibri) -- dibuat versi khusus ber-Carlito
# supaya tidak memengaruhi sheet Trial Balance Bulanan yang sudah benar.
_TOTAL_FONT_LR_BULANAN = Font(name="Carlito", bold=True, color="FF000000")
# [FIX] Header row (Kode/Uraian/Jan-25..Dec-25) di file referensi wrap
# text + border tipis atas-bawah -- sebelumnya cuma center tanpa wrap/border.
_ALIGN_HEADER_LR_BULANAN = Alignment(horizontal="center", vertical="center", wrap_text=False)
_BORDER_HEADER_LR_BULANAN = Border(top=Side(style="thin"), bottom=Side(style="thin"))
# [FIX] Baris kategori (PENDAPATAN/BEBAN LANGSUNG/dst) di referensi punya
# border bawah tipis -- sebelumnya tanpa border sama sekali.
_BORDER_KATEGORI_LR_BULANAN = Border(bottom=Side(style="thin"))
# [FIX] Baris TOTAL/subtotal di referensi punya border atas tipis + bawah
# DOUBLE -- sebelumnya tanpa border sama sekali.
_BORDER_TOTAL_LR_BULANAN = Border(top=Side(style="thin"), bottom=Side(style="double"))
# [FIX] Lebar kolom sebelumnya hasil autofit (beda-beda tergantung isi) --
# diganti FIXED persis lebar di file referensi (kolom A=Kode sempit,
# B=Uraian lebar 37, C-N=kolom bulan ~12.4-13.1).
_LEBAR_KOLOM_LR_BULANAN = {
    "A": 6.43, "B": 37.0, "C": 12.57, "D": 12.86, "E": 12.71, "F": 12.43,
    "G": 13.14, "H": 12.57, "I": 12.43, "J": 12.86, "K": 13.14, "L": 12.71,
    "M": 12.86, "N": 13.29,
}


# [FIX] Gaya khusus sheet "Balance Sheet Bulanan" (sheet ke-10) -- meniru
# PERSIS file model referensi (BALANCE_SHEET_BULANAN___2025.xlsx): font
# "Carlito" size 11 di seluruh sheet (judul size 15), judul & header kolom
# bulan fill navy #17365D + font putih, section utama (ASET/LIABILITAS/
# EKUITAS) fill biru muda #D9EAF7 + font navy #17365D + border bawah tipis,
# SUB-kelompok (Aset Lancar/Aset Tetap/dst) fill LEBIH TERANG #F3F6FA (beda
# dari section utama -- sebelumnya dua level ini kepakai warna sama), baris
# akun font BIRU #0000FF utk kode+uraian & HIJAU #008000 utk angka (pola
# "input vs formula" khas file referensi), baris TOTAL fill #D9E1F2 + label
# hitam bold + angka hijau bold + border atas tipis/bawah dobel. SENGAJA
# dipisah dari konstanta _..._LR_BULANAN (dipakai sheet 9 Laba Rugi
# Bulanan) supaya perubahan di sini TIDAK memengaruhi sheet lain -- HANYA
# sheet "Balance Sheet Bulanan" yang diminta disamakan persis.
_FONT_NAME_BS = "Carlito"
_TITLE_FONT_BS = Font(name=_FONT_NAME_BS, bold=True, size=15, color="FFFFFFFF")
_TITLE_FILL_BS = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_TITLE_ALIGN_BS = Alignment(horizontal="left", vertical="center")

_HEADER_FONT_BS = Font(name=_FONT_NAME_BS, bold=True, size=11, color="FFFFFFFF")
_HEADER_FILL_BS = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_HEADER_ALIGN_BS = Alignment(horizontal="center", vertical="center", wrap_text=False)
_HEADER_BORDER_BS = Border(top=Side(style="thin"), bottom=Side(style="thin"))
_HEADER_BORDER_KANAN_BS = Border(
    top=Side(style="thin"), bottom=Side(style="thin"), right=Side(style="thin")
)

# Section utama (ASET/LIABILITAS/EKUITAS)
_KATEGORI_FONT_BS = Font(name=_FONT_NAME_BS, bold=True, size=11, color="FF17365D")
_KATEGORI_FILL_BS = PatternFill(start_color="FFD9EAF7", end_color="FFD9EAF7", fill_type="solid")
_KATEGORI_BORDER_BS = Border(bottom=Side(style="thin"))

# Sub-kelompok (Aset Lancar/Aset Tetap/dst) -- fill lebih terang drpd
# section utama, TANPA border (persis file referensi).
_SUBKATEGORI_FONT_BS = Font(name=_FONT_NAME_BS, bold=True, size=11, color="FF17365D")
_SUBKATEGORI_FILL_BS = PatternFill(start_color="FFF3F6FA", end_color="FFF3F6FA", fill_type="solid")

# Baris akun (data) -- kode & uraian biru (input), angka hijau (formula/
# hasil hitung).
_ITEM_FONT_LABEL_BS = Font(name=_FONT_NAME_BS, size=11, color="FF0000FF")
_ITEM_FONT_ANGKA_BS = Font(name=_FONT_NAME_BS, size=11, color="FF008000")

# Baris TOTAL -- label hitam bold, angka hijau bold, fill biru muda,
# border atas tipis + bawah dobel.
_TOTAL_FONT_LABEL_BS = Font(name=_FONT_NAME_BS, bold=True, size=11, color="FF000000")
_TOTAL_FONT_ANGKA_BS = Font(name=_FONT_NAME_BS, bold=True, size=11, color="FF008000")
_TOTAL_FILL_BS = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
_TOTAL_BORDER_BS = Border(top=Side(style="thin"), bottom=Side(style="double"))

_LEBAR_KOLOM_BS = {"A": 6.43, "B": 42.14, "C": 14.29}


# [BARU] Gaya khusus sheet "COA" -- meniru PERSIS file model referensi
# (CHART_OF_ACCOUNTS__COA_.xlsx): judul (merge A1:I1) & baris header fill
# navy #17365D + font Carlito putih bold (judul size 15 rata kiri, header
# size 11 center+wrap+border tipis atas-bawah), baris data font Carlito
# size 11 warna biru #0000FF (bukan hitam -- pola "input" khas file
# referensi), lebar kolom MANUAL sama persis file referensi (bukan
# autofit). SENGAJA dipisah dari _HEADER_FONT_14SHEET/_HEADER_FILL_14SHEET
# (dipakai sheet 14-sheet lain) supaya perubahan di sini TIDAK memengaruhi
# sheet lain -- hanya sheet "COA" yang diminta disamakan persis.
_TITLE_FONT_COA = Font(name="Carlito", bold=True, size=15, color="FFFFFFFF")
_TITLE_FILL_COA = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_TITLE_ALIGN_COA = Alignment(horizontal="left", vertical="center")
_HEADER_FONT_COA = Font(name="Carlito", bold=True, size=11, color="FFFFFFFF")
_HEADER_FILL_COA = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_HEADER_ALIGN_COA = Alignment(horizontal="center", vertical="center", wrap_text=False)
_HEADER_BORDER_COA = Border(top=Side(style="thin"), bottom=Side(style="thin"))
_DATA_FONT_COA = Font(name="Carlito", size=11, color="FF0000FF")
_DATA_ALIGN_COA = Alignment(vertical="center")
_LEBAR_KOLOM_COA = {
    "A": 12.0, "B": 42.14, "C": 16.71, "D": 24.14, "E": 14.57,
    "F": 14.71, "G": 18.0, "H": 12.43, "I": 39.86,
}


# [BARU] Gaya khusus sheet "Neraca Saldo Awal" -- meniru PERSIS file model
# referensi (NERACA_SALDO_AWAL___FORMAT_GENERAL_LEDGER___1_JANUARI_2025.xlsx):
# judul (merge A1:M1) & baris header fill navy #17365D + font Carlito
# putih bold (judul size 15 rata kiri, header size 11 center + border tipis
# semua sisi), baris data font Carlito size 11 biru #0000FF utk kolom
# hardcode -- KECUALI kolom "Nama Akun" yang font-nya HIJAU #008000 karena
# isinya formula VLOOKUP (bukan value polos), lebar kolom MANUAL sama
# persis file referensi. SENGAJA dipisah dari konstanta sheet lain (COA,
# 14-sheet generik) supaya perubahan di sini tidak memengaruhi sheet lain.
_TITLE_FONT_NSA = Font(name="Carlito", bold=True, size=15, color="FFFFFFFF")
_TITLE_FILL_NSA = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_TITLE_ALIGN_NSA = Alignment(horizontal="left", vertical="center")
_HEADER_FONT_NSA = Font(name="Carlito", bold=True, size=11, color="FFFFFFFF")
_HEADER_FILL_NSA = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_HEADER_ALIGN_NSA = Alignment(horizontal="center", vertical="center")
_HEADER_BORDER_NSA = Border(
    top=Side(style="thin"), bottom=Side(style="thin"),
    left=Side(style="thin"), right=Side(style="thin"),
)
_DATA_FONT_NSA = Font(name="Carlito", size=11, color="FF0000FF")
_DATA_FONT_FORMULA_NSA = Font(name="Carlito", size=11, color="FF008000")
_DATA_ALIGN_NSA = Alignment(vertical="center")
_FORMAT_TANGGAL_NSA = r"dd\-mmm\-yyyy"
_FORMAT_RUPIAH_NSA = r"#,##0;[Red]\(#,##0\);\-"
_LEBAR_KOLOM_NSA = {
    "A": 3.86, "B": 11.29, "C": 12.57, "D": 14.71, "E": 12.0,
    "F": 37.57, "G": 22.14, "H": 14.29, "J": 11.29, "K": 18.14,
    "L": 19.43, "M": 19.71,
}
# [FIX] Sebelumnya baris data sheet "Neraca Saldo Awal" tidak diberi fill
# sama sekali -- dicek ulang cell-per-cell terhadap screenshot file model
# referensi (NERACA_SALDO_AWAL___FORMAT_GENERAL_LEDGER___1_JANUARI_2025.xlsx,
# disamplekan piksel per pikselnya): baris data ternyata BERSELANG-SELING
# (banded) biru muda #C0E6F5 / putih polos per baris (baris data ke-1,3,5,..
# dari header = biru muda, ke-2,4,6,.. = putih), BUKAN fill rata semua
# baris. Dipisah dari konstanta fill lain (mis. _HEADER_FILL_NSA) supaya
# tidak tertukar dengan fill header/judul.
_BANDING_FILL_NSA = PatternFill(start_color="FFC0E6F5", end_color="FFC0E6F5", fill_type="solid")


# [FIX] Gaya khusus sheet "GL <tahun>" -- SEBELUMNYA pakai gaya generik
# _tulis_header_14sheet (fill biru muda #D9E1F2, tanpa font putih/Carlito,
# tanpa border, tanpa merge judul) + _autofit_14sheet (lebar kolom auto),
# jauh berbeda dari file model referensi GL.xlsx yang user kirim (dianalisis
# ulang cell-per-cell): judul (merge A1:S1) & baris header fill navy
# #17365D + font Carlito putih bold (judul size 15 rata kiri, header size
# 11 center+wrap, border tipis abu #A6A6A6 di sisi atas+bawah semua kolom
# dan sisi kiri/kanan hanya di kolom pertama/terakhir -- pola PERSIS sama
# dengan _border_header_asumsi/_GARIS_ASUMSI di atas, jadi dipakai ulang),
# baris data font Carlito size 11 biru #0000FF utk kolom hardcode, HIJAU
# #008000 utk kolom "Nama Akun" (formula VLOOKUP ke sheet COA DI WORKBOOK
# YANG SAMA -- bukan link file eksternal "[1]COA!" seperti file model yang
# filenya sendiri tidak disertakan, sama seperti fix Neraca Saldo Awal di
# atas), HITAM #000000 utk kolom "Periode" (formula TEXT lokal ke kolom
# Tanggal baris yang sama), align vertical=center (horizontal default,
# sama seperti file referensi), lebar kolom MANUAL sama persis file
# referensi (bukan autofit). Kolom "Status Validasi" (S) SENGAJA
# dipertahankan di luar 18 kolom file referensi (A-R) -- ini fitur audit
# nyata yang sudah ada (lihat _status_validasi_gl()), bukan kosmetik, jadi
# tidak dihapus, cuma ditaruh setelah kolom R supaya 18 kolom pertama
# tetap identik posisi/urutannya dengan file referensi. SENGAJA dipisah
# dari konstanta sheet lain supaya perubahan di sini tidak memengaruhi
# sheet lain.
_TITLE_FONT_GL = Font(name="Carlito", bold=True, size=15, color="FFFFFFFF")
_TITLE_FILL_GL = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_TITLE_ALIGN_GL = Alignment(horizontal="left", vertical="center")
_HEADER_FONT_GL = Font(name="Carlito", bold=True, size=11, color="FFFFFFFF")
_HEADER_FILL_GL = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
# [FIX] wrap_text diganti jadi False -- sebelumnya True, dikombinasikan
# dengan lebar kolom manual yang sempit (lihat _LEBAR_KOLOM_GL lama di
# bawah, sekarang sudah tidak dipakai) bikin teks header seperti "No"/
# "Kode Akun" terpotong ke 2 baris dengan tinggi berbeda-beda per kolom
# -- itu sumber tampilan header "turun"/tidak sejajar yang dikeluhkan
# user. Header GL kata-katanya pendek semua, jadi cukup 1 baris asal
# lebar kolomnya pas (lihat _lebar_kolom_dari_header di bawah).
_HEADER_ALIGN_GL = Alignment(horizontal="center", vertical="center", wrap_text=False)
_DATA_FONT_GL = Font(name="Carlito", size=11, color="FF0000FF")
_DATA_FONT_FORMULA_GL = Font(name="Carlito", size=11, color="FF008000")
_DATA_FONT_PERIODE_GL = Font(name="Carlito", size=11, color="FF000000")
_DATA_ALIGN_GL = Alignment(vertical="center")
_FORMAT_TANGGAL_GL = r"dd\-mmm\-yyyy"
_FORMAT_RUPIAH_GL = r"#,##0;[Red]\(#,##0\);\-"


def _lebar_kolom_dari_header(headers, padding=4, minimum=6):
    """
    [BARU] Lebar kolom dihitung otomatis dari PANJANG TEKS HEADER masing-
    masing kolom -- bukan dari lebar manual/patokan resmi file model
    referensi (_LEBAR_KOLOM_GL versi lama, sudah dihapus). Ini sengaja
    diminta user: dia mau header selalu rapi & pas satu baris (tidak ada
    huruf yang "turun") tanpa lebar kolom terpaku ke ukuran file model.

    `padding` menambah sedikit ruang kosong kiri-kanan supaya teks header
    (bold, jadi sedikit lebih lebar dari teks biasa) tidak mepet ke tepi
    kolom. `minimum` mencegah kolom dengan header sangat pendek (mis. "No")
    jadi terlalu sempit dan kelihatan gepeng.
    """
    lebar = {}
    for i, h in enumerate(headers, 1):
        panjang_teks = len(str(h))
        lebar[get_column_letter(i)] = max(minimum, panjang_teks + padding)
    return lebar


def _lebar_kolom_gabungan_header(*daftar_header, padding=4, minimum=6):
    """
    [BARU] Sama seperti _lebar_kolom_dari_header, tapi untuk sheet yang
    punya LEBIH DARI SATU tabel header berbeda di kolom yang sama (mis.
    sheet "Ringkasan" -- tabel KPI 4 kolom & tabel Tren Bulanan 13 kolom
    sama-sama mulai dari kolom A, dengan teks header yang beda-beda).
    Lebar tiap kolom diambil dari header TERPANJANG di antara semua
    tabel yang memakai kolom itu, supaya header di SEMUA tabel tetap
    tidak ada yang terpotong -- bukan cuma tabel yang ditulis terakhir.
    """
    lebar = {}
    for headers in daftar_header:
        for i, h in enumerate(headers, 1):
            kolom = get_column_letter(i)
            panjang = len(str(h)) + padding
            lebar[kolom] = max(lebar.get(kolom, minimum), panjang, minimum)
    return lebar


def _lebar_kolom_dari_isi(ws, padding=2, minimum=6, lebar_maks=60):
    """
    [BARU] Lebar kolom dihitung dari teks TERPANJANG yang BENAR-BENAR ADA
    di kolom itu -- isi baris data, bukan cuma header (beda dgn
    _lebar_kolom_dari_header/_lebar_kolom_gabungan_header di atas). Ini
    diminta user supaya kolom yang headernya pendek tapi isinya panjang
    (mis. header "Keterangan" tapi isi "Koreksi fiskal diinput pada
    bagian rekonsiliasi sheet Laba Rugi Bulanan") tidak lagi terpotong.
    Dipakai untuk sheet 9-18 di export 18-sheet, GANTI pendekatan lebar-
    dari-header-saja yang sebelumnya dipakai di sheet-sheet tsb.

    Cara pakai: panggil SETELAH semua isi sheet selesai ditulis (bukan
    sebelum, spt beberapa pemanggilan lama), supaya seluruh baris data
    ikut terhitung.

    Sel yang jadi bagian MERGE lebih dari 1 kolom (mis. baris judul sheet
    yang di-merge selebar semua kolom) OTOMATIS DILEWATI -- kalau ikut
    dihitung, kolom pertama akan jadi sangat lebar gara-gara teks judul
    panjang yang sebenarnya melebar ke banyak kolom, bukan representasi
    lebar kolom itu sendiri.

    Keterbatasan yang perlu diketahui: untuk sel berisi FORMULA Excel
    (mis. "=SUM(B35:B38)"), nilai hasil hitungnya baru muncul setelah
    dibuka di Excel -- tidak bisa dibaca dari sini. Untuk sel begitu:
    (a) kalau formatnya format angka/rupiah (number_format mengandung
    "#,##0"), dipakai perkiraan lebar 15 karakter (cukup utk angka
    rupiah sampai belasan digit); (b) kalau formulanya mengandung
    literal teks berkutip (mis. `=IF(ROUND(C22,0)=0,"BALANCE","PERIKSA")`
    -- pola umum utk sel status/keterangan), dipakai panjang literal
    TERPANJANG di antara kutipan itu sbg perkiraan hasil tampil (jauh
    lebih akurat drpd panjang teks formulanya sendiri yg bisa 5x lebih
    panjang dari hasil aslinya); (c) selain itu dipakai panjang teks
    formulanya apa adanya (aman -- lebih lebar drpd perlu, tapi tidak
    akan kepotong).
    """
    import re as _re

    sel_merge_lebar = set()
    for rng in ws.merged_cells.ranges:
        if rng.max_col - rng.min_col + 1 > 1:
            for r in range(rng.min_row, rng.max_row + 1):
                sel_merge_lebar.add((r, rng.min_col))

    lebar_per_kolom: Dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            nilai = cell.value
            if nilai is None or (cell.row, cell.column) in sel_merge_lebar:
                continue
            if isinstance(nilai, bool):
                teks = str(nilai)
            elif isinstance(nilai, (int, float)):
                teks = f"{nilai:,.2f}" if isinstance(nilai, float) else f"{nilai:,}"
            elif isinstance(nilai, str) and nilai.startswith("="):
                fmt = str(cell.number_format or "")
                literal_teks = _re.findall(r'"([^"]*)"', nilai)
                if "#,##0" in fmt:
                    teks = "-999,999,999,999"
                elif literal_teks:
                    teks = max(literal_teks, key=len)
                else:
                    teks = nilai
            else:
                teks = str(nilai)
            panjang = max((len(baris) for baris in teks.split("\n")), default=0)
            surat = cell.column_letter
            if panjang > lebar_per_kolom.get(surat, 0):
                lebar_per_kolom[surat] = panjang

    return {surat: max(minimum, min(panjang + padding, lebar_maks))
            for surat, panjang in lebar_per_kolom.items()}
# [FIX] Sama seperti Neraca Saldo Awal di atas -- baris data sheet
# "GL <tahun>" sebelumnya tidak ada fill sama sekali, padahal file model
# referensi GL.xlsx (dicek ulang lewat sampling piksel screenshot)
# berselang-seling biru muda #C0E6F5 / putih polos PER BARIS FISIK (bukan
# per pasangan debet-kredit satu jurnal) -- baris data ke-1,3,5,.. dari
# header biru muda, ke-2,4,6,.. putih. Warna sama persis dengan
# _BANDING_FILL_NSA, tapi sengaja dibuat konstanta terpisah supaya kedua
# sheet tetap independen kalau salah satu perlu diubah lagi nanti.
_BANDING_FILL_GL = PatternFill(start_color="FFC0E6F5", end_color="FFC0E6F5", fill_type="solid")


# [FIX] Gaya khusus sheet "Buku Bantu Piutang" -- SEBELUMNYA pakai gaya
# generik _tulis_header_14sheet (fill biru muda #D9E1F2, tanpa font putih/
# Carlito, tanpa border, tanpa merge judul) + _autofit_14sheet (lebar kolom
# auto), jauh berbeda dari file model referensi
# BUKU_BANTU_PIUTANG_USAHA_2025.xlsx yang user kirim (dianalisis ulang
# cell-per-cell): judul (merge A1:L1) & baris header fill navy #17365D +
# font Carlito putih bold (judul size 15 rata kiri, header size 11
# center+wrap, border tipis semua sisi -- pola sama dgn _HEADER_BORDER_NSA,
# dipakai ulang), baris data font Carlito size 11 BIRU #0000FF utk 6 kolom
# pertama (No. Invoice s.d. Jatuh Tempo -- hardcode dari df_piutang) dan
# HIJAU #008000 utk 6 kolom terakhir (Nilai Invoice s.d. Status -- hasil
# hitungan _perkaya_piutang_per_baris() di akuntansi_ai.py, setara "formula"
# pada file model referensi walau di sini sudah berupa angka jadi, bukan
# rumus Excel hidup -- pola warna sama persis dgn kolom formula VLOOKUP di
# GL/Neraca Saldo Awal di atas), align vertical=center, lebar kolom MANUAL
# sama persis file referensi (bukan autofit). Baris TOTAL: font hitam bold,
# fill #D9E1F2, border atas tipis + bawah DOUBLE di SELURUH kolom A-L (bukan
# cuma kolom yang berisi angka -- dicek di file referensi, B19:F19 dan
# J19:L19 yang kosong pun tetap kebagian fill+border yang sama).
_TITLE_FONT_PIUTANG = Font(name="Carlito", bold=True, size=15, color="FFFFFFFF")
_TITLE_FILL_PIUTANG = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_TITLE_ALIGN_PIUTANG = Alignment(horizontal="left", vertical="center")
_HEADER_FONT_PIUTANG = Font(name="Carlito", bold=True, size=11, color="FFFFFFFF")
_HEADER_FILL_PIUTANG = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_HEADER_ALIGN_PIUTANG = Alignment(horizontal="center", vertical="center", wrap_text=False)
# [FIX] Border header SEBELUMNYA hitam di 4 sisi tiap sel (jadi ada garis
# vertikal di SEMUA antar-kolom) -- dicek ulang cell-per-cell ke file
# referensi BUKU_BANTU_PIUTANG_USAHA_2025.xlsx, ternyata border aslinya
# ABU-ABU #A6A6A6 dan garis vertikal CUMA di pinggir luar (kiri kolom A,
# kanan kolom L) -- tidak ada garis vertikal di tengah. 3 varian border di
# bawah dipilih per posisi kolom saat menulis header (lihat pemakaiannya).
_WARNA_BORDER_HEADER_PIUTANG = "FFA6A6A6"
_HEADER_BORDER_TENGAH_PIUTANG = Border(
    top=Side(style="thin", color=_WARNA_BORDER_HEADER_PIUTANG),
    bottom=Side(style="thin", color=_WARNA_BORDER_HEADER_PIUTANG),
)
_HEADER_BORDER_KIRI_PIUTANG = Border(
    top=Side(style="thin", color=_WARNA_BORDER_HEADER_PIUTANG),
    bottom=Side(style="thin", color=_WARNA_BORDER_HEADER_PIUTANG),
    left=Side(style="thin", color=_WARNA_BORDER_HEADER_PIUTANG),
)
_HEADER_BORDER_KANAN_PIUTANG = Border(
    top=Side(style="thin", color=_WARNA_BORDER_HEADER_PIUTANG),
    bottom=Side(style="thin", color=_WARNA_BORDER_HEADER_PIUTANG),
    right=Side(style="thin", color=_WARNA_BORDER_HEADER_PIUTANG),
)
_DATA_FONT_HARDCODE_PIUTANG = Font(name="Carlito", size=11, color="FF0000FF")
_DATA_FONT_HITUNG_PIUTANG = Font(name="Carlito", size=11, color="FF008000")
_DATA_ALIGN_PIUTANG = Alignment(vertical="center")
_TOTAL_FONT_PIUTANG = Font(name="Carlito", bold=True, size=11, color="FF000000")
_TOTAL_FILL_PIUTANG = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
# [FIX] Border bawah TOTAL sebelumnya double tanpa warna (default hitam) --
# di file referensi warnanya NAVY #17365D (sama dgn warna judul/header),
# bukan hitam.
_TOTAL_BORDER_PIUTANG = Border(top=Side(style="thin"), bottom=Side(style="double", color="FF17365D"))
_FORMAT_TANGGAL_PIUTANG = r"dd\-mmm\-yyyy"
_FORMAT_RUPIAH_PIUTANG = r"#,##0;[Red]\(#,##0\);\-"
# Kolom hitung (index Excel 1-based) yang nilainya angka & butuh format
# rupiah -- "Hari Tertunggak" (kolom J) sengaja TIDAK dimasukkan karena di
# file referensi kolom itu format General polos (angka hari biasa, bukan
# Rupiah).
_KOLOM_RUPIAH_PIUTANG = (7, 8, 9)  # kolom Excel: G, H, I (Nilai Invoice/Penerimaan/Saldo Piutang)
_LEBAR_KOLOM_PIUTANG = {
    "A": 12.43, "B": 17.43, "C": 23.86, "D": 24.29, "E": 13.14, "F": 14.29,
    "G": 13.57, "H": 13.14, "I": 15.14, "J": 17.71, "K": 20.0, "L": 12.29,
}
# [BARU] Lebar kolom manual khusus sheet "Buku Bantu Hutang", diukur
# persis dari file referensi BUKU_BANTU_HUTANG_USAHA_2025.xlsx (11 kolom
# A-K). Kolom G ("Nilai Tagihan") SENGAJA tidak diberi lebar manual --
# di file referensi kolom itu memang dibiarkan lebar default (~8.43),
# beda dengan sheet Piutang yang G-nya diset manual. Jangan "diperbaiki"
# jadi ikut pola Piutang -- itu justru menyimpang dari file referensi.
_LEBAR_KOLOM_HUTANG = {
    "A": 13.29, "B": 18.14, "C": 21.14, "D": 24.29, "E": 13.14, "F": 14.29,
    "H": 13.43, "I": 14.71, "J": 17.71, "K": 12.29,
}

# [BARU] Style sheet "Buku Bantu Aktiva Tetap", diukur PERSIS dari file
# referensi user (BUKU_BANTU_AKTIVA_TETAP___JADWAL_PENYUSUTAN_2025.xlsx,
# 28 kolom A-AB):
#   - Judul (baris 1, merge A1:AB1): Carlito 15 bold putih, fill navy
#     #17365D, align kiri+tengah, row height 19.5.
#   - Baris 2: HANYA kolom N-Y (bulan) diisi tanggal akhir tiap bulan,
#     font Carlito 11 BIRU #0000FF, align tengah, format dd-mmm-yyyy.
#   - Baris 3 (header kolom): Carlito 11 bold putih, fill navy #17365D,
#     align tengah+tengah+wrap, border tipis atas+bawah semua kolom,
#     kiri HANYA kolom pertama (A), kanan HANYA kolom terakhir (AB).
#   - Baris data: kolom A-K (Asset ID s.d. Metode, input mentah) = font
#     Carlito 11 BIRU #0000FF, TANPA bold. Kolom L-AB (Penyusutan/Bulan
#     s.d. Nilai Buku, hasil formula) = font Carlito 11 HITAM #000000.
#     Border HANYA kolom A (kiri tipis) & kolom AB (kanan tipis) -- tidak
#     ada border lain di baris data.
#   - Baris TOTAL: font Carlito 11 bold hitam, fill biru muda #D9E1F2,
#     border tipis atas+bawah SEMUA kolom, kiri tipis hanya kolom A,
#     kanan tipis hanya kolom AB (border SAMA seperti baris header, beda
#     dengan sheet Piutang/Hutang yang bawahnya double-navy).
_TITLE_FONT_ASET = Font(name="Carlito", bold=True, size=15, color="FFFFFFFF")
_TITLE_FILL_ASET = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_TITLE_ALIGN_ASET = Alignment(horizontal="left", vertical="center")
_HEADER_FONT_ASET = Font(name="Carlito", bold=True, size=11, color="FFFFFFFF")
_HEADER_FILL_ASET = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_HEADER_ALIGN_ASET = Alignment(horizontal="center", vertical="center", wrap_text=False)
_TANGGAL_BULAN_FONT_ASET = Font(name="Carlito", size=11, color="FF0000FF")
_TANGGAL_BULAN_ALIGN_ASET = Alignment(horizontal="center")
_DATA_FONT_INPUT_ASET = Font(name="Carlito", size=11, color="FF0000FF")
_DATA_FONT_CALC_ASET = Font(name="Carlito", size=11, color="FF000000")
_TOTAL_FONT_ASET = Font(name="Carlito", bold=True, size=11, color="FF000000")
_TOTAL_FILL_ASET = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
_BORDER_KIRI_ASET = Border(top=Side(style="thin"), bottom=Side(style="thin"), left=Side(style="thin"))
_BORDER_TENGAH_ASET = Border(top=Side(style="thin"), bottom=Side(style="thin"))
_BORDER_KANAN_ASET = Border(top=Side(style="thin"), bottom=Side(style="thin"), right=Side(style="thin"))
_DATA_BORDER_KIRI_ASET = Border(left=Side(style="thin"))
_DATA_BORDER_KANAN_ASET = Border(right=Side(style="thin"))
_FORMAT_TANGGAL_ASET = r"dd\-mmm\-yyyy"
_FORMAT_RUPIAH_ASET = r"#,##0;[Red]\(#,##0\);\-"
# Lebar kolom manual (28 kolom A-AB), diukur persis dari file referensi.
_LEBAR_KOLOM_ASET = {
    "A": 12.0, "B": 20.29, "C": 17.29, "D": 26.0, "E": 26.57, "F": 20.43,
    "G": 18.0, "H": 18.29, "I": 13.29, "J": 15.0, "K": 11.86, "L": 19.57,
    "M": 25.0, "N": 12.57, "O": 12.86, "P": 12.71, "Q": 12.43, "R": 13.14,
    "S": 12.57, "T": 11.86, "U": 12.86, "V": 13.14, "W": 12.71, "X": 12.86,
    "Y": 13.29, "Z": 18.29, "AA": 25.71, "AB": 22.57,
}
# Urutan huruf kolom A-AB (28 kolom, tetap -- bukan hasil autofit).
_KOLOM_HURUF_ASET = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M",
                      "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y",
                      "Z", "AA", "AB"]
# [BARU] Border header sheet "Buku Bantu Hutang" -- persis file referensi:
# hanya garis pinggir luar (atas+bawah di semua kolom, kiri cuma di kolom
# pertama, kanan cuma di kolom terakhir), TANPA garis vertikal pemisah
# antar kolom di tengah tabel.
_HEADER_BORDER_KIRI_HUTANG = Border(
    top=Side(style="thin"), bottom=Side(style="thin"), left=Side(style="thin"),
)
_HEADER_BORDER_TENGAH_HUTANG = Border(
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_HEADER_BORDER_KANAN_HUTANG = Border(
    top=Side(style="thin"), bottom=Side(style="thin"), right=Side(style="thin"),
)


# ================= Style sheet "Ringkasan" (RINGKASAN_KINERJA_KEUANGAN_2025.xlsx) =================
# [FIX] Dicek cell-per-cell ke file model referensi: judul & header
# tabel KPI/tren pakai gaya PERSIS sama dengan _TITLE_FONT_PIUTANG /
# _HEADER_FONT_PIUTANG / border abu #A6A6A6 di atas -- dipakai ulang
# langsung, tidak didefinisikan ulang. Yang beda dari sheet Piutang
# cuma font baris data (kolom label bold Carlito hitam, kolom
# nilai/angka bold Carlito hijau #008000 di tabel KPI -- TIDAK bold di
# tabel tren, kolom Sumber/Catatan/Metrik Calibri reguler) dan lebar
# kolom A-E yang dipatok manual sesuai file model (bukan autofit).
_LEBAR_KOLOM_RINGKASAN = {"A": 23.14, "B": 14.29, "C": 22.86, "D": 39.14, "E": 14.29}
_LABEL_FONT_RINGKASAN = Font(name="Carlito", bold=True, size=11)
_NILAI_FONT_RINGKASAN = Font(name="Carlito", bold=True, size=11, color="FF008000")
_SUMBER_FONT_RINGKASAN = Font(name="Calibri", size=11)
_METRIK_FONT_RINGKASAN = Font(name="Calibri", size=11)
_NILAI_TREN_FONT_RINGKASAN = Font(name="Carlito", size=11, color="FF008000")
_VCENTER_RINGKASAN = Alignment(vertical="center")
_BORDER_KIRI_RINGKASAN = Border(left=Side(style="thin", color=_WARNA_BORDER_HEADER_PIUTANG))
_BORDER_KANAN_RINGKASAN = Border(right=Side(style="thin", color=_WARNA_BORDER_HEADER_PIUTANG))
_BORDER_KIRI_BAWAH_RINGKASAN = Border(
    left=Side(style="thin", color=_WARNA_BORDER_HEADER_PIUTANG),
    bottom=Side(style="thin", color=_WARNA_BORDER_HEADER_PIUTANG),
)
_BORDER_KANAN_BAWAH_RINGKASAN = Border(
    right=Side(style="thin", color=_WARNA_BORDER_HEADER_PIUTANG),
    bottom=Side(style="thin", color=_WARNA_BORDER_HEADER_PIUTANG),
)
_BORDER_BAWAH_RINGKASAN = Border(bottom=Side(style="thin", color=_WARNA_BORDER_HEADER_PIUTANG))


# ================= Style sheet PLACEHOLDER (Laporan Perubahan Ekuitas,
# Laporan Arus Kas, CALK, Rekonsiliasi Fiskal) =================
# [BARU] 4 sheet ini belum punya file model referensi -- disepakati untuk
# SEMENTARA hanya ditulis judulnya saja dengan gaya PERSIS sama dengan
# sheet lain (fill navy #17365D + font Carlito putih bold, merge baris 1),
# TANPA header kolom/data (kolom & isi menyusul, menunggu penjelasan
# struktur dari user). SENGAJA satu set konstanta dipakai bareng oleh
# ke-4 sheet ini (bukan didup dobel 4x) karena keempatnya identik &
# berstatus placeholder sementara -- boleh dipecah per-sheet nanti begitu
# masing-masing mulai diisi struktur/data sungguhan (ikuti pola COA/NSA/dst).
_TITLE_FONT_PLACEHOLDER_BARU = Font(name="Carlito", bold=True, size=15, color="FFFFFFFF")
_TITLE_FILL_PLACEHOLDER_BARU = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_TITLE_ALIGN_PLACEHOLDER_BARU = Alignment(horizontal="left", vertical="center")
_CATATAN_FONT_PLACEHOLDER_BARU = Font(name="Carlito", italic=True, size=10, color="FF808080")


def _tulis_judul_sheet_placeholder_baru(ws, judul: str, ncols: int = 8) -> None:
    """Tulis baris judul bergaya navy/putih (sama seperti sheet lain),
    lalu satu baris catatan italic abu-abu bahwa struktur kolom & data
    menyusul. TIDAK menulis header kolom apa pun. [CATATAN] Sejak struktur
    kolom sheet 11/12/13/17 sudah dikonfirmasi user, fungsi ini SUDAH TIDAK
    DIPAKAI oleh keempatnya lagi (lihat _tulis_judul_header_baru4() +
    _tulis_sheet_perubahan_ekuitas()/_tulis_sheet_arus_kas()/
    _tulis_sheet_calk()/_tulis_sheet_rekonsiliasi_fiskal() di bawah) --
    sengaja TIDAK dihapus, disimpan sebagai fallback generik kalau suatu
    saat ada sheet baru lain yang strukturnya belum dikonfirmasi lagi."""
    cell_judul = ws.cell(row=1, column=1, value=judul)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell_judul.font = _TITLE_FONT_PLACEHOLDER_BARU
    cell_judul.fill = _TITLE_FILL_PLACEHOLDER_BARU
    cell_judul.alignment = _TITLE_ALIGN_PLACEHOLDER_BARU
    ws.row_dimensions[1].height = 27.95
    ws.cell(row=3, column=1,
             value="Struktur kolom & data menyusul -- menunggu penjelasan lebih lanjut.").font = \
        _CATATAN_FONT_PLACEHOLDER_BARU
    ws.column_dimensions["A"].width = 55


# ================= Style + penulis sheet 11/12/13/17 (Laporan Perubahan
# Ekuitas, Laporan Arus Kas, CALK, Rekonsiliasi Fiskal) =================
# [BARU] Struktur kolom sudah dikonfirmasi user (belum ada file model
# referensi utk 4 sheet ini, beda dgn sheet lain yang selalu dicocokkan
# cell-per-cell ke contoh Excel) -- gaya visual mengikuti pola sheet lain
# supaya satu workbook tetap konsisten: judul navy #17365D merge full
# lebar tabel, header (baris 3, baris 2 kosong) navy+putih bold+border
# tipis semua sisi+wrap, data font Carlito hitam utk teks / biru #0000FF
# utk angka hardcode / hijau #008000 utk angka hasil FORMULA Excel hidup
# (bukan angka Python beku -- supaya kalau salah satu komponen diedit
# manual di Excel, baris Saldo Akhir/Subtotal ikut kereka ulang sendiri),
# baris subtotal/total bold fill #D9E1F2 (sama seperti _PNL_SUBTOTAL_FILL).
_TITLE_FONT_BARU4 = Font(name="Carlito", bold=True, size=15, color="FFFFFFFF")
_TITLE_FILL_BARU4 = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_TITLE_ALIGN_BARU4 = Alignment(horizontal="left", vertical="center")
_HEADER_FONT_BARU4 = Font(name="Carlito", bold=True, size=11, color="FFFFFFFF")
_HEADER_FILL_BARU4 = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_HEADER_ALIGN_BARU4 = Alignment(horizontal="center", vertical="center", wrap_text=True)
_HEADER_BORDER_BARU4 = Border(
    top=Side(style="thin"), bottom=Side(style="thin"),
    left=Side(style="thin"), right=Side(style="thin"),
)
_DATA_FONT_BARU4 = Font(name="Carlito", size=11, color="FF000000")
_DATA_FONT_ANGKA_BARU4 = Font(name="Carlito", size=11, color="FF0000FF")      # angka hardcode
_DATA_FONT_FORMULA_BARU4 = Font(name="Carlito", size=11, color="FF008000")    # angka hasil formula
_DATA_BORDER_BARU4 = Border(bottom=Side(style="thin", color="FFD9D9D9"))
# [FIX] Sebelumnya baris "Subtotal -- <kategori>" di sheet CALK TIDAK
# pakai border sama sekali (cuma font bold + fill) -- beda dgn baris
# TOTAL di sheet 11/12/17 (keluarga gaya yg sama) yang SELALU pakai
# border atas tipis + bawah dobel (_TOTAL_BORDER_PE). Disamakan di sini
# supaya baris subtotal CALK konsisten secara visual dgn baris
# total/subtotal di sheet lain.
_TOTAL_BORDER_BARU4 = Border(top=Side(style="thin"), bottom=Side(style="double"))
_SUBTOTAL_FONT_BARU4 = Font(name="Carlito", bold=True, size=11, color="FF000000")
_SUBTOTAL_FILL_BARU4 = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
_TOTAL_FONT_BARU4 = Font(name="Carlito", bold=True, size=12, color="FFFFFFFF")
_TOTAL_FILL_BARU4 = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_CATATAN_FONT_BARU4 = Font(name="Carlito", italic=True, size=9, color="FF808080")
_FORMAT_RUPIAH_BARU4 = r"#,##0;[Red]\(#,##0\);\-"


def _tulis_judul_header_baru4(ws, judul: str, headers: List[str]) -> int:
    """Tulis judul (row1, merge full lebar tabel) + baris header kolom
    (row3, row2 sengaja kosong -- pola sama dgn sheet PNL/BS Lampiran SPT)
    dgn gaya navy/putih konsisten. Dipakai bareng oleh ke-4 fungsi penulis
    sheet 11/12/13/17 di bawah supaya tidak dobel kode. Return nomor baris
    header (pemanggil mulai menulis data dari header_row + 1)."""
    ncols = len(headers)
    ws.cell(row=1, column=1, value=judul)
    cell_judul = ws.cell(row=1, column=1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell_judul.font = _TITLE_FONT_BARU4
    cell_judul.fill = _TITLE_FILL_BARU4
    cell_judul.alignment = _TITLE_ALIGN_BARU4
    ws.row_dimensions[1].height = 27.95

    header_row = 3
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = _HEADER_FONT_BARU4
        cell.fill = _HEADER_FILL_BARU4
        cell.alignment = _HEADER_ALIGN_BARU4
        cell.border = _HEADER_BORDER_BARU4
    ws.row_dimensions[header_row].height = 30
    for kolom, lebar in _lebar_kolom_dari_header(headers).items():
        ws.column_dimensions[kolom].width = lebar
    return header_row


# ================= Style + penulis sheet 11 (Laporan Perubahan Ekuitas)
# =================
# [FIX] Diganti TOTAL dari versi placeholder 5-kolom (Keterangan/Saldo
# Awal/Penambahan/Pengurangan/Saldo Akhir, TANPA rincian bulanan) menjadi
# PERSIS pola file model referensi LAPORAN_PERUBAHAN_EKUITAS.xlsx yang
# user kirim: tabel ringkasan bulanan (Kode/Uraian/Jan..Des) per komponen
# ekuitas + TOTAL EKUITAS, lalu 2 section rincian roll-forward (RINCIAN
# PERUBAHAN MODAL per akun modal: Saldo Awal -> +Penambahan ->
# -Pengurangan -> Saldo Akhir, ditutup TOTAL SALDO AKHIR MODAL; RINCIAN
# PERUBAHAN SALDO LABA: Saldo Awal -> +Laba Tahun Berjalan (link ke sheet
# "Laba Rugi Bulanan") -> -Pengurangan/Dividen/Prive -> Saldo Akhir),
# ditutup CHECK BALANCE & STATUS "BALANCE"/"PERIKSA" -- semua RUMUS EXCEL
# HIDUP (SUMIFS ke sheet "Neraca Saldo Awal"/"GL <tahun>" LOKAL di
# workbook yang sama, pola sama dgn fix Trial Balance Bulanan/GL <tahun>
# di atas -- BUKAN link file eksternal "[1]..." spt file model aslinya).
#
# [PENTING] Akun modal & saldo laba TIDAK di-hardcode "Modal Tuan A/B"
# spt file model referensi (yang cuma cocok utk client partnership 2
# orang) -- diambil DINAMIS dari kategori/sub_kategori COA client ybs,
# SUMBER SAMA dgn sheet "Balance Sheet Bulanan" (section 10) supaya
# TOTAL EKUITAS di 2 sheet ini selalu tie-out: kategori="ekuitas",
# sub_kategori "Saldo Laba"/"Laba Ditahan" (atau nama akun mengandung itu
# kalau sub_kategori kosong) -> section Saldo Laba; selain itu -> section
# Modal. Client dgn N pemilik modal akan otomatis dapat N blok Saldo
# Awal/Penambahan/Pengurangan/Saldo Akhir di "RINCIAN PERUBAHAN MODAL".
_FONT_NAME_PE = "Carlito"
_TITLE_FONT_PE = Font(name=_FONT_NAME_PE, bold=True, size=15, color="FFFFFFFF")
_TITLE_FILL_PE = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_TITLE_ALIGN_PE = Alignment(horizontal="left", vertical="center")

_HEADER_FONT_PE = Font(name=_FONT_NAME_PE, bold=True, size=11, color="FFFFFFFF")
_HEADER_FILL_PE = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
# [FIX] SEBELUMNYA horizontal="center", vertical="center" saja (tanpa
# wrap_text) -- beda dgn _HEADER_ALIGN_BARU4 (dipakai sheet 13 "CALK" &
# 17 "Rekonsiliasi Fiskal", "keluarga gaya" yang SAMA) yang sudah pakai
# wrap_text=True. Disamakan supaya konsisten satu keluarga -- tidak
# berdampak ke tampilan header bulan (Jan-25 dst, selalu 1 baris) tapi
# menyamakan definisi gaya.
_HEADER_ALIGN_PE = Alignment(horizontal="center", vertical="center", wrap_text=True)

_TANGGAL_FONT_PE = Font(name=_FONT_NAME_PE, size=11, color="FF0000FF")
_TANGGAL_ALIGN_PE = Alignment(horizontal="center", vertical="center")
_FORMAT_TANGGAL_PE = r"dd\-mmm\-yyyy"

_SECTION_FONT_PE = Font(name=_FONT_NAME_PE, bold=True, size=11, color="FF17365D")
_SECTION_FILL_PE = PatternFill(start_color="FFD9EAF7", end_color="FFD9EAF7", fill_type="solid")
_SECTION_ALIGN_PE = Alignment(vertical="center", wrap_text=True)

# [FIX] vertical="top" -- SEBELUMNYA "center". Sheet 13 "CALK" & 17
# "Rekonsiliasi Fiskal" (satu keluarga gaya yang sama) SUDAH pakai
# Alignment(wrap_text=True, vertical="top") utk sel data yang wrap teks
# panjang (kolom "Uraian"). Disamakan di sini supaya kalau nama akun di
# kolom B sampai wrap 2+ baris, teksnya rata ATAS (bukan rata tengah)
# konsisten dgn 2 sheet lain.
_ITEM_ALIGN_PE = Alignment(vertical="top", wrap_text=True)
_ITEM_FONT_LABEL_PE = Font(name=_FONT_NAME_PE, size=11, color="FF0000FF")    # Kode/Uraian akun (tabel ringkasan)
_ITEM_FONT_ANGKA_PE = Font(name=_FONT_NAME_PE, size=11, color="FF008000")    # angka formula LINTAS-SHEET (NSA/GL/Laba Rugi Bulanan)
_ITEM_FONT_LOKAL_PE = Font(name=_FONT_NAME_PE, size=11, color="FF000000")    # label & angka formula LOKAL (roll-forward dlm sheet ini sendiri)

_TOTAL_FONT_PE = Font(name=_FONT_NAME_PE, bold=True, size=11, color="FF000000")
_TOTAL_FILL_PE = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
_STATUS_FILL_PE = PatternFill(start_color="FFD9EAF7", end_color="FFD9EAF7", fill_type="solid")

# [FIX] Sebelumnya sheet ini TANPA border sama sekali (beda dgn semua
# sheet lain -- BS/Neraca/dst selalu pakai border tipis di header +
# border atas tipis/bawah dobel di baris TOTAL). Disamakan di sini.
# [FIX-2] SEBELUMNYA cuma top+bottom thin TANPA left/right sama sekali
# (bahkan di kolom pertama/terakhir) -- beda dgn _HEADER_BORDER_BARU4
# (sheet 13 "CALK" & 17 "Rekonsiliasi Fiskal", KELUARGA GAYA YANG SAMA
# persis disebut bareng di komentar "sheet 11/12/13/17" di atas) yang
# full grid (top+bottom+left+right di SETIAP sel header). Disamakan
# jadi full grid supaya baris header ke-4 sheet keluarga ini identik.
_HEADER_BORDER_PE = Border(
    top=Side(style="thin"), bottom=Side(style="thin"),
    left=Side(style="thin"), right=Side(style="thin"),
)
_TOTAL_BORDER_PE = Border(top=Side(style="thin"), bottom=Side(style="double"))
# [FIX] Border tipis abu-abu di bawah tiap baris data/item -- sheet
# "CALK" (13) & "Rekonsiliasi Fiskal" (17), yang berbagi keluarga gaya
# yang sama (_..._BARU4/_..._PE), SUDAH pakai garis pemisah ini di
# setiap baris data (lihat _DATA_BORDER_BARU4) supaya tabel kelihatan
# rapi "berpenggaris". Sheet "Laporan Perubahan Ekuitas" (11) & "Laporan
# Arus Kas" (12) sebelumnya TIDAK pakai border sama sekali di baris
# item/akun -- disamakan di sini pakai warna & ketebalan identik supaya
# ke-4 sheet keluarga ini konsisten satu sama lain.
_DATA_BORDER_PE = Border(bottom=Side(style="thin", color="FFD9D9D9"))

# [FIX] Sub-judul di dalam satu section (mis. "Perubahan Modal Kerja:"
# di sheet Arus Kas) sebelumnya pakai Font ad-hoc size=10 italic --
# ukuran 10 TIDAK dipakai di sheet lain manapun di seluruh workbook ini
# (badan tabel selalu size 11, judul size 15, catatan size 9) sehingga
# terlihat "asing" dibanding sheet lain. Diganti pakai style sub-section
# resmi: size 11 (konsisten dgn badan tabel), bold, TANPA italic, warna
# navy sama seperti section utama, + fill biru muda senada _SECTION_FILL_PE
# supaya tetap kebeda dari section utama (yg full lebar & fill sama)
# tapi tetap satu keluarga visual, bukan style asing.
_SUBSECTION_FONT_PE = Font(name=_FONT_NAME_PE, bold=True, size=11, color="FF17365D")
_SUBSECTION_FILL_PE = PatternFill(start_color="FFEAF1FA", end_color="FFEAF1FA", fill_type="solid")
_SUBSECTION_ALIGN_PE = Alignment(vertical="center", wrap_text=True)

_FORMAT_RUPIAH_PE = r"#,##0;[Red]\(#,##0\);\-"

_LEBAR_KOLOM_PE = {
    "A": 6.43, "B": 38.14, "C": 12.57, "D": 12.86, "E": 12.71, "F": 12.43,
    "G": 13.14, "H": 12.57, "I": 12.43, "J": 12.86, "K": 13.14, "L": 12.71,
    "M": 12.86, "N": 13.29,
}


def _tulis_sheet_perubahan_ekuitas(
    ws,
    tahun: Any,
    peta_akun: Dict[str, Any],
    tb: Dict[str, Any],
    nsa_baris_awal: int,
    nsa_baris_akhir: int,
    gl_sheet_name: str,
    gl_baris_awal: int,
    gl_baris_akhir: int,
    r_laba_bersih_bulanan_lr: Optional[int],
    r_laba_bersih_ytd_lr: Optional[int] = None,
) -> None:
    """Sheet 11 "Laporan Perubahan Ekuitas" -- lihat catatan panjang di
    atas konstanta _..._PE utk rincian pola & sumber data. `tb`/`peta_akun`
    persis sama dgn yg dipakai sheet "Balance Sheet Bulanan" (section 10)
    supaya TOTAL EKUITAS tie-out ke situ. `r_laba_bersih_bulanan_lr` =
    nomor baris "LABA BERSIH BULANAN" di sheet "Laba Rugi Bulanan" (None
    kalau sheet itu belum ada data -- lihat deklarasi default di section 8,
    baris "Laba Tahun Berjalan" jatuh ke 0 kalau None, bukan #REF!)."""
    try:
        tahun_int = int(tahun)
    except (TypeError, ValueError):
        tahun_int = date.today().year
    thn2 = str(tahun_int)[-2:]
    ncols = 14  # Kode + Uraian + 12 bulan

    for kolom, lebar in _LEBAR_KOLOM_PE.items():
        ws.column_dimensions[kolom].width = lebar

    if not tb:
        ws.cell(row=1, column=1, value=f"LAPORAN PERUBAHAN EKUITAS TAHUN {tahun}")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        ws.cell(row=1, column=1).font = _TITLE_FONT_PE
        ws.cell(row=1, column=1).fill = _TITLE_FILL_PE
        ws.cell(row=1, column=1).alignment = _TITLE_ALIGN_PE
        ws.row_dimensions[1].height = 27.95
        ws.cell(row=3, column=1,
                 value="Belum ada data -- generate laporan bulanan terlebih dahulu.")
        return

    # -- akun ekuitas client ybs (SUMBER SAMA dgn "Balance Sheet Bulanan"
    #    section 10), dipecah "modal" vs "saldo laba/laba ditahan" --
    akun_ekuitas = [no for no in sorted(tb.keys())
                     if str(peta_akun.get(str(no), {}).get("kategori", "")).lower() == "ekuitas"]
    akun_modal: List[str] = []
    akun_saldo_laba: List[str] = []
    for no_akun in akun_ekuitas:
        info = peta_akun.get(str(no_akun), {})
        sub = str(info.get("sub_kategori") or "").strip().lower()
        nama_cek = str(info.get("nama_akun") or tb.get(no_akun, {}).get("nama_akun") or "").strip().lower()
        if sub in ("saldo laba", "laba ditahan") or (
            not sub and ("laba ditahan" in nama_cek or "saldo laba" in nama_cek)
        ):
            akun_saldo_laba.append(no_akun)
        else:
            akun_modal.append(no_akun)

    def _nama(no_akun: str) -> str:
        return (peta_akun.get(str(no_akun), {}).get("nama_akun")
                or tb.get(no_akun, {}).get("nama_akun") or no_akun)

    def _saldo_awal_tahun_formula(no_akun: str) -> str:
        """Saldo awal TAHUN (1 Jan) akun ini -- Kredit-Debit dari sheet
        "Neraca Saldo Awal" (saldo normal akun ekuitas KREDIT, kebalikan
        dari akun Aset/Beban)."""
        return (
            f"SUMIFS('Neraca Saldo Awal'!$I${nsa_baris_awal}:$I${nsa_baris_akhir},"
            f"'Neraca Saldo Awal'!$E${nsa_baris_awal}:$E${nsa_baris_akhir},\"{no_akun}\")"
            f"-SUMIFS('Neraca Saldo Awal'!$H${nsa_baris_awal}:$H${nsa_baris_akhir},"
            f"'Neraca Saldo Awal'!$E${nsa_baris_awal}:$E${nsa_baris_akhir},\"{no_akun}\")"
        )

    def _saldo_kumulatif_formula(no_akun: str, r_kode: int, surat_cutoff: str) -> str:
        """Saldo KUMULATIF (Kredit-Debit) akun ini s.d. akhir bulan pada
        kolom `surat_cutoff` (dari baris tanggal row 2) -- Saldo Awal
        Tahun + seluruh mutasi GL s.d. tanggal itu, dipakai di tabel
        ringkasan atas (baris per akun, kolom Jan..Des). $A{r_kode} = sel
        Kode akun baris ini (ditulis di kolom A tabel ringkasan)."""
        cutoff = f"{surat_cutoff}$2"
        return (
            f"SUMIFS('Neraca Saldo Awal'!$I${nsa_baris_awal}:$I${nsa_baris_akhir},"
            f"'Neraca Saldo Awal'!$E${nsa_baris_awal}:$E${nsa_baris_akhir},$A{r_kode})"
            f"-SUMIFS('Neraca Saldo Awal'!$H${nsa_baris_awal}:$H${nsa_baris_akhir},"
            f"'Neraca Saldo Awal'!$E${nsa_baris_awal}:$E${nsa_baris_akhir},$A{r_kode})"
            f"+SUMIFS('{gl_sheet_name}'!$J${gl_baris_awal}:$J${gl_baris_akhir},"
            f"'{gl_sheet_name}'!$F${gl_baris_awal}:$F${gl_baris_akhir},$A{r_kode},"
            f"'{gl_sheet_name}'!$C${gl_baris_awal}:$C${gl_baris_akhir},\"<=\"&{cutoff})"
            f"-SUMIFS('{gl_sheet_name}'!$I${gl_baris_awal}:$I${gl_baris_akhir},"
            f"'{gl_sheet_name}'!$F${gl_baris_awal}:$F${gl_baris_akhir},$A{r_kode},"
            f"'{gl_sheet_name}'!$C${gl_baris_awal}:$C${gl_baris_akhir},\"<=\"&{cutoff})"
        )

    def _penambahan_bulan_formula(no_akun: str, surat: str) -> str:
        """Mutasi KREDIT bulan pada kolom `surat` saja (bukan kumulatif)
        utk akun ini di sheet "GL <tahun>" -- dipakai baris "Penambahan"
        di section rincian roll-forward."""
        return (
            f"SUMIFS('{gl_sheet_name}'!$J${gl_baris_awal}:$J${gl_baris_akhir},"
            f"'{gl_sheet_name}'!$F${gl_baris_awal}:$F${gl_baris_akhir},\"{no_akun}\","
            f"'{gl_sheet_name}'!$C${gl_baris_awal}:$C${gl_baris_akhir},"
            f"\">=\"&DATE(YEAR({surat}$2),MONTH({surat}$2),1),"
            f"'{gl_sheet_name}'!$C${gl_baris_awal}:$C${gl_baris_akhir},\"<=\"&{surat}$2)"
        )

    def _pengurangan_bulan_formula(no_akun: str, surat: str) -> str:
        """Sama seperti _penambahan_bulan_formula tapi kolom DEBIT (I) --
        dipakai baris "Pengurangan"."""
        return (
            f"SUMIFS('{gl_sheet_name}'!$I${gl_baris_awal}:$I${gl_baris_akhir},"
            f"'{gl_sheet_name}'!$F${gl_baris_awal}:$F${gl_baris_akhir},\"{no_akun}\","
            f"'{gl_sheet_name}'!$C${gl_baris_awal}:$C${gl_baris_akhir},"
            f"\">=\"&DATE(YEAR({surat}$2),MONTH({surat}$2),1),"
            f"'{gl_sheet_name}'!$C${gl_baris_awal}:$C${gl_baris_akhir},\"<=\"&{surat}$2)"
        )

    # ================= Baris 1: judul =================
    ws.cell(row=1, column=1, value=f"LAPORAN PERUBAHAN EKUITAS TAHUN {tahun}")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1).font = _TITLE_FONT_PE
    ws.cell(row=1, column=1).fill = _TITLE_FILL_PE
    ws.cell(row=1, column=1).alignment = _TITLE_ALIGN_PE
    ws.row_dimensions[1].height = 27.95

    # ================= Baris 2: tanggal akhir tiap bulan =================
    for m in range(1, 13):
        col = 2 + m
        tgl = date(tahun_int, m, calendar.monthrange(tahun_int, m)[1])
        c = ws.cell(row=2, column=col, value=tgl)
        c.font = _TANGGAL_FONT_PE
        c.alignment = _TANGGAL_ALIGN_PE
        c.number_format = _FORMAT_TANGGAL_PE

    # ================= Baris 3: header Kode/Uraian/Jan-25..Dec-25 =========
    headers = ["Kode", "Uraian"] + [f"{b}-{thn2}" for b in _BULAN_SINGKAT]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = _HEADER_FONT_PE
        c.fill = _HEADER_FILL_PE
        c.alignment = _HEADER_ALIGN_PE
        c.border = _HEADER_BORDER_PE
    ws.row_dimensions[3].height = 30

    r = 3
    baris_ringkasan: List[int] = []

    # -- baris 4 dst: satu baris per akun ekuitas (modal dulu, lalu saldo
    #    laba), saldo KUMULATIF per akhir bulan --
    for no_akun in akun_modal + akun_saldo_laba:
        r += 1
        ws.cell(row=r, column=1, value=no_akun).font = _ITEM_FONT_LABEL_PE
        ws.cell(row=r, column=2, value=_nama(no_akun)).font = _ITEM_FONT_LABEL_PE
        ws.cell(row=r, column=1).alignment = _ITEM_ALIGN_PE
        ws.cell(row=r, column=2).alignment = _ITEM_ALIGN_PE
        ws.cell(row=r, column=1).border = _DATA_BORDER_PE
        ws.cell(row=r, column=2).border = _DATA_BORDER_PE
        for m in range(1, 13):
            col = 2 + m
            surat = get_column_letter(col)
            cell = ws.cell(row=r, column=col,
                             value=f"={_saldo_kumulatif_formula(no_akun, r, surat)}")
            cell.font = _ITEM_FONT_ANGKA_PE
            cell.alignment = _ITEM_ALIGN_PE
            cell.border = _DATA_BORDER_PE
            cell.number_format = _FORMAT_RUPIAH_PE
        baris_ringkasan.append(r)

    # -- baris "Laba Tahun Berjalan" (bukan akun COA -- link ke "LABA
    #    BERSIH BULANAN" di sheet "Laba Rugi Bulanan", pola sama dgn baris
    #    "Laba Tahun Berjalan" di sheet "Balance Sheet Bulanan") --
    r_laba_tahun_berjalan_top = None
    if r_laba_bersih_bulanan_lr:
        r += 1
        ws.cell(row=r, column=2, value="Laba Tahun Berjalan").font = _ITEM_FONT_LABEL_PE
        ws.cell(row=r, column=2).alignment = _ITEM_ALIGN_PE
        ws.cell(row=r, column=1).border = _DATA_BORDER_PE
        ws.cell(row=r, column=2).border = _DATA_BORDER_PE
        # [FIX BUG] Baris ini komponen TOTAL EKUITAS -- HARUS kumulatif
        # (LABA BERSIH YTD), bukan angka satu bulan (LABA BERSIH BULANAN)
        # -- lihat catatan panjang di deklarasi r_laba_bersih_ytd_lr.
        # Fallback ke r_laba_bersih_bulanan_lr kalau utk suatu alasan
        # baris YTD tidak tersedia (lebih baik ketimbang jatuh ke 0).
        r_sumber_laba_top = r_laba_bersih_ytd_lr or r_laba_bersih_bulanan_lr
        for m in range(1, 13):
            col = 2 + m
            surat = get_column_letter(col)
            cell = ws.cell(row=r, column=col,
                             value=f"='Laba Rugi Bulanan'!{surat}{r_sumber_laba_top}")
            cell.font = _ITEM_FONT_ANGKA_PE
            cell.alignment = _ITEM_ALIGN_PE
            cell.number_format = _FORMAT_RUPIAH_PE
            cell.border = _DATA_BORDER_PE
        r_laba_tahun_berjalan_top = r
        baris_ringkasan.append(r)

    # -- TOTAL EKUITAS --
    r += 1
    r_total_ekuitas = r
    ws.cell(row=r, column=2, value="TOTAL EKUITAS")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = _TOTAL_FONT_PE
        cell.fill = _TOTAL_FILL_PE
        cell.alignment = _ITEM_ALIGN_PE
        cell.border = _TOTAL_BORDER_PE
    for m in range(1, 13):
        col = 2 + m
        surat = get_column_letter(col)
        formula = ("=" + "+".join(f"{surat}{rr}" for rr in baris_ringkasan)
                   if baris_ringkasan else "=0")
        cell = ws.cell(row=r, column=col, value=formula)
        cell.number_format = _FORMAT_RUPIAH_PE
    ws.row_dimensions[r].height = 15.75

    # ================= spacer + RINCIAN PERUBAHAN MODAL =================
    r += 2
    ws.cell(row=r, column=1, value="RINCIAN PERUBAHAN MODAL")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = _SECTION_FONT_PE
        cell.fill = _SECTION_FILL_PE
        cell.alignment = _SECTION_ALIGN_PE

    baris_saldo_akhir_modal: List[int] = []
    for no_akun in akun_modal:
        nama = _nama(no_akun)
        r += 1
        r_awal = r
        ws.cell(row=r, column=2, value=f"Saldo Awal {nama}")
        r += 1
        r_tambah = r
        ws.cell(row=r, column=2, value=f"Penambahan {nama}")
        r += 1
        r_kurang = r
        ws.cell(row=r, column=2, value=f"Pengurangan {nama}")
        r += 1
        r_akhir = r
        ws.cell(row=r, column=2, value=f"Saldo Akhir {nama}")
        baris_saldo_akhir_modal.append(r_akhir)

        for rr in (r_awal, r_tambah, r_kurang, r_akhir):
            ws.cell(row=rr, column=2).font = _ITEM_FONT_LOKAL_PE
            ws.cell(row=rr, column=2).alignment = _ITEM_ALIGN_PE
            ws.cell(row=rr, column=1).border = _DATA_BORDER_PE
            ws.cell(row=rr, column=2).border = _DATA_BORDER_PE

        for m in range(1, 13):
            col = 2 + m
            surat = get_column_letter(col)
            prev_surat = get_column_letter(col - 1)

            f_awal = (f"={_saldo_awal_tahun_formula(no_akun)}" if m == 1
                      else f"={prev_surat}{r_akhir}")
            cell = ws.cell(row=r_awal, column=col, value=f_awal)
            cell.font = _ITEM_FONT_LOKAL_PE
            cell.number_format = _FORMAT_RUPIAH_PE
            cell.alignment = _ITEM_ALIGN_PE
            cell.border = _DATA_BORDER_PE

            cell = ws.cell(row=r_tambah, column=col,
                             value=f"={_penambahan_bulan_formula(no_akun, surat)}")
            cell.font = _ITEM_FONT_ANGKA_PE
            cell.number_format = _FORMAT_RUPIAH_PE
            cell.alignment = _ITEM_ALIGN_PE
            cell.border = _DATA_BORDER_PE

            cell = ws.cell(row=r_kurang, column=col,
                             value=f"={_pengurangan_bulan_formula(no_akun, surat)}")
            cell.font = _ITEM_FONT_ANGKA_PE
            cell.number_format = _FORMAT_RUPIAH_PE
            cell.alignment = _ITEM_ALIGN_PE
            cell.border = _DATA_BORDER_PE

            cell = ws.cell(row=r_akhir, column=col,
                             value=f"={surat}{r_awal}+{surat}{r_tambah}-{surat}{r_kurang}")
            cell.font = _ITEM_FONT_LOKAL_PE
            cell.number_format = _FORMAT_RUPIAH_PE
            cell.alignment = _ITEM_ALIGN_PE
            cell.border = _DATA_BORDER_PE

    # -- TOTAL SALDO AKHIR MODAL --
    r += 1
    r_total_modal = r
    ws.cell(row=r, column=2, value="TOTAL SALDO AKHIR MODAL")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = _TOTAL_FONT_PE
        cell.fill = _TOTAL_FILL_PE
        cell.alignment = _ITEM_ALIGN_PE
        cell.border = _TOTAL_BORDER_PE
    for m in range(1, 13):
        col = 2 + m
        surat = get_column_letter(col)
        formula = ("=" + "+".join(f"{surat}{rr}" for rr in baris_saldo_akhir_modal)
                   if baris_saldo_akhir_modal else "=0")
        cell = ws.cell(row=r, column=col, value=formula)
        cell.number_format = _FORMAT_RUPIAH_PE
    ws.row_dimensions[r].height = 15.75

    # ================= spacer + RINCIAN PERUBAHAN SALDO LABA ============
    r += 2
    ws.cell(row=r, column=1, value="RINCIAN PERUBAHAN SALDO LABA")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = _SECTION_FONT_PE
        cell.fill = _SECTION_FILL_PE
        cell.alignment = _SECTION_ALIGN_PE

    r += 1
    r_sl_awal = r
    ws.cell(row=r, column=2, value="Saldo Awal Saldo Laba")
    r += 1
    r_sl_laba_berjalan = r
    ws.cell(row=r, column=2, value="Laba Tahun Berjalan")
    r += 1
    r_sl_kurang = r
    ws.cell(row=r, column=2, value="Pengurangan Saldo Laba / Dividen / Prive")
    r += 1
    r_sl_akhir = r
    ws.cell(row=r, column=2, value="Saldo Akhir Saldo Laba")

    ws.cell(row=r_sl_awal, column=2).font = _ITEM_FONT_LOKAL_PE
    ws.cell(row=r_sl_laba_berjalan, column=2).font = _ITEM_FONT_ANGKA_PE
    ws.cell(row=r_sl_kurang, column=2).font = _ITEM_FONT_ANGKA_PE
    for rr in (r_sl_awal, r_sl_laba_berjalan, r_sl_kurang):
        ws.cell(row=rr, column=2).alignment = _ITEM_ALIGN_PE
        ws.cell(row=rr, column=1).border = _DATA_BORDER_PE
        ws.cell(row=rr, column=2).border = _DATA_BORDER_PE

    for m in range(1, 13):
        col = 2 + m
        surat = get_column_letter(col)
        prev_surat = get_column_letter(col - 1)

        if m == 1:
            f_awal = ("=" + "+".join(_saldo_awal_tahun_formula(no) for no in akun_saldo_laba)
                      if akun_saldo_laba else "=0")
        else:
            f_awal = f"={prev_surat}{r_sl_akhir}"
        cell = ws.cell(row=r_sl_awal, column=col, value=f_awal)
        cell.font = _ITEM_FONT_LOKAL_PE
        cell.number_format = _FORMAT_RUPIAH_PE
        cell.alignment = _ITEM_ALIGN_PE
        cell.border = _DATA_BORDER_PE

        f_laba = (f"='Laba Rugi Bulanan'!{surat}{r_laba_bersih_bulanan_lr}"
                  if r_laba_bersih_bulanan_lr else "=0")
        cell = ws.cell(row=r_sl_laba_berjalan, column=col, value=f_laba)
        cell.font = _ITEM_FONT_ANGKA_PE
        cell.number_format = _FORMAT_RUPIAH_PE
        cell.alignment = _ITEM_ALIGN_PE
        cell.border = _DATA_BORDER_PE

        if akun_saldo_laba:
            bagian_kredit = "+".join(_penambahan_bulan_formula(no, surat) for no in akun_saldo_laba)
            bagian_debit = "+".join(_pengurangan_bulan_formula(no, surat) for no in akun_saldo_laba)
            f_kurang = f"=-({bagian_debit})+({bagian_kredit})"
        else:
            f_kurang = "=0"
        cell = ws.cell(row=r_sl_kurang, column=col, value=f_kurang)
        cell.font = _ITEM_FONT_ANGKA_PE
        cell.number_format = _FORMAT_RUPIAH_PE
        cell.alignment = _ITEM_ALIGN_PE
        cell.border = _DATA_BORDER_PE

        # [FIX BUG] Sebelumnya rumus ini cuma Saldo Awal + Pengurangan,
        # TIDAK ikut menjumlahkan "Laba Tahun Berjalan" (r_sl_laba_berjalan)
        # -- padahal Laba Tahun Berjalan BUKAN akun COA & tidak pernah
        # tercermin lewat mutasi GL (_penambahan/_pengurangan_bulan_formula
        # yg dipakai r_sl_kurang), jadi angka Saldo Akhir Saldo Laba yang
        # ditampilkan ke user understated sebesar laba tahun berjalan.
        # Bug ini SEBELUMNYA tidak kelihatan dari STATUS "BALANCE" krn baris
        # CHECK BALANCE di bawah menambahkan r_sl_laba_berjalan scr
        # terpisah supaya check-nya nol -- itu menutupi bug, bukan
        # memperbaikinya; angka baris ini sendiri tetap salah.
        cell = ws.cell(row=r_sl_akhir, column=col,
                         value=f"={surat}{r_sl_awal}+{surat}{r_sl_laba_berjalan}+{surat}{r_sl_kurang}")
        cell.number_format = _FORMAT_RUPIAH_PE
        cell.alignment = _ITEM_ALIGN_PE

    # -- "Saldo Akhir Saldo Laba" ditutup sbg baris TOTAL (bold, fill
    #    D9E1F2 -- sama pola dgn TOTAL EKUITAS/TOTAL SALDO AKHIR MODAL di
    #    atas, override style angka yg baru ditulis di loop sebelumnya).
    #    [FIX] number_format sebelumnya "General" (beda dgn semua baris
    #    TOTAL lain di workbook yg selalu pakai format Rupiah) -- disamakan.
    for c in range(1, ncols + 1):
        cell = ws.cell(row=r_sl_akhir, column=c)
        cell.font = _TOTAL_FONT_PE
        cell.fill = _TOTAL_FILL_PE
        cell.alignment = _ITEM_ALIGN_PE
        cell.border = _TOTAL_BORDER_PE
        cell.number_format = _FORMAT_RUPIAH_PE
    ws.row_dimensions[r_sl_akhir].height = 15.75

    # ================= spacer + CHECK BALANCE + STATUS ==================
    r += 2
    r_check = r
    ws.cell(row=r, column=2, value="CHECK BALANCE")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = _TOTAL_FONT_PE
        cell.fill = _TOTAL_FILL_PE
        cell.alignment = _ITEM_ALIGN_PE
        cell.border = _TOTAL_BORDER_PE
    # [FIX] r_sl_akhir SEKARANG sudah ikut menjumlahkan r_sl_laba_berjalan
    # (lihat fix di atas) -- kalau r_sl_laba_berjalan tetap ditambahkan
    # terpisah di sini, jadi dobel hitung & CHECK BALANCE salah flag
    # "PERIKSA" walau laporan sebenarnya balance. Dihapus dari rumus ini.
    for m in range(1, 13):
        col = 2 + m
        surat = get_column_letter(col)
        formula = f"={surat}{r_total_ekuitas}-({surat}{r_total_modal}+{surat}{r_sl_akhir})"
        cell = ws.cell(row=r, column=col, value=formula)
        cell.number_format = _FORMAT_RUPIAH_PE
    ws.row_dimensions[r].height = 15.75

    r += 1
    r_status = r
    ws.cell(row=r, column=2, value="STATUS")
    ws.cell(row=r, column=2).font = _ITEM_FONT_LOKAL_PE
    ws.cell(row=r, column=2).alignment = _ITEM_ALIGN_PE
    for m in range(1, 13):
        col = 2 + m
        surat = get_column_letter(col)
        formula = f'=IF(ROUND({surat}{r_check},0)=0,"BALANCE","PERIKSA")'
        cell = ws.cell(row=r, column=col, value=formula)
        cell.font = _TOTAL_FONT_PE
        cell.fill = _STATUS_FILL_PE
        cell.alignment = _ITEM_ALIGN_PE
    ws.row_dimensions[r].height = 15.75

    # [BARU] Highlight otomatis hijau/merah di baris STATUS -- sheet ini
    # 100% rumus Excel hidup (beda dgn Balance Sheet Bulanan yg nilai
    # check-nya sudah dihitung di Python jadi warnanya bisa langsung
    # ditentukan saat generate), jadi dipakai conditional formatting Excel
    # supaya warnanya tetap ikut update kalau user edit manual di Excel.
    rentang_status = f"C{r_status}:N{r_status}"
    ws.conditional_formatting.add(
        rentang_status,
        CellIsRule(operator="equal", formula=['"BALANCE"'],
                   font=Font(name=_FONT_NAME_PE, bold=True, size=11, color="FF008000")),
    )
    ws.conditional_formatting.add(
        rentang_status,
        CellIsRule(operator="equal", formula=['"PERIKSA"'],
                   font=Font(name=_FONT_NAME_PE, bold=True, size=11, color="FFFF0000")),
    )

def _tulis_sheet_arus_kas(
    ws,
    tahun: Any,
    peta_akun: Dict[str, Any],
    tb: Dict[str, Any],
    nsa_baris_awal: int,
    nsa_baris_akhir: int,
    gl_sheet_name: str,
    gl_baris_awal: int,
    gl_baris_akhir: int,
    r_laba_bersih_bulanan_lr: Optional[int],
    r_total_penyusutan_lr: Optional[int],
) -> None:
    """Sheet 12 "Laporan Arus Kas" -- [FIX/REBUILD] Sebelumnya sheet ini
    cuma daftar mentah transaksi kas per kategori (5 kolom Keterangan/
    Referensi/Jumlah/Kategori/Subtotal, angka Python beku dari
    lapkeu.susun_arus_kas_sederhana()) -- TIDAK konsisten dgn "sheet-sheet
    lainnya" yg sudah dibangun ulang (Trial Balance/Laba Rugi/Balance
    Sheet Bulanan, Laporan Perubahan Ekuitas -- section 8-11): semuanya
    tabel bulanan Kode/Uraian/Jan..Des dgn RUMUS EXCEL HIDUP (bukan angka
    beku) yg saling tie-out via SUMIFS lintas sheet "Neraca Saldo Awal" &
    "GL {tahun}", ditutup baris CHECK/STATUS BALANCE-PERIKSA. Sheet ini
    sekarang disamakan polanya: Laporan Arus Kas METODE TIDAK LANGSUNG
    (indirect method, standar PSAK) dgn 3 aktivitas (Operasi/Investasi/
    Pendanaan), tabel bulanan Jan..Des, DITUTUP verifikasi independen
    (saldo kas hasil arus kas vs saldo kas aktual dari GL) supaya kalau
    ada mutasi kas yg belum terklasifikasi dgn benar, sheet ini otomatis
    kelihatan "PERIKSA" bukan diam-diam salah. Konstanta visual (_..._PE)
    SENGAJA dipakai bareng dgn sheet 11 (bukan didup) supaya kedua sheet
    baru ini (yg sama-sama belum punya file model referensi, strukturnya
    dikonfirmasi user) tetap konsisten satu sama lain & dgn workbook.

    Params sama persis dgn `_tulis_sheet_perubahan_ekuitas()` (dipanggil
    dgn `tb`/`peta_akun`/range NSA & GL yang SAMA supaya akun yg dipakai
    tie-out ke sheet lain), ditambah `r_total_penyusutan_lr` = nomor
    baris "TOTAL PENYUSUTAN" di sheet "Laba Rugi Bulanan" (utk baris
    penyesuaian non-kas "Beban Penyusutan").

    Klasifikasi akun (dari Coa.kategori/sub_kategori, konvensi bebas-isi
    yang SAMA dipakai sheet "Balance Sheet Bulanan"/"Laba Rugi Bulanan"):
    Kas/Bank (dikeluarkan dari modal kerja, jadi baris Awal/Akhir Kas
    tersendiri) -> Piutang/Persediaan/Aset Lancar Lainnya (Operasi) ->
    Aset Tetap (Investasi) -> Hutang Usaha/Liabilitas Lainnya (Operasi)
    -> Modal non-Saldo Laba (Pendanaan). "Saldo Laba"/"Laba Ditahan"
    SENGAJA tidak diikutkan sebagai baris tersendiri (sudah 100% dijelaskan
    lewat "Laba Bersih Tahun Berjalan" -- kalau ada penyesuaian manual ke
    Saldo Laba di luar itu, baris verifikasi di bawah akan menangkapnya
    sbg selisih, bukan diam-diam diabaikan)."""
    ncols = 14  # Kode + Uraian + 12 bulan

    for kolom, lebar in _LEBAR_KOLOM_PE.items():
        ws.column_dimensions[kolom].width = lebar
    # [FIX] Kolom "Uraian" (B) di sheet ini menampung label GABUNGAN
    # "{label baris} -- {nama akun}" (mis. "Kenaikan/(Penurunan) Hutang
    # Usaha -- Hutang Usaha") yang jauh lebih panjang drpd label sheet 11
    # "Laporan Perubahan Ekuitas" (mis. "Saldo Awal <nama>") yang sama-
    # sama pakai lebar kolom B 38.14 dari _LEBAR_KOLOM_PE -- lebar itu
    # kependekan utk sheet ini & bikin teks kepotong/wrap berlebihan.
    # Dilebarkan KHUSUS utk sheet Arus Kas saja (tidak ganggu sheet 11).
    ws.column_dimensions["B"].width = 46

    if not tb:
        ws.cell(row=1, column=1, value=f"LAPORAN ARUS KAS TAHUN {tahun}")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        ws.cell(row=1, column=1).font = _TITLE_FONT_PE
        ws.cell(row=1, column=1).fill = _TITLE_FILL_PE
        ws.cell(row=1, column=1).alignment = _TITLE_ALIGN_PE
        ws.row_dimensions[1].height = 27.95
        ws.cell(row=3, column=1,
                 value="Belum ada data -- generate laporan bulanan terlebih dahulu.")
        return

    try:
        tahun_int = int(tahun)
    except (TypeError, ValueError):
        tahun_int = date.today().year
    thn2 = str(tahun_int)[-2:]

    # ================= Baris 1: judul =================
    ws.cell(row=1, column=1, value=f"LAPORAN ARUS KAS TAHUN {tahun}")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1).font = _TITLE_FONT_PE
    ws.cell(row=1, column=1).fill = _TITLE_FILL_PE
    ws.cell(row=1, column=1).alignment = _TITLE_ALIGN_PE
    ws.row_dimensions[1].height = 27.95

    # ================= Baris 2: tanggal akhir tiap bulan =================
    for m in range(1, 13):
        col = 2 + m
        tgl = date(tahun_int, m, calendar.monthrange(tahun_int, m)[1])
        c = ws.cell(row=2, column=col, value=tgl)
        c.font = _TANGGAL_FONT_PE
        c.alignment = _TANGGAL_ALIGN_PE
        c.number_format = _FORMAT_TANGGAL_PE

    # ================= Baris 3: header Kode/Uraian/Jan-25..Dec-25 =========
    headers = ["Kode", "Uraian"] + [f"{b}-{thn2}" for b in _BULAN_SINGKAT]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = _HEADER_FONT_PE
        c.fill = _HEADER_FILL_PE
        c.alignment = _HEADER_ALIGN_PE
        c.border = _HEADER_BORDER_PE
    ws.row_dimensions[3].height = 30

    # ================= helper rumus (pola sama dgn sheet 11) =============
    def _debit_bulan_formula(no_akun: str, surat: str) -> str:
        return (
            f"SUMIFS('{gl_sheet_name}'!$I${gl_baris_awal}:$I${gl_baris_akhir},"
            f"'{gl_sheet_name}'!$F${gl_baris_awal}:$F${gl_baris_akhir},\"{no_akun}\","
            f"'{gl_sheet_name}'!$C${gl_baris_awal}:$C${gl_baris_akhir},"
            f"\">=\"&DATE(YEAR({surat}$2),MONTH({surat}$2),1),"
            f"'{gl_sheet_name}'!$C${gl_baris_awal}:$C${gl_baris_akhir},\"<=\"&{surat}$2)"
        )

    def _kredit_bulan_formula(no_akun: str, surat: str) -> str:
        return (
            f"SUMIFS('{gl_sheet_name}'!$J${gl_baris_awal}:$J${gl_baris_akhir},"
            f"'{gl_sheet_name}'!$F${gl_baris_awal}:$F${gl_baris_akhir},\"{no_akun}\","
            f"'{gl_sheet_name}'!$C${gl_baris_awal}:$C${gl_baris_akhir},"
            f"\">=\"&DATE(YEAR({surat}$2),MONTH({surat}$2),1),"
            f"'{gl_sheet_name}'!$C${gl_baris_awal}:$C${gl_baris_akhir},\"<=\"&{surat}$2)"
        )

    def _efek_kas_bulan_formula(no_akun: str, surat: str) -> str:
        """Efek kas bulan ini dari mutasi akun ini = Mutasi Kredit -
        Mutasi Debit. SATU rumus ini berlaku utk SEMUA akun non-kas
        (aset lain/liabilitas/modal): akun aset naik (mutasi debit>kredit)
        -> efek kas negatif (kas dipakai); akun liabilitas/modal naik
        (mutasi kredit>debit) -> efek kas positif (kas didapat) -- kedua
        arah sama-sama "kredit-debit" krn arah normal saldo yg berlawanan
        sudah otomatis membalik tandanya."""
        return f"(({_kredit_bulan_formula(no_akun, surat)})-({_debit_bulan_formula(no_akun, surat)}))"

    def _saldo_awal_tahun_formula(no_akun: str, normal: str) -> str:
        debit = (
            f"SUMIFS('Neraca Saldo Awal'!$H${nsa_baris_awal}:$H${nsa_baris_akhir},"
            f"'Neraca Saldo Awal'!$E${nsa_baris_awal}:$E${nsa_baris_akhir},\"{no_akun}\")"
        )
        kredit = (
            f"SUMIFS('Neraca Saldo Awal'!$I${nsa_baris_awal}:$I${nsa_baris_akhir},"
            f"'Neraca Saldo Awal'!$E${nsa_baris_awal}:$E${nsa_baris_akhir},\"{no_akun}\")"
        )
        return f"({debit}-{kredit})" if normal == "debit" else f"({kredit}-{debit})"

    def _saldo_kumulatif_formula(no_akun: str, surat: str, normal: str) -> str:
        """Saldo KUMULATIF (arah normal `normal`) akun ini s.d. akhir
        bulan pada kolom `surat` -- Saldo Awal Tahun + seluruh mutasi GL
        s.d. tanggal itu. Dipakai baris verifikasi saldo kas independen."""
        cutoff = f"{surat}$2"
        awal = _saldo_awal_tahun_formula(no_akun, normal)
        deb_kum = (
            f"SUMIFS('{gl_sheet_name}'!$I${gl_baris_awal}:$I${gl_baris_akhir},"
            f"'{gl_sheet_name}'!$F${gl_baris_awal}:$F${gl_baris_akhir},\"{no_akun}\","
            f"'{gl_sheet_name}'!$C${gl_baris_awal}:$C${gl_baris_akhir},\"<=\"&{cutoff})"
        )
        kred_kum = (
            f"SUMIFS('{gl_sheet_name}'!$J${gl_baris_awal}:$J${gl_baris_akhir},"
            f"'{gl_sheet_name}'!$F${gl_baris_awal}:$F${gl_baris_akhir},\"{no_akun}\","
            f"'{gl_sheet_name}'!$C${gl_baris_awal}:$C${gl_baris_akhir},\"<=\"&{cutoff})"
        )
        mutasi = f"({deb_kum}-{kred_kum})" if normal == "debit" else f"({kred_kum}-{deb_kum})"
        return f"({awal}+{mutasi})"

    # ================= Klasifikasi akun (kategori/sub_kategori COA) ======
    akun_kas_bank: List[str] = []
    akun_piutang: List[str] = []
    akun_persediaan: List[str] = []
    akun_aset_tetap: List[str] = []
    akun_aset_lancar_lain: List[str] = []
    akun_hutang_usaha: List[str] = []
    akun_liabilitas_lain: List[str] = []
    akun_modal: List[str] = []

    for no_akun in sorted(tb.keys()):
        info = peta_akun.get(str(no_akun), {})
        kategori = str(info.get("kategori") or "").strip().lower()
        sub = str(info.get("sub_kategori") or "").strip().lower()
        nama_upper = str(info.get("nama_akun") or tb.get(no_akun, {}).get("nama_akun") or "").upper()
        if kategori == "aset":
            if "akumulasi" in sub or "akum." in nama_upper.lower() or "AKUMULASI" in nama_upper:
                # [PENTING] Akun kontra-aset (mis. Akumulasi Penyusutan)
                # SENGAJA dilewati -- mutasinya (nambah via jurnal
                # penyusutan) 100% non-kas & sudah tercermin lewat baris
                # "Penyesuaian: Beban Penyusutan" di Aktivitas Operasi.
                # Kalau ikut dihitung di sini juga (mis. digabung ke
                # kelompok "aset tetap"), penyusutan akan double-count --
                # sekali via add-back operasi, sekali lagi via efek kas
                # semu di investasi.
                continue
            if sub in ("kas", "bank") or "KAS" in nama_upper or "BANK" in nama_upper:
                akun_kas_bank.append(no_akun)
            elif "piutang" in sub or "PIUTANG" in nama_upper:
                akun_piutang.append(no_akun)
            elif "persediaan" in sub or "inventori" in sub or "PERSEDIAAN" in nama_upper:
                akun_persediaan.append(no_akun)
            elif "tetap" in sub:
                akun_aset_tetap.append(no_akun)
            else:
                akun_aset_lancar_lain.append(no_akun)
        elif kategori in ("liabilitas", "kewajiban"):
            if ("hutang usaha" in sub or "utang usaha" in sub
                    or "HUTANG USAHA" in nama_upper or "UTANG USAHA" in nama_upper):
                akun_hutang_usaha.append(no_akun)
            else:
                akun_liabilitas_lain.append(no_akun)
        elif kategori == "ekuitas":
            nama_cek = nama_upper.lower()
            if sub in ("saldo laba", "laba ditahan") or (
                not sub and ("laba ditahan" in nama_cek or "saldo laba" in nama_cek)
            ):
                pass  # sengaja tidak diikutkan -- lihat catatan docstring
            else:
                akun_modal.append(no_akun)
        # kategori PENDAPATAN/BEBAN diabaikan (sudah tercermin di Laba Bersih)

    def _nama(no_akun: str) -> str:
        return (peta_akun.get(str(no_akun), {}).get("nama_akun")
                or tb.get(no_akun, {}).get("nama_akun") or no_akun)

    r = 3

    def _tulis_section(label: str) -> None:
        nonlocal r
        r += 1
        ws.cell(row=r, column=1, value=label)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = _SECTION_FONT_PE
            cell.fill = _SECTION_FILL_PE
            cell.alignment = _SECTION_ALIGN_PE

    def _tulis_baris_akun(no_akun: str, label: str) -> int:
        nonlocal r
        r += 1
        ws.cell(row=r, column=1, value=no_akun).font = _ITEM_FONT_LABEL_PE
        ws.cell(row=r, column=2, value=f"{label} -- {_nama(no_akun)}").font = _ITEM_FONT_LABEL_PE
        ws.cell(row=r, column=1).alignment = _ITEM_ALIGN_PE
        ws.cell(row=r, column=2).alignment = _ITEM_ALIGN_PE
        ws.cell(row=r, column=1).border = _DATA_BORDER_PE
        ws.cell(row=r, column=2).border = _DATA_BORDER_PE
        for m in range(1, 13):
            col = 2 + m
            surat = get_column_letter(col)
            cell = ws.cell(row=r, column=col, value=f"={_efek_kas_bulan_formula(no_akun, surat)}")
            cell.font = _ITEM_FONT_ANGKA_PE
            cell.number_format = _FORMAT_RUPIAH_PE
            cell.alignment = _ITEM_ALIGN_PE
            cell.border = _DATA_BORDER_PE
        return r

    def _tulis_baris_link(label: str, sheet_sumber: str, r_sumber: Optional[int]) -> int:
        """Baris LOKAL yg nilainya link RUMUS ke sheet lain di workbook
        yang sama (mis. Laba Bersih Bulanan / Total Penyusutan dari sheet
        "Laba Rugi Bulanan"), bukan hasil klasifikasi akun COA."""
        nonlocal r
        r += 1
        ws.cell(row=r, column=2, value=label).font = _ITEM_FONT_LOKAL_PE
        ws.cell(row=r, column=2).alignment = _ITEM_ALIGN_PE
        ws.cell(row=r, column=1).border = _DATA_BORDER_PE
        ws.cell(row=r, column=2).border = _DATA_BORDER_PE
        for m in range(1, 13):
            col = 2 + m
            surat = get_column_letter(col)
            f = f"='{sheet_sumber}'!{surat}{r_sumber}" if r_sumber else "=0"
            cell = ws.cell(row=r, column=col, value=f)
            cell.font = _ITEM_FONT_ANGKA_PE
            cell.number_format = _FORMAT_RUPIAH_PE
            cell.alignment = _ITEM_ALIGN_PE
            cell.border = _DATA_BORDER_PE
        return r

    def _tulis_baris_kosong(label: str) -> int:
        nonlocal r
        r += 1
        ws.cell(row=r, column=2, value=label).font = _ITEM_FONT_LOKAL_PE
        ws.cell(row=r, column=2).alignment = _ITEM_ALIGN_PE
        ws.cell(row=r, column=1).border = _DATA_BORDER_PE
        ws.cell(row=r, column=2).border = _DATA_BORDER_PE
        for m in range(1, 13):
            col = 2 + m
            cell = ws.cell(row=r, column=col, value=0)
            cell.number_format = _FORMAT_RUPIAH_PE
            cell.border = _DATA_BORDER_PE
        return r

    def _tulis_subtotal(label: str, baris_list: List[int]) -> int:
        nonlocal r
        r += 1
        r_sub = r
        ws.cell(row=r, column=2, value=label)
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = _TOTAL_FONT_PE
            cell.fill = _TOTAL_FILL_PE
            cell.alignment = _ITEM_ALIGN_PE
            cell.border = _TOTAL_BORDER_PE
        for m in range(1, 13):
            col = 2 + m
            surat = get_column_letter(col)
            formula = ("=" + "+".join(f"{surat}{rr}" for rr in baris_list)) if baris_list else "=0"
            cell = ws.cell(row=r, column=col, value=formula)
            cell.number_format = _FORMAT_RUPIAH_PE
        ws.row_dimensions[r].height = 15.75
        return r_sub

    # ================= ARUS KAS DARI AKTIVITAS OPERASI ====================
    _tulis_section("ARUS KAS DARI AKTIVITAS OPERASI")
    baris_operasi = [
        _tulis_baris_link("Laba Bersih Tahun Berjalan", "Laba Rugi Bulanan", r_laba_bersih_bulanan_lr),
        _tulis_baris_link("Penyesuaian: Beban Penyusutan", "Laba Rugi Bulanan", r_total_penyusutan_lr),
    ]
    r += 1
    ws.cell(row=r, column=1, value="Perubahan Modal Kerja:")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = _SUBSECTION_FONT_PE
        cell.fill = _SUBSECTION_FILL_PE
        cell.alignment = _SUBSECTION_ALIGN_PE
    for no_akun in akun_piutang:
        baris_operasi.append(_tulis_baris_akun(no_akun, "(Kenaikan)/Penurunan Piutang Usaha"))
    for no_akun in akun_persediaan:
        baris_operasi.append(_tulis_baris_akun(no_akun, "(Kenaikan)/Penurunan Persediaan"))
    for no_akun in akun_aset_lancar_lain:
        baris_operasi.append(_tulis_baris_akun(no_akun, "(Kenaikan)/Penurunan Aset Lancar Lainnya"))
    for no_akun in akun_hutang_usaha:
        baris_operasi.append(_tulis_baris_akun(no_akun, "Kenaikan/(Penurunan) Hutang Usaha"))
    for no_akun in akun_liabilitas_lain:
        baris_operasi.append(_tulis_baris_akun(no_akun, "Kenaikan/(Penurunan) Liabilitas Lancar Lainnya"))
    r_kas_operasi = _tulis_subtotal("KAS BERSIH DARI AKTIVITAS OPERASI", baris_operasi)

    # ================= ARUS KAS DARI AKTIVITAS INVESTASI ==================
    r += 1
    _tulis_section("ARUS KAS DARI AKTIVITAS INVESTASI")
    if akun_aset_tetap:
        baris_investasi = [_tulis_baris_akun(no_akun, "Pembelian/(Penjualan) Aset Tetap")
                            for no_akun in akun_aset_tetap]
    else:
        baris_investasi = [_tulis_baris_kosong("(Tidak ada mutasi aset tetap)")]
    r_kas_investasi = _tulis_subtotal("KAS BERSIH DARI AKTIVITAS INVESTASI", baris_investasi)

    # ================= ARUS KAS DARI AKTIVITAS PENDANAAN ===================
    r += 1
    _tulis_section("ARUS KAS DARI AKTIVITAS PENDANAAN")
    if akun_modal:
        baris_pendanaan = [_tulis_baris_akun(no_akun, "Setoran/(Penarikan) Modal")
                            for no_akun in akun_modal]
    else:
        baris_pendanaan = [_tulis_baris_kosong("(Tidak ada mutasi modal)")]
    r_kas_pendanaan = _tulis_subtotal("KAS BERSIH DARI AKTIVITAS PENDANAAN", baris_pendanaan)

    # ================= KENAIKAN (PENURUNAN) KAS BERSIH =====================
    r += 1
    r_kenaikan = r
    ws.cell(row=r, column=2, value="KENAIKAN (PENURUNAN) KAS BERSIH")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = _TOTAL_FONT_PE
        cell.fill = _TOTAL_FILL_PE
        cell.alignment = _ITEM_ALIGN_PE
        cell.border = _TOTAL_BORDER_PE
    for m in range(1, 13):
        col = 2 + m
        surat = get_column_letter(col)
        formula = f"={surat}{r_kas_operasi}+{surat}{r_kas_investasi}+{surat}{r_kas_pendanaan}"
        cell = ws.cell(row=r, column=col, value=formula)
        cell.number_format = _FORMAT_RUPIAH_PE
    ws.row_dimensions[r].height = 15.75

    # ================= KAS AWAL & AKHIR PERIODE ============================
    r += 1
    r_kas_awal = r
    r_kas_akhir = r + 1
    ws.cell(row=r, column=2, value="Kas dan Setara Kas Awal Periode").font = _ITEM_FONT_LOKAL_PE
    ws.cell(row=r, column=2).alignment = _ITEM_ALIGN_PE
    ws.cell(row=r, column=1).border = _DATA_BORDER_PE
    ws.cell(row=r, column=2).border = _DATA_BORDER_PE
    for m in range(1, 13):
        col = 2 + m
        surat = get_column_letter(col)
        prev_surat = get_column_letter(col - 1)
        if m == 1:
            f_awal = ("=" + "+".join(_saldo_awal_tahun_formula(no, "debit") for no in akun_kas_bank)
                       if akun_kas_bank else "=0")
        else:
            f_awal = f"={prev_surat}{r_kas_akhir}"
        cell = ws.cell(row=r, column=col, value=f_awal)
        cell.font = _ITEM_FONT_ANGKA_PE
        cell.number_format = _FORMAT_RUPIAH_PE
        cell.alignment = _ITEM_ALIGN_PE
        cell.border = _DATA_BORDER_PE

    r += 1
    assert r == r_kas_akhir
    ws.cell(row=r, column=2, value="Kas dan Setara Kas Akhir Periode")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = _TOTAL_FONT_PE
        cell.fill = _TOTAL_FILL_PE
        cell.alignment = _ITEM_ALIGN_PE
        cell.border = _TOTAL_BORDER_PE
    for m in range(1, 13):
        col = 2 + m
        surat = get_column_letter(col)
        formula = f"={surat}{r_kas_awal}+{surat}{r_kenaikan}"
        cell = ws.cell(row=r, column=col, value=formula)
        cell.number_format = _FORMAT_RUPIAH_PE
    ws.row_dimensions[r].height = 15.75

    # ================= VERIFIKASI (saldo kas aktual vs hasil arus kas) =====
    r += 2
    r_verif = r
    ws.cell(row=r, column=2, value="Saldo Kas per Buku Besar (Verifikasi)").font = _ITEM_FONT_LOKAL_PE
    ws.cell(row=r, column=2).alignment = _ITEM_ALIGN_PE
    ws.cell(row=r, column=1).border = _DATA_BORDER_PE
    ws.cell(row=r, column=2).border = _DATA_BORDER_PE
    for m in range(1, 13):
        col = 2 + m
        surat = get_column_letter(col)
        f = ("=" + "+".join(_saldo_kumulatif_formula(no, surat, "debit") for no in akun_kas_bank)
             if akun_kas_bank else "=0")
        cell = ws.cell(row=r, column=col, value=f)
        cell.font = _ITEM_FONT_ANGKA_PE
        cell.number_format = _FORMAT_RUPIAH_PE
        cell.alignment = _ITEM_ALIGN_PE
        cell.border = _DATA_BORDER_PE

    r += 1
    r_selisih = r
    ws.cell(row=r, column=2, value="Selisih (Check)")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = _TOTAL_FONT_PE
        cell.fill = _TOTAL_FILL_PE
        cell.alignment = _ITEM_ALIGN_PE
        cell.border = _TOTAL_BORDER_PE
    for m in range(1, 13):
        col = 2 + m
        surat = get_column_letter(col)
        formula = f"={surat}{r_kas_akhir}-{surat}{r_verif}"
        cell = ws.cell(row=r, column=col, value=formula)
        cell.number_format = _FORMAT_RUPIAH_PE
    ws.row_dimensions[r].height = 15.75

    r += 1
    r_status = r
    ws.cell(row=r, column=2, value="STATUS").font = _ITEM_FONT_LOKAL_PE
    ws.cell(row=r, column=2).alignment = _ITEM_ALIGN_PE
    for m in range(1, 13):
        col = 2 + m
        surat = get_column_letter(col)
        formula = f'=IF(ROUND({surat}{r_selisih},0)=0,"BALANCE","PERIKSA")'
        cell = ws.cell(row=r, column=col, value=formula)
        cell.font = _TOTAL_FONT_PE
        cell.fill = _STATUS_FILL_PE
        cell.alignment = _ITEM_ALIGN_PE
    ws.row_dimensions[r].height = 15.75

    # -- Highlight otomatis hijau/merah di baris STATUS, pola sama dgn
    #    sheet "Laporan Perubahan Ekuitas" --
    rentang_status = f"C{r_status}:N{r_status}"
    ws.conditional_formatting.add(
        rentang_status,
        CellIsRule(operator="equal", formula=['"BALANCE"'],
                   font=Font(name=_FONT_NAME_PE, bold=True, size=11, color="FF008000")),
    )
    ws.conditional_formatting.add(
        rentang_status,
        CellIsRule(operator="equal", formula=['"PERIKSA"'],
                   font=Font(name=_FONT_NAME_PE, bold=True, size=11, color="FFFF0000")),
    )


def _tulis_sheet_calk(ws, calk: Dict[str, Any], tahun: Any) -> None:
    """Sheet 13 "Catatan atas Laporan Keuangan" -- 4 kolom (No./
    Catatan/Rincian-Penjelasan/Jumlah (Rp)).

    [REVISI] Sebelumnya SEMUA akun (Aset/Liabilitas/Ekuitas/Pendapatan/
    Beban) ditempel rata jadi satu blok di bawah Catatan #3 tanpa
    pemisahan Neraca vs Laba Rugi dan tanpa rincian sub-kategori -- tidak
    lazim dibanding format CALK yang umum dipakai (pos Neraca & pos Laba
    Rugi punya catatan terpisah, dan tiap pos besar seperti "Kas dan
    Setara Kas"/"Piutang Usaha"/"Aset Tetap" biasanya dirinci sendiri-
    sendiri, bukan cuma dilabeli kategori besarnya). Sekarang mengikuti
    struktur baru lapkeu.susun_calk_otomatis(): kerangka 7 catatan wajib,
    Catatan #3 ("Rincian akun-akun Neraca") menampilkan "rincian_neraca"
    (ASET -> LIABILITAS -> EKUITAS, tiap kategori dipecah lagi per
    sub_kategori dgn subtotal sub-kategori + TOTAL kategori), Catatan #4
    ("Rincian akun-akun Laba Rugi") menampilkan "rincian_laba_rugi"
    (PENDAPATAN -> BEBAN) dgn pola sama. Akun yang kategorinya tidak
    dikenal di COA (kalau ada) ditulis di blok terpisah di akhir supaya
    tetap kelihatan, bukan hilang diam-diam.

    Baris TOTAL per kategori (mis. "TOTAL ASET") sengaja dihitung dari
    subtotal per sub-kategori yang sudah dikoreksi tanda kontra-akun (lihat
    _nilai_baris_akun() di bawah) supaya SELALU tie-out dgn Total Aset/
    Liabilitas/Ekuitas di sheet Neraca dan Total Pendapatan/Beban di sheet
    Laba Rugi Bulanan -- 1 sumber logika tanda (sinkron dgn
    _nilai_penambah_aset() di laporan_keuangan.py susun_neraca())."""
    headers = ["No.", "Catatan", "Rincian / Penjelasan", "Jumlah (Rp)"]
    header_row = _tulis_judul_header_baru4(ws, f"CATATAN ATAS LAPORAN KEUANGAN TAHUN {tahun}", headers)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 20

    if not calk:
        ws.cell(row=header_row + 2, column=1,
                 value="Belum ada data CALK -- generate laporan keuangan terlebih dahulu.").font = \
            _DATA_FONT_BARU4
        return

    kerangka = calk.get("kerangka_catatan") or []
    rincian_neraca = calk.get("rincian_neraca") or {}
    rincian_laba_rugi = calk.get("rincian_laba_rugi") or {}
    akun_tidak_dikenal = calk.get("akun_tidak_dikenal_kategori") or []
    # [BARU -- integrasi Claude API] Narasi utk catatan naratif (bukan
    # tabel rincian akun, mis. "Dasar Penyusunan", "Kebijakan Akuntansi
    # Signifikan") -- digenerate SEBELUM data sampai di sini lewat
    # claude_client.generate_narasi_calk_claude(), dikirim sebagai
    # calk["narasi_catatan"] = {"1": "...", "2": "...", ...}.
    # SENGAJA hanya dibaca (bukan dipanggil) di sini -- modul ini tetap
    # murni penulis Excel, tidak melakukan panggilan network sendiri.
    # Kalau field ini kosong (belum digenerate/API gagal), kolom
    # "Rincian/Penjelasan" utk catatan naratif tetap kosong seperti
    # sebelumnya -- tidak menggagalkan export.
    narasi_catatan = calk.get("narasi_catatan") or {}

    def _nilai_baris_akun(kategori: str, akun: Dict[str, Any]) -> tuple:
        """Return (nilai_bertanda, label, is_kontra) -- akun KONTRA-ASET
        (kategori ASET, normal_saldo KREDIT, mis. Akumulasi Penyusutan)
        dibalik tandanya supaya mengurangi (bukan menambah) subtotal,
        PERSIS pola _nilai_penambah_aset() di susun_neraca(). Kategori
        lain (LIABILITAS/EKUITAS/PENDAPATAN/BEBAN) TIDAK dibalik --
        susun_neraca()/susun_laba_rugi() sendiri juga tidak melakukan itu
        utk kategori-kategori tsb, jadi disamakan supaya tidak menambah
        logika baru yang tidak ada padanannya di sheet Neraca/Laba Rugi.
        """
        nilai = _angka(akun.get("saldo_akhir"))
        is_kontra = (kategori == "ASET" and akun.get("normal_saldo") == "KREDIT")
        if is_kontra:
            nilai = -nilai
        label = f"{akun.get('no_akun')} -- {akun.get('nama_akun')}"
        if is_kontra:
            label += " (kontra-aset)"
        return nilai, label, is_kontra

    def _tulis_blok_kategori(r: int, judul_blok: Dict[str, Dict[str, List[Dict[str, Any]]]]) -> int:
        """Tulis satu blok (rincian_neraca ATAU rincian_laba_rugi): tiap
        kategori -> tiap sub_kategori (baris item + subtotal sub-kategori)
        -> baris TOTAL kategori. Return baris terakhir yang ditulis."""
        for kategori, per_sub in judul_blok.items():
            r += 1
            ws.cell(row=r, column=2, value=f"Rincian Akun -- {kategori}")
            for c in range(1, 5):
                cell = ws.cell(row=r, column=c)
                cell.font = _SUBTOTAL_FONT_BARU4
                cell.fill = _SUBTOTAL_FILL_BARU4
            total_kategori = 0.0
            for sub_kategori, akun_list in per_sub.items():
                subtotal_sub = 0.0
                for akun in akun_list:
                    r += 1
                    nilai, label, _ = _nilai_baris_akun(kategori, akun)
                    ws.cell(row=r, column=3, value=label).font = _DATA_FONT_BARU4
                    cell_j = ws.cell(row=r, column=4, value=round(nilai, 2))
                    cell_j.font = _DATA_FONT_ANGKA_BARU4
                    cell_j.number_format = _FORMAT_RUPIAH_BARU4
                    ws.cell(row=r, column=3).border = _DATA_BORDER_BARU4
                    ws.cell(row=r, column=4).border = _DATA_BORDER_BARU4
                    subtotal_sub += nilai
                r += 1
                ws.cell(row=r, column=3, value=f"Subtotal -- {sub_kategori}")
                cell_ss = ws.cell(row=r, column=4, value=round(subtotal_sub, 2))
                cell_ss.number_format = _FORMAT_RUPIAH_BARU4
                for c in (3, 4):
                    cell = ws.cell(row=r, column=c)
                    cell.font = Font(name="Carlito", italic=True, size=11, color="FF000000")
                    cell.border = _DATA_BORDER_BARU4
                total_kategori += subtotal_sub
            r += 1
            ws.cell(row=r, column=3, value=f"TOTAL {kategori}")
            cell_t = ws.cell(row=r, column=4, value=round(total_kategori, 2))
            cell_t.number_format = _FORMAT_RUPIAH_BARU4
            for c in (3, 4):
                cell = ws.cell(row=r, column=c)
                cell.font = _SUBTOTAL_FONT_BARU4
                cell.fill = _SUBTOTAL_FILL_BARU4
                cell.border = _TOTAL_BORDER_BARU4
        return r

    r = header_row
    for item in kerangka:
        r += 1
        if ". " in item:
            no_str, sisa = item.split(". ", 1)
        else:
            no_str, sisa = "", item
        ws.cell(row=r, column=1, value=no_str)
        ws.cell(row=r, column=2, value=sisa)
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.font = _DATA_FONT_BARU4
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = _DATA_BORDER_BARU4

        # [BARU -- integrasi Claude API] Catatan naratif (semua nomor
        # SELAIN #3/#4, yang diisi tabel rincian akun otomatis di bawah)
        # -- isi kolom C ("Rincian/Penjelasan") dengan narasi dari
        # narasi_catatan[no_str] kalau ada. Baris tetap dinaikkan tinggi
        # otomatis oleh Excel karena wrap_text=True sudah di-set di atas.
        if no_str.strip() not in ("3", "4"):
            narasi = narasi_catatan.get(no_str.strip())
            if narasi:
                cell_narasi = ws.cell(row=r, column=3, value=narasi)
                cell_narasi.font = _DATA_FONT_BARU4
                cell_narasi.alignment = Alignment(wrap_text=True, vertical="top")
                cell_narasi.border = _DATA_BORDER_BARU4

        # Catatan #3 -- rincian akun-akun Neraca (Aset/Liabilitas/Ekuitas)
        if no_str.strip() == "3" and rincian_neraca:
            r = _tulis_blok_kategori(r, rincian_neraca)
            if akun_tidak_dikenal:
                r += 1
                ws.cell(row=r, column=2,
                         value="Akun terpakai di jurnal tapi belum terdaftar kategorinya di COA -- perlu dilengkapi")
                for c in range(1, 5):
                    cell = ws.cell(row=r, column=c)
                    cell.font = _SUBTOTAL_FONT_BARU4
                    cell.fill = _SUBTOTAL_FILL_BARU4
                for akun in akun_tidak_dikenal:
                    r += 1
                    label = f"{akun.get('no_akun')} -- {akun.get('nama_akun')}"
                    ws.cell(row=r, column=3, value=label).font = _DATA_FONT_BARU4
                    cell_j = ws.cell(row=r, column=4, value=round(_angka(akun.get("saldo_akhir")), 2))
                    cell_j.font = _DATA_FONT_ANGKA_BARU4
                    cell_j.number_format = _FORMAT_RUPIAH_BARU4
                    ws.cell(row=r, column=3).border = _DATA_BORDER_BARU4
                    ws.cell(row=r, column=4).border = _DATA_BORDER_BARU4

        # Catatan #4 -- rincian akun-akun Laba Rugi (Pendapatan/Beban)
        if no_str.strip() == "4" and rincian_laba_rugi:
            r = _tulis_blok_kategori(r, rincian_laba_rugi)

    catatan = calk.get("catatan") or ""
    if catatan:
        ws.cell(row=r + 2, column=1, value=catatan).font = _CATATAN_FONT_BARU4


# ================= Style + penulis sheet 17 "Rekonsiliasi Fiskal" [ROMBAK
# TOTAL] =================
# [FIX -- match 100% referensi REKONSILIASI_FISKAL_TAHUN_2025___LAMPIRAN_
# SPT_TAHUNAN.xlsx, sel demi sel] SEBELUMNYA sheet ini cuma "kloning" 8
# kolom dari `pnl_baku` (data per-akun dinamis yang sama dgn sheet 16) --
# BEDA TOTAL dari file referensi asli yang user kirim, yang ternyata:
#   1. 9 kolom (bukan 8) -- "Keterangan / Bukti" dan "Dasar Umum" adalah
#      DUA kolom terpisah (H dan I), bukan digabung.
#   2. Bagian atas (REV/DIR/GP/OPEX/EBITDA/DEP/OI/OE/NP) BUKAN daftar akun
#      dinamis per client -- ini ringkasan laba-rugi 9-baris BAKU dgn kode
#      tetap, kolom Komersial (C) berupa FORMULA LINTAS-SHEET hidup yang
#      nempel ke subtotal terkait di sheet 16 "PNL Lampiran SPT" (supaya
#      tie-out otomatis, edit manual di satu sisi ikut kereka ulang).
#   3. "RINCIAN KOREKSI FISKAL POSITIF/NEGATIF" adalah CHECKLIST BAKU 10+5
#      pos standar UU PPh (KFP-01..10 / KFN-01..05, sudah termasuk rujukan
#      pasal) -- BUKAN turunan dinamis dari akun COA. Kolom Koreksi
#      Positif/Negatif per pos adalah SEL INPUT MANUAL akuntan (font biru,
#      fill kuning, default 0) -- kecuali baris "lainnya" (KFP-10/KFN-05)
#      yang DIPRA-ISI dgn total agregat dari pph_badan.hitung_pph_pasal_31e()
#      (kalau sudah pernah digenerate) supaya angka yang sudah dihitung
#      sistem tidak hilang -- akuntan tinggal PINDAHKAN ke pos yang lebih
#      spesifik kalau perlu, bukan mulai dari nol.
#   4. TOTAL KOREKSI FISKAL POSITIF/NEGATIF = SUM() 10/5 baris di atasnya
#      (bukan lagi ambil angka agregat pph_badan langsung) -- PENGHASILAN
#      NETO FISKAL -> PKP dihitung FORMULA berantai dari situ, persis file
#      referensi (F35=C35+D35-E35, F37=MAX(0,F35-F36),
#      F38=INT(F37/1000)*1000).
# Konstanta lama (_REKON_LEBAR_KOLOM/_ALIGN_REKON_WRAP/_border_rekon_*)
# SENGAJA TIDAK dihapus di atas fungsi lama -- lihat catatan umum modul
# ini soal kompatibilitas mundur -- tapi tidak lagi dipakai fungsi baru
# di bawah.
_REKON2_TITLE_FONT = Font(name="Carlito", bold=True, size=15, color="FFFFFFFF")
_REKON2_TITLE_FILL = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_REKON2_TITLE_ALIGN = Alignment(horizontal="left", vertical="center")
_REKON2_HEADER_FONT = Font(name="Carlito", bold=True, size=11, color="FFFFFFFF")
_REKON2_HEADER_FILL = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
_REKON2_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_REKON2_LABEL_FONT = Font(name="Calibri", size=11)             # kolom A/B/G/H/I baris item (teks)
_REKON2_FORMULA_LINTAS_FONT = Font(name="Carlito", size=11, color="FF008000")   # hijau -- link ke sheet 16
_REKON2_FORMULA_LOKAL_FONT = Font(name="Carlito", size=11, color="FF000000")    # hitam -- formula sel sendiri
_REKON2_INPUT_FONT = Font(name="Carlito", size=11, color="FF0000FF")            # biru -- input manual
_REKON2_INPUT_FILL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
# [FIX] kolom C (Komersial) di baris subtotal GP/EBITDA/NP -- hijau bold,
# BUKAN hitam seperti kolom subtotal lainnya (lihat catatan di pemanggil).
_REKON2_SUBTOTAL_C_FONT = Font(name="Carlito", bold=True, size=11, color="FF008000")
_REKON2_PLACEHOLDER_FILL = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
_REKON2_FORMAT_RUPIAH = r"#,##0;[Red]\(#,##0\);\-"
_REKON2_LEBAR_KOLOM = {
    "A": 39, "B": 65.86, "C": 23, "D": 22.71, "E": 23.57,
    "F": 12, "G": 11.71, "H": 56.71, "I": 30.29,
}
_GARIS_REKON2 = Side(style="thin")
_GARIS_REKON2_DBL = Side(style="double")


def _border_rekon2(kolom: int, top: bool = False, bottom: bool = False, dbl_bottom: bool = False) -> Border:
    """Border generik 9-kolom sheet 17: dinding kiri hanya kolom A (1),
    dinding kanan hanya kolom I (9) -- top/bottom sesuai parameter,
    persis pola per-baris yang diaudit dari file referensi (lihat
    komentar tiap pemanggil di bawah)."""
    return Border(
        top=_GARIS_REKON2 if top else None,
        bottom=(_GARIS_REKON2_DBL if dbl_bottom else (_GARIS_REKON2 if bottom else None)),
        left=_GARIS_REKON2 if kolom == 1 else None,
        right=_GARIS_REKON2 if kolom == 9 else None,
    )


# Checklist baku 10 pos Koreksi Fiskal Positif -- kode, uraian, keterangan/
# bukti, dasar hukum -- SESUAI teks file referensi kata per kata.
_REKON2_KFP_BAKU: List[Dict[str, str]] = [
    {"kode": "KFP-01", "uraian": "Biaya untuk kepentingan pribadi pemegang saham/sekutu",
     "keterangan": "Isi nilai dan simpan bukti identifikasi biaya", "dasar": "Pasal 9 ayat (1) UU PPh"},
    {"kode": "KFP-02", "uraian": "Pajak Penghasilan dan sanksi administrasi perpajakan",
     "keterangan": "PPh Badan, bunga, denda, dan kenaikan pajak", "dasar": "Pasal 9 ayat (1) UU PPh"},
    {"kode": "KFP-03", "uraian": "Sumbangan/bantuan/hibah yang tidak memenuhi ketentuan",
     "keterangan": "Kecuali yang diperkenankan peraturan", "dasar": "Pasal 9 ayat (1) UU PPh"},
    {"kode": "KFP-04", "uraian": "Cadangan atau provisi yang tidak diperkenankan",
     "keterangan": "Sesuaikan dengan jenis cadangan yang diizinkan", "dasar": "Pasal 9 ayat (1) UU PPh"},
    {"kode": "KFP-05", "uraian": "Biaya tanpa bukti pendukung/nominatif yang memadai",
     "keterangan": "Contoh: entertainment tanpa daftar nominatif", "dasar": "Pasal 6 dan ketentuan pelaksana"},
    {"kode": "KFP-06", "uraian": "Biaya terkait penghasilan yang dikenai PPh final/bukan objek",
     "keterangan": "Tidak boleh mengurangi penghasilan tarif umum", "dasar": "Prinsip matching fiskal"},
    {"kode": "KFP-07", "uraian": "Selisih penyusutan komersial lebih besar dari penyusutan fiskal",
     "keterangan": "Berdasarkan daftar aktiva dan kelompok fiskal", "dasar": "Pasal 11 UU PPh"},
    {"kode": "KFP-08", "uraian": "Natura/kenikmatan yang tidak memenuhi ketentuan pengurang",
     "keterangan": "Verifikasi ketentuan natura tahun pajak terkait", "dasar": "Pasal 6/Pasal 9 UU PPh"},
    {"kode": "KFP-09", "uraian": "Imbalan pihak berelasi yang melebihi kewajaran",
     "keterangan": "Siapkan pembanding kewajaran transaksi", "dasar": "Pasal 9 dan Pasal 18 UU PPh"},
    {"kode": "KFP-10", "uraian": "Koreksi fiskal positif lainnya",
     "keterangan": "Jelaskan dasar dan dokumen pendukung", "dasar": "Sesuai ketentuan yang relevan"},
]
# Checklist baku 5 pos Koreksi Fiskal Negatif.
_REKON2_KFN_BAKU: List[Dict[str, str]] = [
    {"kode": "KFN-01", "uraian": "Penghasilan yang telah dikenai PPh final",
     "keterangan": "Pisahkan dari penghasilan tarif umum", "dasar": "Pasal 4 ayat (2) UU PPh"},
    {"kode": "KFN-02", "uraian": "Penghasilan yang bukan merupakan objek pajak",
     "keterangan": "Lampirkan dasar pengecualian objek pajak", "dasar": "Pasal 4 ayat (3) UU PPh"},
    {"kode": "KFN-03", "uraian": "Selisih penyusutan fiskal lebih besar dari penyusutan komersial",
     "keterangan": "Berdasarkan daftar aktiva dan kelompok fiskal", "dasar": "Pasal 11 UU PPh"},
    {"kode": "KFN-04", "uraian": "Pembalikan koreksi fiskal positif tahun sebelumnya",
     "keterangan": "Cantumkan rekonsiliasi tahun sebelumnya", "dasar": "Sesuai sifat temporer koreksi"},
    {"kode": "KFN-05", "uraian": "Koreksi fiskal negatif lainnya",
     "keterangan": "Jelaskan dasar dan dokumen pendukung", "dasar": "Sesuai ketentuan yang relevan"},
]
# Baris ringkasan laba-rugi baku bagian atas -- (kode, uraian, sifat,
# keterangan, dasar_umum, arah) -- "arah" menentukan formula F=C+D-E
# (pendapatan) vs F=C-D+E (beban); None utk baris subtotal (formula F
# khusus, lihat _tulis_sheet_rekonsiliasi_fiskal()).
_REKON2_KET_STANDAR = "Nilai komersial; koreksi diisi pada rincian di bawah"
_REKON2_RINGKASAN_BAKU = [
    {"kode": "REV", "uraian": "Total Pendapatan Usaha", "sifat": "Penghasilan",
     "dasar": "Pasal 4 & Pasal 6 UU PPh", "arah": "pendapatan", "subtotal": False},
    {"kode": "DIR", "uraian": "Total Beban Langsung", "sifat": "Beban",
     "dasar": "Pasal 6 UU PPh", "arah": "beban", "subtotal": False},
    {"kode": "GP", "uraian": "Laba Kotor", "sifat": "Subtotal", "dasar": None, "subtotal": True},
    {"kode": "OPEX", "uraian": "Total Beban Operasional", "sifat": "Beban",
     "dasar": "Pasal 6 & Pasal 9 UU PPh", "arah": "beban", "subtotal": False},
    {"kode": "EBITDA", "uraian": "EBITDA", "sifat": "Subtotal", "dasar": None, "subtotal": True},
    {"kode": "DEP", "uraian": "Total Beban Penyusutan", "sifat": "Beban",
     "dasar": "Pasal 11 UU PPh", "arah": "beban", "subtotal": False},
    {"kode": "OI", "uraian": "Pendapatan Lain-lain", "sifat": "Penghasilan",
     "dasar": "Pasal 4 UU PPh", "arah": "pendapatan", "subtotal": False},
    {"kode": "OE", "uraian": "Beban Lain-lain", "sifat": "Beban",
     "dasar": "Pasal 6 & Pasal 9 UU PPh", "arah": "beban", "subtotal": False},
    {"kode": "NP", "uraian": "Laba Bersih Komersial", "sifat": "Subtotal", "dasar": None, "subtotal": True},
]


def _tulis_sheet_rekonsiliasi_fiskal(
    ws, pnl_baku: Dict[str, Any], tahun: Any,
    info_sheet_pnl: Optional[Dict[str, Any]] = None,
) -> None:
    """Sheet 17 "Rekonsiliasi Fiskal" -- ROMBAK TOTAL supaya PERSIS sama
    dgn file referensi REKONSILIASI_FISKAL_TAHUN_2025___LAMPIRAN_SPT_
    TAHUNAN.xlsx sel demi sel (lihat komentar blok di atas fungsi ini).

    Args:
        pnl_baku: dipakai HANYA untuk 2 hal sekarang (beda dari versi
            lama yang meng-kloning seluruh baris akunnya): (1) menarik
            angka "KOMPENSASI KERUGIAN FISKAL" & total agregat koreksi
            fiskal positif/negatif dari pph_badan (sbg pra-isi baris
            "lainnya"), (2) fallback pesan kalau laporan belum digenerate.
        info_sheet_pnl: dict return value `_tulis_sheet_pnl_lampiran_spt_
            baku()` -- {"sheet_nama", "subtotal_row", "pl_range",
            "bl_range"} -- WAJIB ada supaya blok ringkasan atas bisa
            dibuat formula lintas-sheet hidup. Kalau None/kosong (mis.
            sheet 16 gagal ditulis), blok atas fallback jadi angka 0
            statis (bukan formula) supaya sheet tetap terbentuk.
    """
    ncols = 9
    ws.cell(row=1, column=1, value=f"REKONSILIASI FISKAL TAHUN {tahun} \u2013 LAMPIRAN SPT TAHUNAN BADAN")
    cell_judul = ws.cell(row=1, column=1)
    cell_judul.font = _REKON2_TITLE_FONT
    cell_judul.fill = _REKON2_TITLE_FILL
    cell_judul.alignment = _REKON2_TITLE_ALIGN
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.row_dimensions[1].height = 19.5

    headers = ["Kode", "Uraian", f"Komersial {tahun}", "Koreksi Fiskal Positif",
               "Koreksi Fiskal Negatif", f"Fiskal {tahun}", "Sifat",
               "Keterangan / Bukti", "Dasar Umum"]
    header_row = 3
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = _REKON2_HEADER_FONT
        cell.fill = _REKON2_HEADER_FILL
        cell.alignment = _REKON2_HEADER_ALIGN
        cell.border = _border_rekon2(col, top=True, bottom=True)
    for kolom, lebar in _REKON2_LEBAR_KOLOM.items():
        ws.column_dimensions[kolom].width = lebar

    if not pnl_baku or not pnl_baku.get("baris"):
        ws.cell(row=header_row + 2, column=1,
                 value="Belum ada data Rekonsiliasi Fiskal -- generate laporan keuangan terlebih dahulu.").font = \
            _REKON2_LABEL_FONT
        return

    sheet_pnl_nama = (info_sheet_pnl or {}).get("sheet_nama")
    subtotal_row = (info_sheet_pnl or {}).get("subtotal_row") or {}
    pl_start, pl_end = (info_sheet_pnl or {}).get("pl_range", (None, None))
    bl_start, bl_end = (info_sheet_pnl or {}).get("bl_range", (None, None))

    def _ref_sheet16(rentang: Optional[str] = None, baris: Optional[int] = None) -> Optional[str]:
        """Bangun referensi lintas-sheet ke 'PNL Lampiran SPT' -- None
        kalau info sheet 16 tidak tersedia (fallback nanti jadi 0)."""
        if not sheet_pnl_nama:
            return None
        if baris is not None:
            return f"'{sheet_pnl_nama}'!C{baris}"
        return f"SUM('{sheet_pnl_nama}'!{rentang})"

    # ---------------- Blok ringkasan REV..NP (baris 4-12) ----------------
    ref_map = {
        "REV": _ref_sheet16(baris=subtotal_row.get("JUMLAH PENDAPATAN USAHA")),
        "DIR": _ref_sheet16(baris=subtotal_row.get("JUMLAH BEBAN LANGSUNG")),
        "OPEX": _ref_sheet16(baris=subtotal_row.get("JUMLAH BEBAN OPERASIONAL")),
        "DEP": _ref_sheet16(baris=subtotal_row.get("JUMLAH BEBAN PENYUSUTAN")),
        "OI": _ref_sheet16(rentang=f"C{pl_start}:C{pl_end}") if pl_start else None,
        "OE": _ref_sheet16(rentang=f"C{bl_start}:C{bl_end}") if bl_start else None,
        "NP": _ref_sheet16(baris=subtotal_row.get("LABA BERSIH KOMERSIAL")),
    }
    r = 3
    row_of = {}  # kode -> nomor baris, dipakai formula subtotal (GP/EBITDA/NP)
    for item in _REKON2_RINGKASAN_BAKU:
        r += 1
        row_of[item["kode"]] = r
        kode, uraian, sifat = item["kode"], item["uraian"], item["sifat"]
        if not item["subtotal"]:
            formula_c = ref_map.get(kode) or 0
            ws.cell(row=r, column=1, value=kode)
            ws.cell(row=r, column=2, value=uraian)
            cell_c = ws.cell(row=r, column=3, value=f"={formula_c}" if formula_c else 0)
            # [FIX] Kolom C baris REV/DIR/OPEX/DEP/OI/OE SELALU hijau
            # (gaya "formula lintas-sheet") di file referensi -- C4/C5/C7/
            # C9/C10/C11 tetap hijau sekalipun sheet 16 belum tersedia dan
            # nilainya fallback ke 0 (dulu jadi hitam kalau formula_c
            # kosong -- salah, sudah dikoreksi).
            cell_c.font = _REKON2_FORMULA_LINTAS_FONT
            cell_c.number_format = _REKON2_FORMAT_RUPIAH
            for col_de in (4, 5):
                cell_de = ws.cell(row=r, column=col_de, value="=0")
                cell_de.font = _REKON2_FORMULA_LOKAL_FONT
                cell_de.fill = _REKON2_PLACEHOLDER_FILL
                cell_de.number_format = _REKON2_FORMAT_RUPIAH
            formula_f = f"=C{r}+D{r}-E{r}" if item["arah"] == "pendapatan" else f"=C{r}-D{r}+E{r}"
            cell_f = ws.cell(row=r, column=6, value=formula_f)
            cell_f.font = _REKON2_FORMULA_LOKAL_FONT
            cell_f.number_format = _REKON2_FORMAT_RUPIAH
            ws.cell(row=r, column=7, value=sifat)
            ws.cell(row=r, column=8, value=_REKON2_KET_STANDAR)
            ws.cell(row=r, column=9, value=item["dasar"])
            for col_label in (1, 2, 7, 8, 9):
                ws.cell(row=r, column=col_label).font = _REKON2_LABEL_FONT
            for col in range(1, 10):
                ws.cell(row=r, column=col).border = _border_rekon2(col, bottom=True)
        else:
            if kode == "GP":
                r_rev, r_dir = row_of["REV"], row_of["DIR"]
                formula_c, formula_f = f"=C{r_rev}-C{r_dir}", f"=F{r_rev}-F{r_dir}"
            elif kode == "EBITDA":
                r_gp, r_opex = row_of["GP"], row_of["OPEX"]
                formula_c, formula_f = f"=C{r_gp}-C{r_opex}", f"=F{r_gp}-F{r_opex}"
            else:  # NP
                r_ebitda, r_dep, r_oi, r_oe = row_of["EBITDA"], row_of["DEP"], row_of["OI"], row_of["OE"]
                formula_c = ref_map.get("NP") or 0
                formula_f = f"=F{r_ebitda}-F{r_dep}+F{r_oi}-F{r_oe}"
            ws.cell(row=r, column=1, value=kode)
            ws.cell(row=r, column=2, value=uraian)
            cell_c = ws.cell(row=r, column=3,
                              value=(f"={formula_c}" if kode == "NP" and formula_c else formula_c))
            cell_c.font = _REKON2_FORMULA_LINTAS_FONT if kode == "NP" and formula_c else _REKON2_FORMULA_LOKAL_FONT
            cell_c.number_format = _REKON2_FORMAT_RUPIAH
            cell_f = ws.cell(row=r, column=6, value=formula_f)
            cell_f.font = _REKON2_FORMULA_LOKAL_FONT
            cell_f.number_format = _REKON2_FORMAT_RUPIAH
            ws.cell(row=r, column=7, value=sifat)
            ws.cell(row=r, column=8, value=_REKON2_KET_STANDAR if kode != "NP" else "")
            for col_label in (1, 2, 7, 8):
                ws.cell(row=r, column=col_label).font = _REKON2_LABEL_FONT
            # [FIX] Font subtotal Carlito bold hitam utk semua kolom KECUALI
            # kolom C (3) -- diverifikasi ulang sel-demi-sel dari file
            # referensi: C6/C8/C12 (GP/EBITDA/NP) TETAP hijau bold
            # (FF008000, gaya "formula lintas-sheet") sekalipun formulanya
            # lokal (=C4-C5 dst) -- hanya F6/F8/F12 dan kolom lain yang
            # hitam. Komentar lama di sini (yang bilang C juga hitam) SALAH
            # -- sudah dikoreksi.
            for col in range(1, 10):
                cell = ws.cell(row=r, column=col)
                cell.font = _SUBTOTAL_FONT_BARU4
                cell.fill = _SUBTOTAL_FILL_BARU4
                cell.border = _border_rekon2(col, top=True, dbl_bottom=True)
                if col in (3, 4, 5, 6):
                    cell.number_format = _REKON2_FORMAT_RUPIAH
            ws.cell(row=r, column=3).font = _REKON2_SUBTOTAL_C_FONT
            if kode in ("GP", "EBITDA"):
                # [FIX] D/E kosong di baris GP & EBITDA tetap memakai gaya
                # "input" (biru bold + fill kuning) di file referensi --
                # sel-sel ini tidak berisi nilai, tapi formatnya begitu di
                # sumber aslinya (beda dari D12/E12 di baris NP yang polos
                # hitam/D9E1F2 seperti subtotal lain) -- direplikasi persis
                # walau tidak lazim, sesuai permintaan "sama persis".
                for col_de in (4, 5):
                    cell_de = ws.cell(row=r, column=col_de)
                    cell_de.font = Font(name="Carlito", bold=True, size=11, color="FF0000FF")
                    cell_de.fill = _REKON2_INPUT_FILL
    row_np = row_of["NP"]

    # [FIX] Baris kosong pemisah (13) juga punya dinding kiri/kanan tipis
    # (kolom A & I) di file referensi -- bukan baris benar-benar polos.
    ws.cell(row=r + 1, column=1).border = _border_rekon2(1)
    ws.cell(row=r + 1, column=9).border = _border_rekon2(9)

    # ------------- RINCIAN KOREKSI FISKAL POSITIF (baris 14-25) ----------
    r += 2  # baris 13 kosong, judul di baris 14
    ws.cell(row=r, column=1, value="RINCIAN KOREKSI FISKAL POSITIF")
    for col in range(1, 10):
        cell = ws.cell(row=r, column=col)
        cell.font = _PNL_JUDUL_FONT
        cell.fill = _PNL_JUDUL_FILL
        cell.border = _border_rekon2(col, bottom=True)

    # Total agregat pph_badan (kalau ada) dipra-isi ke baris "lainnya"
    # (KFP-10) supaya angka yang sudah dihitung sistem tidak hilang --
    # lihat catatan blok komentar di atas fungsi ini.
    total_kfp_agregat = 0.0
    total_kfn_agregat = 0.0
    total_kompensasi = 0.0
    for b in pnl_baku["baris"]:
        if b.get("tipe") != "subtotal":
            continue
        if b["uraian"] == "TOTAL KOREKSI FISKAL POSITIF":
            total_kfp_agregat = _angka(b.get("koreksi_positif"))
        elif b["uraian"] == "TOTAL KOREKSI FISKAL NEGATIF":
            total_kfn_agregat = _angka(b.get("koreksi_negatif"))
        elif b["uraian"] == "KOMPENSASI KERUGIAN FISKAL":
            total_kompensasi = _angka(b.get("komersial"))

    kfp_row_awal = r + 1
    for i, pos in enumerate(_REKON2_KFP_BAKU):
        r += 1
        nilai_default = total_kfp_agregat if pos["kode"] == "KFP-10" else 0
        ws.cell(row=r, column=1, value=pos["kode"])
        ws.cell(row=r, column=2, value=pos["uraian"])
        cell_c = ws.cell(row=r, column=3, value="=0")
        cell_c.font = _REKON2_FORMULA_LOKAL_FONT
        cell_c.number_format = _REKON2_FORMAT_RUPIAH
        cell_d = ws.cell(row=r, column=4, value=nilai_default)
        cell_d.font = _REKON2_INPUT_FONT
        cell_d.fill = _REKON2_INPUT_FILL
        cell_d.number_format = _REKON2_FORMAT_RUPIAH
        cell_e = ws.cell(row=r, column=5, value="=0")
        cell_e.font = _REKON2_FORMULA_LOKAL_FONT
        cell_e.number_format = _REKON2_FORMAT_RUPIAH
        cell_f = ws.cell(row=r, column=6, value=f"=D{r}")
        cell_f.font = _REKON2_FORMULA_LOKAL_FONT
        cell_f.number_format = _REKON2_FORMAT_RUPIAH
        ws.cell(row=r, column=7, value="Positif")
        ws.cell(row=r, column=8, value=pos["keterangan"])
        ws.cell(row=r, column=9, value=pos["dasar"])
        for col_label in (1, 2, 7, 8, 9):
            ws.cell(row=r, column=col_label).font = _REKON2_LABEL_FONT
        ws.cell(row=r, column=1).border = _border_rekon2(1)
        ws.cell(row=r, column=9).border = _border_rekon2(9)
    kfp_row_akhir = r

    r += 1
    ws.cell(row=r, column=2, value="TOTAL KOREKSI FISKAL POSITIF")
    cell_d = ws.cell(row=r, column=4, value=f"=SUM(D{kfp_row_awal}:D{kfp_row_akhir})")
    cell_f = ws.cell(row=r, column=6, value=f"=D{r}")
    for col in range(1, 10):
        cell = ws.cell(row=r, column=col)
        cell.font = _SUBTOTAL_FONT_BARU4
        cell.fill = _SUBTOTAL_FILL_BARU4
        cell.border = _border_rekon2(col, top=True, dbl_bottom=True)
        if col in (3, 4, 5, 6):
            cell.number_format = _REKON2_FORMAT_RUPIAH
    row_total_kfp = r

    # [FIX] Baris kosong pemisah (26) -- dinding kiri/kanan tipis (A & I),
    # sama seperti baris 13/34, lihat catatan di atas.
    ws.cell(row=r + 1, column=1).border = _border_rekon2(1)
    ws.cell(row=r + 1, column=9).border = _border_rekon2(9)

    # ------------- RINCIAN KOREKSI FISKAL NEGATIF (baris 27-33) ----------
    r += 2  # baris kosong, judul
    ws.cell(row=r, column=1, value="RINCIAN KOREKSI FISKAL NEGATIF")
    for col in range(1, 10):
        cell = ws.cell(row=r, column=col)
        cell.font = _PNL_JUDUL_FONT
        cell.fill = _PNL_JUDUL_FILL
        cell.border = _border_rekon2(col, bottom=True)

    kfn_row_awal = r + 1
    for i, neg in enumerate(_REKON2_KFN_BAKU):
        r += 1
        nilai_default = total_kfn_agregat if neg["kode"] == "KFN-05" else 0
        ws.cell(row=r, column=1, value=neg["kode"])
        ws.cell(row=r, column=2, value=neg["uraian"])
        cell_c = ws.cell(row=r, column=3, value="=0")
        cell_c.font = _REKON2_FORMULA_LOKAL_FONT
        cell_c.number_format = _REKON2_FORMAT_RUPIAH
        cell_d = ws.cell(row=r, column=4, value="=0")
        cell_d.font = _REKON2_FORMULA_LOKAL_FONT
        cell_d.number_format = _REKON2_FORMAT_RUPIAH
        cell_e = ws.cell(row=r, column=5, value=nilai_default)
        cell_e.font = _REKON2_INPUT_FONT
        cell_e.fill = _REKON2_INPUT_FILL
        cell_e.number_format = _REKON2_FORMAT_RUPIAH
        cell_f = ws.cell(row=r, column=6, value=f"=E{r}")
        cell_f.font = _REKON2_FORMULA_LOKAL_FONT
        cell_f.number_format = _REKON2_FORMAT_RUPIAH
        ws.cell(row=r, column=7, value="Negatif")
        ws.cell(row=r, column=8, value=neg["keterangan"])
        ws.cell(row=r, column=9, value=neg["dasar"])
        for col_label in (1, 2, 7, 8, 9):
            ws.cell(row=r, column=col_label).font = _REKON2_LABEL_FONT
        ws.cell(row=r, column=1).border = _border_rekon2(1)
        ws.cell(row=r, column=9).border = _border_rekon2(9)
    kfn_row_akhir = r

    r += 1
    ws.cell(row=r, column=2, value="TOTAL KOREKSI FISKAL NEGATIF")
    cell_e = ws.cell(row=r, column=5, value=f"=SUM(E{kfn_row_awal}:E{kfn_row_akhir})")
    cell_f = ws.cell(row=r, column=6, value=f"=E{r}")
    for col in range(1, 10):
        cell = ws.cell(row=r, column=col)
        cell.font = _SUBTOTAL_FONT_BARU4
        cell.fill = _SUBTOTAL_FILL_BARU4
        cell.border = _border_rekon2(col, top=True, dbl_bottom=True)
        if col in (3, 4, 5, 6):
            cell.number_format = _REKON2_FORMAT_RUPIAH
    row_total_kfn = r

    # [FIX] Baris kosong pemisah (34) -- dinding kiri/kanan tipis (A & I),
    # sama seperti baris 13/26, lihat catatan di atas.
    ws.cell(row=r + 1, column=1).border = _border_rekon2(1)
    ws.cell(row=r + 1, column=9).border = _border_rekon2(9)

    # ------------- PENGHASILAN NETO FISKAL -> PKP (baris 35-38) ----------
    r += 2  # baris kosong
    ws.cell(row=r, column=2, value="PENGHASILAN NETO FISKAL")
    ws.cell(row=r, column=3, value=f"=C{row_np}")
    ws.cell(row=r, column=4, value=f"=D{row_total_kfp}")
    ws.cell(row=r, column=5, value=f"=E{row_total_kfn}")
    ws.cell(row=r, column=6, value=f"=C{r}+D{r}-E{r}")
    for col in range(1, 10):
        cell = ws.cell(row=r, column=col)
        cell.font = _SUBTOTAL_FONT_BARU4
        cell.fill = _SUBTOTAL_FILL_BARU4
        cell.border = _border_rekon2(col, top=True, dbl_bottom=True)
        if col in (3, 4, 5, 6):
            cell.number_format = _REKON2_FORMAT_RUPIAH
    row_neto_fiskal = r

    r += 1
    ws.cell(row=r, column=2, value="Kompensasi Kerugian Fiskal")
    ws.cell(row=r, column=2).font = _REKON2_LABEL_FONT
    cell_d = ws.cell(row=r, column=4, value=total_kompensasi)
    cell_d.font = _REKON2_INPUT_FONT
    cell_d.fill = _REKON2_INPUT_FILL
    cell_d.number_format = _REKON2_FORMAT_RUPIAH
    cell_e = ws.cell(row=r, column=5, value=0)
    cell_e.font = _REKON2_LABEL_FONT
    cell_e.number_format = _REKON2_FORMAT_RUPIAH
    cell_f = ws.cell(row=r, column=6, value=f"=D{r}")
    cell_f.font = _REKON2_LABEL_FONT
    cell_f.number_format = _REKON2_FORMAT_RUPIAH
    ws.cell(row=r, column=7, value="Input")
    ws.cell(row=r, column=7).font = _REKON2_LABEL_FONT
    ws.cell(row=r, column=8, value="Isi jika terdapat rugi fiskal yang masih dapat dikompensasikan")
    ws.cell(row=r, column=8).font = _REKON2_LABEL_FONT
    ws.cell(row=r, column=9, value="Pasal 6 ayat (2) UU PPh")
    ws.cell(row=r, column=9).font = _REKON2_LABEL_FONT
    ws.cell(row=r, column=1).border = _border_rekon2(1)
    ws.cell(row=r, column=9).border = _border_rekon2(9)
    row_kompensasi = r

    r += 1
    ws.cell(row=r, column=2, value="Penghasilan Kena Pajak Sebelum Pembulatan")
    ws.cell(row=r, column=6, value=f"=MAX(0,F{row_neto_fiskal}-F{row_kompensasi})")
    for col in range(1, 10):
        cell = ws.cell(row=r, column=col)
        cell.font = _SUBTOTAL_FONT_BARU4
        cell.fill = _SUBTOTAL_FILL_BARU4
        cell.border = _border_rekon2(col, top=True)
        if col in (3, 4, 5, 6):
            cell.number_format = _REKON2_FORMAT_RUPIAH
    row_pkp_sblm = r

    r += 1
    ws.cell(row=r, column=2, value="Penghasilan Kena Pajak \u2013 Dibulatkan ke Bawah Ribuan Rupiah")
    ws.cell(row=r, column=6, value=f"=INT(F{row_pkp_sblm}/1000)*1000")
    for col in range(1, 10):
        cell = ws.cell(row=r, column=col)
        cell.font = _SUBTOTAL_FONT_BARU4
        cell.fill = _SUBTOTAL_FILL_BARU4
        cell.border = _border_rekon2(col, bottom=True)
        if col in (3, 4, 5, 6):
            cell.number_format = _REKON2_FORMAT_RUPIAH

    catatan = ("Baris REV/DIR/OPEX/DEP/OI/OE/NP tertaut langsung (formula lintas-sheet) ke sheet "
               f"'{sheet_pnl_nama}' -- tie-out otomatis." if sheet_pnl_nama else
               "Sheet 'PNL Lampiran SPT' tidak tersedia -- baris ringkasan atas memakai 0, isi manual.")
    if pnl_baku.get("catatan", "").find("FALLBACK") != -1:
        catatan += " Total KFP-10/KFN-05 belum terisi otomatis -- generate PPh Badan 31E dulu."
    ws.cell(row=r + 2, column=1, value=catatan).font = _CATATAN_FONT_BARU4


def _gaya_gridline_ringkasan(chart) -> None:
    """[BARU] Gridline putus-putus abu #CCCCCC + border area chart abu
    #D9D9D9 di sumbu X/Y -- sesuai file model referensi
    RINGKASAN_KINERJA_KEUANGAN_2025.xlsx (dicek lewat chart1.xml/chart2.xml
    di dalam file .xlsx tsb, bukan cuma tampilan)."""
    garis_grid = GraphicalProperties(ln=LineProperties(solidFill="CCCCCC", prstDash="dash"))
    chart.x_axis.majorGridlines = ChartLines(spPr=garis_grid)
    chart.y_axis.majorGridlines = ChartLines(spPr=garis_grid)
    chart.graphical_properties = GraphicalProperties(ln=LineProperties(solidFill="D9D9D9"))


def _tulis_header_14sheet(ws, headers: List[str], row: int = 1) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _HEADER_FONT_14SHEET
        cell.fill = _HEADER_FILL_14SHEET


def _gaya_header_baris(ws, row: int, ncols: int) -> None:
    """[BARU] Tempel gaya header (bold + fill) ke baris yang SUDAH ditulis
    isinya lewat ws.append() -- dipakai untuk sub-tabel tren bulanan di
    sheet Ringkasan, supaya tidak perlu menulis ulang value-nya."""
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _HEADER_FONT_14SHEET
        cell.fill = _HEADER_FILL_14SHEET


def _autofit_14sheet(ws, lebar_maks: int = 45) -> None:
    lebar_per_kolom: Dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            surat = cell.column_letter
            panjang = len(str(cell.value))
            if panjang > lebar_per_kolom.get(surat, 0):
                lebar_per_kolom[surat] = panjang
    for surat, panjang in lebar_per_kolom.items():
        ws.column_dimensions[surat].width = min(panjang + 2, lebar_maks)


def _format_rupiah_kolom(ws, kolom: List[int], baris_mulai: int, baris_selesai: int) -> None:
    for r in range(baris_mulai, baris_selesai + 1):
        for c in kolom:
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = _FORMAT_RUPIAH_14SHEET


def generate_neraca_saldo_awal_virtual(
    coa: List[Dict[str, Any]],
    periode: str,
    tanggal_awal: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """
    [BARU] Sistem HANYA menyimpan satu angka `saldo_awal` per akun COA --
    tidak ada baris jurnal pembukaan bertanggal tersimpan di database.
    Fungsi ini menyusun "jurnal pembukaan" VIRTUAL (dihitung sekali di
    memori untuk keperluan export, TIDAK ditulis ke jurnal_posting)
    semata supaya sheet "Neraca Saldo Awal" berbentuk baris jurnal
    (Tanggal, No. Dokumen, Debet/Kredit) yang konsisten dengan sheet
    GL 2025, bukan cuma tabel 2 kolom akun+saldo.

    Args:
        coa: list COA client (dari db_client.ambil_coa_client()).
        periode: string periode/tahun (mis. "2026" atau "2026-01").
        tanggal_awal: override tanggal virtual (default: 1 Jan tahun periode).

    Returns:
        list of dict siap ditulis ke sheet "Neraca Saldo Awal":
        [{"tanggal", "no_dokumen", "keterangan",
          "no_akun_debet", "nama_akun_debet", "jml_debet",
          "no_akun_kredit", "nama_akun_kredit", "jml_kredit",
          "lawan_transaksi", "project_unit"}, ...]
        [BARU] "lawan_transaksi"/"project_unit" diambil PER AKUN dari
        Coa.lawan_transaksi_saldo_awal/Coa.project_unit_saldo_awal (kalau
        akuntan sudah mengisinya di form COA) -- SEBELUMNYA kolom ini di
        sheet Neraca Saldo Awal hardcode "Pemilik"/"HO" utk SEMUA akun,
        salah kalau data perusahaan sebenarnya beragam (mis. akun
        excavator vs akun modal per pemilik punya lawan
        transaksi/unit berbeda-beda, lihat contoh user). Kalau akuntan
        belum mengisi kolom itu, fallback ke "-" (bukan lagi ditebak).
    """
    if not coa:
        return []

    tahun = (periode or "").split("-")[0] or str(date.today().year)
    tanggal_awal = tanggal_awal or f"{tahun}-01-01"
    keterangan = f"Saldo Awal {periode}"

    jurnal_pembukaan: List[Dict[str, Any]] = []
    for akun in sorted(coa, key=lambda a: str(a.get("no_akun") or "")):
        saldo_awal = _angka(akun.get("saldo_awal"))
        if saldo_awal == 0:
            continue
        normal_saldo = (akun.get("normal_saldo") or "DEBET").upper()
        no_akun = akun.get("no_akun")
        nama_akun = akun.get("nama_akun")

        baris = {
            "tanggal": tanggal_awal,
            "no_dokumen": "SALDO-AWAL",
            "keterangan": keterangan,
            "no_akun_debet": None, "nama_akun_debet": None, "jml_debet": 0.0,
            "no_akun_kredit": None, "nama_akun_kredit": None, "jml_kredit": 0.0,
            # [BARU] per akun, bukan lagi hardcode global
            "lawan_transaksi": akun.get("lawan_transaksi_saldo_awal") or "-",
            "project_unit": akun.get("project_unit_saldo_awal") or "-",
        }
        if normal_saldo == "KREDIT":
            baris["no_akun_kredit"] = no_akun
            baris["nama_akun_kredit"] = nama_akun
            baris["jml_kredit"] = round(saldo_awal, 2)
        else:
            baris["no_akun_debet"] = no_akun
            baris["nama_akun_debet"] = nama_akun
            baris["jml_debet"] = round(saldo_awal, 2)
        jurnal_pembukaan.append(baris)

    return jurnal_pembukaan


def get_tren_saldo_per_bulan(
    riwayat_akun: Dict[str, Dict[str, Any]],
    no_akun_list: List[str],
) -> List[Dict[str, Any]]:
    """
    [BARU] Gabungkan (jumlahkan) tren saldo bulanan untuk sekelompok akun
    (mis. semua akun Piutang) menjadi satu deret 12-bulan. Dipakai untuk
    baris "Tren Piutang Usaha" / "Tren Utang Usaha" pada sheet Ringkasan.

    Args:
        riwayat_akun: output db_client.ambil_riwayat_saldo_bulanan_akun_tren()
            -- {no_akun: {"nama_akun", "kategori", "data": [{"bulan","saldo_akhir"}, ...]}}
        no_akun_list: daftar no_akun yang mau digabung (mis. semua akun
            dengan sub_kategori "Piutang").

    Returns:
        list 12 dict: [{"bulan": 1, "saldo": ...}, ..., {"bulan": 12, "saldo": ...}]
    """
    total_per_bulan = {b: 0.0 for b in range(1, 13)}
    for no_akun in no_akun_list:
        info = riwayat_akun.get(str(no_akun))
        if not info:
            continue
        for item in info.get("data", []):
            bulan = item.get("bulan")
            if bulan in total_per_bulan:
                total_per_bulan[bulan] += _angka(item.get("saldo_akhir"))
    return [{"bulan": b, "saldo": round(total_per_bulan[b], 2)} for b in range(1, 13)]


_BULAN_SINGKAT = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                   "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
_BULAN_PANJANG = ["January", "February", "March", "April", "May", "June",
                   "July", "August", "September", "October", "November", "December"]

# [BARU - sheet GL] Kolom "Sumber" di GL sebelumnya diisi LANGSUNG dari
# kode internal jenis_dokumen (mis. "rekening_koran", "bukti_kas",
# "aset_tetap") -- benar datanya tapi tidak enak dibaca akuntan.
# Mapping ini menerjemahkan ke label ringkas manusiawi. Sengaja dibuat
# TERPISAH dari label panjang di main.py::_PEMROSES_DOKUMEN (yang untuk
# dropdown pilihan jenis dokumen, formatnya lebih verbose/deskriptif) --
# label di sini khusus utk ditampilkan padat dalam satu sel Excel.
# Kalau ada jenis_dokumen baru yang belum masuk daftar, fallback ke kode
# aslinya (title-case) supaya tetap tampil sesuatu, bukan kosong/error.
_LABEL_SUMBER_DOKUMEN = {
    "rekening_koran": "Bank",
    "penjualan": "Invoice penjualan",
    "penilaian_klien": "Penilaian klien",
    "buku_bantu_piutang": "Buku bantu piutang",
    "laporan_keuangan": "Laporan keuangan",
    "faktur_pajak": "Faktur pajak",
    "bukti_potong_pajak": "Bukti potong pajak",
    "spt_masa": "SPT masa",
    "slip_gaji": "Slip gaji",
    "bukti_kas": "Bukti kas",
    "kartu_stok": "Kartu stok",
    "aset_tetap": "Jadwal aktiva tetap",
    "pembelian": "Invoice vendor",
    "rekonsiliasi_bank": "Rekonsiliasi bank",
    "ap_aging": "AP aging",
    "absensi": "Absensi",
}


def _label_sumber(jenis_dokumen: Optional[str]) -> str:
    """Terjemahkan kode jenis_dokumen jadi label ringkas utk kolom "Sumber"."""
    if not jenis_dokumen:
        return "-"
    return _LABEL_SUMBER_DOKUMEN.get(jenis_dokumen, str(jenis_dokumen).replace("_", " ").title())


def _status_validasi_gl(kode_akun: Optional[str]) -> str:
    """
    [BARU] Isi kolom "Status Validasi" di sheet GL <tahun>.

    Sheet GL menulis SATU baris Excel per SISI transaksi (baris debet
    terpisah dari baris kredit -- lihat loop di sheet 4 GL <tahun>),
    jadi status di sini dicek dari kode akun BARIS ITU SENDIRI, bukan
    dari flag gabungan j["sumber_placeholder"] di level jurnal (yang
    bisa True walau cuma salah satu sisi -- debet ATAU kredit -- yang
    sebenarnya bermasalah, sisi lainnya sudah benar).

    Pola "/" di kode akun = penanda placeholder (AI belum yakin akun
    apa) -- pola ini SUDAH DIPAKAI di seluruh sistem sejak lama (lihat
    JurnalPosting.sumber_placeholder & konfirmasi_posting_massal() di
    db_client.py), bukan tebakan baru di sini.

    [KONTEKS] Sejak endpoint /export-14-sheet dipanggil dengan
    hanya_terposting=False, baris jurnal berstatus 'draft' (termasuk yang
    akunnya masih placeholder) ikut masuk apa adanya -- tidak ada lagi
    gate "harus posting manual dulu". Kolom ini menggantikan gate itu:
    murni informasi audit ("apa yang perlu dikoreksi" / "sudah lulus"),
    BUKAN validasi yang memblokir apa pun.
    """
    kode = str(kode_akun or "").strip()
    if not kode or kode.lower() == "none":
        return "Perlu dikoreksi: kode akun kosong"
    if "/" in kode:
        return f'Perlu dikoreksi: kode akun "{kode}" masih placeholder, tentukan akun sebenarnya'
    return "Sudah lulus"


def _tahun_dari_periode(periode: str) -> str:
    """Ambil 4 digit tahun dari string periode (mis. '2025', '2025-01', '2025-06-30')."""
    periode = (periode or "").strip()
    for token in periode.replace("/", "-").split("-"):
        if len(token) == 4 and token.isdigit():
            return token
    return str(date.today().year)


def _ambil_nilai_alias(row: Dict[str, Any], alias_list: List[str]) -> Any:
    """
    Ambil nilai dari dict `row` dengan mencoba beberapa kemungkinan nama
    kolom (alias) berurutan -- dipakai karena nama field asli di
    df_piutang/df_hutang (hasil akuntansi_ai.py) belum tentu 1:1 sama
    dengan nama kolom di template Excel referensi. Kalau semua alias
    tidak ditemukan, kembalikan None (bukan error) supaya sheet tetap
    kebentuk walau sebagian kolom kosong.
    """
    for alias in alias_list:
        if alias in row and row.get(alias) not in (None, ""):
            return row.get(alias)
    return None


def _pecah_per_bulan_debit_kredit(info_akun: Dict[str, Any], normal_saldo: str) -> List[tuple]:
    """
    Ubah 'per_bulan' dari trial_balance_bulanan menjadi 12 pasang
    (debit, kredit) sesuai konvensi Trial Balance -- satu sisi terisi
    per bulan sesuai saldo normal akun. Menerima 2 kemungkinan bentuk:
      a) list of dict [{"debit":.., "kredit":..}, ...] -> dipakai langsung
      b) list of angka (saldo/mutasi bulan itu) -> dipecah otomatis ke
         sisi Debit/Kredit berdasarkan tanda & saldo normal akun

    CATATAN: bentuk asli 'per_bulan' ditentukan oleh
    modules/laporan_keuangan.py::susun_laporan_bulanan_setahun() yang
    belum diupload ke sesi ini -- kalau ternyata bentuknya beda dari 2
    kemungkinan di atas, sesuaikan fungsi ini.
    """
    per_bulan = info_akun.get("per_bulan") or []
    normal_saldo = (normal_saldo or "DEBIT").upper()
    hasil = []
    for i in range(12):
        nilai = per_bulan[i] if i < len(per_bulan) else 0
        if isinstance(nilai, dict):
            hasil.append((_angka(nilai.get("debit")), _angka(nilai.get("kredit"))))
            continue
        nilai = _angka(nilai)
        if normal_saldo.startswith("KREDIT") or normal_saldo.startswith("CREDIT"):
            hasil.append((0.0, nilai) if nilai >= 0 else (abs(nilai), 0.0))
        else:
            hasil.append((nilai, 0.0) if nilai >= 0 else (0.0, abs(nilai)))
    return hasil


def _susun_workbook_18_sheet(data: Dict[str, Any]) -> "openpyxl.Workbook":
    """
    [REFACTOR] Badan asli dari export_18_sheet_lengkap() (isi/urutan sheet
    TIDAK diubah sama sekali) dipindah ke sini apa adanya, supaya bisa
    dipakai ulang oleh export_18_sheet_sebagai_json() -- lihat catatan di
    situ soal kenapa ini penting (satu sumber kebenaran, bukan 2 versi
    logic yang bisa tidak sinkron).

    Susun file Excel 18-sheet PERSIS mengikuti nama & urutan sheet model
    referensi (Model_Laporan_Keuangan_SPT_PPh31E_2025.xlsx), DITAMBAH 4
    sheet baru (11, 12, 13, 17) yang SUDAH diimplementasikan penuh (bukan
    placeholder kosong lagi -- lihat catatan "[FIX] 4 sheet placeholder
    di export 18-sheet SUDAH diimplementasikan" di riwayat perubahan):

      1.  Petunjuk & Asumsi
      2.  COA
      3.  Neraca Saldo Awal
      4.  GL <tahun>
      5.  Buku Bantu Piutang
      6.  Buku Bantu Hutang
      7.  Buku Bantu Aktiva Tetap
      8.  Trial Balance Bulanan
      9.  Laba Rugi Bulanan
      10. Balance Sheet Bulanan
      11. Laporan Perubahan Ekuitas  [BARU -- sudah diisi, lihat _tulis_sheet_perubahan_ekuitas()]
      12. Laporan Arus Kas           [BARU -- sudah diisi, lihat _tulis_sheet_arus_kas()]
      13. Catatan atas Laporan Keuangan (CALK) [BARU -- sudah diisi, lihat _tulis_sheet_calk()]
      14. Ringkasan
      15. BS Lampiran SPT
      16. PNL Lampiran SPT
      17. Rekonsiliasi Fiskal        [BARU -- sudah diisi, lihat _tulis_sheet_rekonsiliasi_fiskal()]
      18. PPh Badan 31E

    PERUBAHAN dari versi sebelumnya:
      - Sheet "Petunjuk & Asumsi" ditambahkan (sebelumnya tidak ada).
      - Sheet "Lampiran SPT Ekuitas" DIHAPUS dari export 14-sheet ini
        (template referensi hanya punya 14 sheet di atas, tidak ada
        sheet Ekuitas terpisah) -- datanya tetap tersedia ringkas di
        sheet "Ringkasan" bagian C (Perubahan Ekuitas).
      - "GL" -> "GL <tahun>", "Lampiran SPT BS"/"PNL" -> "BS Lampiran
        SPT"/"PNL Lampiran SPT" (nama & urutan kata disamakan template).
      - Sheet "GL <tahun>" & "Neraca Saldo Awal" diubah ke format SATU
        BARIS PER SISI JURNAL (Kode Akun + Debit ATAU Kredit per baris,
        dikelompokkan per Journal ID) -- sebelumnya 1 baris = pasangan
        akun debet+kredit digabung, tidak sama dengan format template.
      - Sheet COA, Trial Balance/Laba Rugi/Balance Sheet Bulanan, dan
        Buku Bantu Aktiva Tetap disusun ulang kolomnya supaya sama
        dengan header template (lihat catatan per-sheet di bawah).

    CATATAN PENTING:
      - [FIX] Sheet 7 (Buku Bantu Aktiva Tetap) sebelumnya SELALU
        menampilkan penyusutan rata 12 bulan (fallback) karena mencari
        key "per_bulan"/"penyusutan_bulanan" yang tidak pernah ada di
        output lapkeu.susun_jadwal_penyusutan_bulanan() -- sekarang
        dibaca dari key asli "jadwal_bulanan", jadi aset yang baru
        diperoleh di tengah tahun tampil 0 sebelum mulai disusutkan.
      - [FIX] Sheet 9 (Laba Rugi Bulanan) sebelumnya menampilkan saldo
        KUMULATIF YTD (dari trial_balance_bulanan) apa adanya per kolom
        bulan -- kolom Desember jadi memuat total setahun penuh, bukan
        pergerakan bulan Desember saja. Sekarang dihitung delta antar
        bulan dulu sebelum dijumlahkan ke baris LABA/RUGI BERSIH BULANAN.
      - Sheet 8 (Trial Balance) & 10 (Balance Sheet Bulanan) SENGAJA
        tetap pakai saldo kumulatif apa adanya -- itu memang sifat
        laporan Neraca/Trial Balance (posisi per akhir bulan, bukan
        pergerakan), jadi tidak perlu di-delta seperti sheet 9.
      - [FIX] Sheet Buku Bantu Piutang/Hutang pakai `_ambil_nilai_alias()`
        untuk memetakan nama kolom sumber ke header template. ALIAS_PIUTANG
        & ALIAS_HUTANG SUDAH dicocokkan langsung ke field asli hasil
        akuntansi_ai.py (bukan tebakan lagi) -- sebelumnya beberapa alias
        salah tebak (mis. Piutang: "no_invoice" vs asli "no_transaksi",
        "total"/"jumlah" vs asli "total_akhir"/"sub_total"; Hutang:
        "tanggal_tagihan" vs asli "tanggal_invoice", "nilai_tagihan" vs
        asli "jumlah_utang", "pembayaran" vs asli "jumlah_dibayar") yang
        bikin kolom-kolom itu diam-diam kosong di Excel. Sisi Piutang juga
        sebelumnya tidak pernah punya field umur_hari/bucket_aging/status/
        saldo PER BARIS sama sekali (beda dari AP Aging yang sudah punya) --
        sekarang ditempel lewat akuntansi_ai.py::_perkaya_piutang_per_baris().
        Kalau masih ada kolom kosong di masa depan, tambahkan alias baru di
        daftar `ALIAS_*` di bawah.

    Args (semua key lama tetap dipakai + 1 baru):
        "periode": str
        "tahun_sebelumnya": Optional[int]
        "coa": List[Dict]
        "jurnal": List[Dict]
        "df_piutang": Optional[List[Dict]]
        "df_hutang": Optional[List[Dict]]
        "jadwal_aset": Dict
        "laporan_bulanan": Dict
        "pph_hasil": Dict
        "neraca": Dict, "laba_rugi": Dict, "perubahan_ekuitas": Dict
        "arus_kas": Dict, "calk": Dict -- [BARU] dipakai sheet 12/13, lihat
            lapkeu.generate_5_laporan_keuangan()["arus_kas"/"calk"]
            (susun_arus_kas_sederhana()/susun_calk_otomatis()) -- sebelumnya
            2 key ini dihitung tapi TIDAK PERNAH diteruskan ke fungsi ini
            oleh main.py, sekarang sudah diperbaiki di sisi main.py juga.
        "lampiran_rinci": Dict
        "tren_piutang": Optional[List[Dict]]
        "tren_utang": Optional[List[Dict]]
        "asumsi": Optional[Dict] -- [BARU] data sheet "Petunjuk & Asumsi",
            kunci bebas mis. "nama_perusahaan", "periode_awal",
            "periode_akhir", "tanggal_laporan", "mata_uang",
            "metode_penyusutan", "nilai_residu", dst -- ditulis apa
            adanya sebagai baris Parameter/Nilai.

    Returns:
        openpyxl.Workbook: workbook siap dipakai lewat 2 jalur --
        export_18_sheet_lengkap() (disimpan jadi bytes .xlsx utk download)
        atau export_18_sheet_sebagai_json() (dibaca ulang jadi JSON utk
        ditampilkan di layar). Lihat kedua fungsi tipis di bawah.
    """
    periode = data.get("periode") or ""
    tahun = _tahun_dari_periode(periode)
    coa = data.get("coa") or []
    jurnal = data.get("jurnal") or []
    df_piutang = data.get("df_piutang") or []
    df_hutang = data.get("df_hutang") or []
    jadwal_aset = data.get("jadwal_aset") or {}
    laporan_bulanan = data.get("laporan_bulanan") or {}
    pph_hasil = data.get("pph_hasil") or {}
    neraca = data.get("neraca") or {}
    laba_rugi = data.get("laba_rugi") or {}
    perubahan_ekuitas = data.get("perubahan_ekuitas") or {}
    # [BARU] Dipakai sheet 11/12/13 (Laporan Perubahan Ekuitas/Arus Kas/CALK).
    arus_kas = data.get("arus_kas") or {}
    calk = data.get("calk") or {}
    lampiran_rinci = data.get("lampiran_rinci") or {}
    # [BARU -- integrasi Claude API] Hasil analisis naratif dari
    # claude_client.analisis_ringkasan_keuangan_claude(), digenerate oleh
    # pemanggil SEBELUM data dikirim ke sini -- lihat blok "Temuan
    # Penting & Potensi Masalah" di sheet "Ringkasan" di bawah.
    ringkasan_analisis = data.get("ringkasan_analisis") or {}
    # [BARU] Dipakai sheet "BS Lampiran SPT" (format baku) -- kalau tidak
    # dikirim eksplisit lewat data["tahun_sebelumnya"], fallback tahun-1.
    try:
        tahun_sebelumnya = data.get("tahun_sebelumnya") or (int(tahun) - 1)
    except (TypeError, ValueError):
        tahun_sebelumnya = None
    tren_piutang = data.get("tren_piutang") or []
    tren_utang = data.get("tren_utang") or []
    asumsi = data.get("asumsi") or {}

    # Lookup no_akun -> data COA, dipakai lintas sheet (segment, saldo normal, dst)
    peta_akun = {str(a.get("no_akun")): a for a in coa if a.get("no_akun") is not None}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ================= 1. Petunjuk & Asumsi =================
    # [FIX] Ditulis ulang total supaya PERSIS mengikuti struktur & gaya
    # file model referensi (MODEL_LAPORAN_KEUANGAN_OTOMATIS...xlsx):
    #   - SATU tabel Parameter/Nilai berkelanjutan dari Nama Perusahaan
    #     s.d. Umur Manfaat (bukan 2 tabel terpisah dgn header ganda).
    #   - Judul & semua baris header: fill navy #17365D + font Carlito
    #     putih bold, center + wrap (sebelumnya: fill biru muda D9E1F2,
    #     font hitam, rata kiri, judul sheet tanpa gaya sama sekali).
    #   - Cell nilai/input (Modal, Kas/Bank/Aset Awal, dst): font biru +
    #     fill kuning #FFF2CC, sesuai legenda "Warna biru" di sheet ini
    #     sendiri (sebelumnya polos tanpa warna sama sekali).
    #   - "Keterangan Model" & "Kontrol Saldo Awal" sekarang jadi kolom A
    #     dari baris header itu sendiri (sebelumnya baris judul terpisah
    #     di ATAS baris header -- bikin seluruh tabel legenda geser 1
    #     kolom dibanding file model).
    #   - "Kontrol Saldo Awal" pakai FORMULA EXCEL asli (SUM/IF) yang
    #     mereferensikan baris2 Modal/Kas/Aset di sheet ini sendiri --
    #     bukan angka hasil hitung Python yang beku. Logika kontrolnya
    #     juga dikembalikan ke pola file model (Total Modal vs Total
    #     Kas+Aset Tetap) karena itu satu-satunya perbandingan yang bisa
    #     jadi formula SUM sheet-lokal yang valid (mereferensikan baris
    #     yang memang ada tertulis di sheet ini, bukan seluruh COA).
    #   - Tanggal (Periode Awal/Akhir, Tanggal Laporan) diparse jadi
    #     objek datetime asli + format dd-mmm-yyyy (sebelumnya teks ISO
    #     polos "2025-01-01").
    #   - Lebar kolom disamakan manual persis file model (bukan autofit).
    # Bagian company-specific (Modal/Kas/Bank/Aset Awal, Kontrol Saldo
    # Awal) tetap DIHITUNG DINAMIS dari `coa` & `jadwal_aset` (bukan
    # hardcode 1 contoh perusahaan seperti di file model) supaya otomatis
    # menyesuaikan berapa pun jumlah akun modal/aset yang dikirim client.
    ws = wb.create_sheet("Petunjuk & Asumsi")

    ws.append(["PETUNJUK & ASUMSI LAPORAN KEUANGAN"])
    ws.cell(row=1, column=1).font = _TITLE_FONT_ASUMSI
    ws.cell(row=1, column=1).fill = _TITLE_FILL_ASUMSI
    ws.cell(row=1, column=1).alignment = _TITLE_ALIGN_ASUMSI
    ws.row_dimensions[1].height = 27.95
    # [FIX] Judul di file model referensi merged A1:H1 (navy penuh selebar
    # tabel 8 kolom) -- sebelumnya tidak pernah di-merge sama sekali, jadi
    # kalau dibuka di Excel/Sheets baris judul cuma navy selebar kolom A,
    # sisanya putih polos. Cukup merge saja (dicek langsung ke file model
    # referensi: sel B1-H1 di sana pun tidak diberi fill sendiri -- utk
    # cell yang di-merge, Excel/Sheets selalu merender pakai fill sel
    # pojok kiri-atas/anchor-nya saja, di sini A1).
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    ws.append([None])
    _tulis_header_asumsi(ws, ["Parameter", "Nilai / Asumsi"], row=ws.max_row + 1)

    kunci_baku = ["nama_perusahaan", "periode_awal", "periode_akhir",
                  "tanggal_laporan", "mata_uang", "metode_penyusutan", "nilai_residu"]
    label_baku = {
        "nama_perusahaan": "Nama Perusahaan", "periode_awal": "Periode Awal",
        "periode_akhir": "Periode Akhir", "tanggal_laporan": "Tanggal Laporan / Aging",
        "mata_uang": "Mata Uang", "metode_penyusutan": "Metode Penyusutan",
        "nilai_residu": "Nilai Residu",
    }
    kunci_tanggal = ("periode_awal", "periode_akhir", "tanggal_laporan")
    daftar_aset = jadwal_aset.get("aset") or []

    def _numfmt_untuk(k: str, v: Any) -> Optional[str]:
        if k in kunci_tanggal and isinstance(v, (date, datetime)):
            return _FORMAT_TANGGAL_ASUMSI
        if k == "nilai_residu" and isinstance(v, (int, float)):
            return _FORMAT_RUPIAH_ASUMSI
        return None

    nama_perusahaan_val = asumsi.get("nama_perusahaan") or pph_hasil.get("nama_perusahaan", "")
    _tulis_input_asumsi(ws, ws.max_row + 1, label_baku["nama_perusahaan"], nama_perusahaan_val)
    for k in kunci_baku[1:]:
        if k == "metode_penyusutan":
            default = "Garis Lurus"
        elif k == "mata_uang":
            default = "IDR"
        elif k == "nilai_residu":
            # [BARU] Sebelumnya selalu kosong (tidak ada default & tidak
            # pernah dikirim dari main.py) -- dihitung dari data aset:
            # kalau semua aset punya nilai residu yang sama, tampilkan
            # nilainya; kalau berbeda-beda antar aset, tampilkan
            # "Bervariasi per aset" (tidak masuk akal menampilkan 1 angka
            # tunggal yang menyesatkan).
            nilai_residu_unik = {_angka(a.get("nilai_residu")) for a in daftar_aset}
            default = (nilai_residu_unik.pop() if len(nilai_residu_unik) == 1
                       else ("Bervariasi per aset" if nilai_residu_unik else ""))
        else:
            default = ""
        nilai = asumsi.get(k, default)
        if k in kunci_tanggal:
            nilai = _ke_tanggal_asumsi(nilai)
        _tulis_input_asumsi(ws, ws.max_row + 1, label_baku[k], nilai, _numfmt_untuk(k, nilai))
    for k, v in asumsi.items():
        if k in kunci_baku:
            continue
        _tulis_input_asumsi(ws, ws.max_row + 1, k, v)

    # -- Modal, Kas/Bank & Aset Awal (dinamis dari COA + jadwal aset) --
    # [FIX] SATU tabel yang sama dgn Parameter di atas (tidak ada header
    # "Parameter/Nilai" kedua di tengah) -- persis struktur file model
    # referensi. Baris awal/akhir tiap blok dicatat supaya formula
    # "Kontrol Saldo Awal" di bawah bisa mereferensikan range ini tepat.
    baris_modal_awal = ws.max_row + 1
    for akun in coa:
        kategori = str(akun.get("kategori") or "").strip().upper()
        nama_akun = str(akun.get("nama_akun") or "")
        saldo_awal = akun.get("saldo_awal")
        if kategori == "EKUITAS" and saldo_awal:
            # Hindari label ganda "Modal Awal Modal Tuan A" kalau nama akun
            # di COA client sudah diawali kata "Modal" (pola penamaan yang
            # sangat umum) -- kata itu dilepas dulu sebelum diprefiks.
            nama_bersih = nama_akun
            for prefiks in ("Modal ", "modal "):
                if nama_bersih.startswith(prefiks):
                    nama_bersih = nama_bersih[len(prefiks):]
                    break
            _tulis_input_asumsi(ws, ws.max_row + 1, f"Modal Awal {nama_bersih}",
                                 _angka(saldo_awal), _FORMAT_RUPIAH_ASUMSI)
    baris_modal_akhir = ws.max_row

    baris_kasaset_awal = ws.max_row + 1
    for akun in coa:
        kategori = str(akun.get("kategori") or "").strip().upper()
        nama_akun_upper = str(akun.get("nama_akun") or "").upper()
        sub_kategori = str(akun.get("sub_kategori") or "").upper()
        saldo_awal = akun.get("saldo_awal")
        if kategori == "ASET" and (sub_kategori in ("KAS", "BANK") or "KAS" in nama_akun_upper or "BANK" in nama_akun_upper):
            _tulis_input_asumsi(ws, ws.max_row + 1, f"{akun.get('nama_akun')} Awal",
                                 _angka(saldo_awal), _FORMAT_RUPIAH_ASUMSI)
    for aset in daftar_aset:
        _tulis_input_asumsi(ws, ws.max_row + 1, f"Harga Perolehan {aset.get('nama_aset')}",
                             _angka(aset.get("harga_perolehan")), _FORMAT_RUPIAH_ASUMSI)
    baris_kasaset_akhir = ws.max_row

    # Umur manfaat dikelompokkan per kombinasi (kategori aset, umur_tahun)
    # supaya aset sejenis (mis. 2 unit excavator umur sama) tidak
    # menghasilkan baris duplikat -- label diambil dari kategori aset kalau
    # ada, fallback ke nama aset pertama dalam kelompok itu. Baris ini
    # SENGAJA di luar range baris_kasaset_awal:baris_kasaset_akhir (bukan
    # nilai uang, jangan ikut ke-SUM formula Kontrol Saldo Awal).
    kelompok_umur: Dict[tuple, List[str]] = {}
    for aset in daftar_aset:
        kunci = (aset.get("kategori") or aset.get("nama_aset"), aset.get("umur_tahun"))
        kelompok_umur.setdefault(kunci, []).append(str(aset.get("nama_aset") or ""))
    for (label, umur), nama_list in kelompok_umur.items():
        if umur in (None, ""):
            continue
        label_tampil = label if label and label not in nama_list else " / ".join(dict.fromkeys(nama_list))
        _tulis_input_asumsi(ws, ws.max_row + 1, f"Umur Manfaat {label_tampil} (tahun)",
                             umur, _FORMAT_INT_ASUMSI)

    # -- Keterangan Model (legenda, statis -- sama utk semua client) --
    # [FIX] "Keterangan Model" sekarang kolom A dari baris header 8-kolom
    # itu sendiri (sebelumnya baris judul terpisah di atasnya, bikin
    # tabel legenda geser 1 kolom dibanding posisi di file model).
    ws.append([None])
    for c in (1, 2):
        ws.cell(row=ws.max_row, column=c).border = _border_isi_asumsi(1, 2, c)
    _tulis_header_asumsi(ws, [
        "Keterangan Model", "Warna biru", "Warna hijau", "Warna hitam",
        "Saldo Awal", "GL", "Buku Pembantu", "Laporan",
    ], row=ws.max_row + 1)
    ws.append([
        f"Ringkasan parameter & asumsi laporan keuangan tahun {tahun}. Sel biru dapat diubah sesuai data perusahaan.",
        "Input/hardcode yang dapat diubah.",
        "Formula yang mengambil data dari sheet lain.",
        "Formula/perhitungan dalam sheet yang sama.",
        f"Jurnal pembukaan per 1 Januari {tahun}.",
        f"Seluruh transaksi jurnal {tahun} dengan segment dan lawan transaksi.",
        "Piutang, hutang, dan aktiva tetap terhubung ke GL.",
        f"TB, Laba Rugi, dan Balance Sheet bulanan hingga 31 Desember {tahun}.",
    ])
    baris_ket = ws.max_row
    for c in range(1, 9):
        ws.cell(row=baris_ket, column=c).alignment = _DESC_ALIGN_ASUMSI
    # [FIX] Di file model referensi, baris isi "Keterangan Model" cuma
    # dapat border kotak selebar 2 kolom (A-B) walau headernya di atas
    # selebar 8 kolom -- quirk asli file model, direplikasi persis di sini.
    for c in (1, 2):
        ws.cell(row=baris_ket, column=c).border = _border_isi_asumsi(1, 2, c)

    # -- [BARU -- integrasi Claude API] Catatan Interpretasi Asumsi --
    # SENGAJA ditaruh SETELAH tabel "Keterangan Model" & SEBELUM "Kontrol
    # Saldo Awal", jadi TIDAK mengubah/menggeser struktur baku yang
    # meniru file referensi di atas & di bawahnya. Narasi ini digenerate
    # oleh pemanggil lewat claude_client.generate_narasi_asumsi_claude()
    # SEBELUM data sampai di sini, dikirim sebagai
    # asumsi["narasi_ai"] -- kalau kosong (belum digenerate/API gagal),
    # blok ini cuma dilewati, tidak menggagalkan export. Ditulis italic
    # abu-abu (sama seperti _CATATAN_FONT_PLACEHOLDER_BARU) supaya
    # kelihatan sebagai catatan tambahan, bukan bagian tabel input/formula.
    narasi_asumsi = asumsi.get("narasi_ai") or ""
    if narasi_asumsi:
        ws.append([None])
        baris_judul_narasi = ws.max_row + 1
        ws.cell(row=baris_judul_narasi, column=1, value="Catatan Interpretasi Asumsi (AI)")
        ws.merge_cells(start_row=baris_judul_narasi, start_column=1,
                        end_row=baris_judul_narasi, end_column=8)
        cell_jn = ws.cell(row=baris_judul_narasi, column=1)
        cell_jn.font = _SUBTITLE_FONT_ASUMSI
        cell_jn.fill = _SUBTITLE_FILL_ASUMSI
        baris_isi_narasi = ws.max_row + 1
        ws.cell(row=baris_isi_narasi, column=1, value=narasi_asumsi)
        ws.merge_cells(start_row=baris_isi_narasi, start_column=1,
                        end_row=baris_isi_narasi, end_column=8)
        cell_in = ws.cell(row=baris_isi_narasi, column=1)
        cell_in.font = Font(name="Carlito", italic=True, size=10, color="FF595959")
        cell_in.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[baris_isi_narasi].height = 45

    # -- Kontrol Saldo Awal (formula Excel asli, mereferensikan baris2 di
    # atas) -- struktur & logika kembali PERSIS file model referensi:
    # Total Modal (SUM baris "Modal Awal ...") dibandingkan Total Kas +
    # Aset Tetap (SUM baris "... Awal" / "Harga Perolehan ..."), bukan
    # total seluruh COA -- ini satu-satunya perbandingan yang bisa jadi
    # formula SUM sheet-lokal yang valid (mereferensikan baris yang
    # memang tertulis di sheet ini sendiri, bisa dibuka & dihitung ulang
    # langsung di Excel, bukan angka beku hasil hitung Python).
    ws.append([None])
    for c in (1, 2):
        ws.cell(row=ws.max_row, column=c).border = _border_isi_asumsi(1, 2, c)
    _tulis_header_asumsi(ws, ["Kontrol Saldo Awal", "Formula"], row=ws.max_row + 1)
    baris_total_modal = ws.max_row + 1
    formula_modal = (f"=SUM(B{baris_modal_awal}:B{baris_modal_akhir})"
                      if baris_modal_akhir >= baris_modal_awal else 0)
    _tulis_formula_asumsi(ws, baris_total_modal, "Total Modal", formula_modal, _FORMAT_RUPIAH_ASUMSI)
    baris_total_kasaset = ws.max_row + 1
    formula_kasaset = (f"=SUM(B{baris_kasaset_awal}:B{baris_kasaset_akhir})"
                        if baris_kasaset_akhir >= baris_kasaset_awal else 0)
    _tulis_formula_asumsi(ws, baris_total_kasaset, "Total Kas + Aset Tetap", formula_kasaset, _FORMAT_RUPIAH_ASUMSI)
    baris_selisih = ws.max_row + 1
    _tulis_formula_asumsi(ws, baris_selisih, "Selisih",
                           f"=B{baris_total_kasaset}-B{baris_total_modal}", _FORMAT_RUPIAH_ASUMSI)
    baris_status = ws.max_row + 1
    _tulis_formula_asumsi(ws, baris_status, "Status",
                           f'=IF(B{baris_selisih}=0,"BALANCE","PERIKSA")', tutup_bawah=True)

    # -- Tambahan untuk SPT Tahunan Badan (statis -- sama utk semua client) --
    ws.append([None])
    ws.append(["TAMBAHAN UNTUK SPT TAHUNAN BADAN"])
    baris_judul_spt = ws.max_row
    ws.cell(row=baris_judul_spt, column=1).font = _SUBTITLE_FONT_ASUMSI
    ws.cell(row=baris_judul_spt, column=1).fill = _SUBTITLE_FILL_ASUMSI
    # [FIX] Garis bawah tipis selebar 8 kolom (A-H) di bawah subjudul ini,
    # sesuai file model referensi.
    for c in range(1, 9):
        ws.cell(row=baris_judul_spt, column=c).border = Border(bottom=_GARIS_ASUMSI)
    # [FIX] Header tabel ini di file model referensi border-nya selebar 8
    # kolom (A-H) walau labelnya cuma 5 (Sheet/Fungsi/Input Biru/Formula
    # Hijau/Catatan) -- disamakan lebar kotaknya lewat lebar_kotak=8.
    _tulis_header_asumsi(ws, ["Sheet", "Fungsi", "Input Biru", "Formula Hijau", "Catatan"],
                          row=ws.max_row + 1, lebar_kotak=8)
    daftar_baris_spt = (
        ["BS Lampiran SPT", "Neraca komparatif tahun berjalan dan saldo awal tahun sebelumnya",
         "Akun tambahan jika ada", "Terhubung ke Balance Sheet dan saldo awal",
         "Check balance harus nihil"],
        ["PNL Lampiran SPT", "Laba rugi tahunan dan ringkasan fiskal",
         "Koreksi per akun opsional", "Terhubung ke PNL bulanan",
         "Rincian koreksi utama berada pada PNL bulanan"],
        ["PPh Badan 31E", "Menghitung PPh terutang dan kredit pajak",
         "Skema, omzet tambahan, kredit pajak", "Terhubung ke rekonsiliasi fiskal",
         "Gunakan hanya jika skema tarif umum berlaku"],
    )
    for idx, baris in enumerate(daftar_baris_spt):
        ws.append(baris)
        r = ws.max_row
        for c in range(1, 6):
            ws.cell(row=r, column=c).alignment = _DESC_ALIGN_ASUMSI
        # [FIX] Border kotak baris isi tabel ini selebar 8 kolom (A-H) di
        # file model referensi (beda dari tabel "Keterangan Model" yang
        # cuma 2 kolom) -- baris TERAKHIR (PPh Badan 31E) menutup kotak
        # dgn garis bawah PENUH selebar 8 kolom (beda dari tabel "Kontrol
        # Saldo Awal" yang nutupnya cuma di 2 ujung kolom).
        tutup = idx == len(daftar_baris_spt) - 1
        for c in (1, 8):
            ws.cell(row=r, column=c).border = _border_isi_asumsi(1, 8, c, tutup_bawah=False)
        if tutup:
            for c in range(1, 9):
                b = ws.cell(row=r, column=c).border
                ws.cell(row=r, column=c).border = Border(
                    left=b.left, right=b.right, top=b.top, bottom=_GARIS_ASUMSI)

    # -- Lebar kolom manual, sama persis file model referensi (bukan
    # autofit -- sheet ini satu2nya yang diminta sama persis) --
    for kolom, lebar in _LEBAR_KOLOM_ASUMSI.items():
        ws.column_dimensions[kolom].width = lebar

    # ================= 2. COA =================
    # [FIX] Sheet ini SEBELUMNYA pakai gaya generik _tulis_header_14sheet
    # (fill biru muda #D9E1F2, tanpa font putih, tanpa Carlito, tanpa
    # border) + _autofit_14sheet (lebar kolom auto) -- hasilnya beda jauh
    # dari file model referensi (CHART_OF_ACCOUNTS__COA_.xlsx), yang
    # gaya-nya dianalisis ulang: judul & header fill navy #17365D +
    # font Carlito putih bold, data font Carlito biru #0000FF, lebar
    # kolom manual. Diganti pakai konstanta _..._COA di atas supaya
    # sama persis. "Laporan" tetap diturunkan otomatis dari kategori
    # (kolom itu tidak ada di database) -- lihat _LAPORAN_DARI_KATEGORI.
    ws = wb.create_sheet("COA")

    # [FIX] Baris ditulis dengan NOMOR EKSPLISIT (bukan ws.append() +
    # ws.max_row) -- ws.append([]) untuk baris spacer kosong TIDAK
    # menaikkan ws.max_row (tidak ada cell yang benar2 ditulis), jadi
    # kalau baris header dihitung dari ws.max_row+1 setelahnya, header
    # ikut "nempel" ke baris 2 (persis bug di file yang sudah di-generate
    # sebelumnya -- tidak ada spasi antara judul & header). Struktur baris
    # file model referensi: 1=judul, 2=kosong, 3=header, 4+=data.

    # -- Baris 1: judul, merge A1:I1, fill navy + font putih Carlito 15 --
    cell_judul = ws.cell(row=1, column=1, value="CHART OF ACCOUNTS (COA)")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    cell_judul.font = _TITLE_FONT_COA
    cell_judul.fill = _TITLE_FILL_COA
    cell_judul.alignment = _TITLE_ALIGN_COA
    ws.row_dimensions[1].height = 27.95

    # -- Baris 2: spacer kosong (sama seperti file model referensi) --
    # (tidak ada apa pun ditulis di sini secara sengaja)

    # -- Baris 3: header kolom, fill navy + font putih Carlito 11 bold,
    #    center + wrap + border tipis atas-bawah --
    header_row = 3
    headers_coa = [
        "Kode Akun", "Nama Akun", "Kelompok", "Subkelompok",
        "Saldo Normal", "Laporan", "Arus Kas", "Segment", "Keterangan",
    ]
    for col, h in enumerate(headers_coa, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = _HEADER_FONT_COA
        cell.fill = _HEADER_FILL_COA
        cell.alignment = _HEADER_ALIGN_COA
        cell.border = _HEADER_BORDER_COA
    ws.row_dimensions[header_row].height = 30

    def _laporan_dari_kategori(kategori_raw: Optional[str]) -> Optional[str]:
        """[FIX] Sebelumnya exact-match ke dict 5 key (ASET/LIABILITAS/
        EKUITAS/PENDAPATAN/BEBAN) -- gagal untuk kategori "Pendapatan
        Lain"/"Beban Lain" (2 kategori TAMBAHAN yang ada di file model
        referensi, kolom "Laporan"-nya tetap "Laba Rugi") karena
        kategori.upper() jadi "PENDAPATAN LAIN"/"BEBAN LAIN", tidak sama
        persis dgn key "PENDAPATAN"/"BEBAN". Kolom `kategori` di database
        juga teks bebas (String(20), tanpa enum) -- client bisa menulis
        "Kewajiban" alih-alih "Liabilitas", atau "Modal" alih-alih
        "Ekuitas". Diganti pakai pencocokan substring/kata kunci supaya
        tahan variasi penulisan, bukan exact match."""
        k = (kategori_raw or "").strip().upper()
        if not k:
            return None
        if "ASET" in k or "AKTIVA" in k:
            return "Balance Sheet"
        if "LIABILITAS" in k or "KEWAJIBAN" in k or "UTANG" in k or "HUTANG" in k:
            return "Balance Sheet"
        if "EKUITAS" in k or "MODAL" in k:
            return "Balance Sheet"
        if "PENDAPATAN" in k or "BEBAN" in k or "BIAYA" in k:
            # Menangkap "Pendapatan", "Pendapatan Lain", "Beban",
            # "Beban Lain", dan variasi "Biaya" -- semuanya "Laba Rugi".
            return "Laba Rugi"
        return None

    # -- Baris 4 dst: data, font Carlito biru #0000FF (nomor baris
    #    eksplisit juga, lanjut dari header_row -- konsisten dgn alasan
    #    di atas, tidak pakai ws.append()) --
    for offset, akun in enumerate(coa):
        # [FIX - export 14 sheet] "Laporan" sebelumnya dibaca dari
        # akun.get("laporan") -- key itu TIDAK PERNAH ada di dict hasil
        # db_client.ambil_coa_client() (model Coa tidak punya kolom ini),
        # jadi kolom ini selalu kosong. Diturunkan otomatis dari kategori
        # lewat _laporan_dari_kategori() di atas -- tidak perlu kolom baru
        # di database. Nilai disamakan dgn label file model referensi
        # ("Balance Sheet"/"Laba Rugi", bukan lagi "Neraca"). "Keterangan"
        # tetap apa adanya dari database (kosong kalau memang belum diisi
        # akuntan lewat form COA).
        laporan = _laporan_dari_kategori(akun.get("kategori"))
        row_values = [
            akun.get("no_akun"), akun.get("nama_akun"), akun.get("kategori"),
            akun.get("sub_kategori"), akun.get("normal_saldo"), laporan,
            akun.get("arus_kas"), akun.get("segment"), akun.get("keterangan"),
        ]
        r = header_row + 1 + offset
        for c, v in enumerate(row_values, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = _DATA_FONT_COA
            cell.alignment = _DATA_ALIGN_COA

    # -- Lebar kolom OTOMATIS menyesuaikan panjang teks TERPANJANG yang
    #    benar-benar ada di kolom itu (header MAUPUN seluruh baris data)
    #    -- [FIX] sebelumnya pakai _lebar_kolom_dari_header() yang cuma
    #    menghitung dari panjang teks header, jadi kolom berheader pendek
    #    tapi isinya panjang (mis. "Nama Akun" diisi "Akumulasi Penyusutan
    #    Kendaraan") kepotong. Dipanggil di sini karena harus SETELAH
    #    semua baris data (header_row+1 .. header_row+len(coa)) selesai
    #    ditulis di atas -- _lebar_kolom_dari_isi() butuh isi final untuk
    #    dihitung, memanggilnya sebelum data ditulis akan balik lagi jadi
    #    lebar-dari-header-saja (kolom masih kosong).
    for kolom, lebar in _lebar_kolom_dari_isi(ws).items():
        ws.column_dimensions[kolom].width = lebar

    # ================= 3. Neraca Saldo Awal =================
    # [FIX] Sheet ini SEBELUMNYA pakai _tulis_header_14sheet (fill biru
    # pucat #D9E1F2, tanpa font putih/Carlito, tanpa border) + spacer lewat
    # ws.append([]) -- ws.append([]) TIDAK menaikkan ws.max_row (tidak ada
    # cell yang benar2 ditulis), jadi header ikut "nempel" ke baris 2
    # (bukan baris 3 sesuai file model referensi
    # NERACA_SALDO_AWAL___FORMAT_GENERAL_LEDGER___1_JANUARI_2025.xlsx), dan
    # baris "Kontrol" ikut salah posisi karena baris acuannya (ws.max_row)
    # sudah kacau dari awal. Sama seperti fix sheet COA di atas: baris
    # ditulis dgn NOMOR EKSPLISIT, gaya disamakan navy #17365D + putih
    # Carlito utk judul/header, biru utk data hardcode, HIJAU utk kolom
    # "Nama Akun" (sekarang FORMULA, bukan value polos -- lihat di bawah),
    # lebar kolom manual (bukan autofit).
    ws = wb.create_sheet("Neraca Saldo Awal")
    jurnal_pembukaan = generate_neraca_saldo_awal_virtual(coa, periode)

    # -- Baris 1: judul, merge A1:M1, fill navy + font putih Carlito 15 --
    cell_judul = ws.cell(
        row=1, column=1,
        value=f"NERACA SALDO AWAL – FORMAT GENERAL LEDGER – 1 JANUARI {tahun}",
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    cell_judul.font = _TITLE_FONT_NSA
    cell_judul.fill = _TITLE_FILL_NSA
    cell_judul.alignment = _TITLE_ALIGN_NSA
    ws.row_dimensions[1].height = 27.95

    # -- Baris 2: spacer kosong (sengaja tidak ditulis apa pun -- struktur
    #    file model referensi: 1=judul, 2=kosong, 3=header, 4+=data) --

    # -- Baris 3: header kolom, fill navy + font putih Carlito 11 bold,
    #    center + border tipis semua sisi --
    header_row = 3
    headers_nsa = [
        "No", "Journal ID", "Tanggal", "No. Dokumen", "Kode Akun", "Nama Akun",
        "Keterangan", "Debit", "Kredit", "Segment", "Lawan Transaksi",
        "Project/Asset Unit", "Sumber",
    ]
    for col, h in enumerate(headers_nsa, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = _HEADER_FONT_NSA
        cell.fill = _HEADER_FILL_NSA
        cell.alignment = _HEADER_ALIGN_NSA
        cell.border = _HEADER_BORDER_NSA
    ws.row_dimensions[header_row].height = 30

    # -- Baris 4 dst: data jurnal pembukaan virtual. Kolom "Nama Akun"
    #    ditulis sbg FORMULA VLOOKUP ke sheet "COA" DI WORKBOOK YANG SAMA
    #    (bukan link file eksternal "[1]COA!" seperti file model lama yang
    #    filenya sendiri sudah tidak ikut disertakan) -- sheet "COA" pasti
    #    ada di workbook ini karena dibuat tepat sebelum sheet ini (lihat
    #    "==== 2. COA ====" di atas), jadi formulanya benar2 bisa dihitung
    #    Excel/LibreOffice, bukan #REF!. Range COA disesuaikan otomatis ke
    #    jumlah akun yang sebenarnya (bukan hardcode $A$4:$B$34). --
    baris_awal_coa = 4
    baris_akhir_coa = baris_awal_coa - 1 + max(len(coa), 1)
    for i, b in enumerate(jurnal_pembukaan, 1):
        no_akun = b.get("no_akun_debet") or b.get("no_akun_kredit")
        r = header_row + i
        info_akun = peta_akun.get(str(no_akun), {})
        # [FIX - Neraca Saldo Awal] "Lawan Transaksi"/"Project/Asset Unit"
        # SEBELUMNYA hardcode "Pemilik"/"HO" utk SEMUA baris tanpa
        # memandang akunnya apa -- sekarang diambil per akun (lihat
        # generate_neraca_saldo_awal_virtual), otomatis menyesuaikan
        # kalau data COA client berbeda-beda.
        nilai_baris = [
            i, f"OB-{tahun}", _ke_tanggal_asumsi(b.get("tanggal")), b.get("no_dokumen"),
            no_akun,
            f"=VLOOKUP(E{r},COA!$A${baris_awal_coa}:$B${baris_akhir_coa},2,FALSE)",
            b.get("keterangan"),
            b.get("jml_debet", 0), b.get("jml_kredit", 0),
            info_akun.get("segment") or "-",
            b.get("lawan_transaksi", "-"), b.get("project_unit", "-"),
            f"Saldo awal {periode}",
        ]
        for c, v in enumerate(nilai_baris, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = _DATA_FONT_FORMULA_NSA if c == 6 else _DATA_FONT_NSA
            cell.alignment = _DATA_ALIGN_NSA
            if i % 2 == 1:  # [FIX] banding: baris data ganjil (ke-1,3,5,..) biru muda
                cell.fill = _BANDING_FILL_NSA
            if c == 3:
                cell.number_format = _FORMAT_TANGGAL_NSA
            elif c in (8, 9):
                cell.number_format = _FORMAT_RUPIAH_NSA

    total_debet = round(sum(_angka(b.get("jml_debet")) for b in jurnal_pembukaan), 2)
    total_kredit = round(sum(_angka(b.get("jml_kredit")) for b in jurnal_pembukaan), 2)

    # -- Section "Kontrol": 1 baris kosong setelah data terakhir, lalu
    #    header Kontrol/Debit/Kredit (gaya sama dgn header kolom di atas,
    #    persis file model referensi G11:I11), baris Total (=SUM), baris
    #    Selisih (=Total Debit - Total Kredit) --
    baris_terakhir_data = header_row + len(jurnal_pembukaan)
    baris_header_kontrol = baris_terakhir_data + 2
    for col, h in zip((7, 8, 9), ("Kontrol", "Debit", "Kredit")):
        cell = ws.cell(row=baris_header_kontrol, column=col, value=h)
        cell.font = _HEADER_FONT_NSA
        cell.fill = _HEADER_FILL_NSA
        cell.alignment = _HEADER_ALIGN_NSA
        cell.border = _HEADER_BORDER_NSA

    r_total = baris_header_kontrol + 1
    ws.cell(row=r_total, column=7, value="Total")
    if jurnal_pembukaan:
        ws.cell(row=r_total, column=8,
                 value=f"=SUM(H{header_row + 1}:H{baris_terakhir_data})").number_format = _FORMAT_RUPIAH_NSA
        ws.cell(row=r_total, column=9,
                 value=f"=SUM(I{header_row + 1}:I{baris_terakhir_data})").number_format = _FORMAT_RUPIAH_NSA
    else:
        ws.cell(row=r_total, column=8, value=0).number_format = _FORMAT_RUPIAH_NSA
        ws.cell(row=r_total, column=9, value=0).number_format = _FORMAT_RUPIAH_NSA

    r_selisih = r_total + 1
    ws.cell(row=r_selisih, column=7, value="Selisih")
    ws.cell(row=r_selisih, column=8, value=f"=H{r_total}-I{r_total}").number_format = _FORMAT_RUPIAH_NSA

    # -- Lebar kolom OTOMATIS menyesuaikan panjang teks header (bukan
    #    lebar manual/patokan file model referensi lagi) --
    for kolom, lebar in _lebar_kolom_dari_isi(ws).items():
        ws.column_dimensions[kolom].width = lebar
    # di section ini) supaya sheet "Trial Balance Bulanan" di bawah bisa
    # mereferensikannya lewat formula SUMIFS LOKAL dlm workbook yang sama
    # (bukan link file eksternal "[1]Neraca Saldo Awal!" spt file model
    # referensi aslinya -- link eksternal itu hanya valid kalau file model
    # dibuka bersamaan file sumbernya, sedangkan sheet ini SUDAH ada di
    # workbook yang sama sehingga referensi lokal lebih benar & robust).
    # Kalau tidak ada data pembukaan, tetap pakai rentang 1 baris supaya
    # SUMIFS tidak dapat range terbalik (awal>akhir) -- SUMIFS pada
    # rentang tanpa match tetap aman, hasilnya 0.
    _nsa_baris_awal_data = header_row + 1
    _nsa_baris_akhir_data = max(_nsa_baris_awal_data, baris_terakhir_data)

    # ================= 4. GL <tahun> =================
    # [FIX] Sheet ini SEBELUMNYA pakai gaya generik _tulis_header_14sheet +
    # _autofit_14sheet -- diganti total supaya sama persis dengan file
    # model referensi GL.xlsx yang user kirim (lihat konstanta _..._GL di
    # atas untuk rincian analisisnya). Baris ditulis dgn NOMOR EKSPLISIT
    # (bukan ws.append()), pola sama dengan fix sheet COA/Neraca Saldo Awal
    # di atas -- supaya spacer baris 2 benar2 kosong & tidak menggeser
    # posisi header.
    ws = wb.create_sheet(f"GL {tahun}")

    # -- Baris 1: judul, merge A1:S1, fill navy + font putih Carlito 15 --
    cell_judul = ws.cell(row=1, column=1, value=f"GENERAL LEDGER {tahun} \u2013 TRANSAKSI TERINTEGRASI")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=22)
    cell_judul.font = _TITLE_FONT_GL
    cell_judul.fill = _TITLE_FILL_GL
    cell_judul.alignment = _TITLE_ALIGN_GL
    ws.row_dimensions[1].height = 27.95

    # -- Baris 2: spacer kosong (sengaja tidak ditulis apa pun -- struktur
    #    file model referensi: 1=judul, 2=kosong, 3=header, 4+=data) --

    # -- Baris 3: header kolom, fill navy + font putih Carlito 11 bold,
    #    center+wrap, border tipis abu #A6A6A6 (atas+bawah semua kolom,
    #    kiri hanya kolom pertama, kanan hanya kolom terakhir -- pola sama
    #    dgn _border_header_asumsi, dipakai ulang) --
    header_row = 3
    headers_gl = [
        "No", "Journal ID", "Tanggal", "Periode", "No. Dokumen", "Kode Akun",
        "Nama Akun", "Keterangan", "Debit", "Kredit",
        # [BARU] Diminta user -- 2 kolom baru disisipkan tepat setelah
        # Kredit, sebelum Segment. "Opening Balance" = saldo awal tahun
        # akun ybs (rumus SUMIFS ke sheet "Neraca Saldo Awal", lihat
        # opening_balance_formula di loop data di bawah -- otomatis,
        # sama utk semua baris akun yg sama). "Adjustment" SENGAJA
        # dikosongkan (bukan 0) -- sistem belum punya field yang menandai
        # suatu jurnal sbg "jurnal penyesuaian" (lihat diskusi & keputusan
        # user: opsi 3, diisi manual langsung di Excel oleh akuntan).
        "Opening Balance", "Adjustment",
        "Segment",
        "Lawan Transaksi", "Project/Unit", "Invoice/Referensi", "Jatuh Tempo",
        "Sumber", "Disiapkan Oleh", "Status",
        # [BARU] Kolom ke-19, di LUAR 18 kolom file referensi (A-R) --
        # fitur audit nyata (lihat _status_validasi_gl()), dipertahankan
        # tapi ditaruh setelah kolom "Status" supaya 18 kolom pertama tetap
        # identik posisi dgn file referensi.
        "Status Validasi",
        # [BARU] Diminta user -- "Year to Date" ditaruh PALING AKHIR (bukan
        # dikelompokkan dgn Opening Balance/Adjustment). Rumus saldo
        # berjalan per akun s.d. tanggal baris ybs, lihat ytd_formula di
        # loop data di bawah.
        "Year to Date",
    ]
    n_kolom_gl = len(headers_gl)
    for col, h in enumerate(headers_gl, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = _HEADER_FONT_GL
        cell.fill = _HEADER_FILL_GL
        cell.alignment = _HEADER_ALIGN_GL
        cell.border = _border_header_asumsi(1, n_kolom_gl, col)
    ws.row_dimensions[header_row].height = 30

    # -- Baris 4 dst: data, 1 baris per SISI jurnal (debet terpisah dari
    #    kredit). "Nama Akun" ditulis sbg FORMULA VLOOKUP ke sheet "COA" DI
    #    WORKBOOK YANG SAMA (bukan link file eksternal "[1]COA!" seperti
    #    file model lama), font hijau -- sama pola dgn fix Neraca Saldo
    #    Awal di atas, pakai ulang baris_awal_coa/baris_akhir_coa yang
    #    sudah dihitung di situ. "Periode" ditulis sbg FORMULA TEXT lokal
    #    ke kolom Tanggal baris yang sama (bukan Python string statis),
    #    font hitam -- supaya otomatis ikut berubah kalau Tanggal diedit
    #    manual di Excel, sama seperti file model referensi. --
    # [BARU] Rentang baris data "GL <tahun>" SUDAH BISA dihitung PASTI
    # sebelum loop dimulai -- 1 baris jurnal (`jurnal`) selalu menulis
    # PERSIS 2 baris fisik (debet + kredit terpisah, lihat loop
    # sisi/no_akun/... di bawah), jadi total baris = header_row +
    # 2*len(jurnal), TIDAK tergantung urutan penulisan. Dibutuhkan di sini
    # (bukan cuma setelah loop selesai spt _gl_baris_awal_data/
    # _gl_baris_akhir_data punya utk sheet Trial Balance Bulanan) supaya
    # formula Opening Balance & Year to Date tiap baris (ditulis SAAT loop
    # berjalan) bisa mereferensikan RENTANG PENUH data GL -- termasuk
    # baris-baris SETELAH baris ybs yg belum ditulis -- karena SUMIFS
    # Year to Date perlu menyaring berdasarkan TANGGAL <= tanggal baris
    # ybs, bukan berdasarkan posisi baris, jadi baris "masa depan" dalam
    # loop tetap harus ikut masuk rentang rumus.
    _gl_baris_awal_data_diri = header_row + 1
    _gl_baris_akhir_data_diri = (header_row + 2 * len(jurnal)) if jurnal else header_row

    r = header_row
    no_baris = 0
    for idx, j in enumerate(jurnal, 1):
        journal_id = j.get("voucher") or f"JV-{idx:04d}"
        tanggal = _ke_tanggal_asumsi(j.get("tanggal"))
        jatuh_tempo = _ke_tanggal_asumsi(j.get("jatuh_tempo"))
        seg_debet = peta_akun.get(str(j.get("no_akun_debet")), {}).get("segment")
        seg_kredit = peta_akun.get(str(j.get("no_akun_kredit")), {}).get("segment")
        # [FIX - GL 2025] "No. Dokumen" & "Invoice/Referensi" sebelumnya
        # SALAH diisi pakai nomor voucher (j.get("voucher")) -- voucher itu
        # nomor bukti internal jurnal, bukan nomor dokumen sumber (mis.
        # nomor invoice/faktur). Sekarang pakai no_dokumen yang diisi
        # akuntan saat konfirmasi posting (lihat db_client.py). "Project/
        # Unit" & "Jatuh Tempo" sebelumnya hardcode None -- sekarang ambil
        # dari field yang sama.
        no_dok = j.get("no_dokumen")
        status_tampil = "Posted" if j.get("status") == "terposting" else (j.get("status") or "Posted")
        disiapkan_oleh = j.get("diposting_oleh") or "-"
        sumber_label = _label_sumber(j.get("jenis_dokumen"))  # [FIX - GL] label manusiawi, bukan kode mentah

        for sisi, no_akun, jml_debet, jml_kredit, seg in (
            ("debet", j.get("no_akun_debet"), _angka(j.get("jml_debet")), 0, seg_debet),
            ("kredit", j.get("no_akun_kredit"), 0, _angka(j.get("jml_kredit")), seg_kredit),
        ):
            no_baris += 1
            r += 1

            # [BARU] Opening Balance -- saldo awal TAHUN utk akun di kolom F
            # baris ini (F{r} = Kode Akun baris ybs, ditulis di posisi
            # kolom ke-6 nilai_baris di bawah). Rumus SUMIFS ke sheet
            # "Neraca Saldo Awal" (Debit dikurangi Kredit -- kolom H/I,
            # kriteria Kode Akun kolom E -- posisi PERSIS sama dgn yg
            # dipakai formula Trial Balance Bulanan di section 8 bawah),
            # BUKAN nilai Python statis -- supaya kalau "Neraca Saldo
            # Awal" diedit manual di Excel, kolom ini ikut update
            # otomatis. Sama utk semua baris dgn akun yg sama (isinya
            # berulang per baris, bukan cuma di baris pertama akun ybs --
            # supaya tabel tetap "lengkap" per baris kalau di-filter/pivot).
            opening_balance_formula = (
                f"=SUMIFS('Neraca Saldo Awal'!$H${_nsa_baris_awal_data}:$H${_nsa_baris_akhir_data},"
                f"'Neraca Saldo Awal'!$E${_nsa_baris_awal_data}:$E${_nsa_baris_akhir_data},F{r})"
                f"-SUMIFS('Neraca Saldo Awal'!$I${_nsa_baris_awal_data}:$I${_nsa_baris_akhir_data},"
                f"'Neraca Saldo Awal'!$E${_nsa_baris_awal_data}:$E${_nsa_baris_akhir_data},F{r})"
            )

            # [BARU] Year to Date -- saldo berjalan akun (kolom F baris
            # ini) per TANGGAL baris ini (kolom C baris ini, cutoff "<="):
            # Saldo Awal tahun (dari "Neraca Saldo Awal", sama spt formula
            # Opening Balance di atas) + seluruh mutasi Debit dikurangi
            # Kredit di sheet GL INI SENDIRI (kolom I/J) utk akun yg sama
            # s.d. tanggal itu -- pola rumus sama persis dgn "Ending
            # Balance" di sheet Trial Balance Bulanan (section 8), bedanya
            # cutoff-nya tanggal baris ybs (dinamis per baris), bukan
            # akhir bulan tetap. Referensi ke sheet ini sendiri TIDAK
            # perlu ditulis nama sheet (dalam sheet yg sama).
            ytd_formula = (
                f"=SUMIFS('Neraca Saldo Awal'!$H${_nsa_baris_awal_data}:$H${_nsa_baris_akhir_data},"
                f"'Neraca Saldo Awal'!$E${_nsa_baris_awal_data}:$E${_nsa_baris_akhir_data},F{r})"
                f"-SUMIFS('Neraca Saldo Awal'!$I${_nsa_baris_awal_data}:$I${_nsa_baris_akhir_data},"
                f"'Neraca Saldo Awal'!$E${_nsa_baris_awal_data}:$E${_nsa_baris_akhir_data},F{r})"
                f"+SUMIFS($I${_gl_baris_awal_data_diri}:$I${_gl_baris_akhir_data_diri},"
                f"$F${_gl_baris_awal_data_diri}:$F${_gl_baris_akhir_data_diri},F{r},"
                f"$C${_gl_baris_awal_data_diri}:$C${_gl_baris_akhir_data_diri},\"<=\"&C{r})"
                f"-SUMIFS($J${_gl_baris_awal_data_diri}:$J${_gl_baris_akhir_data_diri},"
                f"$F${_gl_baris_awal_data_diri}:$F${_gl_baris_akhir_data_diri},F{r},"
                f"$C${_gl_baris_awal_data_diri}:$C${_gl_baris_akhir_data_diri},\"<=\"&C{r})"
            )

            nilai_baris = [
                no_baris, journal_id, tanggal,
                f"=TEXT(C{r},\"mmm-yyyy\")",
                no_dok, no_akun,
                f"=VLOOKUP(F{r},COA!$A${baris_awal_coa}:$B${baris_akhir_coa},2,FALSE)",
                j.get("keterangan"), jml_debet, jml_kredit,
                opening_balance_formula,
                None,  # [BARU] Adjustment -- sengaja kosong, diisi manual oleh akuntan di Excel (Opsi 3)
                seg,
                j.get("lawan_transaksi"), j.get("project_unit"), no_dok,
                jatuh_tempo, sumber_label, disiapkan_oleh, status_tampil,
                _status_validasi_gl(no_akun),
                ytd_formula,
            ]
            for c, v in enumerate(nilai_baris, 1):
                cell = ws.cell(row=r, column=c, value=v)
                if c == 4:
                    cell.font = _DATA_FONT_PERIODE_GL
                elif c == 7:
                    cell.font = _DATA_FONT_FORMULA_GL
                elif c in (11, 22):  # Opening Balance & Year to Date -- rumus, samakan gaya dgn kolom formula lain
                    cell.font = _DATA_FONT_FORMULA_GL
                else:
                    cell.font = _DATA_FONT_GL
                cell.alignment = _DATA_ALIGN_GL
                if no_baris % 2 == 1:  # [FIX] banding: baris fisik ganjil (ke-1,3,5,..) biru muda
                    cell.fill = _BANDING_FILL_GL
                if c == 3 or c == 17:  # Tanggal & Jatuh Tempo (geser +2 krn 2 kolom baru disisipkan)
                    cell.number_format = _FORMAT_TANGGAL_GL
                elif c in (9, 10, 11, 12, 22):  # Debit, Kredit, Opening Balance, Adjustment, Year to Date
                    cell.number_format = _FORMAT_RUPIAH_GL

    baris_terakhir_data = r
    if not jurnal:
        r += 1
        ws.cell(row=r, column=1, value="Belum ada jurnal terposting untuk periode ini.")
        baris_terakhir_data = header_row

    # -- Section "Kontrol GL": 1 baris kosong setelah data terakhir, lalu
    #    header Kontrol GL/Debit/Kredit di kolom H/I/J (posisi kolom
    #    Keterangan/Debit/Kredit -- PERSIS sama dgn file model referensi
    #    H163:J163), baris Total (=SUM formula), baris Selisih
    #    (=Total Debit - Total Kredit formula). --
    baris_header_kontrol = baris_terakhir_data + 2
    for col, h in zip((8, 9, 10), ("Kontrol GL", "Debit", "Kredit")):
        cell = ws.cell(row=baris_header_kontrol, column=col, value=h)
        cell.font = _HEADER_FONT_GL
        cell.fill = _HEADER_FILL_GL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    r_total = baris_header_kontrol + 1
    ws.cell(row=r_total, column=8, value="Total")
    if jurnal:
        ws.cell(row=r_total, column=9,
                 value=f"=SUM(I{header_row + 1}:I{baris_terakhir_data})").number_format = _FORMAT_RUPIAH_GL
        ws.cell(row=r_total, column=10,
                 value=f"=SUM(J{header_row + 1}:J{baris_terakhir_data})").number_format = _FORMAT_RUPIAH_GL
    else:
        ws.cell(row=r_total, column=9, value=0).number_format = _FORMAT_RUPIAH_GL
        ws.cell(row=r_total, column=10, value=0).number_format = _FORMAT_RUPIAH_GL

    r_selisih = r_total + 1
    ws.cell(row=r_selisih, column=8, value="Selisih")
    ws.cell(row=r_selisih, column=9, value=f"=I{r_total}-J{r_total}").number_format = _FORMAT_RUPIAH_GL

    # -- Lebar kolom OTOMATIS menyesuaikan panjang teks header (bukan
    #    lebar manual/patokan file model referensi lagi -- lihat
    #    _lebar_kolom_dari_header) --
    for kolom, lebar in _lebar_kolom_dari_isi(ws).items():
        ws.column_dimensions[kolom].width = lebar

    # [FIX Sheet 8] Simpan nama sheet (dinamis ikut variabel `tahun`) +
    # rentang baris data "GL <tahun>" (kolom C=Tanggal, F=Kode Akun,
    # I=Debit, J=Kredit -- PERSIS sama posisi dgn yang ditulis di section
    # ini) supaya sheet "Trial Balance Bulanan" di bawah bisa
    # mereferensikannya lewat formula SUMIFS LOKAL. Sama alasan spt
    # _nsa_baris_awal_data/_nsa_baris_akhir_data di section "3." di atas.
    _gl_sheet_name = f"GL {tahun}"
    _gl_baris_awal_data = header_row + 1
    _gl_baris_akhir_data = max(_gl_baris_awal_data, baris_terakhir_data)

    # ================= 5. Buku Bantu Piutang =================
    # [FIX] Sheet ini SEBELUMNYA pakai gaya generik _tulis_header_14sheet
    # (fill biru muda #D9E1F2, tanpa font putih/Carlito, tanpa border, tanpa
    # merge judul) + _autofit_14sheet (lebar kolom auto) -- diganti total
    # supaya sama persis dengan file model referensi
    # BUKU_BANTU_PIUTANG_USAHA_2025.xlsx yang user kirim (lihat konstanta
    # _..._PIUTANG di atas untuk rincian analisisnya). Baris ditulis dgn
    # NOMOR EKSPLISIT (bukan ws.append()), pola sama dengan fix sheet
    # COA/Neraca Saldo Awal/GL <tahun> di atas -- supaya spacer baris 2
    # benar2 kosong & tidak menggeser posisi header/data.
    ws = wb.create_sheet("Buku Bantu Piutang")

    # -- Baris 1: judul, merge A1:L1, fill navy + font putih Carlito 15 --
    cell_judul = ws.cell(row=1, column=1, value="BUKU BANTU PIUTANG USAHA " + tahun)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    cell_judul.font = _TITLE_FONT_PIUTANG
    cell_judul.fill = _TITLE_FILL_PIUTANG
    cell_judul.alignment = _TITLE_ALIGN_PIUTANG
    ws.row_dimensions[1].height = 27.95

    # -- Baris 2: spacer kosong (sengaja tidak ditulis apa pun -- struktur
    #    file model referensi: 1=judul, 2=kosong, 3=header, 4+=data) --

    # -- Baris 3: header kolom, fill navy + font putih Carlito 11 bold,
    #    center+wrap, border tipis semua sisi --
    header_row = 3
    header_piutang = ["No. Invoice", "Tanggal Invoice", "Pelanggan", "Segment",
                       "Project/Unit", "Jatuh Tempo", "Nilai Invoice", "Penerimaan",
                       "Saldo Piutang", "Hari Tertunggak", "Bucket Aging", "Status"]
    n_kolom_piutang = len(header_piutang)
    for col, h in enumerate(header_piutang, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = _HEADER_FONT_PIUTANG
        cell.fill = _HEADER_FILL_PIUTANG
        cell.alignment = _HEADER_ALIGN_PIUTANG
        if col == 1:
            cell.border = _HEADER_BORDER_KIRI_PIUTANG
        elif col == n_kolom_piutang:
            cell.border = _HEADER_BORDER_KANAN_PIUTANG
        else:
            cell.border = _HEADER_BORDER_TENGAH_PIUTANG
    ws.row_dimensions[header_row].height = 30

    # [FIX] Dicocokkan langsung ke field asli hasil parse_sheet_piutang() /
    # _perkaya_piutang_per_baris() di akuntansi_ai.py (bukan tebakan lagi):
    #   - "No. Invoice" -> field asli "no_transaksi" (bukan "no_invoice").
    #   - "Nilai Invoice" -> field asli "total_akhir"/"sub_total" (bukan
    #     "total"/"jumlah" yang memang tidak pernah ada di df piutang).
    #   - "Jatuh Tempo" -> "tanggal_jatuh_tempo" sekarang BENAR ada (sebelumnya
    #     kolom ini tidak pernah di-parse sama sekali dari sheet sumber).
    #   - "Penerimaan" -> field baru "penerimaan_total" (bayar_tunai +
    #     bayar_kredit, dihitung di _perkaya_piutang_per_baris()).
    #   - "Saldo Piutang"/"Hari Tertunggak"/"Bucket Aging"/"Status" -> field
    #     baru sisa_piutang_hitung/umur_hari/bucket_aging/status, sekarang
    #     BENAR ditempel per baris (sebelumnya cuma agregat, tidak per baris,
    #     jadi selalu kosong walau alias-nya sudah "benar").
    ALIAS_PIUTANG = {
        "No. Invoice": ["no_transaksi", "no_invoice", "nomor_invoice", "No. Invoice"],
        "Tanggal Invoice": ["tanggal_invoice", "tanggal", "Tanggal Invoice"],
        "Pelanggan": ["nama_pelanggan", "pelanggan", "customer", "Pelanggan"],
        "Segment": ["segment", "Segment"],
        "Project/Unit": ["project_unit", "project", "unit", "Project/Unit"],
        "Jatuh Tempo": ["tanggal_jatuh_tempo", "jatuh_tempo", "Jatuh Tempo"],
        "Nilai Invoice": ["total_akhir", "sub_total", "nilai_invoice", "total", "jumlah", "Nilai Invoice"],
        "Penerimaan": ["penerimaan_total", "penerimaan", "sudah_dibayar", "Penerimaan"],
        "Saldo Piutang": ["sisa_piutang_hitung", "sisa_piutang", "sisa_piutang_tertulis",
                           "saldo_piutang", "Saldo Piutang"],
        "Hari Tertunggak": ["umur_hari", "hari_tertunggak", "Hari Tertunggak"],
        "Bucket Aging": ["bucket_aging", "Bucket Aging"],
        "Status": ["status", "Status"],
    }
    # [BARU] Default 0.0 di LUAR blok if -- dipakai sheet "Ringkasan" (KPI
    # "Piutang Usaha") walau df_piutang kosong, supaya tidak NameError.
    total_saldo_piutang = 0.0
    r = header_row
    if df_piutang:
        total_nilai_invoice = total_penerimaan = total_saldo_piutang = 0.0
        for baris in df_piutang:
            nilai = [_ambil_nilai_alias(baris, ALIAS_PIUTANG[h]) for h in header_piutang]
            # [FIX] "Tanggal Invoice" (kolom B) & "Jatuh Tempo" (kolom F)
            # sekarang dikonversi lewat _ke_tanggal_asumsi() -- sebelumnya
            # ditulis apa adanya (kalau sumbernya string ISO, Excel menulis
            # sbg teks biasa, rata kiri, tidak kena format tanggal). Sama
            # pola dgn fix Neraca Saldo Awal/GL <tahun> di atas.
            nilai[1] = _ke_tanggal_asumsi(nilai[1])
            nilai[5] = _ke_tanggal_asumsi(nilai[5])
            r += 1
            for c, v in enumerate(nilai, 1):
                cell = ws.cell(row=r, column=c, value=v)
                # Kolom A-F (No. Invoice s.d. Jatuh Tempo) = data mentah,
                # biru #0000FF. Kolom G-L (Nilai Invoice s.d. Status) =
                # hasil hitungan/turunan, hijau #008000 -- sama persis pola
                # warna "hardcode vs formula/hitungan" di sheet lain.
                cell.font = _DATA_FONT_HARDCODE_PIUTANG if c <= 6 else _DATA_FONT_HITUNG_PIUTANG
                cell.alignment = _DATA_ALIGN_PIUTANG
                if c == 2 or c == 6:
                    cell.number_format = _FORMAT_TANGGAL_PIUTANG
                elif c in _KOLOM_RUPIAH_PIUTANG:
                    cell.number_format = _FORMAT_RUPIAH_PIUTANG
            total_nilai_invoice += _angka(nilai[6])   # Nilai Invoice
            total_penerimaan += _angka(nilai[7])       # Penerimaan
            total_saldo_piutang += _angka(nilai[8])     # Saldo Piutang

        # [BARU] Baris TOTAL di bawah tabel, seperti pada workbook model
        # referensi -- sebelumnya sheet ini berhenti setelah baris data,
        # tidak ada rekap total Nilai Invoice/Penerimaan/Saldo Piutang.
        # 1 baris kosong dulu (sengaja tidak ditulis apa pun) sebelum TOTAL,
        # persis struktur baris 18 (kosong) -> 19 (TOTAL) di file referensi.
        r_total = r + 2
        for c in range(1, 13):
            cell = ws.cell(row=r_total, column=c)
            cell.font = _TOTAL_FONT_PIUTANG
            cell.fill = _TOTAL_FILL_PIUTANG
            cell.border = _TOTAL_BORDER_PIUTANG
        ws.cell(row=r_total, column=1, value="TOTAL")
        # [FIX] Format Rupiah di baris TOTAL DIHAPUS -- dicek ulang ke file
        # referensi, cell G19/H19/I19 di sana formatnya "General" polos
        # (bukan "#,##0..."), beda dengan baris data di atasnya yang memang
        # pakai format Rupiah. Sengaja disamakan persis, walau sekilas
        # tampak tidak konsisten -- itu memang begitu di file model aslinya.
        ws.cell(row=r_total, column=7, value=round(total_nilai_invoice, 2))
        ws.cell(row=r_total, column=8, value=round(total_penerimaan, 2))
        ws.cell(row=r_total, column=9, value=round(total_saldo_piutang, 2))
    else:
        ws.cell(row=r + 1, column=1,
                value="Belum ada data Buku Bantu Piutang -- upload dokumen AR terlebih dahulu.")

    # -- Lebar kolom OTOMATIS menyesuaikan panjang teks header (bukan
    #    lebar manual/patokan file model referensi lagi) --
    for kolom, lebar in _lebar_kolom_dari_isi(ws).items():
        ws.column_dimensions[kolom].width = lebar

    # ================= 6. Buku Bantu Hutang =================
    # [FIX MENYELURUH] Sheet ini sebelumnya cuma pakai helper generik
    # (_tulis_header_14sheet/_autofit_14sheet -- font default, fill biru
    # muda D9E1F2, tanpa border/merge/warna font/format tanggal, baris
    # TOTAL tanpa style sama sekali). Sekarang disamakan PERSIS dengan file
    # referensi BUKU_BANTU_HUTANG_USAHA_2025.xlsx & pola sheet 5 (Buku
    # Bantu Piutang) yang sudah benar -- reuse konstanta style _PIUTANG
    # karena font/warna/format-nya memang identik, cuma beda nama sheet.
    ws = wb.create_sheet("Buku Bantu Hutang")

    # -- Baris 1: judul, merge A1:K1 (11 kolom), fill navy + font putih
    #    Carlito 15, rata kiri-tengah, tinggi baris 19.5 --
    cell_judul = ws.cell(row=1, column=1, value="BUKU BANTU HUTANG USAHA " + tahun)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    cell_judul.font = _TITLE_FONT_PIUTANG
    cell_judul.fill = _TITLE_FILL_PIUTANG
    cell_judul.alignment = _TITLE_ALIGN_PIUTANG
    ws.row_dimensions[1].height = 27.95

    # -- Baris 2: spacer kosong (struktur file referensi: 1=judul,
    #    2=kosong, 3=header, 4+=data) --

    # -- Baris 3: header kolom, fill navy + font putih Carlito 11 bold,
    #    center+wrap, border tipis semua sisi --
    header_row = 3
    header_hutang = ["No. Tagihan", "Tanggal Tagihan", "Vendor", "Segment",
                      "Project/Unit", "Jatuh Tempo", "Nilai Tagihan", "Pembayaran",
                      "Saldo Hutang", "Hari Tertunggak", "Status"]
    # [FIX] Border header sebelumnya pakai _HEADER_BORDER_PIUTANG (thin di
    # 4 sisi) utk SETIAP sel -- itu bikin garis vertikal ekstra antar tiap
    # kolom yang TIDAK ADA di file referensi. File referensi cuma punya
    # garis tipis di pinggir luar tabel header (kolom A dapat border kiri,
    # kolom terakhir/K dapat border kanan, semua kolom dapat atas+bawah --
    # tanpa garis pemisah vertikal di tengah antar kolom).
    n_kolom_hutang = len(header_hutang)
    for col, h in enumerate(header_hutang, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = _HEADER_FONT_PIUTANG
        cell.fill = _HEADER_FILL_PIUTANG
        cell.alignment = _HEADER_ALIGN_PIUTANG
        if col == 1:
            cell.border = _HEADER_BORDER_KIRI_HUTANG
        elif col == n_kolom_hutang:
            cell.border = _HEADER_BORDER_KANAN_HUTANG
        else:
            cell.border = _HEADER_BORDER_TENGAH_HUTANG
    ws.row_dimensions[header_row].height = 30

    # [FIX] Dicocokkan ke field asli hasil parse_sheet_ap_aging()/
    # proses_ap_aging() di akuntansi_ai.py:
    #   - "Tanggal Tagihan" -> field asli "tanggal_invoice" (bukan
    #     "tanggal_tagihan"/"tanggal" yang tidak pernah ada).
    #   - "Nilai Tagihan" -> field asli "jumlah_utang" (bukan "nilai_tagihan"/
    #     "total"/"jumlah").
    #   - "Pembayaran" -> field asli "jumlah_dibayar" (bukan "pembayaran"/
    #     "sudah_dibayar").
    # No. Tagihan, Vendor, Jatuh Tempo, Saldo Hutang, Hari Tertunggak, Status
    # sudah cocok dari awal, tidak diubah.
    ALIAS_HUTANG = {
        "No. Tagihan": ["nomor_invoice", "no_tagihan", "No. Tagihan"],
        "Tanggal Tagihan": ["tanggal_invoice", "tanggal_tagihan", "tanggal", "Tanggal Tagihan"],
        "Vendor": ["nama_supplier", "vendor", "Vendor"],
        "Segment": ["segment", "Segment"],
        "Project/Unit": ["project_unit", "project", "unit", "Project/Unit"],
        "Jatuh Tempo": ["tanggal_jatuh_tempo", "jatuh_tempo", "Jatuh Tempo"],
        "Nilai Tagihan": ["jumlah_utang", "nilai_tagihan", "total", "jumlah", "Nilai Tagihan"],
        "Pembayaran": ["jumlah_dibayar", "pembayaran", "sudah_dibayar", "Pembayaran"],
        "Saldo Hutang": ["sisa_utang_tertulis", "sisa_utang_hitung", "saldo_hutang", "Saldo Hutang"],
        "Hari Tertunggak": ["umur_hari", "hari_tertunggak", "Hari Tertunggak"],
        # [FIX] Sebelumnya alias ini ("status"/"Status") kepentok field
        # status HASIL DATA-QUALITY CHECK ("PERLU REVIEW"/"OK") dari
        # proses_ap_aging() -- makna field itu beda total dari kolom
        # "Status" di sini yang seharusnya status PELUNASAN per invoice
        # (Lunas/Outstanding, sesuai kolom K di file model referensi).
        # Akibatnya SEMUA baris export tertulis "PERLU REVIEW". Sekarang
        # diprioritaskan ke field baru "status_pelunasan" (lihat
        # akuntansi_ai.py::proses_ap_aging()); alias lama dipertahankan
        # di urutan terakhir cuma sbg fallback kalau df_hutang datang dari
        # sumber lama yang belum punya field ini.
        "Status": ["status_pelunasan", "status", "Status"],
    }
    # [BARU] Default 0.0 di LUAR blok if -- dipakai sheet "Ringkasan" (KPI
    # "Hutang Usaha") walau df_hutang kosong, supaya tidak NameError.
    total_saldo_hutang = 0.0
    r = header_row
    if df_hutang:
        total_nilai_tagihan = total_pembayaran = total_saldo_hutang = 0.0
        for baris in df_hutang:
            nilai = [_ambil_nilai_alias(baris, ALIAS_HUTANG[h]) for h in header_hutang]
            # [FIX] "Tanggal Tagihan" (kolom B) & "Jatuh Tempo" (kolom F)
            # sekarang dikonversi lewat _ke_tanggal_asumsi() -- sebelumnya
            # ditulis apa adanya lewat ws.append() (kalau sumbernya string
            # ISO, Excel menulis sbg teks biasa rata-kiri, tidak kena
            # format tanggal dd-mmm-yyyy). Sama pola dgn fix sheet Piutang.
            nilai[1] = _ke_tanggal_asumsi(nilai[1])
            nilai[5] = _ke_tanggal_asumsi(nilai[5])
            r += 1
            for c, v in enumerate(nilai, 1):
                cell = ws.cell(row=r, column=c, value=v)
                # Kolom A-F (No. Tagihan s.d. Jatuh Tempo) = data mentah,
                # biru #0000FF. Kolom G-K (Nilai Tagihan s.d. Status) =
                # hasil hitungan/turunan, hijau #008000 -- [FIX] sebelumnya
                # TIDAK ada pewarnaan font sama sekali (ikut default hitam
                # bawaan ws.append()).
                cell.font = _DATA_FONT_HARDCODE_PIUTANG if c <= 6 else _DATA_FONT_HITUNG_PIUTANG
                cell.alignment = _DATA_ALIGN_PIUTANG
                if c == 2 or c == 6:
                    cell.number_format = _FORMAT_TANGGAL_PIUTANG
                elif c in _KOLOM_RUPIAH_PIUTANG:
                    cell.number_format = _FORMAT_RUPIAH_PIUTANG
            total_nilai_tagihan += _angka(nilai[6])  # Nilai Tagihan
            total_pembayaran += _angka(nilai[7])      # Pembayaran
            total_saldo_hutang += _angka(nilai[8])     # Saldo Hutang

        # [FIX] Baris TOTAL sebelumnya ditulis polos lewat ws.append() --
        # tanpa font/fill/border. Sekarang disamakan persis dgn file
        # referensi & sheet Piutang: bold hitam, fill #D9E1F2, border atas
        # tipis + bawah DOUBLE di SELURUH kolom A-K (bukan cuma kolom yang
        # berisi angka). 1 baris kosong dulu sebelum TOTAL (persis struktur
        # file referensi: baris kosong -> TOTAL).
        r_total = r + 2
        for c in range(1, 12):
            cell = ws.cell(row=r_total, column=c)
            cell.font = _TOTAL_FONT_PIUTANG
            cell.fill = _TOTAL_FILL_PIUTANG
            cell.border = _TOTAL_BORDER_PIUTANG
            cell.alignment = _DATA_ALIGN_PIUTANG  # vertical=center, sama seperti file referensi
        ws.cell(row=r_total, column=1, value="TOTAL")
        ws.cell(row=r_total, column=7, value=round(total_nilai_tagihan, 2)).number_format = _FORMAT_RUPIAH_PIUTANG
        ws.cell(row=r_total, column=8, value=round(total_pembayaran, 2)).number_format = _FORMAT_RUPIAH_PIUTANG
        ws.cell(row=r_total, column=9, value=round(total_saldo_hutang, 2)).number_format = _FORMAT_RUPIAH_PIUTANG
    else:
        ws.cell(row=r + 1, column=1,
                value="Belum ada data Buku Bantu Hutang -- upload dokumen AP Aging terlebih dahulu.")

    # -- Lebar kolom OTOMATIS menyesuaikan panjang teks header (bukan
    #    lebar manual/patokan file model referensi lagi) --
    for kolom, lebar in _lebar_kolom_dari_isi(ws).items():
        ws.column_dimensions[kolom].width = lebar

    # ================= 7. Buku Bantu Aktiva Tetap =================
    # [FIX MENYELURUH] Sheet ini sebelumnya pakai gaya generik
    # (_tulis_header_14sheet: fill biru muda D9E1F2, font default hitam,
    # tanpa border/merge judul + _autofit_14sheet: lebar kolom auto) dan
    # menulis NILAI HASIL HITUNGAN PYTHON (bukan rumus Excel). Diganti
    # total supaya sama persis dengan file model referensi
    # BUKU_BANTU_AKTIVA_TETAP___JADWAL_PENYUSUTAN_2025.xlsx yang user
    # kirim -- termasuk RUMUS-nya (bukan cuma tampilannya):
    #   L (Penyusutan/Bulan)   = (H-I)/(J*12)
    #   N..Y (Jan..Des)        = IF(bulan$2 >= $G_baris, $L_baris, 0)
    #   Z (Penyusutan <tahun>) = SUM(N:Y) per baris
    #   AA (Akum. Penyusutan Akhir) = M + Z
    #   AB (Nilai Buku 31/12)  = H - AA
    #   Baris TOTAL kolom H, L, M, N..Y, Z, AA, AB = SUM(...) per kolom.
    # [CATATAN PENTING] Rumus di atas otomatis membebankan penyusutan
    # MULAI BULAN "Mulai Digunakan" itu sendiri (karena akhir bulan >=
    # tanggal mulai selalu TRUE untuk bulan yang sama) -- BUKAN mulai
    # bulan berikutnya. Ini beda dengan asumsi lama di
    # laporan_keuangan.susun_jadwal_penyusutan_bulanan() (yang menge-nol-
    # kan bulan perolehan/mulai dipakai) -- makanya kolom N-AB di sini
    # SENGAJA dihitung ulang lewat rumus Excel langsung dari kolom
    # dasar (H, I, J, G), bukan lagi memakai jadwal_bulanan Python, supaya
    # persis mengikuti perilaku file referensi.
    ws = wb.create_sheet("Buku Bantu Aktiva Tetap")

    # -- Baris 1: judul, merge A1:AB1, fill navy + font putih Carlito 15 --
    cell_judul = ws.cell(row=1, column=1, value="BUKU BANTU AKTIVA TETAP & JADWAL PENYUSUTAN " + tahun)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=28)
    cell_judul.font = _TITLE_FONT_ASET
    cell_judul.fill = _TITLE_FILL_ASET
    cell_judul.alignment = _TITLE_ALIGN_ASET
    ws.row_dimensions[1].height = 27.95

    # -- Baris 2: HANYA kolom N-Y diisi tanggal akhir tiap bulan, font
    #    biru, format dd-mmm-yyyy -- kolom lain dibiarkan kosong --
    try:
        tahun_int = int(tahun)
    except (TypeError, ValueError):
        tahun_int = date.today().year
    row_tanggal = 2
    for i, m in enumerate(range(1, 13)):
        col = 14 + i  # N=14 .. Y=25
        c = ws.cell(row=row_tanggal, column=col,
                    value=date(tahun_int, m, calendar.monthrange(tahun_int, m)[1]))
        c.font = _TANGGAL_BULAN_FONT_ASET
        c.alignment = _TANGGAL_BULAN_ALIGN_ASET
        c.number_format = _FORMAT_TANGGAL_ASET

    # -- Baris 3: header kolom, fill navy + font putih Carlito 11 bold,
    #    center+wrap, border tipis (kiri hanya A, kanan hanya AB) --
    header_row = 3
    header_aset = ["Asset ID", "Nama Aset", "Kode Akun Aset", "Kode Akum. Penyusutan",
                   "Kode Beban Penyusutan", "Tanggal Perolehan", "Mulai Digunakan",
                   "Harga Perolehan", "Nilai Residu", "Umur (Tahun)", "Metode",
                   "Penyusutan/Bulan", "Akum. Penyusutan Awal"] + \
                  [f"{b}-{tahun[-2:]}" for b in _BULAN_SINGKAT] + \
                  [f"Penyusutan {tahun}", "Akum. Penyusutan Akhir", f"Nilai Buku 31/12/{tahun}"]
    n_kolom_aset = len(header_aset)
    for col, h in enumerate(header_aset, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = _HEADER_FONT_ASET
        cell.fill = _HEADER_FILL_ASET
        cell.alignment = _HEADER_ALIGN_ASET
        if col == 1:
            cell.border = _BORDER_KIRI_ASET
        elif col == n_kolom_aset:
            cell.border = _BORDER_KANAN_ASET
        else:
            cell.border = _BORDER_TENGAH_ASET
    ws.row_dimensions[header_row].height = 30

    daftar_aset = jadwal_aset.get("aset") or []
    metode_default = jadwal_aset.get("metode", "Garis Lurus")
    # [BARU] Default 0.0 di LUAR blok if -- dipakai sheet "Ringkasan" (KPI
    # "Nilai Buku Aset Tetap") walau daftar_aset kosong, supaya tidak
    # NameError. Dihitung ulang di Python (bukan dibaca dari sel) semata
    # untuk KPI ringkasan -- sel Excel-nya sendiri tetap pakai rumus.
    total_nilai_buku = 0.0
    r = header_row
    if daftar_aset:
        for aset in daftar_aset:
            r += 1
            harga_perolehan = _angka(aset.get("harga_perolehan", 0))
            nilai_residu = _angka(aset.get("nilai_residu", 0))
            # [FIX -- BUG NYATA, konsisten dgn field lain di baris ini yang
            # sudah pakai _angka()] Sebelumnya `aset.get("umur_tahun") or 0`
            # -- field2 lain di sekitarnya (harga_perolehan/nilai_residu/
            # akumulasi_awal) SUDAH dilindungi _angka(), tapi umur_tahun
            # kelewat. Kalau NaN (bukan None), pola lama `NaN or 0` tetap
            # NaN -- ditulis ke kolom J, lalu formula Excel
            # "=(H{r}-I{r})/(J{r}*12)" akan #DIV/0!/error karena membagi
            # dgn NaN. Disamakan pola dgn field lain di baris ini.
            umur_tahun = _angka(aset.get("umur_tahun"))
            akumulasi_awal = _angka(aset.get("akumulasi_awal_tahun", 0))
            metode = aset.get("metode") or metode_default
            # [FIX] Field internal "metode" berisi "komersial"/"fiskal"
            # (dari main.py: req.metode_penyusutan, default "komersial" --
            # lihat laporan_keuangan.susun_jadwal_penyusutan_bulanan()),
            # BUKAN label yang ditampilkan file referensi ("Garis Lurus").
            # Rumus L=(H-I)/(J*12) di sheet ini SELALU basis garis lurus
            # komersial (harga - residu), berapa pun nilai field "metode"
            # -- jadi label kolom K dipetakan ke istilah yang sama dgn
            # referensi, tanpa mengubah field "metode" aslinya di tempat
            # lain (dipakai basis pilih kolom fiskal vs komersial).
            metode_label = {
                "komersial": "Garis Lurus",
                "fiskal": "Garis Lurus Fiskal (PMK 96/2009)",
            }.get(str(metode).lower(), metode)

            nilai_baris = [
                aset.get("kode_aset"), aset.get("nama_aset"),
                aset.get("kode_akun_aset", aset.get("kategori")),
                aset.get("kode_akum_penyusutan"), aset.get("kode_beban_penyusutan"),
                # [FIX] Konversi ke objek tanggal asli (sama pola dgn sheet
                # Piutang/Hutang) -- kalau ditulis apa adanya sbg string,
                # Excel menyimpannya sbg teks biasa (rata kiri, tidak kena
                # format dd-mmm-yyyy), beda dari file referensi.
                _ke_tanggal_asumsi(aset.get("tanggal_perolehan")),
                _ke_tanggal_asumsi(aset.get("mulai_digunakan")),
                harga_perolehan, nilai_residu, umur_tahun, metode_label,
                f"=(H{r}-I{r})/(J{r}*12)",       # L: Penyusutan/Bulan
                akumulasi_awal,                    # M: Akum. Penyusutan Awal
            ] + [
                f"=IF({_KOLOM_HURUF_ASET[13 + i]}$2>=$G{r},$L{r},0)"  # N..Y
                for i in range(12)
            ] + [
                f"=SUM(N{r}:Y{r})",                # Z: Penyusutan tahun berjalan
                f"=M{r}+Z{r}",                      # AA: Akum. Penyusutan Akhir
                f"=H{r}-AA{r}",                     # AB: Nilai Buku 31/12
            ]
            for c, v in enumerate(nilai_baris, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font = _DATA_FONT_INPUT_ASET if c <= 11 else _DATA_FONT_CALC_ASET
                if c == 1:
                    cell.border = _DATA_BORDER_KIRI_ASET
                elif c == n_kolom_aset:
                    cell.border = _DATA_BORDER_KANAN_ASET
                if c in (6, 7):
                    cell.number_format = _FORMAT_TANGGAL_ASET
                elif c in (8, 9, 12, 13) or c >= 14:
                    cell.number_format = _FORMAT_RUPIAH_ASET

            # Total nilai buku dihitung ulang murni di Python (nilai
            # akhir setelah penyusutan tahun berjalan) -- dipakai sheet
            # "Ringkasan", tidak memengaruhi rumus di sel Excel.
            dasar = max(harga_perolehan - nilai_residu, 0)
            penyusutan_bulan_py = (dasar / (umur_tahun * 12)) if umur_tahun else 0.0
            tanggal_mulai = aset.get("mulai_digunakan") or aset.get("tanggal_perolehan")
            tgl_parsed = _tanggal_untuk_jadwal_aset(tanggal_mulai)
            penyusutan_tahun_py = 0.0
            for m in range(1, 13):
                akhir_bulan = date(tahun_int, m, calendar.monthrange(tahun_int, m)[1])
                # Sama seperti rumus Excel IF(akhir_bulan>=$G,$L,0): kalau
                # tanggal mulai tidak diketahui atau di tahun sebelumnya,
                # anggap G < awal tahun ini -> semua bulan kena penyusutan.
                if (tgl_parsed is None) or (akhir_bulan >= tgl_parsed):
                    penyusutan_tahun_py += penyusutan_bulan_py
            akum_akhir_py = akumulasi_awal + penyusutan_tahun_py
            total_nilai_buku += harga_perolehan - akum_akhir_py

        # -- Baris TOTAL: bold hitam, fill biru muda, border tipis
        #    atas+bawah semua kolom (kiri hanya A, kanan hanya AB) --
        #    1 baris kosong dulu (sengaja tidak ditulis apa pun) sebelum
        #    TOTAL, persis struktur baris 7 (kosong) -> 8 (TOTAL) di file
        #    referensi.
        r_total = r + 2
        baris_terakhir_data = r
        for c in range(1, n_kolom_aset + 1):
            cell = ws.cell(row=r_total, column=c)
            cell.font = _TOTAL_FONT_ASET
            cell.fill = _TOTAL_FILL_ASET
            if c == 1:
                cell.border = _BORDER_KIRI_ASET
            elif c == n_kolom_aset:
                cell.border = _BORDER_KANAN_ASET
            else:
                cell.border = _BORDER_TENGAH_ASET
        ws.cell(row=r_total, column=1, value="TOTAL")
        for kolom_huruf, col_idx in [("H", 8), ("L", 12), ("M", 13)]:
            c = ws.cell(row=r_total, column=col_idx,
                        value=f"=SUM({kolom_huruf}{header_row + 1}:{kolom_huruf}{baris_terakhir_data})")
            c.number_format = _FORMAT_RUPIAH_ASET
        for i in range(12):
            col_idx = 14 + i
            kh = _KOLOM_HURUF_ASET[col_idx - 1]
            c = ws.cell(row=r_total, column=col_idx,
                        value=f"=SUM({kh}{header_row + 1}:{kh}{baris_terakhir_data})")
            c.number_format = _FORMAT_RUPIAH_ASET
        for kolom_huruf, col_idx in [("Z", 26), ("AA", 27), ("AB", 28)]:
            c = ws.cell(row=r_total, column=col_idx,
                        value=f"=SUM({kolom_huruf}{header_row + 1}:{kolom_huruf}{baris_terakhir_data})")
            c.number_format = _FORMAT_RUPIAH_ASET
    else:
        ws.cell(row=r + 1, column=1,
                value="Belum ada data Aset Tetap -- upload Daftar Aset Tetap terlebih dahulu.")

    # -- Lebar kolom OTOMATIS menyesuaikan panjang teks header (bukan
    #    lebar manual/patokan file model referensi lagi) --
    for kolom, lebar in _lebar_kolom_dari_isi(ws).items():
        ws.column_dimensions[kolom].width = lebar

    # ================= 8. Trial Balance Bulanan =================
    # [FIX] Styling disamakan PERSIS dengan file model referensi
    # (TRIAL_BALANCE_BULANAN___JANUARI_S_D__DESEMBER_2025.xlsx):
    #   - Judul di baris 1, merged A1:Z1, bold size 15, center.
    #   - Baris 2 kosong (pemisah), sama seperti referensi.
    #   - Header 2 baris (baris 3 & 4): "Kode Akun"/"Nama Akun" merge
    #     vertikal 2 baris; tiap nama bulan merge horizontal di atas
    #     pasangan kolom Debit/Kredit-nya (mis. C3:D3 = "January 2025").
    #   - Fill dongker (#17365D) + font putih bold utk semua sel header,
    #     bukan lagi biru muda seperti sheet 14-sheet lainnya.
    ws = wb.create_sheet("Trial Balance Bulanan")
    tb = laporan_bulanan.get("trial_balance_bulanan") or {}
    # [BARU] Ditambah kolom "Ending Balance" per bulan -- jadi 3 kolom per
    # bulan (Debit, Kredit, Ending Balance), bukan 2 lagi. Ending Balance
    # dihitung sesuai saldo normal akun (lihat _BORDER_HEADER_TB8_TENGAH
    # & rumus VLOOKUP ke COA di bawah): akun bersaldo normal DEBET
    # (Aset, Beban) -> Ending = Debit - Kredit; akun bersaldo normal
    # KREDIT (Kontra Aset, Liabilitas, Ekuitas, Pendapatan) -> Ending =
    # Kredit - Debit.
    ncols_tb = 2 + 12 * 3  # Kode Akun + Nama Akun + 12 bulan x (Debit, Kredit, Ending Balance)
    # [BARU] Default aman -- diisi ulang dgn kumulatif LABA BERSIH YTD
    # sungguhan di section 9 (Laba Rugi Bulanan) kalau tb ada. Dideklarasi
    # di sini (bukan cuma di dalam blok section 9) supaya section 10
    # (Balance Sheet Bulanan) selalu punya nilai valid utk baris "Laba
    # Tahun Berjalan", termasuk kalau urutan section berubah di masa depan.
    laba_bersih_ytd: List[float] = [0.0] * 12
    # [BARU] Default aman juga utk sheet 11 (Laporan Perubahan Ekuitas) --
    # nomor baris "LABA BERSIH BULANAN" di sheet "Laba Rugi Bulanan"
    # (diisi ulang di section 9 kalau tb ada). None berarti sheet 11 tidak
    # bisa nge-link ke situ (blm ada data Laba Rugi Bulanan) & baris
    # "Laba Tahun Berjalan"-nya jatuh ke 0, bukan error/#REF!.
    r_laba_bersih_bulanan_lr: Optional[int] = None
    # [FIX BUG] Sheet 11 "Laporan Perubahan Ekuitas" sebelumnya nge-link
    # baris "Laba Tahun Berjalan" di TABEL RINGKASAN ATAS (yang jadi
    # komponen TOTAL EKUITAS) ke r_laba_bersih_bulanan_lr (LABA BERSIH
    # BULANAN -- angka SATU BULAN itu saja) -- padahal baris itu HARUS
    # kumulatif (mis. TOTAL EKUITAS bulan Mei = akumulasi laba Jan..Mei,
    # BUKAN cuma laba bulan Mei sendiri), sama seperti section rincian
    # roll-forward "Saldo Akhir Saldo Laba" di sheet yang sama yang SUDAH
    # benar mengakumulasi. Akibatnya utk bulan ke-2 dst, TOTAL EKUITAS di
    # tabel ringkasan atas SELALU understated & tidak pernah sinkron dgn
    # TOTAL SALDO AKHIR MODAL + Saldo Akhir Saldo Laba di bawahnya -- baris
    # "CHECK BALANCE" salah kasih status "PERIKSA" mulai bulan Februari
    # dst walau laporan sebenarnya balance (dibuktikan lewat rekalkulasi
    # LibreOffice). Ditambah nomor baris "LABA BERSIH YTD" (kumulatif,
    # sudah ada TAPI nomor barisnya sebelumnya tidak pernah ditangkap) --
    # dipakai KHUSUS utk baris "Laba Tahun Berjalan" di tabel ringkasan
    # atas; r_laba_bersih_bulanan_lr (bulanan) tetap dipakai apa adanya di
    # section rincian roll-forward (yang sudah benar mengakumulasi
    # sendiri) & di sheet 12 "Laporan Arus Kas" (metode tidak langsung
    # butuh angka SATU BULAN, bukan YTD).
    r_laba_bersih_ytd_lr: Optional[int] = None
    # [BARU] Sama polanya -- nomor baris "TOTAL PENYUSUTAN" di sheet
    # "Laba Rugi Bulanan" (diisi ulang di section 9 kalau tb ada), dipakai
    # sheet 12 "Laporan Arus Kas" utk baris penyesuaian non-kas "Beban
    # Penyusutan". None -> baris itu jatuh ke 0, bukan error/#REF!.
    r_total_penyusutan_lr: Optional[int] = None

    ws.cell(row=1, column=1, value=f"TRIAL BALANCE BULANAN -- JANUARI S.D. DESEMBER {tahun}")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols_tb)
    ws.cell(row=1, column=1).font = _TITLE_FONT_TB8
    ws.cell(row=1, column=1).fill = _TITLE_FILL_TB8
    ws.cell(row=1, column=1).alignment = _TITLE_ALIGN_TB8
    ws.row_dimensions[1].height = 27.95

    if tb:
        header_row1 = 3
        header_row2 = 4
        ws.row_dimensions[header_row1].height = 30
        ws.row_dimensions[header_row2].height = 30

        ws.cell(row=header_row1, column=1, value="Kode Akun")
        ws.merge_cells(start_row=header_row1, start_column=1, end_row=header_row2, end_column=1)
        ws.cell(row=header_row1, column=2, value="Nama Akun")
        ws.merge_cells(start_row=header_row1, start_column=2, end_row=header_row2, end_column=2)

        for m, nama_bulan in enumerate(_BULAN_PANJANG):
            col_debit = 3 + m * 3
            ws.cell(row=header_row1, column=col_debit, value=f"{nama_bulan} {tahun}")
            ws.merge_cells(start_row=header_row1, start_column=col_debit,
                            end_row=header_row1, end_column=col_debit + 2)
            ws.cell(row=header_row2, column=col_debit, value="Debit")
            ws.cell(row=header_row2, column=col_debit + 1, value="Kredit")
            ws.cell(row=header_row2, column=col_debit + 2, value="Ending Balance")

        # [FIX Sheet 8] Border header -- SEBELUMNYA tidak ada border sama
        # sekali. File referensi punya kotak tipis abu di SETIAP sel
        # header, TAPI TANPA garis pemisah di DALAM blok 3-kolom
        # (Debit/Kredit/Ending Balance) yang sama bulannya -- cuma garis
        # di kiri kolom Debit & kanan kolom Ending Balance tiap blok
        # bulan, kolom Kredit di tengah TANPA garis kiri/kanan sama
        # sekali -- kolom 1 (Kode Akun) & kolom 2 (Nama Akun) masing2
        # dapat garis kiri/kanan (batas tabel), tanpa garis antara
        # keduanya, persis pola blok di atas.
        for r in (header_row1, header_row2):
            for c in range(1, ncols_tb + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = _HEADER_FONT_TB8
                cell.fill = _HEADER_FILL_TB_BULANAN
                cell.alignment = _HEADER_ALIGN_TB8
                if c == 1:
                    cell.border = _BORDER_HEADER_TB8_KIRI
                elif c == 2:
                    cell.border = _BORDER_HEADER_TB8_KANAN
                else:
                    posisi_dlm_blok = (c - 3) % 3
                    if posisi_dlm_blok == 0:
                        cell.border = _BORDER_HEADER_TB8_KIRI
                    elif posisi_dlm_blok == 2:
                        cell.border = _BORDER_HEADER_TB8_KANAN
                    else:
                        cell.border = _BORDER_HEADER_TB8_TENGAH

        # [FIX Sheet 8] Baris data: kolom Kode Akun & Nama Akun font biru
        # #0000FF, kolom Debit/Kredit font hijau #008000 -- sebelumnya
        # TIDAK ada styling sama sekali di baris data (ikut default
        # Calibri hitam), jadi tidak sama dengan file referensi.
        #
        # [FIX Sheet 8 - rumus] Kolom Debit/Kredit tiap bulan SEBELUMNYA
        # ditulis sbg VALUE Python statis dari _pecah_per_bulan_debit_kredit()
        # (hasil laporan_bulanan) -- file model referensi menuliskannya
        # sbg RUMUS Excel live: saldo kumulatif akun per akun (saldo awal
        # dari "Neraca Saldo Awal" + mutasi "GL <tahun>" s.d. akhir bulan
        # ybs, cutoff "<="), lalu dipecah MAX(0,saldo) ke Debit &
        # MAX(0,-saldo) ke Kredit -- supaya kalau Neraca Saldo Awal/GL
        # diedit manual di Excel, Trial Balance ikut update otomatis tanpa
        # perlu re-generate dari aplikasi. Sekarang direplikasi persis,
        # HANYA referensinya diarahkan ke sheet "Neraca Saldo Awal"/"GL
        # <tahun>" LOKAL di workbook yang sama (bukan link "[1]..." file
        # eksternal spt file model aslinya -- lihat catatan di
        # _nsa_baris_awal_data/_gl_baris_awal_data di section 3 & 4 di
        # atas).
        baris_data_awal = header_row2 + 1
        for no_akun in sorted(tb.keys()):
            info = tb[no_akun]
            ws.append([no_akun, info.get("nama_akun")])
            r = ws.max_row
            ws.cell(row=r, column=1).font = _DATA_FONT_KODE_NAMA_TB8
            ws.cell(row=r, column=1).alignment = _DATA_ALIGN_TB8
            ws.cell(row=r, column=2).font = _DATA_FONT_KODE_NAMA_TB8
            ws.cell(row=r, column=2).alignment = _DATA_ALIGN_TB8
            for m in range(12):
                bulan_ke = m + 1
                akhir_hari = calendar.monthrange(int(tahun), bulan_ke)[1]
                cutoff = f"DATE({tahun},{bulan_ke},{akhir_hari})"
                saldo_expr = (
                    f"SUMIFS('Neraca Saldo Awal'!$H${_nsa_baris_awal_data}:$H${_nsa_baris_akhir_data},"
                    f"'Neraca Saldo Awal'!$E${_nsa_baris_awal_data}:$E${_nsa_baris_akhir_data},$A{r})"
                    f"-SUMIFS('Neraca Saldo Awal'!$I${_nsa_baris_awal_data}:$I${_nsa_baris_akhir_data},"
                    f"'Neraca Saldo Awal'!$E${_nsa_baris_awal_data}:$E${_nsa_baris_akhir_data},$A{r})"
                    f"+SUMIFS('{_gl_sheet_name}'!$I${_gl_baris_awal_data}:$I${_gl_baris_akhir_data},"
                    f"'{_gl_sheet_name}'!$F${_gl_baris_awal_data}:$F${_gl_baris_akhir_data},$A{r},"
                    f"'{_gl_sheet_name}'!$C${_gl_baris_awal_data}:$C${_gl_baris_akhir_data},\"<=\"&{cutoff})"
                    f"-SUMIFS('{_gl_sheet_name}'!$J${_gl_baris_awal_data}:$J${_gl_baris_akhir_data},"
                    f"'{_gl_sheet_name}'!$F${_gl_baris_awal_data}:$F${_gl_baris_akhir_data},$A{r},"
                    f"'{_gl_sheet_name}'!$C${_gl_baris_awal_data}:$C${_gl_baris_akhir_data},\"<=\"&{cutoff})"
                )
                col_debit = 3 + m * 3
                col_kredit = col_debit + 1
                col_ending = col_debit + 2
                huruf_debit = get_column_letter(col_debit)
                huruf_kredit = get_column_letter(col_kredit)
                ws.cell(row=r, column=col_debit, value=f"=MAX(0,{saldo_expr})")
                ws.cell(row=r, column=col_kredit, value=f"=MAX(0,-({saldo_expr}))")
                # [BARU] Ending Balance -- rumusnya TERGANTUNG saldo normal
                # akun (lihat sheet COA kolom E "Saldo Normal", VLOOKUP ke
                # baris akun ybs): akun bersaldo normal KREDIT (Kontra
                # Aset, Liabilitas, Ekuitas, Pendapatan) -> Ending =
                # Kredit - Debit; SELAIN itu (termasuk kalau Saldo Normal
                # kosong/belum diisi) default DEBET (Aset, Beban) ->
                # Ending = Debit - Kredit. Nilai di kolom COA!E ditulis
                # "DEBET"/"KREDIT" (lihat _sheet_coa_14 di atas), BUKAN
                # "Debit"/"Kredit" -- makanya dibandingkan ke "KREDIT".
                ws.cell(
                    row=r, column=col_ending,
                    value=(
                        f'=IF(VLOOKUP($A{r},COA!$A:$E,5,0)="KREDIT",'
                        f'{huruf_kredit}{r}-{huruf_debit}{r},'
                        f'{huruf_debit}{r}-{huruf_kredit}{r})'
                    ),
                )
            for col in range(3, ncols_tb + 1):
                ws.cell(row=r, column=col).font = _DATA_FONT_ANGKA_TB8
                ws.cell(row=r, column=col).alignment = _DATA_ALIGN_TB8

        # [BARU] Baris TOTAL -- SUM tiap kolom Debit/Kredit dari baris
        # data pertama s.d. terakhir, sesuai file referensi (jadi
        # pengecekan visual: total Debit harus = total Kredit tiap
        # bulan kalau trial balance-nya sehat/balance).
        baris_data_akhir = ws.max_row
        baris_total = baris_data_akhir + 1
        ws.cell(row=baris_total, column=1, value="TOTAL")
        for col in range(3, ncols_tb + 1):
            surat = ws.cell(row=baris_data_awal, column=col).column_letter
            ws.cell(row=baris_total, column=col,
                    value=f"=SUM({surat}{baris_data_awal}:{surat}{baris_data_akhir})")
        # [FIX Sheet 8] Label "TOTAL" hitam, angka SUM hijau (sama seperti
        # kolom Debit/Kredit di atasnya) -- sebelumnya seluruh baris
        # TOTAL dikasih font hitam polos tanpa warna, tidak sama dengan
        # referensi yang angka SUM-nya tetap hijau.
        ws.cell(row=baris_total, column=1).font = _TOTAL_FONT_LABEL_TB8
        ws.cell(row=baris_total, column=1).fill = _TOTAL_FILL_TB_BULANAN
        ws.cell(row=baris_total, column=1).alignment = _DATA_ALIGN_TB8
        # [FIX Sheet 8] Border baris TOTAL -- garis atas tipis + bawah
        # GANDA warna navy #17365D di SEMUA kolom A-Z, sebelumnya tidak
        # ada border sama sekali.
        ws.cell(row=baris_total, column=1).border = _BORDER_TOTAL_TB8
        for col in range(2, ncols_tb + 1):
            cell = ws.cell(row=baris_total, column=col)
            cell.font = _TOTAL_FONT_ANGKA_TB8
            cell.fill = _TOTAL_FILL_TB_BULANAN
            cell.alignment = _DATA_ALIGN_TB8
            cell.border = _BORDER_TOTAL_TB8

        for col in range(3, ncols_tb + 1):
            for r in range(baris_data_awal, baris_total + 1):
                ws.cell(row=r, column=col).number_format = _FORMAT_AKUNTANSI_TB_BULANAN

        # -- Lebar kolom OTOMATIS menyesuaikan panjang teks header (bukan
        #    lebar manual/patokan file model referensi lagi). Kolom
        #    "Kode Akun"/"Nama Akun" dari panjang teksnya sendiri; kolom
        #    Debit/Kredit tiap bulan dihitung SEPASANG supaya total lebar
        #    keduanya juga cukup menampung judul bulan gabungan (mis.
        #    "September 2025") yang di-merge horizontal di row1 -- kalau
        #    cuma dihitung dari "Debit"/"Kredit" sendiri-sendiri, judul
        #    bulan yang lebih panjang bisa terpotong di sel merge-nya.
        _pad_tb8 = 4
        ws.column_dimensions["A"].width = max(6, len("Kode Akun") + _pad_tb8)
        ws.column_dimensions["B"].width = max(6, len("Nama Akun") + _pad_tb8)
        for m, nama_bulan in enumerate(_BULAN_PANJANG):
            col_debit = 3 + m * 3
            col_kredit = col_debit + 1
            col_ending = col_debit + 2
            judul_bulan = f"{nama_bulan} {tahun}"
            lebar_per_kolom = max(
                len("Debit") + _pad_tb8, len("Kredit") + _pad_tb8,
                len("Ending Balance") + _pad_tb8,
                -(-(len(judul_bulan) + _pad_tb8) // 3),  # ceil, dibagi 3 kolom
            )
            ws.column_dimensions[get_column_letter(col_debit)].width = lebar_per_kolom
            ws.column_dimensions[get_column_letter(col_kredit)].width = lebar_per_kolom
            ws.column_dimensions[get_column_letter(col_ending)].width = lebar_per_kolom

        # [FIX] Lebar di atas dihitung dari header/judul bulan SAJA --
        # kolom "Kode Akun"/"Nama Akun" bisa kepotong kalau isi datanya
        # lebih panjang dari nama headernya sendiri (mis. "Nama Akun"
        # diisi "Akumulasi Penyusutan Kendaraan"), sama seperti bug yang
        # dilaporkan user di sheet "COA". Timpa dengan nilai TERBESAR
        # antara lebar manual (yg menjaga judul bulan gabungan di baris
        # merge row1 tidak kepotong) vs lebar dari isi data aktual --
        # supaya kolom tidak pernah lebih sempit dari salah satu
        # kebutuhan itu (bukan REPLACE total, karena lebar manual di atas
        # juga menjaga alokasi Debit/Kredit/Ending Balance per bulan yg
        # merge-nya tidak bisa dihitung otomatis dari isi cell biasa).
        _lebar_isi_tb8 = _lebar_kolom_dari_isi(ws)
        for _kolom, _lebar_isi in _lebar_isi_tb8.items():
            _lebar_sekarang = ws.column_dimensions[_kolom].width or 0
            if _lebar_isi > _lebar_sekarang:
                ws.column_dimensions[_kolom].width = _lebar_isi
    else:
        ws.cell(row=3, column=1,
                 value="Belum ada data Trial Balance Bulanan -- generate laporan bulanan terlebih dahulu.")

    # ================= 9. Laba Rugi Bulanan =================
    # [FIX] Disusun ulang total agar sama persis file model referensi
    # (LAPORAN_LABA_RUGI_BULANAN___2025.xlsx): sebelumnya sheet ini cuma
    # daftar rata semua akun PENDAPATAN/BEBAN + 1 baris total -- sekarang
    # dikelompokkan PERSIS seperti referensi: PENDAPATAN -> BEBAN LANGSUNG
    # -> LABA KOTOR -> BEBAN OPERASIONAL -> EBITDA -> PENYUSUTAN -> LABA
    # USAHA -> Pendapatan/Beban Lain-lain -> LABA BERSIH BULANAN -> LABA
    # BERSIH YTD. Pengelompokan pakai field `sub_kategori` pada COA
    # (KONVENSI, harus diisi akuntan per akun):
    #   - BEBAN + sub_kategori "HPP"        -> BEBAN LANGSUNG
    #     (dipakai sama seperti L02 Lampiran SPT -- lihat
    #     laporan_keuangan.susun_lampiran_spt_pnl())
    #   - BEBAN + sub_kategori "Penyusutan" -> PENYUSUTAN
    #   - BEBAN + sub_kategori "Lain-lain"  -> Beban Lain-lain (baris
    #     berdiri sendiri, TIDAK masuk TOTAL BEBAN OPERASIONAL)
    #   - BEBAN + sub_kategori lain/kosong  -> BEBAN OPERASIONAL (default)
    #   - PENDAPATAN + sub_kategori "Lain-lain" -> Pendapatan Lain-lain
    #     (baris berdiri sendiri, TIDAK masuk TOTAL PENDAPATAN)
    #   - PENDAPATAN + sub_kategori lain/kosong -> PENDAPATAN (default,
    #     pendapatan usaha utama)
    # Kalau kategori COA client belum diisi sub_kategori sama sekali,
    # semua akun BEBAN otomatis jatuh ke BEBAN OPERASIONAL & semua akun
    # PENDAPATAN ke PENDAPATAN -- sheet tetap kebentuk (tidak error),
    # cuma BEBAN LANGSUNG/PENYUSUTAN/Lain-lain akan kosong (0) sampai
    # akuntan melengkapi sub_kategori di form COA.
    ws = wb.create_sheet("Laba Rugi Bulanan")
    ncols_lr = 2 + 12

    try:
        tahun_int_lr = int(tahun)
    except (TypeError, ValueError):
        tahun_int_lr = date.today().year

    ws.cell(row=1, column=1, value=f"LAPORAN LABA RUGI BULANAN -- {tahun}")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols_lr)
    ws.cell(row=1, column=1).font = _TITLE_FONT_LR_BULANAN
    ws.cell(row=1, column=1).fill = _HEADER_FILL_LR_BULANAN
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")

    if tb:
        # Baris 2: tanggal akhir tiap bulan (dd-mmm-yyyy), kolom A/B kosong.
        baris_tanggal = [None, None] + [
            date(tahun_int_lr, m, calendar.monthrange(tahun_int_lr, m)[1]) for m in range(1, 13)
        ]
        ws.append(baris_tanggal)
        for c in range(3, ncols_lr + 1):
            cell_tgl = ws.cell(row=2, column=c)
            cell_tgl.number_format = "dd-mmm-yyyy"
            # [FIX] Baris tanggal di referensi pakai font Carlito biru
            # #0000FF -- sebelumnya tidak diberi font sama sekali (default).
            cell_tgl.font = _KODE_FONT_LR_BULANAN

        # Baris 3: header Kode/Uraian/Jan-25..Dec-25.
        header_row_lr = 3
        headers_lr = ["Kode", "Uraian"] + [f"{b}-{tahun[-2:]}" for b in _BULAN_SINGKAT]
        _tulis_header_14sheet(ws, headers_lr, row=header_row_lr)
        for c in range(1, ncols_lr + 1):
            cell = ws.cell(row=header_row_lr, column=c)
            cell.font = _HEADER_FONT_LR_BULANAN
            cell.fill = _HEADER_FILL_LR_BULANAN
            cell.alignment = _ALIGN_HEADER_LR_BULANAN
            cell.border = _BORDER_HEADER_LR_BULANAN
        ws.row_dimensions[header_row_lr].height = 30

        def _delta_bulanan_lr(no_akun: str) -> List[float]:
            """Pergerakan bulanan (bukan kumulatif YTD) akun ini, lihat
            catatan [FIX - export 14 sheet] versi sebelumnya di atas."""
            info = tb[no_akun]
            kumulatif = [_angka(v) for v in (info.get("per_bulan") or [0] * 12)][:12]
            kumulatif += [0] * (12 - len(kumulatif))
            baseline = _angka(peta_akun.get(str(no_akun), {}).get("saldo_awal"))
            hasil, sebelumnya = [], baseline
            for v in kumulatif:
                hasil.append(round(v - sebelumnya, 2))
                sebelumnya = v
            return hasil

        # --- Klasifikasi tiap akun PENDAPATAN/BEBAN ke kelompok sheet ini ---
        akun_pendapatan, akun_pendapatan_lain = [], []
        akun_beban_langsung, akun_beban_operasional = [], []
        akun_penyusutan, akun_beban_lain = [], []
        for no_akun in sorted(tb.keys()):
            info_coa = peta_akun.get(str(no_akun), {})
            kategori = str(info_coa.get("kategori") or "").strip().lower()
            sub = str(info_coa.get("sub_kategori") or "").strip().lower()
            if kategori == "pendapatan":
                (akun_pendapatan_lain if sub == "lain-lain" else akun_pendapatan).append(no_akun)
            elif kategori == "beban":
                if sub == "hpp":
                    akun_beban_langsung.append(no_akun)
                elif sub == "penyusutan":
                    akun_penyusutan.append(no_akun)
                elif sub == "lain-lain":
                    akun_beban_lain.append(no_akun)
                else:
                    akun_beban_operasional.append(no_akun)
            # kategori ASET/LIABILITAS/EKUITAS diabaikan di sheet Laba Rugi

        def _sum_kelompok(daftar_akun: List[str]):
            total = [0.0] * 12
            per_akun: Dict[str, List[float]] = {}
            for no_akun in daftar_akun:
                nilai = _delta_bulanan_lr(no_akun)
                per_akun[no_akun] = nilai
                for i in range(12):
                    total[i] += nilai[i]
            return total, per_akun

        def _tulis_kategori(label: str) -> None:
            r = ws.max_row + 1
            ws.cell(row=r, column=2, value=label)
            for c in range(1, ncols_lr + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = _KATEGORI_FONT_LR_BULANAN
                cell.fill = _KATEGORI_FILL_LR_BULANAN
                cell.border = _BORDER_KATEGORI_LR_BULANAN

        def _tulis_item(no_akun: str, nilai: List[float], biru: bool = False) -> None:
            # [FIX] `biru` dipertahankan sebagai parameter (dipanggil dari
            # baris Pendapatan/Beban Lain-lain) tapi TIDAK lagi dipakai utk
            # menimpa semua kolom jadi biru -- pola warna file referensi
            # SAMA di semua baris item (kategori normal maupun lain-lain):
            # kolom Kode & Uraian biru, kolom nilai (C-N) hijau.
            nama = peta_akun.get(str(no_akun), {}).get("nama_akun") or tb[no_akun].get("nama_akun")
            ws.append([no_akun, nama] + [round(v, 2) for v in nilai])
            r = ws.max_row
            ws.cell(row=r, column=1).font = _KODE_FONT_LR_BULANAN
            ws.cell(row=r, column=2).font = _KODE_FONT_LR_BULANAN
            for c in range(3, ncols_lr + 1):
                ws.cell(row=r, column=c).font = _VALUE_FONT_LR_BULANAN

        def _tulis_total(label: str, nilai: List[float]) -> None:
            ws.append(["", label] + [round(v, 2) for v in nilai])
            r = ws.max_row
            for c in range(1, ncols_lr + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = _TOTAL_FONT_LR_BULANAN
                cell.fill = _TOTAL_FILL_TB_BULANAN
                cell.border = _BORDER_TOTAL_LR_BULANAN

        # --- PENDAPATAN ---
        total_pendapatan, nilai_pendapatan = _sum_kelompok(akun_pendapatan)
        _tulis_kategori("PENDAPATAN")
        for no_akun in akun_pendapatan:
            _tulis_item(no_akun, nilai_pendapatan[no_akun])
        _tulis_total("TOTAL PENDAPATAN", total_pendapatan)

        # --- BEBAN LANGSUNG ---
        total_beban_langsung, nilai_beban_langsung = _sum_kelompok(akun_beban_langsung)
        _tulis_kategori("BEBAN LANGSUNG")
        for no_akun in akun_beban_langsung:
            _tulis_item(no_akun, nilai_beban_langsung[no_akun])
        _tulis_total("TOTAL BEBAN LANGSUNG", total_beban_langsung)

        laba_kotor = [total_pendapatan[i] - total_beban_langsung[i] for i in range(12)]
        _tulis_total("LABA KOTOR", laba_kotor)

        # --- BEBAN OPERASIONAL ---
        total_beban_operasional, nilai_beban_operasional = _sum_kelompok(akun_beban_operasional)
        _tulis_kategori("BEBAN OPERASIONAL")
        for no_akun in akun_beban_operasional:
            _tulis_item(no_akun, nilai_beban_operasional[no_akun])
        _tulis_total("TOTAL BEBAN OPERASIONAL", total_beban_operasional)

        ebitda = [laba_kotor[i] - total_beban_operasional[i] for i in range(12)]
        _tulis_total("EBITDA", ebitda)

        # --- PENYUSUTAN ---
        total_penyusutan, nilai_penyusutan = _sum_kelompok(akun_penyusutan)
        _tulis_kategori("PENYUSUTAN")
        for no_akun in akun_penyusutan:
            _tulis_item(no_akun, nilai_penyusutan[no_akun])
        _tulis_total("TOTAL PENYUSUTAN", total_penyusutan)
        # [BARU] Simpan nomor baris ini (kolom C..N = Jan..Des) supaya
        # sheet 12 "Laporan Arus Kas" bisa link ke sini utk baris
        # penyesuaian non-kas "Beban Penyusutan" (bukan angka Python beku).
        r_total_penyusutan_lr = ws.max_row

        laba_usaha = [ebitda[i] - total_penyusutan[i] for i in range(12)]
        _tulis_total("LABA USAHA", laba_usaha)

        # --- Pendapatan & Beban Lain-lain (baris berdiri sendiri, tanpa
        # header kategori/subtotal -- sama seperti file referensi) ---
        total_pendapatan_lain, nilai_pendapatan_lain = _sum_kelompok(akun_pendapatan_lain)
        for no_akun in akun_pendapatan_lain:
            _tulis_item(no_akun, nilai_pendapatan_lain[no_akun], biru=True)
        total_beban_lain, nilai_beban_lain = _sum_kelompok(akun_beban_lain)
        for no_akun in akun_beban_lain:
            _tulis_item(no_akun, nilai_beban_lain[no_akun], biru=True)

        laba_bersih_bulanan = [
            laba_usaha[i] + total_pendapatan_lain[i] - total_beban_lain[i] for i in range(12)
        ]
        _tulis_total("LABA BERSIH BULANAN", laba_bersih_bulanan)
        # [BARU] Simpan nomor baris "LABA BERSIH BULANAN" (kolom C..N =
        # nilai bulan Jan..Des, sama seperti "LABA BERSIH YTD" di bawah)
        # supaya sheet 11 "Laporan Perubahan Ekuitas" bisa nge-link ke
        # baris ini via formula LOKAL ='Laba Rugi Bulanan'!C{...} dst
        # (bukan angka Python beku) utk baris "Laba Tahun Berjalan"-nya.
        r_laba_bersih_bulanan_lr = ws.max_row

        laba_bersih_ytd, akumulasi = [], 0.0
        for v in laba_bersih_bulanan:
            akumulasi += v
            laba_bersih_ytd.append(round(akumulasi, 2))
        _tulis_total("LABA BERSIH YTD", laba_bersih_ytd)
        # [FIX BUG] Tangkap nomor baris ini -- lihat catatan panjang di
        # deklarasi default r_laba_bersih_ytd_lr di section 8 di atas.
        r_laba_bersih_ytd_lr = ws.max_row

        _format_rupiah_kolom(ws, list(range(3, ncols_lr + 1)), header_row_lr + 1, ws.max_row)
        for r in range(header_row_lr + 1, ws.max_row + 1):
            for c in range(3, ncols_lr + 1):
                ws.cell(row=r, column=c).number_format = _FORMAT_AKUNTANSI_TB_BULANAN
        # -- [FIX] Lebar kolom OTOMATIS menyesuaikan panjang teks ISI DATA
        #    terpanjang di kolom (bukan cuma header lagi -- header pendek
        #    spt "Jan-26" tapi angka di bawahnya bisa jauh lebih panjang) --
        for kolom, lebar in _lebar_kolom_dari_isi(ws).items():
            ws.column_dimensions[kolom].width = lebar
        # [FIX] Tinggi baris judul (baris 1) di referensi 19.5 (baris lain
        # dibiarkan auto -- selisih 15.75/16.5 di referensi hanya efek
        # rendering Excel per baris berisi border double, tidak signifikan
        # secara visual dan tidak di-set manual di sini).
        ws.row_dimensions[1].height = 27.95
    else:
        ws.cell(row=3, column=1,
                 value="Belum ada data Laba Rugi Bulanan -- generate laporan bulanan terlebih dahulu.")

    # ================= 10. Balance Sheet Bulanan =================
    # [FIX] Sebelumnya sheet ini cuma daftar rata SEMUA akun ASET/
    # LIABILITAS/EKUITAS tanpa pengelompokan/subtotal. Sekarang disusun
    # PERSIS pola file model referensi (BALANCE_SHEET_BULANAN___2025.xlsx):
    # ASET dipecah ke beberapa kelompok (mis. "Aset Lancar"/"Aset Tetap")
    # dgn baris "TOTAL {KELOMPOK}" tiap kelompok lalu "TOTAL ASET"; sama
    # utk LIABILITAS & EKUITAS; ditutup "TOTAL LIABILITAS & EKUITAS" dan
    # "CHECK BALANCE" (harus 0 kalau neraca sehat).
    #
    # KUNCI supaya sheet ini otomatis MENYESUAIKAN COA client mana pun
    # (bukan hardcode nama akun spt file model referensi, yg daftar
    # akunnya cuma cocok utk SATU perusahaan tertentu): kelompok diambil
    # dari Coa.sub_kategori (field bebas isi per akun -- KONVENSI yg sama
    # dipakai sheet Laba Rugi Bulanan di atas, akuntan isi mis. "Aset
    # Lancar"/"Aset Tetap" per akun lewat form COA). Kalau sub_kategori
    # belum diisi utk suatu akun, akun itu otomatis jatuh ke kelompok
    # fallback "{Kategori} Lainnya" -- sheet tetap kebentuk (tidak error),
    # cuma tidak split rapi sampai akuntan melengkapi sub_kategori.
    ws = wb.create_sheet("Balance Sheet Bulanan")
    ncols_bs = 2 + 12

    ws.cell(row=1, column=1, value=f"BALANCE SHEET BULANAN -- {tahun}")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols_bs)
    ws.cell(row=1, column=1).font = _TITLE_FONT_BS
    ws.cell(row=1, column=1).fill = _TITLE_FILL_BS
    ws.cell(row=1, column=1).alignment = _TITLE_ALIGN_BS
    ws.row_dimensions[1].height = 27.95

    if tb:
        header_row_bs = 3
        headers_bs = ["Kode", "Uraian"] + [f"{b}-{tahun[-2:]}" for b in _BULAN_SINGKAT]
        _tulis_header_14sheet(ws, headers_bs, row=header_row_bs)
        for c in range(1, ncols_bs + 1):
            cell = ws.cell(row=header_row_bs, column=c)
            cell.font = _HEADER_FONT_BS
            cell.fill = _HEADER_FILL_BS
            cell.alignment = _HEADER_ALIGN_BS
            cell.border = _HEADER_BORDER_KANAN_BS if c == ncols_bs else _HEADER_BORDER_BS
        ws.row_dimensions[header_row_bs].height = 30

        def _saldo_bulanan_bs(no_akun: str) -> List[float]:
            """Saldo KUMULATIF per akhir bulan (bukan delta) -- lihat
            catatan [FIX] "Sheet 8 & 10 SENGAJA tetap pakai saldo
            kumulatif" di atas."""
            info = tb[no_akun]
            nilai = [_angka(v) for v in (info.get("per_bulan") or [0] * 12)][:12]
            nilai += [0] * (12 - len(nilai))
            return nilai

        # Urutan kelompok umum akuntansi Indonesia -- label lain (atau
        # sub_kategori kosong) tetap tampil, cuma diurutkan alfabetis
        # SETELAH kelompok yang sudah dikenal ini.
        _PRIORITAS_SUBKATEGORI = {
            "aset": ["aset lancar", "aset tetap", "aset tidak lancar", "aset lain-lain"],
            "liabilitas": ["liabilitas jangka pendek", "liabilitas lancar",
                           "liabilitas jangka panjang", "liabilitas tidak lancar"],
            "ekuitas": ["modal", "saldo laba", "laba ditahan"],
        }

        def _urutan_grup(kategori_key: str, label_lower: str) -> tuple:
            urutan = _PRIORITAS_SUBKATEGORI.get(kategori_key, [])
            if label_lower in urutan:
                return (0, urutan.index(label_lower), label_lower)
            return (1, 0, label_lower)

        def _tulis_grup_header(label: str) -> None:
            """Header section utama (ASET/LIABILITAS/EKUITAS) -- fill biru
            muda #D9EAF7 + border bawah tipis, persis file referensi."""
            r = ws.max_row + 1
            ws.cell(row=r, column=2, value=label)
            for c in range(1, ncols_bs + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = _KATEGORI_FONT_BS
                cell.fill = _KATEGORI_FILL_BS
                cell.border = _KATEGORI_BORDER_BS

        def _tulis_subkategori_header_bs(label: str) -> None:
            """[FIX] Header sub-kelompok (mis. "Aset Lancar"/"Aset Tetap")
            -- SEBELUMNYA baris ini TIDAK PERNAH ditulis sama sekali
            (langsung loncat dari header section ke baris akun), padahal
            file model referensi selalu punya baris label sub-kelompok
            sendiri (fill lebih terang #F3F6FA, tanpa border) sebelum
            daftar akunnya."""
            r = ws.max_row + 1
            ws.cell(row=r, column=2, value=label)
            for c in range(1, ncols_bs + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = _SUBKATEGORI_FONT_BS
                cell.fill = _SUBKATEGORI_FILL_BS

        def _tulis_item_bs(no_akun: str, nilai: List[float]) -> None:
            nama = peta_akun.get(str(no_akun), {}).get("nama_akun") or tb[no_akun].get("nama_akun")
            ws.append([no_akun, nama] + [round(v, 2) for v in nilai])
            r = ws.max_row
            ws.cell(row=r, column=1).font = _ITEM_FONT_LABEL_BS
            ws.cell(row=r, column=2).font = _ITEM_FONT_LABEL_BS
            for c in range(3, ncols_bs + 1):
                ws.cell(row=r, column=c).font = _ITEM_FONT_ANGKA_BS

        def _tulis_total_bs(label: str, nilai: List[float]) -> None:
            ws.append(["", label] + [round(v, 2) for v in nilai])
            r = ws.max_row
            for c in range(1, ncols_bs + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = _TOTAL_FONT_ANGKA_BS if c >= 3 else _TOTAL_FONT_LABEL_BS
                cell.fill = _TOTAL_FILL_BS
                cell.border = _TOTAL_BORDER_BS

        def _kelompokkan(kategori_key: str, label_kategori: str) -> List[float]:
            """Tulis satu bagian (ASET/LIABILITAS/EKUITAS): kelompok per
            sub_kategori dgn subtotal masing2, return total bagian ini
            (list 12 bulan). Kalau tidak ada akun sama sekali di kategori
            ini, tidak menulis apa2 & return nol -- sheet tetap valid utk
            client yg belum punya akun kategori tsb."""
            akun_kategori = [no for no in sorted(tb.keys())
                              if str(peta_akun.get(str(no), {}).get("kategori", "")).lower() == kategori_key]
            if not akun_kategori:
                return [0.0] * 12

            _tulis_grup_header(label_kategori)

            grup: Dict[str, List[str]] = {}
            for no_akun in akun_kategori:
                sub = str(peta_akun.get(str(no_akun), {}).get("sub_kategori") or "").strip()
                label_grup = sub if sub else f"{label_kategori.title()} Lainnya"
                grup.setdefault(label_grup, []).append(no_akun)

            total_bagian = [0.0] * 12
            for label_grup in sorted(grup.keys(), key=lambda lb: _urutan_grup(kategori_key, lb.lower())):
                _tulis_subkategori_header_bs(label_grup)
                total_grup = [0.0] * 12
                for no_akun in grup[label_grup]:
                    nilai = _saldo_bulanan_bs(no_akun)
                    _tulis_item_bs(no_akun, nilai)
                    for i in range(12):
                        total_grup[i] += nilai[i]
                _tulis_total_bs(f"TOTAL {label_grup.upper()}", total_grup)
                for i in range(12):
                    total_bagian[i] += total_grup[i]

            return total_bagian

        # [FIX] Sebelumnya ada baris kosong (ws.append([None])) di antara
        # tiap section -- dihapus supaya SAMA PERSIS file model referensi,
        # yang langsung lanjut dari "TOTAL ASET" ke header "LIABILITAS"
        # tanpa baris pemisah, begitu juga "TOTAL LIABILITAS" -> "EKUITAS".
        total_aset = _kelompokkan("aset", "ASET")
        _tulis_total_bs("TOTAL ASET", total_aset)

        total_liabilitas = _kelompokkan("liabilitas", "LIABILITAS")
        _tulis_total_bs("TOTAL LIABILITAS", total_liabilitas)

        total_ekuitas = _kelompokkan("ekuitas", "EKUITAS")

        # [BARU] "Laba Tahun Berjalan" BUKAN akun COA tersendiri -- diambil
        # dari kumulatif LABA BERSIH YTD yg sudah dihitung di sheet Laba
        # Rugi Bulanan (variabel `laba_bersih_ytd`, section 9 di atas)
        # supaya Ekuitas ikut naik/turun tiap bulan sejalan dgn laba/rugi
        # berjalan -- PERSIS pola file model referensi (baris "Laba Tahun
        # Berjalan" = link ke sheet Laba Rugi Bulanan).
        ws.append(["", "Laba Tahun Berjalan"] + [round(v, 2) for v in laba_bersih_ytd])
        r_laba_ytd = ws.max_row
        ws.cell(row=r_laba_ytd, column=2).font = _ITEM_FONT_LABEL_BS
        for c in range(3, ncols_bs + 1):
            ws.cell(row=r_laba_ytd, column=c).font = _ITEM_FONT_ANGKA_BS
        total_ekuitas = [total_ekuitas[i] + laba_bersih_ytd[i] for i in range(12)]
        _tulis_total_bs("TOTAL EKUITAS", total_ekuitas)

        total_liab_ekuitas = [total_liabilitas[i] + total_ekuitas[i] for i in range(12)]
        _tulis_total_bs("TOTAL LIABILITAS & EKUITAS", total_liab_ekuitas)

        check_balance = [round(total_aset[i] - total_liab_ekuitas[i], 2) for i in range(12)]
        ws.append(["", "CHECK BALANCE"] + check_balance)
        r_check = ws.max_row
        sehat = all(abs(v) < 1 for v in check_balance)
        _font_angka_check = Font(name=_FONT_NAME_BS, bold=True, size=11,
                                  color="FF008000" if sehat else "FFFF0000")
        for c in range(1, ncols_bs + 1):
            cell = ws.cell(row=r_check, column=c)
            cell.font = _font_angka_check if c >= 3 else _TOTAL_FONT_LABEL_BS
            cell.fill = _TOTAL_FILL_BS
            cell.border = _TOTAL_BORDER_BS

        _format_rupiah_kolom(ws, list(range(3, ncols_bs + 1)), header_row_bs + 1, ws.max_row)
        for r in range(header_row_bs + 1, ws.max_row + 1):
            for c in range(3, ncols_bs + 1):
                ws.cell(row=r, column=c).number_format = _FORMAT_AKUNTANSI_TB_BULANAN
        # -- [FIX] Lebar kolom OTOMATIS menyesuaikan panjang teks ISI DATA
        #    terpanjang di kolom (bukan cuma header lagi) --
        for kolom, lebar in _lebar_kolom_dari_isi(ws).items():
            ws.column_dimensions[kolom].width = lebar
    else:
        ws.cell(row=3, column=1,
                 value="Belum ada data Balance Sheet Bulanan -- generate laporan bulanan terlebih dahulu.")

    # ================= 11. Laporan Perubahan Ekuitas [BARU] =================
    # [FIX] Struktur kolom sudah dikonfirmasi user (5 kolom: Keterangan/
    # Saldo Awal/Penambahan/Pengurangan/Saldo Akhir) -- diisi lewat
    # _tulis_sheet_perubahan_ekuitas() (lihat definisi & konstanta _BARU4
    # di atas), GANTI placeholder navy-judul-doang sebelumnya. Data
    # "perubahan_ekuitas" masih tetap dirangkum ringkas juga di sheet
    # "Ringkasan" bagian C (tidak dihapus, dua sheet ini saling melengkapi).
    ws = wb.create_sheet("Laporan Perubahan Ekuitas")
    _tulis_sheet_perubahan_ekuitas(
        ws, tahun, peta_akun, tb,
        _nsa_baris_awal_data, _nsa_baris_akhir_data,
        _gl_sheet_name, _gl_baris_awal_data, _gl_baris_akhir_data,
        r_laba_bersih_bulanan_lr, r_laba_bersih_ytd_lr,
    )
    # [FIX] Timpa lebar kolom manual (_LEBAR_KOLOM_PE) di dalam
    # _tulis_sheet_perubahan_ekuitas() dengan lebar OTOMATIS dari isi
    # data terpanjang, supaya label "{baris} -- {nama akun}" yang panjang
    # tidak terpotong.
    for kolom, lebar in _lebar_kolom_dari_isi(ws).items():
        ws.column_dimensions[kolom].width = lebar

    # ================= 12. Laporan Arus Kas [FIX/REBUILD] =================
    # [FIX] Sebelumnya sheet ini cuma daftar transaksi kas mentah per
    # kategori (angka Python beku dari lapkeu.susun_arus_kas_sederhana(),
    # parameter "arus_kas") -- TIDAK sekelas sheet-sheet lain (Trial
    # Balance/Laba Rugi/Balance Sheet Bulanan, Laporan Perubahan Ekuitas)
    # yg semuanya tabel bulanan RUMUS EXCEL HIDUP tie-out lintas sheet.
    # Diganti Laporan Arus Kas metode tidak langsung (lihat docstring
    # _tulis_sheet_arus_kas()) -- dipanggil dgn parameter PERSIS sama dgn
    # sheet 11 (tb/peta_akun/range NSA & GL yg sama, supaya tetap satu
    # sumber kebenaran) + r_total_penyusutan_lr. Parameter "arus_kas" yg
    # lama TIDAK dipakai lagi oleh sheet ini (sengaja tidak dihapus dari
    # signature export_14_sheet_lengkap() -- lihat catatan di situ --
    # supaya pemanggil lama tidak error, cuma tidak lagi dikonsumsi di
    # sini).
    ws = wb.create_sheet("Laporan Arus Kas")
    _tulis_sheet_arus_kas(
        ws, tahun, peta_akun, tb,
        _nsa_baris_awal_data, _nsa_baris_akhir_data,
        _gl_sheet_name, _gl_baris_awal_data, _gl_baris_akhir_data,
        r_laba_bersih_bulanan_lr, r_total_penyusutan_lr,
    )
    # [FIX] Timpa lebar kolom manual (_LEBAR_KOLOM_PE + kolom B=46 khusus)
    # di dalam _tulis_sheet_arus_kas() dengan lebar OTOMATIS dari isi data
    # terpanjang.
    for kolom, lebar in _lebar_kolom_dari_isi(ws).items():
        ws.column_dimensions[kolom].width = lebar

    # ================= 13. Catatan atas Laporan Keuangan (CALK) [BARU] =================
    # [FIX] Struktur kolom sudah dikonfirmasi user (4 kolom: No./Catatan/
    # Rincian-Penjelasan/Jumlah (Rp)) -- diisi lewat _tulis_sheet_calk(),
    # pakai data "calk" (dari lapkeu.susun_calk_otomatis()).
    # [FIX] Nama lama "Catatan atas Laporan Keuangan (CALK)" = 36 karakter,
    # melebihi batas KERAS Excel 31 karakter untuk nama sheet (satu-satunya
    # dari 18 sheet yang melebihi -- semua sheet lain di bawah 26 karakter).
    # openpyxl sendiri sampai warning "Title is more than 31 characters --
    # Some applications may not be able to read the file" saat file ini
    # disimpan -- di Excel asli ini bisa memicu file dianggap corrupt /
    # perlu "repair" saat dibuka, atau nama dipotong paksa. Dipendekkan
    # jadi 29 karakter (buang akronim "(CALK)" yang redundan, arti tetap
    # sama & masih jelas) supaya aman di semua aplikasi spreadsheet.
    ws = wb.create_sheet("Catatan atas Laporan Keuangan")
    _tulis_sheet_calk(ws, calk, tahun)
    # [FIX] Timpa lebar kolom manual (A=6/B=38/C=55/D=20) di dalam
    # _tulis_sheet_calk() dengan lebar OTOMATIS dari isi data terpanjang
    # -- kolom "Rincian/Penjelasan" (C) sering memuat teks panjang yang
    # sebelumnya kepotong di lebar tetap 55.
    for kolom, lebar in _lebar_kolom_dari_isi(ws).items():
        ws.column_dimensions[kolom].width = lebar

    # ================= 14. Ringkasan =================
    # [FIX] Diganti total mengikuti model referensi baru
    # (RINGKASAN_KINERJA_KEUANGAN_2025.xlsx): sebelumnya sheet ini punya
    # 5 section (A. Neraca, B. Laba Rugi, C. Perubahan Ekuitas, D. PPh
    # Badan, E. Tren Bulanan 8 baris) -- sekarang disederhanakan jadi
    # PERSIS 2 blok seperti model referensi:
    #   1. Tabel KPI 4 kolom (header "KPI 31 Desember <tahun>" / "Nilai" /
    #      "Sumber" / "Catatan") -- 6 baris tetap: Total Pendapatan, Laba
    #      Bersih, Total Aset, Piutang Usaha, Hutang Usaha, Nilai Buku
    #      Aset Tetap.
    #   2. Satu tabel tren bulanan gabungan (baris = Metrik: Pendapatan/
    #      Laba Bersih/Total Aset/Piutang Usaha, kolom = Jan-25..Dec-25).
    # Struktur/urutan kolom SELALU SAMA -- yang berubah cuma angkanya,
    # menyesuaikan data perusahaan (client) yang dikirim.
    #
    # [FIX] Styling (judul/header/border/font/lebar kolom) & 2 chart
    # garis ("Pendapatan dan Laba Bersih Bulanan" + "Tren Total Aset dan
    # Piutang") sekarang dicocokkan cell-per-cell ke file model
    # referensi. SATU deviasi disengaja: di file model referensi, kolom
    # data tabel tren bergeser 1 kolom dari headernya (header bulan ada
    # di kolom B-M tapi angkanya ada di kolom C-N, jadi kolom "Jan-25"
    # selalu kosong dan tiap angka nyambung ke header bulan berikutnya)
    # -- ini kelihatan seperti bug pergeseran kolom di file Excel asal
    # (Model_Laporan_Keuangan_SPT_PPh31E_2025.xlsx), bukan sesuatu yang
    # disengaja. Karena sheet ini laporan keuangan client sungguhan,
    # implementasi di bawah TETAP menyelaraskan header & data (kolom
    # B-M utk keduanya) supaya data bulanan tidak salah label -- semua
    # elemen visual lain (warna, font, border, lebar kolom, chart)
    # tetap dibuat identik dengan file model.
    ws = wb.create_sheet("Ringkasan")
    ws.row_dimensions[1].height = 27.95

    # -------- Judul (merge A1:N1, navy #17365D + Carlito 15 putih bold) --------
    ws.append(["RINGKASAN KINERJA KEUANGAN " + tahun])
    ws.merge_cells("A1:N1")
    cell_judul = ws.cell(row=1, column=1)
    cell_judul.font = _TITLE_FONT_PIUTANG
    cell_judul.fill = _TITLE_FILL_PIUTANG
    cell_judul.alignment = _TITLE_ALIGN_PIUTANG

    # [FIX] ws.append([]) TIDAK menambah baris di openpyxl (list kosong
    # tidak menaikkan max_row) -- pakai ws.append([None]) supaya baris
    # kosong pemisah ini sungguhan ada, sama seperti model referensi
    # (baris 2 kosong sebelum header KPI).
    ws.append([None])
    baris_header_kpi = ws.max_row + 1
    ws.row_dimensions[baris_header_kpi].height = 30
    _tulis_header_14sheet(
        ws, [f"KPI 31 Desember {tahun}", "Nilai", "Sumber", "Catatan"], row=baris_header_kpi
    )
    for c in range(1, 5):
        cell = ws.cell(row=baris_header_kpi, column=c)
        cell.font = _HEADER_FONT_PIUTANG
        cell.fill = _HEADER_FILL_PIUTANG
        cell.alignment = _HEADER_ALIGN_PIUTANG
        cell.border = (
            _HEADER_BORDER_KIRI_PIUTANG if c == 1
            else _HEADER_BORDER_KANAN_PIUTANG if c == 4
            else _HEADER_BORDER_TENGAH_PIUTANG
        )
    baris_awal_kpi = baris_header_kpi + 1

    def _baris_kpi(item: str, nilai, sumber: str, catatan: str = "") -> None:
        ws.append([item, round(_angka(nilai), 2), sumber, catatan])

    total_pendapatan_b = laba_rugi.get("total_pendapatan", 0)
    laba_bersih_b = laba_rugi.get("laba_rugi_bersih", 0)
    total_aset_b = neraca.get("total_aset", 0)

    _baris_kpi("Total Pendapatan", total_pendapatan_b, "Laba Rugi Bulanan",
               "Akumulasi pendapatan Januari-Desember")
    _baris_kpi("Laba Bersih", laba_bersih_b, "Laba Rugi Bulanan",
               "Laba setelah beban penyusutan")
    _baris_kpi("Total Aset", total_aset_b, "Balance Sheet Bulanan",
               f"Saldo per 31 Desember {tahun}")
    _baris_kpi("Piutang Usaha", total_saldo_piutang, "Buku Bantu Piutang",
               "Outstanding pelanggan")
    _baris_kpi("Hutang Usaha", total_saldo_hutang, "Buku Bantu Hutang",
               "Outstanding vendor")
    _baris_kpi("Nilai Buku Aset Tetap", total_nilai_buku, "Buku Bantu Aktiva Tetap",
               f"Setelah penyusutan {tahun}")
    baris_akhir_kpi = ws.max_row
    _format_rupiah_kolom(ws, [2], baris_awal_kpi, baris_akhir_kpi)

    # [FIX] Font/alignment/border baris data KPI sebelumnya tidak
    # di-styling sama sekali (ikut default openpyxl) -- sekarang label
    # (kolom A) bold Carlito, nilai (kolom B) bold Carlito hijau
    # #008000, Sumber/Catatan (kolom C/D) Calibri reguler, plus border
    # abu di pinggir kiri (kolom A)/kanan (kolom D) tabel dan bottom
    # border penutup di baris terakhir -- persis pola file model.
    for r in range(baris_awal_kpi, baris_akhir_kpi + 1):
        is_terakhir = (r == baris_akhir_kpi)
        cell_a, cell_b = ws.cell(row=r, column=1), ws.cell(row=r, column=2)
        cell_c, cell_d = ws.cell(row=r, column=3), ws.cell(row=r, column=4)
        cell_a.font, cell_b.font = _LABEL_FONT_RINGKASAN, _NILAI_FONT_RINGKASAN
        cell_c.font = cell_d.font = _SUMBER_FONT_RINGKASAN
        for cell in (cell_a, cell_b, cell_c, cell_d):
            cell.alignment = _VCENTER_RINGKASAN
        cell_a.border = _BORDER_KIRI_BAWAH_RINGKASAN if is_terakhir else _BORDER_KIRI_RINGKASAN
        cell_d.border = _BORDER_KANAN_BAWAH_RINGKASAN if is_terakhir else _BORDER_KANAN_RINGKASAN
        if is_terakhir:
            cell_b.border = cell_c.border = _BORDER_BAWAH_RINGKASAN

    ws.append([None])
    ws.append([None])

    # -------- [BARU -- integrasi Claude API] Temuan Penting & Potensi
    # Masalah -- digenerate SEBELUM data sampai di sini lewat
    # claude_client.analisis_ringkasan_keuangan_claude(), dikirim sebagai
    # data["ringkasan_analisis"] = {"ringkasan": str, "temuan_penting":
    # [...], "potensi_masalah": [...]}. Modul ini HANYA merender --
    # tidak melakukan panggilan network sendiri, supaya generate Excel
    # tetap cepat & tidak gagal kalau API sedang bermasalah (blok ini
    # cuma dilewati kalau datanya kosong). --------
    if ringkasan_analisis:
        teks_ringkasan = ringkasan_analisis.get("ringkasan") or ""
        temuan_penting = ringkasan_analisis.get("temuan_penting") or []
        potensi_masalah = ringkasan_analisis.get("potensi_masalah") or []

        if teks_ringkasan:
            baris_judul_analisis = ws.max_row + 1
            ws.cell(row=baris_judul_analisis, column=1, value="RINGKASAN ANALISIS")
            ws.merge_cells(start_row=baris_judul_analisis, start_column=1,
                            end_row=baris_judul_analisis, end_column=14)
            cell_j = ws.cell(row=baris_judul_analisis, column=1)
            cell_j.font = _HEADER_FONT_PIUTANG
            cell_j.fill = _HEADER_FILL_PIUTANG
            cell_j.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[baris_judul_analisis].height = 22
            baris_isi = ws.max_row + 1
            ws.cell(row=baris_isi, column=1, value=teks_ringkasan)
            ws.merge_cells(start_row=baris_isi, start_column=1, end_row=baris_isi, end_column=14)
            cell_isi = ws.cell(row=baris_isi, column=1)
            cell_isi.font = _SUMBER_FONT_RINGKASAN
            cell_isi.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[baris_isi].height = 30
            ws.append([None])

        def _tulis_blok_poin(judul: str, poin_list: List[str]) -> None:
            if not poin_list:
                return
            baris_j = ws.max_row + 1
            ws.cell(row=baris_j, column=1, value=judul)
            ws.merge_cells(start_row=baris_j, start_column=1, end_row=baris_j, end_column=14)
            cell_j = ws.cell(row=baris_j, column=1)
            cell_j.font = _LABEL_FONT_RINGKASAN
            cell_j.fill = _SUBSECTION_FILL_PE
            cell_j.alignment = Alignment(horizontal="left", vertical="center")
            for poin in poin_list:
                baris_p = ws.max_row + 1
                ws.cell(row=baris_p, column=1, value=f"\u2022 {poin}")
                ws.merge_cells(start_row=baris_p, start_column=1, end_row=baris_p, end_column=14)
                cell_p = ws.cell(row=baris_p, column=1)
                cell_p.font = _SUMBER_FONT_RINGKASAN
                cell_p.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.append([None])

        _tulis_blok_poin("Temuan Penting", temuan_penting)
        _tulis_blok_poin("Potensi Masalah / Perlu Direview", potensi_masalah)
        ws.append([None])

    # -------- Tren bulanan (1 tabel gabungan, 4 metrik) --------
    bsb = laporan_bulanan.get("balance_sheet_bulanan") or {}
    lrb = laporan_bulanan.get("laba_rugi_bulanan") or {}

    def _seri_12_bulan(nilai_mentah) -> List[float]:
        nilai_mentah = list(nilai_mentah or [])
        nilai_mentah = nilai_mentah[:12] + [0.0] * max(0, 12 - len(nilai_mentah))
        return [round(_angka(v), 2) for v in nilai_mentah]

    seri_pendapatan = _seri_12_bulan(lrb.get("total_pendapatan_bulanan"))
    seri_laba_bersih = _seri_12_bulan(lrb.get("laba_bersih_bulanan"))
    seri_total_aset = _seri_12_bulan(bsb.get("total_aset"))
    # Piutang Usaha bulanan -- ambil dari tren_piutang (snapshot riwayat
    # saldo, sudah 12 elemen Jan-Des sesuai get_tren_saldo_per_bulan()),
    # bukan dari laporan_bulanan (yang tidak menghitung saldo piutang).
    seri_piutang = _seri_12_bulan([item.get("saldo", 0) for item in tren_piutang]) \
        if tren_piutang else [0.0] * 12

    baris_header_tren = ws.max_row + 1
    ws.append(["Metrik"] + [f"{b}-{tahun[-2:]}" for b in _BULAN_SINGKAT])
    ws.append(["Pendapatan"] + seri_pendapatan)
    ws.append(["Laba Bersih"] + seri_laba_bersih)
    ws.append(["Total Aset"] + seri_total_aset)
    ws.append(["Piutang Usaha"] + seri_piutang)
    baris_akhir_tren = ws.max_row

    for c in range(1, 14):
        cell = ws.cell(row=baris_header_tren, column=c)
        cell.font = _HEADER_FONT_PIUTANG
        cell.fill = _HEADER_FILL_PIUTANG
        cell.alignment = _HEADER_ALIGN_PIUTANG
        cell.border = (
            _HEADER_BORDER_KIRI_PIUTANG if c == 1
            else _HEADER_BORDER_KANAN_PIUTANG if c == 13
            else _HEADER_BORDER_TENGAH_PIUTANG
        )
    ws.row_dimensions[baris_header_tren].height = 30
    _format_rupiah_kolom(ws, list(range(2, 14)), baris_header_tren + 1, baris_akhir_tren)

    # [FIX] Font/border baris data tren sebelumnya default (tidak ada
    # styling) -- sekarang label (kolom A) Calibri reguler, nilai
    # (kolom B-M) Carlito hijau #008000 non-bold, plus border abu
    # pinggir kiri/kanan tabel & bottom border penutup di baris
    # terakhir -- persis pola file model.
    for r in range(baris_header_tren + 1, baris_akhir_tren + 1):
        is_terakhir = (r == baris_akhir_tren)
        cell_label = ws.cell(row=r, column=1)
        cell_label.font = _METRIK_FONT_RINGKASAN
        cell_label.alignment = _VCENTER_RINGKASAN
        cell_label.border = _BORDER_KIRI_BAWAH_RINGKASAN if is_terakhir else _BORDER_KIRI_RINGKASAN
        for c in range(2, 14):
            cell = ws.cell(row=r, column=c)
            cell.font = _NILAI_TREN_FONT_RINGKASAN
            cell.alignment = _VCENTER_RINGKASAN
            if c == 13:
                cell.border = _BORDER_KANAN_BAWAH_RINGKASAN if is_terakhir else _BORDER_KANAN_RINGKASAN
            elif is_terakhir:
                cell.border = _BORDER_BAWAH_RINGKASAN

    # -------- 2 chart garis (line chart), sesuai file model referensi --------
    # Chart 1: "Pendapatan dan Laba Bersih Bulanan" -- 1 series per bulan
    # (kolom B-M), 2 titik data per series (baris Pendapatan & Laba Bersih).
    chart1 = LineChart()
    chart1.title = "Pendapatan dan Laba Bersih Bulanan"
    chart1.style = 2
    chart1.legend.position = "b"
    data1 = Reference(ws, min_col=2, max_col=13, min_row=baris_header_tren, max_row=baris_header_tren + 2)
    cat1 = Reference(ws, min_col=1, min_row=baris_header_tren + 1, max_row=baris_header_tren + 2)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cat1)
    for s in chart1.series:
        s.marker.symbol = "circle"
        s.smooth = False
    _gaya_gridline_ringkasan(chart1)
    chart1.width, chart1.height = 20, 8
    ws.add_chart(chart1, "F3")

    # Chart 2: "Tren Total Aset dan Piutang" -- 1 series per bulan,
    # 4 titik data per series (baris Pendapatan/Laba Bersih/Total Aset/
    # Piutang Usaha).
    chart2 = LineChart()
    chart2.title = "Tren Total Aset dan Piutang"
    chart2.style = 2
    chart2.legend.position = "b"
    data2 = Reference(ws, min_col=2, max_col=13, min_row=baris_header_tren, max_row=baris_akhir_tren)
    cat2 = Reference(ws, min_col=1, min_row=baris_header_tren + 1, max_row=baris_akhir_tren)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cat2)
    for s in chart2.series:
        s.marker.symbol = "circle"
        s.smooth = False
    _gaya_gridline_ringkasan(chart2)
    chart2.width, chart2.height = 20, 13
    ws.add_chart(chart2, "F18")

    # [FIX] Lebar kolom OTOMATIS menyesuaikan panjang teks ISI DATA
    # terpanjang di kolom -- dipindah ke SINI (setelah kedua tabel & 2
    # chart selesai ditulis), GANTI pemanggilan _lebar_kolom_gabungan_header
    # yang sebelumnya dijalankan sebelum ada data sama sekali (langsung
    # setelah ws = wb.create_sheet(...), jadi cuma bisa lihat header, tidak
    # bisa lihat isi baris KPI/Tren Bulanan). Baris judul (merge A1:N1)
    # otomatis dilewati oleh _lebar_kolom_dari_isi jadi tidak perlu
    # baris_mulai manual.
    for kolom, lebar in _lebar_kolom_dari_isi(ws).items():
        ws.column_dimensions[kolom].width = lebar

    # ================= 15/16. BS Lampiran SPT / PNL Lampiran SPT =================
    # [FIX] Ditambahkan: (1) kolom "Saldo Tahun Lalu" -- untuk BS diambil
    # dari saldo_awal per akun di COA (yang memang merepresentasikan
    # posisi akhir tahun sebelumnya, lihat generate_neraca_saldo_awal_virtual);
    # untuk PNL belum ada sumber data akun P&L tahun lalu yang tersimpan
    # di sistem (saldo_awal akun Pendapatan/Beban selalu 0 krn ditutup tiap
    # tahun), jadi kolomnya disiapkan lewat data["laba_rugi_tahun_lalu"]
    # opsional ({no_akun: saldo_akhir tahun lalu}) -- kalau belum dikirim,
    # tampil "-" (bukan 0, supaya tidak menyesatkan). (2) baris subtotal
    # per kategori: BS -> "JUMLAH ASET LANCAR"/"JUMLAH ASET TIDAK LANCAR"/
    # "JUMLAH LIABILITAS"/"JUMLAH EKUITAS"/"JUMLAH LIABILITAS DAN EKUITAS";
    # PNL -> "LABA KOTOR" (L01-L02) dan "LABA USAHA" (Laba Kotor-L03).
    lampiran_bs = lampiran_rinci.get("lampiran_bs") or {}
    lampiran_pnl = lampiran_rinci.get("lampiran_pnl") or {}
    laba_rugi_tahun_lalu = data.get("laba_rugi_tahun_lalu") or {}

    def _saldo_tahun_lalu_akun(no_akun) -> float:
        return _angka(peta_akun.get(str(no_akun), {}).get("saldo_awal"))

    def _tulis_baris_akun_bs(ws_lampiran, kode: str, rincian) -> tuple:
        total_ini, total_lalu = 0.0, 0.0
        if not isinstance(rincian, list):
            ws_lampiran.append([kode, "-", "(nilai tunggal)", rincian, ""])
            return _angka(rincian), 0.0
        if not rincian:
            ws_lampiran.append([kode, "-", "(tidak ada akun)", 0, 0])
            return 0.0, 0.0
        for item in rincian:
            no_akun = item.get("no_akun")
            saldo_ini = _angka(item.get("saldo", 0))
            saldo_lalu = _saldo_tahun_lalu_akun(no_akun)
            ws_lampiran.append([kode, no_akun, item.get("nama_akun"), saldo_ini, saldo_lalu])
            total_ini += saldo_ini
            total_lalu += saldo_lalu
        return round(total_ini, 2), round(total_lalu, 2)

    def _tulis_subtotal_bs(ws_lampiran, label: str, nilai_ini, nilai_lalu) -> None:
        ws_lampiran.append(["-", "-", label, round(_angka(nilai_ini), 2), round(_angka(nilai_lalu), 2)])
        r = ws_lampiran.max_row
        for c in (3, 4, 5):
            ws_lampiran.cell(row=r, column=c).font = _SUBTOTAL_FONT_14SHEET

    # [BARU] Sheet 12 dirombak total supaya kolomnya BAKU/identik utk semua
    # client (Kode, Uraian, 31 Desember <tahun>, 31 Desember <tahun lalu>/
    # Saldo Awal, Keterangan) -- persis format template referensi "NERACA
    # -- LAMPIRAN SPT TAHUNAN BADAN (DALAM RUPIAH)". Isinya (jumlah baris
    # liabilitas/ekuitas, nilai tiap bucket aset) tetap menyesuaikan data
    # client masing-masing lewat susun_neraca_lampiran_spt_baku() di
    # modules/laporan_keuangan.py. Fungsi/variabel lama
    # (_tulis_baris_akun_bs, _tulis_subtotal_bs, lampiran_bs) TIDAK dipakai
    # lagi di sini tapi sengaja tidak dihapus -- masih dipakai sheet PNL di
    # bawah (_tulis_baris_akun_pnl beda fungsi) & untuk kompatibilitas kalau
    # ada pemanggil lain yang masih mengandalkan lampiran_rinci["lampiran_bs"].
    # [FIX] Sheet 12 ditulis oleh helper khusus _tulis_sheet_neraca_lampiran_spt_baku()
    # (meniru PERSIS file referensi NERACA___LAMPIRAN_SPT_TAHUNAN.xlsx sel demi
    # sel -- lihat konstanta _NERACA_* di atas) -- GANTI penulisan lama yang
    # masih pakai gaya generik _tulis_header_14sheet/_SUBTOTAL_FONT_14SHEET/
    # _format_rupiah_kolom/_autofit_14sheet (tidak identik dgn file referensi:
    # font/fill/border/warna kolom nilai beda, ada baris ekstra "JUMLAH ASET
    # TIDAK LANCAR" yang tidak ada di template, dan lebar kolom hasil autofit
    # bukan lebar baku file referensi).
    ws_bs = wb.create_sheet("BS Lampiran SPT")
    if not neraca:
        ws_bs.append(["Belum ada data BS Lampiran SPT -- generate laporan keuangan terlebih dahulu."])
    else:
        neraca_baku = susun_neraca_lampiran_spt_baku(neraca, coa, tahun, tahun_sebelumnya)
        _tulis_sheet_neraca_lampiran_spt_baku(ws_bs, neraca_baku, tahun)
        # [FIX] Timpa lebar kolom lama (_lebar_kolom_dari_header, cuma
        # lihat header) dengan lebar OTOMATIS dari isi data terpanjang.
        for kolom, lebar in _lebar_kolom_dari_isi(ws_bs).items():
            ws_bs.column_dimensions[kolom].width = lebar

    # [BARU] Sheet 13 dirombak total supaya kolomnya BAKU/identik utk semua
    # client (Kode, Uraian, Komersial <tahun>, Koreksi Positif, Koreksi
    # Negatif, Fiskal <tahun>, Keterangan) -- persis format template
    # referensi "LABA RUGI & REKONSILIASI FISKAL -- LAMPIRAN SPT TAHUNAN
    # BADAN", pasangan langsung dari rombakan sheet 12 (BS Lampiran SPT) di
    # atas. Isinya (jumlah baris pendapatan/beban, nilai tiap seksi, angka
    # rekonsiliasi fiskal -> PKP) menyesuaikan data client masing-masing
    # lewat susun_laba_rugi_rekonsiliasi_fiskal_lampiran_spt_baku() di
    # modules/laporan_keuangan.py. Fungsi/variabel lama (_tulis_baris_akun_pnl,
    # _tulis_subtotal_pnl, lampiran_pnl, laba_rugi_tahun_lalu) TIDAK dipakai
    # lagi di sini tapi sengaja tidak dihapus -- sama alasannya seperti
    # _tulis_baris_akun_bs/_tulis_subtotal_bs di sheet 12 (kompatibilitas
    # mundur untuk pemanggil lain yang masih mengandalkan lampiran_rinci).
    # [FIX] pnl_baku dihitung SEKALI di sini (bukan di dalam blok if/else
    # sheet 16 lagi) supaya bisa dipakai ULANG oleh sheet 17 "Rekonsiliasi
    # Fiskal" di bawah tanpa menghitung ulang -- 1 sumber angka, kedua
    # sheet dijamin tie-out.
    pnl_baku: Optional[Dict[str, Any]] = None
    info_sheet_pnl: Optional[Dict[str, Any]] = None
    ws_pnl = wb.create_sheet("PNL Lampiran SPT")
    if not laba_rugi:
        ws_pnl.append(["Belum ada data PNL Lampiran SPT -- generate laporan keuangan terlebih dahulu."])
    else:
        rekon_pkp = (pph_hasil or {}).get("rekonsiliasi_fiskal")
        pnl_baku = susun_laba_rugi_rekonsiliasi_fiskal_lampiran_spt_baku(laba_rugi, tahun, rekon_pkp)
        info_sheet_pnl = _tulis_sheet_pnl_lampiran_spt_baku(ws_pnl, pnl_baku, tahun)
        # [FIX] Timpa lebar kolom lama (_lebar_kolom_dari_header, cuma
        # lihat header) dengan lebar OTOMATIS dari isi data terpanjang.
        for kolom, lebar in _lebar_kolom_dari_isi(ws_pnl).items():
            ws_pnl.column_dimensions[kolom].width = lebar

    # ================= 17. Rekonsiliasi Fiskal [ROMBAK TOTAL] =================
    # [FIX -- match 100% referensi] Struktur 9 kolom + checklist baku KFP/
    # KFN + blok ringkasan REV..NP tertaut formula lintas-sheet ke sheet 16
    # -- lihat komentar blok besar di atas _tulis_sheet_rekonsiliasi_fiskal()
    # utk detail lengkap perubahannya. `info_sheet_pnl` (return value baru
    # _tulis_sheet_pnl_lampiran_spt_baku() di atas) WAJIB diteruskan supaya
    # blok ringkasan atas jadi formula hidup, bukan angka 0 statis.
    ws = wb.create_sheet("Rekonsiliasi Fiskal")
    _tulis_sheet_rekonsiliasi_fiskal(ws, pnl_baku or {}, tahun, info_sheet_pnl)
    # [FIX] Timpa lebar kolom lama (_lebar_kolom_dari_header, cuma lihat
    # header) dengan lebar OTOMATIS dari isi data terpanjang.
    for kolom, lebar in _lebar_kolom_dari_isi(ws).items():
        ws.column_dimensions[kolom].width = lebar

    # ================= 18. PPh Badan 31E =================
    # [FIX] Sheet 18 ditulis oleh helper khusus _tulis_sheet_pph_badan_31e()
    # (meniru PERSIS file referensi PERHITUNGAN_PPh_BADAN___TARIF_PASAL_17_
    # DAN_FASILITAS_PASAL_31E.xlsx sel demi sel -- lihat konstanta _PPH31E_*
    # di atas) -- GANTI penulisan lama yang cuma ws.append() polos tanpa
    # gaya/warna/border/formula sama sekali (angka mentah hasil Python, bukan
    # formula Excel hidup).
    ws = wb.create_sheet("PPh Badan 31E")
    if pph_hasil:
        _tulis_sheet_pph_badan_31e(ws, pph_hasil)
        # [FIX] Timpa lebar kolom manual (_PPH31E_LEBAR_KOLOM, patokan file
        # model referensi) dengan lebar OTOMATIS dari isi data terpanjang
        # -- kolom "Keterangan" (E) di sheet ini sering memuat teks
        # penjelasan yang lebih panjang dari lebar patokan referensi.
        for kolom, lebar in _lebar_kolom_dari_isi(ws).items():
            ws.column_dimensions[kolom].width = lebar
    else:
        ws.append(["Belum ada perhitungan PPh Badan 31E -- generate PPh Badan terlebih dahulu."])

    logger.info(f"Export 18-sheet lengkap periode {periode}: {len(wb.sheetnames)} sheet dibuat: {wb.sheetnames}")
    return wb


def export_18_sheet_lengkap(data: Dict[str, Any]) -> bytes:
    """
    Export 18-sheet Excel LENGKAP -- dipakai endpoint download
    (POST .../export-18-sheet). Susunan sheet & kolom: lihat docstring
    _susun_workbook_18_sheet() (badan logicnya ada di situ) -- semua 18
    sheet (termasuk 4 sheet baru 11/12/13/17) sudah terisi lengkap.

    Returns:
        bytes: isi file .xlsx siap dikirim lewat StreamingResponse.
    """
    wb = _susun_workbook_18_sheet(data)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _nilai_json_aman(v: Any) -> Any:
    """Konversi 1 nilai sel openpyxl ke tipe yang aman di-JSON-kan --
    datetime/date -> string ISO, Decimal -> float, sisanya apa adanya."""
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, Decimal):
        return float(v)
    return v


def export_18_sheet_sebagai_json(data: Dict[str, Any]) -> Dict[str, Any]:
    """
    [BARU] Versi JSON dari export_18_sheet_lengkap() -- dipakai endpoint
    preview (GET .../export-18-sheet-json) supaya ke-18 sheet (Petunjuk &
    Asumsi, COA, Neraca Saldo Awal, GL <tahun>, Buku Bantu Piutang/Hutang/
    Aktiva Tetap, Trial Balance/Laba Rugi/Balance Sheet Bulanan, Laporan
    Perubahan Ekuitas, Laporan Arus Kas, CALK, Ringkasan, BS/PNL Lampiran
    SPT, Rekonsiliasi Fiskal, PPh Badan 31E -- semuanya sudah terisi
    lengkap, lihat _susun_workbook_18_sheet()) bisa ditampilkan
    LANGSUNG DI LAYAR, bukan cuma lewat file .xlsx yang di-download.

    SENGAJA memakai ulang _susun_workbook_18_sheet() -- fungsi PERSIS SAMA
    yang dipakai export_18_sheet_lengkap() -- bukan menyusun ulang data
    secara terpisah, supaya versi layar & versi Excel PASTI selalu sinkron
    satu sama lain. Lihat pola bug berulang lain di modul ini (kolom yang
    diam-diam kosong karena 2 tempat berbeda menebak nama field secara
    terpisah) -- caranya dihindari di sini dengan cuma ada 1 sumber
    kebenaran (workbook), dibaca ulang, bukan ditulis dobel.

    Returns:
        {"sheets": [{"nama": str, "rows": [[cell, cell, ...], ...]}, ...]}
        Urutan "sheets" = urutan sheet di file Excel (18 sheet, sesuai
        _susun_workbook_18_sheet()). Setiap "rows" APA ADANYA dari isi
        sheet (termasuk baris judul, baris kosong pemisah, baris header
        tabel, baris TOTAL) -- frontend yang memutuskan cara merender,
        sama seperti orang buka file Excel-nya langsung.
    """
    wb = _susun_workbook_18_sheet(data)
    sheets = []
    for ws in wb.worksheets:
        rows = [[_nilai_json_aman(v) for v in row] for row in ws.iter_rows(values_only=True)]
        sheets.append({"nama": ws.title, "rows": rows})
    return {"sheets": sheets}