"""
modules/excel_export_worker.py
================================
[BARU -- POINT 2] Modul LEAF khusus untuk generate file Excel yang berat
secara CPU (openpyxl susun banyak sheet, iterasi ribuan baris).

KENAPA TIDAK CUKUP asyncio.to_thread SAJA
------------------------------------------
asyncio.to_thread() melempar eksekusi ke THREAD, bukan PROSES baru --
cukup untuk kerja I/O-bound (baca file, query DB), karena GIL dilepas
selagi menunggu I/O. Tapi generate Excel besar itu murni CPU-bound --
Python tetap pegang GIL selama openpyxl menulis sel demi sel. Beberapa
thread yang sama-sama CPU-bound tetap GANTIAN pakai 1 core saja (lihat
benchmark yang sudah ada di ai_file_reader.py: 6 file 8000 baris,
paralel-thread 1.42s vs sequential 1.33s -- nyaris tidak ada percepatan).

ProcessPoolExecutor beda -- tiap proses worker py punya interpreter +
GIL masing-masing, jadi kalau ada beberapa request generate Excel
bersamaan, mereka BENERAN jalan paralel lintas core (bukan cuma
bergantian).

KENAPA HARUS MODUL TERPISAH (bukan fungsi biasa langsung di main.py)
----------------------------------------------------------------------
ProcessPoolExecutor mem-*pickle* fungsi & argumennya untuk dikirim ke
proses worker. Fungsi yang didefinisikan di dalam main.py (yang
meng-import FastAPI, db_client, dsb) berisiko gagal di-pickle, atau --
lebih buruk -- memaksa proses worker baru ikut mengimpor & menjalankan
efek samping level-modul dari main.py (bikin app FastAPI baru, dst).
Modul ini SENGAJA tidak mengimpor FastAPI/db_client sama sekali (LEAF
module) supaya aman dijalankan di proses terpisah.

CARA PAKAI dari main.py (di dalam endpoint async def)
--------------------------------------------------------
    await asyncio.to_thread(
        excel_export_worker.jalankan_buat_excel_hasil_di_proses,
        hasil_semua, str(FOLDER_HASIL), label_per_kode,
    )

Perhatikan: fungsi jalankan_xxx_di_proses() DI SINI masih sync/blocking
(submit ke process pool lalu .result() -- menunggu hasil). Makanya di
sisi main.py, pemanggilannya SENDIRI tetap dibungkus asyncio.to_thread.
Pola dua-lapis ini disengaja: to_thread membebaskan EVENT LOOP dari
blocking-wait-nya, ProcessPoolExecutor membebaskan CPU dari GIL selagi
openpyxl benar-benar bekerja.
"""

from __future__ import annotations

import uuid
from concurrent.futures import ProcessPoolExecutor
from pathlib import Path
from typing import Any, Dict, Optional

import openpyxl
import pandas as pd

# Satu ProcessPoolExecutor dipakai bersama sepanjang umur aplikasi --
# BUKAN dibuat ulang tiap request (spawn proses baru ada overhead-nya
# sendiri). Dibuat lazy (baru dibuat saat dibutuhkan pertama kali) biar
# tidak nge-spawn proses kalau fitur generate Excel tidak pernah dipakai.
_POOL: Optional[ProcessPoolExecutor] = None


def _dapatkan_pool() -> ProcessPoolExecutor:
    global _POOL
    if _POOL is None:
        _POOL = ProcessPoolExecutor(max_workers=2)
    return _POOL


# ============================================================
# WORKER 1 -- dipindah dari _buat_excel_hasil() lama di main.py
# (dipakai endpoint /api/proses-dan-buat-excel)
# ============================================================

def _worker_buat_excel_hasil(
    hasil_per_jenis: Dict[str, Any],
    folder_hasil: str,
    label_per_kode: Dict[str, str],
) -> str:
    """
    Dijalankan DI PROSES TERPISAH -- HARUS fungsi top-level (bukan
    nested/closure) supaya bisa di-pickle oleh ProcessPoolExecutor.
    Isinya identik dengan _buat_excel_hasil() lama di main.py; bedanya
    cuma label sekarang datang lewat parameter label_per_kode (bukan
    baca _PEMROSES_DOKUMEN langsung dari main.py) -- supaya modul ini
    tidak perlu mengimpor apa pun dari main.py (LEAF, aman di-pickle
    lintas proses, dan tidak ada duplikasi/risiko drift dari
    _PEMROSES_DOKUMEN karena label-nya dikirim, bukan disalin ulang).

    Return: string path (str, bukan Path) -- representasi paling aman
    saat dikembalikan lintas proses lewat pickle.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for kode, hasil in hasil_per_jenis.items():
        label = label_per_kode.get(kode, kode)

        ws_ringkasan = wb.create_sheet(f"Ringkasan-{kode}"[:31])
        ws_ringkasan.append(["Ringkasan", label])
        for k, v in (hasil.get("ringkasan") or {}).items():
            ws_ringkasan.append([k, str(v)])

        masalah = hasil.get("masalah") or []
        ws_masalah = wb.create_sheet(f"Perlu Review-{kode}"[:31])
        ws_masalah.append(["Baris", "Alasan", "Rekomendasi"])
        for m in masalah:
            baris = m.get("baris")
            alasan_gabungan = " | ".join(m.get("alasan", []))
            rekomendasi_gabungan = " | ".join(m.get("rekomendasi", []))
            ws_masalah.append([baris, alasan_gabungan, rekomendasi_gabungan])

        draf = hasil.get("draf_jurnal") or []
        if draf:
            ws_jurnal = wb.create_sheet(f"Draf Jurnal-{kode}"[:31])
            kolom = list(draf[0].keys())
            ws_jurnal.append(kolom)
            for baris_jurnal in draf:
                ws_jurnal.append([str(baris_jurnal.get(k, "")) for k in kolom])

        rekon_fiskal = hasil.get("rekonsiliasi_fiskal") or []
        if rekon_fiskal:
            ws_fiskal = wb.create_sheet(f"Rekon Fiskal-{kode}"[:31])
            kolom = list(rekon_fiskal[0].keys())
            ws_fiskal.append(kolom)
            for baris_fiskal in rekon_fiskal:
                ws_fiskal.append([str(baris_fiskal.get(k, "")) for k in kolom])

        di_bawah_kapitalisasi = hasil.get("aset_di_bawah_batas_kapitalisasi") or []
        if di_bawah_kapitalisasi:
            ws_kap = wb.create_sheet(f"Batas Kapitalisasi-{kode}"[:31])
            kolom = list(di_bawah_kapitalisasi[0].keys())
            ws_kap.append(kolom)
            for baris_kap in di_bawah_kapitalisasi:
                ws_kap.append([str(baris_kap.get(k, "")) for k in kolom])

    nama_unik = f"{uuid.uuid4().hex}.xlsx"
    path_file = Path(folder_hasil) / nama_unik
    wb.save(path_file)
    return str(path_file)


def jalankan_buat_excel_hasil_di_proses(
    hasil_per_jenis: Dict[str, Any],
    folder_hasil: str,
    label_per_kode: Dict[str, str],
) -> Path:
    """
    Dipanggil dari main.py lewat asyncio.to_thread (lihat docstring
    modul di atas). Fungsi INI sendiri sync/blocking -- submit kerja ke
    ProcessPoolExecutor lalu .result() (menunggu proses worker
    selesai). Exception dari dalam proses worker (kalau ada) akan
    dilempar ulang di sini secara transparan oleh .result().
    """
    pool = _dapatkan_pool()
    future = pool.submit(_worker_buat_excel_hasil, hasil_per_jenis, folder_hasil, label_per_kode)
    path_str = future.result()
    return Path(path_str)


# ============================================================
# WORKER 2 -- accounting_export.export_rekening_koran_format_akuntan()
# (dipakai endpoint /api/client/{client_id}/rekening-koran/
# export-format-akuntan/{hasil_id})
# ============================================================
# Beda dari worker 1: fungsi aslinya SUDAH lengkap di
# modules/accounting_export.py, jadi di sini TIDAK disalin ulang --
# cukup dipanggil dari dalam worker top-level supaya bisa di-pickle.
# accounting_export.py sendiri sudah tidak menerima client_id / akses
# DB (lihat docstring export_rekening_koran_format_akuntan -- "TIDAK
# LAGI menerima client_id"), jadi aman diimpor & dijalankan di proses
# terpisah.

def _worker_export_rekening_koran_format_akuntan(
    df_hasil: "pd.DataFrame",
    df_coa: Optional["pd.DataFrame"],
    df_piutang: Optional["pd.DataFrame"],
) -> bytes:
    """
    Top-level (picklable). Import accounting_export DI DALAM fungsi
    worker (bukan di top-level modul ini) supaya proses worker yang
    baru di-spawn tidak ikut mengimpor apa pun sebelum benar-benar
    dibutuhkan -- juga menghindari kemungkinan import melingkar kalau
    accounting_export.py suatu saat ikut mengimpor modul ini.
    """
    from . import accounting_export
    return accounting_export.export_rekening_koran_format_akuntan(
        df_hasil, df_coa=df_coa, df_piutang=df_piutang,
    )


def jalankan_export_rekening_koran_format_akuntan_di_proses(
    df_hasil: "pd.DataFrame",
    df_coa: Optional["pd.DataFrame"] = None,
    df_piutang: Optional["pd.DataFrame"] = None,
) -> bytes:
    """
    Versi ProcessPoolExecutor dari
    accounting_export.export_rekening_koran_format_akuntan() -- dipanggil
    dari main.py lewat asyncio.to_thread, pola identik dengan
    jalankan_buat_excel_hasil_di_proses() di atas. ValueError dari fungsi
    asli (mis. df_hasil kosong / kolom 'bank' tidak ada) tetap naik apa
    adanya lewat future.result(), sehingga endpoint pemanggil tetap bisa
    except ValueError persis seperti sebelum dipindah ke process pool.
    """
    pool = _dapatkan_pool()
    future = pool.submit(
        _worker_export_rekening_koran_format_akuntan, df_hasil, df_coa, df_piutang
    )
    return future.result()