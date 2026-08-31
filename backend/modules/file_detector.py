"""
modules/file_detector.py
=========================
Deteksi otomatis jenis file Excel yang diupload — independen dari nama
perusahaan (PT) / nama file, murni berdasarkan pola nama sheet dan header
kolom di dalam file itu sendiri.

Ini MELENGKAPI detektor "jurnal koran" (rekening koran bank) yang sudah ada
sebelumnya. Jenis file baru yang dikenali di modul ini:

- SALES_JOURNAL       : Jurnal Penjualan per outlet
                         (sheet: SALES, REKAP SALES ESB, input xero,
                          REKAP UNPAID, PHR <nama outlet>)
- CASH_RECONCILIATION : Rekonsiliasi Kas Masuk vs Data Sales (ESB)
                         (sheet: "uang masuk vs sales", "Uang Masuk")
- POS_TRANSACTION     : Rekap Transaksi POS / pivot per metode bayar
                         (sheet: "Report Transaction - <tanggal>", atau
                          pivot "Sum of Gross Sales")

Kenapa deteksi dilakukan dari SHEET NAME + HEADER, bukan dari nama file:
nama file berbeda-beda per PT/outlet (mis. "NPI_Sales_Jurnal_Plaza_Marietta"
vs "NPI_Sales_Jurnal_Sibarita"), tapi struktur sheet & header-nya konsisten
karena berasal dari template/sistem kasir (ESB) yang sama. Jadi deteksi ini
otomatis bekerja untuk PT/outlet manapun tanpa perlu hardcode nama outlet.

Didesain murni dengan openpyxl (read_only) — HANYA membaca nama sheet dan
beberapa baris pertama untuk header, TIDAK memuat seluruh isi file, supaya
cepat dipakai sebagai langkah pertama sebelum file diproses lebih lanjut.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Union

import openpyxl

# logging_config.py ada di dalam modules/ (satu folder dengan file ini),
# dan main.py meng-import lewat `from modules import (...)` -- artinya
# modules/ diperlakukan sebagai package (ada __init__.py), bukan folder
# yang ditambahkan langsung ke sys.path. Karena itu HARUS pakai relative
# import; flat import (`from logging_config import ...`) akan gagal
# dengan ModuleNotFoundError karena modules/ sendiri tidak ada di sys.path.
from .logging_config import get_module_logger

logger = get_module_logger("file_detector")


# ============================================================
# JENIS FILE YANG DIKENALI
# ============================================================
JENIS_SALES_JOURNAL = "SALES_JOURNAL"
JENIS_CASH_RECONCILIATION = "CASH_RECONCILIATION"
JENIS_POS_TRANSACTION = "POS_TRANSACTION"
JENIS_TIDAK_DIKENALI = "UNKNOWN"

# Label tampilan (Bahasa Indonesia) untuk tiap jenis - dipakai di UI/log.
LABEL_JENIS: Dict[str, str] = {
    JENIS_SALES_JOURNAL: "Jurnal Penjualan",
    JENIS_CASH_RECONCILIATION: "Rekonsiliasi Kas Masuk vs Sales",
    JENIS_POS_TRANSACTION: "Rekap Transaksi POS",
    JENIS_TIDAK_DIKENALI: "Tidak Dikenali",
}


@dataclass
class HasilDeteksi:
    """Hasil deteksi jenis file untuk satu file Excel."""
    file: str
    jenis: str
    label: str
    confidence: float  # 0.0 - 1.0
    alasan: List[str] = field(default_factory=list)
    sheet_cocok: List[str] = field(default_factory=list)

    def to_dict(self) -> Dict:
        return {
            "file": self.file,
            "jenis": self.jenis,
            "label": self.label,
            "confidence": round(self.confidence, 2),
            "alasan": self.alasan,
            "sheet_cocok": self.sheet_cocok,
        }


# ============================================================
# SIGNATURE / POLA PENANDA PER JENIS FILE
# ============================================================
# Nama sheet dicek case-insensitive & partial match ("contains").
# Ini bagian PALING KUAT karena nama sheet jarang berubah antar outlet/PT.
SIGNATURE_SHEET_NAMES: Dict[str, List[str]] = {
    JENIS_SALES_JOURNAL: [
        "REKAP SALES ESB",
        "INPUT XERO",
        "REKAP UNPAID",
        "PHR ",  # mis. "PHR Plaza", "PHR sibarita"
    ],
    JENIS_CASH_RECONCILIATION: [
        "UANG MASUK",
    ],
    JENIS_POS_TRANSACTION: [
        "REPORT TRANSACTION",
    ],
}

# Header kolom penanda, dicek di beberapa baris pertama tiap sheet
# (case-insensitive). Dipakai sebagai sinyal TAMBAHAN saat nama sheet
# saja belum cukup meyakinkan (mis. sheet "SALES" terlalu generik).
SIGNATURE_HEADERS: Dict[str, List[str]] = {
    JENIS_SALES_JOURNAL: [
        "sales number", "bill number", "menu category",
        "paid/unpaid", "service charge", "dpp", "net revenue",
    ],
    JENIS_CASH_RECONCILIATION: [
        "data esb", "data payment", "payment method name", "selisih",
        "uang masuk",
    ],
    JENIS_POS_TRANSACTION: [
        "gross sales", "receipt number", "collected by", "served by",
        "sum of gross sales", "row labels", "column labels",
    ],
}

MAX_BARIS_HEADER_DICEK = 10  # cek header cuma di N baris pertama tiap sheet
MAX_KOLOM_HEADER_DICEK = 30  # dan N kolom pertama, biar tidak lambat


# ============================================================
# FUNGSI UTAMA
# ============================================================

def deteksi_jenis_file(path: Union[str, Path]) -> HasilDeteksi:
    """
    Deteksi jenis file Excel (Jurnal Penjualan / Rekonsiliasi Kas Masuk /
    Rekap Transaksi POS) berdasarkan struktur sheet & header-nya.

    Args:
        path: path ke file .xlsx

    Returns:
        HasilDeteksi berisi jenis file, tingkat keyakinan (0-1), dan alasan.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"File tidak ditemukan: {path}")

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"❌ Gagal membuka {path.name}: {e}")
        return HasilDeteksi(
            file=path.name, jenis=JENIS_TIDAK_DIKENALI,
            label=LABEL_JENIS[JENIS_TIDAK_DIKENALI], confidence=0.0,
            alasan=[f"Gagal membuka file: {e}"],
        )

    # [FIX] Sebelumnya wb.close() dipanggil SETELAH _kumpulkan_header_teks()
    # tanpa try/finally -- kalau fungsi itu (atau pencocokan di bawahnya)
    # melempar exception, workbook tidak pernah ditutup -> file handle
    # bocor, apalagi ini dipanggil per-upload jadi bisa terakumulasi kalau
    # ada banyak file dengan struktur aneh yang bikin error.
    try:
        sheet_names_upper = [s.upper() for s in wb.sheetnames]
        header_teks_per_sheet = _kumpulkan_header_teks(wb)

        skor: Dict[str, float] = {j: 0.0 for j in SIGNATURE_SHEET_NAMES}
        alasan: Dict[str, List[str]] = {j: [] for j in SIGNATURE_SHEET_NAMES}
        sheet_cocok: Dict[str, List[str]] = {j: [] for j in SIGNATURE_SHEET_NAMES}

        # --- Cocokkan nama sheet (bobot besar: 2 poin per sheet name cocok) ---
        for jenis, pola_list in SIGNATURE_SHEET_NAMES.items():
            for sheet_name, sheet_name_upper in zip(wb.sheetnames, sheet_names_upper):
                for pola in pola_list:
                    if pola.upper() in sheet_name_upper:
                        skor[jenis] += 2.0
                        alasan[jenis].append(f"Nama sheet '{sheet_name}' cocok pola '{pola}'")
                        sheet_cocok[jenis].append(sheet_name)
                        break

        # --- Cocokkan header kolom (bobot lebih kecil: 1 poin per kata kunci) ---
        for jenis, kata_kunci_list in SIGNATURE_HEADERS.items():
            for sheet_name, teks in header_teks_per_sheet.items():
                for kata_kunci in kata_kunci_list:
                    if kata_kunci in teks:
                        skor[jenis] += 1.0
                        alasan[jenis].append(f"Header di sheet '{sheet_name}' memuat '{kata_kunci}'")
    finally:
        wb.close()

    # --- Pilih jenis dengan skor tertinggi ---
    jenis_terpilih = max(skor, key=skor.get)
    skor_tertinggi = skor[jenis_terpilih]

    if skor_tertinggi <= 0:
        return HasilDeteksi(
            file=path.name, jenis=JENIS_TIDAK_DIKENALI,
            label=LABEL_JENIS[JENIS_TIDAK_DIKENALI], confidence=0.0,
            alasan=["Tidak ada pola nama sheet atau header yang cocok"],
        )

    # Normalisasi confidence: skor 6+ dianggap sangat yakin (1.0).
    confidence = min(1.0, skor_tertinggi / 6.0)

    hasil = HasilDeteksi(
        file=path.name,
        jenis=jenis_terpilih,
        label=LABEL_JENIS[jenis_terpilih],
        confidence=confidence,
        alasan=alasan[jenis_terpilih],
        sheet_cocok=sheet_cocok[jenis_terpilih],
    )

    logger.info(
        f"🔍 Deteksi {path.name}: {hasil.label} "
        f"(confidence={confidence:.2f}, skor={skor_tertinggi})"
    )
    return hasil


def deteksi_banyak_file(paths: List[Union[str, Path]]) -> List[HasilDeteksi]:
    """Deteksi jenis untuk banyak file sekaligus (mis. saat batch upload)."""
    hasil = []
    for p in paths:
        try:
            hasil.append(deteksi_jenis_file(p))
        except FileNotFoundError as e:
            logger.warning(f"⚠️ {e}")
    return hasil


# ============================================================
# HELPER INTERNAL
# ============================================================

def _kumpulkan_header_teks(wb) -> Dict[str, str]:
    """
    Kumpulkan semua teks di N baris & M kolom pertama tiap sheet, digabung
    jadi satu string lowercase per sheet (dipakai untuk cek kata kunci).
    """
    hasil: Dict[str, str] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        potongan = []
        for i, row in enumerate(ws.iter_rows(max_row=MAX_BARIS_HEADER_DICEK,
                                              max_col=MAX_KOLOM_HEADER_DICEK,
                                              values_only=True)):
            for cell in row:
                if cell is not None:
                    potongan.append(str(cell))
        hasil[sheet_name] = " ".join(potongan).lower()
    return hasil


# ============================================================
# CATATAN INTEGRASI
# ============================================================
# Modul ini berdiri sendiri (tidak bergantung pada detektor "jurnal koran"
# yang sudah ada). Untuk menggabungkannya ke alur deteksi yang sudah ada
# (mis. di app.py, sebelum file diproses ke akuntansi_ai.proses_dataframe):
#
#   from modules.file_detector import deteksi_jenis_file, JENIS_TIDAK_DIKENALI
#
#   hasil = deteksi_jenis_file(uploaded_path)
#   if hasil.jenis == JENIS_TIDAK_DIKENALI:
#       # fallback ke detektor lama (5 jenis / jurnal koran)
#       hasil_lama = deteksi_jenis_file_lama(uploaded_path)
#       ...
#   else:
#       st.info(f"Terdeteksi: {hasil.label} (confidence {hasil.confidence:.0%})")