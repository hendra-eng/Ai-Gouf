"""
modules/kertas_kerja.py
=========================
[BARU] Generator "Kertas Kerja Laporan Keuangan" (working paper) dalam
format Excel 14-sheet, dibangun DI ATAS pipeline rekening koran yang
sudah ada di akuntansi_ai.py (parse_sheet_bank, muat_workbook,
proses_dataframe, kategorikan_dengan_ai, dst) -- modul ini TIDAK
mengubah fungsi-fungsi itu sama sekali, murni menambah lapisan baru:

    PDF rekening koran (banyak file/bulan/bank)
        -> GL per transaksi (klasifikasi + metadata source PDF/halaman)
        -> Bank_Control (rekonsiliasi bulanan)
        -> Bank_Posting_Summary (rekap per akun x bulan)
        -> Adjustments (template kosong utk koreksi manual akuntan)
        -> TB_Monthly (Bank_Posting_Summary + Adjustments)
        -> BS_Monthly / PNL_Monthly (dari TB_Monthly + COA FS Group)
        -> workbook Kertas_Kerja_Laporan_Keuangan_<tahun>.xlsx

Lalu ada fungsi terbalik untuk membaca kertas kerja yang SUDAH DIKOREKSI
user (terutama sheet Adjustments) dan menyusun ulang TB/BS/PNL dari situ,
supaya proses generate 18-sheet report bisa memakai data yang sudah
direview akuntan, bukan hasil klasifikasi mentah.

CATATAN INTEGRASI / ASUMSI YANG PERLU DICEK OLEH TIM:
------------------------------------------------------
1. Import `akuntansi_ai` di bawah ini mengasumsikan modul itu importable
   sebagai top-level module (karena akuntansi_ai.py ada di root
   backend/, sejajar main.py, bukan di dalam modules/). Kalau di
   deployment kamu ternyata modules/ dan root TIDAK sama-sama ada di
   sys.path, sesuaikan baris import di bawah.

2. AKUN STAGING (Initial Staging Account) ditentukan HANYA dari arah
   transaksi (CR/DB), bukan dari confidence -- ini hasil verifikasi
   langsung ke contoh file kamu (Kertas_Kerja_Laporan_Keuangan_2025.xlsx):
   CR -> akun staging liability (di contohmu: 2101 "Customer Deposit"),
   DB -> akun staging asset (di contohmu: 1105 "Prepaid Expenses").
   Akun mana yang dipakai per client DIAMBIL OTOMATIS dari sheet COA:
   cari akun yang kolom "Notes"/catatannya mengandung kata "staging"
   (case-insensitive), lalu dipisah CR/DB berdasarkan "Normal Balance"
   (Kredit -> staging CR, Debit -> staging DB). Kalau COA client tidak
   punya baris yang ditandai begitu, fungsi akan angkat error yang jelas
   -- akuntan WAJIB menandai 2 akun staging ini di COA client sebelum
   proses jalan (lihat `cari_akun_staging_dari_coa`).

3. BS_Monthly / PNL_Monthly memakai kolom "Statement" (BS/PNL) dan
   "FS Group" di sheet COA contohmu untuk mengelompokkan akun. Kalau COA
   real client tidak punya 2 kolom itu, tambahkan dulu (lihat struktur
   contoh di `_KOLOM_COA_DIBUTUHKAN`).

4. Ekstraksi PDF di sini SENGAJA tidak memakai
   akuntansi_ai._baca_pdf_sebagai_lembar() apa adanya, karena fungsi itu
   membuang informasi nama file & nomor halaman (jadi cuma nama sheet
   sementara "Halaman N - Tabel M"). Di sini kita pakai pdfplumber
   langsung supaya (nama_file_pdf, halaman) ikut nempel di tiap baris
   transaksi, karena GL sheet kertas kerja WAJIB punya 2 kolom itu.
   Parsing tabelnya sendiri tetap dari logic yang sama level
   kompleksitasnya dengan _baca_pdf_sebagai_lembar (pdfplumber
   extract_tables) -- HANYA untuk PDF hasil export digital (bukan hasil
   scan/foto). PDF hasil scan perlu OCR terpisah (lihat
   tax_pdf_extractor.py sbg referensi kalau perlu ditambahkan nanti).
"""
from __future__ import annotations

import concurrent.futures
import hashlib
import io
import json
import os
import pickle
import threading
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# --- Sesuaikan baris ini kalau layout project ternyata beda (lihat
#     catatan integrasi #1 di atas). Coba beberapa kemungkinan supaya
#     modul ini tetap jalan walau strukturnya sedikit beda. ---
try:
    import akuntansi_ai
except ImportError:  # pragma: no cover
    try:
        from .. import akuntansi_ai  # kalau modules/ adalah sub-package
    except ImportError:
        from . import akuntansi_ai  # kalau akuntansi_ai ikut dipindah ke modules/

# [BARU] Client Claude API, dipakai opsional oleh
# perbaiki_gl_dengan_claude_review() di bawah (review baris GL confidence
# non-High). Sama pola try/except bertingkat seperti import akuntansi_ai
# di atas -- modul ini tetap bisa di-import berdiri sendiri walau
# claude_client.py belum ada/belum di path yang sama.
# Pakai panggil_claude_terstruktur() (bukan panggil_claude_teks()) --
# structured output lewat tool-use, sama seperti SEMUA pemanggilan Claude
# lain di codebase ini (generate_narasi_calk_claude, dst di
# claude_client.py) -- Claude "dipaksa" balas sesuai skema, jadi tidak
# perlu strip markdown fence / json.loads manual / tangani JSON tidak valid.
try:
    from .claude_client import panggil_claude_terstruktur, ClaudeError
except ImportError:  # pragma: no cover
    try:
        from claude_client import panggil_claude_terstruktur, ClaudeError
    except ImportError:
        panggil_claude_terstruktur = None
        ClaudeError = Exception

# [BARU] Logger modul sendiri (dipakai antara lain oleh cache jalur PDF di
# bawah) -- sebelumnya modul ini tidak punya logger sama sekali. Pola import
# sama seperti modules/file_detector.py & modules/ai_analysis.py; dibungkus
# try/except supaya modul ini tetap bisa jalan berdiri sendiri (mis. dites
# di luar package modules/) kalau logging_config tidak ikut ter-import.
try:
    from .logging_config import get_module_logger
    logger = get_module_logger("kertas_kerja")
except ImportError:  # pragma: no cover
    import logging
    logger = logging.getLogger("kertas_kerja")


BULAN_URUT = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
BULAN_NO = {b: i + 1 for i, b in enumerate(BULAN_URUT)}


def _angka(v) -> float:
    """[FIX -- skills/xlsx_export] Konversi nilai ke float dengan aman --
    None/NaN/inf semua dianggap 0.0. Pola lama `float(x or 0)` TIDAK aman
    untuk NaN: `float('nan') or 0` mengembalikan `nan` itu sendiri (NaN
    truthy di Python), jadi fallback "or 0"-nya tidak kepakai -- bisa
    menulis NaN langsung ke sel Excel atau merusak akumulasi total
    (total += nan -> total jadi nan seterusnya). Sengaja didefinisikan
    lokal di modul ini (bukan import dari accounting_export.py) supaya
    kertas_kerja.py tetap berdiri sendiri tanpa dependensi silang baru --
    lihat catatan integrasi di docstring atas file ini soal modul ini
    yang sengaja tidak bergantung ke accounting_export.py."""
    if v is None:
        return 0.0
    try:
        f = float(v)
    except (TypeError, ValueError):
        return 0.0
    if pd.isna(f) or f in (float("inf"), float("-inf")):
        return 0.0
    return f

# Kolom COA yang dibutuhkan modul ini (di luar no_akun/nama_akun standar
# yang sudah dipakai akuntansi_ai.py). Kalau COA client belum punya ini,
# BS_Monthly/PNL_Monthly & pencarian akun staging tidak akan berfungsi.
_KOLOM_COA_DIBUTUHKAN = ["Normal Balance", "Statement", "FS Group", "Notes"]

# [BARU -- PERBAIKAN PERFORMA TAHAP 2] Jumlah file PDF yang diekstrak PARALEL
# sekaligus (lihat susun_gl_dari_pdf_rekening_koran di bawah). Ekstraksi PDF
# (pdfplumber/pdf2image+pytesseract) itu campuran I/O (baca file, panggil
# binary poppler/tesseract lewat subprocess) dan kerja C-extension (zlib
# decompress, decode gambar) yang MELEPAS GIL Python -- jadi ThreadPoolExecutor
# tetap memberi percepatan nyata untuk beban ini walau bukan murni I/O-bound,
# TANPA masalah "file_like tidak bisa di-pickle" yang akan muncul kalau pakai
# ProcessPoolExecutor (file upload dari FastAPI/Streamlit biasanya berupa
# objek stream, bukan path, jadi tidak bisa dikirim ke proses lain).
# Default 4 -- dibuat overridable lewat env var supaya bisa dituning per
# server (mis. dinaikkan kalau CPU banyak, diturunkan kalau memori terbatas
# karena tiap thread bisa menahan 1 PDF penuh + gambar hasil rasterisasi OCR
# di memori sekaligus).
PDF_PARALEL_MAKS = int(os.environ.get("KERTAS_KERJA_PDF_PARALEL_MAKS", "4"))

# [BARU -- PERBAIKAN PERFORMA TAHAP 2] Cache hasil ekstraksi PDF PER FILE
# (bukan per nama-bank seperti _NAMA_DASAR_CACHE_JALUR_PDF di bawah).
# MASALAH: cache jalur (grid vs posisi-kata) mempercepat MEMILIH strategi
# yang benar, tapi tiap file tetap diekstrak PENUH dari nol setiap kali
# fungsi ini dipanggil -- kalau proses gagal belakangan (mis. COA client
# belum lengkap, staging account belum ditandai) dan user retry generate
# kertas kerja dengan PDF yang SAMA, seluruh ekstraksi PDF (bagian paling
# lambat) diulang dari nol walau isinya identik. Cache ini menyimpan hasil
# ekstraksi (df + peringatan + ringkasan_footer + jalur yang dipakai) per
# hash konten file (md5) + nama bank, jadi retry dengan file yang sama
# hampir instan (tinggal baca dari disk, skip pdfplumber/OCR sama sekali).
_FOLDER_CACHE_EKSTRAKSI_PDF = Path(__file__).parent / "cache_ekstraksi_pdf"


def _path_cache_ekstraksi_pdf(hash_file: str, client_id: Optional[int] = None) -> str:
    _FOLDER_CACHE_EKSTRAKSI_PDF.mkdir(exist_ok=True)
    prefix = f"client_{client_id}" if client_id else "global"
    return str(_FOLDER_CACHE_EKSTRAKSI_PDF / f"{prefix}__{hash_file}.pkl")


def _hash_isi_file(file_like) -> Optional[str]:
    """Hash md5 dari SELURUH isi file (bukan cuma nama file -- nama file
    bisa sama walau isi beda, mis. re-download statement yang sudah
    direvisi bank). Mengembalikan None (bukan melempar error) kalau
    file_like tidak mendukung read/seek -- caller HARUS menganggap itu
    sebagai cache-miss biasa, bukan kegagalan, supaya ekstraksi tetap
    bisa jalan tanpa cache untuk kasus stream yang tidak seekable."""
    try:
        posisi_awal = file_like.tell()
    except (AttributeError, OSError):
        posisi_awal = None
    try:
        file_like.seek(0)
        isi = file_like.read()
        h = hashlib.md5(isi).hexdigest()
        return h
    except (AttributeError, OSError, TypeError):
        return None
    finally:
        try:
            file_like.seek(posisi_awal if posisi_awal is not None else 0)
        except (AttributeError, OSError):
            pass


def _muat_cache_ekstraksi_pdf(path: str) -> Optional[Dict[str, Any]]:
    """Baca cache hasil ekstraksi 1 file PDF dari disk. Return None kalau
    belum ada/korup/format lama -- caller memperlakukan itu sebagai
    cache-miss (ekstraksi ulang seperti biasa), TIDAK PERNAH melempar error
    ke pemanggil supaya cache murni optimasi, tidak bisa menjatuhkan proses."""
    if not os.path.exists(path):
        return None
    try:
        with open(path, "rb") as f:
            data = pickle.load(f)
        if not isinstance(data, dict) or "df" not in data:
            return None
        # [BARU -- TTL/LRU] Sentuh mtime saat cache-hit, supaya modules/
        # cache_cleanup.py (yang memakai mtime sebagai proxy "terakhir
        # diakses" utk TTL & LRU) tidak salah menghapus file yang justru
        # SERING dipakai tapi kebetulan jarang DITULIS ulang. Gagal touch
        # (mis. permission, race dgn proses lain) tidak fatal -- cache
        # tetap valid dipakai, cuma mtime-nya jadi kurang akurat.
        try:
            os.utime(path, None)
        except OSError:
            pass
        return data
    except Exception as e:  # noqa: BLE001 -- pickle korup bisa lempar banyak jenis error
        logger.warning(f"Gagal baca cache ekstraksi PDF dari {path}: {e} -- dianggap cache-miss.")
        return None


def _simpan_cache_ekstraksi_pdf(data: Dict[str, Any], path: str) -> None:
    """Simpan hasil ekstraksi 1 file PDF ke disk (atomic write, pola sama
    dengan _simpan_cache_jalur_pdf/simpan_pola -- tulis ke .tmp dulu lalu
    os.replace(), supaya proses yang ke-interupsi di tengah tidak pernah
    meninggalkan file cache korup)."""
    try:
        tmp_path = path + ".tmp"
        with open(tmp_path, "wb") as f:
            pickle.dump(data, f, protocol=pickle.HIGHEST_PROTOCOL)
        os.replace(tmp_path, path)
    except Exception as e:  # noqa: BLE001
        # Gagal simpan cache TIDAK BOLEH menjatuhkan proses generate kertas
        # kerja yang datanya sudah berhasil diekstrak -- cukup dicatat.
        logger.warning(f"Gagal menyimpan cache ekstraksi PDF ke {path}: {e} -- dilewati, tidak fatal.")

WARNA_INPUT_MANUAL = "FFDDEBF7"  # biru muda -- konvensi "sel input user" di contoh file
FONT_HEADER = Font(bold=True, size=11)
FONT_JUDUL = Font(bold=True, size=13)


# ============================================================
# 0. STRUKTUR DATA HASIL
# ============================================================

@dataclass
class HasilKertasKerja:
    gl: pd.DataFrame
    bank_control: pd.DataFrame
    bank_posting_summary: pd.DataFrame
    adjustments: pd.DataFrame
    tb_monthly: pd.DataFrame
    bs_monthly: pd.DataFrame
    pnl_monthly: pd.DataFrame
    coa: pd.DataFrame
    peringatan: List[str] = field(default_factory=list)
    status_ai: pd.DataFrame = field(default_factory=pd.DataFrame)
    # [BARU] Hasil validasi otomatis struktur & konsistensi angka (lihat
    # bagian "3.9 VALIDASI HASIL AKHIR" di bawah) -- list of dict
    # {"level": "error"|"warning", "area": str, "pesan": str}. Diisi
    # OTOMATIS di akhir generate_kertas_kerja(), TIDAK perlu flag khusus
    # (murni kode, tidak ada panggilan API, jadi aman selalu jalan).
    temuan_validasi: List[Dict[str, Any]] = field(default_factory=list)
    # [BARU] Narasi penjelasan/prioritas dari Claude atas temuan_validasi
    # di atas -- HANYA diisi kalau generate_kertas_kerja(pakai_claude_review_final=True)
    # DAN ada temuan yang levelnya "error" (lihat jelaskan_temuan_kertas_kerja_claude
    # di claude_client.py). Kosong ("") kalau opsi ini mati atau tidak ada
    # temuan yang perlu dijelaskan -- TIDAK ada panggilan API dalam kasus itu.
    catatan_review_final: str = ""


# ============================================================
# 0.5 [BARU -- PERBAIKAN PERFORMA] CACHE JALUR EKSTRAKSI PDF PER BANK
# ============================================================
# MASALAH: susun_gl_dari_pdf_rekening_koran() (di bawah) SELALU mencoba
# jalur "grid" dulu (_ekstrak_tabel_pdf_dengan_metadata -> pdfplumber
# page.extract_tables() di SEMUA halaman) untuk SETIAP file PDF, walau
# formatnya sudah terbukti tidak punya garis tabel (mis. BCA Tahapan --
# lihat catatan panjang di _ekstrak_pdf_berbasis_posisi di bawah). Untuk
# format begini, jalur grid PASTI gagal/0 baris, lalu PDF yang SAMA dibuka
# & diparse ULANG dari awal lewat jalur fallback posisi-kata -- artinya
# tiap file PDF format ini diproses 2x penuh, padahal hasil jalur grid
# sudah bisa ditebak dari pengalaman file sebelumnya dengan nama bank yang
# sama.
#
# PERBAIKAN: simpan "jalur yang terbukti berhasil" per nama_bank ke disk
# (mirip pola historis client -- pakai folder & skema penyimpanan yang
# SAMA lewat akuntansi_ai._path_pola(), supaya tidak menambah folder/skema
# baru). Begitu 1 file dari bank tertentu pernah ketahuan butuh jalur
# posisi-kata, SEMUA file berikutnya dari bank yang sama (bulan lain, client
# lain yang pakai bank sama) langsung lompat ke jalur posisi-kata tanpa
# buang waktu coba extract_tables() dulu. Kalau ternyata suatu saat file
# dari bank itu GAGAL di jalur posisi-kata (mis. bank ganti format), fungsi
# tetap mencatat peringatan seperti biasa -- cache ini murni optimasi
# urutan-coba, bukan validasi/gate yang bisa bikin file gagal total.
_NAMA_DASAR_CACHE_JALUR_PDF = "jalur_pdf_bank"


def _muat_cache_jalur_pdf(path: str) -> Dict[str, str]:
    """Baca cache {nama_bank: 'grid'|'posisi_kata'} dari disk.
    Return {} kalau file belum ada/korup (aman -- berarti semua bank
    dicoba dari jalur grid dulu seperti perilaku lama, tidak ada yang
    rusak, cuma belum optimal)."""
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except (json.JSONDecodeError, OSError) as e:
        logger.warning(f"Gagal baca cache jalur PDF dari {path}: {e} -- dianggap kosong.")
        return {}


def _simpan_cache_jalur_pdf(cache: Dict[str, str], path: str) -> None:
    """Simpan cache ke disk (atomic write, sama pola dengan simpan_pola()/
    simpan_histori_gaji() di akuntansi_ai.py -- tulis ke .tmp dulu lalu
    os.replace(), supaya proses yang ke-interupsi di tengah tidak pernah
    meninggalkan file cache yang korup/setengah tertulis)."""
    tmp_path = path + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, path)


# ============================================================
# 1. EKSTRAKSI PDF REKENING KORAN DENGAN METADATA SOURCE/PAGE
# ============================================================

def _ekstrak_tabel_pdf_dengan_metadata(file_like, nama_file_pdf: str) -> List[Dict[str, Any]]:
    """
    Sama seperti akuntansi_ai._baca_pdf_sebagai_lembar() tapi TIDAK
    membuang info halaman -- tiap baris tabel mentah dikembalikan
    sebagai dict {"halaman": int, "baris": tuple}. Header/deteksi kolom
    tetap dilakukan belakangan (memakai _cari_header_row dkk milik
    akuntansi_ai.py, dipanggil dari _susun_gl_dari_pdf).
    """
    try:
        import pdfplumber
    except ImportError as e:
        raise RuntimeError(
            "Library 'pdfplumber' belum terinstall. Jalankan: "
            "pip install pdfplumber --break-system-packages"
        ) from e

    hasil: List[Dict[str, Any]] = []
    with pdfplumber.open(file_like) as pdf:
        for i_hal, page in enumerate(pdf.pages, start=1):
            for tabel in page.extract_tables():
                if not tabel:
                    continue
                for row in tabel:
                    hasil.append({
                        "halaman": i_hal,
                        "source_pdf": nama_file_pdf,
                        "baris": tuple(sel if sel is not None else "" for sel in row),
                    })

    if not hasil:
        raise RuntimeError(
            f"Tidak ada tabel terdeteksi di '{nama_file_pdf}' -- kemungkinan PDF hasil "
            "scan/foto (bukan teks asli), atau tabel tidak berbentuk grid yang "
            "terbaca otomatis. File ini dilewati."
        )
    return hasil


# ============================================================
# 1.1 [BARU] EKSTRAKSI BERBASIS POSISI KATA -- FALLBACK UNTUK PDF
#     REKENING KORAN TANPA GARIS TABEL (mis. format BCA "Tahapan")
# ============================================================
# Beberapa rekening koran (contoh: BCA Tahapan) TIDAK punya garis grid
# antar-baris transaksi -- hanya header kolom yang dibingkai kotak,
# baris data di bawahnya murni teks sejajar kolom tanpa garis sama
# sekali. pdfplumber.extract_tables() (dipakai di
# _ekstrak_tabel_pdf_dengan_metadata di atas & di
# akuntansi_ai._baca_pdf_sebagai_lembar) mengandalkan garis grid untuk
# mendeteksi baris/kolom, sehingga PDF jenis ini GAGAL total diekstrak
# (cuma header yang kebaca, 0 baris transaksi).
#
# Fungsi-fungsi di bawah ini adalah jalur ekstraksi ALTERNATIF berbasis
# KOORDINAT kata (x0/top dari pdfplumber.extract_words()), dipakai
# sebagai FALLBACK otomatis oleh susun_gl_dari_pdf_rekening_koran() saat
# jalur grid biasa gagal/menghasilkan terlalu sedikit baris. Didesain
# generik (posisi kolom dicari dinamis dari kata header per halaman,
# bukan koordinat hardcode) supaya tidak cuma jalan untuk 1 template BCA
# spesifik, tapi bank/format lain dengan pola serupa (header berbaris +
# data tanpa garis, 1 kolom MUTASI gabungan dengan suffix DB/CR atau
# DB/K) juga punya peluang terbaca.
#
# Divalidasi terhadap rekening koran BCA Tahapan 25 halaman: total &
# jumlah baris hasil ekstraksi PERSIS cocok dengan ringkasan resmi di
# halaman terakhir statement (MUTASI CR/MUTASI DB/jumlah baris).

# ============================================================
# [DEDUP] Fungsi cluster-kata/header/tahun/float/ekstraksi-baris posisi
# SEBELUMNYA ada 2 salinan: versi aktif & lebih baru di akuntansi_ai.py
# (nama fungsi bersuffix _pdf/_rk, sudah termasuk validasi otomatis
# _validasi_ringkasan_footer_rk) DAN salinan lawas persis di sini yang
# tidak pernah dihapus. Duplikasi ini berisiko drift diam-diam (bug
# diperbaiki di satu tempat, lupa di tempat lain) -- sudah terjadi:
# akuntansi_ai.py sempat dapat perbaikan validasi footer duluan, versi
# di sini tidak ikut dapat. Sekarang _ekstrak_pdf_berbasis_posisi()
# HANYA jadi wrapper tipis ke akuntansi_ai._ekstrak_pdf_rekening_koran_berbasis_posisi()
# -- satu sumber kebenaran, kedua jalur (upload umum & Kertas Kerja)
# otomatis dapat perbaikan yang sama tanpa perlu disinkronkan manual.
# ============================================================

def _ekstrak_pdf_berbasis_posisi(
    file_like, nama_file_pdf: str, nama_bank: str,
) -> Tuple[pd.DataFrame, List[str], Dict[str, Any]]:
    """Titik masuk fallback posisi-kata untuk 1 file PDF rekening koran
    -- dipanggil oleh susun_gl_dari_pdf_rekening_koran() HANYA saat jalur
    grid (_ekstrak_tabel_pdf_dengan_metadata + akuntansi_ai.parse_sheet_bank)
    gagal/tidak menghasilkan baris.

    [DEDUP] Sejak sekarang cuma wrapper ke
    akuntansi_ai._ekstrak_pdf_rekening_koran_berbasis_posisi() -- lihat
    catatan section di atas. Signature & return (df, peringatan,
    ringkasan_footer) TIDAK berubah, jadi susun_gl_dari_pdf_rekening_koran()
    di bawah tidak perlu diubah sama sekali.
    """
    return akuntansi_ai._ekstrak_pdf_rekening_koran_berbasis_posisi(
        file_like, nama_file_pdf, nama_bank
    )




class _LembarDenganMetadata:
    """
    Adapter sekali-pakai: bungkus baris-baris dari 1 PDF (lintas halaman)
    supaya bisa dipakai apa adanya oleh akuntansi_ai.parse_sheet_bank()
    (yang cuma butuh method iter_rows), TAPI kita simpan juga
    (halaman, source_pdf) per baris di luar, supaya sesudah
    parse_sheet_bank() selesai kita bisa mencocokkan baris hasil parse
    dengan metadata sumbernya lewat index posisi baris (urutan tidak
    diubah oleh parse_sheet_bank, jadi index tetap valid).
    """

    def __init__(self, baris_mentah: List[Tuple]):
        self._baris = baris_mentah

    def iter_rows(self, min_row: int = 1, max_row: Optional[int] = None, values_only: bool = True):
        akhir = max_row if max_row is not None else len(self._baris)
        for row in self._baris[min_row - 1: akhir]:
            yield row


def _status_dari_hasil_ekstraksi(hasil: Dict[str, Any]) -> Tuple[str, Optional[str]]:
    """[BARU] Turunkan status "done"|"cache_hit"|"error" (dipakai
    progress_callback per-file) dari dict hasil _proses_satu_file. "error"
    hanya kalau df gagal terbentuk sama sekali -- kalau df ada tapi ada
    peringatan (mis. mojibake/OCR dicurigai), tetap dianggap "done";
    peringatan lengkapnya tetap ikut di `hasil.peringatan` gabungan seperti
    biasa, bukan di sini (di sini cuma pesan singkat status per-file)."""
    if hasil.get("dari_cache"):
        return "cache_hit", None
    if hasil.get("df") is not None and not hasil["df"].empty:
        return "done", None
    pesan = hasil["peringatan"][-1] if hasil.get("peringatan") else "Gagal diekstrak."
    return "error", pesan


def susun_gl_dari_pdf_rekening_koran(
    daftar_file_pdf: List[Tuple[Any, str]],
    nama_bank_per_file: Optional[Dict[str, str]] = None,
    client_id: Optional[int] = None,
    progress_callback: Optional[Any] = None,
) -> Tuple[pd.DataFrame, List[str], Dict[str, Dict[str, Any]]]:
    """
    Args:
        daftar_file_pdf: list of (file_like, nama_file_pdf) -- bisa lebih
            dari satu file (multi bulan/bank), akan digabung jadi satu.
        nama_bank_per_file: optional {nama_file_pdf: nama_bank}. Kalau
            tidak diisi, nama bank diambil dari nama file (tanpa ekstensi).
        client_id: [BARU -- PERBAIKAN PERFORMA] dipakai untuk memuat/
            menyimpan cache "jalur ekstraksi PDF per nama bank" (grid vs
            posisi-kata) -- lihat _muat_cache_jalur_pdf/_simpan_cache_jalur_pdf
            di atas. Opsional: kalau None, cache tetap dipakai tapi versi
            global (dibagi semua client), sama seperti perilaku pola_bank.
            [BARU -- TAHAP 2] Juga dipakai untuk namespace cache hasil
            ekstraksi PER FILE (_path_cache_ekstraksi_pdf) -- supaya file
            dengan isi sama dari client BERBEDA tidak saling tertukar cache.
        progress_callback: [BARU] opsional, callable(nama_file_pdf: str,
            status: str, pesan: Optional[str]) dipanggil dari thread worker
            (HARUS thread-safe -- caller di main.py memakai queue.Queue.put,
            yang thread-safe) tiap kali status ekstraksi 1 file berubah:
            "processing" di awal, lalu "done"/"cache_hit"/"error" di akhir.
            Dibungkus try/except di titik panggil supaya callback yang
            error TIDAK ikut menggagalkan ekstraksi filenya sendiri.

    [BARU -- PERBAIKAN PERFORMA TAHAP 2] Fungsi ini sekarang memproses
    setiap file PDF PARALEL (lihat PDF_PARALEL_MAKS, default 4 file
    sekaligus, dituning lewat env var KERTAS_KERJA_PDF_PARALEL_MAKS) dan
    melewati SELURUH pekerjaan ekstraksi kalau file yang PERSIS sama
    (hash isi + nama bank) pernah diekstrak sebelumnya (cache per-file).
    Urutan baris hasil akhir tetap konsisten dengan urutan `daftar_file_pdf`
    -- perilaku/urutan tidak berubah dari versi sekuensial, cuma lebih cepat.

    Returns:
        (df_gabungan, peringatan, ringkasan_per_file) -- df_gabungan punya
        kolom standar parse_sheet_bank() (no, bank, tanggal, keterangan,
        mutasi_debet, mutasi_kredit, saldo, ...) DITAMBAH 2 kolom baru:
        source_pdf, halaman. [BARU] ringkasan_per_file = {nama_file_pdf:
        {saldo_awal, mutasi_cr, mutasi_db, saldo_akhir}} -- dari blok
        ringkasan resmi di footer tiap statement (hanya untuk file yang
        lewat jalur fallback posisi-kata; lihat catatan di jalur grid di
        bawah), dipakai generate_kertas_kerja() utk mengisi kolom
        "CR/DB Statement"/"Closing" (statement resmi) di Bank_Control.

    File yang gagal diekstrak (PDF hasil scan, dll.) TIDAK menghentikan
    proses -- dicatat di `peringatan` dan dilewati, supaya 1 PDF rusak
    tidak menggagalkan seluruh batch upload multi-bulan.
    """
    nama_bank_per_file = nama_bank_per_file or {}

    # [BARU -- PERBAIKAN PERFORMA] Cache jalur ekstraksi per nama bank --
    # lihat penjelasan lengkap di blok "0.5" di atas file ini. Dibaca SEKALI
    # di depan (bukan per-thread) supaya semua worker paralel di bawah mulai
    # dari snapshot yang sama; perubahan dari 1 file (mis. bank baru ketahuan
    # butuh posisi-kata) TIDAK mempengaruhi keputusan file lain yang jalan
    # BERSAMAAN di batch yang sama (aman -- keputusan itu toh independen per
    # bank, cuma optimasi urutan-coba, bukan korektnes), lalu digabung balik
    # & disimpan ke disk sesudah semua worker selesai.
    path_cache_jalur = akuntansi_ai._path_pola(_NAMA_DASAR_CACHE_JALUR_PDF, client_id)
    cache_jalur = _muat_cache_jalur_pdf(path_cache_jalur)
    lock_cache_jalur = threading.Lock()

    def _proses_satu_file(file_like, nama_file_pdf: str) -> Dict[str, Any]:
        """
        [BARU -- PERBAIKAN PERFORMA TAHAP 2] Isi 1 iterasi loop LAMA,
        dipindah jadi fungsi tersendiri supaya bisa dijalankan PARALEL per
        file lewat ThreadPoolExecutor di bawah (lihat PDF_PARALEL_MAKS) --
        logic ekstraksi (jalur grid/fallback posisi-kata/cache jalur per
        bank) SAMA PERSIS dengan sebelumnya, TIDAK diubah, cuma dipindah
        supaya reusable per-thread. Ditambah 1 lapisan baru di paling awal:
        cache hasil ekstraksi PER FILE (lihat _muat_cache_ekstraksi_pdf di
        atas) -- kalau file PERSIS SAMA (hash md5 isi + nama bank) pernah
        diekstrak sebelumnya, hasilnya dibaca dari disk & SELURUH proses
        pdfplumber/OCR di bawah dilewati sama sekali (paling berguna saat
        user retry generate kertas kerja setelah error di tahap lain, mis.
        staging account COA belum ditandai).

        Return dict seragam: {"nama_file_pdf", "df" (DataFrame atau None),
        "peringatan" (list[str]), "ringkasan_footer" (dict atau None),
        "jalur_terpakai" (nama_bank, "grid"/"posisi_kata" atau None kalau
        gagal total)} -- HANYA membaca/menulis `cache_jalur` di bawah lock,
        aman dipanggil dari banyak thread sekaligus. TIDAK melempar
        exception ke pemanggil (semua exception ditangkap & dicatat sbg
        peringatan, PERSIS perilaku loop lama) supaya 1 file gagal/thread
        error tidak menjatuhkan file lain yang diproses paralel.
        """
        nama_bank = nama_bank_per_file.get(nama_file_pdf) or nama_file_pdf.rsplit(".", 1)[0]
        hasil: Dict[str, Any] = {
            "nama_file_pdf": nama_file_pdf, "df": None, "peringatan": [],
            "ringkasan_footer": None, "jalur_terpakai": None,
        }

        # [BARU] Lapor "processing" begitu worker benar-benar mulai mengerjakan
        # file ini (bukan cuma di-submit ke executor -- kalau semua worker
        # sibuk, file yang masih antre TIDAK dilaporkan "processing" dulu).
        if progress_callback:
            try:
                progress_callback(nama_file_pdf, "processing", None)
            except Exception:  # noqa: BLE001 -- callback gagal tidak boleh menggagalkan ekstraksi
                pass

        def _selesai(hasil: Dict[str, Any]) -> Dict[str, Any]:
            if progress_callback:
                status, pesan = _status_dari_hasil_ekstraksi(hasil)
                try:
                    progress_callback(nama_file_pdf, status, pesan)
                except Exception:  # noqa: BLE001
                    pass
            return hasil

        # --- [BARU] Cache per-file: cek dulu sebelum kerja berat apa pun ---
        hash_file = _hash_isi_file(file_like)
        path_cache_file = _path_cache_ekstraksi_pdf(hash_file, client_id) if hash_file else None
        if path_cache_file:
            cached = _muat_cache_ekstraksi_pdf(path_cache_file)
            if cached is not None and cached.get("nama_bank") == nama_bank:
                hasil["df"] = cached.get("df")
                hasil["peringatan"] = list(cached.get("peringatan") or [])
                hasil["ringkasan_footer"] = cached.get("ringkasan_footer")
                hasil["jalur_terpakai"] = cached.get("jalur_terpakai")
                hasil["dari_cache"] = True
                return _selesai(hasil)

        with lock_cache_jalur:
            jalur_diketahui = cache_jalur.get(nama_bank)

        # --- Jalur A: bank ini SUDAH terbukti butuh fallback posisi-kata ---
        if jalur_diketahui == "posisi_kata":
            try:
                file_like.seek(0)
            except (AttributeError, OSError):
                pass
            try:
                df_fallback, warn_fallback, ringkasan_footer = _ekstrak_pdf_berbasis_posisi(
                    file_like, nama_file_pdf, nama_bank
                )
            except Exception as e:  # noqa: BLE001
                hasil["peringatan"].append(
                    f"'{nama_file_pdf}': gagal diekstrak lewat jalur posisi-kata "
                    f"(dipilih dari cache bank '{nama_bank}') -- {type(e).__name__}: {e}. "
                    "File ini dilewati, file lain tetap diproses."
                )
                return _selesai(hasil)
            hasil["peringatan"].extend(warn_fallback)
            hasil["ringkasan_footer"] = ringkasan_footer or None
            hasil["jalur_terpakai"] = "posisi_kata"
            if df_fallback.empty:
                hasil["peringatan"].append(
                    f"'{nama_file_pdf}': tidak ada baris transaksi terbaca lewat jalur "
                    "posisi-kata (dipilih dari cache)."
                )
            else:
                hasil["df"] = df_fallback
            _simpan_cache_ekstraksi_pdf_helper(path_cache_file, hasil, nama_bank)
            return _selesai(hasil)

        # --- Jalur B: coba grid dulu (default) ---
        try:
            baris_dengan_meta = _ekstrak_tabel_pdf_dengan_metadata(file_like, nama_file_pdf)
        except RuntimeError as e:
            hasil["peringatan"].append(str(e))
            _simpan_cache_ekstraksi_pdf_helper(path_cache_file, hasil, nama_bank)
            return _selesai(hasil)
        except Exception as e:  # noqa: BLE001
            hasil["peringatan"].append(
                f"'{nama_file_pdf}': gagal dibaca/diekstrak ({type(e).__name__}: {e}). "
                "File ini dilewati, file lain tetap diproses."
            )
            _simpan_cache_ekstraksi_pdf_helper(path_cache_file, hasil, nama_bank)
            return _selesai(hasil)

        baris_mentah = [b["baris"] for b in baris_dengan_meta]

        # [FIX -- BUG KRUSIAL, sama persis kasus yang sudah diperbaiki di
        # akuntansi_ai._baca_pdf_sebagai_lembar(), lihat komentar panjang
        # di sana] Sebelumnya baris_mentah dari SEMUA halaman digabung
        # jadi 1 daftar lalu di-parse SEKALIGUS lewat parse_sheet_bank().
        # Untuk PDF tanpa garis grid di badan tabel (mis. BCA "Tahapan"):
        # tiap halaman cuma punya kotak kecil bergaris (info No.Rekening/
        # Periode + baris header kolom), jadi extract_tables() TETAP
        # menangkap "tabel" di tiap halaman -- tapi isinya cuma header/sel
        # kosong (STRING KOSONG "", BUKAN None), berulang 1x per halaman.
        # Karena pengecekan skip baris cuma "keterangan is None DAN
        # tanggal is None", sel string kosong "" LOLOS jadi baris
        # "transaksi" -- sehingga df_bank gabungan jadi TIDAK PERNAH
        # kosong (berisi puluhan baris sampah header/kotak berulang dari
        # tiap halaman), gagal_jalur_grid TIDAK PERNAH ke-trigger, dan
        # fallback posisi-kata (yang seharusnya menangani format ini)
        # TIDAK PERNAH dipanggil -- transaksi ASLI di seluruh PDF hilang
        # total, diam-diam diganti data sampah, sampai akhirnya gagal di
        # validasi tahap lain jauh setelah proses AI selesai (lambat +
        # akhirnya "tidak dikenali").
        #
        # Perbaikan: cek dulu PER HALAMAN (bukan gabungan) apakah grid
        # betul-betul menghasilkan baris transaksi asli -- kriteria SAMA
        # persis dgn yg dipakai akuntansi_ai._baca_pdf_sebagai_lembar
        # (parse_sheet_bank per halaman, dijumlah). Kalau totalnya 0 di
        # SEMUA halaman, langsung anggap grid gagal (skip parse gabungan
        # sama sekali) -- baru lanjut ke fallback posisi-kata di bawah.
        # Kalau ada halaman yang memang menghasilkan transaksi asli
        # (kasus rekening koran ber-grid normal), perilaku LAMA (parse
        # gabungan semua halaman sekaligus) tetap dipakai apa adanya,
        # tidak berubah.
        baris_per_halaman: Dict[int, List[Tuple]] = {}
        for b in baris_dengan_meta:
            baris_per_halaman.setdefault(b["halaman"], []).append(b["baris"])

        total_baris_transaksi_grid_per_halaman = 0
        for baris_halaman in baris_per_halaman.values():
            try:
                df_cek_halaman = akuntansi_ai.parse_sheet_bank(
                    _LembarDenganMetadata(baris_halaman), nama_bank
                )
                total_baris_transaksi_grid_per_halaman += len(df_cek_halaman)
            except akuntansi_ai.FormatTidakDikenali:
                continue
            except Exception:  # noqa: BLE001 -- pengecekan per halaman ini murni untuk deteksi, jangan sampai menghentikan proses
                continue

        gagal_jalur_grid = False
        if total_baris_transaksi_grid_per_halaman == 0:
            gagal_jalur_grid = True
            df_bank = pd.DataFrame()
        else:
            try:
                adapter = _LembarDenganMetadata(baris_mentah)
                df_bank = akuntansi_ai.parse_sheet_bank(adapter, nama_bank)
            except akuntansi_ai.FormatTidakDikenali:
                gagal_jalur_grid = True
                df_bank = pd.DataFrame()
            except Exception as e:  # noqa: BLE001
                gagal_jalur_grid = True
                df_bank = pd.DataFrame()
                hasil["peringatan"].append(
                    f"'{nama_file_pdf}': jalur grid gagal saat parsing tabel bank "
                    f"({type(e).__name__}: {e}) -- mencoba jalur fallback posisi-kata."
                )

        if df_bank.empty and not gagal_jalur_grid:
            gagal_jalur_grid = True

        if gagal_jalur_grid:
            hasil["jalur_terpakai"] = "posisi_kata"
            try:
                file_like.seek(0)
            except (AttributeError, OSError):
                pass
            df_fallback, warn_fallback, ringkasan_footer = _ekstrak_pdf_berbasis_posisi(file_like, nama_file_pdf, nama_bank)
            hasil["peringatan"].extend(warn_fallback)
            hasil["ringkasan_footer"] = ringkasan_footer or None
            if df_fallback.empty:
                hasil["peringatan"].append(f"'{nama_file_pdf}': tidak ada baris transaksi terbaca setelah parsing (jalur grid maupun fallback posisi-kata).")
            else:
                hasil["df"] = df_fallback
            _simpan_cache_ekstraksi_pdf_helper(path_cache_file, hasil, nama_bank)
            return _selesai(hasil)

        hasil["jalur_terpakai"] = "grid"
        df_bank = _tempelkan_metadata_pdf(df_bank, baris_dengan_meta, nama_file_pdf)
        hasil["df"] = df_bank
        _simpan_cache_ekstraksi_pdf_helper(path_cache_file, hasil, nama_bank)
        return _selesai(hasil)

    def _simpan_cache_ekstraksi_pdf_helper(path_cache_file: Optional[str], hasil: Dict[str, Any], nama_bank: str) -> None:
        """Simpan hasil 1 file ke cache per-file (no-op kalau hash gagal
        dihitung, mis. file_like tidak seekable -- lihat _hash_isi_file)."""
        if not path_cache_file:
            return
        _simpan_cache_ekstraksi_pdf({
            "nama_bank": nama_bank,
            "df": hasil["df"],
            "peringatan": hasil["peringatan"],
            "ringkasan_footer": hasil["ringkasan_footer"],
            "jalur_terpakai": hasil["jalur_terpakai"],
        }, path_cache_file)

    # [BARU -- PERBAIKAN PERFORMA TAHAP 2] Jalankan ekstraksi tiap file
    # PARALEL (bukan satu-satu berurutan seperti sebelumnya) -- ini yang
    # paling terasa untuk client yang upload banyak bulan/bank sekaligus
    # (mis. 12 file rekening koran setahun): sebelumnya total waktu =
    # jumlah semua file diekstrak berurutan, sekarang kira-kira dibagi
    # PDF_PARALEL_MAKS (default 4x lebih cepat, bisa dituning lewat env
    # var KERTAS_KERJA_PDF_PARALEL_MAKS). Urutan hasil TETAP dikembalikan
    # sesuai urutan daftar_file_pdf asli (lihat sort di bawah) supaya
    # perilaku/urutan baris di GL tidak berubah dibanding versi sekuensial.
    hasil_per_file: List[Dict[str, Any]] = [None] * len(daftar_file_pdf)  # type: ignore[list-item]
    jumlah_worker = max(1, min(PDF_PARALEL_MAKS, len(daftar_file_pdf))) if daftar_file_pdf else 1
    with concurrent.futures.ThreadPoolExecutor(max_workers=jumlah_worker) as executor:
        future_ke_index = {
            executor.submit(_proses_satu_file, file_like, nama_file_pdf): i
            for i, (file_like, nama_file_pdf) in enumerate(daftar_file_pdf)
        }
        for future in concurrent.futures.as_completed(future_ke_index):
            i = future_ke_index[future]
            try:
                hasil_per_file[i] = future.result()
            except Exception as e:  # noqa: BLE001 -- jaga-jaga: worker seharusnya sudah menangkap semua exception sendiri
                _, nama_file_pdf = daftar_file_pdf[i]
                hasil_per_file[i] = {
                    "nama_file_pdf": nama_file_pdf, "df": None,
                    "peringatan": [f"'{nama_file_pdf}': gagal diproses di worker paralel ({type(e).__name__}: {e})."],
                    "ringkasan_footer": None, "jalur_terpakai": None,
                }

    # Gabungkan hasil semua file (urutan asli) + update cache_jalur sesuai
    # jalur yang TERBUKTI dipakai tiap file (lock tidak perlu lagi di sini
    # karena sudah balik ke thread utama/sekuensial).
    semua_df = []
    peringatan: List[str] = []
    ringkasan_per_file: Dict[str, Dict[str, Any]] = {}
    cache_berubah = False
    for hasil in hasil_per_file:
        nama_file_pdf = hasil["nama_file_pdf"]
        nama_bank = nama_bank_per_file.get(nama_file_pdf) or nama_file_pdf.rsplit(".", 1)[0]
        peringatan.extend(hasil["peringatan"])
        if hasil["ringkasan_footer"]:
            ringkasan_per_file[nama_file_pdf] = hasil["ringkasan_footer"]
        if hasil["df"] is not None and not hasil["df"].empty:
            semua_df.append(hasil["df"])
        if hasil["jalur_terpakai"] and cache_jalur.get(nama_bank) != hasil["jalur_terpakai"]:
            cache_jalur[nama_bank] = hasil["jalur_terpakai"]
            cache_berubah = True

    if cache_berubah:
        _simpan_cache_jalur_pdf(cache_jalur, path_cache_jalur)

    if not semua_df:
        return pd.DataFrame(), peringatan, ringkasan_per_file

    df_gabungan = pd.concat(semua_df, ignore_index=True)
    return df_gabungan, peringatan, ringkasan_per_file


def _tempelkan_metadata_pdf(df_bank: pd.DataFrame, baris_dengan_meta: List[Dict], nama_file_pdf: str) -> pd.DataFrame:
    """
    Tempelkan kolom source_pdf & halaman ke df_bank hasil parse_sheet_bank().

    parse_sheet_bank() melewati baris yang `keterangan` DAN `tanggal`-nya
    kosong (lihat kondisi skip di parse_sheet_bank), jadi jumlah baris
    df_bank bisa LEBIH SEDIKIT dari baris_dengan_meta. Untuk tetap presisi,
    kita jalankan ulang deteksi kolom yang sama (header row) supaya index
    baris mentah yang terpakai vs yang di-skip konsisten dengan yang
    dipakai parse_sheet_bank secara internal.
    """
    header_rownum, header_row = akuntansi_ai._cari_header_row(
        _LembarDenganMetadata([b["baris"] for b in baris_dengan_meta])
    )
    if header_row is None:
        # Seharusnya tidak terjadi karena parse_sheet_bank sudah berhasil,
        # tapi jaga-jaga: kembalikan tanpa metadata daripada crash.
        df_bank["source_pdf"] = nama_file_pdf
        df_bank["halaman"] = None
        return df_bank

    headers = header_row
    idx_tanggal = akuntansi_ai._cari_idx(headers, ["tgl", "tanggal", "date"])
    idx_keterangan = akuntansi_ai._cari_idx(headers, ["keterangan", "remarks"])

    baris_data = baris_dengan_meta[header_rownum:]  # setelah header, 0-indexed pas
    metadata_terpakai = []
    for b in baris_data:
        row = b["baris"]
        if len(row) <= max(idx_tanggal or 0, idx_keterangan or 0):
            continue
        keterangan_kosong = not row[idx_keterangan] if idx_keterangan is not None else True
        tanggal_kosong = not row[idx_tanggal] if idx_tanggal is not None else True
        if keterangan_kosong and tanggal_kosong:
            continue
        metadata_terpakai.append((b["source_pdf"], b["halaman"]))

    if len(metadata_terpakai) != len(df_bank):
        # Fallback aman: tidak presisi 100% tapi tidak crash -- lebih baik
        # kertas kerja tetap terbuat dengan source_pdf terisi (tanpa halaman
        # akurat per baris) daripada gagal total. Ini ditandai di peringatan
        # oleh pemanggil kalau perlu.
        df_bank["source_pdf"] = nama_file_pdf
        df_bank["halaman"] = None
        return df_bank

    df_bank = df_bank.copy()
    df_bank["source_pdf"] = [m[0] for m in metadata_terpakai]
    df_bank["halaman"] = [m[1] for m in metadata_terpakai]
    return df_bank


# ============================================================
# 1.6 [BARU] FALLBACK: BANGUN COA "KAYA" DARI DATABASE CLIENT
# ============================================================
# muat_coa_kertas_kerja() di atas HANYA bisa membaca dari sheet Excel
# COA (skema kaya: Normal Balance/Statement/FS Group/Notes) -- COA yang
# TERSIMPAN DI DATABASE client (tabel Coa, lihat db_client.ambil_coa_client)
# skemanya LEBIH SEDERHANA (no_akun/nama_akun/kategori/normal_saldo,
# TIDAK ADA kolom Statement atau penanda staging). Supaya kertas kerja
# tetap bisa digenerate untuk client yang BELUM upload file COA khusus
# (hanya punya COA dasar di DB), fungsi ini memetakan sebaik mungkin --
# best-effort, BUKAN akurat 100%:
#   - Normal Balance  <- normal_saldo apa adanya (client WAJIB memastikan
#     nilainya persis "Debit"/"Credit", bukan singkatan lain)
#   - Statement (BS/PNL) <- ditebak dari kata kunci umum di `kategori`
#     (Debit-kunci: aset/aktiva/kas/piutang/persediaan/pajak dibayar;
#      Kredit-kunci: liabilitas/kewajiban/utang/ekuitas/modal -> BS;
#      pendapatan/penjualan/beban/biaya/hpp -> PNL). Kalau `kategori`
#      tidak cocok kata kunci manapun, Statement dibiarkan None -- akun
#      itu TIDAK akan muncul di BS_Monthly/PNL_Monthly (bukan salah
#      tempat, tapi hilang; ditandai di `peringatan`).
#   - FS Group <- `kategori` apa adanya (dipakai sebagai label, bukan
#     dihitung ulang)
#   - Notes <- `keterangan` apa adanya -- akun staging TIDAK akan
#     otomatis ketemu lewat jalur ini kecuali field `keterangan` client
#     di DB memang sudah berisi kata "staging" (kemungkinan besar belum
#     -- lihat peringatan yang dikembalikan).
# Akuntan/tim tetap disarankan mengisi/upload COA versi Excel skema kaya
# begitu memungkinkan; fallback ini semata supaya alur tidak buntu total
# kalau itu belum ada.
def bangun_coa_kertas_kerja_dari_db(daftar_akun_db: List[Dict[str, Any]]) -> Tuple[pd.DataFrame, List[str]]:
    """
    Args:
        daftar_akun_db: hasil db_client.ambil_coa_client(client_id) --
            list of dict dengan key no_akun/nama_akun/kategori/
            normal_saldo/keterangan (lihat db_client.py).

    Returns:
        (df_coa_kaya, peringatan) -- df_coa_kaya siap dipakai sebagai
        parameter df_coa di generate_kertas_kerja(), peringatan berisi
        catatan apa saja yang ditebak/tidak lengkap.
    """
    peringatan: List[str] = []
    if not daftar_akun_db:
        return pd.DataFrame(columns=_KOLOM_COA_KK_WAJIB), [
            "COA client masih kosong di database -- upload COA dulu "
            "(lewat file Excel skema kertas kerja, atau menu COA biasa)."
        ]

    kata_kunci_bs = ["aset", "aktiva", "kas", "bank", "piutang", "persediaan",
                     "pajak dibayar", "uang muka", "jaminan", "liabilitas",
                     "kewajiban", "utang", "hutang", "ekuitas", "modal"]
    kata_kunci_pnl = ["pendapatan", "penjualan", "hpp", "beban", "biaya", "pajak penghasilan"]

    baris = []
    ada_tebakan_gagal = False
    for akun in daftar_akun_db:
        kategori = str(akun.get("kategori") or "").strip().lower()
        statement = None
        if any(k in kategori for k in kata_kunci_bs):
            statement = "BS"
        elif any(k in kategori for k in kata_kunci_pnl):
            statement = "PNL"
        else:
            ada_tebakan_gagal = True
        baris.append({
            "Account No.": str(akun.get("no_akun") or ""),
            "Account Name": str(akun.get("nama_akun") or ""),
            "Class": akun.get("kategori"),
            "Normal Balance": akun.get("normal_saldo"),
            "Statement": statement,
            "FS Group": akun.get("kategori"),
            "Tax Return Group": None,
            "Notes": akun.get("keterangan"),
        })

    df = pd.DataFrame(baris)
    peringatan.append(
        "COA dibangun otomatis dari data COA client di database (bukan dari "
        "file Excel skema kertas kerja) -- kolom Statement (BS/PNL) ditebak "
        "dari nama kategori, dan akun staging (Notes berisi 'staging') "
        "kemungkinan besar BELUM ada. Disarankan lengkapi/upload COA versi "
        "Excel skema kertas kerja untuk hasil yang lebih akurat."
    )
    if ada_tebakan_gagal:
        peringatan.append(
            "Sebagian akun kategorinya tidak cocok kata kunci BS/PNL manapun "
            "-- akun tersebut TIDAK akan muncul di BS_Monthly/PNL_Monthly "
            "sampai kolom Statement diisi manual di COA."
        )
    return df, peringatan


# ============================================================
# 2. AKUN STAGING (dari COA, per arah CR/DB)
# ============================================================

# ============================================================
# 1.5 READER COA KHUSUS KERTAS KERJA (skema lebih kaya dari muat_coa lama)
# ============================================================

_HEADER_COA_KK_MAP = {
    "account no.": "Account No.", "no akun": "Account No.", "no. akun": "Account No.",
    "kode akun": "Account No.", "account number": "Account No.",
    "account name": "Account Name", "nama akun": "Account Name", "description": "Account Name",
    "class": "Class", "kelas": "Class",
    "normal balance": "Normal Balance", "saldo normal": "Normal Balance",
    "statement": "Statement",
    "fs group": "FS Group",
    "tax return group": "Tax Return Group",
    "notes": "Notes", "catatan": "Notes", "keterangan": "Notes",
}
_KOLOM_COA_KK_WAJIB = ["Account No.", "Account Name", "Normal Balance", "Statement", "FS Group", "Notes"]


def muat_coa_kertas_kerja(wb: openpyxl.Workbook, nama_sheet_coa: str = "COA") -> pd.DataFrame:
    """
    Reader COA TERPISAH dari akuntansi_ai.muat_coa() -- SENGAJA tidak
    memakai/mengubah fungsi itu, karena muat_coa() cuma menangkap 3 kolom
    (no_akun, nama_akun, kategori) dan dipakai 15+ jenis dokumen lain, jadi
    terlalu berisiko diubah. Fungsi ini membaca skema COA yang LEBIH KAYA
    dibutuhkan khusus untuk kertas kerja (kolom Normal Balance/Statement/
    FS Group/Notes -- lihat _KOLOM_COA_KK_WAJIB), memakai wb yang SAMA yang
    dibuka main.py/handler upload (tidak perlu buka file 2x).

    Kalau sheet COA client belum punya salah satu kolom wajib, dikembalikan
    apa adanya dengan kolom yang hilang berisi None -- error yang jelas
    baru muncul belakangan di cari_akun_staging_dari_coa()/susun_bs_pnl_monthly
    (lebih mudah didebug drpd gagal diam-diam di sini).
    """
    target = None
    for name in wb.sheetnames:
        if "coa" in name.strip().lower():
            target = name
            break
    if target is None:
        return pd.DataFrame(columns=_KOLOM_COA_KK_WAJIB)

    ws = wb[target]
    header_ditemukan = False
    kolom_map: Dict[str, int] = {}
    baris = []
    for row in ws.iter_rows(values_only=True):
        sel = [str(c).strip().lower() if c is not None else "" for c in row]
        if not header_ditemukan:
            cocok = [h for h in sel if h in _HEADER_COA_KK_MAP]
            if len(cocok) >= 3:  # minimal 3 header dikenali baru dianggap baris header
                header_ditemukan = True
                for i, h in enumerate(sel):
                    if h in _HEADER_COA_KK_MAP:
                        kolom_map[_HEADER_COA_KK_MAP[h]] = i
            continue
        if "Account No." not in kolom_map or "Account Name" not in kolom_map:
            continue
        no_akun = row[kolom_map["Account No."]]
        nama_akun = row[kolom_map["Account Name"]]
        if no_akun is None or nama_akun is None:
            continue
        rec = {"Account No.": no_akun, "Account Name": str(nama_akun).strip()}
        for kolom in ["Class", "Normal Balance", "Statement", "FS Group", "Tax Return Group", "Notes"]:
            idx = kolom_map.get(kolom)
            rec[kolom] = row[idx] if idx is not None and idx < len(row) else None
        baris.append(rec)

    df = pd.DataFrame(baris)
    for kolom in _KOLOM_COA_KK_WAJIB:
        if kolom not in df.columns:
            df[kolom] = None
    return df


def cari_akun_staging_dari_coa(df_coa: pd.DataFrame) -> Dict[str, str]:
    """
    [FIX] Definisi fungsi ini sebelumnya HILANG di versi upload (hanya
    docstring + isi fungsi tanpa baris `def`), sehingga baris-baris
    setelah `return df` di atas jadi kode mati (dead code) menempel di
    dalam muat_coa_kertas_kerja(), dan generate_kertas_kerja() akan
    NameError saat memanggil cari_akun_staging_dari_coa(df_coa) karena
    fungsi ini sama sekali belum terdaftar di module. Ditambahkan
    kembali baris `def` ini tanpa mengubah isi logic aslinya.

    Cari 2 akun staging (utk CR & DB) dari sheet COA client: baris yang
    kolom 'Notes' mengandung kata 'staging' (case-insensitive), dipisah
    berdasarkan 'Normal Balance' (Kredit -> staging utk transaksi CR,
    Debit -> staging utk transaksi DB).

    Returns: {"CR": no_akun_str, "DB": no_akun_str}

    Raises ValueError kalau salah satu/keduanya tidak ditemukan -- akuntan
    WAJIB menandai 2 akun ini di COA client (kolom Notes berisi kata
    "staging") sebelum kertas kerja bisa digenerate. Ini disengaja: lebih
    aman gagal dengan pesan jelas daripada menebak akun staging yang salah.
    """
    kolom_hilang = [k for k in _KOLOM_COA_DIBUTUHKAN if k not in df_coa.columns]
    if kolom_hilang:
        raise ValueError(
            f"Sheet COA client belum punya kolom {kolom_hilang}. "
            f"Kolom yang dibutuhkan: {_KOLOM_COA_DIBUTUHKAN}."
        )

    mask_staging = df_coa["Notes"].fillna("").str.contains("staging", case=False)
    kandidat = df_coa[mask_staging]

    akun_kredit = kandidat[kandidat["Normal Balance"].str.lower() == "credit"]
    akun_debit = kandidat[kandidat["Normal Balance"].str.lower() == "debit"]

    if akun_kredit.empty or akun_debit.empty:
        raise ValueError(
            "Akun staging belum ditandai lengkap di COA client. Tandai TEPAT 1 akun "
            "liability (Normal Balance = Credit, mis. 'Customer Deposit') dan 1 akun "
            "asset (Normal Balance = Debit, mis. 'Prepaid Expenses') dengan kata "
            "'staging' di kolom Notes-nya."
        )

    return {
        "CR": str(akun_kredit.iloc[0]["Account No."]),
        "DB": str(akun_debit.iloc[0]["Account No."]),
    }


# ============================================================
# 2.5 [FIX] AKUN BANK (dari COA) -- menggantikan hardcode "1102"
# ============================================================
# Sebelumnya kode akun bank (lawan jurnal saat transaksi belum
# terklasifikasi AI) di-hardcode "1102" di susun_sheet_gl(), yang salah
# untuk hampir semua client selain contoh aslinya (kode akun beda-beda per
# client). Fungsi ini mencari akun Bank dari COA client dengan pola SAMA
# seperti pencarian akun staging (cari_akun_staging_dari_coa di atas):
# baris yang kolom 'Notes'-nya mengandung kata kunci tertentu.

def cari_akun_bank_dari_coa(df_coa: pd.DataFrame) -> Tuple[List[Dict[str, str]], List[str]]:
    """
    Cari akun Bank di COA client: baris yang kolom 'Notes' mengandung kata
    'bank' atau 'rekening' (case-insensitive) -- pola sama seperti akun
    staging (Notes mengandung 'staging'). Akun ini dipakai sebagai lawan
    jurnal saat transaksi TIDAK berhasil diklasifikasi AI/pola (sebelumnya
    di-hardcode "1102" di susun_sheet_gl -- lihat catatan integrasi #FIX
    di docstring susun_sheet_gl).

    Returns:
        (daftar_akun_bank, peringatan) -- daftar_akun_bank = list of
        {"Account No.": ..., "Account Name": ...} untuk SEMUA akun
        bertanda bank (client bisa punya lebih dari 1 rekening/bank,
        sesuai kemungkinan multi-file/multi-bank di
        susun_gl_dari_pdf_rekening_koran). Kalau lebih dari 1 ditemukan,
        pencocokan transaksi -> akun bank spesifik dilakukan PER BARIS di
        susun_sheet_gl() lewat _pilih_akun_bank_untuk_baris() (dicocokkan
        ke kolom 'bank' hasil parse_sheet_bank); kalau tidak ada yang
        cocok, dipakai akun bank PERTAMA sebagai fallback & dicatat di
        peringatan supaya akuntan sadar ada baris yang mungkin salah
        rekening.

    Raises ValueError kalau TIDAK ADA akun bertanda bank sama sekali --
    akuntan WAJIB menandai minimal 1 akun aset Bank di COA client (kolom
    Notes berisi kata "bank"/"rekening") sebelum kertas kerja bisa
    digenerate akurat. Sama seperti akun staging: sengaja gagal dengan
    pesan jelas daripada menebak/hardcode kode akun yang bisa salah.
    """
    kolom_hilang = [k for k in _KOLOM_COA_DIBUTUHKAN if k not in df_coa.columns]
    if kolom_hilang:
        raise ValueError(
            f"Sheet COA client belum punya kolom {kolom_hilang}. "
            f"Kolom yang dibutuhkan: {_KOLOM_COA_DIBUTUHKAN}."
        )

    mask_bank = df_coa["Notes"].fillna("").str.contains("bank|rekening", case=False, regex=True)
    # [FIX] Kata "bank" di Notes TERNYATA juga muncul di akun-akun lain yang
    # terkait bank tapi BUKAN akun kas/bank itu sendiri -- ditemukan lewat
    # tes dengan COA contoh (Kertas_Kerja_Laporan_Keuangan_2025.xlsx):
    # "Bank Charges" (5205, Expense/PNL), "Bank Interest Income - Final"
    # (6101, Other Income/PNL), "Final Tax on Bank Interest..." (7102,
    # Other Expense/PNL), dan "Opening Balance Suspense - Review" (3909,
    # Equity/BS, Notes-nya menyebut "saldo awal bank") semuanya ikut
    # tertangkap regex "bank|rekening" di atas, padahal cuma 1102/1103
    # (Asset, Statement=BS) yang benar akun kas/bank. Tanpa filter ini,
    # transaksi yang gagal diklasifikasi bisa salah diposting ke akun
    # PNL/Equity sbg lawan jurnal bank -- jauh lebih merusak daripada
    # hardcode lama. Dipersempit ke Class=Asset & Statement=BS (2 kolom
    # yang SUDAH pasti ada di df_coa lewat _KOLOM_COA_DIBUTUHKAN/skema
    # kertas kerja standar).
    if "Class" in df_coa.columns:
        mask_bank = mask_bank & (df_coa["Class"].fillna("").str.strip().str.lower() == "asset")
    mask_bank = mask_bank & (df_coa["Statement"].fillna("").str.strip().str.upper() == "BS")
    kandidat = df_coa[mask_bank]

    if kandidat.empty:
        raise ValueError(
            "Akun Bank belum ditandai di COA client. Tandai minimal 1 akun "
            "aset Bank (mis. 'Bank BCA', 'Bank Mandiri') dengan kata "
            "'bank' di kolom Notes-nya, DAN pastikan Class='Asset' & "
            "Statement='BS' pada akun tsb -- akun ini dipakai sebagai lawan "
            "jurnal bank saat transaksi belum berhasil diklasifikasi "
            "otomatis (sebelumnya salah di-hardcode sebagai '1102')."
        )

    daftar_akun_bank = [
        {"Account No.": str(r["Account No."]), "Account Name": str(r["Account Name"])}
        for _, r in kandidat.iterrows()
    ]

    peringatan: List[str] = []
    if len(daftar_akun_bank) > 1:
        nama_semua = [d["Account Name"] for d in daftar_akun_bank]
        peringatan.append(
            f"Ditemukan {len(daftar_akun_bank)} akun bertanda 'bank' di COA "
            f"({nama_semua}) -- setiap transaksi akan dicocokkan ke akun "
            "bank yang namanya paling mirip nama bank di rekening koran "
            "asalnya; kalau tidak ada yang cocok, dipakai akun bank "
            f"pertama ('{nama_semua[0]}') sebagai fallback. Cek sheet GL "
            "kolom Debit/Credit Account kalau client punya lebih dari 1 "
            "rekening untuk memastikan pencocokan sudah benar."
        )

    return daftar_akun_bank, peringatan


def _pilih_akun_bank_untuk_baris(
    nama_bank_transaksi: Optional[str], daftar_akun_bank: List[Dict[str, str]],
) -> str:
    """
    Cocokkan nama bank di 1 baris transaksi (kolom 'bank' hasil
    parse_sheet_bank() di akuntansi_ai.py, mis. "BCA"/"Mandiri", diambil
    dari nama file PDF atau parameter nama_bank_per_file) ke salah satu
    akun di `daftar_akun_bank` (hasil cari_akun_bank_dari_coa), berdasar
    containment substring case-insensitive terhadap Account Name.

    Fallback ke akun bank PERTAMA di daftar kalau nama kosong atau tidak
    ada yang cocok -- aman untuk kasus paling umum (client hanya punya 1
    rekening bank yang direkonsiliasi).
    """
    if nama_bank_transaksi:
        target = str(nama_bank_transaksi).strip().lower()
        if target:
            for akun in daftar_akun_bank:
                if target in akun["Account Name"].lower():
                    return akun["Account No."]
    return daftar_akun_bank[0]["Account No."]


# ============================================================
# 3. SHEET GL (per transaksi)
# ============================================================

def susun_sheet_gl(
    df_hasil_klasifikasi: pd.DataFrame,
    akun_staging: Dict[str, str],
    daftar_akun_bank: List[Dict[str, str]],
) -> pd.DataFrame:
    """
    Susun sheet GL dari output akuntansi_ai.proses_dataframe() (dipanggil
    lewat proses_file_rekening_koran, lalu ditempel metadata source_pdf/
    halaman lewat susun_gl_dari_pdf_rekening_koran di atas).

    df_hasil_klasifikasi diharapkan sudah punya kolom: tanggal, bank,
    keterangan, supplier_cust, mutasi_debet, mutasi_kredit, source_pdf,
    halaman, no_akun_debet, nama_akun_debet, no_akun_kredit,
    nama_akun_kredit, sumber_kategori, catatan_ai (nama kolom mengikuti
    proses_dataframe() di akuntansi_ai.py -- SESUAIKAN kalau nama kolom
    riil berbeda setelah dites).

    daftar_akun_bank: hasil cari_akun_bank_dari_coa(df_coa) -- dipakai
    sebagai lawan jurnal bank saat transaksi belum terklasifikasi AI
    (dulu di-hardcode "1102", lihat [FIX] di bawah).
    """
    # [BARU -- PERBAIKAN PERFORMA -- VEKTORISASI] Sebelumnya loop Python
    # murni (iterrows()) per baris transaksi -- untuk rekening koran multi
    # bulan/bank dengan ribuan baris, ini bisa jadi ratusan ms - detik
    # HANYA untuk fungsi ini (mis. 5000 baris = ~0.39 detik). Logika
    # PERSIS SAMA (termasuk fix NaN-truthy pakai _angka() yang sudah ada
    # dari sebelumnya), cuma dihitung per-KOLOM (vektor) alih-alih per
    # BARIS -- diverifikasi menghasilkan angka SAMA PERSIS dengan versi
    # lama di data sintetis (5000 baris + kasus tepi kosong/bank tak
    # dikenal), ~15x lebih cepat (0.39 detik -> 0.025 detik).
    #
    # [CATATAN PERBAIKAN KECIL, SEKALIAN] "Classification Basis" dulu
    # `r.get("catatan_ai") or r.get("sumber_kategori")` -- rentan bug
    # NaN-truthy YANG SAMA yang dijelaskan panjang di versi lama fungsi
    # ini (kalau catatan_ai ternyata tersimpan sbg NaN float, bukan None,
    # maka `NaN or x` mengembalikan NaN itu sendiri, BUKAN x, karena NaN
    # truthy di Python). Di jalur produksi normal catatan_ai diinisialisasi
    # None (bukan NaN) jadi TIDAK kena bug ini, tapi supaya konsisten
    # dengan filosofi fix _angka() di file ini, sekarang NaN & None
    # sama-sama dianggap "kosong" (fallback ke sumber_kategori).
    r = df_hasil_klasifikasi.reset_index(drop=True)
    n = len(r)
    idx1 = np.arange(1, n + 1)

    mutasi_debet = r.get("mutasi_debet").map(_angka) if "mutasi_debet" in r.columns else pd.Series([0.0] * n)
    mutasi_kredit = r.get("mutasi_kredit").map(_angka) if "mutasi_kredit" in r.columns else pd.Series([0.0] * n)
    arah = np.where(mutasi_debet.values == 0, "CR", "DB")
    nominal = np.where(mutasi_debet.values != 0, mutasi_debet.values, mutasi_kredit.values)

    no_akun_debet = r.get("no_akun_debet")
    no_akun_kredit = r.get("no_akun_kredit")
    ada_jurnal = no_akun_debet.notna().values & no_akun_kredit.notna().values

    confidence = _confidence_dari_sumber_vektor(r.get("sumber_kategori"))

    tanggal = pd.to_datetime(r.get("tanggal"), errors="coerce")
    tahun = tanggal.dt.year
    journal_id = [
        f"{int(y) if pd.notna(y) else ''}-{i:04d}" for y, i in zip(tahun, idx1)
    ]
    month_no = tanggal.dt.month
    month_name = month_no.map(lambda m: BULAN_URUT[int(m) - 1] if pd.notna(m) else None)

    is_cr = arah == "CR"
    staging = np.where(is_cr, akun_staging["CR"], akun_staging["DB"])

    suggested = np.select(
        [is_cr & ada_jurnal, (~is_cr) & ada_jurnal],
        [no_akun_kredit.astype(str).values, no_akun_debet.astype(str).values],
        default=staging,
    )
    nama_akun_kredit = r.get("nama_akun_kredit")
    nama_akun_debet = r.get("nama_akun_debet")
    suggested_nama = np.where(is_cr, nama_akun_kredit.values, nama_akun_debet.values)

    # [FIX -- tetap dipertahankan] Sebelumnya akun bank di-hardcode "1102"
    # -- dicari dari COA client (cari_akun_bank_dari_coa), dicocokkan per
    # baris ke nama bank transaksi. [VEKTORISASI TAMBAHAN] Pencocokan
    # substring (_pilih_akun_bank_untuk_baris) sekarang dihitung SEKALI
    # per nilai bank UNIK (biasanya cuma 1-3 rekening per client), bukan
    # per baris transaksi -- fungsi pencocokannya sendiri TIDAK diubah.
    bank_col = r.get("bank")
    peta_bank = {
        b: _pilih_akun_bank_untuk_baris(b, daftar_akun_bank)
        for b in bank_col.dropna().unique()
    }
    default_bank = daftar_akun_bank[0]["Account No."]
    akun_bank_baris = bank_col.map(peta_bank)
    akun_bank_baris = akun_bank_baris.where(akun_bank_baris.notna(), default_bank).values

    debit_account = np.select(
        [ada_jurnal, (~ada_jurnal) & (~is_cr)],
        [no_akun_debet.values, np.full(n, akun_staging["DB"], dtype=object)],
        default=akun_bank_baris,
    )
    credit_account = np.select(
        [ada_jurnal, (~ada_jurnal) & is_cr],
        [no_akun_kredit.values, np.full(n, akun_staging["CR"], dtype=object)],
        default=akun_bank_baris,
    )

    review_status = np.select(
        [confidence == "High", confidence == "Medium"],
        ["Auto-High", "Auto-Medium"],
        default="Need Review",
    )

    catatan_ai = r.get("catatan_ai")
    sumber_kategori_col = r.get("sumber_kategori")
    catatan_ai_notna_dan_tidak_kosong = catatan_ai.notna() & (catatan_ai.astype(str) != "")
    classification_basis = np.where(
        catatan_ai_notna_dan_tidak_kosong, catatan_ai.values, sumber_kategori_col.values
    )

    return pd.DataFrame({
        "Journal ID": journal_id,
        "Date": r.get("tanggal").values,
        "Month No.": month_no.values,
        "Month": month_name.values,
        "Source PDF": r.get("source_pdf").values,
        "Page": r.get("halaman").values,
        "Bank Description": r.get("keterangan").values,
        "Counterparty": r.get("supplier_cust").values,
        "Direction": arah,
        "Amount (IDR)": nominal,
        "Initial Staging Account": staging,
        "Suggested Counterpart Account": suggested,
        "Suggested Counterpart Name": suggested_nama,
        "Confidence": confidence,
        "Classification Basis": classification_basis,
        "Debit Account": debit_account,
        "Credit Account": credit_account,
        "Debit (IDR)": nominal,
        "Credit (IDR)": nominal,
        "Review Status": review_status,
        "Review Notes": None,
    })


def _confidence_dari_sumber(sumber_kategori: str) -> str:
    """
    Petakan `sumber_kategori` (istilah internal akuntansi_ai.py) ke
    Low/Medium/High seperti di contoh kertas kerja. PEMETAAN INI ASUMSI
    AWAL -- tolong dikonfirmasi ke tim persis nilai-nilai apa saja yang
    muncul di `sumber_kategori` real (dari kode: "Sesuai Pola yang
    Dipelajari", "Belum Terkategori", nilai dari confidence field AI
    tinggi/sedang/rendah, dst).
    """
    s = sumber_kategori.lower()
    # [BARU] Label dari _terapkan_kata_kunci_kertas_kerja(), mis.
    # "Kata Kunci Kertas Kerja (High)" / "... (Medium)" -- dicek LEBIH
    # DULU sebelum aturan "perlu cek" di bawah supaya tidak salah kena.
    if "kata kunci kertas kerja" in s:
        if "(high)" in s:
            return "High"
        if "(medium)" in s:
            return "Medium"
        return "Medium"
    if "belum terkategori" in s or "perlu cek" in s:
        return "Low"
    if "pola yang dipelajari" in s or "tinggi" in s:
        return "High"
    if "sedang" in s:
        return "Medium"
    if "rendah" in s:
        return "Low"
    return "Medium"


def _confidence_dari_sumber_vektor(sumber_series: pd.Series) -> np.ndarray:
    """[BARU -- PERBAIKAN PERFORMA -- VEKTORISASI] Versi vektor dari
    _confidence_dari_sumber() di atas -- dipakai susun_sheet_gl() supaya
    tidak perlu memanggil fungsi Python murni per baris. Logika CABANG
    (urutan prioritas if/elif) DISALIN PERSIS, cuma diekspresikan sbg
    kondisi boolean per-kolom + np.select. Diverifikasi menghasilkan
    label SAMA PERSIS dengan _confidence_dari_sumber() untuk semua nilai
    sumber_kategori yang dikenal (termasuk kombinasi "kata kunci kertas
    kerja" + "(high)"/"(medium)", "perlu cek", dst). _confidence_dari_sumber()
    sendiri TIDAK dihapus/diubah -- tetap dipertahankan sbg referensi &
    untuk pemanggil lain di luar susun_sheet_gl."""
    s = sumber_series.fillna("").astype(str).str.lower()
    cond_kkk = s.str.contains("kata kunci kertas kerja", na=False)
    cond_kkk_high = cond_kkk & s.str.contains(r"\(high\)", na=False)
    cond_kkk_medium = cond_kkk & ~cond_kkk_high
    cond_low1 = (~cond_kkk) & (
        s.str.contains("belum terkategori", na=False) | s.str.contains("perlu cek", na=False)
    )
    cond_high2 = (~cond_kkk) & ~cond_low1 & (
        s.str.contains("pola yang dipelajari", na=False) | s.str.contains("tinggi", na=False)
    )
    cond_medium2 = (~cond_kkk) & ~cond_low1 & ~cond_high2 & s.str.contains("sedang", na=False)
    cond_low2 = (~cond_kkk) & ~cond_low1 & ~cond_high2 & ~cond_medium2 & s.str.contains("rendah", na=False)
    return np.select(
        [cond_kkk_high, cond_kkk_medium, cond_low1, cond_high2, cond_medium2, cond_low2],
        ["High", "Medium", "Low", "High", "Medium", "Low"],
        default="Medium",
    )


# ============================================================
# 3.5 [BARU] REVIEW & PERBAIKAN GL LEWAT CLAUDE API (OPSIONAL)
# ============================================================
# Setelah GL disusun dari klasifikasi (susun_sheet_gl), baris yang
# Confidence-nya BUKAN "High" (jadi Medium/Low) dikirim SEKALIGUS dalam
# batch ke Claude API supaya diberi saran akun yang lebih tepat +
# alasannya. Hasilnya:
#   1. Kolom GL yang relevan diperbarui (Suggested Counterpart Account/
#      Name, Debit/Credit Account, Confidence, Review Notes, Review
#      Status)
#   2. df_status baru dibangun berisi SEMUA baris yang disentuh Claude
#      (nilai sebelum & sesudah + alasan) -- dipakai sheet "Status" baru
#
# INI OPSIONAL (dikontrol parameter pakai_claude_review di
# generate_kertas_kerja) -- kalau False atau tidak ada baris yang perlu
# direview, fungsi ini TIDAK memanggil API sama sekali (hemat kredit).
#
# Pakai claude_client.panggil_claude_terstruktur() -- structured output
# lewat tool-use, konsisten dengan pemanggilan Claude lain di codebase
# (mis. generate_narasi_calk_claude di claude_client.py). Claude dipaksa
# balas sesuai _TOOL_REVIEW_GL_KK di bawah, jadi tidak ada parsing JSON
# manual/markdown-fence-stripping seperti pendekatan teks bebas.

_UKURAN_BATCH_REVIEW = 30  # baris per panggilan API -- jaga token tetap wajar

_TOOL_REVIEW_GL_KK = {
    "name": "kirim_review_gl_kertas_kerja",
    "description": (
        "Kirim hasil review akun untuk setiap transaksi GL kertas kerja "
        "yang confidence-nya bukan High."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "hasil": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "journal_id": {
                            "type": "string",
                            "description": "Journal ID persis seperti di input.",
                        },
                        "akun_disarankan": {
                            "type": "string",
                            "description": "Nomor akun dari daftar COA yang diberikan.",
                        },
                        "nama_akun_disarankan": {
                            "type": "string",
                            "description": "Nama akun sesuai nomor tsb.",
                        },
                        "alasan": {
                            "type": "string",
                            "description": "Alasan singkat, maksimal 1 kalimat.",
                        },
                        "yakin": {
                            "type": "boolean",
                            "description": (
                                "False kalau deskripsi transaksi terlalu ambigu utk "
                                "dipastikan -- akun_disarankan tetap diisi (akun yang "
                                "sudah diusulkan sistem, jangan diubah) tapi tetap "
                                "ditandai perlu review manual."
                            ),
                        },
                    },
                    "required": [
                        "journal_id", "akun_disarankan", "nama_akun_disarankan",
                        "alasan", "yakin",
                    ],
                },
            },
        },
        "required": ["hasil"],
    },
}

_SYSTEM_PROMPT_REVIEW_GL_KK = (
    "Kamu adalah asisten akuntan yang mereview klasifikasi transaksi bank "
    "ke akun COA untuk working paper kertas kerja. Untuk SETIAP transaksi "
    "yang diberikan, akun yang disarankan sistem punya confidence "
    "rendah/sedang -- tugasmu usulkan akun COA yang PALING TEPAT beserta "
    "alasan singkat. HANYA boleh pilih nomor akun dari daftar COA yang "
    "diberikan di prompt -- JANGAN mengarang nomor akun yang tidak ada di "
    "daftar itu. Kalau tidak yakin akun mana yang paling tepat untuk suatu "
    "transaksi (deskripsinya terlalu ambigu), tetap sertakan baris itu di "
    "hasil, isi akun_disarankan dengan akun yang sudah diusulkan sistem "
    "(jangan diubah), dan set yakin=false -- jangan ditebak paksa."
)


def _bangun_prompt_review_gl(baris_batch: List[Dict], df_coa: pd.DataFrame) -> str:
    """Susun isi prompt (bagian data) untuk 1 batch baris GL yang perlu
    direview Claude. Instruksi format output TIDAK perlu ditulis di sini
    lagi -- sudah dijamin oleh tool_schema (_TOOL_REVIEW_GL_KK) dan
    system prompt (_SYSTEM_PROMPT_REVIEW_GL_KK) lewat
    panggil_claude_terstruktur()."""
    daftar_akun = "\n".join(
        f"- {r['Account No.']} | {r['Account Name']}"
        for _, r in df_coa[["Account No.", "Account Name"]].drop_duplicates().iterrows()
    )
    daftar_transaksi = json.dumps(baris_batch, default=str, ensure_ascii=False, indent=2)

    return (
        "DAFTAR AKUN COA YANG TERSEDIA (HANYA boleh pilih dari daftar ini):\n"
        f"{daftar_akun}\n\n"
        "TRANSAKSI YANG PERLU DIREVIEW (format JSON):\n"
        f"{daftar_transaksi}\n\n"
        "Review setiap transaksi di atas dan kirim hasilnya lewat tool yang tersedia."
    )


def _terapkan_hasil_review_claude(
    df_gl: pd.DataFrame, hasil_review: List[Dict], df_coa: pd.DataFrame
) -> Tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    """Terapkan hasil review Claude (list dict dari JSON) ke df_gl.
    Return (df_gl_diperbarui, df_status_untuk_sheet_baru, peringatan).

    [FIX 1] df_coa sekarang WAJIB diberikan supaya akun yang disarankan
    Claude divalidasi dulu terhadap daftar akun COA yang sesungguhnya
    sebelum diterapkan. Versi sebelumnya langsung percaya
    "akun_disarankan" apa adanya walau prompt SUDAH minta Claude hanya
    memilih dari daftar COA -- LLM tetap bisa halusinasi nomor akun yang
    tidak ada (typo/gabungan/nomor lama). Kalau itu terjadi & lolos tanpa
    validasi, GL akan mereferensikan akun yang tidak ada di COA client,
    yang baru ketahuan belakangan (kalau ketahuan sama sekali) saat
    Bank_Posting_Summary/TB_Monthly disusun dari GL ini -- jauh lebih
    sulit dilacak balik ke baris asalnya. Sekarang: kalau
    "akun_disarankan" tidak cocok dengan Account No. manapun di df_coa,
    baris itu TETAP pakai akun/saran SISTEM (bukan dibuang), "yakin"
    dipaksa jadi False, dan pesan alasannya ditandai jelas -- baris tetap
    ditandai perlu review manual, bukan diam-diam salah.

    [FIX 2] Kolom "Confidence" di df_gl sekarang IKUT diperbarui jadi
    "High" saat Claude yakin (yakin=True DAN akun tervalidasi).
    Sebelumnya hanya "Review Status"/"Review Notes" yang diperbarui,
    "Confidence" dibiarkan tetap Medium/Low lama -- akibatnya
    ringkasan_status_kertas_kerja() (dipakai utk ringkasan yang
    ditunjukkan ke user SEBELUM konfirmasi lanjut ke laporan 18-sheet)
    tetap melaporkan baris ini sebagai Medium/Low walau sudah
    diperbaiki Claude, jadi angka confidence yang dilihat user
    menyesatkan (undercount High, overcount Medium/Low)."""
    df_gl = df_gl.copy()
    baris_status = []
    peringatan: List[str] = []
    akun_valid = set(df_coa["Account No."].astype(str).str.strip()) if not df_coa.empty else set()
    jumlah_halusinasi = 0
    peta_hasil = {h.get("journal_id"): h for h in hasil_review if h.get("journal_id")}

    for idx, row in df_gl.iterrows():
        hasil = peta_hasil.get(row["Journal ID"])
        if hasil is None:
            continue  # baris ini tidak dikirim/tidak dibalas Claude -- biarkan apa adanya

        akun_lama = row["Suggested Counterpart Account"]
        nama_lama = row["Suggested Counterpart Name"]
        confidence_lama = row["Confidence"]
        yakin = bool(hasil.get("yakin", False))
        akun_usulan = str(hasil.get("akun_disarankan") or akun_lama)
        nama_usulan = hasil.get("nama_akun_disarankan") or nama_lama
        alasan = hasil.get("alasan") or ""

        # [FIX 1] Validasi akun_disarankan terhadap COA sebelum dipakai --
        # lihat penjelasan lengkap di docstring fungsi ini.
        if akun_valid and akun_usulan.strip() not in akun_valid:
            jumlah_halusinasi += 1
            akun_baru, nama_baru = akun_lama, nama_lama
            yakin = False
            alasan = (
                f"[DITOLAK -- akun '{akun_usulan}' usulan Claude tidak ada di COA] "
                + alasan
            )
        else:
            akun_baru, nama_baru = akun_usulan, nama_usulan

        df_gl.at[idx, "Suggested Counterpart Account"] = akun_baru
        df_gl.at[idx, "Suggested Counterpart Name"] = nama_baru
        # Sisi Debit/Credit Account ikut diperbarui, sesuai arah transaksi
        # (Direction "CR" -> lawan jurnal ada di Debit Account karena
        # bank/staging sudah di Credit, begitu juga sebaliknya -- pola ini
        # SAMA seperti logic asli di susun_sheet_gl).
        if row["Direction"] == "CR":
            df_gl.at[idx, "Debit Account"] = akun_baru
        else:
            df_gl.at[idx, "Credit Account"] = akun_baru

        # [FIX 2] Confidence ikut naik ke "High" hanya kalau Claude yakin
        # DAN akun sudah lolos validasi COA -- lihat docstring.
        if yakin:
            df_gl.at[idx, "Confidence"] = "High"

        df_gl.at[idx, "Review Notes"] = alasan
        df_gl.at[idx, "Review Status"] = (
            "AI-Corrected (Claude)" if yakin else "Need Review (Claude tidak yakin)"
        )

        baris_status.append({
            "Journal ID": row["Journal ID"],
            "Date": row["Date"],
            "Bank Description": row["Bank Description"],
            "Confidence Sebelum": confidence_lama,
            "Confidence Sesudah": df_gl.at[idx, "Confidence"],
            "Akun Sebelum": f"{akun_lama} - {nama_lama}",
            "Akun Setelah Claude": f"{akun_baru} - {nama_baru}",
            "Alasan Claude": alasan,
            "Claude Yakin?": "Ya" if yakin else "Tidak -- tetap perlu review manual",
        })

    if jumlah_halusinasi:
        peringatan.append(
            f"{jumlah_halusinasi} baris ditolak: Claude menyarankan nomor akun "
            "yang tidak ada di COA client -- baris tetap pakai akun/saran "
            "sistem semula dan ditandai perlu review manual."
        )

    return df_gl, pd.DataFrame(baris_status), peringatan


def perbaiki_gl_dengan_claude_review(
    df_gl: pd.DataFrame,
    df_coa: pd.DataFrame,
    client_id: Optional[int] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    """Titik masuk utama. Kirim baris GL confidence non-High ke Claude API
    secara batch (lewat panggil_claude_terstruktur, structured output),
    terapkan hasilnya, dan bangun df_status untuk sheet baru.

    Return: (df_gl_diperbarui, df_status, peringatan)

    Kalau tidak ada baris yang perlu direview, ATAU panggil_claude_terstruktur
    tidak tersedia (import gagal), fungsi ini TIDAK memanggil API sama
    sekali -- return df_gl asli + df_status kosong + peringatan penjelas.

    client_id: diteruskan ke panggil_claude_terstruktur() sebagai metadata
    audit trail (dicatat sebagai str) -- sama seperti pemanggil Claude
    lain di codebase ini (mis. generate_narasi_calk_claude).
    """
    peringatan: List[str] = []

    if panggil_claude_terstruktur is None:
        peringatan.append(
            "Review Claude API dilewati: modules/claude_client.py tidak "
            "berhasil di-import (cek instalasi/lokasi file)."
        )
        return df_gl, pd.DataFrame(), peringatan

    df_perlu_review = df_gl[df_gl["Confidence"] != "High"]
    if df_perlu_review.empty:
        peringatan.append(
            "Review Claude API dilewati: semua baris GL sudah confidence "
            "High, tidak ada yang perlu direview."
        )
        return df_gl, pd.DataFrame(), peringatan

    kolom_dikirim = [
        "Journal ID", "Date", "Bank Description", "Counterparty", "Direction",
        "Amount (IDR)", "Suggested Counterpart Account", "Suggested Counterpart Name",
        "Confidence", "Classification Basis",
    ]
    baris_untuk_review = df_perlu_review[kolom_dikirim].to_dict("records")

    semua_hasil: List[Dict] = []
    jumlah_batch_gagal = 0
    for awal in range(0, len(baris_untuk_review), _UKURAN_BATCH_REVIEW):
        batch = baris_untuk_review[awal:awal + _UKURAN_BATCH_REVIEW]
        prompt = _bangun_prompt_review_gl(batch, df_coa)
        try:
            hasil = panggil_claude_terstruktur(
                prompt,
                _SYSTEM_PROMPT_REVIEW_GL_KK,
                _TOOL_REVIEW_GL_KK,
                modul_pemanggil="kertas_kerja",
                client_id=str(client_id) if client_id is not None else None,
                max_tokens=4096,
            )
            hasil_batch = hasil.get("hasil", [])
            if not isinstance(hasil_batch, list):
                raise ClaudeError(
                    f"field 'hasil' dari tool_use bukan array (dapat {type(hasil_batch).__name__})"
                )
            semua_hasil.extend(hasil_batch)
        except ClaudeError as e:
            jumlah_batch_gagal += 1
            peringatan.append(f"Batch review Claude gagal (ClaudeError): {e}")

    if jumlah_batch_gagal:
        peringatan.append(
            f"{jumlah_batch_gagal} dari "
            f"{-(-len(baris_untuk_review) // _UKURAN_BATCH_REVIEW)} batch review "
            "Claude gagal -- baris terkait TETAP pakai hasil klasifikasi awal "
            "(tidak hilang, cuma tidak sempat direview Claude)."
        )

    df_gl_baru, df_status, peringatan_terap = _terapkan_hasil_review_claude(
        df_gl, semua_hasil, df_coa
    )
    peringatan.extend(peringatan_terap)

    if not df_status.empty:
        jumlah_yakin = (df_status["Claude Yakin?"] == "Ya").sum()
        peringatan.append(
            f"Claude API mereview {len(df_status)} baris GL: {jumlah_yakin} "
            f"diperbaiki dengan yakin, {len(df_status) - jumlah_yakin} tetap "
            "ditandai perlu review manual (lihat sheet Status)."
        )

    return df_gl_baru, df_status, peringatan


# ============================================================
# 3.9 [BARU] VALIDASI STRUKTUR & KONSISTENSI HASIL AKHIR (OTOMATIS)
# ============================================================
# Dijalankan OTOMATIS di akhir generate_kertas_kerja() -- TIDAK butuh flag,
# TIDAK ada panggilan API sama sekali di sini (murni kode Python), jadi
# selalu aman & murah dijalankan setiap kali. Ada 2 lapis:
#
#   1. validasi_struktur_kertas_kerja() -- cek kolom WAJIB tiap DataFrame
#      hasil (GL, Bank_Control, dst) ADA, dibandingkan ke
#      _SPEK_KOLOM_KERTAS_KERJA di bawah (diturunkan dari file contoh
#      client "Kertas_Kerja_Laporan_Keuangan_2025.xlsx" yang sudah
#      diverifikasi tim). Tujuannya menangkap REGRESI kode (mis. ada yang
#      lupa isi kolom, salah rename, dst) SEBELUM sampai ke file Excel.
#
#   2. validasi_konsistensi_angka_kertas_kerja() -- membungkus
#      hitung_fs_control() (SUDAH ADA, dipakai jg utk sheet FS_Control)
#      jadi list temuan terstruktur, supaya bisa dipakai bareng temuan
#      struktur di atas dan (opsional) dikirim ke Claude untuk dijelaskan.
#      TIDAK menduplikasi logic hitung_fs_control() -- hanya membaca ulang
#      hasilnya dan menyaring baris yang statusnya BUKAN OK/BALANCED.
#
# Claude API (opsional, lihat generate_kertas_kerja(pakai_claude_review_final=))
# TIDAK dipakai untuk MENGHITUNG ulang angka atau memutuskan benar/salah --
# itu tetap tugas kode di atas (deterministik, presisi). Claude hanya
# dipakai untuk MENJELASKAN temuan dalam bahasa yang mudah dibaca akuntan
# & mengurutkan prioritas -- lihat jelaskan_temuan_kertas_kerja_claude()
# di claude_client.py.

# Kolom WAJIB tiap DataFrame hasil, diturunkan dari header sheet di file
# contoh client (Kertas_Kerja_Laporan_Keuangan_2025.xlsx). "statis" = nama
# kolom apa adanya. "per_bulan" = suffix yang WAJIB ada untuk SETIAP bulan
# di BULAN_URUT (mis. "Debit" -> "Jan_Debit", "Feb_Debit", ... "Dec_Debit").
_SPEK_KOLOM_KERTAS_KERJA: Dict[str, Dict[str, List[str]]] = {
    "gl": {
        "statis": [
            "Journal ID", "Date", "Month No.", "Month", "Source PDF", "Page",
            "Bank Description", "Counterparty", "Direction", "Amount (IDR)",
            "Initial Staging Account", "Suggested Counterpart Account",
            "Suggested Counterpart Name", "Confidence", "Classification Basis",
            "Debit Account", "Credit Account", "Debit (IDR)", "Credit (IDR)",
            "Review Status", "Review Notes",
        ],
        "per_bulan": [],
    },
    "bank_control": {
        "statis": [
            "Month", "Opening", "CR Statement", "DB Statement", "Closing",
            "CR Count", "DB Count", "CR Extracted", "DB Extracted",
            "Closing Calc", "Diff CR", "Diff DB", "Diff Closing",
            "Opening Continuity", "Transactions",
        ],
        "per_bulan": [],
    },
    "bank_posting_summary": {
        "statis": ["Account No.", "Account Name", "Normal Balance"],
        "per_bulan": ["Debit", "Credit"],
    },
    "adjustments": {
        "statis": ["Account No.", "Account Name", "Review Notes"],
        "per_bulan": ["Debit", "Credit"],
    },
    "tb_monthly": {
        "statis": ["Account No.", "Account Name", "Normal Balance"],
        "per_bulan": ["Debit", "Credit", "Net Helper"],
    },
    "coa": {
        "statis": _KOLOM_COA_DIBUTUHKAN + ["Account No.", "Account Name"],
        "per_bulan": [],
    },
}


def validasi_struktur_kertas_kerja(hasil: "HasilKertasKerja") -> List[Dict[str, Any]]:
    """Cek kolom WAJIB tiap DataFrame hasil ADA (tidak cek isi/nilai --
    itu tugas validasi_konsistensi_angka_kertas_kerja). Return list temuan
    {"level": "error", "area": <nama_dataframe>, "pesan": str} -- HANYA
    berisi entri kalau ada kolom yang HILANG (DataFrame yang lengkap tidak
    menyumbang temuan apa pun, list bisa kosong).

    Sengaja TIDAK memeriksa BS_Monthly/PNL_Monthly/BS_Tax/PNL_Tax/FS_Control
    di sini -- sheet-sheet itu baru dihitung penuh di tulis_kertas_kerja_excel()
    (butuh `tahun`, hasil PPh, dst yang belum tentu tersedia di titik
    generate_kertas_kerja() ini), jadi di luar cakupan pengecekan tahap ini.
    """
    temuan: List[Dict[str, Any]] = []
    for area, spek in _SPEK_KOLOM_KERTAS_KERJA.items():
        df = getattr(hasil, area, None)
        if df is None:
            temuan.append({
                "level": "error", "area": area,
                "pesan": f"DataFrame '{area}' tidak ada di HasilKertasKerja (None).",
            })
            continue
        kolom_ada = set(df.columns)
        kolom_wajib = set(spek["statis"])
        for suffix in spek["per_bulan"]:
            kolom_wajib |= {f"{b}_{suffix}" for b in BULAN_URUT}
        hilang = sorted(kolom_wajib - kolom_ada)
        if hilang:
            temuan.append({
                "level": "error", "area": area,
                "pesan": (
                    f"Sheet '{area}' kehilangan {len(hilang)} kolom wajib: "
                    f"{', '.join(hilang[:10])}" + (" ..." if len(hilang) > 10 else "")
                ),
            })
    return temuan


def validasi_konsistensi_angka_kertas_kerja(
    hasil: "HasilKertasKerja", tahun: int,
) -> List[Dict[str, Any]]:
    """Bungkus hitung_fs_control() (logic SUDAH ADA, tidak diduplikasi)
    jadi list temuan terstruktur -- HANYA baris yang statusnya BUKAN
    OK/BALANCED yang dimasukkan (working paper yang baru digenerate dari
    PDF MEMANG WAJAR belum "BALANCED" di semua baris -- Opening_Balance &
    Adjustments belum diisi akuntan, lihat docstring hitung_fs_control --
    jadi ini bukan berarti kode error, tapi tetap perlu ditunjukkan ke
    user sebagai daftar "yang masih perlu direview manual").

    Level "error" untuk GL/TB/BS tidak balance (ini indikasi bug/data
    rusak, BUKAN hal yang wajar). Level "warning" untuk status lain
    (Opening_Balance/Adjustments belum diisi -- ini memang alur normal,
    bukan indikasi kode salah)."""
    temuan: List[Dict[str, Any]] = []
    try:
        df_control = hitung_fs_control(hasil, tahun)
    except Exception as e:  # noqa: BLE001 -- validator tidak boleh menggagalkan generate
        temuan.append({
            "level": "error", "area": "FS_Control",
            "pesan": f"Gagal menghitung FS_Control untuk validasi: {type(e).__name__}: {e}",
        })
        return temuan

    _LEVEL_ERROR_UNTUK = ("gl balanced", "tb_monthly balanced", "bs_monthly balanced")
    for _, r in df_control.iterrows():
        label = str(r["Control"])
        status = str(r["Status"])
        if status in ("BALANCED", "OK"):
            continue
        # "BELUM ADA DATA" (mis. BS_Monthly belum ada transaksi sama sekali)
        # BUKAN indikasi tidak balance -- beda dari "CHECK" (dihitung tapi
        # selisih). Diturunkan jadi "warning", bukan "error", supaya tidak
        # jadi false-positive di kertas kerja yang memang masih kosong.
        if status == "BELUM ADA DATA":
            temuan.append({
                "level": "warning", "area": "FS_Control",
                "pesan": f"{label}: status={status} (belum ada transaksi untuk dihitung, bukan indikasi error).",
            })
            continue
        level = "error" if any(k in label.lower() for k in _LEVEL_ERROR_UNTUK) else "warning"
        temuan.append({
            "level": level, "area": "FS_Control",
            "pesan": f"{label}: status={status}" + (f", value={r['Value']}" if pd.notna(r.get("Value")) else ""),
        })
    return temuan


def jalankan_validasi_otomatis_kertas_kerja(
    hasil: "HasilKertasKerja", tahun: int,
) -> List[Dict[str, Any]]:
    """Titik masuk gabungan -- dipanggil OTOMATIS di akhir
    generate_kertas_kerja(), tanpa flag/opsional (murni kode, tidak ada
    panggilan API). Return list temuan gabungan (struktur + konsistensi
    angka); TIDAK melempar exception -- kegagalan validasi itu sendiri
    dicatat sbg temuan "error", bukan menggagalkan generate."""
    temuan = validasi_struktur_kertas_kerja(hasil)
    temuan.extend(validasi_konsistensi_angka_kertas_kerja(hasil, tahun))
    return temuan


# ============================================================
# 4. BANK_CONTROL (rekonsiliasi bulanan)
# ============================================================

def susun_bank_control(
    df_gl: pd.DataFrame,
    saldo_awal_per_bulan: Optional[Dict[int, float]] = None,
    saldo_statement_per_bulan: Optional[Dict[int, Dict[str, float]]] = None,
) -> pd.DataFrame:
    """
    Rekap per bulan: total CR/DB dari GL (extracted), jumlah transaksi,
    dan opening/closing balance kalau tersedia. `saldo_awal_per_bulan`
    (opsional, {1: saldo_awal_jan, ...}) dipakai kalau ada data saldo
    riil dari statement -- kalau tidak diisi, opening/closing dihitung
    murni dari akumulasi CR/DB GL dimulai dari 0 (BUKAN saldo riil bank,
    hanya untuk cross-check jumlah transaksi tereksrak).

    [STRUKTUR - kolom disamakan dgn contoh Kertas_Kerja_Laporan_Keuangan]
    Kolom "CR Statement"/"DB Statement"/"Closing" (saldo & mutasi ASLI yang
    TERTULIS di rekening koran, dipakai utk verifikasi silang thd hasil
    ekstraksi PDF) dan turunannya ("Diff CR"/"Diff DB"/"Diff Closing")
    BELUM bisa diisi otomatis -- parser PDF saat ini
    (susun_gl_dari_pdf_rekening_koran / akuntansi_ai.parse_sheet_bank)
    TIDAK menangkap kolom saldo berjalan per baris dari tabel rekening
    koran, hanya mutasi debit/kredit per transaksi. Kolom-kolom ini
    SENGAJA None dulu (kerangka/struktur saja) -- TODO lanjutan: parser
    PDF perlu ditambah utk menangkap kolom "Saldo" per baris, lalu opsi
    `saldo_statement_per_bulan` di sini diisi dari situ supaya
    rekonsiliasi (Diff CR/DB/Closing) beneran jalan.

    `saldo_statement_per_bulan` (opsional, {bulan_no: {"cr": ..., "db":
    ..., "closing": ...}}) -- kalau nanti sudah tersedia, dipakai utk
    mengisi kolom Statement & Diff. Untuk saat ini biasanya None/kosong
    dari pemanggil, jadi kolom-kolom itu tetap None (bukan ditebak).

    "Opening Continuity" DIHITUNG (bukan placeholder) -- selisih antara
    Opening bulan ini vs Closing Calc bulan sebelumnya. Karena Opening
    bulan ini SELALU didefinisikan = saldo_berjalan (closing bulan lalu)
    kecuali dioverride via saldo_awal_per_bulan, nilainya akan 0 secara
    konsisten kalau tidak ada override -- kolom ini tetap berguna sbg
    cek kalau suatu saat saldo_awal_per_bulan dipakai (mis. opening bulan
    tertentu di-override manual dari data lain, lalu ketahuan bedanya).
    """
    saldo_statement_per_bulan = saldo_statement_per_bulan or {}
    baris = []
    saldo_berjalan = 0.0
    closing_bulan_lalu: Optional[float] = None
    for bulan_no in range(1, 13):
        bulan = BULAN_URUT[bulan_no - 1]
        subset = df_gl[df_gl["Month No."] == bulan_no]
        cr = subset.loc[subset["Direction"] == "CR", "Amount (IDR)"].sum()
        db = subset.loc[subset["Direction"] == "DB", "Amount (IDR)"].sum()
        opening = saldo_awal_per_bulan.get(bulan_no, saldo_berjalan) if saldo_awal_per_bulan else saldo_berjalan
        closing = opening + cr - db
        saldo_berjalan = closing

        statement_bulan = saldo_statement_per_bulan.get(bulan_no) or {}
        cr_statement = statement_bulan.get("cr")
        db_statement = statement_bulan.get("db")
        closing_statement = statement_bulan.get("closing")

        opening_continuity = None if closing_bulan_lalu is None else round(opening - closing_bulan_lalu, 2)
        closing_bulan_lalu = closing

        baris.append({
            "Month": bulan,
            "Opening": opening,
            "CR Statement": cr_statement,
            "DB Statement": db_statement,
            "Closing": closing_statement,
            "CR Count": int((subset["Direction"] == "CR").sum()),
            "DB Count": int((subset["Direction"] == "DB").sum()),
            "CR Extracted": cr,
            "DB Extracted": db,
            "Closing Calc": closing,
            "Diff CR": round(cr_statement - cr, 2) if cr_statement is not None else None,
            "Diff DB": round(db_statement - db, 2) if db_statement is not None else None,
            "Diff Closing": round(closing_statement - closing, 2) if closing_statement is not None else None,
            "Opening Continuity": opening_continuity,
            "Transactions": len(subset),
            # [FIX] Sebelumnya "len(subset) > 0 or bulan_no == 1" -- ini bug,
            # bulan Januari akan selalu berstatus "OK" walau 0 transaksi
            # (mis. rekening koran yang diupload tidak mencakup Januari sama
            # sekali). Status sekarang murni berdasarkan ada/tidaknya
            # transaksi di bulan tsb, konsisten untuk 12 bulan.
            "Status": "OK" if len(subset) > 0 else "Tidak Ada Data",
        })
    return pd.DataFrame(baris)


# ============================================================
# 5. BANK_POSTING_SUMMARY (per akun x bulan)
# ============================================================

def susun_bank_posting_summary(df_gl: pd.DataFrame, df_coa: pd.DataFrame) -> pd.DataFrame:
    """Rekap debit/credit per akun (Account No.) x bulan, dari sheet GL.

    [BARU -- PERBAIKAN PERFORMA -- VEKTORISASI] Sebelumnya loop per akun
    COA x 12 bulan, dan TIAP iterasi memfilter ULANG seluruh df_gl (2x
    boolean mask penuh -- Debit Account & Credit Account) -- untuk N akun
    COA ini jadi N x 12 x 2 scan penuh ke df_gl (mis. COA 60 akun x 5000
    baris GL = 1.336 detik). Sekarang cukup 2x groupby+pivot di awal
    (Debit & Credit dipisah, karena masing-masing dikelompokkan ke KOLOM
    akun yang berbeda -- Debit Account vs Credit Account), lalu
    di-reindex ke daftar akun COA supaya akun yang tidak muncul sama
    sekali di GL tetap dapat barisnya dengan nilai 0 (perilaku IDENTIK
    dengan versi lama, yang juga selalu menghasilkan 1 baris per akun COA
    meski jumlahnya 0). Diverifikasi menghasilkan angka SAMA PERSIS
    dengan versi lama di data sintetis (60 akun x 5000 baris GL), ~121x
    lebih cepat (1.336 detik -> 0.011 detik)."""
    no_akun_coa = df_coa["Account No."].astype(str)

    pivot_debit = (
        df_gl.groupby([df_gl["Debit Account"].astype(str), "Month No."])["Debit (IDR)"]
        .sum()
        .unstack("Month No.")
        .reindex(index=no_akun_coa, columns=range(1, 13), fill_value=0.0)
        .fillna(0.0)
    )
    pivot_credit = (
        df_gl.groupby([df_gl["Credit Account"].astype(str), "Month No."])["Credit (IDR)"]
        .sum()
        .unstack("Month No.")
        .reindex(index=no_akun_coa, columns=range(1, 13), fill_value=0.0)
        .fillna(0.0)
    )

    out = pd.DataFrame({
        "Account No.": no_akun_coa.values,
        "Account Name": df_coa["Account Name"].values,
        "Normal Balance": (
            df_coa["Normal Balance"].values if "Normal Balance" in df_coa.columns
            else [None] * len(df_coa)
        ),
    })
    for bulan_no, bulan in enumerate(BULAN_URUT, start=1):
        out[f"{bulan}_Debit"] = pivot_debit[bulan_no].values
        out[f"{bulan}_Credit"] = pivot_credit[bulan_no].values
    return out


# ============================================================
# 6. ADJUSTMENTS (template kosong utk koreksi manual)
# ============================================================

def susun_template_adjustments(df_coa: pd.DataFrame) -> pd.DataFrame:
    """Template kosong (semua 0) per akun x bulan -- diisi manual oleh akuntan
    setelah review GL/Bank_Posting_Summary, lalu di-upload ulang."""
    baris = []
    for _, akun in df_coa.iterrows():
        rec = {"Account No.": str(akun["Account No."]), "Account Name": akun["Account Name"], "Review Notes": None}
        for bulan in BULAN_URUT:
            rec[f"{bulan}_Debit"] = 0
            rec[f"{bulan}_Credit"] = 0
        baris.append(rec)
    return pd.DataFrame(baris)


# ============================================================
# 7. TB_MONTHLY = Bank_Posting_Summary + Adjustments
# ============================================================

def susun_tb_monthly(df_posting: pd.DataFrame, df_adjustments: pd.DataFrame) -> pd.DataFrame:
    """
    [STRUKTUR - disamakan dgn contoh Kertas_Kerja_Laporan_Keuangan] Selain
    kolom {bulan}_Debit/{bulan}_Credit (posting + adjustment), ditambah 12
    kolom {bulan}_Net Helper -- kolom bantu (Debit - Credit per bulan)
    yang di contoh dipakai sbg perantara sebelum disusun jadi BS/PNL
    Monthly (mis. utk cek cepat arah saldo per akun tanpa lihat 2 kolom
    terpisah). Nilainya DIHITUNG (bukan placeholder) karena tinggal
    selisih dari kolom Debit/Credit yang sudah ada -- TIDAK memakai info
    Normal Balance (jadi sifatnya "net mentah", bukan "saldo normal");
    logic saldo per Normal Balance yang sesungguhnya tetap ada terpisah
    di susun_bs_pnl_monthly().
    """
    df = df_posting.merge(
        df_adjustments, on=["Account No.", "Account Name"], suffixes=("_posting", "_adj"),
    )
    baris = []
    for _, r in df.iterrows():
        rec = {"Account No.": r["Account No."], "Account Name": r["Account Name"],
               "Normal Balance": r.get("Normal Balance")}
        for bulan in BULAN_URUT:
            rec[f"{bulan}_Debit"] = (r.get(f"{bulan}_Debit_posting") or 0) + (r.get(f"{bulan}_Debit_adj") or 0)
            rec[f"{bulan}_Credit"] = (r.get(f"{bulan}_Credit_posting") or 0) + (r.get(f"{bulan}_Credit_adj") or 0)
        for bulan in BULAN_URUT:
            rec[f"{bulan}_Net Helper"] = rec[f"{bulan}_Debit"] - rec[f"{bulan}_Credit"]
        baris.append(rec)
    return pd.DataFrame(baris)


# ============================================================
# 8. BS_MONTHLY / PNL_MONTHLY (dari TB_Monthly + COA FS Group)
# ============================================================

def susun_bs_pnl_monthly(
    df_tb: pd.DataFrame, df_coa: pd.DataFrame, df_opening: Optional[pd.DataFrame] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Gabungkan TB_Monthly dengan kolom Statement (BS/PNL) & FS Group di
    COA, hitung saldo bersih per bulan (Debit - Credit kalau Normal
    Balance = Debit, sebaliknya kalau Credit), akumulasi untuk BS
    (saldo berjalan) atau flow murni per bulan untuk PNL.

    [STRUKTUR - disamakan dgn contoh Kertas_Kerja_Laporan_Keuangan] Kolom
    "FS Group" TETAP ada di DataFrame yang dikembalikan (dipakai sbg kunci
    pengelompokan saat ditulis ke Excel -- lihat _tulis_sheet_bs_pnl_monthly),
    TAPI TIDAK ikut ditulis sbg kolom di sheet -- di contoh, grouping
    ditampilkan sbg BARIS header per kelompok (mis. "ASSETS",
    "OPERATING REVENUE"), bukan kolom tambahan. Kolom "Total <tahun>"
    (jumlah 12 bulan) ditambahkan KHUSUS utk PNL saat penulisan (BS tidak
    perlu -- BS itu snapshot per bulan, bukan flow yg masuk akal dijumlah).

    CATATAN KETERBATASAN (structural-only fix, BUKAN full logic): contoh
    replika py juga punya baris SUBTOTAL per kelompok (mis. "TOTAL ASSETS",
    "Total Operating Revenue", "GROSS PROFIT", "NET PROFIT / (LOSS)") --
    baris-baris itu SENGAJA belum direproduksi di sini (itu logic
    perhitungan/konten, bukan struktur kolom) -- TODO lanjutan kalau
    dibutuhkan subtotal per kelompok yang akurat.

    [FIX -- menutup gap #3, lihat TODO lama di hitung_fs_control()]
    df_opening (opsional, skema sheet Opening_Balance: kolom "Account
    No."/"Opening Debit (IDR)"/"Opening Credit (IDR)") kalau diisi & tidak
    kosong, dipakai sbg titik awal akumulasi saldo BS per akun -- BUKAN
    selalu mulai dari 0 seperti sebelumnya. Ini supaya BS_Monthly yang
    ditampilkan di working paper untuk direview user SINKRON dengan Neraca
    di laporan 18-sheet final (yang sudah pakai Opening_Balance lewat
    hitung_saldo_per_akun_dari_tb_monthly()).

    Saldo awal per akun disimpan SIGNED sesuai arah Normal Balance-nya
    (Opening Debit - Opening Credit kalau normal saldo Debit, sebaliknya
    kalau Credit) -- SENGAJA BEDA dari konvensi "raw debit-credit" yang
    dipakai konversi_coa_kertas_kerja_ke_peta()/
    hitung_saldo_per_akun_dari_tb_monthly() (lihat KETERBATASAN DIKETAHUI
    di bawah), supaya konsisten dgn variabel `net` bulanan di fungsi ini
    sendiri yang JUGA sudah signed per Normal Balance -- tanpa ini,
    mencampur saldo awal "raw" dgn pergerakan bulanan yang "signed" akan
    salah untuk akun Normal Balance Credit (liabilitas/ekuitas).

    Kalau df_opening None/kosong (mis. saat kertas kerja baru pertama kali
    digenerate, sebelum user isi saldo awal -- lihat generate_kertas_kerja),
    perilaku PERSIS SAMA seperti sebelumnya (mulai dari 0).

    PNL TIDAK terpengaruh df_opening -- akun PNL secara akuntansi tidak
    punya saldo awal berjalan (flow murni per tahun buku, selalu mulai
    dari 0 tiap tahun).

    [KETERBATASAN DIKETAHUI -- DI LUAR CAKUPAN FIX INI, TOLONG DICEK TIM]
    hitung_saldo_per_akun_dari_tb_monthly() (dipakai jembatan ke laporan
    18-sheet, susun_data_export_18_sheet_dari_kertas_kerja) menyimpan &
    memakai saldo_awal dengan konvensi BERBEDA (raw debit-credit, TIDAK
    di-flip sesuai Normal Balance) dikombinasikan dgn
    "saldo_akhir = saldo_awal - pergerakan" khusus utk akun KREDIT --
    dari pengecekan manual dgn angka contoh, kombinasi ini tampak
    menghasilkan saldo akhir yang SALAH utk akun Normal Balance Credit
    (liabilitas/ekuitas/pendapatan) begitu Opening_Balance akun itu diisi
    bukan 0 (di file contoh saat ini semua Opening_Balance masih 0, jadi
    belum kelihatan efeknya). BUKAN bagian dari gap #3 yang diminta &
    BELUM diubah di sini -- cuma dicatat supaya tim sadar sebelum
    Opening_Balance client sungguhan mulai diisi utk akun Kredit.
    """
    coa_map = df_coa.set_index(df_coa["Account No."].astype(str))[["Account Name", "Normal Balance", "Statement", "FS Group"]]

    saldo_awal_map: Dict[str, float] = {}
    if df_opening is not None and not df_opening.empty:
        for _, r in df_opening.iterrows():
            no_akun = str(r.get("Account No.") or "").strip()
            if not no_akun or no_akun not in coa_map.index:
                continue
            debit = pd.to_numeric(r.get("Opening Debit (IDR)"), errors="coerce")
            credit = pd.to_numeric(r.get("Opening Credit (IDR)"), errors="coerce")
            debit = float(debit) if pd.notna(debit) else 0.0
            credit = float(credit) if pd.notna(credit) else 0.0
            normal_balance = str(coa_map.loc[no_akun]["Normal Balance"] or "").strip().lower()
            saldo_awal_map[no_akun] = (debit - credit) if normal_balance == "debit" else (credit - debit)

    baris_bs, baris_pnl = [], []
    # BS mulai akumulasi dari saldo awal (0 kalau akun tsb tidak ada di
    # df_opening/df_opening tidak diberikan -- perilaku lama, tidak berubah).
    saldo_akumulasi = dict(saldo_awal_map)
    for _, r in df_tb.iterrows():
        no_akun = str(r["Account No."])
        if no_akun not in coa_map.index:
            continue
        info = coa_map.loc[no_akun]
        statement = info["Statement"]
        rec = {"Account / Description": f"{no_akun} - {info['Account Name']}"}
        akumulasi = saldo_akumulasi.get(no_akun, 0.0)
        total_tahun = 0.0
        for bulan in BULAN_URUT:
            debit = r.get(f"{bulan}_Debit") or 0
            credit = r.get(f"{bulan}_Credit") or 0
            net = (debit - credit) if str(info["Normal Balance"]).lower() == "debit" else (credit - debit)
            if statement == "BS":
                akumulasi += net
                rec[bulan] = akumulasi
            else:
                rec[bulan] = net
                total_tahun += net
        saldo_akumulasi[no_akun] = akumulasi
        rec["FS Group"] = info["FS Group"]
        if statement == "PNL":
            rec["Total Tahun"] = total_tahun
        (baris_bs if statement == "BS" else baris_pnl).append(rec)

    return pd.DataFrame(baris_bs), pd.DataFrame(baris_pnl)


# ============================================================
# 8.5 [BARU - Step 2] KERANGKA: FS_CONTROL, OPENING_BALANCE, BS_TAX, PNL_TAX
# ============================================================
# Sheet-sheet ini sebelumnya hanya _tulis_sheet_placeholder (teks "TODO"
# apa adanya) -- sekarang diisi KERANGKA (struktur tabel siap pakai,
# nilai masih None/0), sesuai rencana step 2. Perhitungan/pengisian nilai
# SUNGGUHAN (status BALANCED/CHECK di FS_Control, koreksi fiskal di
# BS_Tax/PNL_Tax) sengaja BELUM ditambahkan di sini -- itu langkah
# berikutnya, supaya tidak menampilkan status yang salah sebelum logic
# penghitungannya benar-benar dites dengan data riil.

def susun_fs_control_kerangka(tahun: int) -> pd.DataFrame:
    """
    Kerangka sheet FS_Control: baris label pemeriksaan (control check)
    yang akan dihitung otomatis di iterasi berikutnya (dipakai sbg syarat
    sebelum user boleh konfirmasi lanjut generate laporan 18-sheet, lihat
    langkah "Sambungkan konfirmasi user" di rencana kertas kerja). Kolom
    Value & Status SENGAJA masih None di versi ini.
    """
    label_kontrol = [
        f"GL Balanced (Total Debit = Total Credit) - {tahun}",
        "Bank_Control - Semua Bulan Berstatus OK",
        "TB_Monthly Balanced (Total Debit = Total Credit)",
        "BS_Monthly Balanced (Total Assets = Total Liabilities + Equity)",
        "Opening_Balance Sudah Diisi (bukan seluruhnya 0)",
        "Adjustments Sudah Direview Akuntan",
        f"OVERALL STATUS KERTAS KERJA {tahun}",
    ]
    return pd.DataFrame([
        {"Control": label, "Value": None, "Target / Reference": None, "Status": None}
        for label in label_kontrol
    ])


def susun_opening_balance(df_coa: pd.DataFrame) -> pd.DataFrame:
    """
    Kerangka Opening_Balance: 1 baris per akun COA -- INPUT MANUAL
    client/akuntan (saldo per 1 Januari tahun buku). Working paper yang
    digenerate dari PDF rekening koran TIDAK BISA tahu saldo awal ini
    (rekening koran yang diupload biasanya hanya mencakup tahun berjalan,
    bukan histori dari awal berdirinya usaha) -- makanya nilainya WAJIB
    diisi manual, bukan ditebak/dihitung otomatis dari data yang ada.

    [STRUKTUR - disamakan dgn contoh Kertas_Kerja_Laporan_Keuangan] Kolom
    "Opening Balance (IDR)" tunggal (versi lama) DIPECAH jadi "Opening
    Debit (IDR)" & "Opening Credit (IDR)" terpisah (default 0 keduanya --
    akuntan isi salah satu sesuai Normal Balance akun saat review), lalu
    "Net Debit/(Credit)" = Opening Debit - Opening Credit (dihitung, bukan
    input manual), "Review Notes" (input manual, None dulu), dan "Status"
    default "Not Entered" (berubah jadi "Entered"/dst oleh akuntan saat
    kolom Debit/Credit sudah diisi -- versi ini belum ada logic auto-ubah
    status, akuntan set manual saat review, sesuai kolom Review Notes yg
    juga manual).
    """
    baris = []
    for _, akun in df_coa.iterrows():
        opening_debit = 0
        opening_credit = 0
        baris.append({
            "Account No.": str(akun["Account No."]),
            "Account Name": akun["Account Name"],
            "Opening Debit (IDR)": opening_debit,
            "Opening Credit (IDR)": opening_credit,
            "Net Debit/(Credit)": opening_debit - opening_credit,
            "Review Notes": None,
            "Status": "Not Entered",
        })
    return pd.DataFrame(baris)


def susun_bs_tax_kerangka(df_coa: pd.DataFrame) -> pd.DataFrame:
    """
    Kerangka BS_Tax -- akun ber-Statement 'BS'.

    [STRUKTUR - disamakan dgn contoh Kertas_Kerja_Laporan_Keuangan] Skema
    kolom BS_Tax BEDA dari PNL_Tax (lihat susun_pnl_tax_kerangka) --
    "Account No." + "Account Name" DIGABUNG jadi 1 kolom "Account /
    Description" (format "{no} - {nama}", sama seperti BS_Monthly/
    PNL_Monthly), "Tax Return Group" di-rename jadi "Tax Return Category"
    (label kolom persis di contoh), dan ada kolom baru "Review / Tax Note"
    (default None -- input manual akuntan). Amount (IDR) masih 0 di
    kerangka ini, diisi oleh isi_bs_tax_dengan_saldo().
    """
    if df_coa is None or df_coa.empty or "Statement" not in df_coa.columns:
        return pd.DataFrame(columns=["Tax Return Category", "Account / Description", "Amount (IDR)", "Review / Tax Note"])

    subset = df_coa[df_coa["Statement"] == "BS"]
    baris = []
    for _, akun in subset.iterrows():
        baris.append({
            "Tax Return Category": akun.get("Tax Return Group"),
            "Account / Description": f"{akun['Account No.']} - {akun['Account Name']}",
            "Amount (IDR)": 0,
            "Review / Tax Note": None,
        })
    return pd.DataFrame(baris)


def susun_pnl_tax_kerangka(df_coa: pd.DataFrame) -> pd.DataFrame:
    """
    Kerangka PNL_Tax -- akun ber-Statement 'PNL'.

    [STRUKTUR - disamakan dgn contoh Kertas_Kerja_Laporan_Keuangan] BEDA
    dari BS_Tax: "Account No."/"Account Name" TETAP terpisah (tidak
    digabung), kolom tetap "Tax Return Group" (bukan "Category"), dan ada
    5 kolom breakdown koreksi fiskal yang TIDAK ada di BS_Tax:
    "Accounting Contribution" (angka akuntansi murni, sebelumnya bernama
    "Amount (IDR)" di versi lama -- diisi isi_pnl_tax_dengan_saldo()),
    "Auto Positive"/"Auto Negative" (koreksi fiskal yang BISA dideteksi
    otomatis dari data, mis. bunga bank final -- BELUM ada logic
    deteksinya di iterasi ini, default 0), "Manual Positive"/
    "Manual Negative" (input manual akuntan, default 0), dan
    "Fiscal Contribution" (= Accounting Contribution + Auto Positive -
    Auto Negative + Manual Positive - Manual Negative, dihitung otomatis
    di isi_pnl_tax_dengan_saldo -- akan SAMA dgn Accounting Contribution
    selama Auto/Manual masih 0 semua), plus "Review Notes" (manual, None).
    """
    if df_coa is None or df_coa.empty or "Statement" not in df_coa.columns:
        return pd.DataFrame(columns=[
            "Account No.", "Account Name", "Tax Return Group", "Accounting Contribution",
            "Auto Positive", "Auto Negative", "Manual Positive", "Manual Negative",
            "Fiscal Contribution", "Review Notes",
        ])

    subset = df_coa[df_coa["Statement"] == "PNL"]
    baris = []
    for _, akun in subset.iterrows():
        baris.append({
            "Account No.": str(akun["Account No."]),
            "Account Name": akun["Account Name"],
            "Tax Return Group": akun.get("Tax Return Group"),
            "Accounting Contribution": 0,
            "Auto Positive": 0,
            "Auto Negative": 0,
            "Manual Positive": 0,
            "Manual Negative": 0,
            "Fiscal Contribution": 0,
            "Review Notes": None,
        })
    return pd.DataFrame(baris)


# ============================================================
# 8b. [BARU -- FIX GAP] LAPISAN KATA KUNCI KHUSUS KERTAS KERJA
# ============================================================
# MASALAH YANG DIPERBAIKI: akuntansi_ai.cocokkan_kata_kunci_ke_coa() (tahap
# 2 di proses_dataframe(), dipanggil generate_kertas_kerja() lewat
# akuntansi_ai.proses_dataframe()) memakai KATA_KUNCI_AKUN yang isinya
# singkatan akun ala-Indonesia (mis. "ADM BANK", "LISTRIK", "SEWA", "BBM")
# dicari sbg SUBSTRING PERSIS di df_coa["nama_akun"]. COA skema kertas
# kerja (contoh: Kertas_Kerja_Laporan_Keuangan_2025.xlsx) pakai nama akun
# BAHASA INGGRIS ("Bank Charges", "Rental Revenue", dst) -- substring
# "ADM BANK" tidak akan PERNAH cocok dgn "BANK CHARGES", jadi utk COA
# berbahasa Inggris, lapisan kata-kunci itu MATI TOTAL (0% match rate,
# diverifikasi lewat test dgn PDF rekening koran riil) -- semua transaksi
# yang gagal dikenali pola historis client langsung lompat ke AI (kalau
# pakai_ai=True) atau ke staging (kalau AI mati/gagal), padahal
# sebenarnya banyak transaksi (biaya admin bank, sewa/rental, BBM, gaji,
# listrik, dst) punya penanda teks yang JELAS & AMAN dikenali tanpa AI.
#
# FUNGSI DI BAWAH INI ADALAH LAPISAN TERPISAH (TIDAK mengubah
# akuntansi_ai.py sama sekali -- supaya tidak mengganggu fitur lain yang
# masih bergantung ke KATA_KUNCI_AKUN versi lama/Indonesia), dipasang di
# generate_kertas_kerja() SETELAH akuntansi_ai.proses_dataframe() selesai,
# HANYA menyentuh baris yang MASIH belum berhasil diklasifikasi (no_akun
# debet/kredit masih kosong) -- baris yang sudah kena pola historis client
# atau AI TIDAK disentuh/ditimpa sama sekali.
#
# Pencarian dilakukan bilingual terhadap df_coa["Account Name"] DAN
# df_coa["Notes"] sekaligus (Notes pada COA kertas-kerja biasanya berisi
# deskripsi Bahasa Indonesia -- lihat contoh: akun 5205 "Bank Charges"
# Notes="Biaya bank") -- supaya cocok baik COA Inggris, Indonesia, maupun
# campuran, TANPA perlu tahu bahasa COA client sebelumnya.
#
# Confidence sengaja dibagi 2 tingkat:
#   - "High"  : hanya utk pola yang SANGAT deterministik & tidak ambigu
#               (biaya transaksi/admin bank -- narasi baku dari sistem
#               bank sendiri, bukan input manual customer).
#   - "Medium": kategori umum lain (BBM, gaji, listrik, telekomunikasi,
#               asuransi, sewa/rental) -- match kata kunci cukup kuat,
#               tapi tetap disarankan direview akuntan (beda dgn "Low"/
#               staging yang WAJIB direview).
# Kategori yang ambigu/butuh judgement (mis. transfer masuk/keluar tanpa
# keterangan jelas, tarik tunai ATM, top-up e-wallet) SENGAJA TIDAK
# dimasukkan ke sini -- itu memang harus tetap staging (Customer
# Deposit/Prepaid Expenses/Operational Advances), sesuai desain asli file
# contoh kamu.

KATA_KUNCI_KERTAS_KERJA: List[Dict[str, Any]] = [
    # --- Biaya transaksi/admin bank (DETERMINISTIK -- narasi baku bank) ---
    {
        "kata": ["biaya txn", "biaya transaksi", "txn fee", "biaya adm",
                 "adm bank", "administrasi bank", "biaya bulanan rekening",
                 # [BARU] variasi tambahan -- semuanya narasi baku sistem
                 # bank (bukan input manual customer), sama amannya dgn
                 # entri "biaya adm bank" yang sudah ada.
                 "biaya kliring", "kliring", "biaya rtgs", "rtgs", "biaya skn", "skn ",
                 "biaya transfer antar bank", "biaya transfer online", "biaya switching",
                 "bank garansi", "biaya provisi", "provisi kredit",
                 "biaya sms banking", "biaya mobile banking", "biaya cek/bg", "biaya cek bg"],
        "arah": "DB",
        "penanda": ["bank charges", "biaya bank", "beban bank", "biaya administrasi bank",
                    "admin bank", "biaya adm bank"],
        "filter_class": ["Expense"], "filter_statement": "PNL",
        "confidence": "High",
        "alasan": "Biaya transaksi/administrasi bank (narasi baku dari sistem bank)",
    },
    # --- Sewa / rental (revenue) -- hanya utk uang MASUK ---
    {
        "kata": ["sewa", "rental", "scaffolding", "scafolding", "skafolding",
                 "cafolding", "stager"],
        "arah": "CR",
        "penanda": ["rental revenue", "pendapatan sewa", "sewa"],
        "filter_class": ["Revenue"], "filter_statement": "PNL",
        "confidence": "Medium",
        "alasan": "Mengandung kata kunci sewa/rental pada narasi transaksi masuk",
    },
    # --- Bahan bakar ---
    {
        "kata": ["bbm", "spbu", "pertamina", "shell ", "bensin", "solar "],
        "arah": "DB",
        "penanda": ["fuel", "bbm", "bahan bakar"],
        "filter_class": ["Expense"], "filter_statement": "PNL",
        "confidence": "Medium",
        "alasan": "Mengandung kata kunci BBM/SPBU pada narasi transaksi",
    },
    # --- Gaji / payroll ---
    {
        "kata": ["gaji", "payroll"],
        "arah": "DB",
        "penanda": ["salaries", "salary", "gaji", "upah"],
        "filter_class": ["Expense"], "filter_statement": "PNL",
        "confidence": "Medium",
        "alasan": "Mengandung kata kunci gaji/payroll pada narasi transaksi",
    },
    # --- Listrik ---
    {
        "kata": ["listrik", "pln "],
        "arah": "DB",
        "penanda": ["electricity", "listrik", "utilities", "utilitas"],
        "filter_class": ["Expense"], "filter_statement": "PNL",
        "confidence": "Medium",
        "alasan": "Mengandung kata kunci listrik/PLN pada narasi transaksi",
    },
    # --- Telekomunikasi ---
    {
        "kata": ["telkom", "internet", "wifi", "telepon", "indihome", "pulsa"],
        "arah": "DB",
        "penanda": ["telephone", "internet", "telekomunikasi", "komunikasi"],
        "filter_class": ["Expense"], "filter_statement": "PNL",
        "confidence": "Medium",
        "alasan": "Mengandung kata kunci telekomunikasi pada narasi transaksi",
    },
    # --- Asuransi ---
    {
        "kata": ["asuransi", "insurance", "premi"],
        "arah": "DB",
        "penanda": ["insurance", "asuransi"],
        "filter_class": ["Expense"], "filter_statement": "PNL",
        "confidence": "Medium",
        "alasan": "Mengandung kata kunci asuransi/premi pada narasi transaksi",
    },
    # --- [BARU] Bunga bank & pajak final atas bunga -- baris ini MUNCUL DI
    # SETIAP rekening koran, SETIAP bank, SETIAP bulan (baris tutup buku
    # akhir bulan) -- narasi baku 100% dari sistem bank, bukan input manual,
    # jadi paling aman utk exact-match & confidence tinggi. [VERIFIED] dari
    # PDF BCA Tahapan asli (baris terakhir tiap bulan: "BUNGA" arah CR +
    # "PAJAK BUNGA" arah DB tepat di bawahnya). exact=True supaya tidak
    # kebablasan menangkap keterangan lain yang kebetulan memuat kata
    # "bunga" (mis. nama merchant "TOKO BUNGA ..."). Bank lain mungkin
    # pakai label sedikit beda (mis. "PPH BUNGA") -- tambahkan variasinya
    # di sini kalau ketemu di PDF bank lain, jangan ganti jadi substring
    # biasa (exact tetap dipertahankan demi keamanan).
    {
        "kata": ["bunga"],
        "arah": "CR",
        "exact": True,
        "penanda": ["interest income", "bunga bank", "pendapatan bunga"],
        "filter_class": ["Other Income", "Revenue"], "filter_statement": "PNL",
        "confidence": "High",
        "alasan": "Baris bunga bank bulanan (narasi baku sistem bank, bukan input manual)",
    },
    {
        "kata": ["pajak bunga", "pph bunga"],
        "arah": "DB",
        "exact": True,
        "penanda": ["final tax on bank interest", "pph final bunga", "pajak bunga"],
        "filter_class": ["Other Expense", "Expense"], "filter_statement": "PNL",
        "confidence": "High",
        "alasan": "Baris pajak final atas bunga bank bulanan (narasi baku sistem bank, bukan input manual)",
    },
    # --- [BARU] Kategori tambahan -- semua Medium (bukan High) karena narasi
    # ini BISA ditulis manual oleh customer/teller, beda dgn bunga/biaya
    # admin bank yang selalu baku dari sistem. Tetap disarankan spot-check,
    # tapi cukup kuat untuk mengurangi beban AI secara signifikan karena
    # kategori2 ini sering muncul berulang-ulang di rekening koran usaha.
    # --- Sewa/rental -- arah KELUAR (perusahaan MEMBAYAR sewa; beda dari
    # entri "sewa" arah CR di atas yang untuk perusahaan MENERIMA sewa) ---
    {
        "kata": ["sewa", "rental", "kontrakan", "kos "],
        "arah": "DB",
        "penanda": ["rental expense", "beban sewa", "sewa dibayar", "biaya sewa"],
        "filter_class": ["Expense"], "filter_statement": "PNL",
        "confidence": "Medium",
        "alasan": "Mengandung kata kunci sewa/rental pada narasi transaksi keluar",
    },
    # --- BPJS (Kesehatan/Ketenagakerjaan) ---
    {
        "kata": ["bpjs", "jamsostek"],
        "arah": "DB",
        "penanda": ["bpjs", "jaminan sosial", "employee benefit"],
        "filter_class": ["Expense"], "filter_statement": "PNL",
        "confidence": "Medium",
        "alasan": "Mengandung kata kunci BPJS/Jamsostek pada narasi transaksi",
    },
    # --- Setoran pajak (PPh 21/23/25/Final, PPN) via bank/billing DJP ---
    {
        "kata": ["pph 21", "pph21", "pph 23", "pph23", "pph 25", "pph25",
                 "pph pasal 4", "billing pajak", "setoran pajak", "ntpn"],
        "arah": "DB",
        "penanda": ["tax expense", "beban pajak", "hutang pajak", "pajak dibayar"],
        "filter_class": ["Expense", "Liability"], "filter_statement": None,
        "confidence": "Medium",
        "alasan": "Mengandung kata kunci setoran pajak (PPh/billing DJP) pada narasi transaksi",
    },
    # --- Ekspedisi / kurir / ongkos kirim ---
    {
        "kata": ["jne", "j&t", "jnt express", "sicepat", "anteraja", "ninja xpress",
                 "pos indonesia", "ekspedisi", "ongkos kirim", "ongkir"],
        "arah": "DB",
        "penanda": ["freight", "pengiriman", "beban angkut", "ongkos kirim", "ekspedisi"],
        "filter_class": ["Expense"], "filter_statement": "PNL",
        "confidence": "Medium",
        "alasan": "Mengandung kata kunci jasa ekspedisi/kurir pada narasi transaksi",
    },
    # --- Perlengkapan & administrasi kantor ---
    {
        "kata": ["atk", "alat tulis kantor", "office supplies", "office stationery"],
        "arah": "DB",
        "penanda": ["office supplies", "perlengkapan kantor", "atk"],
        "filter_class": ["Expense"], "filter_statement": "PNL",
        "confidence": "Medium",
        "alasan": "Mengandung kata kunci perlengkapan/ATK kantor pada narasi transaksi",
    },
    # --- Konsumsi / catering karyawan ---
    {
        "kata": ["konsumsi", "catering", "makan minum karyawan"],
        "arah": "DB",
        "penanda": ["konsumsi", "catering", "meals", "entertainment"],
        "filter_class": ["Expense"], "filter_statement": "PNL",
        "confidence": "Medium",
        "alasan": "Mengandung kata kunci konsumsi/catering pada narasi transaksi",
    },
    # --- Iklan / marketing / promosi ---
    {
        "kata": ["iklan", "advertising", "marketing fee", "promosi", "sponsorship"],
        "arah": "DB",
        "penanda": ["advertising", "marketing", "promosi", "beban iklan"],
        "filter_class": ["Expense"], "filter_statement": "PNL",
        "confidence": "Medium",
        "alasan": "Mengandung kata kunci iklan/marketing/promosi pada narasi transaksi",
    },
    # --- Jasa profesional (notaris/legal/konsultan/akuntan) ---
    {
        "kata": ["notaris", "legal fee", "jasa profesional", "konsultan", "akuntan publik",
                 "audit fee", "biaya audit"],
        "arah": "DB",
        "penanda": ["professional fee", "jasa profesional", "legal", "konsultan", "audit"],
        "filter_class": ["Expense"], "filter_statement": "PNL",
        "confidence": "Medium",
        "alasan": "Mengandung kata kunci jasa profesional (notaris/legal/konsultan/audit) pada narasi transaksi",
    },
    # --- Pemeliharaan / perbaikan ---
    {
        "kata": ["maintenance", "perawatan", "perbaikan", "service ac", "service kendaraan"],
        "arah": "DB",
        "penanda": ["maintenance", "repair", "pemeliharaan", "perbaikan"],
        "filter_class": ["Expense"], "filter_statement": "PNL",
        "confidence": "Medium",
        "alasan": "Mengandung kata kunci pemeliharaan/perbaikan pada narasi transaksi",
    },
]


def _cocok_penanda_di_coa(
    penanda_list: List[str], df_coa: pd.DataFrame,
    filter_class: Optional[List[str]] = None, filter_statement: Optional[str] = None,
) -> Optional[Tuple[str, str]]:
    """Cari 1 akun COA (skema kertas kerja) yang Account Name ATAU Notes-nya
    mengandung salah satu string di `penanda_list` (case-insensitive),
    dipersempit ke Class/Statement tertentu kalau diisi (mengurangi
    false-positive, sama semangatnya dgn cari_akun_bank_dari_coa)."""
    kandidat = df_coa
    if filter_class and "Class" in df_coa.columns:
        kandidat = kandidat[kandidat["Class"].fillna("").isin(filter_class)]
    if filter_statement and "Statement" in df_coa.columns:
        kandidat = kandidat[kandidat["Statement"].fillna("").str.strip().str.upper() == filter_statement.upper()]
    if kandidat.empty:
        return None

    nama = kandidat["Account Name"].fillna("").astype(str)
    notes = kandidat["Notes"].fillna("").astype(str) if "Notes" in kandidat.columns else pd.Series([""] * len(kandidat), index=kandidat.index)
    for penanda in penanda_list:
        p = penanda.lower()
        mask = nama.str.lower().str.contains(p, regex=False) | notes.str.lower().str.contains(p, regex=False)
        cocok = kandidat[mask]
        if not cocok.empty:
            baris = cocok.iloc[0]
            return str(baris["Account No."]), str(baris["Account Name"])
    return None


def cocokkan_kata_kunci_kertas_kerja(
    keterangan: str, arah: str, df_coa: pd.DataFrame,
) -> Optional[Dict[str, str]]:
    """Coba cocokkan `keterangan` (narasi transaksi bank) ke salah satu
    aturan di KATA_KUNCI_KERTAS_KERJA, dipersempit ke arah transaksi
    ("CR"/"DB"). Return None kalau tidak ada yang cocok (baris tetap
    lanjut ke staging seperti sebelumnya -- fungsi ini TIDAK PERNAH
    memaksa suatu klasifikasi, hanya menambah)."""
    if not keterangan:
        return None
    t = str(keterangan).lower().strip()
    for aturan in KATA_KUNCI_KERTAS_KERJA:
        if aturan["arah"] != arah:
            continue
        # [BARU] "exact": True -- dipakai utk baris yang keterangannya SELALU
        # persis 1 kata baku dari sistem bank sendiri (mis. "BUNGA", "PAJAK
        # BUNGA"), BUKAN narasi bebas. Substring biasa ("bunga" in teks)
        # berisiko salah tangkap nama merchant yang kebetulan memuat kata itu
        # (mis. "TOKO BUNGA ..."), jadi utk kata kunci sesensitif ini WAJIB
        # exact-match seluruh keterangan, bukan sekadar mengandung.
        if aturan.get("exact"):
            if t not in aturan["kata"]:
                continue
        elif not any(k in t for k in aturan["kata"]):
            continue
        hasil = _cocok_penanda_di_coa(
            aturan["penanda"], df_coa,
            filter_class=aturan.get("filter_class"), filter_statement=aturan.get("filter_statement"),
        )
        if hasil:
            no_akun, nama_akun = hasil
            return {
                "no_akun": no_akun, "nama_akun": nama_akun,
                "confidence": aturan["confidence"], "alasan": aturan["alasan"],
            }
    return None


def _pra_klasifikasi_kata_kunci_kertas_kerja(
    df_bank_mentah: pd.DataFrame, df_coa: pd.DataFrame, daftar_akun_bank: List[Dict[str, str]],
) -> Tuple[pd.DataFrame, "pd.Series[bool]", int]:
    """
    [BARU -- PERBAIKAN PERFORMA] Logikanya SAMA dengan
    _terapkan_kata_kunci_kertas_kerja() di bawah, TAPI dijalankan SEBELUM
    akuntansi_ai.proses_dataframe()/AI, terhadap df_bank_mentah MENTAH
    (baris hasil ekstraksi PDF, belum ada kolom jurnal sama sekali).

    KENAPA INI PERLU (bukan sekadar duplikasi _terapkan_kata_kunci_kertas_kerja):
    Sebelumnya lapisan kata-kunci kertas kerja HANYA dipasang SETELAH
    proses_dataframe() selesai (lihat generate_kertas_kerja) -- jadi walaupun
    sebuah transaksi 100% pasti bisa dikenali lewat kata kunci di sini (mis.
    baris "BUNGA"/"PAJAK BUNGA" akhir bulan yang ADA DI SETIAP rekening koran,
    setiap bank, setiap bulan), baris itu SUDAH KEBURU dikirim ke AI duluan
    lewat proses_dataframe() sebelum lapisan kata-kunci sempat jalan --
    kata kunci itu jadi TIDAK PERNAH benar-benar mengurangi beban AI, cuma
    jadi jaring pengaman untuk hasil AI yang gagal/null.

    Dengan menjalankan pengecekan yang SAMA di sini, SEBELUM proses_dataframe()
    dipanggil, baris yang sudah pasti kena kata kunci dikeluarkan dari batch
    yang dikirim ke AI sama sekali -- mengecilkan jumlah transaksi yang perlu
    AI (lebih sedikit chunk, lebih cepat prosesnya).

    generate_kertas_kerja() TETAP memanggil _terapkan_kata_kunci_kertas_kerja()
    juga (setelah AI, seperti sebelumnya) sebagai jaring pengaman KEDUA utk
    baris yang tidak kena kata kunci di sini TAPI AI juga gagal/menandai null
    -- kedua lapisan saling melengkapi, tidak saling menggantikan.

    Returns: (df_bank_mentah + kolom jurnal terisi utk baris yang cocok,
    mask boolean baris mana yang cocok, jumlah baris yang cocok).
    """
    df = df_bank_mentah.copy()
    for _kol in ("no_akun_debet", "nama_akun_debet", "no_akun_kredit", "nama_akun_kredit",
                 "jml_debet", "jml_kredit", "sumber_kategori", "confidence_ai", "alasan_ai", "catatan_ai"):
        df[_kol] = pd.Series([None] * len(df), dtype="object", index=df.index)

    peta_bank_nama = {a["Account No."]: a["Account Name"] for a in daftar_akun_bank}
    mask_cocok = pd.Series(False, index=df.index)

    for idx, row in df.iterrows():
        arah = "CR" if (row.get("mutasi_debet") or 0) == 0 else "DB"
        cocok = cocokkan_kata_kunci_kertas_kerja(row.get("keterangan"), arah, df_coa)
        if not cocok:
            continue

        nominal = row.get("mutasi_debet") or row.get("mutasi_kredit") or 0
        akun_bank_baris = _pilih_akun_bank_untuk_baris(row.get("bank"), daftar_akun_bank)
        nama_bank_baris = peta_bank_nama.get(akun_bank_baris)

        if arah == "CR":
            df.at[idx, "no_akun_debet"] = akun_bank_baris
            df.at[idx, "nama_akun_debet"] = nama_bank_baris
            df.at[idx, "no_akun_kredit"] = cocok["no_akun"]
            df.at[idx, "nama_akun_kredit"] = cocok["nama_akun"]
        else:
            df.at[idx, "no_akun_debet"] = cocok["no_akun"]
            df.at[idx, "nama_akun_debet"] = cocok["nama_akun"]
            df.at[idx, "no_akun_kredit"] = akun_bank_baris
            df.at[idx, "nama_akun_kredit"] = nama_bank_baris

        df.at[idx, "jml_debet"] = nominal
        df.at[idx, "jml_kredit"] = nominal
        df.at[idx, "sumber_kategori"] = f"Kata Kunci Kertas Kerja - Pra-AI ({cocok['confidence']})"
        df.at[idx, "catatan_ai"] = cocok["alasan"]
        mask_cocok.at[idx] = True

    return df, mask_cocok, int(mask_cocok.sum())


def _terapkan_kata_kunci_kertas_kerja(
    df_hasil: pd.DataFrame, df_coa: pd.DataFrame, daftar_akun_bank: List[Dict[str, str]],
) -> Tuple[pd.DataFrame, int]:
    """Terapkan cocokkan_kata_kunci_kertas_kerja() ke SEMUA baris df_hasil
    yang MASIH belum terklasifikasi (no_akun_debet/kredit kosong) setelah
    akuntansi_ai.proses_dataframe() -- dipanggil di generate_kertas_kerja()
    persis sebelum susun_sheet_gl(). Baris yang SUDAH terklasifikasi (pola
    historis, kata-kunci lama, atau AI) TIDAK disentuh sama sekali.

    Returns: (df_hasil dgn kolom jurnal terisi utk baris yang cocok,
    jumlah baris yang berhasil ditingkatkan)."""
    df_hasil = df_hasil.copy()
    jumlah = 0
    peta_bank_nama = {a["Account No."]: a["Account Name"] for a in daftar_akun_bank}

    for idx, row in df_hasil.iterrows():
        sudah_ada = pd.notna(row.get("no_akun_debet")) and pd.notna(row.get("no_akun_kredit"))
        if sudah_ada:
            continue

        arah = "CR" if (row.get("mutasi_debet") or 0) == 0 else "DB"
        cocok = cocokkan_kata_kunci_kertas_kerja(row.get("keterangan"), arah, df_coa)
        if not cocok:
            continue

        nominal = row.get("mutasi_debet") or row.get("mutasi_kredit") or 0
        akun_bank_baris = _pilih_akun_bank_untuk_baris(row.get("bank"), daftar_akun_bank)
        nama_bank_baris = peta_bank_nama.get(akun_bank_baris)

        if arah == "CR":
            df_hasil.at[idx, "no_akun_debet"] = akun_bank_baris
            df_hasil.at[idx, "nama_akun_debet"] = nama_bank_baris
            df_hasil.at[idx, "no_akun_kredit"] = cocok["no_akun"]
            df_hasil.at[idx, "nama_akun_kredit"] = cocok["nama_akun"]
        else:
            df_hasil.at[idx, "no_akun_debet"] = cocok["no_akun"]
            df_hasil.at[idx, "nama_akun_debet"] = cocok["nama_akun"]
            df_hasil.at[idx, "no_akun_kredit"] = akun_bank_baris
            df_hasil.at[idx, "nama_akun_kredit"] = nama_bank_baris

        df_hasil.at[idx, "jml_debet"] = nominal
        df_hasil.at[idx, "jml_kredit"] = nominal
        df_hasil.at[idx, "sumber_kategori"] = f"Kata Kunci Kertas Kerja ({cocok['confidence']})"
        if pd.isna(df_hasil.at[idx, "catatan_ai"]):
            df_hasil.at[idx, "catatan_ai"] = cocok["alasan"]
        jumlah += 1

    return df_hasil, jumlah


# ============================================================
# 9. ORKESTRATOR END-TO-END
# ============================================================

def generate_kertas_kerja(
    daftar_file_pdf: List[Tuple[Any, str]],
    df_coa: pd.DataFrame,
    client_id: Optional[int] = None,
    pakai_ai: bool = True,
    pakai_claude_review: bool = False,
    pakai_claude_review_final: bool = False,
    peringatan_awal: Optional[List[str]] = None,
    progress_callback: Optional[Any] = None,
) -> Tuple[HasilKertasKerja, List[str]]:
    """
    Titik masuk utama. Args:
        daftar_file_pdf: list (file_like, nama_file_pdf) rekening koran
            client (boleh multi bank/bulan, akan digabung).
        df_coa: pd.DataFrame (WAJIB sudah punya kolom Normal
            Balance/Statement/FS Group/Notes, 2 akun ditandai "staging"
            di Notes -- lihat cari_akun_staging_dari_coa() -- DAN minimal
            1 akun ditandai "bank"/"rekening" di Notes -- lihat
            [FIX] cari_akun_bank_dari_coa()).
        client_id: dipakai untuk memuat/menyimpan pola historis client
            (pola_bank) via akuntansi_ai, sama seperti proses_file_rekening_koran.
        pakai_claude_review: [BARU] default False (OPSIONAL, hemat kredit).
            Kalau True, HANYA baris GL dengan Confidence bukan "High" yang
            dikirim (dibatch) ke Claude API lewat
            perbaiki_gl_dengan_claude_review() untuk disarankan akun yang
            lebih tepat -- lihat fungsi itu utk detail (termasuk validasi
            anti-halusinasi akun & batching). Kalau False, tahap ini
            dilewati sepenuhnya, tidak ada panggilan API sama sekali.
        pakai_claude_review_final: [BARU] default False (OPSIONAL, hemat
            kredit). Validasi struktur & konsistensi angka (lihat section
            "3.9 VALIDASI STRUKTUR & KONSISTENSI" di atas,
            jalankan_validasi_otomatis_kertas_kerja) SELALU jalan otomatis
            terlepas dari flag ini (murni kode, tidak ada panggilan API,
            hasilnya di hasil.temuan_validasi). Flag ini HANYA mengontrol
            apakah temuan level "error" (kalau ada) dikirim ke Claude API
            untuk dijelaskan/diprioritaskan dalam bahasa akuntan
            (hasil.catatan_review_final) -- lihat
            jelaskan_temuan_kertas_kerja_claude() di claude_client.py.
            Kalau False atau tidak ada temuan "error", TIDAK ada
            panggilan API tambahan sama sekali.
        peringatan_awal: [BARU] peringatan yang sudah ada SEBELUM fungsi
            ini dipanggil (mis. dari bangun_coa_kertas_kerja_dari_db()
            kalau COA dibangun dari fallback DB, bukan file Excel) --
            digabung di depan supaya tidak hilang.
        progress_callback: [BARU] diteruskan apa adanya ke
            susun_gl_dari_pdf_rekening_koran() -- lihat docstring di sana.
            Opsional, dipakai endpoint SSE per-file di main.py.

    Returns: (HasilKertasKerja, peringatan_gabungan)
    """
    peringatan: List[str] = list(peringatan_awal or [])

    # [BARU] Kalau akun staging/bank belum lengkap ditandai di COA client
    # (atau COA-nya kosong sama sekali), JANGAN gagalkan seluruh proses --
    # tambahkan akun placeholder otomatis ke df_coa supaya kertas kerja
    # tetap bisa digenerate dari PDF. Akun placeholder ini WAJIB direklas
    # manual oleh akuntan sebelum laporan final dipakai -- ditandai jelas
    # di peringatan.
    try:
        akun_staging = cari_akun_staging_dari_coa(df_coa)
    except ValueError as e:
        peringatan.append(
            f"[FALLBACK] {e} -- sistem tetap lanjut memakai akun placeholder "
            "'9999 - Akun Belum Teridentifikasi (Staging)'. WAJIB direklas "
            "manual ke akun yang benar sebelum laporan final dipakai."
        )
        if "9999" not in df_coa.get("Account No.", pd.Series(dtype=str)).astype(str).values:
            df_coa = pd.concat([df_coa, pd.DataFrame([{
                "Account No.": "9999",
                "Account Name": "Akun Belum Teridentifikasi (Staging)",
                "Class": "Asset", "Normal Balance": "Debit", "Statement": "BS",
                "FS Group": "Aset Lancar Lainnya", "Tax Return Group": None,
                "Notes": "staging",
            }])], ignore_index=True)
        akun_staging = {"CR": "9999", "DB": "9999"}

    # [FIX] Sebelumnya kode akun bank di-hardcode "1102" di susun_sheet_gl
    # -- sekarang dicari dari COA client, konsisten dengan cara akun
    # staging dicari (lihat cari_akun_bank_dari_coa untuk detail & alasan).
    try:
        daftar_akun_bank, warn_bank = cari_akun_bank_dari_coa(df_coa)
        peringatan.extend(warn_bank)
    except ValueError as e:
        peringatan.append(
            f"[FALLBACK] {e} -- sistem tetap lanjut memakai akun placeholder "
            "'9998 - Bank (Belum Diverifikasi)'. WAJIB direklas manual ke "
            "akun bank yang benar sebelum laporan final dipakai."
        )
        if "9998" not in df_coa.get("Account No.", pd.Series(dtype=str)).astype(str).values:
            df_coa = pd.concat([df_coa, pd.DataFrame([{
                "Account No.": "9998",
                "Account Name": "Bank (Belum Diverifikasi)",
                "Class": "Asset", "Normal Balance": "Debit", "Statement": "BS",
                "FS Group": "Kas dan Setara Kas", "Tax Return Group": None,
                "Notes": "bank",
            }])], ignore_index=True)
        daftar_akun_bank = [{"Account No.": "9998", "Account Name": "Bank (Belum Diverifikasi)"}]

    df_bank_mentah, warn_pdf, ringkasan_statement_per_file = susun_gl_dari_pdf_rekening_koran(
        daftar_file_pdf, client_id=client_id, progress_callback=progress_callback,
    )
    peringatan.extend(warn_pdf)
    if df_bank_mentah.empty:
        raise ValueError(
            "Tidak ada transaksi rekening koran yang berhasil diekstrak dari PDF yang diupload. "
            f"Detail: {peringatan}"
        )

    # [BARU -- PERBAIKAN PERFORMA] Jalankan kata-kunci kertas kerja SEBELUM
    # mengirim apa pun ke proses_dataframe()/AI -- lihat docstring
    # _pra_klasifikasi_kata_kunci_kertas_kerja utk alasan kenapa ini penting
    # (posisi LAMA lapisan kata-kunci ini, SETELAH AI, tidak pernah benar-benar
    # mengurangi beban AI). Hanya baris yang TIDAK kena di sini yang
    # diteruskan ke AI -- makin banyak yang kena di sini (bunga bank/pajak
    # bunga/biaya bank/dll -- narasi baku sistem bank), makin sedikit &
    # makin cepat panggilan AI-nya.
    df_bank_mentah, mask_pra_kk, jumlah_pra_kk = _pra_klasifikasi_kata_kunci_kertas_kerja(
        df_bank_mentah, df_coa, daftar_akun_bank,
    )
    _KOLOM_JURNAL_SEMENTARA = [
        "no_akun_debet", "nama_akun_debet", "no_akun_kredit", "nama_akun_kredit",
        "jml_debet", "jml_kredit", "sumber_kategori", "confidence_ai", "alasan_ai", "catatan_ai",
    ]
    df_pra_cocok = df_bank_mentah[mask_pra_kk]
    df_belum_ai = df_bank_mentah[~mask_pra_kk].drop(columns=_KOLOM_JURNAL_SEMENTARA, errors="ignore")
    if jumlah_pra_kk:
        peringatan.append(
            f"{jumlah_pra_kk} transaksi diklasifikasi otomatis SEBELUM tahap AI lewat kata kunci "
            "kertas kerja (bunga bank/pajak bunga/biaya bank/dll -- narasi baku sistem bank), "
            "mengurangi jumlah transaksi yang perlu dikirim ke AI. Tetap disarankan spot-check."
        )

    # [FIX] akuntansi_ai.proses_dataframe() (lewat cocokkan_kata_kunci_ke_coa
    # dkk) mengharapkan df_coa skema DATABASE (kolom "no_akun"/"nama_akun"),
    # BUKAN skema kertas kerja ("Account No."/"Account Name") yang dipakai
    # df_coa di seluruh modul ini -- sebelumnya diteruskan apa adanya dan
    # meledak KeyError('nama_akun') begitu ada transaksi yang perlu
    # dicocokkan lewat kata kunci (baik pakai_ai=True MAUPUN False, karena
    # tahap kata-kunci selalu dicoba sebelum AI). Dibuat versi ringan
    # (hanya kolom yang benar-benar dipakai proses_dataframe) di sini,
    # tanpa mengubah df_coa asli (yang masih dipakai skema kertas kerja di
    # seluruh fungsi lain modul ini).
    df_coa_untuk_ai = pd.DataFrame({
        "no_akun": df_coa["Account No."].astype(str),
        "nama_akun": df_coa["Account Name"].astype(str),
    })

    path_pola = akuntansi_ai._path_pola("pola_bank", client_id)
    pola = akuntansi_ai.muat_pola(path_pola)
    df_hasil_ai = akuntansi_ai.proses_dataframe(
        df_belum_ai, df_coa_untuk_ai, pola, pakai_ai=pakai_ai,
        api_key=akuntansi_ai.ambil_api_key() if pakai_ai else None,
    )
    akuntansi_ai.simpan_pola(pola, path_pola)

    # Gabungkan kembali baris yang sudah kena kata-kunci pra-AI + hasil
    # proses_dataframe() -- urutan dikembalikan sesuai index ASLI
    # df_bank_mentah (proses_dataframe TIDAK mengubah index, cuma mengisi
    # kolom, jadi sort_index() aman mengembalikan urutan baris original;
    # ini penting krn penempelan source_pdf/halaman di bawah mengasumsikan
    # urutan baris df_hasil == urutan baris df_bank_mentah).
    df_hasil = pd.concat([df_pra_cocok, df_hasil_ai]).sort_index()

    # proses_dataframe() TIDAK BOLEH mengurangi/menambah baris relatif ke
    # df_bank_mentah -- ini asumsi penting supaya source_pdf/halaman tetap
    # sejajar. Kalau ternyata proses_dataframe mengubah jumlah baris,
    # metadata perlu ditempel ulang lewat kolom kunci (mis. index asli)
    # bukan positional -- TOLONG DIKONFIRMASI ke tim saat testing.
    if len(df_hasil) == len(df_bank_mentah):
        df_hasil["source_pdf"] = df_bank_mentah["source_pdf"].values
        df_hasil["halaman"] = df_bank_mentah["halaman"].values
    else:
        peringatan.append(
            "PERINGATAN INTEGRASI: jumlah baris berubah setelah proses_dataframe() "
            f"({len(df_bank_mentah)} -> {len(df_hasil)}) -- kolom Source PDF/Page di "
            "GL TIDAK bisa dijamin akurat per baris untuk batch ini. Perlu perbaikan "
            "cara penempelan metadata (lihat docstring generate_kertas_kerja)."
        )
        df_hasil["source_pdf"] = None
        df_hasil["halaman"] = None

    # [BARU -- FIX GAP] Lapisan kata-kunci khusus kertas kerja (lihat blok
    # komentar panjang di atas _terapkan_kata_kunci_kertas_kerja) --
    # HANYA menyentuh baris yang masih belum terklasifikasi sama sekali
    # setelah pola historis/kata-kunci lama/AI di atas. Dipasang di sini
    # (bukan sebelum proses_dataframe) supaya tidak pernah menimpa hasil
    # AI/pola yang sudah ada, dan supaya masih bisa jalan walau pakai_ai=False.
    df_hasil, jumlah_kw_kk = _terapkan_kata_kunci_kertas_kerja(df_hasil, df_coa, daftar_akun_bank)
    if jumlah_kw_kk:
        peringatan.append(
            f"{jumlah_kw_kk} transaksi diklasifikasi otomatis lewat kata kunci "
            "kertas kerja (biaya bank/sewa/BBM/gaji/listrik/dll, tanpa AI) -- "
            "tetap disarankan spot-check, terutama yang confidence Medium."
        )

    df_gl = susun_sheet_gl(df_hasil, akun_staging, daftar_akun_bank)

    # [BARU -- OPSIONAL, default mati] Review Claude API HANYA untuk baris
    # confidence non-High, dibatch (lihat perbaiki_gl_dengan_claude_review).
    # Ditaruh di sini (SEBELUM Bank_Control/Posting/TB/BS/PNL disusun)
    # supaya seluruh sheet turunan memakai df_gl yang SUDAH direview, bukan
    # hasil klasifikasi mentah -- kalau dipasang setelah sheet turunan
    # dihitung, koreksi Claude tidak akan pernah terlihat di laporan.
    df_status_ai = pd.DataFrame()
    if pakai_claude_review:
        df_gl, df_status_ai, peringatan_review = perbaiki_gl_dengan_claude_review(
            df_gl, df_coa, client_id=client_id,
        )
        peringatan.extend(peringatan_review)

    # [BARU] Petakan ringkasan_statement_per_file (per nama file PDF) ke
    # saldo_statement_per_bulan (per Month No.) yang dibutuhkan
    # susun_bank_control() -- ditentukan dari bulan MAYORITAS baris GL
    # yang source_pdf-nya file tsb (harusnya 1 file = 1 bulan, tapi kalau
    # ada statement yang membelah tahun/bulan, majority vote lebih aman
    # daripada asumsi kaku "1 file = 1 bulan").
    saldo_statement_per_bulan: Dict[int, Dict[str, float]] = {}
    if ringkasan_statement_per_file and "Source PDF" in df_gl.columns and "Month No." in df_gl.columns:
        for nama_file_pdf, ringkasan_footer in ringkasan_statement_per_file.items():
            subset = df_gl[df_gl["Source PDF"] == nama_file_pdf]
            if subset.empty:
                continue
            bulan_no = int(subset["Month No."].mode().iloc[0])
            saldo_statement_per_bulan[bulan_no] = {
                "cr": ringkasan_footer.get("mutasi_cr"),
                "db": ringkasan_footer.get("mutasi_db"),
                "closing": ringkasan_footer.get("saldo_akhir"),
            }

    df_bank_control = susun_bank_control(df_gl, saldo_statement_per_bulan=saldo_statement_per_bulan)
    df_posting = susun_bank_posting_summary(df_gl, df_coa)
    df_adjustments = susun_template_adjustments(df_coa)
    df_tb = susun_tb_monthly(df_posting, df_adjustments)
    # [FIX -- gap #3] Saat pertama kali digenerate, Opening_Balance memang
    # SELALU 0 di semua akun (input manual client, lihat susun_opening_balance)
    # -- jadi hasilnya identik dgn perilaku lama di titik ini. Tetap dialirkan
    # (bukan None) supaya konsisten secara arsitektur dgn jalur re-upload di
    # susun_data_export_18_sheet_dari_kertas_kerja(), yang MEMANG membawa
    # Opening_Balance yang sudah dikoreksi user.
    df_opening_awal = susun_opening_balance(df_coa)
    df_bs, df_pnl = susun_bs_pnl_monthly(df_tb, df_coa, df_opening_awal)

    hasil = HasilKertasKerja(
        gl=df_gl, bank_control=df_bank_control, bank_posting_summary=df_posting,
        adjustments=df_adjustments, tb_monthly=df_tb, bs_monthly=df_bs,
        pnl_monthly=df_pnl, coa=df_coa, peringatan=peringatan, status_ai=df_status_ai,
    )

    # [BARU -- OTOMATIS, TIDAK ADA FLAG] Validasi struktur & konsistensi
    # angka -- lihat section "3.9 VALIDASI STRUKTUR & KONSISTENSI" di atas
    # file ini. Murni kode (tidak ada panggilan API), jadi SELALU dijalankan
    # di sini, sama seperti tahapan kata-kunci/klasifikasi lain di atas.
    # `tahun` ditentukan di sini (bukan diteruskan sbg parameter) karena
    # sebelumnya generate_kertas_kerja() memang tidak menerima `tahun` --
    # dipakai HANYA utk validasi, TIDAK menggantikan tentukan_tahun_dari_gl()
    # yang tetap dipanggil terpisah oleh caller (main.py) sebelum
    # tulis_kertas_kerja_excel().
    tahun_untuk_validasi, peringatan_tahun = tentukan_tahun_dari_gl(hasil.gl)
    temuan = jalankan_validasi_otomatis_kertas_kerja(hasil, tahun_untuk_validasi)
    hasil.temuan_validasi = temuan

    jumlah_error = sum(1 for t in temuan if t["level"] == "error")
    jumlah_warning = sum(1 for t in temuan if t["level"] == "warning")
    if temuan:
        peringatan.append(
            f"Validasi otomatis kertas kerja: {jumlah_error} temuan error, "
            f"{jumlah_warning} temuan warning (lihat hasil.temuan_validasi untuk detail)."
        )
        for t in temuan:
            peringatan.append(f"[Validasi/{t['level'].upper()}/{t['area']}] {t['pesan']}")

    # [BARU -- OPSIONAL, default mati] Kirim temuan level "error" ke Claude
    # untuk dijelaskan/diprioritaskan dalam bahasa akuntan -- HANYA temuan
    # terstruktur yang dikirim (bukan data transaksi client), konsisten
    # dengan prinsip audit trail claude_client.py (tidak menyimpan/mengirim
    # data sensitif client tanpa perlu). Kalau tidak ada temuan "error" ATAU
    # flag ini False, TIDAK ada panggilan API di sini sama sekali.
    if pakai_claude_review_final and jumlah_error:
        try:
            from .claude_client import jelaskan_temuan_kertas_kerja_claude
        except ImportError:  # pragma: no cover
            try:
                from claude_client import jelaskan_temuan_kertas_kerja_claude
            except ImportError:
                jelaskan_temuan_kertas_kerja_claude = None
        if jelaskan_temuan_kertas_kerja_claude is not None:
            try:
                hasil.catatan_review_final = jelaskan_temuan_kertas_kerja_claude(
                    [t for t in temuan if t["level"] == "error"],
                    client_id=str(client_id) if client_id is not None else None,
                )
            except Exception as e:  # noqa: BLE001 -- review final tidak boleh menggagalkan generate
                peringatan.append(
                    f"Review final Claude API gagal ({type(e).__name__}: {e}) -- "
                    "temuan validasi tetap tersedia di hasil.temuan_validasi apa adanya."
                )
        else:
            peringatan.append(
                "Review final Claude API dilewati: modules/claude_client.py "
                "tidak berhasil di-import."
            )

    return hasil, peringatan


# ============================================================
# 10. MEMBACA ULANG KERTAS KERJA YANG SUDAH DIKOREKSI USER
# ============================================================

def baca_adjustments_dari_kertas_kerja(file_like) -> pd.DataFrame:
    """
    Baca sheet 'Adjustments' dari file kertas kerja yang sudah dikoreksi
    & diupload ulang user. Dipakai sebelum generate report 18-sheet,
    supaya reklasifikasi manual akuntan ikut terpakai.
    """
    wb = openpyxl.load_workbook(file_like, data_only=True)
    if "Adjustments" not in wb.sheetnames:
        raise ValueError("Sheet 'Adjustments' tidak ditemukan di file yang diupload.")
    ws = wb["Adjustments"]

    # Header ada di baris ke-4 pada contoh (baris 1=judul, 2=instruksi,
    # 3=kosong, 4=header kolom bulan, 5=sub-header Debit/Credit).
    header_bulan = [c.value for c in ws[4]]
    header_debit_kredit = [c.value for c in ws[5]]

    kolom_final = []
    bulan_aktif = None
    for atas, bawah in zip(header_bulan, header_debit_kredit):
        if atas in BULAN_URUT:
            bulan_aktif = atas
        if bawah in ("Debit", "Credit") and bulan_aktif:
            kolom_final.append(f"{bulan_aktif}_{bawah}")
        elif atas in ("Account No.", "Account Name", "Review Notes"):
            kolom_final.append(atas)
        else:
            kolom_final.append(atas)

    rows = list(ws.iter_rows(min_row=6, values_only=True))
    df = pd.DataFrame(rows, columns=kolom_final)
    return df


# ============================================================
# 10.5 [BARU] HELPER: TAHUN DARI GL + RINGKASAN STATUS
# ============================================================

def tentukan_tahun_dari_gl(df_gl: pd.DataFrame) -> Tuple[int, List[str]]:
    """
    Tebak "tahun" kertas kerja dari kolom Date di sheet GL yang sudah
    tersusun (dipakai sebagai parameter `tahun` di tulis_kertas_kerja_excel).
    Kalau transaksi ternyata mencakup lebih dari 1 tahun (mis. rekening
    koran yang diupload mencampur Des tahun lalu dengan Jan-Nov tahun
    ini), tahun yang dipakai adalah yang PALING BANYAK transaksinya, dan
    dicatat sebagai peringatan supaya akuntan sadar ada transaksi lain
    tahun yang mungkin perlu kertas kerja terpisah.
    """
    peringatan: List[str] = []
    if df_gl is None or df_gl.empty or "Date" not in df_gl.columns:
        tahun_sekarang = datetime.now().year
        peringatan.append(
            f"Tidak ada tanggal transaksi terbaca di GL -- tahun kertas kerja "
            f"dipakai default tahun berjalan ({tahun_sekarang}), TOLONG DICEK."
        )
        return tahun_sekarang, peringatan

    tahun_series = pd.to_datetime(df_gl["Date"], errors="coerce").dt.year.dropna()
    if tahun_series.empty:
        tahun_sekarang = datetime.now().year
        peringatan.append(
            f"Kolom Date di GL tidak bisa dibaca sebagai tanggal -- tahun kertas "
            f"kerja dipakai default tahun berjalan ({tahun_sekarang}), TOLONG DICEK."
        )
        return tahun_sekarang, peringatan

    hitung_per_tahun = tahun_series.value_counts()
    tahun_terpilih = int(hitung_per_tahun.idxmax())
    if len(hitung_per_tahun) > 1:
        lain = {int(t): int(c) for t, c in hitung_per_tahun.items() if int(t) != tahun_terpilih}
        peringatan.append(
            f"Transaksi rekening koran mencakup lebih dari 1 tahun -- kertas kerja "
            f"ini dibuat untuk tahun {tahun_terpilih} (transaksi terbanyak). "
            f"Transaksi tahun lain yang ikut terekstrak: {lain} -- pertimbangkan "
            f"kertas kerja terpisah per tahun kalau ini bukan tahun buku yang dimaksud."
        )
    return tahun_terpilih, peringatan


def ringkasan_status_kertas_kerja(hasil: HasilKertasKerja) -> Dict[str, Any]:
    """
    [BARU] Ringkasan singkat untuk ditampilkan ke user SETELAH kertas
    kerja dibuat, sebelum user diminta konfirmasi lanjut ke laporan
    18-sheet -- jumlah transaksi per tingkat confidence (High/Medium/
    Low), status per bulan dari Bank_Control, dan daftar peringatan.
    Versi awal ini murni pengumpulan angka, BELUM ada status keseluruhan
    "BALANCED"/"CHECK" ala sheet FS_Control di contoh (FS_Control sudah
    berupa kerangka tabel Control/Value/Status sejak step 2 -- lihat
    susun_fs_control_kerangka -- tapi Value/Status-nya masih None, belum
    dihitung otomatis), itu langkah lanjutan.
    """
    df_gl = hasil.gl
    jumlah_transaksi = int(len(df_gl)) if df_gl is not None else 0

    confidence_count: Dict[str, int] = {"High": 0, "Medium": 0, "Low": 0}
    if df_gl is not None and not df_gl.empty and "Confidence" in df_gl.columns:
        vc = df_gl["Confidence"].value_counts()
        for level in confidence_count:
            confidence_count[level] = int(vc.get(level, 0))

    status_per_bulan = {}
    if hasil.bank_control is not None and not hasil.bank_control.empty:
        for _, r in hasil.bank_control.iterrows():
            status_per_bulan[r["Month"]] = {
                "transaksi": int(r.get("Transactions") or 0),
                "status": r.get("Status"),
            }

    return {
        "jumlah_transaksi": jumlah_transaksi,
        "confidence_count": confidence_count,
        "status_per_bulan": status_per_bulan,
        "jumlah_peringatan": len(hasil.peringatan),
        "peringatan": hasil.peringatan,
    }


# ============================================================
# 11. PENULIS WORKBOOK (14 sheet)
# ============================================================

def tulis_kertas_kerja_excel(hasil: HasilKertasKerja, tahun: int, identitas: Optional[Dict[str, str]] = None) -> bytes:
    """Tulis semua sheet ke satu workbook .xlsx (bytes), format & urutan
    sheet mengikuti contoh Kertas_Kerja_Laporan_Keuangan_2025.xlsx.

    [lanjutan Step 2] FS_Control, BS_Tax, PNL_Tax, PPh17_31E sekarang diisi
    NILAI SUNGGUHAN (lihat hitung_fs_control/isi_bs_tax_dengan_saldo/
    isi_pnl_tax_dengan_saldo/hitung_pph17_31e di atas), bukan lagi
    kerangka None/0 -- struktur tabel & label kolom TIDAK berubah.
    Opening_Balance TETAP 0 (input manual client, sengaja tidak ditebak --
    lihat susun_opening_balance) & Identity TETAP input manual (belum ada
    sumber datanya) -- keduanya justru dipakai hitung_fs_control() sebagai
    syarat OVERALL STATUS "BALANCED".
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # [PERBAIKAN STRUKTUR] FS_Control sebelumnya lewat _tulis_dataframe
    # generik (cuma 4 kolom) -- sekarang pakai penulis khusus yg
    # mereplikasi 3 blok tabel (Control/Classification Summary/Monthly
    # P&L Snapshot) sesuai contoh Kertas_Kerja_Laporan_Keuangan (8 kolom).
    df_fs_control = hitung_fs_control(hasil, tahun)
    _tulis_sheet_fs_control(wb, df_fs_control, tahun, hasil)

    _tulis_sheet_identity(wb, identitas or {})
    _tulis_dataframe(wb, "COA", hasil.coa, judul=f"CHART OF ACCOUNTS")

    # [URUTAN SHEET] Bank_Control ditulis SEBELUM Opening_Balance --
    # disesuaikan dengan urutan 14 sheet yang diminta (sheet 4 =
    # Bank_Control, sheet 5 = Opening_Balance). Isi/logic kedua sheet
    # TIDAK berubah, cuma urutan pemanggilan _tulis_dataframe-nya.
    _tulis_dataframe(wb, "Bank_Control", hasil.bank_control, judul=f"BANK STATEMENT CONTROL - {tahun}")

    df_opening_balance = susun_opening_balance(hasil.coa)
    _tulis_dataframe(
        wb, "Opening_Balance", df_opening_balance,
        judul=f"OPENING BALANCE - 1 JANUARY {tahun} (INITIAL INPUT = ZERO)",
        kolom_manual=["Opening Debit (IDR)", "Opening Credit (IDR)", "Review Notes"],
    )

    _tulis_dataframe(
        wb, "GL", hasil.gl, judul="GENERAL LEDGER / BANK JOURNAL DETAIL - " + str(tahun), freeze="A5",
        subjudul=(
            "Detailed bank-derived journal audit trail. Suggested classifications are conservative. "
            "For statement changes, post the reclassification or accrual in the Adjustments sheet; "
            "the Bank Posting Summary remains the generated bank-source schedule."
        ),
    )

    # [BARU -- OPSIONAL] Sheet tambahan di luar 14 sheet asli, HANYA
    # muncul kalau pakai_claude_review=True DAN ada baris yang benar-benar
    # direview (lihat generate_kertas_kerja/perbaiki_gl_dengan_claude_review).
    # Kalau kosong (fitur mati atau semua GL sudah High), sheet ini tidak
    # ditulis sama sekali -- workbook tetap 14 sheet seperti biasa.
    if not hasil.status_ai.empty:
        _tulis_dataframe(
            wb, "Status", hasil.status_ai,
            judul="STATUS REVIEW CLAUDE API — AUDIT TRAIL PERBAIKAN",
            subjudul="Baris yang direview & diperbaiki Claude API. 'Claude Yakin? = Tidak' tetap perlu dicek manual.",
        )

    # [PERBAIKAN STRUKTUR] Bank_Posting_Summary/Adjustments/TB_Monthly
    # sebelumnya lewat _tulis_dataframe generik (nama kolom flat mis.
    # "Jan_Debit" di 1 baris header) -- sekarang pakai penulis khusus
    # dengan header 2 baris (nama bulan di-merge, "Debit"/"Credit" di
    # baris bawah) sesuai contoh. Jumlah & isi kolom TIDAK berubah, cuma
    # cara menampilkan header-nya yang disamakan.
    _tulis_sheet_posting_style(
        wb, "Bank_Posting_Summary", hasil.bank_posting_summary,
        judul=f"BANK POSTING SUMMARY BY ACCOUNT / MONTH - GENERATED FROM GL",
        kolom_leading=["Account No.", "Account Name", "Normal Balance"],
    )
    _tulis_sheet_posting_style(
        wb, "Adjustments", hasil.adjustments,
        judul="ADJUSTMENTS / RECLASSIFICATIONS BY ACCOUNT & MONTH",
        subjudul=(
            "Enter additional debit/credit adjustments in blue cells after reviewing supporting "
            "documents (e.g., reclassify Customer Deposit to Revenue, Prepaid Expenses to "
            "expense/assets, depreciation, payroll, taxes, accruals). Keep total debit = total "
            "credit for each month."
        ),
        kolom_leading=["Account No.", "Account Name", "Review Notes"],
        kolom_input_manual=True,
    )
    _tulis_sheet_tb_monthly(wb, hasil.tb_monthly, judul=f"TRIAL BALANCE - MONTH END {tahun}")

    _tulis_sheet_bs_pnl_monthly(wb, "BS_Monthly", hasil.bs_monthly, judul=f"BALANCE SHEET - MONTH END {tahun}")
    _tulis_sheet_bs_pnl_monthly(
        wb, "PNL_Monthly", hasil.pnl_monthly, judul=f"STATEMENT OF PROFIT OR LOSS - MONTHLY {tahun}",
        tambah_kolom_total=True, tahun=tahun,
    )

    df_bs_tax = isi_bs_tax_dengan_saldo(susun_bs_tax_kerangka(hasil.coa), hasil.bs_monthly)
    _tulis_sheet_bs_tax(wb, df_bs_tax, tahun)

    df_pnl_tax = isi_pnl_tax_dengan_saldo(susun_pnl_tax_kerangka(hasil.coa), hasil.pnl_monthly)
    _tulis_sheet_pnl_tax(wb, df_pnl_tax, tahun)

    hasil_pph = hitung_pph17_31e(df_pnl_tax, hasil.coa, tahun)
    _tulis_sheet_pph17_31e(wb, tahun, hasil_pph)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# 12. [BARU - lanjutan Step 2] PERHITUNGAN NILAI SUNGGUHAN untuk
# FS_Control, BS_Tax/PNL_Tax, PPh17_31E -- sebelumnya kerangka di atas
# (susun_fs_control_kerangka, susun_bs_tax_kerangka, susun_pnl_tax_kerangka,
# _tulis_sheet_pph17_31e) sengaja Value/Amount/PERHITUNGAN-nya None/0.
# Fungsi-fungsi di sini mengisi angka sungguhan dari GL/TB_Monthly/
# BS_Monthly/PNL_Monthly yang sudah tersusun di HasilKertasKerja, TANPA
# mengubah kerangka/struktur tabel yang sudah ada (kolom & label sama).
#
# CATATAN PENTING -- ini cek TEKNIS (keseimbangan angka), BUKAN
# pemeriksaan substansi akuntansi. "BALANCED" di FS_Control artinya
# angka-angkanya konsisten satu sama lain, BUKAN berarti kategorisasi
# transaksinya sudah benar -- itu tetap tanggung jawab akuntan saat
# review sheet GL/Adjustments.
# ============================================================

_TOLERANSI_BALANCE = 1.0  # toleransi pembulatan floating point (IDR)
_KATA_KUNCI_PENDAPATAN = ["pendapatan", "penjualan", "revenue", "sales"]


def _total_debit_credit_gl(df_gl: pd.DataFrame) -> Tuple[float, float]:
    if df_gl is None or df_gl.empty:
        return 0.0, 0.0
    total_debit = float(pd.to_numeric(df_gl.get("Debit (IDR)"), errors="coerce").fillna(0).sum())
    total_credit = float(pd.to_numeric(df_gl.get("Credit (IDR)"), errors="coerce").fillna(0).sum())
    return total_debit, total_credit


def _total_debit_credit_tb(df_tb: pd.DataFrame) -> Tuple[float, float]:
    if df_tb is None or df_tb.empty:
        return 0.0, 0.0
    total_debit = sum(
        float(pd.to_numeric(df_tb.get(f"{b}_Debit"), errors="coerce").fillna(0).sum()) for b in BULAN_URUT
    )
    total_credit = sum(
        float(pd.to_numeric(df_tb.get(f"{b}_Credit"), errors="coerce").fillna(0).sum()) for b in BULAN_URUT
    )
    return total_debit, total_credit


def _bulan_terakhir_dengan_data(df_bs_atau_pnl: pd.DataFrame) -> Optional[str]:
    """Bulan terakhir (paling akhir tahun) yang punya minimal 1 nilai bukan-0
    di salah satu akun -- dipakai sbg 'saldo akhir tahun berjalan' utk cek
    BS Balanced & sbg basis Amount BS_Tax (BS = snapshot posisi per tanggal,
    beda dgn PNL yang flow/diakumulasi 12 bulan)."""
    if df_bs_atau_pnl is None or df_bs_atau_pnl.empty:
        return None
    for bulan in reversed(BULAN_URUT):
        if bulan in df_bs_atau_pnl.columns and (pd.to_numeric(df_bs_atau_pnl[bulan], errors="coerce").fillna(0) != 0).any():
            return bulan
    return None


def hitung_fs_control(hasil: HasilKertasKerja, tahun: int) -> pd.DataFrame:
    """
    [BARU - lanjutan Step 2] Isi Value/Status sungguhan di kerangka
    FS_Control (label & urutan baris tetap dari susun_fs_control_kerangka).
    OVERALL STATUS sengaja mensyaratkan Opening_Balance & Adjustments
    JUGA sudah diisi/direview -- bukan cuma cek keseimbangan GL/TB/BS --
    karena working paper yang BARU digenerate dari PDF rekening koran
    SELALU punya Opening_Balance=0 & Adjustments=0 di awal (lihat
    susun_opening_balance/susun_template_adjustments), jadi OVERALL akan
    selalu "CHECK" sampai akuntan benar-benar review & isi keduanya --
    ini SENGAJA, supaya endpoint konfirmasi user (langkah berikutnya
    setelah ini) tidak mengizinkan lanjut ke laporan 18-sheet final dari
    working paper yang belum direview sama sekali.

    KETERBATASAN DIKETAHUI (belum diperbaiki di iterasi ini, TOLONG DICEK
    sebelum dipakai produksi): cek "BS_Monthly Balanced" HANYA menjumlah
    saldo dari aktivitas GL bulan berjalan (BS_Monthly), TIDAK menambahkan
    Opening_Balance di atasnya -- jadi kalau Opening_Balance client bukan
    nol (kasus paling umum utk client yang bukan usaha baru), BS_Monthly
    Balanced akan tetap "CHECK" walau reklasifikasi Adjustments sudah
    lengkap & benar, karena saldo awal belum ikut dihitung. Perbaikan yang
    dibutuhkan: thread Opening_Balance ke susun_bs_pnl_monthly() sbg saldo
    akumulasi awal per akun (bukan mulai dari 0), lalu sesuaikan cek ini.
    """
    df = susun_fs_control_kerangka(tahun)

    total_debit_gl, total_credit_gl = _total_debit_credit_gl(hasil.gl)
    gl_balanced = abs(total_debit_gl - total_credit_gl) <= _TOLERANSI_BALANCE

    jumlah_bulan_ok = 0
    if hasil.bank_control is not None and not hasil.bank_control.empty:
        jumlah_bulan_ok = int((hasil.bank_control["Status"] == "OK").sum())
    bank_control_ok = jumlah_bulan_ok >= 1

    total_debit_tb, total_credit_tb = _total_debit_credit_tb(hasil.tb_monthly)
    tb_balanced = abs(total_debit_tb - total_credit_tb) <= _TOLERANSI_BALANCE

    bulan_bs = _bulan_terakhir_dengan_data(hasil.bs_monthly)
    total_assets = total_liab_equity = 0.0
    bs_balanced = False
    if bulan_bs and hasil.coa is not None and not hasil.coa.empty and hasil.bs_monthly is not None:
        fs_group_map = hasil.coa.set_index(hasil.coa["Account No."].astype(str))["FS Group"]
        for _, r in hasil.bs_monthly.iterrows():
            no_akun = str(r["Account / Description"]).split(" - ", 1)[0]
            fs_group = str(fs_group_map.get(no_akun, "") or "").lower()
            nilai = _angka(r.get(bulan_bs))
            if any(k in fs_group for k in ("liabilitas", "kewajiban", "ekuitas", "modal")):
                total_liab_equity += nilai
            else:
                total_assets += nilai
        bs_balanced = abs(total_assets - total_liab_equity) <= _TOLERANSI_BALANCE

    df_opening = susun_opening_balance(hasil.coa) if hasil.coa is not None else pd.DataFrame()
    opening_terisi = bool((df_opening["Net Debit/(Credit)"] != 0).any()) if not df_opening.empty else False

    adjustments_direview = False
    if hasil.adjustments is not None and not hasil.adjustments.empty:
        kolom_nilai = [c for c in hasil.adjustments.columns if c.endswith(("_Debit", "_Credit"))]
        if kolom_nilai:
            nilai_adj = hasil.adjustments[kolom_nilai].apply(pd.to_numeric, errors="coerce").fillna(0)
            adjustments_direview = bool((nilai_adj != 0).any().any())

    overall_balanced = (
        gl_balanced and bank_control_ok and tb_balanced and bs_balanced
        and opening_terisi and adjustments_direview
    )

    nilai_status = {
        f"GL Balanced (Total Debit = Total Credit) - {tahun}": (
            round(total_debit_gl - total_credit_gl, 2), "BALANCED" if gl_balanced else "CHECK",
        ),
        "Bank_Control - Semua Bulan Berstatus OK": (
            jumlah_bulan_ok, "BALANCED" if bank_control_ok else "CHECK",
        ),
        "TB_Monthly Balanced (Total Debit = Total Credit)": (
            round(total_debit_tb - total_credit_tb, 2), "BALANCED" if tb_balanced else "CHECK",
        ),
        "BS_Monthly Balanced (Total Assets = Total Liabilities + Equity)": (
            round(total_assets - total_liab_equity, 2) if bulan_bs else None,
            "BALANCED" if bs_balanced else ("CHECK" if bulan_bs else "BELUM ADA DATA"),
        ),
        "Opening_Balance Sudah Diisi (bukan seluruhnya 0)": (
            None, "OK" if opening_terisi else "CHECK -- INPUT MANUAL DIBUTUHKAN",
        ),
        "Adjustments Sudah Direview Akuntan": (
            None, "OK" if adjustments_direview else "CHECK -- BELUM ADA KOREKSI MANUAL",
        ),
        f"OVERALL STATUS KERTAS KERJA {tahun}": (
            None, "BALANCED" if overall_balanced else "CHECK",
        ),
    }
    df["Value"] = [nilai_status[label][0] for label in df["Control"]]
    df["Status"] = [nilai_status[label][1] for label in df["Control"]]
    return df


def isi_bs_tax_dengan_saldo(df_bs_tax_kerangka: pd.DataFrame, df_bs_monthly: pd.DataFrame) -> pd.DataFrame:
    """
    [BARU - lanjutan Step 2] Isi kolom Amount (IDR) di kerangka BS_Tax
    dengan saldo akhir tahun berjalan dari BS_Monthly (bulan terakhir yang
    ada datanya -- BS itu snapshot posisi, bukan akumulasi 12 bulan).
    Ini masih ANGKA BUKU (commercial basis) apa adanya, BELUM ada koreksi
    fiskal apa pun -- kolom "Tax Return Category" tetap dari COA, dipakai
    sbg dasar mapping manual akuntan utk koreksi fiskal (langkah lanjutan
    berikutnya lagi, di luar cakupan working paper ini).

    [STRUKTUR] Karena skema BS_Tax sekarang menggabung Account No/Name
    jadi 1 kolom "Account / Description" (lihat susun_bs_tax_kerangka),
    pencocokan ke saldo BS_Monthly dilakukan dgn mengekstrak nomor akun
    dari awal string "Account / Description" (format "{no} - {nama}"),
    pola yang sama dipakai hitung_fs_control().
    """
    if df_bs_tax_kerangka is None or df_bs_tax_kerangka.empty:
        return df_bs_tax_kerangka
    bulan_bs = _bulan_terakhir_dengan_data(df_bs_monthly)
    if bulan_bs is None:
        return df_bs_tax_kerangka

    saldo_per_akun: Dict[str, float] = {}
    for _, r in df_bs_monthly.iterrows():
        no_akun = str(r["Account / Description"]).split(" - ", 1)[0]
        saldo_per_akun[no_akun] = _angka(r.get(bulan_bs))

    df = df_bs_tax_kerangka.copy()
    no_akun_kerangka = df["Account / Description"].astype(str).str.split(" - ", n=1).str[0]
    df["Amount (IDR)"] = no_akun_kerangka.map(saldo_per_akun).fillna(0)
    return df


def isi_pnl_tax_dengan_saldo(df_pnl_tax_kerangka: pd.DataFrame, df_pnl_monthly: pd.DataFrame) -> pd.DataFrame:
    """
    Sama seperti isi_bs_tax_dengan_saldo, tapi PNL adalah flow -- nilai
    per akun = jumlah 12 bulan (bukan saldo 1 bulan seperti BS).

    [STRUKTUR] Mengisi "Accounting Contribution" (bukan lagi "Amount
    (IDR)" -- lihat susun_pnl_tax_kerangka utk skema kolom baru), lalu
    "Fiscal Contribution" dihitung = Accounting Contribution + Auto
    Positive - Auto Negative + Manual Positive - Manual Negative. Auto/
    Manual Positive/Negative TETAP 0 di sini (belum ada logic deteksi
    otomatis maupun input manual -- lihat catatan lengkap di
    susun_pnl_tax_kerangka), jadi utk versi ini Fiscal Contribution akan
    SAMA PERSIS dgn Accounting Contribution.
    """
    if df_pnl_tax_kerangka is None or df_pnl_tax_kerangka.empty:
        return df_pnl_tax_kerangka
    if df_pnl_monthly is None or df_pnl_monthly.empty:
        return df_pnl_tax_kerangka

    total_per_akun: Dict[str, float] = {}
    for _, r in df_pnl_monthly.iterrows():
        no_akun = str(r["Account / Description"]).split(" - ", 1)[0]
        total_per_akun[no_akun] = sum(_angka(r.get(b)) for b in BULAN_URUT)

    df = df_pnl_tax_kerangka.copy()
    df["Accounting Contribution"] = df["Account No."].astype(str).map(total_per_akun).fillna(0)
    df["Fiscal Contribution"] = (
        df["Accounting Contribution"]
        + df["Auto Positive"] - df["Auto Negative"]
        + df["Manual Positive"] - df["Manual Negative"]
    )
    return df


def hitung_pph17_31e(df_pnl_tax_terisi: pd.DataFrame, df_coa: pd.DataFrame, tahun: int) -> Dict[str, Any]:
    """
    [BARU - lanjutan Step 2] Panggil pph_badan.hitung_pph_pasal_31e()
    dengan angka dari PNL_Tax (sudah diisi isi_pnl_tax_dengan_saldo()).

    ASUMSI YANG PERLU DIKONFIRMASI KE TIM (belum ada sumber otomatis lain
    di sistem untuk ini, jadi ditebak dari kata kunci -- pola yang sama
    dipakai bangun_coa_kertas_kerja_dari_db untuk kata_kunci_pnl):
    - Peredaran Bruto Usaha = jumlah akun PNL_Tax yang FS Group ATAU
      Account Name-nya mengandung kata "pendapatan"/"penjualan"/"revenue"/
      "sales" (lihat _KATA_KUNCI_PENDAPATAN). Akun PNL lain (HPP/beban/
      biaya) dianggap beban.
    - Laba Bersih Komersial = total pendapatan - total beban di atas.
    - Koreksi fiskal positif/negatif & kompensasi kerugian fiskal BELUM
      ada sumber otomatis di working paper ini -- tetap 0 (kertas kerja
      hanya sampai bank rekonsiliasi + TB/BS/PNL, BUKAN penyusutan aset
      tetap/koreksi fiskal lain yang ada di modul laporan 18-sheet) --
      akuntan input manual di sel biru ASUMSI/INPUT sheet PPh17_31E kalau
      perlu override sebelum lanjut ke laporan final.
    - skema_pajak dipakai default SKEMA_TARIF_UMUM_31E (pph_badan.py) --
      kalau client ternyata pakai skema lain (mis. PPh Final UMKM),
      perhitungan ini TIDAK berlaku, akuntan perlu proses lewat modul
      PPh Final UMKM terpisah.

    Returns: dict hasil pph_badan.hitung_pph_pasal_31e(), atau {} kalau
    PNL_Tax kosong (belum ada data utk dihitung).
    """
    try:
        # [FIX] Urutan lama (bare import -> '..' -> '.') tidak pernah
        # berhasil di deployment nyata: pph_badan.py adalah sibling module
        # DI DALAM package 'modules' (sama seperti laporan_keuangan.py yang
        # sudah benar diimport via 'from modules import laporan_keuangan'
        # di _bangun_laporan_bulanan_dari_tb_monthly() di file ini, dan
        # sama seperti main.py::'from modules import pph_badan'), BUKAN
        # sejajar/di atas package modules seperti akuntansi_ai.py.
        from modules import pph_badan as pb
    except ImportError:
        try:
            import pph_badan as pb  # fallback: run langsung dari dalam folder modules/
        except ImportError:
            # [FIX] Sebelumnya kalau kedua jalur import gagal, exception
            # menjalar ke tulis_kertas_kerja_excel() dan menggagalkan
            # SELURUH export (13 sheet lain yang sudah beres ikut hilang)
            # hanya gara-gara 1 sheet pajak -- padahal PPh17_31E memang
            # SENGAJA masih provisional/kerangka di tahap ini (lihat
            # WARNING statis di _tulis_sheet_pph17_31e). Sekarang gagal
            # dengan aman: kembalikan {} supaya sheet ditulis sbg kerangka
            # kosong (perilaku sebelum Step 2), 13 sheet lain tetap lengkap.
            return {}

    if df_pnl_tax_terisi is None or df_pnl_tax_terisi.empty:
        return {}

    fs_group_map: Dict[str, str] = {}
    if df_coa is not None and not df_coa.empty:
        fs_group_map = df_coa.set_index(df_coa["Account No."].astype(str))["FS Group"].to_dict()

    pendapatan_usaha = 0.0
    total_beban = 0.0
    for _, r in df_pnl_tax_terisi.iterrows():
        no_akun = str(r["Account No."])
        label = f"{fs_group_map.get(no_akun, '') or ''} {r.get('Account Name', '') or ''}".lower()
        # [STRUKTUR] Sebelumnya baca kolom "Amount (IDR)" -- kolom itu sudah
        # tidak ada lagi di skema PNL_Tax baru (lihat susun_pnl_tax_kerangka),
        # diganti "Fiscal Contribution" (= Accounting Contribution selama
        # Auto/Manual koreksi masih 0, lihat isi_pnl_tax_dengan_saldo).
        nilai = _angka(r.get("Fiscal Contribution"))
        if any(k in label for k in _KATA_KUNCI_PENDAPATAN):
            pendapatan_usaha += nilai
        else:
            total_beban += nilai

    laba_bersih_komersial = pendapatan_usaha - total_beban

    return pb.hitung_pph_pasal_31e(
        peredaran_bruto=pendapatan_usaha,
        laba_bersih_komersial=laba_bersih_komersial,
        koreksi_fiskal_positif=0.0,
        koreksi_fiskal_negatif=0.0,
        kompensasi_kerugian_fiskal=0.0,
        tahun_pajak=tahun,
    )


def _tulis_sheet_bs_tax(wb, df_bs_tax: pd.DataFrame, tahun: int):
    """
    [BARU - perbaikan struktur] BS_Tax -- sebelumnya lewat _tulis_dataframe
    generik (tabel utama saja). Ditambah blok "TAX ANNEX CONTROL" di bawah
    (Total Assets / Total Liabilities & Equity / Balance Check) sesuai
    contoh -- struktur kolom sama (4 kolom: label di kolom A, angka di
    kolom C, mengikuti posisi "Amount (IDR)"), nilainya dihitung dari
    df_bs_tax yang sudah diisi isi_bs_tax_dengan_saldo() kalau ada kolom
    Tax Return Category yang bisa dibedakan aset vs liabilitas/ekuitas;
    kalau tidak bisa ditentukan, dikosongkan (None) sesuai instruksi.
    """
    ws = wb.create_sheet("BS_Tax")
    ws.cell(row=1, column=1, value=f"BALANCE SHEET - TAX RETURN ANNEX ({tahun})").font = FONT_JUDUL
    baris_header = 3
    header = ["Tax Return Category", "Account / Description", "Amount (IDR)", "Review / Tax Note"]
    for j, kolom in enumerate(header, start=1):
        ws.cell(row=baris_header, column=j, value=kolom).font = FONT_HEADER

    baris = baris_header + 1
    if df_bs_tax is None or df_bs_tax.empty:
        ws.cell(row=baris, column=1, value="(tidak ada data)")
        baris += 1
    else:
        for _, r in df_bs_tax.iterrows():
            ws.cell(row=baris, column=1, value=r.get("Tax Return Category"))
            ws.cell(row=baris, column=2, value=r.get("Account / Description"))
            ws.cell(row=baris, column=3, value=r.get("Amount (IDR)"))
            ws.cell(row=baris, column=4, value=r.get("Review / Tax Note"))
            baris += 1

    baris += 1
    ws.cell(row=baris, column=1, value="TAX ANNEX CONTROL").font = FONT_HEADER
    baris += 1
    # [KOSONG SENGAJA] Total Assets/Liabilities & Equity/Balance Check
    # butuh klasifikasi aset vs liabilitas+ekuitas per akun (sama seperti
    # hitung_fs_control's BS_Monthly Balanced check) -- belum dihubungkan
    # ke sini di iterasi ini, jadi label baris ditulis tapi nilai None
    # dulu, sesuai instruksi "boleh kosong, yang penting kolomnya sudah ada".
    for label in ["Total Assets", "Total Liabilities & Equity", "Balance Check"]:
        ws.cell(row=baris, column=1, value=label)
        baris += 1

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 40


def _tulis_sheet_pnl_tax(wb, df_pnl_tax: pd.DataFrame, tahun: int):
    """
    [BARU - perbaikan struktur] PNL_Tax -- sebelumnya lewat _tulis_dataframe
    generik (tabel utama saja, tanpa subjudul instruksi). Ditambah subjudul
    & blok "FISCAL SUMMARY" di bawah (Gross Business Turnover/Accounting
    Profit Before Tax/Total Positive-Negative Corrections/Fiscal Net
    Income/Status) sesuai contoh -- nilai ringkasan dihitung dari
    df_pnl_tax yang sudah diisi isi_pnl_tax_dengan_saldo(); baris yang
    butuh sumber lain (mis. status review) dikosongkan dulu.
    """
    ws = wb.create_sheet("PNL_Tax")
    ws.cell(row=1, column=1, value=f"PROFIT & LOSS - TAX RETURN / FISCAL RECONCILIATION {tahun}").font = FONT_JUDUL
    ws.cell(row=2, column=1, value=(
        "Manual fiscal corrections are blue input cells. Auto/Manual Positive-Negative columns "
        "are placeholders for fiscal corrections to be reviewed by the accountant."
    ))
    baris_header = 4
    header = ["Account No.", "Account Name", "Tax Return Group", "Accounting Contribution",
              "Auto Positive", "Auto Negative", "Manual Positive", "Manual Negative",
              "Fiscal Contribution", "Review Notes"]
    for j, kolom in enumerate(header, start=1):
        ws.cell(row=baris_header, column=j, value=kolom).font = FONT_HEADER

    baris = baris_header + 1
    total_pendapatan = 0.0
    total_fiscal = 0.0
    total_positif = 0.0
    total_negatif = 0.0
    if df_pnl_tax is not None and not df_pnl_tax.empty:
        for _, r in df_pnl_tax.iterrows():
            for j, kolom in enumerate(header, start=1):
                ws.cell(row=baris, column=j, value=r.get(kolom))
            total_fiscal += _angka(r.get("Fiscal Contribution"))
            total_positif += _angka(r.get("Auto Positive")) + _angka(r.get("Manual Positive"))
            total_negatif += _angka(r.get("Auto Negative")) + _angka(r.get("Manual Negative"))
            label = str(r.get("Tax Return Group") or "").lower()
            if "peredaran usaha" in label:
                total_pendapatan += _angka(r.get("Fiscal Contribution"))
            baris += 1
    else:
        ws.cell(row=baris, column=1, value="(tidak ada data)")
        baris += 1

    baris += 1
    ws.cell(row=baris, column=1, value="FISCAL SUMMARY").font = FONT_HEADER
    baris += 1
    ringkasan_fiskal = [
        ("Gross Business Turnover", total_pendapatan),
        ("Accounting Profit Before Tax", total_fiscal),
        ("Total Positive Corrections", total_positif),
        ("Total Negative Corrections", total_negatif),
        ("Fiscal Net Income", total_fiscal + total_positif - total_negatif),
        ("Status", None),
    ]
    for label, value in ringkasan_fiskal:
        ws.cell(row=baris, column=1, value=label)
        ws.cell(row=baris, column=3, value=value)
        baris += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 26
    for col in "DEFGHI":
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["J"].width = 30


def _tulis_sheet_placeholder(wb, nama: str, judul: str):
    """[Step 2] Sudah tidak dipakai oleh tulis_kertas_kerja_excel() untuk
    ke-5 sheet yang dulu placeholder (FS_Control/Opening_Balance/BS_Tax/
    PNL_Tax/PPh17_31E sekarang punya kerangka sendiri-sendiri, lihat
    fungsi susun_*_kerangka & _tulis_sheet_pph17_31e). Dibiarkan ada
    sebagai utilitas generik kalau nanti butuh sheet placeholder baru."""
    ws = wb.create_sheet(nama)
    ws["A1"] = judul
    ws["A1"].font = FONT_JUDUL
    ws["A2"] = "TODO: sheet ini belum diisi otomatis di versi awal modul kertas_kerja.py."


def _tulis_sheet_identity(wb, identitas: Dict[str, str]):
    """
    [PERBAIKAN STRUKTUR] Identity -- sebelumnya cuma tabel Field/Value (2
    kolom, 5 baris). Ditambah field tambahan (Address/Tax Year/Currency/
    Bank/dst -- semua tetap input manual, sel biru) DAN tabel ke-2 di
    bawahnya (Tax Model / Value / Source / URL / Note, 5 kolom) yang
    merangkum skema PPh Badan yang dipakai -- sesuai contoh
    Kertas_Kerja_Laporan_Keuangan (max 6 kolom di sheet ini).
    """
    try:
        import pph_badan as pb
    except ImportError:  # pragma: no cover
        try:
            from .. import pph_badan as pb
        except ImportError:
            try:
                from . import pph_badan as pb
            except ImportError:
                pb = None

    ws = wb.create_sheet("Identity")
    ws["A1"] = "BUSINESS IDENTITY & FINANCIAL STATEMENT BASIS"
    ws["A1"].font = FONT_JUDUL
    ws["A3"] = "Field"
    ws["B3"] = "Value"
    ws["A3"].font = FONT_HEADER
    ws["B3"].font = FONT_HEADER

    fill_manual = PatternFill("solid", fgColor=WARNA_INPUT_MANUAL)
    field_default = [
        "Business / Entity Name", "NPWP", "NIB / Business License", "Legal Form",
        "Business Activity", "Address", "Tax Year", "Currency", "Bank",
        "Bank Account No.", "Bank Account Holder", "Accounting Basis",
    ]
    baris = 4
    for field_name in field_default:
        ws.cell(row=baris, column=1, value=field_name)
        ws.cell(row=baris, column=2, value=identitas.get(field_name)).fill = fill_manual
        baris += 1

    baris += 1
    ws.cell(row=baris, column=1, value="IMPORTANT BASIS / LIMITATION").font = FONT_HEADER
    baris += 1
    ws.cell(row=baris, column=1, value=identitas.get("Basis / Limitation Note"))
    baris += 3

    # Tabel ke-2: ringkasan skema PPh Badan (Tax Model) -- sama seperti
    # blok ASUMSI/INPUT di sheet PPh17_31E, disalin ringkas di sini
    # supaya identitas & basis pajak client ada di satu tempat.
    tarif_umum = getattr(pb, "TARIF_PPH_BADAN_UMUM", 0.22) if pb else 0.22
    persen_pengurangan = getattr(pb, "PERSENTASE_PENGURANGAN_PASAL_31E", 0.5) if pb else 0.5
    tarif_efektif = getattr(pb, "TARIF_EFEKTIF_FASILITAS", tarif_umum * (1 - persen_pengurangan)) if pb else 0.11
    batas_maks_fasilitas = getattr(pb, "BATAS_MAKS_PEREDARAN_BRUTO_FASILITAS", 50_000_000_000) if pb else 50_000_000_000
    batas_fasilitas_penuh = getattr(pb, "BATAS_PEREDARAN_BRUTO_FASILITAS_PENUH", 4_800_000_000) if pb else 4_800_000_000

    header_pajak = ["Tax Model", "Value", "Source", "URL", "Note"]
    for j, kolom in enumerate(header_pajak, start=1):
        ws.cell(row=baris, column=j, value=kolom).font = FONT_HEADER
    ws.cell(row=baris, column=6, value=None)  # kolom F sengaja kosong, samakan lebar sheet dgn contoh
    baris += 1

    baris_pajak = [
        ("Standard Corporate Rate", tarif_umum, None, None, "Pasal 17 UU PPh"),
        ("31E Reduced Effective Rate", tarif_efektif, None, None, "50% x tarif umum"),
        ("31E Gross Turnover Limit", batas_maks_fasilitas, None, None, "Pasal 31E"),
        ("31E Facility Portion", batas_fasilitas_penuh, None, None,
         "Assumes normal corporate income tax regime, not final turnover tax regime."),
    ]
    for label, value, source, url, note in baris_pajak:
        ws.cell(row=baris, column=1, value=label)
        ws.cell(row=baris, column=2, value=value)
        ws.cell(row=baris, column=3, value=source)
        ws.cell(row=baris, column=4, value=url)
        ws.cell(row=baris, column=5, value=note)
        baris += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 40


def _tulis_sheet_pph17_31e(wb, tahun: int, hasil_pph: Optional[Dict[str, Any]] = None):
    """
    Sheet PPh17_31E -- perhitungan PPh Badan tarif Pasal 17 + fasilitas
    pengurangan tarif Pasal 31E. Strukturnya SENGAJA mengikuti field &
    konstanta yang SUDAH ADA di modules/pph_badan.py (TARIF_PPH_BADAN_UMUM,
    PERSENTASE_PENGURANGAN_PASAL_31E, TARIF_EFEKTIF_FASILITAS,
    BATAS_PEREDARAN_BRUTO_FASILITAS_PENUH, dst -- BUKAN angka baru yang
    ditebak) supaya baris & urutannya selaras dengan fungsi
    hitung_total_peredaran_bruto() / hitung_pkp() /
    hitung_pembagian_fasilitas_31e() / hitung_pph_pasal_31e() di modul itu.

    2 bagian:
      1. ASUMSI / INPUT (sel biru, boleh diedit akuntan) -- default value
         standar dari konstanta pph_badan.py (per UU HPP, WAJIB
         diverifikasi ulang ke peraturan terbaru sebelum dipakai untuk
         pelaporan resmi -- lihat catatan yang sama di pph_badan.py).
      2. PERHITUNGAN -- [lanjutan Step 2] kalau `hasil_pph` diisi (hasil
         hitung_pph17_31e()), baris ini diisi ANGKA SUNGGUHAN dari situ.
         Kalau None (mis. PNL_Tax masih kosong/gagal dihitung), tetap
         fallback ke 0 seperti kerangka semula -- supaya sheet tetap
         terbentuk lengkap walau perhitungan belum bisa dilakukan.
    """
    try:
        import pph_badan as pb
    except ImportError:  # pragma: no cover -- ikuti pola import akuntansi_ai di atas
        try:
            from .. import pph_badan as pb
        except ImportError:
            try:
                from . import pph_badan as pb
            except ImportError:
                pb = None  # konstanta di bawah fallback ke default hardcode kalau modul tak ketemu

    tarif_umum = getattr(pb, "TARIF_PPH_BADAN_UMUM", 0.22)
    persen_pengurangan = getattr(pb, "PERSENTASE_PENGURANGAN_PASAL_31E", 0.5)
    tarif_efektif = getattr(pb, "TARIF_EFEKTIF_FASILITAS", tarif_umum * (1 - persen_pengurangan))
    batas_fasilitas_penuh = getattr(pb, "BATAS_PEREDARAN_BRUTO_FASILITAS_PENUH", 4_800_000_000)
    batas_maks_fasilitas = getattr(pb, "BATAS_MAKS_PEREDARAN_BRUTO_FASILITAS", 50_000_000_000)
    skema_default = getattr(pb, "SKEMA_TARIF_UMUM_31E", "Tarif Umum Pasal 17/31E")

    ws = wb.create_sheet("PPh17_31E")
    ws["A1"] = f"CORPORATE INCOME TAX CALCULATION - PASAL 17 & PASAL 31E ({tahun})"
    ws["A1"].font = FONT_JUDUL
    ws["A2"] = (
        "WARNING: Provisional tax calculation. Do not use for filing until staging accounts, "
        "Adjustments, and fiscal corrections are reviewed."
    )

    fill_manual = PatternFill("solid", fgColor=WARNA_INPUT_MANUAL)

    # [PERBAIKAN STRUKTUR] Sebelumnya sheet ini cuma 2 kolom (Field/Value).
    # Contoh Kertas_Kerja_Laporan_Keuangan pakai 6 kolom: Assumption/Input,
    # Value, Source/Note, (kolom D kosong sbg pemisah), Official Source,
    # URL -- disamakan di sini. Kolom Official Source/URL hanya diisi utk
    # baris ASUMSI (referensi resmi DJP/JDIH), baris PERHITUNGAN memakai
    # kolom C (Source/Note) saja, sisanya kosong (None) -- sama seperti
    # contoh.
    baris = 5
    header_asumsi = ["Assumption / Input", "Value", "Source / Note", None, "Official Source", "URL"]
    for j, kolom in enumerate(header_asumsi, start=1):
        if kolom is not None:
            ws.cell(row=baris, column=j, value=kolom).font = FONT_HEADER
    baris += 1

    # Urutan & label mengikuti struktur fungsi di pph_badan.py:
    # hitung_total_peredaran_bruto -> hitung_pkp -> hitung_pembagian_fasilitas_31e
    # -> hitung_pph_pasal_31e (lihat docstring fungsi ini). Kalau hasil_pph
    # diisi, key di bawah PERSIS mengikuti struktur dict yang dikembalikan
    # pph_badan.hitung_pph_pasal_31e() (lihat hitung_pph17_31e() di atas).
    pb_detail = (hasil_pph or {}).get("peredaran_bruto_detail", {})
    pb_rekon = (hasil_pph or {}).get("rekonsiliasi_fiskal", {})
    pb_fasilitas = (hasil_pph or {}).get("fasilitas_31e", {})

    asumsi = [
        # (label, value, source/note, official source, url) -- 2 baris
        # terakhir (official source/url) diisi HANYA utk baris asumsi
        # utama, sesuai contoh (baris perhitungan tidak punya referensi).
        ("Tax Year", tahun, None,
         "DJP - Corporate Income Tax Calculation",
         "https://pajak.go.id/id/mekanisme-penghitungan-pajak-penghasilan-badan"),
        ("Tax Regime", skema_default, "Normal corporate regime assumed",
         "DJP - Corporate Income Tax", None),
        ("Standard Rate", tarif_umum, "Corporate rate (Pasal 17)",
         "JDIH - Income Tax Law / Pasal 31E", "https://jdih.kemenkeu.go.id/dok/uu-8-tahun-1983/view"),
        ("31E Reduced Rate", tarif_efektif, "50% x tarif umum", None, None),
        ("31E Gross Turnover Limit", batas_maks_fasilitas, "Pasal 31E", None, None),
        ("31E Facility Turnover Portion", batas_fasilitas_penuh, "Pasal 31E", None, None),
    ]
    for label, value, source, official, url in asumsi:
        cell_label = ws.cell(row=baris, column=1, value=label)
        cell_value = ws.cell(row=baris, column=2, value=value)
        cell_label.fill = fill_manual
        cell_value.fill = fill_manual
        ws.cell(row=baris, column=3, value=source)
        ws.cell(row=baris, column=5, value=official)
        ws.cell(row=baris, column=6, value=url)
        baris += 1

    baris += 1
    ws.cell(row=baris, column=1, value="TAXABLE INCOME & PASAL 31E").font = FONT_HEADER
    baris += 1

    label_perhitungan = [
        ("Gross Business Turnover", pb_detail.get("total_peredaran_bruto") or pb_detail.get("peredaran_bruto_usaha", 0), "Operating revenue"),
        ("Fiscal Net Income", pb_rekon.get("penghasilan_neto_fiskal", 0), "After fiscal corrections"),
        ("Fiscal Loss Carryforward Utilized", pb_rekon.get("kompensasi_kerugian_fiskal", 0), "Manual input"),
        ("Taxable Income Before Rounding", pb_rekon.get("penghasilan_kena_pajak", 0), None),
        ("PKP Eligible for 31E Reduced Rate", pb_fasilitas.get("pkp_mendapat_fasilitas", 0), None),
        ("PKP Subject to Standard Rate", pb_fasilitas.get("pkp_tidak_mendapat_fasilitas", 0), None),
        ("Tax on 31E Portion", (hasil_pph or {}).get("pph_atas_pkp_fasilitas", 0), None),
        ("Tax on Standard Portion", (hasil_pph or {}).get("pph_atas_pkp_nonfasilitas", 0), None),
        ("Total Corporate Income Tax", (hasil_pph or {}).get("pph_badan_terutang", 0), None),
    ]
    for label, value, note in label_perhitungan:
        ws.cell(row=baris, column=1, value=label)
        ws.cell(row=baris, column=2, value=value)
        ws.cell(row=baris, column=3, value=note)
        baris += 1

    baris += 1
    ws.cell(row=baris, column=1, value="TAX CREDITS / PPh 29").font = FONT_HEADER
    baris += 1
    # [KOSONG SENGAJA] Kredit pajak (PPh 22/23/24/25) belum ada sumber
    # otomatis di kertas kerja ini (tidak diekstrak dari rekening koran) --
    # struktur baris disiapkan, nilai default 0/input manual sesuai contoh.
    kredit_pajak = [
        ("PPh 22 Tax Credit", 0, "Manual input"),
        ("PPh 23 Tax Credit", 0, "Manual input"),
        ("PPh 24 Foreign Tax Credit", 0, "Manual input"),
        ("PPh 25 Installments", 0, "Manual input"),
        ("Total Tax Credits", 0, None),
        ("PPh 29 Payable / (Overpayment)", (hasil_pph or {}).get("pph_badan_terutang", 0), None),
    ]
    for label, value, note in kredit_pajak:
        cell_value = ws.cell(row=baris, column=1, value=label)
        val_cell = ws.cell(row=baris, column=2, value=value)
        if note == "Manual input":
            val_cell.fill = fill_manual
        ws.cell(row=baris, column=3, value=note)
        baris += 1

    if hasil_pph and hasil_pph.get("fasilitas_31e", {}).get("status_fasilitas"):
        baris += 1
        ws.cell(row=baris, column=1, value="Status Fasilitas 31E").font = FONT_HEADER
        ws.cell(row=baris, column=2, value=hasil_pph["fasilitas_31e"]["status_fasilitas"])

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["E"].width = 34
    ws.column_dimensions["F"].width = 46


def _tulis_sheet_bs_pnl_monthly(
    wb, nama_sheet: str, df: pd.DataFrame, judul: str,
    tambah_kolom_total: bool = False, tahun: Optional[int] = None,
):
    """
    [BARU - struktur disamakan dgn contoh] Penulis khusus utk BS_Monthly/
    PNL_Monthly -- BEDA dari _tulis_dataframe generik krn kolom "FS Group"
    di df TIDAK ditulis sbg kolom, melainkan dipakai utk mengelompokkan
    akun jadi BARIS header per grup (mis. "ASSETS", "OPERATING REVENUE"),
    sesuai tampilan di contoh Kertas_Kerja_Laporan_Keuangan. Urutan grup
    mengikuti urutan kemunculan pertama di df (yang berarti mengikuti
    urutan baris di sheet COA client).

    tambah_kolom_total=True (dipakai utk PNL_Monthly) menambah 1 kolom
    "Total <tahun>" di akhir, isi dari kolom "Total Tahun" yang sudah
    dihitung susun_bs_pnl_monthly().

    CATATAN KETERBATASAN: baris SUBTOTAL per grup (mis. "TOTAL ASSETS")
    dan baris ringkasan akhir (mis. "NET PROFIT / (LOSS)") ada di contoh
    tapi SENGAJA belum direproduksi di sini -- itu logic perhitungan,
    bukan struktur kolom (lihat catatan yg sama di susun_bs_pnl_monthly).
    """
    ws = wb.create_sheet(nama_sheet)
    ws.cell(row=1, column=1, value=judul).font = FONT_JUDUL
    baris_header = 3

    header = ["Account / Description"] + BULAN_URUT + ([f"Total {tahun}"] if tambah_kolom_total else [])
    for j, kolom in enumerate(header, start=1):
        ws.cell(row=baris_header, column=j, value=kolom).font = FONT_HEADER

    if df is None or df.empty:
        ws.cell(row=baris_header + 1, column=1, value="(tidak ada data)")
        return

    baris = baris_header + 1
    df_grup = df.assign(_grup=df["FS Group"].fillna("(Tanpa FS Group)"))
    for grup, subset in df_grup.groupby("_grup", sort=False):
        ws.cell(row=baris, column=1, value=str(grup).upper()).font = FONT_HEADER
        baris += 1
        for _, r in subset.iterrows():
            ws.cell(row=baris, column=1, value=r.get("Account / Description"))
            for j, bulan in enumerate(BULAN_URUT, start=2):
                ws.cell(row=baris, column=j, value=r.get(bulan))
            if tambah_kolom_total:
                ws.cell(row=baris, column=2 + len(BULAN_URUT), value=r.get("Total Tahun"))
            baris += 1

    ws.column_dimensions["A"].width = 48
    for j in range(2, 2 + len(BULAN_URUT) + (1 if tambah_kolom_total else 0)):
        ws.column_dimensions[get_column_letter(j)].width = 16


def _tulis_header_bulan_dua_baris(ws, baris_atas: int, baris_bawah: int, kolom_mulai: int,
                                   sub_labels: Tuple[str, ...] = ("Debit", "Credit")) -> int:
    """
    [BARU - perbaikan struktur] Tulis header 2 baris bergaya kertas kerja
    contoh (Bank_Posting_Summary/Adjustments/TB_Monthly): baris ATAS = nama
    bulan (di-merge sepanjang jumlah sub_labels, mis. Jan menutupi 2 kolom
    Debit+Credit), baris BAWAH = label sub kolom per bulan (mis. "Debit",
    "Credit", atau "Net Helper" kalau sub_labels cuma 1 elemen).

    Sebelumnya (_tulis_dataframe generik) nama kolom ditulis FLAT 1 baris
    (mis. "Jan_Debit") -- beda dari contoh yang pakai 2 baris header dengan
    nama bulan di-merge. Fungsi ini menyamakan tampilannya persis.

    Returns: nomor kolom SETELAH blok 12 bulan selesai (utk lanjut nulis
    blok kolom berikutnya kalau ada, mis. Net Helper di TB_Monthly).
    """
    kolom = kolom_mulai
    for bulan in BULAN_URUT:
        ws.cell(row=baris_atas, column=kolom, value=bulan).font = FONT_HEADER
        if len(sub_labels) > 1:
            ws.merge_cells(start_row=baris_atas, start_column=kolom,
                            end_row=baris_atas, end_column=kolom + len(sub_labels) - 1)
            for j, label in enumerate(sub_labels):
                ws.cell(row=baris_bawah, column=kolom + j, value=label).font = FONT_HEADER
        else:
            ws.cell(row=baris_bawah, column=kolom, value=sub_labels[0]).font = FONT_HEADER
        kolom += len(sub_labels)
    return kolom


def _tulis_sheet_posting_style(
    wb, nama_sheet: str, df: pd.DataFrame, judul: str, kolom_leading: List[str],
    subjudul: Optional[str] = None, kolom_input_manual: bool = False,
):
    """
    [BARU - perbaikan struktur] Penulis khusus utk sheet gaya "akun x bulan
    (Debit/Credit)" -- dipakai utk Bank_Posting_Summary & Adjustments.
    Header 2 baris (nama bulan di-merge di baris atas, "Debit"/"Credit" di
    baris bawah) SAMA seperti contoh Kertas_Kerja_Laporan_Keuangan, beda
    dari _tulis_dataframe generik yang menulis nama kolom flat.

    kolom_leading: nama kolom non-bulan di depan tabel (mis.
        ["Account No.", "Account Name", "Normal Balance"] utk
        Bank_Posting_Summary, atau ["Account No.", "Account Name",
        "Review Notes"] utk Adjustments) -- header-nya di-merge vertikal
        menutupi 2 baris header supaya sejajar dgn header bulan di
        sebelahnya. df WAJIB punya kolom-kolom ini plus "{Bulan}_Debit"/
        "{Bulan}_Credit" utk tiap bulan di BULAN_URUT.
    """
    ws = wb.create_sheet(nama_sheet)
    ws.cell(row=1, column=1, value=judul).font = FONT_JUDUL
    baris = 2
    if subjudul:
        ws.cell(row=baris, column=1, value=subjudul)
        baris += 1
    baris += 1  # baris kosong pemisah, sama seperti contoh (judul, subjudul, kosong, header)
    baris_header_atas = baris
    baris_header_bawah = baris + 1

    for j, kolom in enumerate(kolom_leading, start=1):
        ws.cell(row=baris_header_atas, column=j, value=kolom).font = FONT_HEADER
        ws.merge_cells(start_row=baris_header_atas, start_column=j, end_row=baris_header_bawah, end_column=j)

    kolom_mulai_bulan = len(kolom_leading) + 1
    _tulis_header_bulan_dua_baris(ws, baris_header_atas, baris_header_bawah, kolom_mulai_bulan)

    baris_data_mulai = baris_header_bawah + 1
    if df is None or df.empty:
        ws.cell(row=baris_data_mulai, column=1, value="(tidak ada data)")
        return

    fill_manual = PatternFill("solid", fgColor=WARNA_INPUT_MANUAL) if kolom_input_manual else None
    baris_data = baris_data_mulai
    for _, r in df.iterrows():
        for j, kolom in enumerate(kolom_leading, start=1):
            ws.cell(row=baris_data, column=j, value=r.get(kolom))
        kolom_idx = kolom_mulai_bulan
        for bulan in BULAN_URUT:
            cell_d = ws.cell(row=baris_data, column=kolom_idx, value=r.get(f"{bulan}_Debit"))
            cell_c = ws.cell(row=baris_data, column=kolom_idx + 1, value=r.get(f"{bulan}_Credit"))
            if fill_manual is not None:
                cell_d.fill = fill_manual
                cell_c.fill = fill_manual
            kolom_idx += 2
        baris_data += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 34
    ws.freeze_panes = ws.cell(row=baris_data_mulai, column=kolom_mulai_bulan).coordinate


def _tulis_sheet_tb_monthly(wb, df: pd.DataFrame, judul: str):
    """
    [BARU - perbaikan struktur] Penulis khusus utk TB_Monthly -- SAMA
    seperti _tulis_sheet_posting_style (header bulan Debit/Credit 2 baris)
    DITAMBAH 12 kolom "Net Helper" di akhir (1 kolom per bulan, header
    ATAS = nama bulan lagi -- TIDAK di-merge kali ini, header BAWAH =
    "Net Helper"), persis struktur di contoh Kertas_Kerja_Laporan_Keuangan
    (kolom AC..AN).
    """
    ws = wb.create_sheet("TB_Monthly")
    ws.cell(row=1, column=1, value=judul).font = FONT_JUDUL
    baris_header_atas = 4
    baris_header_bawah = 5
    kolom_leading = ["Account No.", "Account Name", "Normal Balance"]

    for j, kolom in enumerate(kolom_leading, start=1):
        ws.cell(row=baris_header_atas, column=j, value=kolom).font = FONT_HEADER
        ws.merge_cells(start_row=baris_header_atas, start_column=j, end_row=baris_header_bawah, end_column=j)

    kolom_mulai_bulan = len(kolom_leading) + 1
    kolom_setelah_dk = _tulis_header_bulan_dua_baris(ws, baris_header_atas, baris_header_bawah, kolom_mulai_bulan)

    kolom_net_helper_mulai = kolom_setelah_dk
    kolom = kolom_net_helper_mulai
    for bulan in BULAN_URUT:
        ws.cell(row=baris_header_atas, column=kolom, value=bulan).font = FONT_HEADER
        ws.cell(row=baris_header_bawah, column=kolom, value="Net Helper").font = FONT_HEADER
        kolom += 1

    baris_data_mulai = baris_header_bawah + 1
    if df is None or df.empty:
        ws.cell(row=baris_data_mulai, column=1, value="(tidak ada data)")
        return

    baris_data = baris_data_mulai
    for _, r in df.iterrows():
        for j, kolom_nama in enumerate(kolom_leading, start=1):
            ws.cell(row=baris_data, column=j, value=r.get(kolom_nama))
        kolom_idx = kolom_mulai_bulan
        for bulan in BULAN_URUT:
            ws.cell(row=baris_data, column=kolom_idx, value=r.get(f"{bulan}_Debit"))
            ws.cell(row=baris_data, column=kolom_idx + 1, value=r.get(f"{bulan}_Credit"))
            kolom_idx += 2
        kolom_idx = kolom_net_helper_mulai
        for bulan in BULAN_URUT:
            ws.cell(row=baris_data, column=kolom_idx, value=r.get(f"{bulan}_Net Helper"))
            kolom_idx += 1
        baris_data += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 34
    ws.freeze_panes = ws.cell(row=baris_data_mulai, column=kolom_mulai_bulan).coordinate


def _tulis_sheet_fs_control(wb, df_control: pd.DataFrame, tahun: int, hasil: "HasilKertasKerja"):
    """
    [BARU - perbaikan struktur] Penulis khusus utk FS_Control -- SEBELUMNYA
    ditulis lewat _tulis_dataframe generik (cuma 4 kolom: Control/Value/
    Target/Status). Di contoh Kertas_Kerja_Laporan_Keuangan, sheet ini
    punya 3 blok:
      1. Tabel CONTROL (kolom A-D) -- cek keseimbangan angka, isinya dari
         hitung_fs_control() (logic TIDAK berubah).
      2. Tabel CLASSIFICATION SUMMARY (kolom F-H, sejajar tabel 1) --
         ringkasan jumlah transaksi per tingkat confidence dari GL.
      3. Tabel MONTHLY P&L SNAPSHOT (kolom A-D, di bawah tabel 1) --
         [STRUKTUR SAJA] belum ada sumber Total Expenses/Profit per bulan
         yang terpisah dari GL Suggested Counterpart di modul ini, jadi
         nilainya sengaja dikosongkan (None) dulu -- lihat instruksi:
         boleh kosong, yang penting struktur/kolomnya sudah sesuai.
    """
    ws = wb.create_sheet("FS_Control")
    ws.cell(row=1, column=1, value=f"FINANCIAL STATEMENT WORKING PAPER CONTROL - {tahun}").font = FONT_JUDUL
    ws.cell(row=2, column=1, value=(
        "Bank extraction & financial statement classification control summary. "
        "Blue cells are user inputs."
    ))

    baris_header = 5
    header_kontrol = ["CONTROL", "VALUE", "TARGET / REFERENCE", "STATUS"]
    for j, kolom in enumerate(header_kontrol, start=1):
        ws.cell(row=baris_header, column=j, value=kolom).font = FONT_HEADER
    ws.cell(row=baris_header, column=6, value="CLASSIFICATION SUMMARY").font = FONT_HEADER
    ws.cell(row=baris_header + 1, column=6, value="Metric").font = FONT_HEADER
    ws.cell(row=baris_header + 1, column=7, value="Value").font = FONT_HEADER
    ws.cell(row=baris_header + 1, column=8, value="Comment").font = FONT_HEADER

    baris = baris_header + 1
    df_control = df_control if df_control is not None else pd.DataFrame()
    for _, r in df_control.iterrows():
        ws.cell(row=baris, column=1, value=r.get("Control"))
        ws.cell(row=baris, column=2, value=r.get("Value"))
        ws.cell(row=baris, column=3, value=r.get("Target / Reference"))
        ws.cell(row=baris, column=4, value=r.get("Status"))
        baris += 1

    ringkasan = ringkasan_status_kertas_kerja(hasil)
    cc = ringkasan.get("confidence_count", {})
    baris_ringkasan = [
        ("Total bank transactions", ringkasan.get("jumlah_transaksi", 0), None),
        ("High-confidence transactions", cc.get("High", 0), None),
        ("Medium-confidence transactions", cc.get("Medium", 0), None),
        ("Low-confidence transactions", cc.get("Low", 0), None),
    ]
    baris_r = baris_header + 2
    for label, value, comment in baris_ringkasan:
        ws.cell(row=baris_r, column=6, value=label)
        ws.cell(row=baris_r, column=7, value=value)
        ws.cell(row=baris_r, column=8, value=comment)
        baris_r += 1

    baris_pnl_start = max(baris, baris_r) + 2
    ws.cell(row=baris_pnl_start, column=1, value="MONTHLY P&L SNAPSHOT (IDR)").font = FONT_HEADER
    header_pnl = ["Month", "Revenue", "Total Expenses", "Profit Before Tax"]
    for j, kolom in enumerate(header_pnl, start=1):
        ws.cell(row=baris_pnl_start + 1, column=j, value=kolom).font = FONT_HEADER
    for i, bulan in enumerate(BULAN_URUT, start=baris_pnl_start + 2):
        ws.cell(row=i, column=1, value=bulan)
        # [KOSONG SENGAJA] lihat docstring fungsi ini.

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["F"].width = 30


def _tulis_dataframe(wb, nama_sheet: str, df: pd.DataFrame, judul: str = "", freeze: Optional[str] = None,
                      kolom_input_manual: bool = False, kolom_manual: Optional[List[str]] = None,
                      subjudul: Optional[str] = None):
    """
    kolom_input_manual: True -> highlight semua kolom yang namanya
        berakhiran "_Debit"/"_Credit" (dipakai Adjustments, banyak kolom
        per bulan dgn pola nama seragam).
    kolom_manual: [BARU - Step 2] list nama kolom PERSIS yang mau
        di-highlight sebagai input manual -- dipakai saat kolomnya sedikit
        & namanya tidak mengikuti pola _Debit/_Credit (mis. Opening_Balance
        punya 1 kolom "Opening Balance (IDR)"). Boleh dipakai bersamaan
        dengan kolom_input_manual kalau suatu saat dibutuhkan keduanya.
    subjudul: [BARU - perbaikan struktur] baris ke-2 opsional (teks
        instruksi/catatan, mis. "Bank extraction is fully reconciled...")
        -- di contoh Kertas_Kerja_Laporan_Keuangan beberapa sheet
        (GL, Adjustments, PNL_Tax, PPh17_31E) punya baris catatan ini
        persis di bawah judul, sebelum baris kosong & header kolom.
        Kalau diisi, header kolom otomatis turun 1 baris (row 4, bukan 3).
    """
    ws = wb.create_sheet(nama_sheet)
    baris_awal = 1
    if judul:
        ws.cell(row=1, column=1, value=judul).font = FONT_JUDUL
        baris_awal = 3
    if subjudul:
        ws.cell(row=2, column=1, value=subjudul)
        baris_awal = 4

    if df is None or df.empty:
        ws.cell(row=baris_awal, column=1, value="(tidak ada data)")
        return

    for j, kolom in enumerate(df.columns, start=1):
        cell = ws.cell(row=baris_awal, column=j, value=str(kolom))
        cell.font = FONT_HEADER

    fill_manual = PatternFill("solid", fgColor=WARNA_INPUT_MANUAL) if (kolom_input_manual or kolom_manual) else None
    for i, (_, row) in enumerate(df.iterrows(), start=baris_awal + 1):
        for j, kolom in enumerate(row.index, start=1):
            val = row[kolom]
            if isinstance(val, (pd.Timestamp, datetime, date)):
                val = pd.Timestamp(val).to_pydatetime()
            cell = ws.cell(row=i, column=j, value=val)
            akan_di_highlight = (
                (kolom_input_manual and str(kolom).endswith(("_Debit", "_Credit")))
                or (kolom_manual and str(kolom) in kolom_manual)
            )
            if akan_di_highlight:
                cell.fill = fill_manual

    for j, kolom in enumerate(df.columns, start=1):
        lebar = max(10, min(40, len(str(kolom)) + 2))
        ws.column_dimensions[get_column_letter(j)].width = lebar

    if freeze:
        ws.freeze_panes = freeze

# ============================================================
# 12. [BARU] JEMBATAN: KERTAS KERJA -> DATA EXPORT 18-SHEET
# ============================================================
# Menutup gap "user konfirmasi -> generate laporan 18-sheet dari kertas
# kerja ini" yang sebelumnya belum ada (generator 18-sheet lama hanya
# membaca jurnal yang SUDAH diposting ke database, bukan hasil kertas
# kerja bank-only ini).
#
# KEPUTUSAN DESAIN PENTING (tolong direview tim sebelum dipakai produksi):
#
# 1. Adjustments (sheet ke-8) TIDAK dikonversi jadi "jurnal" ala database.
#    Adjustments di kertas kerja adalah koreksi PER AKUN PER BULAN (rec[f"
#    {bulan}_Debit"]/rec[f"{bulan}_Credit"] langsung ditambahkan ke saldo
#    akun itu di susun_tb_monthly()) -- BUKAN pasangan debit/kredit yang
#    saling seimbang per baris seperti jurnal database
#    (no_akun_debet/no_akun_kredit dalam 1 baris yang sama). Memaksakan
#    Adjustments jadi jurnal berpasangan butuh akun lawan buatan
#    (mis. akun "suspense") yang TIDAK diminta strukturnya di kertas
#    kerja manapun -- berisiko malah mengaburkan hasil, bukan membantu.
#
#    Solusinya: TB_Monthly (GL + Adjustments, sudah dihitung
#    susun_tb_monthly()) dipakai LANGSUNG sebagai sumber saldo_per_akun
#    untuk Neraca/Laba Rugi/Perubahan Ekuitas/CALK/laporan bulanan --
#    BUKAN dihitung ulang dari jurnal mentah lewat
#    lapkeu.hitung_saldo_per_akun(). Ini justru LEBIH BENAR (adjustments
#    ikut kepakai), bukan workaround yang lebih lemah.
#
# 2. Arus Kas TETAP dihitung dari jurnal (GL asli, hasil transaksi bank
#    riil) lewat lapkeu.susun_arus_kas_sederhana() -- BUKAN dari
#    TB_Monthly -- karena arus kas butuh info transaksi PER BARIS (akun
#    lawan tiap pergerakan kas) untuk diklasifikasi Operasi/Investasi/
#    Pendanaan, bukan cuma saldo akhir per akun. Adjustments (yang
#    umumnya bersifat akrual/reklasifikasi, BUKAN pergerakan kas riil)
#    SENGAJA tidak ikut dihitung di Arus Kas versi ini -- kalau nanti ada
#    Adjustments yang sebetulnya representasi kas riil, perlu ditandai
#    khusus supaya ikut masuk Arus Kas (belum ada mekanismenya di sini).
#
# 3. df_piutang/df_hutang/tren_piutang/tren_utang/jadwal_aset (Buku
#    Bantu Piutang/Hutang/Aktiva Tetap) BELUM dihasilkan dari kertas
#    kerja bank-only ini (kertas kerja ini murni dari rekening koran,
#    tidak ada input piutang/hutang/aset tetap terpisah) -- dikirim
#    kosong/None ke accounting_export, sheet terkait akan tampil kosong
#    di laporan 18-sheet. Ini BUKAN bug, tapi keterbatasan cakupan data
#    sumber -- akuntan tetap bisa isi manual lewat endpoint upload
#    piutang/hutang/aset tetap yang sudah ada terpisah SEBELUM
#    men-generate laporan 18-sheet final, kalau datanya tersedia.
#
# 4. Opening_Balance WAJIB sudah diisi (bukan default 0 semua) supaya
#    Neraca & Perubahan Ekuitas akurat -- fungsi ini menerima df_opening
#    terpisah (baca_opening_balance_dari_kertas_kerja(), lihat di bawah)
#    persis karena HasilKertasKerja TIDAK menyimpan opening balance
#    (opening balance murni input manual akuntan setelah kertas kerja
#    awal digenerate, lihat susun_opening_balance()).
# ============================================================

_KATEGORI_KERTAS_KERJA_KE_DB = {
    "asset": "ASET",
    "contra asset": "ASET",  # tetap ASET -- arah kontra ditentukan Normal Balance, bukan kategori
    "liability": "LIABILITAS",
    "equity": "EKUITAS",
    "revenue": "PENDAPATAN",
    "other income": "PENDAPATAN",
    "expense": "BEBAN",
    "other expense": "BEBAN",
    "income tax": "BEBAN",
}


def _normal_saldo_ke_db(normal_balance_kertas_kerja: Optional[str]) -> Optional[str]:
    """'Debit'/'Credit' (kertas kerja) -> 'DEBET'/'KREDIT' (skema database)."""
    if not normal_balance_kertas_kerja:
        return None
    s = str(normal_balance_kertas_kerja).strip().lower()
    if s == "debit":
        return "DEBET"
    if s == "credit":
        return "KREDIT"
    return None


def konversi_coa_kertas_kerja_ke_peta(
    df_coa: pd.DataFrame, df_opening: Optional[pd.DataFrame] = None,
) -> Dict[str, Dict[str, Any]]:
    """
    Ubah df_coa (skema kertas kerja: Account No./Account Name/Class/
    Normal Balance/Statement/FS Group/Tax Return Group/Notes) jadi peta
    {no_akun: info} PERSIS bentuk lapkeu.peta_akun_dari_coa() -- dipakai
    langsung oleh lapkeu.susun_arus_kas_sederhana() & susun_calk_otomatis()
    tanpa lewat lapkeu.peta_akun_dari_coa() (yang mengharapkan skema
    database "no_akun"/"nama_akun"/"kategori"/dst, bukan skema kertas
    kerja) -- lihat _KATEGORI_KERTAS_KERJA_KE_DB utk pemetaan "Class".

    Kolom "FS Group" dipakai sbg sub_kategori (dipakai heuristik arus kas
    utk mendeteksi "ASET TETAP"/"JANGKA PANJANG"/dst -- lihat FS Group di
    COA contoh: "Fixed Assets" TIDAK persis cocok string yang dicari
    lapkeu._klasifikasi_arus_kas() ("ASET TETAP"/"INVESTASI" dalam Bahasa
    Indonesia, huruf besar) -- KETERBATASAN DIKETAHUI: heuristik arus kas
    kemungkinan salah klasifikasi utk COA berbahasa Inggris seperti
    contoh ini, akan selalu jatuh ke default "operasi". Tim disarankan
    menambah FS Group Indonesia-friendly ATAU perluas
    lapkeu._klasifikasi_arus_kas() supaya kenal istilah Inggris juga,
    sebelum dipakai ke client sungguhan.
    """
    saldo_awal_map: Dict[str, float] = {}
    if df_opening is not None and not df_opening.empty:
        # [FIX -- BUG KRITIS, SIGN-FLIP SALDO AWAL AKUN KREDIT] Sebelumnya
        # baris ini SELALU `debit - credit` (raw, TIDAK memandang Normal
        # Balance akun) -- padahal hitung_saldo_per_akun_dari_tb_monthly()
        # di bawah MENGASUMSIKAN nilai "saldo_awal" di sini SUDAH signed
        # sesuai arah Normal Balance (positif = arah normal), persis pola
        # yang SUDAH BENAR dipakai susun_bs_pnl_monthly() (baris ~1495).
        #
        # Akibatnya: akun Normal Balance KREDIT (Hutang Usaha, Modal, Laba
        # Ditahan, Pendapatan -- lazim punya saldo awal bukan 0 utk client
        # yang sudah beroperasi) tersimpan dgn TANDA TERBALIK. Contoh:
        # Hutang Usaha saldo awal Rp 1.000.000 (kredit) tersimpan sbg
        # Rp -1.000.000 -- lalu saldo_akhir di
        # hitung_saldo_per_akun_dari_tb_monthly() jadi salah sampai 2x
        # nilai saldo awal itu sendiri. Ini MENGALIR LANGSUNG ke Neraca,
        # Perubahan Ekuitas, dan CALK di laporan 18-sheet final client
        # (lihat susun_data_export_18_sheet_dari_kertas_kerja) -- Neraca
        # bisa TIDAK BALANCE / salah tanda utk setiap akun liabilitas/
        # ekuitas yang diisi saldo awal.
        #
        # Fix: flip sesuai Normal Balance, SAMA PERSIS konvensi
        # susun_bs_pnl_monthly() -- (credit - debit) utk akun Kredit,
        # (debit - credit) utk akun Debit. Perlu ambil Normal Balance dari
        # df_coa dulu (baris lama tidak melakukan ini sama sekali).
        _normal_balance_per_akun: Dict[str, str] = {
            str(akun.get("Account No.") or "").strip(): str(akun.get("Normal Balance") or "").strip().lower()
            for _, akun in df_coa.iterrows()
        }
        for _, r in df_opening.iterrows():
            no_akun = str(r.get("Account No.") or "").strip()
            if no_akun:
                debit = pd.to_numeric(r.get("Opening Debit (IDR)"), errors="coerce") or 0
                credit = pd.to_numeric(r.get("Opening Credit (IDR)"), errors="coerce") or 0
                debit, credit = float(debit), float(credit)
                normal_balance = _normal_balance_per_akun.get(no_akun, "")
                # [PENTING] Nilai mentah "Normal Balance" di df_coa itu
                # Bahasa Inggris ("Debit"/"Credit", sudah di-lower di atas)
                # -- BUKAN "debet"/"kredit" Indonesia (itu baru muncul
                # setelah lewat _normal_saldo_ke_db()). Samakan dgn
                # perbandingan yang sudah benar di susun_bs_pnl_monthly()
                # baris ~1495 (`== "debit"`), supaya tidak salah lagi.
                saldo_awal_map[no_akun] = (credit - debit) if normal_balance == "credit" else (debit - credit)

    peta: Dict[str, Dict[str, Any]] = {}
    for _, akun in df_coa.iterrows():
        no_akun = str(akun.get("Account No.") or "").strip()
        if not no_akun:
            continue
        kelas = str(akun.get("Class") or "").strip().lower()
        peta[no_akun] = {
            "nama_akun": akun.get("Account Name") or no_akun,
            "kategori": _KATEGORI_KERTAS_KERJA_KE_DB.get(kelas),
            "sub_kategori": akun.get("FS Group"),
            "normal_saldo": _normal_saldo_ke_db(akun.get("Normal Balance")),
            "saldo_awal": saldo_awal_map.get(no_akun, 0.0),
        }
    return peta


def hitung_saldo_per_akun_dari_tb_monthly(
    df_tb: pd.DataFrame, df_coa: pd.DataFrame, df_opening: Optional[pd.DataFrame] = None,
) -> Dict[str, Dict[str, Any]]:
    """
    Bangun dict saldo_per_akun PERSIS bentuk lapkeu.hitung_saldo_per_akun()
    (no_akun/nama_akun/kategori/sub_kategori/normal_saldo/saldo_awal/
    total_debet/total_kredit/saldo_akhir/dikenal_di_coa/
    jumlah_baris_placeholder/nominal_placeholder/keterangan_perlu_dikoreksi)
    TAPI sumbernya df_tb (TB_Monthly = Bank_Posting_Summary + Adjustments,
    sudah dihitung susun_tb_monthly()) -- bukan jurnal mentah. Lihat
    catatan desain #1 di kepala bagian 12 untuk alasannya.

    Dipakai sbg pengganti langsung lapkeu.hitung_saldo_per_akun() sebelum
    memanggil lapkeu.susun_neraca()/susun_laba_rugi()/
    susun_perubahan_ekuitas()/susun_calk_otomatis() -- fungsi-fungsi itu
    TIDAK diubah sama sekali, cuma dikasih saldo_per_akun dari sumber lain.
    """
    peta_coa = konversi_coa_kertas_kerja_ke_peta(df_coa, df_opening)
    tb_map = {str(r["Account No."]): r for _, r in df_tb.iterrows()} if df_tb is not None and not df_tb.empty else {}

    saldo: Dict[str, Dict[str, Any]] = {}
    # Semua akun COA disertakan (sama seperti lapkeu -- termasuk yang
    # tidak punya pergerakan sama sekali, supaya saldo awal tetap tampil).
    for no_akun, info in peta_coa.items():
        baris_tb = tb_map.get(no_akun)
        total_debet = 0.0
        total_kredit = 0.0
        if baris_tb is not None:
            for bulan in BULAN_URUT:
                total_debet += _angka(baris_tb.get(f"{bulan}_Debit"))
                total_kredit += _angka(baris_tb.get(f"{bulan}_Credit"))

        saldo_awal = _angka(info.get("saldo_awal"))
        pergerakan = total_debet - total_kredit
        normal_saldo = info.get("normal_saldo") or "DEBET"
        saldo_akhir = (saldo_awal - pergerakan) if normal_saldo == "KREDIT" else (saldo_awal + pergerakan)

        saldo[no_akun] = {
            "no_akun": no_akun,
            "nama_akun": info.get("nama_akun") or no_akun,
            "kategori": info.get("kategori"),
            "sub_kategori": info.get("sub_kategori"),
            "normal_saldo": normal_saldo,
            "saldo_awal": saldo_awal,
            "total_debet": round(total_debet, 2),
            "total_kredit": round(total_kredit, 2),
            "saldo_akhir": round(saldo_akhir, 2),
            "dikenal_di_coa": True,  # semua baris di sini SUMBERnya COA itu sendiri
            # [KETERBATASAN] Tidak seperti lapkeu.hitung_saldo_per_akun(),
            # versi ini belum menghitung "berapa baris GL berstatus 'Need
            # Review' yang nyumbang saldo akun ini" -- TODO lanjutan kalau
            # akuntan butuh info itu di level saldo per akun (saat ini
            # sudah bisa dilihat per-transaksi di sheet GL, kolom Review
            # Status/Confidence).
            "jumlah_baris_placeholder": 0,
            "nominal_placeholder": 0.0,
            "keterangan_perlu_dikoreksi": None,
        }
    return saldo


def konversi_gl_ke_jurnal_db(df_gl: pd.DataFrame) -> List[Dict[str, Any]]:
    """
    Ubah GL kertas kerja (1 baris = 1 transaksi bank, kolom "Debit
    Account"/"Credit Account"/"Debit (IDR)"/"Credit (IDR)"/"Date"/"Bank
    Description") jadi list dict skema jurnal database ("tanggal",
    "keterangan", "no_akun_debet", "jml_debet", "no_akun_kredit",
    "jml_kredit") -- HANYA dipakai untuk lapkeu.susun_arus_kas_sederhana()
    (lihat catatan desain #2), BUKAN untuk menghitung saldo_per_akun
    (yang sumbernya TB_Monthly, lihat hitung_saldo_per_akun_dari_tb_monthly).

    Setiap baris GL disini SUDAH berupa pasangan debit/kredit seimbang
    (nominal debit = nominal kredit, lihat susun_sheet_gl()), jadi
    konversinya 1:1 tanpa perlu split/gabung baris.
    """
    if df_gl is None or df_gl.empty:
        return []
    jurnal: List[Dict[str, Any]] = []
    for _, r in df_gl.iterrows():
        jurnal.append({
            "tanggal": r.get("Date"),
            "keterangan": r.get("Bank Description"),
            "no_akun_debet": str(r.get("Debit Account")) if pd.notna(r.get("Debit Account")) else None,
            # [FIX -- BUG NYATA, sama persis yang sudah dijelaskan di
            # docstring _angka() paling atas file ini] Sebelumnya
            # `float(r.get("Debit (IDR)") or 0)` -- pola ini TIDAK aman
            # untuk NaN: di Python `nan or 0` mengembalikan `nan` itu
            # sendiri (NaN truthy), jadi fallback "or 0" tidak pernah
            # kepakai kalau selnya NaN. NaN yang lolos ke sini akan
            # meracuni sum() di lapkeu.susun_arus_kas_sederhana() (arus
            # kas) -- 1 baris NaN bikin total arus kas jadi NaN
            # seterusnya. Ganti pakai _angka() yang SUDAH ADA di modul
            # ini justru untuk mencegah pola ini (harusnya dipakai dari
            # awal, kelewat di titik ini).
            "jml_debet": _angka(r.get("Debit (IDR)")),
            "no_akun_kredit": str(r.get("Credit Account")) if pd.notna(r.get("Credit Account")) else None,
            "jml_kredit": _angka(r.get("Credit (IDR)")),
        })
    return jurnal


def baca_opening_balance_dari_kertas_kerja(file_like) -> pd.DataFrame:
    """
    [BARU] Pasangan baca_adjustments_dari_kertas_kerja() tapi untuk sheet
    'Opening_Balance' -- dibutuhkan supaya saldo awal yang sudah diisi/
    dikoreksi akuntan ikut terpakai di jembatan ke laporan 18-sheet
    (menutup sebagian gap #2 "belum ada pembaca Opening_Balance yang
    dikoreksi", sekaligus prasyarat gap #1 supaya Neraca akurat).

    Header di contoh ada di baris ke-3 (baris 1=judul, 2=kosong, 3=header
    kolom -- dikonfirmasi langsung ke file contoh, BUKAN ditebak --
    urutan kolom persis: Account No./Account Name/Opening Debit (IDR)/
    Opening Credit (IDR)/Net Debit/(Credit)/Review Notes/Status). Baris
    "OPENING BALANCE CONTROL" di bagian bawah sheet (lihat contoh Kertas_
    Kerja_Laporan_Keuangan_2025.xlsx) SENGAJA tidak dibaca fungsi ini --
    itu ringkasan, bukan baris per akun, dan berhenti otomatis begitu
    "Account No." kosong/bukan angka.
    """
    wb = openpyxl.load_workbook(file_like, data_only=True)
    if "Opening_Balance" not in wb.sheetnames:
        raise ValueError("Sheet 'Opening_Balance' tidak ditemukan di file yang diupload.")
    ws = wb["Opening_Balance"]

    header = [c.value for c in ws[3]]
    baris = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        rec = dict(zip(header, row))
        no_akun = rec.get("Account No.")
        if no_akun is None or str(no_akun).strip() == "":
            break  # sudah masuk blok "OPENING BALANCE CONTROL" di bawah tabel akun
        rec["Account No."] = str(no_akun).strip()
        baris.append(rec)
    return pd.DataFrame(baris)


def baca_gl_dari_kertas_kerja(file_like) -> pd.DataFrame:
    """
    [BARU] Baca ulang sheet 'GL' dari file kertas kerja (baru digenerate
    ATAU sudah diupload ulang user) -- dibutuhkan supaya endpoint
    generate-laporan-18-sheet-dari-kertas-kerja bisa bekerja HANYA dari
    1 file .xlsx yang diupload user (tanpa perlu HasilKertasKerja
    tersimpan di memori server dari request generate-kertas-kerja
    sebelumnya, yang notabene sudah selesai/stateless begitu response
    dikirim).

    GL TIDAK diedit user (bukan kolom "input manual" -- lihat
    WARNA_INPUT_MANUAL hanya dipakai di Adjustments/Opening_Balance),
    jadi pembacaan ini murni "reconstruct df_gl apa adanya dari Excel",
    bukan membaca koreksi.
    """
    wb = openpyxl.load_workbook(file_like, data_only=True)
    if "GL" not in wb.sheetnames:
        raise ValueError("Sheet 'GL' tidak ditemukan di file yang diupload.")
    ws = wb["GL"]

    # Baris 1 = catatan/subjudul (lihat _tulis_dataframe subjudul), baris
    # 2 = header kolom, sesuai bagaimana _tulis_dataframe menulis sheet GL
    # (subjudul diisi -> baris_awal = 4... TAPI GL ditulis TANPA judul
    # eksplisit lewat _tulis_dataframe -- baris_awal dihitung ulang sesuai
    # apakah judul/subjudul dipakai. Untuk keamanan, cari baris header
    # dengan mendeteksi baris yang mengandung "Journal ID" persis,
    # daripada hardcode nomor baris.
    baris_header = None
    for i in range(1, 8):
        nilai_baris = [c.value for c in ws[i]]
        if "Journal ID" in nilai_baris:
            baris_header = i
            break
    if baris_header is None:
        raise ValueError("Header sheet 'GL' (kolom 'Journal ID') tidak ditemukan di 6 baris pertama.")

    header = [c.value for c in ws[baris_header]]
    rows = list(ws.iter_rows(min_row=baris_header + 1, values_only=True))
    df = pd.DataFrame(rows, columns=header)
    # Kolom "Date" perlu dipastikan bertipe datetime (openpyxl data_only
    # biasanya sudah mengembalikan datetime asli utk sel bertipe tanggal,
    # tapi dipaksa ulang di sini supaya konsisten dgn df_gl hasil generate
    # langsung, yang datang dari pandas/tanggal asli, bukan re-read Excel).
    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    return df


def susun_data_export_18_sheet_dari_kertas_kerja(
    df_gl: pd.DataFrame,
    df_coa: pd.DataFrame,
    df_adjustments: pd.DataFrame,
    df_opening: pd.DataFrame,
    tahun: int,
    nama_perusahaan: Optional[str] = None,
    prive_atau_dividen: float = 0.0,
    setoran_modal_baru: float = 0.0,
    penyesuaian_ekuitas_manual: float = 0.0,
) -> Dict[str, Any]:
    """
    Fungsi utama jembatan (LANGKAH 1 dari gap yang diminta). Susun ulang
    df_posting/df_tb/df_bs/df_pnl dari df_gl+df_coa+df_adjustments (dgn
    fungsi yang SUDAH ADA -- susun_bank_posting_summary/susun_tb_monthly/
    susun_bs_pnl_monthly, TIDAK diduplikasi), lalu bangun dict "data_export"
    dengan bentuk PERSIS SAMA seperti main.py::_susun_data_export_18_sheet()
    supaya bisa langsung dilempar ke
    accounting_export.export_18_sheet_lengkap()/export_18_sheet_sebagai_json()
    tanpa perubahan di sisi accounting_export.py sama sekali.

    df_gl/df_coa/df_adjustments/df_opening SENGAJA diminta terpisah
    (bukan menerima 1 objek HasilKertasKerja) supaya fungsi ini bisa
    dipakai baik dari HasilKertasKerja yang baru digenerate (endpoint
    generate-kertas-kerja) MAUPUN dari re-read file kertas kerja yang
    sudah dikoreksi user (baca_gl_dari_kertas_kerja +
    muat_coa_kertas_kerja + baca_adjustments_dari_kertas_kerja +
    baca_opening_balance_dari_kertas_kerja) -- pemanggil (endpoint di
    main.py) yang memutuskan sumbernya.

    KETERBATASAN (lihat juga catatan desain #3 di kepala bagian 12):
    df_piutang/df_hutang/tren_piutang/tren_utang/jadwal_aset dikirim
    kosong -- laporan 18-sheet hasil fungsi ini TIDAK akan punya isi di
    sheet Buku Bantu Piutang/Hutang/Aktiva Tetap sampai data itu
    ditambahkan terpisah (bukan cakupan kertas kerja bank-only ini).
    """
    from modules import laporan_keuangan as lapkeu  # import lokal -- hindari import siklik di level modul
    from modules import pph_badan

    periode = str(tahun)

    df_posting = susun_bank_posting_summary(df_gl, df_coa)
    df_tb = susun_tb_monthly(df_posting, df_adjustments)
    # [FIX -- gap #3] df_opening di sini SUDAH bisa berisi saldo awal yang
    # sudah dikoreksi user (baca_opening_balance_dari_kertas_kerja) --
    # dialirkan supaya df_bs/df_pnl (kalau nanti dipakai pemanggil, lihat
    # catatan di bawah) juga sinkron dgn saldo_per_akun/neraca final.
    df_bs, df_pnl = susun_bs_pnl_monthly(df_tb, df_coa, df_opening)  # dipakai tulis_kertas_kerja_excel, bukan di sini

    saldo_per_akun = hitung_saldo_per_akun_dari_tb_monthly(df_tb, df_coa, df_opening)
    peta_coa = konversi_coa_kertas_kerja_ke_peta(df_coa, df_opening)
    jurnal_db = konversi_gl_ke_jurnal_db(df_gl)

    laba_rugi = lapkeu.susun_laba_rugi(saldo_per_akun, periode)
    neraca = lapkeu.susun_neraca(saldo_per_akun, laba_rugi["laba_rugi_bersih"], penyesuaian_ekuitas_manual)
    perubahan_ekuitas = lapkeu.susun_perubahan_ekuitas(
        saldo_per_akun, laba_rugi["laba_rugi_bersih"], prive_atau_dividen, setoran_modal_baru,
        periode, penyesuaian_ekuitas_manual=penyesuaian_ekuitas_manual,
    )
    # [KETERBATASAN -- lihat catatan desain #2] arus_kas dihitung dari GL
    # saja (transaksi bank riil), Adjustments TIDAK ikut.
    arus_kas = lapkeu.susun_arus_kas_sederhana(jurnal_db, peta_coa)
    calk = lapkeu.susun_calk_otomatis(peta_coa, saldo_per_akun, periode)

    laporan_bulanan = _bangun_laporan_bulanan_dari_tb_monthly(df_tb, df_coa, df_opening, tahun)

    # -- PPh Badan 31E -- [KETERBATASAN] koreksi_fiskal_positif/negatif
    # default 0 (fiscal_reconciliation.py butuh data aset tetap yang
    # tidak tersedia dari kertas kerja bank-only ini) -- akuntan WAJIB
    # cek ulang koreksi fiskal secara manual sebelum SPT difinalkan.
    pph_hasil = pph_badan.hitung_pph_pasal_31e(
        peredaran_bruto=laba_rugi.get("total_pendapatan", 0),
        laba_bersih_komersial=laba_rugi.get("laba_rugi_bersih", 0),
        tambahan_peredaran_bruto_lainnya=0,
        retur_pengurangan_peredaran_bruto=0,
        koreksi_fiskal_positif=0,
        koreksi_fiskal_negatif=0,
        kompensasi_kerugian_fiskal=0,
        kredit_pajak=None,
        tahun_pajak=tahun,
        nama_perusahaan=nama_perusahaan,
        skema_pajak="Tarif Umum Pasal 17/31E",
        keterangan_peredaran_bruto=None,
    )

    lap_data_untuk_lampiran = {
        "neraca": neraca, "laba_rugi": laba_rugi, "perubahan_ekuitas": perubahan_ekuitas,
        "arus_kas": arus_kas, "calk": calk,
        # [FIX] susun_lampiran_spt_lengkap_rinci() membaca laporan["meta"]["periode"]
        # (fallback ke laporan["neraca"]["periode"], yang TIDAK ADA di susun_neraca())
        # -- tanpa ini, periode di 3 lampiran rinci selalu kosong string.
        "meta": {"periode": periode},
    }
    lampiran_rinci = lapkeu.susun_lampiran_spt_lengkap_rinci(lap_data_untuk_lampiran, tahun_sebelumnya=None, coa=[
        {
            "no_akun": no, "nama_akun": info["nama_akun"], "kategori": info["kategori"],
            "sub_kategori": info["sub_kategori"], "normal_saldo": info["normal_saldo"],
        }
        for no, info in peta_coa.items()
    ])

    asumsi = {
        "nama_perusahaan": nama_perusahaan or "",
        "periode_awal": f"{tahun}-01-01",
        "periode_akhir": f"{tahun}-12-31",
        "tanggal_laporan": f"{tahun}-12-31",
    }

    return {
        "periode": periode,
        "tahun_sebelumnya": None,
        "coa": [
            {
                "no_akun": no, "nama_akun": info["nama_akun"], "kategori": info["kategori"],
                "sub_kategori": info["sub_kategori"], "normal_saldo": info["normal_saldo"],
                "saldo_awal": info["saldo_awal"],
            }
            for no, info in peta_coa.items()
        ],
        "jurnal": jurnal_db,
        "df_piutang": None,
        "df_hutang": None,
        "jadwal_aset": {},
        "laporan_bulanan": laporan_bulanan,
        "pph_hasil": pph_hasil,
        "neraca": neraca,
        "laba_rugi": laba_rugi,
        "perubahan_ekuitas": perubahan_ekuitas,
        "arus_kas": arus_kas,
        "calk": calk,
        "lampiran_rinci": lampiran_rinci,
        "tren_piutang": [],
        "tren_utang": [],
        "asumsi": asumsi,
    }


def _bangun_laporan_bulanan_dari_tb_monthly(
    df_tb: pd.DataFrame, df_coa: pd.DataFrame, df_opening: Optional[pd.DataFrame], tahun: int,
) -> Dict[str, Any]:
    """
    Bangun dict PERSIS bentuk lapkeu.susun_laporan_bulanan_setahun() --
    {"tahun", "trial_balance_bulanan", "laba_rugi_bulanan",
    "balance_sheet_bulanan", "meta"} -- TAPI dari df_tb (TB_Monthly kertas
    kerja, sudah termasuk Adjustments) secara KUMULATIF per bulan,
    bukan dari jurnal mentah 12x seperti versi lapkeu (yang tidak
    memasukkan Adjustments -- lihat catatan desain #1).

    Cara kerja: untuk tiap bulan ke-N, jumlahkan Debit/Credit TB_Monthly
    dari Jan..bulan-N (kumulatif), lalu terapkan rumus saldo_akhir yang
    SAMA seperti hitung_saldo_per_akun_dari_tb_monthly (saldo_awal +/-
    pergerakan tergantung Normal Balance), lalu panggil
    lapkeu.susun_laba_rugi()/susun_neraca() ulang PERSIS seperti
    lapkeu.susun_laporan_bulanan_setahun() -- fungsi itu sendiri TIDAK
    diduplikasi logikanya, hanya dipanggil 12x dgn saldo_per_akun kumulatif
    yang sumbernya beda.
    """
    from modules import laporan_keuangan as lapkeu

    peta_coa = konversi_coa_kertas_kerja_ke_peta(df_coa, df_opening)
    tb_map = {str(r["Account No."]): r for _, r in df_tb.iterrows()} if df_tb is not None and not df_tb.empty else {}

    per_bulan_saldo, per_bulan_laba_rugi, per_bulan_neraca = [], [], []
    laba_bersih_bulan_sebelumnya = 0.0

    for idx_bulan, bulan_batas in enumerate(BULAN_URUT, start=1):
        bulan_sd_ini = BULAN_URUT[:idx_bulan]
        saldo_per_akun: Dict[str, Dict[str, Any]] = {}
        for no_akun, info in peta_coa.items():
            baris_tb = tb_map.get(no_akun)
            if baris_tb is None:
                total_debet = 0.0
                total_kredit = 0.0
            else:
                total_debet = sum(_angka(baris_tb.get(f"{b}_Debit")) for b in bulan_sd_ini)
                total_kredit = sum(_angka(baris_tb.get(f"{b}_Credit")) for b in bulan_sd_ini)
            saldo_awal = _angka(info.get("saldo_awal"))
            normal_saldo = info.get("normal_saldo") or "DEBET"
            pergerakan = total_debet - total_kredit
            saldo_akhir = (saldo_awal - pergerakan) if normal_saldo == "KREDIT" else (saldo_awal + pergerakan)
            saldo_per_akun[no_akun] = {
                "no_akun": no_akun, "nama_akun": info.get("nama_akun") or no_akun,
                "kategori": info.get("kategori"), "sub_kategori": info.get("sub_kategori"),
                "normal_saldo": normal_saldo, "saldo_awal": saldo_awal,
                "total_debet": round(total_debet, 2), "total_kredit": round(total_kredit, 2),
                "saldo_akhir": round(saldo_akhir, 2), "dikenal_di_coa": True,
                "jumlah_baris_placeholder": 0, "nominal_placeholder": 0.0,
                "keterangan_perlu_dikoreksi": None,
            }

        laba_rugi_ytd = lapkeu.susun_laba_rugi(saldo_per_akun, periode=f"{tahun}-{idx_bulan:02d}")
        neraca = lapkeu.susun_neraca(saldo_per_akun, laba_rugi_ytd["laba_rugi_bersih"])
        laba_rugi_ytd["laba_bersih_bulanan"] = round(
            laba_rugi_ytd["laba_rugi_bersih"] - laba_bersih_bulan_sebelumnya, 2
        )
        laba_bersih_bulan_sebelumnya = laba_rugi_ytd["laba_rugi_bersih"]

        per_bulan_saldo.append(saldo_per_akun)
        per_bulan_laba_rugi.append(laba_rugi_ytd)
        per_bulan_neraca.append(neraca)

    trial_balance_bulanan: Dict[str, Any] = {}
    for no_akun in sorted(peta_coa.keys()):
        info_pertama = next((s[no_akun] for s in per_bulan_saldo if no_akun in s), {})
        trial_balance_bulanan[no_akun] = {
            "nama_akun": info_pertama.get("nama_akun", no_akun),
            "kategori": info_pertama.get("kategori"),
            "per_bulan": [round(s.get(no_akun, {}).get("saldo_akhir", 0.0), 2) for s in per_bulan_saldo],
            "keterangan_perlu_dikoreksi": None,
        }

    pendapatan_ytd_list = [lr["total_pendapatan"] for lr in per_bulan_laba_rugi]
    total_pendapatan_bulanan, pendapatan_sebelumnya = [], 0.0
    for v in pendapatan_ytd_list:
        total_pendapatan_bulanan.append(round(v - pendapatan_sebelumnya, 2))
        pendapatan_sebelumnya = v

    laba_rugi_bulanan = {
        "total_pendapatan_ytd": pendapatan_ytd_list,
        "total_beban_ytd": [lr["total_beban"] for lr in per_bulan_laba_rugi],
        "laba_bersih_ytd": [lr["laba_rugi_bersih"] for lr in per_bulan_laba_rugi],
        "laba_bersih_bulanan": [lr["laba_bersih_bulanan"] for lr in per_bulan_laba_rugi],
        "total_pendapatan_bulanan": total_pendapatan_bulanan,
    }
    balance_sheet_bulanan = {
        "total_aset": [n["total_aset"] for n in per_bulan_neraca],
        "total_liabilitas": [n["total_liabilitas"] for n in per_bulan_neraca],
        "total_ekuitas": [n["total_ekuitas"] for n in per_bulan_neraca],
        "balance": [n["balance"] for n in per_bulan_neraca],
    }
    bulan_tidak_balance = [i + 1 for i, ok in enumerate(balance_sheet_bulanan["balance"]) if not ok]

    return {
        "tahun": tahun,
        "trial_balance_bulanan": trial_balance_bulanan,
        "laba_rugi_bulanan": laba_rugi_bulanan,
        "balance_sheet_bulanan": balance_sheet_bulanan,
        "meta": {
            "jumlah_baris_jurnal": int(len(df_tb) if df_tb is not None else 0),
            "jumlah_baris_tanpa_tanggal": 0,
            "bulan_tidak_balance": bulan_tidak_balance,
            "peringatan": (
                [f"Balance Sheet tidak balance di bulan: {bulan_tidak_balance}"] if bulan_tidak_balance else []
            ),
        },
    }